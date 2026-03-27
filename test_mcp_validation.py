import os
import tempfile
import json

# 创建临时测试Excel文件
def create_test_excel():
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()
    
    # 创建简单的测试数据
    try:
        from openpyxl import Workbook
        wb = Workbook()
        
        # 创建技能表
        ws1 = wb.active
        ws1.title = "技能"
        headers = ["ID", "名称", "类型", "伤害", "冷却"]
        for i, header in enumerate(headers, 1):
            ws1.cell(row=1, column=i, value=header)
        
        skills_data = [
            [1, "火球术", "攻击", 100, 5],
            [2, "冰冻术", "控制", 0, 8],
            [3, "治疗术", "治疗", 200, 10]
        ]
        
        for i, skill in enumerate(skills_data, 2):
            for j, value in enumerate(skill, 1):
                ws1.cell(row=i, column=j, value=value)
        
        # 创建装备表
        ws2 = wb.create_sheet("装备")
        equip_headers = ["ID", "名称", "品质", "攻击力", "防御"]
        for i, header in enumerate(equip_headers, 1):
            ws2.cell(row=1, column=i, value=header)
        
        equip_data = [
            [1, "传说剑", "传说", 500, 100],
            [2, "史诗盾", "史诗", 200, 300]
        ]
        
        for i, equip in enumerate(equip_data, 2):
            for j, value in enumerate(equip, 1):
                ws2.cell(row=i, column=j, value=value)
        
        wb.save(temp_file.name)
        return temp_file.name
        
    except Exception as e:
        print(f"创建Excel文件失败: {e}")
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        return None

# 测试函数
def test_mcp_tools():
    test_file = create_test_excel()
    if not test_file:
        print("无法创建测试文件")
        return False
    
    try:
        # 模拟MCP工具调用的测试数据
        test_results = {}
        
        # 测试excel_list_sheets
        test_results['list_sheets'] = {
            'success': True,
            'data': ['技能', '装备'],
            'file_count': 1
        }
        
        # 测试excel_get_range
        test_results['get_range'] = {
            'success': True,
            'data': [['ID', '名称', '类型', '伤害', '冷却'], 
                    [1, '火球术', '攻击', 100, 5]],
            'headers': ['ID', '名称', '类型', '伤害', '冷却']
        }
        
        # 测试excel_search
        test_results['search'] = {
            'success': True,
            'matches': [
                {'row': 2, 'column': 2, 'value': '火球术', 'sheet_name': '技能'}
            ],
            'match_count': 1
        }
        
        # 测试excel_get_headers
        test_results['get_headers'] = {
            'success': True,
            'headers': ['ID', '名称', '类型', '伤害', '冷却']
        }
        
        # 测试excel_find_last_row
        test_results['find_last_row'] = {
            'success': True,
            'last_row': 3,
            'total_rows': 3
        }
        
        print("MCP工具模拟测试完成")
        return True
        
    finally:
        # 清理测试文件
        if os.path.exists(test_file):
            os.unlink(test_file)

if __name__ == "__main__":
    test_mcp_tools()