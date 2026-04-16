# -*- coding: utf-8 -*-
"""
format_cells R63 迭代测试 - 深度边缘 case 第三轮

新增测试场景（覆盖 R57-R62 未涉及的场景）:
  1. merge + format_cells 组合操作：先合并再格式化 / 格式化含合并单元格的范围
  2. wrap_text 显式关闭（False）验证
  3. text_rotation 边界值：0, 90, -45→45, 200→180, 字符串输入
  4. shrink_to_fit 参数独立测试
  5. strikethrough 开关切换
  6. 无效 preset 名称容错
  7. range 表达式含 '!' 分隔符
  8. 单元格格式化幂等性（同一格式应用两次）
  9. border diagonal 方向设置
  10. number_format 重置为 General
  11. 混合扁平+嵌套键（同时传 bold 和 font.size）
  12. format_cells 对整行/整列范围
  13. 中文字体名 + 中文 sheet 名组合
  14. format_cells 后读取 merged_cell_range 行为
"""

import os
import pytest
import tempfile
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from excel_mcp_server_fastmcp.core.excel_writer import ExcelWriter
from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations


# ==================== Helper ====================

def _create_test_xlsx(file_path: str, rows: int = 5, cols: int = 4, sheet_name: str = "Sheet1"):
    """创建测试用 xlsx 文件，含数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=r * 10 + c)
    wb.save(file_path)
    wb.close()


def _read_cell_style(file_path: str, cell_ref: str = "A1", sheet_name: str = "Sheet1") -> dict:
    """读取单元格的样式属性用于断言"""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    cell = ws[cell_ref]
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    result = {
        "bold": font.bold,
        "italic": font.italic,
        "underline": font.underline,
        "strikethrough": font.strikethrough,
        "size": font.size,
        "font_name": font.name,
        "color": str(font.color) if font.color else None,
        "fill_type": fill.fill_type if fill else None,
        "fgColor": str(fill.fgColor) if fill and hasattr(fill, 'fgColor') and fill.fgColor else None,
        "alignment_h": alignment.horizontal,
        "alignment_v": alignment.vertical,
        "wrap_text": bool(alignment.wrap_text) if alignment.wrap_text is not None else False,
        "text_rotation": alignment.text_rotation,
        "number_format": cell.number_format,
        "border_left": border.left.style if border and border.left else None,
        "border_right": border.right.style if border and border.right else None,
        "border_top": border.top.style if border and border.top else None,
        "border_bottom": border.bottom.style if border and border.bottom else None,
        "value": cell.value,
        "indent": alignment.indent,
        "shrink_to_fit": alignment.shrink_to_fit,
    }
    wb.close()
    return result


# ==================== Test Class ====================

class TestFormatCellsR63:
    """format_cells R63 第九轮测试套件 — 深度边缘 case 第三轮"""

    # ---------- 1. merge + format_cells 组合操作 ----------

    def test_merge_then_format_merged_range(self, tmp_path):
        """先合并 A1:C1，再对合并区域格式化"""
        fp = str(tmp_path / "merge_fmt.xlsx")
        _create_test_xlsx(fp, 3, 5)

        # 先合并
        r_merge = ExcelOperations.merge_cells(fp, "Sheet1", "A1:C1")
        assert r_merge["success"], f"合并失败: {r_merge.get('message')}"

        # 再格式化合并区域
        r_fmt = ExcelOperations.format_cells(fp, "Sheet1", "A1:C1",
            formatting={"bold": True, "bg_color": "FF0000", "alignment": "center"})
        assert r_fmt["success"], f"格式化失败: {r_fmt.get('message')}"

        # 验证左上角单元格（合并区域的代表）的样式
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True, "合并区域应加粗"
        assert style["alignment_h"] == "center", "合并区域应居中"

    def test_format_range_containing_merged_cells(self, tmp_path):
        """格式化一个包含已合并单元格的大范围"""
        fp = str(tmp_path / "fmt_contains_merge.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for r in range(1, 6):
            for c in range(1, 6):
                ws.cell(row=r, column=c, value=r * 10 + c)
        # 合并 B2:C3
        ws.merge_cells("B2:C3")
        wb.save(fp)
        wb.close()

        # 格式化包含合并区域的大范围 A1:E5
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:E5",
            formatting={"italic": True, "font_size": 12})
        assert result["success"], f"失败: {result.get('message')}"
        # 验证非合并区域正常
        style = _read_cell_style(fp, "A1")
        assert style["italic"] is True
        assert style["size"] == 12

    def test_format_then_merge_preserves_format(self, tmp_path):
        """先格式化再合并，格式应保留在合并区域代表格"""
        fp = str(tmp_path / "fmt_then_merge.xlsx")
        _create_test_xlsx(fp, 3, 3)

        # 先格式化
        ExcelOperations.format_cells(fp, "Sheet1", "A1:B1",
            formatting={"bold": True, "bg_color": "00FF00"})
        # 再合并
        ExcelOperations.merge_cells(fp, "Sheet1", "A1:B1")

        # 验证合并后左上角保留格式
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True, "合并后应保留加粗"

    # ---------- 2. wrap_text 显式关闭 ----------

    def test_wrap_text_false_turns_off(self, tmp_path):
        """wrap_text=False 应关闭自动换行"""
        fp = str(tmp_path / "wrap_off.xlsx")
        _create_test_xlsx(fp)

        # 先开启
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"wrap_text": True})
        style_on = _read_cell_style(fp, "A1")
        assert style_on["wrap_text"] is True, "开启 wrap_text 应生效"

        # 再显式关闭
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"wrap_text": False})
        style_off = _read_cell_style(fp, "A1")
        assert style_off["wrap_text"] is False, "wrap_text=False 应关闭换行"

    # ---------- 3. text_rotation 边界值 ----------

    def test_text_rotation_zero(self, tmp_path):
        """text_rotation=0 水平文本"""
        fp = str(tmp_path / "rot_zero.xlsx")
        _create_test_xlsx(fp)
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"text_rotation": 0})
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 0

    def test_text_rotation_90_vertical(self, tmp_path):
        """text_rotation=90 垂直文本"""
        fp = str(tmp_path / "rot_90.xlsx")
        _create_test_xlsx(fp)
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"text_rotation": 90})
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 90

    def test_text_rotation_negative_clamped(self, tmp_path):
        """负数 text_rotation 取绝对值（-45 → 45）"""
        fp = str(tmp_path / "rot_neg.xlsx")
        _create_test_xlsx(fp)
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"text_rotation": -45})
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 45, "负数旋转角度应取绝对值"

    def test_text_rotation_over_180_clamped(self, tmp_path):
        """超过 180 的 text_rotation 被限制为 180"""
        fp = str(tmp_path / "rot_over.xlsx")
        _create_test_xlsx(fp)
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"text_rotation": 270})
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 180, "超过180应clamp到180"

    def test_text_rotation_string_input(self, tmp_path):
        """text_rotation 字符串输入 '45' 应被接受"""
        fp = str(tmp_path / "rot_str.xlsx")
        _create_test_xlsx(fp)
        # openpyxl Alignment 接受 int，字符串可能触发异常或被转换
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"text_rotation": "45"})
        # 不崩溃即可，具体行为取决于 openpyxl
        assert result["success"] or "rotation" in result.get("message", "").lower()

    # ---------- 4. shrink_to_fit 参数 ----------

    def test_shrink_to_fit_true(self, tmp_path):
        """shrink_to_fit=True 启用缩小字体填充"""
        fp = str(tmp_path / "shrink_on.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"shrink_to_fit": True})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["shrink_to_fit"] is True

    def test_shrink_to_fit_false(self, tmp_path):
        """shrink_to_fit=False 关闭缩小字体填充"""
        fp = str(tmp_path / "shrink_off.xlsx")
        _create_test_xlsx(fp)
        # 先开启
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"shrink_to_fit": True})
        # 再关闭
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"shrink_to_fit": False})
        assert result["success"]
        style = _read_cell_style(fp, "A1")
        # openpyxl 可能将 False 存储为 None（默认值）
        assert style["shrink_to_fit"] is False or style["shrink_to_fit"] is None

    # ---------- 5. strikethrough 开关切换 ----------

    def test_strikethrough_toggle_on_off(self, tmp_path):
        """strikethrough 先开后关"""
        fp = str(tmp_path / "strike_toggle.xlsx")
        _create_test_xlsx(fp)

        # 开启删除线
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"strikethrough": True})
        assert _read_cell_style(fp, "A1")["strikethrough"] is True

        # 关闭删除线
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"strikethrough": False})
        assert _read_cell_style(fp, "A1")["strikethrough"] is False

    # ---------- 6. 无效 preset 容错 ----------

    def test_invalid_preset_name_fallback(self, tmp_path):
        """不存在的 preset 名称应回退为空格式或报错"""
        fp = str(tmp_path / "bad_preset.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            preset="nonexistent_preset",
            formatting={"bold": True})
        # 无效 preset 应不影响用户提供的 formatting
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True, "无效 preset 不应影响用户 formatting"

    def test_preset_only_no_formatting(self, tmp_path):
        """只传 preset 不传 formatting"""
        fp = str(tmp_path / "preset_only.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", preset="title")
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True, "title preset 应加粗"
        assert style["size"] == 14, "title preset 字号应为 14"

    # ---------- 7. range 含 '!' 分隔符 ----------

    def test_range_with_sheet_separator(self, tmp_path):
        """range 已含 Sheet!A1 格式"""
        fp = str(tmp_path / "range_sep.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "Sheet1!A1:B2",
            formatting={"bg_color": "FF00FF"})
        assert result["success"], f"失败: {result.get('message')}"

    def test_range_with_different_sheet_in_separator(self, tmp_path):
        """range 中的 sheet 名与 sheet_name 参数不一致时以 range 为准"""
        fp = str(tmp_path / "range_diff_sheet.xlsx")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        ws1["A1"] = "s1"
        ws2["A1"] = "s2"
        wb.save(fp)
        wb.close()

        # range 指定 Sheet2!A1，sheet_name 参数是 Sheet1
        result = ExcelOperations.format_cells(fp, "Sheet1", "Sheet2!A1",
            formatting={"bold": True})
        assert result["success"], f"失败: {result.get('message')}"
        # 验证 Sheet2 的 A1 被格式化了
        style = _read_cell_style(fp, "A1", sheet_name="Sheet2")
        assert style["bold"] is True, "range 中的 sheet 名应优先"

    # ---------- 8. 格式化幂等性 ----------

    def test_idempotent_same_format_twice(self, tmp_path):
        """同一格式应用两次结果一致"""
        fp = str(tmp_path / "idem_fmt.xlsx")
        _create_test_xlsx(fp)

        fmt = {"bold": True, "font_size": 16, "bg_color": "AABBCC", "alignment": "right"}
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting=fmt)
        style1 = _read_cell_style(fp, "A1")

        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting=fmt)
        style2 = _read_cell_style(fp, "A1")

        assert style1["bold"] == style2["bold"]
        assert style1["size"] == style2["size"]
        assert style1["alignment_h"] == style2["alignment_h"]

    # ---------- 9. border diagonal 设置 ----------

    def test_border_with_diagonal(self, tmp_path):
        """border 包含 diagonal 设置"""
        fp = str(tmp_path / "bdr_diag.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"border": {
                "top": "thin",
                "bottom": "thin",
                "diagonal": "thin",
                "diagonal_direction": 1,  # 从左上到右下
            }})
        assert result["success"], f"失败: {result.get('message')}"

    def test_border_all_sides_detailed(self, tmp_path):
        """border 四边 + diagonal 全部详细设置"""
        fp = str(tmp_path / "bdr_full.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:B2",
            formatting={"border": {
                "left": {"style": "thin", "color": "FF0000"},
                "right": {"style": "medium", "color": "00FF00"},
                "top": {"style": "thick", "color": "0000FF"},
                "bottom": {"style": "dashed", "color": "FFFF00"},
                "diagonal": {"style": "double", "color": "FF00FF"},
                "color": "000000",  # 默认边框色
            }})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["border_top"] == "thick"
        assert style["border_bottom"] == "dashed"

    # ---------- 10. number_format 重置为 General ----------

    def test_number_format_reset_to_general(self, tmp_path):
        """number_format 设为 'General' 重置为默认"""
        fp = str(tmp_path / "nf_general.xlsx")
        _create_test_xlsx(fp)

        # 先设为百分比
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"number_format": "0.00%"})
        assert "%" in _read_cell_style(fp, "A1")["number_format"]

        # 重置为 General
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"number_format": "General"})
        style = _read_cell_style(fp, "A1")
        assert style["number_format"] == "General"

    def test_number_format_empty_string(self, tmp_path):
        """number_format 空字符串行为"""
        fp = str(tmp_path / "nf_empty.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"number_format": ""})
        # 空字符串可能被 openpyxl 接受或拒绝
        assert result["success"]

    # ---------- 11. 混合扁平+嵌套键 ----------

    def test_mixed_flat_and_nested_keys(self, tmp_path):
        """同时传扁平键 (bold) 和嵌套键 (font.size)"""
        fp = str(tmp_path / "mixed_keys.xlsx")
        _create_test_xlsx(fp)
        # 由于 normalize 检测到嵌套键就直接返回原样，扁平键会被透传
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={
                "bold": True,           # 扁平键
                "font": {"size": 18},   # 嵌套键
                "bg_color": "FF0000",   # 扁平键
            })
        assert result["success"], f"失败: {result.get('message')}"

    # ---------- 12. 整行/整列范围格式化 ----------

    def test_format_entire_row(self, tmp_path):
        """格式化整行范围 1:1"""
        fp = str(tmp_path / "fmt_row.xlsx")
        _create_test_xlsx(fp, 5, 5)
        result = ExcelOperations.format_cells(fp, "Sheet1", "1:1",
            formatting={"bold": True, "bg_color": "EEEEEE"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True

    def test_format_entire_column(self, tmp_path):
        """格式化整列范围 A:A"""
        fp = str(tmp_path / "fmt_col.xlsx")
        _create_test_xlsx(fp, 5, 5)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A:A",
            formatting={"italic": True, "font_color": "0000FF"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["italic"] is True

    # ---------- 13. 中文字体名 + 中文 sheet 名组合 ----------

    def test_chinese_font_and_chinese_sheet(self, tmp_path):
        """中文字体名 + 中文工作表名组合场景"""
        fp = str(tmp_path / "zh_combo.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "数据表"
        ws["A1"] = "测试"
        wb.save(fp)
        wb.close()

        result = ExcelOperations.format_cells(fp, "数据表", "A1",
            formatting={
                "font_name": "微软雅黑",
                "bold": True,
                "font_size": 12,
                "bg_color": "FFF2CC",
                "alignment": "center",
            })
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1", sheet_name="数据表")
        assert style["font_name"] == "微软雅黑"
        assert style["bold"] is True
        assert style["size"] == 12
        assert style["alignment_h"] == "center"

    # ---------- 14. format_cells 后 merge/unmerge 数据保持 ----------

    def test_unmerge_then_format_individual(self, tmp_path):
        """先取消合并再逐个格式化"""
        fp = str(tmp_path / "unmerge_fmt.xlsx")
        _create_test_xlsx(fp, 3, 3)

        # 合并
        ExcelOperations.merge_cells(fp, "Sheet1", "A1:B1")
        # 取消合并
        r_unmerge = ExcelOperations.unmerge_cells(fp, "Sheet1", "A1:B1")
        assert r_unmerge["success"] or "not merged" in r_unmerge.get("message", "").lower() or "no" in r_unmerge.get("message", "").lower()

        # 分别格式化原来的合并区域
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"bold": True})
        ExcelOperations.format_cells(fp, "Sheet1", "B1", formatting={"italic": True})

        style_a = _read_cell_style(fp, "A1")
        style_b = _read_cell_style(fp, "B1")
        assert style_a["bold"] is True
        assert style_b["italic"] is True

    # ---------- 15. preset=data 验证 ----------

    def test_preset_data_default_font(self, tmp_path):
        """preset='data' 应用默认数据字体"""
        fp = str(tmp_path / "preset_data.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:C3", preset="data")
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["font_name"] == "微软雅黑", "data preset 字体应为微软雅黑"
        assert style["size"] == 10, "data preset 字号应为 10"

    # ---------- 16. format_cells 单元格值为 None/空 ----------

    def test_format_empty_cell(self, tmp_path):
        """格式化空值单元格不应崩溃"""
        fp = str(tmp_path / "fmt_empty.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # A1 不设值（None）
        wb.save(fp)
        wb.close()

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": True, "bg_color": "FF0000"})
        assert result["success"], f"失败: {result.get('message')}"

    def test_format_cell_with_long_text(self, tmp_path):
        """格式化含长文本的单元格"""
        fp = str(tmp_path / "fmt_long_text.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "这是一段很长的中文文本用于测试自动换行功能是否正常工作" * 5
        wb.save(fp)
        wb.close()

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"wrap_text": True, "font_size": 11, "font_name": "微软雅黑"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["wrap_text"] is True

    # ---------- 17. border_style=dict 类型传入 ----------

    def test_border_style_as_dict_value(self, tmp_path):
        """border_style 传入 dict 而非字符串"""
        fp = str(tmp_path / "bdr_style_dict.xlsx")
        _create_test_xlsx(fp)
        # border_style 正常只接受字符串，但代码有 dict 分支
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"border_style": {"top": "medium", "custom": "value"}})
        # 不管成功失败，不崩溃就行
        assert result["success"] or result["success"] is False

    # ---------- 18. italic 独立开关 ----------

    def test_italic_toggle(self, tmp_path):
        """italic 开关切换"""
        fp = str(tmp_path / "italic_toggle.xlsx")
        _create_test_xlsx(fp)

        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"italic": True})
        assert _read_cell_style(fp, "A1")["italic"] is True

        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"italic": False})
        assert _read_cell_style(fp, "A1")["italic"] is False

    # ---------- 19. font_color RGB 颜色深度验证 ----------

    def test_font_color_rgb_without_hash(self, tmp_path):
        """font_color 不带 # 的 6 位 HEX"""
        fp = str(tmp_path / "fc_hex.xlsx")
        _create_test_xlsx(fp)
        for color in ["FF0000", "00FF00", "0000FF", "123ABC", "abcdef"]:
            result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
                formatting={"font_color": color})
            assert result["success"], f"颜色 {color} 失败"

    def test_bg_color_various_formats(self, tmp_path):
        """bg_color 各种格式：3位、6位、带#"""
        fp = str(tmp_path / "bgc_various.xlsx")
        _create_test_xlsx(fp)
        # 3 位短格式
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"bg_color": "F00"})
        # 6 位标准
        ExcelOperations.format_cells(fp, "Sheet1", "A2", formatting={"bg_color": "AABBCC"})
        # 带 #
        ExcelOperations.format_cells(fp, "Sheet1", "A3", formatting={"bg_color": "#112233"})
        # 全部成功即可
        style = _read_cell_style(fp, "A3")
        assert style["fill_type"] is not None

    # ---------- 20. vertical_alignment 各选项 ----------

    def test_vertical_alignment_options(self, tmp_path):
        """vertical_alignment 所有有效选项"""
        fp = str(tmp_path / "valign_opts.xlsx")
        _create_test_xlsx(fp, 1, 5)

        for i, val in enumerate(["top", "center", "bottom", "middle", "justify"], start=1):
            col = get_column_letter(i)
            result = ExcelOperations.format_cells(fp, "Sheet1", f"{col}1",
                formatting={"vertical_alignment": val})
            assert result["success"], f"vertical_alignment={val} 失败: {result.get('message')}"

    # ---------- 21. alignment horizontal 各选项 ----------

    def test_horizontal_alignment_all_options(self, tmp_path):
        """horizontal alignment 所有选项"""
        fp = str(tmp_path / "halign_opts.xlsx")
        _create_test_xlsx(fp, 1, 7)

        options = ["left", "center", "right", "justify", "centerContinuous", "distributed", "general"]
        for i, val in enumerate(options, start=1):
            col = get_column_letter(i)
            result = ExcelOperations.format_cells(fp, "Sheet1", f"{col}1",
                formatting={"alignment": val})
            assert result["success"], f"alignment={val} 失败: {result.get('message')}"

        # 验证 center 生效
        style = _read_cell_style(fp, "B1")
        assert style["alignment_h"] in ("center", "centerContinuous"), "center 应生效"
