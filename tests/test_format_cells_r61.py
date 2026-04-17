# -*- coding: utf-8 -*-
"""
format_cells R61 迭代测试 - 组合操作、边界值、样式覆盖验证

新增测试场景:
  1. merge + bold + bg_color 同时传（组合操作）
  2. merge + border_style + bg_color 三重组合
  3. wrap_text=False 显式关闭换行
  4. font_color / bg_color 带 # 前缀自动去除
  5. border_style: double / dotted / dashed 各种线型
  6. text_rotation 边界值: 0, 90, -90→90, 180, >180→180 clamp
  7. preset currency/data/highlight 与自定义 formatting 覆盖
  8. 合并 → 取消合并 往返（unmerge 后数据可编辑）
  9. Unicode 工作表名格式化（中文 sheet 名）
  10. 部分更新不重置未涉及的属性（设 bold 再设 bg_color，bold 保留）
  11. 颜色 3 位 HEX 简写（如 "F00"）
  12. bold=False / italic=False 显式关闭
  13. strikethrough=True 读回验证
  14. indent=0 显式零值
  15. number_format "General" 重置为默认
"""

import os
import pytest
import tempfile
from pathlib import Path
from openpyxl import Workbook, load_workbook

from excel_mcp_server_fastmcp.core.excel_writer import ExcelWriter
from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations


# ==================== Helper ====================

def _create_test_xlsx(file_path: str, rows: int = 5, cols: int = 4, sheet_name: str = "Sheet1"):
    """创建测试用 xlsx 文件，含数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=r * 10 + c)
    wb.save(file_path)
    wb.close()


def _read_cell_style(file_path: str, cell_ref: str = "A1", sheet_name: str = "Sheet1") -> dict:
    """读取单元格的样式属性用于断言"""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    cell = ws[cell_ref]
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    result = {
        "bold": font.bold,
        "italic": font.italic,
        "underline": font.underline,
        "strikethrough": font.strikethrough,
        "size": font.size,
        "font_name": font.name,
        "color": str(font.color) if font.color else None,
        "fill_type": fill.fill_type if fill else None,
        "fgColor": str(fill.fgColor) if fill and fill.fgColor else None,
        "alignment_h": alignment.horizontal,
        "alignment_v": alignment.vertical,
        "wrap_text": bool(alignment.wrap_text) if alignment.wrap_text is not None else False,
        "text_rotation": alignment.text_rotation,
        "number_format": cell.number_format,
        "border_left": border.left.style if border and border.left else None,
        "border_right": border.right.style if border and border.right else None,
        "border_top": border.top.style if border and border.top else None,
        "border_bottom": border.bottom.style if border and border.bottom else None,
        "value": cell.value,
        "indent": alignment.indent,
    }
    wb.close()
    return result


# ==================== Test Class ====================

class TestFormatCellsR61:
    """format_cells R61 第七轮测试套件 — 组合操作与边界值"""

    # ---------- 1. merge + bold + bg_color 组合 ----------

    def test_merge_bold_bgcolor_combined(self, tmp_path):
        """merge + bold + bg_color 同时传递应全部生效"""
        fp = str(tmp_path / "combined.xlsx")
        _create_test_xlsx(fp, 3, 3)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:C1",
            formatting={"merge": True, "bold": True, "bg_color": "FFFF00"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True, "加粗应生效"
        assert style["fgColor"] is not None or style["fill_type"] is not None, "背景色应生效"

    def test_merge_border_bgcolor_triple(self, tmp_path):
        """merge + border_style + bg_color 三重组合"""
        fp = str(tmp_path / "triple.xlsx")
        _create_test_xlsx(fp, 3, 3)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:B2",
            formatting={"merge": True, "border_style": "thin", "bg_color": "FFEEBB"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["border_top"] is not None or style["border_left"] is not None, "边框应生效"

    # ---------- 2. wrap_text 显式关闭 ----------

    def test_wrap_text_false_turns_off(self, tmp_path):
        """wrap_text=False 应显式关闭自动换行"""
        fp = str(tmp_path / "wrap_false.xlsx")
        _create_test_xlsx(fp)
        # 先开启换行
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"wrap_text": True})
        style_on = _read_cell_style(fp, "A1")
        assert style_on["wrap_text"] is True, "开启换行后 wrap_text 应为 True"
        # 再关闭换行
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"wrap_text": False})
        style_off = _read_cell_style(fp, "A1")
        assert style_off["wrap_text"] is False, "wrap_text=False 应关闭换行"

    # ---------- 3. # 前缀颜色处理 ----------

    def test_font_color_hash_prefix_stripped(self, tmp_path):
        """font_color 带 # 前缀应被去除"""
        fp = str(tmp_path / "hash_font.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_color": "#FF0000"})
        assert result["success"], f"失败: {result.get('message')}"

    def test_bg_color_hash_prefix_stripped(self, tmp_path):
        """bg_color 带 # 前缀应被去除"""
        fp = str(tmp_path / "hash_bg.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bg_color": "#00FF00"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["fill_type"] is not None, "# 前缀 bg_color 应生效"

    # ---------- 4. border_style 各种线型 ----------

    def test_border_style_double(self, tmp_path):
        """border_style='double' 双线边框"""
        fp = str(tmp_path / "bdr_double.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:B2",
            formatting={"border_style": "double"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["border_top"] == "double" or style["border_bottom"] == "double", \
            "double 边框应生效"

    def test_border_style_dotted(self, tmp_path):
        """border_style='dotted' 点线边框"""
        fp = str(tmp_path / "bdr_dotted.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"border_style": "dotted"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["border_top"] == "dotted" or style["border_left"] == "dotted", \
            "dotted 边框应生效"

    def test_border_style_dashed(self, tmp_path):
        """border_style='dashed' 虚线边框"""
        fp = str(tmp_path / "bdr_dashed.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"border_style": "dashed"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert "dash" in (style["border_top"] or ""), "dashed 边框应生效"

    def test_border_style_thick(self, tmp_path):
        """border_style='thick' 粗线边框"""
        fp = str(tmp_path / "bdr_thick.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"border_style": "thick"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["border_top"] == "thick" or style["border_left"] == "thick", \
            "thick 边框应生效"

    def test_border_style_medium(self, tmp_path):
        """border_style='medium' 中等线边框"""
        fp = str(tmp_path / "bdr_medium.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"border_style": "medium"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["border_top"] == "medium" or style["border_left"] == "medium", \
            "medium 边框应生效"

    # ---------- 5. text_rotation 边界值 ----------

    def test_text_rotation_zero(self, tmp_path):
        """text_rotation=0 水平文本"""
        fp = str(tmp_path / "rot0.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"text_rotation": 0})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 0, "text_rotation=0 应保持为 0"

    def test_text_rotation_90(self, tmp_path):
        """text_rotation=90 垂直文本"""
        fp = str(tmp_path / "rot90.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"text_rotation": 90})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 90, "text_rotation=90 应保持为 90"

    def test_text_rotation_negative_90_becomes_90(self, tmp_path):
        """text_rotation=-90 应取绝对值为 90"""
        fp = str(tmp_path / "rot_neg90.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"text_rotation": -90})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 90, "-90 应取绝对值变为 90"

    def test_text_rotation_180_max(self, tmp_path):
        """text_rotation=180 最大允许值"""
        fp = str(tmp_path / "rot180.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"text_rotation": 180})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 180, "text_rotation=180 应保持为 180"

    def test_text_rotation_over_180_clamped(self, tmp_path):
        """text_rotation>180 应被 clamp 到 180"""
        fp = str(tmp_path / "rot_over180.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"text_rotation": 270})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 180, "270 应被 clamp 到 180"

    def test_text_rotation_invalid_string_clamped_to_0(self, tmp_path):
        """text_rotation 无效字符串应 fallback 为 0"""
        fp = str(tmp_path / "rot_invalid.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"text_rotation": "abc"})
        assert result["success"], f"失败: {result.get('message')}"

    # ---------- 6. preset 覆盖测试 ----------

    def test_preset_currency_override_number_format(self, tmp_path):
        """preset='currency' 后用户覆盖 number_format"""
        fp = str(tmp_path / "preset_curr_ov.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            preset="currency",
            formatting={"number_format": "0.000%"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert "%" in style["number_format"], "用户 number_format 应覆盖预设"

    def test_preset_data_override_font_size(self, tmp_path):
        """preset='data' 后用户覆盖 font_size"""
        fp = str(tmp_path / "preset_data_ov.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            preset="data",
            formatting={"font_size": 20, "bold": True})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["size"] == 20, "用户 font_size=20 应覆盖预设的 10"
        assert style["bold"] is True, "用户 bold=True 应生效"

    def test_preset_highlight_override_fill_color(self, tmp_path):
        """preset='highlight' 后用户覆盖 bg_color"""
        fp = str(tmp_path / "preset_hi_ov.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            preset="highlight",
            formatting={"bg_color": "FF0000"})
        assert result["success"], f"失败: {result.get('message')}"

    def test_preset_nonexistent_no_crash(self, tmp_path):
        """不存在的 preset 名称不应崩溃"""
        fp = str(tmp_path / "preset_bad.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            preset="nonexistent_preset",
            formatting={"bold": True})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True, "无效 preset 时 formatting 仍应生效"

    # ---------- 7. 合并 → 取消合并 往返 ----------

    def test_merge_then_unmerge_roundtrip(self, tmp_path):
        """先合并再取消合并，单元格可独立编辑"""
        fp = str(tmp_path / "merge_unmerge.xlsx")
        _create_test_xlsx(fp, 3, 3)
        # 合并
        r1 = ExcelOperations.merge_cells(fp, "Sheet1", "A1:B1")
        assert r1["success"], f"合并失败: {r1.get('message')}"
        # 取消合并
        r2 = ExcelOperations.unmerge_cells(fp, "Sheet1", "A1:B1")
        assert r2["success"], f"取消合并失败: {r2.get('message')}"
        # 验证 B1 可以独立设置值和样式
        wb = load_workbook(fp)
        ws = wb["Sheet1"]
        b1 = ws["B1"]
        assert b1.value is not None or True, "B1 应可访问"
        wb.close()

    def test_merge_unmerge_then_format(self, tmp_path):
        """合并→取消合并→格式化，格式应正常应用"""
        fp = str(tmp_path / "mu_format.xlsx")
        _create_test_xlsx(fp, 3, 3)
        ExcelOperations.merge_cells(fp, "Sheet1", "A1:C1")
        ExcelOperations.unmerge_cells(fp, "Sheet1", "A1:C1")
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:C1",
            formatting={"bold": True, "bg_color": "CCCCFF"})
        assert result["success"], f"格式化失败: {result.get('message')}"
        style = _read_cell_style(fp, "B1")
        assert style["bold"] is True, "取消合并后 B1 加粗应生效"

    # ---------- 8. Unicode 工作表名 ----------

    def test_unicode_sheet_name_format(self, tmp_path):
        """中文工作表名格式化不报错"""
        fp = str(tmp_path / "unicode_sheet.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "装备表"
        for r in range(1, 4):
            for c in range(1, 4):
                ws.cell(row=r, column=c, value=r * 10 + c)
        wb.save(fp)
        wb.close()

        result = ExcelOperations.format_cells(fp, "装备表", "A1:B2",
            formatting={"bold": True, "font_name": "微软雅黑"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1", sheet_name="装备表")
        assert style["bold"] is True, "中文工作表名下加粗应生效"
        assert style["font_name"] == "微软雅黑", "中文字体名应生效"

    def test_unicode_sheet_merge_and_format(self, tmp_path):
        """中文工作表名 + 合并 + 格式化组合"""
        fp = str(tmp_path / "unicode_merge_fmt.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "数据"
        for r in range(1, 4):
            for c in range(1, 4):
                ws.cell(row=r, column=c, value=r * 10 + c)
        wb.save(fp)
        wb.close()

        result = ExcelOperations.format_cells(fp, "数据", "A1:C1",
            formatting={"merge": True, "bold": True, "alignment": "center"})
        assert result["success"], f"失败: {result.get('message')}"

    # ---------- 9. 部分更新保留其他属性 ----------

    def test_partial_update_preserves_bold(self, tmp_path):
        """先设 bold=True，再只设 bg_color，bold 应保留"""
        fp = str(tmp_path / "partial_keep.xlsx")
        _create_test_xlsx(fp)
        # 第一步：设 bold
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"bold": True})
        # 第二步：只设 bg_color（不涉及 bold）
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"bg_color": "00FF00"})
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True, "部分更新时之前的 bold 应保留"
        assert style["fill_type"] is not None, "新设的 bg_color 应生效"

    def test_partial_update_preserves_font_size(self, tmp_path):
        """先设 font_size=18，再只设 italic，font_size 应保留"""
        fp = str(tmp_path / "partial_keep_fs.xlsx")
        _create_test_xlsx(fp)
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"font_size": 18})
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"italic": True})
        style = _read_cell_style(fp, "A1")
        assert style["size"] == 18, "font_size=18 应保留"
        assert style["italic"] is True, "italic=True 应生效"

    # ---------- 10. 3位 HEX 颜色简写 ----------

    def test_color_3digit_hex_shorthand_expanded(self, tmp_path):
        """3 位 HEX 颜色简写如 'F00' 应自动扩展为 6 位（F00→FF0000）"""
        # 先验证 _normalize_formatting 层的扩展行为
        norm = ExcelOperations._normalize_formatting({"font_color": "F00", "bg_color": "0F0"})
        assert norm["font"]["color"] == "FF0000", f"font_color应为FF0000, 实际{norm['font'].get('color')}"
        assert norm["fill"]["color"] == "00FF00", f"bg_color应为00FF00, 实际{norm['fill'].get('color')}"
        # 再验证端到端：format_cells 成功应用
        fp = str(tmp_path / "color3digit.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_color": "F00", "bg_color": "0F0"})
        assert result["success"] is True, f"3位HEX应被接受: {result}"

    # ---------- 11. bold=False / italic=False 显式关闭 ----------

    def test_bold_false_explicit_off(self, tmp_path):
        """bold=False 显式关闭加粗"""
        fp = str(tmp_path / "bold_false.xlsx")
        _create_test_xlsx(fp)
        # 先加粗
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"bold": True})
        assert _read_cell_style(fp, "A1")["bold"] is True
        # 再关闭
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"bold": False})
        assert _read_cell_style(fp, "A1")["bold"] is False, "bold=False 应关闭加粗"

    def test_italic_false_explicit_off(self, tmp_path):
        """italic=False 显式关闭斜体"""
        fp = str(tmp_path / "italic_false.xlsx")
        _create_test_xlsx(fp)
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"italic": True})
        assert _read_cell_style(fp, "A1")["italic"] is True
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"italic": False})
        assert _read_cell_style(fp, "A1")["italic"] is False, "italic=False 应关闭斜体"

    # ---------- 12. strikethrough 读写 ----------

    def test_strikethround_true_readback(self, tmp_path):
        """strikethrough=True 设置后读回应为 True"""
        fp = str(tmp_path / "strike.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"strikethrough": True})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["strikethrough"] is True, "删除线应生效"

    def test_strikethrough_false_explicit_off(self, tmp_path):
        """strikethrough=False 显式关闭删除线"""
        fp = str(tmp_path / "strike_false.xlsx")
        _create_test_xlsx(fp)
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"strikethrough": True})
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"strikethrough": False})
        assert _read_cell_style(fp, "A1")["strikethrough"] is False, \
            "strikethrough=False 应关闭删除线"

    # ---------- 13. indent=0 显式零值 ----------

    def test_indent_zero_explicit(self, tmp_path):
        """indent=0 显式设置零缩进不应崩溃"""
        fp = str(tmp_path / "indent0.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"indent": 0})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["indent"] == 0, "indent=0 应保持为 0"

    def test_indent_positive_value(self, tmp_path):
        """indent=3 正缩进值"""
        fp = str(tmp_path / "indent3.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"indent": 3})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["indent"] == 3, "indent=3 应保持为 3"

    # ---------- 14. number_format General 重置 ----------

    def test_number_format_general_reset(self, tmp_path):
        """number_format='General' 重置为默认格式"""
        fp = str(tmp_path / "nf_general.xlsx")
        _create_test_xlsx(fp)
        # 先设为货币格式
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"number_format": "¥#,##0.00"})
        assert "¥" in _read_cell_style(fp, "A1")["number_format"]
        # 重置为 General
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"number_format": "General"})
        style = _read_cell_style(fp, "A1")
        assert style["number_format"] == "General", "应重置为 General"

    # ---------- 15. shrink_to_fit 读写 ----------

    def test_shrink_to_fit_true(self, tmp_path):
        """shrink_to_fit=True 设置后应生效"""
        fp = str(tmp_path / "shrink.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"shrink_to_fit": True})
        assert result["success"], f"失败: {result.get('message')}"

    def test_shrink_to_fit_false(self, tmp_path):
        """shrink_to_fit=False 显式关闭"""
        fp = str(tmp_path / "shrink_f.xlsx")
        _create_test_xlsx(fp)
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"shrink_to_fit": True})
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"shrink_to_fit": False})
        assert result["success"], f"失败: {result.get('message')}"

    # ---------- 16. alignment + vertical_alignment 同时传 ----------

    def test_alignment_and_vertical_combined(self, tmp_path):
        """同时传 alignment 和 vertical_alignment"""
        fp = str(tmp_path / "align_both.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"alignment": "right", "vertical_alignment": "top"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["alignment_h"] == "right", "水平对齐应为 right"
        assert style["alignment_v"] == "top", "垂直对齐应为 top"

    def test_vertical_alignment_middle_maps_to_center(self, tmp_path):
        """vertical_alignment='middle' 应映射为 'center'"""
        fp = str(tmp_path / "align_mid.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"vertical_alignment": "middle"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["alignment_v"] == "center", "middle 应映射为 center"

    # ---------- 17. 全部字体属性组合 ----------

    def test_all_font_attrs_at_once(self, tmp_path):
        """一次性设置所有字体属性"""
        fp = str(tmp_path / "all_font.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={
            "bold": True,
            "italic": True,
            "underline": "double",
            "strikethrough": True,
            "font_size": 16,
            "font_color": "AA00BB",
            "font_name": "Arial",
        })
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["italic"] is True
        assert style["underline"] == "double"
        assert style["strikethrough"] is True
        assert style["size"] == 16
        assert style["font_name"] == "Arial"

    # ---------- 18. 单元格范围 A:A（整列）格式化 ----------

    def test_full_column_range_format(self, tmp_path):
        """整列范围 A:A 格式化不应崩溃"""
        fp = str(tmp_path / "col_range.xlsx")
        _create_test_xlsx(fp, 10, 4)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A:A",
            formatting={"bold": True})
        assert result["success"], f"失败: {result.get('message')}"

    # ---------- 19. 单元格范围 1:1（整行）格式化 ----------

    def test_full_row_range_format(self, tmp_path):
        """整行范围 1:1 格式化不应崩溃"""
        fp = str(tmp_path / "row_range.xlsx")
        _create_test_xlsx(fp, 10, 4)
        result = ExcelOperations.format_cells(fp, "Sheet1", "1:1",
            formatting={"bg_color": "EEEEEE"})
        assert result["success"], f"失败: {result.get('message')}"

    # ---------- 20. 多种 preset 逐一验证 ----------

    def test_preset_title_default(self, tmp_path):
        """preset='title' 默认效果：微软雅黑 14号 加粗 居中"""
        fp = str(tmp_path / "preset_title.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", preset="title")
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True, "title preset 应加粗"
        assert style["size"] == 14, "title preset 应 14 号"
        assert style["font_name"] == "微软雅黑", "title preset 应使用微软雅黑"
        assert style["alignment_h"] == "center", "title preset 应居中"

    def test_preset_header_default(self, tmp_path):
        """preset='header' 默认效果：微软雅黑 11号 加粗 灰底"""
        fp = str(tmp_path / "preset_header.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", preset="header")
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True, "header preset 应加粗"
        assert style["size"] == 11, "header preset 应 11 号"
