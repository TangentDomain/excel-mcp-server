"""excel_describe_table 测试套件"""

import sys
from pathlib import Path

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root / "src"))

import pytest
from openpyxl import Workbook

from excel_mcp_server_fastmcp.server import excel_describe_table


def _make_simple_wb() -> Workbook:
    """普通单行表头：5 行数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"
    ws.append(["ID", "Name", "Price", "Type"])
    ws.append([1, "Sword", 100, "Weapon"])
    ws.append([2, "Shield", 200, "Armor"])
    ws.append([3, "Potion", 50, "Consumable"])
    ws.append([4, "Ring", 500, "Accessory"])
    ws.append([5, "Staff", 300, "Weapon"])
    return wb


def _make_dual_header_wb() -> Workbook:
    """双行表头（游戏配置常见格式）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(["技能ID", "技能名称", "伤害值", "冷却时间"])
    ws.append(["skill_id", "skill_name", "damage", "cooldown"])
    ws.append([101, "火球术", 80, 5])
    ws.append([102, "冰箭", 60, 3])
    ws.append([103, "治疗术", 0, 8])
    return wb


def _make_empty_wb() -> Workbook:
    """空表：只有表头"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"
    ws.append(["ID", "Name"])
    return wb


def _make_mixed_types_wb() -> Workbook:
    """混合类型表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed"
    ws.append(["ID", "Price", "Name", "Active", "Rate"])
    ws.append([1, 99.5, "A", True, 0.1])
    ws.append([2, None, "B", False, 0.5])
    ws.append([3, 150, "C", True, None])
    return wb


@pytest.fixture
def simple_file(tmp_path) -> str:
    path = tmp_path / "simple.xlsx"
    _make_simple_wb().save(path)
    return str(path)


@pytest.fixture
def dual_file(tmp_path) -> str:
    path = tmp_path / "dual.xlsx"
    _make_dual_header_wb().save(path)
    return str(path)


@pytest.fixture
def empty_file(tmp_path) -> str:
    path = tmp_path / "empty.xlsx"
    _make_empty_wb().save(path)
    return str(path)


@pytest.fixture
def mixed_file(tmp_path) -> str:
    path = tmp_path / "mixed.xlsx"
    _make_mixed_types_wb().save(path)
    return str(path)


class TestDescribeTable:
    """excel_describe_table 基础功能测试"""

    def test_describe_simple_table(self, simple_file):
        result = excel_describe_table(simple_file)
        assert result["success"] is True
        data = result["data"]
        assert data["sheet_name"] == "Items"
        assert data["row_count"] == 5
        assert data["column_count"] == 4
        col_names = [c["name"] for c in data["columns"]]
        assert col_names == ["ID", "Name", "Price", "Type"]

    def test_describe_dual_header(self, dual_file):
        result = excel_describe_table(dual_file)
        assert result["success"] is True
        data = result["data"]
        assert data["header_type"] == "dual"
        assert data["sheet_name"] == "Skills"
        assert data["row_count"] == 3
        # 双行表头应返回英文名（第2行）
        col_names = [c["name"] for c in data["columns"]]
        assert col_names == ["skill_id", "skill_name", "damage", "cooldown"]

    def test_describe_empty_table(self, empty_file):
        result = excel_describe_table(empty_file)
        assert result["success"] is True
        data = result["data"]
        assert data["row_count"] == 0

    def test_describe_column_types(self, mixed_file):
        result = excel_describe_table(mixed_file)
        assert result["success"] is True
        columns = {c["name"]: c for c in result["data"]["columns"]}
        # Price 列有 float 值
        assert columns["Price"]["type"] in ("float", "number", "mixed")

    def test_describe_nonexistent_sheet(self, simple_file):
        result = excel_describe_table(simple_file, sheet_name="Nope")
        assert result["success"] is False
        assert "error_code" in result.get("meta", {})

    def test_describe_invalid_file(self):
        result = excel_describe_table("/nonexistent/path.xlsx")
        assert result["success"] is False

    def test_describe_structure_consistent(self, simple_file):
        """验证返回值结构包含所有必要字段"""
        result = excel_describe_table(simple_file)
        assert result["success"] is True
        data = result["data"]
        assert "sheet_name" in data
        assert "header_type" in data
        assert "row_count" in data
        assert "column_count" in data
        assert "columns" in data
        assert isinstance(data["columns"], list)
        for col in data["columns"]:
            assert "name" in col
            assert "type" in col

    def test_describe_sample_values(self, simple_file):
        result = excel_describe_table(simple_file)
        assert result["success"] is True
        columns = {(c["name"], c["type"]): c for c in result["data"]["columns"]}
        # Name 列应有 sample_values
        for (name, _), col in columns.items():
            if name == "Name":
                assert "sample_values" in col
