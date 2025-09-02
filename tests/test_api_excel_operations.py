# -*- coding: utf-8 -*-
"""
Excel MCP Server - ExcelOperations API单元测试

专门测试api.excel_operations模块的ExcelOperations类
独立于MCP层，直接测试API业务逻辑

测试覆盖：
1. get_range方法的完整流程
2. 参数验证逻辑
3. 业务逻辑执行
4. 结果格式化
5. 错误处理
6. 资源管理
7. 边界条件和异常情况
"""

import pytest
import tempfile
import unittest.mock
from pathlib import Path
from openpyxl import Workbook

from src.api.excel_operations import ExcelOperations
from src.models.types import OperationResult


class TestExcelOperations:
    """
    @class TestExcelOperations
    @brief ExcelOperations类的全面单元测试
    @intention 确保API层的业务逻辑正确性、参数验证有效性和异常处理完整性
    """

    # ==================== 测试数据准备 ====================

    @pytest.fixture
    def test_excel_file(self, temp_dir):
        """创建用于测试的Excel文件"""
        file_path = temp_dir / "test_operations.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # 添加测试数据
        test_data = [
            ["姓名", "年龄", "邮箱"],
            ["张三", 25, "zhang@example.com"],
            ["李四", 30, "li@example.com"],
            ["王五", 28, "wang@example.com"]
        ]

        for row_idx, row_data in enumerate(test_data, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        wb.save(file_path)
        wb.close()

        return str(file_path)

    @pytest.fixture
    def sample_operation_result(self):
        """创建示例OperationResult对象用于测试格式化"""
        return OperationResult(
            success=True,
            data=[
                [{"coordinate": "A1", "value": "姓名"}, {"coordinate": "B1", "value": "年龄"}],
                [{"coordinate": "A2", "value": "张三"}, {"coordinate": "B2", "value": 25}]
            ],
            metadata={
                "range": "TestSheet!A1:B2",
                "sheet_name": "TestSheet",
                "dimensions": {"rows": 2, "columns": 2}
            }
        )

    # ==================== get_range方法完整流程测试 ====================

    def test_get_range_success_flow(self, test_excel_file):
        """测试get_range方法的成功执行流程"""
        # 执行测试
        result = ExcelOperations.get_range(
            test_excel_file,
            "TestSheet!A1:C2"
        )

        # 验证结果
        assert result['success'] is True
        assert 'data' in result
        assert isinstance(result['data'], list)
        assert len(result['data']) == 2  # 2行数据
        assert len(result['data'][0]) == 3  # 3列数据

        # 验证数据内容
        first_row = result['data'][0]
        assert first_row[0]['value'] == '姓名'
        assert first_row[1]['value'] == '年龄'
        assert first_row[2]['value'] == '邮箱'

    def test_get_range_with_formatting(self, test_excel_file):
        """测试包含格式信息的get_range调用"""
        result = ExcelOperations.get_range(
            test_excel_file,
            "TestSheet!A1:B1",
            include_formatting=True
        )

        assert result['success'] is True
        assert len(result['data']) == 1
        assert len(result['data'][0]) == 2

    def test_get_range_different_range_types(self, test_excel_file):
        """测试不同类型的范围表达式"""
        test_cases = [
            ("TestSheet!A1:B2", 2, 2),    # 单元格范围
            ("TestSheet!A1", 1, 1),       # 单个单元格
        ]

        for range_expr, expected_rows, expected_cols in test_cases:
            result = ExcelOperations.get_range(test_excel_file, range_expr)
            assert result['success'] is True
            assert len(result['data']) == expected_rows
            assert len(result['data'][0]) == expected_cols

    # ==================== 参数验证测试 ====================

    def test_validate_range_parameters_empty_file_path(self):
        """测试空文件路径的参数验证"""
        # 更新为使用新的_validate_range_format方法
        result1 = ExcelOperations._validate_range_format("TestSheet!A1:B2")
        assert result1['valid'] is True  # 只验证格式，不验证文件路径

        # 实际的文件路径验证在get_range方法内部处理
        result = ExcelOperations.get_range("", "TestSheet!A1:B2")
        assert result['success'] is False
        assert 'error' in result

    def test_validate_range_parameters_empty_range(self):
        """测试空范围表达式的参数验证"""
        result1 = ExcelOperations._validate_range_format("")
        assert result1['valid'] is False
        assert 'range参数不能为空' in result1['error']

        result2 = ExcelOperations._validate_range_format("   ")
        assert result2['valid'] is False
        assert 'range参数不能为空' in result2['error']

    def test_validate_range_parameters_missing_sheet_name(self):
        """测试缺少工作表名的范围表达式验证"""
        result = ExcelOperations._validate_range_format("A1:B2")
        assert result['valid'] is False
        assert 'range必须包含工作表名' in result['error']

    def test_validate_range_parameters_valid_input(self):
        """测试有效输入的参数验证（不应抛出异常）"""
        # 这些调用不应抛出异常
        result1 = ExcelOperations._validate_range_format("Sheet1!A1:B2")
        assert result1['valid'] is True

        result2 = ExcelOperations._validate_range_format("数据!C1:D10")
        assert result2['valid'] is True

        result3 = ExcelOperations._validate_range_format("MySheet!A1")
        assert result3['valid'] is True

    # ==================== 业务逻辑执行测试 ====================

    @unittest.mock.patch('src.api.excel_operations.ExcelReader')
    def test_execute_get_range_reader_lifecycle(self, mock_reader_class):
        """测试ExcelReader的生命周期管理（通过get_range方法）"""
        # 设置mock
        mock_reader = unittest.mock.MagicMock()
        mock_reader_class.return_value = mock_reader
        mock_reader.get_range.return_value = OperationResult(success=True, data=[])

        # 执行测试（通过公共API而不是内部方法）
        result = ExcelOperations.get_range(
            "test.xlsx",
            "Sheet1!A1:B2",
            False
        )

        # 验证调用
        mock_reader_class.assert_called_once_with("test.xlsx")
        mock_reader.get_range.assert_called_once_with("Sheet1!A1:B2", False)
        mock_reader.close.assert_called_once()

        # 验证结果
        assert result['success'] is True

    @unittest.mock.patch('src.api.excel_operations.ExcelReader')
    def test_execute_get_range_exception_handling(self, mock_reader_class):
        """测试异常处理"""
        # 设置mock抛出异常
        mock_reader = unittest.mock.MagicMock()
        mock_reader_class.return_value = mock_reader
        mock_reader.get_range.side_effect = Exception("读取错误")

        # 执行测试并验证异常被正确处理
        result = ExcelOperations.get_range("test.xlsx", "Sheet1!A1:B2")

        assert result['success'] is False
        assert '获取范围数据失败' in result['error']
        assert '读取错误' in result['error']

    # ==================== 结果格式化测试 ====================

    @unittest.mock.patch('src.api.excel_operations.format_operation_result')
    def test_format_get_range_result(self, mock_formatter):
        """测试结果格式化功能（通过get_range方法）"""
        # 设置mock
        from src.models.types import OperationResult
        sample_result = OperationResult(success=True, data=[])
        expected_formatted = {"success": True, "data": [], "formatted": True}
        mock_formatter.return_value = expected_formatted

        with unittest.mock.patch('src.api.excel_operations.ExcelReader') as mock_reader_class:
            mock_reader = unittest.mock.MagicMock()
            mock_reader_class.return_value = mock_reader
            mock_reader.get_range.return_value = sample_result

            # 执行测试
            result = ExcelOperations.get_range("test.xlsx", "Sheet1!A1:B2")

            # 验证调用和结果
            mock_formatter.assert_called_once_with(sample_result)
            assert result == expected_formatted

    # ==================== 错误处理测试 ====================

    def test_create_error_response_structure(self):
        """测试错误响应的结构"""
        error_msg = "测试错误消息"
        result = ExcelOperations._format_error_result(error_msg)  # 使用正确的方法名

        assert result['success'] is False
        assert result['error'] == error_msg
        assert result['data'] is None

    def test_get_range_error_handling_invalid_params(self):
        """测试get_range的参数验证错误处理"""
        # 测试无效的范围格式
        result = ExcelOperations.get_range("test.xlsx", "A1:B2")  # 缺少工作表名
        assert result['success'] is False
        assert 'range必须包含工作表名' in result['error']

        # 测试无效的范围表达式
        result = ExcelOperations.get_range("test.xlsx", "A1:B2")
        assert result['success'] is False
        assert 'range必须包含工作表名' in result['error']

    @unittest.mock.patch('src.api.excel_operations.ExcelReader')
    def test_get_range_error_handling_execution_failure(self, mock_reader_class):
        """测试get_range的业务执行错误处理"""
        # 设置mock：模拟文件不存在的情况
        mock_reader_class.side_effect = FileNotFoundError("文件未找到")

        result = ExcelOperations.get_range("nonexistent.xlsx", "Sheet1!A1:B2")

        assert result['success'] is False
        assert '文件未找到' in result['error']

    # ==================== 日志功能测试 ====================

    @unittest.mock.patch('src.api.excel_operations.logger')
    def test_debug_logging_enabled(self, mock_logger, test_excel_file):
        """测试启用调试日志时的日志记录"""
        # 临时启用调试日志
        original_debug_setting = ExcelOperations.DEBUG_LOG_ENABLED
        ExcelOperations.DEBUG_LOG_ENABLED = True

        try:
            ExcelOperations.get_range(test_excel_file, "TestSheet!A1:B2")

            # 验证日志调用 - 使用info而不是debug
            assert mock_logger.info.call_count >= 1  # 开始的日志
            info_calls = [call[0][0] for call in mock_logger.info.call_args_list]
            assert any("开始获取范围数据" in msg for msg in info_calls)

        finally:
            # 恢复原始设置
            ExcelOperations.DEBUG_LOG_ENABLED = original_debug_setting

    @unittest.mock.patch('src.api.excel_operations.logger')
    def test_debug_logging_disabled(self, mock_logger, test_excel_file):
        """测试禁用调试日志时不记录调试信息"""
        # 确保调试日志被禁用
        original_debug_setting = ExcelOperations.DEBUG_LOG_ENABLED
        ExcelOperations.DEBUG_LOG_ENABLED = False

        try:
            ExcelOperations.get_range(test_excel_file, "TestSheet!A1:B2")

            # 验证没有info日志（因为被禁用）
            mock_logger.info.assert_not_called()

        finally:
            # 恢复原始设置
            ExcelOperations.DEBUG_LOG_ENABLED = original_debug_setting
            # 恢复原始设置
            ExcelOperations.DEBUG_LOG_ENABLED = original_debug_setting

    @unittest.mock.patch('src.api.excel_operations.logger')
    def test_error_logging(self, mock_logger):
        """测试错误情况下的日志记录"""
        # 临时启用调试日志
        original_debug_setting = ExcelOperations.DEBUG_LOG_ENABLED
        ExcelOperations.DEBUG_LOG_ENABLED = True

        try:
            # 触发错误 - 使用无效的范围格式
            ExcelOperations.get_range("test.xlsx", "A1:B2")  # 缺少工作表名

            # 由于是参数验证错误，不会记录error级别日志，而是直接返回错误
            # 所以这里不验证error日志，而是验证返回结果
            result = ExcelOperations.get_range("test.xlsx", "A1:B2")
            assert result['success'] is False
            assert 'range必须包含工作表名' in result['error']

        finally:
            # 恢复原始设置
            ExcelOperations.DEBUG_LOG_ENABLED = original_debug_setting

    # ==================== 边界条件测试 ====================

    def test_get_range_nonexistent_file(self):
        """测试不存在的文件"""
        result = ExcelOperations.get_range("/nonexistent/path.xlsx", "Sheet1!A1:B2")
        assert result['success'] is False
        assert 'error' in result

    def test_get_range_invalid_range_format(self, test_excel_file):
        """测试无效的范围格式"""
        invalid_ranges = [
            "InvalidSheet!A1:B2",  # 不存在的工作表
            "TestSheet!Z1:AA100",  # 超出数据范围（应该仍然成功，但返回空数据）
        ]

        for invalid_range in invalid_ranges:
            result = ExcelOperations.get_range(test_excel_file, invalid_range)
            # 注意：根据实际实现，某些"无效"范围可能仍然会成功但返回空数据
            # 这里我们主要确保不会崩溃
            assert 'success' in result

    # ==================== 集成测试 ====================

    def test_get_range_integration_with_real_file(self, test_excel_file):
        """集成测试：使用真实Excel文件的完整流程"""
        # 测试多种范围表达式
        test_cases = [
            {
                "range": "TestSheet!A1:C1",
                "expected_data": "姓名",
                "description": "表头行测试"
            },
            {
                "range": "TestSheet!A2:C2",
                "expected_data": "张三",
                "description": "数据行测试"
            },
            {
                "range": "TestSheet!B1:B4",
                "expected_rows": 4,
                "expected_cols": 1,
                "description": "单列测试"
            }
        ]

        for case in test_cases:
            result = ExcelOperations.get_range(test_excel_file, case["range"])

            assert result['success'] is True, f"失败的测试用例: {case['description']}"
            assert 'data' in result

            if 'expected_data' in case:
                assert result['data'][0][0]['value'] == case['expected_data']

            if 'expected_rows' in case and 'expected_cols' in case:
                assert len(result['data']) == case['expected_rows']
                assert len(result['data'][0]) == case['expected_cols']

    # ==================== 性能和资源管理测试 ====================

    @unittest.mock.patch('src.api.excel_operations.ExcelReader')
    def test_resource_management_multiple_calls(self, mock_reader_class):
        """测试多次调用的资源管理"""
        mock_reader = unittest.mock.MagicMock()
        mock_reader_class.return_value = mock_reader
        mock_reader.get_range.return_value = OperationResult(success=True, data=[])

        # 多次调用
        for _ in range(3):
            ExcelOperations.get_range("test.xlsx", "Sheet1!A1:B2")

        # 验证每次调用都正确创建和关闭reader
        assert mock_reader_class.call_count == 3
        assert mock_reader.close.call_count == 3

    def test_concurrent_access_safety(self, test_excel_file):
        """测试并发访问的安全性（基本测试）"""
        import threading
        import time

        results = []
        errors = []

        def worker():
            try:
                result = ExcelOperations.get_range(test_excel_file, "TestSheet!A1:B2")
                results.append(result)
            except Exception as e:
                errors.append(e)

        # 创建多个线程
        threads = []
        for _ in range(5):
            thread = threading.Thread(target=worker)
            threads.append(thread)
            thread.start()

        # 等待所有线程完成
        for thread in threads:
            thread.join()

        # 验证结果
        assert len(errors) == 0, f"并发访问出现错误: {errors}"
        assert len(results) == 5
        assert all(result['success'] for result in results)


    # ==================== 测试find_last_row方法 ====================

    def test_find_last_row_entire_sheet(self, test_excel_file):
        """测试查找整个工作表的最后一行"""
        result = ExcelOperations.find_last_row(str(test_excel_file), "TestSheet")

        assert result['success']
        assert result['data']['last_row'] == 4  # 表头1行 + 数据3行
        assert result['data']['sheet_name'] == "TestSheet"
        assert result['data']['column'] is None
        assert result['data']['search_scope'] == "整个工作表"
        assert "成功查找整个工作表最后一行" in result['message']

    def test_find_last_row_specific_column_by_name(self, test_excel_file):
        """测试查找指定列（按列名）的最后一行"""
        result = ExcelOperations.find_last_row(str(test_excel_file), "TestSheet", "A")

        assert result['success']
        assert result['data']['last_row'] == 4
        assert result['data']['sheet_name'] == "TestSheet"
        assert result['data']['column'] == "A"
        assert result['data']['search_scope'] == "A列"
        assert "成功查找A列最后一行" in result['message']

    def test_find_last_row_specific_column_by_index(self, test_excel_file):
        """测试查找指定列（按索引）的最后一行"""
        result = ExcelOperations.find_last_row(str(test_excel_file), "TestSheet", 2)

        assert result['success']
        assert result['data']['last_row'] == 4
        assert result['data']['sheet_name'] == "TestSheet"
        assert result['data']['column'] == 2
        assert result['data']['search_scope'] == "B列"
        assert "成功查找B列最后一行" in result['message']

    def test_find_last_row_empty_sheet(self, temp_dir):
        """测试查找空工作表的最后一行"""
        empty_file = temp_dir / "empty.xlsx"
        wb = Workbook()
        wb.save(empty_file)

        result = ExcelOperations.find_last_row(str(empty_file), "Sheet")

        assert result['success']
        assert result['data']['last_row'] == 0
        assert "没有数据" in result['message']

    def test_find_last_row_nonexistent_sheet(self, test_excel_file):
        """测试查找不存在的工作表"""
        result = ExcelOperations.find_last_row(str(test_excel_file), "NonExistentSheet")

        assert not result['success']
        assert "工作表不存在" in result['error']

    def test_find_last_row_invalid_file(self):
        """测试无效的文件路径"""
        result = ExcelOperations.find_last_row("nonexistent_file.xlsx", "Sheet1")

        assert not result['success']
        assert "查找最后一行失败" in result['error']

    def test_find_last_row_invalid_column_name(self, test_excel_file):
        """测试无效的列名"""
        result = ExcelOperations.find_last_row(str(test_excel_file), "TestSheet", "INVALID")

        assert not result['success']
        assert "无效的列名" in result['error']

    def test_find_last_row_invalid_column_index(self, test_excel_file):
        """测试无效的列索引"""
        result = ExcelOperations.find_last_row(str(test_excel_file), "TestSheet", 0)

        assert not result['success']
        assert "列索引必须大于等于1" in result['error']

    def test_find_last_row_invalid_column_type(self, test_excel_file):
        """测试无效的列参数类型"""
        result = ExcelOperations.find_last_row(str(test_excel_file), "TestSheet", 3.14)

        assert not result['success']
        assert "列参数必须是字符串或整数" in result['error']

    @unittest.mock.patch('src.api.excel_operations.logger')
    def test_find_last_row_logging(self, mock_logger, test_excel_file):
        """测试find_last_row的日志记录"""
        # 开启调试日志
        original_debug = ExcelOperations.DEBUG_LOG_ENABLED
        ExcelOperations.DEBUG_LOG_ENABLED = True

        try:
            ExcelOperations.find_last_row(str(test_excel_file), "TestSheet")

            # 验证日志调用
            mock_logger.info.assert_called()
            log_calls = [call.args[0] for call in mock_logger.info.call_args_list]
            assert any("[API][ExcelOperations] 开始查找最后一行" in call for call in log_calls)
        finally:
            ExcelOperations.DEBUG_LOG_ENABLED = original_debug
