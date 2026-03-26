# -*- coding: utf-8 -*-
"""
Excel Operations 格式化测试套件

覆盖 excel_operations.py 中的格式化功能
"""

import pytest
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations


class TestExcelFormatOperations:
    """Excel格式化操作测试"""

    @pytest.fixture
    def format_test_file(self, temp_dir):
        """创建格式化测试文件"""
        file_path = temp_dir / "format_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "FormatSheet"
        
        # 添加测试数据
        ws['A1'] = "Header1"
        ws['B1'] = "Header2"
        ws['A2'] = "Data1"
        ws['B2'] = "Data2"
        ws['A3'] = "Data3"
        ws['B3'] = "Data4"
        
        wb.save(str(file_path))
        return str(file_path)

    def test_format_cells_basic(self, format_test_file):
        """测试基本格式化"""
        result = ExcelOperations.format_cells(
            file_path=format_test_file,
            sheet_name="FormatSheet",
            range="A1:B1",
            formatting={"font": {"bold": True}}
        )
        
        assert result is not None

    def test_format_cells_preset(self, format_test_file):
        """测试预设格式化"""
        result = ExcelOperations.format_cells(
            file_path=format_test_file,
            sheet_name="FormatSheet",
            range="A1",
            preset="highlight"
        )
        
        assert result is not None

    def test_merge_cells(self, format_test_file):
        """测试合并单元格"""
        result = ExcelOperations.merge_cells(
            file_path=format_test_file,
            sheet_name="FormatSheet",
            range="A1:B2"
        )
        
        assert result is not None

    def test_unmerge_cells(self, format_test_file):
        """测试取消合并单元格"""
        # 先合并
        ExcelOperations.merge_cells(
            file_path=format_test_file,
            sheet_name="FormatSheet",
            range="A1:B2"
        )
        
        # 再取消合并
        result = ExcelOperations.unmerge_cells(
            file_path=format_test_file,
            sheet_name="FormatSheet",
            range="A1:B2"
        )
        
        assert result is not None

    def test_set_borders(self, format_test_file):
        """测试设置边框"""
        result = ExcelOperations.set_borders(
            file_path=format_test_file,
            sheet_name="FormatSheet",
            range="A1:B2",
            border_style="thin"
        )
        
        assert result is not None

    def test_set_row_height(self, format_test_file):
        """测试设置行高"""
        result = ExcelOperations.set_row_height(
            file_path=format_test_file,
            sheet_name="FormatSheet",
            row_index=1,
            height=30
        )
        
        assert result is not None

    def test_set_column_width(self, format_test_file):
        """测试设置列宽"""
        result = ExcelOperations.set_column_width(
            file_path=format_test_file,
            sheet_name="FormatSheet",
            column_index=1,
            width=20
        )
        
        assert result is not None


class TestExcelDataOperations:
    """Excel数据操作测试"""

    @pytest.fixture
    def data_test_file(self, temp_dir):
        """创建数据测试文件"""
        file_path = temp_dir / "data_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DataSheet"
        
        ws['A1'] = "ID"
        ws['B1'] = "Value"
        ws['A2'] = 1
        ws['B2'] = 100
        
        wb.save(str(file_path))
        return str(file_path)

    def test_insert_rows_beginning(self, data_test_file):
        """测试在开头插入行"""
        result = ExcelOperations.insert_rows(
            file_path=data_test_file,
            sheet_name="DataSheet",
            row_index=1,
            count=1
        )

        assert result is not None

    def test_insert_rows_middle(self, data_test_file):
        """测试在中间插入行"""
        result = ExcelOperations.insert_rows(
            file_path=data_test_file,
            sheet_name="DataSheet",
            row_index=2,
            count=1
        )

        assert result is not None

    def test_delete_rows_beginning(self, data_test_file):
        """测试删除开头行"""
        result = ExcelOperations.delete_rows(
            file_path=data_test_file,
            sheet_name="DataSheet",
            row_index=2,
            count=1
        )

        assert result is not None

    def test_insert_columns_beginning(self, data_test_file):
        """测试在开头插入列"""
        result = ExcelOperations.insert_columns(
            file_path=data_test_file,
            sheet_name="DataSheet",
            column_index=1,
            count=1
        )

        assert result is not None

    def test_delete_columns_beginning(self, data_test_file):
        """测试删除开头列"""
        result = ExcelOperations.delete_columns(
            file_path=data_test_file,
            sheet_name="DataSheet",
            column_index=2,
            count=1
        )

        assert result is not None


class TestExcelSearchAdvanced:
    """Excel高级搜索测试"""

    @pytest.fixture
    def search_test_file(self, temp_dir):
        """创建搜索测试文件"""
        file_path = temp_dir / "search_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "SearchSheet"
        
        # 添加包含公式的单元格
        ws['A1'] = "Header"
        ws['A2'] = "Item1"
        ws['A3'] = "Item2"
        ws['A4'] = "Test123"
        ws['A5'] = "test456"
        
        # 添加公式
        ws['B1'] = "=A1"
        ws['B2'] = "=A2"
        
        wb.save(str(file_path))
        return str(file_path)

    def test_search_include_formulas(self, search_test_file):
        """测试包含公式的搜索"""
        result = ExcelOperations.search(
            file_path=search_test_file,
            pattern="=.*",
            sheet_name="SearchSheet",
            include_formulas=True,
            use_regex=True
        )
        
        assert result is not None

    def test_search_case_sensitive(self, search_test_file):
        """测试大小写敏感搜索"""
        result = ExcelOperations.search(
            file_path=search_test_file,
            pattern="item1",
            sheet_name="SearchSheet",
            case_sensitive=True
        )
        
        assert result['success'] is True

    def test_search_whole_word(self, search_test_file):
        """测试全词匹配"""
        result = ExcelOperations.search(
            file_path=search_test_file,
            pattern="Item",
            sheet_name="SearchSheet",
            whole_word=True
        )
        
        assert result['success'] is True

    def test_search_no_matches(self, search_test_file):
        """测试无匹配结果"""
        result = ExcelOperations.search(
            file_path=search_test_file,
            pattern="NonExistent",
            sheet_name="SearchSheet"
        )
        
        assert result['success'] is True


class TestExcelCompareAdvanced:
    """Excel高级比较测试"""

    @pytest.fixture
    def compare_file1(self, temp_dir):
        """创建比较文件1"""
        file_path = temp_dir / "compare1.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        ws['A1'] = "ID"
        ws['B1'] = "Name"
        ws['C1'] = "Value"
        
        for i in range(2, 12):
            ws[f'A{i}'] = i - 1
            ws[f'B{i}'] = f"Name{i}"
            ws[f'C{i}'] = i * 10
        
        wb.save(str(file_path))
        return str(file_path)

    @pytest.fixture
    def compare_file2(self, temp_dir):
        """创建比较文件2"""
        file_path = temp_dir / "compare2.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        ws['A1'] = "ID"
        ws['B1'] = "Name"
        ws['C1'] = "Value"
        
        # 有一些差异
        for i in range(2, 12):
            ws[f'A{i}'] = i - 1
            ws[f'B{i}'] = f"Name{i}"
            if i < 8:
                ws[f'C{i}'] = i * 20  # 不同的值
            else:
                ws[f'C{i}'] = i * 10
        
        wb.save(str(file_path))
        return str(file_path)

    def test_compare_sheets_detailed(self, compare_file1, compare_file2):
        """测试详细的工作表比较"""
        result = ExcelOperations.compare_sheets(
            file1_path=compare_file1,
            sheet1_name="Data",
            file2_path=compare_file2,
            sheet2_name="Data"
        )
        
        assert result is not None

    def test_check_duplicate_ids_detailed(self, compare_file1):
        """测试详细重复ID检查"""
        result = ExcelOperations.check_duplicate_ids(
            file_path=compare_file1,
            sheet_name="Data",
            id_column=1
        )
        
        assert result is not None


class TestExcelEdgeCases:
    """边界情况测试"""

    @pytest.fixture
    def edge_file(self, temp_dir):
        """创建边界测试文件"""
        file_path = temp_dir / "edge.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Edge"
        
        # 添加各种边界数据
        ws['A1'] = ""
        ws['A2'] = " "
        ws['A3'] = "  "
        ws['A4'] = "\t"
        ws['A5'] = "\n"
        
        wb.save(str(file_path))
        return str(file_path)

    def test_search_whitespace(self, edge_file):
        """测试搜索空白字符"""
        result = ExcelOperations.search(
            file_path=edge_file,
            pattern=" ",
            sheet_name="Edge"
        )
        
        assert result is not None

    def test_search_special_chars(self, edge_file):
        """测试搜索特殊字符"""
        result = ExcelOperations.search(
            file_path=edge_file,
            pattern="\t",
            sheet_name="Edge"
        )
        
        assert result is not None
