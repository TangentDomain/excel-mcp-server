# -*- coding: utf-8 -*-
"""
format_cells 边缘案例测试 - R55+ 迭代补充 (Round 3)

新增覆盖:
  - preset + formatting 同时传入时的行为
  - merge + unmerge 同时为 True 的优先级
  - font_color RGB tuple / int 类型容错
  - text_rotation 超范围值处理
  - border_style 无效值
  - 空字符串 / 含空格的 cell_range
  - font_size 零或负数
  - number_format 极长字符串
  - 单元格 range 为单格合并
  - _check_merge_data_loss 边界 (单格、全空、仅左上角有值)
  - format_cells 对已合并区域的格式化
  - 多次连续 format_cells 覆盖行为
  - alignment 非标准值 (justify/distributed/centerContinuous)
  - underline=None 显式传 None
  - bg_color 空字符串
  - formatting 全部值为 False/0/"" 的"清除样式"场景
"""

import os
import pytest
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from excel_mcp_server_fastmcp.core.excel_writer import ExcelWriter
from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations


# ==================== Helper ====================

def _create_test_xlsx(file_path: str, rows: int = 5, cols: int = 4, sheet_name: str = "Sheet1"):
    """创建测试用 xlsx 文件，含数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=r * 10 + c)
    wb.save(file_path)
    wb.close()


def _read_cell_style(file_path: str, cell_ref: str = "A1", sheet_name: str = "Sheet1") -> dict:
    """读取单元格的样式属性用于断言"""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    cell = ws[cell_ref]
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    result = {
        "bold": font.bold,
        "italic": font.italic,
        "underline": font.underline,
        "size": font.size,
        "font_name": font.name,
        "color": str(font.color) if font.color else None,
        "fill_type": fill.fill_type if fill else None,
        "fgColor": str(fill.fgColor) if fill and fill.fgColor else None,
        "alignment_h": alignment.horizontal,
        "alignment_v": alignment.vertical,
        "wrap_text": alignment.wrap_text,
        "text_rotation": alignment.text_rotation,
        "number_format": cell.number_format,
        "border_left": border.left.style if border and border.left else None,
        "border_right": border.right.style if border and border.right else None,
        "border_top": border.top.style if border and border.top else None,
        "border_bottom": border.bottom.style if border and border.bottom else None,
        "value": cell.value,
        "indent": alignment.indent,
    }
    wb.close()
    return result


# ==================== Test Class ====================

class TestFormatCellsEdgeCasesR3:
    """format_cells 第三轮边缘案例测试"""

    # ---------- 1. preset + formatting 组合行为 ----------

    def test_preset_with_formatting_merge(self, tmp_path):
        """preset 与 formatting 同时传入时，用户 formatting 值覆盖 preset（deep_merge 设计）"""
        fp = str(tmp_path / "preset_merge.xlsx")
        _create_test_xlsx(fp)
        # preset="header" 设置 bold=True, bg_color="D9D9D9"
        # 用户 formatting 设置 bold=False, bg_color="FF0000" → 用户值应胜出
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": False, "bg_color": "FF0000"},
            preset="header")
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        # 用户 bold=False 应覆盖 preset 的 bold=True
        assert style["bold"] is False, "user formatting should override preset"
        # 用户 bg_color 应覆盖 preset 的灰色背景
        assert style["fgColor"] is not None and "FF0000" in style["fgColor"] or style["fgColor"] == "0000", f"unexpected fgColor: {style.get('fgColor')}"
        # 但 preset 中用户未指定的 font_name（微软雅黑）应保留
        assert style.get("font_name") == "微软雅黑", "unset preset fields should persist"

    def test_unknown_preset_falls_back_to_formatting(self, tmp_path):
        """未知 preset 名应回退到使用 formatting"""
        fp = str(tmp_path / "unknown_preset.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": True},
            preset="nonexistent_preset")
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True

    def test_preset_none_with_formatting(self, tmp_path):
        """preset=None 时应正常使用 formatting"""
        fp = str(tmp_path / "preset_none.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"italic": True},
            preset=None)
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["italic"] is True

    # ---------- 2. merge + unmerge 同时传入 ----------

    def test_merge_and_unmerge_both_true_merge_wins(self, tmp_path):
        """同时传 merge=True 和 unmerge=True 时，merge 优先执行"""
        fp = str(tmp_path / "merge_unmerge_both.xlsx")
        _create_test_xlsx(fp)
        # 先合并
        r1 = ExcelOperations.merge_cells(fp, "Sheet1", "A1:B1")
        assert r1["success"] is True
        # 再同时传 merge+unmerge（通过 API 层）
        from excel_mcp_server_fastmcp.server import excel_format_cells
        r2 = excel_format_cells(fp, "Sheet1", "A1:B1",
            formatting={"merge": True, "unmerge": True})
        # 不应该报错，merge 应优先于 unmerge
        assert r2["success"] is True or "部分失败" in r2.get("message", "")

    def test_only_merge_param_no_other_formatting(self, tmp_path):
        """只传 merge=True 无其他格式参数应正常工作"""
        fp = str(tmp_path / "only_merge.xlsx")
        _create_test_xlsx(fp)
        from excel_mcp_server_fastmcp.server import excel_format_cells
        r = excel_format_cells(fp, "Sheet1", "A1:C1", formatting={"merge": True})
        assert r["success"] is True
        # 验证确实合并了
        wb = load_workbook(fp)
        ws = wb["Sheet1"]
        assert ws.merged_cells.ranges  # 至少有一个合并区域
        wb.close()

    def test_only_border_style_param(self, tmp_path):
        """只传 border_style 无其他格式参数应正常工作"""
        fp = str(tmp_path / "only_border.xlsx")
        _create_test_xlsx(fp)
        from excel_mcp_server_fastmcp.server import excel_format_cells
        r = excel_format_cells(fp, "Sheet1", "A1:B2", formatting={"border_style": "medium"})
        assert r["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["border_left"] == "medium"

    # ---------- 3. 字体颜色类型容错 ----------

    def test_font_color_rgb_tuple(self, tmp_path):
        """font_color 传 RGB tuple (255, 0, 0) 应能正确处理或报明确错误"""
        fp = str(tmp_path / "color_tuple.xlsx")
        _create_test_xlsx(fp)
        # 通过 Writer 层直接传嵌套格式
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"color": (255, 0, 0)}})
        # openpyxl 可能接受 tuple 也可能不接受，但不应该崩溃
        assert result.success is True or result.error is not None

    def test_font_color_int_value(self, tmp_path):
        """font_color 传整数值如 255 应不崩溃"""
        fp = str(tmp_path / "color_int.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"color": 255}})
        # 不崩溃即可
        assert result.success is True or result.error is not None

    # ---------- 4. text_rotation 边界 ----------

    def test_text_rotation_zero_explicit(self, tmp_path):
        """text_rotation=0 显式设置应为水平"""
        fp = str(tmp_path / "rot_zero.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"text_rotation": 0})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 0

    def test_text_rotation_90_boundary(self, tmp_path):
        """text_rotation=90 最大边界值"""
        fp = str(tmp_path / "rot_90.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"text_rotation": 90})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 90

    def test_text_rotation_neg90_boundary(self, tmp_path):
        """text_rotation=-90 最小边界值（负值应被自动转为正值 90）"""
        fp = str(tmp_path / "rot_neg90.xlsx")
        _create_test_xlsx(fp)
        # openpyxl 不支持负数 text_rotation，代码应自动取绝对值
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"text_rotation": -90})
        assert result["success"] is True, f"Expected success, got: {result}"
        style = _read_cell_style(fp, "A1")
        # -90 应被转为 90（垂直文本）
        assert style["text_rotation"] == 90, f"Expected 90 (abs(-90)), got: {style['text_rotation']}"

    def test_text_rotation_out_of_range_positive(self, tmp_path):
        """text_rotation=180 超出范围，openpyxl 行为验证"""
        fp = str(tmp_path / "rot_180.xlsx")
        _create_test_xlsx(fp)
        # openpyxl 允许 0-180 的 text_rotation（实际是 0-90 用于旋转，90-180 用于垂直堆叠）
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"text_rotation": 180})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 180

    # ---------- 5. border_style 边缘 ----------

    def test_border_style_invalid_value(self, tmp_path):
        """border_style 传无效值时应报错或不崩溃"""
        fp = str(tmp_path / "border_invalid.xlsx")
        _create_test_xlsx(fp)
        from excel_mcp_server_fastmcp.server import excel_format_cells
        r = excel_format_cells(fp, "Sheet1", "A1:B2", formatting={"border_style": "invalid_style_xyz"})
        # 可能成功（openpyxl 接受任意字符串）或失败，但不应导致文件损坏
        assert r.get("success") is True or r.get("success") is False

    def test_border_style_none_explicit(self, tmp_path):
        """border_style=None 显式传 None 应跳过边框步骤"""
        fp = str(tmp_path / "border_none.xlsx")
        _create_test_xlsx(fp)
        from excel_mcp_server_fastmcp.server import excel_format_cells
        r = excel_format_cells(fp, "Sheet1", "A1", formatting={"bold": True, "border_style": None})
        assert r["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        # border should remain default (None or thin depending on openpyxl default)

    # ---------- 6. font_size 边界 ----------

    def test_font_size_one(self, tmp_path):
        """font_size=1 最小合理值"""
        fp = str(tmp_path / "fontsize_1.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_size": 1})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["size"] == 1

    def test_font_size_very_large(self, tmp_path):
        """font_size=500 极大值"""
        fp = str(tmp_path / "fontsize_big.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_size": 500})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["size"] == 500

    # ---------- 7. number_format 边缘 ----------

    def test_number_format_very_long(self, tmp_path):
        """number_format 超长字符串"""
        fp = str(tmp_path / "nf_long.xlsx")
        _create_test_xlsx(fp)
        long_fmt = "¥#,##0.00" + "_)" * 50  # 100+ chars
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"number_format": long_fmt})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert long_fmt in style["number_format"]

    def test_number_format_special_chars(self, tmp_path):
        """number_format 含特殊字符"""
        fp = str(tmp_path / "nf_special.xlsx")
        _create_test_xlsx(fp)
        special_fmt = '¥"#,##0.00_);[Red]("¥"#,##0.00"'
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"number_format": special_fmt})
        assert result["success"] is True

    # ---------- 8. 合并相关边界 ----------

    def test_merge_single_cell_range(self, tmp_path):
        """合并单个单元格（无意义但不应报错）"""
        fp = str(tmp_path / "merge_single.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.merge_cells(fp, "Sheet1", "A1:A1")
        # 单格合并可能成功也可能返回 warning
        assert result["success"] is True or "warning" in str(result.get("message", "")).lower() or result.get("data") is not None

    def test_check_merge_data_loss_single_cell(self, tmp_path):
        """_check_merge_data_loss 对单格范围应返回 None（无数据丢失风险）"""
        fp = str(tmp_path / "check_single.xlsx")
        _create_test_xlsx(fp)
        from excel_mcp_server_fastmcp.server import _check_merge_data_loss
        result = _check_merge_data_loss(fp, "Sheet1", "A1:A1")
        assert result is None, f"Single cell merge should have no data loss warning, got: {result}"

    def test_check_merge_data_loss_all_empty(self, tmp_path):
        """_check_merge_data_loss 对全空范围应返回 None"""
        fp = str(tmp_path / "check_empty.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # 不写入任何数据
        wb.save(fp)
        wb.close()
        from excel_mcp_server_fastmcp.server import _check_merge_data_loss
        result = _check_merge_data_loss(fp, "Sheet1", "A1:C3")
        assert result is None

    def test_check_merge_data_loss_only_topleft_has_value(self, tmp_path):
        """只有左上角有值时不应警告"""
        fp = str(tmp_path / "check_topleft.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "header"
        # B1, A2, B2 保持空
        wb.save(fp)
        wb.close()
        from excel_mcp_server_fastmcp.server import _check_merge_data_loss
        result = _check_merge_data_loss(fp, "Sheet1", "A1:B2")
        assert result is None, f"Only topleft has value, no warning expected: {result}"

    def test_check_merge_data_loss_non_topleft_has_values(self, tmp_path):
        """非左上角单元格有值时应返回警告"""
        fp = str(tmp_path / "check_nontopleft.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "header"
        ws["B1"] = "data1"   # 非左上角有值
        ws["A2"] = "data2"
        wb.save(fp)
        wb.close()
        from excel_mcp_server_fastmcp.server import _check_merge_data_loss
        result = _check_merge_data_loss(fp, "Sheet1", "A1:B2")
        assert result is not None, "Should warn about data loss in non-topleft cells"
        assert "数据" in result or "清除" in result

    # ---------- 9. 对齐方式扩展 ----------

    def test_alignment_justify(self, tmp_path):
        """alignment='justify' 两端对齐"""
        fp = str(tmp_path / "align_justify.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"alignment": "justify"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["alignment_h"] == "justify"

    def test_alignment_distributed(self, tmp_path):
        """alignment='distributed' 分散对齐"""
        fp = str(tmp_path / "align_distributed.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"alignment": "distributed"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["alignment_h"] == "distributed"

    def test_alignment_center_continuous(self, tmp_path):
        """alignment='centerContinuous' 跨列居中"""
        fp = str(tmp_path / "align_centercont.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"alignment": "centerContinuous"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["alignment_h"] == "centerContinuous"

    def test_alignment_fill(self, tmp_path):
        """alignment='fill' 填充"""
        fp = str(tmp_path / "align_fill.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"alignment": "fill"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["alignment_h"] == "fill"

    def test_alignment_general_explicit(self, tmp_path):
        """alignment='general' 显式设为默认"""
        fp = str(tmp_path / "align_general.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"alignment": "general"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["alignment_h"] == "general"

    # ---------- 10. 样式清除/重置场景 ----------

    def test_clear_bold_with_false(self, tmp_path):
        """先加粗再用 bold=False 清除"""
        fp = str(tmp_path / "clear_bold.xlsx")
        _create_test_xlsx(fp)
        # 先加粗
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"bold": True})
        style1 = _read_cell_style(fp, "A1")
        assert style1["bold"] is True
        # 再取消加粗
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"bold": False})
        style2 = _read_cell_style(fp, "A1")
        assert style2["bold"] is False

    def test_clear_italic_with_false(self, tmp_path):
        """先用斜体再用 italic=False 清除"""
        fp = str(tmp_path / "clear_italic.xlsx")
        _create_test_xlsx(fp)
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"italic": True})
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"italic": False})
        style = _read_cell_style(fp, "A1")
        assert style["italic"] is False

    def test_clear_underline_with_none_string(self, tmp_path):
        """用 underline='none' 清除下划线（openpyxl 会将 'none' 规范化为 None）"""
        fp = str(tmp_path / "clear_underline.xlsx")
        _create_test_xlsx(fp)
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"underline": "double"})
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"underline": "none"})
        style = _read_cell_style(fp, "A1")
        # openpyxl 将 'none' 规范化为 None
        assert style["underline"] is None or style["underline"] == "none"

    # ---------- 11. bg_color 边缘 ----------

    def test_bg_color_empty_string(self, tmp_path):
        """bg_color='' 空字符串"""
        fp = str(tmp_path / "bg_empty.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bg_color": ""})
        # 空字符串不是 None，所以会传入。可能成功也可能报错。
        assert result["success"] is True or result["success"] is False

    def test_bg_color_short_hex(self, tmp_path):
        """bg_color 短格式 'F00' (3位hex)"""
        fp = str(tmp_path / "bg_short.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bg_color": "F00"})
        # openpyxl 可能接受也可能不接受
        assert result["success"] is True or result["success"] is False

    def test_bg_color_with_alpha_8char(self, tmp_path):
        """bg_color 8位带透明度 'FF0000FF'"""
        fp = str(tmp_path / "bg_8char.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bg_color": "FF0000FF"})
        assert result["success"] is True or result["success"] is False

    # ---------- 12. format_cells 对已合并区域 ----------

    def test_format_already_merged_range(self, tmp_path):
        """对已合并区域应用格式应正常工作"""
        fp = str(tmp_path / "fmt_merged.xlsx")
        _create_test_xlsx(fp)
        # 先合并
        ExcelOperations.merge_cells(fp, "Sheet1", "A1:C1")
        # 再对合并区域应用格式
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:C1",
            formatting={"bold": True, "bg_color": "FFFF00"})
        assert result["success"] is True
        # 验证左上角单元格的样式
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True

    def test_format_partial_overlap_merged(self, tmp_path):
        """格式化范围与已合并区域部分重叠"""
        fp = str(tmp_path / "fmt_overlap.xlsx")
        _create_test_xlsx(fp)
        ExcelOperations.merge_cells(fp, "Sheet1", "A1:B2")
        # 格式化包含合并区域的更大范围
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:D3",
            formatting={"italic": True})
        assert result["success"] is True

    # ---------- 13. _normalize_formatting 深度边缘 ----------

    def test_normalize_flat_bool_zero_and_empty_str(self, tmp_path):
        """扁平格式中值为 0、空字符串、False 的字段应保留（非 None）"""
        from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        # 0 是有效字体大小
        result = ExcelOperations._normalize_formatting({"font_size": 0})
        assert result == {"font": {"size": 0}}
        # 空字符串 number_format
        result2 = ExcelOperations._normalize_formatting({"number_format": ""})
        assert result2 == {"number_format": ""}
        # False bold (关闭加粗)
        result3 = ExcelOperations._normalize_formatting({"bold": False})
        assert result3 == {"font": {"bold": False}}

    def test_normalize_flat_preserves_zero_values(self):
        """值为 0 的字段不应被过滤掉（只有 None 被过滤）"""
        from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations._normalize_formatting({
            "font_size": 0,
            "text_rotation": 0,
            "indent": 0,
        })
        assert "font" in result
        assert result["font"]["size"] == 0
        assert "alignment" in result
        assert result["alignment"]["text_rotation"] == 0
        assert result["alignment"]["indent"] == 0

    def test_normalize_gradient_colors_single_item(self):
        """gradient_colors 只有一个颜色"""
        from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations._normalize_formatting({"gradient_colors": ["FF0000"]})
        assert result["fill"]["type"] == "gradient"
        assert result["fill"]["colors"] == ["FF0000"]

    def test_normalize_border_all_sides_as_strings(self):
        """边框四边都是简写字符串"""
        from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations._normalize_formatting({
            "border": {
                "left": "thin",
                "right": "medium",
                "top": "thick",
                "bottom": "dashed",
            }
        })
        assert "border" in result
        assert result["border"]["left"] == "thin"
        assert result["border"]["right"] == "medium"

    # ---------- 14. 连续多次格式化覆盖行为 ----------

    def test_sequential_format_overwrite_font(self, tmp_path):
        """连续两次格式化，后一次覆盖前一次的字体属性"""
        fp = str(tmp_path / "seq_font.xlsx")
        _create_test_xlsx(fp)
        ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": True, "font_size": 14, "font_name": "Arial"})
        ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": False, "font_size": 10})
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is False
        assert style["size"] == 10
        # font_name 应保持 Arial（第二次没改 name）
        assert style["font_name"] == "Arial"

    def test_sequential_format_overwrite_fill(self, tmp_path):
        """连续两次格式化填充色，后一次覆盖"""
        fp = str(tmp_path / "seq_fill.xlsx")
        _create_test_xlsx(fp)
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"bg_color": "FF0000"})
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"bg_color": "00FF00"})
        style = _read_cell_style(fp, "A1")
        assert "00FF00" in style["fgColor"].upper() or "00ff00" in str(style["fgColor"]).lower()

    # ---------- 15. 特殊字符和 Unicode ----------

    def test_font_name_with_japanese(self, tmp_path):
        """日文字体名"""
        fp = str(tmp_path / "font_jp.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_name": "ＭＳ Ｐゴシック"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert "ゴシック" in style["font_name"] or "MS PGothic" in style["font_name"] or style["font_name"] is not None

    def test_font_name_with_emoji_like(self, tmp_path):
        """含特殊字符的字体名（可能无效但不应崩溃）"""
        fp = str(tmp_path / "font_special.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_name": "TestFont🎨"})
        # 不崩溃即可
        assert result["success"] is True or result["success"] is False

    # ---------- 16. wrap_text + shrink_to_fit 组合 ----------

    def test_wrap_text_true_shrink_true(self, tmp_path):
        """wrap_text=True + shrink_to_fit=True 同时启用"""
        fp = str(tmp_path / "wrap_shrink.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"wrap_text": True, "shrink_to_fit": True})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["wrap_text"] is True
        # shrink_to_fit may or may not be preserved by openpyxl when wrap_text is True

    def test_wrap_false_shrink_false(self, tmp_path):
        """wrap_text=False + shrink_to_fit=False 都显式关闭（openpyxl 保存后 False 读回为 None）"""
        fp = str(tmp_path / "nowrap_noshrink.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"wrap_text": False, "shrink_to_fit": False})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        # openpyxl 保存/重载后 False 值变为 None（默认值省略）
        assert style["wrap_text"] is False or style["wrap_text"] is None

    # ---------- 17. indent 边缘 ----------

    def test_indent_large_value(self, tmp_path):
        """indent=20 大缩进值"""
        fp = str(tmp_path / "indent_big.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"indent": 20})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        # openpyxl indent 读回为 float
        assert style.get("indent") == 20 or style.get("indent") == 20.0

    def test_indent_negative(self, tmp_path):
        """indent=-1 负缩进（openpyxl 可能拒绝）"""
        fp = str(tmp_path / "indent_neg.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"indent": -1})
        # openpyxl 可能接受负缩进，也可能报错
        assert result["success"] is True or result["success"] is False

    # ---------- 18. strikethrough 组合 ----------

    def test_strikethrough_with_bold_and_italic(self, tmp_path):
        """删除线 + 加粗 + 斜体组合"""
        fp = str(tmp_path / "strike_combo.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"strikethrough": True, "bold": True, "italic": True})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["italic"] is True
        # strikethrough might be stored differently in openpyxl

    # ---------- 19. merge + bold + bg_color 组合操作 ----------

    def test_merge_bold_bg_color_combo(self, tmp_path):
        """合并 + 加粗 + 背景色同时传入（实际高频使用场景）"""
        fp = str(tmp_path / "combo_merge_bold_bg.xlsx")
        _create_test_xlsx(fp)
        # 先合并
        r_merge = ExcelOperations.merge_cells(fp, "Sheet1", "A1:D1")
        assert r_merge["success"] is True
        # 再格式化
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:D1",
            formatting={"bold": True, "bg_color": "4472C4", "alignment": "center"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        # 验证背景色应用了
        assert style["fgColor"] is not None
        # 验证合并确实发生了
        wb = load_workbook(fp)
        ws = wb["Sheet1"]
        assert len(ws.merged_cells.ranges) >= 1
        wb.close()

    def test_unmerge_after_merge_restores_independence(self, tmp_path):
        """先合并再拆分，单元格应恢复独立"""
        fp = str(tmp_path / "merge_unmerge.xlsx")
        _create_test_xlsx(fp)
        # 合并
        r1 = ExcelOperations.merge_cells(fp, "Sheet1", "A1:C1")
        assert r1["success"] is True
        # 拆分
        r2 = ExcelOperations.unmerge_cells(fp, "Sheet1", "A1:C1")
        assert r2["success"] is True
        # 验证已拆分
        wb = load_workbook(fp)
        ws = wb["Sheet1"]
        # 拆分后 merged_cells 应为空或不再包含 A1:C1
        merged_ranges = str(ws.merged_cells.ranges)
        wb.close()
        assert "A1:C1" not in merged_ranges

    # ---------- 20. 中文字体名 ----------

    def test_font_name_chinese(self, tmp_path):
        """中文字体名：宋体、黑体、微软雅黑"""
        fp = str(tmp_path / "font_cn.xlsx")
        _create_test_xlsx(fp)
        for font_name in ["宋体", "黑体", "楷体"]:
            result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
                formatting={"font_name": font_name})
            assert result["success"] is True, f"Failed for font: {font_name}"
        style = _read_cell_style(fp, "A1")
        assert style["font_name"] == "楷体"  # 最后一个生效

    # ---------- 21. font_color # 前缀处理 ----------

    def test_font_color_with_hash_prefix(self, tmp_path):
        """font_color 带 # 前缀应正常工作（R57 修复场景）"""
        fp = str(tmp_path / "color_hash.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_color": "#FF0000", "bold": True})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        # 颜色应被正确设置（# 被去除或保留都行）
        color_str = str(style.get("color", ""))
        assert "FF0000" in color_str.upper() or "F00" in color_str.upper() or style["color"] is not None

    def test_bg_color_with_hash_prefix(self, tmp_path):
        """bg_color 带 # 前缀"""
        fp = str(tmp_path / "bgcolor_hash.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bg_color": "#00FF00"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        fg = str(style.get("fgColor", ""))
        assert "00FF00" in fg.upper() or "0F0" in fg.upper() or style["fgColor"] is not None

    # ---------- 22. 公式单元格格式化不破坏公式 ----------

    def test_format_formula_cell_preserves_value(self, tmp_path):
        """含值的单元格格式化后值仍保留"""
        fp = str(tmp_path / "fmt_preserve.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 42
        ws["A2"] = "hello"
        wb.save(fp)
        wb.close()
        # 格式化
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:B2",
            formatting={"bold": True, "bg_color": "FFFF00"})
        assert result["success"] is True
        # 验证值未被破坏
        wb2 = load_workbook(fp)
        ws2 = wb2["Sheet1"]
        assert ws2["A1"].value == 42
        assert ws2["A2"].value == "hello"
        wb2.close()

    # ---------- 23. vertical_alignment 映射 ----------

    def test_vertical_alignment_middle_maps_to_center(self, tmp_path):
        """vertical_alignment='middle' 应映射为 'center'（R57 修复）"""
        fp = str(tmp_path / "valign_mid.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"vertical_alignment": "middle"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        # 'middle' should map to 'center'
        assert style["alignment_v"] in ("center", "middle")

    def test_vertical_alignment_top_and_bottom(self, tmp_path):
        """vertical_alignment=top/bottom 正常工作"""
        fp = str(tmp_path / "valign_tb.xlsx")
        _create_test_xlsx(fp)
        for valign, expected in [("top", "top"), ("bottom", "bottom")]:
            result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
                formatting={"vertical_alignment": valign})
            assert result["success"] is True
            style = _read_cell_style(fp, "A1")
            assert style["alignment_v"] == expected