# -*- coding: utf-8 -*-
"""
Excel Operations 边缘情况测试

覆盖更多边缘情况和错误处理路径
"""

import pytest
from pathlib import Path
from openpyxl import Workbook
from src.api.excel_operations import ExcelOperations


class TestExcelOpsEdgeMore:
    """更多边缘情况测试"""

    @pytest.fixture
    def edge_file(self, temp_dir):
        """创建边缘测试文件"""
        file_path = temp_dir / "edge_more.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "EdgeSheet"
        
        # 各种特殊值
        ws['A1'] = "①"  # 特殊字符
        ws['A2'] = "①"  # 重复
        ws['A3'] = " "   # 空格
        ws['A4'] = "\t"  # Tab
        ws['A5'] = "\n" # 换行
        
        wb.save(file_path)
        return str(file_path)

    def test_search_special_chars(self, edge_file):
        """测试搜索特殊字符"""
        result = ExcelOperations.search(
            file_path=edge_file,
            pattern="①",
            sheet_name="EdgeSheet"
        )
        assert result is not None

    def test_search_whitespace_only(self, edge_file):
        """测试仅搜索空格"""
        result = ExcelOperations.search(
            file_path=edge_file,
            pattern=" ",
            sheet_name="EdgeSheet"
        )
        assert result is not None


class TestExcelOpsMoreErrors:
    """更多错误处理测试"""

    def test_file_not_found_get_range(self):
        """测试文件不存在的错误处理"""
        result = ExcelOperations.get_range(
            file_path="/nonexistent/file.xlsx",
            range_expression="Sheet1!A1"
        )
        assert result['success'] is False

    def test_file_not_found_update(self):
        """测试更新不存在的文件"""
        result = ExcelOperations.update_range(
            file_path="/nonexistent/file.xlsx",
            range_expression="Sheet1!A1",
            data=[[1]]
        )
        assert result['success'] is False


class TestExcelOpsFormatting:
    """格式化测试"""

    @pytest.fixture
    def format_file(self, temp_dir):
        """创建格式化文件"""
        file_path = temp_dir / "format_more.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "FormatSheet"
        
        ws['A1'] = "Data"
        
        wb.save(file_path)
        return str(file_path)

    def test_format_cells_bold(self, format_file):
        """测试粗体格式化"""
        result = ExcelOperations.format_cells(
            file_path=format_file,
            sheet_name="FormatSheet",
            range="A1",
            formatting={"font": {"bold": True}}
        )
        assert result is not None

    def test_format_cells_fill(self, format_file):
        """测试背景色格式化"""
        result = ExcelOperations.format_cells(
            file_path=format_file,
            sheet_name="FormatSheet",
            range="A1",
            formatting={"fill": {"patternType": "solid", "fgColor": "FFFF00"}}
        )
        assert result is not None


class TestExcelOpsConvert:
    """格式转换测试"""

    @pytest.fixture
    def convert_file(self, temp_dir):
        """创建转换测试文件"""
        file_path = temp_dir / "convert_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws['A1'] = "Test"
        
        wb.save(file_path)
        return str(file_path)

    def test_convert_to_csv(self, convert_file, temp_dir):
        """测试转换为CSV"""
        csv_path = temp_dir / "output.csv"
        
        result = ExcelOperations.convert_format(
            input_path=convert_file,
            output_path=str(csv_path),
            target_format="csv"
        )
        
        assert result is not None


class TestExcelOpsMoreCompare:
    """更多比较测试"""

    @pytest.fixture
    def comp1(self, temp_dir):
        """创建比较文件1"""
        file_path = temp_dir / "c1.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        
        ws['A1'] = "H"
        ws['A2'] = 1
        ws['A3'] = 2
        
        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def comp2(self, temp_dir):
        """创建比较文件2"""
        file_path = temp_dir / "c2.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        
        ws['A1'] = "H"
        ws['A2'] = 1
        ws['A3'] = 3
        
        wb.save(file_path)
        return str(file_path)

    def test_compare_sheets_value_only(self, comp1, comp2):
        """测试仅比较值"""
        result = ExcelOperations.compare_sheets(
            file1_path=comp1,
            sheet1_name="S",
            file2_path=comp2,
            sheet2_name="S"
        )
        
        assert result is not None
