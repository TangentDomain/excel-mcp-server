# -*- coding: utf-8 -*-
"""
Excel Operations 深度测试套件

覆盖 excel_operations.py 中的更多函数和边缘情况
"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from src.api.excel_operations import ExcelOperations


class TestExcelOperationsDataValidation:
    """ExcelOperations 数据验证测试"""

    @pytest.fixture
    def validation_file(self, temp_dir):
        """创建验证测试文件"""
        file_path = temp_dir / "validation_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ValidationSheet"
        
        # 添加各种类型的数据
        ws['A1'] = "ID"
        ws['A2'] = 1
        ws['A3'] = 2
        ws['A4'] = 3
        
        ws['B1'] = "Name"
        ws['B2'] = "Alice"
        ws['B3'] = "Bob"
        ws['B4'] = "Charlie"
        
        ws['C1'] = "Age"
        ws['C2'] = 25
        ws['C3'] = 30
        ws['C4'] = 35
        
        wb.save(file_path)
        return str(file_path)

    def test_find_last_row_basic(self, validation_file):
        """测试查找最后一行基本功能"""
        result = ExcelOperations.find_last_row(
            file_path=validation_file,
            sheet_name="ValidationSheet"
        )
        
        assert result['success'] is True

    def test_find_last_row_with_column_letter(self, validation_file):
        """测试使用列字母查找最后一行"""
        result = ExcelOperations.find_last_row(
            file_path=validation_file,
            sheet_name="ValidationSheet",
            column="A"
        )
        
        assert result['success'] is True

    def test_find_last_row_with_column_number(self, validation_file):
        """测试使用列号查找最后一行"""
        result = ExcelOperations.find_last_row(
            file_path=validation_file,
            sheet_name="ValidationSheet",
            column=2
        )
        
        assert result['success'] is True

    def test_find_last_row_empty_column(self, temp_dir):
        """测试查找空列的最后一行"""
        file_path = temp_dir / "empty_column.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "EmptySheet"
        
        # 只有表头，没有数据
        ws['A1'] = "Header"
        
        wb.save(file_path)
        
        result = ExcelOperations.find_last_row(
            file_path=str(file_path),
            sheet_name="EmptySheet"
        )
        
        assert result['success'] is True


class TestExcelOperationsUpdateRange:
    """ExcelOperations 更新范围测试"""

    @pytest.fixture
    def update_file(self, temp_dir):
        """创建更新测试文件"""
        file_path = temp_dir / "update_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "UpdateSheet"
        
        ws['A1'] = "ID"
        ws['A2'] = 1
        ws['A3'] = 2
        
        wb.save(file_path)
        return str(file_path)

    def test_update_range_basic(self, update_file):
        """测试基本范围更新"""
        result = ExcelOperations.update_range(
            file_path=update_file,
            range_expression="UpdateSheet!B1:B3",
            data=[["Name1"], ["Name2"], ["Name3"]],
            preserve_formulas=True
        )
        
        assert result['success'] is True

    def test_update_range_with_formulas(self, update_file):
        """测试更新包含公式的单元格"""
        result = ExcelOperations.update_range(
            file_path=update_file,
            range_expression="UpdateSheet!C1:C3",
            data=[["=A1"], ["=A2"], ["=A3"]],
            preserve_formulas=True
        )
        
        assert result['success'] is True


class TestExcelOperationsSearchAdvanced:
    """ExcelOperations 高级搜索测试"""

    @pytest.fixture
    def advanced_search_file(self, temp_dir):
        """创建高级搜索测试文件"""
        file_path = temp_dir / "advanced_search.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "SearchSheet"
        
        # 包含各种搜索模式
        ws['A1'] = "Item123"
        ws['A2'] = "item456"
        ws['A3'] = "TEST789"
        ws['A4'] = "123abc"
        ws['A5'] = "ABC123"
        
        wb.save(file_path)
        return str(file_path)

    def test_search_whole_word_match(self, advanced_search_file):
        """测试全词匹配"""
        result = ExcelOperations.search(
            file_path=advanced_search_file,
            pattern="Item",
            sheet_name="SearchSheet",
            whole_word=True
        )
        
        assert result['success'] is True

    def test_search_include_values_only(self, advanced_search_file):
        """测试只搜索值"""
        result = ExcelOperations.search(
            file_path=advanced_search_file,
            pattern="123",
            sheet_name="SearchSheet",
            include_values=True,
            include_formulas=False
        )
        
        assert result['success'] is True


class TestExcelOperationsSheetManagement:
    """ExcelOperations 工作表管理测试"""

    @pytest.fixture
    def sheet_mgmt_file(self, temp_dir):
        """创建工作表管理测试文件"""
        file_path = temp_dir / "sheet_mgmt.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        wb.save(file_path)
        return str(file_path)

    def test_create_sheet_with_index(self, sheet_mgmt_file):
        """测试在指定位置创建工作表"""
        result = ExcelOperations.create_sheet(
            file_path=sheet_mgmt_file,
            sheet_name="NewSheet",
            index=0
        )
        
        assert result is not None

    def test_rename_sheet_basic(self, sheet_mgmt_file):
        """测试重命名工作表"""
        result = ExcelOperations.rename_sheet(
            file_path=sheet_mgmt_file,
            old_name="Sheet1",
            new_name="RenamedSheet"
        )
        
        assert result is not None


class TestExcelOperationsCompareAdvanced:
    """ExcelOperations 高级比较测试"""

    @pytest.fixture
    def compare_data_file1(self, temp_dir):
        """创建比较数据文件1"""
        file_path = temp_dir / "compare_data1.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        ws['A1'] = "ID"
        ws['B1'] = "Value"
        
        for i in range(2, 12):
            ws[f'A{i}'] = i - 1
            ws[f'B{i}'] = i * 10
        
        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def compare_data_file2(self, temp_dir):
        """创建比较数据文件2"""
        file_path = temp_dir / "compare_data2.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        ws['A1'] = "ID"
        ws['B1'] = "Value"
        
        # 有差异的数据
        for i in range(2, 12):
            ws[f'A{i}'] = i - 1
            if i < 8:
                ws[f'B{i}'] = i * 20  # 不同的值
            else:
                ws[f'B{i}'] = i * 10
        
        wb.save(file_path)
        return str(file_path)

    def test_compare_files_advanced(self, compare_data_file1, compare_data_file2):
        """测试高级文件比较"""
        result = ExcelOperations.compare_files(
            file1_path=compare_data_file1,
            file2_path=compare_data_file2
        )
        
        assert result is not None


class TestExcelOperationsHeadersAdvanced:
    """ExcelOperations 高级表头测试"""

    @pytest.fixture
    def header_file(self, temp_dir):
        """创建表头测试文件"""
        file_path = temp_dir / "header_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "HeaderSheet"
        
        # 多列表头
        ws['A1'] = "Col1"
        ws['B1'] = "Col2"
        ws['C1'] = "Col3"
        ws['D1'] = "Col4"
        ws['E1'] = "Col5"
        
        ws['A2'] = "Field1"
        ws['B2'] = "Field2"
        ws['C2'] = "Field3"
        ws['D2'] = "Field4"
        ws['E2'] = "Field5"
        
        wb.save(file_path)
        return str(file_path)

    def test_get_headers_with_row(self, header_file):
        """测试指定表头行"""
        result = ExcelOperations.get_headers(
            file_path=header_file,
            sheet_name="HeaderSheet",
            header_row=1
        )
        
        assert result['success'] is True

    def test_get_headers_with_max_columns(self, header_file):
        """测试限制最大列数"""
        result = ExcelOperations.get_headers(
            file_path=header_file,
            sheet_name="HeaderSheet",
            max_columns=3
        )
        
        assert result['success'] is True


class TestExcelOperationsEdgeCases:
    """ExcelOperations 边缘情况测试"""

    @pytest.fixture
    def edge_file(self, temp_dir):
        """创建边缘测试文件"""
        file_path = temp_dir / "edge_ops.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "EdgeSheet"
        
        # 特殊值
        ws['A1'] = None
        ws['A2'] = ""
        ws['A3'] = 0
        ws['A4'] = False
        ws['A5'] = "None"
        ws['A6'] = "NULL"
        
        wb.save(file_path)
        return str(file_path)

    def test_search_none_values(self, edge_file):
        """测试搜索None值"""
        result = ExcelOperations.search(
            file_path=edge_file,
            pattern="None",
            sheet_name="EdgeSheet"
        )
        
        assert result['success'] is True

    def test_search_empty_string(self, edge_file):
        """测试搜索空字符串"""
        result = ExcelOperations.search(
            file_path=edge_file,
            pattern="",
            sheet_name="EdgeSheet"
        )
        
        assert result['success'] is True

    def test_get_range_with_special_values(self, edge_file):
        """测试获取包含特殊值的范围"""
        result = ExcelOperations.get_range(
            file_path=edge_file,
            range_expression="EdgeSheet!A1:A6"
        )
        
        assert result['success'] is True
