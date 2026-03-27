# -*- coding: utf-8 -*-
"""
Excel Operations 高级功能测试套件

覆盖 excel_operations.py 中的高级功能 - 修正版
"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations


class TestExcelOperationsAdvanced:
    """ExcelOperations 高级功能测试"""

    @pytest.fixture
    def test_file(self, temp_dir):
        """创建测试文件"""
        file_path = temp_dir / "advanced_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        test_data = [
            ["ID", "Name", "Value", "Category", "Status"],
            [1, "Item1", 100, "A", "Active"],
            [2, "Item2", 200, "B", "Inactive"],
            [3, "Item3", 300, "A", "Active"],
            [4, "Item4", 400, "C", "Inactive"],
            [5, "Item5", 500, "B", "Active"],
        ]
        
        for row_idx, row_data in enumerate(test_data, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def multi_sheet_file(self, temp_dir):
        """创建多工作表文件"""
        file_path = temp_dir / "multi_sheet_test.xlsx"
        
        wb = Workbook()
        
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1['A1'] = "ID"
        ws1['B1'] = "Name"
        ws1['A2'] = 1
        ws1['B2'] = "First"
        
        ws2 = wb.create_sheet(title="Sheet2")
        ws2['A1'] = "ID"
        ws2['B1'] = "Value"
        ws2['A2'] = 100
        ws2['B2'] = 200
        
        ws3 = wb.create_sheet(title="Sheet3")
        ws3['A1'] = "Name"
        ws3['B1'] = "Data"
        ws3['A2'] = "Test"
        ws3['B2'] = "Sample"
        
        wb.save(file_path)
        return str(file_path)

    # ==================== 搜索增强测试 ====================

    def test_search_with_multiple_options(self, test_file):
        """测试多种搜索选项组合"""
        result = ExcelOperations.search(
            file_path=test_file,
            pattern="Item",
            sheet_name="TestSheet",
            case_sensitive=False,
            whole_word=False,
            use_regex=False,
            include_values=True,
            include_formulas=False
        )
        
        assert result['success'] is True

    def test_search_specific_columns(self, test_file):
        """测试搜索特定列"""
        result = ExcelOperations.search(
            file_path=test_file,
            pattern="Item",
            sheet_name="TestSheet",
            range="B1:B10"
        )
        
        assert result['success'] is True

    # ==================== 目录搜索增强测试 ====================

    def test_search_directory_with_all_options(self, temp_dir_with_excel_files):
        """测试目录搜索所有选项"""
        result = ExcelOperations.search_directory(
            directory_path=temp_dir_with_excel_files,
            pattern="test",
            case_sensitive=False,
            whole_word=False,
            use_regex=False,
            include_values=True,
            include_formulas=False,
            recursive=True,
            file_extensions=[".xlsx", ".xlsm"],
            file_pattern=None,
            max_files=50
        )
        
        assert result['success'] is True

    # ==================== 工作表管理增强测试 ====================

    def test_delete_first_sheet(self, multi_sheet_file):
        """测试删除第一个工作表"""
        result = ExcelOperations.delete_sheet(
            file_path=multi_sheet_file,
            sheet_name="Sheet1"
        )
        
        assert result['success'] is True

    def test_list_sheets_info(self, multi_sheet_file):
        """测试列出工作表详细信息"""
        result = ExcelOperations.list_sheets(multi_sheet_file)
        
        assert result['success'] is True
        assert 'sheets' in result['data']

    # ==================== 数据查找增强测试 ====================

    def test_find_last_row_with_column(self, test_file):
        """测试查找特定列的最后一行"""
        result = ExcelOperations.find_last_row(
            file_path=test_file,
            sheet_name="TestSheet",
            column="A"
        )
        
        assert result['success'] is True

    # ==================== 错误处理增强测试 ====================

    def test_search_invalid_directory(self):
        """测试搜索无效目录"""
        result = ExcelOperations.search_directory(
            directory_path="/nonexistent/directory",
            pattern="test"
        )
        
        assert result['success'] is False

    def test_delete_nonexistent_sheet(self, test_file):
        """测试删除不存在的工作表"""
        result = ExcelOperations.delete_sheet(
            file_path=test_file,
            sheet_name="NonExistentSheet"
        )
        
        assert result['success'] is False

    # ==================== 边界条件测试 ====================

    def test_search_empty_pattern(self, test_file):
        """测试空搜索模式"""
        result = ExcelOperations.search(
            file_path=test_file,
            pattern=""
        )
        
        assert result is not None


class TestExcelCompareOperations:
    """Excel比较操作测试"""

    @pytest.fixture
    def file1(self, temp_dir):
        """创建第一个测试文件"""
        file_path = temp_dir / "compare1.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        ws['A1'] = "ID"
        ws['B1'] = "Name"
        ws['C1'] = "Value"
        ws['A2'] = 1
        ws['B2'] = "Item1"
        ws['C2'] = 100
        ws['A3'] = 2
        ws['B3'] = "Item2"
        ws['C3'] = 200
        
        wb.save(str(file_path))
        return str(file_path)

    def test_check_duplicate_ids(self, file1):
        """测试重复ID检查"""
        result = ExcelOperations.check_duplicate_ids(
            file_path=file1,
            sheet_name="Data",
            id_column=1
        )
        
        assert result is not None


class TestExcelFileInfoOperations:
    """Excel文件信息操作测试"""

    @pytest.fixture
    def info_test_file(self, temp_dir):
        """创建测试文件"""
        file_path = temp_dir / "info_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws2 = wb.create_sheet(title="Sheet2")
        ws2['A1'] = "Data"
        
        wb.save(str(file_path))
        return str(file_path)

    def test_get_file_info_detailed(self, info_test_file):
        """测试获取详细文件信息"""
        result = ExcelOperations.get_file_info(info_test_file)
        
        assert result['success'] is True
        assert 'data' in result

    def test_get_file_info_nonexistent(self):
        """测试获取不存在文件的信息"""
        result = ExcelOperations.get_file_info("/nonexistent/file.xlsx")
        
        assert result['success'] is False


class TestExcelHeadersOperations:
    """Excel表头操作测试"""

    @pytest.fixture
    def headers_test_file(self, temp_dir):
        """创建带表头的测试文件"""
        file_path = temp_dir / "headers_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        ws['A1'] = "ID描述"
        ws['B1'] = "名称描述"
        ws['C1'] = "数值描述"
        ws['A2'] = "id"
        ws['B2'] = "name"
        ws['C2'] = "value"
        ws['A3'] = 1
        ws['B3'] = "Item1"
        ws['C3'] = 100
        
        wb.save(str(file_path))
        return str(file_path)

    def test_get_headers_dual_row(self, headers_test_file):
        """测试获取双行表头"""
        result = ExcelOperations.get_headers(
            file_path=headers_test_file,
            sheet_name="TestSheet",
            header_row=1
        )
        
        assert result['success'] is True
        assert 'descriptions' in result or 'headers' in result

    def test_get_headers_custom_row(self, headers_test_file):
        """测试获取自定义行表头"""
        result = ExcelOperations.get_headers(
            file_path=headers_test_file,
            sheet_name="TestSheet",
            header_row=2,
            max_columns=3
        )
        
        assert result['success'] is True

    def test_get_sheet_headers(self, headers_test_file):
        """测试获取所有工作表表头"""
        result = ExcelOperations.get_sheet_headers(headers_test_file)
        
        assert result['success'] is True
