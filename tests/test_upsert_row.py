"""Tests for upsert_row functionality."""
import os
import pytest
from openpyxl import Workbook


FIXTURE_PATH = os.path.join(os.path.dirname(__file__), 'test_data', 'upsert_test.xlsx')


@pytest.fixture
def workbook():
    """Create a test workbook with sample data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "技能配置"

    # 单行表头
    ws.append(["skill_id", "skill_name", "skill_type", "damage", "cooldown"])
    ws.append([1001, "火球术", "法师", 150, 3.0])
    ws.append([1002, "冰冻术", "法师", 120, 4.0])
    ws.append([1003, "斩击", "战士", 200, 2.0])

    os.makedirs(os.path.dirname(FIXTURE_PATH), exist_ok=True)
    wb.save(FIXTURE_PATH)
    yield FIXTURE_PATH
    if os.path.exists(FIXTURE_PATH):
        os.remove(FIXTURE_PATH)


@pytest.fixture
def dual_header_workbook():
    """Create a test workbook with dual-header (Chinese + English)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "装备配置"

    # 双行表头
    ws.append(["装备ID", "装备名称", "品质", "攻击力"])
    ws.append(["equip_id", "equip_name", "rarity", "attack"])
    ws.append([2001, "铁剑", "普通", 50])
    ws.append([2002, "魔杖", "稀有", 120])
    ws.append([2003, "圣剑", "传说", 300])

    os.makedirs(os.path.dirname(FIXTURE_PATH), exist_ok=True)
    wb.save(FIXTURE_PATH)
    yield FIXTURE_PATH
    if os.path.exists(FIXTURE_PATH):
        os.remove(FIXTURE_PATH)


class TestUpsertUpdate:
    """Tests for upsert_row when the key exists (UPDATE path)."""

    def test_update_existing_row(self, workbook):
        """Update an existing row by key."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.upsert_row(
            workbook, "技能配置",
            key_column="skill_id", key_value=1001,
            updates={"damage": 180, "cooldown": 2.5}
        )
        assert result['success'] is True
        assert result['data']['action'] == 'update'
        assert result['data']['updated_count'] == 2

    def test_update_persists(self, workbook):
        """Updated values are persisted in the file."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        ExcelOperations.upsert_row(
            workbook, "技能配置",
            key_column="skill_id", key_value=1002,
            updates={"damage": 999}
        )
        from openpyxl import load_workbook
        wb = load_workbook(workbook)
        # Row 3 is 1002 (row 1=header, row 2=1001, row 3=1002)
        assert wb["技能配置"].cell(row=3, column=4).value == 999
        wb.close()

    def test_update_does_not_affect_other_rows(self, workbook):
        """Updating one row doesn't change other rows."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        ExcelOperations.upsert_row(
            workbook, "技能配置",
            key_column="skill_id", key_value=1001,
            updates={"damage": 0}
        )
        from openpyxl import load_workbook
        wb = load_workbook(workbook)
        # Row 4 (1003 斩击) should be unchanged
        assert wb["技能配置"].cell(row=4, column=4).value == 200
        wb.close()

    def test_update_string_key(self, workbook):
        """Upsert with string key value."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.upsert_row(
            workbook, "技能配置",
            key_column="skill_name", key_value="斩击",
            updates={"damage": 250}
        )
        assert result['success'] is True
        assert result['data']['action'] == 'update'


class TestUpsertInsert:
    """Tests for upsert_row when the key doesn't exist (INSERT path)."""

    def test_insert_new_row(self, workbook):
        """Insert a new row when key doesn't exist."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.upsert_row(
            workbook, "技能配置",
            key_column="skill_id", key_value=1004,
            updates={"skill_id": 1004, "skill_name": "雷电术", "skill_type": "法师", "damage": 180, "cooldown": 3.5}
        )
        assert result['success'] is True
        assert result['data']['action'] == 'insert'
        assert result['data']['row'] == 5  # After 1 header + 3 data rows

    def test_insert_persists(self, workbook):
        """Inserted row is persisted in the file."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        ExcelOperations.upsert_row(
            workbook, "技能配置",
            key_column="skill_id", key_value=1005,
            updates={"skill_id": 1005, "skill_name": "治愈术", "damage": 0}
        )
        from openpyxl import load_workbook
        wb = load_workbook(workbook)
        assert wb["技能配置"].cell(row=5, column=1).value == 1005
        assert wb["技能配置"].cell(row=5, column=2).value == "治愈术"
        wb.close()

    def test_insert_key_auto_filled(self, workbook):
        """Key column value is auto-filled even if not in updates dict."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.upsert_row(
            workbook, "技能配置",
            key_column="skill_id", key_value=1006,
            updates={"skill_name": "新风术"}  # skill_id not in updates
        )
        assert result['success'] is True
        assert result['data']['action'] == 'insert'
        from openpyxl import load_workbook
        wb = load_workbook(workbook)
        assert wb["技能配置"].cell(row=5, column=1).value == 1006
        wb.close()


class TestUpsertDualHeader:
    """Tests for upsert_row with dual-header (header_row=2)."""

    def test_update_dual_header(self, dual_header_workbook):
        """Update with dual-header (header_row=2)."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.upsert_row(
            dual_header_workbook, "装备配置",
            key_column="equip_id", key_value=2001,
            updates={"attack": 60},
            header_row=2
        )
        assert result['success'] is True
        assert result['data']['action'] == 'update'

    def test_insert_dual_header(self, dual_header_workbook):
        """Insert with dual-header."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.upsert_row(
            dual_header_workbook, "装备配置",
            key_column="equip_id", key_value=2004,
            updates={"equip_id": 2004, "equip_name": "暗影匕首", "rarity": "史诗", "attack": 180},
            header_row=2
        )
        assert result['success'] is True
        assert result['data']['action'] == 'insert'


class TestUpsertErrors:
    """Tests for upsert_row error cases."""

    def test_nonexistent_sheet(self, workbook):
        """Error when sheet doesn't exist."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.upsert_row(
            workbook, "不存在的表",
            key_column="skill_id", key_value=1001,
            updates={"damage": 100}
        )
        assert result['success'] is False
        assert "不存在" in result.get('message', '')

    def test_nonexistent_key_column(self, workbook):
        """Error when key column doesn't exist."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.upsert_row(
            workbook, "技能配置",
            key_column="不存在的列", key_value=1001,
            updates={"damage": 100}
        )
        assert result['success'] is False
        assert "不存在" in result.get('message', '')

    def test_empty_updates(self, workbook):
        """Error when updates dict is empty."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.upsert_row(
            workbook, "技能配置",
            key_column="skill_id", key_value=1001,
            updates={}
        )
        assert result['success'] is False

    def test_none_key_value(self, workbook):
        """Error when key_value is None."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.upsert_row(
            workbook, "技能配置",
            key_column="skill_id", key_value=None,
            updates={"damage": 100}
        )
        assert result['success'] is False


class TestBatchInsertRows:
    """Tests for batch_insert_rows functionality."""

    @pytest.fixture
    def batch_workbook(self):
        """Create a test workbook for batch insert tests."""
        wb = Workbook()
        ws = wb.active
        ws.title = "怪物配置"
        ws.append(["monster_id", "monster_name", "level", "hp"])
        ws.append([3001, "哥布林", 5, 100])
        ws.append([3002, "狼人", 10, 300])

        path = os.path.join(os.path.dirname(__file__), 'test_data', 'batch_insert_test.xlsx')
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb.save(path)
        yield path
        if os.path.exists(path):
            os.remove(path)

    def test_batch_insert_basic(self, batch_workbook):
        """Batch insert multiple rows."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        data = [
            {"monster_id": 3003, "monster_name": "巨龙", "level": 50, "hp": 5000},
            {"monster_id": 3004, "monster_name": "骷髅兵", "level": 3, "hp": 50},
        ]
        result = ExcelOperations.batch_insert_rows(batch_workbook, "怪物配置", data)
        assert result['success'] is True
        assert result['data']['action'] == 'batch_insert'
        assert result['data']['inserted_count'] == 2
        assert result['data']['start_row'] == 4

    def test_batch_insert_persists(self, batch_workbook):
        """Batch inserted rows are persisted."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        data = [
            {"monster_id": 3005, "monster_name": "恶魔", "hp": 800},
        ]
        ExcelOperations.batch_insert_rows(batch_workbook, "怪物配置", data)
        from openpyxl import load_workbook
        wb = load_workbook(batch_workbook)
        assert wb["怪物配置"].cell(row=4, column=1).value == 3005
        assert wb["怪物配置"].cell(row=4, column=2).value == "恶魔"
        wb.close()

    def test_batch_insert_single_row(self, batch_workbook):
        """Batch insert with single row."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        data = [{"monster_id": 3006, "monster_name": "精灵"}]
        result = ExcelOperations.batch_insert_rows(batch_workbook, "怪物配置", data)
        assert result['success'] is True
        assert result['data']['inserted_count'] == 1

    def test_batch_insert_empty_data(self, batch_workbook):
        """Error when data is empty."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.batch_insert_rows(batch_workbook, "怪物配置", [])
        assert result['success'] is False

    def test_batch_insert_nonexistent_sheet(self, batch_workbook):
        """Error when sheet doesn't exist."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.batch_insert_rows(
            batch_workbook, "不存在的表",
            [{"monster_id": 1}]
        )
        assert result['success'] is False
        assert "不存在" in result.get('message', '')

    def test_batch_insert_unknown_columns(self, batch_workbook):
        """Unknown columns are ignored and reported."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        data = [
            {"monster_id": 3007, "monster_name": "幽灵", "unknown_col": "ignored"},
        ]
        result = ExcelOperations.batch_insert_rows(batch_workbook, "怪物配置", data)
        assert result['success'] is True
        assert "unknown_col" in result['data']['unknown_columns']
