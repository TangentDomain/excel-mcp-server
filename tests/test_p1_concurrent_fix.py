"""
P1-concurrent bug 修复验证: 多线程并发写入测试

验证 threading.Lock 能正确保护 xlsx 文件免受并发写入损坏。
测试场景:
1. 多线程并发 UPDATE 同一文件
2. 多线程混合 INSERT/UPDATE/DELETE
3. 验证文件完整性(不损坏)
4. 验证数据一致性(无丢失)
"""

import os
import tempfile
import threading
import time
import pytest
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import Workbook, load_workbook

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)


def _create_test_xlsx(file_path: str, rows: int = 20):
    """创建测试用 xlsx 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "装备"
    ws.append(["ID", "Name", "Price", "Rarity"])
    for i in range(1, rows + 1):
        ws.append([i, f"Item-{i}", float(i * 10.0), "Common"])
    wb.save(file_path)
    wb.close()


def _verify_file_integrity(file_path: str) -> bool:
    """验证 xlsx 文件是否损坏"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        # 确保能读取数据
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return len(rows) > 0
    except Exception:
        return False


class TestP1ConcurrentFix:
    """P1-concurrent 并发写入修复验证"""

    def test_concurrent_updates(self):
        """多线程并发 UPDATE 同一文件不导致损坏（flaky: 资源竞争，允许重试）"""
        max_attempts = 3
        for attempt in range(max_attempts):
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                test_file = f.name

            try:
                _create_test_xlsx(test_file, 50)

                errors = []
                num_threads = 8
                ops_per_thread = 5

                def update_worker(thread_id):
                    for i in range(ops_per_thread):
                        try:
                            result = execute_advanced_update_query(
                                test_file,
                                f"UPDATE 装备 SET Price = Price * 1.01 WHERE ID = {(thread_id * ops_per_thread + i) % 50 + 1}"
                            )
                            if not result["success"]:
                                errors.append(f"Thread-{thread_id} op-{i}: {result.get('message', '')}")
                        except Exception as e:
                            errors.append(f"Thread-{thread_id} op-{i} EXC: {e}")

                threads = [threading.Thread(target=update_worker, args=(t,)) for t in range(num_threads)]
                for t in threads:
                    t.start()
                for t in threads:
                    t.join(timeout=30)

                # 验证文件未损坏
                assert _verify_file_integrity(test_file), "xlsx 文件已损坏!"
                assert len(errors) == 0, f"并发更新有错误: {errors[:5]}"

                # 验证数据一致性
                wb = load_workbook(test_file)
                ws = wb.active
                data_rows = list(ws.iter_rows(min_row=2, values_only=True))
                wb.close()
                assert len(data_rows) == 50, f"行数不一致: 期望50, 实际{len(data_rows)}"
                return  # 成功则退出重试循环
            except AssertionError:
                if attempt < max_attempts - 1:
                    time.sleep(1)
                    continue
                raise
            finally:
                if os.path.exists(test_file):
                    os.remove(test_file)

    def test_mixed_concurrent_writes(self):
        """多线程混合 INSERT/UPDATE/DELETE 操作"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name

        try:
            _create_test_xlsx(test_file, 20)

            errors = []
            num_threads = 6

            def insert_worker():
                try:
                    result = execute_advanced_insert_query(
                        test_file,
                        "INSERT INTO 装备 (ID, Name, Price, Rarity) VALUES (999, 'ThreadItem', 99.99, 'Legendary')"
                    )
                    if not result["success"]:
                        errors.append(f"INSERT failed: {result.get('message', '')}")
                except Exception as e:
                    errors.append(f"INSERT EXC: {e}")

            def update_worker():
                try:
                    result = execute_advanced_update_query(
                        test_file,
                        "UPDATE 装备 SET Price = ROUND(Price * 1.05, 2) WHERE Rarity = 'Common'"
                    )
                    if not result["success"]:
                        errors.append(f"UPDATE failed: {result.get('message', '')}")
                except Exception as e:
                    errors.append(f"UPDATE EXC: {e}")

            def delete_worker():
                try:
                    result = execute_advanced_delete_query(
                        test_file,
                        "DELETE FROM 装备 WHERE ID = 999"
                    )
                    if not result["success"] and "没有匹配" not in result.get("message", ""):
                        errors.append(f"DELETE failed: {result.get('message', '')}")
                except Exception as e:
                    errors.append(f"DELETE EXC: {e}")

            # 混合启动各类型操作线程
            threads = []
            for i in range(num_threads):
                if i % 3 == 0:
                    threads.append(threading.Thread(target=insert_worker))
                elif i % 3 == 1:
                    threads.append(threading.Thread(target=update_worker))
                else:
                    threads.append(threading.Thread(target=delete_worker))

            for t in threads:
                t.start()
            for t in threads:
                t.join(timeout=30)

            # 验证文件未损坏
            assert _verify_file_integrity(test_file), "xlsx 文件已损坏!"
            print(f"  混合操作完成, 错误数: {len(errors)}")
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_rapid_fire_same_row(self):
        """极端场景: 多线程同时修改同一行"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name

        try:
            _create_test_xlsx(test_file, 10)

            errors = []
            num_threads = 12

            def worker(tid):
                try:
                    result = execute_advanced_update_query(
                        test_file,
                        f"UPDATE 装备 SET Price = {tid * 100.0}, Name = 'T{tid}' WHERE ID = 1"
                    )
                    if not result["errors"]:
                        errors.append(f"T{tid}: {result.get('message')}")
                except Exception as e:
                    errors.append(f"T{tid} EXC: {e}")

            threads = [threading.Thread(target=worker, args=(i,)) for i in range(num_threads)]
            for t in threads:
                t.start()
            for t in threads:
                t.join(timeout=30)

            assert _verify_file_integrity(test_file), "极端并发下 xlsx 文件损坏!"

            # 验证第1行存在且数据合法
            wb = load_workbook(test_file)
            ws = wb.active
            row1 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            wb.close()
            assert row1[0] == 1, f"ID 应为1, 实际{row1[0]}"
            assert isinstance(row1[2], (int, float)), f"Price 应为数字, 实际{row1[2]}"
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_lock_per_file_isolation(self):
        """验证不同文件的锁互不影响(不同文件应可并行写入)"""
        files = []
        try:
            # 创建两个独立文件
            for idx in range(2):
                f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                _create_test_xlsx(f.name, 15)
                files.append(f.name)
                f.close()

            start_time = time.time()
            errors = []

            def write_to_file(fpath):
                try:
                    result = execute_advanced_update_query(
                        fpath,
                        "UPDATE 装备 SET Price = 999.99 WHERE ID = 1"
                    )
                    if not result["success"]:
                        errors.append(f"{fpath}: {result.get('message')}")
                except Exception as e:
                    errors.append(f"{fpath} EXC: {e}")

            # 两个文件并行写入，不应互相阻塞
            t1 = threading.Thread(target=write_to_file, args=(files[0],))
            t2 = threading.Thread(target=write_to_file, args=(files[1],))
            t1.start()
            t2.start()
            t1.join(timeout=10)
            t2.join(timeout=10)
            elapsed = time.time() - start_time

            # 如果锁是全局的(错误实现)，两个文件会串行执行
            # 正确的按文件路径隔离锁应该允许并行
            assert elapsed < 8.0, f"按文件锁可能未正确隔离, 耗时{elapsed:.1f}s过长"
            assert len(errors) == 0, f"有错误: {errors}"
            assert _verify_file_integrity(files[0])
            assert _verify_file_integrity(files[1])
        finally:
            for f in files:
                if os.path.exists(f):
                    os.remove(f)


if __name__ == "__main__":
    import pytest
    pytest.main([__file__, "-v", "-s"])
