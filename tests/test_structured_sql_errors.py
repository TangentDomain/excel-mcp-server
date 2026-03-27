"""测试结构化SQL错误 — REQ-025

验证StructuredSQLError为AI提供可自动修复的错误信息：
- error_code: 机器可读错误分类
- hint: AI修复建议
- context: 可用列/表等上下文
"""
import pytest
from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    StructuredSQLError,
    _classify_value_error,
    _generate_value_error_hint,
)


class TestStructuredSQLError:
    """StructuredSQLError异常类测试"""

    def test_basic_creation(self):
        err = StructuredSQLError("column_not_found", "列不存在", "请检查拼写")
        assert err.error_code == "column_not_found"
        assert err.message == "列不存在"
        assert err.hint == "请检查拼写"
        assert err.context == {}
        assert str(err) == "列不存在"

    def test_with_context(self):
        err = StructuredSQLError(
            "column_not_found",
            "列 'xxx' 不存在",
            hint="用excel_get_headers查看",
            context={"available_columns": ["id", "name"]}
        )
        assert err.context == {"available_columns": ["id", "name"]}

    def test_is_exception(self):
        err = StructuredSQLError("test", "test")
        assert isinstance(err, Exception)


class TestClassifyValueError:
    """ValueError分类测试"""

    def test_column_not_found(self):
        assert _classify_value_error("列 'xxx' 不存在") == "column_not_found"
        assert _classify_value_error("COLUMN not found") == "column_not_found"

    def test_table_not_found(self):
        assert _classify_value_error("表 'xxx' 不存在") == "table_not_found"
        assert _classify_value_error("TABLE not found") == "table_not_found"

    def test_unsupported(self):
        assert _classify_value_error("不支持的功能") == "unsupported_feature"

    def test_window_function(self):
        """窗口函数相关错误优先匹配为window_function_error"""
        assert _classify_value_error("不支持的窗口函数: XXX") == "window_function_error"

    def test_join(self):
        # "JOIN表 'xxx'" contains "表 '" so it matches table_not_found first (correct)
        assert _classify_value_error("JOIN表 'xxx' 不存在") == "table_not_found"
        assert _classify_value_error("JOIN缺少ON条件") == "join_error"

    def test_union(self):
        assert _classify_value_error("UNION 查询中未找到有效的 SELECT") == "union_error"

    def test_cte(self):
        assert _classify_value_error("CTE 'xxx' 执行失败") == "cte_error"

    def test_expression(self):
        assert _classify_value_error("不支持的表达式") == "unsupported_feature"

    def test_generic(self):
        assert _classify_value_error("未知错误") == "execution_error"


class TestGenerateValueErrorHint:
    """ValueError智能提示生成测试"""

    def test_column_not_found_hint(self):
        hint = _generate_value_error_hint("列 'xxx' 不存在。可用列: ['id', 'name']")
        assert "excel_get_headers" in hint

    def test_table_not_found_hint(self):
        hint = _generate_value_error_hint("表 'xxx' 不存在。可用表: ['a', 'b']")
        assert "excel_list_sheets" in hint

    def test_from_subquery_hint(self):
        hint = _generate_value_error_hint("不支持FROM子查询")
        assert "WHERE col IN" in hint

    def test_join_table_hint(self):
        hint = _generate_value_error_hint("JOIN表 'xxx' 不存在")
        assert "excel_list_sheets" in hint

    def test_join_no_on_hint(self):
        hint = _generate_value_error_hint("JOIN缺少ON条件")
        assert "ON" in hint

    def test_join_column_hint(self):
        hint = _generate_value_error_hint("左表 'a' 没有列 'xxx'")
        assert "ON条件" in hint

    def test_window_function_hint(self):
        hint = _generate_value_error_hint("不支持的窗口函数: NTILE")
        assert "ROW_NUMBER" in hint

    def test_string_function_hint(self):
        hint = _generate_value_error_hint("不支持的字符串函数: FORMAT")
        assert "UPPER" in hint

    def test_generic_no_hint(self):
        hint = _generate_value_error_hint("某个随机错误")
        assert hint == ""


class TestStructuredErrorInSQLQuery:
    """SQL查询中结构化错误集成测试"""

    def test_column_not_found_returns_structured(self, tmp_path):
        """列不存在的错误返回结构化响应"""
        import openpyxl
        from excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "test"
        ws.append(["ID", "Name"])
        ws.append([1, "Alice"])
        f = tmp_path / "test.xlsx"
        wb.save(f)

        result = execute_advanced_sql_query(str(f), "SELECT xxx FROM test")
        assert result['success'] is False
        qi = result['query_info']
        assert qi['error_type'] == 'column_not_found'
        assert 'hint' in qi or 'details' in qi

    def test_table_not_found_returns_structured(self, tmp_path):
        """表不存在的错误返回结构化响应"""
        import openpyxl
        from excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["ID"])
        f = tmp_path / "test.xlsx"
        wb.save(f)

        result = execute_advanced_sql_query(str(f), "SELECT * FROM nonexistent")
        assert result['success'] is False
        qi = result['query_info']
        assert qi['error_type'] == 'table_not_found'
        assert 'hint' in qi

    def test_syntax_error_returns_hint(self, tmp_path):
        """SQL语法错误返回hint字段"""
        import openpyxl
        from excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "test"
        ws.append(["ID"])
        f = tmp_path / "test.xlsx"
        wb.save(f)

        result = execute_advanced_sql_query(str(f), "SELEC * FROM test")
        assert result['success'] is False
        qi = result['query_info']
        assert qi['error_type'] == 'syntax_error'
        assert 'hint' in qi

    def test_unsupported_sql_returns_structured(self, tmp_path):
        """不支持的SQL功能返回结构化响应"""
        import openpyxl
        from excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "test"
        ws.append(["ID"])
        f = tmp_path / "test.xlsx"
        wb.save(f)

        result = execute_advanced_sql_query(str(f), "INSERT INTO test VALUES (1)")
        assert result['success'] is False
        qi = result['query_info']
        assert qi['error_type'] in ('unsupported_sql', 'unsupported_feature')
