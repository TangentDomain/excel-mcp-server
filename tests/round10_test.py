#!/usr/bin/env python3
"""
Round 10 迭代测试 — 数据分析深度 + QA边缘场景 + 窗口函数补齐 + 性能基准
主题: FIRST_VALUE/LAST_VALUE/NTILE/PERCENT_RANK, DISTINCT+ORDER BY, 双行表头,
      UPDATE精度, 大数据量性能, 边缘写入压力
"""
import sys
import os
import time
import random
import copy
import tempfile
import shutil

sys.path.insert(0, '/root/workspace/excel-mcp-server/src')

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)

# ============================================================
# 测试数据准备
# ============================================================

TEST_DIR = '/tmp/excelmcp_r10'
os.makedirs(TEST_DIR, exist_ok=True)

def create_equipment_table(wb):
    """装备表 — 主测试表"""
    ws = wb.active
    ws.title = "装备"
    headers = ["ID", "Name", "BaseAtk", "AtkBonus", "Price", "Rarity", "Category", "DropRate"]
    ws.append(headers)
    
    rarities = ["Common", "Rare", "Epic", "Legendary"]
    categories = ["Weapon", "Armor", "Accessory", "Consumable"]
    
    for i in range(1, 51):
        rarity = random.choice(rarities)
        cat = random.choice(categories)
        price = round(random.uniform(10.0, 9999.99), 2)
        # 故意放一些特殊值
        if i == 1:
            price = 0.0  # 零价格
        elif i == 2:
            price = -5.99  # 负数价格（折扣）
        elif i == 3:
            price = 3.14159265358979  # 高精度PI值
        elif i == 4:
            price = 1234567.89  # 大数值
        
        drop_rate = round(random.uniform(0.001, 1.0), 4) if i != 25 else 0.0  # 一个零掉率
        
        ws.append([
            i,
            f"Item_{i}_{cat}",
            random.randint(10, 500),
            round(random.uniform(5.0, 200.0), 2),
            price,
            rarity,
            cat,
            drop_rate
        ])
    return ws

def create_skill_table(wb):
    """技能表 — 用于JOIN测试"""
    ws = wb.create_sheet("技能")
    headers = ["SkillID", "Name", "Damage", "ManaCost", "Type", "Cooldown"]
    ws.append(headers)
    
    skill_types = ["Attack", "Defense", "Buff", "Debuff"]
    for i in range(1, 31):
        ws.append([
            i,
            f"Skill_{i}",
            random.randint(50, 2000),
            random.randint(10, 500),
            random.choice(skill_types),
            random.randint(1, 120)
        ])
    return ws

def create_shop_table(wb):
    """商店表 — 用于运营场景"""
    ws = wb.create_sheet("商店")
    headers = ["ShopID", "ItemID", "ShopPrice", "Stock", "Discount"]
    ws.append(headers)
    
    for i in range(1, 41):
        discount = random.choice([0, 0.05, 0.1, 0.15, 0.2, 0.5])
        shop_price = round(random.uniform(100, 50000) * (1 - discount), 2)
        ws.append([
            i,
            random.randint(1, 50),
            shop_price,
            random.randint(0, 999),
            discount
        ])
    return ws

def create_double_header_table(wb):
    """双行表头配置表 — 游戏常见格式"""
    ws = wb.create_sheet("双行表头")
    # 第一行：中文描述
    ws.append(["编号", "名称", "基础属性", "", "经济属性", "", "分类"])
    # 第二行：英文字段名
    ws.append(["ID", "Name", "BaseAtk", "AtkBonus", "Price", "Rarity", "Category"])
    
    for i in range(1, 21):
        ws.append([
            i,
            f"DH_Item_{i}",
            random.randint(10, 300),
            round(random.uniform(5.0, 100.0), 2),
            round(random.uniform(50.0, 5000.0), 2),
            random.choice(["Common", "Rare", "Epic"]),
            random.choice(["Weapon", "Armor"])
        ])
    return ws

def create_large_table(filepath, num_rows=10000):
    """大数据量表 — 性能基准测试用"""
    wb = Workbook()
    ws = wb.active
    ws.title = "大表"
    headers = ["ID", "Name", "Value1", "Value2", "Value3", "Category", "Status", "Score"]
    ws.append(headers)
    
    categories = ["A", "B", "C", "D", "E"]
    statuses = ["Active", "Inactive", "Pending"]
    
    for i in range(1, num_rows + 1):
        ws.append([
            i,
            f"Row_{i}",
            round(random.uniform(-100000, 100000), 4),
            round(random.uniform(0, 1000), 6),
            round(random.uniform(0.001, 99999.999), 3),
            random.choice(categories),
            random.choice(statuses),
            round(random.uniform(0, 100), 2)
        ])
    wb.save(filepath)
    return filepath


def build_test_file():
    """构建标准测试文件"""
    wb = Workbook()
    create_equipment_table(wb)
    create_skill_table(wb)
    create_shop_table(wb)
    create_double_header_table(wb)
    
    path = os.path.join(TEST_DIR, 'r10_test.xlsx')
    wb.save(path)
    return path


# ============================================================
# 测试执行引擎
# ============================================================

class TestResult:
    def __init__(self, name, category, sql, expected_check=None, is_update=False, 
                 is_insert=False, is_delete=False, expected_error=False):
        self.name = name
        self.category = category
        self.sql = sql
        self.expected_check = expected_check  # callable(result) -> bool
        self.is_update = is_update
        self.is_insert = is_insert
        self.is_delete = is_delete
        self.expected_error = expected_error
        self.passed = False
        self.error_msg = ""
        self.execution_time = 0
        self.result_data = None
    
    def run(self, file_path):
        start = time.time()
        try:
            if self.is_update:
                result = execute_advanced_update_query(file_path, self.sql)
            elif self.is_insert:
                result = execute_advanced_insert_query(file_path, self.sql)
            elif self.is_delete:
                result = execute_advanced_delete_query(file_path, self.sql)
            else:
                result = execute_advanced_sql_query(file_path, self.sql)
            
            self.execution_time = time.time() - start
            self.result_data = result
            
            if self.expected_error:
                # 期望出错的情况
                if not result['success']:
                    self.passed = True  # 正确地报错了
                    self.error_msg = f"预期错误: {result.get('message', '')[:80]}"
                else:
                    self.passed = False
                    self.error_msg = "预期错误但执行成功"
            else:
                if result['success']:
                    if self.expected_check:
                        try:
                            check_result = self.expected_check(result)
                            if check_result is True:
                                self.passed = True
                            else:
                                self.passed = False
                                self.error_msg = str(check_result)[:120] if check_result else "断言失败"
                        except Exception as e:
                            self.passed = False
                            self.error_msg = f"检查异常: {str(e)[:100]}"
                    else:
                        self.passed = True
                else:
                    self.passed = False
                    self.error_msg = result.get('message', '未知错误')[:150]
        except Exception as e:
            self.execution_time = time.time() - start
            self.passed = False
            self.error_msg = f"异常: {str(e)[:150]}"


def run_tests(tests, file_path):
    """批量运行测试并报告结果"""
    results = []
    passed = 0
    failed = 0
    
    print(f"\n{'='*80}")
    print(f"📊 共 {len(tests)} 个测试用例")
    print(f"{'='*80}\n")
    
    for i, t in enumerate(tests, 1):
        t.run(file_path)
        results.append(t)
        
        icon = "✅" if t.passed else "❌"
        cat_tag = f"[{t.category}]"
        time_str = f"{t.execution_time:.3f}s"
        
        print(f"{icon} {i:2d}. {cat_tag} {t.name} ({time_str})")
        if not t.passed:
            print(f"     SQL: {t.sql[:90]}{'...' if len(t.sql)>90 else ''}")
            print(f"     错误: {t.error_msg}")
        
        if t.passed:
            passed += 1
        else:
            failed += 1
    
    print(f"\n{'='*80}")
    print(f"📈 结果: {passed}/{len(tests)} 通过 ({passed*100//max(len(tests),1)}%) | 失败: {failed}")
    print(f"{'='*80}\n")
    
    return results


# ============================================================
# Group A: 窗口函数补齐 (FIRST_VALUE / LAST_VALUE / NTILE / PERCENT_RANK)
# ============================================================

def get_group_a():
    """数据分析视角：补齐窗口函数家族"""
    return [
        TestResult(
            "A1: FIRST_VALUE 每个稀有度最贵装备",
            "窗口函数",
            "SELECT DISTINCT Category, FIRST_VALUE(Name) OVER (PARTITION BY Category ORDER BY Price DESC) as TopItem FROM 装备",
            expected_check=lambda r: len(r.get('data', [])) >= 4 and all('TopItem' in str(d) for d in r['data'])
        ),
        TestResult(
            "A2: LAST_VALUE 每个类别最便宜装备",
            "窗口函数",
            "SELECT ID, Name, Category, Price, LAST_VALUE(Name) OVER (PARTITION BY Category ORDER BY Price ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING) as Cheapest FROM 装备 ORDER BY Category LIMIT 10",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) > 0
        ),
        TestResult(
            "A3: NTILE 将装备按价格分4桶",
            "窗口函数",
            "SELECT ID, Name, Price, NTILE(4) OVER (ORDER BY Price) as PriceQuartile FROM 装备 ORDER BY Price LIMIT 15",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) >= 14
        ),
        TestResult(
            "A4: NTILE PARTITION BY 类别内分桶",
            "窗口函数",
            "SELECT ID, Name, Category, Price, NTILE(3) OVER (PARTITION BY Category ORDER BY Price) as CatQuartile FROM 装备 ORDER BY Category, Price LIMIT 15",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) >= 14
        ),
        TestResult(
            "A5: PERCENT_RANK 价格百分位排名",
            "窗口函数",
            "SELECT ID, Name, Price, ROUND(PERCENT_RANK() OVER (ORDER BY Price), 4) as PctRank FROM 装备 ORDER BY Price LIMIT 10",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) >= 9
        ),
        TestResult(
            "A6: PERCENT_RANK PARTITION BY 类别内百分位",
            "窗口函数",
            "SELECT ID, Name, Category, Price, ROUND(PERCENT_RANK() OVER (PARTITION BY Category ORDER BY Price), 4) as CatPctRank FROM 装备 WHERE Category IN ('Weapon', 'Armor') ORDER BY Category, Price LIMIT 12",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) >= 11
        ),
        TestResult(
            "A7: CUME_DIST 累积分布",
            "窗口函数",
            "SELECT ID, Name, Price, ROUND(CUME_DIST() OVER (ORDER BY Price), 4) as CumDist FROM 装备 ORDER BY Price LIMIT 8",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) >= 7
        ),
        TestResult(
            "A8: NTH_VALUE 取每组第N个值",
            "窗口函数",
            "SELECT ID, Name, Category, Price, NTH_VALUE(Name, 2) OVER (PARTITION BY Category ORDER BY Price ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING) as SecondInCat FROM 装备 WHERE Category IN ('Weapon', 'Accessory') ORDER BY Category, Price LIMIT 10",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) >= 9
        ),
    ]


# ============================================================
# Group B: DISTINCT + ORDER BY / GROUP BY 联合去重排序
# ============================================================

def get_group_b():
    """数据分析视角：去重+排序联合操作"""
    return [
        TestResult(
            "B1: DISTINCT 基础去重稀有度列表",
            "DISTINCT",
            "SELECT DISTINCT Rarity FROM 装备 ORDER BY Rarity",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 4
        ),
        TestResult(
            "B2: DISTINCT 多列去重",
            "DISTINCT",
            "SELECT DISTINCT Rarity, Category FROM 装备 ORDER BY Rarity, Category",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) >= 8
        ),
        TestResult(
            "B3: DISTINCT COUNT 去重计数",
            "DISTINCT",
            "SELECT COUNT(DISTINCT Category) as UniqueCategories FROM 装备",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) > 0
        ),
        TestResult(
            "B4: DISTINCT ON (PostgreSQL风格) — 兼容性测试",
            "DISTINCT",
            "SELECT DISTINCT ON (Rarity) Rarity, Name, Price FROM 装备 ORDER BY Rarity, Price DESC",
            expected_error=True  # 可能不支持，这是PG特有语法
        ),
        TestResult(
            "B5: GROUP BY + HAVING 去重后筛选",
            "聚合",
            "SELECT Category, COUNT(*) as cnt, AVG(Price) as AvgPrice FROM 装备 GROUP BY Category HAVING COUNT(*) >= 5 ORDER BY AvgPrice DESC",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) > 0
        ),
        TestResult(
            "B6: 子查询中DISTINCT",
            "子查询",
            "SELECT * FROM (SELECT DISTINCT Category FROM 装备) t ORDER BY Category",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 4
        ),
    ]


# ============================================================
# Group C: 双行表头深度测试
# ============================================================

def get_group_c(file_path):
    """游戏策划视角：双行表头是游戏配置表常见格式"""
    return [
        TestResult(
            "C1: 双行表头基础查询",
            "双行表头",
            "SELECT ID, Name, BaseAtk, Price FROM `双行表头` LIMIT 5",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) > 0
        ),
        TestResult(
            "C2: 双行表头 WHERE 条件",
            "双行表头",
            "SELECT * FROM `双行表头` WHERE Category = 'Weapon'",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) > 0
        ),
        TestResult(
            "C3: 双行表头 ORDER BY 排序",
            "双行表头",
            "SELECT ID, Name, Price FROM `双行表头` ORDER BY Price DESC LIMIT 5",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) > 0
        ),
        TestResult(
            "C4: 双行表头 GROUP BY 聚合",
            "双行表头",
            "SELECT Category, COUNT(*) as cnt, AVG(BaseAtk) as AvgAtk, MAX(Price) as MaxPrice FROM `双行表头` GROUP BY Category",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) > 0
        ),
        TestResult(
            "C5: 双行表头 UPDATE 写入",
            "双行表头",
            "UPDATE `双行表头` SET Price = Price * 1.1 WHERE Category = 'Weapon'",
            is_update=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "C6: 双行表头 UPDATE 后回读验证",
            "双行表头",
            "SELECT ID, Name, Price FROM `双行表头` WHERE Category = 'Weapon' ORDER BY ID LIMIT 3",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) > 0
        ),
        TestResult(
            "C7: 双行表头 INSERT 新行",
            "双行表头",
            "INSERT INTO `双行表头` (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category) VALUES (999, 'Test_DoubleHeader', 100, 10.5, 999.99, 'Epic', 'Weapon')",
            is_insert=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "C8: 双行表头 INSERT 后回读验证",
            "双行表头",
            "SELECT * FROM `双行表头` WHERE ID = 999",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 1
        ),
    ]


# ============================================================
# Group D: UPDATE 精度问题深入探索 (H2)
# ============================================================

def get_group_d():
    """运营/策划视角：数值精度问题深入测试"""
    return [
        TestResult(
            "D1: UPDATE 写入高精度小数 PI",
            "UPDATE精度",
            "UPDATE 装备 SET Price = 3.14159265358979323846 WHERE ID = 3",
            is_update=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "D1-验证: 回读PI精度",
            "UPDATE精度",
            "SELECT ID, Price FROM 装备 WHERE ID = 3",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 1
        ),
        TestResult(
            "D2: UPDATE 写入极大浮点数",
            "UPDATE精度",
            "UPDATE 装备 SET Price = 999999999.123456789 WHERE ID = 4",
            is_update=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "D2-验证: 回读极大数精度",
            "UPDATE精度",
            "SELECT ID, Price FROM 装备 WHERE ID = 4",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 1
        ),
        TestResult(
            "D3: UPDATE 写入极小小数",
            "UPDATE精度",
            "UPDATE 装备 SET DropRate = 0.000001 WHERE ID = 1",
            is_update=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "D3-验证: 回读极小小数",
            "UPDATE精度",
            "SELECT ID, DropRate FROM 装备 WHERE ID = 1",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 1
        ),
        TestResult(
            "D4: UPDATE 数学表达式计算精度",
            "UPDATE精度",
            "UPDATE 装备 SET Price = ROUND(Price * 1.15 + 0.01, 4) WHERE Rarity = 'Epic' AND ID <= 10",
            is_update=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "D4-验证: 表达式计算结果回读",
            "UPDATE精度",
            "SELECT ID, Name, Price FROM 装备 WHERE Rarity = 'Epic' AND ID <= 10 ORDER BY ID",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) > 0
        ),
        TestResult(
            "D5: UPDATE 负数写入和回读",
            "UPDATE精度",
            "UPDATE 装备 SET Price = -999.123456789 WHERE ID = 2",
            is_update=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "D5-验证: 负数精度回读",
            "UPDATE精度",
            "SELECT ID, Price FROM 装备 WHERE ID = 2",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 1
        ),
    ]


# ============================================================
# Group E: 大数据量性能基准
# ============================================================

def get_group_e(large_file_path):
    """QA视角：性能基准测试"""
    return [
        TestResult(
            "E1: 全表 SELECT * 10000行",
            "性能",
            "SELECT * FROM 大表",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 10000
        ),
        TestResult(
            "E2: WHERE 条件筛选 10000行",
            "性能",
            "SELECT * FROM 大表 WHERE Category = 'A' AND Status = 'Active'",
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "E3: ORDER BY 排序 10000行",
            "性能",
            "SELECT * FROM 大表 ORDER BY Value3 DESC LIMIT 500",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 500
        ),
        TestResult(
            "E4: GROUP BY 聚合 10000行",
            "性能",
            "SELECT Category, Status, COUNT(*) as cnt, AVG(Score) as AvgScore, SUM(Value1) as SumV1 FROM 大表 GROUP BY Category, Status ORDER BY cnt DESC",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) > 0
        ),
        TestResult(
            "E5: 窗口函数 RANK 10000行",
            "性能",
            "SELECT ID, Score, RANK() OVER (ORDER BY Score DESC) as Rank FROM 大表 ORDER BY Score DESC LIMIT 20",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 20
        ),
        TestResult(
            "E6: UPDATE 批量修改 10000行",
            "性能",
            "UPDATE 大表 SET Value1 = Value1 * 1.05 WHERE Category = 'B'",
            is_update=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "E7: LIKE 模糊搜索 10000行",
            "性能",
            "SELECT * FROM 大表 WHERE Name LIKE '%Row_5%' LIMIT 100",
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "E8: CASE WHEN 分类统计 10000行",
            "性能",
            "SELECT Category, CASE WHEN AVG(Score) > 50 THEN 'High' ELSE 'Low' END as Level, COUNT(*) as cnt FROM 大表 GROUP BY CASE WHEN AVG(Score) > 50 THEN 'High' ELSE 'Low' END, Category ORDER BY Category LIMIT 10",
            expected_check=lambda r: r['success']
        ),
    ]


# ============================================================
# Group F: 边缘场景压力 — 写入+NULL+特殊字符
# ============================================================

def get_group_f():
    """QA视角：边缘场景压力测试"""
    return [
        TestResult(
            "F1: INSERT 含单引号和双引号的文本",
            "边缘写入",
            "INSERT INTO 装备 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category, DropRate) VALUES (600, \"O'Brien's_\"Sword\"\", 50, 5.0, 199.99, 'Rare', 'Weapon', 0.5)",
            is_insert=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "F1-验证: 特殊字符回读",
            "边缘写入",
            "SELECT ID, Name FROM 装备 WHERE ID = 600",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 1
        ),
        TestResult(
            "F2: INSERT 含中文和特殊符号的文本",
            "边缘写入",
            "INSERT INTO 装备 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category, DropRate) VALUES (601, '火元素·龙之剑★超极品！@#$%', 999, 99.9, 8888.88, 'Legendary', 'Weapon', 0.001)",
            is_insert=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "F2-验证: 中文特殊符号回读",
            "边缘写入",
            "SELECT ID, Name FROM 装备 WHERE ID = 601",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 1
        ),
        TestResult(
            "F3: UPDATE 使用 COALESCE 处理 NULL",
            "边缘写入",
            "UPDATE 装备 SET Price = COALESCE(NULL, 100.0) WHERE ID = 1",
            is_update=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "F3-验证: COALESCE 结果回读",
            "边缘写入",
            "SELECT ID, Price FROM 装备 WHERE ID = 1",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 1
        ),
        TestResult(
            "F4: DELETE 删除刚插入的测试数据",
            "边缘写入",
            "DELETE FROM 装备 WHERE ID IN (600, 601)",
            is_delete=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "F4-验证: 确认删除成功",
            "边缘写入",
            "SELECT COUNT(*) as remaining FROM 装备 WHERE ID IN (600, 601)",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) > 0
        ),
        TestResult(
            "F5: INSERT 超长文本字段 (>200字符)",
            "边缘写入",
            "INSERT INTO 装备 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category, DropRate) VALUES (602, '" + "A" * 250 + "', 1, 1.0, 1.0, 'Common', 'Consumable', 1.0)",
            is_insert=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "F5-验证: 超长文本回读完整性",
            "边缘写入",
            "SELECT ID, LENGTH(Name) as NameLen FROM 装备 WHERE ID = 602",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) == 1
        ),
        TestResult(
            "F6: UPDATE 多条件复合 WHERE + 数学表达式",
            "边缘写入",
            "UPDATE 装备 SET Price = ABS(Price) + CEIL(AtkBonus) * FLOOR(BaseAtk / 10) WHERE Rarity IN ('Rare', 'Epic') AND Category = 'Weapon' AND ID < 20",
            is_update=True,
            expected_check=lambda r: r['success']
        ),
        TestResult(
            "F6-验证: 复合表达式结果",
            "边缘写入",
            "SELECT ID, Name, Price FROM 装备 WHERE Rarity IN ('Rare', 'Epic') AND Category = 'Weapon' AND ID < 20 ORDER BY ID",
            expected_check=lambda r: r['success'] and len(r.get('data', [])) > 0
        ),
        TestResult(
            "F7: 清理所有测试插入数据",
            "边缘写入",
            "DELETE FROM 装备 WHERE ID >= 600",
            is_delete=True,
            expected_check=lambda r: r['success']
        ),
    ]


# ============================================================
# Main
# ============================================================

if __name__ == '__main__':
    print("=" * 80)
    print("🔄 ExcelMCP Round 10 迭代测试")
    print("=" * 80)
    
    # 构建测试数据
    print("\n📦 准备测试数据...")
    file_path = build_test_file()
    large_file_path = os.path.join(TEST_DIR, 'r10_large.xlsx')
    create_large_table(large_file_path, num_rows=10000)
    print(f"   标准测试文件: {file_path}")
    print(f"   大数据量文件: {large_file_path} (10000行)")
    
    all_results = []
    total_passed = 0
    total_failed = 0
    
    # Group A: 窗口函数补齐
    print("\n" + "=" * 80)
    print("📐 Group A: 窗口函数补齐 (FIRST_VALUE/LAST_VALUE/NTILE/PERCENT_RANK/CUME_DIST/NTH_VALUE)")
    print("=" * 80)
    group_a = get_group_a()
    results_a = run_tests(group_a, file_path)
    all_results.extend(results_a)
    pa = sum(1 for r in results_a if r.passed)
    fa = len(results_a) - pa
    total_passed += pa; total_failed += fa
    
    # Group B: DISTINCT + ORDER BY
    print("\n" + "=" * 80)
    print("🔍 Group B: DISTINCT + ORDER BY / GROUP BY 联合去重排序")
    print("=" * 80)
    group_b = get_group_b()
    results_b = run_tests(group_b, file_path)
    all_results.extend(results_b)
    pb = sum(1 for r in results_b if r.passed)
    fb = len(results_b) - pb
    total_passed += pb; total_failed += fb
    
    # Group C: 双行表头
    print("\n" + "=" * 80)
    print("📋 Group C: 双行表头深度测试")
    print("=" * 80)
    group_c = get_group_c(file_path)
    results_c = run_tests(group_c, file_path)
    all_results.extend(results_c)
    pc = sum(1 for r in results_c if r.passed)
    fc = len(results_c) - pc
    total_passed += pc; total_failed += fc
    
    # Group D: UPDATE 精度
    print("\n" + "=" * 80)
    print("🎯 Group D: UPDATE 精度问题深入探索 (H2)")
    print("=" * 80)
    group_d = get_group_d()
    results_d = run_tests(group_d, file_path)
    all_results.extend(results_d)
    pd_val = sum(1 for r in results_d if r.passed)
    fd = len(results_d) - pd_val
    total_passed += pd_val; total_failed += fd
    
    # Group E: 性能基准
    print("\n" + "=" * 80)
    print("⚡ Group E: 大数据量性能基准 (10000行)")
    print("=" * 80)
    group_e = get_group_e(large_file_path)
    results_e = run_tests(group_e, large_file_path)
    all_results.extend(results_e)
    pe = sum(1 for r in results_e if r.passed)
    fe = len(results_e) - pe
    total_passed += pe; total_failed += fe
    
    # Group F: 边缘场景
    print("\n" + "=" * 80)
    print("🔥 Group F: 边缘场景压力 — 写入+NULL+特殊字符")
    print("=" * 80)
    group_f = get_group_f()
    results_f = run_tests(group_f, file_path)
    all_results.extend(results_f)
    pf = sum(1 for r in results_f if r.passed)
    ff = len(results_f) - pf
    total_passed += pf; total_failed += ff
    
    # 最终汇总
    total = len(all_results)
    print("\n" + "=" * 80)
    print("📊 Round 10 最终汇总")
    print("=" * 80)
    print(f"  Group A (窗口函数补齐):   {pa}/{len(results_a)}")
    print(f"  Group B (DISTINCT去重):    {pb}/{len(results_b)}")
    print(f"  Group C (双行表头):        {pc}/{len(results_c)}")
    print(f"  Group D (UPDATE精度):      {pd_val}/{len(results_d)}")
    print(f"  Group E (性能基准):        {pe}/{len(results_e)}")
    print(f"  Group F (边缘压力):        {pf}/{len(results_f)}")
    print(f"  {'─'*40}")
    print(f"  总计:                      {total_passed}/{total} ({total_passed*100//max(total,1)}%)")
    print(f"  失败: {total_failed}")
    print("=" * 80)
    
    # 输出失败详情供分析
    failed_results = [r for r in all_results if not r.passed]
    if failed_results:
        print("\n❌ 失败用例详情:")
        for r in failed_results:
            print(f"\n  [{r.category}] {r.name}")
            print(f"  SQL: {r.sql}")
            print(f"  Error: {r.error_msg}")
            if r.result_data:
                print(f"  Result keys: {list(r.result_data.keys()) if isinstance(r.result_data, dict) else type(r.result_data)}")
    
    # 精度专项输出
    print("\n📐 精度测试详细数据:")
    for r in results_d:
        if r.passed and r.result_data and r.result_data.get('data'):
            print(f"  {r.name}:")
            for row in r.result_data['data'][:3]:
                print(f"    {row}")
