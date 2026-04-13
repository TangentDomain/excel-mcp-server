"""对比 calamine vs openpyxl 引擎读取极小浮点数"""

import sys

sys.path.insert(0, "/root/workspace/excel-mcp-server/src")
import pandas as pd
from openpyxl import Workbook

# 创建测试文件
wb = Workbook()
ws = wb.active
ws.title = "test"
ws.append(["ID", "Val"])
ws.append([1, 0.000001])
ws.append([2, 8.88e-2])
wb.save("/tmp/r5_float_test.xlsx")

fp = "/tmp/r5_float_test.xlsx"

print("=== openpyxl 直接读 ===")
from openpyxl import load_workbook

wb2 = load_workbook(fp)
ws2 = wb2["test"]
for row in ws2.iter_rows(min_row=2, values_only=True):
    print(f"  Val={row[1]!r} (type={type(row[1]).__name__})")

print("\n=== calamine 引擎 (项目实际使用) ===")
try:
    df_c = pd.read_excel(fp, sheet_name="test", engine="calamine")
    print(f"  Val dtype: {df_c['Val'].dtype}")
    for i, v in enumerate(df_c["Val"].values):
        print(f"  row {i}: Val={v!r}")
except Exception as e:
    print(f"  ERROR: {e}")

print("\n=== openpyxl 引擎 ===")
try:
    df_o = pd.read_excel(fp, sheet_name="test", engine="openpyxl")
    print(f"  Val dtype: {df_o['Val'].dtype}")
    for i, v in enumerate(df_o["Val"].values):
        print(f"  row {i}: Val={v!r}")
except Exception as e:
    print(f"  ERROR: {e}")

print("\n=== 结论: 哪个引擎丢失精度? ===")
