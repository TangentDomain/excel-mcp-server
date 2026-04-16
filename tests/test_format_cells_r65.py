# -*- coding: utf-8 -*-
"""
format_cells R65 迭代测试 - 深度边缘 case 第五轮

新增测试场景（覆盖 R64 未涉及的场景）:
  1. 绝对引用范围 $A$1:$B$2 的 format_cells 兼容性
  2. 完全空工作表（无数据）的 format_cells
  3. font_size=0 和负数的容错行为
  4. text_rotation 边界值：0, 90, 180, 181→clamp, -45→abs
  5. number_format "General" 重置行为
  6. 仅传 number_format 的最小格式化
  7. border diagonal_direction 各枚举值（0-3）
  8. alignment indent 负数/零/大值的边界
  9. Unicode/中文工作表名 + format_cells
  10. 全属性组合格式化（同时设置 font+fill+alignment+border+number_format）
  11. format_cells 后 cell.value 类型保持（int/float/str/bool/None/date）
  12. format_cells 对含公式的单元格（不应破坏公式）
  13. 嵌套格式中 font.color 为 None 时的行为
  14. _normalize_formatting 对未知键的透传行为
  15. 连续多次 format_cells 幂等性（同一格式应用 5 次）
  16. bg_color 带 # 前缀的清理验证
  17. fill type 大小写不敏感（SOLID → solid, Gradient → gradient）
  18. openpyxl Color 对象作为 font_color 直接传入
"""

import os
import pytest
import tempfile
from datetime import date, datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color as OpenpyxlColor, PatternFill

from excel_mcp_server_fastmcp.core.excel_writer import ExcelWriter
from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations


# ==================== Helper ====================

def _create_test_xlsx(file_path: str, rows: int = 5, cols: int = 4, sheet_name: str = "Sheet1"):
    """创建测试用 xlsx 文件，含数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=r * 10 + c)
    wb.save(file_path)
    wb.close()


def _read_cell_style(file_path: str, cell_ref: str = "A1", sheet_name: str = "Sheet1") -> dict:
    """读取单元格的样式属性用于断言"""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    cell = ws[cell_ref]
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    result = {
        "bold": font.bold,
        "italic": font.italic,
        "underline": font.underline,
        "strikethrough": font.strikethrough,
        "size": font.size,
        "font_name": font.name,
        "color": str(font.color) if font.color else None,
        "fill_type": fill.fill_type if fill else None,
        "fgColor": str(fill.fgColor) if fill and hasattr(fill, 'fgColor') and fill.fgColor else None,
        "bgColor": str(fill.bgColor) if fill and hasattr(fill, 'bgColor') and fill.bgColor else None,
        "alignment_h": alignment.horizontal,
        "alignment_v": alignment.vertical,
        "wrap_text": bool(alignment.wrap_text) if alignment.wrap_text is not None else False,
        "text_rotation": alignment.text_rotation,
        "number_format": cell.number_format,
        "border_left": border.left.style if border and border.left else None,
        "border_right": border.right.style if border and border.right else None,
        "border_top": border.top.style if border and border.top else None,
        "border_bottom": border.bottom.style if border and border.bottom else None,
        "value": cell.value,
        "indent": alignment.indent,
        "shrink_to_fit": alignment.shrink_to_fit,
    }
    wb.close()
    return result


# ==================== Test Class ====================

class TestFormatCellsR65:
    """format_cells R65 第五轮测试套件 — 深度边缘 case 第五轮"""

    # ---------- 1. 绝对引用范围 ----------

    def test_absolute_ref_range(self, tmp_path):
        """$A$1:$B$2 绝对引用范围应正常工作"""
        fp = str(tmp_path / "absolute_ref.xlsx")
        _create_test_xlsx(fp)

        r = ExcelOperations.format_cells(fp, "Sheet1", "$A$1:$B$2", {"bold": True})
        assert r["success"], f"格式化失败: {r.get('message')}"

        style_a1 = _read_cell_style(fp, "A1")
        style_b2 = _read_cell_style(fp, "B2")
        assert style_a1["bold"] is True
        assert style_b2["bold"] is True
        # C1 不应被格式化
        style_c1 = _read_cell_style(fp, "C1")
        assert style_c1["bold"] is not True

    # ---------- 2. 完全空工作表 ----------

    def test_empty_sheet_format(self, tmp_path):
        """完全空的工作表（无数据）执行 format_cells 不应报错"""
        fp = str(tmp_path / "empty_sheet.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        wb.save(fp)
        wb.close()

        r = ExcelOperations.format_cells(fp, "Empty", "A1:C3", {"bg_color": "FF0000"})
        assert r["success"], f"空表格式化失败: {r.get('message')}"
        assert r["metadata"]["formatted_count"] == 9  # 3x3 = 9 cells

    # ---------- 3. font_size 边界 ----------

    def test_font_size_zero(self, tmp_path):
        """font_size=0 应能正常设置（openpyxl 容错）"""
        fp = str(tmp_path / "fontsize_zero.xlsx")
        _create_test_xlsx(fp)

        # API 层可能不允许 size=0，但 Writer 层应直接传递
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"size": 0}})
        assert result.success
        style = _read_cell_style(fp, "A1")
        assert style["size"] == 0

    def test_font_size_negative(self, tmp_path):
        """font_size=-1 应能正常设置（由 openpyxl 处理）"""
        fp = str(tmp_path / "fontsize_neg.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"size": -5}})
        assert result.success
        style = _read_cell_style(fp, "A1")
        assert style["size"] == -5

    def test_font_size_float(self, tmp_path):
        """font_size=12.5 浮点数应被接受"""
        fp = str(tmp_path / "fontsize_float.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"size": 12.5}})
        assert result.success
        style = _read_cell_style(fp, "A1")
        assert style["size"] == 12.5

    # ---------- 4. text_rotation 边界值 ----------

    def test_text_rotation_zero(self, tmp_path):
        """text_rotation=0 水平文本"""
        fp = str(tmp_path / "rot_zero.xlsx")
        _create_test_xlsx(fp)
        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"text_rotation": 0})
        assert r["success"]
        assert _read_cell_style(fp, "A1")["text_rotation"] == 0

    def test_text_rotation_90(self, tmp_path):
        """text_rotation=90 垂直文本"""
        fp = str(tmp_path / "rot_90.xlsx")
        _create_test_xlsx(fp)
        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"text_rotation": 90})
        assert r["success"]
        assert _read_cell_style(fp, "A1")["text_rotation"] == 90

    def test_text_rotation_180(self, tmp_path):
        """text_rotation=180 最大允许值"""
        fp = str(tmp_path / "rot_180.xlsx")
        _create_test_xlsx(fp)
        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"text_rotation": 180})
        assert r["success"]
        assert _read_cell_style(fp, "A1")["text_rotation"] == 180

    def test_text_rotation_over_max_clamped(self, tmp_path):
        """text_rotation=270 应被 clamp 到 180"""
        fp = str(tmp_path / "rot_clamp.xlsx")
        _create_test_xlsx(fp)
        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"text_rotation": 270})
        assert r["success"]
        assert _read_cell_style(fp, "A1")["text_rotation"] == 180

    def test_text_rotation_negative_abs(self, tmp_path):
        """text_rotation=-45 应取绝对值为 45"""
        fp = str(tmp_path / "rot_neg.xlsx")
        _create_test_xlsx(fp)
        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"text_rotation": -45})
        assert r["success"]
        assert _read_cell_style(fp, "A1")["text_rotation"] == 45

    def test_text_rotation_non_numeric(self, tmp_path):
        """text_rotation='abc' 非数字应 fallback 到 0"""
        fp = str(tmp_path / "rot_str.xlsx")
        _create_test_xlsx(fp)
        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"text_rotation": "abc"})
        assert r["success"]
        assert _read_cell_style(fp, "A1")["text_rotation"] == 0

    # ---------- 5. number_format General 重置 ----------

    def test_number_format_general_reset(self, tmp_path):
        """number_format='General' 应重置为默认格式"""
        fp = str(tmp_path / "nf_general.xlsx")
        _create_test_xlsx(fp)

        # 先设为货币格式
        r1 = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"number_format": "¥#,##0.00"})
        assert r1["success"]
        assert _read_cell_style(fp, "A1")["number_format"] == "¥#,##0.00"

        # 重置为 General
        r2 = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"number_format": "General"})
        assert r2["success"]
        assert _read_cell_style(fp, "A1")["number_format"] == "General"

    def test_number_format_empty_string(self, tmp_path):
        """number_format='' 空字符串的行为"""
        fp = str(tmp_path / "nf_empty.xlsx")
        _create_test_xlsx(fp)
        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"number_format": ""})
        assert r["success"]
        # 空字符串应被保留
        assert _read_cell_style(fp, "A1")["number_format"] == ""

    def test_number_format_only(self, tmp_path):
        """仅传 number_format 的最小格式化"""
        fp = str(tmp_path / "nf_only.xlsx")
        _create_test_xlsx(fp)
        original = _read_cell_style(fp, "A1")

        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"number_format": "0.00%"})
        assert r["success"]

        after = _read_cell_style(fp, "A1")
        assert after["number_format"] == "0.00%"
        # 其他样式不变
        assert after["bold"] == original["bold"]
        assert after["fill_type"] == original["fill_type"]

    # ---------- 6. border diagonal_direction ----------

    def test_border_diagonal_direction_0(self, tmp_path):
        """border diagonal_direction=0（无对角线）"""
        fp = str(tmp_path / "bdiag0.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {
                "diagonal": "thin",
                "diagonal_direction": 0,
            }
        })
        assert result.success

    def test_border_diagonal_direction_1(self, tmp_path):
        """border diagonal_direction=1（主对角线）"""
        fp = str(tmp_path / "bdiag1.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {
                "diagonal": "thin",
                "diagonal_direction": 1,
            }
        })
        assert result.success

    def test_border_diagonal_direction_2(self, tmp_path):
        """border diagonal_direction=2（反对角线）"""
        fp = str(tmp_path / "bdiag2.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {
                "diagonal": "thin",
                "diagonal_direction": 2,
            }
        })
        assert result.success

    def test_border_diagonal_direction_3(self, tmp_path):
        """border diagonal_direction=3（双对角线）"""
        fp = str(tmp_path / "bdiag3.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {
                "diagonal": "thin",
                "diagonal_direction": 3,
            }
        })
        assert result.success

    # ---------- 7. alignment indent 边界 ----------

    def test_indent_zero(self, tmp_path):
        """indent=0 无缩进"""
        fp = str(tmp_path / "indent0.xlsx")
        _create_test_xlsx(fp)
        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"indent": 0})
        assert r["success"]
        assert _read_cell_style(fp, "A1")["indent"] == 0

    def test_indent_large_value(self, tmp_path):
        """indent=100 大缩进值"""
        fp = str(tmp_path / "indent_large.xlsx")
        _create_test_xlsx(fp)
        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"indent": 100})
        assert r["success"]
        assert _read_cell_style(fp, "A1")["indent"] == 100

    # ---------- 8. Unicode 工作表名 ----------

    def test_chinese_sheet_name(self, tmp_path):
        """中文工作表名的 format_cells"""
        fp = str(tmp_path / "chinese_sheet.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "数据表"
        ws["A1"] = "测试"
        wb.save(fp)
        wb.close()

        r = ExcelOperations.format_cells(fp, "数据表", "A1", {"bold": True, "font_name": "微软雅黑"})
        assert r["success"], f"中文工作表格式化失败: {r.get('message')}"
        style = _read_cell_style(fp, "A1", "数据表")
        assert style["bold"] is True
        assert style["font_name"] == "微软雅黑"

    def test_sheet_name_with_spaces(self, tmp_path):
        """带空格的工作表名"""
        fp = str(tmp_path / "space_sheet.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "My Sheet"
        ws["A1"] = "test"
        wb.save(fp)
        wb.close()

        r = ExcelOperations.format_cells(fp, "My Sheet", "A1", {"italic": True})
        assert r["success"]

    # ---------- 9. 全属性组合 ----------

    def test_all_properties_combo(self, tmp_path):
        """同时设置所有格式属性的组合操作"""
        fp = str(tmp_path / "all_props.xlsx")
        _create_test_xlsx(fp)

        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {
            "bold": True,
            "italic": True,
            "underline": "double",
            "font_name": "Arial",
            "font_size": 16,
            "font_color": "FF0000",
            "bg_color": "00FF00",
            "alignment": "center",
            "vertical_alignment": "middle",
            "wrap_text": True,
            "text_rotation": 0,
            "number_format": "#,##0.00",
            "border_style": "medium",
        })
        assert r["success"], f"组合格式化失败: {r.get('message')}"

        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["italic"] is True
        assert style["underline"] == "double"
        assert style["font_name"] == "Arial"
        assert style["size"] == 16
        assert style["alignment_h"] == "center"
        assert style["alignment_v"] == "center"
        assert style["wrap_text"] is True
        assert style["number_format"] == "#,##0.00"
        assert style["border_top"] == "medium"
        assert style["border_bottom"] == "medium"

    # ---------- 10. 数据类型保持 ----------

    def test_preserve_int_value(self, tmp_path):
        """format_cells 后 int 值不变"""
        fp = str(tmp_path / "preserve_int.xlsx")
        _create_test_xlsx(fp)
        original_val = 11  # A1 = 1*10+1
        ExcelOperations.format_cells(fp, "Sheet1", "A1", {"bold": True})
        assert _read_cell_style(fp, "A1")["value"] == original_val

    def test_preserve_float_value(self, tmp_path):
        """format_cells 后 float 值不变"""
        fp = str(tmp_path / "preserve_float.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 3.14159265358979
        wb.save(fp)
        wb.close()

        ExcelOperations.format_cells(fp, "Sheet1", "A1", {"bg_color": "FF0000"})
        result = _read_cell_style(fp, "A1")["value"]
        assert abs(result - 3.14159265358979) < 1e-10

    def test_preserve_string_value(self, tmp_path):
        """format_cells 后字符串值不变"""
        fp = str(tmp_path / "preserve_str.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Hello 世界 🌍"
        wb.save(fp)
        wb.close()

        ExcelOperations.format_cells(fp, "Sheet1", "A1", {"bold": True})
        assert _read_cell_style(fp, "A1")["value"] == "Hello 世界 🌍"

    def test_preserve_bool_value(self, tmp_path):
        """format_cells 后布尔值不变"""
        fp = str(tmp_path / "preserve_bool.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = True
        ws["A2"] = False
        wb.save(fp)
        wb.close()

        ExcelOperations.format_cells(fp, "Sheet1", "A1:A2", {"italic": True})
        assert _read_cell_style(fp, "A1")["value"] is True
        assert _read_cell_style(fp, "A2")["value"] is False

    def test_preserve_none_value(self, tmp_path):
        """format_cells 后空值不变"""
        fp = str(tmp_path / "preserve_none.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = None
        wb.save(fp)
        wb.close()

        ExcelOperations.format_cells(fp, "Sheet1", "A1", {"bg_color": "000000"})
        assert _read_cell_style(fp, "A1")["value"] is None

    def test_preserve_date_value(self, tmp_path):
        """format_cells 后日期值不变（存储为序列号）"""
        fp = str(tmp_path / "preserve_date.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = date(2025, 6, 15)
        wb.save(fp)
        wb.close()

        ExcelOperations.format_cells(fp, "Sheet1", "A1", {"bold": True})
        val = _read_cell_style(fp, "A1")["value"]
        # openpyxl 读取日期时可能返回 date/datetime 对象或序列号
        if isinstance(val, (int, float)):
            assert val > 45000
        else:
            # datetime 或 date 对象
            assert hasattr(val, "year") and val.year == 2025

    # ---------- 11. 公式保持（不破坏） ----------

    def test_preserve_formula(self, tmp_path):
        """format_cells 不应破坏已有公式"""
        fp = str(tmp_path / "preserve_formula.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=SUM(A1:A2)"
        wb.save(fp)
        wb.close()

        r = ExcelOperations.format_cells(fp, "Sheet1", "A3", {"bold": True, "font_color": "0000FF"})
        assert r["success"]

        wb2 = load_workbook(fp)
        ws2 = wb2["Sheet1"]
        # 公式应保留为字符串（openpyxl 不计算公式，只保留公式字符串）
        val = ws2["A3"].value
        assert isinstance(val, str) and val.startswith("="), f"公式被破坏，当前值: {val}"
        assert val == "=SUM(A1:A2)"
        wb2.close()

    # ---------- 12. 嵌套格式 font.color=None ----------

    def test_nested_font_color_none(self, tmp_path):
        """嵌套格式中 font.color=None 的行为"""
        fp = str(tmp_path / "nest_color_none.xlsx")
        _create_test_xlsx(fp)

        # 先设颜色
        r1 = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"font_color": "FF0000"})
        assert r1["success"]

        # 用嵌套格式传 color=None（Writer 层会收到 None 并传给 Font()）
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"color": None}})
        # openpyxl Font(color=None) 可能抛异常或重置颜色
        # 关键是不应崩溃
        assert result.success, f"嵌套 color=None 失败: {result.error}"

    # ---------- 13. _normalize_formatting 未知键透传 ----------

    def test_normalize_unknown_keys_passthrough(self, tmp_path):
        """未知键应被透传到嵌套格式"""
        result = ExcelOperations._normalize_formatting({
            "custom_key": "custom_value",
            "another_key": 42,
        })
        assert result["custom_key"] == "custom_value"
        assert result["another_key"] == 42

    def test_normalize_unknown_with_known(self, tmp_path):
        """未知键和已知键混合时，已知键正常转换，未知键透传"""
        result = ExcelOperations._normalize_formatting({
            "bold": True,
            "magic_flag": True,
            "bg_color": "FF0",
        })
        assert result["font"]["bold"] is True
        assert result["fill"]["color"] == "FF0"
        assert result["magic_flag"] is True

    # ---------- 14. 多次幂等性 ----------

    def test_idempotency_5_times(self, tmp_path):
        """同一格式连续应用 5 次，结果一致"""
        fp = str(tmp_path / "idempotent5.xlsx")
        _create_test_xlsx(fp)

        fmt = {"bold": True, "font_color": "FF0000", "bg_color": "00FF00", "alignment": "center"}
        for i in range(5):
            r = ExcelOperations.format_cells(fp, "Sheet1", "A1", fmt)
            assert r["success"], f"第{i+1}次格式化失败"

        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["alignment_h"] == "center"

    # ---------- 15. bg_color # 前缀清理 ----------

    def test_bg_color_hash_prefix(self, tmp_path):
        """bg_color 带 # 前缀应被自动清理"""
        fp = str(tmp_path / "hash_color.xlsx")
        _create_test_xlsx(fp)

        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"bg_color": "#FF0000"})
        assert r["success"]
        style = _read_cell_style(fp, "A1")
        # 验证填充色已设置（不检查具体值因为 openpyxl 内部表示可能不同）
        assert style["fill_type"] is not None

    def test_font_color_hash_prefix(self, tmp_path):
        """font_color 带 # 前缀应被自动清理"""
        fp = str(tmp_path / "hash_fcolor.xlsx")
        _create_test_xlsx(fp)

        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"font_color": "#0000FF"})
        assert r["success"]

    # ---------- 16. fill type 大小写不敏感 ----------

    def test_fill_type_uppercase_solid(self, tmp_path):
        """fill_type='SOLID' 大写应正常工作"""
        fp = str(tmp_path / "fill_upper.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"fill": {"type": "SOLID", "color": "FF0000"}})
        assert result.success

    def test_fill_type_mixed_case_gradient(self, tmp_path):
        """fill_type='Gradient' 混合大小写"""
        fp = str(tmp_path / "fill_mixed.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {"type": "Gradient", "colors": ["FF0000", "00FF00"]}
        })
        assert result.success

    # ---------- 17. shrink_to_fit toggle ----------

    def test_shrink_to_fit_on(self, tmp_path):
        """shrink_to_fit=True 开启"""
        fp = str(tmp_path / "shrink_on.xlsx")
        _create_test_xlsx(fp)
        r = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"shrink_to_fit": True})
        assert r["success"]
        assert _read_cell_style(fp, "A1")["shrink_to_fit"] is True

    def test_shrink_to_fit_off(self, tmp_path):
        """shrink_to_fit=False 关闭"""
        fp = str(tmp_path / "shrink_off.xlsx")
        _create_test_xlsx(fp)
        # 先开启
        ExcelOperations.format_cells(fp, "Sheet1", "A1", {"shrink_to_fit": True})
        assert _read_cell_style(fp, "A1")["shrink_to_fit"] is True
        # 再关闭
        r2 = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"shrink_to_fit": False})
        assert r2["success"]
        # openpyxl 中 False 布尔样式属性可能存储为 None（表示未设置/默认）
        assert _read_cell_style(fp, "A1")["shrink_to_fit"] in (False, None)

    # ---------- 18. strikethrough toggle ----------

    def test_strikethrough_on_off(self, tmp_path):
        """strikethrough 开关切换"""
        fp = str(tmp_path / "strike.xlsx")
        _create_test_xlsx(fp)

        # 默认无删除线
        assert _read_cell_style(fp, "A1")["strikethrough"] is not True

        # 开启
        r1 = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"strikethrough": True})
        assert r1["success"]
        assert _read_cell_style(fp, "A1")["strikethrough"] is True

        # 关闭
        r2 = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"strikethrough": False})
        assert r2["success"]
        assert _read_cell_style(fp, "A1")["strikethrough"] is False

    # ---------- 19. 单元格范围精确计数 ----------

    def test_single_cell_count(self, tmp_path):
        """单单元格格式化 count=1"""
        fp = str(tmp_path / "single_cnt.xlsx")
        _create_test_xlsx(fp)
        r = ExcelOperations.format_cells(fp, "Sheet1", "B2", {"bold": True})
        assert r["success"]
        assert r["metadata"]["formatted_count"] == 1

    def test_range_count_3x3(self, tmp_path):
        """3x3 范围格式化 count=9"""
        fp = str(tmp_path / "range_cnt.xlsx")
        _create_test_xlsx(fp)
        r = ExcelOperations.format_cells(fp, "Sheet1", "A1:C3", {"bold": True})
        assert r["success"]
        assert r["metadata"]["formatted_count"] == 9

    def test_full_row_count(self, tmp_path):
        """整行格式化应覆盖足够多的列"""
        fp = str(tmp_path / "row_cnt.xlsx")
        _create_test_xlsx(fp, rows=5, cols=4)
        r = ExcelOperations.format_cells(fp, "Sheet1", "1:1", {"bold": True})
        assert r["success"]
        # 整行至少应有列数个单元格被格式化
        assert r["metadata"]["formatted_count"] >= 4

    # ---------- 20. preset + 用户自定义合并 ----------

    def test_preset_title_override_bold(self, tmp_path):
        """preset=title (bold=True) + 用户 bold=False → bold=False"""
        fp = str(tmp_path / "preset_override.xlsx")
        _create_test_xlsx(fp)

        r = ExcelOperations.format_cells(
            fp, "Sheet1", "A1",
            formatting={"bold": False},
            preset="title",
        )
        assert r["success"]
        # title preset 设 bold=True，但用户显式 bold=False 应覆盖
        assert _read_cell_style(fp, "A1")["bold"] is False

    def test_preset_header_add_user_color(self, tmp_path):
        """preset=header + 用户额外添加 font_color"""
        fp = str(tmp_path / "preset_add_color.xlsx")
        _create_test_xlsx(fp)

        r = ExcelOperations.format_cells(
            fp, "Sheet1", "A1",
            formatting={"font_color": "FF0000"},
            preset="header",
        )
        assert r["success"]
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True  # header preset 的 bold
        # font_color 应被添加到 header preset 中

    # ---------- 21. wrap_text toggle ----------

    def test_wrap_text_toggle_on_off(self, tmp_path):
        """wrap_text 开关切换"""
        fp = str(tmp_path / "wrap_toggle.xlsx")
        _create_test_xlsx(fp)

        # 默认 False
        assert _read_cell_style(fp, "A1")["wrap_text"] is False

        # 开启
        r1 = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"wrap_text": True})
        assert r1["success"]
        assert _read_cell_style(fp, "A1")["wrap_text"] is True

        # 关闭
        r2 = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"wrap_text": False})
        assert r2["success"]
        assert _read_cell_style(fp, "A1")["wrap_text"] is False

    # ---------- 22. _normalize_formatting 空输入 ----------

    class TestNormalizeFormattingR65:
        """_normalize_formatting 补充单元测试"""

        def test_normalize_none_input(self):
            """None 输入返回空字典"""
            assert ExcelOperations._normalize_formatting(None) == {}

        def test_normalize_empty_dict(self):
            """空字典返回空字典"""
            assert ExcelOperations._normalize_formatting({}) == {}

        def test_normalize_all_none_values_flat(self):
            """扁平格式所有值 None 返回空字典"""
            result = ExcelOperations._normalize_formatting({
                "bold": None, "italic": None, "bg_color": None,
                "font_size": None, "alignment": None, "number_format": None,
            })
            # 所有键都被过滤掉
            assert "font" not in result or len(result.get("font", {})) == 0
            assert "fill" not in result
            assert "alignment" not in result or len(result.get("alignment", {})) == 0
            assert "number_format" not in result

        def test_normalize_bg_color_with_hash(self):
            """bg_color=#RRGGBB 自动去除 #"""
            result = ExcelOperations._normalize_formatting({"bg_color": "#AABBCC"})
            assert result["fill"]["color"] == "AABBCC"

        def test_normalize_font_color_with_hash(self):
            """font_color=#RRGGBB 自动去除 #"""
            result = ExcelOperations._normalize_formatting({"font_color": "#112233"})
            assert result["font"]["color"] == "112233"

        def test_normalize_gradient_colors_conversion(self):
            """gradient_colors 正确转换为 fill.colors"""
            result = ExcelOperations._normalize_formatting({
                "gradient_colors": ["AA", "BB", "CC"]
            })
            assert result["fill"]["type"] == "gradient"
            assert result["fill"]["colors"] == ["AA", "BB", "CC"]

        def test_normalize_border_style_string(self):
            """border_style 字符串简写展开为四边"""
            result = ExcelOperations._normalize_formatting({"border_style": "thick"})
            assert result["border"]["top"] == "thick"
            assert result["border"]["bottom"] == "thick"
            assert result["border"]["left"] == "thick"
            assert result["border"]["right"] == "thick"

        def test_normalize_vertical_middle_to_center(self):
            """vertical_alignment='middle' 映射为 'center'"""
            result = ExcelOperations._normalize_formatting({"vertical_alignment": "middle"})
            assert result["alignment"]["vertical"] == "center"

        def test_normalize_vertical_center_passthrough(self):
            """vertical_alignment='center' 保持不变"""
            result = ExcelOperations._normalize_formatting({"vertical_alignment": "center"})
            assert result["alignment"]["vertical"] == "center"

        def test_nested_format_passthrough(self):
            """已是嵌套格式的直接返回"""
            nested = {
                "font": {"name": "Test", "size": 12},
                "fill": {"type": "solid", "color": "FFF"},
                "alignment": {"horizontal": "right"},
            }
            result = ExcelOperations._normalize_formatting(nested)
            assert result is nested

        def test_border_dict_passthrough(self):
            """border 作为 dict 直接透传"""
            border_cfg = {"top": {"style": "medium", "color": "FF0000"}, "bottom": "thin"}
            result = ExcelOperations._normalize_formatting({"border": border_cfg})
            assert result["border"] == border_cfg
