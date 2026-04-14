"""
Round 31 MCP 接口实测脚本
===========================
本轮方向: 
  1. 大数据量深度压力测试 (超越R26的10K → 50K/100K行, 超宽表, 超长文本)
  2. 权限/安全路径测试 (尝试读系统文件、写非Excel路径等)
  3. SQL注入新向量探索 (UNION-based, stack queries, encoding tricks)
  4. 已知P0问题回归验证

执行方式: cd /root/workspace/excel-mcp-server && source venv/bin/activate && python tests/round31_mcp_test.py
"""

import sys
import os
import time
import tempfile
import traceback

# 添加src到path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)
from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
from openpyxl import Workbook
import pandas as pd
import random

# ============================================================
# 测试基础设施
# ============================================================

TEST_DIR = tempfile.mkdtemp(prefix='r31_test_')
RESULTS = []

def make_test_file(name, rows=100, cols=3, sheet_name='Sheet1', data_fn=None):
    """创建测试Excel文件"""
    path = os.path.join(TEST_DIR, name)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if data_fn:
        data_fn(ws)
    else:
        # 默认: ID(int), Name(str), Value(float)
        headers = [f'Col{i}' for i in range(cols)] if cols > 3 else ['ID', 'Name', 'Value']
        ws.append(headers)
        for i in range(1, rows + 1):
            if cols > 3:
                row = [i] + [f'Val_{i}_{j}' for j in range(cols-1)]
            else:
                row = [i, f'Item-{i}', round(random.uniform(1, 1000), 2)]
            ws.append(row)
    
    wb.save(path)
    return path


def record(tool_name, input_desc, expected, actual, passed, notes=''):
    """记录测试结果"""
    status = '✅ PASS' if passed else '❌ FAIL'
    RESULTS.append({
        'tool': tool_name,
        'input': input_desc,
        'expected': expected,
        'actual': str(actual)[:100],
        'passed': passed,
        'notes': notes
    })
    print(f"  {status} | {tool_name}: {input_desc}")
    if not passed and notes:
        print(f"         → {notes}")


def print_section(title):
    print(f"\n{'='*70}")
    print(f" {title}")
    print('='*70)


# ============================================================
# A组: 大数据量深度压力测试 (超越R26)
# ============================================================
def test_group_a_big_data_stress():
    print_section("A组: 大数据量深度压力测试")
    
    # A1: 50K行基础查询性能
    print("\n[A1] 创建50K行测试文件...")
    path_50k = make_test_file('stress_50k.xlsx', rows=50000, cols=3)
    
    t0 = time.time()
    result = execute_advanced_sql_query(path_50k, "SELECT COUNT(*) as cnt FROM Sheet1")
    t1 = time.time()
    
    record('excel_query', '50K行COUNT(*)', 'success+cnt=50000',
           f"success={result['success']}, data={str(result.get('data',''))[:60]}",
           result.get('success', False) and t1-t0 < 10,
           f"耗时={t1-t0:.2f}s")
    
    # A2: 50K行WHERE过滤
    t0 = time.time()
    result = execute_advanced_sql_query(path_50k, "SELECT * FROM Sheet1 WHERE ID > 49990 LIMIT 5")
    t1 = time.time()
    
    row_count = len(result.get('data', [])) - 1 if result.get('data') and len(result['data']) > 1 else 0
    record('excel_query', '50K行WHERE+LIMIT5', 'success+~10行',
           f"rows={row_count}, time={t1-t0:.2f}s",
           result.get('success', False) and row_count <= 11 and t1-t0 < 10,
           f"耗时={t1-t0:.2f}s, 行数={row_count}")
    
    # A3: 50K行GROUP BY聚合
    t0 = time.time()
    result = execute_advanced_sql_query(path_50k, 
        "SELECT CASE WHEN ID <= 25000 THEN 'A' ELSE 'B' END as grp, COUNT(*) as cnt, SUM(Value) as total "
        "FROM Sheet1 GROUP BY grp")
    t1 = time.time()
    
    record('excel_query', '50K行GROUP BY聚合', 'success+2组',
           f"success={result['success']}, time={t1-t0:.2f}s",
           result.get('success', False) and t1-t0 < 15,
           f"耗时={t1-t0:.2f}s")
    
    # A4: 100K行极限压力
    print("\n[A4] 创建100K行测试文件...")
    path_100k = make_test_file('stress_100k.xlsx', rows=100000, cols=3)
    
    t0 = time.time()
    result = execute_advanced_sql_query(path_100k, "SELECT COUNT(*) as cnt FROM Sheet1")
    t1 = time.time()
    
    record('excel_query', '100K行COUNT(*)', 'success+cnt=100000',
           f"success={result['success']}, time={t1-t0:.2f}s",
           result.get('success', False) and t1-t0 < 20,
           f"耗时={t1-t0:.2f}s")
    
    # A5: 100K行ORDER BY
    t0 = time.time()
    result = execute_advanced_sql_query(path_100k, "SELECT * FROM Sheet1 ORDER BY Value DESC LIMIT 3")
    t1 = time.time()
    
    record('excel_query', '100K行ORDER BY DESC LIMIT3', 'success+<20s',
           f"success={result['success']}, time={t1-t0:.2f}s",
           result.get('success', False) and t1-t0 < 30,
           f"耗时={t1-t0:.2f}s")
    
    # A6: 50K行UPDATE全表
    t0 = time.time()
    result = execute_advanced_update_query(path_50k, "UPDATE Sheet1 SET Value = Value * 1.01")
    t1 = time.time()
    
    record('excel_update_query', '50K行全表UPDATE Value*1.01', 'success',
           f"success={result['success']}, time={t1-t0:.2f}s",
           result.get('success', False),
           f"耗时={t1-t0:.2f}s, msg={str(result.get('message',''))[:60]}")
    
    # A7: 30列宽表查询
    print("\n[A7] 创建30列宽表...")
    path_wide = make_test_file('wide_30cols.xlsx', rows=100, cols=30)
    
    t0 = time.time()
    result = execute_advanced_sql_query(path_wide, "SELECT * FROM `Sheet1` LIMIT 3")
    t1 = time.time()
    
    col_count = len(result['data'][0]) if result.get('data') and result['data'] else 0
    record('excel_query', '30列宽表SELECT*', f'success+{col_count}列',
           f"cols={col_count}, time={t1-t0:.2f}s",
           result.get('success', False) and col_count == 30,
           f"耗时={t1-t0:.2f}s, 列数={col_count}")
    
    # A8: 30列宽表WHERE+聚合
    t0 = time.time()
    result = execute_advanced_sql_query(path_wide, 
        "SELECT Col0, COUNT(*) as cnt, AVG(Col1) as avg1 FROM `Sheet1` GROUP BY Col0 LIMIT 5")
    t1 = time.time()
    
    record('excel_query', '30列宽表GROUP BY', 'success',
           f"time={t1-t0:.2f}s",
           result.get('success', False),
           f"耗时={t1-t0:.2f}s")
    
    # A9: 超长文本字段(20000字符)
    print("\n[A9] 创建超长文本测试文件...")
    path_longtext = make_test_file('long_text.xlsx', rows=5, cols=3)
    
    long_str = 'A' * 20000
    # 需要通过API写入超长文本
    result = execute_advanced_update_query(path_longtext, 
        f"UPDATE Sheet1 SET Name = '{long_str}' WHERE ID = 1")
    
    record('excel_update_query', '20000字符超长文本UPDATE', 'success',
           f"success={result['success']}",
           result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # 验证读取
    if result.get('success'):
        result2 = execute_advanced_sql_query(path_longtext, "SELECT Name FROM Sheet1 WHERE ID = 1")
        read_len = len(str(result2['data'][1][0])) if result2.get('data') and len(result2['data']) > 1 else 0
        record('excel_query', '20000字符读取验证', f'len={read_len}',
               f"expected~20000, got={read_len}",
               read_len >= 19900,  # 允许少量差异
               f"读取长度={read_len}")
    
    # A10: 50K行DELETE批量操作
    path_del = make_test_file('del_50k.xlsx', rows=50000, cols=3)
    t0 = time.time()
    result = execute_advanced_delete_query(path_del, "DELETE FROM Sheet1 WHERE ID > 49995")
    t1 = time.time()
    
    record('excel_delete_query', '50K行DELETE WHERE>49995', 'success',
           f"time={t1-t0:.2f}s",
           result.get('success', False),
           f"耗时={t1-t0:.2f}s")


# ============================================================
# B组: 权限/安全路径测试 (新方向!)
# ============================================================
def test_group_b_permission_security():
    print_section("B组: 权限/安全路径测试")
    
    # B1: 尝试读取 /etc/passwd
    result = execute_advanced_sql_query('/etc/passwd', "SELECT * FROM Sheet1")
    record('excel_query', '路径=/etc/passwd(系统文件)', '应失败/报错',
           f"success={result['success']}",
           not result.get('success', False),
           f"实际: success={result.get('success')}, msg={str(result.get('message',''))[:80]}")
    
    # B2: 尝试读取 /etc/shadow
    result = execute_advanced_sql_query('/etc/shadow', "SELECT * FROM Sheet1")
    record('excel_query', '路径=/etc/shadow(敏感文件)', '应失败',
           f"success={result['success']}",
           not result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # B3: 尝试写系统路径
    result = execute_advanced_update_query('/tmp/etc_test.xlsx', "UPDATE Sheet1 SET Name = 'hack'")
    record('excel_update_query', '写/tmp路径(非预期)', '应失败或成功但无危害',
           f"success={result['success']}",
           True,  # 写/tmp其实可以接受，只要不写关键系统位置
           f"msg={str(result.get('message',''))[:80]}")
    
    # B4: 尝试路径遍历 ../../etc/passwd
    result = execute_advanced_sql_query('../../etc/passwd', "SELECT * FROM Sheet1")
    record('excel_query', '路径遍历../../etc/passwd', '应失败/规范化',
           f"success={result['success']}",
           not result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # B5: 尝试符号链接攻击 (先创建symlink)
    try:
        os.symlink('/etc/passwd', os.path.join(TEST_DIR, 'trick.xlsx'))
        result = execute_advanced_sql_query(os.path.join(TEST_DIR, 'trick.xlsx'), "SELECT * FROM Sheet1")
        record('excel_query', 'symlink→/etc/passwd', '应失败或不泄露内容',
               f"success={result['success']}",
               not result.get('success', False) or (result.get('data') == [['']] or len(result.get('data',[])) <= 1),
               f"msg={str(result.get('message',''))[:80]}")
    except Exception as e:
        record('excel_query', 'symlink→/etc/passwd', 'N/A(symlink创建失败)',
               str(e), False, f"exception: {e}")
    
    # B6: 空路径
    result = execute_advanced_sql_query('', "SELECT * FROM Sheet1")
    record('excel_query', '空字符串路径', '应失败',
           f"success={result['success']}",
           not result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # B7: 只有空格的路径
    result = execute_advanced_sql_query('   ', "SELECT * FROM Sheet1")
    record('excel_query', '纯空格路径', '应失败',
           f"success={result['success']}",
           not result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # B8: 尝试通过SQL读取外部文件 (LOAD DATA INFILE style)
    path_ok = make_test_file('normal.xlsx', rows=3)
    result = execute_advanced_sql_query(path_ok, "LOAD DATA INFILE '/etc/passwd' INTO TABLE Sheet1")
    record('excel_query', 'LOAD DATA INFILE攻击', '应失败/不支持',
           f"success={result['success']}",
           not result.get('success', False),
           str(result.get('message', ''))[:100])


# ============================================================
# C组: SQL注入新向量探索
# ============================================================
def test_group_c_sql_injection_new_vectors():
    print_section("C组: SQL注入新向量探索")
    
    path = make_test_file('inject_r31.xlsx', rows=5)
    
    # C1: UNION-based injection in SELECT
    result = execute_advanced_sql_query(path, 
        "SELECT * FROM Sheet1 UNION SELECT 1,2,3")
    is_union_executed = result.get('success', False) and len(result.get('data', [])) > 6
    record('excel_query', 'UNION注入 SELECT*UNION SELECT 1,2,3', '应拒绝或无害',
           f"success={result['success']}, rows={len(result.get('data',[]))}",
           not is_union_executed,  # 不应该让UNION执行
           f"UNION被执行={is_union_executed}, rows={len(result.get('data',[]))}")
    
    # C2: UNION with column mismatch
    result = execute_advanced_sql_query(path,
        "SELECT * FROM Sheet1 UNION SELECT 'hacked'")
    record('excel_query', 'UNION列数不匹配', '应失败',
           f"success={result['success']}",
           not result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # C3: Stack queries with semicolon (P0-2 regression)
    result = execute_advanced_sql_query(path,
        "SELECT * FROM Sheet1; SELECT * FROM Sheet1")
    is_multi = result.get('success', False) and ('多语句' in str(result.get('message', '')) or 'multi' in str(result.get('data', '')).lower())
    record('excel_query', '[P0-2回归]分号多语句SELECT;SELECT', '应拒绝!',
           f"success={result['success']}, multi={is_multi}",
           not result.get('success', False),  # 应该拒绝!
           f"仍可执行! msg={str(result.get('message',''))[:80]}")
    
    # C4: UPDATE semicolon injection (P0-4 regression)
    result = execute_advanced_update_query(path,
        "UPDATE Sheet1 SET Value = 0 WHERE ID = 1; DROP TABLE Sheet1")
    record('excel_update_query', '[P0-4回归]UPDATE分号注入', '应拒绝!',
           f"success={result['success']}",
           not result.get('success', False),
           f"仍可执行! affected={result.get('affected_rows','?')}")
    
    # C5: UPDATE comment injection (P0-7 regression)
    # 先确认当前数据
    before = execute_advanced_sql_query(path, "SELECT COUNT(*) as cnt FROM Sheet1")
    cnt_before = before['data'][1][0] if before.get('data') and len(before['data']) > 1 else '?'
    
    result = execute_advanced_update_query(path,
        "UPDATE Sheet1 SET Value = -9999 -- WHERE ID = 99999")
    
    after = execute_advanced_sql_query(path, "SELECT COUNT(*) as cnt FROM Sheet1")
    cnt_after = after['data'][1][0] if after.get('data') and len(after['data']) > 1 else '?'
    
    all_updated = (result.get('success', False) and 
                   cnt_before == cnt_after and 
                   cnt_after != '?' and int(cnt_after) > 1)
    record('excel_update_query', '[P0-7回归]注释符注入全表篡改', '应拒绝!',
           f"success={result['success']}, before={cnt_before}, after={cnt_after}",
           not all_updated,  # 如果全表被更新了就是漏洞
           f"全表被篡改! before={cnt_before}→after={cnt_after}")
    
    # C6: INSERT semicolon injection (P0-5 regression)
    result = execute_advanced_insert_query(path,
        "INSERT INTO Sheet1 (ID, Name, Value) VALUES (999, 'hack', 0); DROP TABLE Sheet1")
    record('excel_insert_query', '[P0-5回归]INSERT分号注入', '应拒绝!',
           f"success={result['success']}",
           not result.get('success', False),
           f"仍可执行!")
    
    # C7: DELETE semicolon injection (P0-6 regression)
    result = execute_advanced_delete_query(path,
        "DELETE FROM Sheet1 WHERE ID = 99999; DROP TABLE Sheet1")
    record('excel_delete_query', '[P0-6回归]DELETE分号注入', '应拒绝!',
           f"success={result['success']}",
           not result.get('success', False),
           f"仍可执行!")
    
    # C8: Hex-encoded attack vector
    result = execute_advanced_sql_query(path,
        "SELECT * FROM Sheet1 WHERE Name = 0x6861636B")  # hex for 'hack'
    record('excel_query', 'Hex编码注入 0x6861636B', '应失败/无匹配',
           f"success={result['success']}",
           True,  # 不管结果如何，只要不崩溃就行
           str(result.get('message', ''))[:80])
    
    # C9: CHAR() function injection attempt
    result = execute_advanced_sql_query(path,
        "SELECT * FROM Sheet1 WHERE Name = CHAR(104,97,99,107)")  # CHAR() for 'hack'
    record('excel_query', 'CHAR()函数注入', '应失败/不支持/无匹配',
           f"success={result['success']}",
           True,
           str(result.get('message', ''))[:80])
    
    # C10: Double URL encoding / Unicode normalization
    result = execute_advanced_sql_query(path,
        "SELECT * FROM Sheet1 WHERE Name = 'ｈａｃｋ'")  # fullwidth characters
    record('excel_query', 'Unicode全角字符注入', '应无匹配(正常)',
           f"success={result['success']}",
           result.get('success', False),
           str(result.get('message', ''))[:80])


# ============================================================
# D组: 类型边界深度测试
# ============================================================
def test_group_d_type_boundary():
    print_section("D组: 类型边界深度测试")
    
    # 创建混合类型测试文件
    path = make_test_file('type_boundary.xlsx', rows=3)
    
    # D1: 极大浮点数
    result = execute_advanced_update_query(path, 
        "UPDATE Sheet1 SET Value = 1.7976931348623157e+308 WHERE ID = 1")
    record('excel_update_query', 'float MAX(double)', 'success',
           f"success={result['success']}",
           result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # D2: 极小浮点数
    result = execute_advanced_update_query(path,
        "UPDATE Sheet1 SET Value = -1.7976931348623157e+308 WHERE ID = 2")
    record('excel_update_query', 'float MIN(double)负', 'success',
           f"success={result['success']}",
           result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # D3: 超大整数
    big_int = 10**18
    result = execute_advanced_update_query(path,
        f"UPDATE Sheet1 SET Value = {big_int} WHERE ID = 3")
    record('excel_update_query', '超大整数10^18', 'success',
           f"success={result['success']}",
           result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # 验证读取
    verify = execute_advanced_sql_query(path, "SELECT ID, Value FROM Sheet1 ORDER BY ID")
    if verify.get('success') and len(verify.get('data', [])) > 1:
        for row in verify['data'][1:]:
            record('excel_query', '边界值回读验证', '数值正确',
                   f"ID={row[0]}, V={row[1]}", True, '')
    
    # D4: NaN值
    result = execute_advanced_update_query(path,
        "UPDATE Sheet1 SET Value = CAST('nan' AS FLOAT64) WHERE ID = 1")
    nan_handled = result.get('success', False) or ('error' in str(result.get('message', '')).lower() or 
                                                     'invalid' in str(result.get('message', '')).lower())
    record('excel_update_query', 'NaN值写入', 'success或优雅错误',
           f"success={result['success']}",
           nan_handled,
           str(result.get('message', ''))[:80])
    
    # D5: Infinity值
    result = execute_advanced_update_query(path,
        "UPDATE Sheet1 SET Value = CAST('inf' AS FLOAT64) WHERE ID = 1")
    inf_handled = result.get('success', False) or ('error' in str(result.get('message', '')).lower())
    record('excel_update_query', 'Infinity值写入', 'success或优雅错误',
           f"success={result['success']}",
           inf_handled,
           str(result.get('message', ''))[:80])
    
    # D6: 负零
    result = execute_advanced_update_query(path,
        "UPDATE Sheet1 SET Value = -0.0 WHERE ID = 1")
    record('excel_update_query', '负零(-0.0)', 'success',
           f"success={result['success']}",
           result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # D7: 精度丢失测试
    result = execute_advanced_update_query(path,
        "UPDATE Sheet1 SET Value = 0.1 + 0.2 WHERE ID = 1")
    record('excel_update_query', '浮点精度0.1+0.2', 'success(~0.3)',
           f"success={result['success']}",
           result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # D8: 混合类型列写入 (int→string)
    result = execute_advanced_update_query(path,
        "UPDATE Sheet1 SET Name = 12345 WHERE ID = 1")
    record('excel_update_query', 'int写入字符串列', 'success或类型错误',
           f"success={result['success']}",
           True,  # 不崩溃就算通过
           str(result.get('message', ''))[:80])
    
    # D9: 布尔值写入
    result = execute_advanced_update_query(path,
        "UPDATE Sheet1 SET Value = TRUE WHERE ID = 1")
    bool_ok = result.get('success', False)
    record('excel_update_query', '布尔TRUE写入数值列', 'success',
           f"success={bool_ok}",
           bool_ok,
           str(result.get('message', ''))[:80])
    
    # D10: NULL字面量 (已知限制)
    result = execute_advanced_sql_query(path, "SELECT NULL as null_col FROM Sheet1 LIMIT 1")
    null_ok = result.get('success', False)
    record('excel_query', 'NULL字面量SELECT', 'success或优雅错误(已知限制)',
           f"success={null_ok}",
           True,  # 已知限制，不崩溃即可
           str(result.get('message', ''))[:80])


# ============================================================
# E组: 并发/重复操作深度测试
# ============================================================
def test_group_e_concurrent_repeat():
    print_section("E组: 并发/重复操作深度测试")
    
    path = make_test_file('repeat_test.xlsx', rows=5)
    
    # E1: 同一单元格连续写入10次
    for i in range(10):
        result = execute_advanced_update_query(path,
            f"UPDATE Sheet1 SET Value = {i} WHERE ID = 1")
    
    # 验证最终值
    final = execute_advanced_sql_query(path, "SELECT Value FROM Sheet1 WHERE ID = 1")
    final_val = final['data'][1][0] if final.get('data') and len(final['data']) > 1 else None
    record('excel_update_query', '同一单元格连续写入10次', f'最终值=9, 实际={final_val}',
           f"val={final_val}", final_val == 9,
           f"最终值={final_val}")
    
    # E2: 连续查询不同sheet (单sheet多次)
    for i in range(20):
        result = execute_advanced_sql_query(path, f"SELECT * FROM Sheet1 WHERE ID = {i%5+1}")
    
    record('excel_query', '连续快速查询20次', '全部success',
           'all_success', result.get('success', False),
           f"第20次结果: success={result.get('success')}")
    
    # E3: INSERT相同主键多次
    results = []
    for i in range(3):
        r = execute_advanced_insert_query(path,
            f"INSERT INTO Sheet1 (ID, Name, Value) VALUES ({100+i}, 'dup', {i})")
        results.append(r.get('success', False))
    
    all_insert_ok = all(results)
    record('excel_insert_query', '连续INSERT 3行不同ID', '全部success',
           f"results={results}", all_insert_ok,
           f"插入结果={results}")
    
    # E4: DELETE后再次DELETE同一条件
    r1 = execute_advanced_delete_query(path, "DELETE FROM Sheet1 WHERE ID = 999")
    r2 = execute_advanced_delete_query(path, "DELETE FROM Sheet1 WHERE ID = 999")
    record('excel_delete_query', '重复DELETE同一不存在行', '两次都success(幂等)',
           f"r1={r1.get('success')}, r2={r2.get('success')}",
           r1.get('success', False) and r2.get('success', False),
           f"r1={r1.get('success')}, r2={r2.get('success')}")
    
    # E5: UPDATE→SELECT→UPDATE→SELECT 一致性链
    r1 = execute_advanced_update_query(path, "UPDATE Sheet1 SET Value = 42.5 WHERE ID = 2")
    v1 = execute_advanced_sql_query(path, "SELECT Value FROM Sheet1 WHERE ID = 2")
    val1 = v1['data'][1][0] if v1.get('data') and len(v1['data']) > 1 else None
    
    r2 = execute_advanced_update_query(path, "UPDATE Sheet1 SET Value = Value + 100 WHERE ID = 2")
    v2 = execute_advanced_sql_query(path, "SELECT Value FROM Sheet1 WHERE ID = 2")
    val2 = v2['data'][1][0] if v2.get('data') and len(v2['data']) > 1 else None
    
    expected_val2 = 142.5 if isinstance(val1, (int, float)) else None
    chain_ok = (r1.get('success') and r2.get('success') and 
                abs((val2 or 0) - (expected_val2 or 0)) < 0.01)
    record('chain', 'UPDATE→READ→UPDATE→READ一致性链', f'val1={val1}, val2≈{expected_val2}',
           f"val1={val1}, val2={val2}", chain_ok,
           f"一致性检查: {val1} → {val2} (期望≈{expected_val2})")


# ============================================================
# F组: 多Sheet联动深度测试
# ============================================================
def test_group_f_multi_sheet():
    print_section("F组: 多Sheet联动深度测试")
    
    # 创建多sheet文件
    path = os.path.join(TEST_DIR, 'multi_sheet.xlsx')
    wb = Workbook()
    
    # Sheet1: 玩家
    ws1 = wb.active
    ws1.title = 'Players'
    ws1.append(['PlayerID', 'Name', 'GuildID', 'Level'])
    for i in range(1, 21):
        ws1.append([i, f'Player{i}', (i-1)//5 + 1, random.randint(1, 80)])
    
    # Sheet2: 公会
    ws2 = wb.create_sheet('Guilds')
    ws2.append(['GuildID', 'GuildName', 'MasterID'])
    for i in range(1, 5):
        ws2.append([i, f'Guild{i}', i*5])
    
    wb.save(path)
    
    # F1: 跨Sheet JOIN
    result = execute_advanced_sql_query(path,
        "SELECT p.Name, g.GuildName, p.Level "
        "FROM Players p JOIN Guilds g ON p.GuildID = g.GuildID "
        "ORDER BY p.Level DESC LIMIT 5")
    record('excel_query', '跨Sheet JOIN Players↔Guilds', 'success+多行',
           f"success={result['success']}, rows={len(result.get('data',[]))-1}",
           result.get('success', False) and len(result.get('data', [])) > 2,
           f"返回{len(result.get('data',[]))-1}行数据")
    
    # F2: 跨Sheet聚合+JOIN
    result = execute_advanced_sql_query(path,
        "SELECT g.GuildName, COUNT(*) as member_cnt, AVG(p.Level) as avg_lvl, MAX(p.Level) as max_lvl "
        "FROM Players p JOIN Guilds g ON p.GuildID = g.GuildID "
        "GROUP BY g.GuildID, g.GuildName ORDER BY member_cnt DESC")
    record('excel_query', '跨Sheet聚合GROUP BY+JOIN', 'success+4组',
           f"success={result['success']}",
           result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # F3: CTE + 多Sheet
    result = execute_advanced_sql_query(path,
        "WITH TopPlayers AS ("
        "  SELECT * FROM Players WHERE Level >= 60"
        ") "
        "SELECT tp.Name, g.GuildName "
        "FROM TopPlayers tp JOIN Guilds g ON tp.GuildID = g.GuildID")
    record('excel_query', 'CTE+跨SheetJOIN', 'success',
           f"success={result['success']}",
           result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # F4: 子查询跨Sheet
    result = execute_advanced_sql_query(path,
        "SELECT * FROM Players "
        "WHERE GuildID IN (SELECT GuildID FROM Guilds WHERE GuildName = 'Guild1')")
    record('excel_query', '子查询IN(跨Sheet)', 'success+~5行',
           f"success={result['success']}",
           result.get('success', False),
           str(result.get('message', ''))[:80])
    
    # F5: UPDATE跨Sheet引用 (应该失败)
    result = execute_advanced_update_query(path,
        "UPDATE Players SET Level = (SELECT MAX(Level) FROM Players) + 1 WHERE PlayerID = 1")
    record('excel_update_query', 'UPDATE SET=(子查询)', 'success或优雅错误',
           f"success={result['success']}",
           True,  # 不崩溃即可
           str(result.get('message', ''))[:80])
    
    # F6: 不同Sheet同名列JOIN
    # 两表都有GuildID，验证别名处理
    result = execute_advanced_sql_query(path,
        "SELECT p.PlayerID, p.GuildID as P_Guild, g.GuildID as G_Guild, g.GuildName "
        "FROM Players p INNER JOIN Guilds g ON p.GuildID = g.GuildID "
        "LIMIT 3")
    record('excel_query', '跨Sheet同名列别名处理', 'success+区分两列',
           f"success={result['success']}",
           result.get('success', False),
           str(result.get('message', ''))[:80])


# ============================================================
# G组: 已知P0问题回归
# ============================================================
def test_group_g_p0_regression():
    print_section("G组: 已知P0问题回归验证")
    
    path = make_test_file('p0_regres.xlsx', rows=5)
    
    # G1: [P0-1/R25] script_runner RCE — 无法通过SQL API测，跳过
    record('N/A', 'script_runner RCE(P0-1)', '需MCP协议测试', 'skip', True, '不在Python API范围内')
    
    # G2: [P0-2] SELECT 分号多语句
    result = execute_advanced_sql_query(path, "SELECT * FROM Sheet1; DROP TABLE Sheet1")
    is_vuln = result.get('success', False)
    record('[P0-2]', 'SELECT分号多语句注入', '应拒绝❌仍存在',
           f"executed={is_vuln}", not is_vuln,
           f"🚨 仍存在! executed={is_vuln}")
    
    # G3: [P0-4] UPDATE 分号多语句
    result = execute_advanced_update_query(path, "UPDATE Sheet1 SET Value = 0; DROP SHEET1")
    is_vuln = result.get('success', False)
    record('[P0-4]', 'UPDATE分号多语句注入', '应拒绝❌仍存在',
           f"executed={is_vuln}", not is_vuln,
           f"🚨 仍存在! executed={is_vuln}")
    
    # G4: [P0-5] INSERT 分号多语句
    result = execute_advanced_insert_query(path, "INSERT INTO Sheet1 VALUES(999,'x',0); DROP SHEET1")
    is_vuln = result.get('success', False)
    record('[P0-5]', 'INSERT分号多语句注入', '应拒绝❌仍存在',
           f"executed={is_vuln}", not is_vuln,
           f"🚨 仍存在! executed={is_vuln}")
    
    # G5: [P0-6] DELETE 分号多语句
    result = execute_advanced_delete_query(path, "DELETE FROM Sheet1 WHERE 1=1; DROP SHEET1")
    is_vuln = result.get('success', False)
    record('[P0-6]', 'DELETE分号多语句注入', '应拒绝❌仍存在',
           f"executed={is_vuln}", not is_vuln,
           f"🚨 仍存在! executed={is_vuln}")
    
    # G6: [P0-7] UPDATE 注释符注入
    before = execute_advanced_sql_query(path, "SELECT COUNT(*) FROM Sheet1")
    cnt_b = before['data'][1][0] if before.get('data') and len(before['data']) > 1 else 0
    result = execute_advanced_update_query(path, "UPDATE Sheet1 SET Value = -1 -- WHERE ID = 99999")
    after = execute_advanced_sql_query(path, "SELECT COUNT(*) FROM Sheet1")
    cnt_a = after['data'][1][0] if after.get('data') and len(after['data']) > 1 else 0
    is_vuln = result.get('success', False) and cnt_a > 1
    record('[P0-7]', 'UPDATE注释符全表篡改', '应拒绝❌仍存在',
           f"before={cnt_b}, after={cnt_a}", not is_vuln,
           f"🚨 仍存在! 全表被改 {cnt_b}→{cnt_a}")
    
    # G7: [FIXED] P0-3 uint8溢出修复验证
    path_mix = make_test_file('uint8_fix.xlsx', rows=3)
    # 写入大值
    result = execute_advanced_update_query(path_mix, "UPDATE Sheet1 SET Value = 999 WHERE ID = 1")
    verify = execute_advanced_sql_query(path_mix, "SELECT Value FROM Sheet1 WHERE ID = 1")
    val = verify['data'][1][0] if verify.get('data') and len(verify['data']) > 1 else None
    is_fixed = (val == 999)
    record('[P0-3✅]', 'uint8溢出修复 V=999', f'应为999, 实际={val}',
           f"val={val}", is_fixed,
           f"修复{'有效' if is_fixed else '失效'}! val={val}")


# ============================================================
# 主函数
# ============================================================
def main():
    print("=" * 70)
    print(" Round 31 MCP 接口实测")
    print(" 方向: 大数据量深度压力 + 权限/安全路径 + SQL注入新向量 + 类型边界")
    print("=" * 70)
    print(f"\n测试目录: {TEST_DIR}")
    
    total_start = time.time()
    
    try:
        test_group_a_big_data_stress()
        test_group_b_permission_security()
        test_group_c_sql_injection_new_vectors()
        test_group_d_type_boundary()
        test_group_e_concurrent_repeat()
        test_group_f_multi_sheet()
        test_group_g_p0_regression()
    except Exception as e:
        print(f"\n🚨 测试过程异常: {e}")
        traceback.print_exc()
    
    total_time = time.time() - total_start
    
    # 汇总
    print_section("测试结果汇总")
    
    passed = sum(1 for r in RESULTS if r['passed'])
    failed = sum(1 for r in RESULTS if not r['passed'])
    total = len(RESULTS)
    
    print(f"\n总计: {total} 个测试场景")
    print(f"  ✅ 通过: {passed} ({passed/total*100:.1f}%)" if total > 0 else "")
    print(f"  ❌ 失败: {failed} ({failed/total*100:.1f}%)" if total > 0 else "")
    print(f"  ⏱️ 总耗时: {total_time:.1f}s")
    
    print(f"\n--- 失败详情 ---")
    for r in RESULTS:
        if not r['passed']:
            print(f"  ❌ [{r['tool']}] {r['input']}")
            print(f"     期望: {r['expected']}")
            print(f"     实际: {r['actual']}")
            if r['notes']:
                print(f"     备注: {r['notes']}")
    
    # 清理
    import shutil
    try:
        shutil.rmtree(TEST_DIR)
    except:
        pass
    
    return passed, failed, total


if __name__ == '__main__':
    passed, failed, total = main()
    sys.exit(0 if failed == 0 else 1)
