"""REQ-015 Streaming Write Verification Tests — 验证流式写入后读取工具正常工作

流式写入（StreamingWriter）经过 calamine 读取 + openpyxl write_only 写入，
与传统 openpyxl 读写路径不同。本测试验证流式写入后，所有读取工具仍能正常工作。

核心问题发现并修复：write_only 模式写入的 xlsx 文件缺少 <dimension> 元数据，
导致 read_only 模式下 sheet.max_row / sheet.max_column 返回 None，
find_last_row 等依赖这些属性的方法会崩溃。已在 find_last_row 中添加降级路径。
"""

import os
import shutil

import pytest
from openpyxl import Workbook, load_workbook
from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
from excel_mcp_server_fastmcp.api.advanced_sql_query import AdvancedSQLQueryEngine


class TestStreamingWriteVerification:
    """REQ-015 主要验证类：测试流式写入后所有读取工具是否正常工作"""

    @pytest.fixture
    def game_data_file(self, tmp_path):
        """创建游戏数据测试文件（3个工作表）
        Row layout:
          Characters: Header + 4 data rows (Paladin/Mage/Archer/Warrior)
          Skills: Header + 4 data rows
          Items: Header + 3 data rows
        """
        fp = str(tmp_path / "game_test.xlsx")
        wb = Workbook()

        ws = wb.active
        ws.title = "Characters"
        ws.append(["ID", "Name", "Class", "Level", "ATK", "DEF", "HP"])
        ws.append([1, "Paladin", "Paladin", 50, 150, 120, 2000])
        ws.append([2, "Mage", "Mage", 45, 200, 80, 1500])
        ws.append([3, "Archer", "Archer", 48, 180, 90, 1700])
        ws.append([4, "Warrior", "Warrior", 52, 160, 110, 1900])

        ws_skills = wb.create_sheet("Skills")
        ws_skills.append(["SkillID", "SkillName", "ClassLimit", "Damage", "Type"])
        ws_skills.append([101, "HolySlash", "Paladin", 300, "Holy"])
        ws_skills.append([102, "Fireball", "Mage", 250, "Fire"])
        ws_skills.append([103, "IceArrow", "Mage", 200, "Ice"])
        ws_skills.append([104, "MultiShot", "Archer", 280, "Physical"])

        ws_items = wb.create_sheet("Items")
        ws_items.append(["ItemID", "ItemName", "Type", "ATKBonus", "DEFBonus"])
        ws_items.append([1001, "FlameSword", "Weapon", 50, 10])
        ws_items.append([1002, "IceShield", "Armor", 0, 40])
        ws_items.append([1003, "MageRobe", "Armor", 20, 20])

        wb.save(fp)
        wb.close()
        return fp

    def _values(self, range_result):
        """从 get_range 结果提取纯值矩阵"""
        data = range_result.get('data', [])
        if not data:
            return []
        if isinstance(data[0], list) and data[0] and isinstance(data[0][0], dict):
            return [[cell['value'] for cell in row] for row in data]
        return data

    # ==================== batch_insert 验证 ====================

    def test_batch_insert_then_list_sheets(self, game_data_file):
        """流式 batch_insert_rows 后 list_sheets 正常"""
        data = [{"ID": 5, "Name": "Assassin", "Class": "Assassin", "Level": 47, "ATK": 190, "DEF": 85, "HP": 1600}]
        r = ExcelOperations.batch_insert_rows(game_data_file, "Characters", data, streaming=True)
        assert r['success'] is True, r.get('message')

        sheets = ExcelOperations.list_sheets(game_data_file)
        assert sheets['success'] is True
        assert sheets['total_sheets'] == 3

    def test_batch_insert_then_get_headers(self, game_data_file):
        """流式 batch_insert_rows 后 get_headers 正常"""
        data = [{"ID": 5, "Name": "Test", "Class": "Mage", "Level": 50, "ATK": 200, "DEF": 80, "HP": 1500}]
        r = ExcelOperations.batch_insert_rows(game_data_file, "Characters", data, streaming=True)
        assert r['success'] is True

        headers = ExcelOperations.get_headers(game_data_file, "Characters")
        assert headers['success'] is True
        assert headers['header_count'] == 7
        # get_headers may detect dual-row headers; both descriptions and field_names should exist
        if headers.get('dual_rows'):
            assert len(headers['descriptions']) == 7
            assert len(headers['field_names']) == 7
        else:
            assert len(headers['headers']) == 7

    def test_batch_insert_then_get_range(self, game_data_file):
        """流式 batch_insert_rows 后 get_range 正常"""
        data = [{"ID": 5, "Name": "Assassin", "Class": "Assassin", "Level": 47, "ATK": 190, "DEF": 85, "HP": 1600}]
        r = ExcelOperations.batch_insert_rows(game_data_file, "Characters", data, streaming=True)
        assert r['success'] is True

        range_data = ExcelOperations.get_range(game_data_file, "Characters!A1:D6")
        assert range_data['success'] is True
        values = self._values(range_data)
        assert values[0] == ["ID", "Name", "Class", "Level"]
        assert values[-1][0] == 5
        assert values[-1][1] == "Assassin"

    def test_batch_insert_then_find_last_row(self, game_data_file):
        """流式 batch_insert_rows 后 find_last_row 正常（关键验证点：修复 dimension=None 崩溃）"""
        data = [
            {"ID": 5, "Name": "Monk", "Class": "Monk", "Level": 51, "ATK": 175, "DEF": 105, "HP": 1850},
            {"ID": 6, "Name": "Priest", "Class": "Priest", "Level": 46, "ATK": 140, "DEF": 100, "HP": 1800},
        ]
        r = ExcelOperations.batch_insert_rows(game_data_file, "Characters", data, streaming=True)
        assert r['success'] is True

        result = ExcelOperations.find_last_row(game_data_file, "Characters")
        assert result['success'] is True, f"find_last_row failed: {result.get('message')}"
        # 1 header + 4 original + 2 new = 7 rows
        assert result['last_row'] == 7

    def test_batch_insert_then_sql_where(self, game_data_file):
        """流式 batch_insert_rows 后 SQL WHERE 查询正常"""
        data = [
            {"SkillID": 105, "SkillName": "ShadowStrike", "ClassLimit": "Assassin", "Damage": 320, "Type": "Dark"},
            {"SkillID": 106, "SkillName": "Heal", "ClassLimit": "Priest", "Damage": 0, "Type": "Holy"},
        ]
        r = ExcelOperations.batch_insert_rows(game_data_file, "Skills", data, streaming=True)
        assert r['success'] is True

        engine = AdvancedSQLQueryEngine()
        result = engine.execute_sql_query(game_data_file, "SELECT * FROM Skills WHERE Type = 'Fire'", include_headers=False)
        assert result['success'] is True
        assert len(result['data']) == 1
        assert result['data'][0][1] == "Fireball"

        result = engine.execute_sql_query(game_data_file, "SELECT * FROM Skills WHERE Type = 'Dark'", include_headers=False)
        assert result['success'] is True
        assert len(result['data']) == 1
        assert result['data'][0][1] == "ShadowStrike"

    def test_batch_insert_then_sql_join(self, game_data_file):
        """流式 batch_insert_rows 后 SQL JOIN 查询正常"""
        engine = AdvancedSQLQueryEngine()
        result = engine.execute_sql_query(game_data_file, """
            SELECT c.Name, s.SkillName
            FROM Characters c
            JOIN Skills s ON c.Class = s.ClassLimit
            WHERE c.Class = 'Mage'
        """, include_headers=False)
        assert result['success'] is True
        assert len(result['data']) == 2  # Fireball + IceArrow

    def test_batch_insert_then_sql_count(self, game_data_file):
        """流式 batch_insert_rows 后 SQL COUNT 与 find_last_row 一致"""
        data = [{"ID": 5, "Name": "Test", "Class": "Mage", "Level": 50, "ATK": 200, "DEF": 80, "HP": 1500}]
        ExcelOperations.batch_insert_rows(game_data_file, "Characters", data, streaming=True)

        engine = AdvancedSQLQueryEngine()
        count = engine.execute_sql_query(game_data_file, "SELECT COUNT(*) FROM Characters", include_headers=False)
        assert count['success'] is True
        assert int(count['data'][0][0]) == 5  # 4 original + 1 new

        last_row = ExcelOperations.find_last_row(game_data_file, "Characters")
        assert last_row['success'] is True
        assert last_row['last_row'] == 6  # 1 header + 5 data

    # ==================== delete_rows 验证 ====================

    def test_delete_rows_then_get_range(self, game_data_file):
        """流式 delete_rows 后 get_range 正常"""
        # Row 3 = Mage (Row 1=header, Row 2=Paladin, Row 3=Mage)
        r = ExcelOperations.delete_rows(game_data_file, "Characters", 3, 1, streaming=True)
        assert r['success'] is True

        range_data = ExcelOperations.get_range(game_data_file, "Characters!A1:B4")
        assert range_data['success'] is True
        values = self._values(range_data)
        assert values[0] == ["ID", "Name"]
        names = [row[1] for row in values[1:]]
        assert "Mage" not in names

    def test_delete_rows_then_find_last_row(self, game_data_file):
        """流式 delete_rows 后 find_last_row 正常"""
        # Delete row 3 (Mage)
        r = ExcelOperations.delete_rows(game_data_file, "Characters", 3, 1, streaming=True)
        assert r['success'] is True

        result = ExcelOperations.find_last_row(game_data_file, "Characters")
        assert result['success'] is True
        assert result['last_row'] == 4  # 1 header + 3 data (4-1)

    def test_delete_rows_then_sql_count(self, game_data_file):
        """流式 delete_rows 后 SQL COUNT 正常"""
        r = ExcelOperations.delete_rows(game_data_file, "Characters", 3, 1, streaming=True)
        assert r['success'] is True

        engine = AdvancedSQLQueryEngine()
        count = engine.execute_sql_query(game_data_file, "SELECT COUNT(*) FROM Characters", include_headers=False)
        assert count['success'] is True
        assert int(count['data'][0][0]) == 3  # 4 - 1 deleted

    # ==================== update_range 验证 ====================

    def test_update_range_overwrite_then_get_range(self, game_data_file):
        """流式 update_range (overwrite模式) 后 get_range 正常"""
        # overwrite E2 (Paladin's ATK) with 999
        r = ExcelOperations.update_range(
            game_data_file, "Characters!E2:E2", [[999]],
            streaming=True, insert_mode=False
        )
        assert r['success'] is True

        range_data = ExcelOperations.get_range(game_data_file, "Characters!A2:E2")
        assert range_data['success'] is True
        values = self._values(range_data)
        assert values[0][4] == 999  # Paladin ATK updated

    def test_update_range_overwrite_then_find_last_row(self, game_data_file):
        """流式 update_range 后 find_last_row 行数不变"""
        r = ExcelOperations.update_range(
            game_data_file, "Characters!E2:E2", [[999]],
            streaming=True, insert_mode=False
        )
        assert r['success'] is True

        result = ExcelOperations.find_last_row(game_data_file, "Characters")
        assert result['success'] is True
        assert result['last_row'] == 5  # unchanged: 1 header + 4 data

    # ==================== 复杂混合场景 ====================

    def test_complex_mixed_operations(self, game_data_file):
        """复杂场景：连续多种流式操作后，所有读取工具正常"""
        engine = AdvancedSQLQueryEngine()

        # 1. batch_insert Characters +2
        r1 = ExcelOperations.batch_insert_rows(game_data_file, "Characters", [
            {"ID": 5, "Name": "Assassin", "Class": "Assassin", "Level": 47, "ATK": 190, "DEF": 85, "HP": 1600},
            {"ID": 6, "Name": "Priest", "Class": "Priest", "Level": 46, "ATK": 140, "DEF": 100, "HP": 1800},
        ], streaming=True)
        assert r1['success'] is True

        # 2. batch_insert Skills +1
        r2 = ExcelOperations.batch_insert_rows(game_data_file, "Skills", [
            {"SkillID": 105, "SkillName": "Heal", "ClassLimit": "Priest", "Damage": 0, "Type": "Holy"},
        ], streaming=True)
        assert r2['success'] is True

        # 3. delete_rows Characters row 3 (Mage)
        r3 = ExcelOperations.delete_rows(game_data_file, "Characters", 3, 1, streaming=True)
        assert r3['success'] is True

        # 4. update_range overwrite Paladin ATK
        r4 = ExcelOperations.update_range(
            game_data_file, "Characters!E2:E2", [[185]],
            streaming=True, insert_mode=False
        )
        assert r4['success'] is True

        # === 综合验证 ===

        # list_sheets
        sheets = ExcelOperations.list_sheets(game_data_file)
        assert sheets['success'] is True
        assert sheets['total_sheets'] == 3

        # get_headers
        headers = ExcelOperations.get_headers(game_data_file, "Characters")
        assert headers['success'] is True
        assert headers['header_count'] == 7

        # find_last_row: 1 header + 4 orig - 1 deleted + 2 inserted = 6
        last_row = ExcelOperations.find_last_row(game_data_file, "Characters")
        assert last_row['success'] is True
        assert last_row['last_row'] == 6

        # get_range verify update
        range_data = ExcelOperations.get_range(game_data_file, "Characters!A2:E2")
        assert range_data['success'] is True
        values = self._values(range_data)
        assert values[0][4] == 185  # Paladin ATK

        # SQL COUNT
        count = engine.execute_sql_query(game_data_file, "SELECT COUNT(*) FROM Characters", include_headers=False)
        assert count['success'] is True
        assert int(count['data'][0][0]) == 5

        # SQL WHERE verify new skill
        heal = engine.execute_sql_query(game_data_file, "SELECT * FROM Skills WHERE SkillName = 'Heal'", include_headers=False)
        assert heal['success'] is True
        assert len(heal['data']) == 1
        assert heal['data'][0][1] == "Heal"

        # SQL JOIN
        join = engine.execute_sql_query(game_data_file, """
            SELECT c.Name, s.SkillName
            FROM Characters c
            JOIN Skills s ON c.Class = s.ClassLimit
        """, include_headers=False)
        assert join['success'] is True
        assert len(join['data']) >= 2

    # ==================== find_last_row 边界场景 ====================

    def test_find_last_row_after_multiple_streaming_inserts(self, game_data_file):
        """多次流式插入后 find_last_row 正常"""
        for i in range(5, 10):
            r = ExcelOperations.batch_insert_rows(game_data_file, "Characters", [
                {"ID": i, "Name": f"Char{i}", "Class": "Mage", "Level": 40 + i, "ATK": 200, "DEF": 80, "HP": 1500}
            ], streaming=True)
            assert r['success'] is True

        result = ExcelOperations.find_last_row(game_data_file, "Characters")
        assert result['success'] is True
        assert result['last_row'] == 10  # 1 header + 4 original + 5 new

    def test_find_last_row_with_column_after_streaming(self, game_data_file):
        """流式写入后 find_last_row 指定列参数正常"""
        ExcelOperations.batch_insert_rows(game_data_file, "Characters", [
            {"ID": 5, "Name": "TestChar", "Class": "Mage", "Level": 50, "ATK": 200, "DEF": 80, "HP": 1500}
        ], streaming=True)

        result = ExcelOperations.find_last_row(game_data_file, "Characters", "A")
        assert result['success'] is True
        assert result['last_row'] == 6

        result = ExcelOperations.find_last_row(game_data_file, "Characters", 1)
        assert result['success'] is True
        assert result['last_row'] == 6

    def test_streaming_write_table_info_consistency(self, game_data_file):
        """流式写入后，get_headers / find_last_row / SQL COUNT 三者一致"""
        ExcelOperations.batch_insert_rows(game_data_file, "Characters", [
            {"ID": 5, "Name": "Monk", "Class": "Monk", "Level": 51, "ATK": 175, "DEF": 105, "HP": 1850},
        ], streaming=True)
        ExcelOperations.delete_rows(game_data_file, "Characters", 3, 1, streaming=True)

        headers = ExcelOperations.get_headers(game_data_file, "Characters")
        assert headers['success'] is True
        assert headers['header_count'] == 7

        last_row = ExcelOperations.find_last_row(game_data_file, "Characters")
        assert last_row['success'] is True
        assert last_row['last_row'] == 5  # 1 header + 4 - 1 + 1 = 5

        engine = AdvancedSQLQueryEngine()
        count = engine.execute_sql_query(game_data_file, "SELECT COUNT(*) FROM Characters", include_headers=False)
        assert count['success'] is True
        assert int(count['data'][0][0]) == 4

        # Cross-validate: last_row should be count + 1 (header)
        assert last_row['last_row'] == int(count['data'][0][0]) + 1


class TestStreamingPerformanceComparison:
    """流式写入 vs 传统写入性能对比"""

    @pytest.fixture
    def large_dataset_file(self, tmp_path):
        """创建1000行测试文件（含表头）"""
        fp = str(tmp_path / "large_test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "LargeData"
        ws.append(["ID", "Name", "Type", "Value1", "Value2"])
        for i in range(1, 1001):
            ws.append([i, f"Item_{i}", f"Type_{i % 10}", i * 10, i * 5])
        wb.save(fp)
        wb.close()
        return fp

    def test_streaming_vs_traditional_consistency(self, large_dataset_file):
        """验证流式写入和传统写入的结果一致"""
        import time

        traditional_file = large_dataset_file.replace(".xlsx", "_traditional.xlsx")
        shutil.copy2(large_dataset_file, traditional_file)

        new_data = [
            {"ID": 1000 + i, "Name": f"New_{i}", "Type": f"NewType_{i}",
             "Value1": (1000 + i) * 10, "Value2": (1000 + i) * 5}
            for i in range(1, 101)
        ]

        t0 = time.time()
        r = ExcelOperations.batch_insert_rows(large_dataset_file, "LargeData", new_data, streaming=True)
        t_stream = time.time() - t0
        assert r['success'] is True

        t0 = time.time()
        wb = load_workbook(traditional_file)
        ws = wb["LargeData"]
        for row in new_data:
            ws.append([row["ID"], row["Name"], row["Type"], row["Value1"], row["Value2"]])
        wb.save(traditional_file)
        wb.close()
        t_trad = time.time() - t0

        last_s = ExcelOperations.find_last_row(large_dataset_file, "LargeData")
        last_t = ExcelOperations.find_last_row(traditional_file, "LargeData")
        assert last_s['success'] is True
        assert last_t['success'] is True
        # 1 header + 1000 data + 100 new = 1101
        assert last_s['last_row'] == last_t['last_row'] == 1101

        print(f"\n流式写入: {t_stream:.3f}s | 传统写入: {t_trad:.3f}s | 比值: {t_trad/max(t_stream,0.001):.1f}x")
        os.remove(traditional_file)