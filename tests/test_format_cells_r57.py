# -*- coding: utf-8 -*-
"""
format_cells 第四轮边缘案例测试 - R57 迭代

新增覆盖:
  - bg_color / font_color 带 # 前缀的容错
  - 中文 Sheet 名格式化
  - 全组合操作: merge + bold + bg_color + alignment + border_style 同时传
  - unmerge 后数据完整性验证
  - vertical_alignment 扁平格式
  - border 字典每边独立颜色配置
  - 整列/整行范围格式化 (A:A, 1:1)
  - number_format 日期/时间/百分比/科学计数法
  - text_rotation=45 对角文本
  - font_name 中文（微软雅黑）
  - formatting 含未知键不报错（透传兼容性）
  - openpyxl Color 对象作为 font_color 值
  - 多 cell_range 格式化后逐个验证一致性
"""

import os
import pytest
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from excel_mcp_server_fastmcp.core.excel_writer import ExcelWriter
from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations


# ==================== Helper ====================

def _create_test_xlsx(file_path: str, rows: int = 5, cols: int = 4, sheet_name: str = "Sheet1"):
    """创建测试用 xlsx 文件，含数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=r * 10 + c)
    wb.save(file_path)
    wb.close()


def _read_cell_style(file_path: str, cell_ref: str = "A1", sheet_name: str = "Sheet1") -> dict:
    """读取单元格的样式属性用于断言"""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    cell = ws[cell_ref]
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    result = {
        "bold": font.bold,
        "italic": font.italic,
        "underline": font.underline,
        "size": font.size,
        "font_name": font.name,
        "color": str(font.color) if font.color else None,
        "fill_type": fill.fill_type if fill else None,
        "fgColor": str(fill.fgColor) if fill and fill.fgColor else None,
        "alignment_h": alignment.horizontal,
        "alignment_v": alignment.vertical,
        "wrap_text": alignment.wrap_text,
        "text_rotation": alignment.text_rotation,
        "number_format": cell.number_format,
        "border_left": border.left.style if border and border.left else None,
        "border_right": border.right.style if border and border.right else None,
        "border_top": border.top.style if border and border.top else None,
        "border_bottom": border.bottom.style if border and border.bottom else None,
        "value": cell.value,
        "indent": alignment.indent,
    }
    wb.close()
    return result


# ==================== Test Class ====================

class TestFormatCellsR57:
    """format_cells R57 第四轮测试套件"""

    # ---------- 1. bg_color / font_color 带 # 前缀 ----------

    def test_bg_color_with_hash_prefix(self, tmp_path):
        """bg_color='#FF0000' 带 # 前缀应被自动清理"""
        fp = str(tmp_path / "bg_hash.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bg_color": "#FF0000"})
        assert result["success"] is True, f"Expected success, got: {result}"
        style = _read_cell_style(fp, "A1")
        assert style["fill_type"] == "solid"

    def test_font_color_with_hash_prefix(self, tmp_path):
        """font_color='#00FF00' 带 # 前缀应被自动清理"""
        fp = str(tmp_path / "fc_hash.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_color": "#0000FF"})
        assert result["success"] is True, f"Expected success, got: {result}"

    def test_bg_color_hash_upper_lower(self, tmp_path):
        """bg_color='#ff0000' 小写 # 前缀"""
        fp = str(tmp_path / "bg_hash_low.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bg_color": "#ff0000"})
        assert result["success"] is True, f"Expected success, got: {result}"

    # ---------- 2. 中文 Sheet 名 ----------

    def test_format_chinese_sheet_name(self, tmp_path):
        """对中文 Sheet 名进行格式化"""
        fp = str(tmp_path / "cn_sheet.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "装备表"
        for r in range(1, 4):
            for c in range(1, 4):
                ws.cell(row=r, column=c, value=r * 10 + c)
        wb.save(fp)
        wb.close()

        result = ExcelOperations.format_cells(fp, "装备表", "A1:B2",
            formatting={"bold": True, "bg_color": "FFFF00"})
        assert result["success"] is True, f"Chinese sheet format failed: {result}"
        style = _read_cell_style(fp, "A1", sheet_name="装备表")
        assert style["bold"] is True

    def test_format_chinese_sheet_merge_and_format(self, tmp_path):
        """中文 Sheet 名 + 合并 + 格式化组合"""
        fp = str(tmp_path / "cn_sheet_merge.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "数据表"
        ws.append(["名称", "数值", "备注"])
        ws.append(["测试", 100, "OK"])
        wb.save(fp)
        wb.close()

        from excel_mcp_server_fastmcp.server import excel_format_cells
        r = excel_format_cells(fp, "数据表", "A1:C1",
            formatting={"merge": True, "bold": True, "alignment": "center"})
        assert r["success"] is True, f"Failed: {r}"

    # ---------- 3. 全组合操作 ----------

    def test_full_combo_merge_bold_bg_alignment_border(self, tmp_path):
        """merge + bold + bg_color + alignment + border_style 五合一"""
        fp = str(tmp_path / "full_combo.xlsx")
        _create_test_xlsx(fp)
        from excel_mcp_server_fastmcp.server import excel_format_cells
        r = excel_format_cells(fp, "Sheet1", "A1:D1",
            formatting={
                "merge": True,
                "bold": True,
                "bg_color": "4472C4",
                "alignment": "center",
                "border_style": "medium",
            })
        assert r["success"] is True, f"Full combo failed: {r}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["alignment_h"] == "center"
        assert style["border_top"] == "medium"

    def test_full_combo_with_font_details(self, tmp_path):
        """完整字体+填充+对齐+数字格式+边框（不含 merge）"""
        fp = str(tmp_path / "full_nomerge.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:C3",
            formatting={
                "bold": True,
                "italic": True,
                "underline": "double",
                "font_size": 14,
                "font_name": "Arial",
                "font_color": "FFFFFF",
                "bg_color": "000000",
                "alignment": "center",
                "vertical_alignment": "middle",
                "wrap_text": True,
                "number_format": "0.00%",
                "border": {"top": "thin", "bottom": "thick", "left": "medium", "right": "dashed"},
            })
        assert result["success"] is True, f"Full no-merge combo failed: {result}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["italic"] is True
        assert style["size"] == 14
        assert style["alignment_h"] == "center"
        assert "%" in style["number_format"]

    # ---------- 4. unmerge 数据完整性 ----------

    def test_unmerge_preserves_topleft_data(self, tmp_path):
        """合并后再拆分，左上角数据应保留"""
        fp = str(tmp_path / "unmerge_data.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "header_a"
        ws["B1"] = "header_b"
        ws["C1"] = "header_c"
        wb.save(fp)
        wb.close()

        # 合并
        r1 = ExcelOperations.merge_cells(fp, "Sheet1", "A1:C1")
        assert r1["success"] is True

        # 拆分
        r2 = ExcelOperations.unmerge_cells(fp, "Sheet1", "A1:C1")
        assert r2["success"] is True

        # 验证左上角数据还在
        wb2 = load_workbook(fp)
        ws2 = wb2["Sheet1"]
        assert ws2["A1"].value == "header_a", f"Expected 'header_a', got {ws2['A1'].value}"
        wb2.close()

    def test_unmerge_then_reformat(self, tmp_path):
        """拆分后重新格式化各单元格应正常工作"""
        fp = str(tmp_path / "unmerge_refmt.xlsx")
        _create_test_xlsx(fp, rows=3, cols=3)
        ExcelOperations.merge_cells(fp, "Sheet1", "A1:C1")
        ExcelOperations.unmerge_cells(fp, "Sheet1", "A1:C1")
        # 拆分后分别设置不同格式
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": True})
        assert result["success"] is True
        result2 = ExcelOperations.format_cells(fp, "Sheet1", "B1",
            formatting={"italic": True, "bg_color": "FF0000"})
        assert result2["success"] is True
        style_a = _read_cell_style(fp, "A1")
        style_b = _read_cell_style(fp, "B1")
        assert style_a["bold"] is True
        assert style_b["italic"] is True

    # ---------- 5. vertical_alignment 扁平格式 ----------

    def test_vertical_alignment_top(self, tmp_path):
        """vertical_alignment='top'"""
        fp = str(tmp_path / "valign_top.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"vertical_alignment": "top"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["alignment_v"] == "top"

    def test_vertical_alignment_bottom(self, tmp_path):
        """vertical_alignment='bottom'"""
        fp = str(tmp_path / "valign_bot.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"vertical_alignment": "bottom"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["alignment_v"] == "bottom"

    def test_vertical_alignment_center(self, tmp_path):
        """vertical_alignment='center'（垂直居中）"""
        fp = str(tmp_path / "valign_center.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"vertical_alignment": "center", "alignment": "center"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["alignment_v"] == "center"
        assert style["alignment_h"] == "center"

    # ---------- 6. border 详细配置 ----------

    def test_border_per_side_with_colors(self, tmp_path):
        """边框四边不同颜色"""
        fp = str(tmp_path / "border_colors.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={
                "border": {
                    "left": {"style": "thin", "color": "FF0000"},
                    "right": {"style": "medium", "color": "00FF00"},
                    "top": {"style": "thick", "color": "0000FF"},
                    "bottom": {"style": "dashed", "color": "FFFF00"},
                }
            })
        assert result["success"] is True

    def test_border_only_two_sides(self, tmp_path):
        """只设两边边框（bottom + right），其他边保持默认"""
        fp = str(tmp_path / "border_2side.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={
                "border": {
                    "bottom": "double",
                    "right": "medium",
                }
            })
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["border_bottom"] == "double"
        assert style["border_right"] == "medium"

    # ---------- 7. 整列/整行范围 ----------

    def test_format_entire_column(self, tmp_path):
        """格式化整列 A:A"""
        fp = str(tmp_path / "fmt_col.xlsx")
        _create_test_xlsx(fp, rows=10, cols=4)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A:A",
            formatting={"bold": True})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        style_a5 = _read_cell_style(fp, "A5")
        assert style_a5["bold"] is True

    def test_format_entire_row(self, tmp_path):
        """格式化整行 1:1"""
        fp = str(tmp_path / "fmt_row.xlsx")
        _create_test_xlsx(fp, rows=5, cols=6)
        result = ExcelOperations.format_cells(fp, "Sheet1", "1:1",
            formatting={"bg_color": "D9D9D9"})
        assert result["success"] is True
        style_a1 = _read_cell_style(fp, "A1")
        assert style_a1["fill_type"] == "solid"
        style_d1 = _read_cell_style(fp, "D1")
        assert style_d1["fill_type"] == "solid"

    # ---------- 8. number_format 特殊格式 ----------

    def test_number_format_date(self, tmp_path):
        """number_format 日期格式"""
        fp = str(tmp_path / "nf_date.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"number_format": "YYYY-MM-DD"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["number_format"] == "YYYY-MM-DD"

    def test_number_format_percent(self, tmp_path):
        """number_format 百分比格式"""
        fp = str(tmp_path / "nf_pct.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"number_format": "0.0%"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["number_format"] == "0.0%"

    def test_number_format_scientific(self, tmp_path):
        """number_format 科学计数法"""
        fp = str(tmp_path / "nf_sci.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"number_format": "0.00E+00"})
        assert result["success"] is True

    def test_number_format_currency(self, tmp_path):
        """number_format 货币格式"""
        fp = str(tmp_path / "nf_cur.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"number_format": "¥#,##0.00"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert "¥" in style["number_format"]

    def test_number_format_fraction(self, tmp_path):
        """number_format 分数格式"""
        fp = str(tmp_path / "nf_frac.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"number_format": "# ?/?"})
        assert result["success"] is True

    # ---------- 9. text_rotation 对角线 ----------

    def test_text_rotation_45_diagonal(self, tmp_path):
        """text_rotation=45 对角文本"""
        fp = str(tmp_path / "rot_45.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"text_rotation": 45})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 45

    # ---------- 10. 中文字体名 ----------

    def test_font_name_chinese_simhei(self, tmp_path):
        """中文字体：黑体"""
        fp = str(tmp_path / "font_simhei.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_name": "黑体"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["font_name"] is not None

    def test_font_name_chinese_simyahei(self, tmp_path):
        """中文字体：微软雅黑"""
        fp = str(tmp_path / "font_yahei.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_name": "微软雅黑"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["font_name"] is not None

    # ---------- 11. 未知键透传 ----------

    def test_unknown_keys_passthrough(self, tmp_path):
        """formatting 含未知键不应报错（向后兼容）"""
        fp = str(tmp_path / "unknown_keys.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": True, "custom_key_123": "should_be_ignored"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True

    # ---------- 12. 多单元格一致性验证 ----------

    def test_multi_cell_consistency(self, tmp_path):
        """范围格式化后所有单元格样式一致"""
        fp = str(tmp_path / "multi_consist.xlsx")
        _create_test_xlsx(fp, rows=4, cols=5)
        result = ExcelOperations.format_cells(fp, "Sheet1", "B2:D4",
            formatting={"bold": True, "italic": True, "bg_color": "00FF00"})
        assert result["success"] is True
        # 抽查几个单元格
        for ref in ["B2", "C3", "D4"]:
            style = _read_cell_style(fp, ref)
            assert style["bold"] is True, f"{ref} should be bold"
            assert style["italic"] is True, f"{ref} should be italic"

    # ---------- 13. pattern fill ----------

    def test_pattern_fill_gray125(self, tmp_path):
        """pattern fill: gray125（Excel 默认背景）"""
        fp = str(tmp_path / "pat_gray125.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"fill_type": "pattern"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        # pattern fill 的 fill_type 应为 pattern 或类似值
        assert style["fill_type"] is not None

    def test_pattern_fill_with_custom_fgcolor(self, tmp_path):
        """自定义图案前景色"""
        fp = str(tmp_path / "pat_custom.xlsx")
        _create_test_xlsx(fp)
        # 通过嵌套格式传入详细 pattern 配置
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {
                "type": "pattern",
                "patternType": "lightDown",
                "fgColor": "FF0000",
                "bgColor": "00FF00",
            }
        })
        assert result.success is True

    # ---------- 14. preset header 验证 ----------

    def test_preset_header_applies_correctly(self, tmp_path):
        """preset='header' 应应用：微软雅黑 11号 加粗 + 灰色背景"""
        fp = str(tmp_path / "preset_hdr.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            preset="header")
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["size"] == 11

    def test_preset_title_verification(self, tmp_path):
        """preset='title' 应应用：微软雅黑 14号 加粗 居中"""
        fp = str(tmp_path / "preset_title.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            preset="title")
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["size"] == 14
        assert style["alignment_h"] == "center"

    # ---------- 15. shrink_to_fit 单独使用 ----------

    def test_shrink_to_fit_true_alone(self, tmp_path):
        """shrink_to_fit=True 单独使用"""
        fp = str(tmp_path / "shrink_alone.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"shrink_to_fit": True})
        assert result["success"] is True

    # ---------- 16. double underline + accounting ----------

    def test_underline_double_accounting(self, tmp_path):
        """双会计下划线 doubleAccounting"""
        fp = str(tmp_path / "ul_dba.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"underline": "doubleAccounting"})
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["underline"] == "doubleAccounting"

    # ---------- 17. 超大范围格式化 ----------

    def test_large_range_formatting(self, tmp_path):
        """较大范围 (20x20) 格式化性能和正确性"""
        fp = str(tmp_path / "large_fmt.xlsx")
        _create_test_xlsx(fp, rows=25, cols=25)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:X25",
            formatting={"bold": True, "font_size": 9})
        assert result["success"] is True
        meta = result.get("metadata", {})
        assert meta.get("formatted_count", 0) > 0

    # ---------- 18. 合并区域部分重叠再格式化 ----------

    def test_merge_overlapping_regions(self, tmp_path):
        """两个相邻合并区域分别格式化互不影响"""
        fp = str(tmp_path / "overlap_merge.xlsx")
        _create_test_xlsx(fp, rows=3, cols=6)
        # 合并 A1-C1 和 D1-F1
        ExcelOperations.merge_cells(fp, "Sheet1", "A1:C1")
        ExcelOperations.merge_cells(fp, "Sheet1", "D1:F1")
        # 分别格式化
        r1 = ExcelOperations.format_cells(fp, "Sheet1", "A1:C1",
            formatting={"bg_color": "FF0000", "bold": True})
        r2 = ExcelOperations.format_cells(fp, "Sheet1", "D1:F1",
            formatting={"bg_color": "0000FF", "italic": True})
        assert r1["success"] is True
        assert r2["success"] is True
        style_a = _read_cell_style(fp, "A1")
        style_d = _read_cell_style(fp, "D1")
        assert style_a["bold"] is True
        assert style_d["italic"] is True
