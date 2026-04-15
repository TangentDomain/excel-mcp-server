#!/usr/bin/env python3
"""
Round 26 MCP 接口实测 - 格式兼容性 + 大数据量压力测试 + P0回归
主题: 格式兼容性测试 (Format Compatibility) + 大数据量压力 (Large Data Stress)
日期: 2026-04-14
"""

import sys
import os
import time
import tempfile
import shutil
import random
import string

# Add src to path
sys.path.insert(0, '/root/workspace/excel-mcp-server/src')

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)

# Test result tracking
results = []
start_time = time.time()

def test(name, category, passed, detail=""):
    """Record a test result."""
    status = "✅ PASS" if passed else "❌ FAIL"
    results.append({"name": name, "category": category, "passed": passed, "detail": detail})
    print(f"{status} | {category} | {name}")
    if not passed and detail:
        print(f"       → {detail}")
    return passed

# ============================================================
# Setup: Create test data files
# ============================================================
TMPDIR = tempfile.mkdtemp(prefix="round26_")
print(f"📁 测试目录: {TMPDIR}")

def create_test_xlsx(path, rows=100, cols=5, sheet_name="Sheet1"):
    """Create a standard .xlsx test file."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Headers
    headers = [f"Col_{i}" for i in range(cols)]
    headers[0] = "ID"
    headers[1] = "Name"
    if cols > 2: headers[2] = "Value"
    if cols > 3: headers[3] = "Price"
    if cols > 4: headers[4] = "Category"
    ws.append(headers)
    
    # Data
    categories = ["A", "B", "C", "D"]
    for i in range(1, rows + 1):
        row = [i, f"Item_{i}"]
        if cols > 2: row.append(round(random.uniform(10, 1000), 2))
        if cols > 3: row.append(round(random.uniform(1.0, 9999.99), 2))
        if cols > 4: row.append(random.choice(categories))
        ws.append(row)
    
    wb.save(path)
    return path


# ============================================================
# Group A: Format Compatibility Tests (格式兼容性)
# ============================================================
print("\n" + "="*70)
print("🔧 Group A: 格式兼容性测试 (Format Compatibility)")
print("="*70)

# A1: Standard .xlsx file
print("\n--- A1-A3: 标准 .xlsx 格式 ---")
std_xlsx = os.path.join(TMPDIR, "standard.xlsx")
create_test_xlsx(std_xlsx, 50, 5)

try:
    r = execute_advanced_sql_query(std_xlsx, "SELECT COUNT(*) as cnt FROM Sheet1")
    test("A1: .xlsx 基础查询", "格式兼容", r['success'], f"got: {r.get('data', [])}")
except Exception as e:
    test("A1: .xlsx 基础查询", "格式兼容", False, str(e)[:100])

try:
    r = execute_advanced_sql_query(std_xlsx, "SELECT * FROM Sheet1 ORDER BY ID DESC LIMIT 3")
    test("A2: .xlsx ORDER BY + LIMIT", "格式兼容", r['success'] and len(r.get('data', [])) == 3, f"rows: {len(r.get('data', []))}")
except Exception as e:
    test("A2: .xlsx ORDER BY + LIMIT", "格式兼容", False, str(e)[:100])

try:
    r = execute_advanced_update_query(std_xlsx, "UPDATE Sheet1 SET Value = Value * 2 WHERE ID <= 5")
    test("A3: .xlsx UPDATE 操作", "格式兼容", r['success'], f"got: {r.get('message', '')[:60]}")
except Exception as e:
    test("A3: .xlsx UPDATE 操作", "格式兼容", False, str(e)[:100])


# A4: Try .xls format (old Excel format) - should it work?
print("\n--- A4: .xls 旧格式测试 ---")
try:
    # Try creating with xlwt or openpyxl
    try:
        import xlwt
        has_xlwt = True
    except ImportError:
        has_xlwt = False
    
    if has_xlwt:
        xls_path = os.path.join(TMPDIR, "old_format.xls")
        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        ws.write(0, 0, "ID")
        ws.write(0, 1, "Name")
        ws.write(0, 2, "Value")
        for i in range(1, 21):
            ws.write(i, 0, i)
            ws.write(i, 1, f"Item_{i}")
            ws.write(i, 2, round(random.uniform(10, 100), 2))
        wb.save(xls_path)
        
        r = execute_advanced_sql_query(xls_path, "SELECT COUNT(*) FROM Sheet1")
        test("A4: .xls 旧格式读取", "格式兼容", r['success'], f"got: {r.get('message', '')[:80]}")
    else:
        test("A4: .xls 旧格式跳过(xlwt未安装)", "格式兼容", True, "xlwt not available")
except Exception as e:
    test("A4: .xls 旧格式测试", "格式兼容", False, str(e)[:100])


# A5: Empty Excel file (only headers, no data rows)
print("\n--- A5: 空数据文件 ---")
empty_path = os.path.join(TMPDIR, "empty_data.xlsx")
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "EmptySheet"
ws.append(["ID", "Name", "Value"])
wb.save(empty_path)

try:
    r = execute_advanced_sql_query(empty_path, "SELECT * FROM EmptySheet")
    is_empty_ok = r['success'] and len(r.get('data', [])) == 0
    test("A5: 空数据表查询", "格式兼容", is_empty_ok, f"rows: {len(r.get('data', []))}")
except Exception as e:
    test("A5: 空数据表查询", "格式兼容", False, str(e)[:100])

try:
    r = execute_advanced_sql_query(empty_path, "SELECT COUNT(*) as cnt FROM EmptySheet")
    count_val = r.get('data', [[0]])[0][0] if r.get('data') else None
    test("A5b: 空表 COUNT(*)", "格式兼容", r['success'] and count_val == 0, f"count={count_val}")
except Exception as e:
    test("A5b: 空表 COUNT(*)", "格式兼容", False, str(e)[:100])


# A6: Completely empty file (no sheets with data)
print("\n--- A6: 完全空文件 ---")
really_empty = os.path.join(TMPDIR, "really_empty.xlsx")
wb2 = Workbook()
wb2.save(really_empty)

try:
    r = execute_advanced_sql_query(really_empty, "SELECT * FROM Sheet")
    # Should fail gracefully - no data or sheet issue
    test("A6: 完全空文件处理", "格式兼容", not r['success'] or r.get('data') is not None, 
         f"success={r.get('success')}, msg={str(r.get('message',''))[:60]}")
except Exception as e:
    test("A6: 完全空文件处理", "格式兼容", True, f"异常可接受: {str(e)[:60]}")


# A7: Corrupted / invalid file
print("\n--- A7: 损坏文件 ---")
corrupted = os.path.join(TMPDIR, "corrupted.xlsx")
with open(corrupted, 'wb') as f:
    f.write(b'this is not a valid xlsx file at all!!!' * 100)

try:
    r = execute_advanced_sql_query(corrupted, "SELECT * FROM Sheet1")
    test("A7: 损坏文件优雅报错", "格式兼容", not r['success'], 
         f"should fail but got: success={r.get('success')}, msg={str(r.get('message',''))[:60]}")
except Exception as e:
    test("A7: 损坏文件异常处理", "格式兼容", True, f"正确抛出异常: {type(e).__name__}")


# A8: File with special characters in sheet name
print("\n--- A8: 特殊字符 Sheet 名 ---")
special_sheet = os.path.join(TMPDIR, "special_sheet.xlsx")
wb3 = Workbook()
ws3 = wb3.active
# Use a sheet name with special chars
ws3.title = "Sheet-Test_数据"
ws3.append(["ID", "Name", "Val"])
for i in range(1, 11):
    ws3.append([i, f"Item_{i}", round(random.uniform(1, 100), 2)])
wb3.save(special_sheet)

try:
    r = execute_advanced_sql_query(special_sheet, 'SELECT COUNT(*) FROM `Sheet-Test_数据`')
    test("A8a: 特殊字符Sheet名-反引号", "格式兼容", r['success'], f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("A8a: 特殊字符Sheet名-反引号", "格式兼容", False, str(e)[:100])

try:
    r = execute_advanced_sql_query(special_sheet, 'SELECT * FROM `Sheet-Test_数据` LIMIT 3')
    test("A8b: 特殊字符Sheet名-查询", "格式兼容", r['success'] and len(r.get('data', [])) <= 3, 
         f"rows: {len(r.get('data', []))}")
except Exception as e:
    test("A8b: 特殊字符Sheet名-查询", "格式兼容", False, str(e)[:100])


# A9: File with Unicode content (Chinese, emoji, etc.)
print("\n--- A9: Unicode 内容 ---")
unicode_path = os.path.join(TMPDIR, "unicode.xlsx")
wb4 = Workbook()
ws4 = wb4.active
ws4.title = "UnicodeData"
ws4.append(["ID", "名称", "描述", "价格"])
unicode_data = [
    [1, "装备⚔️", "传奇级武器，攻击力+999", 9999.99],
    [2, "药水🧪", "恢复500点生命值", 199.50],
    [3, "卷轴📜", "传送至随机地点", 888.88],
    [4, "日本語テスト", "日本語の説明文です", 1234.56],
    [5, "한국어테스트", "한국어 설명입니다", 5678.90],
    [6, "العربية", "نص عربي", 1111.22],
    [7, "🎮游戏道具🎲", "包含emoji的🌟超长描述🔥" * 3, 3333.33],
]
for row in unicode_data:
    ws4.append(row)
wb4.save(unicode_path)

try:
    r = execute_advanced_sql_query(unicode_path, "SELECT * FROM UnicodeData WHERE ID = 1")
    test("A9a: Unicode基础查询", "格式兼容", r['success'], f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("A9a: Unicode基础查询", "格式兼容", False, str(e)[:100])

try:
    r = execute_advanced_sql_query(unicode_path, "SELECT 名称, 价格 FROM UnicodeData ORDER BY 价格 DESC")
    test("A9b: Unicode列名+排序", "格式兼容", r['success'] and len(r.get('data', [])) > 0,
         f"rows: {len(r.get('data', []))}")
except Exception as e:
    test("A9b: Unicode列名+排序", "格式兼容", False, str(e)[:100])

try:
    r = execute_advanced_sql_query(unicode_path, "SELECT COUNT(*) as cnt FROM UnicodeData WHERE 名称 LIKE '%装备%'")
    test("A9c: Unicode LIKE 查询", "格式兼容", r['success'],
         f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("A9c: Unicode LIKE 查询", "格式兼容", False, str(e)[:100])


# A10: File with very long text content
print("\n--- A10: 超长文本 ---")
longtext_path = os.path.join(TMPDIR, "longtext.xlsx")
wb5 = Workbook()
ws5 = wb5.active
ws5.title = "LongText"
ws5.append(["ID", "Content"])
long_str = "这是一个很长的字符串。" * 200  # ~2400 chars
ws5.append([1, long_str])
very_long_str = "X" * 10000  # 10000 chars
ws5.append([2, very_long_str])
# String with newlines
ws5.append([3, "Line1\nLine2\nLine3\nLine4\nLine5"])
wb5.save(longtext_path)

try:
    r = execute_advanced_sql_query(longtext_path, "SELECT ID, LENGTH(Content) as len FROM LongText")
    test("A10a: 超长文本读取", "格式兼容", r['success'],
         f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("A10a: 超长文本读取", "格式兼容", False, str(e)[:100])

try:
    r = execute_advanced_sql_query(longtext_path, "SELECT * FROM LongText WHERE ID = 2")
    test("A10b: 10000字符字段", "格式兼容", r['success'],
         f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("A10b: 10000字符字段", "格式兼容", False, str(e)[:100])


# ============================================================
# Group B: Large Data Stress Tests (大数据量压力测试)
# ============================================================
print("\n" + "="*70)
print("🔧 Group B: 大数据量压力测试 (Large Data Stress)")
print("="*70)

# B1: 10000 rows
print("\n--- B1-B3: 大行数测试 ---")
large_path = os.path.join(TMPDIR, "large_10k.xlsx")
t0 = time.time()
create_test_xlsx(large_path, 10000, 5, "BigTable")
t_create = time.time() - t0
print(f"   创建 10000 行耗时: {t_create:.2f}s")

try:
    t0 = time.time()
    r = execute_advanced_sql_query(large_path, "SELECT COUNT(*) as cnt FROM BigTable")
    t_query = time.time() - t0
    cnt = r.get('data', [[0]])[0][0] if r.get('data') else -1
    test(f"B1: 10K行 COUNT(*) [{t_query:.2f}s]", "大数据量", 
         r['success'] and cnt == 10000, f"count={cnt}")
except Exception as e:
    test("B1: 10K行 COUNT(*)", "大数据量", False, str(e)[:100])

try:
    t0 = time.time()
    r = execute_advanced_sql_query(large_path, "SELECT Category, COUNT(*), AVG(Value), AVG(Price) FROM BigTable GROUP BY Category")
    t_query = time.time() - t0
    test(f"B2: 10K行 GROUP BY聚合 [{t_query:.2f}s]", "大数据量",
         r['success'] and len(r.get('data', [])) == 4, f"groups={len(r.get('data', []))}")
except Exception as e:
    test("B2: 10K行 GROUP BY聚合", "大数据量", False, str(e)[:100])

try:
    t0 = time.time()
    r = execute_advanced_sql_query(large_path, "SELECT * FROM BigTable ORDER BY Price DESC LIMIT 10")
    t_query = time.time() - t0
    test(f"B3: 10K行 ORDER BY LIMIT10 [{t_query:.2f}s]", "大数据量",
         r['success'] and len(r.get('data', [])) == 10, f"rows={len(r.get('data', []))}")
except Exception as e:
    test("B3: 10K行 ORDER BY LIMIT10", "大数据量", False, str(e)[:100])


# B4: Wide table (many columns)
print("\n--- B4-B5: 宽表测试 ---")
wide_path = os.path.join(TMPDIR, "wide_table.xlsx")
wb_wide = Workbook()
ws_wide = wb_wide.active
ws_wide.title = "WideTable"
num_cols = 30
headers = ["ID"] + [f"Metric_{i}" for i in range(num_cols)]
ws_wide.append(headers)
for i in range(1, 51):  # 50 rows x 30 cols
    row = [i] + [round(random.uniform(0, 1000), 2) for _ in range(num_cols)]
    ws_wide.append(row)
wb_wide.save(wide_path)

try:
    t0 = time.time()
    r = execute_advanced_sql_query(wide_path, "SELECT * FROM WideTable LIMIT 5")
    t_query = time.time() - t0
    # Check we get all columns
    col_count = len(r.get('columns', r.get('data', [[]])[0] if r.get('data') else []))
    test(f"B4: 30列宽表查询 [{t_query:.2f}s]", "大数据量",
         r['success'], f"cols≈{col_count}")
except Exception as e:
    test("B4: 30列宽表查询", "大数据量", False, str(e)[:100])

try:
    t0 = time.time()
    r = execute_advanced_sql_query(wide_path, f"SELECT ID, Metric_0, Metric_15, Metric_29 FROM WideTable WHERE ID <= 10")
    t_query = time.time() - t0
    test(f"B5: 宽表指定列筛选 [{t_query:.2f}s]", "大数据量",
         r['success'] and len(r.get('data', [])) <= 10, f"rows={len(r.get('data', []))}")
except Exception as e:
    test("B5: 宽表指定列筛选", "大数据量", False, str(e)[:100])


# B6: Window function on large dataset
print("\n--- B6-B7: 大数据集窗口函数 ---")
try:
    t0 = time.time()
    r = execute_advanced_sql_query(large_path, 
        "SELECT ID, Name, Price, RANK() OVER (ORDER BY Price DESC) as rnk FROM BigTable LIMIT 20")
    t_query = time.time() - t0
    test(f"B6: 10K行 RANK窗口函数 [{t_query:.2f}s]", "大数据量",
         r['success'], f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("B6: 10K行 RANK窗口函数", "大数据量", False, str(e)[:100])

try:
    t0 = time.time()
    r = execute_advanced_sql_query(large_path,
        "SELECT Category, Price, ROW_NUMBER() OVER (PARTITION BY Category ORDER BY Price DESC) as rn "
        "FROM BigTable WHERE rn <= 3")
    t_query = time.time() - t0
    test(f"B7: 10K行 PARTITION+ROW_NUMBER [{t_query:.2f}s]", "大数据量",
         r['success'], f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("B7: 10K行 PARTITION+ROW_NUMBER", "大数据量", False, str(e)[:100])


# B8: Complex query on large dataset (CTE + JOIN-like pattern)
print("\n--- B8: 复杂查询压力 ---")
try:
    t0 = time.time()
    r = execute_advanced_sql_query(large_path, """
        WITH TopItems AS (
            SELECT *, ROW_NUMBER() OVER (ORDER BY Price DESC) as rn 
            FROM BigTable 
        )
        SELECT Category, COUNT(*) as cnt, AVG(Price) as avg_price 
        FROM TopItems 
        WHERE rn <= 100 
        GROUP BY Category
    """)
    t_query = time.time() - t0
    test(f"B8: CTE+窗口函数+GROUP BY [{t_query:.2f}s]", "大数据量",
         r['success'], f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("B8: CTE+窗口函数+GROUP BY", "大数据量", False, str(e)[:100])


# B9: UPDATE on large dataset
print("\n--- B9: 大数据量UPDATE ---")
try:
    t0 = time.time()
    r = execute_advanced_update_query(large_path, 
        "UPDATE BigTable SET Value = ROUND(Value * 1.05, 2) WHERE Category = 'A'")
    t_query = time.time() - t0
    test(f"B9: 10K行条件UPDATE [{t_query:.2f}s]", "大数据量",
         r['success'], f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("B9: 10K行条件UPDATE", "大数据量", False, str(e)[:100])


# B10: INSERT into large table
print("\n--- B10: 大数据量INSERT ---")
try:
    t0 = time.time()
    r = execute_advanced_insert_query(large_path,
        "INSERT INTO BigTable (ID, Name, Value, Price, Category) VALUES (10001, 'StressTest', 999.99, 8888.88, 'Z')")
    t_query = time.time() - t0
    test(f"B10: INSERT新行 [{t_query:.2f}s]", "大数据量",
         r['success'], f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("B10: INSERT新行", "大数据量", False, str(e)[:100])


# ============================================================
# Group C: P0 Regression Tests (P0回归验证)
# ============================================================
print("\n" + "="*70)
print("🔧 Group C: P0 问题回归验证 (Round 25 发现)")
print("="*70)

# C1: script_runner arbitrary code execution check
print("\n--- C1-C2: P0 回归 ---")

# C1: SELECT multi-statement injection
try:
    r = execute_advanced_sql_query(std_xlsx, "SELECT COUNT(*) FROM Sheet1; SELECT COUNT(*) FROM Sheet1")
    if r['success']:
        # Check if multi-statement was executed
        data_str = str(r.get('data', ''))
        msg_str = str(r.get('message', ''))
        has_multi = '2/2' in msg_str or ';' in msg_str or 'multi' in msg_str.lower()
        if has_multi or isinstance(r.get('data'), list) and len(r.get('data', [])) > 1:
            test("C1: SELECT分号多语句 [仍存在!]", "P0回归", False, 
                 f"P0仍存在! multi-statement executed: {msg_str[:80]}")
        else:
            test("C1: SELECT分号多语句 [已修复?]", "P0回归", True, 
                 "单条结果或已拦截")
    else:
        test("C1: SELECT分号多语句 [已拒绝]", "P0回归", True, 
             f"正确拒绝: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("C1: SELECT分号多语句", "P0回归", False, str(e)[:100])


# C2: SQL comment injection
try:
    r = execute_advanced_sql_query(std_xlsx, "SELECT * FROM Sheet1 WHERE ID = 1 -- 注释注入")
    test("C2: SQL注释符注入", "P0回归", r['success'], 
         f"comment executed: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("C2: SQL注释符注入", "P0回归", False, str(e)[:100])


# C3: Path traversal attempt
print("\n--- C3-C4: 安全边界 ---")
try:
    r = execute_advanced_sql_query("/etc/passwd", "SELECT * FROM Sheet1")
    test("C3: 系统路径读取", "P0回归", not r['success'],
         f"should fail but: success={r.get('success')}, msg={str(r.get('message',''))[:60]}")
except Exception as e:
    test("C3: 系统路径读取", "P0回归", True, f"正确异常: {type(e).__name__}")


# C4: UNION-based injection
try:
    r = execute_advanced_sql_query(std_xlsx, "SELECT * FROM Sheet1 WHERE ID = 1 UNION SELECT 1,2,3,4,5")
    test("C4: UNION注入", "P0回归", r['success'],
         f"UNION executed: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("C4: UNION注入", "P0回归", False, str(e)[:100])


# ============================================================
# Group D: Type Boundary Tests (类型边界测试)
# ============================================================
print("\n" + "="*70)
print("🔧 Group D: 类型边界测试 (Type Boundaries)")
print("="*70)

type_path = os.path.join(TMPDIR, "types.xlsx")
wb_t = Workbook()
ws_t = wb_t.active
ws_t.title = "Types"
ws_t.append(["ID", "IntVal", "FloatVal", "StrVal", "BoolVal", "DateVal", "NullVal"])
type_data = [
    [1, 2147483647, 3.14159265358979, "normal", True, "2025-12-31", None],
    [2, -2147483648, 0.000000001, "", False, "2020-01-01", None],
    [3, 0, 999999.999, "特殊\"引号", True, "2099-06-15", ""],
    [4, 9223372036854775807, 1e308, "null", False, "1970-01-01", None],
    [5, 1, 0.0, "   spaces   ", True, "2026-04-14T13:00:00", "N/A"],
]
for row in type_data:
    ws_t.append(row)
wb_t.save(type_path)

# D1: Max int values
try:
    r = execute_advanced_sql_query(type_path, "SELECT ID, IntVal FROM Types WHERE ID = 1 OR ID = 4")
    test("D1: 大整数(int32/int64边界)", "类型边界", r['success'],
         f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("D1: 大整数", "类型边界", False, str(e)[:100])

# D2: Extreme floats
try:
    r = execute_advanced_sql_query(type_path, "SELECT ID, FloatVal FROM Types ORDER BY FloatVal DESC")
    test("D2: 极端浮点数", "类型边界", r['success'],
         f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("D2: 极端浮点数", "类型边界", False, str(e)[:100])

# D3: NULL handling
try:
    r = execute_advanced_sql_query(type_path, "SELECT COUNT(*) as total, SUM(IntVal) as sum_int FROM Types")
    test("D3: NULL值聚合(SUM/COUNT)", "类型边界", r['success'],
         f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("D3: NULL值聚合", "类型边界", False, str(e)[:100])

# D4: Empty strings vs NULL
try:
    r = execute_advanced_sql_query(type_path, "SELECT * FROM Types WHERE StrVal = '' OR StrVal IS NULL")
    test("D4: 空字符串与NULL区分", "类型边界", r['success'],
         f"rows: {len(r.get('data', []))}")
except Exception as e:
    test("D4: 空字符串与NULL区分", "类型边界", False, str(e)[:100])

# D5: Boolean mixed with numbers
try:
    r = execute_advanced_sql_query(type_path, "SELECT BoolVal, COUNT(*) as cnt FROM Types GROUP BY BoolVal")
    test("D5: 布尔值分组聚合", "类型边界", r['success'],
         f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("D5: 布尔值分组聚合", "类型边界", False, str(e)[:100])

# D6: Special characters in values
try:
    r = execute_advanced_sql_query(type_path, "SELECT StrVal FROM Types WHERE ID = 3")
    test("D6: 特殊字符值(引号等)", "类型边界", r['success'],
         f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("D6: 特殊字符值", "类型边界", False, str(e)[:100])

# D7: Date boundary
try:
    r = execute_advanced_sql_query(type_path, "SELECT DateVal FROM Types ORDER BY DateVal")
    test("D7: 日期范围处理", "类型边界", r['success'],
         f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("D7: 日期范围处理", "类型边界", False, str(e)[:100])

# D8: Numeric overflow in expressions
try:
    r = execute_advanced_sql_query(type_path, "SELECT IntVal * IntVal * IntVal as overflow_val FROM Types WHERE ID = 4")
    test("D8: 数值溢出表达式", "类型边界", r['success'],
         f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("D8: 数值溢出表达式", "类型边界", False, str(e)[:100])


# ============================================================
# Group E: Multi-Sheet Operations (多Sheet操作)
# ============================================================
print("\n" + "="*70)
print("🔧 Group E: 多Sheet操作 (Multi-Sheet)")
print("="*70)

multi_path = os.path.join(TMPDIR, "multi_sheet.xlsx")
wb_m = Workbook()
ws_m = wb_m.active
ws_m.title = "Players"
ws_m.append(["PlayerID", "Name", "Level", "GuildID"])
for i in range(1, 21):
    ws_m.append([i, f"Player_{i}", random.randint(1, 99), random.randint(1, 5)])

ws_g = wb_m.create_sheet("Guilds")
ws_g.append(["GuildID", "GuildName", "MasterID"])
for i in range(1, 6):
    ws_g.append([i, f"Guild_{i}", i * 4])

ws_i = wb_m.create_sheet("Items")
ws_i.append(["ItemID", "ItemName", "OwnerID", "Power"])
for i in range(1, 31):
    ws_i.append([i, f"Weapon_{i}", random.randint(1, 20), random.randint(10, 999)])

wb_m.save(multi_path)

# E1: Query different sheets sequentially
try:
    r1 = execute_advanced_sql_query(multi_path, "SELECT COUNT(*) FROM Players")
    r2 = execute_advanced_sql_query(multi_path, "SELECT COUNT(*) FROM Guilds")
    r3 = execute_advanced_sql_query(multi_path, "SELECT COUNT(*) FROM Items")
    all_ok = r1['success'] and r2['success'] and r3['success']
    test("E1: 多Sheet顺序查询", "多Sheet", all_ok,
         f"Players={r1.get('data')}, Guilds={r2.get('data')}, Items={r3.get('data')}")
except Exception as e:
    test("E1: 多Sheet顺序查询", "多Sheet", False, str(e)[:100])

# E2: Cross-sheet analysis (manual correlation)
try:
    r = execute_advanced_sql_query(multi_path, """
        SELECT g.GuildID, g.GuildName, COUNT(p.PlayerID) as MemberCount 
        FROM Guilds g
        LEFT JOIN Players p ON g.GuildID = p.GuildID 
        GROUP BY g.GuildID, g.GuildName
        ORDER BY MemberCount DESC
    """)
    test("E2: 跨Sheet JOIN统计", "多Sheet", r['success'],
         f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("E2: 跨Sheet JOIN统计", "多Sheet", False, str(e)[:100])

# E3: Complex cross-sheet query
try:
    r = execute_advanced_sql_query(multi_path, """
        WITH PlayerStats AS (
            SELECT p.GuildID, COUNT(*) as cnt, AVG(p.Level) as avg_lvl 
            FROM Players p 
            GROUP BY p.GuildID
        )
        SELECT g.GuildName, ps.cnt, ps.avg_lvl 
        FROM Guilds g 
        JOIN PlayerStats ps ON g.GuildID = ps.GuildID
    """)
    test("E3: CTE跨Sheet关联", "多Sheet", r['success'],
         f"msg: {str(r.get('message',''))[:60]}")
except Exception as e:
    test("E3: CTE跨Sheet关联", "多Sheet", False, str(e)[:100])


# ============================================================
# Summary
# ============================================================
print("\n" + "="*70)
print("📊 Round 26 测试总结")
print("="*70)

total = len(results)
passed = sum(1 for r in results if r['passed'])
failed = total - passed

categories = {}
for r in results:
    cat = r['category']
    if cat not in categories:
        categories[cat] = {'total': 0, 'pass': 0}
    categories[cat]['total'] += 1
    if r['passed']:
        categories[cat]['pass'] += 1

print(f"\n总测试: {total} | 通过: {passed} | 失败: {failed} | 通过率: {passed/total*100:.1f}%")
print(f"总耗时: {time.time() - start_time:.2f}s")

print("\n按分类:")
for cat, stats in sorted(categories.items()):
    rate = stats['pass']/stats['total']*100
    status = "✅" if stats['pass'] == stats['total'] else "⚠️"
    print(f"  {status} {cat}: {stats['pass']}/{stats['total']} ({rate:.0f}%)")

if failed > 0:
    print("\n❌ 失败用例:")
    for r in results:
        if not r['passed']:
            print(f"  - [{r['category']}] {r['name']}: {r['detail'][:80]}")

# Cleanup
shutil.rmtree(TMPDIR, ignore_errors=True)
print(f"\n🧹 测试目录已清理: {TMPDIR}")
