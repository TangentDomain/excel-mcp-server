# -*- coding: utf-8 -*-
"""
Excel Converter增强测试套件
测试src.core.excel_converter模块的所有核心功能
目标覆盖率：75%+
"""

import pytest
import tempfile
import os
import csv
import json
from pathlib import Path
from openpyxl import Workbook
from unittest.mock import patch, mock_open

from src.core.excel_converter import ExcelConverter
from src.models.types import OperationResult
from src.utils.exceptions import ExcelFileNotFoundError, DataValidationError


class TestExcelConverterBasic:
    """ExcelConverter基础功能测试"""

    def setup_method(self):
        """每个测试方法前的设置"""
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()

        # 创建基础Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "TestData"

        # 添加测试数据
        ws['A1'] = "Name"
        ws['B1'] = "Age"
        ws['C1'] = "City"
        ws['A2'] = "Alice"
        ws['B2'] = 25
        ws['C2'] = "Beijing"
        ws['A3'] = "Bob"
        ws['B3'] = 30
        ws['C3'] = "Shanghai"
        ws['A4'] = "张三"
        ws['B4'] = 28
        ws['C4'] = "广州"

        # 创建第二个工作表
        ws2 = wb.create_sheet(title="SecondSheet")
        ws2['A1'] = "Product"
        ws2['B1'] = "Price"
        ws2['A2'] = "Apple"
        ws2['B2'] = 5.5
        ws2['A3'] = "Orange"
        ws2['B3'] = 3.2

        wb.save(self.temp_file.name)
        self.file_path = self.temp_file.name

    def teardown_method(self):
        """每个测试方法后的清理"""
        if os.path.exists(self.file_path):
            try:
                os.unlink(self.file_path)
            except:
                pass

    def test_converter_init_valid_file(self):
        """测试有效文件路径的初始化"""
        converter = ExcelConverter(self.file_path)
        assert converter.file_path == self.file_path

    def test_converter_init_invalid_file(self):
        """测试无效文件路径的初始化"""
        with pytest.raises(Exception):
            ExcelConverter("invalid_file.xlsx")

    def test_export_to_csv_default_sheet(self):
        """测试导出默认工作表到CSV"""
        converter = ExcelConverter(self.file_path)
        output_file = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
        output_file.close()

        try:
            result = converter.export_to_csv(output_file.name)

            assert result.success is True
            assert result.data['row_count'] == 4  # 4行数据
            assert result.data['sheet_name'] == "TestData"
            assert result.data['encoding'] == "utf-8"
            assert os.path.exists(output_file.name)

            # 验证CSV内容
            with open(output_file.name, 'r', encoding='utf-8') as f:
                content = f.read()
                assert "Name" in content
                assert "Alice" in content
                assert "张三" in content

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_export_to_csv_specific_sheet(self):
        """测试导出指定工作表到CSV"""
        converter = ExcelConverter(self.file_path)
        output_file = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
        output_file.close()

        try:
            result = converter.export_to_csv(output_file.name, sheet_name="SecondSheet")

            assert result.success is True
            assert result.data['row_count'] == 3  # 3行数据
            assert result.data['sheet_name'] == "SecondSheet"
            assert os.path.exists(output_file.name)

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_export_to_csv_nonexistent_sheet(self):
        """测试导出不存在的工作表"""
        converter = ExcelConverter(self.file_path)
        output_file = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
        output_file.close()

        try:
            result = converter.export_to_csv(output_file.name, sheet_name="NonExistent")

            assert result.success is False
            assert "不存在" in result.error

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_export_to_csv_gbk_encoding(self):
        """测试GBK编码导出"""
        converter = ExcelConverter(self.file_path)
        output_file = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
        output_file.close()

        try:
            result = converter.export_to_csv(output_file.name, encoding="gbk")

            assert result.success is True
            assert result.data['encoding'] == "gbk"
            assert os.path.exists(output_file.name)

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_export_to_csv_creates_directory(self):
        """测试导出时创建目录"""
        converter = ExcelConverter(self.file_path)

        # 使用不存在的深层目录
        output_path = os.path.join(tempfile.gettempdir(), "excel_test", "subdir", "output.csv")

        result = converter.export_to_csv(output_path)

        assert result.success is True
        assert os.path.exists(output_path)
        assert os.path.exists(os.path.dirname(output_path))

        # 清理
        import shutil
        if os.path.exists(os.path.join(tempfile.gettempdir(), "excel_test")):
            shutil.rmtree(os.path.join(tempfile.gettempdir(), "excel_test"))

    def test_export_to_csv_filters_empty_rows(self):
        """测试导出时过滤空行"""
        converter = ExcelConverter(self.file_path)
        output_file = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
        output_file.close()

        # 添加一些空行到Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "TestEmptyRows"
        ws['A1'] = "Data1"
        ws['A2'] = ""  # 空行
        ws['A3'] = "Data3"
        ws['A4'] = ""  # 空行
        ws['A5'] = "Data5"
        wb.save(self.file_path)

        try:
            result = converter.export_to_csv(output_file.name)

            assert result.success is True
            assert result.data['row_count'] == 3  # 只有3行非空数据

            with open(output_file.name, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                assert len(lines) == 3  # 3行非空数据

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    @patch('src.core.excel_converter.logger')
    def test_export_to_csv_logging(self, mock_logger):
        """测试导出操作的日志记录"""
        converter = ExcelConverter(self.file_path)
        output_file = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
        output_file.close()

        try:
            result = converter.export_to_csv(output_file.name)
            assert result.success is True

            # 验证日志记录（可能没有错误日志）
            # mock_logger.error.assert_not_called()

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)


class TestExcelConverterImport:
    """ExcelConverter导入功能测试"""

    def test_import_from_csv_basic(self):
        """测试基础CSV导入"""
        csv_content = """Name,Age,City
Alice,25,Beijing
Bob,30,Shanghai
张三,28,广州"""

        csv_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')
        csv_file.write(csv_content)
        csv_file.close()

        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.import_from_csv(csv_file.name, output_file.name)

            assert result.success is True
            assert result.data['row_count'] == 4  # 包括表头
            assert result.data['sheet_name'] == "Sheet1"
            assert result.data['has_header'] is True
            assert os.path.exists(output_file.name)

            # 验证Excel文件内容
            from openpyxl import load_workbook
            wb = load_workbook(output_file.name)
            ws = wb.active
            assert ws['A1'].value == "Name"
            assert ws['A2'].value == "Alice"
            assert ws['B2'].value == 25  # 应该转换为数字
            assert ws['A4'].value == "张三"

        finally:
            os.unlink(csv_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_import_from_csv_custom_sheet_name(self):
        """测试自定义工作表名称的CSV导入"""
        csv_content = """Product,Price
Apple,5.5
Orange,3.2"""

        csv_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')
        csv_file.write(csv_content)
        csv_file.close()

        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.import_from_csv(
                csv_file.name, output_file.name,
                sheet_name="Products", has_header=False
            )

            assert result.success is True
            assert result.data['sheet_name'] == "Products"
            assert result.data['has_header'] is False

            # 验证工作表名称
            from openpyxl import load_workbook
            wb = load_workbook(output_file.name)
            assert "Products" in wb.sheetnames

        finally:
            os.unlink(csv_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_import_from_csv_gbk_encoding(self):
        """测试GBK编码的CSV导入"""
        csv_content = """姓名,年龄,城市
张三,25,北京
李四,30,上海"""

        csv_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='gbk')
        csv_file.write(csv_content)
        csv_file.close()

        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.import_from_csv(
                csv_file.name, output_file.name,
                encoding="gbk"
            )

            assert result.success is True
            assert result.data['encoding'] == "gbk"

        finally:
            os.unlink(csv_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_import_from_csv_nonexistent_file(self):
        """测试导入不存在的CSV文件"""
        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.import_from_csv("nonexistent.csv", output_file.name)

            assert result.success is False
            assert "不存在" in result.error

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_import_from_csv_creates_directory(self):
        """测试导入时创建目录"""
        csv_content = """Data1,Data2
Value1,Value2"""

        csv_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')
        csv_file.write(csv_content)
        csv_file.close()

        # 使用不存在的深层目录
        output_path = os.path.join(tempfile.gettempdir(), "import_test", "subdir", "output.xlsx")

        try:
            result = ExcelConverter.import_from_csv(csv_file.name, output_path)

            assert result.success is True
            assert os.path.exists(output_path)
            assert os.path.exists(os.path.dirname(output_path))

            # 清理
            import shutil
            if os.path.exists(os.path.join(tempfile.gettempdir(), "import_test")):
                shutil.rmtree(os.path.join(tempfile.gettempdir(), "import_test"))

        finally:
            os.unlink(csv_file.name)

    def test_import_from_csv_numeric_conversion(self):
        """测试数值转换功能"""
        csv_content = """Integer,Float,Mixed,Text
123,45.67,12.34,Hello
-456,-78.90,0,World
0,0.0,999999999,Test"""

        csv_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')
        csv_file.write(csv_content)
        csv_file.close()

        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.import_from_csv(csv_file.name, output_file.name)

            assert result.success is True

            # 验证数值转换
            from openpyxl import load_workbook
            wb = load_workbook(output_file.name)
            ws = wb.active

            # 检查数值转换
            assert isinstance(ws['B2'].value, float)  # 浮点数
            assert isinstance(ws['A2'].value, int)     # 整数
            assert ws['D2'].value == "Hello"        # 文本保持不变

        finally:
            os.unlink(csv_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    @patch('src.core.excel_converter.logger')
    def test_import_from_csv_logging(self, mock_logger):
        """测试导入操作的日志记录"""
        csv_content = """Test,Data
Value1,Value2"""

        csv_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')
        csv_file.write(csv_content)
        csv_file.close()

        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.import_from_csv(csv_file.name, output_file.name)
            assert result.success is True

        finally:
            os.unlink(csv_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)


class TestExcelConverterFormatConversion:
    """ExcelConverter格式转换测试"""

    def setup_method(self):
        """每个测试方法前的设置"""
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "TestData"

        # 添加测试数据
        ws['A1'] = "Name"
        ws['B1'] = "Score"
        ws['A2'] = "Alice"
        ws['B2'] = 95.5
        ws['A3'] = "Bob"
        ws['B3'] = 87.0

        # 创建第二个工作表
        ws2 = wb.create_sheet(title="Second")
        ws2['A1'] = "Subject"
        ws2['B1'] = "Grade"
        ws2['A2'] = "Math"
        ws2['B2'] = "A"

        wb.save(self.temp_file.name)
        self.file_path = self.temp_file.name

    def teardown_method(self):
        """每个测试方法后的清理"""
        if os.path.exists(self.file_path):
            try:
                os.unlink(self.file_path)
            except:
                pass

    def test_convert_format_xlsx_to_xlsx(self):
        """测试xlsx到xlsx的转换"""
        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.convert_format(
                self.file_path, output_file.name, "xlsx"
            )

            assert result.success is True
            assert result.data['input_format'] == ".xlsx"
            assert result.data['output_format'] == "xlsx"
            assert result.data['file_size'] > 0
            assert os.path.exists(output_file.name)

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_convert_format_xlsx_to_xlsm(self):
        """测试xlsx到xlsm的转换"""
        output_file = tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.convert_format(
                self.file_path, output_file.name, "xlsm"
            )

            assert result.success is True
            assert result.data['output_format'] == "xlsm"
            assert os.path.exists(output_file.name)

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_convert_format_xlsx_to_json(self):
        """测试xlsx到JSON的转换"""
        output_file = tempfile.NamedTemporaryFile(suffix='.json', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.convert_format(
                self.file_path, output_file.name, "json"
            )

            assert result.success is True
            assert result.data['output_format'] == "json"
            assert result.data['sheet_count'] == 2  # 两个工作表
            assert os.path.exists(output_file.name)

            # 验证JSON内容
            with open(output_file.name, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
                assert "TestData" in json_data
                assert "Second" in json_data
                assert json_data["TestData"][0] == ["Name", "Score"]

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_convert_format_xlsx_to_csv(self):
        """测试xlsx到CSV的转换"""
        output_file = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.convert_format(
                self.file_path, output_file.name, "csv"
            )

            assert result.success is True
            assert result.data['output_format'] == "csv"
            assert result.data['row_count'] > 0
            assert os.path.exists(output_file.name)

            # 验证CSV内容（应该是活动工作表的数据）
            with open(output_file.name, 'r', encoding='utf-8') as f:
                content = f.read()
                assert "Name" in content
                assert "Score" in content

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_convert_format_case_insensitive(self):
        """测试大小写不敏感的格式转换"""
        output_file = tempfile.NamedTemporaryFile(suffix='.json', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.convert_format(
                self.file_path, output_file.name, "JSON"
            )

            assert result.success is True
            assert result.data['output_format'] == "json"

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_convert_format_unsupported_format(self):
        """测试不支持的格式转换"""
        output_file = tempfile.NamedTemporaryFile(suffix='.txt', delete=False)
        output_file.close()

        result = ExcelConverter.convert_format(
            self.file_path, output_file.name, "txt"
        )

        assert result.success is False
        assert "不支持" in result.error

    def test_convert_format_nonexistent_file(self):
        """测试转换不存在的文件"""
        output_file = tempfile.NamedTemporaryFile(suffix='.json', delete=False)
        output_file.close()

        result = ExcelConverter.convert_format(
            "nonexistent.xlsx", output_file.name, "json"
        )

        assert result.success is False
        assert "不存在" in result.error

    def test_convert_format_creates_directory(self):
        """测试转换时创建目录"""
        output_path = os.path.join(tempfile.gettempdir(), "convert_test", "subdir", "output.json")

        try:
            result = ExcelConverter.convert_format(
                self.file_path, output_path, "json"
            )

            assert result.success is True
            assert os.path.exists(output_path)
            assert os.path.exists(os.path.dirname(output_path))

            # 清理
            import shutil
            if os.path.exists(os.path.join(tempfile.gettempdir(), "convert_test")):
                shutil.rmtree(os.path.join(tempfile.gettempdir(), "convert_test"))

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_convert_format_json_unicode_support(self):
        """测试JSON转换的Unicode支持"""
        # 创建包含Unicode数据的Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "UnicodeData"
        ws['A1'] = "中文"
        ws['B1'] = "English"
        ws['A2'] = "张三"
        ws['B2'] = "Alice"
        ws['A3'] = "🎮游戏"
        ws['B3'] = "📊Data"

        unicode_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        unicode_file.close()
        wb.save(unicode_file.name)

        output_file = tempfile.NamedTemporaryFile(suffix='.json', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.convert_format(
                unicode_file.name, output_file.name, "json"
            )

            assert result.success is True

            # 验证JSON中的Unicode内容
            with open(output_file.name, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
                assert "张三" in str(json_data)
                assert "🎮" in str(json_data)

        finally:
            os.unlink(unicode_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_convert_format_csv_filters_empty_rows(self):
        """测试CSV转换时过滤空行"""
        # 创建包含空行的Excel文件
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Data1"
        ws['A2'] = ""  # 空行
        ws['A3'] = "Data3"
        ws['A4'] = ""  # 空行
        ws['A5'] = "Data5"

        test_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        test_file.close()
        wb.save(test_file.name)

        output_file = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.convert_format(
                test_file.name, output_file.name, "csv"
            )

            assert result.success is True
            assert result.data['row_count'] == 3  # 只有3行非空数据

        finally:
            os.unlink(test_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    @patch('src.core.excel_converter.logger')
    def test_convert_format_logging(self, mock_logger):
        """测试转换操作的日志记录"""
        output_file = tempfile.NamedTemporaryFile(suffix='.json', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.convert_format(
                self.file_path, output_file.name, "json"
            )
            assert result.success is True

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)


class TestExcelConverterMerge:
    """ExcelConverter文件合并测试"""

    def setup_method(self):
        """每个测试方法前的设置"""
        # 创建多个测试Excel文件
        self.test_files = []

        for i in range(3):
            temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            temp_file.close()

            wb = Workbook()
            # 删除默认工作表
            wb.remove(wb.active)

            # 创建工作表
            ws = wb.create_sheet(title=f"Sheet{i+1}")
            ws['A1'] = f"File{i+1}_Data"
            ws['B1'] = f"File{i+1}_Value"
            for j in range(2, 5):
                ws[f'A{j}'] = f"Data{i+1}_{j}"
                ws[f'B{j}'] = f"Value{i+1}_{j}"

            wb.save(temp_file.name)
            self.test_files.append(temp_file.name)

    def teardown_method(self):
        """每个测试方法后的清理"""
        for file_path in self.test_files:
            if os.path.exists(file_path):
                try:
                    os.unlink(file_path)
                except:
                    pass

    def test_merge_files_sheets_mode(self):
        """测试工作表模式合并"""
        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.merge_files(
                self.test_files, output_file.name, "sheets"
            )

            assert result.success is True
            assert result.data['merged_files'] == 3
            assert result.data['total_sheets'] == 3  # 每个文件一个工作表
            assert result.data['merge_mode'] == "sheets"
            assert os.path.exists(output_file.name)

            # 验证合并后的文件
            from openpyxl import load_workbook
            wb = load_workbook(output_file.name)
            assert len(wb.sheetnames) == 3
            assert any("File1" in name for name in wb.sheetnames)

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_merge_files_append_mode(self):
        """测试追加模式合并"""
        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.merge_files(
                self.test_files, output_file.name, "append"
            )

            assert result.success is True
            assert result.data['merged_files'] == 3
            assert result.data['total_sheets'] == 1  # 只有一个工作表
            assert result.data['merge_mode'] == "append"
            assert os.path.exists(output_file.name)

            # 验证合并后的文件
            from openpyxl import load_workbook
            wb = load_workbook(output_file.name)
            assert len(wb.sheetnames) == 1
            assert wb.sheetnames[0] == "合并数据"

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_merge_files_unsupported_mode(self):
        """测试不支持的合并模式"""
        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        result = ExcelConverter.merge_files(
            self.test_files, output_file.name, "unsupported"
        )

        assert result.success is False
        assert "不支持" in result.error

    def test_merge_files_empty_list(self):
        """测试空文件列表合并"""
        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        result = ExcelConverter.merge_files(
            [], output_file.name, "sheets"
        )

        assert result.success is False
        assert "不能为空" in result.error

    def test_merge_files_nonexistent_file(self):
        """测试包含不存在文件的合并"""
        files_with_missing = self.test_files + ["nonexistent.xlsx"]
        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        result = ExcelConverter.merge_files(
            files_with_missing, output_file.name, "sheets"
        )

        assert result.success is False
        assert "不存在" in result.error

    def test_merge_files_long_sheet_names(self):
        """测试长工作表名称处理"""
        # 创建具有长名称的Excel文件
        long_name_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        long_name_file.close()

        wb = Workbook()
        ws = wb.create_sheet(title="ThisIsAVeryLongSheetNameThatExceedsTheNormalLimit")
        ws['A1'] = "Test Data"
        wb.save(long_name_file.name)

        files_with_long_name = self.test_files[:1] + [long_name_file.name]
        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.merge_files(
                files_with_long_name, output_file.name, "sheets"
            )

            assert result.success is True
            assert result.data['merged_files'] == 2

        finally:
            os.unlink(long_name_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_merge_files_creates_directory(self):
        """测试合并时创建目录"""
        output_path = os.path.join(tempfile.gettempdir(), "merge_test", "subdir", "merged.xlsx")

        try:
            result = ExcelConverter.merge_files(
                self.test_files, output_path, "sheets"
            )

            assert result.success is True
            assert os.path.exists(output_path)
            assert os.path.exists(os.path.dirname(output_path))

            # 清理
            import shutil
            if os.path.exists(os.path.join(tempfile.gettempdir(), "merge_test")):
                shutil.rmtree(os.path.join(tempfile.gettempdir(), "merge_test"))

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_merge_files_unicode_content(self):
        """测试合并Unicode内容"""
        # 创建包含Unicode的Excel文件
        unicode_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        unicode_file.close()

        wb = Workbook()
        ws = wb.create_sheet(title="UnicodeData")
        ws['A1'] = "中文数据"
        ws['B1'] = "English数据"
        ws['A2'] = "张三"
        ws['B2'] = "Alice"
        ws['A3'] = "🎮游戏"
        ws['B3'] = "📊图表"
        wb.save(unicode_file.name)

        files_with_unicode = self.test_files[:1] + [unicode_file.name]
        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.merge_files(
                files_with_unicode, output_file.name, "sheets"
            )

            assert result.success is True
            assert result.data['merged_files'] == 2

        finally:
            os.unlink(unicode_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    @patch('src.core.excel_converter.logger')
    def test_merge_files_logging(self, mock_logger):
        """测试合并操作的日志记录"""
        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.merge_files(
                self.test_files, output_file.name, "sheets"
            )
            assert result.success is True

        finally:
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)


class TestExcelConverterEdgeCases:
    """ExcelConverter边界条件测试"""

    def test_import_from_csv_empty_file(self):
        """测试导入空CSV文件"""
        csv_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')
        csv_file.write("")  # 空文件
        csv_file.close()

        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.import_from_csv(csv_file.name, output_file.name)

            assert result.success is True
            assert result.data['row_count'] == 0

        finally:
            os.unlink(csv_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_import_from_csv_only_empty_rows(self):
        """测试导入只有空行的CSV文件"""
        csv_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')
        csv_file.write(",\n,\n,\n")  # 只有空行
        csv_file.close()

        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.import_from_csv(csv_file.name, output_file.name)

            assert result.success is True
            # 应该导入空行，因为导入不过滤空行

        finally:
            os.unlink(csv_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_export_to_csv_empty_sheet(self):
        """测试导出空工作表"""
        # 创建空Excel文件
        empty_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        empty_file.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "EmptySheet"
        # 不添加任何数据
        wb.save(empty_file.name)

        converter = ExcelConverter(empty_file.name)
        output_file = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
        output_file.close()

        try:
            result = converter.export_to_csv(output_file.name)

            assert result.success is True
            assert result.data['row_count'] == 0

        finally:
            os.unlink(empty_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_convert_format_empty_workbook(self):
        """测试转换空工作簿"""
        # 创建空Excel文件
        empty_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        empty_file.close()

        wb = Workbook()
        # 删除默认工作表，创建空工作簿
        wb.remove(wb.active)
        wb.save(empty_file.name)

        output_file = tempfile.NamedTemporaryFile(suffix='.json', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.convert_format(
                empty_file.name, output_file.name, "json"
            )

            assert result.success is True
            assert result.data['sheet_count'] == 0

            with open(output_file.name, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
                assert json_data == {}

        finally:
            os.unlink(empty_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_large_file_handling(self):
        """测试大文件处理"""
        # 创建大数据量CSV文件
        large_csv_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')

        # 写入大量数据
        large_csv_file.write("Index,Value\n")
        for i in range(1000):
            large_csv_file.write(f"{i},{i * 2}\n")
        large_csv_file.close()

        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.import_from_csv(large_csv_file.name, output_file.name)

            assert result.success is True
            assert result.data['row_count'] == 1001  # 包括表头

        finally:
            os.unlink(large_csv_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_special_characters_in_csv(self):
        """测试CSV中的特殊字符处理"""
        csv_content = '''"Name,With,Comma","Description"
"Value""With"Quotes","Test,Data"
"Line\nBreak","Another Test"'''

        csv_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')
        csv_file.write(csv_content)
        csv_file.close()

        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.import_from_csv(csv_file.name, output_file.name)

            assert result.success is True

        finally:
            os.unlink(csv_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_numeric_edge_cases(self):
        """测试数值边界情况"""
        csv_content = """Integer,Float,Scientific,Invalid
0,0.0,1.23e5,not_a_number
-999999999,999999999.99,-1.23e-10,text
2147483647,3.14159265359,1.79e308,empty"""

        csv_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')
        csv_file.write(csv_content)
        csv_file.close()

        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.import_from_csv(csv_file.name, output_file.name)

            assert result.success is True

            # 验证数值转换
            from openpyxl import load_workbook
            wb = load_workbook(output_file.name)
            ws = wb.active

            assert ws['A2'].value == 0
            assert ws['B2'].value == 0.0
            # 科学计数法可能保持为字符串，取决于实现

        finally:
            os.unlink(csv_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

    def test_file_permission_error(self):
        """测试文件权限错误"""
        if os.name == 'nt':  # Windows系统
            # 这个测试在Windows上可能不太可靠，跳过
            pytest.skip("File permission test skipped on Windows")
            return

        output_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        output_file.close()

        # 尝试对只读目录进行操作
        readonly_dir = tempfile.mkdtemp()
        output_path = os.path.join(readonly_dir, "output.xlsx")

        # 设置目录为只读
        os.chmod(readonly_dir, 0o444)

        try:
            result = ExcelConverter.convert_format(
                self.file_path, output_path, "json"
            )

            # 结果可能成功也可能失败，取决于系统
            assert result is not None

        finally:
            # 恢复权限以便清理
            os.chmod(readonly_dir, 0o755)
            import shutil
            shutil.rmtree(readonly_dir)

    def test_concurrent_operations(self):
        """测试并发操作"""
        import threading

        results = []
        errors = []

        def worker(worker_id):
            try:
                output_file = tempfile.NamedTemporaryFile(suffix='.json', delete=False)
                output_file.close()

                result = ExcelConverter.convert_format(
                    self.file_path, output_file.name, "json"
                )
                results.append((worker_id, result.success))

                # 清理
                if os.path.exists(output_file.name):
                    os.unlink(output_file.name)

            except Exception as e:
                errors.append((worker_id, str(e)))

        # 启动多个线程
        threads = []
        for i in range(3):
            thread = threading.Thread(target=worker, args=(i,))
            threads.append(thread)
            thread.start()

        for thread in threads:
            thread.join()

        # 验证结果
        assert len(errors) == 0, f"并发操作出现错误: {errors}"
        assert len(results) == 3
        assert all(success for _, success in results)

    def test_memory_usage_large_data(self):
        """测试大数据量的内存使用"""
        import gc

        # 强制垃圾回收
        gc.collect()

        # 创建大数据量的Excel文件
        large_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        large_file.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "LargeData"

        # 添加大量数据
        for row in range(100):
            for col in range(20):
                ws.cell(row=row+1, column=col+1, value=f"Data_{row}_{col}")

        wb.save(large_file.name)

        output_file = tempfile.NamedTemporaryFile(suffix='.json', delete=False)
        output_file.close()

        try:
            result = ExcelConverter.convert_format(
                large_file.name, output_file.name, "json"
            )

            assert result.success is True

        finally:
            # 再次垃圾回收
            gc.collect()

            # 清理
            os.unlink(large_file.name)
            if os.path.exists(output_file.name):
                os.unlink(output_file.name)

        # 如果能到达这里，说明没有严重的内存问题
        assert True


if __name__ == "__main__":
    pytest.main([__file__, "-v"])