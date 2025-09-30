# -*- coding: utf-8 -*-
"""
Excel MCP Server - ID查重功能单元测试

专门测试excel_check_duplicate_ids功能的完整性和正确性
用于游戏配置表ID重复检查的全面测试

测试覆盖：
1. 无重复ID的正常情况
2. 存在重复ID的检测
3. 多种重复模式测试
4. 参数验证和边界条件
5. 错误处理和异常情况
6. 不同数据类型的ID处理
"""

import pytest
import tempfile
import unittest.mock
from pathlib import Path
from openpyxl import Workbook

from src.api.excel_operations import ExcelOperations


class TestExcelCheckDuplicateIds:
    """
    @class TestExcelCheckDuplicateIds
    @brief excel_check_duplicate_ids功能的全面单元测试
    @intention 确保ID查重功能在各种场景下的正确性和稳定性
    """

    # ==================== 测试数据准备 ====================

    @pytest.fixture
    def temp_dir(self):
        """创建临时目录"""
        import tempfile
        from pathlib import Path
        import gc
        import shutil
        import time
        import logging

        temp_path = Path(tempfile.mkdtemp())

        try:
            yield temp_path
        finally:
            # 使用与conftest.py相同的清理机制
            def safe_rmtree(path, max_retries=5, delay=0.1):
                for attempt in range(max_retries):
                    try:
                        gc.collect()
                        shutil.rmtree(path)
                        return
                    except PermissionError as e:
                        if attempt == max_retries - 1:
                            try:
                                gc.collect()
                                for file_path in Path(path).rglob("*"):
                                    if file_path.is_file():
                                        try:
                                            file_path.unlink(missing_ok=True)
                                        except PermissionError:
                                            pass
                                shutil.rmtree(path, ignore_errors=True)
                                return
                            except Exception:
                                logging.warning(f"Could not remove temp directory {path}: {e}")
                                return
                        time.sleep(delay)
                        delay *= 2
                        gc.collect()

            safe_rmtree(temp_path)

    @pytest.fixture
    def no_duplicate_file(self, temp_dir):
        """创建无重复ID的测试文件"""
        file_path = temp_dir / "no_duplicates.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "技能配置表"

        # 技能配置数据：ID, 技能名, 类型, 伤害
        test_data = [
            ["技能ID", "技能名称", "技能类型", "伤害值"],
            [1001, "火球术", "攻击", 100],
            [1002, "冰冻术", "控制", 80],
            [1003, "治疗术", "治疗", 50],
            [1004, "闪电术", "攻击", 120],
            [1005, "护盾术", "防御", 0]
        ]

        for row_data in test_data:
            ws.append(row_data)

        wb.save(file_path)
        return file_path

    @pytest.fixture
    def duplicate_file(self, temp_dir):
        """创建有重复ID的测试文件"""
        file_path = temp_dir / "with_duplicates.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "装备配置表"

        # 装备配置数据：包含重复ID
        test_data = [
            ["装备ID", "装备名称", "类型", "品质"],
            [2001, "钢剑", "武器", "普通"],
            [2002, "法杖", "武器", "稀有"],
            [2001, "钢剑改", "武器", "普通"],  # 重复ID 2001
            [2003, "皮甲", "防具", "普通"],
            [2002, "高级法杖", "武器", "史诗"],  # 重复ID 2002
            [2004, "钢盾", "防具", "普通"],
            [2001, "钢剑精炼版", "武器", "精良"]  # 重复ID 2001 (第3次)
        ]

        for row_data in test_data:
            ws.append(row_data)

        wb.save(file_path)
        return file_path

    @pytest.fixture
    def empty_file(self, temp_dir):
        """创建空的测试文件（只有表头）"""
        file_path = temp_dir / "empty.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "空配置表"

        # 只有表头
        ws.append(["ID", "名称", "类型"])

        wb.save(file_path)
        return file_path

    @pytest.fixture
    def single_row_file(self, temp_dir):
        """创建只有一行数据的测试文件"""
        file_path = temp_dir / "single_row.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "单行配置表"

        test_data = [
            ["ID", "名称"],
            [9001, "唯一道具"]
        ]

        for row_data in test_data:
            ws.append(row_data)

        wb.save(file_path)
        return file_path

    @pytest.fixture
    def string_id_file(self, temp_dir):
        """创建字符串ID的测试文件"""
        file_path = temp_dir / "string_ids.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "字符串ID表"

        test_data = [
            ["技能代码", "技能名称", "描述"],
            ["SKILL_FIREBALL", "火球术", "基础火系攻击魔法"],
            ["SKILL_HEAL", "治疗术", "基础治疗魔法"],
            ["SKILL_FIREBALL", "高级火球术", "进阶火系攻击魔法"],  # 重复
            ["SKILL_SHIELD", "护盾术", "防御魔法"]
        ]

        for row_data in test_data:
            ws.append(row_data)

        wb.save(file_path)
        return file_path

    # ==================== 正常功能测试 ====================

    def test_no_duplicates_detection(self, no_duplicate_file):
        """测试无重复ID的正确识别"""
        result = ExcelOperations.check_duplicate_ids(
            str(no_duplicate_file),
            "技能配置表",
            id_column=1,
            header_row=1
        )

        assert result['success'] is True
        assert result['has_duplicates'] is False
        assert result['duplicate_count'] == 0
        assert result['total_ids'] == 5
        assert result['unique_ids'] == 5
        assert len(result['duplicates']) == 0
        assert "无重复ID" in result['message']

    def test_duplicates_detection(self, duplicate_file):
        """测试重复ID的正确检测"""
        result = ExcelOperations.check_duplicate_ids(
            str(duplicate_file),
            "装备配置表",
            id_column=1,
            header_row=1
        )

        assert result['success'] is True
        assert result['has_duplicates'] is True
        assert result['duplicate_count'] == 2  # 2001和2002两个重复ID
        assert result['total_ids'] == 7
        assert result['unique_ids'] == 4  # 2001, 2002, 2003, 2004
        assert len(result['duplicates']) == 2

        # 验证重复详情
        duplicates = result['duplicates']

        # 查找ID 2001的重复信息
        id_2001 = next(d for d in duplicates if d['id_value'] == 2001)
        assert id_2001['count'] == 3
        assert len(id_2001['rows']) == 3
        assert sorted(id_2001['rows']) == [2, 4, 8]  # 第2, 4, 8行（绝对行号）

        # 查找ID 2002的重复信息
        id_2002 = next(d for d in duplicates if d['id_value'] == 2002)
        assert id_2002['count'] == 2
        assert len(id_2002['rows']) == 2
        assert sorted(id_2002['rows']) == [3, 6]  # 第3, 6行

    def test_string_id_duplicates(self, string_id_file):
        """测试字符串ID的重复检测"""
        result = ExcelOperations.check_duplicate_ids(
            str(string_id_file),
            "字符串ID表",
            id_column=1,
            header_row=1
        )

        assert result['success'] is True
        assert result['has_duplicates'] is True
        assert result['duplicate_count'] == 1

        # 验证字符串ID重复
        duplicate = result['duplicates'][0]
        assert duplicate['id_value'] == "SKILL_FIREBALL"
        assert duplicate['count'] == 2
        assert sorted(duplicate['rows']) == [2, 4]

    # ==================== 边界条件测试 ====================

    def test_empty_sheet_handling(self, empty_file):
        """测试空工作表的处理"""
        result = ExcelOperations.check_duplicate_ids(
            str(empty_file),
            "空配置表",
            id_column=1,
            header_row=1
        )

        assert result['success'] is True
        assert result['has_duplicates'] is False
        assert result['duplicate_count'] == 0
        assert result['total_ids'] == 0
        assert result['unique_ids'] == 0
        assert len(result['duplicates']) == 0

    def test_single_row_handling(self, single_row_file):
        """测试单行数据的处理"""
        result = ExcelOperations.check_duplicate_ids(
            str(single_row_file),
            "单行配置表",
            id_column=1,
            header_row=1
        )

        assert result['success'] is True
        assert result['has_duplicates'] is False
        assert result['duplicate_count'] == 0
        assert result['total_ids'] == 1
        assert result['unique_ids'] == 1
        assert len(result['duplicates']) == 0

    # ==================== 参数验证测试 ====================

    def test_different_id_columns(self, duplicate_file):
        """测试不同ID列的处理"""
        # 使用第2列作为ID列（装备名称）
        result = ExcelOperations.check_duplicate_ids(
            str(duplicate_file),
            "装备配置表",
            id_column=2,  # 装备名称列
            header_row=1
        )

        assert result['success'] is True
        # 装备名称应该有重复（钢剑类的几个变种）

    def test_column_by_letter(self, no_duplicate_file):
        """测试使用字母指定列"""
        result = ExcelOperations.check_duplicate_ids(
            str(no_duplicate_file),
            "技能配置表",
            id_column="A",  # A列即第1列
            header_row=1
        )

        assert result['success'] is True
        assert result['has_duplicates'] is False

    def test_different_header_row(self, temp_dir):
        """测试不同的表头行设置"""
        file_path = temp_dir / "multi_header.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "多表头表"

        # 多行表头数据
        test_data = [
            ["游戏配置表", "", ""],  # 第1行：标题
            ["技能ID", "技能名称", "伤害值"],  # 第2行：实际表头
            [5001, "火球", 100],
            [5002, "冰冻", 80],
            [5001, "火球升级", 120]  # 重复ID
        ]

        for row_data in test_data:
            ws.append(row_data)

        wb.save(file_path)

        # 使用第2行作为表头
        result = ExcelOperations.check_duplicate_ids(
            str(file_path),
            "多表头表",
            id_column=1,
            header_row=2
        )

        assert result['success'] is True
        assert result['has_duplicates'] is True
        assert result['duplicate_count'] == 1

    # ==================== 错误处理测试 ====================

    def test_file_not_found(self):
        """测试文件不存在的错误处理"""
        result = ExcelOperations.check_duplicate_ids(
            "nonexistent_file.xlsx",
            "Sheet1",
            id_column=1,
            header_row=1
        )

        assert result['success'] is False
        assert "文件不存在" in result['message'] or "FileNotFoundError" in result['message']

    def test_sheet_not_found(self, no_duplicate_file):
        """测试工作表不存在的错误处理"""
        result = ExcelOperations.check_duplicate_ids(
            str(no_duplicate_file),
            "不存在的工作表",
            id_column=1,
            header_row=1
        )

        assert result['success'] is False
        assert "工作表不存在" in result['message'] or "Worksheet" in result['message']

    def test_invalid_column(self, no_duplicate_file):
        """测试无效列的错误处理"""
        result = ExcelOperations.check_duplicate_ids(
            str(no_duplicate_file),
            "技能配置表",
            id_column=999,  # 不存在的列
            header_row=1
        )

        assert result['success'] is False
        assert "列不存在" in result['message'] or "索引超出范围" in result['message']

    def test_invalid_header_row(self, no_duplicate_file):
        """测试无效表头行的错误处理"""
        result = ExcelOperations.check_duplicate_ids(
            str(no_duplicate_file),
            "技能配置表",
            id_column=1,
            header_row=999  # 不存在的行
        )

        assert result['success'] is False
        assert "表头行不存在" in result['message'] or "索引超出范围" in result['message']

    # ==================== 性能和压力测试 ====================

    def test_large_dataset_performance(self, temp_dir):
        """测试大数据集的性能表现"""
        file_path = temp_dir / "large_dataset.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "大数据表"

        # 创建大数据集（1000行，包含一些重复）
        ws.append(["ID", "数据"])

        # 添加999行数据，其中包含重复
        for i in range(1, 1000):
            # 每100个创建一个重复
            id_value = i if i % 100 != 0 else i - 1
            ws.append([id_value, f"数据{i}"])

        wb.save(file_path)

        # 测试大数据集处理
        result = ExcelOperations.check_duplicate_ids(
            str(file_path),
            "大数据表",
            id_column=1,
            header_row=1
        )

        assert result['success'] is True
        assert result['total_ids'] == 999
        assert result['has_duplicates'] is True
        # 验证重复检测的正确性
        assert result['duplicate_count'] > 0

    # ==================== 综合集成测试 ====================

    def test_comprehensive_workflow(self, temp_dir):
        """测试完整的工作流程"""
        file_path = temp_dir / "comprehensive.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "综合测试表"

        # 综合数据：包含数字ID、字符串数据、重复、边界情况
        test_data = [
            ["角色ID", "角色名", "职业", "等级"],
            [1, "战士A", "战士", 10],
            [2, "法师B", "法师", 8],
            [1, "战士A克隆", "战士", 10],  # 重复
            [3, "盗贼C", "盗贼", 12],
            [None, "未知角色", "未知", 1],  # 空ID
            [4, "牧师D", "牧师", 9],
            [2, "法师B变种", "法师", 8],  # 重复
            [5, "骑士E", "骑士", 15]
        ]

        for row_data in test_data:
            ws.append(row_data)

        wb.save(file_path)

        result = ExcelOperations.check_duplicate_ids(
            str(file_path),
            "综合测试表",
            id_column=1,
            header_row=1
        )

        assert result['success'] is True
        assert result['has_duplicates'] is True

        # 验证综合结果的合理性
        assert result['total_ids'] >= 0
        assert result['unique_ids'] <= result['total_ids']
        assert result['duplicate_count'] >= 0
        assert len(result['duplicates']) == result['duplicate_count']
