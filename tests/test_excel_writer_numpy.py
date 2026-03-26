# -*- coding: utf-8 -*-
"""
Excel Writer Numpy统计函数测试套件

覆盖 excel_writer.py 中的numpy统计函数
"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from src.core.excel_writer import ExcelWriter


class TestExcelWriterNumpyFunctions:
    """ExcelWriter Numpy统计函数测试"""

    @pytest.fixture
    def numpy_test_file(self, temp_dir):
        """创建数值测试文件"""
        file_path = temp_dir / "numpy_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "NumpySheet"
        
        # 创建数值数据用于测试统计函数
        # A列: 1-20
        for i in range(1, 21):
            ws.cell(row=i, column=1, value=i)
        
        # B列: 10-29
        for i in range(1, 21):
            ws.cell(row=i, column=2, value=i + 9)
        
        # C列: 偶数
        for i in range(1, 21):
            ws.cell(row=i, column=3, value=i * 2)
        
        # D列: 带小数的数值
        for i in range(1, 21):
            ws.cell(row=i, column=4, value=i * 1.5)
        
        wb.save(file_path)
        return str(file_path)

    # ==================== AVERAGE函数测试 ====================

    def test_formula_average_column(self, numpy_test_file):
        """测试AVERAGE公式按列"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E1",
            [["=AVERAGE(A1:A20)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_average_range(self, numpy_test_file):
        """测试AVERAGE公式多列范围"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E2",
            [["=AVERAGE(A1:D20)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== MIN/MAX函数测试 ====================

    def test_formula_min_column(self, numpy_test_file):
        """测试MIN公式按列"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E3",
            [["=MIN(A1:A20)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_max_column(self, numpy_test_file):
        """测试MAX公式按列"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E4",
            [["=MAX(A1:A20)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== MEDIAN函数测试 ====================

    def test_formula_median(self, numpy_test_file):
        """测试MEDIAN公式"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E5",
            [["=MEDIAN(A1:A20)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== STDEV函数测试 ====================

    def test_formula_stdev(self, numpy_test_file):
        """测试STDEV公式"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E6",
            [["=STDEV(A1:A20)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== PERCENTILE函数测试 ====================

    def test_formula_percentile(self, numpy_test_file):
        """测试PERCENTILE公式"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E7",
            [["=PERCENTILE(A1:A20,0.5)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_percentile_25(self, numpy_test_file):
        """测试PERCENTILE 25%"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E8",
            [["=PERCENTILE(A1:A20,0.25)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== QUARTILE函数测试 ====================

    def test_formula_quartile(self, numpy_test_file):
        """测试QUARTILE公式"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E9",
            [["=QUARTILE(A1:A20,1)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== COUNT函数测试 ====================

    def test_formula_count_numbers(self, numpy_test_file):
        """测试COUNT公式"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E10",
            [["=COUNT(A1:A20)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_counta(self, numpy_test_file):
        """测试COUNTA公式"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E11",
            [["=COUNTA(A1:D20)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== SUM函数测试 ====================

    def test_formula_sum_column(self, numpy_test_file):
        """测试SUM公式按列"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E12",
            [["=SUM(A1:A20)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== 条件计数测试 ====================

    def test_formula_countif_greater(self, numpy_test_file):
        """测试COUNTIF大于条件"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E13",
            [["=COUNTIF(A1:A20,\">10\")"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_countif_less_or_equal(self, numpy_test_file):
        """测试COUNTIF小于等于条件"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E14",
            [["=COUNTIF(A1:A20,\"<=10\")"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_countif_greater_or_equal(self, numpy_test_file):
        """测试COUNTIF大于等于条件"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E15",
            [["=COUNTIF(A1:A20,\">=15\")"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== 条件求和测试 ====================

    def test_formula_sumif_greater(self, numpy_test_file):
        """测试SUMIF大于条件"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E16",
            [["=SUMIF(A1:A20,\">10\")"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_sumif_less(self, numpy_test_file):
        """测试SUMIF小于条件"""
        writer = ExcelWriter(numpy_test_file)
        result = writer.update_range(
            "NumpySheet!E17",
            [["=SUMIF(A1:A20,\"<10\")"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== 复杂公式组合测试 ====================

    def test_formula_combined_stats(self, numpy_test_file):
        """测试组合统计公式"""
        writer = ExcelWriter(numpy_test_file)
        
        formulas = [
            ["=SUM(A1:A20)+AVERAGE(B1:B20)"],
            ["=MAX(A1:A20)-MIN(A1:A20)"],
            ["=COUNTIF(A1:A20,\">10\")+COUNTIF(B1:B20,\">10\")"]
        ]
        
        result = writer.update_range(
            "NumpySheet!F1:F3",
            formulas,
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True


class TestExcelWriterEdgeCasesNumpy:
    """ExcelWriter Numpy边界情况测试"""

    @pytest.fixture
    def edge_file(self, temp_dir):
        """创建边界测试文件"""
        file_path = temp_dir / "edge_numpy.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "EdgeSheet"
        
        # 包含各种边界值
        ws['A1'] = 0
        ws['A2'] = -1
        ws['A3'] = 1
        ws['A4'] = 100
        ws['A5'] = 0.001
        
        wb.save(file_path)
        return str(file_path)

    def test_formula_min_edge(self, edge_file):
        """测试MIN边界值"""
        writer = ExcelWriter(edge_file)
        result = writer.update_range(
            "EdgeSheet!B1",
            [["=MIN(A1:A5)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_max_edge(self, edge_file):
        """测试MAX边界值"""
        writer = ExcelWriter(edge_file)
        result = writer.update_range(
            "EdgeSheet!B2",
            [["=MAX(A1:A5)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_average_edge(self, edge_file):
        """测试AVERAGE边界值"""
        writer = ExcelWriter(edge_file)
        result = writer.update_range(
            "EdgeSheet!B3",
            [["=AVERAGE(A1:A5)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True


class TestExcelWriterConcatenate:
    """ExcelWriter CONCATENATE函数测试"""

    @pytest.fixture
    def concat_file(self, temp_dir):
        """创建连接测试文件"""
        file_path = temp_dir / "concat_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ConcatSheet"
        
        ws['A1'] = "Hello"
        ws['A2'] = "World"
        ws['A3'] = "Test"
        
        wb.save(file_path)
        return str(file_path)

    def test_formula_concatenate(self, concat_file):
        """测试CONCATENATE公式"""
        writer = ExcelWriter(concat_file)
        result = writer.update_range(
            "ConcatSheet!B1",
            [["=CONCATENATE(A1,\" \",A2)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_concatenate_multiple(self, concat_file):
        """测试多单元格连接"""
        writer = ExcelWriter(concat_file)
        result = writer.update_range(
            "ConcatSheet!B2",
            [["=CONCATENATE(A1,\",\",A2,\",\",A3)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True


class TestExcelWriterDecimalOperations:
    """ExcelWriter 小数操作测试"""

    @pytest.fixture
    def decimal_file(self, temp_dir):
        """创建小数测试文件"""
        file_path = temp_dir / "decimal_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DecimalSheet"
        
        # 包含小数的值
        ws['A1'] = 3.14159
        ws['A2'] = 2.71828
        ws['A3'] = 1.41421
        
        wb.save(file_path)
        return str(file_path)

    def test_formula_decimal_operations(self, decimal_file):
        """测试小数运算"""
        writer = ExcelWriter(decimal_file)
        
        formulas = [
            ["=A1*A2"],
            ["=A1/A2"],
            ["=A1+A2+A3"],
            ["=A1-A2"]
        ]
        
        result = writer.update_range(
            "DecimalSheet!B1:B4",
            formulas,
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True
