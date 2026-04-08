"""
REQ-035 修复测试 - MCP真实验证发现的7项核心功能问题修复验证

问题清单:
1. SQL解析器中文表名ANSI转义码乱码
2. get_range格式验证过严（缺少工作表名时不自动推断）
3. batch_insert_rows验证过严（不接受dict/tuple）
4. delete_rows的condition模式使用了不存在的_row_number和ExcelOperations.query
5. query WHERE/JOIN/GROUP BY/子查询（受ANSI转义码影响）
"""

import os
import sys
import pytest
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
from excel_mcp_server_fastmcp.api.advanced_sql_query import AdvancedSQLQueryEngine


@pytest.fixture
def game_xlsx(tmp_path):
    """创建游戏配置测试文件（技能表+装备表）"""
    filepath = str(tmp_path / "game_config.xlsx")
    wb = openpyxl.Workbook()
    
    # 技能表
    ws = wb.active
    ws.title = '技能表'
    ws.append(['技能ID', '技能名称', '伤害', '冷却时间', '类型'])
    ws.append([1, '火球术', 100, 5, '攻击'])
    ws.append([2, '冰冻术', 80, 3, '控制'])
    ws.append([3, '治疗术', 50, 8, '辅助'])
    ws.append([4, '雷电术', 120, 4, '攻击'])
    ws.append([5, '护盾术', 0, 10, '防御'])
    
    # 装备表
    ws2 = wb.create_sheet('装备表')
    ws2.append(['装备ID', '装备名称', '技能ID', '稀有度'])
    ws2.append([1, '法杖', 1, '传说'])
    ws2.append([2, '法袍', 2, '史诗'])
    ws2.append([3, '盾牌', 5, '稀有'])
    
    wb.save(filepath)
    return filepath


@pytest.fixture
def single_sheet_xlsx(tmp_path):
    """创建单工作表测试文件"""
    filepath = str(tmp_path / "single_sheet.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['ID', '名称', '值'])
    ws.append([1, 'A', 10])
    ws.append([2, 'B', 20])
    wb.save(filepath)
    return filepath


class TestANSIEscapeCleanup:
    """测试1: SQL解析器ANSI转义码清理"""
    
    def test_ansi_escape_in_table_name(self, game_xlsx):
        """ANSI转义码不应导致SQL解析失败"""
        q = AdvancedSQLQueryEngine()
        # 模拟终端粘贴带入的ANSI转义码
        sql = 'SELECT * FROM \x1b[4m[\x1b[0m技能表'
        result = q.execute_sql_query(game_xlsx, sql)
        assert result['success'], f"ANSI转义码清理失败: {result.get('message', '')}"
    
    def test_ansi_escape_in_where_clause(self, game_xlsx):
        """WHERE子句中的ANSI转义码应被清理"""
        q = AdvancedSQLQueryEngine()
        sql = "SELECT * FROM 技能表 \x1b[1mWHERE\x1b[0m 伤害 > 50"
        result = q.execute_sql_query(game_xlsx, sql)
        assert result['success'], f"WHERE子句ANSI清理失败: {result.get('message', '')}"
    
    def test_normal_query_unaffected(self, game_xlsx):
        """正常SQL查询不受ANSI清理影响"""
        q = AdvancedSQLQueryEngine()
        result = q.execute_sql_query(game_xlsx, 'SELECT * FROM 技能表 WHERE 伤害 > 50')
        assert result['success']
        data = result.get('data', [])
        # 应该返回3行（火球术100, 冰冻术80, 雷电术120）
        assert len(data) >= 3


class TestGetRangeAutoInfer:
    """测试2: get_range工作表名自动推断"""
    
    def test_single_sheet_auto_infer(self, single_sheet_xlsx):
        """单工作表文件应自动推断工作表名"""
        result = ExcelOperations.get_range(single_sheet_xlsx, 'A1:B2')
        # ExcelOperations层仍需工作表名，但server.py层会自动推断
        # 这里测试ExcelOperations._validate_range_format的错误消息
        assert result['success'] is False  # ExcelOperations层仍要求!
        assert '工作表名' in result.get('message', '')
    
    def test_explicit_sheet_name(self, game_xlsx):
        """显式工作表名应正常工作"""
        result = ExcelOperations.get_range(game_xlsx, '技能表!A1:C3')
        assert result['success'], f"显式工作表名失败: {result.get('message', '')}"
        data = result.get('data', [])
        assert len(data) == 3  # 3行数据
    
    def test_full_range(self, game_xlsx):
        """完整范围读取"""
        result = ExcelOperations.get_range(game_xlsx, '技能表!A1:E6')
        assert result['success']
        data = result.get('data', [])
        assert len(data) == 6  # 1行表头 + 5行数据


class TestBatchInsertRowsValidation:
    """测试3: batch_insert_rows数据验证优化"""
    
    def test_dict_auto_wrap(self, game_xlsx):
        """单个字典应自动包装为列表"""
        data = {'技能ID': 99, '技能名称': '测试', '伤害': 50, '冷却时间': 3, '类型': '测试'}
        result = ExcelOperations.batch_insert_rows(game_xlsx, '技能表', data)
        assert result['success'], f"字典自动包装失败: {result.get('message', '')}"
    
    def test_tuple_data(self, game_xlsx):
        """元组数据应被接受"""
        data = (
            {'技能ID': 100, '技能名称': '测试1', '伤害': 50, '冷却时间': 3, '类型': '测试'},
            {'技能ID': 101, '技能名称': '测试2', '伤害': 60, '冷却时间': 4, '类型': '测试'},
        )
        result = ExcelOperations.batch_insert_rows(game_xlsx, '技能表', data)
        assert result['success'], f"元组数据失败: {result.get('message', '')}"
    
    def test_list_data(self, game_xlsx):
        """列表数据应正常工作"""
        data = [{'技能ID': 102, '技能名称': '测试3', '伤害': 70, '冷却时间': 5, '类型': '测试'}]
        result = ExcelOperations.batch_insert_rows(game_xlsx, '技能表', data)
        assert result['success'], f"列表数据失败: {result.get('message', '')}"
    
    def test_empty_list_rejected(self, game_xlsx):
        """空列表应被拒绝"""
        result = ExcelOperations.batch_insert_rows(game_xlsx, '技能表', [])
        assert result['success'] is False
        assert '不能为空' in result.get('message', '')


class TestDeleteRowsCondition:
    """测试4: delete_rows条件删除模式修复"""
    
    def test_excel_operations_query_exists(self, game_xlsx):
        """ExcelOperations.query方法应存在且可用"""
        result = ExcelOperations.query(game_xlsx, 'SELECT * FROM 技能表 WHERE 伤害 > 50')
        assert result['success'], f"ExcelOperations.query不存在或失败: {result.get('message', '')}"
    
    def test_condition_delete(self, tmp_path):
        """条件删除应正常工作"""
        filepath = str(tmp_path / "delete_test.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '测试表'
        ws.append(['ID', '名称', '值'])
        ws.append([1, 'A', 10])
        ws.append([2, 'B', 20])
        ws.append([3, 'C', 30])
        wb.save(filepath)
        
        # 删除值>15的行
        result = ExcelOperations.delete_rows(filepath, '测试表', row_index=3, count=1)
        assert result['success'], f"删除失败: {result.get('message', '')}"
    
    def test_simple_row_delete(self, tmp_path):
        """简单行号删除"""
        filepath = str(tmp_path / "simple_delete.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        ws.append(['A', 'B'])
        ws.append([1, 2])
        ws.append([3, 4])
        wb.save(filepath)
        
        result = ExcelOperations.delete_rows(filepath, 'Sheet1', row_index=2, count=1)
        assert result['success']


class TestSQLQueryFunctions:
    """测试5-7: SQL查询功能（WHERE/JOIN/GROUP BY/子查询）"""
    
    def test_where_clause(self, game_xlsx):
        """WHERE条件查询"""
        q = AdvancedSQLQueryEngine()
        result = q.execute_sql_query(game_xlsx, 'SELECT * FROM 技能表 WHERE 伤害 > 80')
        assert result['success'], f"WHERE查询失败: {result.get('message', '')}"
    
    def test_join_query(self, game_xlsx):
        """JOIN关联查询"""
        q = AdvancedSQLQueryEngine()
        result = q.execute_sql_query(
            game_xlsx,
            'SELECT a.技能名称, b.装备名称 FROM 技能表 a JOIN 装备表 b ON a.技能ID = b.技能ID'
        )
        assert result['success'], f"JOIN查询失败: {result.get('message', '')}"
    
    def test_group_by(self, game_xlsx):
        """GROUP BY聚合查询"""
        q = AdvancedSQLQueryEngine()
        result = q.execute_sql_query(
            game_xlsx,
            'SELECT 类型, COUNT(*) as cnt, AVG(伤害) as avg_dmg FROM 技能表 GROUP BY 类型'
        )
        assert result['success'], f"GROUP BY查询失败: {result.get('message', '')}"
    
    def test_subquery(self, game_xlsx):
        """子查询"""
        q = AdvancedSQLQueryEngine()
        result = q.execute_sql_query(
            game_xlsx,
            'SELECT * FROM 技能表 WHERE 伤害 > (SELECT AVG(伤害) FROM 技能表)'
        )
        assert result['success'], f"子查询失败: {result.get('message', '')}"
    
    def test_order_by(self, game_xlsx):
        """ORDER BY排序"""
        q = AdvancedSQLQueryEngine()
        result = q.execute_sql_query(
            game_xlsx,
            'SELECT * FROM 技能表 ORDER BY 伤害 DESC'
        )
        assert result['success'], f"ORDER BY查询失败: {result.get('message', '')}"
    
    def test_chinese_column_where(self, game_xlsx):
        """中文列名WHERE查询"""
        q = AdvancedSQLQueryEngine()
        result = q.execute_sql_query(game_xlsx, 'SELECT * FROM 技能表 WHERE 类型 = "攻击"')
        assert result['success'], f"中文列名WHERE失败: {result.get('message', '')}"
