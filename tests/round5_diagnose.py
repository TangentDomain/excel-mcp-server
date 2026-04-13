"""诊断 Round 5 两个失败用例的根因"""
import sys
sys.path.insert(0, '/root/workspace/excel-mcp-server/src')
from openpyxl import load_workbook
import pandas as pd

print("=" * 60)
print("诊断1: 极小浮点数精度 (A4)")
print("=" * 60)

wb = load_workbook('/tmp/excelmcp_round5_test.xlsx')
ws = wb['装备']

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] == 12:
        print(f"openpyxl 原始数据: {row}")
        print(f"AtkBonus type={type(row[3])}, value={row[3]!r}")
        print(f"Price type={type(row[4])}, value={row[4]!r}")
        break

df = pd.read_excel('/tmp/excelmcp_round5_test.xlsx', sheet_name='装备', engine='openpyxl')
row12 = df[df['ID'] == 12]
print(f"\nPandas 读取:")
print(row12.to_dict('records'))
print(f"AtkBonus dtype: {df['AtkBonus'].dtype}")

# 检查所有 AtkBonus 值
print(f"\n所有 AtkBonus 值:")
for i, v in enumerate(df['AtkBonus'].values):
    print(f"  row {i}: {v!r} (type={type(v).__name__})")

print("\n" + "=" * 60)
print("诊断2: DELETE 后 COUNT 返回空 (C4b)")
print("=" * 60)

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_delete_query,
)

file_path = '/tmp/excelmcp_round5_test.xlsx'

# 先看看当前有哪些 '边界测试' 行
r1 = execute_advanced_sql_query(file_path, "SELECT ID, Name FROM 装备 WHERE Name LIKE '边界测试%'")
print(f"\nDELETE 前 - 边界测试行:")
print(f"  success={r1['success']}, data={r1['data']}")

# 执行 DELETE
r2 = execute_advanced_delete_query(file_path, "DELETE FROM 装备 WHERE Name LIKE '边界测试%'")
print(f"\nDELETE 结果: success={r2['success']}, message={r2.get('message', '')}")

# 再 COUNT
r3 = execute_advanced_sql_query(file_path, "SELECT COUNT(*) as cnt FROM 装备 WHERE Name LIKE '边界测试%'")
print(f"\nDELETE 后 COUNT: success={r3['success']}, data={r3['data']}")
print(f"  message={r3.get('message', '')}")

# 也试试不带 WHERE 的 COUNT
r4 = execute_advanced_sql_query(file_path, "SELECT COUNT(*) as total_cnt FROM 装备")
print(f"\n全表 COUNT: success={r4['success']}, data={r4['data']}")
