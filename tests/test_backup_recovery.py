"""
Excel MCP 服务器备份恢复机制测试

本文件包含备份创建、管理和恢复功能的全面测试。
"""

import pytest
import os
import tempfile
import time
import json
import hashlib
from pathlib import Path
from unittest.mock import patch, MagicMock
import shutil

# 导入要测试的模块
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'src'))

from src.api.excel_operations import ExcelOperations
from src.utils.exceptions import SecurityError, OperationCancelledError


class TestBackupCreation:
    """备份创建测试类"""

    @pytest.fixture
    def temp_excel_file(self):
        """创建临时Excel文件用于测试"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            file_path = tmp.name

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "TestData"

        # 添加测试数据
        ws.append(["ID", "名称", "类型", "数值"])
        ws.append([1, "项目1", "A", 100])
        ws.append([2, "项目2", "B", 200])
        ws.append([3, "项目3", "C", 300])

        wb.save(file_path)
        wb.close()

        yield file_path

        # 清理
        if os.path.exists(file_path):
            os.unlink(file_path)

    @pytest.fixture
    def backup_dir(self):
        """创建临时备份目录"""
        temp_dir = tempfile.mkdtemp(prefix="excel_backup_test_")
        yield temp_dir

        # 清理
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)

    def test_create_basic_backup(self, temp_excel_file, backup_dir):
        """测试基本备份创建"""
        with patch('os.makedirs') as mock_makedirs:
            # 模拟备份目录创建成功
            mock_makedirs.return_value = None

            result = ExcelOperations.create_auto_backup(
                file_path=temp_excel_file,
                backup_name="basic_test",
                backup_reason="基本测试备份",
                user_id="test_user"
            )

            assert result['success'] is True
            assert 'backup_path' in result
            assert 'backup_id' in result
            assert result['backup_name'] == "basic_test"
            assert result['backup_reason'] == "基本测试备份"
            assert result['user_id'] == "test_user"
            assert 'timestamp' in result
            assert 'checksum' in result

    def test_create_backup_with_checksum(self, temp_excel_file):
        """测试创建带校验和的备份"""
        # 计算原始文件的校验和
        with open(temp_excel_file, 'rb') as f:
            original_checksum = hashlib.sha256(f.read()).hexdigest()

        result = ExcelOperations.create_auto_backup(
            file_path=temp_excel_file,
            backup_name="checksum_test",
            backup_reason="校验和测试",
            user_id="test_user"
        )

        assert result['success'] is True
        assert 'checksum' in result
        assert len(result['checksum']) == 64  # SHA-256 长度

        # 验证校验和正确
        assert result['checksum'] == original_checksum

        # 验证备份文件存在且校验和一致
        if os.path.exists(result['backup_path']):
            with open(result['backup_path'], 'rb') as f:
                backup_checksum = hashlib.sha256(f.read()).hexdigest()
            assert backup_checksum == original_checksum

            # 清理
            os.unlink(result['backup_path'])

    def test_create_backup_metadata(self, temp_excel_file):
        """测试备份元数据创建"""
        result = ExcelOperations.create_auto_backup(
            file_path=temp_excel_file,
            backup_name="metadata_test",
            backup_reason="元数据测试",
            user_id="test_user"
        )

        assert result['success'] is True

        # 检查备份元数据文件
        metadata_file = result['backup_path'].replace('.xlsx', '_metadata.json')
        if os.path.exists(metadata_file):
            with open(metadata_file, 'r', encoding='utf-8') as f:
                metadata = json.load(f)

            assert metadata['backup_name'] == "metadata_test"
            assert metadata['backup_reason'] == "元数据测试"
            assert metadata['user_id'] == "test_user"
            assert metadata['original_file'] == temp_excel_file
            assert metadata['checksum'] == result['checksum']
            assert 'created_at' in metadata

            # 清理
            os.unlink(metadata_file)

        # 清理备份文件
        if os.path.exists(result['backup_path']):
            os.unlink(result['backup_path'])

    def test_create_backup_file_not_found(self):
        """测试备份不存在的文件"""
        result = ExcelOperations.create_auto_backup(
            file_path="不存在的文件.xlsx",
            backup_name="error_test",
            backup_reason="错误测试",
            user_id="test_user"
        )

        assert result['success'] is False
        assert "文件不存在" in result['error']

    def test_create_backup_invalid_file(self):
        """测试备份无效文件"""
        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as tmp:
            tmp.write("不是Excel文件".encode('utf-8'))
            tmp_path = tmp.name

        try:
            result = ExcelOperations.create_auto_backup(
                file_path=tmp_path,
                backup_name="invalid_test",
                backup_reason="无效文件测试",
                user_id="test_user"
            )

            assert result['success'] is False
            assert "不是有效的Excel文件" in result['error']
        finally:
            os.unlink(tmp_path)

    def test_create_multiple_backups(self, temp_excel_file):
        """测试创建多个备份"""
        backup_paths = []

        try:
            # 创建多个备份
            for i in range(3):
                result = ExcelOperations.create_auto_backup(
                    file_path=temp_excel_file,
                    backup_name=f"multi_test_{i}",
                    backup_reason=f"多备份测试 {i}",
                    user_id="test_user"
                )

                assert result['success'] is True
                backup_paths.append(result['backup_path'])
                assert os.path.exists(result['backup_path'])

                # 检查备份ID唯一性
                backup_id = result['backup_id']
                assert backup_id is not None

                # 添加小延迟确保时间戳不同
                time.sleep(0.1)

        finally:
            # 清理所有备份文件
            for backup_path in backup_paths:
                if os.path.exists(backup_path):
                    os.unlink(backup_path)
                metadata_path = backup_path.replace('.xlsx', '_metadata.json')
                if os.path.exists(metadata_path):
                    os.unlink(metadata_path)


class TestBackupRecovery:
    """备份恢复测试类"""

    @pytest.fixture
    def temp_excel_file(self):
        """创建临时Excel文件用于测试"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            file_path = tmp.name

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "TestData"

        # 添加原始数据
        ws.append(["ID", "名称", "类型", "数值"])
        ws.append([1, "原始项目1", "A", 100])
        ws.append([2, "原始项目2", "B", 200])
        ws.append([3, "原始项目3", "C", 300])

        wb.save(file_path)
        wb.close()

        yield file_path

        # 清理
        if os.path.exists(file_path):
            os.unlink(file_path)

    @pytest.fixture
    def backup_file(self, temp_excel_file):
        """创建备份文件用于恢复测试"""
        result = ExcelOperations.create_auto_backup(
            file_path=temp_excel_file,
            backup_name="recovery_test",
            backup_reason="恢复测试备份",
            user_id="test_user"
        )

        yield result['backup_path']

        # 清理
        if os.path.exists(result['backup_path']):
            os.unlink(result['backup_path'])
        metadata_path = result['backup_path'].replace('.xlsx', '_metadata.json')
        if os.path.exists(metadata_path):
            os.unlink(metadata_path)

    def test_restore_from_backup_success(self, temp_excel_file, backup_file):
        """测试成功从备份恢复"""
        # 首先修改原文件
        from openpyxl import Workbook, load_workbook
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        ws.append([4, "修改项目", "D", 999])  # 添加新行
        wb.save(temp_excel_file)
        wb.close()

        # 验证文件已被修改
        wb_modified = load_workbook(temp_excel_file)
        ws_modified = wb_modified.active
        modified_rows = list(ws_modified.iter_rows(values_only=True))
        wb_modified.close()
        assert len(modified_rows) == 5  # 应该有5行

        # 从备份恢复
        result = ExcelOperations.restore_from_backup(
            original_file_path=temp_excel_file,
            backup_path=backup_file
        )

        assert result['success'] is True
        assert 'restored_at' in result

        # 验证文件已恢复
        wb_restored = load_workbook(temp_excel_file)
        ws_restored = wb_restored.active
        restored_rows = list(ws_restored.iter_rows(values_only=True))
        wb_restored.close()
        assert len(restored_rows) == 4  # 应该恢复到4行

    def test_restore_from_backup_file_not_found(self, temp_excel_file):
        """测试从不存在的备份文件恢复"""
        result = ExcelOperations.restore_from_backup(
            original_file_path=temp_excel_file,
            backup_path="不存在的备份.xlsx"
        )

        assert result['success'] is False
        assert "备份文件不存在" in result['error']

    def test_restore_from_backup_invalid_backup(self, temp_excel_file):
        """测试从无效备份文件恢复"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write("无效的Excel文件".encode('utf-8'))
            invalid_backup = tmp.name

        try:
            result = ExcelOperations.restore_from_backup(
                original_file_path=temp_excel_file,
                backup_path=invalid_backup
            )

            assert result['success'] is False
            assert "备份文件已损坏" in result['error']
        finally:
            os.unlink(invalid_backup)

    def test_restore_from_backup_with_checksum_verification(self, temp_excel_file, backup_file):
        """测试带校验和验证的备份恢复"""
        # 获取备份的校验和
        metadata_path = backup_file.replace('.xlsx', '_metadata.json')
        with open(metadata_path, 'r', encoding='utf-8') as f:
            metadata = json.load(f)
        original_checksum = metadata['checksum']

        # 尝试恢复（应该成功）
        result = ExcelOperations.restore_from_backup(
            original_file_path=temp_excel_file,
            backup_path=backup_file
        )

        assert result['success'] is True

        # 验证恢复后文件的校验和
        with open(temp_excel_file, 'rb') as f:
            restored_checksum = hashlib.sha256(f.read()).hexdigest()
        assert restored_checksum == original_checksum

    def test_restore_from_backup_original_file_missing(self, backup_file):
        """测试原文件不存在时的恢复"""
        # 删除原文件
        original_path = "temp_original.xlsx"

        result = ExcelOperations.restore_from_backup(
            original_file_path=original_path,
            backup_path=backup_file
        )

        assert result['success'] is True
        assert os.path.exists(original_path)

        # 验证恢复的文件内容正确
        from openpyxl import load_workbook
        wb = load_workbook(original_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        assert len(rows) >= 4  # 至少有4行数据

        # 清理
        if os.path.exists(original_path):
            os.unlink(original_path)


class TestBackupManagement:
    """备份管理测试类"""

    @pytest.fixture
    def temp_excel_file(self):
        """创建临时Excel文件用于测试"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            file_path = tmp.name

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "TestData"

        ws.append(["ID", "名称", "类型", "数值"])
        ws.append([1, "项目1", "A", 100])

        wb.save(file_path)
        wb.close()

        yield file_path

        # 清理
        if os.path.exists(file_path):
            os.unlink(file_path)

    def test_list_backups_empty(self, temp_excel_file):
        """测试列出空备份列表"""
        result = ExcelOperations.list_backups(temp_excel_file)

        assert result['success'] is True
        assert len(result['backups']) == 0

    def test_list_backups_with_data(self, temp_excel_file):
        """测试列出有备份的列表"""
        backup_paths = []

        try:
            # 创建几个备份
            for i in range(3):
                result = ExcelOperations.create_auto_backup(
                    file_path=temp_excel_file,
                    backup_name=f"list_test_{i}",
                    backup_reason=f"列表测试 {i}",
                    user_id="test_user"
                )
                backup_paths.append(result['backup_path'])

            # 列出备份
            list_result = ExcelOperations.list_backups(temp_excel_file)

            assert list_result['success'] is True
            assert len(list_result['backups']) == 3

            # 验证备份信息
            for backup in list_result['backups']:
                assert 'backup_id' in backup
                assert 'backup_name' in backup
                assert 'backup_reason' in backup
                assert 'created_at' in backup
                assert 'file_size' in backup
                assert 'checksum' in backup

        finally:
            # 清理备份文件
            for backup_path in backup_paths:
                if os.path.exists(backup_path):
                    os.unlink(backup_path)
                metadata_path = backup_path.replace('.xlsx', '_metadata.json')
                if os.path.exists(metadata_path):
                    os.unlink(metadata_path)

    def test_list_backups_nonexistent_file(self):
        """测试列出不存在文件的备份"""
        result = ExcelOperations.list_backups("不存在的文件.xlsx")

        assert result['success'] is True  # 应该成功，但返回空列表
        assert len(result['backups']) == 0

    def test_cleanup_old_backups(self, temp_excel_file):
        """测试清理旧备份"""
        backup_paths = []

        try:
            # 创建备份
            for i in range(5):
                result = ExcelOperations.create_auto_backup(
                    file_path=temp_excel_file,
                    backup_name=f"cleanup_test_{i}",
                    backup_reason=f"清理测试 {i}",
                    user_id="test_user"
                )
                backup_paths.append(result['backup_path'])
                time.sleep(0.1)  # 确保时间戳不同

            # 清理，保留最近3个
            result = ExcelOperations.cleanup_old_backups(
                file_path=temp_excel_file,
                keep_count=3
            )

            assert result['success'] is True
            assert result['cleaned_count'] == 2
            assert result['remaining_count'] == 3

            # 验证剩余的备份文件
            remaining_backups = ExcelOperations.list_backups(temp_excel_file)
            assert len(remaining_backups['backups']) == 3

        finally:
            # 清理剩余的备份文件
            for backup_path in backup_paths:
                if os.path.exists(backup_path):
                    os.unlink(backup_path)
                metadata_path = backup_path.replace('.xlsx', '_metadata.json')
                if os.path.exists(metadata_path):
                    os.unlink(metadata_path)


class TestBackupIntegration:
    """备份集成测试类"""

    @pytest.fixture
    def temp_excel_file(self):
        """创建临时Excel文件用于测试"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            file_path = tmp.name

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "GameData"

        # 游戏数据示例
        ws.append(["技能ID", "技能名称", "技能类型", "技能伤害", "冷却时间"])
        ws.append([1001, "火球术", "火系", 150, 3.0])
        ws.append([1002, "冰箭", "冰系", 120, 2.5])
        ws.append([1003, "雷击", "雷系", 180, 4.0])

        wb.save(file_path)
        wb.close()

        yield file_path

        # 清理
        if os.path.exists(file_path):
            os.unlink(file_path)

    def test_backup_before_high_risk_operation(self, temp_excel_file):
        """测试高风险操作前自动创建备份"""
        # 模拟高风险操作
        new_data = [[f"修改数据{i}" for i in range(1, 11)] for _ in range(50)]

        # 评估操作影响（应该是高风险）
        impact = ExcelOperations.assess_operation_impact(
            file_path=temp_excel_file,
            range_expression="GameData!A1:E50",
            operation_type="update",
            preview_data=new_data
        )

        assert impact['success'] is True
        assert impact['impact_analysis']['operation_risk_level'] in ['high', 'critical']

        # 为高风险操作创建备份
        backup_result = ExcelOperations.create_auto_backup(
            file_path=temp_excel_file,
            backup_name="high_risk_operation_backup",
            backup_reason="高风险操作前自动备份",
            user_id="integration_test"
        )

        assert backup_result['success'] is True
        backup_path = backup_result['backup_path']

        try:
            # 执行高风险操作（模拟）
            # 注意：这里只是测试集成，实际的高风险操作会被安全检查阻止

            # 验证备份文件存在且完整
            assert os.path.exists(backup_path)
            assert backup_result['checksum'] is not None

            # 测试从备份恢复
            restore_result = ExcelOperations.restore_from_backup(
                original_file_path=temp_excel_file,
                backup_path=backup_path
            )

            assert restore_result['success'] is True

        finally:
            # 清理备份文件
            if os.path.exists(backup_path):
                os.unlink(backup_path)
            metadata_path = backup_path.replace('.xlsx', '_metadata.json')
            if os.path.exists(metadata_path):
                os.unlink(metadata_path)

    def test_complete_backup_workflow(self, temp_excel_file):
        """测试完整的备份工作流程"""
        backup_paths = []

        try:
            # 1. 创建初始备份
            backup1 = ExcelOperations.create_auto_backup(
                file_path=temp_excel_file,
                backup_name="initial_backup",
                backup_reason="初始备份",
                user_id="workflow_test"
            )
            assert backup1['success'] is True
            backup_paths.append(backup1['backup_path'])

            # 2. 修改文件
            from openpyxl import load_workbook
            wb = load_workbook(temp_excel_file)
            ws = wb.active
            ws.append([1004, "新技能", "风系", 100, 2.0])  # 添加新技能
            wb.save(temp_excel_file)
            wb.close()

            # 3. 创建修改后备份
            backup2 = ExcelOperations.create_auto_backup(
                file_path=temp_excel_file,
                backup_name="modified_backup",
                backup_reason="修改后备份",
                user_id="workflow_test"
            )
            assert backup2['success'] is True
            backup_paths.append(backup2['backup_path'])

            # 4. 列出所有备份
            backups = ExcelOperations.list_backups(temp_excel_file)
            assert backups['success'] is True
            assert len(backups['backups']) == 2

            # 5. 恢复到初始状态
            restore_result = ExcelOperations.restore_from_backup(
                original_file_path=temp_excel_file,
                backup_path=backup1['backup_path']
            )
            assert restore_result['success'] is True

            # 6. 验证恢复结果
            wb_restored = load_workbook(temp_excel_file)
            ws_restored = wb_restored.active
            rows = list(ws_restored.iter_rows(values_only=True))
            wb_restored.close()

            assert len(rows) == 4  # 应该回到初始的4行（包括表头）

            # 7. 清理旧备份，只保留最新的
            cleanup_result = ExcelOperations.cleanup_old_backups(
                file_path=temp_excel_file,
                keep_count=1
            )
            assert cleanup_result['success'] is True
            assert cleanup_result['cleaned_count'] == 1

        finally:
            # 清理所有备份文件
            for backup_path in backup_paths:
                if os.path.exists(backup_path):
                    os.unlink(backup_path)
                metadata_path = backup_path.replace('.xlsx', '_metadata.json')
                if os.path.exists(metadata_path):
                    os.unlink(metadata_path)

    def test_backup_error_handling(self, temp_excel_file):
        """测试备份错误处理"""
        # 测试权限错误（模拟）
        with patch('builtins.open', side_effect=PermissionError("权限不足")):
            result = ExcelOperations.create_auto_backup(
                file_path=temp_excel_file,
                backup_name="permission_test",
                backup_reason="权限测试",
                user_id="error_test"
            )

            assert result['success'] is False
            assert "权限" in result['error'].lower() or "无法" in result['error']

        # 测试磁盘空间不足（模拟）
        with patch('shutil.copy2', side_effect=OSError("磁盘空间不足")):
            result = ExcelOperations.create_auto_backup(
                file_path=temp_excel_file,
                backup_name="space_test",
                backup_reason="空间测试",
                user_id="error_test"
            )

            assert result['success'] is False
            assert "空间" in result['error'] or "磁盘" in result['error']


if __name__ == "__main__":
    # 运行测试
    pytest.main([__file__, "-v"])