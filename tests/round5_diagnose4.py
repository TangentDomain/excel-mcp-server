"""追踪极小浮点数 0.000001 在管道中哪里变成 0.0"""

import sys

sys.path.insert(0, "/root/workspace/excel-mcp-server/src")
from openpyxl import Workbook

from excel_mcp_server_fastmcp.api.advanced_sql_query import AdvancedSQLQueryEngine

# 创建测试文件
wb = Workbook()
ws = wb.active
ws.title = "test"
ws.append(["ID", "Val"])
ws.append([1, 0.000001])
ws.append([2, 8.88e-2])  # 0.0888
wb.save("/tmp/r5_float_test.xlsx")

fp = "/tmp/r5_float_test.xlsx"
engine = AdvancedSQLQueryEngine()

# Step 1: openpyxl 直接读
print("=== Step 1: openpyxl 直接读取 ===")
from openpyxl import load_workbook

wb2 = load_workbook(fp)
ws2 = wb2["test"]
for row in ws2.iter_rows(min_row=2, values_only=True):
    print(f"  raw: Val={row[1]!r} (type={type(row[1]).__name__})")

# Step 2: engine 的 _load_workbook
print("\n=== Step 2: _load_workbook ===")
worksheets_data = engine._load_workbook(fp)
df = worksheets_data["test"]
print(f"  df:\n{df}")
print(f"  Val dtype: {df['Val'].dtype}")
for i, v in enumerate(df["Val"].values):
    print(f"  row {i}: Val={v!r}")

# Step 3: 查询结果
print("\n=== Step 3: SELECT 结果 ===")
from excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

r = execute_advanced_sql_query(fp, "SELECT ID, Val FROM test")
print(f"  result: {r['data']}")
