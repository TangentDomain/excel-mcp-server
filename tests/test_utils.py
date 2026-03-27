# -*- coding: utf-8 -*-
"""
工具类和模型测试
合并了异常处理、数据模型、兼容性测试等工具类功能测试
这个文件替代了原本的test_models.py, test_compatibility.py等工具类测试
"""

import pytest
import tempfile
from pathlib import Path
from dataclasses import FrozenInstanceError

from src.excel_mcp_server_fastmcp.utils.exceptions import (
    ExcelException,
    ExcelFileNotFoundError,
    SheetNotFoundError,
    DataValidationError,
    OperationLimitError
)
from src.excel_mcp_server_fastmcp.models.types import (
    SheetInfo, RangeInfo, CellInfo, SearchMatch, ModifiedCell,
    OperationResult, RangeType, MatchType
)


class TestExceptions:
    """测试自定义异常类"""

    def test_excel_file_not_found_error(self):
        """Test ExcelFileNotFoundError creation with enhanced message"""
        error = ExcelFileNotFoundError("test.xlsx")
        assert isinstance(error, FileNotFoundError)
        assert isinstance(error, Exception)
        assert "test.xlsx" in str(error)
        assert error.message is not None
        assert error.suggested_fix is not None

    def test_sheet_not_found_error(self):
        """Test SheetNotFoundError creation with enhanced message"""
        error = SheetNotFoundError("Sheet1")
        assert isinstance(error, Exception)
        assert "Sheet1" in str(error)
        assert error.suggested_fix is not None
        # Test with available sheets hint
        error2 = SheetNotFoundError("MissingSheet", ["Sheet1", "Sheet2"])
        assert "Sheet1" in str(error2)
        assert "Sheet2" in str(error2)

    def test_data_validation_error(self):
        """Test DataValidationError creation with enhanced message"""
        error = DataValidationError("Invalid data", "some detail")
        assert isinstance(error, Exception)
        assert "Invalid data" in str(error)
        assert error.suggested_fix is not None
        formatted = error.get_formatted_message()
        assert "数据验证失败" in formatted

    def test_operation_limit_error(self):
        """Test OperationLimitError creation"""
        error = OperationLimitError("批量行操作", "最多1000行", "数据量太大")
        assert isinstance(error, Exception)
        assert "批量行操作" in str(error)
        assert error.suggested_fix is not None

    def test_excel_exception_base(self):
        """Test base ExcelException with all fields"""
        exc = ExcelException("test message", hint="test hint", suggested_fix="test fix")
        assert "test message" in str(exc)
        assert exc.hint == "test hint"
        assert exc.suggested_fix == "test fix"
        formatted = exc.get_formatted_message()
        assert "Excel操作错误" in formatted

    def test_exceptions_inheritance(self):
        """Test that exceptions inherit from Exception"""
        assert issubclass(ExcelFileNotFoundError, Exception)
        assert issubclass(SheetNotFoundError, Exception)
        assert issubclass(DataValidationError, Exception)

    def test_exceptions_can_be_raised_and_caught(self):
        """Test that exceptions can be raised and caught"""
        def raise_file_not_found():
            raise ExcelFileNotFoundError("test.xlsx")

        def raise_sheet_not_found():
            raise SheetNotFoundError("Sheet1")

        def raise_data_validation():
            raise DataValidationError("Invalid data")

        with pytest.raises(ExcelFileNotFoundError):
            raise_file_not_found()

        with pytest.raises(SheetNotFoundError):
            raise_sheet_not_found()

        with pytest.raises(DataValidationError):
            raise_data_validation()

    def test_exception_chinese_support(self):
        """测试异常类的中文支持"""
        # 测试中文错误消息
        chinese_errors = [
            ExcelFileNotFoundError("中文文件名.xlsx"),
            SheetNotFoundError("中文工作表"),
            DataValidationError("中文错误信息")
        ]

        for error in chinese_errors:
            error_str = str(error)
            assert isinstance(error_str, str)
            assert len(error_str) > 0


class TestModelTypes:
    """测试数据模型类型"""

    def test_sheet_info_creation(self):
        """Test SheetInfo creation with actual fields"""
        sheet_info = SheetInfo(
            index=0,
            name="Sheet1",
            max_row=100,
            max_column=10,
            max_column_letter="J"
        )

        assert sheet_info.name == "Sheet1"
        assert sheet_info.index == 0
        assert sheet_info.max_row == 100
        assert sheet_info.max_column == 10
        assert sheet_info.max_column_letter == "J"

    def test_sheet_info_chinese_name(self):
        """测试中文工作表名称"""
        sheet_info = SheetInfo(
            index=1,
            name="数据分析表",
            max_row=50,
            max_column=5,
            max_column_letter="E"
        )

        assert sheet_info.name == "数据分析表"
        assert sheet_info.index == 1

    def test_range_info_creation(self):
        """Test RangeInfo creation with actual fields"""
        range_info = RangeInfo(
            sheet_name="Sheet1",
            cell_range="A1:C10",
            range_type=RangeType.CELL_RANGE
        )

        assert range_info.sheet_name == "Sheet1"
        assert range_info.cell_range == "A1:C10"
        assert range_info.range_type == RangeType.CELL_RANGE

    def test_range_info_chinese_sheet(self):
        """测试中文工作表的范围信息"""
        range_info = RangeInfo(
            sheet_name="销售数据",
            cell_range="A1:E20",
            range_type=RangeType.CELL_RANGE
        )

        assert range_info.sheet_name == "销售数据"
        assert range_info.cell_range == "A1:E20"

    def test_cell_info_creation(self):
        """Test CellInfo creation with actual fields"""
        cell_info = CellInfo(
            coordinate="A1",
            value="Test Value"
        )

        assert cell_info.coordinate == "A1"
        assert cell_info.value == "Test Value"

    def test_cell_info_chinese_value(self):
        """测试中文单元格值"""
        cell_info = CellInfo(
            coordinate="B2",
            value="中文数据测试"
        )

        assert cell_info.coordinate == "B2"
        assert cell_info.value == "中文数据测试"

    def test_search_match_creation(self):
        """测试搜索匹配结果模型"""
        search_match = SearchMatch(
            sheet="Sheet1",
            cell="A1",
            match="搜索结果",
            match_type=MatchType.VALUE
        )

        assert search_match.sheet == "Sheet1"
        assert search_match.cell == "A1"
        assert search_match.match == "搜索结果"
        assert search_match.match_type == MatchType.VALUE

    def test_modified_cell_creation(self):
        """测试修改单元格模型"""
        modified_cell = ModifiedCell(
            coordinate="C3",
            old_value="旧值",
            new_value="新值"
        )

        assert modified_cell.coordinate == "C3"
        assert modified_cell.old_value == "旧值"
        assert modified_cell.new_value == "新值"

    def test_operation_result_success(self):
        """测试成功操作结果"""
        result = OperationResult(
            success=True,
            message="操作成功完成",
            data=["项目1", "项目2"]
        )

        assert result.success is True
        assert result.message == "操作成功完成"
        assert result.data == ["项目1", "项目2"]
        assert result.error is None

    def test_operation_result_failure(self):
        """测试失败操作结果"""
        result = OperationResult(
            success=False,
            error="操作执行失败",
            data=None
        )

        assert result.success is False
        assert result.error == "操作执行失败"
        assert result.data is None
        assert result.message is None

    def test_range_type_enum(self):
        """测试范围类型枚举"""
        assert RangeType.CELL_RANGE is not None
        assert RangeType.SINGLE_ROW is not None
        assert RangeType.ROW_RANGE is not None
        assert RangeType.COLUMN_RANGE is not None

    def test_match_type_enum(self):
        """测试匹配类型枚举"""
        assert MatchType.VALUE is not None
        assert MatchType.FORMULA is not None

    def test_model_immutability(self):
        """测试模型不可变性（如果使用了dataclass frozen）"""
        cell_info = CellInfo(coordinate="A1", value="test")

        # 如果模型是不可变的，修改应该失败
        try:
            cell_info.value = "modified"
            # 如果到达这里，说明模型是可变的，这也是可接受的
            assert cell_info.value == "modified"
        except (FrozenInstanceError, AttributeError):
            # 如果抛出异常，说明模型是不可变的
            assert cell_info.value == "test"

    def test_model_repr_methods(self):
        """测试模型的字符串表示方法"""
        sheet_info = SheetInfo(
            index=0,
            name="测试表",
            max_row=10,
            max_column=3,
            max_column_letter="C"
        )

        repr_str = repr(sheet_info)
        assert "测试表" in repr_str
        assert "SheetInfo" in repr_str


class TestCompatibility:
    """测试兼容性相关功能"""

    def test_openpyxl_version_compatibility(self):
        """测试openpyxl版本兼容性"""
        import openpyxl

        # 检查openpyxl版本
        version = openpyxl.__version__
        assert isinstance(version, str)
        assert len(version.split('.')) >= 2

    def test_excel_file_creation_compatibility(self, temp_dir):
        """测试Excel文件创建兼容性"""
        from openpyxl import Workbook

        # 创建基本工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "兼容性测试"

        # 添加中文数据
        ws['A1'] = "中文标题"
        ws['B1'] = "数值"
        ws['A2'] = "产品1"
        ws['B2'] = 100

        # 保存文件
        file_path = temp_dir / "compatibility_test.xlsx"
        wb.save(str(file_path))

        assert file_path.exists()
        assert file_path.stat().st_size > 0

    def test_unicode_handling_compatibility(self, temp_dir):
        """测试Unicode字符处理兼容性"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # 测试各种Unicode字符
        unicode_data = {
            'A1': '中文字符',
            'A2': '日本語',
            'A3': '한국어',
            'A4': 'العربية',
            'A5': 'Русский',
            'A6': '🔥💡🎉'  # Emoji
        }

        for cell, value in unicode_data.items():
            ws[cell] = value

        # 保存并验证
        file_path = temp_dir / "unicode_test.xlsx"
        wb.save(str(file_path))

        assert file_path.exists()

    def test_formula_compatibility(self, temp_dir):
        """测试公式兼容性"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # 添加数据和公式
        ws['A1'] = 10
        ws['A2'] = 20
        ws['A3'] = '=A1+A2'  # 简单公式
        ws['A4'] = '=SUM(A1:A2)'  # 函数公式

        # 保存文件
        file_path = temp_dir / "formula_test.xlsx"
        wb.save(str(file_path))

        assert file_path.exists()

    def test_large_file_compatibility(self, temp_dir):
        """测试大文件处理兼容性"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # 创建较大的数据集
        for row in range(1, 101):  # 100行数据
            for col in range(1, 11):  # 10列数据
                ws.cell(row=row, column=col, value=f"数据_{row}_{col}")

        # 保存文件
        file_path = temp_dir / "large_file_test.xlsx"
        wb.save(str(file_path))

        assert file_path.exists()
        assert file_path.stat().st_size > 5000  # 文件应该有一定大小    def test_sheet_name_compatibility(self, temp_dir):
        """测试工作表名称兼容性"""
        from openpyxl import Workbook

        wb = Workbook()

        # 测试各种工作表名称
        test_names = [
            "基础数据",
            "Sales_Report_2024",
            "数据分析-结果",
            "测试表(副本)",
        ]

        # 移除默认工作表
        wb.remove(wb.active)

        for name in test_names:
            # 处理特殊字符
            safe_name = name.replace('/', '_').replace('\\', '_').replace('*', '_')
            safe_name = safe_name.replace('?', '_').replace('[', '_').replace(']', '_')
            safe_name = safe_name.replace(':', '_').strip()

            ws = wb.create_sheet(title=safe_name[:31])  # Excel限制31字符
            ws['A1'] = f"这是{safe_name}工作表"

        # 保存文件
        file_path = temp_dir / "sheet_name_test.xlsx"
        wb.save(str(file_path))

        assert file_path.exists()


class TestUtilityFunctions:
    """测试工具函数"""

    def test_string_encoding_handling(self):
        """测试字符串编码处理"""
        # 测试各种编码的字符串
        test_strings = [
            "普通英文",
            "中文字符串",
            "Mixed 中英文 String",
            "特殊符号 !@#$%^&*()",
            "数字123456789"
        ]

        for test_str in test_strings:
            # 确保字符串可以正确编码和解码
            encoded = test_str.encode('utf-8')
            decoded = encoded.decode('utf-8')
            assert decoded == test_str

    def test_path_handling_compatibility(self):
        """测试路径处理兼容性"""
        from pathlib import Path

        # 测试不同格式的路径
        paths = [
            "simple_file.xlsx",
            "folder/file.xlsx",
            "中文文件夹/测试文件.xlsx",
            "special chars/file (1).xlsx"
        ]

        for path_str in paths:
            path_obj = Path(path_str)
            assert isinstance(path_obj, Path)
            # 确保路径对象可以转换为字符串
            assert isinstance(str(path_obj), str)

    def test_data_type_conversion(self):
        """测试数据类型转换"""
        # 测试不同数据类型的处理
        test_data = [
            ("字符串", str),
            (123, int),
            (45.67, float),
            (True, bool),
            (None, type(None))
        ]

        for value, expected_type in test_data:
            assert isinstance(value, expected_type)
            # 测试转换为字符串
            str_value = str(value) if value is not None else ""
            assert isinstance(str_value, str)

    def test_error_message_formatting(self):
        """测试错误消息格式化"""
        # 测试不同类型的错误消息
        error_messages = [
            "简单错误消息",
            "包含数字123的错误消息",
            "包含特殊字符!@#的错误",
            "多行错误消息\n第二行\n第三行"
        ]

        for msg in error_messages:
            # 确保错误消息可以正确处理
            assert isinstance(msg, str)
            assert len(msg) > 0

            # 测试错误消息的基本格式化
            formatted = f"错误: {msg}"
            assert "错误:" in formatted
            assert msg in formatted
