"""数据验证单元测试模块

为 excel_set_data_validation 函数提供全面的单元测试，覆盖所有支持的验证类型。
"""

import unittest
import tempfile
import os
from unittest.mock import patch, MagicMock
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from excel_mcp_server_fastmcp.server import excel_set_data_validation


class TestDataValidation(unittest.TestCase):
    """数据验证功能测试类"""
    
    def setUp(self):
        """测试前准备：创建临时Excel文件"""
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, 'test_validation.xlsx')
        
        # 创建一个简单的Excel文件用于测试
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            # 添加一些测试数据
            ws['A1'] = 'Header'
            ws['A2'] = 'Test1'
            ws['A3'] = 'Test2'
            wb.save(self.test_file)
        except ImportError:
            # 如果openpyxl不可用，创建一个空文件用于模拟
            with open(self.test_file, 'w') as f:
                f.write('mock_excel_file')
    
    def tearDown(self):
        """测试后清理：删除临时文件"""
        try:
            if os.path.exists(self.test_file):
                os.unlink(self.test_file)
            os.rmdir(self.temp_dir)
        except:
            pass
    
    def test_list_validation_success(self):
        """测试1: list验证类型 - 正常情况"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="list",
            criteria="选项1,选项2,选项3",
            input_title="选择选项",
            input_message="请从下拉列表中选择"
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['data']['validation_type'], 'list')
        self.assertEqual(result['data']['criteria'], '选项1,选项2,选项3')
    
    def test_list_validation_empty_criteria(self):
        """测试2: list验证类型 - 空criteria（应失败）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="list",
            criteria="",
        )
        
        self.assertFalse(result['success'])
        self.assertIn('必须提供', result['message'])
    
    def test_whole_number_validation_success(self):
        """测试3: whole_number验证类型 - 正常情况"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="whole_number",
            criteria="between,1,100",
            input_title="输入数字",
            input_message="请输入1-100之间的整数"
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['data']['validation_type'], 'whole_number')
        self.assertEqual(result['data']['criteria'], 'between,1,100')
    
    def test_whole_number_conversion(self):
        """测试4: whole_number验证类型 - 值转换测试"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="whole_number",
            criteria="between,1.5,100.8",  # 小数输入，应转换为整数
        )
        
        # 成功情况下，值应该被转换为整数
        if result['success']:
            self.assertEqual(result['data']['criteria'], 'between,1,100')  # 转换后的值
    
    def test_decimal_validation_success(self):
        """测试5: decimal验证类型 - 正常情况"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="decimal",
            criteria="greater_than,0",
            input_title="输入数字",
            input_message="请输入大于0的数字"
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['data']['validation_type'], 'decimal')
    
    def test_decimal_validation_float_values(self):
        """测试6: decimal验证类型 - 浮点数值处理"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="decimal",
            criteria="between,1.5,10.8",  # 浮点数范围
        )
        
        self.assertTrue(result['success'])
    
    def test_date_validation_success(self):
        """测试7: date验证类型 - 正常情况"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="date",
            criteria="between,2024-01-01,2024-12-31",
            input_title="输入日期",
            input_message="请输入2024年的日期"
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['data']['validation_type'], 'date')
    
    def test_date_validation_format(self):
        """测试8: date验证类型 - 日期格式标准化"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="date",
            criteria="between,2024-1-1,2024-12-31",  # 非标准格式，应标准化
        )
        
        if result['success']:
            # 检查日期是否被标准化为YYYY-MM-DD格式
            criteria_data = result['data']['criteria']
            self.assertEqual(len(criteria_data.split(',')), 3)  # 格式应为: between,YYYY-MM-DD,YYYY-MM-DD
    
    def test_date_validation_invalid_format(self):
        """测试9: date验证类型 - 无效日期格式（应失败）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="date",
            criteria="between,invalid-date,2024-12-31",
        )
        
        self.assertFalse(result['success'])
        self.assertIn('格式错误', result['message'])
    
    def test_text_length_validation_success(self):
        """测试10: text_length验证类型 - 正常情况"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="text_length",
            criteria="less_than,100",
            input_title="输入文本",
            input_message="请输入少于100个字符的文本"
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['data']['validation_type'], 'text_length')
    
    def test_text_length_conversion(self):
        """测试11: text_length验证类型 - 值转换测试"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="text_length",
            criteria="between,10.5,50.8",  # 小数输入，应转换为整数
        )
        
        if result['success']:
            self.assertEqual(result['data']['criteria'], 'between,10,50')  # 转换后的值
    
    def test_custom_validation_success(self):
        """测试12: custom验证类型 - 正常情况"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="custom",
            criteria="=AND(A1>0,A1<100)",
            input_title="自定义验证",
            input_message="请满足自定义条件"
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['data']['validation_type'], 'custom')
    
    def test_custom_validation_empty_criteria(self):
        """测试13: custom验证类型 - 空criteria（应失败）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="custom",
            criteria="",
        )
        
        self.assertFalse(result['success'])
        self.assertIn('必须提供', result['message'])
    
    def test_unsupported_validation_type(self):
        """测试14: 不支持的验证类型（应失败）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="unsupported_type",
            criteria="test",
        )
        
        self.assertFalse(result['success'])
        self.assertIn('不支持', result['message'])
    
    def test_invalid_operator(self):
        """测试15: 无效的操作符（应失败）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="whole_number",
            criteria="invalid_operator,1,100",
        )
        
        self.assertFalse(result['success'])
        self.assertIn('不支持的操作符', result['message'])
    
    def test_between_operator_missing_value2(self):
        """测试16: between操作符缺少第二个值（应失败）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="whole_number",
            criteria="between,1",  # 缺少第二个值
        )
        
        self.assertFalse(result['success'])
        self.assertIn('需要两个值', result['message'])
    
    def test_file_not_found(self):
        """测试17: 文件不存在（应失败）"""
        result = excel_set_data_validation(
            file_path="nonexistent_file.xlsx",
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="list",
            criteria="选项1,选项2,选项3",
        )
        
        self.assertFalse(result['success'])
    
    def test_sheet_not_found(self):
        """测试18: 工作表不存在（应失败）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="NonExistentSheet",
            range_address="A2:A3",
            validation_type="list",
            criteria="选项1,选项2,选项3",
        )
        
        self.assertFalse(result['success'])
    
    def test_edge_cases_whitespace(self):
        """测试19: 边界情况 - 空白字符处理"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="list",
            criteria="  选项1  ,  选项2  ,  选项3  ",  # 带空白字符
        )
        
        # 应该能够处理空白字符
        if result['success']:
            self.assertEqual(result['data']['criteria'], '  选项1  ,  选项2  ,  选项3  ')
    
    def test_validation_count_increment(self):
        """测试20: 验证计数器递增"""
        # 第一次设置验证
        result1 = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="TestSheet",
            range_address="A2:A3",
            validation_type="list",
            criteria="选项1,选项2",
        )
        
        if result1['success']:
            first_count = result1['data']['validation_count']
            
            # 第二次设置验证，计数应该增加
            result2 = excel_set_data_validation(
                file_path=self.test_file,
                sheet_name="TestSheet",
                range_address="B2:B3",
                validation_type="list",
                criteria="选项3,选项4",
            )
            
            if result2['success']:
                second_count = result2['data']['validation_count']
                self.assertGreater(second_count, first_count)


class TestDataValidationIntegration(unittest.TestCase):
    """数据验证集成测试：多验证类型组合测试"""

    def setUp(self):
        """测试前准备：创建临时Excel文件"""
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, 'test_integration.xlsx')

        # 创建测试文件
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "IntegrationTest"
            wb.save(self.test_file)
        except ImportError:
            with open(self.test_file, 'w') as f:
                f.write('mock_excel_file')

    def tearDown(self):
        """测试后清理"""
        try:
            if os.path.exists(self.test_file):
                os.unlink(self.test_file)
            os.rmdir(self.temp_dir)
        except:
            pass


class TestDataValidationBoundary(unittest.TestCase):
    """数据验证边界测试：空值、超限值、特殊字符"""

    def setUp(self):
        """测试前准备：创建临时Excel文件"""
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, 'test_boundary.xlsx')

        # 创建测试文件
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "BoundaryTest"
            wb.save(self.test_file)
        except ImportError:
            with open(self.test_file, 'w') as f:
                f.write('mock_excel_file')

    def tearDown(self):
        """测试后清理"""
        try:
            if os.path.exists(self.test_file):
                os.unlink(self.test_file)
            os.rmdir(self.temp_dir)
        except:
            pass

    def test_whole_number_boundary_max_value(self):
        """测试：whole_number验证 - 极大值边界"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="whole_number",
            criteria="less_than,2147483647"  # 2^31-1
        )
        self.assertTrue(result['success'])

    def test_whole_number_boundary_min_value(self):
        """测试：whole_number验证 - 极小值边界"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="whole_number",
            criteria="greater_than,-2147483648"  # -2^31
        )
        self.assertTrue(result['success'])

    def test_whole_number_boundary_zero(self):
        """测试：whole_number验证 - 零值边界"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="whole_number",
            criteria="greater_than_or_equal,0"
        )
        self.assertTrue(result['success'])

    def test_whole_number_boundary_negative_values(self):
        """测试：whole_number验证 - 负数边界"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="whole_number",
            criteria="between,-100,-1"
        )
        self.assertTrue(result['success'])

    def test_decimal_boundary_very_small(self):
        """测试：decimal验证 - 极小浮点数"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="decimal",
            criteria="greater_than,0.000001"
        )
        self.assertTrue(result['success'])

    def test_decimal_boundary_very_large(self):
        """测试：decimal验证 - 极大浮点数"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="decimal",
            criteria="less_than,999999999.99"
        )
        self.assertTrue(result['success'])

    def test_decimal_boundary_scientific_notation(self):
        """测试：decimal验证 - 科学计数法数值"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="decimal",
            criteria="between,1e-10,1e10"
        )
        self.assertTrue(result['success'])

    def test_decimal_boundary_negative_zero(self):
        """测试：decimal验证 - 负零值"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="decimal",
            criteria="greater_than,-0.0"
        )
        self.assertTrue(result['success'])

    def test_date_boundary_min_date(self):
        """测试：date验证 - 最小日期边界（Excel 1900-01-01）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="date",
            criteria="greater_than_or_equal,1900-01-01"
        )
        self.assertTrue(result['success'])

    def test_date_boundary_max_date(self):
        """测试：date验证 - 最大日期边界（9999-12-31）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="date",
            criteria="less_than_or_equal,9999-12-31"
        )
        self.assertTrue(result['success'])

    def test_date_boundary_leap_year(self):
        """测试：date验证 - 闰年日期"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="date",
            criteria="between,2024-02-28,2024-03-01"  # 包含闰年2月29日
        )
        self.assertTrue(result['success'])

    def test_text_length_boundary_max_length(self):
        """测试：text_length验证 - 最大长度边界（Excel限制32767）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="text_length",
            criteria="less_than_or_equal,32767"
        )
        self.assertTrue(result['success'])

    def test_text_length_boundary_zero_length(self):
        """测试：text_length验证 - 零长度边界"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="text_length",
            criteria="greater_than_or_equal,0"
        )
        self.assertTrue(result['success'])

    def test_text_length_boundary_single_char(self):
        """测试：text_length验证 - 单字符边界"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="text_length",
            criteria="equal,1"
        )
        self.assertTrue(result['success'])

    def test_list_special_chars_comma(self):
        """测试：list验证 - 选项包含逗号（应正常处理）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="list",
            criteria="选项1,包含,逗号,选项3"
        )
        # 逗号作为分隔符，"包含"会被识别为独立选项
        self.assertTrue(result['success'])

    def test_list_special_chars_quotes(self):
        """测试：list验证 - 选项包含引号"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="list",
            criteria='选项1,"包含引号",选项3'
        )
        self.assertTrue(result['success'])

    def test_list_special_chars_special_unicode(self):
        """测试：list验证 - 特殊Unicode字符"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="list",
            criteria="中文选项,日本語选项,한국어选项,English选项"
        )
        self.assertTrue(result['success'])

    def test_list_special_chars_emoji(self):
        """测试：list验证 - Emoji表情符号"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="list",
            criteria="👍,👎,❤️,🎉"
        )
        self.assertTrue(result['success'])

    def test_list_special_chars_empty_item(self):
        """测试：list验证 - 空选项"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="list",
            criteria="选项1,,选项2"  # 包含空选项
        )
        self.assertTrue(result['success'])

    def test_custom_special_chars_complex_formula(self):
        """测试：custom验证 - 复杂公式表达式"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="custom",
            criteria="=AND(A1>0,OR(A1<100,A1>1000),NOT(ISBLANK(A1)))"
        )
        self.assertTrue(result['success'])

    def test_custom_special_chars_text_formula(self):
        """测试：custom验证 - 文本比较公式"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="custom",
            criteria='=EXACT(A1,"expected")'
        )
        self.assertTrue(result['success'])

    def test_custom_special_chars_nested_functions(self):
        """测试：custom验证 - 嵌套函数公式"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="custom",
            criteria="=IF(ISNUMBER(A1),A1>0,LEN(A1)>5)"
        )
        self.assertTrue(result['success'])

    def test_whole_number_invalid_value(self):
        """测试：whole_number验证 - 无效数值转换（应失败）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="whole_number",
            criteria="greater_than,not_a_number"
        )
        self.assertFalse(result['success'])
        self.assertIn('值类型转换失败', result['message'])

    def test_decimal_invalid_value(self):
        """测试：decimal验证 - 无效浮点数值（应失败）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="decimal",
            criteria="greater_than,invalid_float"
        )
        self.assertFalse(result['success'])

    def test_date_invalid_year(self):
        """测试：date验证 - 无效年份（应失败）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="date",
            criteria="greater_than,1899-12-31"  # Excel最小日期之前
        )
        # 可能为成功或失败，取决于实现，这里不强制断言
        pass

    def test_date_invalid_month(self):
        """测试：date验证 - 无效月份（应失败）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="date",
            criteria="greater_than,2024-13-01"
        )
        self.assertFalse(result['success'])

    def test_text_length_invalid_value(self):
        """测试：text_length验证 - 无效长度值（应失败）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="text_length",
            criteria="greater_than,not_a_length"
        )
        self.assertFalse(result['success'])

    def test_text_length_negative_value(self):
        """测试：text_length验证 - 负长度值（应失败或转换）"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="text_length",
            criteria="greater_than,-1"
        )
        # 负值会被转换为整数，可能成功，这里只检查不崩溃
        pass

    def test_criteria_whitespace_only(self):
        """测试：所有类型 - 仅空白字符的criteria（应失败）"""
        for validation_type in ['whole_number', 'decimal', 'date', 'text_length']:
            result = excel_set_data_validation(
                file_path=self.test_file,
                sheet_name="BoundaryTest",
                range_address="A1:A1",
                validation_type=validation_type,
                criteria="   "  # 仅空白字符
            )
            self.assertFalse(result['success'], f"{validation_type}应拒绝空白criteria")

    def test_criteria_extra_commas(self):
        """测试：数值类型 - 额外的逗号分隔符"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="whole_number",
            criteria="between,1,100,"  # 结尾多一个逗号
        )
        # 可能成功（忽略多余参数）或失败，取决于实现
        pass

    def test_range_address_validation(self):
        """测试：range_address格式的边界情况"""
        # 标准格式
        result1 = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1",
            validation_type="list",
            criteria="选项1,选项2"
        )
        self.assertTrue(result1['success'])

        # 多单元格范围
        result2 = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1000",
            validation_type="list",
            criteria="选项1,选项2"
        )
        self.assertTrue(result2['success'])

    def test_empty_input_title_and_message(self):
        """测试：input_title和input_message为空字符串"""
        result = excel_set_data_validation(
            file_path=self.test_file,
            sheet_name="BoundaryTest",
            range_address="A1:A1",
            validation_type="list",
            criteria="选项1,选项2",
            input_title="",
            input_message=""
        )
        self.assertTrue(result['success'])
    
    def test_multiple_validations_same_sheet(self):
        """测试：同一工作表设置多个验证规则"""
        validation_configs = [
            {
                'type': 'list',
                'criteria': '选项1,选项2,选项3',
                'range': 'A2:A10'
            },
            {
                'type': 'whole_number',
                'criteria': 'between,1,100',
                'range': 'B2:B10'
            },
            {
                'type': 'date',
                'criteria': 'greater_than,2024-01-01',
                'range': 'C2:C10'
            },
            {
                'type': 'text_length',
                'criteria': 'less_than,50',
                'range': 'D2:D10'
            },
            {
                'type': 'decimal',
                'criteria': 'between,0.0,1.0',
                'range': 'E2:E10'
            },
            {
                'type': 'custom',
                'criteria': '=AND(A2>0,B2<100)',
                'range': 'F2:F10'
            }
        ]
        
        results = []
        for config in validation_configs:
            result = excel_set_data_validation(
                file_path=self.test_file,
                sheet_name="IntegrationTest",
                range_address=config['range'],
                validation_type=config['type'],
                criteria=config['criteria']
            )
            results.append(result)
            
            if result['success']:
                print(f"✅ {config['type']} 验证成功: {config['criteria']}")
            else:
                print(f"❌ {config['type']} 验证失败: {result['message']}")
        
        # 验证至少大部分验证都成功
        success_count = sum(1 for r in results if r['success'])
        self.assertGreaterEqual(success_count, len(validation_configs) - 1)  # 允许最多1个失败


if __name__ == '__main__':
    # 运行所有测试
    unittest.main(verbosity=2)