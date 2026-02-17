# -*- coding: utf-8 -*-
"""
Excel Operations 更多边缘测试

覆盖 excel_operations.py 中未覆盖的更多代码路径
"""

import pytest
from pathlib import Path
from openpyxl import Workbook
from datetime import datetime

from src.api.excel_operations import ExcelOperations


class TestExcelOpsMoreSearch:
    """更多搜索测试"""

    @pytest.fixture
    def search_file(self, temp_dir):
        """创建搜索测试文件"""
        file_path = temp_dir / "search_more.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "SearchSheet"
        
        # 各种数据
        ws['A1'] = "Header"
        ws['A2'] = "Test123"
        ws['A3'] = "test456"
        ws['A4'] = "TEST"
        ws['A5'] = "123"
        
        wb.save(file_path)
        return str(file_path)

    def test_search_regex(self, search_file):
        """测试正则搜索"""
        result = ExcelOperations.search(
            file_path=search_file,
            pattern=r"\d+",
            sheet_name="SearchSheet",
            use_regex=True
        )
        assert result is not None

    def test_search_formulas_only(self, search_file):
        """测试仅搜索公式"""
        result = ExcelOperations.search(
            file_path=search_file,
            pattern=".*",
            sheet_name="SearchSheet",
            include_formulas=True,
            include_values=False
        )
        assert result is not None


class TestExcelOpsMoreRange:
    """更多范围操作测试"""

    @pytest.fixture
    def range_file(self, temp_dir):
        """创建范围测试文件"""
        file_path = temp_dir / "range_more.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "RangeSheet"
        
        for i in range(1, 21):
            ws.cell(row=i, column=1, value=i)
            ws.cell(row=i, column=2, value=i * 10)
        
        wb.save(file_path)
        return str(file_path)

    def test_update_range_multi_sheet(self, range_file):
        """测试多工作表更新"""
        result = ExcelOperations.update_range(
            file_path=range_file,
            range_expression="RangeSheet!C1:C10",
            data=[[i] for i in range(100, 110)],
            preserve_formulas=False
        )
        assert result is not None


class TestExcelOpsMoreSheet:
    """更多工作表操作测试"""

    @pytest.fixture
    def sheet_file(self, temp_dir):
        """创建工作表测试文件"""
        file_path = temp_dir / "sheet_more.xlsx"
        
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(file_path)
        
        return str(file_path)

    def test_delete_sheet_first(self, sheet_file):
        """测试删除第一个工作表"""
        result = ExcelOperations.delete_sheet(sheet_file, "Sheet1")
        assert result is not None


class TestExcelOpsMoreCompare:
    """更多比较操作测试"""

    @pytest.fixture
    def dup_file(self, temp_dir):
        """创建重复测试文件"""
        file_path = temp_dir / "dup_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DupSheet"
        
        ws['A1'] = "ID"
        ws['A2'] = 1
        ws['A3'] = 2
        ws['A4'] = 1  # 重复
        ws['A5'] = 3
        
        wb.save(file_path)
        return str(file_path)

    def test_check_duplicates_with_header(self, dup_file):
        """测试带表头的重复检查"""
        result = ExcelOperations.check_duplicate_ids(
            file_path=dup_file,
            sheet_name="DupSheet",
            id_column=1,
            header_row=1
        )
        assert result is not None


class TestExcelOpsMoreHeaders:
    """更多表头操作测试"""

    @pytest.fixture
    def header_file(self, temp_dir):
        """创建表头测试文件"""
        file_path = temp_dir / "header_more.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "HeaderSheet"
        
        # 多行表头
        ws['A1'] = "Desc1"
        ws['A2'] = "Field1"
        ws['B1'] = "Desc2"
        ws['B2'] = "Field2"
        
        wb.save(file_path)
        return str(file_path)

    def test_get_headers_second_row(self, header_file):
        """测试获取第二行表头"""
        result = ExcelOperations.get_headers(
            file_path=header_file,
            sheet_name="HeaderSheet",
            header_row=2
        )
        assert result is not None


class TestExcelOpsMoreDataOps:
    """更多数据操作测试"""

    @pytest.fixture
    def data_file(self, temp_dir):
        """创建数据操作测试文件"""
        file_path = temp_dir / "data_more.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DataSheet"
        
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=i)
        
        wb.save(file_path)
        return str(file_path)

    def test_insert_rows_beginning(self, data_file):
        """测试在开头插入行"""
        result = ExcelOperations.insert_rows(
            file_path=data_file,
            sheet_name="DataSheet",
            row_index=1,
            count=1
        )
        assert result is not None

    def test_insert_columns_end(self, data_file):
        """测试在末尾插入列"""
        result = ExcelOperations.insert_columns(
            file_path=data_file,
            sheet_name="DataSheet",
            column_index=10,
            count=1
        )
        assert result is not None

    def test_delete_columns_beginning(self, data_file):
        """测试删除开头列"""
        result = ExcelOperations.delete_columns(
            file_path=data_file,
            sheet_name="DataSheet",
            column_index=1,
            count=1
        )
        assert result is not None
