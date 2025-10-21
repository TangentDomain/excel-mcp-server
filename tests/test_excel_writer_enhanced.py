# -*- coding: utf-8 -*-
"""
Excel Writerå¢å¼ºæµ‹è¯•å¥—ä»¶
æµ‹è¯•src.core.excel_writeræ¨¡å—çš„æ‰€æœ‰æ ¸å¿ƒåŠŸèƒ½
ç›®æ ‡è¦†ç›–ç‡ï¼š75%+
"""

import pytest
import tempfile
import os
import time
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment

from src.core.excel_writer import ExcelWriter
from src.models.types import OperationResult, RangeType
from src.utils.exceptions import SheetNotFoundError, DataValidationError


class TestExcelWriterBasic:
    """ExcelWriteråŸºç¡€åŠŸèƒ½æµ‹è¯•"""

    def setup_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•å‰çš„è®¾ç½®"""
        # åˆ›å»ºä¸´æ—¶Excelæ–‡ä»¶
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()

        # åˆ›å»ºåŸºç¡€Excelæ–‡ä»¶
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # æ·»åŠ ä¸€äº›æµ‹è¯•æ•°æ®
        ws['A1'] = "Name"
        ws['B1'] = "Age"
        ws['C1'] = "City"
        ws['A2'] = "Alice"
        ws['B2'] = 25
        ws['C2'] = "Beijing"
        ws['A3'] = "Bob"
        ws['B3'] = 30
        ws['C3'] = "Shanghai"

        wb.save(self.temp_file.name)
        self.file_path = self.temp_file.name

    def teardown_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•åçš„æ¸…ç†"""
        if os.path.exists(self.file_path):
            try:
                os.unlink(self.file_path)
            except:
                pass

    def test_writer_initialization_valid_file(self):
        """æµ‹è¯•æœ‰æ•ˆæ–‡ä»¶è·¯å¾„çš„åˆå§‹åŒ–"""
        writer = ExcelWriter(self.file_path)
        assert writer.file_path == self.file_path

    def test_writer_initialization_invalid_file(self):
        """æµ‹è¯•æ— æ•ˆæ–‡ä»¶è·¯å¾„çš„åˆå§‹åŒ–"""
        with pytest.raises(Exception):
            ExcelWriter("invalid_path.xlsx")

    def test_update_range_single_cell(self):
        """æµ‹è¯•æ›´æ–°å•ä¸ªå•å…ƒæ ¼"""
        writer = ExcelWriter(self.file_path)
        data = [["New Value"]]

        result = writer.update_range("TestSheet!A2", data)

        assert result.success is True
        assert result.metadata['modified_cells_count'] == 1
        assert result.metadata['sheet_name'] == "TestSheet"

    def test_update_range_multiple_cells(self):
        """æµ‹è¯•æ›´æ–°å¤šä¸ªå•å…ƒæ ¼"""
        writer = ExcelWriter(self.file_path)
        data = [["New1", "New2"], ["New3", "New4"]]

        result = writer.update_range("TestSheet!A2:B3", data)

        assert result.success is True
        assert result.metadata['modified_cells_count'] == 4

    def test_update_range_preserve_formulas(self):
        """æµ‹è¯•ä¿ç•™å…¬å¼"""
        writer = ExcelWriter(self.file_path)

        # å…ˆæ·»åŠ ä¸€ä¸ªå…¬å¼
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        ws = wb["TestSheet"]
        ws['D2'] = "=SUM(B2:B3)"
        wb.save(self.file_path)

        # æ›´æ–°å…¶ä»–å•å…ƒæ ¼ï¼Œä¿ç•™å…¬å¼ï¼ˆä½¿ç”¨è¦†ç›–æ¨¡å¼é¿å…è¡Œæ’å…¥å½±å“å…¬å¼ä½ç½®ï¼‰
        data = [["Updated"]]
        result = writer.update_range("TestSheet!A2", data, preserve_formulas=True, insert_mode=False)

        assert result.success is True

        # éªŒè¯å…¬å¼ä»ç„¶å­˜åœ¨
        wb = load_workbook(self.file_path)
        ws = wb["TestSheet"]
        assert ws['D2'].data_type == 'f'  # å…¬å¼ç±»å‹

    def test_update_range_overwrite_formulas(self):
        """æµ‹è¯•è¦†ç›–å…¬å¼"""
        writer = ExcelWriter(self.file_path)

        # å…ˆæ·»åŠ ä¸€ä¸ªå…¬å¼
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        ws = wb["TestSheet"]
        ws['D2'] = "=SUM(B2:B3)"
        wb.save(self.file_path)

        # æ›´æ–°å…¬å¼å•å…ƒæ ¼ï¼Œä¸ä¿ç•™å…¬å¼
        data = [["New Value"]]
        result = writer.update_range("TestSheet!D2", data, preserve_formulas=False)

        assert result.success is True

        # éªŒè¯å…¬å¼è¢«è¦†ç›–
        wb = load_workbook(self.file_path)
        ws = wb["TestSheet"]
        assert ws['D2'].value == "New Value"

    def test_update_range_insert_mode(self):
        """æµ‹è¯•æ’å…¥æ¨¡å¼"""
        writer = ExcelWriter(self.file_path)
        original_rows = 3  # åˆå§‹æœ‰3è¡Œæ•°æ®

        data = [["Insert1", "Insert2"], ["Insert3", "Insert4"]]
        result = writer.update_range("TestSheet!A2", data, insert_mode=True)

        assert result.success is True
        assert result.metadata['insert_mode'] is True
        assert result.metadata['rows_inserted'] == 2

    def test_update_range_invalid_sheet(self):
        """æµ‹è¯•æ— æ•ˆå·¥ä½œè¡¨"""
        writer = ExcelWriter(self.file_path)
        data = [["Test"]]

        result = writer.update_range("NonExistentSheet!A1", data)

        assert result.success is False
        assert "å·¥ä½œè¡¨ä¸å­˜åœ¨" in result.error or "ä¸å­˜åœ¨" in result.error

    def test_convert_to_cell_range_row_range_error(self):
        """æµ‹è¯•è¡ŒèŒƒå›´æ ¼å¼è½¬æ¢é”™è¯¯"""
        writer = ExcelWriter(self.file_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        sheet = wb["TestSheet"]

        with pytest.raises(ValueError) as exc_info:
            writer._convert_to_cell_range("2:5", RangeType.ROW_RANGE, sheet, [["test"]])

        assert "ä¸æ”¯æŒçº¯è¡ŒèŒƒå›´æ ¼å¼" in str(exc_info.value)

    def test_convert_to_cell_range_single_row_error(self):
        """æµ‹è¯•å•è¡ŒèŒƒå›´æ ¼å¼è½¬æ¢é”™è¯¯"""
        writer = ExcelWriter(self.file_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        sheet = wb["TestSheet"]

        with pytest.raises(ValueError) as exc_info:
            writer._convert_to_cell_range("3:3", RangeType.SINGLE_ROW, sheet, [["test"]])

        assert "ä¸æ”¯æŒå•è¡ŒèŒƒå›´æ ¼å¼" in str(exc_info.value)

    def test_convert_to_cell_range_column_range(self):
        """æµ‹è¯•åˆ—èŒƒå›´æ ¼å¼è½¬æ¢"""
        writer = ExcelWriter(self.file_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        sheet = wb["TestSheet"]

        result = writer._convert_to_cell_range("A:C", RangeType.COLUMN_RANGE, sheet, [["test"]])
        assert result == "A:C"

    def test_convert_to_cell_range_single_column(self):
        """æµ‹è¯•å•åˆ—èŒƒå›´æ ¼å¼è½¬æ¢"""
        writer = ExcelWriter(self.file_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        sheet = wb["TestSheet"]

        result = writer._convert_to_cell_range("B:B", RangeType.SINGLE_COLUMN, sheet, [["test"]])
        assert result == "B:B"

    def test_get_worksheet_valid_sheet(self):
        """æµ‹è¯•è·å–æœ‰æ•ˆå·¥ä½œè¡¨"""
        writer = ExcelWriter(self.file_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)

        sheet = writer._get_worksheet(wb, "TestSheet")
        assert sheet.title == "TestSheet"

    def test_get_worksheet_empty_name(self):
        """æµ‹è¯•ç©ºå·¥ä½œè¡¨åç§°"""
        writer = ExcelWriter(self.file_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)

        with pytest.raises(SheetNotFoundError) as exc_info:
            writer._get_worksheet(wb, "")

        assert "å·¥ä½œè¡¨åç§°ä¸èƒ½ä¸ºç©º" in str(exc_info.value)

    def test_get_worksheet_nonexistent_sheet(self):
        """æµ‹è¯•ä¸å­˜åœ¨çš„å·¥ä½œè¡¨"""
        writer = ExcelWriter(self.file_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)

        with pytest.raises(SheetNotFoundError) as exc_info:
            writer._get_worksheet(wb, "NonExistent")

        assert "å·¥ä½œè¡¨ä¸å­˜åœ¨" in str(exc_info.value)

    def test_write_data_basic(self):
        """æµ‹è¯•åŸºç¡€æ•°æ®å†™å…¥"""
        writer = ExcelWriter(self.file_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        sheet = wb["TestSheet"]

        data = [["New1", "New2"], ["New3", "New4"]]
        modified_cells = writer._write_data(sheet, data, 2, 1, False)

        assert len(modified_cells) == 4
        assert modified_cells[0].coordinate == "A2"
        assert modified_cells[0].new_value == "New1"

    def test_write_data_preserve_existing_formula(self):
        """æµ‹è¯•ä¿ç•™ç°æœ‰å…¬å¼"""
        writer = ExcelWriter(self.file_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        sheet = wb["TestSheet"]

        # è®¾ç½®ä¸€ä¸ªå…¬å¼
        sheet['D2'].value = "=SUM(B2:C2)"

        data = [["ShouldNotOverwrite"]]
        modified_cells = writer._write_data(sheet, data, 2, 4, True)  # preserve_formulas=True

        # åº”è¯¥æ²¡æœ‰ä¿®æ”¹ï¼Œå› ä¸ºæ˜¯å…¬å¼
        assert len(modified_cells) == 0

    def test_write_data_overwrite_formula(self):
        """æµ‹è¯•è¦†ç›–å…¬å¼"""
        writer = ExcelWriter(self.file_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        sheet = wb["TestSheet"]

        # è®¾ç½®ä¸€ä¸ªå…¬å¼
        sheet['D2'].value = "=SUM(B2:C2)"

        data = [["NewValue"]]
        modified_cells = writer._write_data(sheet, data, 2, 4, False)  # preserve_formulas=False

        # åº”è¯¥ä¿®æ”¹äº†å…¬å¼
        assert len(modified_cells) == 1
        assert modified_cells[0].new_value == "NewValue"


class TestExcelWriterRowColumnOperations:
    """ExcelWriterè¡Œåˆ—æ“ä½œæµ‹è¯•"""

    def setup_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•å‰çš„è®¾ç½®"""
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # æ·»åŠ æµ‹è¯•æ•°æ®
        for i in range(1, 6):
            for j in range(1, 4):
                ws.cell(row=i, column=j, value=f"R{i}C{j}")

        wb.save(self.temp_file.name)
        self.file_path = self.temp_file.name

    def teardown_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•åçš„æ¸…ç†"""
        if os.path.exists(self.file_path):
            try:
                os.unlink(self.file_path)
            except:
                pass

    def test_insert_rows_basic(self):
        """æµ‹è¯•åŸºç¡€è¡Œæ’å…¥"""
        writer = ExcelWriter(self.file_path)

        result = writer.insert_rows("TestSheet", 3, 2)

        assert result.success is True
        assert result.metadata['inserted_at_row'] == 3
        assert result.metadata['inserted_count'] == 2
        assert result.metadata['new_max_row'] > result.metadata['original_max_row']

    def test_insert_rows_single_row(self):
        """æµ‹è¯•æ’å…¥å•è¡Œ"""
        writer = ExcelWriter(self.file_path)

        result = writer.insert_rows("TestSheet", 2, 1)

        assert result.success is True
        assert result.metadata['inserted_count'] == 1

    def test_insert_columns_basic(self):
        """æµ‹è¯•åŸºç¡€åˆ—æ’å…¥"""
        writer = ExcelWriter(self.file_path)

        result = writer.insert_columns("TestSheet", 2, 2)

        assert result.success is True
        assert result.metadata['inserted_at_column'] == 2
        assert result.metadata['inserted_count'] == 2
        assert result.metadata['new_max_column'] > result.metadata['original_max_column']

    def test_insert_columns_single_column(self):
        """æµ‹è¯•æ’å…¥å•åˆ—"""
        writer = ExcelWriter(self.file_path)

        result = writer.insert_columns("TestSheet", 1, 1)

        assert result.success is True
        assert result.metadata['inserted_count'] == 1

    def test_delete_rows_basic(self):
        """æµ‹è¯•åŸºç¡€è¡Œåˆ é™¤"""
        writer = ExcelWriter(self.file_path)

        result = writer.delete_rows("TestSheet", 3, 2)

        assert result.success is True
        assert result.metadata['deleted_start_row'] == 3
        assert result.metadata['actual_deleted_count'] == 2

    def test_delete_rows_beyond_range(self):
        """æµ‹è¯•åˆ é™¤è¶…å‡ºèŒƒå›´çš„è¡Œ"""
        writer = ExcelWriter(self.file_path)

        result = writer.delete_rows("TestSheet", 100, 5)

        assert result.success is False
        assert "è¶…è¿‡å·¥ä½œè¡¨æœ€å¤§è¡Œæ•°" in result.error

    def test_delete_rows_partial_range(self):
        """æµ‹è¯•éƒ¨åˆ†èŒƒå›´åˆ é™¤"""
        writer = ExcelWriter(self.file_path)

        # ä»ç¬¬4è¡Œå¼€å§‹åˆ é™¤3è¡Œï¼Œä½†åªæœ‰2è¡Œå¯ç”¨
        result = writer.delete_rows("TestSheet", 4, 3)

        assert result.success is True
        assert result.metadata['actual_deleted_count'] == 2  # åªèƒ½åˆ é™¤2è¡Œ

    def test_delete_columns_basic(self):
        """æµ‹è¯•åŸºç¡€åˆ—åˆ é™¤"""
        writer = ExcelWriter(self.file_path)

        result = writer.delete_columns("TestSheet", 2, 1)

        assert result.success is True
        assert result.metadata['deleted_start_column'] == 2
        assert result.metadata['actual_deleted_count'] == 1

    def test_delete_columns_beyond_range(self):
        """æµ‹è¯•åˆ é™¤è¶…å‡ºèŒƒå›´çš„åˆ—"""
        writer = ExcelWriter(self.file_path)

        result = writer.delete_columns("TestSheet", 100, 5)

        assert result.success is False
        assert "è¶…è¿‡å·¥ä½œè¡¨æœ€å¤§åˆ—æ•°" in result.error

    def test_delete_columns_partial_range(self):
        """æµ‹è¯•éƒ¨åˆ†èŒƒå›´åˆ—åˆ é™¤"""
        writer = ExcelWriter(self.file_path)

        # ä»ç¬¬2åˆ—å¼€å§‹åˆ é™¤5åˆ—ï¼Œä½†åªæœ‰3åˆ—å¯ç”¨
        result = writer.delete_columns("TestSheet", 2, 5)

        assert result.success is True
        assert result.metadata['actual_deleted_count'] == 2  # åªèƒ½åˆ é™¤2åˆ—


class TestExcelWriterFormulaOperations:
    """ExcelWriterå…¬å¼æ“ä½œæµ‹è¯•"""

    def setup_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•å‰çš„è®¾ç½®"""
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # æ·»åŠ ä¸€äº›æ•°æ®ç”¨äºå…¬å¼è®¡ç®—
        ws['A1'] = 10
        ws['A2'] = 20
        ws['A3'] = 30
        ws['B1'] = 5
        ws['B2'] = 15
        ws['B3'] = 25

        wb.save(self.temp_file.name)
        self.file_path = self.temp_file.name

    def teardown_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•åçš„æ¸…ç†"""
        if os.path.exists(self.file_path):
            try:
                os.unlink(self.file_path)
            except:
                pass

    def test_set_formula_basic(self):
        """æµ‹è¯•åŸºç¡€å…¬å¼è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_formula("C1", "SUM(A1:A3)", "TestSheet")

        assert result.success is True
        assert result.metadata['cell_address'] == "C1"
        assert result.metadata['formula'] == "SUM(A1:A3)"

    def test_set_formula_with_equals_sign(self):
        """æµ‹è¯•å¸¦ç­‰å·çš„å…¬å¼è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_formula("C1", "=SUM(A1:A3)", "TestSheet")

        assert result.success is True
        assert result.metadata['formula'] == "SUM(A1:A3)"  # ç­‰å·è¢«ç§»é™¤

    def test_set_formula_empty_formula(self):
        """æµ‹è¯•ç©ºå…¬å¼"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_formula("C1", "", "TestSheet")

        assert result.success is False
        assert "å…¬å¼ä¸èƒ½ä¸ºç©º" in result.error

    def test_set_formula_invalid_cell_address(self):
        """æµ‹è¯•æ— æ•ˆå•å…ƒæ ¼åœ°å€"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_formula("INVALID", "SUM(A1:A3)", "TestSheet")

        assert result.success is False
        assert "Invalid cell coordinates" in result.error

    def test_set_formula_nonexistent_sheet(self):
        """æµ‹è¯•ä¸å­˜åœ¨çš„å·¥ä½œè¡¨"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_formula("C1", "SUM(A1:A3)", "NonExistent")

        assert result.success is False
        assert "å·¥ä½œè¡¨ä¸å­˜åœ¨" in result.error or "ä¸å­˜åœ¨" in result.error

    @patch('src.core.excel_writer.logger')
    def test_set_formula_logging(self, mock_logger):
        """æµ‹è¯•å…¬å¼è®¾ç½®æ—¥å¿—è®°å½•"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_formula("C1", "SUM(A1:A3)", "TestSheet")

        assert result.success is True
        # éªŒè¯æ—¥å¿—è®°å½•
        mock_logger.info.assert_called()

    def test_evaluate_formula_basic(self):
        """æµ‹è¯•åŸºç¡€å…¬å¼è®¡ç®—"""
        writer = ExcelWriter(self.file_path)

        result = writer.evaluate_formula("SUM(A1:A3)", "TestSheet")

        # ç»“æœåº”è¯¥æ˜¯10+20+30=60
        assert result.success is True
        assert result.data == 60
        assert result.metadata['result_type'] == "number"

    def test_evaluate_formula_empty_formula(self):
        """æµ‹è¯•ç©ºå…¬å¼è®¡ç®—"""
        writer = ExcelWriter(self.file_path)

        result = writer.evaluate_formula("", "TestSheet")

        assert result.success is False
        assert "å…¬å¼ä¸èƒ½ä¸ºç©º" in result.error

    def test_evaluate_formula_with_equals_sign(self):
        """æµ‹è¯•å¸¦ç­‰å·çš„å…¬å¼è®¡ç®—"""
        writer = ExcelWriter(self.file_path)

        result = writer.evaluate_formula("=SUM(A1:A3)", "TestSheet")

        assert result.success is True
        assert result.data == 60

    def test_evaluate_formula_average(self):
        """æµ‹è¯•å¹³å‡å€¼å…¬å¼"""
        writer = ExcelWriter(self.file_path)

        result = writer.evaluate_formula("AVERAGE(A1:A3)", "TestSheet")

        # (10+20+30)/3 = 20
        assert result.success is True
        assert result.data == 20

    def test_evaluate_formula_count(self):
        """æµ‹è¯•è®¡æ•°å…¬å¼"""
        writer = ExcelWriter(self.file_path)

        result = writer.evaluate_formula("COUNT(A1:A3)", "TestSheet")

        assert result.success is True
        assert result.data == 3

    def test_evaluate_formula_min_max(self):
        """æµ‹è¯•æœ€å€¼å…¬å¼"""
        writer = ExcelWriter(self.file_path)

        min_result = writer.evaluate_formula("MIN(A1:A3)", "TestSheet")
        max_result = writer.evaluate_formula("MAX(A1:A3)", "TestSheet")

        assert min_result.success is True
        assert min_result.data == 10
        assert max_result.success is True
        assert max_result.data == 30

    def test_evaluate_formula_text_concatenation(self):
        """æµ‹è¯•æ–‡æœ¬è¿æ¥å…¬å¼"""
        writer = ExcelWriter(self.file_path)

        # å…ˆæ·»åŠ ä¸€äº›æ–‡æœ¬æ•°æ®
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        ws = wb["TestSheet"]
        ws['D1'] = "Hello"
        ws['D2'] = "World"
        wb.save(self.file_path)

        result = writer.evaluate_formula('CONCATENATE("Hello", " ", "World")', "TestSheet")

        assert result.success is True
        assert result.data == "Hello World"

    def test_evaluate_formula_basic_math(self):
        """æµ‹è¯•åŸºç¡€æ•°å­¦è¿ç®—"""
        writer = ExcelWriter(self.file_path)

        result = writer.evaluate_formula("10 + 20 * 2", "TestSheet")

        assert result.success is True
        assert result.data == 50  # 20*2 + 10

    def test_evaluate_formula_cache_hit(self):
        """æµ‹è¯•å…¬å¼ç¼“å­˜å‘½ä¸­"""
        writer = ExcelWriter(self.file_path)

        # ç¬¬ä¸€æ¬¡è®¡ç®—
        result1 = writer.evaluate_formula("SUM(A1:A3)", "TestSheet")
        assert result1.success is True
        assert result1.metadata['cached'] is False

        # ç¬¬äºŒæ¬¡è®¡ç®—ç›¸åŒå…¬å¼ï¼ˆåº”è¯¥å‘½ä¸­ç¼“å­˜ï¼‰
        result2 = writer.evaluate_formula("SUM(A1:A3)", "TestSheet")
        assert result2.success is True
        assert result2.metadata['cached'] is True

    def test_evaluate_formula_result_types(self):
        """æµ‹è¯•ä¸åŒç»“æœç±»å‹"""
        writer = ExcelWriter(self.file_path)

        # æ•°å­—ç±»å‹
        num_result = writer.evaluate_formula("SUM(A1:A3)", "TestSheet")
        assert num_result.success is True
        assert num_result.metadata['result_type'] == "number"

        # æ–‡æœ¬ç±»å‹
        text_result = writer.evaluate_formula('CONCATENATE("A", "B")', "TestSheet")
        assert text_result.success is True
        assert text_result.metadata['result_type'] == "text"

    def test_evaluate_formula_invalid_formula(self):
        """æµ‹è¯•æ— æ•ˆå…¬å¼"""
        writer = ExcelWriter(self.file_path)

        result = writer.evaluate_formula("INVALID_FUNCTION(A1:A3)", "TestSheet")

        # åº”è¯¥è¿”å›Noneæˆ–é”™è¯¯
        assert result.success is True  # å…¬å¼è®¡ç®—æœ¬èº«æˆåŠŸï¼Œä½†ç»“æœå¯èƒ½æ˜¯None
        assert result.data is None

    def test_create_temp_workbook(self):
        """æµ‹è¯•åˆ›å»ºä¸´æ—¶å·¥ä½œç°¿"""
        writer = ExcelWriter(self.file_path)

        with patch('src.utils.formula_cache.get_formula_cache') as mock_cache:
            mock_cache_instance = Mock()
            mock_cache.return_value = mock_cache_instance

            temp_workbook, temp_file_path = writer._create_temp_workbook("TestSheet", mock_cache_instance)

            assert temp_workbook is not None
            assert temp_file_path is not None
            assert os.path.exists(temp_file_path)

            # æ¸…ç†
            try:
                os.unlink(temp_file_path)
            except:
                pass

    def test_basic_formula_parse_sum(self):
        """æµ‹è¯•åŸºç¡€å…¬å¼è§£æ - SUMå‡½æ•°"""
        writer = ExcelWriter(self.file_path)

        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        ws = wb["TestSheet"]

        result = writer._basic_formula_parse("SUM(A1:A3)", ws)

        assert result == 60  # 10+20+30

    def test_basic_formula_parse_average(self):
        """æµ‹è¯•åŸºç¡€å…¬å¼è§£æ - AVERAGEå‡½æ•°"""
        writer = ExcelWriter(self.file_path)

        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        ws = wb["TestSheet"]

        result = writer._basic_formula_parse("AVERAGE(A1:A3)", ws)

        assert result == 20  # (10+20+30)/3

    def test_get_range_values(self):
        """æµ‹è¯•è·å–èŒƒå›´å€¼"""
        writer = ExcelWriter(self.file_path)

        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        ws = wb["TestSheet"]

        values = writer._get_range_values(ws, "A1", "A3")

        assert len(values) == 3
        assert 10 in values
        assert 20 in values
        assert 30 in values

    def test_numpy_average_fallback(self):
        """æµ‹è¯•numpyå¹³å‡å€¼å›é€€å®ç°"""
        writer = ExcelWriter(self.file_path)

        values = [10, 20, 30]
        result = writer._numpy_average(values)

        assert result == 20

    def test_numpy_min_max(self):
        """æµ‹è¯•numpyæœ€å€¼å‡½æ•°"""
        writer = ExcelWriter(self.file_path)

        values = [10, 20, 30, 5, 100]

        min_result = writer._numpy_min(values)
        max_result = writer._numpy_max(values)

        assert min_result == 5
        assert max_result == 100

    def test_numpy_median(self):
        """æµ‹è¯•numpyä¸­ä½æ•°å‡½æ•°"""
        writer = ExcelWriter(self.file_path)

        # å¥‡æ•°ä¸ªå€¼
        odd_values = [10, 20, 30]
        assert writer._numpy_median(odd_values) == 20

        # å¶æ•°ä¸ªå€¼
        even_values = [10, 20, 30, 40]
        assert writer._numpy_median(even_values) == 25

    def test_numpy_stdev_var(self):
        """æµ‹è¯•numpyæ ‡å‡†å·®å’Œæ–¹å·®å‡½æ•°"""
        writer = ExcelWriter(self.file_path)

        values = [10, 20, 30, 40, 50]

        stdev_result = writer._numpy_stdev(values)
        var_result = writer._numpy_var(values)

        # éªŒè¯æ ‡å‡†å·® = æ–¹å·®çš„å¹³æ–¹æ ¹
        assert abs(stdev_result - (var_result ** 0.5)) < 0.001

    def test_numpy_countif(self):
        """æµ‹è¯•numpyæ¡ä»¶è®¡æ•°å‡½æ•°"""
        writer = ExcelWriter(self.file_path)

        values = [10, 20, 30, 40, 50]

        # å¤§äº25çš„è®¡æ•°
        count_gt = writer._numpy_countif(values, ">25")
        assert count_gt == 3  # 30, 40, 50

        # ç­‰äº20çš„è®¡æ•°
        count_eq = writer._numpy_countif(values, "=20")
        assert count_eq == 1

    def test_numpy_sumif(self):
        """æµ‹è¯•numpyæ¡ä»¶æ±‚å’Œå‡½æ•°"""
        writer = ExcelWriter(self.file_path)

        values = [10, 20, 30, 40, 50]

        # å¤§äº25çš„æ±‚å’Œ
        sum_gt = writer._numpy_sumif(values, ">25")
        assert sum_gt == 120  # 30+40+50

        # ç­‰äº20çš„æ±‚å’Œ
        sum_eq = writer._numpy_sumif(values, "=20")
        assert sum_eq == 20

    def test_numpy_averageif(self):
        """æµ‹è¯•numpyæ¡ä»¶å¹³å‡å€¼å‡½æ•°"""
        writer = ExcelWriter(self.file_path)

        values = [10, 20, 30, 40, 50]

        # å¤§äº25çš„å¹³å‡å€¼
        avg_gt = writer._numpy_averageif(values, ">25")
        assert avg_gt == 40  # (30+40+50)/3

    def test_numpy_mode(self):
        """æµ‹è¯•numpyä¼—æ•°å‡½æ•°"""
        writer = ExcelWriter(self.file_path)

        values = [10, 20, 20, 30, 30, 30, 40]
        result = writer._numpy_mode(values)

        assert result == 30  # 30å‡ºç°æ¬¡æ•°æœ€å¤š

    def test_calculate_range_sum(self):
        """æµ‹è¯•èŒƒå›´æ±‚å’Œè®¡ç®—"""
        writer = ExcelWriter(self.file_path)

        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        ws = wb["TestSheet"]

        result = writer._calculate_range_sum(ws, "A1", "A3")

        assert result == 60  # 10+20+30

    def test_calculate_range_count(self):
        """æµ‹è¯•èŒƒå›´è®¡æ•°è®¡ç®—"""
        writer = ExcelWriter(self.file_path)

        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        ws = wb["TestSheet"]

        result = writer._calculate_range_count(ws, "A1", "A3")

        assert result == 3  # 3ä¸ªæ•°å€¼

    def test_get_result_type(self):
        """æµ‹è¯•ç»“æœç±»å‹åˆ¤æ–­"""
        writer = ExcelWriter(self.file_path)

        # æ•°å­—ç±»å‹
        assert writer._get_result_type(42) == "number"
        assert writer._get_result_type(3.14) == "number"

        # æ–‡æœ¬ç±»å‹
        assert writer._get_result_type("hello") == "text"

        # å¸ƒå°”ç±»å‹
        assert writer._get_result_type(True) == "boolean"

        # ç©ºå€¼ç±»å‹
        assert writer._get_result_type(None) == "null"

        # æœªçŸ¥ç±»å‹
        assert writer._get_result_type([]) == "unknown"


class TestExcelWriterFormatting:
    """ExcelWriteræ ¼å¼åŒ–æµ‹è¯•"""

    def setup_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•å‰çš„è®¾ç½®"""
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # æ·»åŠ æµ‹è¯•æ•°æ®
        for i in range(1, 4):
            for j in range(1, 4):
                ws.cell(row=i, column=j, value=f"R{i}C{j}")

        wb.save(self.temp_file.name)
        self.file_path = self.temp_file.name

    def teardown_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•åçš„æ¸…ç†"""
        if os.path.exists(self.file_path):
            try:
                os.unlink(self.file_path)
            except:
                pass

    def test_format_cells_basic(self):
        """æµ‹è¯•åŸºç¡€å•å…ƒæ ¼æ ¼å¼åŒ–"""
        writer = ExcelWriter(self.file_path)

        formatting = {
            'font': {'bold': True, 'color': 'FF0000'},
            'fill': {'color': 'FFFF00'}
        }

        result = writer.format_cells("TestSheet!A1:B2", formatting)

        assert result.success is True
        assert result.metadata['formatted_count'] > 0

    def test_format_cells_font_only(self):
        """æµ‹è¯•ä»…å­—ä½“æ ¼å¼åŒ–"""
        writer = ExcelWriter(self.file_path)

        formatting = {
            'font': {'name': 'Arial', 'size': 12, 'bold': True}
        }

        result = writer.format_cells("TestSheet!A1", formatting)

        assert result.success is True
        assert result.metadata['formatted_count'] == 1

    def test_format_cells_fill_only(self):
        """æµ‹è¯•ä»…èƒŒæ™¯å¡«å……æ ¼å¼åŒ–"""
        writer = ExcelWriter(self.file_path)

        formatting = {
            'fill': {'color': '00FF00'}
        }

        result = writer.format_cells("TestSheet!A1:C1", formatting)

        assert result.success is True
        assert result.metadata['formatted_count'] == 3

    def test_format_cells_alignment(self):
        """æµ‹è¯•å¯¹é½æ ¼å¼åŒ–"""
        writer = ExcelWriter(self.file_path)

        formatting = {
            'alignment': {'horizontal': 'center', 'vertical': 'center'}
        }

        result = writer.format_cells("TestSheet!A1:B2", formatting)

        assert result.success is True

    def test_format_cells_invalid_range(self):
        """æµ‹è¯•æ— æ•ˆèŒƒå›´çš„æ ¼å¼åŒ–"""
        writer = ExcelWriter(self.file_path)

        formatting = {'font': {'bold': True}}

        result = writer.format_cells("NonExistentSheet!A1", formatting)

        assert result.success is False

    def test_apply_cell_format_font(self):
        """æµ‹è¯•åº”ç”¨å­—ä½“æ ¼å¼"""
        writer = ExcelWriter(self.file_path)

        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        ws = wb["TestSheet"]
        cell = ws['A1']

        formatting = {
            'font': {'bold': True, 'italic': True, 'size': 14}
        }

        writer._apply_cell_format(cell, formatting)

        assert cell.font.bold is True
        assert cell.font.italic is True
        assert cell.font.size == 14

    def test_apply_cell_format_fill(self):
        """æµ‹è¯•åº”ç”¨èƒŒæ™¯å¡«å……"""
        writer = ExcelWriter(self.file_path)

        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        ws = wb["TestSheet"]
        cell = ws['A1']

        formatting = {
            'fill': {'color': 'FF0000'}
        }

        writer._apply_cell_format(cell, formatting)

        assert cell.fill.start_color.rgb in ['FFFF0000', '00FF0000']  # æ¥å—ä¸¤ç§æ ¼å¼

    def test_apply_cell_format_alignment(self):
        """æµ‹è¯•åº”ç”¨å¯¹é½æ ¼å¼"""
        writer = ExcelWriter(self.file_path)

        from openpyxl import load_workbook
        wb = load_workbook(self.file_path)
        ws = wb["TestSheet"]
        cell = ws['A1']

        formatting = {
            'alignment': {'horizontal': 'right', 'vertical': 'bottom'}
        }

        writer._apply_cell_format(cell, formatting)

        assert cell.alignment.horizontal == 'right'
        assert cell.alignment.vertical == 'bottom'


class TestExcelWriterMergeOperations:
    """ExcelWriteråˆå¹¶æ“ä½œæµ‹è¯•"""

    def setup_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•å‰çš„è®¾ç½®"""
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # æ·»åŠ æµ‹è¯•æ•°æ®
        for i in range(1, 4):
            for j in range(1, 4):
                ws.cell(row=i, column=j, value=f"R{i}C{j}")

        wb.save(self.temp_file.name)
        self.file_path = self.temp_file.name

    def teardown_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•åçš„æ¸…ç†"""
        if os.path.exists(self.file_path):
            try:
                os.unlink(self.file_path)
            except:
                pass

    def test_merge_cells_basic(self):
        """æµ‹è¯•åŸºç¡€å•å…ƒæ ¼åˆå¹¶"""
        writer = ExcelWriter(self.file_path)

        result = writer.merge_cells("TestSheet!A1:C1")

        assert result.success is True
        assert result.data['merged_range'] == "A1:C1"
        assert result.data['sheet_name'] == "TestSheet"

    def test_merge_cells_with_sheet_name_parameter(self):
        """æµ‹è¯•å¸¦å·¥ä½œè¡¨å‚æ•°çš„åˆå¹¶"""
        writer = ExcelWriter(self.file_path)

        result = writer.merge_cells("A1:B1", "TestSheet")

        assert result.success is True
        assert result.data['merged_range'] == "A1:B1"

    def test_merge_cells_full_range_expression(self):
        """æµ‹è¯•å®Œæ•´èŒƒå›´è¡¨è¾¾å¼åˆå¹¶"""
        writer = ExcelWriter(self.file_path)

        result = writer.merge_cells("TestSheet!B2:D3")

        assert result.success is True
        assert result.data['merged_range'] == "B2:D3"

    def test_merge_cells_nonexistent_sheet(self):
        """æµ‹è¯•ä¸å­˜åœ¨å·¥ä½œè¡¨çš„åˆå¹¶"""
        writer = ExcelWriter(self.file_path)

        result = writer.merge_cells("NonExistentSheet!A1:C1")

        assert result.success is False
        assert "å·¥ä½œè¡¨ä¸å­˜åœ¨" in result.error or "ä¸å­˜åœ¨" in result.error

    def test_unmerge_cells_basic(self):
        """æµ‹è¯•åŸºç¡€å–æ¶ˆåˆå¹¶"""
        writer = ExcelWriter(self.file_path)

        # å…ˆåˆå¹¶
        merge_result = writer.merge_cells("TestSheet!A1:C1")
        assert merge_result.success is True

        # å†å–æ¶ˆåˆå¹¶
        result = writer.unmerge_cells("TestSheet!A1:C1")

        assert result.success is True
        assert result.data['unmerged_range'] == "A1:C1"

    def test_unmerge_cells_nonexistent_sheet(self):
        """æµ‹è¯•ä¸å­˜åœ¨å·¥ä½œè¡¨çš„å–æ¶ˆåˆå¹¶"""
        writer = ExcelWriter(self.file_path)

        result = writer.unmerge_cells("NonExistentSheet!A1:C1")

        assert result.success is False
        assert "å·¥ä½œè¡¨ä¸å­˜åœ¨" in result.error or "ä¸å­˜åœ¨" in result.error


class TestExcelWriterBorderOperations:
    """ExcelWriterè¾¹æ¡†æ“ä½œæµ‹è¯•"""

    def setup_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•å‰çš„è®¾ç½®"""
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # æ·»åŠ æµ‹è¯•æ•°æ®
        for i in range(1, 4):
            for j in range(1, 4):
                ws.cell(row=i, column=j, value=f"R{i}C{j}")

        wb.save(self.temp_file.name)
        self.file_path = self.temp_file.name

    def teardown_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•åçš„æ¸…ç†"""
        if os.path.exists(self.file_path):
            try:
                os.unlink(self.file_path)
            except:
                pass

    def test_set_borders_basic(self):
        """æµ‹è¯•åŸºç¡€è¾¹æ¡†è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_borders("TestSheet!A1:C3", "thin")

        assert result.success is True
        assert result.data['border_style'] == "thin"
        assert result.data['cell_count'] == 9  # 3x3 = 9 cells

    def test_set_borders_thick(self):
        """æµ‹è¯•ç²—è¾¹æ¡†è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_borders("TestSheet!A1:B2", "thick")

        assert result.success is True
        assert result.data['border_style'] == "thick"

    def test_set_borders_double(self):
        """æµ‹è¯•åŒè¾¹æ¡†è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_borders("TestSheet!A1", "double")

        assert result.success is True
        assert result.data['border_style'] == "double"

    def test_set_borders_with_sheet_name_parameter(self):
        """æµ‹è¯•å¸¦å·¥ä½œè¡¨å‚æ•°çš„è¾¹æ¡†è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_borders("A1:C3", "thin", "TestSheet")

        assert result.success is True
        assert result.data['sheet_name'] == "TestSheet"

    def test_set_borders_nonexistent_sheet(self):
        """æµ‹è¯•ä¸å­˜åœ¨å·¥ä½œè¡¨çš„è¾¹æ¡†è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_borders("NonExistentSheet!A1:C1", "thin")

        assert result.success is False
        assert "å·¥ä½œè¡¨ä¸å­˜åœ¨" in result.error or "ä¸å­˜åœ¨" in result.error

    def test_set_borders_single_cell(self):
        """æµ‹è¯•å•å•å…ƒæ ¼è¾¹æ¡†"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_borders("TestSheet!B2", "dashed")

        assert result.success is True
        assert result.data['cell_count'] == 1

    def test_set_borders_large_range(self):
        """æµ‹è¯•å¤§èŒƒå›´è¾¹æ¡†"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_borders("TestSheet!A1:Z100", "thin")

        assert result.success is True
        assert result.data['cell_count'] == 2600  # 26*100


class TestExcelWriterDimensionOperations:
    """ExcelWriterå°ºå¯¸æ“ä½œæµ‹è¯•"""

    def setup_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•å‰çš„è®¾ç½®"""
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # æ·»åŠ æµ‹è¯•æ•°æ®
        for i in range(1, 6):
            for j in range(1, 4):
                ws.cell(row=i, column=j, value=f"R{i}C{j}")

        wb.save(self.temp_file.name)
        self.file_path = self.temp_file.name

    def teardown_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•åçš„æ¸…ç†"""
        if os.path.exists(self.file_path):
            try:
                os.unlink(self.file_path)
            except:
                pass

    def test_set_row_height_basic(self):
        """æµ‹è¯•åŸºç¡€è¡Œé«˜è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_row_height(2, 25.5, "TestSheet")

        assert result.success is True
        assert result.data['row_number'] == 2
        assert result.data['height'] == 25.5
        assert result.data['sheet_name'] == "TestSheet"

    def test_set_row_height_default_sheet(self):
        """æµ‹è¯•é»˜è®¤å·¥ä½œè¡¨çš„è¡Œé«˜è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_row_height(3, 30.0)

        assert result.success is True
        assert result.data['row_number'] == 3
        assert result.data['height'] == 30.0

    def test_set_row_height_nonexistent_sheet(self):
        """æµ‹è¯•ä¸å­˜åœ¨å·¥ä½œè¡¨çš„è¡Œé«˜è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_row_height(2, 25.0, "NonExistent")

        assert result.success is False
        assert "å·¥ä½œè¡¨ä¸å­˜åœ¨" in result.error or "ä¸å­˜åœ¨" in result.error

    def test_set_column_width_basic(self):
        """æµ‹è¯•åŸºç¡€åˆ—å®½è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_column_width("B", 15.5, "TestSheet")

        assert result.success is True
        assert result.data['column'] == "B"
        assert result.data['width'] == 15.5
        assert result.data['sheet_name'] == "TestSheet"

    def test_set_column_width_lowercase(self):
        """æµ‹è¯•å°å†™åˆ—æ ‡è¯†ç¬¦"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_column_width("c", 20.0, "TestSheet")

        assert result.success is True
        assert result.data['column'] == "C"  # åº”è¯¥è½¬æ¢ä¸ºå¤§å†™

    def test_set_column_width_default_sheet(self):
        """æµ‹è¯•é»˜è®¤å·¥ä½œè¡¨çš„åˆ—å®½è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_column_width("A", 12.0)

        assert result.success is True
        assert result.data['column'] == "A"
        assert result.data['width'] == 12.0

    def test_set_column_width_nonexistent_sheet(self):
        """æµ‹è¯•ä¸å­˜åœ¨å·¥ä½œè¡¨çš„åˆ—å®½è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        result = writer.set_column_width("B", 15.0, "NonExistent")

        assert result.success is False
        assert "å·¥ä½œè¡¨ä¸å­˜åœ¨" in result.error or "ä¸å­˜åœ¨" in result.error

    def test_set_column_width_multiple_columns(self):
        """æµ‹è¯•å¤šåˆ—å®½è®¾ç½®"""
        writer = ExcelWriter(self.file_path)

        # è®¾ç½®å¤šä¸ªåˆ—çš„å®½åº¦
        result_a = writer.set_column_width("A", 10.0)
        result_b = writer.set_column_width("B", 15.0)
        result_c = writer.set_column_width("C", 20.0)

        assert result_a.success is True
        assert result_b.success is True
        assert result_c.success is True

        assert result_a.data['width'] == 10.0
        assert result_b.data['width'] == 15.0
        assert result_c.data['width'] == 20.0


class TestExcelWriterPerformanceAndErrorHandling:
    """ExcelWriteræ€§èƒ½å’Œé”™è¯¯å¤„ç†æµ‹è¯•"""

    def setup_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•å‰çš„è®¾ç½®"""
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        wb.save(self.temp_file.name)
        self.file_path = self.temp_file.name

    def teardown_method(self):
        """æ¯ä¸ªæµ‹è¯•æ–¹æ³•åçš„æ¸…ç†"""
        if os.path.exists(self.file_path):
            try:
                os.unlink(self.file_path)
            except:
                pass

    def test_large_data_update_performance(self):
        """æµ‹è¯•å¤§æ•°æ®é‡æ›´æ–°æ€§èƒ½"""
        writer = ExcelWriter(self.file_path)

        # åˆ›å»ºå¤§é‡æ•°æ®
        large_data = [[f"Value_{i}_{j}" for j in range(10)] for i in range(100)]

        start_time = time.time()
        result = writer.update_range("TestSheet!A1", large_data)
        end_time = time.time()

        assert result.success is True
        assert result.metadata['modified_cells_count'] == 1000  # 100*10
        assert end_time - start_time < 5.0  # åº”è¯¥åœ¨5ç§’å†…å®Œæˆ

    def test_concurrent_operations(self):
        """æµ‹è¯•å¹¶å‘æ“ä½œå®‰å…¨æ€§"""
        import threading

        results = []
        errors = []

        def worker(worker_id):
            try:
                writer = ExcelWriter(self.file_path)
                data = [[f"Worker_{worker_id}_Cell_{i}" for i in range(3)]]
                result = writer.update_range(f"TestSheet!A{worker_id + 1}", data)
                results.append((worker_id, result.success))
            except Exception as e:
                errors.append((worker_id, str(e)))

        # å¯åŠ¨å¤šä¸ªçº¿ç¨‹
        threads = []
        for i in range(3):
            thread = threading.Thread(target=worker, args=(i,))
            threads.append(thread)
            thread.start()

        for thread in threads:
            thread.join()

        # éªŒè¯ç»“æœ
        assert len(errors) == 0, f"å¹¶å‘æ“ä½œå‡ºç°é”™è¯¯: {errors}"
        assert len(results) == 3
        assert all(success for _, success in results)

    def test_memory_usage_large_data(self):
        """æµ‹è¯•å¤§æ•°æ®é‡çš„å†…å­˜ä½¿ç”¨"""
        import gc

        writer = ExcelWriter(self.file_path)

        # å¼ºåˆ¶åƒåœ¾å›æ”¶
        gc.collect()

        # åˆ›å»ºéå¸¸å¤§çš„æ•°æ®
        very_large_data = [[f"Cell_{i}_{j}" for j in range(50)] for i in range(200)]

        # æ‰§è¡Œæ“ä½œ
        result = writer.update_range("TestSheet!A1", very_large_data)

        assert result.success is True
        assert result.metadata['modified_cells_count'] == 10000  # 200*50

        # å†æ¬¡åƒåœ¾å›æ”¶
        gc.collect()

        # å¦‚æœèƒ½åˆ°è¾¾è¿™é‡Œï¼Œè¯´æ˜æ²¡æœ‰ä¸¥é‡çš„å†…å­˜é—®é¢˜
        assert True

    def test_error_handling_corrupted_file(self):
        """æµ‹è¯•æŸåæ–‡ä»¶çš„å¤„ç†"""
        # åˆ›å»ºä¸€ä¸ªæŸåçš„æ–‡ä»¶ï¼ˆå†™å…¥æ— æ•ˆå†…å®¹ï¼‰
        with open(self.file_path, 'w') as f:
            f.write("This is not a valid Excel file")

        writer = ExcelWriter(self.file_path)
        data = [["Test"]]

        result = writer.update_range("TestSheet!A1", data)

        assert result.success is False
        # æ£€æŸ¥é”™è¯¯ä¿¡æ¯
        error_msg = str(result.error) if hasattr(result, 'error') and result.error else ""
        assert "å¤±è´¥" in error_msg or "error" in error_msg.lower() or "zip" in error_msg.lower()

    def test_error_handling_invalid_range_format(self):
        """æµ‹è¯•æ— æ•ˆèŒƒå›´æ ¼å¼çš„å¤„ç†"""
        writer = ExcelWriter(self.file_path)
        data = [["Test"]]

        # æµ‹è¯•å„ç§æ— æ•ˆæ ¼å¼
        invalid_ranges = [
            "InvalidRange",
            "Sheet1!Invalid",
            "A1:ZZ999",  # è¶…å‡ºExcelé™åˆ¶
            "",
        ]

        for invalid_range in invalid_ranges:
            result = writer.update_range(invalid_range, data)
            # æŸäº›å¯èƒ½ä¼šæˆåŠŸï¼ŒæŸäº›å¯èƒ½ä¼šå¤±è´¥ï¼Œä¸»è¦æµ‹è¯•ä¸ä¼šå´©æºƒ
            assert result is not None

    def test_error_handling_permission_denied(self):
        """æµ‹è¯•æƒé™æ‹’ç»çš„å¤„ç†"""
        # å°è¯•å¯¹åªè¯»æ–‡ä»¶è¿›è¡Œæ“ä½œ
        if os.name == 'nt':  # Windowsç³»ç»Ÿ
            # è®¾ç½®æ–‡ä»¶ä¸ºåªè¯»
            os.chmod(self.file_path, 0o444)

            writer = ExcelWriter(self.file_path)
            data = [["Test"]]

            result = writer.update_range("TestSheet!A1", data)

            # æ¢å¤æ–‡ä»¶æƒé™ä»¥ä¾¿æ¸…ç†
            os.chmod(self.file_path, 0o666)

            # ç»“æœå¯èƒ½æˆåŠŸä¹Ÿå¯èƒ½å¤±è´¥ï¼Œå–å†³äºç³»ç»Ÿ
            assert result is not None

    def test_recovery_after_error(self):
        """æµ‹è¯•é”™è¯¯åçš„æ¢å¤èƒ½åŠ›"""
        writer = ExcelWriter(self.file_path)

        # å…ˆæ‰§è¡Œä¸€ä¸ªä¼šå¤±è´¥çš„æ“ä½œ
        invalid_result = writer.update_range("NonExistentSheet!A1", [["Test"]])
        assert invalid_result.success is False

        # å†æ‰§è¡Œä¸€ä¸ªæ­£å¸¸çš„æ“ä½œ
        valid_result = writer.update_range("TestSheet!A1", [["Recovery"]])
        assert valid_result.success is True

    def test_data_type_handling(self):
        """æµ‹è¯•å„ç§æ•°æ®ç±»å‹çš„å¤„ç†"""
        from datetime import datetime

        writer = ExcelWriter(self.file_path)

        # æµ‹è¯•å„ç§æ•°æ®ç±»å‹
        test_data = [
            [123, 45.67, True, None, "Text"],
            ["ä¸­æ–‡", "ğŸ®", "", 0, []],
            [{"key": "value"}, (1, 2), 3+4j, datetime.now(), "Special"]
        ]

        result = writer.update_range("TestSheet!A1", test_data)

        assert result.success is True
        assert result.metadata['modified_cells_count'] == 15  # 3 rows * 5 cols

    def test_unicode_and_special_characters(self):
        """æµ‹è¯•Unicodeå’Œç‰¹æ®Šå­—ç¬¦å¤„ç†"""
        writer = ExcelWriter(self.file_path)

        # æµ‹è¯•å„ç§Unicodeå­—ç¬¦
        unicode_data = [
            ["ä¸­æ–‡æµ‹è¯•", "English", "EspaÃ±ol", "FranÃ§ais"],
            ["ğŸ®æ¸¸æˆ", "ğŸ“Šæ•°æ®", "ğŸ”§å·¥å…·", "ğŸš€æ€§èƒ½"],
            ["Î±Î²Î³", "âˆ‘âˆâˆ«", "â„ƒâ„‰", "â™ â™¥â™¦â™£"]
        ]

        result = writer.update_range("TestSheet!A1", unicode_data)

        assert result.success is True
        assert result.metadata['modified_cells_count'] == 12


if __name__ == "__main__":
    pytest.main([__file__, "-v"])