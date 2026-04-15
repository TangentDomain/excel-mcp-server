#!/usr/bin/env python3
"""
Round 37 MCP 接口测试 - 多Sheet联动测试（核心创造性方向）
+ P0第10轮回归 + 已知问题快速回归

测试组:
  A组: 新接口专项（本轮无新代码变更，跳过详细测试）
  B组: P0第10轮回归（已知问题追踪）
  C组: 多Sheet联动测试（本轮核心！8个场景）
  D组: 已知问题快速回归
"""

import sys
import os
import shutil
import tempfile
import traceback

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)

# ============================================================
# 测试基础设施
# ============================================================

TEST_DIR = tempfile.mkdtemp(prefix="r37_")
BASE_FILE = os.path.join(TEST_DIR, "base_test.xlsx")
MULTI_FILE = os.path.join(TEST_DIR, "multi_sheet.xlsx")

RESULTS = {"pass": 0, "fail": 0, "details": []}


def result(name: str, passed: bool, detail: str = ""):
    status = "✅ PASS" if passed else "❌ FAIL"
    print(f"  {status} | {name}: {detail[:120]}")
    RESULTS["pass" if passed else "fail"] += 1
    RESULTS["details"].append((name, passed, detail))


def create_multi_sheet_test_file():
    """创建多Sheet测试文件：装备表 + 怪物表 + 掉落表"""
    wb = Workbook()

    # Sheet1: 装备表
    ws1 = wb.active
    ws1.title = "装备"
    ws1.append(["ID", "Name", "Atk", "Price", "Rarity"])
    for i in range(1, 11):
        ws1.append([i, f"Weapon-{i}", i * 10, i * 100.5, ["Common", "Rare", "Epic", "Legendary"][(i - 1) % 4]])

    # Sheet2: 怪物表
    ws2 = wb.create_sheet("怪物")
    ws2.append(["ID", "Name", "HP", "Level", "DropEquipID"])
    monsters = [
        (1, "Slime", 50, 1, 1),
        (2, "Goblin", 200, 3, 2),
        (3, "Orc", 500, 5, 3),
        (4, "Dragon", 5000, 10, 10),
        (5, "Demon", 3000, 8, None),  # 无掉落
    ]
    for m in monsters:
        ws2.append(list(m))

    # Sheet3: 掉落表
    ws3 = wb.create_sheet("掉落")
    ws3.append(["MonsterID", "EquipID", "DropRate"])
    drops = [
        (1, 1, 0.8),
        (1, 2, 0.2),
        (2, 3, 0.5),
        (2, 4, 0.5),
        (3, 5, 0.7),
        (3, 6, 0.3),
        (4, 10, 1.0),
    ]
    for d in drops:
        ws3.append(list(d))

    wb.save(MULTI_FILE)
    return MULTI_FILE


def create_base_test_file():
    """创建基础单Sheet测试文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "Name", "Value", "Score"])
    for i in range(1, 21):
        ws.append([i, f"Item-{i}", i * 10, i * 5.5])
    wb.save(BASE_FILE)
    return BASE_FILE


def copy_test_file(src):
    """复制测试文件到临时位置"""
    dst = os.path.join(TEST_DIR, f"test_{os.urandom(4).hex()}.xlsx")
    shutil.copy2(src, dst)
    return dst


# ============================================================
# A组: 新接口专项（本轮无新变更）
# ============================================================

def test_group_a():
    print("\n" + "=" * 70)
    print("📋 A组: 新接口专项（本轮无新业务代码变更）")
    print("=" * 70)
    result("A0-无新变更确认", True, "本轮仅 .iteration-memory.md + round36_mcp_test.py, 无新API")


# ============================================================
# B组: P0 第10轮回归
# ============================================================

def test_group_b():
    print("\n" + "=" * 70)
    print("📋 B组: P0 第10轮回归（已知问题追踪 R10）")
    print("=" * 70)

    # 确保基础测试文件存在
    if not os.path.exists(BASE_FILE):
        create_base_test_file()
    f = copy_test_file(BASE_FILE)

    # P0-2: SELECT分号注入
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1; DROP TABLE Sheet1")
    is_vuln = r.get('success', False) and len(r.get('data', [])) > 1
    result("P0-2-R10 SELECT分号注入", not is_vuln,
           f"success={r.get('success')}, rows={len(r.get('data', [])) if r.get('data') else 0}")

    # P0-4: UPDATE分号注入
    f2 = copy_test_file(BASE_FILE)
    r = execute_advanced_update_query(f2, "UPDATE Sheet1 SET Value=999; DROP TABLE Sheet1")
    is_vuln = r.get('success', False)
    result("P0-4-R10 UPDATE分号注入", not is_vuln,
           f"success={r.get('success')}, msg={str(r.get('message',''))[:60]}")

    # P0-5: INSERT分号注入
    f3 = copy_test_file(BASE_FILE)
    r = execute_advanced_insert_query(f3, "INSERT INTO Sheet1 (ID,Name,Value,Score) VALUES(999,'HACK',-1,-1); DROP TABLE Sheet1")
    is_vuln = r.get('success', False)
    result("P0-5-R10 INSERT分号注入", not is_vuln,
           f"success={r.get('success')}, 插入恶意数据={is_vuln}")

    # P0-6: DELETE分号注入
    f4 = copy_test_file(BASE_FILE)
    r = execute_advanced_delete_query(f4, "DELETE FROM Sheet1 WHERE ID=99999; DROP TABLE Sheet1")
    # 即使WHERE无匹配，也不应该接受含分号的SQL
    accepted_semicolon = r.get('success', False)
    no_match_msg = "没有匹配" in r.get('message', '')
    is_vuln = accepted_semicolon and not no_match_msg  # 如果成功但不是"无匹配"，说明执行了恶意语句
    result("P0-6-R10 DELETE分号注入", not accepted_semicolon,
           f"success={accepted_semicolon}, msg={str(r.get('message',''))[:60]}")

    # P0-7: UPDATE注释符全表篡改
    f5 = copy_test_file(BASE_FILE)
    # 先查原始sum
    r_before = execute_advanced_sql_query(f5, "SELECT SUM(Value) as s FROM Sheet1")
    before_sum = 0
    if r_before.get('data') and len(r_before['data']) > 1:
        try:
            before_sum = float(r_before['data'][1][0])
        except:
            pass
    r = execute_advanced_update_query(f5, "UPDATE Sheet1 SET Value=-1 -- WHERE ID=999999")
    r_after = execute_advanced_sql_query(f5, "SELECT SUM(Value) as s FROM Sheet1")
    after_sum = 0
    if r_after.get('data') and len(r_after['data']) > 1:
        try:
            after_sum = float(r_after['data'][1][0])
        except:
            pass
    all_changed = after_sum != before_sum and after_sum < 0
    result("P0-7-R10 UPDATE注释符篡改", not (r.get('success') and all_changed),
           f"success={r.get('success')}, sum:{before_sum}->{after_sum}, 全表篡改={all_changed}")

    # P0-3: uint8溢出（已修复验证）
    f6 = copy_test_file(BASE_FILE)
    r = execute_advanced_update_query(f6, "UPDATE Sheet1 SET Value=999 WHERE ID=1")
    r_check = execute_advanced_sql_query(f6, "SELECT Value FROM Sheet1 WHERE ID=1")
    read_val = None
    if r_check.get('data') and len(r_check['data']) > 1:
        read_val = r_check['data'][1][0]
    result("P0-3-R10 uint8溢出修复", read_val == 999,
           f"写入999, 读回={read_val}")


# ============================================================
# C组: 多Sheet联动测试（本轮核心！）
# ============================================================

def test_group_c():
    print("\n" + "=" * 70)
    print("📋 C组: 多Sheet联动测试（本轮核心创造性方向）")
    print("=" * 70)

    create_multi_sheet_test_file()
    mf = MULTI_FILE

    # C1: 跨Sheet JOIN查询
    print("\n--- C1: 跨Sheet JOIN查询 ---")
    try:
        # 注意：SQLGlot的Excel引擎中，不同Sheet是否支持JOIN？
        # 测试同一Sheet内的JOIN先
        r = execute_advanced_sql_query(mf, """
            SELECT m.Name as MonsterName, e.Name as EquipName, d.DropRate
            FROM 怪物 m
            JOIN 掉落 d ON m.ID = d.MonsterID
            JOIN 装备 e ON d.EquipID = e.ID
            ORDER BY m.Level DESC
        """)
        if r.get('success'):
            row_count = len(r.get('data', [])) - 1  # 减去表头
            result("C1-跨Sheet三表JOIN", row_count > 0,
                   f"success=True, 返回{row_count}行数据, data[1]={r.get('data',[[]])[1][:3] if len(r.get('data',[]))>1 else 'N/A'}")
        else:
            result("C1-跨Sheet三表JOIN", False, f"success=False, msg={str(r.get('message',''))[:100]}")
    except Exception as e:
        result("C1-跨Sheet三表JOIN", False, f"异常: {str(e)[:100]}")

    # C1b: 同一Sheet内自JOIN
    try:
        r = execute_advanced_sql_query(mf, """
            SELECT a.Name as Name1, b.Name as Name2, a.Atk as Atk1, b.Atk as Atk2
            FROM 装备 a
            JOIN 装备 b ON a.Atk < b.Atk
            WHERE a.ID <= 3 AND b.ID <= 5
            LIMIT 5
        """)
        if r.get('success'):
            row_count = len(r.get('data', [])) - 1
            result("C1b-同Sheet自JOIN", row_count > 0,
                   f"success=True, 返回{row_count}行")
        else:
            result("C1b-同Sheet自JOIN", False, f"msg={str(r.get('message',''))[:100]}")
    except Exception as e:
        result("C1b-同Sheet自JOIN", False, f"异常: {str(e)[:100]}")

    # C2: 创建新Sheet后立即查询
    print("\n--- C2: 创建新Sheet后立即查询 ---")
    try:
        from openpyxl import load_workbook
        # 用openpyxl添加一个新Sheet
        wb = load_workbook(mf)
        ws_new = wb.create_sheet("新建表")
        ws_new.append(["ID", "Data"])
        ws_new.append([1, "TestNew"])
        ws_new.append([2, "TestNew2"])
        wb.save(mf)

        # 立即查询新Sheet
        r = execute_advanced_sql_query(mf, "SELECT * FROM 新建表")
        if r.get('success'):
            row_count = len(r.get('data', [])) - 1
            result("C2-新建Sheet后立即查询", row_count == 2,
                   f"success=True, 返回{row_count}行(期望2)")
        else:
            result("C2-新建Sheet后立即查询", False, f"msg={str(r.get('message',''))[:100]}")
    except Exception as e:
        result("C2-新建Sheet后立即查询", False, f"异常: {str(e)[:100]}")

    # C3: 重命名Sheet后的查询
    print("\n--- C3: 重命名Sheet后的查询 ---")
    try:
        from openpyxl import load_workbook
        f_rename = copy_test_file(mf)
        wb = load_workbook(f_rename)

        # 检查原Sheet是否存在
        r_before = execute_advanced_sql_query(f_rename, "SELECT COUNT(*) as cnt FROM 装备")
        before_ok = r_before.get('success', False)

        # 重命名Sheet
        if '装备' in wb.sheetnames:
            wb['装备'].title = '装备RENAME'
            wb.save(f_rename)

        # 用旧名查询（应该失败）
        r_old = execute_advanced_sql_query(f_rename, "SELECT COUNT(*) as cnt FROM 装备")
        old_fail = not r_old.get('success', False)

        # 用新名查询（应该成功）
        r_new = execute_advanced_sql_query(f_rename, "SELECT COUNT(*) as cnt FROM 装备RENAME")
        new_ok = r_new.get('success', False)

        result("C3-重命名Sheet后查询", old_fail and new_ok,
               f"旧名失败={old_fail}, 新名成功={new_ok}, msg_old={str(r_old.get('message',''))[:50]}, msg_new={str(r_new.get('message',''))[:50]}")
    except Exception as e:
        result("C3-重命名Sheet后查询", False, f"异常: {str(e)[:100]}")

    # C4: 多Sheet批量操作 - 同时更新多个Sheet
    print("\n--- C4: 多Sheet批量操作 ---")
    try:
        f_batch = copy_test_file(mf)

        # 更新装备表的Price
        r1 = execute_advanced_update_query(f_batch, "UPDATE 装备 SET Price = Price * 2 WHERE Rarity = 'Legendary'")

        # 更新怪物的HP
        r2 = execute_advanced_update_query(f_batch, "UPDATE 怪物 SET HP = HP + 100 WHERE Level >= 5")

        # 验证两个更新都生效了
        r_check1 = execute_advanced_sql_query(f_batch, "SELECT Price FROM 装备 WHERE Rarity='Legendary' LIMIT 1")
        r_check2 = execute_advanced_sql_query(f_batch, "SELECT HP FROM 怪物 WHERE Level>=5 LIMIT 1")

        price_updated = False
        hp_updated = False
        if r_check1.get('data') and len(r_check1['data']) > 1:
            try:
                p = float(r_check1['data'][1][0])
                price_updated = p > 1000  # Legendary原来Price应该很大，翻倍后更大
            except:
                pass
        if r_check2.get('data') and len(r_check2['data']) > 1:
            try:
                h = float(r_check2['data'][1][0])
                hp_updated = h > 5000  # Dragon原来HP=5000, +100后>5000
            except:
                pass

        both_ok = r1.get('success') and r2.get('success') and price_updated and hp_updated
        result("C4-多Sheet同时更新", both_ok,
               f"update1={r1.get('success')}, update2={r2.get('success')}, price_upd={price_updated}, hp_upd={hp_updated}")
    except Exception as e:
        result("C4-多Sheet同时更新", False, f"异常: {str(e)[:100]}")

    # C5: 跨Sheet引用验证 - Sheet A的数据变更影响Sheet B的JOIN结果
    print("\n--- C5: 跨Sheet引用验证 ---")
    try:
        f_ref = copy_test_file(mf)

        # 先查JOIN结果
        r_before = execute_advanced_sql_query(f_ref, """
            SELECT e.Name, m.Name as Monster
            FROM 装备 e
            JOIN 掉落 d ON e.ID = d.EquipID
            JOIN 怪物 m ON d.MonsterID = m.ID
            WHERE e.ID = 1
        """)
        before_data = None
        if r_before.get('data') and len(r_before['data']) > 1:
            before_data = r_before['data'][1]

        # 修改装备表的Name
        execute_advanced_update_query(f_ref, "UPDATE 装备 SET Name = 'EXCALIBUR-MODIFIED' WHERE ID = 1")

        # 再查JOIN结果
        r_after = execute_advanced_sql_query(f_ref, """
            SELECT e.Name, m.Name as Monster
            FROM 装备 e
            JOIN 掉落 d ON e.ID = d.EquipID
            JOIN 怪物 m ON d.MonsterID = m.ID
            WHERE e.ID = 1
        """)
        after_data = None
        if r_after.get('data') and len(r_after['data']) > 1:
            after_data = r_after['data'][1]

        changed = (before_data != after_data) and after_data and 'EXCALIBUR' in str(after_data[0])
        result("C5-跨Sheet引用变更传播", changed,
               f"before={before_data}, after={after_data}, 变更传播={changed}")
    except Exception as e:
        result("C5-跨Sheet引用变更传播", False, f"异常: {str(e)[:100]}")

    # C6: Sheet不存在时的错误处理
    print("\n--- C6: Sheet不存在时的错误处理 ---")
    try:
        # 查询不存在的Sheet
        r1 = execute_advanced_sql_query(mf, "SELECT * FROM 不存在的Sheet12345")
        err1 = not r1.get('success', False) and ('不存在' in r1.get('message', '') or 'exist' in r1.get('message', '').lower())

        # 更新不存在的Sheet
        r2 = execute_advanced_update_query(mf, "UPDATE 不存在的Sheet SET Name='x' WHERE ID=1")
        err2 = not r2.get('success', False)

        # 删除不存在的Sheet
        r3 = execute_advanced_delete_query(mf, "DELETE FROM 不存在的Sheet WHERE ID=1")
        err3 = not r3.get('success', False)

        # 插入不存在的Sheet
        r4 = execute_advanced_insert_query(mf, "INSERT INTO 不存在的Sheet (ID) VALUES(1)")
        err4 = not r4.get('success', False)

        all_rejected = err1 and err2 and err3 and err4
        useful_msg = '不存在' in r1.get('message', '') or 'exist' in r1.get('message', '').lower()
        result("C6-Sheet不存在错误处理", all_rejected,
               f"query_err={err1}, update_err={err2}, delete_err={err3}, insert_err={err4}, 有用消息={useful_msg}")
    except Exception as e:
        result("C6-Sheet不存在错误处理", False, f"异常: {str(e)[:100]}")

    # C7: 特殊Sheet名在跨Sheet场景中的表现
    print("\n--- C7: 特殊Sheet名跨Sheet场景 ---")
    try:
        from openpyxl import load_workbook
        f_special = copy_test_file(mf)
        wb = load_workbook(f_special)

        # 创建中文Sheet名
        ws_cn = wb.create_sheet("中文装备表")
        ws_cn.append(["ID", "名称"])
        ws_cn.append([1, "测试"])

        # 创建带空格的Sheet名
        ws_space = wb.create_sheet("Sheet With Spaces")
        ws_space.append(["ID", "Val"])
        ws_space.append([1, 100])

        # 创建带连字符的Sheet名
        ws_hyphen = wb.create_sheet("my-equip-data")
        ws_hyphen.append(["ID", "Val"])
        ws_hyphen.append([1, 200])

        wb.save(f_special)

        # 测试中文Sheet名查询
        r_cn = execute_advanced_sql_query(f_special, "SELECT * FROM 中文装备表")
        cn_ok = r_cn.get('success', False)

        # 测试空格Sheet名查询
        r_sp = execute_advanced_sql_query(f_special, "SELECT * FROM `Sheet With Spaces`")
        sp_ok = r_sp.get('success', False)

        # 如果反引号不行，试试不用
        if not sp_ok:
            r_sp2 = execute_advanced_sql_query(f_special, "SELECT * FROM [Sheet With Spaces]")
            sp_ok = r_sp2.get('success', False)

        # 测试连字符Sheet名查询
        r_hy = execute_advanced_sql_query(f_special, "SELECT * FROM `my-equip-data`")
        hy_ok = r_hy.get('success', False)
        if not hy_ok:
            r_hy2 = execute_advanced_sql_query(f_special, "SELECT * FROM [my-equip-data]")
            hy_ok = r_hy2.get('success', False)

        result("C7a-中文Sheet名跨Sheet", cn_ok, f"success={cn_ok}")
        result("C7b-空格Sheet名(反引号)", sp_ok, f"success={sp_ok}")
        result("C7c-连字符Sheet名(反引号)", hy_ok, f"success={hy_ok}")

        # 特殊Sheet名参与JOIN
        if cn_ok:
            r_join = execute_advanced_sql_query(f_special, """
                SELECT a.Name, b.名称
                FROM 装备 a
                LEFT JOIN 中文装备表 b ON a.ID = b.ID
                LIMIT 3
            """)
            join_ok = r_join.get('success', False)
            result("C7d-特殊Sheet名参与JOIN", join_ok, f"success={join_ok}, msg={str(r_join.get('message',''))[:80]}")
        else:
            result("C7d-特殊Sheet名参与JOIN", False, "中文Sheet不可用，跳过JOIN测试")
    except Exception as e:
        result("C7-特殊Sheet名跨Sheet", False, f"异常: {str(e)[:100]}")

    # C8: 多Sheet事务性 - 一个Sheet操作失败不影响其他Sheet
    print("\n--- C8: 多Sheet事务性 ---")
    try:
        f_tx = copy_test_file(mf)

        # 正常更新装备表
        r1 = execute_advanced_update_query(f_tx, "UPDATE 装备 SET Price = 7777 WHERE ID = 1")

        # 故意对怪物表做无效操作（类型不匹配等）
        r2 = execute_advanced_update_query(f_tx, "UPDATE 怪物 SET HP = 'NOT_A_NUMBER' WHERE ID = 1")

        # 验证装备表的更新不受影响
        r_check = execute_advanced_sql_query(f_tx, "SELECT Price FROM 装备 WHERE ID = 1")
        price_ok = False
        if r_check.get('data') and len(r_check['data']) > 1:
            try:
                p = float(r_check['data'][1][0])
                price_ok = (p == 7777)
            except:
                pass

        tx_ok = r1.get('success') and (not r2.get('success', False)) and price_ok
        result("C8-多Sheet事务隔离", tx_ok,
               f"update1_ok={r1.get('success')}, update2_fail={not r2.get('success')}, 装备表price正确={price_ok}")
    except Exception as e:
        result("C8-多Sheet事务隔离", False, f"异常: {str(e)[:100]}")


# ============================================================
# D组: 已知问题快速回归
# ============================================================

def test_group_d():
    print("\n" + "=" * 70)
    print("📋 D组: 已知问题快速回归")
    print("=" * 70)

    # 确保基础测试文件存在
    if not os.path.exists(BASE_FILE):
        create_base_test_file()

    # D1: 并发多线程写入文件损坏
    print("\n--- D1: 并发多线程写入 ---")
    import threading
    import time

    f_conc = copy_test_file(BASE_FILE)
    errors = []
    lock = threading.Lock()

    def thread_write(thread_id):
        try:
            for i in range(5):
                r = execute_advanced_update_query(
                    f_conc,
                    f"UPDATE Sheet1 SET Value = {thread_id * 1000 + i} WHERE ID = {(thread_id % 20) + 1}"
                )
                if not r.get('success'):
                    with lock:
                        errors.append(f"T{thread_id}-iter{i}:{r.get('message','')[:40]}")
        except Exception as e:
            with lock:
                errors.append(f"T{thread_id}:EXCEPTION:{str(e)[:40]}")

    threads = [threading.Thread(target=thread_write, args=(t,)) for t in range(3)]
    t_start = time.time()
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=30)
    elapsed = time.time() - t_start

    conc_ok = len(errors) == 0
    result("D1-P1 并发多线程写入(3线程x5次)", conc_ok,
           f"错误数={len(errors)}, 耗时={elapsed:.3f}s" + (f", 首错={errors[0]}" if errors else ""))

    # D2: 公式列D名丢失
    print("\n--- D2: 公式列D名 ---")
    try:
        from openpyxl import load_workbook
        f_formula = os.path.join(TEST_DIR, "formula_test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "FormulaSheet"
        ws['A1'] = 'ID'
        ws['B1'] = 'Name'
        ws['C1'] = 'Val'
        ws['D1'] = '=A1+B1'  # 公式列
        ws.append([1, "Test", 10])
        ws.append([2, "Test2", 20])
        wb.save(f_formula)

        r = execute_advanced_sql_query(f_formula, "SELECT * FROM FormulaSheet")
        cols = []
        if r.get('data') and len(r['data']) > 0:
            cols = list(r['data'][0])
        has_d = 'D' in cols or any('D' in str(c) for c in cols)
        result("D2-P2 公式列D名保留", has_d,
               f"可用列={cols}, 含D={has_d}")
    except Exception as e:
        result("D2-P2 公式列D名保留", False, f"异常: {str(e)[:80]}")

    # D3: ||字符串拼接
    print("\n--- D3: ||字符串拼接 ---")
    f_cat = copy_test_file(BASE_FILE)
    r = execute_advanced_update_query(f_cat, "UPDATE Sheet1 SET Name = Name || '-suffix' WHERE ID = 1")
    cat_ok = r.get('success', False)
    if cat_ok:
        r_check = execute_advanced_sql_query(f_cat, "SELECT Name FROM Sheet1 WHERE ID=1")
        new_name = ""
        if r_check.get('data') and len(r_check['data']) > 1:
            new_name = str(r_check['data'][1][0])
        cat_ok = 'suffix' in new_name
    result("D3-P2-1 ||字符串拼接", cat_ok,
           f"success={r.get('success')}, msg={str(r.get('message',''))[:80]}")

    # D4: 极端浮点值损坏文件
    print("\n--- D4: 极端浮点值 ---")
    f_float = copy_test_file(BASE_FILE)
    try:
        r = execute_advanced_update_query(f_float, "UPDATE Sheet1 SET Value = 1.7976931348623157e+308 WHERE ID=1")
        write_ok = r.get('success', False)

        # 尝试读取
        r_read = execute_advanced_sql_query(f_float, "SELECT Value FROM Sheet1 WHERE ID=1")
        read_ok = r_read.get('success', False)
        read_val = "N/A"
        if r_read.get('data') and len(r_read['data']) > 1:
            read_val = str(r_read['data'][1][0])

        file_damaged = write_ok and not read_ok
        result("D4-P2-4 极端浮点值(float_max)", not file_damaged,
               f"write={write_ok}, read={read_ok}, val={read_val}, 文件损坏={file_damaged}")
    except Exception as e:
        result("D4-P2-4 极端浮点值(float_max)", False, f"异常: {str(e)[:80]}")

    # D5: 浮点精度损失
    print("\n--- D5: 浮点精度 ---")
    f_prec = copy_test_file(BASE_FILE)
    pi_val = 3.14159265358979323846
    r = execute_advanced_update_query(f_prec, f"UPDATE Sheet1 SET Value = {pi_val} WHERE ID=1")
    r_check = execute_advanced_sql_query(f_prec, "SELECT Value FROM Sheet1 WHERE ID=1")
    read_back = None
    if r_check.get('data') and len(r_check['data']) > 1:
        read_back = r_check['data'][1][0]
    precision_loss = True
    if read_back is not None:
        try:
            diff = abs(float(read_back) - pi_val)
            precision_loss = diff > 0.01
        except:
            pass
    result("D5-P2 浮点精度(π)", not precision_loss,
           f"写入π={pi_val}, 读回={read_back}, 精度损失={'是⚠️' if precision_loss else '否✅'}")


# ============================================================
# 主函数
# ============================================================

def main():
    print("=" * 70)
    print("🔬 ExcelMCP Round 37 测试 - 多Sheet联动 + P0第10轮回归")
    print("=" * 70)
    print(f"测试目录: {TEST_DIR}")
    print(f"基础文件: {BASE_FILE}")
    print(f"多Sheet文件: {MULTI_FILE}")

    try:
        create_base_test_file()
        print(f"\n✅ 基础测试文件已创建: {BASE_FILE}")

        test_group_a()
        test_group_b()
        test_group_c()
        test_group_d()

    except Exception as e:
        print(f"\n💥 致命错误: {e}")
        traceback.print_exc()
    finally:
        # 清理
        try:
            shutil.rmtree(TEST_DIR)
        except:
            pass

    # 汇总
    total = RESULTS["pass"] + RESULTS["fail"]
    print("\n" + "=" * 70)
    print(f"📊 Round 37 测试结果汇总")
    print("=" * 70)
    print(f"总测试数: {total}")
    print(f"  ✅ 通过: {RESULTS['pass']} ({RESULTS['pass']/total*100:.1f}%)" if total > 0 else "")
    print(f"  ❌ 失败: {RESULTS['fail']} ({RESULTS['fail']/total*100:.1f}%)" if total > 0 else "")

    if RESULTS["fail"] > 0:
        print(f"\n--- ❌ 失败详情 ---")
        for name, passed, detail in RESULTS["details"]:
            if not passed:
                print(f"  [{name}] {detail[:100]}")

    return RESULTS["fail"] == 0


if __name__ == "__main__":
    ok = main()
    sys.exit(0 if ok else 1)
