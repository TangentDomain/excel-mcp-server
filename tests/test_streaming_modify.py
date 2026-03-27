"""StreamingWriter 修改操作测试 — delete_rows / delete_columns / update_range"""

import shutil
import tempfile
import time

import pytest
from openpyxl import Workbook, load_workbook

from excel_mcp_server_fastmcp.core.streaming_writer import StreamingWriter


@pytest.fixture
def sample_file(tmp_path):
    """创建带数据的测试文件（多工作表）"""
    fp = str(tmp_path / "test.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Skills"

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25

    ws.append(["ID", "Name", "Damage", "Type"])
    ws.append([1, "Fireball", 100, "Fire"])
    ws.append([2, "Ice Spike", 80, "Ice"])
    ws.append([3, "Thunder", 120, "Lightning"])
    ws.append([4, "Heal", 50, "Holy"])

    ws2 = wb.create_sheet("Items")
    ws2.append(["ItemID", "ItemName"])
    ws2.append([101, "Sword"])
    ws2.append([102, "Shield"])

    wb.save(fp)
    wb.close()
    return fp


class TestDeleteRowsStreaming:
    def test_delete_single_row(self, sample_file):
        """删除单行"""
        success, msg, meta = StreamingWriter.delete_rows(sample_file, "Skills", 3, 1)
        assert success is True
        assert meta['action'] == 'delete_rows_streaming'
        assert meta['actual_count'] == 1
        assert meta['original_rows'] == 5
        assert meta['new_rows'] == 4

        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert rows[2][0] == 3  # 第3行变成了原来的Thunder

    def test_delete_multiple_rows(self, sample_file):
        """删除多行"""
        success, msg, meta = StreamingWriter.delete_rows(sample_file, "Skills", 2, 2)
        assert success is True
        assert meta['actual_count'] == 2
        assert meta['new_rows'] == 3

        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # 表头 + Fireball被删 + Heal + 保留Items
        assert rows[0][0] == "ID"
        assert rows[1][0] == 3  # Thunder

    def test_delete_last_row(self, sample_file):
        """删除最后一行"""
        success, msg, meta = StreamingWriter.delete_rows(sample_file, "Skills", 5, 1)
        assert success is True
        assert meta['new_rows'] == 4

    def test_delete_header_row(self, sample_file):
        """删除表头行（允许，用户自行决定）"""
        success, msg, meta = StreamingWriter.delete_rows(sample_file, "Skills", 1, 1)
        assert success is True
        assert meta['new_rows'] == 4

    def test_delete_more_than_exists(self, sample_file):
        """删除行数超过实际行数"""
        success, msg, meta = StreamingWriter.delete_rows(sample_file, "Skills", 4, 10)
        assert success is True
        assert meta['actual_count'] == 2  # 只有2行可删
        assert meta['new_rows'] == 3  # header + 2 remaining

    def test_delete_invalid_row(self, sample_file):
        """起始行号超过总行数"""
        success, msg, meta = StreamingWriter.delete_rows(sample_file, "Skills", 100, 1)
        assert success is False
        assert "超过" in msg

    def test_delete_preserves_other_sheets(self, sample_file):
        """删除行不影响其他工作表"""
        StreamingWriter.delete_rows(sample_file, "Skills", 2, 1)
        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Items"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert len(rows) == 3  # header + 2 items
        assert rows[1] == (101, "Sword")

    def test_delete_preserves_col_widths(self, sample_file):
        """删除行保留列宽"""
        StreamingWriter.delete_rows(sample_file, "Skills", 2, 1)
        wb = load_workbook(sample_file)
        ws = wb["Skills"]
        assert ws.column_dimensions['A'].width == 20
        assert ws.column_dimensions['B'].width == 30
        wb.close()

    def test_delete_sheet_not_found(self, sample_file):
        """工作表不存在"""
        success, msg, meta = StreamingWriter.delete_rows(sample_file, "NotExist", 1, 1)
        assert success is False


class TestDeleteColumnsStreaming:
    def test_delete_single_column(self, sample_file):
        """删除单列"""
        success, msg, meta = StreamingWriter.delete_columns(sample_file, "Skills", 3, 1)
        assert success is True
        assert meta['action'] == 'delete_columns_streaming'
        assert meta['actual_count'] == 1
        assert meta['original_columns'] == 4
        assert meta['new_columns'] == 3

        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # Damage列（第3列）被删
        assert rows[0] == ("ID", "Name", "Type")  # Damage消失
        assert rows[1] == (1, "Fireball", "Fire")

    def test_delete_multiple_columns(self, sample_file):
        """删除多列"""
        success, msg, meta = StreamingWriter.delete_columns(sample_file, "Skills", 2, 2)
        assert success is True
        assert meta['actual_count'] == 2
        assert meta['new_columns'] == 2

        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert rows[0] == ("ID", "Type")  # Name+Damage被删，保留ID和Type

    def test_delete_first_column(self, sample_file):
        """删除第一列"""
        success, msg, meta = StreamingWriter.delete_columns(sample_file, "Skills", 1, 1)
        assert success is True
        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert rows[0] == ("Name", "Damage", "Type")

    def test_delete_invalid_column(self, sample_file):
        """起始列号超过最大列数"""
        success, msg, meta = StreamingWriter.delete_columns(sample_file, "Skills", 100, 1)
        assert success is False
        assert "超过" in msg

    def test_delete_preserves_other_sheets(self, sample_file):
        """删除列不影响其他工作表"""
        StreamingWriter.delete_columns(sample_file, "Skills", 1, 1)
        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Items"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert rows[0] == ("ItemID", "ItemName")

    def test_delete_sheet_not_found(self, sample_file):
        """工作表不存在"""
        success, msg, meta = StreamingWriter.delete_columns(sample_file, "NotExist", 1, 1)
        assert success is False


class TestUpdateRangeStreaming:
    def test_overwrite_single_cell(self, sample_file):
        """覆盖单个单元格"""
        success, msg, meta = StreamingWriter.update_range(
            sample_file, "Skills", 2, 2, [["Ice Lance"]]
        )
        assert success is True
        assert meta['action'] == 'update_range_streaming'
        assert meta['updated_cells'] == 1

        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert rows[1][1] == "Ice Lance"

    def test_overwrite_range(self, sample_file):
        """覆盖范围区域"""
        success, msg, meta = StreamingWriter.update_range(
            sample_file, "Skills", 2, 1,
            [[99, "Meteor", 500, "Fire"], [100, "Blizzard", 300, "Ice"]]
        )
        assert success is True
        assert meta['updated_cells'] == 8  # 2行 × 4列
        assert meta['rows_written'] == 2

        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert rows[1] == (99, "Meteor", 500, "Fire")
        assert rows[2] == (100, "Blizzard", 300, "Ice")

    def test_overwrite_preserves_other_rows(self, sample_file):
        """覆盖不影响其他行"""
        StreamingWriter.update_range(sample_file, "Skills", 2, 1, [[99, "X", 0, "Y"]])
        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # 第3行（Ice Spike）不变
        assert rows[2] == (2, "Ice Spike", 80, "Ice")

    def test_overwrite_partial_row(self, sample_file):
        """覆盖部分列"""
        success, msg, meta = StreamingWriter.update_range(
            sample_file, "Skills", 3, 2, [["New Ice Spike"]]
        )
        assert success is True
        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert rows[2][1] == "New Ice Spike"
        assert rows[2][0] == 2  # ID不变
        assert rows[2][2] == 80  # Damage不变

    def test_overwrite_empty_data(self, sample_file):
        """空数据报错"""
        success, msg, meta = StreamingWriter.update_range(
            sample_file, "Skills", 2, 1, []
        )
        assert success is False

    def test_overwrite_extends_rows(self, sample_file):
        """写入超出当前行数时自动扩展"""
        success, msg, meta = StreamingWriter.update_range(
            sample_file, "Skills", 7, 1, [[999, "Ultima", 9999, "Ultimate"]]
        )
        assert success is True
        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert len(rows) == 7
        assert rows[6][0] == 999

    def test_overwrite_preserves_other_sheets(self, sample_file):
        """覆盖不影响其他工作表"""
        StreamingWriter.update_range(sample_file, "Skills", 2, 1, [[99, "X", 0, "Y"]])
        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Items"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert rows[1] == (101, "Sword")

    def test_overwrite_preserves_col_widths(self, sample_file):
        """覆盖保留列宽"""
        StreamingWriter.update_range(sample_file, "Skills", 2, 1, [[99, "X", 0, "Y"]])
        wb = load_workbook(sample_file)
        ws = wb["Skills"]
        assert ws.column_dimensions['A'].width == 20
        wb.close()

    def test_overwrite_sheet_not_found(self, sample_file):
        """工作表不存在"""
        success, msg, meta = StreamingWriter.update_range(
            sample_file, "NotExist", 1, 1, [["x"]]
        )
        assert success is False


class TestPerformanceCompare:
    """性能对比：streaming vs openpyxl（修改操作）"""

    def test_delete_rows_large_file(self, tmp_path):
        """大文件删除行性能对比"""
        # 创建大文件（10000行）
        fp = str(tmp_path / "large.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["ID", "Name", "Value", "Desc", "Extra"])
        for i in range(1, 10001):
            ws.append([i, f"item_{i}", i * 10, f"description_{i}", f"extra_{i}"])
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        wb.save(fp)
        wb.close()

        # streaming 删除
        start = time.time()
        success, msg, meta = StreamingWriter.delete_rows(fp, "Data", 5000, 100)
        streaming_time = time.time() - start
        assert success is True

        # openpyxl 删除（重建文件）
        fp2 = str(tmp_path / "large2.xlsx")
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Data"
        ws2.append(["ID", "Name", "Value", "Desc", "Extra"])
        for i in range(1, 10001):
            ws2.append([i, f"item_{i}", i * 10, f"description_{i}", f"extra_{i}"])
        ws2.column_dimensions['A'].width = 10
        ws2.column_dimensions['B'].width = 20
        wb2.save(fp2)
        wb2.close()

        start = time.time()
        from excel_mcp_server_fastmcp.core.excel_writer import ExcelWriter
        writer = ExcelWriter(fp2)
        result = writer.delete_rows("Data", 5000, 100)
        openpyxl_time = time.time() - start
        assert result.success is True

        print(f"\n📊 删除行性能对比 (10000行文件，删除100行):")
        print(f"  Streaming: {streaming_time:.3f}s")
        print(f"  OpenPyXL:  {openpyxl_time:.3f}s")
        print(f"  比率: {openpyxl_time / max(streaming_time, 0.001):.1f}x")
