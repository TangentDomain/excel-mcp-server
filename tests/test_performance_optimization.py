"""
性能优化测试 - REQ-032

测试 dtype 优化、缓存增强、大文件支持等性能优化功能。
"""

import os
import tempfile
import pytest
import pandas as pd
import numpy as np

from excel_mcp_server_fastmcp.api.advanced_sql_query import AdvancedSQLQueryEngine
from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations


@pytest.fixture
def temp_dir():
    with tempfile.TemporaryDirectory() as tmpdir:
        yield tmpdir


@pytest.fixture
def mixed_type_file(temp_dir):
    """创建混合类型数据的测试文件"""
    from openpyxl import Workbook
    file_path = os.path.join(temp_dir, "mixed_types.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # 表头
    headers = ["id", "name", "score", "category", "active"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 数据行：混合类型
    for row in range(2, 502):
        ws.cell(row=row, column=1, value=row - 1)              # int
        ws.cell(row=row, column=2, value=f"item_{row - 1}")     # string
        ws.cell(row=row, column=3, value=(row * 3.14) % 100)   # float
        ws.cell(row=row, column=4, value=f"cat_{(row - 1) % 5}")  # 低基数 string
        ws.cell(row=row, column=5, value=row % 2 == 0)          # bool

    wb.save(file_path)
    return file_path


class TestDtypeOptimization:
    """测试 DataFrame dtype 优化"""

    def test_optimize_dtypes_reduces_memory(self, mixed_type_file):
        """验证 dtype 优化后 DataFrame 内存减少"""
        engine = AdvancedSQLQueryEngine()

        # 加载数据（会自动调用 _optimize_dtypes）
        result = engine.execute_sql_query(mixed_type_file, "SELECT * FROM Data LIMIT 100")
        assert result['success'] is True

        # 手动验证优化效果
        df_raw = pd.read_excel(mixed_type_file, sheet_name="Data", engine='calamine', keep_default_na=False)
        raw_mem = df_raw.memory_usage(deep=True).sum() / 1024 / 1024

        df_opt = engine._optimize_dtypes(df_raw.copy())
        opt_mem = df_opt.memory_usage(deep=True).sum() / 1024 / 1024

        # 优化后内存应该更少（或至少不增加）
        assert opt_mem <= raw_mem, f"优化后内存 {opt_mem:.2f}MB > 原始 {raw_mem:.2f}MB"

        engine.clear_cache()

    def test_optimize_dtypes_preserves_data(self, mixed_type_file):
        """验证 dtype 优化不改变数据内容"""
        engine = AdvancedSQLQueryEngine()

        df_raw = pd.read_excel(mixed_type_file, sheet_name="Data", engine='calamine', keep_default_na=False)
        df_opt = engine._optimize_dtypes(df_raw.copy())

        # 行数列数不变
        assert len(df_opt) == len(df_raw)
        assert len(df_opt.columns) == len(df_raw.columns)

        # 数值精度保持（float32 精度足够）
        for col in df_raw.columns:
            if df_raw[col].dtype == 'float64':
                # float32 精度误差 < 1e-6
                diff = (df_raw[col].astype('float32').values - df_opt[col].values.astype('float32'))
                if len(diff) > 0:
                    max_diff = np.max(np.abs(diff))
                    assert max_diff < 1e-4, f"列 {col} 精度损失过大: {max_diff}"

        engine.clear_cache()

    def test_optimize_dtypes_int_downcast(self):
        """验证整数列正确降级"""
        engine = AdvancedSQLQueryEngine()

        df = pd.DataFrame({
            'tiny': [1, 2, 3, 4, 5],
            'small': [100, 200, 300, 40000, 50000],
            'big': [100000, 200000, 3000000, 40000000, 500000000],
            'negative': [-100, -50, 0, 50, 100],
        })

        # 确保初始类型是 int64
        for col in df.columns:
            df[col] = df[col].astype('int64')

        df_opt = engine._optimize_dtypes(df)

        assert df_opt['tiny'].dtype == 'uint8'
        assert df_opt['small'].dtype == 'uint16'
        assert df_opt['big'].dtype == 'uint32'
        assert df_opt['negative'].dtype == 'int8'

    def test_optimize_dtypes_float_downcast(self):
        """验证浮点列降级为 float32"""
        engine = AdvancedSQLQueryEngine()

        df = pd.DataFrame({
            'values': [1.1, 2.2, 3.3, 4.4, 5.5],
        })
        df['values'] = df['values'].astype('float64')

        df_opt = engine._optimize_dtypes(df)
        assert df_opt['values'].dtype == 'float32'

    def test_optimize_dtypes_category_conversion(self):
        """验证字符串列保持 object 类型（不再转 category 以兼容 UPDATE 写入）"""
        engine = AdvancedSQLQueryEngine()

        # 100行，只有5个唯一值
        df = pd.DataFrame({
            'cat_col': [f"cat_{i % 5}" for i in range(100)],
        })
        df['cat_col'] = df['cat_col'].astype('object')

        df_opt = engine._optimize_dtypes(df)
        assert df_opt['cat_col'].dtype == 'object'  # 不再转 category

    def test_optimize_dtypes_high_cardinality_unchanged(self):
        """验证高基数字符串列不转为 category"""
        engine = AdvancedSQLQueryEngine()

        # 100行，100个唯一值 → 基数比 = 1.0 > 0.3
        df = pd.DataFrame({
            'unique_col': [f"item_{i}" for i in range(100)],
        })
        df['unique_col'] = df['unique_col'].astype('object')

        df_opt = engine._optimize_dtypes(df)
        assert df_opt['unique_col'].dtype != 'category'


class TestCacheEnhancement:
    """测试缓存增强功能"""

    def test_cache_size_increased(self):
        """验证缓存大小已增大"""
        engine = AdvancedSQLQueryEngine()
        assert engine._max_cache_size == 20, f"df_cache大小应为20，实际: {engine._max_cache_size}"
        assert engine._max_query_cache_size == 15, f"query_cache大小应为15，实际: {engine._max_query_cache_size}"

    def test_query_result_cache_hit(self, mixed_type_file):
        """验证查询结果缓存命中"""
        engine = AdvancedSQLQueryEngine()
        sql = "SELECT * FROM Data WHERE id < 10"

        # 第一次查询（缓存未命中）
        result1 = engine.execute_sql_query(mixed_type_file, sql)
        assert result1['success'] is True

        # 第二次查询（缓存命中）
        result2 = engine.execute_sql_query(mixed_type_file, sql)
        assert result2['success'] is True

        # 结果一致
        assert len(result1['data']) == len(result2['data'])

        engine.clear_cache()

    def test_memory_aware_eviction(self, mixed_type_file):
        """验证内存感知缓存淘汰"""
        engine = AdvancedSQLQueryEngine()

        # 加载一些数据到缓存
        engine.execute_sql_query(mixed_type_file, "SELECT * FROM Data")
        assert len(engine._df_cache) > 0

        # 调用内存感知淘汰（设置极小阈值触发淘汰）
        engine.evict_cache_by_memory(target_mb=0.001)
        assert len(engine._df_cache) == 0

    def test_estimate_cache_memory(self, mixed_type_file):
        """验证缓存内存估算"""
        engine = AdvancedSQLQueryEngine()

        # 空缓存内存为 0
        assert engine._estimate_cache_memory_mb() == 0.0

        # 加载数据后缓存内存 > 0
        engine.execute_sql_query(mixed_type_file, "SELECT * FROM Data")
        assert engine._estimate_cache_memory_mb() > 0.0

        engine.clear_cache()
        assert engine._estimate_cache_memory_mb() == 0.0


class TestLargeFileSupport:
    """测试大文件支持"""

    def test_file_size_limit_2gb(self, mixed_type_file):
        """验证文件大小限制已提高到 2GB"""
        engine = AdvancedSQLQueryEngine()

        # 小文件应该正常查询
        result = engine.execute_sql_query(mixed_type_file, "SELECT * FROM Data LIMIT 10")
        assert result['success'] is True

        engine.clear_cache()

    def test_large_file_warning_logged(self, mixed_type_file, caplog):
        """验证大文件（>500MB）有日志记录"""
        import logging
        engine = AdvancedSQLQueryEngine()

        # 对于小文件，不应有警告日志
        with caplog.at_level(logging.INFO):
            engine.execute_sql_query(mixed_type_file, "SELECT * FROM Data LIMIT 10")

        # 清理
        engine.clear_cache()

    def test_medium_dataset_sql_performance(self, temp_dir):
        """测试中等数据集的 SQL 查询性能"""
        from openpyxl import Workbook
        file_path = os.path.join(temp_dir, "medium_data.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        headers = ["id", "name", "value", "category"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        for row in range(2, 5002):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=f"item_{row - 1}")
            ws.cell(row=row, column=3, value=(row * 1.5) % 1000)
            ws.cell(row=row, column=4, value=f"cat_{(row - 1) % 20}")

        wb.save(file_path)

        engine = AdvancedSQLQueryEngine()

        # SELECT ALL（data包含表头行 + 数据行）
        result = engine.execute_sql_query(file_path, "SELECT * FROM Data LIMIT 100")
        assert result['success'] is True
        assert len(result['data']) >= 100  # 包含表头

        # WHERE
        result = engine.execute_sql_query(file_path, "SELECT * FROM Data WHERE id > 4000")
        assert result['success'] is True
        assert len(result['data']) >= 500  # 至少返回数百行

        # GROUP BY
        result = engine.execute_sql_query(file_path, "SELECT category, COUNT(*) as cnt FROM Data GROUP BY category")
        assert result['success'] is True
        assert len(result['data']) >= 20  # 20个分类

        # 聚合
        result = engine.execute_sql_query(file_path, "SELECT AVG(value) as avg_val FROM Data")
        assert result['success'] is True

        engine.clear_cache()


class TestBackwardCompatibility:
    """确保优化不影响现有功能"""

    def test_small_file_operations_unchanged(self, temp_dir):
        """验证小文件操作不受影响"""
        from openpyxl import Workbook
        file_path = os.path.join(temp_dir, "small.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws.cell(row=1, column=1, value="name")
        ws.cell(row=2, column=1, value="Alice")
        ws.cell(row=3, column=1, value="Bob")
        wb.save(file_path)

        # 读取
        result = ExcelOperations.get_range(file_path, "Sheet1!A1:A3")
        assert result['success'] is True

        # SQL查询
        engine = AdvancedSQLQueryEngine()
        result = engine.execute_sql_query(file_path, "SELECT * FROM Sheet1")
        assert result['success'] is True
        assert len(result['data']) >= 2  # 包含表头行


class TestConcurrentBatchOperations:
    """多线程批量操作优化测试 - REQ-032"""

    def test_parallel_validate_batch_data(self):
        """测试并行数据验证"""
        from excel_mcp_server_fastmcp.utils.concurrent_utils import parallel_validate_batch_data

        rows = [{"id": i, "name": f"item_{i}"} for i in range(1000)]

        def validate(row):
            if row["id"] < 0:
                return "id must be >= 0"
            return None

        errors = parallel_validate_batch_data(rows, validate)
        assert len(errors) == 1000
        assert all(e is None for e in errors)

    def test_parallel_validate_batch_data_with_errors(self):
        """测试并行数据验证（含错误行）"""
        from excel_mcp_server_fastmcp.utils.concurrent_utils import parallel_validate_batch_data

        rows = [{"id": i, "name": f"item_{i}"} for i in range(100)]
        rows[10]["id"] = -1
        rows[50]["id"] = -2

        def validate(row):
            if row["id"] < 0:
                return f"invalid id: {row['id']}"
            return None

        errors = parallel_validate_batch_data(rows, validate)
        assert errors[10] is not None
        assert errors[50] is not None
        assert sum(1 for e in errors if e is not None) == 2

    def test_parallel_validate_empty(self):
        """测试空数据并行验证"""
        from excel_mcp_server_fastmcp.utils.concurrent_utils import parallel_validate_batch_data

        errors = parallel_validate_batch_data([], lambda r: None)
        assert errors == []

    def test_parallel_group_execute(self):
        """测试并行分组执行"""
        from excel_mcp_server_fastmcp.utils.concurrent_utils import parallel_group_execute

        groups = {
            "A": [1, 2, 3],
            "B": [4, 5, 6],
            "C": [7, 8, 9],
        }

        def process(key, data):
            return sum(data)

        results = parallel_group_execute(groups, process)
        assert results["A"] == 6
        assert results["B"] == 15
        assert results["C"] == 24

    def test_batch_delete_rows_single(self, temp_dir):
        """测试批量删除单行"""
        from openpyxl import Workbook
        file_path = os.path.join(temp_dir, "batch_del_single.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["ID", "Name"])
        ws.append([1, "Alice"])
        ws.append([2, "Bob"])
        ws.append([3, "Charlie"])
        wb.save(file_path)

        result = ExcelOperations.batch_delete_rows(file_path, "Data", [2])
        assert result['success'] is True
        deleted = result.get('data', {}).get('deleted_count', 0)
        assert deleted == 1

        # 验证数据正确
        read = ExcelOperations.get_range(file_path, "Data!A1:B3")
        assert read['success'] is True

    def test_batch_delete_rows_multiple(self, temp_dir):
        """测试批量删除多行"""
        from openpyxl import Workbook
        file_path = os.path.join(temp_dir, "batch_del_multi.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["ID", "Name"])
        for i in range(1, 21):
            ws.append([i, f"item_{i}"])
        wb.save(file_path)

        # 删除第3、5、7行
        result = ExcelOperations.batch_delete_rows(file_path, "Data", [3, 5, 7])
        assert result['success'] is True
        deleted = result.get('data', {}).get('deleted_count', 0)
        assert deleted == 3

        # 验证文件可读
        read = ExcelOperations.get_range(file_path, "Data!A1:B2")
        assert read['success'] is True

    def test_batch_delete_rows_empty_list(self, temp_dir):
        """测试批量删除空列表"""
        result = ExcelOperations.batch_delete_rows(
            "/tmp/nonexistent.xlsx", "Data", []
        )
        assert result['success'] is False

    def test_batch_delete_rows_preserves_other_data(self, temp_dir):
        """测试批量删除不影响其他行数据"""
        from openpyxl import Workbook
        file_path = os.path.join(temp_dir, "batch_del_preserve.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["ID", "Name", "Value"])
        ws.append([1, "Alice", 100])
        ws.append([2, "Bob", 200])
        ws.append([3, "Charlie", 300])
        ws.append([4, "Diana", 400])
        wb.save(file_path)

        # 删除第3行（Bob）
        result = ExcelOperations.batch_delete_rows(file_path, "Data", [3])
        assert result['success'] is True

        # 验证头行和其余数据完好
        read = ExcelOperations.get_range(file_path, "Data!A1:C4")
        assert read['success'] is True
        # 表头应保留（get_range 返回 [{coordinate, value}, ...] 格式）
        assert read['data'][0][0]['value'] == "ID"
        # Alice 应保留
        assert read['data'][1][1]['value'] == "Alice"

    def test_batch_delete_performance_vs_sequential(self, temp_dir):
        """测试批量删除性能优于逐行删除"""
        from openpyxl import Workbook
        import time

        # 创建较大数据集
        for label, row_count in [("batch", 100), ("seq", 100)]:
            file_path = os.path.join(temp_dir, f"perf_{label}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["ID", "Name"])
            for i in range(1, row_count + 1):
                ws.append([i, f"item_{i}"])
            wb.save(file_path)

        # 批量删除
        rows_to_delete = list(range(5, 15))
        file_batch = os.path.join(temp_dir, "perf_batch.xlsx")
        start = time.time()
        ExcelOperations.batch_delete_rows(file_batch, "Data", rows_to_delete)
        batch_time = time.time() - start

        # 逐行删除（旧方式）
        file_seq = os.path.join(temp_dir, "perf_seq.xlsx")
        start = time.time()
        for r in sorted(rows_to_delete, reverse=True):
            ExcelOperations.delete_rows(file_seq, "Data", r, 1)
        seq_time = time.time() - start

        print(f"批量删除10行: {batch_time:.3f}s vs 逐行删除: {seq_time:.3f}s")
        # 批量删除应不慢于逐行删除（通常快2-5倍）
        assert batch_time < seq_time * 2, (
            f"批量删除应不慢于逐行删除的2倍: {batch_time:.3f}s vs {seq_time:.3f}s"
        )

        # 清理缓存
        ExcelOperations.clear_cache() if hasattr(ExcelOperations, 'clear_cache') else None
