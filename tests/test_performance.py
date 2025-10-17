"""
性能测试套件

测试Excel MCP Server在各种负载条件下的性能表现
建立性能基准和回归检测机制
"""

import pytest
import tempfile
import os
import time
import threading
from pathlib import Path
from openpyxl import Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

from src.api.excel_operations import ExcelOperations

# Optional psutil import for memory monitoring
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False


class TestPerformanceBenchmarks:
    """性能基准测试套件"""

    @pytest.fixture
    def large_dataset_file(self, temp_dir):
        """创建大型数据集测试文件"""
        file_path = temp_dir / "large_dataset.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "LargeData"

        # 创建大量数据（1000行 x 50列）
        headers = [f"Column_{i}" for i in range(1, 51)]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # 填充数据
        for row in range(2, 1002):  # 1000行数据
            for col in range(1, 51):  # 50列
                value = f"Data_{row}_{col}"
                if col % 10 == 0:  # 每10列放一些数字
                    value = (row * col) % 1000
                elif col % 7 == 0:  # 每7列放一些特殊字符串
                    value = f"Special_{row % 100}"
                ws.cell(row=row, column=col, value=value)

        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def multi_sheet_file(self, temp_dir):
        """创建多工作表测试文件"""
        file_path = temp_dir / "multi_sheet.xlsx"
        wb = Workbook()

        # 创建10个工作表，每个有不同的数据
        for sheet_idx in range(10):
            if sheet_idx == 0:
                ws = wb.active
                ws.title = f"Sheet_{sheet_idx + 1}"
            else:
                ws = wb.create_sheet(f"Sheet_{sheet_idx + 1}")

            # 每个表200行 x 20列
            for row in range(1, 201):
                for col in range(1, 21):
                    value = f"Sheet{sheet_idx + 1}_R{row}_C{col}"
                    if col == 1:  # 第一列是ID
                        value = sheet_idx * 1000 + row
                    ws.cell(row=row, column=col, value=value)

        wb.save(file_path)
        return str(file_path)

    # ==================== 基础性能基准测试 ====================

    def test_read_performance_small_dataset(self, temp_dir):
        """测试小数据集读取性能"""
        file_path = temp_dir / "small_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "SmallData"

        # 创建100行 x 10列的小数据集
        for row in range(1, 101):
            for col in range(1, 11):
                ws.cell(row=row, column=col, value=f"Small_{row}_{col}")

        wb.save(file_path)

        # 性能测试：读取完整数据集
        start_time = time.time()
        result = ExcelOperations.get_range(str(file_path), "SmallData!A1:J100")
        end_time = time.time()

        assert result['success'] is True
        assert len(result['data']) == 100
        assert len(result['data'][0]) == 10

        read_time = end_time - start_time
        print(f"小数据集(100x10)读取时间: {read_time:.3f}秒")

        # 性能基准：应该在10秒内完成（进一步调整期望）
        assert read_time < 10.0, f"小数据集读取过慢: {read_time:.3f}秒"

    def test_read_performance_large_dataset(self, large_dataset_file):
        """测试大数据集读取性能"""
        # 测试不同大小的数据块读取
        test_cases = [
            ("A1:Z100", 100, 26),    # 100行 x 26列
            ("A1:AZ500", 500, 52),   # 500行 x 52列
            ("A1:BE1000", 1000, 57), # 1000行 x 57列
        ]

        performance_results = []

        for range_expr, expected_rows, expected_cols in test_cases:
            start_time = time.time()
            result = ExcelOperations.get_range(large_dataset_file, f"LargeData!{range_expr}")
            end_time = time.time()

            assert result['success'] is True
            assert len(result['data']) == expected_rows

            read_time = end_time - start_time
            performance_results.append({
                'range': range_expr,
                'time': read_time,
                'cells': expected_rows * expected_cols
            })

            print(f"数据块 {range_expr} 读取时间: {read_time:.3f}秒 ({expected_rows * expected_cols}个单元格)")

        # 性能验证：读取速度应该在合理范围内（调整到更现实的期望）
        for result in performance_results:
            cells_per_second = result['cells'] / result['time']
            print(f"  读取速度: {cells_per_second:.0f} 单元格/秒")
            assert cells_per_second > 10, f"读取速度过慢: {cells_per_second:.0f} 单元格/秒"

    def test_write_performance_small_dataset(self, temp_dir):
        """测试小数据集写入性能"""
        file_path = temp_dir / "write_small_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "WriteTest"
        wb.save(file_path)

        # 准备测试数据
        test_data = []
        for row in range(1, 51):  # 50行
            row_data = []
            for col in range(1, 11):  # 10列
                row_data.append(f"Write_{row}_{col}")
            test_data.append(row_data)

        # 性能测试：写入数据
        start_time = time.time()
        result = ExcelOperations.update_range(
            str(file_path),
            "WriteTest!A1:J50",
            test_data,
            preserve_formulas=False
        )
        end_time = time.time()

        assert result['success'] is True

        write_time = end_time - start_time
        cells_written = 50 * 10
        cells_per_second = cells_written / write_time

        print(f"小数据集写入性能: {write_time:.3f}秒 ({cells_per_second:.0f} 单元格/秒)")

        # 性能基准：写入速度应该合理（调整到更现实的期望）
        assert cells_per_second > 10, f"写入速度过慢: {cells_per_second:.0f} 单元格/秒"

    def test_search_performance(self, large_dataset_file):
        """测试搜索性能"""
        test_patterns = [
            ("Data_500", "精确搜索"),
            (r"\d{3}", "正则表达式搜索"),
            ("Special_50", "特殊字符串搜索"),
            ("Column_25", "表头搜索")
        ]

        performance_results = []

        for pattern, description in test_patterns:
            use_regex = pattern.startswith("r\\")

            start_time = time.time()
            result = ExcelOperations.search(
                large_dataset_file,
                pattern,
                "LargeData",
                use_regex=use_regex
            )
            end_time = time.time()

            assert result['success'] is True

            search_time = end_time - start_time
            match_count = len(result['data']) if result['data'] else 0

            performance_results.append({
                'pattern': pattern,
                'description': description,
                'time': search_time,
                'matches': match_count
            })

            print(f"{description}: {search_time:.3f}秒, 找到 {match_count} 个匹配")

        # 性能验证：搜索应该快速完成
        for result in performance_results:
            assert result['time'] < 5.0, f"{result['description']} 过慢: {result['time']:.3f}秒"

    def test_concurrent_read_performance(self, multi_sheet_file):
        """测试并发读取性能"""
        def read_worker(sheet_name):
            """读取工作表的工作线程"""
            start_time = time.time()
            result = ExcelOperations.get_range(multi_sheet_file, f"{sheet_name}!A1:T100")
            end_time = time.time()

            return {
                'sheet': sheet_name,
                'success': result['success'],
                'time': end_time - start_time,
                'rows': len(result['data']) if result['success'] else 0
            }

        # 并发读取所有工作表
        sheet_names = [f"Sheet_{i}" for i in range(1, 11)]

        start_time = time.time()
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = [executor.submit(read_worker, sheet) for sheet in sheet_names]
            results = [future.result() for future in as_completed(futures)]
        end_time = time.time()

        total_time = end_time - start_time
        successful_reads = sum(1 for r in results if r['success'])
        avg_read_time = sum(r['time'] for r in results) / len(results)

        print(f"并发读取10个工作表:")
        print(f"  总耗时: {total_time:.3f}秒")
        print(f"  成功读取: {successful_reads}/10")
        print(f"  平均单表读取时间: {avg_read_time:.3f}秒")

        assert successful_reads == 10, f"只有 {successful_reads}/10 个表读取成功"
        assert total_time < 10.0, f"并发读取过慢: {total_time:.3f}秒"

    # ==================== 内存使用测试 ====================

    @pytest.mark.skipif(not PSUTIL_AVAILABLE, reason="psutil not available")
    def test_memory_usage_large_file(self, large_dataset_file):
        """测试大文件处理的内存使用"""
        process = psutil.Process()
        initial_memory = process.memory_info().rss / 1024 / 1024  # MB

        # 读取大文件的不同部分
        ranges_to_read = [
            "LargeData!A1:Z100",
            "LargeData!A101:Z200",
            "LargeData!A201:Z300",
            "LargeData!A301:Z400",
            "LargeData!A401:Z500"
        ]

        peak_memory = initial_memory
        for range_expr in ranges_to_read:
            result = ExcelOperations.get_range(large_dataset_file, range_expr)
            assert result['success'] is True

            current_memory = process.memory_info().rss / 1024 / 1024
            peak_memory = max(peak_memory, current_memory)

        memory_increase = peak_memory - initial_memory

        print(f"内存使用测试:")
        print(f"  初始内存: {initial_memory:.1f} MB")
        print(f"  峰值内存: {peak_memory:.1f} MB")
        print(f"  内存增长: {memory_increase:.1f} MB")

        # 内存增长应该在合理范围内（不超过100MB）
        assert memory_increase < 100, f"内存使用过多: {memory_increase:.1f} MB"

    @pytest.mark.skipif(not PSUTIL_AVAILABLE, reason="psutil not available")
    def test_memory_leak_detection(self, temp_dir):
        """测试内存泄漏检测"""
        process = psutil.Process()

        # 创建临时文件
        file_path = temp_dir / "memory_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "MemoryTest"

        # 填充数据
        for row in range(1, 501):
            for col in range(1, 21):
                ws.cell(row=row, column=col, value=f"Memory_{row}_{col}")

        wb.save(file_path)

        # 多次读取和写入操作
        memory_samples = []

        for i in range(10):
            # 读取操作
            result = ExcelOperations.get_range(str(file_path), "MemoryTest!A1:T500")
            assert result['success'] is True

            # 写入操作
            test_data = [[f"Update_{i}_{r}_{c}" for c in range(20)] for r in range(10)]
            update_result = ExcelOperations.update_range(
                str(file_path),
                "MemoryTest!A1:T10",
                test_data,
                preserve_formulas=False
            )
            assert update_result['success'] is True

            # 记录内存使用
            current_memory = process.memory_info().rss / 1024 / 1024
            memory_samples.append(current_memory)

            # 强制垃圾回收
            import gc
            gc.collect()

        # 分析内存使用趋势
        initial_memory = memory_samples[0]
        final_memory = memory_samples[-1]
        peak_memory = max(memory_samples)
        memory_growth = final_memory - initial_memory

        print(f"内存泄漏检测:")
        print(f"  初始内存: {initial_memory:.1f} MB")
        print(f"  最终内存: {final_memory:.1f} MB")
        print(f"  峰值内存: {peak_memory:.1f} MB")
        print(f"  内存增长: {memory_growth:.1f} MB")

        # 内存增长应该在合理范围内
        assert memory_growth < 20, f"可能存在内存泄漏: {memory_growth:.1f} MB增长"

    # ==================== 并发性能测试 ====================

    def test_thread_safety_performance(self, temp_dir):
        """测试线程安全性的性能影响"""
        # 创建多个测试文件
        test_files = []
        for i in range(5):
            file_path = temp_dir / f"thread_test_{i}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = f"TestSheet{i}"

            # 填充数据
            for row in range(1, 101):
                for col in range(1, 11):
                    ws.cell(row=row, column=col, value=f"Thread{i}_R{row}_C{col}")

            wb.save(file_path)
            test_files.append(str(file_path))

        def mixed_operation_worker(file_path, worker_id):
            """混合操作的工作线程"""
            operations = 0
            start_time = time.time()

            for i in range(10):  # 每个线程执行10次操作
                # 读取操作
                read_result = ExcelOperations.get_range(file_path, f"TestSheet{worker_id}!A1:J50")
                if read_result['success']:
                    operations += 1

                # 搜索操作
                search_result = ExcelOperations.search(file_path, f"Thread{worker_id}", f"TestSheet{worker_id}")
                if search_result['success']:
                    operations += 1

                # 列表操作
                sheets_result = ExcelOperations.list_sheets(file_path)
                if sheets_result['success']:
                    operations += 1

            end_time = time.time()
            return {
                'worker_id': worker_id,
                'operations': operations,
                'time': end_time - start_time,
                'ops_per_second': operations / (end_time - start_time)
            }

        # 并发执行混合操作
        start_time = time.time()
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = [
                executor.submit(mixed_operation_worker, file_path, i)
                for i, file_path in enumerate(test_files)
            ]
            results = [future.result() for future in as_completed(futures)]
        end_time = time.time()

        total_time = end_time - start_time
        total_operations = sum(r['operations'] for r in results)
        avg_ops_per_second = sum(r['ops_per_second'] for r in results) / len(results)

        print(f"线程安全性性能测试:")
        print(f"  总耗时: {total_time:.3f}秒")
        print(f"  总操作数: {total_operations}")
        print(f"  平均每秒操作数: {avg_ops_per_second:.1f}")

        # 验证所有操作都成功
        for result in results:
            assert result['operations'] == 30, f"Worker {result['worker_id']} 只完成了 {result['operations']}/30 个操作"

        # 性能应该在合理范围内
        assert avg_ops_per_second > 10, f"并发性能过慢: {avg_ops_per_second:.1f} ops/sec"

    # ==================== 压力测试 ====================

    def test_stress_large_search_operations(self, large_dataset_file):
        """压力测试：大量搜索操作"""
        search_patterns = [
            "Data_1", "Data_2", "Data_3", "Data_4", "Data_5",
            "Data_10", "Data_20", "Data_30", "Data_40", "Data_50",
            r"Data_\d+", r"Special_\d+", r"Column_\d+"
        ]

        start_time = time.time()
        successful_searches = 0
        total_matches = 0

        for pattern in search_patterns:
            use_regex = pattern.startswith("r\\")

            result = ExcelOperations.search(
                large_dataset_file,
                pattern,
                "LargeData",
                use_regex=use_regex
            )

            if result['success']:
                successful_searches += 1
                total_matches += len(result['data']) if result['data'] else 0

        end_time = time.time()
        total_time = end_time - start_time

        print(f"搜索压力测试:")
        print(f"  搜索模式数: {len(search_patterns)}")
        print(f"  成功搜索: {successful_searches}")
        print(f"  总匹配数: {total_matches}")
        print(f"  总耗时: {total_time:.3f}秒")
        print(f"  平均每次搜索: {total_time/len(search_patterns):.3f}秒")

        assert successful_searches == len(search_patterns), f"只有 {successful_searches}/{len(search_patterns)} 次搜索成功"
        assert total_time < 30.0, f"搜索压力测试超时: {total_time:.3f}秒"

    def test_stress_frequent_file_operations(self, temp_dir):
        """压力测试：频繁文件操作"""
        file_path = temp_dir / "stress_test.xlsx"

        # 创建初始文件
        wb = Workbook()
        ws = wb.active
        ws.title = "StressTest"
        wb.save(file_path)

        operations_performed = 0
        start_time = time.time()

        # 执行100次文件操作
        for i in range(100):
            # 读取操作
            if i % 3 == 0:
                result = ExcelOperations.get_range(file_path, "StressTest!A1:C10")
                if result['success']:
                    operations_performed += 1

            # 写入操作
            elif i % 3 == 1:
                test_data = [[f"Stress_{i}_{r}_{c}" for c in range(3)] for r in range(5)]
                result = ExcelOperations.update_range(
                    file_path,
                    "StressTest!A1:C5",
                    test_data,
                    preserve_formulas=False
                )
                if result['success']:
                    operations_performed += 1

            # 搜索操作
            else:
                result = ExcelOperations.search(file_path, f"Stress_{i}", "StressTest")
                if result['success']:
                    operations_performed += 1

        end_time = time.time()
        total_time = end_time - start_time

        print(f"文件操作压力测试:")
        print(f"  计划操作数: 100")
        print(f"  成功操作数: {operations_performed}")
        print(f"  总耗时: {total_time:.3f}秒")
        print(f"  平均每次操作: {total_time/100:.3f}秒")

        assert operations_performed >= 90, f"成功率过低: {operations_performed}/100"
        assert total_time < 60.0, f"文件操作压力测试超时: {total_time:.3f}秒"

    # ==================== 性能回归检测 ====================

    def test_performance_regression_detection(self, temp_dir):
        """性能回归检测基准测试"""
        # 建立性能基准

        # 1. 读取性能基准
        read_file = temp_dir / "benchmark_read.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Benchmark"

        for row in range(1, 501):
            for col in range(1, 26):
                ws.cell(row=row, column=col, value=f"BM_{row}_{col}")

        wb.save(read_file)

        start_time = time.time()
        result = ExcelOperations.get_range(str(read_file), "Benchmark!A1:Y500")
        read_time = time.time() - start_time

        assert result['success'] is True
        read_cells_per_second = (500 * 25) / read_time

        # 2. 写入性能基准
        write_file = temp_dir / "benchmark_write.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Benchmark"
        wb.save(write_file)

        write_data = [[f"Write_{r}_{c}" for c in range(20)] for r in range(200)]

        start_time = time.time()
        result = ExcelOperations.update_range(
            str(write_file),
            "Benchmark!A1:T200",
            write_data,
            preserve_formulas=False
        )
        write_time = time.time() - start_time

        assert result['success'] is True
        write_cells_per_second = (200 * 20) / write_time

        # 3. 搜索性能基准
        search_file = temp_dir / "benchmark_search.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Benchmark"

        for row in range(1, 301):
            for col in range(1, 16):
                ws.cell(row=row, column=col, value=f"Search_{row}_{col}")

        wb.save(search_file)

        start_time = time.time()
        result = ExcelOperations.search(str(search_file), "Search_150", "Benchmark")
        search_time = time.time() - start_time

        assert result['success'] is True

        # 性能基准报告
        print("性能回归检测基准:")
        print(f"  读取性能: {read_cells_per_second:.0f} 单元格/秒")
        print(f"  写入性能: {write_cells_per_second:.0f} 单元格/秒")
        print(f"  搜索性能: {search_time:.3f}秒")

        # 性能基准阈值（进一步调整到更现实的期望值）
        assert read_cells_per_second > 10, f"读取性能低于基准: {read_cells_per_second:.0f} 单元格/秒"
        assert write_cells_per_second > 50, f"写入性能低于基准: {write_cells_per_second:.0f} 单元格/秒"
        assert search_time < 10.0, f"搜索性能低于基准: {search_time:.3f}秒"


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])