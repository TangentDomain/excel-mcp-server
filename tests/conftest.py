"""
Test configuration and fixtures for Excel MCP Server tests
"""

import logging
import shutil
import tempfile
import time
import uuid
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# 注册自定义标记（消除 PytestUnknownMarkWarning）
def pytest_configure(config):
    config.addinivalue_line("markers", "slow: 标记测试为慢速测试")
    config.addinivalue_line("markers", "integration: 标记集成测试")
    config.addinivalue_line("markers", "security: 标记安全功能测试")
    config.addinivalue_line("markers", "unit: 标记单元测试")
    config.addinivalue_line("markers", "performance: 标记性能测试")
    config.addinivalue_line("markers", "api: 标记API测试")
    config.addinivalue_line("markers", "core: 标记核心模块测试")
    config.addinivalue_line("markers", "utils: 标记工具模块测试")
    config.addinivalue_line("markers", "xdist_group: pytest-xdist 并行分组")
    config.addinivalue_line("markers", "timeout: 超时限制")


# Set up logging to suppress warnings
logging.getLogger("openpyxl").setLevel(logging.ERROR)


@pytest.fixture(autouse=True)
def clear_sql_engine_cache():
    """在每个测试之前清除SQL引擎缓存，避免并行测试时的缓存污染"""
    try:
        from excel_mcp_server_fastmcp.api.advanced_sql_query import _get_engine

        engine = _get_engine()
        engine.clear_cache()
    except ImportError:
        # 如果sqlglot未安装，跳过
        pass


def safe_rmtree(path, max_retries=5, delay=0.1):
    """Safely remove directory tree with retry mechanism for Windows file locking"""
    import gc

    for attempt in range(max_retries):
        try:
            # Force garbage collection before attempting removal
            gc.collect()

            # Try to remove the directory
            shutil.rmtree(path)
            return

        except PermissionError as e:
            if attempt == max_retries - 1:
                # Last attempt failed, try to remove individual files first
                try:
                    gc.collect()
                    for file_path in Path(path).rglob("*"):
                        if file_path.is_file():
                            try:
                                file_path.unlink(missing_ok=True)
                            except PermissionError:
                                pass
                    # Try directory removal again
                    shutil.rmtree(path, ignore_errors=True)
                    return
                except Exception:
                    # If all else fails, just log warning
                    logging.warning(f"Could not remove temp directory {path}: {e}")
                    return

            # Wait and retry
            time.sleep(delay)
            delay *= 2  # Exponential backoff
            gc.collect()  # Force GC between attempts


@pytest.fixture
def temp_dir():
    """Create a temporary directory for test files"""
    temp_path = Path(tempfile.mkdtemp())

    try:
        yield temp_path
    finally:
        # Ensure all Excel file handles are closed by forcing garbage collection
        import gc
        import warnings

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=ResourceWarning)
            gc.collect()
        safe_rmtree(temp_path)


@pytest.fixture
def temp_dir_with_excel_files(temp_dir):
    """Create a temporary directory with sample Excel files"""
    # Create some sample Excel files
    for i in range(3):
        wb = Workbook()
        ws = wb.active
        ws.title = f"Sheet{i + 1}"

        # Add some test data
        ws["A1"] = f"标题{i + 1}"
        ws["B1"] = f"数据{i + 1}"
        ws["A2"] = f"内容{i + 1}"
        ws["B2"] = i * 100

        file_path = temp_dir / f"test_file_{i + 1}.xlsx"
        wb.save(str(file_path))

    yield str(temp_dir)


@pytest.fixture
def sample_excel_file(temp_dir, request):
    """Create a sample Excel file for testing with unique name"""
    # Generate unique filename for each test
    test_id = str(uuid.uuid4())[:8]
    test_name = request.node.name
    file_path = temp_dir / f"test_sample_{test_name}_{test_id}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add sample data - dual header format for excel_get_headers compatibility
    # 第1行：字段描述（descriptions）
    ws.append(["姓名描述", "年龄描述", "部门描述", "薪资描述", "总计描述"])
    # 第2行：字段名（field_names）
    ws.append(["name", "age", "department", "salary", "total"])
    # 第3行开始：实际数据
    data = [
        ["张三", 25, "技术部", 8000],
        ["李四", 30, "市场部", 9000],
        ["王五", 28, "技术部", 8500],
        ["赵六", 35, "人事部", 9500],
    ]

    for row in data:
        ws.append(row)

    # Add some formulas
    ws["E3"] = "=SUM(D3:D6)"

    # Add formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Create second sheet with dual header format
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["产品描述", "销量描述", "单价描述"])  # Row 1: descriptions
    ws2.append(["product", "sales", "price"])  # Row 2: field_names
    ws2.append(["A", 100, 50])  # Row 3+: actual data
    ws2.append(["B", 200, 30])

    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def empty_excel_file(temp_dir, request):
    """Create an empty Excel file for testing with unique name"""
    # Generate unique filename for each test
    test_id = str(uuid.uuid4())[:8]
    test_name = request.node.name
    file_path = temp_dir / f"test_empty_{test_name}_{test_id}.xlsx"

    wb = Workbook()
    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def multi_sheet_excel_file(temp_dir, request):
    """Create an Excel file with multiple sheets for testing with unique name"""
    # Generate unique filename for each test
    test_id = str(uuid.uuid4())[:8]
    test_name = request.node.name
    file_path = temp_dir / f"test_multi_sheet_{test_name}_{test_id}.xlsx"

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create multiple sheets with dual header format
    sheet_names = ["数据", "图表", "汇总", "分析"]
    for name in sheet_names:
        ws = wb.create_sheet(name)
        ws.append(["测试数据描述", "值描述"])  # Row 1: descriptions
        ws.append(["test_data", "value"])  # Row 2: field_names
        ws.append(["项目1", 100])  # Row 3+: actual data
        ws.append(["项目2", 200])

    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def formula_excel_file(temp_dir):
    """Create an Excel file with various formulas for testing"""
    file_path = temp_dir / "test_formulas.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Formulas"

    # Add data for formulas
    data = [
        ["数值A", "数值B", "和", "积", "平均"],
        [10, 20, "=A2+B2", "=A2*B2", "=AVERAGE(A2:B2)"],
        [30, 40, "=A3+B3", "=A3*B3", "=AVERAGE(A3:B3)"],
        [50, 60, "=A4+B4", "=A4*B4", "=AVERAGE(A4:B4)"],
    ]

    for row in data:
        ws.append(row)

    # Add summary formulas
    ws["A6"] = "总计"
    ws["B6"] = "=SUM(A2:A4)"
    ws["C6"] = "=SUM(C2:C4)"

    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def game_config_file():
    """Provide path to the game config test file with dual-row headers"""
    return str(Path(__file__).parent / "test_data" / "game_config.xlsx")
