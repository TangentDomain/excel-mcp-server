#!/usr/bin/env python3
"""
Round 32 MCP 接口实测 - 错误注入与容错恢复测试 + P0第5轮回归验证

创造性方向: 🔄 错误注入与容错恢复测试（之前未做过的方向！）
核心问题: 各种错误发生后，系统是否能正常继续工作？数据是否一致？

测试组:
  A组: P0漏洞第5轮回归确认（确定性验证）
  B组: 错误后恢复能力（错误→正确查询是否正常）
  C组: 错误消息质量审计（错误信息是否有用）
  D组: 部分失败场景（批量操作中的部分失败）
  E组: 文件状态一致性（错误后文件是否可读）
  F组: SQL语法容错（畸形SQL的优雅处理）
  G组: 极端输入容错（空字符串、超长、特殊字符组合）
"""

import os
import sys
import tempfile
import shutil
import traceback
from datetime import datetime

# 确保能导入项目模块
sys.path.insert(0, '/root/workspace/excel-mcp-server/src')
os.chdir('/root/workspace/excel-mcp-server')

from openpyxl import Workbook
import pandas as pd

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)

# ==================== 测试基础设施 ====================

TEST_DIR = tempfile.mkdtemp(prefix='r32_test_')
TEST_FILE = os.path.join(TEST_DIR, 'test_recovery.xlsx')

import pytest

pytestmark = pytest.mark.skip(reason="legacy test runner pattern, not for direct pytest")

class TestResult:
    def __init__(self):
        self.passed = 0
        self.failed = 0
        self.results = []
    
    def ok(self, name, detail=""):
        self.passed += 1
        self.results.append((name, True, detail))
        print(f"  ✅ {name}")
    
    def fail(self, name, detail=""):
        self.failed += 1
        self.results.append((name, False, detail))
        print(f"  ❌ {name}")
        if detail:
            print(f"     → {detail}")

def create_test_file(path=TEST_FILE):
    """创建标准测试文件: Sheet1(ID, Name, Value, Price), 5行数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "Name", "Value", "Price"])
    for i in range(1, 6):
        ws.append([i, f"Item-{i}", i * 10, round(i * 99.9, 2)])
    wb.save(path)
    return path

def create_multi_sheet_file(path=os.path.join(TEST_DIR, 'multi.xlsx')):
    """创建多Sheet测试文件"""
    wb = Workbook()
    # Sheet1
    ws1 = wb.active
    ws1.title = "Players"
    ws1.append(["ID", "Name", "Level", "GuildID"])
    for i in range(1, 6):
        ws1.append([i, f"Player-{i}", i * 10, (i % 3) + 1])
    # Sheet2
    ws2 = wb.create_sheet("Guilds")
    ws2.append(["ID", "Name", "MasterID"])
    ws2.append([1, "Guild-A", 1])
    ws2.append([2, "Guild-B", 2])
    ws2.append([3, "Guild-C", 3])
    wb.save(path)
    return path

def fresh_file():
    """获取一个全新的测试文件路径"""
    path = os.path.join(TEST_DIR, f'fresh_{id(os.urandom(4))}.xlsx')
    return create_test_file(path)

# ==================== A组: P0漏洞第5轮回归确认 ====================
def test_group_a_p0_regression(tr):
    """A组: P0安全漏洞第5轮确定性回归验证"""
    print("\n" + "="*60)
    print("🔴 A组: P0漏洞第5轮回归确认")
    print("="*60)
    
    f = fresh_file()
    
    # P0-2: SELECT 分号多语句
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1; SELECT * FROM Sheet1")
    if r.get('success') and '多语句' in str(r.get('message', '')):
        tr.fail("A1 [P0-2] SELECT分号多语句", f"仍可执行! msg={str(r.get('message',''))[:80]}")
    elif not r.get('success'):
        tr.ok("A1 [P0-2] SELECT分号多语句", "已拦截/拒绝 ✅")
    else:
        tr.fail("A1 [P0-2] SELECT分号多语句", f"异常状态: success={r.get('success')}, msg={str(r.get('message',''))[:80]}")
    
    # P0-4: UPDATE 分号多语句
    f2 = fresh_file()
    r = execute_advanced_update_query(f2, "UPDATE Sheet1 SET Value = 999 WHERE ID = 1; DROP TABLE Sheet1")
    if r.get('success'):
        # 验证是否真的执行了更新
        r2 = execute_advanced_sql_query(f2, "SELECT Value FROM Sheet1 WHERE ID = 1")
        if r2.get('success') and r2.get('data'):
            val = r2['data'][1][0] if len(r2['data']) > 1 else None  # skip header
            if val == 999:
                tr.fail("A2 [P0-4] UPDATE分号多语句", f"更新成功且值变为999! 分号后语句被忽略但未拒绝!")
            else:
                tr.ok("A2 [P0-4] UPDATE分号多语句", f"值={val}, 可能未实际执行")
        else:
            tr.fail("A2 [P0-4] UPDATE分号多语句", f"返回success但无法验证结果")
    else:
        tr.ok("A2 [P0-4] UPDATE分号多语句", "已拒绝 ❌→✅ FIXED?")
    
    # P0-5: INSERT 分号多语句
    f3 = fresh_file()
    r = execute_advanced_insert_query(f3, "INSERT INTO Sheet1 (ID, Name, Value, Price) VALUES (100, 'Hack', 0, 0); DROP TABLE Sheet1")
    if r.get('success'):
        r2 = execute_advanced_sql_query(f3, "SELECT COUNT(*) FROM Sheet1")
        if r2.get('success') and r2.get('data'):
            cnt = r2['data'][1][0] if len(r2['data']) > 1 else None
            if cnt and cnt > 5:  # 原来只有5行
                tr.fail("A3 [P0-5] INSERT分号多语句", f"插入成功! 总行数={cnt}")
            else:
                tr.ok("A3 [P0-5] INSERT分号多语句", f"行数未变={cnt}")
        else:
            tr.fail("A3 [P0-5] INSERT分号多语句", "返回success但无法验证")
    else:
        tr.ok("A3 [P0-5] INSERT分号多语句", "已拒绝 ✅?")
    
    # P0-6: DELETE 分号多语句
    f4 = fresh_file()
    r = execute_advanced_delete_query(f4, "DELETE FROM Sheet1 WHERE ID = 999; DROP TABLE Sheet1")
    if r.get('success'):
        # 检查原始数据是否还在
        r2 = execute_advanced_sql_query(f4, "SELECT COUNT(*) FROM Sheet1")
        if r2.get('success') and r2.get('data'):
            cnt = r2['data'][1][0] if len(r2['data']) > 1 else None
            tr.fail("A4 [P0-6] DELETE分号多语句", f"执行成功! 剩余行数={cnt} (应为5)")
        else:
            tr.fail("A4 [P0-6] DELETE分号多语句", "执行成功但后续查询失败!")
    else:
        tr.ok("A4 [P0-6] DELETE分号多语句", "已拒绝 ✅?")
    
    # P0-7: UPDATE 注释符全表篡改
    f5 = fresh_file()
    r = execute_advanced_update_query(f5, "UPDATE Sheet1 SET Value = -1 -- WHERE ID = 1")
    if r.get('success'):
        r2 = execute_advanced_sql_query(f5, "SELECT SUM(Value) FROM Sheet1")
        if r2.get('success') and r2.get('data'):
            total = r2['data'][1][0] if len(r2['data']) > 1 else None
            # 正常SUM应该是 10+20+30+40+50=150; 如果全表被改为-1则是-5
            if total == -5 or total == -1 * 5:
                tr.fail("A5 [P0-7] UPDATE注释符全表篡改", f"🚨 全表被篡改! SUM(Value)={total} (预期150)")
            elif total == 150:
                tr.ok("A5 [P0-7] UPDATE注释符全表篡改", f"WHERE条件生效! SUM={total} ✅ FIXED!")
            else:
                tr.fail("A5 [P0-7] UPDATE注释符全表篡改", f"异常SUM值: {total}")
        else:
            tr.fail("A5 [P0-7] UPDATE注释符全表篡改", "无法验证结果")
    else:
        tr.ok("A5 [P0-7] UPDATE注释符全表篡改", "已拒绝 ✅?")


# ==================== B组: 错误后恢复能力 ====================
def test_group_b_error_recovery(tr):
    """B组: 错误发生后系统是否能正常恢复"""
    print("\n" + "="*60)
    print("🔄 B组: 错误后恢复能力测试")
    print("="*60)
    
    # B1: 语法错误后正常查询能否工作
    f = fresh_file()
    r_err = execute_advanced_sql_query(f, "SELECTTTT * FROMM Sheet1")  # 故意语法错误
    r_ok = execute_advanced_sql_query(f, "SELECT * FROM Sheet1")
    if r_ok.get('success') and r_ok.get('data') and len(r_ok['data']) == 6:  # 1 header + 5 rows
        tr.ok("B1 语法错误后SELECT恢复", "语法错误后正常查询可用")
    else:
        tr.fail("B1 语法错误后SELECT恢复", f"错误后查询失败: success={r_ok.get('success')}")
    
    # B2: 表名错误后正确查询
    f2 = fresh_file()
    r_err = execute_advanced_sql_query(f2, "SELECT * FROM NonExistentTable")
    r_ok = execute_advanced_sql_query(f2, "SELECT * FROM Sheet1")
    if r_ok.get('success') and r_ok.get('data'):
        tr.ok("B2 表名错误后恢复", "查不存在的表后正常查询可用")
    else:
        tr.fail("B2 表名错误后恢复", "错误后查询不可用")
    
    # B3: 列名错误后正确查询
    f3 = fresh_file()
    r_err = execute_advanced_sql_query(f3, "SELECT NonExistentCol FROM Sheet1")
    r_ok = execute_advanced_sql_query(f3, "SELECT ID, Name FROM Sheet1")
    if r_ok.get('success') and r_ok.get('data'):
        tr.ok("B3 列名错误后恢复", "列名错误后正常查询可用")
    else:
        tr.fail("B3 列名错误后恢复", "错误后查询不可用")
    
    # B4: UPDATE失败(语法错)后SELECT
    f4 = fresh_file()
    r_err = execute_advanced_update_query(f4, "UPDATTE Sheet1 SETT Value = 999")
    r_ok = execute_advanced_sql_query(f4, "SELECT * FROM Sheet1")
    if r_ok.get('success') and r_ok.get('data'):
        tr.ok("B4 UPDATE语法错误后SELECT恢复", "UPDATE失败后SELECT正常")
    else:
        tr.fail("B4 UPDATE语法错误后SELECT恢复", "UPDATE失败后SELECT异常")
    
    # B5: 连续多次不同类型错误后的恢复
    f5 = fresh_file()
    errors = [
        ("SELECT", "SELEC * FROM Sheet1"),
        ("UPDATE", "UPDATE Sheet1 SET Value = 'x' WHER ID = 1"),  # WHERE typo
        ("INSERT", "INSERT INTOO Sheet1 ..."),
        ("DELETE", "DELETEE FROM Sheet1 WHERE ID = 1"),
    ]
    all_recovered = True
    for op, bad_sql in errors:
        if op == "SELECT":
            execute_advanced_sql_query(f5, bad_sql)
        elif op == "UPDATE":
            execute_advanced_update_query(f5, bad_sql)
        elif op == "INSERT":
            execute_advanced_insert_query(f5, bad_sql)
        elif op == "DELETE":
            execute_advanced_delete_query(f5, bad_sql)
    
    r_final = execute_advanced_sql_query(f5, "SELECT COUNT(*) FROM Sheet1")
    if r_final.get('success') and r_final.get('data'):
        tr.ok("B5 连续4次错误后恢复", "多种错误后系统仍正常工作")
    else:
        tr.fail("B5 连续4次错误后恢复", "连续错误后系统异常")
        all_recovered = False
    
    # B6: 空文件路径错误后正常文件操作
    r_err = execute_advanced_sql_query("", "SELECT * FROM Sheet1")
    f6 = fresh_file()
    r_ok = execute_advanced_sql_query(f6, "SELECT * FROM Sheet1")
    if r_ok.get('success') and r_ok.get('data'):
        tr.ok("B6 空路径错误后恢复", "空路径错误后正常文件操作OK")
    else:
        tr.fail("B6 空路径错误后恢复", "空路径错误后异常")


# ==================== C组: 错误消息质量审计 ====================
def test_group_c_error_quality(tr):
    """C组: 错误消息是否有用、友好、一致"""
    print("\n" + "="*60)
    print("📋 C组: 错误消息质量审计")
    print("="*60)
    
    f = TEST_FILE
    create_test_file(f)
    
    test_cases = [
        # (sql, operation, expected_keywords)
        ("SELECT * FROM NoExist", "query", ["找不到", "不存在", "not found", "NoExist", "sheet"]),
        ("SELECTT * FROM Sheet1", "query", ["语法", "syntax", "error", "无效", "invalid"]),
        ("SELECT FakeColumn FROM Sheet1", "query", ["列", "column", "FakeColumn", "找不到"]),
    ]
    
    for i, (sql, op, expected) in enumerate(test_cases, 1):
        if op == "query":
            r = execute_advanced_sql_query(f, sql)
        
        msg = str(r.get('message', '')).lower()
        msg_orig = r.get('message', '')
        
        # 检查: 是否有错误消息（非空）
        has_msg = bool(msg_orig and len(msg_orig) > 0)
        # 检查: 是否包含有用信息
        has_useful = any(kw.lower() in msg for kw in expected)
        # 检查: 是否非空消息或非通用消息
        is_generic = msg in ['', 'error', 'failed', '未知错误']
        
        if has_msg and (has_useful or not is_generic):
            tr.ok(f"C{i} 错误消息质量: {sql[:30]}", f"msg='{str(msg_orig)[:60]}'")
        else:
            tr.fail(f"C{i} 错误消息质量: {sql[:30]}", 
                   f"msg='{str(msg_orig)[:60]}' (has_msg={has_msg}, useful={has_useful})")
    
    # C4: 成功时消息是否也合理
    r = execute_advanced_sql_query(f, "SELECT COUNT(*) FROM Sheet1")
    if r.get('success') and r.get('message'):
        msg = str(r.get('message', ''))
        if len(msg) > 0:
            tr.ok("C4 成功消息质量", f"msg='{msg[:60]}'")
        else:
            tr.fail("C4 成功消息质量", "成功但消息为空")
    else:
        tr.fail("C4 成功消息质量", "基础查询失败!")


# ==================== D组: 部分失败场景 ====================
def test_group_d_partial_failure(tr):
    """D组: 批量/复杂操作中的部分失败"""
    print("\n" + "="*60)
    print("⚠️ D组: 部分失败场景测试")
    print("="*60)
    
    # D1: CASE WHEN 中部分分支可能失败
    f = fresh_file()
    r = execute_advanced_sql_query(f, """
        SELECT ID, 
               CASE WHEN ID = 1 THEN 999 
                    WHEN ID = 2 THEN Value * 100
                    ELSE Value 
               END as NewValue 
        FROM Sheet1
    """)
    if r.get('success'):
        tr.ok("D1 CASE WHEN多分支", "CASE WHEN多分支正常执行")
    else:
        tr.fail("D1 CASE WHEN多分支", f"失败: {str(r.get('message',''))[:80]}")
    
    # D2: WHERE IN 中混合有效和无效值
    f2 = fresh_file()
    r = execute_advanced_sql_query(f2, "SELECT * FROM Sheet1 WHERE ID IN (1, 2, 999)")
    if r.get('success') and r.get('data'):
        row_count = len(r['data']) - 1  # 减去header
        if row_count == 2:  # 只有ID=1和2存在
            tr.ok("D2 WHERE IN混合有效无效值", f"正确返回{row_count}行")
        else:
            tr.fail("D2 WHERE IN混合有效无效值", f"返回{row_count}行, 预期2行")
    else:
        tr.fail("D2 WHERE IN混合有效无效值", f"查询失败: {str(r.get('message',''))[:60]}")
    
    # D3: 数学表达式中有NULL/缺失值的行
    f3 = fresh_file()
    # 先把某行的Value设为可能引起问题的值
    r = execute_advanced_update_query(f3, "UPDATE Sheet1 SET Value = NULL WHERE ID = 3")
    r2 = execute_advanced_sql_query(f3, "SELECT ID, Value * 2 AS DoubleValue FROM Sheet1")
    if r2.get('success') and r2.get('data'):
        tr.ok("D3 含NULL行的数学表达式", "表达式计算处理了NULL行")
    else:
        tr.fail("D3 含NULL行的数学表达式", f"失败: {str(r2.get('message',''))[:80]}")
    
    # D4: ORDER BY 对含特殊值的排序
    f4 = fresh_file()
    r = execute_advanced_update_query(f4, "UPDATE Sheet1 SET Price = 0 WHERE ID = 1")
    r = execute_advanced_update_query(f4, "UPDATE Sheet1 SET Price = 9999.99 WHERE ID = 5")
    r2 = execute_advanced_sql_query(f4, "SELECT ID, Price FROM Sheet1 ORDER BY Price ASC")
    if r2.get('success') and r2.get('data'):
        prices = [row[1] for row in r2['data'][1:]]  # skip header
        is_sorted = prices == sorted(prices)
        if is_sorted:
            tr.ok("D4 ORDER BY含极值排序", f"排序正确: {prices}")
        else:
            tr.fail("D4 ORDER BY含极值排序", f"排序异常: {prices}")
    else:
        tr.fail("D4 ORDER BY含极值排序", f"查询失败")
    
    # D5: GROUP BY 对单行分组
    f5 = fresh_file()
    r = execute_advanced_sql_query(f5, "SELECT ID, COUNT(*), SUM(Value) FROM Sheet1 GROUP BY ID")
    if r.get('success') and r.get('data'):
        row_count = len(r['data']) - 1
        if row_count == 5:  # 5个不同的ID
            tr.ok("D5 GROUP BY每行唯一值", f"正确分为{row_count}组")
        else:
            tr.fail("D5 GROUP BY每行唯一值", f"分为{row_count}组, 预期5组")
    else:
        tr.fail("D5 GROUP BY每行唯一值", f"失败: {str(r.get('message',''))[:60]}")


# ==================== E组: 文件状态一致性 ====================
def test_group_e_file_consistency(tr):
    """E组: 错误操作后文件是否仍然可读"""
    print("\n" + "="*60)
    print("📁 E组: 文件状态一致性测试")
    print("="*60)
    
    # E1: 失败的UPDATE后文件可读
    f = fresh_file()
    r = execute_advanced_update_query(f, "UPDATE nonexistent SET x = 1")  # 不存在的表
    r2 = execute_advanced_sql_query(f, "SELECT * FROM Sheet1")
    if r2.get('success') and r2.get('data') and len(r2['data']) == 6:
        tr.ok("E1 失败UPDATE后文件可读", "对不存在表UPDATE后原文件完好")
    else:
        tr.fail("E1 失败UPDATE后文件可读", "文件可能损坏或不可读")
    
    # E2: 多次交替读写后数据一致
    f2 = fresh_file()
    ops_success = True
    for i in range(5):
        r = execute_advanced_update_query(f2, f"UPDATE Sheet1 SET Value = {i*100} WHERE ID = 1")
        if not r.get('success'):
            ops_success = False
        r2 = execute_advanced_sql_query(f2, f"SELECT Value FROM Sheet1 WHERE ID = 1")
        if not (r2.get('success') and r2.get('data')):
            ops_success = False
    
    # 最终验证
    r_final = execute_advanced_sql_query(f2, "SELECT Value FROM Sheet1 WHERE ID = 1")
    if r_final.get('success') and r_final.get('data'):
        final_val = r_final['data'][1][0]
        if final_val == 400:  # 最后一次设为400
            tr.ok("E2 多次交替读写一致性", f"最终值={final_val}, 一致性OK")
        else:
            tr.fail("E2 多次交替读写一致性", f"最终值={final_val}, 预期400")
    else:
        tr.fail("E2 多次交替读写一致性", "最终读取失败")
    
    # E3: pandas直接读取验证一致性
    f3 = fresh_file()
    execute_advanced_update_query(f3, "UPDATE Sheet1 SET Value = 7777 WHERE ID = 2")
    try:
        df = pd.read_excel(f3)
        val_at_id2 = df.loc[df['ID'] == 2, 'Value'].values[0]
        if val_at_id2 == 7777:
            tr.ok("E3 pandas交叉验证一致性", f"pandas读取确认Value={val_at_id2}")
        else:
            tr.fail("E3 pandas交叉验证一致性", f"pandas读取Value={val_at_id2}, 预期7777")
    except Exception as e:
        tr.fail("E3 pandas交叉验证一致性", f"pandas读取失败: {str(e)[:60]}")
    
    # E4: 大数值写入后文件可读
    f4 = fresh_file()
    r = execute_advanced_update_query(f4, "UPDATE Sheet1 SET Value = 999999 WHERE ID = 1")
    try:
        df = pd.read_excel(f4)
        val = df.loc[df['ID'] == 1, 'Value'].values[0]
        tr.ok("E4 大数值写入后pandas可读", f"Value={val}, 文件完整")
    except Exception as e:
        tr.fail("E4 大数值写入后pandas可读", f"读取失败: {str(e)[:80]}")
    
    # E5: 负数写入后文件可读
    f5 = fresh_file()
    r = execute_advanced_update_query(f5, "UPDATE Sheet1 SET Value = -99999 WHERE ID = 3")
    try:
        df = pd.read_excel(f5)
        val = df.loc[df['ID'] == 3, 'Value'].values[0]
        if val == -99999:
            tr.ok("E5 负数写入后pandas可读", f"Value={val}, 一致")
        else:
            tr.fail("E5 负数写入后pandas可读", f"Value={val}, 预期-99999")
    except Exception as e:
        tr.fail("E5 负数写入后pandas可读", f"读取失败: {str(e)[:80]}")


# ==================== F组: SQL语法容错 ====================
def test_group_f_sql_tolerance(tr):
    """F组: 畸形SQL的优雅处理"""
    print("\n" + "="*60)
    print("🔤 F组: SQL语法容错测试")
    print("="*60)
    
    f = TEST_FILE
    create_test_file(f)
    
    tolerance_tests = [
        # (sql, op, should_fail_gracefully, description)
        ("", "query", True, "空SQL"),
        ("   ", "query", True, "纯空格SQL"),
        ("-- 这只是注释", "query", True, "纯注释SQL"),
        ("SELECT", "query", True, "不完整的SELECT"),
        ("SELECT * FROM", "query", True, "无表名的FROM"),
        ("SELECT * FROM Sheet1 WHERE", "query", True, "无条件的WHERE"),
        ("SELECT * FROM Sheet1 ORDER BY", "query", True, "无表达式的ORDER BY"),
        ("SELECT ,,, FROM Sheet1", "query", True, "多余逗号"),
        ("SELECT * FROM Sheet1 LIMIT -1", "query", True, "负数LIMIT"),
        ("SELECT * FROM Sheet1 LIMIT abc", "query", True, "非数字LIMIT"),
    ]
    
    for i, (sql, op, should_fail, desc) in enumerate(tolerance_tests, 1):
        try:
            if op == "query":
                r = execute_advanced_sql_query(f, sql)
            
            if should_fail:
                # 应该优雅地失败（不是crash）
                if not r.get('success'):
                    # 有合理的错误消息
                    msg = str(r.get('message', ''))
                    if len(msg) > 0:
                        tr.ok(f"F{i} 容错: {desc}", f"优雅拒绝: '{msg[:50]}'")
                    else:
                        tr.fail(f"F{i} 容错: {desc}", "失败但无错误消息")
                else:
                    # 意外成功了
                    tr.fail(f"F{i} 容错: {desc}", f"意外成功! data={str(r.get('data',''))[:40]}")
            else:
                if r.get('success'):
                    tr.ok(f"F{i} 容错: {desc}", "正确执行")
                else:
                    tr.fail(f"F{i} 容错: {desc}", f"意外失败: {str(r.get('message',''))[:50]}")
        except Exception as e:
            tr.fail(f"F{i} 容错: {desc}", f"抛出异常! {type(e).__name__}: {str(e)[:50]}")
    
    # F12: 多余括号
    r = execute_advanced_sql_query(f, "SELECT ((ID + 1)) FROM Sheet1")
    if r.get('success'):
        tr.ok("F12 容错: 多余嵌套括号", "正确解析嵌套括号")
    else:
        tr.fail("F12 容错: 多余嵌套括号", f"失败: {str(r.get('message',''))[:50]}")
    
    # F13: 关键字大小写混合
    r = execute_advanced_sql_query(f, "select id, name from Sheet1 where value > 20 order by id desc limit 3")
    if r.get('success') and r.get('data'):
        tr.ok("F13 容错: 小写关键字", "小写关键字可执行")
    else:
        tr.fail("F13 容错: 小写关键字", f"失败: {str(r.get('message',''))[:50]}")


# ==================== G组: 极端输入容错 ====================
def test_group_g_extreme_input(tr):
    """G组: 极端/边界输入的处理"""
    print("\n" + "="*60)
    print("🎯 G组: 极端输入容错测试")
    print("="*60)
    
    # G1: 超长字符串作为WHERE值
    f = fresh_file()
    long_str = "A" * 5000
    r = execute_advanced_update_query(f, f"UPDATE Sheet1 SET Name = '{long_str}' WHERE ID = 1")
    if r.get('success'):
        r2 = execute_advanced_sql_query(f, "SELECT Name FROM Sheet1 WHERE ID = 1")
        if r2.get('success') and r2.get('data'):
            stored_name = str(r2['data'][1][0])
            if len(stored_name) == 5000:
                tr.ok("G1 超长字符串(5000字符)", f"完整存储并读取, 长度={len(stored_name)}")
            else:
                tr.fail("G1 超长字符串(5000字符)", f"存储长度={len(stored_name)}, 预期5000")
        else:
            tr.fail("G1 超长字符串(5000字符)", "写入成功但读取失败")
    else:
        tr.fail("G1 超长字符串(5000字符)", f"写入失败: {str(r.get('message',''))[:60]}")
    
    # G2: 特殊字符在字符串中
    f2 = fresh_file()
    special_names = [
        "It's a test",          # 单引号
        'Test "quoted" value',  # 双引号
        "Test\nwith\nnewlines", # 换行符
        "Tab\there",           # Tab
        "Path=C:\\Users\\x",   # 反斜杠
    ]
    all_ok = True
    for j, name in enumerate(special_names):
        safe_name = name.replace("'", "''")  # SQL转义单引号
        r = execute_advanced_update_query(f2, f"UPDATE Sheet1 SET Name = '{safe_name}' WHERE ID = {j+1}")
        if not r.get('success'):
            all_ok = False
            break
    
    if all_ok:
        tr.ok("G2 特殊字符字符串", "5种特殊字符全部写入成功")
    else:
        tr.fail("G2 特殊字符字符串", f"第{j+1}种字符失败: {name[:20]}")
    
    # G3: Unicode极端字符
    f3 = fresh_file()
    unicode_texts = [
        "中文测试名称",
        "日本語テスト",
        "한국어테스트",
        "العربية اختبار",
        "🎮🔥💎Emoji Test",
        "ខ្មែរខ្មែរ",  # 高棉文
        "אברית טסט",  # 希伯来文
    ]
    uni_ok = True
    for k, text in enumerate(unicode_texts):
        r = execute_advanced_update_query(f3, f"UPDATE Sheet1 SET Name = '{text}' WHERE ID = {(k % 5) + 1}")
        if not r.get('success'):
            uni_ok = False
            break
    
    if uni_ok:
        # 验证回读
        r2 = execute_advanced_sql_query(f3, "SELECT Name FROM Sheet1")
        if r2.get('success'):
            tr.ok("G3 Unicode极端字符(7种语言+Emoji)", "全部写入并可读")
        else:
            tr.fail("G3 Unicode极端字符(7种语言+Emoji)", "写入成功但读取失败")
    else:
        tr.fail("G3 Unicode极端字符(7种语言+Emoji)", f"第{k+1}种文字失败: {text[:20]}")
    
    # G4: 数值精度边界
    f4 = fresh_file()
    precision_tests = [
        (0.1 + 0.2, "浮点精度0.1+0.2"),
        (1.0 / 3.0, "1/3循环小数"),
        (0.000001, "极小正数"),
        (123456789.123456789, "高精度大数"),
        (-0.000001, "极小负数"),
    ]
    prec_ok = True
    for m, (val, desc) in enumerate(precision_tests):
        r = execute_advanced_update_query(f4, f"UPDATE Sheet1 SET Price = {val} WHERE ID = {(m % 5) + 1}")
        if not r.get('success'):
            prec_ok = False
            break
    
    if prec_ok:
        tr.ok("G4 数值精度边界(5种)", "各种精度数值写入成功")
    else:
        tr.fail("G4 数值精度边界(5种)", f"第{m+1}种失败: {desc}")
    
    # G5: 空字符串 vs NULL
    f5 = fresh_file()
    r = execute_advanced_update_query(f5, "UPDATE Sheet1 SET Name = '' WHERE ID = 1")  # 空字符串
    r2 = execute_advanced_update_query(f5, "UPDATE Sheet1 SET Name = NULL WHERE ID = 2")  # NULL
    if r.get('success') and r2.get('success'):
        r3 = execute_advanced_sql_query(f5, "SELECT Name FROM Sheet1 WHERE ID IN (1, 2) ORDER BY ID")
        if r3.get('success') and r3.get('data'):
            name1 = r3['data'][1][0] if len(r3['data']) > 1 else None
            name2 = r3['data'][2][0] if len(r3['data']) > 2 else None
            tr.ok("G5 空字符串vs NULL", f"空串='{name1}', NULL={name2} (均接受)")
        else:
            tr.fail("G5 空字符串vs NULL", "读取失败")
    else:
        result1 = "OK" if r.get('success') else f"FAIL:{str(r.get('message',''))[:30]}"
        result2 = "OK" if r2.get('success') else f"FAIL:{str(r2.get('message',''))[:30]}"
        tr.fail("G5 空字符串vs NULL", f"空串={result1}, NULL={result2}")


# ==================== H组: 已知P1/P2回归验证 ====================
def test_group_h_known_issues_regression(tr):
    """H组: 已知P1/P2问题回归检查"""
    print("\n" + "="*60)
    print("📋 H组: 已知P1/P2回归验证")
    print("="*60)
    
    # P2-1: UPDATE SET || 字符串拼接
    f = fresh_file()
    r = execute_advanced_update_query(f, "UPDATE Sheet1 SET Name = 'Prefix-' || Name WHERE ID = 1")
    if not r.get('success'):
        tr.ok("H1 [P2-1] UPDATE SET ||拼接", "确认不支持 (已知限制)")
    else:
        tr.fail("H1 [P2-1] UPDATE SET ||拼接", "现在支持了? 或静默忽略?")
    
    # P2-2: CASE WHEN 作为算术操作数
    f2 = fresh_file()
    r = execute_advanced_sql_query(f2, "SELECT Name, Value * CASE WHEN ID = 1 THEN 2 ELSE 1 END AS DoubleV FROM Sheet1")
    if not r.get('success'):
        tr.ok("H2 [P2-2] CASE算术操作数", "确认不支持 (已知限制)")
    else:
        tr.ok("H2 [P2-2] CASE算术操作数", "现在支持了! ✅ FIXED?")
    
    # P2-4: 极端浮点值(R31发现)
    f3 = fresh_file()
    extreme_floats = [
        ("float('inf')", "Infinity"),
        ("float('-inf')", "-Infinity"),
        ("float('nan')", "NaN"),
    ]
    for expr, label in extreme_floats:
        # 通过CAST尝试写入
        r = execute_advanced_update_query(f3, f"UPDATE Sheet1 SET Value = CAST('{expr.split('(')[1].split(')')[0]}' AS FLOAT64) WHERE ID = 1")
        # 不管成功失败，检查文件是否还可读
        try:
            df = pd.read_excel(f3)
            _ = df.shape  # 触发完整读取
            tr.ok(f"H3 [P2-4] {label}后文件可读", f"{label}写入后文件仍可读")
        except Exception as e:
            err_str = str(e)[:60]
            if 'infinity' in err_str.lower() or 'float' in err_str.lower() or 'nan' in err_str.lower():
                tr.fail(f"H3 [P2-4] {label}后文件可读", f"🔄 R31确认仍存在: {err_str}")
            else:
                tr.fail(f"H3 [P2-4] {label}后文件可读", f"其他错误: {err_str}")


# ==================== 主函数 ====================
def main():
    print("=" * 70)
    print("🔄 Round 32 MCP 接口实测")
    print("主题: 错误注入与容错恢复测试 + P0第5轮回归")
    print(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"测试目录: {TEST_DIR}")
    print("=" * 70)
    
    tr = TestResult()
    
    try:
        test_group_a_p0_regression(tr)
        test_group_b_error_recovery(tr)
        test_group_c_error_quality(tr)
        test_group_d_partial_failure(tr)
        test_group_e_file_consistency(tr)
        test_group_f_sql_tolerance(tr)
        test_group_g_extreme_input(tr)
        test_group_h_known_issues_regression(tr)
    except Exception as e:
        print(f"\n🚨 测试过程发生未捕获异常: {e}")
        traceback.print_exc()
    
    # 输出汇总
    total = tr.passed + tr.failed
    print("\n" + "=" * 70)
    print("📊 Round 32 MCP 测试汇总")
    print("=" * 70)
    print(f"  总计: {total} 个场景")
    print(f"  通过: ✅ {tr.passed} ({tr.passed/total*100:.1f}%)" if total > 0 else "  通过: N/A")
    print(f"  失败: ❌ {tr.failed} ({tr.failed/total*100:.1f}%)" if total > 0 else "  失败: N/A")
    print()
    
    # 失败详情
    if tr.failed > 0:
        print("❌ 失败场景列表:")
        for name, passed, detail in tr.results:
            if not passed:
                print(f"  - {name}: {detail}")
        print()
    
    # 清理
    try:
        shutil.rmtree(TEST_DIR)
    except:
        pass
    
    return tr


if __name__ == "__main__":
    result = main()
    sys.exit(0 if result.failed == 0 else 1)
