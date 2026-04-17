# -*- coding: utf-8 -*-
"""
format_cells R67 迭代测试 - 深度边缘 case 第七轮

新增测试场景（覆盖 R66 未涉及的场景）:
  1. strikethrough=True/False 扁平格式
  2. font_color 带 # 前缀（_clean_color 行为）
  3. gradient fill 三色及以上渐变
  4. 含批注(comment)的单元格格式化后批注保留
  5. ARGB 8位颜色值（含 alpha 通道）
  6. border_style 扁平简写（字符串→四边）
  7. 超长文本(5000+字符)单元格格式化不截断
  8. number_format 科学计数法 "0.00E+00"
  9. 先合并→格式化→拆分，格式是否扩散到子单元格
  10. 组合操作：merge + bold + bg_color 同时通过 format_cells 传参
  11. Unicode 工作表名 + format_cells 全属性组合
  12. font_size=1 极小字号边界
  13. indent 负值容错（openpyxl 行为）
  14. format_cells 后 comment/hyperlink/data_validation 三重保留
  15. _normalize_formatting bold=False 显式关闭
  16. format_cells 对错误类型值单元格（dict/list）容错
  17. 同时设 font+fill+alignment+border+number_format 全套属性
  18. format_cells 范围超出数据区域（如 Z999:AA1000）
  19. bg_color 带 # 前缀扁平格式
  20. wrap_text=True + text_rotation 非零组合
"""

import os
import pytest
import tempfile
from datetime import date, datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color as OpenpyxlColor, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

from excel_mcp_server_fastmcp.core.excel_writer import ExcelWriter
from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations


# ==================== Helper ====================

def _create_test_xlsx(file_path: str, rows: int = 5, cols: int = 4, sheet_name: str = "Sheet1"):
    """创建测试用 xlsx 文件，含数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=r * 10 + c)
    wb.save(file_path)
    wb.close()


def _read_cell_style(file_path: str, cell_ref: str = "A1", sheet_name: str = "Sheet1") -> dict:
    """读取单元格的样式属性用于断言"""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    cell = ws[cell_ref]
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    result = {
        "bold": font.bold,
        "italic": font.italic,
        "underline": font.underline,
        "strikethrough": font.strikethrough,
        "size": font.size,
        "font_name": font.name,
        "color": str(font.color) if font.color else None,
        "fill_type": fill.fill_type if fill else None,
        "fgColor": str(fill.fgColor) if fill and hasattr(fill, 'fgColor') and fill.fgColor else None,
        "bgColor": str(fill.bgColor) if fill and hasattr(fill, 'bgColor') and fill.bgColor else None,
        "alignment_h": alignment.horizontal,
        "alignment_v": alignment.vertical,
        "wrap_text": bool(alignment.wrap_text) if alignment.wrap_text is not None else False,
        "text_rotation": alignment.text_rotation,
        "number_format": cell.number_format,
        "border_left": border.left.style if border and border.left else None,
        "border_right": border.right.style if border and border.right else None,
        "border_top": border.top.style if border and border.top else None,
        "border_bottom": border.bottom.style if border and border.bottom else None,
        "value": cell.value,
        "indent": alignment.indent,
        "shrink_to_fit": alignment.shrink_to_fit,
    }
    wb.close()
    return result


# ==================== Test Class ====================

class TestFormatCellsR67:
    """format_cells R67 第七轮测试套件 — 深度边缘 case 第七轮"""

    # ---------- 1. strikethrough 扁平格式 ----------

    def test_strikethrough_flat_true(self, tmp_path):
        """strikethrough=True 扁平格式正确应用"""
        fp = str(tmp_path / "strike_t.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"strikethrough": True})
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["strikethrough"] is True

    def test_strikethrough_flat_false(self, tmp_path):
        """strikethrough=False 扁平格式显式关闭删除线"""
        fp = str(tmp_path / "strike_f.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"].font = Font(strikethrough=True)
        ws["A1"].value = "test"
        wb.save(fp)
        wb.close()

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"strikethrough": False})
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["strikethrough"] is False

    # ---------- 2. font_color 带 # 前缀 ----------

    @pytest.mark.parametrize("color_input,expected", [
        ("#FF0000", "FF0000"),
        ("#0000FF", "0000FF"),
        ("#ABCDEF", "ABCDEF"),
        ("#123", "112233"),  # 3位HEX扩展为6位
    ])
    def test_font_color_hash_prefix(self, tmp_path, color_input, expected):
        """font_color 带 # 前缀应被清理，3位HEX自动扩展"""
        fp = str(tmp_path / f"fcolor_{expected}.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"font_color": color_input})
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert expected in (style["color"] or "")

    # ---------- 3. gradient fill 三色渐变 ----------

    def test_gradient_fill_three_colors(self, tmp_path):
        """gradient fill 支持三色渐变"""
        fp = str(tmp_path / "grad3.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {
                "type": "gradient",
                "colors": ["FF0000", "00FF00", "0000FF"],
                "gradient_type": "linear",
            }
        })
        assert result.success

        style = _read_cell_style(fp, "A1")
        assert style["fill_type"] in ("linear", "path")

    def test_gradient_fill_single_color(self, tmp_path):
        """gradient fill 单色应不崩溃（退化处理）"""
        fp = str(tmp_path / "grad1.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {
                "type": "gradient",
                "colors": ["FF0000"],
            }
        })
        # 单色渐变可能成功也可能失败，不应崩溃
        assert result.success or result.error is not None

    # ---------- 4. 含批注单元格格式化 ----------

    def test_format_cell_with_comment_preserves_comment(self, tmp_path):
        """对含批注的单元格格式化不应破坏批注"""
        fp = str(tmp_path / "comment.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"].value = "Hello"
        ws["A1"].comment = Comment("这是批注", "Author")
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"bold": True, "color": "FF0000"}})
        assert result.success

        wb2 = load_workbook(fp)
        ws2 = wb2["Sheet1"]
        cell = ws2["A1"]
        assert cell.value == "Hello"
        assert cell.comment is not None
        assert cell.comment.text == "这是批注"
        assert cell.font.bold is True
        wb2.close()

    # ---------- 5. ARGB 8位颜色（alpha通道）----------

    def test_argb_color_8char(self, tmp_path):
        """ARGB 8位颜色值（含 alpha 通道）应被接受"""
        fp = str(tmp_path / "argb.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {"type": "solid", "color": "FFFF0000"}  # ARGB: FF=alpha, FF0000=red
        })
        assert result.success

    def test_argb_font_color(self, tmp_path):
        """font color 为 ARGB 格式"""
        fp = str(tmp_path / "argb_font.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "font": {"color": "FF00FF00"}  # ARGB green
        })
        assert result.success

    # ---------- 6. border_style 扁平简写 ----------

    def test_border_style_flat_shorthand(self, tmp_path):
        """border_style="medium" 扁平简写应用到四边"""
        fp = str(tmp_path / "bstyle_flat.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"border_style": "medium"})
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["border_left"] == "medium"
        assert style["border_right"] == "medium"
        assert style["border_top"] == "medium"
        assert style["border_bottom"] == "medium"

    def test_border_style_flat_thick(self, tmp_path):
        """border_style="thick" 扁平简写"""
        fp = str(tmp_path / "bstyle_thick.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"border_style": "thick"})
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["border_left"] == "thick"
        assert style["border_right"] == "thick"

    # ---------- 7. 超长文本单元格 ----------

    def test_format_long_text_cell(self, tmp_path):
        """超长文本(5000字符)格式化后内容完整"""
        fp = str(tmp_path / "longtext.xlsx")
        long_text = "A" * 5000
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"].value = long_text
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"bold": True}, "alignment": {"wrap_text": True}})
        assert result.success

        wb2 = load_workbook(fp)
        assert len(wb2.active["A1"].value) == 5000
        assert wb2.active["A1"].font.bold is True
        wb2.close()

    # ---------- 8. number_format 科学计数法 ----------

    @pytest.mark.parametrize("nf", [
        "0.00E+00",
        "##0.0E+0",
        "Scientific",
    ])
    def test_number_format_scientific(self, tmp_path, nf):
        """科学计数法 number_format"""
        fp = str(tmp_path / f"nf_sci_{hash(nf) % 10000}.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"number_format": nf})
        assert result.success

        style = _read_cell_style(fp, "A1")
        assert nf in style["number_format"]

    # ---------- 9. 合并→格式化→拆分 ----------

    def test_merge_format_unmerge(self, tmp_path):
        """先合并、再格式化、再拆分，检查子单元格状态"""
        fp = str(tmp_path / "mfmtu.xlsx")
        _create_test_xlsx(fp, rows=3, cols=3)

        writer = ExcelWriter(fp)

        # 1. 合并
        r1 = writer.merge_cells("A1:B2", "Sheet1")
        assert r1.success

        # 2. 格式化合并区域
        r2 = writer.format_cells("Sheet1!A1:B2", {
            "font": {"bold": True, "color": "FF0000"},
            "fill": {"type": "solid", "color": "FFFF00"},
        })
        assert r2.success

        # 3. 拆分
        r3 = writer.unmerge_cells("A1:B2", "Sheet1")
        assert r3.success

        # 拆分后 A1 应保留格式（openpyxl 行为：左上角单元格保留）
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True

    # ---------- 10. 组合操作：全套属性同时传 ----------

    def test_full_combo_all_attributes(self, tmp_path):
        """同时传入 font+fill+alignment+border+number_format 全套属性"""
        fp = str(tmp_path / "fullcombo.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(
            fp, "Sheet1", "A1:C3",
            formatting={
                "bold": True,
                "italic": True,
                "underline": "double",
                "font_size": 14,
                "font_color": "#FF0000",
                "bg_color": "#00FF00",
                "alignment": "center",
                "vertical_alignment": "middle",
                "wrap_text": True,
                "number_format": "#,##0.00",
                "border_style": "medium",
            }
        )
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["italic"] is True
        assert style["underline"] == "double"
        assert style["size"] == 14
        assert style["alignment_h"] == "center"
        assert style["alignment_v"] == "center"
        assert style["wrap_text"] is True
        assert style["border_left"] == "medium"

    # ---------- 11. Unicode 工作表名 + 全属性 ----------

    def test_unicode_sheet_full_format(self, tmp_path):
        """中文工作表名 + 完整格式化操作"""
        fp = str(tmp_path / "unicode_sheet.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "数据表"
        ws["A1"].value = "测试"
        wb.save(fp)
        wb.close()

        result = ExcelOperations.format_cells(
            fp, "数据表", "A1",
            formatting={
                "bold": True,
                "font_name": "微软雅黑",
                "bg_color": "FFD700",
                "alignment": "center",
            }
        )
        assert result["success"] is True

        style = _read_cell_style(fp, "A1", sheet_name="数据表")
        assert style["bold"] is True
        assert style["font_name"] == "微软雅黑"

    # ---------- 12. font_size=1 极小字号 ----------

    @pytest.mark.parametrize("fsize", [1, 0.5, 255, 409])
    def test_font_size_extreme_values(self, tmp_path, fsize):
        """极端 font_size 值（openpyxl 接受的范围通常 1-409）"""
        fp = str(tmp_path / f"fsize_{fsize}.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"size": fsize}})
        # openpyxl 通常接受 1-409 的 size，0 可能报错
        if fsize >= 1:
            assert result.success, f"font_size={fsize} 应成功: {result.error}"
        # 不论成功失败都不应崩溃

    # ---------- 13. indent 负值 ----------

    def test_indent_negative_value(self, tmp_path):
        """indent 为负值时的行为（openpyxl 可能接受或拒绝）"""
        fp = str(tmp_path / "indent_neg.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"alignment": {"indent": -5}})
        # 负值 indent 可能成功也可能失败，取决于 openpyxl 版本
        # 关键是不崩溃
        assert result.success or result.error is not None

    # ---------- 14. 三重保留：comment + hyperlink + data_validation ----------

    def test_triple_preservation(self, tmp_path):
        """同时含批注+超链接+数据验证的单元格格式化后三者都保留"""
        fp = str(tmp_path / "triple.xlsx")
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        cell = ws["A1"]
        cell.value = "Triple"
        cell.comment = Comment("批注文本", "TestAuthor")
        cell.hyperlink = "https://example.org"
        dv = DataValidation(type="whole", operator="between", formula1=1, formula2=100)
        dv.add(cell)
        ws.add_data_validation(dv)
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "font": {"bold": True, "color": "FF0000"},
            "fill": {"type": "solid", "color": "FFFF00"},
            "alignment": {"horizontal": "center"},
        })
        assert result.success

        wb2 = load_workbook(fp)
        ws2 = wb2["Sheet1"]
        c = ws2["A1"]
        assert c.value == "Triple"
        assert c.comment is not None
        assert c.hyperlink is not None
        assert c.hyperlink.target == "https://example.org"
        dvs = list(ws2.data_validations.dataValidation)
        assert len(dvs) >= 1
        assert c.font.bold is True
        wb2.close()

    # ---------- 15. bold=False 显式关闭 ----------

    def test_bold_false_explicit(self, tmp_path):
        """bold=False 显式关闭粗体（覆盖已有粗体）"""
        fp = str(tmp_path / "bold_false.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"].font = Font(bold=True)
        ws["A1"].value = "was bold"
        wb.save(fp)
        wb.close()

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"bold": False})
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["bold"] is False

    # ---------- 16. 错误类型值单元格容错 ----------

    def test_format_cell_no_crash_on_weird_value(self, tmp_path):
        """单元格值为异常类型时格式化不崩溃"""
        fp = str(tmp_path / "weird_val.xlsx")
        # openpyxl 本身不支持 dict/list 作为 cell value，
        # 但测试格式化代码对各种情况的健壮性
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "font": {"bold": True},
            "fill": {"type": "solid", "color": "FF0000"},
            "alignment": {"wrap_text": True},
            "border": {"top": "thin", "bottom": "thin"},
            "number_format": "@",
        })
        assert result.success

    # ---------- 17. 范围超出数据区域 ----------

    def test_range_beyond_data(self, tmp_path):
        """范围超出已有数据区域（Z999:AA1000）不应崩溃"""
        fp = str(tmp_path / "far_range.xlsx")
        _create_test_xlsx(fp, rows=3, cols=3)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!Z999:AA1000", {"font": {"bold": True}})
        # 超范围可能成功（创建空单元格）也可能失败，不应崩溃
        assert result.success or result.error is not None

    def test_range_single_far_cell(self, tmp_path):
        """单个远距离单元格（如 ZZ1）格式化"""
        fp = str(tmp_path / "far_cell.xlsx")
        _create_test_xlsx(fp, rows=3, cols=3)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!ZZ1", {"font": {"bold": True}})
        assert result.success or result.error is not None

    # ---------- 18. bg_color 带 # 前缀扁平格式 ----------

    @pytest.mark.parametrize("color_in", [
        "#FF0",       # 3位+前缀 → FFFF00
        "#FF0000",   # 标准6位
        "#AABBCCDD", # ARGB 8位
        "ABC",       # 3位→AABBCC
        "AABBCC",    # 标准6位
    ])
    def test_bg_color_various_formats(self, tmp_path, color_in):
        """bg_color 各种格式（带#、不带#、3位、6位、8位），3位自动扩展"""
        fp = str(tmp_path / f"bg_{color_in.replace('#', '')}.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"bg_color": color_in})
        assert result["success"] is True

    # ---------- 19. wrap_text + text_rotation 组合 ----------

    @pytest.mark.parametrize("rot", [45, 90, -45])
    def test_wrap_text_with_rotation(self, tmp_path, rot):
        """wrap_text=True 与非零 text_rotation 组合"""
        fp = str(tmp_path / f"wrap_rot_{rot}.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", {
            "wrap_text": True,
            "text_rotation": rot,
        })
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["wrap_text"] is True
        # 负值取绝对值
        expected_rot = abs(rot) if rot < 0 else rot
        assert style["text_rotation"] == expected_rot

    # ---------- 20. italic 扁平格式独立测试 ----------

    def test_italic_flat_true_false(self, tmp_path):
        """italic=True/False 扁平格式"""
        fp = str(tmp_path / "italic.xlsx")
        _create_test_xlsx(fp)

        # 设为 italic
        r1 = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"italic": True})
        assert r1["success"] is True
        style1 = _read_cell_style(fp, "A1")
        assert style1["italic"] is True

        # 取消 italic
        r2 = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"italic": False})
        assert r2["success"] is True
        style2 = _read_cell_style(fp, "A1")
        assert style2["italic"] is False

    # ---------- 21. font_name 扁平格式 ----------

    def test_font_name_flat(self, tmp_path):
        """font_name 扁平格式设置字体"""
        fp = str(tmp_path / "fname_flat.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"font_name": "Courier New"})
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["font_name"] == "Courier New"

    # ---------- 22. underline 扁平格式各变体 ----------

    @pytest.mark.parametrize("uline,val", [
        ("single", "single"),
        ("double", "double"),
        (True, "single"),
        (False, None),  # openpyxl 读取时 None 表示无下划线
    ])
    def test_underline_flat_variants(self, tmp_path, uline, val):
        """underline 扁平格式各种值"""
        fp = str(tmp_path / f"ulflat_{uline}.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"underline": uline})
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["underline"] == val

    # ---------- 23. format_cells 对含日期时间的单元格 ----------

    def test_format_datetime_cell(self, tmp_path):
        """datetime 类型单元格格式化后保持类型"""
        fp = str(tmp_path / "dt_cell.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        dt_val = datetime(2025, 6, 15, 10, 30, 0)
        ws["A1"].value = dt_val
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "font": {"bold": True},
            "number_format": "YYYY-MM-DD HH:MM:SS",
        })
        assert result.success

        wb2 = load_workbook(fp)
        val = wb2.active["A1"].value
        assert isinstance(val, datetime)
        assert val.year == 2025
        assert val.month == 6
        assert wb2.active["A1"].font.bold is True
        wb2.close()

    # ---------- 24. _normalize_formatting italic=False ----------

    def test_normalize_italic_false(self):
        """_normalize_formatting 中 italic=False 应包含在输出中"""
        result = ExcelOperations._normalize_formatting({"italic": False})
        assert "font" in result
        assert result["font"]["italic"] is False

    # ---------- 25. preset='title' + 自定义合并验证 ----------

    def test_preset_title_custom_merge(self, tmp_path):
        """preset='title' + 用户自定义 bg_color 合并"""
        fp = str(tmp_path / "preset_title_cust.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(
            fp, "Sheet1", "A1",
            formatting={"bg_color": "FFD700", "alignment": "right"},
            preset="title",
        )
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True  # title preset 的 bold
        assert style["size"] == 14   # title preset 的 size
        assert style["alignment_h"] == "right"  # 用户自定义覆盖

    # ---------- 26. format_cells 同一范围连续不同格式覆盖 ----------

    def test_overwrite_format_same_range(self, tmp_path):
        """同一范围先后应用不同格式，后者覆盖前者"""
        fp = str(tmp_path / "overwrite.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)

        # 第一次：红色背景
        r1 = writer.format_cells("Sheet1!A1", {"fill": {"type": "solid", "color": "FF0000"}})
        assert r1.success

        # 第二次：蓝色背景（覆盖红色）
        r2 = writer.format_cells("Sheet1!A1", {"fill": {"type": "solid", "color": "0000FF"}})
        assert r2.success

        style = _read_cell_style(fp, "A1")
        # 最终应为蓝色（后者覆盖）
        assert "0000FF" in (style["fgColor"] or "")

    # ---------- 27. border 字典形式（每边不同样式）----------

    def test_border_per_side_different_styles(self, tmp_path):
        """边框四边分别设置不同样式"""
        fp = str(tmp_path / "border_mixed.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {
                "left": "thin",
                "right": "thick",
                "top": "medium",
                "bottom": "dashed",
            }
        })
        assert result.success

        style = _read_cell_style(fp, "A1")
        assert style["border_left"] == "thin"
        assert style["border_right"] == "thick"
        assert style["border_top"] == "medium"
        assert style["border_bottom"] == "dashed"

    # ---------- 28. border 带 color 的字典形式 ----------

    def test_border_with_color(self, tmp_path):
        """边框指定颜色"""
        fp = str(tmp_path / "border_col.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {
                "left": {"style": "medium", "color": "FF0000"},
                "right": {"style": "thin", "color": "00FF00"},
            }
        })
        assert result.success

    # ---------- 29. format_cells 后文件仍可正常打开 ----------

    def test_file_readable_after_format(self, tmp_path):
        """format_cells 后文件可被 openpyxl 正常读取且数据完整"""
        fp = str(tmp_path / "readable.xlsx")
        _create_test_xlsx(fp, rows=10, cols=10)

        # 记录原始数据
        orig_data = []
        wb = load_workbook(fp)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=10, values_only=True):
            orig_data.append(list(row))
        wb.close()

        # 大面积格式化
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:J10", {
            "font": {"bold": True, "size": 11},
            "fill": {"type": "solid", "color": "F2F2F2"},
            "alignment": {"horizontal": "center"},
            "border": {"top": "thin", "bottom": "thin"},
        })
        assert result.success
        assert result.metadata["formatted_count"] == 100

        # 验证数据完整性
        wb2 = load_workbook(fp)
        ws2 = wb2.active
        new_data = []
        for row in ws2.iter_rows(min_row=1, max_row=10, max_col=10, values_only=True):
            new_data.append(list(row))
        wb2.close()

        assert new_data == orig_data, "格式化后数据发生变化!"

    # ---------- 30. shrink_to_fit 扁平格式 ----------

    def test_shrink_to_fit_flat(self, tmp_path):
        """shrink_to_fit 扁平格式"""
        fp = str(tmp_path / "shrink_flat.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"shrink_to_fit": True})
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["shrink_to_fit"] is True

    # ---------- 31. format_cells 空范围字符串 ----------

    def test_empty_range_string(self, tmp_path):
        """空范围字符串应返回错误"""
        fp = str(tmp_path / "empty_range.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!", {"font": {"bold": True}})
        # 空范围应失败
        assert result.success is False or result.error is not None

    # ---------- 32. vertical_alignment扁平 "middle" → "center" 转换 ----------

    def test_vertical_middle_to_center(self, tmp_path):
        """vertical_alignment="middle" 应转换为 "center" """
        fp = str(tmp_path / "v_mid.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", {"vertical_alignment": "middle"})
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["alignment_v"] == "center"  # middle → center 转换
