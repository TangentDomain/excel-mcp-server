"""
Round 31 D组问题深度诊断 - 极端值写入导致文件损坏?
"""
import sys, os, tempfile, shutil
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
)
from openpyxl import Workbook

TEST_DIR = tempfile.mkdtemp(prefix='r31_diag_')

def test_extreme_value_corruption():
    """测试各种极端值是否导致文件损坏"""
    
    test_cases = [
        ('neg_inf', "UPDATE Sheet1 SET Value = -1.7976931348623157e+308 WHERE ID = 1"),
        ('big_int', "UPDATE Sheet1 SET Value = 1000000000000000000 WHERE ID = 1"),
        ('nan_val', "UPDATE Sheet1 SET Value = CAST('nan' AS FLOAT64) WHERE ID = 1"),
        ('inf_val', "UPDATE Sheet1 SET Value = CAST('inf' AS FLOAT64) WHERE ID = 1"),
        ('neg_zero', "UPDATE Sheet1 SET Value = -0.0 WHERE ID = 1"),
        ('precision', "UPDATE Sheet1 SET Value = 0.1 + 0.2 WHERE ID = 1"),
        ('bool_true', "UPDATE Sheet1 SET Value = TRUE WHERE ID = 1"),
    ]
    
    print("=== 极端值写入文件损坏诊断 ===\n")
    
    for name, sql in test_cases:
        # 每次创建新文件
        path = os.path.join(TEST_DIR, f'{name}.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Name', 'Value'])
        ws.append([1, 'test', 100.5])
        wb.save(path)
        
        # 写入极端值
        r1 = execute_advanced_update_query(path, sql)
        
        # 尝试读取
        r2 = execute_advanced_sql_query(path, "SELECT Value FROM Sheet1 WHERE ID = 1")
        
        # 直接用openpyxl尝试打开
        try:
            wb2 = Workbook()
            wb2.load_workbook(path)
            openpyxl_ok = True
            openpyxl_err = ''
        except Exception as e:
            openpyxl_ok = False
            openpyxl_err = str(e)[:80]
        
        write_ok = r1.get('success', False)
        read_ok = r2.get('success', False)
        read_msg = str(r2.get('message', ''))[:60]
        read_data = str(r2.get('data', ''))[:60]
        
        status = '✅' if (read_ok or openpyxl_ok) else '🔴 文件损坏!'
        print(f"{status} [{name}] write={write_ok} read={read_ok} openpyxl={openpyxl_ok}")
        if not read_ok:
            print(f"   read_error: {read_msg}")
            print(f"   read_data: {read_data}")
        if not openpyxl_ok:
            print(f"   openpyxl_error: {openpyxl_err}")
        if write_ok and not read_ok:
            print(f"   🚨 写入成功但读取失败 → 文件可能已损坏!")
        print()
    
    # 清理
    shutil.rmtree(TEST_DIR)

if __name__ == '__main__':
    test_extreme_value_corruption()
