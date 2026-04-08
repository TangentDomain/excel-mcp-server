# -*- coding: utf-8 -*-
"""
核心Excel操作功能测试
合并了ExcelReader, ExcelWriter, ExcelManager的测试
这个文件替代了原本分散在多个文件中的核心功能测试
"""

import pytest
import tempfile
from pathlib import Path

from src.excel_mcp_server_fastmcp.core.excel_reader import ExcelReader
from src.excel_mcp_server_fastmcp.core.excel_writer import ExcelWriter
from src.excel_mcp_server_fastmcp.core.excel_manager import ExcelManager
from src.excel_mcp_server_fastmcp.models.types import OperationResult, SheetInfo, CellInfo, ModifiedCell
from src.excel_mcp_server_fastmcp.utils.exceptions import ExcelFileNotFoundError, SheetNotFoundError
from src.excel_mcp_server_fastmcp.utils.formula_cache import get_formula_cache


class TestExcelCore:
    """核心Excel操作功能的综合测试"""

    # ==================== Excel Reader 测试 ====================

    def test_reader_init_valid_file(self, sample_excel_file):
        """Test ExcelReader initialization with valid file"""
        reader = ExcelReader(sample_excel_file)
        assert reader.file_path == sample_excel_file

    def test_reader_init_invalid_file(self):
        """Test ExcelReader initialization with invalid file"""
        with pytest.raises(FileNotFoundError):
            ExcelReader("nonexistent_file.xlsx")

    def test_reader_list_sheets(self, sample_excel_file):
        """Test listing sheets"""
        reader = ExcelReader(sample_excel_file)
        result = reader.list_sheets()

        assert isinstance(result, OperationResult)
        assert result.success is True
        assert isinstance(result.data, list)
        assert len(result.data) >= 1

        # Check first sheet
        sheet1 = result.data[0]
        assert isinstance(sheet1, SheetInfo)
        assert hasattr(sheet1, 'name')
        assert hasattr(sheet1, 'index')

    def test_reader_get_range_cell_range(self, sample_excel_file):
        """Test getting a range of cells"""
        reader = ExcelReader(sample_excel_file)
        result = reader.get_range("Sheet1!A1:C5")

        assert isinstance(result, OperationResult)
        assert result.success is True
        assert isinstance(result.data, list)
        assert len(result.data) == 5
        assert len(result.data[0]) == 3

    def test_reader_get_range_single_cell(self, sample_excel_file):
        """Test getting a single cell"""
        reader = ExcelReader(sample_excel_file)
        result = reader.get_range("Sheet1!A1")

        assert result.success is True
        assert isinstance(result.data, list)
        assert len(result.data) == 1
        assert len(result.data[0]) == 1

    def test_reader_get_range_with_sheet_name(self, sample_excel_file):
        """Test getting range with explicit sheet name"""
        reader = ExcelReader(sample_excel_file)
        result = reader.get_range("Sheet1!A1:B2")

        assert result.success is True
        assert isinstance(result.data, list)
        assert len(result.data) == 2
        assert len(result.data[0]) == 2

    def test_reader_get_range_invalid_sheet(self, sample_excel_file):
        """Test getting range from non-existent sheet"""
        reader = ExcelReader(sample_excel_file)
        result = reader.get_range("NonExistentSheet!A1:A5")

        assert result.success is False
        assert "工作表不存在" in result.error or "不存在" in result.error

    def test_reader_get_range_with_formatting(self, sample_excel_file):
        """Test getting range with formatting information"""
        reader = ExcelReader(sample_excel_file)
        result = reader.get_range("Sheet1!A1:B2", include_formatting=True)

        assert result.success is True
        assert isinstance(result.data, list)

    def test_reader_get_range_unicode_content(self, sample_excel_file):
        """Test getting range with unicode content"""
        reader = ExcelReader(sample_excel_file)
        result = reader.get_range("Sheet1!A1:A5")

        assert result.success is True
        assert isinstance(result.data, list)
        assert len(result.data) == 5

        # Check that unicode content is handled properly
        for row in result.data:
            assert len(row) == 1

    # ==================== Excel Writer 测试 ====================

    def test_writer_init_valid_file(self, sample_excel_file):
        """Test ExcelWriter initialization with valid file"""
        writer = ExcelWriter(sample_excel_file)
        assert writer.file_path == sample_excel_file

    def test_writer_init_invalid_file(self):
        """Test ExcelWriter initialization with invalid file"""
        with pytest.raises(ExcelFileNotFoundError):
            ExcelWriter("nonexistent_file.xlsx")

    def test_writer_update_range_single_cell(self, sample_excel_file):
        """Test updating a single cell"""
        writer = ExcelWriter(sample_excel_file)
        result = writer.update_range("Sheet1!A1", [["新标题"]])

        assert result.success is True
        assert isinstance(result.data, list)
        assert len(result.data) == 1
        assert isinstance(result.data[0], ModifiedCell)

    def test_writer_update_range_multiple_cells(self, sample_excel_file):
        """Test updating multiple cells"""
        writer = ExcelWriter(sample_excel_file)
        new_data = [
            ["新产品", "新价格"],
            ["产品A", 100],
            ["产品B", 200]
        ]
        result = writer.update_range("Sheet1!A1:B3", new_data)

        assert result.success is True
        assert isinstance(result.data, list)
        assert len(result.data) == 6  # 3 rows * 2 columns

    def test_writer_update_range_with_sheet_name(self, sample_excel_file):
        """Test updating with explicit sheet name"""
        writer = ExcelWriter(sample_excel_file)
        result = writer.update_range("Sheet2!A1", [["新产品"]])

        assert result.success is True
        assert len(result.data) == 1

    def test_writer_update_range_preserve_formulas(self, sample_excel_file):
        """Test updating while preserving formulas"""
        writer = ExcelWriter(sample_excel_file)
        result = writer.update_range("Sheet1!A6", [["总计行"]], preserve_formulas=True)

        assert result.success is True
        assert len(result.data) >= 1

    def test_writer_update_range_overwrite_formulas(self, sample_excel_file):
        """Test updating and overwriting formulas"""
        writer = ExcelWriter(sample_excel_file)
        result = writer.update_range("Sheet1!E2", [["手动值"]], preserve_formulas=False)

        assert result.success is True
        assert len(result.data) >= 1

    def test_writer_update_range_invalid_sheet(self, sample_excel_file):
        """Test updating non-existent sheet"""
        writer = ExcelWriter(sample_excel_file)
        result = writer.update_range("NonExistentSheet!A1", [["测试"]])

        assert result.success is False
        assert "工作表" in result.error

    def test_writer_insert_rows(self, sample_excel_file):
        """Test inserting rows"""
        writer = ExcelWriter(sample_excel_file)
        result = writer.insert_rows("Sheet1", 2, 2)

        assert result.success is True
        assert 'inserted_count' in result.metadata

    def test_writer_insert_columns(self, sample_excel_file):
        """Test inserting columns"""
        writer = ExcelWriter(sample_excel_file)
        result = writer.insert_columns("Sheet1", 2, 1)

        assert result.success is True
        assert 'inserted_count' in result.metadata

    def test_writer_delete_rows(self, sample_excel_file):
        """Test deleting rows"""
        writer = ExcelWriter(sample_excel_file)
        result = writer.delete_rows("Sheet1", 3, 1)

        assert result.success is True
        assert 'actual_deleted_count' in result.metadata

    def test_writer_delete_columns(self, sample_excel_file):
        """Test deleting columns"""
        writer = ExcelWriter(sample_excel_file)
        result = writer.delete_columns("Sheet1", 3, 1)

        assert result.success is True
        assert 'actual_deleted_count' in result.metadata

    def test_writer_format_cells(self, sample_excel_file):
        """Test formatting cells"""
        writer = ExcelWriter(sample_excel_file)
        formatting = {
            'font': {'name': 'Arial', 'size': 14, 'bold': True}
        }
        result = writer.format_cells("Sheet1!A1:D1", formatting)

        assert result.success is True
        assert 'formatted_count' in result.metadata

    def test_writer_update_range_mixed_data_types(self, sample_excel_file):
        """Test updating with mixed data types"""
        writer = ExcelWriter(sample_excel_file)
        mixed_data = [
            ["文本", 123, 45.67, True],
            ["更多文本", 456, 78.90, False]
        ]
        result = writer.update_range("Sheet1!A1:D2", mixed_data)

        assert result.success is True
        assert len(result.data) == 8  # 2 rows * 4 columns

    # ==================== Excel Manager 测试 ====================

    def test_manager_init_valid_file(self, sample_excel_file):
        """Test ExcelManager initialization"""
        manager = ExcelManager(sample_excel_file)
        assert manager.file_path == sample_excel_file

    def test_manager_create_file(self, temp_dir):
        """Test creating a new Excel file"""
        file_path = temp_dir / "new_test_file.xlsx"
        result = ExcelManager.create_file(str(file_path), ["Sheet1", "Sheet2"])

        assert result.success is True
        assert file_path.exists()
        assert "成功创建" in result.message

    def test_manager_create_file_default_sheets(self, temp_dir):
        """Test creating file with default sheets"""
        file_path = temp_dir / "default_sheets.xlsx"
        result = ExcelManager.create_file(str(file_path))

        assert result.success is True
        assert file_path.exists()

    def test_manager_create_sheet(self, sample_excel_file):
        """Test creating a new sheet"""
        manager = ExcelManager(sample_excel_file)
        result = manager.create_sheet("新工作表")

        assert result.success is True
        assert result.data.name == "新工作表"

    def test_manager_create_sheet_duplicate_name(self, sample_excel_file):
        """Test creating sheet with duplicate name"""
        manager = ExcelManager(sample_excel_file)
        result = manager.create_sheet("Sheet1")  # Already exists

        assert result.success is False
        assert "已存在" in result.error or "exist" in result.error.lower()

    def test_manager_delete_sheet(self, sample_excel_file):
        """Test deleting a sheet"""
        manager = ExcelManager(sample_excel_file)
        # First create a sheet to delete
        manager.create_sheet("临时工作表")

        result = manager.delete_sheet("临时工作表")
        assert result.success is True

    def test_manager_delete_sheet_nonexistent(self, sample_excel_file):
        """Test deleting non-existent sheet"""
        manager = ExcelManager(sample_excel_file)
        result = manager.delete_sheet("不存在的工作表")

        assert result.success is False
        assert "不存在" in result.error or "not found" in result.error.lower()

    def test_manager_rename_sheet(self, sample_excel_file):
        """Test renaming a sheet"""
        manager = ExcelManager(sample_excel_file)
        result = manager.rename_sheet("Sheet1", "重命名工作表")

        assert result.success is True
        assert result.data.name == "重命名工作表"

    def test_manager_rename_sheet_nonexistent(self, sample_excel_file):
        """Test renaming non-existent sheet"""
        manager = ExcelManager(sample_excel_file)
        result = manager.rename_sheet("不存在的工作表", "新名称")

        assert result.success is False
        assert "不存在" in result.error or "not found" in result.error.lower()

    def test_manager_list_sheets(self, sample_excel_file):
        """Test listing all sheets through manager"""
        manager = ExcelManager(sample_excel_file)
        result = manager.list_sheets()

        assert result.success is True
        assert isinstance(result.data, list)
        assert len(result.data) >= 1

    # ==================== 综合测试 ====================

    def test_core_workflow_integration(self, temp_dir):
        """Test integrated workflow: create -> write -> read -> manage"""
        # 1. Create file
        file_path = temp_dir / "integration_test.xlsx"
        result = ExcelManager.create_file(str(file_path), ["数据表", "汇总表"])
        assert result.success is True

        # 2. Write data
        writer = ExcelWriter(str(file_path))
        test_data = [
            ["项目", "金额", "状态"],
            ["项目A", 1000, "完成"],
            ["项目B", 2000, "进行中"]
        ]
        result = writer.update_range("数据表!A1:C3", test_data)
        assert result.success is True

        # 3. Read data back
        reader = ExcelReader(str(file_path))
        result = reader.get_range("数据表!A1:C3")
        assert result.success is True
        assert len(result.data) == 3

        # 4. Manage sheets
        manager = ExcelManager(str(file_path))
        result = manager.create_sheet("临时表")
        assert result.success is True

    def test_core_error_handling_consistency(self, sample_excel_file):
        """Test that all core components handle errors consistently"""
        # Reader error handling
        reader = ExcelReader(sample_excel_file)
        result = reader.get_range("不存在的工作表!A1:A1")
        assert result.success is False
        assert isinstance(result.error, str)

        # Writer error handling
        writer = ExcelWriter(sample_excel_file)
        result = writer.update_range("不存在的工作表!A1", [["测试"]])
        assert result.success is False
        assert isinstance(result.error, str)

        # Manager error handling
        manager = ExcelManager(sample_excel_file)
        result = manager.delete_sheet("不存在的工作表")
        assert result.success is False
        assert isinstance(result.error, str)

    def test_core_chinese_support(self, sample_excel_file):
        """Test Chinese character support across all core components"""
        # Test Chinese data writing
        writer = ExcelWriter(sample_excel_file)
        chinese_data = [["中文标题", "数值"], ["产品名称", 100]]
        result = writer.update_range("Sheet1!A1:B2", chinese_data)
        assert result.success is True

        # Test Chinese data reading
        reader = ExcelReader(sample_excel_file)
        result = reader.get_range("Sheet1!A1:B2")
        assert result.success is True
        assert len(result.data) == 2

        # Test Chinese sheet management
        manager = ExcelManager(sample_excel_file)
        result = manager.create_sheet("中文工作表名称")
        assert result.success is True


class TestFormulaDispatch:
    """公式分发表重构测试 — 验证_range_formulals分发表+条件统计修复"""

    @pytest.fixture
    def writer_with_data(self, sample_excel_file):
        """创建带数据的writer，A1:A10 = [10,20,30,40,50,60,70,80,90,100]"""
        writer = ExcelWriter(sample_excel_file)
        data = [[i * 10] for i in range(1, 11)]
        writer.update_range("Sheet1!A1:A10", data)
        return writer

    def test_dispatch_table_integrity(self):
        """分发表包含所有18个公式函数"""
        table = ExcelWriter._RANGE_FORMULAS
        assert len(table) == 18
        names = ['SUM', 'AVERAGE', 'COUNT', 'MIN', 'MAX', 'MEDIAN',
                 'STDEV', 'VAR', 'PERCENTILE', 'QUARTILE',
                 'MODE', 'SKEW', 'KURT', 'GEOMEAN', 'HARMEAN',
                 'COUNTIF', 'SUMIF', 'AVERAGEIF']
        for i, expected in enumerate(names):
            import re
            assert re.search(expected, table[i][0]), f"第{i}项缺少{expected}"

    def test_formula_sum_range(self, writer_with_data):
        writer = writer_with_data
        from openpyxl import load_workbook
        wb = load_workbook(writer.file_path)
        sheet = wb.active
        result = writer._basic_formula_parse("SUM(A1:A10)", sheet)
        assert result == 550.0
        wb.close()

    def test_formula_average_range(self, writer_with_data):
        writer = writer_with_data
        from openpyxl import load_workbook
        wb = load_workbook(writer.file_path)
        sheet = wb.active
        result = writer._basic_formula_parse("AVERAGE(A1:A10)", sheet)
        assert result == 55.0
        wb.close()

    def test_formula_count_range(self, writer_with_data):
        writer = writer_with_data
        from openpyxl import load_workbook
        wb = load_workbook(writer.file_path)
        sheet = wb.active
        result = writer._basic_formula_parse("COUNT(A1:A10)", sheet)
        assert result == 10
        wb.close()

    def test_formula_min_max(self, writer_with_data):
        writer = writer_with_data
        from openpyxl import load_workbook
        wb = load_workbook(writer.file_path)
        sheet = wb.active
        assert writer._basic_formula_parse("MIN(A1:A10)", sheet) == 10.0
        assert writer._basic_formula_parse("MAX(A1:A10)", sheet) == 100.0
        wb.close()

    def test_formula_median_stdev(self, writer_with_data):
        writer = writer_with_data
        from openpyxl import load_workbook
        wb = load_workbook(writer.file_path)
        sheet = wb.active
        median = writer._basic_formula_parse("MEDIAN(A1:A10)", sheet)
        assert median == 55.0  # (50+60)/2
        stdev = writer._basic_formula_parse("STDEV(A1:A10)", sheet)
        assert stdev > 0
        wb.close()

    def test_formula_countif_all_operators(self, writer_with_data):
        """COUNTIF全部条件运算符（修复前fallback仅支持>）"""
        from openpyxl import load_workbook
        wb = load_workbook(writer_with_data.file_path)
        sheet = wb.active
        # >50: 60,70,80,90,100 = 5
        assert writer_with_data._basic_formula_parse("COUNTIF(A1:A10,\">50\")", sheet) == 5
        # <50: 10,20,30,40 = 4
        assert writer_with_data._basic_formula_parse("COUNTIF(A1:A10,\"<50\")", sheet) == 4
        # >=50: 50,60,70,80,90,100 = 6
        assert writer_with_data._basic_formula_parse("COUNTIF(A1:A10,\">=50\")", sheet) == 6
        # <=50: 10,20,30,40,50 = 5
        assert writer_with_data._basic_formula_parse("COUNTIF(A1:A10,\"<=50\")", sheet) == 5
        # =50: 1
        assert writer_with_data._basic_formula_parse("COUNTIF(A1:A10,\"=50\")", sheet) == 1
        # 隐式等于
        assert writer_with_data._basic_formula_parse("COUNTIF(A1:A10,\"50\")", sheet) == 1
        wb.close()

    def test_formula_sumif_all_operators(self, writer_with_data):
        """SUMIF全部条件运算符（修复前fallback仅支持>）"""
        from openpyxl import load_workbook
        wb = load_workbook(writer_with_data.file_path)
        sheet = wb.active
        # >50: 60+70+80+90+100 = 400
        assert writer_with_data._basic_formula_parse("SUMIF(A1:A10,\">50\")", sheet) == 400.0
        # <50: 10+20+30+40 = 100
        assert writer_with_data._basic_formula_parse("SUMIF(A1:A10,\"<50\")", sheet) == 100.0
        # >=50: 50+60+70+80+90+100 = 450
        assert writer_with_data._basic_formula_parse("SUMIF(A1:A10,\">=50\")", sheet) == 450.0
        # <=50: 10+20+30+40+50 = 150
        assert writer_with_data._basic_formula_parse("SUMIF(A1:A10,\"<=50\")", sheet) == 150.0
        wb.close()

    def test_formula_averageif_all_operators(self, writer_with_data):
        """AVERAGEIF全部条件运算符（修复前fallback仅支持>）"""
        from openpyxl import load_workbook
        wb = load_workbook(writer_with_data.file_path)
        sheet = wb.active
        # >50: avg(60,70,80,90,100) = 80
        assert writer_with_data._basic_formula_parse("AVERAGEIF(A1:A10,\">50\")", sheet) == 80.0
        # <50: avg(10,20,30,40) = 25
        assert writer_with_data._basic_formula_parse("AVERAGEIF(A1:A10,\"<50\")", sheet) == 25.0
        # 无匹配: 返回0
        assert writer_with_data._basic_formula_parse("AVERAGEIF(A1:A10,\">999\")", sheet) == 0
        wb.close()

    def test_formula_concatenate(self, sample_excel_file):
        """CONCATENATE函数"""
        from openpyxl import load_workbook
        writer = ExcelWriter(sample_excel_file)
        wb = load_workbook(sample_excel_file)
        sheet = wb.active
        result = writer._basic_formula_parse(
            'CONCATENATE("Hello", " ", "World")', sheet)
        assert result == "Hello World"
        wb.close()

    def test_formula_if(self, sample_excel_file):
        """IF函数>和<比较"""
        from openpyxl import load_workbook
        writer = ExcelWriter(sample_excel_file)
        wb = load_workbook(sample_excel_file)
        sheet = wb.active
        assert writer._basic_formula_parse("IF(10>5,Y,N)", sheet) == "Y"
        assert writer._basic_formula_parse("IF(5>10,Y,N)", sheet) == "N"
        assert writer._basic_formula_parse("IF(3<7,A,B)", sheet) == "A"
        wb.close()

    def test_formula_simple_math(self, sample_excel_file):
        """简单数学表达式"""
        from openpyxl import load_workbook
        writer = ExcelWriter(sample_excel_file)
        wb = load_workbook(sample_excel_file)
        sheet = wb.active
        assert writer._basic_formula_parse("1+2+3", sheet) == 6
        assert writer._basic_formula_parse("(2+3)*4", sheet) == 20
        assert writer._basic_formula_parse("10/3", sheet) == pytest.approx(10/3, rel=1e-9)
        wb.close()

    def test_formula_list_sum_average(self, sample_excel_file):
        """数字列表SUM/AVERAGE"""
        from openpyxl import load_workbook
        writer = ExcelWriter(sample_excel_file)
        wb = load_workbook(sample_excel_file)
        sheet = wb.active
        assert writer._basic_formula_parse("SUM(1,2,3,4,5)", sheet) == 15
        assert writer._basic_formula_parse("AVERAGE(10,20,30)", sheet) == 20
        wb.close()

    def test_apply_condition_static(self):
        """_apply_condition纯Python实现不依赖numpy"""
        values = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100]
        assert ExcelWriter._apply_condition(values, ">50", "count") == 5
        assert ExcelWriter._apply_condition(values, "<50", "count") == 4
        assert ExcelWriter._apply_condition(values, ">=50", "count") == 6
        assert ExcelWriter._apply_condition(values, "<=50", "count") == 5
        assert ExcelWriter._apply_condition(values, "=50", "count") == 1
        assert ExcelWriter._apply_condition(values, ">50", "sum") == 400.0
        assert ExcelWriter._apply_condition(values, "<50", "average") == 25.0
        assert ExcelWriter._apply_condition(values, ">999", "average") == 0

    # ==================== 智能追加优化测试 ====================

    def test_smart_append_at_end(self, sample_excel_file):
        """智能追加：目标行在数据末尾之后时，跳过insert_rows"""
        from openpyxl import load_workbook
        writer = ExcelWriter(sample_excel_file)

        # 获取当前数据末尾行
        wb = load_workbook(sample_excel_file)
        max_row = wb.active.max_row
        wb.close()

        # 在数据末尾之后写入（min_row > max_row）
        target_row = max_row + 2
        target_range = f"Sheet1!A{target_row}:C{target_row}"
        result = writer.update_range(target_range, [["追加行1", "追加行2", "追加行3"]], insert_mode=True)

        assert result.success
        assert result.metadata.get('smart_append') is True
        assert result.metadata.get('mode_description') == '智能追加模式'

        # 验证数据确实写入了
        wb = load_workbook(sample_excel_file)
        cell_value = wb.active.cell(row=target_row, column=1).value
        wb.close()
        assert cell_value == '追加行1'

    def test_smart_append_multiple_rows(self, sample_excel_file):
        """智能追加：多行追加到末尾"""
        from openpyxl import load_workbook
        writer = ExcelWriter(sample_excel_file)

        wb = load_workbook(sample_excel_file)
        max_row = wb.active.max_row
        wb.close()

        target_row = max_row + 5
        data = [
            ["行1-A", "行1-B"],
            ["行2-A", "行2-B"],
            ["行3-A", "行3-B"],
        ]
        result = writer.update_range(
            f"Sheet1!A{target_row}:B{target_row + 2}", data, insert_mode=True
        )

        assert result.success
        assert result.metadata.get('smart_append') is True
        assert result.metadata.get('modified_cells_count') == 6

    def test_no_smart_append_in_data_range(self, sample_excel_file):
        """非追加场景：目标行在数据范围内，不应触发智能追加"""
        from openpyxl import load_workbook
        writer = ExcelWriter(sample_excel_file)

        wb = load_workbook(sample_excel_file)
        max_row = wb.active.max_row
        wb.close()

        # 在数据范围内写入（min_row <= max_row）
        if max_row >= 3:
            target_row = 3
        else:
            target_row = max_row

        result = writer.update_range(
            f"Sheet1!A{target_row}", [["替换值"]], insert_mode=True
        )

        assert result.success
        assert result.metadata.get('smart_append') is False
        assert result.metadata.get('mode_description') == '插入模式'

    def test_overwrite_mode_no_smart_append(self, sample_excel_file):
        """覆盖模式不应触发智能追加"""
        from openpyxl import load_workbook
        writer = ExcelWriter(sample_excel_file)

        wb = load_workbook(sample_excel_file)
        max_row = wb.active.max_row
        wb.close()

        result = writer.update_range(
            f"Sheet1!A{max_row + 5}", [["覆盖值"]], insert_mode=False
        )

        assert result.success
        assert result.metadata.get('smart_append') is False


class TestFormulaCalculation:
    """公式计算功能测试 - 验证不同文件格式处理"""

    @pytest.fixture
    def excel_file_with_formulas(self, temp_dir):
        """创建包含公式的测试文件"""
        from openpyxl import Workbook
        file_path = temp_dir / "formulas_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # 写入测试数据
        ws['A1'] = 10
        ws['A2'] = 20
        ws['A3'] = 30
        ws['A4'] = 40
        ws['A5'] = 50

        # 写入公式
        ws['B1'] = '=SUM(A1:A5)'
        ws['B2'] = '=AVERAGE(A1:A5)'
        ws['B3'] = '=MAX(A1:A5)'
        ws['B4'] = '=MIN(A1:A5)'
        ws['B5'] = '=COUNT(A1:A5)'

        wb.save(file_path)
        wb.close()
        return file_path

    @pytest.fixture
    def xlsm_file_with_formulas(self, temp_dir):
        """创建包含公式的xlsm文件（带宏的Excel文件）"""
        from openpyxl import Workbook
        file_path = temp_dir / "formulas_test.xlsm"
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # 写入测试数据
        ws['A1'] = 100
        ws['A2'] = 200
        ws['A3'] = 300

        # 写入公式
        ws['B1'] = '=SUM(A1:A3)'
        ws['B2'] = '=AVERAGE(A1:A3)'

        wb.save(file_path)
        wb.close()
        return file_path

    def test_detect_file_format_xlsx(self, excel_file_with_formulas):
        """检测xlsx文件格式"""
        writer = ExcelWriter(str(excel_file_with_formulas))
        detected_format = writer._detect_file_format()
        assert detected_format == 'xlsx'

    def test_detect_file_format_xlsm(self, xlsm_file_with_formulas):
        """检测xlsm文件格式"""
        writer = ExcelWriter(str(xlsm_file_with_formulas))
        detected_format = writer._detect_file_format()
        assert detected_format == 'xlsm'

    def test_detect_file_format_nonexistent(self):
        """检测不存在的文件，应抛出异常"""
        with pytest.raises(ExcelFileNotFoundError):
            ExcelWriter("nonexistent.xlsx")

    def test_create_temp_workbook_xlsx(self, excel_file_with_formulas):
        """为xlsx文件创建临时工作簿进行公式计算"""
        from openpyxl import load_workbook
        writer = ExcelWriter(str(excel_file_with_formulas))
        cache = get_formula_cache()

        temp_wb, temp_path = writer._create_temp_workbook("Data", cache)

        assert temp_wb is not None
        assert temp_path.endswith('.xlsx')

        # 验证临时工作簿包含原始数据
        temp_sheet = temp_wb.active
        assert temp_sheet['A1'].value == 10
        assert temp_sheet['A5'].value == 50

        temp_wb.close()

    def test_create_temp_workbook_xlsm(self, xlsm_file_with_formulas):
        """为xlsm文件创建临时工作簿进行公式计算"""
        from openpyxl import load_workbook
        writer = ExcelWriter(str(xlsm_file_with_formulas))
        cache = get_formula_cache()

        temp_wb, temp_path = writer._create_temp_workbook("Data", cache)

        assert temp_wb is not None
        assert temp_path.endswith('.xlsx')

        # 验证临时工作簿包含原始数据
        temp_sheet = temp_wb.active
        assert temp_sheet['A1'].value == 100
        assert temp_sheet['A3'].value == 300

        temp_wb.close()

    def test_formula_calculation_xlsx(self, excel_file_with_formulas):
        """在xlsx文件中计算公式结果"""
        from openpyxl import load_workbook
        writer = ExcelWriter(str(excel_file_with_formulas))
        wb = load_workbook(excel_file_with_formulas)
        sheet = wb.active

        # 计算SUM公式
        sum_result = writer._basic_formula_parse("SUM(A1:A5)", sheet)
        assert sum_result == 150.0

        # 计算AVERAGE公式
        avg_result = writer._basic_formula_parse("AVERAGE(A1:A5)", sheet)
        assert avg_result == 30.0

        # 计算MAX公式
        max_result = writer._basic_formula_parse("MAX(A1:A5)", sheet)
        assert max_result == 50.0

        # 计算MIN公式
        min_result = writer._basic_formula_parse("MIN(A1:A5)", sheet)
        assert min_result == 10.0

        # 计算COUNT公式
        count_result = writer._basic_formula_parse("COUNT(A1:A5)", sheet)
        assert count_result == 5

        wb.close()

    def test_formula_calculation_xlsm(self, xlsm_file_with_formulas):
        """在xlsm文件中计算公式结果"""
        from openpyxl import load_workbook
        writer = ExcelWriter(str(xlsm_file_with_formulas))
        wb = load_workbook(xlsm_file_with_formulas)
        sheet = wb.active

        # 计算SUM公式
        sum_result = writer._basic_formula_parse("SUM(A1:A3)", sheet)
        assert sum_result == 600.0

        # 计算AVERAGE公式
        avg_result = writer._basic_formula_parse("AVERAGE(A1:A3)", sheet)
        assert avg_result == 200.0

        wb.close()

    def test_temp_workbook_preserves_format(self, excel_file_with_formulas):
        """临时工作簿创建时保留原始文件格式信息"""
        writer = ExcelWriter(str(excel_file_with_formulas))
        cache = get_formula_cache()

        # 获取检测到的格式
        file_format = writer._detect_file_format()

        # 创建临时工作簿
        temp_wb, temp_path = writer._create_temp_workbook("Data", cache)

        # 验证临时文件扩展名与检测格式一致
        assert file_format == 'xlsx'
        assert temp_path.endswith('.xlsx')

        temp_wb.close()

    def test_formula_with_multiple_formats(self, temp_dir):
        """验证公式计算在不同格式文件中的一致性"""
        from openpyxl import Workbook, load_workbook

        # 创建xlsx文件
        xlsx_path = temp_dir / "test_xlsx.xlsx"
        wb_xlsx = Workbook()
        ws_xlsx = wb_xlsx.active
        ws_xlsx['A1'] = 5
        ws_xlsx['A2'] = 15
        ws_xlsx['A3'] = 25
        wb_xlsx.save(xlsx_path)
        wb_xlsx.close()

        # 创建xlsm文件
        xlsm_path = temp_dir / "test_xlsm.xlsm"
        wb_xlsm = Workbook()
        ws_xlsm = wb_xlsm.active
        ws_xlsm['A1'] = 5
        ws_xlsm['A2'] = 15
        ws_xlsm['A3'] = 25
        wb_xlsm.save(xlsm_path)
        wb_xlsm.close()

        # 在两种格式文件中计算相同公式，结果应一致
        writer_xlsx = ExcelWriter(str(xlsx_path))
        wb_xlsx = load_workbook(xlsx_path)
        sheet_xlsx = wb_xlsx.active
        sum_xlsx = writer_xlsx._basic_formula_parse("SUM(A1:A3)", sheet_xlsx)
        avg_xlsx = writer_xlsx._basic_formula_parse("AVERAGE(A1:A3)", sheet_xlsx)
        wb_xlsx.close()

        writer_xlsm = ExcelWriter(str(xlsm_path))
        wb_xlsm = load_workbook(xlsm_path)
        sheet_xlsm = wb_xlsm.active
        sum_xlsm = writer_xlsm._basic_formula_parse("SUM(A1:A3)", sheet_xlsm)
        avg_xlsm = writer_xlsm._basic_formula_parse("AVERAGE(A1:A3)", sheet_xlsm)
        wb_xlsm.close()

        assert sum_xlsx == sum_xlsm == 45.0
        assert avg_xlsx == avg_xlsm == 15.0
