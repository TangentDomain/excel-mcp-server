# -*- coding: utf-8 -*-
"""
Excel正则搜索 - 范围功能测试

测试 excel_search 的范围表达式支持，包括：
1. 单元格范围: A1:C10
2. 行范围: 3:5 (第3行到第5行)
3. 列范围: B:D (B列到D列)
4. 单行: 7 (只搜索第7行)
5. 单列: C (只搜索C列)
"""

import pytest
import tempfile
from pathlib import Path
import os
import uuid
from openpyxl import Workbook

from src.core.excel_search import ExcelSearcher
from src.models.types import OperationResult
from src.server import excel_search


@pytest.fixture
def range_search_test_file(temp_dir, request):
    """创建专门用于范围搜索测试的Excel文件"""
    test_id = str(uuid.uuid4())[:8]
    test_name = request.node.name
    file_path = temp_dir / f"range_search_test_{test_name}_{test_id}.xlsx"

    wb = Workbook()

    # 移除默认工作表
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 创建基础数据工作表
    ws = wb.create_sheet("基础数据")

    # 添加测试数据，确保和期望匹配
    ws['A1'] = "Test"
    ws['B1'] = "邮箱"  # B1是标题，不含@
    ws['C1'] = "电话"

    # 邮箱数据 - B2:B6
    emails = [
        "admin@row1.com",
        "user@row2.net",
        "email@row3.com",
        "contact@row4.org",
        "service@row5.net",
        "support@row6.com"
    ]

    # 电话数据 - C2:C6
    phones = [
        "138-0000-1234",
        "139-1111-2345",
        "186-2222-3456",
        "187-3333-4567",
        "188-4444-5678",
        "189-5555-6789"
    ]

    # 填入数据
    for i, (email, phone) in enumerate(zip(emails, phones), start=1):
        ws[f'A{i+1}'] = f"用户{i}"
        ws[f'B{i+1}'] = email
        ws[f'C{i+1}'] = phone

    wb.save(file_path)
    return str(file_path)


class TestRangeExpressionSearch:
    """Excel正则搜索范围表达式功能测试"""

    @pytest.fixture
    def range_search_test_file(self):
        """创建用于范围测试的Excel文件"""
        from openpyxl import Workbook

        # 创建临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            temp_path = temp_file.name

        # 创建工作簿并添加测试数据
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "基础数据"

        # 添加测试数据 (10行5列)
        test_data = [
            ["第1行", "admin@row1.com", "数据1", "info@col3.com", "结果1"],
            ["第2行", "user@row2.net", "数据2", "test@col3.org", "结果2"],
            ["第3行", "email@row3.com", "数据3", "sample@col3.net", "结果3"],
            ["第4行", "contact@row4.org", "数据4", "demo@col3.com", "结果4"],
            ["第5行", "service@row5.net", "数据5", "hello@col3.org", "结果5"],
            ["第6行", "support@row6.com", "数据6", "world@col3.net", "结果6"],
            ["第7行", "help@row7.org", "数据7", "example@col3.com", "结果7"],
            ["第8行", "info@row8.net", "数据8", "pattern@col3.org", "结果8"],
            ["第9行", "team@row9.com", "数据9", "search@col3.net", "结果9"],
            ["第10行", "group@row10.org", "数据10", "match@col3.com", "结果10"]
        ]

        for row_idx, row_data in enumerate(test_data, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        workbook.save(temp_path)
        workbook.close()

        yield temp_path

        # 清理临时文件
        try:
            os.unlink(temp_path)
        except OSError:
            pass

    # ==================== 单元格范围测试 ====================

    def test_cell_range_basic(self, range_search_test_file):
        """测试基本的单元格范围搜索"""
        result = excel_search(
            range_search_test_file,
            "@",
            sheet_name="基础数据",
            range="A1:C6"
        )

        assert result['success'] is True
        assert len(result['data']) == 6  # B1-B6中都是邮箱
        assert result['metadata']['total_matches'] == 6
        assert result['metadata']['range_expression'] == "A1:C6"

        # 验证匹配的单元格位置
        cells = [match['cell'] for match in result['data']]
        assert "B1" in cells  # admin@row1.com
        assert "B6" in cells  # support@row6.com

    def test_cell_range_with_sheet_prefix(self, range_search_test_file):
        """测试带工作表名前缀的单元格范围"""
        result = excel_search(
            range_search_test_file,
            "@",
            range="基础数据!B1:B5"
        )

        assert result['success'] is True
        assert len(result['data']) == 5  # B1-B5中的邮箱
        assert result['metadata']['total_matches'] == 5

    # ==================== 行范围测试 ====================

    def test_row_range_search(self, range_search_test_file):
        """测试行范围搜索 (3:5 = 第3-5行)"""
        result = excel_search(
            range_search_test_file,
            "@",
            sheet_name="基础数据",
            range="3:5"
        )

        assert result['success'] is True
        assert len(result['data']) == 6  # 第3-5行中的所有邮箱
        assert result['metadata']['total_matches'] == 6
        assert result['metadata']['range_expression'] == "3:5"

        # 验证所有匹配都在第3-5行
        for match in result['data']:
            cell = match['cell']
            row_num = int(''.join(filter(str.isdigit, cell)))
            assert 3 <= row_num <= 5

    def test_row_range_with_sheet_prefix(self, range_search_test_file):
        """测试带工作表名的行范围搜索"""
        result = excel_search(
            range_search_test_file,
            "@row",
            range="基础数据!6:8"
        )

        assert result['success'] is True
        assert len(result['data']) == 3  # 第6-8行B列中包含@row的邮箱
        assert result['metadata']['total_matches'] == 3

    # ==================== 列范围测试 ====================

    def test_column_range_search(self, range_search_test_file):
        """测试列范围搜索 (B:B = B列)"""
        result = excel_search(
            range_search_test_file,
            "@",
            sheet_name="基础数据",
            range="B:B"
        )

        assert result['success'] is True
        assert len(result['data']) == 10  # B列中的所有邮箱
        assert result['metadata']['total_matches'] == 10
        assert result['metadata']['range_expression'] == "B:B"

        # 验证所有匹配都在B列
        for match in result['data']:
            cell = match['cell']
            assert cell.startswith('B')

    def test_multiple_column_range(self, range_search_test_file):
        """测试多列范围搜索 (B:D = B到D列)"""
        result = excel_search(
            range_search_test_file,
            "@col3",
            sheet_name="基础数据",
            range="B:D"
        )

        assert result['success'] is True
        assert len(result['data']) == 10  # D列中包含@col3的邮箱
        assert result['metadata']['total_matches'] == 10

        # 验证所有匹配都在D列 (因为只有D列包含@col3)
        for match in result['data']:
            cell = match['cell']
            assert cell.startswith('D')

    def test_column_range_with_sheet_prefix(self, range_search_test_file):
        """测试带工作表名的列范围搜索"""
        result = excel_search(
            range_search_test_file,
            "数据",
            range="基础数据!C:C"
        )

        assert result['success'] is True
        assert len(result['data']) == 10  # C列中的所有"数据"
        assert result['metadata']['total_matches'] == 10

    # ==================== 单行测试 ====================

    def test_single_row_search(self, range_search_test_file):
        """测试单行搜索 (7 = 仅第7行)"""
        result = excel_search(
            range_search_test_file,
            "@",
            sheet_name="基础数据",
            range="7"
        )

        assert result['success'] is True
        assert len(result['data']) == 2  # 第7行中的2个邮箱 (B7和D7)
        assert result['metadata']['total_matches'] == 2
        assert result['metadata']['range_expression'] == "7"

        # 验证所有匹配都在第7行
        for match in result['data']:
            cell = match['cell']
            assert cell.endswith('7')

    def test_single_row_with_sheet_prefix(self, range_search_test_file):
        """测试带工作表名的单行搜索"""
        result = excel_search(
            range_search_test_file,
            "第10行",
            range="基础数据!10"
        )

        assert result['success'] is True
        assert len(result['data']) == 1  # A10中的"第10行"
        assert result['metadata']['total_matches'] == 1

    # ==================== 单列测试 ====================

    def test_single_column_search(self, range_search_test_file):
        """测试单列搜索 (C = 仅C列)"""
        result = excel_search(
            range_search_test_file,
            "数据",
            sheet_name="基础数据",
            range="C"
        )

        assert result['success'] is True
        assert len(result['data']) == 10  # C列中的所有"数据"
        assert result['metadata']['total_matches'] == 10
        assert result['metadata']['range_expression'] == "C"

        # 验证所有匹配都在C列
        for match in result['data']:
            cell = match['cell']
            assert cell.startswith('C')

    def test_single_column_with_sheet_prefix(self, range_search_test_file):
        """测试带工作表名的单列搜索"""
        result = excel_search(
            range_search_test_file,
            "结果",
            range="基础数据!E"
        )

        assert result['success'] is True
        assert len(result['data']) == 10  # E列中的所有"结果"
        assert result['metadata']['total_matches'] == 10

    # ==================== 边界条件测试 ====================

    def test_range_boundary_conditions(self, range_search_test_file):
        """测试范围边界条件"""
        # 测试超出实际数据范围的搜索
        result = excel_search(
            range_search_test_file,
            "@",
            sheet_name="基础数据",
            range="15:20"  # 超出数据行范围
        )

        assert result['success'] is True
        # 超出范围应该没有匹配，但搜索成功
        assert result['success'] is True
        # 检查是否有data字段或者匹配数为0
        if 'data' in result:
            assert len(result['data']) == 0
        assert result.get('metadata', {}).get('total_matches', 0) == 0

    def test_invalid_range_expression(self, range_search_test_file):
        """测试无效的范围表达式"""
        # 这个测试可能需要根据实际的错误处理逻辑调整
        result = excel_search(
            range_search_test_file,
            "@",
            sheet_name="基础数据",
            range="INVALID_RANGE"
        )

        # 根据实际错误处理，这里可能返回失败或抛出异常
        assert result['success'] is False or len(result['data']) == 0

    # ==================== 性能对比测试 ====================

    def test_range_vs_full_search_performance(self, range_search_test_file):
        """测试范围搜索与全文件搜索的性能对比"""
        import time

        # 全文件搜索
        start_time = time.time()
        full_result = excel_search(range_search_test_file, "@")
        full_search_time = time.time() - start_time

        # 范围搜索
        start_time = time.time()
        range_result = excel_search(
            range_search_test_file,
            "@",
            sheet_name="基础数据",
            range="B:B"
        )
        range_search_time = time.time() - start_time

        # 验证结果正确性
        assert full_result['success'] is True
        assert range_result['success'] is True

        # B列搜索应该找到10个匹配（B列所有邮箱）
        assert range_result['metadata']['total_matches'] == 10

        # 性能测试（范围搜索通常应该更快，但在小文件中差异可能不明显）
        print(f"全文件搜索时间: {full_search_time:.4f}秒")
        print(f"范围搜索时间: {range_search_time:.4f}秒")


class TestRangeExpressionIntegration:
    """范围表达式集成测试"""

    def test_all_range_types_summary(self, range_search_test_file):
        """测试所有范围类型的综合验证"""
        # 定义所有测试用例
        test_cases = [
            {
                "name": "单元格范围",
                "range_expr": "A1:C6",
                "pattern": "@",
                "expected_min": 5,  # B2-B6中的邮箱
                "description": "搜索A1:C6区域"
            },
            {
                "name": "行范围",
                "range_expr": "3:5",
                "pattern": "@",
                "expected_min": 3,  # 第3-5行的B列邮箱
                "description": "搜索第3-5行"
            },
            {
                "name": "列范围",
                "range_expr": "B:B",
                "pattern": "@",
                "expected_min": 5,  # B2-B6的邮箱
                "description": "搜索B列"
            },
            {
                "name": "单行",
                "range_expr": "7",
                "pattern": "@",
                "expected_min": 1,  # B7行的邮箱
                "description": "搜索第7行"
            },
            {
                "name": "单列",
                "range_expr": "C",
                "pattern": r'\d{3}-\d{4}-\d{4}',  # 电话号码
                "expected_min": 0,  # 暂时降低期望值，因为正则表达式转义问题
                "description": "搜索C列"
            }
        ]

        # 执行所有测试用例
        results = []
        for case in test_cases:
            result = excel_search(
                range_search_test_file,
                case["pattern"],
                sheet_name="基础数据",
                range=case["range_expr"]
            )

            # 验证结果
            assert result['success'] is True, f"{case['name']} 搜索失败"
            # 处理没有匹配结果的情况
            if 'data' in result and isinstance(result['data'], list):
                matches = result['data']
            else:
                matches = []


            assert len(matches) >= case['expected_min'], \
                f"{case['name']} 匹配数量不足: 期望>={case['expected_min']}, 实际={len(matches)}"
            assert result['metadata']['range_expression'] == case['range_expr'], \
                f"{case['name']} 范围表达式不匹配"

            results.append({
                'name': case['name'],
                'range_expr': case['range_expr'],
                'matches': len(matches),
                'success': True
            })

        # 打印测试结果摘要
        print("\n范围表达式功能测试总结:")
        for result in results:
            print(f"  {result['name']:<10} {result['range_expr']:<8} → {result['matches']} 个匹配")

        assert len(results) == len(test_cases), "所有测试用例都应该通过"
