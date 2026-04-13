"""
Round 6 Bug 诊断脚本
"""
import os, sys, tempfile, shutil
sys.path.insert(0, '/root/workspace/excel-mcp-server/src')

from openpyxl import Workbook
from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
)

TEST_DIR = tempfile.mkdtemp(prefix='excelmcp_diag_')
TEST_FILE = os.path.join(TEST_DIR, 'diag.xlsx')

wb = Workbook()
ws = wb.active; ws.title = "装备"
ws.append(["ID", "Name", "Type", "BaseAtk", "Price", "Rarity"])
for r in [[1,"铁剑","Weapon",10,100,"Common"],[2,"钢剑","Weapon",25,500,"Rare"]]:
    ws.append(r)
ws2 = wb.create_sheet("技能")
ws2.append(["ID", "Name", "Type", "Damage", "ManaCost", "ClassReq"])
for r in [[101,"火焰斩","Active",150,20,"Warrior"],[102,"冰霜箭","Active",120,15,"Archer"]]:
    ws2.append(r)
ws3 = wb.create_sheet("空表")
ws3.append(["ID", "Name", "Value"])  # only header
wb.save(TEST_FILE)

print("=" * 60)
print("Bug A7 诊断: UNION子查询在FROM中")
print("=" * 60)

# 测试 sqlglot 如何解析 UNION 子查询
import sqlglot
sql_a7 = """SELECT Type, COUNT(*) as cnt FROM (
    SELECT Type FROM 装备
    UNION ALL
    SELECT ClassReq as Type FROM 技能
) t GROUP BY Type"""
parsed = sqlglot.parse(sql_a7)[0]
print(f"顶层类型: {type(parsed).__name__}")
from_clause = parsed.args.get('from') or parsed.args.get('from_')
print(f"FROM类型: {type(from_clause.this).__name__ if from_clause else 'None'}")
if from_clause and hasattr(from_clause, 'this'):
    inner = from_clause.this
    print(f"FROM.this类型: {type(inner).__name__}")
    if isinstance(inner, sqlglot.exp.Subquery):
        print(f"  Subquery.this类型: {type(inner.this).__name__}")

# 实际执行
r = execute_advanced_sql_query(TEST_FILE, sql_a7)
print(f"\n执行结果: success={r.get('success')}")
print(f"message: {r.get('message', '')[:200]}")

print("\n" + "=" * 60)
print("Bug E2 诊断: 空Sheet聚合")
print("=" * 60)

sql_e2 = "SELECT COUNT(*) as cnt, AVG(Value) as avg_val FROM 空表"
r2 = execute_advanced_sql_query(TEST_FILE, sql_e2)
print(f"执行结果: success={r2.get('success')}")
print(f"message: {r2.get('message', '')[:200]}")
if r2.get('data'):
    print(f"data: {r2['data']}")

# 检查空表的DataFrame结构
import pandas as pd
df_empty = pd.read_excel(TEST_FILE, sheet_name='空表', header=0)
print(f"\n空表DataFrame: shape={df_empty.shape}, columns={list(df_empty.columns)}")

print("\n" + "=" * 60)
print("Bug C2 诊断: 三表JOIN链式解析")
print("=" * 60)

# 添加掉落和商店sheet
ws4 = wb.create_sheet("掉落")
ws4.append(["MonsterID", "MonsterName", "DropItemID", "DropRate", "MinQty", "MaxQty"])
ws4.append([1001, "史莱姆", 1, 50.0, 1, 3])
ws5 = wb.create_sheet("商店")
ws5.append(["ID", "ItemID", "ItemType", "Stock", "Discount", "Currency"])
ws5.append([1, 1, "装备", 999, 0, "Gold"])
wb.save(TEST_FILE)

sql_c2 = """SELECT 掉落.MonsterName, 装备.Name as ItemName, 装备.Price, 商店.Stock
FROM 掉落 
INNER JOIN 装备 ON 掉落.DropItemID = 装备.ID
LEFT JOIN 商店 ON 装备.ID = 商店.ItemID"""
r3 = execute_advanced_sql_query(TEST_FILE, sql_c2)
print(f"执行结果: success={r3.get('success')}")
print(f"message: {r3.get('message', '')[:300]}")

shutil.rmtree(TEST_DIR)
print("\n✅ 诊断完成")
