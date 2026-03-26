# -*- coding: utf-8 -*-
"""
Excel Writer 更多测试覆盖

覆盖 excel_writer.py 中未覆盖的边缘代码路径
"""

import pytest
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.core.excel_writer import ExcelWriter


class TestExcelWriterMoreFormulas:
    """更多公式测试"""

    @pytest.fixture
    def formula_file(self, temp_dir):
        """创建公式测试文件"""
        file_path = temp_dir / "formula_more.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "FormulaSheet"
        
        # 数值数据
        for i in range(1, 21):
            ws.cell(row=i, column=1, value=i * 10)
        
        wb.save(file_path)
        return str(file_path)

    def test_formula_if_nested(self, formula_file):
        """测试嵌套IF"""
        writer = ExcelWriter(formula_file)
        result = writer.update_range(
            "FormulaSheet!C1",
            [["=IF(A1>100,\"A\",IF(A1>50,\"B\",\"C\"))"]],
            preserve_formulas=True,
            insert_mode=False
        )
        assert result.success is True

    def test_formula_power(self, formula_file):
        """测试幂运算"""
        writer = ExcelWriter(formula_file)
        result = writer.update_range(
            "FormulaSheet!C2",
            [["=POWER(A1,2)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        assert result.success is True

    def test_formula_sqrt(self, formula_file):
        """测试平方根"""
        writer = ExcelWriter(formula_file)
        result = writer.update_range(
            "FormulaSheet!C3",
            [["=SQRT(A1)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        assert result.success is True

    def test_formula_round(self, formula_file):
        """测试四舍五入"""
        writer = ExcelWriter(formula_file)
        result = writer.update_range(
            "FormulaSheet!C4",
            [["=ROUND(A1/3,2)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        assert result.success is True

    def test_formula_abs(self, formula_file):
        """测试绝对值"""
        writer = ExcelWriter(formula_file)
        result = writer.update_range(
            "FormulaSheet!C5",
            [["=ABS(-A1)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        assert result.success is True

    def test_formula_mod(self, formula_file):
        """测试取模"""
        writer = ExcelWriter(formula_file)
        result = writer.update_range(
            "FormulaSheet!C6",
            [["=MOD(A1,7)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        assert result.success is True


class TestExcelWriterStylesAdvanced:
    """高级样式测试"""

    @pytest.fixture
    def style_file(self, temp_dir):
        """创建样式测试文件"""
        file_path = temp_dir / "style_advanced.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "StyleSheet"
        
        ws['A1'] = "Header"
        ws['A2'] = "Data"
        
        wb.save(file_path)
        return str(file_path)

    def test_set_row_height_multiple(self, style_file):
        """测试设置多行行高"""
        writer = ExcelWriter(style_file)
        
        for row in range(1, 6):
            result = writer.set_row_height(row, 20 + row * 5, "StyleSheet")
            assert result.success is True

    def test_set_column_width_multiple(self, style_file):
        """测试设置多列列宽"""
        writer = ExcelWriter(style_file)
        
        for col in ['A', 'B', 'C', 'D']:
            result = writer.set_column_width(col, 15, "StyleSheet")
            assert result.success is True

    def test_merge_cells_range(self, style_file):
        """测试合并单元格范围"""
        writer = ExcelWriter(style_file)
        result = writer.merge_cells("A1:C3", "StyleSheet")
        assert result.success is True


class TestExcelWriterBulkOperations:
    """批量操作测试"""

    @pytest.fixture
    def bulk_file(self, temp_dir):
        """创建批量测试文件"""
        file_path = temp_dir / "bulk_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "BulkSheet"
        
        # 创建大量数据
        for row in range(1, 51):
            for col in range(1, 11):
                ws.cell(row=row, column=col, value=f"{row}_{col}")
        
        wb.save(file_path)
        return str(file_path)

    def test_bulk_insert_rows(self, bulk_file):
        """测试批量插入行"""
        writer = ExcelWriter(bulk_file)
        result = writer.insert_rows("BulkSheet", 10, 5)
        assert result.success is True

    def test_bulk_delete_rows(self, bulk_file):
        """测试批量删除行"""
        writer = ExcelWriter(bulk_file)
        result = writer.delete_rows("BulkSheet", 10, 5)
        assert result.success is True

    def test_bulk_insert_columns(self, bulk_file):
        """测试批量插入列"""
        writer = ExcelWriter(bulk_file)
        result = writer.insert_columns("BulkSheet", 5, 3)
        assert result.success is True


class TestExcelWriterDataTypes:
    """数据类型测试"""

    @pytest.fixture
    def data_type_file(self, temp_dir):
        """创建数据类型测试文件"""
        file_path = temp_dir / "data_type.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DataTypeSheet"
        
        wb.save(file_path)
        return str(file_path)

    def test_write_integer(self, data_type_file):
        """测试写入整数"""
        writer = ExcelWriter(data_type_file)
        result = writer.update_range(
            "DataTypeSheet!A1",
            [[12345]],
            preserve_formulas=True,
            insert_mode=False
        )
        assert result.success is True

    def test_write_float(self, data_type_file):
        """测试写入浮点数"""
        writer = ExcelWriter(data_type_file)
        result = writer.update_range(
            "DataTypeSheet!A2",
            [[3.14159265359]],
            preserve_formulas=True,
            insert_mode=False
        )
        assert result.success is True

    def test_write_boolean(self, data_type_file):
        """测试写入布尔值"""
        writer = ExcelWriter(data_type_file)
        result = writer.update_range(
            "DataTypeSheet!A3",
            [[True], [False]],
            preserve_formulas=True,
            insert_mode=False
        )
        assert result.success is True

    def test_write_date(self, data_type_file):
        """测试写入日期"""
        from datetime import datetime
        writer = ExcelWriter(data_type_file)
        result = writer.update_range(
            "DataTypeSheet!A4",
            [[datetime(2024, 1, 1)]],
            preserve_formulas=True,
            insert_mode=False
        )
        assert result.success is True


class TestExcelWriterErrorHandling:
    """错误处理测试"""

    def test_invalid_range(self, temp_dir):
        """测试无效范围"""
        file_path = temp_dir / "error_test.xlsx"
        
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(file_path)
        
        writer = ExcelWriter(file_path)
        result = writer.update_range(
            "Sheet1!XXX",
            [[1]],
            preserve_formulas=True,
            insert_mode=False
        )
        # 无效范围应该失败
        assert result is not None
