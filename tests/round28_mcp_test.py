"""
Round 28 MCP 接口实测
主题: UPDATE/INSERT/DELETE 深度测试 + CASE WHEN 复杂表达式 + 公式单元格 + P0全回归
日期: 2026-04-14
"""

import sys
import os
import tempfile
import shutil
import random
import numpy as np

# 添加 src 到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openpyxl import Workbook
from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)

# ============================================================
# 测试数据准备
# ============================================================
TEMP_DIR = tempfile.mkdtemp(prefix='r28_test_')
TEST_FILE = os.path.join(TEMP_DIR, 'test_round28.xlsx')
TEST_FILE_MIXED = os.path.join(TEMP_DIR, 'test_mixed_r28.xlsx')
TEST_FILE_FORMULA = os.path.join(TEMP_DIR, 'test_formula_r28.xlsx')

def create_base_test_file():
    """创建基础测试文件 - 装备配置表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "装备配置"
    headers = ["ID", "Name", "BaseAtk", "AtkBonus", "Price", "Rarity", "Category"]
    ws.append(headers)
    random.seed(42)
    for i in range(1, 31):
        ws.append([
            i,
            f"Item-{i}",
            random.randint(10, 100),
            round(random.uniform(5.5, 45.8), 2),
            round(random.uniform(50.5, 9999.99), 2),
            random.choice(["Common", "Rare", "Epic", "Legendary"]),
            random.choice(["Weapon", "Armor", "Accessory"])
        ])
    wb.save(TEST_FILE)
    return TEST_FILE

def create_mixed_type_file():
    """创建混合类型测试文件（用于P0回归: uint8溢出）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "MixedData"
    ws.append(["ID", "V", "F"])  # V=int列, F=float列 → 触发uint8推断
    for i in range(1, 11):
        ws.append([i, i * 3, float(i) * 1.5])
    wb.save(TEST_FILE_MIXED)
    return TEST_FILE_MIXED

def create_formula_file():
    """创建含公式的测试文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "FormulaSheet"
    ws.append(["ID", "A", "B", "Sum", "Product"])
    for i in range(1, 11):
        ws.append([i, i * 2, i * 3, None, None])
        # 写入公式
        row = i + 1
        ws.cell(row=row, column=4).value = f"=B{row}+C{row}"  # Sum = A+B
        ws.cell(row=row, column=5).value = f"=B{row}*C{row}"  # Product = A*B
    
    # 第二个sheet: 纯数据
    ws2 = wb.create_sheet("PureData")
    ws2.append(["ID", "Value", "Status"])
    for i in range(1, 21):
        ws2.append([i, i * 10, "active" if i % 2 == 0 else "inactive"])
    
    wb.save(TEST_FILE_FORMULA)
    return TEST_FILE_FORMULA


class TestResult:
    def __init__(self):
        self.results = []
        self.passed = 0
        self.failed = 0
    
    def add(self, name, passed, detail=""):
        status = "✅" if passed else "❌"
        self.results.append((name, passed, detail))
        if passed:
            self.passed += 1
        else:
            self.failed += 1
        print(f"  {status} {name}" + (f" — {detail}" if detail else ""))
    
    def summary(self):
        total = self.passed + self.failed
        print(f"\n  📊 合计: {self.passed}/{total} 通过 ({self.passed/total*100:.0f}%)" if total > 0 else "")
        return self.passed, self.failed


# ============================================================
# A组: UPDATE 深度测试（P0修复后重点验证）
# ============================================================
def test_update_deep(tr):
    print("\n" + "="*60)
    print("🔧 A组: UPDATE 深度测试")
    print("="*60)
    
    create_base_test_file()
    
    # A1: 基础UPDATE
    r = execute_advanced_update_query(TEST_FILE, "UPDATE 装备配置 SET Price = 100 WHERE ID = 1")
    tr.add("A1: 基础UPDATE SET常量", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # 验证更新结果
    r2 = execute_advanced_sql_query(TEST_FILE, "SELECT Price FROM 装备配置 WHERE ID = 1")
    tr.add("A1-验证: Price=100", 
           r2['success'] and len(r2.get('data', [])) > 1 and r2['data'][1][0] == 100,
           f"data={r2.get('data', [])}")
    
    # A2: UPDATE带数学表达式
    r = execute_advanced_update_query(TEST_FILE, "UPDATE 装备配置 SET Price = ROUND(Price * 1.2, 2) WHERE Rarity = 'Legendary'")
    tr.add("A2: UPDATE SET数学表达式(ROUND)", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # A3: UPDATE SET多列
    r = execute_advanced_update_query(TEST_FILE, "UPDATE 装备配置 SET BaseAtk = 999, AtkBonus = 88.8 WHERE ID = 5")
    tr.add("A3: UPDATE SET多列", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # A4: UPDATE带CASE WHEN
    r = execute_advanced_update_query(TEST_FILE, 
        "UPDATE 装备配置 SET Price = CASE WHEN Rarity = 'Legendary' THEN 9999 WHEN Rarity = 'Epic' THEN 4999 ELSE 99 END")
    tr.add("A4: UPDATE SET CASE WHEN表达式", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # 验证CASE WHEN结果
    r2 = execute_advanced_sql_query(TEST_FILE, "SELECT DISTINCT Price FROM 装备配置 ORDER BY Price")
    if r2['success'] and len(r2.get('data', [])) > 1:
        prices = set(str(x[0]) for x in r2['data'][1:])  # skip header
        expected = {'9999', '4999', '99'}
        tr.add("A4-验证: CASE WHEN三档价格", prices == expected, f"prices={prices}")
    else:
        tr.add("A4-验证: CASE WHEN三档价格", False, f"查询失败: {r2.get('message', '')[:60]}")
    
    # A5: UPDATE无WHERE（全表更新）
    r = execute_advanced_update_query(TEST_FILE, "UPDATE 装备配置 SET Category = 'General'")
    tr.add("A5: UPDATE无WHERE(全表更新)", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # A6: UPDATE WHERE IN
    r = execute_advanced_update_query(TEST_FILE, "UPDATE 装备配置 SET BaseAtk = 777 WHERE ID IN (1, 2, 3)")
    tr.add("A6: UPDATE WHERE IN子句", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # A7: UPDATE WHERE BETWEEN
    r = execute_advanced_update_query(TEST_FILE, "UPDATE 装备配置 SET AtkBonus = 0.0 WHERE ID BETWEEN 10 AND 20")
    tr.add("A7: UPDATE WHERE BETWEEN", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # A8: UPDATE WHERE LIKE
    r = execute_advanced_update_query(TEST_FILE, "UPDATE 装备配置 SET Name = 'Special-' || Name WHERE Name LIKE 'Item-1%'")
    tr.add("A8: UPDATE WHERE LIKE+字符串拼接", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # A9: UPDATE不匹配任何行
    r = execute_advanced_update_query(TEST_FILE, "UPDATE 装备配置 SET Price = 0 WHERE ID = 99999")
    tr.add("A9: UPDATE零行影响(无匹配)", r['success'], r.get('message', '')[:80] if not r['success'] else '')


# ============================================================
# B组: P0 回归验证（uint8溢出修复）
# ============================================================
def test_p0_regression(tr):
    print("\n" + "="*60)
    print("🔧 B组: P0 回归验证 (UPDATE uint8溢出)")
    print("="*60)
    
    create_mixed_type_file()
    
    # B1: P0核心场景: SET V=999 (>255, 应该不被截断)
    r = execute_advanced_update_query(TEST_FILE_MIXED, "UPDATE MixedData SET V = 999 WHERE ID = 1")
    tr.add("B1: [P0核心] Mixed列SET V=999", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    r2 = execute_advanced_sql_query(TEST_FILE_MIXED, "SELECT V FROM MixedData WHERE ID = 1")
    if r2['success'] and len(r2.get('data', [])) > 1:
        val = r2['data'][1][0]
        tr.add("B1-验证: V=999(非231!)", val == 999, f"实际值={val} (旧bug会返回231)")
    else:
        tr.add("B1-验证: V=999(非231!)", False, f"查询失败")
    
    # B2: 边界值256
    create_mixed_type_file()  # 重置
    r = execute_advanced_update_query(TEST_FILE_MIXED, "UPDATE MixedData SET V = 256 WHERE ID = 2")
    r2 = execute_advanced_sql_query(TEST_FILE_MIXED, "SELECT V FROM MixedData WHERE ID = 2")
    if r2['success'] and len(r2.get('data', [])) > 1:
        val = r2['data'][1][0]
        tr.add("B2: [P0边界] V=256(uint8边界+1)", val == 256, f"实际值={val} (旧bug会返回0)")
    else:
        tr.add("B2: [P0边界] V=256", False, "查询失败")
    
    # B3: 大值1000
    create_mixed_type_file()
    r = execute_advanced_update_query(TEST_FILE_MIXED, "UPDATE MixedData SET V = 1000 WHERE ID = 3")
    r2 = execute_advanced_sql_query(TEST_FILE_MIXED, "SELECT V FROM MixedData WHERE ID = 3")
    if r2['success'] and len(r2.get('data', [])) > 1:
        val = r2['data'][1][0]
        tr.add("B3: [P0大值] V=1000", val == 1000, f"实际值={val} (旧bug会返回232)")
    else:
        tr.add("B3: [P0大值] V=1000", False, "查询失败")
    
    # B4: 超大值99999
    create_mixed_type_file()
    r = execute_advanced_update_query(TEST_FILE_MIXED, "UPDATE MixedData SET V = 99999 WHERE ID = 4")
    r2 = execute_advanced_sql_query(TEST_FILE_MIXED, "SELECT V FROM MixedData WHERE ID = 4")
    if r2['success'] and len(r2.get('data', [])) > 1:
        val = r2['data'][1][0]
        tr.add("B4: [P0超大] V=99999", val == 99999, f"实际值={val}")
    else:
        tr.add("B4: [P0超大] V=99999", False, "查询失败")
    
    # B5: 负数（uint8不能表示负数）
    create_mixed_type_file()
    r = execute_advanced_update_query(TEST_FILE_MIXED, "UPDATE MixedData SET V = -1 WHERE ID = 5")
    tr.add("B5: [P0负数] V=-1(负数写入)", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    r2 = execute_advanced_sql_query(TEST_FILE_MIXED, "SELECT V FROM MixedData WHERE ID = 5")
    if r2['success'] and len(r2.get('data', [])) > 1:
        val = r2['data'][1][0]
        tr.add("B5-验证: V=-1", val == -1, f"实际值={val}")
    else:
        tr.add("B5-验证: V=-1", False, "查询失败")
    
    # B6: 0值（边界）
    create_mixed_type_file()
    r = execute_advanced_update_query(TEST_FILE_MIXED, "UPDATE MixedData SET V = 0 WHERE ID = 6")
    r2 = execute_advanced_sql_query(TEST_FILE_MIXED, "SELECT V FROM MixedData WHERE ID = 6")
    if r2['success'] and len(r2.get('data', [])) > 1:
        val = r2['data'][1][0]
        tr.add("B6: [P0零值] V=0", val == 0, f"实际值={val}")
    else:
        tr.add("B6: [P0零值] V=0", False, "查询失败")
    
    # B7: 浮点数写入整数列
    create_mixed_type_file()
    r = execute_advanced_update_query(TEST_FILE_MIXED, "UPDATE MixedData SET V = 123.456 WHERE ID = 7")
    r2 = execute_advanced_sql_query(TEST_FILE_MIXED, "SELECT V FROM MixedData WHERE ID = 7")
    if r2['success'] and len(r2.get('data', [])) > 1:
        val = r2['data'][1][0]
        tr.add("B7: [P0类型混入] V=123.456(浮点→整列)", isinstance(val, (int, float, np.integer, np.floating)), f"实际值={val}, 类型={type(val)}")
    else:
        tr.add("B7: [P0类型混入] V=123.456", False, "查询失败")


# ============================================================
# C组: INSERT 深度测试
# ============================================================
def test_insert_deep(tr):
    print("\n" + "="*60)
    print("🔧 C组: INSERT 深度测试")
    print("="*60)
    
    create_base_test_file()
    
    # C1: 基础INSERT单行
    r = execute_advanced_insert_query(
        TEST_FILE,
        "INSERT INTO 装备配置 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category) VALUES (100, 'Excalibur', 150, 50.5, 9999.99, 'Legendary', 'Weapon')"
    )
    tr.add("C1: INSERT单行完整数据", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # 验证插入
    r2 = execute_advanced_sql_query(TEST_FILE, "SELECT Name, Price FROM 装备配置 WHERE ID = 100")
    if r2['success'] and len(r2.get('data', [])) > 1:
        tr.add("C1-验证: 插入数据可查", r2['data'][1][0] == 'Excalibur' and float(r2['data'][1][1]) == 9999.99, f"data={r2['data'][1]}")
    else:
        tr.add("C1-验证: 插入数据可查", False, f"查询失败: {r2.get('message', '')[:60]}")
    
    # C2: INSERT含特殊字符
    r = execute_advanced_insert_query(
        TEST_FILE,
        "INSERT INTO 装备配置 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category) VALUES (101, '日本語アイテム「特殊」', 200, 99.9, 888.88, 'Epic', 'Accessory')"
    )
    tr.add("C2: INSERT含Unicode特殊字符", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # C3: INSERT含单引号
    r = execute_advanced_insert_query(
        TEST_FILE,
        "INSERT INTO 装备配置 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category) VALUES (102, 'It''s Magic', 100, 30.0, 500.0, 'Rare', 'Weapon')"
    )
    tr.add("C3: INSERT含转义单引号(It''s)", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # C4: INSERT空字符串值
    r = execute_advanced_insert_query(
        TEST_FILE,
        "INSERT INTO 装备配置 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category) VALUES (103, '', 0, 0.0, 0.0, 'Common', '')"
    )
    tr.add("C4: INSERT空字符串值", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # C5: INSERT负数值
    r = execute_advanced_insert_query(
        TEST_FILE,
        "INSERT INTO 装备配置 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category) VALUES (104, 'DebtItem', -100, -50.5, -999.99, 'Common', 'Weapon')"
    )
    tr.add("C5: INSERT负数值", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # C6: INSERT极大数值
    r = execute_advanced_insert_query(
        TEST_FILE,
        "INSERT INTO 装备配置 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category) VALUES (105, 'BigNum', 2147483647, 999999.999, 999999999.99, 'Legendary', 'Armor')"
    )
    tr.add("C6: INSERT极大数值(INT_MAX级)", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # C7: INSERT极小浮点
    r = execute_advanced_insert_query(
        TEST_FILE,
        "INSERT INTO 装备配置 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category) VALUES (106, 'TinyFloat', 1, 0.000001, 0.000000001, 'Common', 'Accessory')"
    )
    tr.add("C7: INSERT极小浮点数", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # C8: INSERT重复主键(看是否允许)
    r = execute_advanced_insert_query(
        TEST_FILE,
        "INSERT INTO 装备配置 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category) VALUES (1, 'Duplicate', 0, 0.0, 0.0, 'Common', 'Test')"
    )
    tr.add("C8: INSERT重复ID(冲突处理)", r['success'], r.get('message', '')[:80])


# ============================================================
# D组: DELETE 深度测试
# ============================================================
def test_delete_deep(tr):
    print("\n" + "="*60)
    print("🔧 D组: DELETE 深度测试")
    print("="*60)
    
    create_base_test_file()  # 重新创建，确保有30条数据
    
    # 先确认初始行数
    r0 = execute_advanced_sql_query(TEST_FILE, "SELECT COUNT(*) FROM 装备配置")
    initial_count = None
    if r0['success'] and len(r0.get('data', [])) > 1:
        initial_count = int(r0['data'][1][0])
    
    # D1: DELETE单行
    r = execute_advanced_delete_query(TEST_FILE, "DELETE FROM 装备配置 WHERE ID = 1")
    tr.add("D1: DELETE单行(WHERE ID=1)", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    r2 = execute_advanced_sql_query(TEST_FILE, "SELECT COUNT(*) FROM 装备配置")
    if r2['success'] and initial_count is not None and len(r2.get('data', [])) > 1:
        new_count = int(r2['data'][1][0])
        tr.add("D1-验证: 行数减1", new_count == initial_count - 1, f"{initial_count} → {new_count}")
    
    # D2: DELETE WHERE条件(IN)
    r = execute_advanced_delete_query(TEST_FILE, "DELETE FROM 装备配置 WHERE ID IN (2, 3, 4)")
    tr.add("D2: DELETE WHERE IN (多条)", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # D3: DELETE WHERE LIKE
    r = execute_advanced_delete_query(TEST_FILE, "DELETE FROM 装备配置 WHERE Name LIKE 'Item-5%'")
    tr.add("D3: DELETE WHERE LIKE模式", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # D4: DELETE WHERE BETWEEN
    r = execute_advanced_delete_query(TEST_FILE, "DELETE FROM 装备配置 WHERE ID BETWEEN 6 AND 10")
    tr.add("D4: DELETE WHERE BETWEEN范围", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # D5: DELETE WHERE比较运算符
    r = execute_advanced_delete_query(TEST_FILE, "DELETE FROM 装备配置 WHERE Price > 5000")
    tr.add("D5: DELETE WHERE Price>5000", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # D6: DELETE无WHERE（清空表！危险操作测试）
    # 用一个临时sheet测
    r = execute_advanced_delete_query(TEST_FILE, "DELETE FROM 装备配置 WHERE Rarity = 'Common'")
    tr.add("D6: DELETE按类别批量删除", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # D7: DELETE零匹配
    r = execute_advanced_delete_query(TEST_FILE, "DELETE FROM 装备配置 WHERE ID = 99999")
    tr.add("D7: DELETE零匹配(安全无影响)", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # D8: DELETE WHERE复合条件(AND/OR)
    r = execute_advanced_delete_query(TEST_FILE, "DELETE FROM 装备配置 WHERE Rarity = 'Legendary' AND Category = 'Weapon'")
    tr.add("D8: DELETE WHERE复合AND条件", r['success'], r.get('message', '')[:80] if not r['success'] else '')


# ============================================================
# E组: CASE WHEN 复杂表达式测试
# ============================================================
def test_case_when_complex(tr):
    print("\n" + "="*60)
    print("🔧 E组: CASE WHEN 复杂表达式测试")
    print("="*60)
    
    create_base_test_file()
    
    # E1: 简单CASE WHEN SELECT
    r = execute_advanced_sql_query(TEST_FILE, 
        "SELECT Name, CASE WHEN Rarity = 'Legendary' THEN 'S' WHEN Rarity = 'Epic' THEN 'A' ELSE 'B' END as Grade FROM 装备配置 LIMIT 5")
    tr.add("E1: CASE WHEN简单分级(SELECT)", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    if r['success'] and len(r.get('data', [])) > 1:
        grades = set(row[-1] for row in r['data'][1:])
        tr.add("E1-验证: Grade只含S/A/B", grades.issubset({'S', 'A', 'B'}), f"grades={grades}")
    
    # E2: CASE WHEN嵌套表达式
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT Name, CASE WHEN Price > 5000 THEN CASE WHEN BaseAtk > 50 THEN 'TopTier' ELSE 'Expensive' END ELSE 'Normal' END as Tier FROM 装备配置 LIMIT 5")
    tr.add("E2: CASE WHEN嵌套(SELECT)", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # E3: CASE WHEN与聚合函数
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT CASE WHEN COUNT(*) > 5 THEN 'Many' ELSE 'Few' END as CntGrade FROM 装备配置")
    tr.add("E3: CASE WHEN包裹聚合COUNT", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # E4: CASE WHEN在WHERE中
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT * FROM 装备配置 WHERE CASE WHEN Rarity = 'Legendary' THEN 1 ELSE 0 END = 1 LIMIT 3")
    tr.add("E4: CASE WHEN在WHERE条件中", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # E5: CASE WHEN与ORDER BY组合
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT Name, Rarity, CASE Rarity WHEN 'Legendary' THEN 1 WHEN 'Epic' THEN 2 WHEN 'Rare' THEN 3 ELSE 4 END as ROrder FROM 装备配置 ORDER BY ROrder LIMIT 5")
    tr.add("E5: CASE WHEN简写+ORDER BY", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # E6: CASE WHEN NULL处理(COALESCE替代)
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT Name, COALESCE(NULL, Price, 0) as SafePrice FROM 装备配置 LIMIT 3")
    tr.add("E6: COALESCE NULL处理", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # E7: CASE WHEN与算术混合
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT Name, BaseAtk * CASE WHEN Rarity = 'Legendary' THEN 2.0 ELSE 1.0 END as AdjAtk FROM 装备配置 LIMIT 5")
    tr.add("E7: CASE WHEN与算术乘法混合", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # E8: 多个CASE WHEN列
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT Name, CASE WHEN BaseAtk > 50 THEN 'High' ELSE 'Low' END as AtkTier, CASE WHEN Price > 1000 THEN 'Rich' ELSE 'Cheap' END as Wealth FROM 装备配置 LIMIT 5")
    tr.add("E8: 多个CASE WHEN列同时查询", r['success'], r.get('message', '')[:80] if not r['success'] else '')


# ============================================================
# F组: 公式单元格交互测试
# ============================================================
def test_formula_interaction(tr):
    print("\n" + "="*60)
    print("🔧 F组: 公式单元格交互测试")
    print("="*60)
    
    create_formula_file()
    
    # F1: 查询含公式的sheet（公式应被当作值读取）
    r = execute_advanced_sql_query(TEST_FILE_FORMULA, "SELECT * FROM FormulaSheet LIMIT 5")
    tr.add("F1: 查询含公式Sheet", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    if r['success']:
        print(f"    公式列数据示例: {r.get('data', [])[:3]}")
    
    # F2: 对含公式的数据进行SQL过滤
    r = execute_advanced_sql_query(TEST_FILE_FORMULA, "SELECT ID, A, B FROM FormulaSheet WHERE A > 10 ORDER BY ID")
    tr.add("F2: 过滤公式Sheet普通列", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # F3: 聚合公式Sheet的数据
    r = execute_advanced_sql_query(TEST_FILE_FORMULA, "SELECT COUNT(*), SUM(A), AVG(B) FROM FormulaSheet")
    tr.add("F3: 聚合公式Sheet数据", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # F4: 更新公式Sheet的普通列
    r = execute_advanced_update_query(TEST_FILE_FORMULA, "UPDATE FormulaSheet SET A = 100 WHERE ID = 1")
    tr.add("F4: UPDATE公式Sheet普通列", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # F5: 跨Sheet查询（纯数据Sheet）
    r = execute_advanced_sql_query(TEST_FILE_FORMULA, "SELECT Status, COUNT(*) as cnt FROM PureData GROUP BY Status")
    tr.add("F5: 公式文件的纯数据Sheet聚合", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # F6: JOIN公式Sheet和纯数据Sheet
    r = execute_advanced_sql_query(TEST_FILE_FORMULA, 
        "SELECT f.ID, f.A, p.Value FROM FormulaSheet f JOIN PureData p ON f.ID = p.ID LIMIT 5")
    tr.add("F6: JOIN公式Sheet与纯数据Sheet", r['success'], r.get('message', '')[:80] if not r['success'] else '')


# ============================================================
# G组: 已知P0/P1问题回归
# ============================================================
def test_known_issues_regression(tr):
    print("\n" + "="*60)
    print("🔧 G组: 已知P0/P1问题回归")
    print("="*60)
    
    create_base_test_file()
    
    # G1: [P0-2] SELECT分号多语句注入
    r = execute_advanced_sql_query(TEST_FILE, "SELECT COUNT(*) FROM 装备配置; SELECT COUNT(*) FROM 装备配置")
    is_injection_working = r['success'] and '多语句' in str(r.get('message', ''))
    tr.add("G1: [P0-2回归] SELECT分号多语句", is_injection_working, 
           "仍可注入⚠️" if is_injection_working else "已修复✅")
    
    # G2: 边界值: 超长字符串
    long_str = "A" * 10000
    # 通过API可能不好传超长SQL，改用插入方式
    r = execute_advanced_insert_query(
        TEST_FILE,
        f"INSERT INTO 装备配置 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category) VALUES (200, '{long_str}', 1, 1.0, 1.0, 'Common', 'Test')"
    )
    tr.add("G2: 超长字符串INSERT(10000字符)", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # G3: SQL注入尝试(单引号逃逸)
    r = execute_advanced_sql_query(TEST_FILE, "SELECT * FROM 装备配置 WHERE Name = '' OR 1=1 --'")
    tr.add("G3: SQL注入(OR 1=1注释)", r['success'], f"返回{len(r.get('data', []))-1}行" if r['success'] else r.get('message', '')[:60])
    
    # G4: 特殊Sheet名(反引号)
    # 创建特殊名称sheet的文件
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet-Special_数据"
    ws.append(["ID", "Val"])
    ws.append([1, "test"])
    special_file = os.path.join(TEMP_DIR, 'special.xlsx')
    wb.save(special_file)
    
    r = execute_advanced_sql_query(special_file, "SELECT * FROM `Sheet-Special_数据`")
    tr.add("G4: 特殊Sheet名反引号引用", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # G5: NULL比较
    r = execute_advanced_sql_query(TEST_FILE, "SELECT * FROM 装备配置 WHERE Name = NULL LIMIT 3")
    tr.add("G5: NULL等值比较(应为空)", r['success'], f"返回{len(r.get('data',[]))-1}行" if r['success'] else r.get('message','')[:60])
    
    # G6: IS NULL
    r = execute_advanced_sql_query(TEST_FILE, "SELECT * FROM 装备配置 WHERE Name IS NULL LIMIT 3")
    tr.add("G6: IS NULL语法", r['success'], f"返回{len(r.get('data',[]))-1}行" if r['success'] else r.get('message','')[:60])
    
    # G7: NOT IN空集
    r = execute_advanced_sql_query(TEST_FILE, "SELECT * FROM 装备配置 WHERE ID NOT IN (99999, 88888) LIMIT 3")
    tr.add("G7: NOT IN空集(应返回全部)", r['success'], f"返回{len(r.get('data',[]))-1}行" if r['success'] else r.get('message','')[:60])
    
    # G8: ORDER BY多列
    r = execute_advanced_sql_query(TEST_FILE, "SELECT * FROM 装备配置 ORDER BY Rarity DESC, Price ASC LIMIT 5")
    tr.add("G8: ORDER BY多列(DESC+ASC)", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # G9: DISTINCT
    r = execute_advanced_sql_query(TEST_FILE, "SELECT DISTINCT Rarity, Category FROM 装备配置")
    tr.add("G9: DISTINCT多列去重", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # G10: HAVING
    r = execute_advanced_sql_query(TEST_FILE, "SELECT Rarity, COUNT(*) as cnt FROM 装备配置 GROUP BY Rarity HAVING COUNT(*) > 3")
    tr.add("G10: HAVING过滤分组", r['success'], r.get('message', '')[:80] if not r['success'] else '')


# ============================================================
# H组: 子查询深度测试
# ============================================================
def test_subquery_deep(tr):
    print("\n" + "="*60)
    print("🔧 H组: 子查询深度测试")
    print("="*60)
    
    create_base_test_file()
    
    # H1: WHERE子查询(标量)
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT * FROM 装备配置 WHERE Price > (SELECT AVG(Price) FROM 装备配置) LIMIT 5")
    tr.add("H1: WHERE标量子查询(AVG)", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # H2: WHERE IN子查询
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT * FROM 装备配置 WHERE ID IN (SELECT ID FROM 装备配置 WHERE Rarity = 'Legendary') LIMIT 5")
    tr.add("H2: WHERE IN子查询", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # H3: FROM子查询(派生表)
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT * FROM (SELECT Rarity, AVG(Price) as AvgP FROM 装备配置 GROUP BY Rarity) t WHERE AvgP > 100")
    tr.add("H3: FROM派生表子查询", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # H4: SELECT中的标量子查询
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT Name, Price, (SELECT AVG(Price) FROM 装备配置) as OverallAvg FROM 装备配置 LIMIT 3")
    tr.add("H4: SELECT标量子查询", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # H5: CTE + UPDATE (如果支持)
    # 先测试CTE基本功能
    r = execute_advanced_sql_query(TEST_FILE,
        "WITH HighPrice AS (SELECT * FROM 装备配置 WHERE Price > 5000) SELECT COUNT(*) FROM HighPrice")
    tr.add("H5: CTE基础查询", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # H6: EXISTS相关子查询
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT * FROM 装备配置 e WHERE EXISTS (SELECT 1 FROM 装备配置 e2 WHERE e2.Rarity = 'Legendary' AND e2.ID = e.ID) LIMIT 3")
    tr.add("H6: EXISTS相关子查询", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # H7: NOT EXISTS
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT * FROM 装备配置 e WHERE NOT EXISTS (SELECT 1 FROM 装备配置 e2 WHERE e2.ID = 99999 AND e2.ID = e.ID) LIMIT 3")
    tr.add("H7: NOT EXISTS子查询", r['success'], r.get('message', '')[:80] if not r['success'] else '')
    
    # H8: 嵌套多层子查询
    r = execute_advanced_sql_query(TEST_FILE,
        "SELECT * FROM (SELECT Rarity, AVG(Price) a FROM 装备配置 GROUP BY Rarity) t WHERE a > (SELECT MIN(Price) FROM 装备配置)")
    tr.add("H8: 多层嵌套子查询", r['success'], r.get('message', '')[:80] if not r['success'] else '')


# ============================================================
# 主流程
# ============================================================
def main():
    print("=" * 70)
    print("🔄 Round 28: ExcelMCP 迭代测试")
    print("   主题: UPDATE/INSERT/DELETE深度 + CASE WHEN复杂式 + 公式单元格 + P0回归")
    print("=" * 70)
    
    tr = TestResult()
    
    try:
        # A组: UPDATE深度
        test_update_deep(tr)
        
        # B组: P0回归
        test_p0_regression(tr)
        
        # C组: INSERT深度
        test_insert_deep(tr)
        
        # D组: DELETE深度
        test_delete_deep(tr)
        
        # E组: CASE WHEN复杂
        test_case_when_complex(tr)
        
        # F组: 公式交互
        test_formula_interaction(tr)
        
        # G组: 已知问题回归
        test_known_issues_regression(tr)
        
        # H组: 子查询深度
        test_subquery_deep(tr)
        
    except Exception as e:
        print(f"\n❌ 测试过程异常: {e}")
        import traceback
        traceback.print_exc()
    finally:
        # 清理临时文件
        shutil.rmtree(TEMP_DIR, ignore_errors=True)
    
    # 最终汇总
    print("\n" + "=" * 70)
    print("📊 Round 28 最终汇总")
    print("=" * 70)
    passed, failed = tr.summary()
    
    print(f"\n📝 详细结果:")
    for name, passed_flag, detail in tr.results:
        status = "✅" if passed_flag else "❌"
        detail_str = f"  └─ {detail}" if detail else ""
        print(f"  {status} {name}{detail_str}")
    
    return passed, failed


if __name__ == "__main__":
    passed, failed = main()
    sys.exit(0 if failed == 0 else 1)
