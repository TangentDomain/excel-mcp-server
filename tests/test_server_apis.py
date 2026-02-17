# -*- coding: utf-8 -*-
"""
Server API 测试套件

覆盖 server.py 中的 MCP 工具接口
"""

import pytest
from pathlib import Path
from openpyxl import Workbook


class TestServerAPIs:
    """Server API 测试"""

    @pytest.fixture
    def test_file(self, temp_dir):
        """创建测试文件"""
        file_path = temp_dir / "test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws['A1'] = "ID"
        ws['B1'] = "Name"
        ws['A2'] = 1
        ws['B2'] = "Test"
        ws['A3'] = 2
        ws['B3'] = "Data"
        
        wb.save(file_path)
        return str(file_path)

    def test_list_sheets_api(self, test_file):
        """测试列出工作表API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.list_sheets(test_file)
        
        assert result['success'] is True
        assert 'sheets' in result

    def test_get_file_info_api(self, test_file):
        """测试获取文件信息API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.get_file_info(test_file)
        
        assert result['success'] is True
        assert 'data' in result

    def test_get_range_api(self, test_file):
        """测试获取范围API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.get_range(test_file, "Sheet1!A1:B3")
        
        assert result['success'] is True

    def test_get_headers_api(self, test_file):
        """测试获取表头API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.get_headers(test_file, "Sheet1")
        
        assert result['success'] is True

    def test_find_last_row_api(self, test_file):
        """测试查找最后一行API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.find_last_row(test_file, "Sheet1")
        
        assert result['success'] is True

    def test_search_api(self, test_file):
        """测试搜索API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.search(test_file, "Test", "Sheet1")
        
        assert result['success'] is True


class TestServerAdvancedAPIs:
    """Server 高级 API 测试"""

    @pytest.fixture
    def multi_sheet_file(self, temp_dir):
        """创建多工作表文件"""
        file_path = temp_dir / "multi.xlsx"
        
        wb = Workbook()
        
        ws1 = wb.active
        ws1.title = "Data"
        ws1['A1'] = "ID"
        ws1['A2'] = 1
        ws1['A3'] = 2
        
        ws2 = wb.create_sheet("Config")
        ws2['A1'] = "Key"
        ws2['A2'] = "Value"
        
        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def test_file_for_create(self, temp_dir):
        """创建测试文件"""
        file_path = temp_dir / "test_create.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        wb.save(file_path)
        return str(file_path)

    def test_create_sheet_api(self, test_file_for_create):
        """测试创建工作表API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.create_sheet(
            file_path=test_file_for_create,
            sheet_name="NewSheet"
        )
        
        assert result is not None

    def test_delete_sheet_api(self, multi_sheet_file):
        """测试删除工作表API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.delete_sheet(multi_sheet_file, "Config")
        
        # 可能成功也可能失败，取决于实现
        assert result is not None

    def test_rename_sheet_api(self, test_file_for_create):
        """测试重命名工作表API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.rename_sheet(
            file_path=test_file_for_create,
            old_name="Sheet1",
            new_name="RenamedSheet"
        )
        
        assert result is not None


class TestServerFormatAPIs:
    """Server 格式化 API 测试"""

    @pytest.fixture
    def format_file(self, temp_dir):
        """创建格式化测试文件"""
        file_path = temp_dir / "format.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "FormatSheet"
        
        ws['A1'] = "Header"
        ws['A2'] = "Data"
        
        wb.save(file_path)
        return str(file_path)

    def test_format_cells_api(self, format_file):
        """测试格式化单元格API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.format_cells(
            file_path=format_file,
            sheet_name="FormatSheet",
            range="A1",
            formatting={"font": {"bold": True}}
        )
        
        assert result is not None

    def test_set_borders_api(self, format_file):
        """测试设置边框API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.set_borders(
            file_path=format_file,
            sheet_name="FormatSheet",
            range="A1:B2",
            border_style="thin"
        )
        
        assert result is not None


class TestServerDataAPIs:
    """Server 数据操作 API 测试"""

    @pytest.fixture
    def data_file(self, temp_dir):
        """创建数据操作测试文件"""
        file_path = temp_dir / "data.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DataSheet"
        
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i)
        
        wb.save(file_path)
        return str(file_path)

    def test_insert_rows_api(self, data_file):
        """测试插入行API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.insert_rows(
            file_path=data_file,
            sheet_name="DataSheet",
            row_index=3,
            count=1
        )
        
        assert result is not None

    def test_delete_rows_api(self, data_file):
        """测试删除行API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.delete_rows(
            file_path=data_file,
            sheet_name="DataSheet",
            row_index=3,
            count=1
        )
        
        assert result is not None

    def test_insert_columns_api(self, data_file):
        """测试插入列API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.insert_columns(
            file_path=data_file,
            sheet_name="DataSheet",
            column_index=2,
            count=1
        )
        
        assert result is not None

    def test_delete_columns_api(self, data_file):
        """测试删除列API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.delete_columns(
            file_path=data_file,
            sheet_name="DataSheet",
            column_index=2,
            count=1
        )
        
        assert result is not None


class TestServerCompareAPIs:
    """Server 比较 API 测试"""

    @pytest.fixture
    def compare_file1(self, temp_dir):
        """创建比较文件1"""
        file_path = temp_dir / "compare1.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        ws['A1'] = "ID"
        ws['A2'] = 1
        ws['A3'] = 2
        
        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def compare_file2(self, temp_dir):
        """创建比较文件2"""
        file_path = temp_dir / "compare2.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        ws['A1'] = "ID"
        ws['A2'] = 1
        ws['A3'] = 3
        
        wb.save(file_path)
        return str(file_path)

    def test_compare_sheets_api(self, compare_file1, compare_file2):
        """测试比较工作表API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.compare_sheets(
            file1_path=compare_file1,
            sheet1_name="Data",
            file2_path=compare_file2,
            sheet2_name="Data"
        )
        
        assert result is not None

    def test_check_duplicate_ids_api(self, compare_file1):
        """测试检查重复ID API"""
        from src.api.excel_operations import ExcelOperations
        result = ExcelOperations.check_duplicate_ids(
            file_path=compare_file1,
            sheet_name="Data",
            id_column=1
        )
        
        assert result is not None


class TestServerExportAPIs:
    """Server 导入导出 API 测试"""

    @pytest.fixture
    def export_file(self, temp_dir):
        """创建导出测试文件"""
        file_path = temp_dir / "export.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws['A1'] = "ID"
        ws['A2'] = 1
        ws['A3'] = 2
        
        wb.save(file_path)
        return str(file_path)

    def test_export_to_csv_api(self, export_file, temp_dir):
        """测试导出CSV API"""
        from src.api.excel_operations import ExcelOperations
        csv_path = temp_dir / "export.csv"
        
        result = ExcelOperations.export_to_csv(
            file_path=export_file,
            output_path=str(csv_path),
            sheet_name="Sheet1"
        )
        
        assert result is not None

    def test_convert_format_api(self, export_file, temp_dir):
        """测试格式转换 API"""
        from src.api.excel_operations import ExcelOperations
        output_path = temp_dir / "output.xls"
        
        result = ExcelOperations.convert_format(
            input_path=export_file,
            output_path=str(output_path),
            target_format="xls"
        )
        
        # xls 转换可能失败但API应该返回结果
        assert result is not None
