# -*- coding: utf-8 -*-
"""
Excel Writer 高级功能测试套件

覆盖 excel_writer.py 中的高级公式计算功能
"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from src.core.excel_writer import ExcelWriter


class TestExcelWriterAdvancedFormulas:
    """ExcelWriter 高级公式测试"""

    @pytest.fixture
    def formula_test_file(self, temp_dir):
        """创建公式测试文件"""
        file_path = temp_dir / "formula_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "FormulaSheet"
        
        # 创建数值数据用于测试公式
        # A列: 数值
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=i * 10)
        
        # B列: 数值
        for i in range(1, 11):
            ws.cell(row=i, column=2, value=i * 5)
        
        # C列: 文本
        ws['C1'] = "Apple"
        ws['C2'] = "Banana"
        ws['C3'] = "Apple"
        ws['C4'] = "Cherry"
        ws['C5'] = "Banana"
        
        wb.save(file_path)
        return str(file_path)

    # ==================== SUM函数测试 ====================

    def test_formula_sum_basic(self, formula_test_file):
        """测试SUM公式基础功能"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!D1",
            [["=SUM(A1:A10)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_sum_multiple_ranges(self, formula_test_file):
        """测试SUM公式多范围"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!D2",
            [["=SUM(A1:A5,B1:B5)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== AVERAGE函数测试 ====================

    def test_formula_average_basic(self, formula_test_file):
        """测试AVERAGE公式"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!D3",
            [["=AVERAGE(A1:A10)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== MIN/MAX函数测试 ====================

    def test_formula_min_max(self, formula_test_file):
        """测试MIN/MAX公式"""
        writer = ExcelWriter(formula_test_file)
        
        result1 = writer.update_range(
            "FormulaSheet!D4",
            [["=MIN(A1:A10)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        result2 = writer.update_range(
            "FormulaSheet!D5",
            [["=MAX(A1:A10)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result1.success is True
        assert result2.success is True

    # ==================== COUNT函数测试 ====================

    def test_formula_count(self, formula_test_file):
        """测试COUNT公式"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!D6",
            [["=COUNT(A1:A10)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_countif(self, formula_test_file):
        """测试COUNTIF公式"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!D7",
            [["=COUNTIF(A1:A10,\">50\")"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_countif_less(self, formula_test_file):
        """测试COUNTIF小于条件"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!D8",
            [["=COUNTIF(A1:A10,\"<50\")"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_countif_equal(self, formula_test_file):
        """测试COUNTIF等于条件"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!D9",
            [["=COUNTIF(A1:A10,\"=50\")"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== SUMIF函数测试 ====================

    def test_formula_sumif(self, formula_test_file):
        """测试SUMIF公式"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!D10",
            [["=SUMIF(A1:A10,\">30\")"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_sumif_less(self, formula_test_file):
        """测试SUMIF小于条件"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!E1",
            [["=SUMIF(A1:A10,\"<50\")"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== IF函数测试 ====================

    def test_formula_if_basic(self, formula_test_file):
        """测试IF公式基础功能"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!E2",
            [["=IF(A1>50,\"High\",\"Low\")"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_if_greater(self, formula_test_file):
        """测试IF大于条件"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!E3",
            [["=IF(A1>100,1,0)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_if_less(self, formula_test_file):
        """测试IF小于条件"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!E4",
            [["=IF(A1<50,1,0)"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== 嵌套公式测试 ====================

    def test_formula_nested_if(self, formula_test_file):
        """测试嵌套IF"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!E5",
            [["=IF(A1>80,\">80\",IF(A1>60,\">60\",\"<60\"))"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== 数学运算测试 ====================

    def test_formula_addition(self, formula_test_file):
        """测试加法"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!E6",
            [["=A1+B1"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_subtraction(self, formula_test_file):
        """测试减法"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!E7",
            [["=A1-B1"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_multiplication(self, formula_test_file):
        """测试乘法"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!E8",
            [["=A1*B1"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_division(self, formula_test_file):
        """测试除法"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!E9",
            [["=A1/B1"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_formula_combined_operations(self, formula_test_file):
        """测试组合运算"""
        writer = ExcelWriter(formula_test_file)
        result = writer.update_range(
            "FormulaSheet!E10",
            [["=(A1+B1)*C1"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True


class TestExcelWriterStyles:
    """ExcelWriter 样式测试"""

    @pytest.fixture
    def style_test_file(self, temp_dir):
        """创建样式测试文件"""
        file_path = temp_dir / "style_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "StyleSheet"
        
        ws['A1'] = "Header"
        ws['A2'] = "Data1"
        ws['A3'] = "Data2"
        
        wb.save(file_path)
        return str(file_path)

    def test_set_row_height_style(self, style_test_file):
        """测试设置行高"""
        writer = ExcelWriter(style_test_file)
        result = writer.set_row_height(1, 30, "StyleSheet")
        
        assert result.success is True

    def test_set_column_width_style(self, style_test_file):
        """测试设置列宽"""
        writer = ExcelWriter(style_test_file)
        result = writer.set_column_width("A", 20, "StyleSheet")
        
        assert result.success is True

    def test_merge_cells_style(self, style_test_file):
        """测试合并单元格"""
        writer = ExcelWriter(style_test_file)
        result = writer.merge_cells("A1:B1", "StyleSheet")
        
        assert result.success is True

    def test_unmerge_cells_style(self, style_test_file):
        """测试取消合并"""
        writer = ExcelWriter(style_test_file)
        
        # 先合并
        writer.merge_cells("A1:B1", "StyleSheet")
        
        # 再取消合并
        result = writer.unmerge_cells("A1:B1", "StyleSheet")
        
        assert result.success is True


class TestExcelWriterComplex:
    """ExcelWriter 复杂场景测试"""

    @pytest.fixture
    def complex_test_file(self, temp_dir):
        """创建复杂测试文件"""
        file_path = temp_dir / "complex_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ComplexSheet"
        
        # 创建复杂数据
        # 数值列
        for i in range(1, 21):
            ws.cell(row=i, column=1, value=i)
        
        # 带小数的数值列
        for i in range(1, 21):
            ws.cell(row=i, column=2, value=i * 1.5)
        
        # 文本列
        for i in range(1, 21):
            ws.cell(row=i, column=3, value=f"Item{i}")
        
        wb.save(file_path)
        return str(file_path)

    def test_bulk_formula_updates(self, complex_test_file):
        """测试批量公式更新"""
        writer = ExcelWriter(complex_test_file)
        
        # 批量更新多行公式
        formulas = [[f"=A{i}+B{i}"] for i in range(1, 11)]
        
        result = writer.update_range(
            "ComplexSheet!D1:D10",
            formulas,
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_mixed_data_formula(self, complex_test_file):
        """测试混合数据公式"""
        writer = ExcelWriter(complex_test_file)
        
        result = writer.update_range(
            "ComplexSheet!E1:E5",
            [
                ["=A1*B1"],
                ["=SUM(A1:A5)"],
                ["=AVERAGE(B1:B5)"],
                ["=IF(A1>10,\"Yes\",\"No\")"],
                ["=COUNTIF(A1:A10,\">5\")"]
            ],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True


class TestExcelWriterDataOperations:
    """ExcelWriter 数据操作测试"""

    @pytest.fixture
    def data_test_file(self, temp_dir):
        """创建数据操作测试文件"""
        file_path = temp_dir / "data_ops_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DataOpsSheet"
        
        # 添加测试数据
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=i)
            ws.cell(row=i, column=2, value=f"Data{i}")
        
        wb.save(file_path)
        return str(file_path)

    def test_insert_rows(self, data_test_file):
        """测试插入行"""
        writer = ExcelWriter(data_test_file)
        result = writer.insert_rows("DataOpsSheet", 5, 2)
        
        assert result.success is True

    def test_delete_rows(self, data_test_file):
        """测试删除行"""
        writer = ExcelWriter(data_test_file)
        result = writer.delete_rows("DataOpsSheet", 5, 1)
        
        assert result.success is True

    def test_insert_columns(self, data_test_file):
        """测试插入列"""
        writer = ExcelWriter(data_test_file)
        result = writer.insert_columns("DataOpsSheet", 3, 1)
        
        assert result.success is True

    def test_delete_columns(self, data_test_file):
        """测试删除列"""
        writer = ExcelWriter(data_test_file)
        result = writer.delete_columns("DataOpsSheet", 2, 1)
        
        assert result.success is True
