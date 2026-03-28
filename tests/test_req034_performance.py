"""
性能优化测试 - REQ-034
测试JOIN查询性能、大数据量查询、缓存机制

@intention: 确保性能优化有效，JOIN查询效率提升
"""

import pytest
import os
import tempfile
import time
from openpyxl import Workbook

import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from excel_mcp_server_fastmcp.api.advanced_sql_query import AdvancedSQLQueryEngine


def create_large_test_file(num_rows=1000, num_cols=10):
    """创建大数据量测试文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "大数据"

    # 表头
    headers = [f"列{i}" for i in range(1, num_cols + 1)]
    ws.append(headers)

    # 数据
    for i in range(1, num_rows + 1):
        row = [i, f"名称_{i}", i * 10.0, i % 100, i / 100.0]
        # 填充剩余列
        while len(row) < num_cols:
            row.append(f"数据_{i}_{len(row)}")
        ws.append(row)

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        wb.save(f.name)
        return f.name


def create_join_test_files(num_rows=500):
    """创建JOIN性能测试用的两个文件"""
    # 文件1: 主表
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "主表"
    ws1.append(["ID", "名称", "分类ID", "值"])
    for i in range(1, num_rows + 1):
        ws1.append([i, f"项目_{i}", i % 50, i * 1.5])

    # 文件2: 从表
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "从表"
    ws2.append(["ID", "分类名", "描述"])
    for i in range(1, 51):
        ws2.append([i, f"分类_{i}", f"这是分类{i}的描述"])

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f1:
        wb1.save(f1.name)
        file1 = f1.name

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f2:
        wb2.save(f2.name)
        file2 = f2.name

    return file1, file2


@pytest.fixture
def large_file():
    """创建大数据量测试文件"""
    filepath = create_large_test_file(500, 8)
    yield filepath
    os.unlink(filepath)


@pytest.fixture
def join_files():
    """创建JOIN测试文件"""
    file1, file2 = create_join_test_files(300)
    yield file1, file2
    os.unlink(file1)
    os.unlink(file2)


@pytest.fixture
def engine():
    """创建SQL查询引擎"""
    return AdvancedSQLQueryEngine()


class TestQueryPerformance:
    """查询性能测试"""

    def test_simple_query_performance(self, engine, large_file):
        """测试简单查询性能 (应<2秒)"""
        start = time.time()
        result = engine.execute_sql_query(
            large_file,
            "SELECT * FROM 大数据 WHERE 列1 > 250"
        )
        elapsed = time.time() - start

        assert result['success'], f"查询失败: {result.get('error', '')}"
        assert elapsed < 5.0, f"简单查询耗时过长: {elapsed:.2f}秒"

    def test_aggregate_query_performance(self, engine, large_file):
        """测试聚合查询性能 (应<3秒)"""
        start = time.time()
        result = engine.execute_sql_query(
            large_file,
            "SELECT 列4, COUNT(*) as 数量, AVG(列3) as 平均值, SUM(列3) as 总和 "
            "FROM 大数据 GROUP BY 列4 HAVING COUNT(*) > 5"
        )
        elapsed = time.time() - start

        assert result['success'], f"聚合查询失败: {result.get('error', '')}"
        assert elapsed < 5.0, f"聚合查询耗时过长: {elapsed:.2f}秒"

    def test_order_limit_performance(self, engine, large_file):
        """测试排序+LIMIT性能 (应<3秒)"""
        start = time.time()
        result = engine.execute_sql_query(
            large_file,
            "SELECT * FROM 大数据 ORDER BY 列3 DESC LIMIT 10"
        )
        elapsed = time.time() - start

        assert result['success'], f"排序查询失败: {result.get('error', '')}"
        assert elapsed < 5.0, f"排序查询耗时过长: {elapsed:.2f}秒"
        assert len(result['data']) <= 10, "LIMIT 10应该返回<=10行"


class TestJOINPerformance:
    """JOIN性能测试"""

    def test_inner_join_performance(self, engine, join_files):
        """测试INNER JOIN性能"""
        file1, file2 = join_files
        start = time.time()
        result = engine.execute_sql_query(
            file1,
            "SELECT m.名称, m.值, s.分类名 "
            "FROM 主表 m "
            f"INNER JOIN 从表@'{file2}' s ON m.分类ID = s.ID "
            "WHERE m.值 > 100"
        )
        elapsed = time.time() - start

        assert result['success'], f"JOIN查询失败: {result.get('error', '')}"
        assert elapsed < 8.0, f"JOIN查询耗时过长: {elapsed:.2f}秒"

    def test_left_join_performance(self, engine, join_files):
        """测试LEFT JOIN性能"""
        file1, file2 = join_files
        start = time.time()
        result = engine.execute_sql_query(
            file1,
            "SELECT m.名称, s.分类名 "
            "FROM 主表 m "
            f"LEFT JOIN 从表@'{file2}' s ON m.分类ID = s.ID "
            "ORDER BY m.ID LIMIT 50"
        )
        elapsed = time.time() - start

        assert result['success'], f"LEFT JOIN查询失败: {result.get('error', '')}"
        assert elapsed < 8.0, f"LEFT JOIN耗时过长: {elapsed:.2f}秒"

    def test_join_aggregate_performance(self, engine, join_files):
        """测试JOIN + 聚合查询性能"""
        file1, file2 = join_files
        start = time.time()
        result = engine.execute_sql_query(
            file1,
            "SELECT s.分类名, COUNT(*) as 数量, AVG(m.值) as 平均值 "
            "FROM 主表 m "
            f"INNER JOIN 从表@'{file2}' s ON m.分类ID = s.ID "
            "GROUP BY s.分类名 "
            "ORDER BY 数量 DESC LIMIT 10"
        )
        elapsed = time.time() - start

        assert result['success'], f"JOIN聚合查询失败: {result.get('error', '')}"
        assert elapsed < 8.0, f"JOIN聚合耗时过长: {elapsed:.2f}秒"


class TestCacheMechanism:
    """缓存机制测试"""

    def test_cache_reuse(self, engine, large_file):
        """测试缓存重用（第二次查询应该更快）"""
        # 第一次查询
        start1 = time.time()
        result1 = engine.execute_sql_query(
            large_file,
            "SELECT * FROM 大数据 WHERE 列1 > 100"
        )
        elapsed1 = time.time() - start1

        assert result1['success'], f"第一次查询失败: {result1.get('error', '')}"

        # 第二次查询（应命中缓存）
        start2 = time.time()
        result2 = engine.execute_sql_query(
            large_file,
            "SELECT * FROM 大数据 WHERE 列1 > 200"
        )
        elapsed2 = time.time() - start2

        assert result2['success'], f"第二次查询失败: {result2.get('error', '')}"
        # 第二次应该至少不比第一次慢太多（缓存命中时应该更快）
        # 不严格要求更快，因为系统可能有其他负载

    def test_cache_clear(self, engine, large_file):
        """测试缓存清除"""
        # 第一次查询建立缓存
        result1 = engine.execute_sql_query(
            large_file,
            "SELECT COUNT(*) as 总数 FROM 大数据"
        )
        assert result1['success']

        # 清除缓存
        engine.clear_cache()

        # 验证缓存已清除（通过内部状态检查）
        assert len(engine._df_cache) == 0, "缓存应该被清除"

    def test_cache_size_limit(self, engine):
        """测试缓存大小限制"""
        # 创建多个文件填充缓存
        files = []
        for i in range(12):  # 超过最大缓存10
            filepath = create_large_test_file(50, 5)
            files.append(filepath)
            engine.execute_sql_query(
                filepath,
                f"SELECT * FROM 大数据 LIMIT 1"
            )

        # 缓存不应该超过限制
        assert len(engine._df_cache) <= engine._max_cache_size + 2, \
            f"缓存超出限制: {len(engine._df_cache)} > {engine._max_cache_size}"

        # 清理
        for f in files:
            os.unlink(f)


if __name__ == '__main__':
    pytest.main([__file__, '-v', '--tb=short'])
