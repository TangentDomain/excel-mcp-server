"""L1 外部真值不变量测试（INV-1 ~ INV-4）。
INV-1: 结果结构一致性 — result 必须包含 success/data/message
INV-2: SQL-SQLite 结果对齐 — 同一 SQL 在 ExcelMCP 和 SQLite 结果一致
INV-3: 文件完整性守恒 — SELECT 不修改文件；写操作只改目标 sheet
INV-4: 行数守恒 — COUNT(*) 返回的行数 = 实际数据行数（不含表头）
"""

from __future__ import annotations

import os
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_delete_query,
    execute_advanced_insert_query,
    execute_advanced_sql_query,
    execute_advanced_update_query,
)
from excel_mcp_server_fastmcp.calibrator.core import cmd_import, cmd_query

from .conftest import (
    assert_failure_safe,
    assert_result_structure,
    dual_header_file,
    empty_file,
    get_data_rows,
    get_headers,
    multi_sheet_file,
    simple_file,
    single_row_file,
    special_char_file,
)

# ============================================================
# INV-1: 结果结构一致性
# ============================================================


class TestINV1ResultStructure:
    """INV-1: result 必须包含 success(bool), data(list), message(str)"""

    def test_select_success_structure(self, simple_file):
        result = execute_advanced_sql_query(simple_file, "SELECT * FROM 数据")
        assert_result_structure(result)

    def test_select_failure_structure(self, simple_file):
        result = execute_advanced_sql_query(simple_file, "SELECT * FROM 不存在的表")
        assert_result_structure(result)

    def test_update_success_structure(self, simple_file):
        result = execute_advanced_update_query(simple_file, "UPDATE 数据 SET Price = 999 WHERE ID = 1")
        # UPDATE 返回 {success, affected_rows, changes, message}
        assert isinstance(result, dict)
        assert "success" in result and isinstance(result["success"], bool)
        assert "message" in result and isinstance(result["message"], str)

    def test_update_failure_structure(self, simple_file):
        result = execute_advanced_update_query(simple_file, "UPDATE 不存在的表 SET X = 1")
        assert isinstance(result, dict)
        assert result["success"] is False
        assert len(result.get("message", "")) > 0

    def test_insert_success_structure(self, simple_file):
        result = execute_advanced_insert_query(
            simple_file,
            "INSERT INTO 数据 (ID, Name, Price, Active, Tags) VALUES (99, '测试', 1, '否', '标签')",
        )
        # INSERT 返回 {success, affected_rows, message}
        assert isinstance(result, dict)
        assert "success" in result and isinstance(result["success"], bool)
        assert "message" in result

    def test_delete_success_structure(self, simple_file):
        result = execute_advanced_delete_query(simple_file, "DELETE FROM 数据 WHERE ID = 99")
        # DELETE 返回 {success, affected_rows, message}
        assert isinstance(result, dict)
        assert "success" in result and isinstance(result["success"], bool)
        assert "message" in result

    def test_delete_failure_structure(self, simple_file):
        result = execute_advanced_delete_query(simple_file, "DELETE FROM 不存在的表 WHERE ID = 1")
        assert isinstance(result, dict)
        assert result["success"] is False
        assert len(result.get("message", "")) > 0


# ============================================================
# INV-2: SQL-SQLite 结果对齐
# ============================================================

# calibrator 使用的临时数据库名
_CAL_DB = "inv_l2_crossval"


def _align_result(excel_result: dict, sqlite_result: dict, tol: float = 0.01) -> bool:
    """比较 ExcelMCP 和 SQLite 结果是否一致。

    - 跳过 SQLite 的 _rowid_ 列（calibrator 自动添加）
    - 浮点容差 tol
    - 空值/None 统一处理
    - 处理行顺序差异（排序后比较）
    """
    import math

    if not excel_result["success"] or not sqlite_result.get("success"):
        return False

    excel_data = excel_result["data"]
    sqlite_rows = sqlite_result.get("rows", [])
    sqlite_headers = sqlite_result.get("headers", [])

    if len(excel_data) == 0 and len(sqlite_rows) == 0:
        return True
    if len(excel_data) == 0 or len(sqlite_rows) == 0:
        return False

    # 跳过 _rowid_ 列
    rowid_idx = None
    for idx, h in enumerate(sqlite_headers):
        if h == "_rowid_":
            rowid_idx = idx
            break

    # 构建不含 _rowid_ 的 SQLite 数据
    sqlite_rows_clean = []
    for row in sqlite_rows:
        clean_row = [v for i, v in enumerate(row) if i != rowid_idx]
        sqlite_rows_clean.append(clean_row)

    # 行数一致（Excel 数据含表头，SQLite 数据不含表头）
    if len(excel_data) - 1 != len(sqlite_rows_clean):
        return False

    def _value_key(v):
        """用于排序的值归一化：None < number < string"""
        if v is None:
            return (0, 0)
        try:
            f = float(v)
            if math.isnan(f):
                return (0, 0)
            return (1, float(f))
        except (ValueError, TypeError):
            return (2, str(v))

    def _row_key(row):
        return tuple(_value_key(v) for v in row)

    # 提取数据行并排序
    excel_rows = excel_data[1:]
    sorted_excel = sorted(excel_rows, key=_row_key)
    sorted_sqlite = sorted(sqlite_rows_clean, key=_row_key)

    # 逐值比较
    for erow, srow in zip(sorted_excel, sorted_sqlite):
        if len(erow) != len(srow):
            return False
        for ev, sv in zip(erow, srow):
            if ev is None and sv is None:
                continue
            if ev is None or sv is None:
                ev_str = str(ev).strip() if ev is not None else ""
                sv_str = str(sv).strip() if sv is not None else ""
                if ev_str == "" and sv_str == "":
                    continue
                return False
            try:
                ef = float(ev)
                sf = float(sv)
                if abs(ef - sf) > tol:
                    return False
            except (ValueError, TypeError):
                if str(ev).strip() != str(sv).strip():
                    return False
    return True


class TestINV2SQLSQLiteAlignment:
    """INV-2: 同一 SQL 在 ExcelMCP 和 SQLite 上的结果一致"""

    @pytest.fixture(autouse=True)
    def _setup_calibrator(self, simple_file):
        """导入测试文件到 SQLite"""
        self.file_path = simple_file
        import_result = cmd_import(simple_file, _CAL_DB)
        assert import_result["success"], f"calibrator 导入失败: {import_result}"

    def test_simple_select(self):
        sql = "SELECT * FROM 数据"
        excel_result = execute_advanced_sql_query(self.file_path, sql)
        sqlite_result = cmd_query(_CAL_DB, sql)
        assert _align_result(excel_result, sqlite_result), f"ExcelMCP 和 SQLite 结果不一致\nExcel: {excel_result['data'][:3]}\nSQLite: {sqlite_result.get('rows', [])[:3]}"

    def test_where_clause(self):
        sql = "SELECT * FROM 数据 WHERE ID = 1"
        excel_result = execute_advanced_sql_query(self.file_path, sql)
        sqlite_result = cmd_query(_CAL_DB, sql)
        assert _align_result(excel_result, sqlite_result)

    def test_order_by(self):
        sql = "SELECT * FROM 数据 ORDER BY Price DESC"
        excel_result = execute_advanced_sql_query(self.file_path, sql)
        sqlite_result = cmd_query(_CAL_DB, sql)
        assert _align_result(excel_result, sqlite_result)

    def test_aggregation_count(self):
        sql = "SELECT COUNT(*) FROM 数据"
        excel_result = execute_advanced_sql_query(self.file_path, sql)
        sqlite_result = cmd_query(_CAL_DB, sql)
        assert _align_result(excel_result, sqlite_result)

    def test_aggregation_sum_avg(self):
        sql = "SELECT SUM(Price), AVG(Price) FROM 数据"
        excel_result = execute_advanced_sql_query(self.file_path, sql)
        sqlite_result = cmd_query(_CAL_DB, sql)
        assert _align_result(excel_result, sqlite_result)

    def test_group_by(self):
        sql = "SELECT Active, COUNT(*) FROM 数据 GROUP BY Active"
        excel_result = execute_advanced_sql_query(self.file_path, sql)
        sqlite_result = cmd_query(_CAL_DB, sql)
        assert _align_result(excel_result, sqlite_result)

    def test_limit(self):
        sql = "SELECT * FROM 数据 LIMIT 2"
        excel_result = execute_advanced_sql_query(self.file_path, sql)
        sqlite_result = cmd_query(_CAL_DB, sql)
        assert _align_result(excel_result, sqlite_result)

    def test_like(self):
        sql = "SELECT * FROM 数据 WHERE Name LIKE '%剑%'"
        excel_result = execute_advanced_sql_query(self.file_path, sql)
        sqlite_result = cmd_query(_CAL_DB, sql)
        assert _align_result(excel_result, sqlite_result)

    def test_in_clause(self):
        sql = "SELECT * FROM 数据 WHERE ID IN (1, 3, 5)"
        excel_result = execute_advanced_sql_query(self.file_path, sql)
        sqlite_result = cmd_query(_CAL_DB, sql)
        assert _align_result(excel_result, sqlite_result)

    def test_between(self):
        sql = "SELECT * FROM 数据 WHERE Price BETWEEN 50 AND 200"
        excel_result = execute_advanced_sql_query(self.file_path, sql)
        sqlite_result = cmd_query(_CAL_DB, sql)
        assert _align_result(excel_result, sqlite_result)

    def test_case_when(self):
        sql = "SELECT Name, CASE WHEN Price > 100 THEN '贵' ELSE '便宜' END AS 等级 FROM 数据"
        excel_result = execute_advanced_sql_query(self.file_path, sql)
        sqlite_result = cmd_query(_CAL_DB, sql)
        assert _align_result(excel_result, sqlite_result)


# ============================================================
# INV-3: 文件完整性守恒
# ============================================================


class TestINV3FileIntegrity:
    """INV-3: SELECT 不修改文件；写操作只改目标 sheet"""

    def test_select_does_not_modify_file(self, simple_file, tmp_path):
        """SELECT 不应修改文件的任何内容"""
        # 读取原始文件内容
        wb_before = load_workbook(simple_file)
        rows_before = []
        for row in wb_before.active.iter_rows(values_only=True):
            rows_before.append(row)
        sheet_names_before = set(wb_before.sheetnames)

        # 执行 SELECT
        result = execute_advanced_sql_query(simple_file, "SELECT * FROM 数据 WHERE ID = 1")
        assert result["success"]

        # 重新读取文件
        wb_after = load_workbook(simple_file)
        rows_after = []
        for row in wb_after.active.iter_rows(values_only=True):
            rows_after.append(row)
        sheet_names_after = set(wb_after.sheetnames)

        assert rows_before == rows_after, "SELECT 修改了文件内容"
        assert sheet_names_before == sheet_names_after, "SELECT 修改了 sheet 列表"

    def test_update_only_changes_target_sheet(self, multi_sheet_file, tmp_path):
        """UPDATE 只应修改目标 sheet，其他 sheet 不变"""
        # 读取副表原始内容
        wb_before = load_workbook(multi_sheet_file)
        secondary_rows_before = list(wb_before["副表"].iter_rows(values_only=True))

        # UPDATE 主表
        result = execute_advanced_update_query(multi_sheet_file, "UPDATE 主表 SET Name = 'Charlie' WHERE ID = 1")
        assert result["success"]

        # 验证副表未变
        wb_after = load_workbook(multi_sheet_file)
        secondary_rows_after = list(wb_after["副表"].iter_rows(values_only=True))
        assert secondary_rows_before == secondary_rows_after, "UPDATE 影响了非目标 sheet"

        # 验证主表确实被修改
        primary_data = list(wb_after["主表"].iter_rows(values_only=True))
        # 第二行（索引1）应该是修改后的数据
        assert primary_data[1] == (1, "Charlie"), f"主表未被正确修改: {primary_data[1]}"

    def test_insert_preserves_other_sheets(self, multi_sheet_file, tmp_path):
        """INSERT 不应影响其他 sheet"""
        wb_before = load_workbook(multi_sheet_file)
        secondary_before = list(wb_before["副表"].iter_rows(values_only=True))

        result = execute_advanced_insert_query(
            multi_sheet_file,
            "INSERT INTO 主表 (ID, Name) VALUES (99, 'New')",
        )
        assert result["success"]

        wb_after = load_workbook(multi_sheet_file)
        secondary_after = list(wb_after["副表"].iter_rows(values_only=True))
        assert secondary_before == secondary_after, "INSERT 影响了非目标 sheet"

    def test_delete_preserves_other_sheets(self, multi_sheet_file, tmp_path):
        """DELETE 不应影响其他 sheet"""
        wb_before = load_workbook(multi_sheet_file)
        secondary_before = list(wb_before["副表"].iter_rows(values_only=True))

        result = execute_advanced_delete_query(multi_sheet_file, "DELETE FROM 主表 WHERE ID = 2")
        assert result["success"]

        wb_after = load_workbook(multi_sheet_file)
        secondary_after = list(wb_after["副表"].iter_rows(values_only=True))
        assert secondary_before == secondary_after, "DELETE 影响了非目标 sheet"


# ============================================================
# INV-4: 行数守恒
# ============================================================


class TestINV4RowCount:
    """INV-4: COUNT(*) 返回的行数 = 实际数据行数（不含表头）"""

    def test_count_matches_actual_rows_simple(self, simple_file):
        result = execute_advanced_sql_query(simple_file, "SELECT COUNT(*) FROM 数据")
        assert result["success"]
        # simple_file 有 5 行数据（含 1 行 NULL）
        count_val = result["data"][1][0]
        assert count_val == 5, f"COUNT(*)={count_val}，期望 5"

    def test_count_matches_actual_rows_empty(self, empty_file):
        result = execute_advanced_sql_query(empty_file, "SELECT COUNT(*) FROM 空表")
        assert result["success"]
        count_val = result["data"][1][0]
        assert count_val == 0, f"空表 COUNT(*)={count_val}，期望 0"

    def test_count_matches_actual_rows_single(self, single_row_file):
        result = execute_advanced_sql_query(single_row_file, "SELECT COUNT(*) FROM 单行表")
        assert result["success"]
        count_val = result["data"][1][0]
        assert count_val == 1, f"单行表 COUNT(*)={count_val}，期望 1"

    def test_count_star_vs_count_col(self, simple_file):
        """COUNT(*) >= COUNT(col)，因为 NULL 不计入 COUNT(col)"""
        result_star = execute_advanced_sql_query(simple_file, "SELECT COUNT(*) FROM 数据")
        result_col = execute_advanced_sql_query(simple_file, "SELECT COUNT(Name) FROM 数据")
        assert result_star["success"] and result_col["success"]
        count_star = result_star["data"][1][0]
        count_name = result_col["data"][1][0]
        count_price = execute_advanced_sql_query(simple_file, "SELECT COUNT(Price) FROM 数据")["data"][1][0]
        assert count_star >= count_name, f"COUNT(*)={count_star} < COUNT(Name)={count_name}"
        assert count_star >= count_price, f"COUNT(*)={count_star} < COUNT(Price)={count_price}"

    def test_select_star_row_count(self, simple_file):
        """SELECT * 返回的行数（不含表头）应等于 COUNT(*)"""
        count_result = execute_advanced_sql_query(simple_file, "SELECT COUNT(*) FROM 数据")
        star_result = execute_advanced_sql_query(simple_file, "SELECT * FROM 数据")
        assert count_result["success"] and star_result["success"]
        count_val = count_result["data"][1][0]
        actual_rows = len(star_result["data"]) - 1  # 减去表头
        assert actual_rows == count_val, f"SELECT * 返回 {actual_rows} 行，COUNT(*)={count_val}"
