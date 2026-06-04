"""不变量测试共享 fixture。

提供各类 Excel 测试文件：
- 简单表：单行表头，基本数据类型
- 空表：0 行数据
- 单行表：1 行数据
- 全 NULL 表：所有值为空
- 特殊字符表：中文/emoji/单引号/反斜杠
- 双行表头表：游戏配置常见格式
"""

from __future__ import annotations

import math
import shutil
from pathlib import Path

import pytest
from openpyxl import Workbook

# ============================================================
# 辅助函数
# ============================================================

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "test_data"


def _make_simple_wb() -> Workbook:
    """创建简单测试表：ID(int), Name(str), Price(float), Active(bool), Tags(str)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws.append(["ID", "Name", "Price", "Active", "Tags"])
    ws.append([1, "铁剑", 100.5, "是", "武器,近战"])
    ws.append([2, "火球术", 250.0, "否", "技能,魔法"])
    ws.append([3, "生命药水", 50.0, "是", "消耗品"])
    ws.append([4, None, None, None, None])
    ws.append([5, "O'Brien's Sword", 999.99, "是", "武器,稀有"])
    return wb


def _make_empty_wb() -> Workbook:
    """创建空表：只有表头，0 行数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = "空表"
    ws.append(["ID", "Name", "Value"])
    return wb


def _make_single_row_wb() -> Workbook:
    """创建单行表：1 行数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = "单行表"
    ws.append(["ID", "Name", "Score"])
    ws.append([1, "唯一", 100])
    return wb


def _make_all_null_wb() -> Workbook:
    """创建全 NULL 表：5 行数据但 ColA 值全为空"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Null表"
    ws.append(["ColA", "ColB", "ColC"])
    for i in range(1, 6):
        ws.cell(row=i + 1, column=1, value=None)  # ColA 显式为 None
        ws.cell(row=i + 1, column=2, value=f"b{i}")  # ColB 有值确保行被保留
        ws.cell(row=i + 1, column=3, value=f"c{i}")  # ColC 有值确保行被保留
    return wb


def _make_special_char_wb() -> Workbook:
    """创建特殊字符表：中文/日文/韩文/emoji/单引号/反斜杠"""
    wb = Workbook()
    ws = wb.active
    ws.title = "特殊字符"
    ws.append(["名称", "描述", "备注"])
    ws.append(["武器⚔️", "火球の術", "한국어"])
    ws.append(["O'Brien", "路径\\test", "正常"])
    ws.append(["中文列名测试", '包含"双引号"', "换行\n测试"])
    ws.append(["超级长的字符串" + "X" * 5000, "极小值", 0.000000001])
    ws.append(["负数测试", "极大值", 1e15])
    return wb


def _make_dual_header_wb() -> Workbook:
    """创建双行表头表：游戏配置常见格式"""
    wb = Workbook()
    ws = wb.active
    ws.title = "技能配置"
    # 第一行：中文描述
    ws.append(["技能标识", "技能名称", "基础伤害", "冷却时间", "技能类型"])
    # 第二行：英文字段名
    ws.append(["skill_id", "skill_name", "base_damage", "cooldown", "skill_type"])
    # 第三行开始：数据
    ws.append(["SK001", "烈焰斩", 150, 5, "物理"])
    ws.append(["SK002", "冰霜新星", 200, 8, "魔法"])
    ws.append(["SK003", "治愈之光", 0, 3, "治疗"])
    ws.append(["SK004", "雷霆一击", 300, 10, "物理"])
    return wb


def _make_numbers_wb() -> Workbook:
    """创建数值表：用于聚合/窗口函数测试"""
    wb = Workbook()
    ws = wb.active
    ws.title = "数值"
    ws.append(["Category", "Value", "Rank"])
    ws.append(["A", 10, 1])
    ws.append(["A", 20, 2])
    ws.append(["A", 20, 2])  # 并列值
    ws.append(["B", 30, 1])
    ws.append(["B", 40, 2])
    ws.append(["B", 50, 3])
    return wb


def _make_multi_sheet_wb() -> Workbook:
    """创建多 sheet 表：用于文件完整性测试"""
    wb = Workbook()
    # Sheet1
    ws1 = wb.active
    ws1.title = "主表"
    ws1.append(["ID", "Name"])
    ws1.append([1, "Alice"])
    ws1.append([2, "Bob"])
    # Sheet2
    ws2 = wb.create_sheet("副表")
    ws2.append(["ID", "Extra"])
    ws2.append([1, "extra1"])
    ws2.append([2, "extra2"])
    return wb


def _make_formula_wb() -> Workbook:
    """创建含公式的测试表：A列数值，B列=A*2公式，C列文本"""
    wb = Workbook()
    ws = wb.active
    ws.title = "公式表"
    ws.append(["原始值", "计算值", "名称"])
    ws.append([10, None, "行1"])  # B2 将设公式
    ws.append([20, None, "行2"])
    ws.append([30, None, "行3"])
    ws.append([40, None, "行4"])
    ws.append([50, None, "行5"])
    # 设置公式
    for row in range(2, 7):
        ws.cell(row=row, column=2).value = f"=A{row}*2"
    return wb


def _make_writable_wb() -> Workbook:
    """创建可写测试表：ID(int), Name(str), Price(float), Stock(int), Active(str)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "商品"
    ws.append(["ID", "Name", "Price", "Stock", "Active"])
    ws.append([1, "铁剑", 100.0, 50, "是"])
    ws.append([2, "火球术", 250.0, 30, "否"])
    ws.append([3, "生命药水", 50.0, 100, "是"])
    ws.append([4, "魔法盾", 180.0, 0, "是"])
    ws.append([5, "加速靴", 75.5, 25, "否"])
    ws.append([6, "传说之剑", 999.99, 5, "是"])
    return wb


def _make_skills_wb() -> Workbook:
    """创建技能配置表（跨文件 JOIN 左表）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "技能配置"
    ws.append(["技能ID", "技能名称", "伤害", "类型"])
    ws.append(["S001", "火球术", 150, "法师"])
    ws.append(["S002", "冰霜新星", 200, "法师"])
    ws.append(["S003", "治愈之光", 0, "牧师"])
    ws.append(["S004", "烈焰斩", 300, "战士"])
    ws.append(["S005", "雷霆一击", 250, "战士"])
    return wb


def _make_drops_wb() -> Workbook:
    """创建掉落配置表（跨文件 JOIN 右表）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "掉落配置"
    ws.append(["掉落ID", "关联技能", "掉落物品", "数量"])
    ws.append(["D001", "S001", "火焰精华", 2])
    ws.append(["D002", "S001", "灰烬", 5])
    ws.append(["D003", "S002", "冰晶碎片", 3])
    ws.append(["D004", "S004", "剑刃碎片", 1])
    ws.append(["D005", "S006", "虚无之尘", 10])  # S006 不存在于技能表
    return wb


def _save_wb(wb: Workbook, tmp_dir: Path, name: str) -> str:
    """保存 Workbook 到临时目录，返回路径"""
    path = tmp_dir / name
    wb.save(str(path))
    return str(path)


# ============================================================
# Fixtures
# ============================================================


@pytest.fixture
def inv_tmp_dir(tmp_path):
    """不变量测试专用临时目录"""
    yield tmp_path
    # tmp_path 由 pytest 自动清理


@pytest.fixture
def simple_file(inv_tmp_dir) -> str:
    """简单测试表：5 行数据，含 NULL 行"""
    return _save_wb(_make_simple_wb(), inv_tmp_dir, "simple.xlsx")


@pytest.fixture
def empty_file(inv_tmp_dir) -> str:
    """空表：只有表头"""
    return _save_wb(_make_empty_wb(), inv_tmp_dir, "empty.xlsx")


@pytest.fixture
def single_row_file(inv_tmp_dir) -> str:
    """单行表：1 行数据"""
    return _save_wb(_make_single_row_wb(), inv_tmp_dir, "single_row.xlsx")


@pytest.fixture
def all_null_file(inv_tmp_dir) -> str:
    """全 NULL 表：5 行数据但值全为空"""
    return _save_wb(_make_all_null_wb(), inv_tmp_dir, "all_null.xlsx")


@pytest.fixture
def special_char_file(inv_tmp_dir) -> str:
    """特殊字符表：中文/emoji/日文/韩文/单引号/反斜杠/超长字符串"""
    return _save_wb(_make_special_char_wb(), inv_tmp_dir, "special_char.xlsx")


@pytest.fixture
def dual_header_file(inv_tmp_dir) -> str:
    """双行表头表：游戏配置格式"""
    return _save_wb(_make_dual_header_wb(), inv_tmp_dir, "dual_header.xlsx")


@pytest.fixture
def numbers_file(inv_tmp_dir) -> str:
    """数值表：用于聚合/窗口函数测试（含并列值）"""
    return _save_wb(_make_numbers_wb(), inv_tmp_dir, "numbers.xlsx")


@pytest.fixture
def multi_sheet_file(inv_tmp_dir) -> str:
    """多 sheet 表：用于文件完整性测试"""
    return _save_wb(_make_multi_sheet_wb(), inv_tmp_dir, "multi_sheet.xlsx")


@pytest.fixture
def formula_file(inv_tmp_dir) -> str:
    """含公式的测试表：A列数值，B列=A*2公式，C列文本"""
    return _save_wb(_make_formula_wb(), inv_tmp_dir, "formula.xlsx")


@pytest.fixture
def writable_file(inv_tmp_dir) -> str:
    """可写测试表：ID/Name/Price/Stock/Active，6 行数据"""
    return _save_wb(_make_writable_wb(), inv_tmp_dir, "writable.xlsx")


@pytest.fixture
def skills_file(inv_tmp_dir) -> str:
    """技能配置表（跨文件 JOIN 左表）：技能ID/技能名称/伤害/类型，5 行"""
    return _save_wb(_make_skills_wb(), inv_tmp_dir, "skills.xlsx")


@pytest.fixture
def drops_file(inv_tmp_dir) -> str:
    """掉落配置表（跨文件 JOIN 右表）：掉落ID/关联技能/掉落物品/数量，5 行"""
    return _save_wb(_make_drops_wb(), inv_tmp_dir, "drops.xlsx")


# ============================================================
# 辅助函数（供测试文件 import）
# ============================================================


def assert_result_structure(result: dict) -> None:
    """INV-1: 验证返回值结构 {success, data, message}"""
    assert isinstance(result, dict), f"result 应为 dict，实际为 {type(result)}"
    assert "success" in result, "result 缺少 'success' 键"
    assert "data" in result, "result 缺少 'data' 键"
    assert "message" in result, "result 缺少 'message' 键"
    assert isinstance(result["success"], bool), f"success 应为 bool，实际为 {type(result['success'])}"
    assert isinstance(result["data"], list), f"data 应为 list，实际为 {type(result['data'])}"
    assert isinstance(result["message"], str), f"message 应为 str，实际为 {type(result['message'])}"


def assert_failure_safe(result: dict) -> None:
    """INV-5: success=False 时 data 为空列表，message 非空且无堆栈"""
    if result["success"]:
        return
    assert result["data"] == [], f"失败时 data 应为 []，实际为 {result['data']}"
    assert len(result["message"]) > 0, "失败时 message 不能为空"
    # 不含 Python 堆栈信息
    stack_patterns = ["Traceback", 'File "', "line ", "Exception"]
    msg = result["message"]
    for pattern in stack_patterns:
        assert pattern not in msg, f"错误消息不应包含堆栈信息，发现 '{pattern}'"


def rows_equal(a: list, b: list, tol: float = 0.01) -> bool:
    """比较两行数据是否相等（浮点容差）"""
    if len(a) != len(b):
        return False
    for va, vb in zip(a, b):
        if isinstance(va, float) and isinstance(vb, float):
            if math.isnan(va) and math.isnan(vb):
                continue
            if abs(va - vb) > tol:
                return False
        elif va != vb:
            return False
    return True


def get_data_rows(result: dict) -> list[list]:
    """从 result 中提取数据行（跳过表头行）"""
    if not result["success"] or len(result["data"]) <= 1:
        return []
    return result["data"][1:]


def get_headers(result: dict) -> list:
    """从 result 中提取表头"""
    if not result["success"] or len(result["data"]) == 0:
        return []
    return result["data"][0]


__all__ = [
    "assert_result_structure",
    "assert_failure_safe",
    "rows_equal",
    "get_data_rows",
    "get_headers",
]
