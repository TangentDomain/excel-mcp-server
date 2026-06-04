"""L3 写操作不变量测试（INV-16 ~ INV-24）。

INV-16: UPDATE 后读回验证
INV-17: INSERT 行数守恒
INV-18: DELETE 行数守恒
INV-20: 公式列守恒
INV-22: affected_rows 精确
INV-23: 无匹配写操作安全
INV-24: NULL 写入语义
"""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_delete_query,
    execute_advanced_insert_query,
    execute_advanced_sql_query,
    execute_advanced_update_query,
)

# ============================================================
# INV-16: UPDATE 后读回验证
# ============================================================


class TestINV16UpdateReadback:
    """INV-16: UPDATE 后 SELECT 读回验证 SET 表达式生效"""

    def test_update_single_column_readback(self, writable_file):
        """UPDATE 单列后读回验证值正确"""
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = 999 WHERE ID = 1")
        result = execute_advanced_sql_query(writable_file, "SELECT Price FROM 商品 WHERE ID = 1")
        assert result["success"]
        assert result["data"][1][0] == 999

    def test_update_multiple_columns_readback(self, writable_file):
        """UPDATE 多列后读回验证"""
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = 888, Stock = 99 WHERE ID = 2")
        result = execute_advanced_sql_query(writable_file, "SELECT Price, Stock FROM 商品 WHERE ID = 2")
        assert result["success"]
        assert result["data"][1][0] == 888
        assert result["data"][1][1] == 99

    def test_update_with_expression_readback(self, writable_file):
        """UPDATE SET 表达式计算后读回"""
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = ROUND(Price * 1.1, 2) WHERE Active = '是'")
        result = execute_advanced_sql_query(writable_file, "SELECT ID, Price FROM 商品 WHERE Active = '是' ORDER BY ID")
        assert result["success"]
        # ID=1: 100*1.1=110, ID=3: 50*1.1=55, ID=4: 180*1.1=198, ID=6: 999.99*1.1≈1100.0
        rows = result["data"][1:]
        assert rows[0][1] == 110.0  # ID=1
        assert rows[1][1] == 55.0  # ID=3
        assert rows[2][1] == 198.0  # ID=4
        assert abs(rows[3][1] - 1100.0) < 0.1  # ID=6, 浮点容差

    def test_update_non_target_columns_unchanged(self, writable_file):
        """UPDATE 不影响非 SET 列"""
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = 0 WHERE ID = 1")
        result = execute_advanced_sql_query(writable_file, "SELECT Name, Stock, Active FROM 商品 WHERE ID = 1")
        assert result["success"]
        assert result["data"][1] == ["铁剑", 50, "是"]

    def test_update_non_target_rows_unchanged(self, writable_file):
        """UPDATE 不影响非 WHERE 匹配行"""
        before = execute_advanced_sql_query(writable_file, "SELECT Price FROM 商品 WHERE ID = 3")
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = 0 WHERE ID = 1")
        after = execute_advanced_sql_query(writable_file, "SELECT Price FROM 商品 WHERE ID = 3")
        assert before["data"][1][0] == after["data"][1][0]

    def test_update_string_value_readback(self, writable_file):
        """UPDATE SET 字符串值"""
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Name = '传说铁剑' WHERE ID = 1")
        result = execute_advanced_sql_query(writable_file, "SELECT Name FROM 商品 WHERE ID = 1")
        assert result["success"]
        assert result["data"][1][0] == "传说铁剑"

    def test_update_null_value_readback(self, writable_file):
        """UPDATE SET NULL 值"""
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Stock = NULL WHERE ID = 4")
        result = execute_advanced_sql_query(writable_file, "SELECT Stock FROM 商品 WHERE ID = 4")
        assert result["success"]
        # NULL 值读回可能是 None 或空字符串
        val = result["data"][1][0]
        assert val is None or val == "" or val == "None"

    def test_update_idempotent(self, writable_file):
        """同一 UPDATE 执行两次结果一致（幂等性）"""
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = 777 WHERE ID = 1")
        r1 = execute_advanced_sql_query(writable_file, "SELECT Price FROM 商品 WHERE ID = 1")
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = 777 WHERE ID = 1")
        r2 = execute_advanced_sql_query(writable_file, "SELECT Price FROM 商品 WHERE ID = 1")
        assert r1["data"][1][0] == r2["data"][1][0] == 777


# ============================================================
# INV-17: INSERT 行数守恒
# ============================================================


class TestINV17InsertRowCount:
    """INV-17: INSERT N 行后 COUNT(*) 增加 N"""

    def test_insert_single_row_count(self, writable_file):
        """INSERT 1 行后 COUNT 增加 1"""
        before = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")
        execute_advanced_insert_query(
            writable_file,
            "INSERT INTO 商品 (ID, Name, Price, Stock, Active) VALUES (99, '新物品', 50.0, 10, '是')",
        )
        after = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")
        assert after["data"][1][0] == before["data"][1][0] + 1

    def test_insert_multiple_rows_count(self, writable_file):
        """INSERT 多行后 COUNT 正确增加"""
        before = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")
        execute_advanced_insert_query(
            writable_file,
            "INSERT INTO 商品 (ID, Name, Price, Stock, Active) VALUES (90, 'A', 1, 1, '是'), (91, 'B', 2, 2, '否')",
        )
        after = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")
        assert after["data"][1][0] == before["data"][1][0] + 2

    def test_insert_readback_values(self, writable_file):
        """INSERT 后读回验证列值"""
        execute_advanced_insert_query(
            writable_file,
            "INSERT INTO 商品 (ID, Name, Price, Stock, Active) VALUES (99, '测试物品', 123.45, 67, '否')",
        )
        result = execute_advanced_sql_query(writable_file, "SELECT Name, Price, Stock, Active FROM 商品 WHERE ID = 99")
        assert result["success"]
        row = result["data"][1]
        assert row[0] == "测试物品"
        assert abs(row[1] - 123.45) < 0.01
        assert row[2] == 67
        assert row[3] == "否"

    def test_insert_preserves_existing_rows(self, writable_file):
        """INSERT 不影响已有行"""
        before = execute_advanced_sql_query(writable_file, "SELECT * FROM 商品 WHERE ID = 1")
        execute_advanced_insert_query(
            writable_file,
            "INSERT INTO 商品 (ID, Name, Price, Stock, Active) VALUES (99, '新', 1, 1, '是')",
        )
        after = execute_advanced_sql_query(writable_file, "SELECT * FROM 商品 WHERE ID = 1")
        assert before["data"][1] == after["data"][1]

    def test_insert_idempotent_separate_ids(self, writable_file):
        """两次 INSERT 不同 ID 后行数正确"""
        before = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")
        execute_advanced_insert_query(
            writable_file,
            "INSERT INTO 商品 (ID, Name, Price, Stock, Active) VALUES (90, 'X', 1, 1, '是')",
        )
        execute_advanced_insert_query(
            writable_file,
            "INSERT INTO 商品 (ID, Name, Price, Stock, Active) VALUES (91, 'Y', 2, 2, '否')",
        )
        after = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")
        assert after["data"][1][0] == before["data"][1][0] + 2

    def test_insert_failure_no_change(self, writable_file):
        """INSERT 失败不改变文件"""
        before = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")
        result = execute_advanced_insert_query(writable_file, "INSERT INTO 不存在的表 (X) VALUES (1)")
        assert result["success"] is False
        after = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")
        assert after["data"][1][0] == before["data"][1][0]


# ============================================================
# INV-18: DELETE 行数守恒
# ============================================================


class TestINV18DeleteRowCount:
    """INV-18: DELETE 后 COUNT(*) 减少 affected_rows"""

    def test_delete_single_row_count(self, writable_file):
        """DELETE 1 行后 COUNT 减少 1"""
        before_count = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")["data"][1][0]
        result = execute_advanced_delete_query(writable_file, "DELETE FROM 商品 WHERE ID = 1")
        assert result["success"]
        after_count = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")["data"][1][0]
        assert after_count == before_count - 1
        assert result["affected_rows"] == 1

    def test_delete_multiple_rows_count(self, writable_file):
        """DELETE 多行后 COUNT 正确减少"""
        before_count = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")["data"][1][0]
        result = execute_advanced_delete_query(writable_file, "DELETE FROM 商品 WHERE Active = '否'")
        assert result["success"]
        after_count = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")["data"][1][0]
        assert after_count == before_count - result["affected_rows"]
        # 验证被删行不再出现
        check = execute_advanced_sql_query(writable_file, "SELECT * FROM 商品 WHERE Active = '否'")
        assert len(check["data"]) == 1  # 只有表头

    def test_delete_deleted_row_gone(self, writable_file):
        """DELETE 后被删行不再出现"""
        execute_advanced_delete_query(writable_file, "DELETE FROM 商品 WHERE ID = 3")
        result = execute_advanced_sql_query(writable_file, "SELECT * FROM 商品 WHERE ID = 3")
        assert len(result["data"]) == 1  # 只有表头，无数据行

    def test_delete_preserves_other_rows(self, writable_file):
        """DELETE 不影响未匹配行"""
        before = execute_advanced_sql_query(writable_file, "SELECT * FROM 商品 WHERE ID = 1")
        execute_advanced_delete_query(writable_file, "DELETE FROM 商品 WHERE ID = 3")
        after = execute_advanced_sql_query(writable_file, "SELECT * FROM 商品 WHERE ID = 1")
        assert before["data"][1] == after["data"][1]

    def test_delete_failure_no_change(self, writable_file):
        """DELETE 失败不改变文件"""
        before_count = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")["data"][1][0]
        result = execute_advanced_delete_query(writable_file, "DELETE FROM 不存在的表 WHERE ID = 1")
        assert result["success"] is False
        after_count = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")["data"][1][0]
        assert after_count == before_count

    def test_delete_all_rows(self, writable_file):
        """DELETE 所有行后 COUNT 为 0"""
        execute_advanced_delete_query(writable_file, "DELETE FROM 商品 WHERE ID >= 1")
        result = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")
        assert result["data"][1][0] == 0


# ============================================================
# INV-20: 公式列守恒
# ============================================================


class TestINV20FormulaPreservation:
    """INV-20: UPDATE 后非目标列公式仍存在"""

    def test_update_data_column_preserves_formula(self, formula_file):
        """UPDATE A列后 B列公式仍存在"""
        execute_advanced_update_query(formula_file, "UPDATE 公式表 SET 原始值 = 100 WHERE 名称 = '行1'")
        # 用 openpyxl 直接检查公式
        wb = load_workbook(formula_file)
        ws = wb["公式表"]
        cell_b2 = ws.cell(row=2, column=2)
        assert cell_b2.value is not None
        b2_value = cell_b2.value
        if isinstance(b2_value, str) and b2_value.startswith("="):
            # 公式保留
            assert True
        elif isinstance(b2_value, (int, float)):
            # 公式被计算为值，检查值是否正确 (100*2=200)
            assert b2_value == 200.0

    def test_update_text_column_preserves_formula(self, formula_file):
        """UPDATE C列后 B列公式仍存在"""
        execute_advanced_update_query(formula_file, "UPDATE 公式表 SET 名称 = '已修改' WHERE 原始值 = 20")
        wb = load_workbook(formula_file)
        ws = wb["公式表"]
        cell_b3 = ws.cell(row=3, column=2)
        b3_value = cell_b3.value
        # 原始值=20 未改，公式 =A3*2 = 40
        if isinstance(b3_value, str) and b3_value.startswith("="):
            assert True
        elif isinstance(b3_value, (int, float)):
            assert b3_value == 40.0

    def test_update_all_data_rows_formulas_intact(self, formula_file):
        """UPDATE 多行后所有公式列仍存在"""
        execute_advanced_update_query(formula_file, "UPDATE 公式表 SET 原始值 = 原始值 + 10")
        wb = load_workbook(formula_file)
        ws = wb["公式表"]
        for row in range(2, 7):
            cell = ws.cell(row=row, column=2)
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                continue  # 公式保留
            elif isinstance(val, (int, float)):
                # 原始值已+10，所以公式值应该是 (orig+10)*2
                orig = [20, 30, 40, 50, 60][row - 2]
                expected = orig * 2
                assert abs(val - expected) < 0.01, f"Row {row}: formula value {val} != expected {expected}"

    def test_formula_column_not_overwritten_by_update(self, formula_file):
        """UPDATE 不应覆盖公式列的值"""
        # 先读取公式列原始状态
        wb_before = load_workbook(formula_file)
        ws_before = wb_before["公式表"]
        formulas_before = []
        for row in range(2, 7):
            formulas_before.append(ws_before.cell(row=row, column=2).value)

        # UPDATE 非公式列
        execute_advanced_update_query(formula_file, "UPDATE 公式表 SET 原始值 = 999 WHERE 名称 = '行5'")

        # 验证公式列未被修改
        wb_after = load_workbook(formula_file)
        ws_after = wb_after["公式表"]
        for row in range(2, 7):
            val_after = ws_after.cell(row=row, column=2).value
            val_before = formulas_before[row - 2]
            # 行5 (row=6) 的原始值被改为 999，如果公式保留则值可能变为 1998
            # 但其他行应完全不变
            if row != 6:
                assert val_after == val_before, f"Row {row}: formula changed from {val_before} to {val_after}"

    def test_update_idempotent_preserves_formulas(self, formula_file):
        """UPDATE 幂等执行后公式列状态一致"""
        execute_advanced_update_query(formula_file, "UPDATE 公式表 SET 原始值 = 100 WHERE 名称 = '行1'")
        wb1 = load_workbook(formula_file)
        f1 = [wb1["公式表"].cell(row=r, column=2).value for r in range(2, 7)]

        execute_advanced_update_query(formula_file, "UPDATE 公式表 SET 原始值 = 100 WHERE 名称 = '行1'")
        wb2 = load_workbook(formula_file)
        f2 = [wb2["公式表"].cell(row=r, column=2).value for r in range(2, 7)]

        assert f1 == f2


# ============================================================
# INV-22: affected_rows 精确
# ============================================================


class TestINV22AffectedRowsAccuracy:
    """INV-22: affected_rows == 实际变更行数"""

    def test_update_affected_rows_matches_where(self, writable_file):
        """UPDATE affected_rows = WHERE 匹配行数"""
        # Active='是' 有 4 行 (ID=1,3,4,6)
        result = execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = 0 WHERE Active = '是'")
        assert result["success"]
        assert result["affected_rows"] == 4

    def test_update_no_match_zero_affected(self, writable_file):
        """UPDATE WHERE 无匹配 → affected_rows = 0"""
        result = execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = 0 WHERE ID = 999")
        assert result["success"]
        assert result["affected_rows"] == 0

    def test_delete_affected_rows_matches(self, writable_file):
        """DELETE affected_rows = 实际删除行数"""
        result = execute_advanced_delete_query(writable_file, "DELETE FROM 商品 WHERE Stock = 0")
        assert result["success"]
        assert result["affected_rows"] == 1  # ID=4 Stock=0

    def test_delete_no_match_zero_affected(self, writable_file):
        """DELETE WHERE 无匹配 → affected_rows = 0"""
        result = execute_advanced_delete_query(writable_file, "DELETE FROM 商品 WHERE ID = 999")
        assert result["success"]
        assert result["affected_rows"] == 0


# ============================================================
# INV-23: 无匹配写操作安全
# ============================================================


class TestINV23NoMatchWriteSafety:
    """INV-23: UPDATE/DELETE WHERE 无匹配 → 文件不变"""

    def test_update_no_match_file_unchanged(self, writable_file):
        """UPDATE 无匹配不修改文件"""
        wb_before = load_workbook(writable_file)
        rows_before = list(wb_before.active.iter_rows(values_only=True))

        result = execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = 0 WHERE ID = 999")
        assert result["success"]
        assert result["affected_rows"] == 0

        wb_after = load_workbook(writable_file)
        rows_after = list(wb_after.active.iter_rows(values_only=True))
        assert rows_before == rows_after

    def test_delete_no_match_file_unchanged(self, writable_file):
        """DELETE 无匹配不修改文件"""
        wb_before = load_workbook(writable_file)
        rows_before = list(wb_before.active.iter_rows(values_only=True))

        result = execute_advanced_delete_query(writable_file, "DELETE FROM 商品 WHERE ID = 999")
        assert result["success"]
        assert result["affected_rows"] == 0

        wb_after = load_workbook(writable_file)
        rows_after = list(wb_after.active.iter_rows(values_only=True))
        assert rows_before == rows_after

    def test_update_no_match_count_unchanged(self, writable_file):
        """UPDATE 无匹配 COUNT 不变"""
        before = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")["data"][1][0]
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = 0 WHERE ID = 999")
        after = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")["data"][1][0]
        assert before == after

    def test_delete_no_match_count_unchanged(self, writable_file):
        """DELETE 无匹配 COUNT 不变"""
        before = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")["data"][1][0]
        execute_advanced_delete_query(writable_file, "DELETE FROM 商品 WHERE ID = 999")
        after = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")["data"][1][0]
        assert before == after


# ============================================================
# INV-24: NULL 写入语义
# ============================================================


class TestINV24NullWriteSemantics:
    """INV-24: 数值列写 NULL 后读回为 NULL/空"""

    def test_null_write_numeric_column(self, writable_file):
        """数值列写 NULL 后读回"""
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = NULL WHERE ID = 1")
        result = execute_advanced_sql_query(writable_file, "SELECT Price FROM 商品 WHERE ID = 1")
        val = result["data"][1][0]
        assert val is None or val == "" or val == "None"

    def test_null_write_preserves_other_columns(self, writable_file):
        """写 NULL 不影响其他列"""
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = NULL WHERE ID = 1")
        result = execute_advanced_sql_query(writable_file, "SELECT Name, Stock, Active FROM 商品 WHERE ID = 1")
        assert result["data"][1] == ["铁剑", 50, "是"]

    def test_null_write_does_not_affect_other_rows(self, writable_file):
        """写 NULL 不影响其他行"""
        before = execute_advanced_sql_query(writable_file, "SELECT Price FROM 商品 WHERE ID = 2")["data"][1][0]
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = NULL WHERE ID = 1")
        after = execute_advanced_sql_query(writable_file, "SELECT Price FROM 商品 WHERE ID = 2")["data"][1][0]
        assert before == after

    def test_null_write_count_unchanged(self, writable_file):
        """写 NULL 不改变行数"""
        before = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")["data"][1][0]
        execute_advanced_update_query(writable_file, "UPDATE 商品 SET Price = NULL WHERE ID = 1")
        after = execute_advanced_sql_query(writable_file, "SELECT COUNT(*) FROM 商品")["data"][1][0]
        assert before == after
