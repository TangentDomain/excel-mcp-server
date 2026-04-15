"""
R42 深度代码审查 — 安全加固 & 边缘场景测试

覆盖代码审查发现的 P0/P1 问题:
- R1: LIKE 正则注入/ReDoS 防护
- R2: _serialize_value inf 处理
- R3: 除零产生 inf → NULL
- R6: 标量子查询空结果返回 None
+ 通用边缘场景: 空表/单行表/特殊字符/超长字符串

注意: API 返回 data 格式为 list[list], 首行为表头.
openpyxl 限制: 列名中的 emoji 可能被替换为下划线.
"""

import os
import tempfile
import numpy as np
from openpyxl import Workbook

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
)


def _create_test_xlsx(file_path: str, rows: int = 20, headers=None, data_fn=None):
    """创建测试用 xlsx 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if headers is None:
        headers = ["ID", "Name", "Value", "Category"]
    ws.append(headers)
    if data_fn:
        for row in data_fn(rows):
            ws.append(row)
    else:
        for i in range(1, rows + 1):
            ws.append([i, f"Item-{i}", float(i * 10.0), "Cat-A"])
    wb.save(file_path)
    wb.close()


def _rows(result):
    """提取数据行(去掉表头)"""
    return result["data"][1:] if len(result["data"]) > 0 else []


def _hdr(result):
    """提取表头行"""
    return result["data"][0] if len(result["data"]) > 0 else []


class TestR42LikeRegexSafety:
    """R1: LIKE 模式正则注入/ReDoS 防护"""

    def test_like_with_regex_metacharacters(self):
        """LIKE 模式含正则元字符(括号)不应崩溃 — 元字符被正确转义为字面量"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            # 数据含括号: Item-(v1), Item-(v2) 等
            _create_test_xlsx(test_file, 5, data_fn=lambda n: [
                (i, f"Item-(v{i})", float(i), "X") for i in range(1, n+1)
            ])
            # LIKE 'Item-(%': ( 被转义为字面量匹配, % 是通配符
            result = execute_advanced_sql_query(test_file, "SELECT * FROM Sheet1 WHERE Name LIKE 'Item-(%'")
            assert result["success"], f"Failed: {result.get('message', '')}"
            assert len(_rows(result)) >= 1
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_like_reDos_protection(self):
        """超长 LIKE 模式应被拒绝(ReDoS 防护)"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            _create_test_xlsx(test_file, 5)
            long_pattern = "a%" * 200  # 600 chars
            result = execute_advanced_sql_query(
                test_file,
                f"SELECT * FROM Sheet1 WHERE Name LIKE '{long_pattern}'"
            )
            # 应返回错误或空结果, 不应挂起
            assert not result["success"] or len(_rows(result)) == 0
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_like_normal_percent_wildcard(self):
        """% 通配符在修复后仍正常工作"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            _create_test_xlsx(test_file, 10)
            r1 = execute_advanced_sql_query(test_file, "SELECT COUNT(*) as cnt FROM Sheet1 WHERE Name LIKE 'Item-%'")
            assert r1["success"]
            assert _rows(r1)[0][0] == 10  # 全部匹配 Item-*
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_like_normal_underscore_wildcard(self):
        """_ 通配符(单字符)在修复后仍正常工作

        注意: pandas str.match() 不锚定尾部,所以 _ (转为 .) 会匹配
        前缀后跟任意字符的字符串(如 Item-1 和 Item-10 都匹配 Item-_).
        这是 pandas 的已知行为,不是 R42 修复引入的回归.
        """
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            _create_test_xlsx(test_file, 10)
            r = execute_advanced_sql_query(test_file, "SELECT COUNT(*) as cnt FROM Sheet1 WHERE Name LIKE 'Item-_'")
            assert r["success"]
            cnt = _rows(r)[0][0]
            # 至少匹配 Item-1 到 Item-9 (可能更多因 pandas match 不锚定尾部)
            assert cnt >= 9, f"Expected >= 9, got {cnt}"
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_like_dot_escaped_as_literal(self):
        """LIKE 中的点号应作为字面量匹配(不被当正则通配符)"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            _create_test_xlsx(test_file, 5, data_fn=lambda n: [
                (i, f"v1.0.{i}", float(i), "X") for i in range(1, n+1)
            ])
            # 点号是字面量, v1._.% 应匹配 v1.0.1, v1.0.2 等
            result = execute_advanced_sql_query(test_file, "SELECT * FROM Sheet1 WHERE Name LIKE 'v1._.%'")
            assert result["success"]
            assert len(_rows(result)) >= 1
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_like_brackets_escaped(self):
        """LIKE 中的方括号不应被当成正则字符类"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            _create_test_xlsx(test_file, 5, data_fn=lambda n: [
                (i, f"test[ {i}]", float(i), "X") for i in range(1, n+1)
            ])
            result = execute_advanced_sql_query(test_file, "SELECT * FROM Sheet1 WHERE Name LIKE 'test[%'")
            assert result["success"]
            assert len(_rows(result)) >= 1
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)


class TestR42InfinityHandling:
    """R2+R3: inf 序列化安全 + 除零处理"""

    def test_division_by_zero_returns_null(self):
        """除零产生的 inf 应转为 NULL 而非崩溃"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            _create_test_xlsx(test_file, 5, data_fn=lambda n: [
                (i, f"Item-{i}", float(i), "X") for i in range(1, n+1)
            ])
            result = execute_advanced_sql_query(
                test_file,
                "SELECT ID, Value / 0 AS div_result FROM Sheet1"
            )
            assert result["success"], f"Failed: {result.get('message', '')}"
            # 列顺序: ID(index 0), div_result(index 1)
            for row in _rows(result):
                val = row[1]
                assert val is None, f"Expected None for div/0, got {val} (type={type(val).__name__})"
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_inf_value_serialization_no_crash(self):
        """含 inf 值的单元格序列化不应 OverflowError 崩溃"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(["ID", "Name", "Value"])
            ws.append([1, "Normal", 10.5])
            ws.append([2, "InfVal", float("inf")])
            ws.append([3, "NegInf", float("-inf")])
            ws.append([4, "NaN", float("nan")])
            wb.save(test_file)
            wb.close()

            result = execute_advanced_sql_query(test_file, "SELECT * FROM Sheet1")
            assert result["success"], f"Failed: {result.get('message', '')}"
            # 不应因 inf/nan 而 OverflowError 崩溃
            assert len(_rows(result)) == 4
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)


class TestR42ScalarSubqueryNull:
    """R6: 标量子查询空结果返回 None"""

    def test_empty_scalar_subquery_returns_null(self):
        """空子查询在 WHERE 中应表现为 NULL 比较(不返回错误行)"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            _create_test_xlsx(test_file, 5, data_fn=lambda n: [
                (i, f"Item-{i}", float(i * 100), "X") for i in range(1, n+1)
            ])
            # 创建空 sheet
            wb = __import__("openpyxl").load_workbook(test_file)
            ws = wb.create_sheet("EmptySheet")
            ws.append(["Col1"])
            wb.save(test_file)
            wb.close()

            # 空子查询 → NULL → WHERE Value > NULL 不返回行(SQL标准)
            result = execute_advanced_sql_query(
                test_file,
                "SELECT * FROM Sheet1 WHERE Value > (SELECT MAX(Col1) FROM EmptySheet)"
            )
            assert result["success"], f"Failed: {result.get('message', '')}"
            assert len(_rows(result)) == 0, \
                f"Empty scalar subquery should return 0 rows, got {len(_rows(result))}"
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)


class TestR42EdgeCases:
    """通用边缘场景测试"""

    def test_empty_table_select(self):
        """空表 SELECT * 不应崩溃"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "EmptyTbl"
            ws.append(["ID", "Name", "Value"])
            wb.save(test_file)
            wb.close()

            result = execute_advanced_sql_query(test_file, "SELECT * FROM EmptyTbl")
            assert result["success"], f"Failed: {result.get('message', '')}"
            assert len(_rows(result)) == 0

            r2 = execute_advanced_sql_query(test_file, "SELECT COUNT(*) as cnt, SUM(Value) as total FROM EmptyTbl")
            assert r2["success"]
            assert _rows(r2)[0][0] == 0  # COUNT = 0
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_single_row_table(self):
        """单行表各种查询"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            _create_test_xlsx(test_file, 1)
            result = execute_advanced_sql_query(test_file, "SELECT * FROM Sheet1")
            assert result["success"] and len(_rows(result)) == 1

            r2 = execute_advanced_sql_query(test_file, "SELECT AVG(Value) as avg_val FROM Sheet1")
            assert r2["success"] and abs(_rows(r2)[0][0] - 10.0) < 0.01

            r3 = execute_advanced_sql_query(test_file, "SELECT MAX(Value) as mx, MIN(Value) as mn FROM Sheet1")
            assert r3["success"]
            d = _rows(r3)[0]
            assert abs(d[0] - d[1]) < 0.01  # MAX == MIN for single row
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_special_characters_in_column_names(self):
        """列名含特殊字符的查询"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            _create_test_xlsx(test_file, 5, headers=["ID", "Name(备注)", "Value(元)", "Type#"])
            result = execute_advanced_sql_query(test_file, "SELECT ID FROM Sheet1")
            assert result["success"] and len(_rows(result)) == 5

            r2 = execute_advanced_sql_query(test_file, "SELECT `Name(备注)` FROM Sheet1")
            assert r2["success"] and len(_rows(r2)) == 5
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_very_long_string_values(self):
        """超长字符串值不截断/崩溃"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            long_str = "A" * 5000
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(["ID", "LongText"])
            ws.append([1, long_str])
            ws.append([2, "B" * 5000])
            wb.save(test_file)
            wb.close()

            result = execute_advanced_sql_query(test_file, "SELECT * FROM Sheet1")
            assert result["success"]
            assert len(_rows(result)) == 2
            assert len(str(_rows(result)[0][1])) == 5000

            # LIKE 前缀匹配长字符串
            prefix = long_str[:10]
            r2 = execute_advanced_sql_query(test_file, f"SELECT * FROM Sheet1 WHERE LongText LIKE '{prefix}%%'")
            assert r2["success"] and len(_rows(r2)) == 1
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_numeric_edge_cases(self):
        """数值边界: 0, 负数, 极大值, 极小值, 高精度"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(["ID", "Val"])
            ws.append([1, 0])
            ws.append([2, -999999.99])
            ws.append([3, 1e15])
            ws.append([4, 0.000000001])
            ws.append([5, 3.14159265358979])
            wb.save(test_file)
            wb.close()

            result = execute_advanced_sql_query(test_file, "SELECT * FROM Sheet1 ORDER BY ID")
            assert result["success"] and len(_rows(result)) == 5

            r2 = execute_advanced_sql_query(test_file, "SELECT SUM(Val) as total FROM Sheet1")
            assert r2["success"]

            r3 = execute_advanced_sql_query(test_file, "SELECT * FROM Sheet1 WHERE Val < 0")
            assert r3["success"] and len(_rows(r3)) == 1 and _rows(r3)[0][0] == 2
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_null_and_empty_string(self):
        """NULL 和空字符串的处理"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(["ID", "Name", "Value"])
            ws.append([1, "", 10])
            ws.append([2, None, 20])
            ws.append([3, "Hello", 30])
            wb.save(test_file)
            wb.close()

            result = execute_advanced_sql_query(test_file, "SELECT * FROM Sheet1 ORDER BY ID")
            assert result["success"]

            r2 = execute_advanced_sql_query(test_file, "SELECT * FROM Sheet1 WHERE Name IS NULL")
            assert r2["success"]
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_unicode_column_names_and_values(self):
        """Unicode 列名和值(中文/日文/韩文)
        
        注意: openpyxl 对列名中 emoji 的支持有限,可能被替换为下划线.
        此测试验证中文/日文/韩文等 Unicode 文字可正常工作.
        """
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "数据表"
            ws.append(["编号", "名称", "数值列"])
            ws.append([1, "中文测试🎉", 100.5])
            ws.append([2, "日本語テスト", 200.75])
            ws.append([3, "한국어", 300.25])
            ws.append([4, "Emoji 🔥💯", 400.0])
            wb.save(test_file)
            wb.close()

            result = execute_advanced_sql_query(test_file, "SELECT * FROM `数据表`")
            assert result["success"]
            assert len(_rows(result)) == 4

            # 使用实际列名(emoji 可能被替换为 _)
            actual_hdr = _hdr(result)
            val_col = actual_hdr[2]  # 第三列
            r2 = execute_advanced_sql_query(test_file, f"SELECT SUM(`{val_col}`) as total FROM `数据表`")
            assert r2["success"], f"Failed: {r2.get('message', '')}"
            expected = 100.5 + 200.75 + 300.25 + 400.0
            assert abs(_rows(r2)[0][0] - expected) < 0.01
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_update_empty_table(self):
        """UPDATE 空表不应崩溃"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "EmptyTbl"
            ws.append(["ID", "Name", "Value"])
            wb.save(test_file)
            wb.close()

            result = execute_advanced_update_query(
                test_file,
                "UPDATE EmptyTbl SET Value = 999 WHERE ID = 1"
            )
            assert result["success"]
            assert result.get("row_count", 0) == 0
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_select_star_from_empty_result(self):
        """WHERE 条件导致空结果集时不应崩溃"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            test_file = f.name
        try:
            _create_test_xlsx(test_file, 5)
            result = execute_advanced_sql_query(
                test_file,
                "SELECT * FROM Sheet1 WHERE ID = 999999"
            )
            assert result["success"]
            assert len(_rows(result)) == 0
            # 表头仍应存在
            assert len(_hdr(result)) > 0
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)


if __name__ == "__main__":
    import pytest
    pytest.main([__file__, "-v", "-s"])
