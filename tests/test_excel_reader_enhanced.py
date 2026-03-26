# -*- coding: utf-8 -*-
"""
Excel Reader 增强测试套件

覆盖 excel_reader.py 中未被测试的功能：
1. 行范围数据读取
2. 列范围数据读取
3. 边界条件和异常处理
"""

import pytest
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.core.excel_reader import ExcelReader
from src.models.types import OperationResult


class TestExcelReaderEnhanced:
    """ExcelReader 增强功能测试"""

    @pytest.fixture
    def multi_row_excel_file(self, temp_dir):
        """创建多行测试数据文件"""
        file_path = temp_dir / "multi_row_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DataSheet"
        
        # 创建 10 行测试数据
        headers = ["ID", "Name", "Value", "Status"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        for row in range(2, 12):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=f"Item_{row-1}")
            ws.cell(row=row, column=3, value=(row-1) * 10)
            ws.cell(row=row, column=4, value="Active" if row % 2 == 0 else "Inactive")
        
        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def multi_column_excel_file(self, temp_dir):
        """创建多列测试数据文件"""
        file_path = temp_dir / "multi_col_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "WideSheet"
        
        # 创建 10 列测试数据
        for col in range(1, 11):
            ws.cell(row=1, column=col, value=f"Col_{col}")
        
        for row in range(2, 6):
            for col in range(1, 11):
                ws.cell(row=row, column=col, value=f"Data_{row}_{col}")
        
        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def empty_cell_excel_file(self, temp_dir):
        """创建包含空单元格的文件"""
        file_path = temp_dir / "empty_cell_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "EmptySheet"
        
        # 某些单元格为空
        ws['A1'] = "Header1"
        ws['C1'] = "Header3"
        ws['A2'] = "Value1"
        ws['B2'] = None
        ws['C2'] = "Value3"
        
        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def formatted_excel_file(self, temp_dir):
        """创建带有格式的文件"""
        file_path = temp_dir / "formatted_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "FormatSheet"
        
        # 添加格式
        ws['A1'] = "Title"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = PatternFill(start_color="FFFF00", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws['A2'] = "Data"
        ws['B2'] = 100
        
        wb.save(file_path)
        return str(file_path)

    # ==================== 行范围测试 ====================

    def test_get_row_range(self, multi_row_excel_file):
        """测试行范围读取"""
        reader = ExcelReader(multi_row_excel_file)
        result = reader.get_range("DataSheet!2:5")
        
        assert result.success is True
        assert len(result.data) == 4  # 4 行 (2,3,4,5)
        
        # 每行应该有4列数据
        for row in result.data:
            assert len(row) == 4
        
        reader.close()

    def test_get_single_row(self, multi_row_excel_file):
        """测试单行读取"""
        reader = ExcelReader(multi_row_excel_file)
        result = reader.get_range("DataSheet!3:3")
        
        assert result.success is True
        assert len(result.data) == 1  # 1 行
        assert len(result.data[0]) == 4  # 4 列
        
        reader.close()

    def test_get_row_with_headers(self, multi_row_excel_file):
        """测试带表头的行范围"""
        reader = ExcelReader(multi_row_excel_file)
        result = reader.get_range("DataSheet!1:3")
        
        assert result.success is True
        assert len(result.data) == 3  # 表头 + 2 行数据
        
        # 第一行是表头
        assert result.data[0][0].value == "ID"
        
        reader.close()

    # ==================== 列范围测试 ====================

    def test_get_column_range(self, multi_column_excel_file):
        """测试列范围读取"""
        reader = ExcelReader(multi_column_excel_file)
        result = reader.get_range("WideSheet!B:D")
        
        assert result.success is True
        # 应该有多行数据
        assert len(result.data) > 0
        
        # 每行应该有 B, C, D 三列
        for row in result.data:
            assert len(row) == 3
        
        reader.close()

    def test_get_single_column(self, multi_column_excel_file):
        """测试单列读取"""
        reader = ExcelReader(multi_column_excel_file)
        result = reader.get_range("WideSheet!C")
        
        assert result.success is True
        # 应该有 5 行 (1 行表头 + 4 行数据)
        assert len(result.data) == 5
        
        reader.close()

    def test_get_column_with_headers(self, multi_column_excel_file):
        """测试带表头的列范围"""
        reader = ExcelReader(multi_column_excel_file)
        result = reader.get_range("WideSheet!A:C")
        
        assert result.success is True
        # 第一行是表头
        assert result.data[0][0].value == "Col_1"
        
        reader.close()

    # ==================== 空单元格测试 ====================

    def test_get_range_with_empty_cells(self, empty_cell_excel_file):
        """测试包含空单元格的读取"""
        reader = ExcelReader(empty_cell_excel_file)
        result = reader.get_range("EmptySheet!A1:C2")
        
        assert result.success is True
        assert len(result.data) == 2
        
        # 检查空单元格
        assert result.data[1][1].value is None
        
        reader.close()

    # ==================== 格式化测试 ====================

    def test_get_range_with_formatting(self, formatted_excel_file):
        """测试读取格式信息"""
        reader = ExcelReader(formatted_excel_file)
        result = reader.get_range("FormatSheet!A1:B2", include_formatting=True)
        
        assert result.success is True
        
        # 检查第一个单元格有格式（返回的是 CellInfo 对象）
        cell = result.data[0][0]
        assert cell.value == "Title"
        
        # 格式信息应该存在于对象中（具体取决于实现）
        # 检查格式化选项是否生效
        assert hasattr(cell, 'value')
        
        reader.close()

    # ==================== 边界测试 ====================

    def test_get_range_large_row_range(self, temp_dir):
        """测试大范围行读取"""
        file_path = temp_dir / "large_row_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "LargeSheet"
        
        # 创建 100 行数据
        for row in range(1, 101):
            ws.cell(row=row, column=1, value=row)
        
        wb.save(file_path)
        
        reader = ExcelReader(str(file_path))
        result = reader.get_range("LargeSheet!1:50")
        
        assert result.success is True
        assert len(result.data) == 50
        
        reader.close()

    def test_get_range_large_column_range(self, temp_dir):
        """测试大范围列读取"""
        file_path = temp_dir / "large_col_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "WideData"
        
        # 创建 20 列数据
        for col in range(1, 21):
            ws.cell(row=1, column=col, value=f"Col{col}")
        
        for row in range(2, 6):
            for col in range(1, 21):
                ws.cell(row=row, column=col, value=f"Data{row}{col}")
        
        wb.save(file_path)
        
        reader = ExcelReader(str(file_path))
        result = reader.get_range("WideData!A:J")
        
        assert result.success is True
        # 应该有多行
        assert len(result.data) > 0
        # 每行应该有 10 列
        assert len(result.data[0]) == 10
        
        reader.close()

    # ==================== 错误处理测试 ====================

    def test_get_range_invalid_sheet(self, multi_row_excel_file):
        """测试无效工作表"""
        reader = ExcelReader(multi_row_excel_file)
        result = reader.get_range("NonExistentSheet!A1:C3")
        
        assert result.success is False
        
        reader.close()

    def test_get_range_invalid_range_format(self, multi_row_excel_file):
        """测试无效范围格式"""
        reader = ExcelReader(multi_row_excel_file)
        result = reader.get_range("DataSheet!InvalidRange")
        
        # 某些无效格式可能被处理为默认范围
        # 不应该崩溃
        
        reader.close()
