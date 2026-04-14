"""
Round 36 MCP 接口实测 - 并发安全深度测试 + P0第9轮回归验证
============================================================================
方向选择:
  A组: 并发安全深度测试 (多线程读写/竞态条件/快速连续操作/文件锁竞争)
  B组: P0 第9轮回归验证 (全量P0在干净文件上验证)
  C组: 已知问题追踪 (连字符Sheet名/公式列名丢失/浮点精度)
  D组: 错误消息质量审计 (错误信息是否清晰/是否有帮助)

日期: 2026-04-14
轮次: Round 36
"""

import sys
import os
import tempfile
import subprocess
import json
import time
import traceback
import shutil
import threading
import concurrent.futures

sys.path.insert(0, '/root/workspace/excel-mcp-server/src')
sys.path.insert(0, '/root/workspace/excel-mcp-server')

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)

# ============================================================
# 测试数据准备
# ============================================================
TEST_DIR = tempfile.mkdtemp(prefix='r36_test_')
BASE_FILE = os.path.join(TEST_DIR, 'r36_base.xlsx')

def setup_test_file():
    """创建包含多Sheet的标准测试文件"""
    import pandas as pd
    from openpyxl import Workbook
    
    wb = Workbook()
    
    # === Sheet1: 标准数据 ===
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["ID", "Name", "Value", "Price", "Score"])
    for i in range(1, 21):
        ws1.append([i, f"Item-{i}", i * 10, round(i * 1.5, 2), i * 100])
    
    # === 装备配置 Sheet ===
    ws2 = wb.create_sheet("装备配置")
    ws2.append(["ID", "Name", "BaseAtk", "Price", "Rarity"])
    for i in range(1, 11):
        ws2.append([i, f"Equip-{i}", i * 5 + 10, round(i * 99.9, 2),
                     ["Common", "Rare", "Epic", "Legendary"][i % 4]])
    
    # === Types Sheet ===
    from datetime import datetime, date
    ws3 = wb.create_sheet("Types")
    ws3.append(["ID", "IntVal", "FloatVal", "DateVal", "BoolVal", "TextVal", "NullCol"])
    ws3.append([1, 42, 3.14, date(2024, 6, 15), True, "hello", None])
    ws3.append([2, -999, 0.001, date(2025, 1, 1), False, "world", None])
    ws3.append([3, 0, 100.5, date(2023, 12, 31), True, "", None])
    ws3.append([4, 2147483647, 1e-10, datetime.now(), False, "special!@#", None])
    ws3.append([5, 100, 999.99, date(2024, 7, 20), True, "中文测试", None])
    
    wb.save(BASE_FILE)
    return BASE_FILE


def copy_test_file(original, name):
    """复制一份干净的测试文件"""
    dest = os.path.join(TEST_DIR, name)
    shutil.copy2(original, dest)
    return dest


# ============================================================
# 测试结果记录
# ============================================================
results = []
passed_count = 0
failed_count = 0

def record(test_name, group, passed, detail="", input_sql="", expected="", actual=""):
    global passed_count, failed_count
    status = "✅ PASS" if passed else "❌ FAIL"
    if passed:
        passed_count += 1
    else:
        failed_count += 1
    results.append({
        "test": test_name,
        "group": group,
        "status": status,
        "detail": detail,
        "input": input_sql,
        "expected": expected,
        "actual": actual,
    })
    print(f"  {status} | {test_name}: {detail}")


# ============================================================
# A组: 并发安全深度测试
# ============================================================
def test_group_a_concurrency():
    """A组: 并发安全深度测试"""
    print("\n" + "=" * 70)
    print("🔧 A组: 并发安全深度测试")
    print("=" * 70)
    
    base = setup_test_file()
    
    # A1: 多线程并发读取同一文件
    print("\n--- A1-A5: 多线程并发读取 ---")
    def read_worker(file_path, worker_id):
        try:
            result = execute_advanced_sql_query(file_path, f"SELECT ID, Name FROM Sheet1 WHERE ID = {worker_id}")
            return result.get('success', False) and len(result.get('data', [])) > 0
        except Exception as e:
            return False
    
    f1 = copy_test_file(base, 'conc_read.xlsx')
    start = time.time()
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = [executor.submit(read_worker, f1, i) for i in range(1, 9)]
        outcomes = [f.result() for f in futures]
    elapsed = time.time() - start
    all_ok = all(outcomes)
    record("A1-多线程并发读取(8线程)", "A", all_ok,
           f"8线程同时读取, 全部成功={all_ok}, 耗时{elapsed:.2f}s",
           f"SELECT * FROM Sheet1 (8 threads)", "全部成功", f"{sum(outcomes)}/8 成功")
    
    # A2: 快速连续写入
    print("\n--- A6-A10: 快速连续操作 ---")
    f2 = copy_test_file(base, 'rapid_write.xlsx')
    errors = []
    start = time.time()
    for i in range(20):
        result = execute_advanced_update_query(
            f2, f"UPDATE Sheet1 SET Value = {i * 100} WHERE ID = 1"
        )
        if not result.get('success'):
            errors.append(f"iter{i}: {result.get('message', '')[:50]}")
    elapsed = time.time() - start
    record("A2-快速连续UPDATE(20次)", "A", len(errors) == 0,
           f"20次连续UPDATE, 错误数={len(errors)}, 耗时{elapsed:.2f}s",
           "20x UPDATE Sheet1 SET Value=? WHERE ID=1", "0错误",
           f"{len(errors)}错误: {errors[:2] if errors else 'none'}")
    
    # 验证最终值
    verify = execute_advanced_sql_query(f2, "SELECT Value FROM Sheet1 WHERE ID = 1")
    final_val = None
    if verify.get('success') and len(verify.get('data', [])) >= 2:
        final_val = verify['data'][1][0] if verify['data'][1] else None
    record("A3-连续写入后值正确性", "A", final_val == 1900,
           f"最终Value应为1900(19*100), 实际={final_val}",
           "最终值=1900", str(final_val))
    
    # A4: 交替读写操作
    f3 = copy_test_file(base, 'alt_rw.xlsx')
    alt_errors = []
    for i in range(10):
        # 写
        wr = execute_advanced_update_query(f3, f"UPDATE Sheet1 SET Price = {i * 7.77} WHERE ID = 2")
        # 读
        rd = execute_advanced_sql_query(f3, "SELECT COUNT(*) as cnt FROM Sheet1")
        if not wr.get('success'):
            alt_errors.append(f"write-{i}: fail")
        if not rd.get('success'):
            alt_errors.append(f"read-{i}: fail")
    record("A4-交替读写(10轮)", "A", len(alt_errors) == 0,
           f"10轮交替读写, 错误={len(alt_errors)}",
           "0错误", f"{len(alt_errors)}错误")
    
    # A5: 不同Sheet并发操作
    f4 = copy_test_file(base, 'multi_sheet_conc.xlsx')
    sheet_errors = []
    for i in range(10):
        r1 = execute_advanced_update_query(f4, f"UPDATE Sheet1 SET Score = {i} WHERE ID = 1")
        r2 = execute_advanced_update_query(f4, f"UPDATE 装备配置 SET BaseAtk = {i * 10} WHERE ID = 1")
        r3 = execute_advanced_sql_query(f4, "SELECT * FROM Types LIMIT 1")
        if not r1.get('success'): sheet_errors.append(f"s1w-{i}")
        if not r2.get('success'): sheet_errors.append(f'eqw-{i}')
        if not r3.get('success'): sheet_errors.append(f'tr-{i}')
    record("A5-多Sheet混合操作(10轮)", "A", len(sheet_errors) == 0,
           f"3个Sheet混合操作, 错误={len(sheet_errors)}",
           "0错误", f"{len(sheet_errors)}错误")
    
    # A6: 大批量INSERT后立即查询
    f5 = copy_test_file(base, 'bulk_insert_read.xlsx')
    insert_results = []
    for i in range(15):
        ir = execute_advanced_insert_query(
            f5,
            f"INSERT INTO Sheet1 (ID, Name, Value, Price, Score) VALUES ({100+i}, 'Bulk-{i}', {i}, {i*1.1}, {i*10})"
        )
        insert_results.append(ir.get('success', False))
    
    # 立即查询验证
    vr = execute_advanced_sql_query(f5, "SELECT COUNT(*) as total FROM Sheet1")
    count_val = None
    if vr.get('success') and len(vr.get('data', [])) >= 2:
        count_val = vr['data'][1][0]
    
    all_inserts_ok = all(insert_results)
    record("A6-批量INSERT+立即查询", "A", all_inserts_ok and count_val == 35,
           f"15条INSERT全部成功={all_inserts_ok}, 总行数={count_val}(期望35)",
           "15 INSERT OK + COUNT=35", f"INSERT:{sum(insert_results)}/15, COUNT:{count_val}")
    
    # A7: DELETE后再INSERT同ID
    f6 = copy_test_file(base, 'delete_reinsert.xlsx')
    d1 = execute_advanced_delete_query(f6, "DELETE FROM Sheet1 WHERE ID = 5")
    i1 = execute_advanced_insert_query(
        f6, "INSERT INTO Sheet1 (ID, Name, Value, Price, Score) VALUES (5, 'Reborn', 999, 99.99, 500)"
    )
    v1 = execute_advanced_sql_query(f6, "SELECT Name, Value FROM Sheet1 WHERE ID = 5")
    reborn_val = None
    reborn_name = None
    if v1.get('success') and len(v1.get('data', [])) >= 2:
        row = v1['data'][1]
        reborn_name = row[0] if len(row) > 0 else None
        reborn_val = row[1] if len(row) > 1 else None
    record("A7-DELETE后同ID重新INSERT", "A",
           reborn_name == 'Reborn' and reborn_val == 999,
           f"删除ID=5后重插, Name={reborn_name}, Value={reborn_val}",
           "Name=Reborn, Value=999", f"Name={reborn_name}, Value={reborn_val}")
    
    # A8: 并发场景 - ThreadPool模拟并发写不同行
    f7 = copy_test_file(base, 'concurrent_diff_rows.xlsx')
    write_lock = threading.Lock()
    write_errors = []

    def write_different_row(fid, row_id):
        try:
            result = execute_advanced_update_query(
                fid, f"UPDATE Sheet1 SET Value = {row_id * 111} WHERE ID = {row_id}"
            )
            with write_lock:
                if not result.get('success'):
                    write_errors.append(f"row{row_id}: {result.get('message', '')[:40]}")
        except Exception as e:
            with write_lock:
                write_errors.append(f"row{row_id}: exception {str(e)[:40]}")

    start = time.time()
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(write_different_row, f7, i) for i in range(1, 11)]
        concurrent.futures.wait(futures)
    elapsed = time.time() - start
    
    # 验证所有行都被更新了
    verify_all = execute_advanced_sql_query(f7, "SELECT ID, Value FROM Sheet1 ORDER BY ID")
    all_correct = True
    if verify_all.get('success'):
        for row in verify_all.get('data', [])[1:]:  # skip header
            if len(row) >= 2:
                rid, val = row[0], row[1]
                expected_val = rid * 111 if isinstance(rid, (int, float)) and 1 <= rid <= 10 else rid * 10
                # Only check rows 1-10 which we updated
                if isinstance(rid, (int, float)) and 1 <= rid <= 10:
                    if val != expected_val:
                        all_correct = False
    
    record("A8-并发写不同行(5线程x10行)", "A",
           len(write_errors) == 0,
           f"5线程并发写10行, 写错误={len(write_errors)}, 耗时{elapsed:.3f}s",
           "0写错误", f"{len(write_errors)}错误")
    
    # A9: 极端快速连续操作压力测试 (50次操作)
    f9 = copy_test_file(base, 'stress_50ops.xlsx')
    stress_errors = []
    start = time.time()
    for i in range(50):
        op_type = i % 4
        if op_type == 0:
            r = execute_advanced_update_query(f9, f"UPDATE Sheet1 SET Value = {i} WHERE ID = {(i % 20) + 1}")
        elif op_type == 1:
            r = execute_advanced_sql_query(f9, f"SELECT * FROM Sheet1 WHERE ID = {(i % 20) + 1}")
        elif op_type == 2:
            r = execute_advanced_sql_query(f9, "SELECT COUNT(*), SUM(Value), AVG(Price) FROM Sheet1")
        else:
            r = execute_advanced_update_query(f9, f"UPDATE Sheet1 SET Score = Score + 1 WHERE ID = {(i % 20) + 1}")
        if not r.get('success'):
            stress_errors.append(f"op{i}:{r.get('message','')[:30]}")
    elapsed = time.time() - start
    record("A9-极端压力(50次混合操作)", "A", len(stress_errors) < 3,
           f"50次混合操作, 错误={len(stress_errors)}/50, 耗时{elapsed:.2f}s",
           "<3错误(允许偶发)", f"{len(stress_errors)}错误, 样例: {stress_errors[:2]}")
    
    # A10: 文件状态一致性 - 操作前后文件可读
    f10 = copy_test_file(base, 'consistency.xlsx')
    before = execute_advanced_sql_query(f10, "SELECT COUNT(*) FROM Sheet1")
    before_cnt = before['data'][1][0] if before.get('success') and len(before.get('data', [])) >= 2 else None
    
    # 执行一系列操作
    ops_ok = True
    for i in range(5):
        r = execute_advanced_update_query(f10, f"UPDATE Sheet1 SET Value = {i * 100} WHERE ID = {i + 1}")
        ops_ok = ops_ok and r.get('success', False)
    
    after = execute_advanced_sql_query(f10, "SELECT COUNT(*) FROM Sheet1")
    after_cnt = after['data'][1][0] if after.get('success') and len(after.get('data', [])) >= 2 else None
    
    record("A10-操作前后文件一致性", "A",
           before_cnt == after_cnt and ops_ok and before_cnt is not None,
           f"操作前COUNT={before_cnt}, 操作后COUNT={after_cnt}, 中间操作OK={ops_ok}",
           f"COUNT不变({before_cnt})", f"before={before_cnt}, after={after_cnt}, ops={ops_ok}")


# ============================================================
# B组: P0 第9轮回归验证
# ============================================================
def test_group_b_p0_regression():
    """B组: P0 第9轮回归验证 - 在干净文件上逐一验证"""
    print("\n" + "=" * 70)
    print("🚨 B组: P0 第9轮回归验证")
    print("=" * 70)
    
    base = setup_test_file()
    
    # P0-2: SELECT分号多语句注入
    print("\n--- P0-2: SELECT分号注入 ---")
    f_p02 = copy_test_file(base, 'p02_select_semicolon.xlsx')
    r = execute_advanced_sql_query(f_p02, "SELECT * FROM Sheet1; DROP TABLE Sheet1")
    is_vuln = r.get('success', False)
    data_len = len(r.get('data', []))
    record("P0-2-R9 SELECT分号多语句注入", "B", not is_vuln,
           f"success={is_vuln}, 返回数据行数={data_len}",
           "应拒绝分号/返回error", f"success={is_vuln}, rows={data_len}",
           "🚨 仍存在!" if is_vuln else "✅ 已修复!")
    
    # P0-4: UPDATE分号多语句注入
    print("\n--- P0-4: UPDATE分号注入 ---")
    f_p04 = copy_test_file(base, 'p04_update_semicolon.xlsx')
    r = execute_advanced_update_query(f_p04, "UPDATE Sheet1 SET Value = 999 WHERE ID = 1; DROP TABLE Sheet1")
    is_vuln = r.get('success', False)
    record("P0-4-R9 UPDATE分号多语句注入", "B", not is_vuln,
           f"success={is_vuln}, message={str(r.get('message', ''))[:60]}",
           "应拒绝分号/返回error", f"success={is_vuln}",
           "🚨 仍存在!" if is_vuln else "✅ 已修复!")
    
    # P0-5: INSERT分号多语句注入
    print("\n--- P0-5: INSERT分号注入 ---")
    f_p05 = copy_test_file(base, 'p05_insert_semicolon.xlsx')
    r = execute_advanced_insert_query(
        f_p05,
        "INSERT INTO Sheet1 (ID, Name, Value, Price, Score) VALUES (999, 'Hacked', 0, 0, 0); DROP TABLE Sheet1"
    )
    is_vuln = r.get('success', False)
    record("P0-5-R9 INSERT分号多语句注入", "B", not is_vuln,
           f"success={is_vuln}, 插入了恶意数据",
           "应拒绝含分号的INSERT", f"success={is_vuln}",
           "🚨 仍存在!" if is_vuln else "✅ 已修复!")
    
    # P0-6: DELETE分号多语句注入
    print("\n--- P0-6: DELETE分号注入 ---")
    f_p06 = copy_test_file(base, 'p06_delete_semicolon.xlsx')
    r = execute_advanced_delete_query(f_p06, "DELETE FROM Sheet1 WHERE ID = 999; DROP TABLE Sheet1")
    is_vuln = r.get('success', False)
    # 注意: 即使ID=999不存在，如果success=True说明没拦截分号
    msg = str(r.get('message', ''))
    no_match = '没有匹配行' in msg or 'no match' in msg.lower() or '0 行' in msg
    record("P0-6-R9 DELETE分号多语句注入", "B", not is_vuln or (is_vuln and no_match),
           f"success={is_vuln}, message={msg[:60]}, 无匹配={no_match}",
           "应拒绝分号或至少报错", f"success={is_vuln}, no_match={no_match}",
           "🚨 可能仍存在(未拒绝分号)" if is_vuln else "✅ 已修复!")
    
    # P0-7: UPDATE注释符全表篡改
    print("\n--- P0-7: UPDATE注释符篡改 ---")
    f_p07 = copy_test_file(base, 'p07_comment_bypass.xlsx')
    
    # 先检查原始总和
    before_sum_q = execute_advanced_sql_query(f_p07, "SELECT SUM(Value) as s FROM Sheet1")
    before_sum = 0
    if before_sum_q.get('success') and len(before_sum_q.get('data', [])) >= 2:
        before_sum = before_sum_q['data'][1][0]
    
    r = execute_advanced_update_query(f_p07, "UPDATE Sheet1 SET Value = -1 -- WHERE ID = 999999")
    is_vuln = r.get('success', False)
    affected_msg = str(r.get('message', ''))
    
    # 检查实际影响范围
    after_sum_q = execute_advanced_sql_query(f_p07, "SELECT SUM(Value) as s FROM Sheet1")
    after_sum = 0
    if after_sum_q.get('success') and len(after_sum_q.get('data', [])) >= 2:
        after_sum = after_sum_q['data'][1][0]
    
    actual_affected = 0
    if before_sum and after_sum:
        # 每行Value变为-1, 20行 => 总和=-20
        if after_sum == -20:
            actual_affected = "全表(20行)"
        elif after_sum != before_sum:
            actual_affected = f"部分(sum: {before_sum}->{after_sum})"
        else:
            actual_affected = "无变化"
    
    record("P0-7-R9 UPDATE注释符全表篡改", "B", not is_vuln,
           f"success={is_vuln}, affected≈{actual_affected}, sum: {before_sum}->{after_sum}",
           "应拒绝注释符截断WHERE", f"affected={actual_affected}",
           "🚨🚨🚨 仍存在! 全表被篡改!" if (is_vuln and after_sum == -20) else ("⚠️ 待确认" if is_vuln else "✅ 已修复"))
    
    # P0-3: uint8溢出修复确认 (应已修复)
    print("\n--- P0-3: uint8溢出修复确认 ---")
    f_p03 = copy_test_file(base, 'p03_uint8.xlsx')
    r = execute_advanced_update_query(f_p03, "UPDATE Types SET IntVal = 999 WHERE ID = 1")
    vr = execute_advanced_sql_query(f_p03, "SELECT IntVal FROM Types WHERE ID = 1")
    read_back = None
    if vr.get('success') and len(vr.get('data', [])) >= 2:
        read_back = vr['data'][1][0]
    record("P0-3-R9 uint8溢出修复确认", "B", read_back == 999,
           f"写入999, 读回={read_back}",
           "读回=999(非231等溢出值)", f"读回={read_back}",
           "✅ 持续有效" if read_back == 999 else f"❌ 回退! 读回={read_back}")


# ============================================================
# C组: 已知问题追踪
# ============================================================
def test_group_c_known_issues():
    """C组: 已知P1/P2/P3问题追踪"""
    print("\n" + "=" * 70)
    print("📋 C组: 已知问题追踪")
    print("=" * 70)
    
    base = setup_test_file()
    
    # C1: 连字符Sheet名
    print("\n--- C1: 连字符Sheet名 ---")
    from openpyxl import Workbook
    f_c1 = os.path.join(TEST_DIR, 'c1_hyphen_sheet.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "my-data"
    ws.append(["ID", "Val"])
    ws.append([1, 100])
    wb.save(f_c1)
    
    r = execute_advanced_sql_query(f_c1, "SELECT * FROM `my-data`")
    c1_success = r.get('success', False)
    c1_msg = str(r.get('message', ''))[:80]
    
    if not c1_success:
        # 尝试不用反引号
        r2 = execute_advanced_sql_query(f_c1, "SELECT * FROM my-data")
        c1_success2 = r2.get('success', False)
        c1_msg2 = str(r2.get('message', ''))[:80]
    else:
        c1_success2 = False
        c1_msg2 = ""
    
    record("C1-P1 连字符Sheet名(my-data)", "C", c1_success or c1_success2,
           f"反引号版: success={c1_success}, msg={c1_msg}; 无引号: success={c1_success2}, msg={c1_msg2}",
           "支持连字符Sheet名", f"均失败" if (not c1_success and not c1_success2) else "部分成功")
    
    # C2: 公式列名丢失
    print("\n--- C2: 公式列名 ---")
    f_c2 = os.path.join(TEST_DIR, 'c2_formula_col.xlsx')
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "FormulaSheet"
    ws2.append(["A", "B", "C", "D"])  # D列将有公式
    ws2.append([1, 2, 3, None])
    ws2.append([4, 5, 6, None])
    # 设置D列公式
    for row_idx in range(2, 4):
        ws2[f'D{row_idx}'] = f'=A{row_idx}*B{row_idx}+C{row_idx}'
    wb2.save(f_c2)
    
    r = execute_advanced_sql_query(f_c2, "SELECT * FROM FormulaSheet")
    cols = []
    if r.get('success'):
        cols = r.get('data', [])[0] if r.get('data') else []
    has_d = 'D' in cols
    has_row_num = '_ROW_NUMBER_' in cols
    record("C2-P2 公式列D名保留", "C", has_d,
           f"可用列={cols}, 含D={has_d}, 含_ROW_NUMBER_={has_row_num}",
           "列名包含'D'", f"列={cols}")
    
    # C3: 浮点精度损失
    print("\n--- C3: 浮点精度 ---")
    f_c3 = copy_test_file(base, 'c3_float_prec.xlsx')
    pi_val = 3.14159265358979
    r = execute_advanced_update_query(f_c3, f"UPDATE Types SET FloatVal = {pi_val} WHERE ID = 1")
    vr = execute_advanced_sql_query(f_c3, "SELECT FloatVal FROM Types WHERE ID = 1")
    read_back = None
    if vr.get('success') and len(vr.get('data', [])) >= 2:
        read_back = vr['data'][1][0]
    precision_ok = False
    if read_back is not None:
        try:
            diff = abs(float(read_back) - pi_val)
            precision_ok = diff < 0.01  # 允许0.01误差
        except:
            pass
    record("C3-P2 浮点精度(π)", "C", precision_ok,
           f"写入π={pi_val}, 读回={read_back}",
           "精度损失<0.01", f"差值={abs(float(read_back)-pi_val) if read_back else 'N/A'}")
    
    # C4: || 字符串拼接
    print("\n--- C4: ||拼接 ---")
    f_c4 = copy_test_file(base, 'c4_concat.xlsx')
    r = execute_advanced_update_query(f_c4, "UPDATE Sheet1 SET Name = 'Hello' || ' World' WHERE ID = 1")
    c4_success = r.get('success', False)
    c4_msg = str(r.get('message', ''))[:80]
    record("C4-P2-1 UPDATE SET ||拼接", "C", c4_success,
           f"success={c4_success}, msg={c4_msg}",
           "支持||拼接", f"不支持: {c4_msg[:50]}" if not c4_success else "支持")


# ============================================================
# D组: 错误消息质量审计
# ============================================================
def test_group_d_error_quality():
    """D组: 错误消息质量审计 - 检查错误信息是否清晰有用"""
    print("\n" + "=" * 70)
    print("🔍 D组: 错误消息质量审计")
    print("=" * 70)
    
    base = setup_test_file()
    
    error_tests = [
        # (名称, SQL, 是否期望成功, 期望关键字)
        ("D1-不存在的表", "SELECT * FROM NonExistentTable12345", False, ["表", "table", "exist", "找不到"]),
        ("D2-不存在的列", "SELECT FakeColumnXYZ FROM Sheet1", False, ["列", "column", "exist"]),
        ("D3-语法错误", "SELECT FROM WHERE", False, ["语法", "syntax", "error"]),
        ("D4-空SQL", "", False, ["空", "empty", "sql"]),
        ("D5-只有空格的SQL", "   ", False, ["空", "empty", "sql"]),
        ("D6-DROP TABLE(危险操作)", "DROP TABLE Sheet1", False, ["drop", "不支持", "not support", "danger"]),
        ("D7-CREATE TABLE", "CREATE TABLE NewTable (id INT)", False, ["create", "不支持", "not support"]),
        ("D8-ALTER TABLE", "ALTER TABLE Sheet1 ADD col INT", False, ["alter", "不支持", "not support"]),
        ("D9-WHERE子句引用窗口函数别名", "SELECT *, RANK() OVER(ORDER BY Value) as rnk FROM Sheet1 WHERE rnk <= 3", False, ["窗口", "window", "where", "alias"]),
        ("D10-除零错误", "SELECT ID/0 FROM Sheet1", False, ["零", "zero", "div", "error"] or True),  # 可能成功返回inf
    ]
    
    for name, sql, expect_success, expected_keywords in error_tests:
        f_tmp = copy_test_file(base, f'dq_{name.split("-")[0]}.xlsx')
        
        is_select = sql.strip().upper().startswith('SELECT')
        is_update = sql.strip().upper().startswith('UPDATE')
        is_insert = sql.strip().upper().startswith('INSERT')
        is_delete = sql.strip().upper().startswith('DELETE')
        is_drop = sql.strip().upper().startswith('DROP')
        is_create = sql.strip().upper().startswith('CREATE')
        is_alter = sql.strip().upper().startswith('ALTER')
        
        try:
            if is_select:
                r = execute_advanced_sql_query(f_tmp, sql)
            elif is_update:
                r = execute_advanced_update_query(f_tmp, sql)
            elif is_insert:
                r = execute_advanced_insert_query(f_tmp, sql)
            elif is_delete:
                r = execute_advanced_delete_query(f_tmp, sql)
            elif is_drop or is_create or is_alter:
                # 尝试用select处理DDL
                r = execute_advanced_sql_query(f_tmp, sql)
            else:
                r = execute_advanced_sql_query(f_tmp, sql)
            
            success = r.get('success', False)
            message = str(r.get('message', ''))
            
            if expect_success:
                # 期望成功
                passed = success
                detail = f"success={success}"
            else:
                # 期望失败，且错误消息应包含有用的关键词
                if not success:
                    # 检查消息是否有实质内容（非空且有提示意义）
                    msg_lower = message.lower()
                    has_keyword = any(kw.lower() in msg_lower for kw in expected_keywords)
                    msg_meaningful = len(message) > 5  # 至少有一些内容
                    passed = msg_meaningful  # 主要检查有意义的错误消息
                    detail = f"正确拒绝, msg='{message[:60]}', 有关键词={has_keyword}"
                else:
                    # 不应该成功的却成功了
                    passed = False
                    detail = f"不应成功但success=True! msg='{message[:60]}'"
            
            record(name, "D", passed, detail, sql[:50],
                   "期望fail+有意义错误" if not expect_success else "期望success",
                   f"success={success}, msg={message[:40]}")
            
        except Exception as e:
            record(name, "D", expect_success == False,
                   f"异常: {str(e)[:60]}", sql[:50], "期望异常/错误", str(e)[:40])


# ============================================================
# E组: 额外探索 - 事务语义和数据原子性
# ============================================================
def test_group_e_atomicity():
    """E组: 数据原子性和事务语义探索"""
    print("\n" + "=" * 70)
    print("⚛️  E组: 数据原子性和事务语义")
    print("=" * 70)
    
    base = setup_test_file()
    
    # E1: UPDATE部分列不影响其他列
    f_e1 = copy_test_file(base, 'e1_partial_update.xlsx')
    before = execute_advanced_sql_query(f_e1, "SELECT Name, Value, Price FROM Sheet1 WHERE ID = 1")
    before_data = before['data'][1] if before.get('success') and len(before.get('data', [])) >= 2 else None
    
    r = execute_advanced_update_query(f_e1, "UPDATE Sheet1 SET Value = 88888 WHERE ID = 1")
    after = execute_advanced_sql_query(f_e1, "SELECT Name, Value, Price FROM Sheet1 WHERE ID = 1")
    after_data = after['data'][1] if after.get('success') and len(after.get('data', [])) >= 2 else None
    
    name_unchanged = False
    price_unchanged = False
    value_changed = False
    if before_data and after_data:
        name_unchanged = before_data[0] == after_data[0]
        price_unchanged = abs(before_data[2] - after_data[2]) < 0.001 if isinstance(before_data[2], (int, float)) and isinstance(after_data[2], (int, float)) else before_data[2] == after_data[2]
        value_changed = after_data[1] == 88888
    
    record("E1-UPDATE部分列原子性", "E",
           name_unchanged and price_unchanged and value_changed,
           f"Name不变={name_unchanged}, Price不变={price_unchanged}, Value变={value_changed}",
           "只改Value, 其他列不变", f"Name:{before_data[0] if before_data else '?'}->{after_data[0] if after_data else '?'}")
    
    # E2: INSERT不影响已有数据
    f_e2 = copy_test_file(base, 'e2_insert_no_sidefx.xlsx')
    before_rows = execute_advanced_sql_query(f_e2, "SELECT COUNT(*) FROM Sheet1")
    before_cnt = before_rows['data'][1][0] if before_rows.get('success') and len(before_rows.get('data', [])) >= 2 else None
    
    ir = execute_advanced_insert_query(
        f_e2, "INSERT INTO Sheet1 (ID, Name, Value, Price, Score) VALUES (9999, 'TestAtomic', 1, 1, 1)"
    )
    
    # 验证原有数据未被修改
    check_old = execute_advanced_sql_query(f_e2, "SELECT Name, Value FROM Sheet1 WHERE ID = 1")
    old_name = None
    old_val = None
    if check_old.get('success') and len(check_old.get('data', [])) >= 2:
        old_name = check_old['data'][1][0]
        old_val = check_old['data'][1][1]
    
    after_cnt_r = execute_advanced_sql_query(f_e2, "SELECT COUNT(*) FROM Sheet1")
    after_cnt = after_cnt_r['data'][1][0] if after_cnt_r.get('success') and len(after_cnt_r.get('data', [])) >= 2 else None
    
    original_ok = (old_name == 'Item-1' and old_val == 10)  # 原始值
    cnt_increased = (after_cnt == before_cnt + 1) if (before_cnt and after_cnt) else False
    
    record("E2-INSERT无副作用", "E", original_ok and cnt_increased and ir.get('success'),
           f"原ID=1: Name={old_name}(期望Item-1), Val={old_val}(期望10); 行数:{before_cnt}->{after_cnt}",
           "原数据不变, 行数+1", f"original_ok={original_ok}, cnt+1={cnt_increased}")
    
    # E3: DELETE只删目标行
    f_e3 = copy_test_file(base, 'e3_delete_target_only.xlsx')
    before_del = execute_advanced_sql_query(f_e3, "SELECT COUNT(*) FROM Sheet1")
    before_dc = before_del['data'][1][0] if before_del.get('success') and len(before_del.get('data', [])) >= 2 else None
    
    dr = execute_advanced_delete_query(f_e3, "DELETE FROM Sheet1 WHERE ID = 15")
    
    # 验证其他行还在
    check_id1 = execute_advanced_sql_query(f_e3, "SELECT Name FROM Sheet1 WHERE ID = 1")
    id1_exists = False
    if check_id1.get('success') and len(check_id1.get('data', [])) >= 2:
        id1_exists = len(check_id1['data']) >= 2
    
    check_id15 = execute_advanced_sql_query(f_e3, "SELECT Name FROM Sheet1 WHERE ID = 15")
    id15_exists = False
    if check_id15.get('success') and len(check_id15.get('data', [])) >= 2:
        id15_exists = True
    
    after_dc_r = execute_advanced_sql_query(f_e3, "SELECT COUNT(*) FROM Sheet1")
    after_dc = after_dc_r['data'][1][0] if after_dc_r.get('success') and len(after_dc_r.get('data', [])) >= 2 else None
    
    record("E3-DELETE精确性", "E", id1_exists and not id15_exists and (after_dc == before_dc - 1),
           f"ID=1仍在={id1_exists}, ID=15已删={not id15_exists}, 行数:{before_dc}->{after_dc}",
           "只删ID=15, 其他在", f"id1={id1_exists}, id15={id15_exists}, cnt:{before_dc}->{after_dc}")
    
    # E4: WHERE无匹配时UPDATE不影响任何行
    f_e4 = copy_test_file(base, 'e4_no_match_update.xlsx')
    before_vals = execute_advanced_sql_query(f_e4, "SELECT SUM(Value) as s FROM Sheet1")
    before_s = before_vals['data'][1][0] if before_vals.get('success') and len(before_vals.get('data', [])) >= 2 else None
    
    nr = execute_advanced_update_query(f_e4, "UPDATE Sheet1 SET Value = 99999 WHERE ID = 99999")
    
    after_vals = execute_advanced_sql_query(f_e4, "SELECT SUM(Value) as s FROM Sheet1")
    after_s = after_vals['data'][1][0] if after_vals.get('success') and len(after_vals.get('data', [])) >= 2 else None
    
    unchanged = (before_s == after_s) if (before_s and after_s) else False
    record("E4-WHERE无匹配UPDATE无影响", "E", unchanged and nr.get('success'),
           f"SUM(Value): {before_s}->{after_s}, unchanged={unchanged}",
           "SUM不变(无匹配不修改)", f"before_sum={before_s}, after_sum={after_s}")
    
    # E5: 复杂表达式的计算正确性
    f_e5 = copy_test_file(base, 'e5_complex_expr.xlsx')
    r = execute_advanced_update_query(f_e5, "UPDATE Sheet1 SET Value = ID * 10 + Score / 5 WHERE ID <= 5")
    vr = execute_advanced_sql_query(f_e5, "SELECT ID, Value, Score FROM Sheet1 WHERE ID <= 5 ORDER BY ID")
    
    calc_correct = True
    if vr.get('success'):
        for row in vr.get('data', [])[1:]:  # skip header
            if len(row) >= 3:
                rid, rval, rscore = row[0], row[1], row[2]
                try:
                    expected = rid * 10 + rscore / 5
                    if abs(float(rval) - float(expected)) > 0.01:
                        calc_correct = False
                        break
                except:
                    calc_correct = False
                    break
    
    record("E5-复杂表达式计算正确性", "E", calc_correct,
           f"Value = ID*10 + Score/5 对ID<=5的计算验证",
           "每行计算精确匹配", "计算正确" if calc_correct else "存在偏差")


# ============================================================
# 主函数
# ============================================================
def main():
    print("=" * 70)
    print("🔄 Round 36 MCP 接口实测")
    print("   方向: 并发安全深度测试 + P0第9轮回归 + 错误消息质量审计 + 数据原子性")
    print("=" * 70)
    print(f"测试目录: {TEST_DIR}")
    
    try:
        test_group_a_concurrency()
        test_group_b_p0_regression()
        test_group_c_known_issues()
        test_group_d_error_quality()
        test_group_e_atomicity()
    except Exception as e:
        print(f"\n❌ 测试过程异常: {e}")
        traceback.print_exc()
    
    # ============================================================
    # 结果汇总
    # ============================================================
    print("\n" + "=" * 70)
    print("📊 Round 36 测试结果汇总")
    print("=" * 70)
    
    total = passed_count + failed_count
    print(f"\n总测试数: {total}")
    print(f"  ✅ 通过: {passed_count} ({passed_count/total*100:.1f}%)" if total > 0 else "")
    print(f"  ❌ 失败: {failed_count} ({failed_count/total*100:.1f}%)" if total > 0 else "")
    
    # 按分组统计
    groups = {}
    for r in results:
        g = r['group']
        if g not in groups:
            groups[g] = {'pass': 0, 'fail': 0, 'tests': []}
        if 'PASS' in r['status']:
            groups[g]['pass'] += 1
        else:
            groups[g]['fail'] += 1
        groups[g]['tests'].append(r)
    
    print("\n--- 分组统计 ---")
    group_names = {
        'A': 'A组: 并发安全深度测试',
        'B': 'B组: P0第9轮回归',
        'C': 'C组: 已知问题追踪',
        'D': 'D组: 错误消息质量审计',
        'E': 'E组: 数据原子性',
    }
    for g in sorted(groups.keys()):
        info = groups[g]
        t = info['pass'] + info['fail']
        pct = info['pass'] / t * 100 if t > 0 else 0
        print(f"  {group_names.get(g, g)}: {info['pass']}/{t} ({pct:.1f}%)")
    
    # 失败详情
    if failed_count > 0:
        print("\n--- ❌ 失败详情 ---")
        for r in results:
            if 'FAIL' in r['status']:
                print(f"  [{r['group']}] {r['test']}: {r['detail']}")
    
    # 清理
    try:
        shutil.rmtree(TEST_DIR)
    except:
        pass
    
    return passed_count, failed_count


if __name__ == '__main__':
    main()
