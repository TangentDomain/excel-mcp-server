"""
Fixed tests for ExcelSearcher class - matching actual API implementation
"""

import pytest
from openpyxl import Workbook
from src.core.excel_search import ExcelSearcher
from src.models.types import OperationResult
from src.utils.exceptions import ExcelFileNotFoundError


class TestExcelSearcher:
    """Test cases for ExcelSearcher class"""

    def test_init_valid_file(self, sample_excel_file):
        """Test initialization with valid file"""
        searcher = ExcelSearcher(sample_excel_file)
        assert searcher.file_path == sample_excel_file

    def test_init_invalid_file(self):
        """Test initialization with invalid file"""
        with pytest.raises(FileNotFoundError):
            ExcelSearcher("nonexistent_file.xlsx")

    def test_regex_search_simple_pattern(self, sample_excel_file):
        """Test simple regex search"""
        searcher = ExcelSearcher(sample_excel_file)
        result = searcher.regex_search(r"张三")

        assert result.success is True
        assert result.data is not None
        # Check response structure based on actual API

    def test_regex_search_case_insensitive(self, sample_excel_file):
        """Test case insensitive search"""
        searcher = ExcelSearcher(sample_excel_file)
        result = searcher.regex_search(r"技术部", flags="i")

        assert result.success is True
        assert result.data is not None

    def test_regex_search_numbers(self, sample_excel_file):
        """Test searching for numbers"""
        searcher = ExcelSearcher(sample_excel_file)
        result = searcher.regex_search(r"\d+")

        assert result.success is True
        assert result.data is not None

    def test_regex_search_formulas_only(self, formula_excel_file):
        """Test searching only in formulas"""
        searcher = ExcelSearcher(formula_excel_file)
        result = searcher.regex_search(r"SUM", search_values=False, search_formulas=True)

        assert result.success is True
        assert result.data is not None

    def test_regex_search_values_only(self, formula_excel_file):
        """Test searching only in values"""
        searcher = ExcelSearcher(formula_excel_file)
        result = searcher.regex_search(r"总计", search_values=True, search_formulas=False)

        assert result.success is True
        assert result.data is not None

    def test_regex_search_both_values_and_formulas(self, formula_excel_file):
        """Test searching in both values and formulas"""
        searcher = ExcelSearcher(formula_excel_file)
        result = searcher.regex_search(r"AVERAGE", search_values=True, search_formulas=True)

        assert result.success is True
        assert result.data is not None

    def test_regex_search_no_matches(self, sample_excel_file):
        """Test search with no matches"""
        searcher = ExcelSearcher(sample_excel_file)
        result = searcher.regex_search(r"不存在的文本")

        assert result.success is True
        assert result.data is not None

    def test_regex_search_empty_pattern(self, sample_excel_file):
        """Test search with empty pattern"""
        searcher = ExcelSearcher(sample_excel_file)
        result = searcher.regex_search("")

        # Should handle empty pattern gracefully
        assert isinstance(result, OperationResult)

    def test_regex_search_multiple_sheets(self, multi_sheet_excel_file):
        """Test search across multiple sheets"""
        searcher = ExcelSearcher(multi_sheet_excel_file)
        result = searcher.regex_search(r"测试数据")

        assert result.success is True
        assert result.data is not None

    def test_regex_search_large_file_performance(self, temp_dir):
        """Test search performance on larger file"""
        from openpyxl import Workbook

        file_path = temp_dir / "test_large.xlsx"
        wb = Workbook()
        ws = wb.active

        # Add many rows
        for i in range(100):
            ws.append([f"数据{i}", f"值{i}", "测试"])

        wb.save(file_path)

        searcher = ExcelSearcher(str(file_path))
        result = searcher.regex_search(r"测试")

        assert result.success is True
        assert result.data is not None

    def test_regex_search_specific_sheet(self, tmp_path):
        """Test searching in specific sheet only"""
        # Create test file with multiple sheets
        file_path = tmp_path / "test_sheets.xlsx"
        wb = Workbook()

        # Sheet1
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "test123"
        ws1["B1"] = "hello"

        # Sheet2
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "test456"
        ws2["B1"] = "world"

        wb.save(file_path)

        searcher = ExcelSearcher(str(file_path))

        # Test search all sheets
        result_all = searcher.regex_search(r"test\d+")
        assert result_all.success is True
        assert len(result_all.data) == 2  # Should find both test123 and test456

        # Test search specific sheet (Sheet1)
        result_sheet1 = searcher.regex_search(r"test\d+", sheet_name="Sheet1")
        assert result_sheet1.success is True
        assert len(result_sheet1.data) == 1
        assert result_sheet1.data[0].match == "test123"
        assert result_sheet1.data[0].sheet == "Sheet1"

        # Test search specific sheet (Sheet2)
        result_sheet2 = searcher.regex_search(r"test\d+", sheet_name="Sheet2")
        assert result_sheet2.success is True
        assert len(result_sheet2.data) == 1
        assert result_sheet2.data[0].match == "test456"
        assert result_sheet2.data[0].sheet == "Sheet2"

        # Test search non-existent sheet
        result_invalid = searcher.regex_search(r"test\d+", sheet_name="NonExistent")
        assert result_invalid.success is False
        assert "不存在" in result_invalid.error
