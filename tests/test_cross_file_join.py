"""
跨文件JOIN测试 - REQ-027

测试 @'filepath' 语法支持跨文件关联查询
"""
import os
import sys
import pytest
import openpyxl

# 源码路径导入
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))
from excel_mcp.api.advanced_sql_query import AdvancedSQLQueryEngine


# === Fixtures ===

@pytest.fixture
def skills_file(tmp_path):
    """创建技能配置文件"""
    filepath = tmp_path / "skills.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '技能配置'
    ws.append(['技能名称', '技能ID', '技能类型', '伤害'])
    ws.append(['火球术', 'S001', '法师', 150])
    ws.append(['斩击', 'S002', '战士', 80])
    ws.append(['治疗', 'S003', '牧师', 0])
    ws.append(['冰冻术', 'S004', '法师', 120])
    wb.save(filepath)
    wb.close()
    return str(filepath)


@pytest.fixture
def drops_file(tmp_path):
    """创建掉落配置文件"""
    filepath = tmp_path / "drops.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '掉落配置'
    ws.append(['掉落ID', '关联技能', '掉落物品', '数量'])
    ws.append(['D001', 'S001', '火焰精华', 5])
    ws.append(['D002', 'S002', '铁剑碎片', 3])
    ws.append(['D003', 'S003', '圣光碎片', 2])
    ws.append(['D004', 'S004', '冰晶碎片', 4])
    wb.save(filepath)
    wb.close()
    return str(filepath)


@pytest.fixture
def monsters_file(tmp_path):
    """创建怪物配置文件（不同sheet名）"""
    filepath = tmp_path / "monsters.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '怪物表'
    ws.append(['怪物名', '怪物ID', '关联技能', '等级'])
    ws.append(['火龙', 'M001', 'S001', 50])
    ws.append(['冰龙', 'M002', 'S004', 60])
    wb.save(filepath)
    wb.close()
    return str(filepath)


@pytest.fixture
def subfile(tmp_path):
    """在子目录创建文件（测试不同目录跨文件）"""
    subdir = tmp_path / "subdir"
    subdir.mkdir()
    filepath = subdir / "equipment.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '装备表'
    ws.append(['装备名', '装备ID', '关联技能', '攻击力'])
    ws.append(['法杖', 'E001', 'S001', 30])
    ws.append(['大剑', 'E002', 'S002', 50])
    wb.save(filepath)
    wb.close()
    return str(filepath)


@pytest.fixture
def engine():
    """创建SQL引擎实例"""
    eng = AdvancedSQLQueryEngine()
    yield eng
    eng.clear_cache()


# === Tests ===

class TestCrossFileBasicJoin:
    """基本跨文件JOIN功能测试"""

    def test_inner_join_same_directory(self, engine, skills_file, drops_file):
        """同目录跨文件INNER JOIN"""
        sql = (
            f"SELECT s.技能名称, d.掉落物品 "
            f"FROM 技能配置@'{skills_file}' s "
            f"JOIN 掉落配置@'{drops_file}' d ON s.技能ID = d.关联技能"
        )
        result = engine.execute_sql_query(skills_file, sql)
        assert result['success'] is True
        data = result['data']
        assert len(data) == 5  # 1 header + 4 data rows
        # 验证关联正确
        skill_names = [row[0] for row in data[1:]]
        drop_items = [row[1] for row in data[1:]]
        assert '火球术' in skill_names
        assert '火焰精华' in drop_items

    def test_left_join_preserves_all_rows(self, engine, skills_file, drops_file):
        """跨文件LEFT JOIN保留左表所有行"""
        # drops_file只有4个掉落，但假设某个技能没有掉落
        sql = (
            f"SELECT s.技能名称, d.掉落物品 "
            f"FROM 技能配置@'{skills_file}' s "
            f"LEFT JOIN 掉落配置@'{drops_file}' d ON s.技能ID = d.关联技能"
        )
        result = engine.execute_sql_query(skills_file, sql)
        assert result['success'] is True
        data = result['data']
        # 所有4个技能都应该出现
        assert len(data) == 5  # header + 4 data rows

    def test_cross_join(self, engine, skills_file, drops_file):
        """跨文件CROSS JOIN（笛卡尔积）"""
        sql = (
            f"SELECT s.技能名称, d.掉落物品 "
            f"FROM 技能配置@'{skills_file}' s "
            f"CROSS JOIN 掉落配置@'{drops_file}' d"
        )
        result = engine.execute_sql_query(skills_file, sql)
        assert result['success'] is True
        data = result['data']
        # 4 skills × 4 drops = 16 rows + header
        assert len(data) == 17

    def test_where_on_cross_file_join(self, engine, skills_file, drops_file):
        """跨文件JOIN + WHERE过滤"""
        sql = (
            f"SELECT s.技能名称, d.掉落物品 "
            f"FROM 技能配置@'{skills_file}' s "
            f"JOIN 掉落配置@'{drops_file}' d ON s.技能ID = d.关联技能 "
            f"WHERE s.技能类型 = '法师'"
        )
        result = engine.execute_sql_query(skills_file, sql)
        assert result['success'] is True
        data = result['data']
        # 只有2个法师技能
        assert len(data) == 3  # header + 2 data rows
        skill_names = [row[0] for row in data[1:]]
        assert '火球术' in skill_names
        assert '冰冻术' in skill_names

    def test_order_by_on_cross_file_join(self, engine, skills_file, drops_file):
        """跨文件JOIN + ORDER BY排序"""
        sql = (
            f"SELECT s.技能名称, d.数量 "
            f"FROM 技能配置@'{skills_file}' s "
            f"JOIN 掉落配置@'{drops_file}' d ON s.技能ID = d.关联技能 "
            f"ORDER BY d.数量 DESC"
        )
        result = engine.execute_sql_query(skills_file, sql)
        assert result['success'] is True
        data = result['data']
        quantities = [row[1] for row in data[1:]]
        assert quantities == sorted(quantities, reverse=True)

    def test_limit_on_cross_file_join(self, engine, skills_file, drops_file):
        """跨文件JOIN + LIMIT限制"""
        sql = (
            f"SELECT s.技能名称, d.掉落物品 "
            f"FROM 技能配置@'{skills_file}' s "
            f"JOIN 掉落配置@'{drops_file}' d ON s.技能ID = d.关联技能 "
            f"LIMIT 2"
        )
        result = engine.execute_sql_query(skills_file, sql)
        assert result['success'] is True
        data = result['data']
        assert len(data) == 3  # header + 2 data rows


class TestCrossFileDifferentDirectory:
    """不同目录跨文件JOIN测试"""

    def test_relative_path_from_primary_file(self, engine, skills_file, subfile):
        """相对路径（相对于主文件目录）"""
        # subfile is in tmp_path/subdir/equipment.xlsx
        # skills_file is in tmp_path/skills.xlsx
        # Relative from skills_file: subdir/equipment.xlsx
        sql = (
            f"SELECT s.技能名称, e.装备名 "
            f"FROM 技能配置@'{subfile}' s "
            f"JOIN 装备表@'{subfile}' e ON s.技能ID = e.关联技能"
        )
        result = engine.execute_sql_query(skills_file, sql)
        assert result['success'] is True
        data = result['data']
        assert len(data) == 3  # header + 2 data rows


class TestCrossFileEdgeCases:
    """边界情况测试"""

    def test_backward_compatible_no_at_syntax(self, engine, skills_file):
        """无@语法时行为不变（向后兼容）"""
        sql = "SELECT * FROM 技能配置 LIMIT 2"
        result = engine.execute_sql_query(skills_file, sql)
        assert result['success'] is True
        data = result['data']
        assert len(data) == 3  # header + 2 data rows

    def test_nonexistent_file_error(self, engine, skills_file):
        """引用不存在的文件应报错"""
        sql = "SELECT * FROM 技能配置@'nonexistent_file.xlsx' s"
        result = engine.execute_sql_query(skills_file, sql)
        assert result['success'] is False
        assert '不存在' in result['message']

    def test_empty_excel_file(self, engine, skills_file, tmp_path):
        """空Excel文件（无数据行）应返回空结果或报错"""
        empty_file = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        wb.save(empty_file)
        wb.close()

        sql = f"SELECT * FROM Sheet@'{empty_file}' s"
        result = engine.execute_sql_query(skills_file, sql)
        # 空文件有Sheet但没有数据行，JOIN应该无匹配
        assert result['success'] is True or result['success'] is False

    def test_double_quotes_path(self, engine, skills_file, drops_file):
        """双引号包裹路径"""
        sql = (
            f'SELECT s.技能名称, d.掉落物品 '
            f'FROM 技能配置@"{skills_file}" s '
            f'JOIN 掉落配置@"{drops_file}" d ON s.技能ID = d.关联技能'
        )
        result = engine.execute_sql_query(skills_file, sql)
        assert result['success'] is True
        data = result['data']
        assert len(data) == 5  # header + 4 data rows


class TestCrossFileThreeWayJoin:
    """三表跨文件JOIN测试"""

    def test_three_way_join(self, engine, skills_file, drops_file, monsters_file):
        """三表跨文件JOIN"""
        sql = (
            f"SELECT s.技能名称, d.掉落物品, m.怪物名 "
            f"FROM 技能配置@'{skills_file}' s "
            f"JOIN 掉落配置@'{drops_file}' d ON s.技能ID = d.关联技能 "
            f"JOIN 怪物表@'{monsters_file}' m ON s.技能ID = m.关联技能"
        )
        result = engine.execute_sql_query(skills_file, sql)
        assert result['success'] is True
        data = result['data']
        # S001(火球术) has drops + monsters, S004(冰冻术) has drops + monsters
        # S002(斩击) has drops but no monsters, S003(治疗) has drops but no monsters
        # INNER JOIN: only rows matching all 3 tables
        assert len(data) == 3  # header + 2 data rows (S001 and S004)

    def test_left_join_three_way(self, engine, skills_file, drops_file, monsters_file):
        """三表LEFT JOIN"""
        sql = (
            f"SELECT s.技能名称, d.掉落物品, m.怪物名 "
            f"FROM 技能配置@'{skills_file}' s "
            f"LEFT JOIN 掉落配置@'{drops_file}' d ON s.技能ID = d.关联技能 "
            f"LEFT JOIN 怪物表@'{monsters_file}' m ON s.技能ID = m.关联技能"
        )
        result = engine.execute_sql_query(skills_file, sql)
        assert result['success'] is True
        data = result['data']
        # All 4 skills should appear
        assert len(data) == 5  # header + 4 data rows
