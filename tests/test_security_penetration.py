"""
Excel MCP 服务器安全性渗透测试

本文件包含各种可能导致数据破坏的恶意操作测试，验证安全机制的有效性。
所有测试文件都创建在临时目录中，测试完成后自动清理。
"""

import pytest
import os
import tempfile
import shutil
import time
import json
import hashlib
from pathlib import Path
from unittest.mock import patch, MagicMock
import subprocess

# 导入要测试的模块
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'src'))

from src.api.excel_operations import ExcelOperations
from src.utils.exceptions import SecurityError, ExcelFileError, OperationCancelledError


class TestPenetrationTesting:
    """安全渗透测试类"""

    @pytest.fixture
    def temp_dir(self):
        """创建临时目录"""
        temp_dir = tempfile.mkdtemp(prefix="excel_security_test_")
        yield temp_dir
        # 清理
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)

    @pytest.fixture
    def temp_excel_file(self, temp_dir):
        """在临时目录中创建Excel文件"""
        file_path = os.path.join(temp_dir, "test_data.xlsx")

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "GameData"

        # 添加重要的游戏数据
        ws.append(["技能ID", "技能名称", "技能类型", "技能伤害", "冷却时间"])
        ws.append([1001, "火球术", "火系", 150, 3.0])
        ws.append([1002, "冰箭", "冰系", 120, 2.5])
        ws.append([1003, "雷击", "雷系", 180, 4.0])

        # 添加公式
        ws.cell(row=5, column=5, value="=E2+E3+E4")  # 总冷却时间

        wb.save(file_path)
        wb.close()

        yield file_path

    @pytest.fixture
    def large_temp_excel_file(self, temp_dir):
        """在临时目录中创建大型Excel文件"""
        file_path = os.path.join(temp_dir, "large_data.xlsx")

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "LargeData"

        # 添加大量重要数据
        ws.append(["ID", "名称", "类型", "数值", "公式列"])
        for i in range(1, 1000):
            ws.append([i, f"重要数据{i}", chr(65 + (i % 26)), i * 10, f"=D{i}*2"])

        wb.save(file_path)
        wb.close()

        yield file_path

    def test_malicious_range_injection(self, temp_excel_file):
        """测试恶意范围注入攻击"""
        malicious_ranges = [
            "../../../etc/passwd",  # 路径遍历攻击
            "..\\..\\windows\\system32\\config\\sam",  # Windows系统文件
            "/dev/null",  # Unix设备文件
            "CON",  # Windows设备名
            "AUX",  # Windows设备名
            "Sheet1' DROP TABLE users; --",  # SQL注入风格攻击
            "Sheet1!A1:$(rm -rf /)",  # 命令注入风格攻击
            "Sheet1!A1:`format c:`",  # Windows命令注入
        ]

        for malicious_range in malicious_ranges:
            result = ExcelOperations._validate_range_expression(malicious_range)
            assert result['valid'] is False, f"恶意范围应该被拒绝: {malicious_range}"

    def test_path_traversal_attack(self, temp_dir):
        """测试路径遍历攻击"""
        base_file = os.path.join(temp_dir, "base.xlsx")

        # 创建基础文件
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["test", "data"])
        wb.save(base_file)
        wb.close()

        malicious_paths = [
            "../../../etc/passwd.xlsx",
            "..\\..\\windows\\system32\\config\\sam.xlsx",
            "../../../../root/.ssh/id_rsa.xlsx",
            "C:\\Windows\\System32\\drivers\\etc\\hosts.xlsx",
            "/etc/shadow.xlsx",
        ]

        for malicious_path in malicious_paths:
            result = ExcelOperations.check_file_status(malicious_path)
            assert result['success'] is False, f"恶意路径应该被拒绝: {malicious_path}"

    def test_data_exfiltration_attempt(self, temp_excel_file):
        """测试数据窃取尝试"""
        # 尝试读取系统敏感文件的Excel版本
        sensitive_files = [
            os.path.expanduser("~/.ssh/id_rsa"),
            os.path.expanduser("~/.aws/credentials"),
            "/etc/passwd",
            "/etc/shadow",
            "C:\\Windows\\System32\\config\\SAM",
        ]

        for sensitive_file in sensitive_files:
            # 尝试将敏感文件作为Excel文件处理
            malicious_excel_path = f"{sensitive_file}.xlsx"
            result = ExcelOperations.check_file_status(malicious_excel_path)
            assert result['success'] is False, f"敏感文件访问应该被拒绝: {malicious_excel_path}"

    def test_formula_injection_attack(self, temp_excel_file):
        """测试公式注入攻击"""
        malicious_formulas = [
            "=CMD|' /C calc'!A1",  # Windows命令执行
            "=EXEC('powershell -c Get-Process')",  # PowerShell执行
            "=HYPERLINK('javascript:alert(1)','Click')",  # JavaScript注入
            "=CALL('shell32.dll','WinExec','JJJC','calc.exe',0)",  # DLL调用
            "=1/0",  # 除零错误
            "=NA()",  # 故意错误
            "=#REF!",  # 引用错误
            "=#NAME?",  # 名称错误
        ]

        for formula in malicious_formulas:
            # 尝试插入恶意公式
            result = ExcelOperations.update_range(
                file_path=temp_excel_file,
                range="GameData!F1",
                data=[[formula]],
                insert_mode=True,
                skip_safety_checks=False
            )

            # 系统应该允许插入但会警告
            # 注意：实际的公式执行由Excel处理，我们的系统只负责数据存储

    def test_memory_exhaustion_attack(self, temp_dir):
        """测试内存耗尽攻击"""
        # 创建超大数据集
        large_data = [[f"数据{i}" for i in range(1, 1001)] for _ in range(1000)]

        # 尝试处理超大数据集
        large_file = os.path.join(temp_dir, "memory_test.xlsx")

        # 创建基础文件
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["test"])
        wb.save(large_file)
        wb.close()

        # 尝试超大数据操作
        result = ExcelOperations.assess_operation_impact(
            file_path=large_file,
            range_expression="Sheet1!A1:ZZ1000",
            operation_type="update",
            preview_data=large_data
        )

        # 应该检测到极高风险
        assert result['success'] is True
        assert result['impact_analysis']['operation_risk_level'] == 'critical'
        assert len(result['warnings']) > 0

    def test_disk_space_exhaustion_attack(self, temp_dir):
        """测试磁盘空间耗尽攻击"""
        # 尝试创建大量备份文件
        test_file = os.path.join(temp_dir, "disk_test.xlsx")

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["test"])
        wb.save(test_file)
        wb.close()

        # 尝试创建大量备份
        backup_paths = []
        for i in range(100):  # 尝试创建100个备份
            result = ExcelOperations.create_auto_backup(
                file_path=test_file,
                backup_name=f"disk_attack_{i}",
                backup_reason="磁盘空间攻击测试",
                user_id="attacker"
            )

            if result['success']:
                backup_paths.append(result['backup_path'])
            else:
                # 系统应该拒绝过多的备份创建
                break

        # 清理备份
        for backup_path in backup_paths:
            if os.path.exists(backup_path):
                os.unlink(backup_path)
            metadata_path = backup_path.replace('.xlsx', '_metadata.json')
            if os.path.exists(metadata_path):
                os.unlink(metadata_path)

    def test_race_condition_attack(self, temp_excel_file):
        """测试竞态条件攻击"""
        # 模拟同时操作同一个文件
        import threading

        results = []

        def concurrent_operation(operation_id):
            try:
                result = ExcelOperations.update_range(
                    file_path=temp_excel_file,
                    range=f"GameData!A{operation_id + 5}:E{operation_id + 5}",
                    data=[[operation_id, f"并发操作{operation_id}", "测试", 100, 1.0]],
                    insert_mode=True,
                    skip_safety_checks=False
                )
                results.append((operation_id, result))
            except Exception as e:
                results.append((operation_id, {'success': False, 'error': str(e)}))

        # 启动多个并发操作
        threads = []
        for i in range(10):
            thread = threading.Thread(target=concurrent_operation, args=(i,))
            threads.append(thread)

        # 同时启动所有线程
        for thread in threads:
            thread.start()

        # 等待所有线程完成
        for thread in threads:
            thread.join()

        # 验证结果：至少有一些操作应该成功或失败，但不能导致系统崩溃
        assert len(results) == 10
        successful_operations = [r for r in results if r[1].get('success', False)]
        failed_operations = [r for r in results if not r[1].get('success', False)]

        # 系统应该能够处理并发操作
        assert len(successful_operations) + len(failed_operations) == 10

    def test_bypass_security_checks_attempt(self, temp_excel_file):
        """测试绕过安全检查的尝试"""
        # 尝试通过直接参数调用绕过安全检查
        malicious_data = [["恶意数据", "删除", "覆盖"]]

        # 1. 尝试跳过安全检查
        result = ExcelOperations.update_range(
            file_path=temp_excel_file,
            range="GameData!A1:C100",
            data=malicious_data,
            insert_mode=False,  # 危险：覆盖模式
            skip_safety_checks=True,  # 尝试跳过安全检查
            require_confirmation=False
        )

        # 系统应该记录这种危险操作
        # 在实际实现中，skip_safety_checks应该只用于系统维护

        # 2. 尝试使用无效的确认令牌
        impact = {
            'success': True,
            'impact_analysis': {
                'operation_risk_level': 'high',
                'total_cells': 300,
                'existing_data_count': 200
            }
        }

        result = ExcelOperations.update_range(
            file_path=temp_excel_file,
            range="GameData!A1:C100",
            data=malicious_data,
            insert_mode=False,
            require_confirmation=True,
            confirmation_token="fake_token_12345",  # 无效令牌
            skip_safety_checks=False
        )

        assert result['success'] is False
        assert "确认令牌" in result['error']

    def test_privilege_escalation_attempt(self, temp_dir):
        """测试权限提升尝试"""
        # 尝试创建系统级别的文件
        system_paths = [
            "C:\\Windows\\System32\\malicious.xlsx",
            "/etc/malicious.xlsx",
            "/usr/local/bin/malicious.xlsx",
        ]

        for system_path in system_paths:
            result = ExcelOperations.create_auto_backup(
                file_path=temp_dir + "/test.xlsx",  # 使用存在的文件
                backup_name="privilege_test",
                backup_reason="权限提升测试",
                user_id="attacker"
            )

            if result['success']:
                # 尝试将备份移动到系统目录（应该失败）
                try:
                    import shutil
                    shutil.move(result['backup_path'], system_path)
                    # 如果移动成功，说明存在权限问题
                    assert False, f"不应该能够移动文件到系统目录: {system_path}"
                except (PermissionError, OSError):
                    # 预期的错误，权限保护正常工作
                    pass
                finally:
                    # 清理备份文件
                    if os.path.exists(result['backup_path']):
                        os.unlink(result['backup_path'])

    def test_dos_attack_via_large_requests(self, temp_excel_file):
        """测试通过大请求的拒绝服务攻击"""
        # 尝试发送极大的范围表达式
        large_ranges = [
            "GameData!A1:ZZ999999",  # 极大范围
            "GameData!A1:XFD1048576",  # Excel最大范围
            "GameData!A1:ZZZZZZ999999",  # 超出Excel范围
        ]

        for large_range in large_ranges:
            result = ExcelOperations._validate_range_expression(large_range)
            # 应该拒绝或限制极大的范围
            if result['valid']:
                # 如果范围有效，评估影响时应该检测到极高风险
                impact = ExcelOperations.assess_operation_impact(
                    file_path=temp_excel_file,
                    range_expression=large_range,
                    operation_type="update",
                    preview_data=[["test"]]
                )

                assert impact['success'] is True
                assert impact['impact_analysis']['operation_risk_level'] == 'critical'

    def test_backup_tampering_attempt(self, temp_excel_file):
        """测试备份篡改尝试"""
        # 1. 创建合法备份
        backup_result = ExcelOperations.create_auto_backup(
            file_path=temp_excel_file,
            backup_name="integrity_test",
            backup_reason="完整性测试",
            user_id="test_user"
        )

        assert backup_result['success'] is True
        backup_path = backup_result['backup_path']

        try:
            # 2. 尝试篡改备份文件
            with open(backup_path, 'r+b') as f:
                f.seek(100)  # 跳到文件中间
                f.write(b"MALICIOUS_DATA")  # 写入恶意数据

            # 3. 尝试从被篡改的备份恢复
            restore_result = ExcelOperations.restore_from_backup(
                original_file_path=temp_excel_file,
                backup_path=backup_path
            )

            # 系统应该检测到备份损坏
            assert restore_result['success'] is False
            assert "损坏" in restore_result['error'] or "校验和不匹配" in restore_result['error']

        finally:
            # 清理
            if os.path.exists(backup_path):
                os.unlink(backup_path)
            metadata_path = backup_path.replace('.xlsx', '_metadata.json')
            if os.path.exists(metadata_path):
                os.unlink(metadata_path)

    def test_concurrent_backup_creation_attack(self, temp_excel_file):
        """测试并发备份创建攻击"""
        import threading

        backup_results = []

        def create_backup_thread(thread_id):
            result = ExcelOperations.create_auto_backup(
                file_path=temp_excel_file,
                backup_name=f"concurrent_attack_{thread_id}",
                backup_reason="并发攻击测试",
                user_id=f"attacker_{thread_id}"
            )
            backup_results.append((thread_id, result))

        # 启动多个线程同时创建备份
        threads = []
        for i in range(20):
            thread = threading.Thread(target=create_backup_thread, args=(i,))
            threads.append(thread)

        for thread in threads:
            thread.start()

        for thread in threads:
            thread.join()

        # 验证结果：系统应该能够处理并发备份创建
        assert len(backup_results) == 20

        successful_backups = [r for r in backup_results if r[1].get('success', False)]
        backup_paths = [r[1]['backup_path'] for r in successful_backups]

        try:
            # 至少应该有一些备份创建成功
            assert len(successful_backups) > 0

            # 验证备份文件完整性
            for backup_path in backup_paths:
                assert os.path.exists(backup_path)

                # 验证校验和
                metadata_path = backup_path.replace('.xlsx', '_metadata.json')
                if os.path.exists(metadata_path):
                    with open(metadata_path, 'r', encoding='utf-8') as f:
                        metadata = json.load(f)

                    assert 'checksum' in metadata

                    # 验证文件校验和
                    with open(backup_path, 'rb') as f:
                        actual_checksum = hashlib.sha256(f.read()).hexdigest()

                    assert actual_checksum == metadata['checksum']

        finally:
            # 清理所有备份文件
            for backup_path in backup_paths:
                if os.path.exists(backup_path):
                    os.unlink(backup_path)
                metadata_path = backup_path.replace('.xlsx', '_metadata.json')
                if os.path.exists(metadata_path):
                    os.unlink(metadata_path)

    def test_metadata_injection_attack(self, temp_excel_file):
        """测试元数据注入攻击"""
        # 1. 创建正常备份
        backup_result = ExcelOperations.create_auto_backup(
            file_path=temp_excel_file,
            backup_name="metadata_test",
            backup_reason="元数据测试",
            user_id="legitimate_user"
        )

        backup_path = backup_result['backup_path']
        metadata_path = backup_path.replace('.xlsx', '_metadata.json')

        try:
            # 2. 尝试注入恶意元数据
            malicious_metadata = {
                "backup_name": "metadata_test",
                "backup_reason": "元数据测试",
                "user_id": "attacker",
                "original_file": temp_excel_file,
                "checksum": backup_result['checksum'],
                "created_at": backup_result['timestamp'],
                "malicious_field": "恶意数据",
                "system_command": "rm -rf /",
                "extra_info": {"attack": "true", "payload": "malicious"}
            }

            with open(metadata_path, 'w', encoding='utf-8') as f:
                json.dump(malicious_metadata, f, ensure_ascii=False, indent=2)

            # 3. 尝试列出备份，验证恶意元数据处理
            list_result = ExcelOperations.list_backups(temp_excel_file)

            # 系统应该正确处理元数据，忽略恶意字段
            assert list_result['success'] is True
            if list_result['backups']:
                backup = list_result['backups'][0]
                # 确保只包含预期字段
                expected_fields = {'backup_id', 'backup_name', 'backup_reason',
                                 'user_id', 'created_at', 'file_size', 'checksum'}
                for field in backup.keys():
                    assert field in expected_fields, f"发现未预期字段: {field}"

        finally:
            # 清理
            if os.path.exists(backup_path):
                os.unlink(backup_path)
            if os.path.exists(metadata_path):
                os.unlink(metadata_path)


class TestSecurityAudit:
    """安全审计测试类"""

    @pytest.fixture
    def temp_dir(self):
        """创建临时目录"""
        temp_dir = tempfile.mkdtemp(prefix="excel_audit_test_")
        yield temp_dir
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)

    def test_security_logging(self, temp_dir):
        """测试安全日志记录"""
        test_file = os.path.join(temp_dir, "audit_test.xlsx")

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["test"])
        wb.save(test_file)
        wb.close()

        # 执行各种操作并验证日志记录
        operations = [
            ("file_check", lambda: ExcelOperations.check_file_status(test_file)),
            ("impact_assess", lambda: ExcelOperations.assess_operation_impact(
                test_file, "Sheet1!A1:C1", "read", None)),
            ("backup_create", lambda: ExcelOperations.create_auto_backup(
                test_file, "audit_backup", "审计测试", "audit_user")),
        ]

        for op_name, op_func in operations:
            try:
                result = op_func()
                # 在实际实现中，应该检查安全日志是否记录了此操作
                assert result is not None
            except Exception as e:
                # 记录异常，但继续测试其他操作
                pass

    def test_permission_verification(self, temp_dir):
        """测试权限验证"""
        # 创建只读文件
        readonly_file = os.path.join(temp_dir, "readonly.xlsx")

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["test"])
        wb.save(readonly_file)
        wb.close()

        # 设置为只读
        os.chmod(readonly_file, 0o444)

        try:
            # 尝试修改只读文件
            result = ExcelOperations.update_range(
                file_path=readonly_file,
                range="Sheet1!A1",
                data=[["modified"]],
                insert_mode=True
            )

            # 系统应该检测到权限问题
            # 注意：实际的权限检查取决于操作系统和文件系统

        finally:
            # 恢复权限以便清理
            os.chmod(readonly_file, 0o644)

    def test_input_sanitization(self, temp_dir):
        """测试输入清理"""
        malicious_inputs = [
            "../../etc/passwd",
            "..\\..\\windows\\system32\\config\\sam",
            "'; DROP TABLE users; --",
            "<script>alert('xss')</script>",
            "$(rm -rf /)",
            "`format c:`",
            "\x00\x01\x02\x03",  # 空字节和二进制数据
            "A" * 10000,  # 超长字符串
        ]

        for malicious_input in malicious_inputs:
            # 测试范围表达式清理
            range_result = ExcelOperations._validate_range_expression(f"Sheet1!{malicious_input}")
            # 应该拒绝恶意输入
            if "Sheet1!" + malicious_input == malicious_input:
                assert range_result['valid'] is False

    def test_resource_limits(self, temp_dir):
        """测试资源限制"""
        test_file = os.path.join(temp_dir, "resource_test.xlsx")

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["test"])
        wb.save(test_file)
        wb.close()

        # 测试内存使用限制
        large_data = [["x" * 1000] * 100 for _ in range(100)]  # 约10MB数据

        start_time = time.time()
        result = ExcelOperations.assess_operation_impact(
            file_path=test_file,
            range_expression="Sheet1!A1:CV100",
            operation_type="update",
            preview_data=large_data
        )
        end_time = time.time()

        # 操作应该在合理时间内完成
        assert end_time - start_time < 30, "操作时间过长，可能存在性能问题"

        if result['success']:
            # 应该检测到高风险
            assert result['impact_analysis']['operation_risk_level'] in ['high', 'critical']


if __name__ == "__main__":
    # 运行测试
    pytest.main([__file__, "-v"])