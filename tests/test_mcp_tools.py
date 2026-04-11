#!/usr/bin/env python3
"""
MCP工具验证测试
验证Excel MCP服务器的核心功能
"""
import json
import sys
import os
from pathlib import Path

# 添加项目路径
project_root = Path(__file__).parent
src_path = project_root / "src"
sys.path.insert(0, str(src_path))

import pytest
from openpyxl import Workbook


@pytest.fixture(scope="module")
def game_excel_file(tmp_path_factory):
    """创建游戏配置测试Excel文件（含skills/items/characters三表）"""
    file_path = tmp_path_factory.mktemp("game_data") / "test_game_data.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # 删除默认Sheet

    # === skills 工作表 ===
    ws_skills = wb.create_sheet("skills")
    ws_skills.append(["技能ID", "名称", "职业", "消耗MP", "冷却时间", "伤害值"])
    ws_skills.append([101, "火球术", "法师", 20, 5, 80])
    ws_skills.append([102, "冰箭", "法师", 15, 3, 60])
    ws_skills.append([103, "剑击", "战士", 0, 0, 50])
    ws_skills.append([104, "治疗术", "牧师", 30, 8, 0])
    ws_skills.append([105, "射击", "弓箭手", 10, 2, 40])

    # === items 工作表 ===
    ws_items = wb.create_sheet("items")
    ws_items.append(["物品ID", "名称", "类型", "稀有度", "价格"])
    ws_items.append([1001, "铁剑", "武器", "普通", 100])
    ws_items.append([1002, "法杖", "武器", "稀有", 500])
    ws_items.append([1003, "皮甲", "防具", "普通", 200])
    ws_items.append([1004, "魔法袍", "防具", "史诗", 1200])
    ws_items.append([1005, "药水", "消耗品", "普通", 50])

    # === characters 工作表 ===
    ws_chars = wb.create_sheet("characters")
    ws_chars.append(["角色ID", "名称", "职业", "等级", "生命值", "魔法值"])
    ws_chars.append([1, "艾莉娅", "法师", 10, 50, 80])
    ws_chars.append([2, "托尔", "战士", 15, 100, 0])
    ws_chars.append([3, "莉娜", "牧师", 12, 70, 60])
    ws_chars.append([4, "巴纳", "弓箭手", 8, 60, 40])

    wb.save(str(file_path))
    return str(file_path)


def test_list_sheets(game_excel_file):
    """测试1: 列出工作表"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
    ops = ExcelOperations()
    result = ops.list_sheets(game_excel_file)
    sheets = result['sheets']
    expected = ['skills', 'items', 'characters']
    for e in expected:
        assert e in sheets, f"缺少工作表: {e}"


def test_get_headers(game_excel_file):
    """测试2: 获取表头"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
    ops = ExcelOperations()
    result = ops.get_headers(game_excel_file, "skills")
    descriptions = result['descriptions']
    assert any("技能ID" in desc for desc in descriptions), "缺少技能ID字段描述"


def test_query_data(game_excel_file):
    """测试3: 查询数据"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
    ops = ExcelOperations()
    query_result = ops.query(game_excel_file, "skills", "职业='法师'")
    assert len(query_result) >= 2, "应该至少有2个法师技能"


def test_get_range(game_excel_file):
    """测试4: 获取范围数据"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
    ops = ExcelOperations()
    result = ops.get_range(game_excel_file, "skills", "A1:F5")
    assert len(result) >= 2, "应该至少有2行数据"


def test_find_last_row(game_excel_file):
    """测试5: 查找最后一行"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
    ops = ExcelOperations()
    result = ops.find_last_row(game_excel_file, "skills")
    last_row = result['last_row']
    assert last_row >= 5, f"期望至少5行，实际{last_row}行"


def test_get_file_info(game_excel_file):
    """测试6: 表格描述"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
    ops = ExcelOperations()
    file_info = ops.get_file_info(game_excel_file)
    data = file_info.get('data', {})
    assert 'sheet_count' in data, "缺少工作表计数"
    assert data['sheet_count'] >= 3, f"期望至少3个工作表，实际{data['sheet_count']}个"


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))
