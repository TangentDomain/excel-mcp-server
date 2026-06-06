"""
Tests for previously untested tools:
- excel_compare_sheets
- excel_search_directory
"""

import os

import pytest

from excel_mcp_server_fastmcp.server import (
    excel_compare_sheets,
    excel_create_file,
    excel_search_directory,
    excel_update_range,
)


class TestExcelSearchDirectory:
    """Test directory search functionality"""

    def test_search_directory_basic(self, temp_dir_with_excel_files):
        """Test basic directory search"""
        result = excel_search_directory(temp_dir_with_excel_files, "标题")

        assert result["success"] is True
        assert "data" in result
        assert isinstance(result["data"], list)
        assert result["meta"]["total_matches"] > 0

    def test_search_directory_case_sensitive(self, temp_dir_with_excel_files):
        """Test case-sensitive search"""
        result = excel_search_directory(temp_dir_with_excel_files, "标题", case_sensitive=True)

        assert result["success"] is True

    def test_search_directory_no_match(self, temp_dir_with_excel_files):
        """Test search with no matches"""
        result = excel_search_directory(temp_dir_with_excel_files, "ZZZNONEXISTENT123")

        assert result["success"] is True
        assert result["meta"]["total_matches"] == 0

    def test_search_directory_regex(self, temp_dir_with_excel_files):
        """Test regex search"""
        result = excel_search_directory(temp_dir_with_excel_files, r"标题\d", use_regex=True)

        assert result["success"] is True
        assert isinstance(result["data"], list)

    def test_search_directory_nonexistent(self, temp_dir):
        """Test searching non-existent directory"""
        result = excel_search_directory("/nonexistent/path", "test")

        assert result["success"] is False

    def test_search_directory_whole_word(self, temp_dir_with_excel_files):
        """Test whole word matching"""
        result = excel_search_directory(temp_dir_with_excel_files, "标题", whole_word=True)

        assert result["success"] is True

    def test_search_directory_file_extensions(self, temp_dir):
        """Test search with file extension filter"""
        from openpyxl import Workbook

        path = os.path.join(str(temp_dir), "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "searchterm"
        wb.save(path)

        result = excel_search_directory(str(temp_dir), "searchterm", file_extensions=[".xlsx"])
        assert result["success"] is True
        assert result["meta"]["total_matches"] > 0


class TestExcelCompareSheets:
    """Test sheet comparison functionality"""

    def test_compare_identical_sheets(self, sample_excel_file, temp_dir):
        """Test comparing identical sheets"""
        import shutil

        copy_path = os.path.join(str(temp_dir), "copy.xlsx")
        shutil.copy2(sample_excel_file, copy_path)

        result = excel_compare_sheets(sample_excel_file, "Sheet1", copy_path, "Sheet1")

        assert result["success"] is True
        assert result["data"]["total_differences"] == 0

    def test_compare_different_sheets(self, sample_excel_file, temp_dir):
        """Test comparing sheets with different data"""
        import shutil

        from openpyxl import load_workbook

        copy_path = os.path.join(str(temp_dir), "modified.xlsx")
        shutil.copy2(sample_excel_file, copy_path)

        # Modify the copy
        wb = load_workbook(copy_path)
        ws = wb["Sheet1"]
        ws["A3"] = "MODIFIED_VALUE"
        wb.save(copy_path)
        wb.close()

        result = excel_compare_sheets(sample_excel_file, "Sheet1", copy_path, "Sheet1")

        assert result["success"] is True
        assert result["data"]["total_differences"] > 0

    def test_compare_nonexistent_sheet(self, sample_excel_file, temp_dir):
        """Test comparing non-existent sheet"""
        import shutil

        copy_path = os.path.join(str(temp_dir), "copy.xlsx")
        shutil.copy2(sample_excel_file, copy_path)

        result = excel_compare_sheets(sample_excel_file, "NoSuchSheet", copy_path, "Sheet1")

        assert result["success"] is False
