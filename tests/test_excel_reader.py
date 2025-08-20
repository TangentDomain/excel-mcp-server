#!/usr/bin/env python3
"""
Excel MCP Server - ExcelReader模块测试

测试excel_reader模块的所有功能，包括正常场景、边界条件和错误处理
"""

import pytest
from pathlib import Path
import tempfile
from openpyxl import Workbook

from excel_mcp.core.excel_reader import ExcelReader
from excel_mcp.utils.exceptions import FileNotFoundError, SheetNotFoundError
from excel_mcp.models.types import RangeType


class TestExcelReader:
    """测试ExcelReader类的所有功能"""

    def test_init_with_valid_file(self, sample_xlsx_file):
        """测试使用有效文件初始化"""
        reader = ExcelReader(sample_xlsx_file)
        assert reader.file_path == str(Path(sample_xlsx_file).absolute())

    def test_init_with_invalid_file(self, nonexistent_file_path):
        """测试使用无效文件初始化"""
        with pytest.raises(FileNotFoundError):
            ExcelReader(nonexistent_file_path)

    def test_list_sheets_basic(self, sample_xlsx_file):
        """测试基本的工作表列表功能"""
        reader = ExcelReader(sample_xlsx_file)
        result = reader.list_sheets()

        assert result.success is True
        assert 'sheets' in result.data.__dict__
        assert 'active_sheet' in result.data.__dict__
        assert len(result.data.sheets) > 0
        assert result.data.active_sheet in result.data.sheets

    def test_list_sheets_multi_sheet(self, multi_sheet_xlsx_file):
        """测试多工作表文件"""
        reader = ExcelReader(multi_sheet_xlsx_file)
        result = reader.list_sheets()

        assert result.success is True
        assert len(result.data.sheets) == 2
        assert "Data" in result.data.sheets
        assert "Summary" in result.data.sheets
        assert result.data.active_sheet == "Data"

    def test_get_range_basic_cell_range(self, sample_xlsx_file):
        """测试基本单元格范围读取"""
        reader = ExcelReader(sample_xlsx_file)
        result = reader.get_range("A1:C2")

        assert result.success is True
        assert len(result.data.data) == 2  # 2行
        assert len(result.data.data[0]) == 3  # 3列
        assert result.data.data[0] == ["Name", "Age", "Email"]
        assert result.data.data[1][0] == "Alice"

    def test_get_range_with_sheet_name(self, multi_sheet_xlsx_file):
        """测试指定工作表名的范围读取"""
        reader = ExcelReader(multi_sheet_xlsx_file)
        result = reader.get_range("Data!A1:B3")

        assert result.success is True
        assert len(result.data.data) == 3
        assert result.data.data[0] == ["ID", "Value"]
        assert result.data.data[1] == [1, "First"]
        assert result.data.data[2] == [2, "Second"]

    def test_get_range_full_row(self, sample_xlsx_file):
        """测试整行读取"""
        reader = ExcelReader(sample_xlsx_file)
        result = reader.get_range("1:1")

        assert result.success is True
        assert len(result.data.data) == 1
        assert result.data.data[0][:3] == ["Name", "Age", "Email"]

    def test_get_range_full_column(self, sample_xlsx_file):
        """测试整列读取"""
        reader = ExcelReader(sample_xlsx_file)
        result = reader.get_range("A:A")

        assert result.success is True
        assert len(result.data.data) >= 5  # 至少有5行数据
        assert result.data.data[0][0] == "Name"
        assert result.data.data[1][0] == "Alice"

    def test_get_range_empty_cells(self, empty_xlsx_file):
        """测试空单元格读取"""
        reader = ExcelReader(empty_xlsx_file)
        result = reader.get_range("A1:C3")

        assert result.success is True
        assert len(result.data.data) == 3
        assert all(cell is None for row in result.data.data for cell in row)

    def test_get_range_invalid_sheet(self, sample_xlsx_file):
        """测试无效工作表名"""
        reader = ExcelReader(sample_xlsx_file)
        result = reader.get_range("NonExistentSheet!A1:B2")

        assert result.success is False
        assert "NonExistentSheet" in result.error

    def test_get_range_invalid_range_format(self, sample_xlsx_file):
        """测试无效范围格式"""
        reader = ExcelReader(sample_xlsx_file)
        result = reader.get_range("INVALID_RANGE")

        assert result.success is False
        assert "无效的范围" in result.error or "格式错误" in result.error

    def test_get_range_with_formatting(self, sample_xlsx_file):
        """测试包含格式信息的读取"""
        reader = ExcelReader(sample_xlsx_file)
        result = reader.get_range("A1:B2", include_formatting=True)

        assert result.success is True
        assert hasattr(result.data, 'formatting') or 'formatting' in result.metadata

    def test_get_range_large_range(self, temp_dir):
        """测试大范围读取性能"""
        # 创建大数据文件
        file_path = temp_dir / "large_data.xlsx"
        workbook = Workbook()
        sheet = workbook.active

        # 生成1000行数据
        for row in range(1, 1001):
            sheet.cell(row=row, column=1, value=f"Data_{row}")
            sheet.cell(row=row, column=2, value=row)

        workbook.save(file_path)

        reader = ExcelReader(str(file_path))
        result = reader.get_range("A1:B1000")

        assert result.success is True
        assert len(result.data.data) == 1000
        assert result.data.data[0] == ["Data_1", 1]
        assert result.data.data[-1] == ["Data_1000", 1000]

    def test_get_range_edge_cases(self, sample_xlsx_file):
        """测试边界情况"""
        reader = ExcelReader(sample_xlsx_file)

        # 测试单个单元格
        result = reader.get_range("A1:A1")
        assert result.success is True
        assert len(result.data.data) == 1
        assert len(result.data.data[0]) == 1

        # 测试超出数据范围
        result = reader.get_range("Z1:AA100")
        assert result.success is True
        # 应该返回空数据或None


@pytest.fixture
def large_data_file(temp_dir):
    """创建大数据测试文件"""
    file_path = temp_dir / "large_test_data.xlsx"
    workbook = Workbook()
    sheet = workbook.active

    # 创建100行×50列的数据
    for row in range(1, 101):
        for col in range(1, 51):
            sheet.cell(row=row, column=col, value=f"R{row}C{col}")

    workbook.save(file_path)
    return str(file_path)


class TestExcelReaderPerformance:
    """测试ExcelReader性能"""

    def test_large_file_performance(self, large_data_file):
        """测试大文件处理性能"""
        reader = ExcelReader(large_data_file)

        import time
        start_time = time.time()
        result = reader.get_range("A1:AX100")  # 50列 x 100行
        end_time = time.time()

        assert result.success is True
        assert len(result.data.data) == 100
        assert len(result.data.data[0]) == 50
        # 性能要求：小于5秒
        assert (end_time - start_time) < 5.0

    def test_multiple_operations_performance(self, sample_xlsx_file):
        """测试多次操作性能"""
        reader = ExcelReader(sample_xlsx_file)

        import time
        start_time = time.time()

        # 执行100次读取操作
        for i in range(100):
            result = reader.get_range(f"A{i%5+1}:C{i%5+1}")
            assert result.success is True

        end_time = time.time()
        # 性能要求：小于2秒
        assert (end_time - start_time) < 2.0
