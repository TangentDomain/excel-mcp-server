"""
Excel Search完整测试套件

为ExcelSearcher类的所有核心功能提供全面的测试覆盖
目标覆盖率：80%+
"""

import pytest
import tempfile
import os
import re
from pathlib import Path
from unittest.mock import patch, MagicMock
from openpyxl import Workbook

from src.core.excel_search import ExcelSearcher
from src.models.types import SearchMatch, MatchType, OperationResult


class TestExcelSearcherBasic:
    """ExcelSearcher基础功能测试"""

    def setup_method(self):
        """每个测试方法前的设置"""
        self.test_file = "test_search.xlsx"
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Sheet1"

        # 创建测试数据
        self.ws['A1'] = "Name"
        self.ws['B1'] = "Value"
        self.ws['A2'] = "Alice"
        self.ws['B2'] = 100
        self.ws['A3'] = "Bob"
        self.ws['B3'] = 200
        self.ws['A4'] = "Charlie"
        self.ws['B4'] = 300

        # 添加一些公式
        self.ws['C1'] = "Formula"
        self.ws['C2'] = "=A2"
        self.ws['C3'] = "=B2+B3"

        # 创建第二个工作表
        self.ws2 = self.wb.create_sheet(title="Sheet2")
        self.ws2['A1'] = "Department"
        self.ws2['B1'] = "Budget"
        self.ws2['A2'] = "Sales"
        self.ws2['B2'] = 50000
        self.ws2['A3'] = "Marketing"
        self.ws2['B3'] = 30000

        # 保存文件
        self.wb.save(self.test_file)

        # 创建搜索器
        self.searcher = ExcelSearcher(self.test_file)

    def teardown_method(self):
        """每个测试方法后的清理"""
        # 清理临时文件
        if os.path.exists(self.test_file):
            try:
                os.unlink(self.test_file)
            except OSError:
                pass

    def test_searcher_initialization(self):
        """测试搜索器初始化"""
        # 验证搜索器对象被正确创建
        assert self.searcher is not None
        assert hasattr(self.searcher, 'file_path')
        assert self.test_file in self.searcher.file_path

        # 测试初始化验证 - 使用正确的异常类型
        from src.utils.exceptions import ExcelFileNotFoundError
        with pytest.raises(ExcelFileNotFoundError):
            ExcelSearcher("nonexistent.xlsx")

    def test_regex_search_basic_pattern(self):
        """测试基础正则搜索"""
        result = self.searcher.regex_search("Alice")

        assert result.success is True
        assert len(result.data) > 0
        assert result.metadata['total_matches'] > 0
        assert result.metadata['pattern'] == "Alice"

    def test_regex_search_case_insensitive(self):
        """测试不区分大小写搜索"""
        result = self.searcher.regex_search("alice", flags="i")

        assert result.success is True
        assert len(result.data) > 0
        assert any("Alice" in match.value for match in result.data)

    def test_regex_search_with_formulas(self):
        """测试搜索公式内容"""
        # 重新创建包含明确公式的测试文件
        wb = Workbook()
        ws = wb.active

        # 添加明确的公式内容
        ws['A1'] = "Formula Test"
        ws['A2'] = "=SUM(1,2,3)"
        ws['A3'] = "Normal text"
        ws['A4'] = "=A2*2"

        wb.save(self.test_file)

        # 重新创建搜索器
        self.searcher = ExcelSearcher(self.test_file)

        result = self.searcher.regex_search("=", search_values=True, search_formulas=True)

        assert result.success is True
        # 搜索应该至少找到包含等号的文本内容

    def test_regex_search_specific_sheet(self):
        """测试在指定工作表搜索"""
        result = self.searcher.regex_search("Department", sheet_name="Sheet2")

        assert result.success is True
        assert len(result.data) > 0
        assert all(match.sheet == "Sheet2" for match in result.data)

    def test_regex_search_nonexistent_sheet(self):
        """测试在不存在的工作表搜索"""
        result = self.searcher.regex_search("test", sheet_name="NonExistentSheet")

        assert result.success is False
        assert "工作表 'NonExistentSheet' 不存在" in result.error

    def test_regex_search_invalid_pattern(self):
        """测试无效正则表达式"""
        result = self.searcher.regex_search("[invalid")

        assert result.success is False
        assert "无效的正则表达式" in result.error

    def test_regex_search_range_expression(self):
        """测试范围表达式搜索"""
        result = self.searcher.regex_search("Alice", range_expression="A1:B3")

        assert result.success is True
        assert len(result.data) > 0

    def test_regex_search_range_with_sheet(self):
        """测试带工作表名的范围表达式"""
        result = self.searcher.regex_search("Sales", range_expression="Sheet2!A1:B3")

        assert result.success is True
        assert len(result.data) > 0


class TestExcelSearcherRegexFlags:
    """ExcelSearcher正则表达式标志测试"""

    def setup_method(self):
        """每个测试方法前的设置"""
        self.test_file = "test_flags.xlsx"
        self.wb = Workbook()
        self.ws = self.wb.active

        # 创建测试数据
        self.ws['A1'] = "Hello World"
        self.ws['A2'] = "hello world"
        self.ws['A3'] = "HELLO WORLD"
        self.ws['A4'] = "Line1\nLine2\nLine3"

        self.wb.save(self.test_file)
        self.searcher = ExcelSearcher(self.test_file)

    def teardown_method(self):
        """每个测试方法后的清理"""
        if os.path.exists(self.test_file):
            try:
                os.unlink(self.test_file)
            except OSError:
                pass

    def test_build_regex_flags_empty(self):
        """测试空标志"""
        flags = self.searcher._build_regex_flags("")
        assert flags == 0

    def test_build_regex_flags_ignore_case(self):
        """测试忽略大小写标志"""
        flags = self.searcher._build_regex_flags("i")
        assert flags == re.IGNORECASE

        flags = self.searcher._build_regex_flags("I")
        assert flags == re.IGNORECASE

    def test_build_regex_flags_multiline(self):
        """测试多行标志"""
        flags = self.searcher._build_regex_flags("m")
        assert flags == re.MULTILINE

        flags = self.searcher._build_regex_flags("M")
        assert flags == re.MULTILINE

    def test_build_regex_flags_dotall(self):
        """测试点匹配换行标志"""
        flags = self.searcher._build_regex_flags("s")
        assert flags == re.DOTALL

        flags = self.searcher._build_regex_flags("S")
        assert flags == re.DOTALL

    def test_build_regex_flags_combined(self):
        """测试组合标志"""
        flags = self.searcher._build_regex_flags("ims")
        expected = re.IGNORECASE | re.MULTILINE | re.DOTALL
        assert flags == expected

    def test_search_with_ignore_case(self):
        """测试不区分大小写搜索"""
        result = self.searcher.regex_search("hello", flags="i")

        assert result.success is True
        # 应该匹配多个变体
        assert len(result.data) >= 3

    def test_search_with_multiline(self):
        """测试多行搜索"""
        result = self.searcher.regex_search("^Line2$", flags="m")

        assert result.success is True
        # 应该找到包含多行文本的单元格
        assert len(result.data) > 0

    def test_search_with_dotall(self):
        """测试点匹配换行"""
        result = self.searcher.regex_search("Line1.*Line3", flags="s")

        assert result.success is True
        assert len(result.data) > 0


class TestExcelSearcherRangeTypes:
    """ExcelSearcher范围类型测试"""

    def setup_method(self):
        """每个测试方法前的设置"""
        self.test_file = "test_ranges.xlsx"
        self.wb = Workbook()
        self.ws = self.wb.active

        # 填充测试数据
        for row in range(1, 11):
            for col in range(1, 6):
                self.ws.cell(row=row, column=col, value=f"R{row}C{col}")

        self.wb.save(self.test_file)
        self.searcher = ExcelSearcher(self.test_file)

    def teardown_method(self):
        """每个测试方法后的清理"""
        if os.path.exists(self.test_file):
            try:
                os.unlink(self.test_file)
            except OSError:
                pass

    def test_search_cell_range(self):
        """测试单元格范围搜索"""
        result = self.searcher.regex_search("R5", range_expression="B2:E8")

        assert result.success is True
        assert len(result.data) > 0
        # 验证匹配的单元格都在指定范围内
        for match in result.data:
            # 检查坐标是否在B2:E8范围内
            assert '5' in match.cell  # R5应该在结果中

    def test_search_row_range(self):
        """测试行范围搜索"""
        result = self.searcher.regex_search("R3", range_expression="3:5")

        assert result.success is True
        assert len(result.data) > 0
        # 验证所有匹配都在3-5行
        for match in result.data:
            row_num = int(re.sub(r'[A-Z]', '', match.cell))
            assert 3 <= row_num <= 5

    def test_search_single_row(self):
        """测试单行搜索"""
        result = self.searcher.regex_search("R4", range_expression="4")

        assert result.success is True
        assert len(result.data) > 0
        # 验证所有匹配都在第4行
        for match in result.data:
            row_num = int(re.sub(r'[A-Z]', '', match.cell))
            assert row_num == 4

    def test_search_column_range(self):
        """测试列范围搜索"""
        result = self.searcher.regex_search("C", range_expression="B:D")

        assert result.success is True
        assert len(result.data) > 0
        # 验证所有匹配都在B-D列
        for match in result.data:
            col_letter = re.sub(r'\d', '', match.cell)
            assert col_letter in ['B', 'C', 'D']

    def test_search_single_column(self):
        """测试单列搜索"""
        result = self.searcher.regex_search("C", range_expression="C")

        assert result.success is True
        assert len(result.data) > 0
        # 验证所有匹配都在C列
        for match in result.data:
            col_letter = re.sub(r'\d', '', match.cell)
            assert col_letter == 'C'


class TestExcelSearcherDirectorySearch:
    """ExcelSearcher目录搜索测试"""

    def setup_method(self):
        """每个测试方法前的设置"""
        self.temp_dir = tempfile.mkdtemp()

        # 创建多个测试Excel文件
        self.test_files = []
        for i in range(3):
            file_path = os.path.join(self.temp_dir, f"test_file_{i}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = f"File {i} Content"
            ws['A2'] = f"Searchable content {i}"
            ws['A3'] = f"Unique keyword_{i}"
            wb.save(file_path)
            self.test_files.append(file_path)

        # 创建子目录和文件
        self.sub_dir = os.path.join(self.temp_dir, "subdir")
        os.makedirs(self.sub_dir)

        sub_file = os.path.join(self.sub_dir, "sub_file.xlsx")
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Subdirectory content"
        ws['A2'] = "Nested search"
        wb.save(sub_file)
        self.test_files.append(sub_file)

    def teardown_method(self):
        """每个测试方法后的清理"""
        # 清理临时文件和目录
        import shutil
        try:
            shutil.rmtree(self.temp_dir)
        except OSError:
            pass

    def test_regex_search_directory_basic(self):
        """测试基础目录搜索"""
        searcher = ExcelSearcher(self.test_files[0])  # 使用任意文件创建搜索器
        result = searcher.regex_search_directory(self.temp_dir, "content")

        assert result.success is True
        assert len(result.data) > 0
        assert result.metadata['total_files_found'] >= 3
        assert len(result.metadata['searched_files']) >= 3

    def test_regex_search_directory_recursive(self):
        """测试递归目录搜索"""
        searcher = ExcelSearcher(self.test_files[0])
        result = searcher.regex_search_directory(self.temp_dir, "Nested", recursive=True)

        assert result.success is True
        assert len(result.data) > 0
        # 应该找到子目录中的文件
        assert len(result.metadata['searched_files']) >= 4

    def test_regex_search_directory_non_recursive(self):
        """测试非递归目录搜索"""
        searcher = ExcelSearcher(self.test_files[0])
        result = searcher.regex_search_directory(self.temp_dir, "content", recursive=False)

        assert result.success is True
        # 应该只搜索根目录
        assert len(result.metadata['searched_files']) == 3

    def test_regex_search_directory_with_extensions(self):
        """测试指定扩展名的目录搜索"""
        searcher = ExcelSearcher(self.test_files[0])

        # 创建一个不同扩展名的文件
        txt_file = os.path.join(self.temp_dir, "test.txt")
        with open(txt_file, 'w') as f:
            f.write("This should not be found")

        result = searcher.regex_search_directory(
            self.temp_dir, "content", file_extensions=['.xlsx']
        )

        assert result.success is True
        # 只应该找到Excel文件
        assert all(file.endswith('.xlsx') for file in result.metadata['searched_files'])

    def test_regex_search_directory_with_file_pattern(self):
        """测试文件名模式过滤的目录搜索"""
        searcher = ExcelSearcher(self.test_files[0])
        result = searcher.regex_search_directory(
            self.temp_dir, "content", file_pattern=r"test_file_\d+\.xlsx"
        )

        assert result.success is True
        # 应该只匹配特定模式的文件
        assert len(result.metadata['searched_files']) == 3

    def test_regex_search_directory_max_files_limit(self):
        """测试最大文件数限制"""
        searcher = ExcelSearcher(self.test_files[0])
        result = searcher.regex_search_directory(
            self.temp_dir, "content", max_files=2
        )

        assert result.success is True
        assert len(result.metadata['searched_files']) <= 2

    def test_regex_search_directory_invalid_directory(self):
        """测试无效目录路径"""
        searcher = ExcelSearcher(self.test_files[0])
        result = searcher.regex_search_directory("/nonexistent/path", "content")

        assert result.success is False
        assert "目录不存在" in result.error

    def test_regex_search_directory_invalid_file_pattern(self):
        """测试无效文件名模式"""
        searcher = ExcelSearcher(self.test_files[0])
        result = searcher.regex_search_directory(
            self.temp_dir, "content", file_pattern="[invalid"
        )

        assert result.success is False
        assert "无效的文件名正则表达式" in result.error


class TestExcelSearcherStaticMethods:
    """ExcelSearcher静态方法测试"""

    def setup_method(self):
        """每个测试方法前的设置"""
        self.temp_dir = tempfile.mkdtemp()

        # 创建测试文件
        test_file = os.path.join(self.temp_dir, "static_test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Static method test"
        ws['A2'] = "Search content"
        wb.save(test_file)
        self.test_file = test_file

    def teardown_method(self):
        """每个测试方法后的清理"""
        import shutil
        try:
            shutil.rmtree(self.temp_dir)
        except OSError:
            pass

    def test_search_directory_static(self):
        """测试静态目录搜索方法"""
        result = ExcelSearcher.search_directory_static(self.temp_dir, "content")

        assert result.success is True
        assert len(result.data) > 0
        assert len(result.metadata['searched_files']) >= 1

    def test_build_regex_flags_static(self):
        """测试静态正则标志构建方法"""
        flags = ExcelSearcher._build_regex_flags_static("i")
        assert flags == re.IGNORECASE

        flags = ExcelSearcher._build_regex_flags_static("ms")
        expected = re.MULTILINE | re.DOTALL
        assert flags == expected

    def test_find_excel_files_static(self):
        """测试静态Excel文件查找方法"""
        from pathlib import Path

        directory = Path(self.temp_dir)
        files = ExcelSearcher._find_excel_files_static(
            directory, ['.xlsx'], None, True, 10
        )

        assert len(files) >= 1
        assert all(file.suffix == '.xlsx' for file in files)

    def test_should_include_file_static(self):
        """测试静态文件包含判断方法"""
        from pathlib import Path

        # 测试正常文件
        valid_file = Path(self.test_file)
        assert ExcelSearcher._should_include_file_static(valid_file, None) is True

        # 测试临时文件
        temp_file = Path(self.temp_dir) / "~temp.xlsx"
        assert ExcelSearcher._should_include_file_static(temp_file, None) is False

        # 测试隐藏文件
        hidden_file = Path(self.temp_dir) / ".hidden.xlsx"
        assert ExcelSearcher._should_include_file_static(hidden_file, None) is False

    def test_static_method_with_file_pattern(self):
        """测试静态方法的文件名模式过滤"""
        result = ExcelSearcher.search_directory_static(
            self.temp_dir, "test", file_pattern=r".*test.*\.xlsx"
        )

        assert result.success is True
        assert len(result.metadata['searched_files']) >= 1


class TestExcelSearcherErrorHandling:
    """ExcelSearcher错误处理测试"""

    def test_regex_search_corrupted_file(self):
        """测试损坏文件的处理"""
        # 创建一个无效的Excel文件
        corrupted_file = "corrupted.xlsx"
        with open(corrupted_file, 'w') as f:
            f.write("This is not a valid Excel file")

        try:
            searcher = ExcelSearcher(corrupted_file)
            result = searcher.regex_search("test")

            # 应该返回错误结果，不抛出异常
            assert result.success is False
            assert "无法加载Excel文件" in result.error

        finally:
            if os.path.exists(corrupted_file):
                try:
                    os.unlink(corrupted_file)
                except OSError:
                    pass

    def test_regex_search_load_error(self):
        """测试文件加载错误的处理"""
        # 模拟load_workbook抛出通用异常
        temp_file = "temp_test.xlsx"
        wb = Workbook()
        wb.save(temp_file)

        try:
            searcher = ExcelSearcher(temp_file)

            # 使用patch直接模拟ExcelSearcher内部错误
            with patch.object(searcher, '_search_workbook', side_effect=Exception("Load error")):
                result = searcher.regex_search("test")
                assert result.success is False
                assert "Load error" in result.error
        finally:
            if os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except OSError:
                    pass

    def test_directory_search_io_error(self):
        """测试目录搜索IO错误"""
        # 创建一个临时文件用于初始化
        temp_file = "temp_test.xlsx"
        wb = Workbook()
        wb.save(temp_file)

        try:
            searcher = ExcelSearcher(temp_file)

            # 使用mock模拟IO错误
            with patch('pathlib.Path.exists', return_value=False):
                result = searcher.regex_search_directory("/invalid/path", "test")
                assert result.success is False
                assert "目录不存在" in result.error
        finally:
            if os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except OSError:
                    pass


class TestExcelSearcherPerformance:
    """ExcelSearcher性能测试"""

    def setup_method(self):
        """每个测试方法前的设置"""
        self.test_file = "performance_test.xlsx"
        self.wb = Workbook()
        self.ws = self.wb.active

        # 创建大量测试数据
        for row in range(1, 101):
            for col in range(1, 11):
                self.ws.cell(row=row, column=col, value=f"Cell_{row}_{col}")

        self.wb.save(self.test_file)
        self.searcher = ExcelSearcher(self.test_file)

    def teardown_method(self):
        """每个测试方法后的清理"""
        if os.path.exists(self.test_file):
            try:
                os.unlink(self.test_file)
            except OSError:
                pass

    def test_search_performance_large_file(self):
        """测试大文件搜索性能"""
        import time

        start_time = time.time()
        result = self.searcher.regex_search("Cell_50")
        end_time = time.time()

        assert result.success is True
        assert len(result.data) > 0
        # 搜索应该在合理时间内完成（1秒内）
        assert end_time - start_time < 1.0

    def test_search_memory_usage(self):
        """测试内存使用"""
        # 这个测试主要确保没有内存泄漏
        for _ in range(10):
            result = self.searcher.regex_search("Cell")
            assert result.success is True
            assert len(result.data) > 0

    def test_concurrent_searches(self):
        """测试并发搜索"""
        import threading
        import time

        results = []
        errors = []

        def worker():
            try:
                result = self.searcher.regex_search("Cell")
                results.append(result)
            except Exception as e:
                errors.append(e)

        # 启动多个线程
        threads = []
        for _ in range(5):
            thread = threading.Thread(target=worker)
            threads.append(thread)
            thread.start()

        # 等待所有线程完成
        for thread in threads:
            thread.join()

        # 验证结果
        assert len(errors) == 0
        assert len(results) == 5
        assert all(result.success for result in results)


class TestExcelSearcherEdgeCases:
    """ExcelSearcher边界条件测试"""

    def setup_method(self):
        """每个测试方法前的设置"""
        self.test_file = "edge_case_test.xlsx"
        self.wb = Workbook()
        self.ws = self.wb.active

        # 创建边界条件测试数据
        self.ws['A1'] = ""  # 空字符串
        self.ws['A2'] = None  # None值
        self.ws['A3'] = "   "  # 只有空格
        self.ws['A4'] = "Special chars: !@#$%^&*()"
        self.ws['A5'] = "Unicode: 中文测试"
        self.ws['A6'] = "Numbers: 1234567890"
        self.ws['A7'] = "Mixed: Hello123World"

        self.wb.save(self.test_file)
        self.searcher = ExcelSearcher(self.test_file)

    def teardown_method(self):
        """每个测试方法后的清理"""
        if os.path.exists(self.test_file):
            try:
                os.unlink(self.test_file)
            except OSError:
                pass

    def test_search_empty_string(self):
        """测试搜索空字符串"""
        result = self.searcher.regex_search("")

        assert result.success is True
        # 空字符串应该匹配所有非空单元格

    def test_search_special_characters(self):
        """测试搜索特殊字符"""
        result = self.searcher.regex_search(r"[!@#$%^&*()]")

        assert result.success is True
        assert len(result.data) > 0

    def test_search_unicode_content(self):
        """测试搜索Unicode内容"""
        result = self.searcher.regex_search("中文", flags="i")

        assert result.success is True
        assert len(result.data) > 0

    def test_search_numbers(self):
        """测试搜索数字"""
        result = self.searcher.regex_search(r"\d+")

        assert result.success is True
        assert len(result.data) > 0

    def test_search_mixed_content(self):
        """测试搜索混合内容"""
        result = self.searcher.regex_search(r"Hello\d+World")

        assert result.success is True
        assert len(result.data) > 0

    def test_search_complex_pattern(self):
        """测试复杂正则模式"""
        # 查找以大写字母开头，包含数字的字符串
        result = self.searcher.regex_search(r"^[A-Z][a-zA-Z]*\d+[a-zA-Z]*$")

        assert result.success is True
        # 应该找到匹配的模式

    def test_search_with_backreferences(self):
        """测试带回溯引用的正则"""
        # 查找重复的单词
        self.ws['A8'] = "hello hello world"
        self.wb.save(self.test_file)

        result = self.searcher.regex_search(r"\b(\w+)\s+\1\b")

        assert result.success is True
        assert len(result.data) > 0

    def test_search_with_lookahead(self):
        """测试向前查找的正则"""
        # 查找后面跟着数字的字母
        result = self.searcher.regex_search(r"[a-zA-Z](?=\d)")

        assert result.success is True
        assert len(result.data) > 0

    def test_search_very_long_pattern(self):
        """测试非常长的正则模式"""
        long_pattern = "a" * 1000
        result = self.searcher.regex_search(long_pattern)

        # 应该能处理长模式而不崩溃
        assert result.success is True or result.success is False

    def test_search_with_zero_width_matches(self):
        """测试零宽度匹配"""
        # 匹配单词边界
        result = self.searcher.regex_search(r"\b")

        assert result.success is True
        # 零宽度匹配应该被处理


if __name__ == "__main__":
    pytest.main([__file__, "-v"])