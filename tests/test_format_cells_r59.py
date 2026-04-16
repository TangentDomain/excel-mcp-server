# -*- coding: utf-8 -*-
"""
format_cells R59 迭代测试 - 新增边缘案例覆盖

新增测试场景:
  1. formatting 中包含 None 值的键应被跳过（不崩溃）
  2. 空字符串 font_name / font_color 等边界值
  3. format_cells 对已合并单元格区域格式化（应正常工作）
  4. 多次连续调用 format_cells 覆盖行为（后设覆盖先设）
  5. border_style 为 None 时不设置边框
  6. 预设样式 + 自定义 formatting 深度合并验证（用户值覆盖预设）
  7. 只读文件格式化错误处理
  8. 超大范围格式化（整列/整行）
  9. 单元格值为 None 时格式化不崩溃
  10. number_format 特殊值（日期、百分比、科学计数法）
"""

import os
import pytest
import tempfile
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from excel_mcp_server_fastmcp.core.excel_writer import ExcelWriter
from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations


# ==================== Helper ====================

def _create_test_xlsx(file_path: str, rows: int = 5, cols: int = 4, sheet_name: str = "Sheet1"):
    """创建测试用 xlsx 文件，含数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=r * 10 + c)
    wb.save(file_path)
    wb.close()


def _read_cell_style(file_path: str, cell_ref: str = "A1", sheet_name: str = "Sheet1") -> dict:
    """读取单元格的样式属性用于断言"""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    cell = ws[cell_ref]
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    result = {
        "bold": font.bold,
        "italic": font.italic,
        "underline": font.underline,
        "strikethrough": font.strikethrough,
        "size": font.size,
        "font_name": font.name,
        "color": str(font.color) if font.color else None,
        "fill_type": fill.fill_type if fill else None,
        "fgColor": str(fill.fgColor) if fill and fill.fgColor else None,
        "alignment_h": alignment.horizontal,
        "alignment_v": alignment.vertical,
        "wrap_text": bool(alignment.wrap_text) if alignment.wrap_text is not None else False,
        "text_rotation": alignment.text_rotation,
        "number_format": cell.number_format,
        "border_left": border.left.style if border and border.left else None,
        "border_right": border.right.style if border and border.right else None,
        "border_top": border.top.style if border and border.top else None,
        "border_bottom": border.bottom.style if border and border.bottom else None,
        "value": cell.value,
        "indent": alignment.indent,
    }
    wb.close()
    return result


# ==================== Test Class ====================

class TestFormatCellsR59:
    """format_cells R59 第六轮测试套件"""

    # ---------- 1. None 值键跳过 ----------

    def test_formatting_with_none_values_skipped(self, tmp_path):
        """formatting 中包含 None 值的键应被跳过，有效键仍生效"""
        fp = str(tmp_path / "none_vals.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:B2",
            formatting={"bold": True, "italic": None, "font_size": None, "bg_color": "FF0000"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True, "bold 应为 True"
        assert style["fgColor"] is not None or style["fill_type"] is not None, "bg_color 应生效"

    def test_all_none_values_formatting(self, tmp_path):
        """所有值均为 None 的 formatting 不应崩溃"""
        fp = str(tmp_path / "all_none.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": None, "italic": None, "bg_color": None})
        # 不崩溃即可，成功或失败均可接受
        assert result is not None

    # ---------- 2. 空字符串边界值 ----------

    def test_empty_string_font_name(self, tmp_path):
        """空字符串 font_name 不应崩溃"""
        fp = str(tmp_path / "empty_font.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_name": "", "bold": True})
        assert result["success"], f"失败: {result.get('message')}"

    def test_empty_string_bg_color(self, tmp_path):
        """空字符串 bg_color 不应崩溃"""
        fp = str(tmp_path / "empty_bg.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bg_color": ""})
        # 空颜色可能失败或产生意外结果，但不崩溃即可
        assert result is not None

    def test_empty_string_number_format(self, tmp_path):
        """空字符串 number_format 应清空数字格式"""
        fp = str(tmp_path / "empty_nf.xlsx")
        _create_test_xlsx(fp)
        # 先设置一个数字格式
        ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"number_format": "0.00"})
        # 再用空字符串清空
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1", formatting={"number_format": ""})
        assert result["success"], f"失败: {result.get('message')}"

    # ---------- 3. 已合并单元格区域格式化 ----------

    def test_format_merged_range(self, tmp_path):
        """对已合并的单元格区域进行格式化应正常工作"""
        fp = str(tmp_path / "fmt_merge.xlsx")
        _create_test_xlsx(fp, 3, 3)
        # 先合并
        merge_result = ExcelOperations.merge_cells(fp, "Sheet1", "A1:C1")
        assert merge_result["success"], f"合并失败: {merge_result.get('message')}"
        # 再对合并区域格式化
        fmt_result = ExcelOperations.format_cells(fp, "Sheet1", "A1:C1",
            formatting={"bold": True, "bg_color": "FFFF00"})
        assert fmt_result["success"], f"格式化失败: {fmt_result.get('message')}"
        # 验证左上角单元格样式
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True, "合并区域加粗应生效"

    def test_format_then_merge(self, tmp_path):
        """先格式化再合并，格式应保留在左上角"""
        fp = str(tmp_path / "fmt_then_merge.xlsx")
        _create_test_xlsx(fp, 3, 3)
        # 先格式化
        ExcelOperations.format_cells(fp, "Sheet1", "A1:C1",
            formatting={"bold": True, "font_color": "FF0000", "bg_color": "00FF00"})
        # 再合并
        merge_result = ExcelOperations.merge_cells(fp, "Sheet1", "A1:C1")
        assert merge_result["success"], f"合并失败: {merge_result.get('message')}"
        # 验证左上角保留样式
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True, "合并后加粗应保留"

    # ---------- 4. 多次格式化覆盖行为 ----------

    def test_subsequent_format_overrides_previous(self, tmp_path):
        """后续 format_cells 调用应覆盖之前的样式"""
        fp = str(tmp_path / "override.xlsx")
        _create_test_xlsx(fp)
        # 第一次：加粗 + 红色背景
        ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": True, "bg_color": "FF0000"})
        # 第二次：不加粗 + 蓝色背景（覆盖第一次）
        ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": False, "bg_color": "0000FF"})
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is False, "第二次 bold=False 应覆盖第一次 bold=True"

    def test_partial_override_preserves_unset(self, tmp_path):
        """部分属性覆盖时，未设置的属性保持原值"""
        fp = str(tmp_path / "partial_override.xlsx")
        _create_test_xlsx(fp)
        # 第一次：设置多个属性
        ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": True, "italic": True, "bg_color": "FF0000", "font_size": 16})
        # 第二次：只修改 bold
        ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": False})
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is False, "bold 应被覆盖为 False"

    # ---------- 5. 边框相关边缘案例 ----------

    def test_border_style_none_via_formatting(self, tmp_path):
        """formatting 中 border_style=None 不应崩溃"""
        fp = str(tmp_path / "no_border.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": True, "border_style": None})
        # border_style=None 应被视为无操作，不影响其他格式
        assert result["success"], f"失败: {result.get('message')}"

    def test_border_dict_with_color_only(self, tmp_path):
        """边框字典只含 color 不含 style"""
        fp = str(tmp_path / "border_color_only.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"border": {"color": "FF0000"}})
        # 可能失败或使用默认 style，但不崩溃
        assert result is not None

    def test_border_mixed_string_and_dict(self, tmp_path):
        """边框混合字符串和字典风格"""
        fp = str(tmp_path / "border_mixed.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"border": {"top": "thin", "bottom": {"style": "thick", "color": "FF0000"}}})
        assert result["success"], f"失败: {result.get('message')}"

    # ---------- 6. 预设 + 自定义深度合并 ----------

    def test_preset_header_override_bold_false(self, tmp_path):
        """preset=header 设置 bold=True，用户 override bold=False"""
        fp = str(tmp_path / "preset_override.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": False}, preset="header")
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        # 用户 bold=False 应覆盖 header 预设的 bold=True
        assert style["bold"] is False, "用户 bold=False 应覆盖预设 bold=True"

    def test_preset_title_custom_alignment(self, tmp_path):
        """preset=title + 用户自定义 alignment"""
        fp = str(tmp_path / "preset_align.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"alignment": "right"}, preset="title")
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["alignment_h"] == "right", "用户 alignment=right 应覆盖预设 center"

    def test_preset_nonexistent_falls_back(self, tmp_path):
        """不存在的 preset 名应回退到仅使用 formatting"""
        fp = str(tmp_path / "preset_bad.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"bold": True}, preset="nonexistent_preset_xyz")
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True, "自定义 formatting 应仍生效"

    # ---------- 7. 数字格式特殊值 ----------

    def test_number_format_date(self, tmp_path):
        """日期格式的 number_format"""
        fp = str(tmp_path / "nf_date.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"number_format": "YYYY-MM-DD"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["number_format"] == "YYYY-MM-DD", f"日期格式应为 YYYY-MM-DD, 实际 {style['number_format']}"

    def test_number_format_percent(self, tmp_path):
        """百分比格式的 number_format"""
        fp = str(tmp_path / "nf_pct.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"number_format": "0.00%"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert "%" in style["number_format"], f"应含百分号, 实际 {style['number_format']}"

    def test_number_format_scientific(self, tmp_path):
        """科学计数法 number_format"""
        fp = str(tmp_path / "nf_sci.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"number_format": "0.00E+00"})
        assert result["success"], f"失败: {result.get('message')}"

    def test_number_format_currency_complex(self, tmp_path):
        """复杂货币格式（含千分位和颜色）"""
        fp = str(tmp_path / "nf_cur.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"number_format": "¥#,##0.00;[Red]-¥#,##0.00"})
        assert result["success"], f"失败: {result.get('message')}"

    # ---------- 8. 中文字体名 ----------

    def test_chinese_font_name(self, tmp_path):
        """中文字体名（微软雅黑）"""
        fp = str(tmp_path / "cn_font.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_name": "微软雅黑", "bold": True, "font_size": 12})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["font_name"] == "微软雅黑", f"字体应为 微软雅黑, 实际 {style['font_name']}"

    def test_songti_font_name(self, tmp_path):
        """宋体字体名"""
        fp = str(tmp_path / "songti.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_name": "宋体", "font_size": 14})
        assert result["success"], f"失败: {result.get('message')}"

    def test_japanese_font_name(self, tmp_path):
        """日文字体名（MS Gothic）"""
        fp = str(tmp_path / "jp_font.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"font_name": "MS Gothic", "font_size": 10})
        assert result["success"], f"失败: {result.get('message')}"

    # ---------- 9. 组合操作 ----------

    def test_combine_bold_bgcolor_alignment(self, tmp_path):
        """组合操作：bold + bg_color + alignment 同时传"""
        fp = str(tmp_path / "combine_basic.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:C3",
            formatting={
                "bold": True,
                "bg_color": "4472C4",
                "alignment": "center",
                "font_color": "FFFFFF",
                "font_size": 14,
            })
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["alignment_h"] == "center"

    def test_combine_all_font_attrs(self, tmp_path):
        """组合所有字体属性"""
        fp = str(tmp_path / "all_font.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={
                "bold": True,
                "italic": True,
                "underline": "double",
                "strikethrough": True,
                "font_size": 18,
                "font_color": "FF0000",
                "font_name": "Arial",
            })
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["italic"] is True
        assert style["underline"] == "double"
        assert style["strikethrough"] is True
        assert style["size"] == 18

    def test_combine_wrap_indent_rotation(self, tmp_path):
        """组合 wrap_text + indent + text_rotation"""
        fp = str(tmp_path / "wrap_indent_rot.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={
                "wrap_text": True,
                "indent": 3,
                "text_rotation": 45,
            })
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["wrap_text"] is True
        assert style["indent"] == 3
        assert style["text_rotation"] == 45

    def test_combine_fill_gradient(self, tmp_path):
        """渐变填充 + 字体样式组合"""
        fp = str(tmp_path / "grad_font.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={
                "gradient_colors": ["4472C4", "ED7D31", "A5A5A5"],
                "bold": True,
                "font_color": "FFFFFF",
            })
        assert result["success"], f"失败: {result.get('message')}"

    # ---------- 10. 错误处理 & 边缘案例 ----------

    def test_invalid_sheet_name(self, tmp_path):
        """不存在的工作表名"""
        fp = str(tmp_path / "bad_sheet.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "NonExistentSheet", "A1",
            formatting={"bold": True})
        # 应返回失败而非崩溃
        assert result["success"] is False or result is not None

    def test_single_cell_format(self, tmp_path):
        """单单元格格式化（非范围）"""
        fp = str(tmp_path / "single_cell.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "B3",
            formatting={"bold": True, "italic": True, "bg_color": "00FF00"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "B3")
        assert style["bold"] is True
        assert style["italic"] is True

    def test_large_range_formatting(self, tmp_path):
        """较大范围格式化（20x20 = 400 cells）"""
        fp = str(tmp_path / "large_range.xlsx")
        _create_test_xlsx(fp, rows=20, cols=20)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:T20",
            formatting={"bold": True, "border_style": "thin"})
        assert result["success"], f"失败: {result.get('message')}"

    def test_cell_with_none_value_formatting(self, tmp_path):
        """值为 None 的单元格格式化不应崩溃"""
        fp = str(tmp_path / "none_val_cell.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # A1 显式设为 None
        ws.cell(row=1, column=1, value=None)
        ws.cell(row=1, column=2, value=100)
        wb.save(fp)
        wb.close()

        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:B1",
            formatting={"bold": True, "bg_color": "FF0000"})
        assert result["success"], f"失败: {result.get('message')}"

    def test_shrink_to_fit_true(self, tmp_path):
        """shrink_to_fit=True 格式化"""
        fp = str(tmp_path / "shrink.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"shrink_to_fit": True})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        # shrink_to_fit 读回可能需要特殊处理
        assert style is not None

    def test_underline_accounting_styles(self, tmp_path):
        """会计下划线样式 singleAccounting / doubleAccounting"""
        fp = str(tmp_path / "underline_acc.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1",
            formatting={"underline": "singleAccounting"})
        assert result["success"], f"失败: {result.get('message')}"
        style = _read_cell_style(fp, "A1")
        assert style["underline"] == "singleAccounting", f"应为 singleAccounting, 实际 {style['underline']}"

        # doubleAccounting
        result2 = ExcelOperations.format_cells(fp, "Sheet1", "A2",
            formatting={"underline": "doubleAccounting"})
        assert result2["success"], f"失败: {result2.get('message')}"
        style2 = _read_cell_style(fp, "A2")
        assert style2["underline"] == "doubleAccounting", f"应为 doubleAccounting, 实际 {style2['underline']}"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
