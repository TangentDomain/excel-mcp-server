"""
Round 30 MCP 接口实测 - SQL注入防护深度回归 + 边界组合 + P0全量验证
=====================================================================
本轮重点方向:
1. R29发现的P0×4新漏洞深度确认和变体测试
2. 注入防护绕过技术尝试（编码、嵌套、混淆）
3. 边界组合：特殊字符Sheet名 + 超长列名 + 公式单元格 + 数值溢出
4. 已知全部P0/P1问题回归验证
5. UPDATE/DELETE/INSERT 安全边界探索

日期: 2026-04-14
轮次: Round 30
"""

import sys
import os
import tempfile
import shutil
import traceback

# 确保能导入项目模块
sys.path.insert(0, '/root/workspace/excel-mcp-server/src')
sys.path.insert(0, '/root/workspace/excel-mcp-server')

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)

# ============================================================
# 测试基础设施
# ============================================================
TEST_RESULTS = []
TEST_DIR = "/tmp/round30_test"

import pytest

pytestmark = pytest.mark.skip(reason="legacy test runner pattern, not for direct pytest")

def setup_test_env():
    """创建测试环境"""
    if os.path.exists(TEST_DIR):
        shutil.rmtree(TEST_DIR)
    os.makedirs(TEST_DIR, exist_ok=True)

def record(name, sql, expected, actual, status, detail=""):
    """记录测试结果"""
    TEST_RESULTS.append({
        "name": name,
        "sql": sql,
        "expected": expected,
        "actual": actual,
        "status": status,
        "detail": detail
    })
    icon = "✅" if status else "❌"
    print(f"  {icon} {name}")
    if not status and detail:
        print(f"     详情: {detail[:150]}")

def create_basic_test_file(filepath):
    """创建基础测试Excel文件"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "Name", "Price", "Rarity"])
    ws.append([1, "Sword", 100.5, "Common"])
    ws.append([2, "Shield", 250.0, "Rare"])
    ws.append([3, "Staff", 999.99, "Epic"])
    ws.append([4, "Bow", 50.0, "Legendary"])
    wb.save(filepath)
    return filepath

def create_special_sheet_test_file(filepath):
    """创建含特殊字符Sheet名的测试文件"""
    from openpyxl import Workbook
    wb = Workbook()
    
    # Sheet名含特殊字符
    ws = wb.active
    ws.title = "Sheet-Test_数据"
    ws.append(["ID", "User_Name", "Email_Address_ExtraLongColumnName", "Value"])
    ws.append([1, "Alice", "alice@test.com", 100])
    ws.append([2, "Bob O'Brien", "bob@test.com", 200])
    ws.append([3, "Charlie", "charlie@test.com", 300])
    
    # 含公式的Sheet
    ws2 = wb.create_sheet("FormulaSheet")
    ws2.append(["ID", "A", "B", "C"])
    ws2.append([1, 10, 20, None])
    ws2.append([2, 30, 40, None])
    ws2.append([3, 50, 60, None])
    for row in range(2, 5):
        ws2[f'C{row}'] = f'=A{row}*B{row}'
    
    # 超长列名Sheet
    ws3 = wb.create_sheet("WideColumns")
    headers = ["ID"] + [f"VeryLongColumnName_{i}_ForTestingPurpose" for i in range(5)]
    ws3.append(headers)
    ws3.append([1, 10, 20, 30, 40, 50])
    ws3.append([2, 11, 21, 31, 41, 51])
    
    # 中文Sheet名
    ws4 = wb.create_sheet("装备配置")
    ws4.append(["ID", "名称", "攻击力", "价格", "稀有度"])
    ws4.append([1, "圣剑", 150, 999.99, "传说"])
    ws4.append([2, "铁盾", 80, 250.0, "稀有"])
    ws4.append([3, "法杖", 120, 500.5, "史诗"])
    
    # emoji Sheet名
    try:
        ws5 = wb.create_sheet("Data🔥Test")
        ws5.append(["ID", "Val"])
        ws5.append([1, 42])
        ws5.append([2, 99])
    except Exception:
        # 某些环境不支持emoji sheet名
        pass
    
    # 数值边界测试Sheet
    ws6 = wb.create_sheet("NumBoundary")
    ws6.append(["ID", "IntVal", "FloatVal", "MixedVal"])
    ws6.append([1, 127, 3.402823466e38, 255])  # int8 max, float32 max, uint8 max
    ws6.append([2, 128, -3.402823466e38, 256])   # int8 overflow boundary
    ws6.append([3, 32767, 1.7976931348623157e308, 65535])  # int16 max, float64 max, uint16 max
    ws6.append([4, 32768, -1.7976931348623157e308, 65536])  # int16 overflow
    ws6.append([5, 2147483647, 0.0, 0])  # int32 max
    ws6.append([6, -2147483648, float('inf'), -1])  # int32 min, inf
    
    wb.save(filepath)
    return filepath


# ============================================================
# A组: SELECT路径SQL注入防护 (R29 P0-2 回归 + 变体)
# ============================================================
def test_group_a_select_injection(file_path):
    """SELECT路径注入防护测试"""
    print("\n" + "="*70)
    print("A组: SELECT路径SQL注入防护 (P0-2回归+变体)")
    print("="*70)
    
    tests = [
        # A1: 基础分号多语句 (P0-2经典)
        ("A1: SELECT分号多语句(P0-2回归)",
         "SELECT * FROM Sheet1; DROP TABLE Sheet1",
         False, "应拒绝含分号的多语句"),
        
        # A2: UNION注入
        ("A2: UNION注入尝试",
         "SELECT * FROM Sheet1 UNION SELECT * FROM Sheet1",
         True, "UNION是合法SQL语法，应支持"),
        
        # A3: 注释符注入
        ("A3: SELECT注释符注入",
         "SELECT * FROM Sheet1 -- 这是注释",
         True, "尾随注释可能被接受"),
        
        # A4: NULL字节注入 (R29 E1发现)
        ("A4: NULL字节作为分隔符",
         f"SELECT * FROM Sheet1\x00; DROP TABLE Sheet1",
         False, "NULL字节不应作为语句分隔符"),
        
        # A5: 堆叠查询DROP
        ("A5: SELECT后跟DROP TABLE(P0-2)",
         "SELECT * FROM Sheet1; DROP TABLE Sheet1",
         False, "必须拒绝DROP"),
        
        # A6: 布尔盲注模式
        ("A6: 布尔盲注模式",
         "SELECT * FROM Sheet1 WHERE 1=1 OR '1'='1'",
         True, "合法WHERE条件"),
        
        # A7: LIKE注入
        ("A7: LIKE注入尝试",
         "SELECT * FROM Sheet1 WHERE Name LIKE '%'; DROP TABLE Sheet1 --'",
         False, "LIKE中的注入应被检测"),
        
        # A8: 子查询注入
        ("A8: 子查询注入",
         "SELECT * FROM (SELECT * FROM Sheet1) t; DROP TABLE Sheet1",
         False, "子查询后的分号应被拒绝"),
        
        # A9: 编码绕过 - URL编码风格
        ("A9: URL编码%3B代替分号",
         "SELECT * FROM Sheet1%3B DROP TABLE Sheet1",
         True, "%3B不是真正的分号，可能被当字符串"),
        
        # A10: 十六进制编码
        ("A10: 十六进制/CHAR()注入",
         "SELECT * FROM Sheet1 WHERE Name = CHAR(58)",
         True, "CHAR()函数调用"),
        
        # A11: 双分号
        ("A11: 双分号多语句",
         "SELECT * FROM Sheet1;; DROP TABLE Sheet1",
         False, "双分号也是多语句"),
        
        # A12: 分号+空格变体
         ("A12: 分号+换行+DROP",
         "SELECT * FROM Sheet1;\nDROP TABLE Sheet1",
         False, "换行不改变多语句性质"),
    ]
    
    passed = 0
    for name, sql, should_succeed, reason in tests:
        try:
            result = execute_advanced_sql_query(file_path, sql)
            is_success = result.get('success', False)
            
            if should_succeed:
                if is_success:
                    record(name, sql, "成功执行", "成功执行", True)
                    passed += 1
                else:
                    record(name, sql, "成功执行", f"失败: {result.get('message', '')[:80]}", False, 
                           f"意外失败: {result.get('message', '')[:100]}")
            else:
                if is_success:
                    # 检查是否真的执行了多条语句
                    data_str = str(result.get('data', ''))
                    if '多语句' in data_str or 'multi' in data_str.lower():
                        record(name, sql, "应拒绝", f"执行了多语句! data={str(result.get('data',''))[:80]}", False,
                               "安全漏洞: 多语句被执行!")
                    else:
                        # 可能只执行了第一条
                        record(name, sql, "应拒绝", f"未正确拦截: {result.get('message', '')[:80]}", False,
                               "未拒绝含危险关键词的语句")
                else:
                    record(name, sql, "应拒绝", "正确拒绝", True)
                    passed += 1
        except Exception as e:
            if not should_succeed:
                record(name, sql, "应拒绝(异常)", f"异常: {str(e)[:60]}", True, "异常=被拦截")
                passed += 1
            else:
                record(name, sql, "成功执行", f"异常: {str(e)[:60]}", False, f"异常: {traceback.format_exc()[-100:]}")
    
    print(f"\n  A组结果: {passed}/{len(tests)} 通过")
    return passed, len(tests)


# ============================================================
# B组: UPDATE路径SQL注入防护 (R29 P0-4/P0-7 回归 + 变体)
# ============================================================
def test_group_b_update_injection(file_path):
    """UPDATE路径注入防护测试"""
    print("\n" + "="*70)
    print("B组: UPDATE路径SQL注入防护 (P0-4/P0-7回归+变体)")
    print("="*70)
    
    # 使用独立副本避免测试间干扰
    import copy
    
    tests = [
        # B1: UPDATE分号多语句 (P0-4)
        ("B1: UPDATE分号+DROP(P0-4回归)",
         "UPDATE Sheet1 SET Price = 0 WHERE ID = 1; DROP TABLE Sheet1",
         False, "UPDATE后分号多语句必须拒绝"),
        
        # B2: UPDATE注释符吃掉WHERE (P0-7)
        ("B2: UPDATE注释符吃掉WHERE(P0-7回归)",
         "UPDATE Sheet1 SET Price = 0 -- 注释 WHERE ID = 1",
         False, "注释符导致全表篡改必须拒绝"),
        
        # B3: UPDATE分号+INSERT
        ("B3: UPDATE分号+INSERT",
         "UPDATE Sheet1 SET Price = 999 WHERE ID = 1; INSERT INTO Sheet1 VALUES(99,'Hack',0,'X')",
         False, "UPDATE+INSERT堆叠必须拒绝"),
        
        # B4: UPDATE内联注释 /* */
        ("B4: UPDATE内联/* */注释",
         "UPDATE Sheet1 SET /* 注释 */ Price = 0 WHERE ID = 1",
         True, "内联注释在SET前可能是合法的"),
        
        # B5: UPDATE SET中注入
        ("B5: UPDATE SET值中注入",
         "UPDATE Sheet1 SET Price = 0; DROP TABLE Sheet1 WHERE ID = 1",
         False, "SET值中的分号必须拒绝"),
        
        # B6: UPDATE #注释(MySQL风格)
        ("B6: UPDATE #MySQL风格注释",
         "UPDATE Sheet1 SET Price = 0 # WHERE ID = 1",
         False, "#注释也可能吃掉WHERE条件"),
        
        # B7: UPDATE 正常操作(基线)
        ("B7: UPDATE正常操作(基线)",
         "UPDATE Sheet1 SET Price = 777 WHERE ID = 1",
         True, "正常UPDATE必须成功"),
        
        # B8: UPDATE CASE WHEN (基线)
        ("B8: UPDATE CASE WHEN(基线)",
         "UPDATE Sheet1 SET Price = CASE WHEN Rarity='Epic' THEN 888 ELSE Price END",
         True, "CASE WHEN SET应支持"),
        
        # B9: UPDATE 分号+DELETE
        ("B9: UPDATE分号+DELETE",
         "UPDATE Sheet1 SET Price = 0 WHERE ID = 1; DELETE FROM Sheet1 WHERE ID > 0",
         False, "UPDATE+DELETE堆叠必须拒绝"),
        
        # B10: UPDATE 字符串中嵌套引号
        ("B10: UPDATE字符串中引号转义",
         "UPDATE Sheet1 SET Name = 'O''Brien''s Sword' WHERE ID = 1",
         True, "SQL标准双引号转义应支持"),
    ]
    
    passed = 0
    for name, sql, should_succeed, reason in tests:
        try:
            result = execute_advanced_update_query(file_path, sql)
            is_success = result.get('success', False)
            
            if should_succeed:
                if is_success:
                    updated = result.get('updated_count', result.get('data', '?'))
                    record(name, sql, "成功执行", f"成功, updated={updated}", True)
                    passed += 1
                else:
                    msg = result.get('message', str(result))[:120]
                    record(name, sql, "成功执行", f"失败: {msg}", False, f"意外失败: {msg}")
            else:
                if is_success:
                    updated = result.get('updated_count', '?')
                    detail = f"安全漏洞! 成功更新了{updated}个单元格"
                    record(name, sql, "应拒绝", f"未被拦截! updated={updated}", False, detail)
                else:
                    msg = result.get('message', '被拒绝')[:100]
                    record(name, sql, "应拒绝", f"正确拒绝: {msg}", True)
                    passed += 1
        except Exception as e:
            if not should_succeed:
                record(name, sql, "应拒绝(异常)", f"异常(=被拦截): {str(e)[:60]}", True)
                passed += 1
            else:
                record(name, sql, "成功执行", f"异常: {str(e)[:80]}", False, str(e)[:100])
    
    print(f"\n  B组结果: {passed}/{len(tests)} 通过")
    return passed, len(tests)


# ============================================================
# C组: INSERT路径SQL注入防护 (R29 P0-5 回归 + 变体)
# ============================================================
def test_group_c_insert_injection(file_path):
    """INSERT路径注入防护测试"""
    print("\n" + "="*70)
    print("C组: INSERT路径SQL注入防护 (P0-5回归+变体)")
    print("="*70)
    
    tests = [
        # C1: INSERT分号多语句 (P0-5)
        ("C1: INSERT分号+DROP(P0-5回归)",
         "INSERT INTO Sheet1 (ID,Name,Price,Rarity) VALUES(99,'Hack',0,'X'); DROP TABLE Sheet1",
         False, "INSERT后分号必须拒绝"),
        
        # C2: INSERT注释符
        ("C2: INSERT注释符变体",
         "INSERT INTO Sheet1 (ID,Name,Price,Rarity) VALUES(98,'Hack',0,'X') -- 尾随注释",
         True, "INSERT尾随注释通常可接受"),
        
        # C3: INSERT分号+UPDATE
        ("C3: INSERT分号+UPDATE",
         "INSERT INTO Sheet1 (ID,Name,Price,Rarity) VALUES(97,'Hack',0,'X'); UPDATE Sheet1 SET Price=0",
         False, "INSERT+UPDATE堆叠必须拒绝"),
        
        # C4: INSERT 正常(基线)
        ("C4: INSERT正常(基线)",
         "INSERT INTO Sheet1 (ID,Name,Price,Rarity) VALUES(10,'NewItem',99.9,'Common')",
         True, "正常INSERT必须成功"),
        
        # C5: INSERT 含特殊字符
        ("C5: INSERT含特殊字符",
         "INSERT INTO Sheet1 (ID,Name,Price,Rarity) VALUES(11,'It''s a ''test''',1.1,'Epic')",
         True, "转义引号应支持"),
        
        # C6: INSERT Unicode
        ("C6: INSERT中文Unicode",
         "INSERT INTO Sheet1 (ID,Name,Price,Rarity) Values(12,'日本語テスト',200,'Rare')",
         True, "Unicode插入应支持"),
        
        # C7: INSERT 负数/零
        ("C7: INSERT负数值",
         "INSERT INTO Sheet1 (ID,Name,Price,Rarity) VALUES(13,'NegTest',-99.99,'Common')",
         True, "负数应支持"),
        
        # C8: INSERT 极大值
        ("C8: INSERT极大浮点值",
         "INSERT INTO Sheet1 (ID,Name,Price,Rarity) VALUES(14,'BigNum',1.797e308,'Legendary')",
         True, "大浮点数应支持或优雅报错"),
    ]
    
    passed = 0
    for name, sql, should_succeed, reason in tests:
        try:
            result = execute_advanced_insert_query(file_path, sql)
            is_success = result.get('success', False)
            
            if should_succeed:
                if is_success:
                    inserted = result.get('inserted_count', result.get('data', '?'))
                    record(name, sql, "成功执行", f"成功, inserted={inserted}", True)
                    passed += 1
                else:
                    msg = result.get('message', str(result))[:120]
                    record(name, sql, "成功执行", f"失败: {msg}", False, msg)
            else:
                if is_success:
                    inserted = result.get('inserted_count', '?')
                    record(name, sql, "应拒绝", f"未被拦截! inserted={inserted}", False,
                           f"安全漏洞! 插入了{inserted}行")
                else:
                    msg = result.get('message', '被拒绝')[:100]
                    record(name, sql, "应拒绝", f"正确拒绝: {msg}", True)
                    passed += 1
        except Exception as e:
            if not should_succeed:
                record(name, sql, "应拒绝(异常)", f"异常(=被拦截): {str(e)[:60]}", True)
                passed += 1
            else:
                record(name, sql, "成功执行", f"异常: {str(e)[:80]}", False, str(e)[:100])
    
    print(f"\n  C组结果: {passed}/{len(tests)} 通过")
    return passed, len(tests)


# ============================================================
# D组: DELETE路径SQL注入防护 (R29 P0-6 回归 + 变体)
# ============================================================
def test_group_d_delete_injection(file_path):
    """DELETE路径注入防护测试"""
    print("\n" + "="*70)
    print("D组: DELETE路径SQL注入防护 (P0-6回归+变体)")
    print("="*70)
    
    tests = [
        # D1: DELETE分号多语句 (P0-6)
        ("D1: DELETE分号+DROP(P0-6回归)",
         "DELETE FROM Sheet1 WHERE ID = 999; DROP TABLE Sheet1",
         False, "DELETE后分号必须拒绝"),
        
        # D2: DELETE注释符
        ("D2: DELETE注释符吃掉WHERE",
         "DELETE FROM Sheet1 -- WHERE ID = 999",
         False, "注释符导致全表删除必须拒绝"),
        
        # D3: DELETE分号+INSERT
        ("D3: DELETE分号+INSERT",
         "DELETE FROM Sheet1 WHERE ID = 999; INSERT INTO Sheet1 VALUES(0,'Hacked',0,'X')",
         False, "DELETE+INSERT堆叠必须拒绝"),
        
        # D4: DELETE 正常(基线)
        ("D4: DELETE正常(基线)",
         "DELETE FROM Sheet1 WHERE ID = 999",
         True, "正常DELETE(无匹配)应成功"),
        
        # D5: DELETE OR 1=1
        ("D5: DELETE OR条件恒真",
         "DELETE FROM Sheet1 WHERE ID = 999 OR 1=1",
         True, "OR 1=1语法合法(虽然逻辑危险)"),
        
        # D6: DELETE 正常匹配
        ("D6: DELETE正常匹配一行",
         "DELETE FROM Sheet1 WHERE ID = 4",
         True, "删除Bow这一行"),
        
        # D7: DELETE #注释
        ("D7: DELETE #MySQL风格注释",
         "DELETE FROM Sheet1 # WHERE ID = 999",
         False, "#注释导致全表删除"),
        
        # D8: DELETE 分号空语句
        ("D8: DELETE后跟空语句",
         "DELETE FROM Sheet1 WHERE ID = 999; ;",
         False, "多余分号也属于多语句"),
    ]
    
    passed = 0
    for name, sql, should_succeed, reason in tests:
        try:
            result = execute_advanced_delete_query(file_path, sql)
            is_success = result.get('success', False)
            
            if should_succeed:
                if is_success:
                    deleted = result.get('deleted_count', result.get('data', '?'))
                    record(name, sql, "成功执行", f"成功, deleted={deleted}", True)
                    passed += 1
                else:
                    msg = result.get('message', str(result))[:120]
                    record(name, sql, "成功执行", f"失败: {msg}", False, msg)
            else:
                if is_success:
                    deleted = result.get('deleted_count', '?')
                    record(name, sql, "应拒绝", f"未被拦截! deleted={deleted}", False,
                           f"安全漏洞! 删除了{deleted}行")
                else:
                    msg = result.get('message', '被拒绝')[:100]
                    record(name, sql, "应拒绝", f"正确拒绝: {msg}", True)
                    passed += 1
        except Exception as e:
            if not should_succeed:
                record(name, sql, "应拒绝(异常)", f"异常(=被拦截): {str(e)[:60]}", True)
                passed += 1
            else:
                record(name, sql, "成功执行", f"异常: {str(e)[:80]}", False, str(e)[:100])
    
    print(f"\n  D组结果: {passed}/{len(tests)} 通过")
    return passed, len(tests)


# ============================================================
# E组: 边界组合测试
# ============================================================
def test_group_e_boundary_combination(file_path):
    """边界组合测试"""
    print("\n" + "="*70)
    print("E组: 边界组合测试 (特殊Sheet名+超长列名+公式+数值边界)")
    print("="*70)
    
    tests = []
    
    # E1-E4: 特殊字符Sheet名
    tests.append(("E1: 特殊字符Sheet名查询",
                  "SELECT * FROM `Sheet-Test_数据`",
                  True, "反引号引用特殊字符Sheet名"))
    
    tests.append(("E2: 中文Sheet名查询",
                  "SELECT * FROM `装备配置`",
                  True, "中文Sheet名"))
    
    tests.append(("E3: 特殊Sheet名过滤",
                  "SELECT * FROM `Sheet-Test_数据` WHERE Value >= 200",
                  True, "特殊Sheet名+WHERE"))
    
    tests.append(("E4: 特殊Sheet名聚合",
                  "SELECT COUNT(*), SUM(Value) FROM `Sheet-Test_数据`",
                  True, "特殊Sheet名+聚合"))
    
    # E5-E7: 超长列名
    tests.append(("E5: 超长列名查询",
                  "SELECT ID, VeryLongColumnName_0_ForTestingPurpose FROM WideColumns",
                  True, "超长列名选择"))
    
    tests.append(("E6: 超长列名WHERE",
                  "SELECT * FROM WideColumns WHERE VeryLongColumnName_2_ForTestingPurpose > 25",
                  True, "超长列名+WHERE"))
    
    tests.append(("E7: 超长列名UPDATE",
                  "UPDATE WideColumns SET VeryLongColumnName_0_ForTestingPurpose = 999 WHERE ID = 1",
                  True, "超长列名UPDATE"))
    
    # E8-E11: 公式单元格
    tests.append(("E8: 公式Sheet查询",
                  "SELECT * FROM FormulaSheet",
                  True, "公式Sheet基础查询"))
    
    tests.append(("E9: 公式列聚合SUM",
                  "SELECT SUM(C), AVG(C) FROM FormulaSheet",
                  True, "公式列聚合(R29 C7已知问题回归)"))
    
    tests.append(("E10: 公式列过滤",
                  "SELECT * FROM FormulaSheet WHERE C > 400",
                  True, "公式列值过滤"))
    
    tests.append(("E11: 公式Sheet UPDATE普通列",
                  "UPDATE FormulaSheet SET A = 100 WHERE ID = 1",
                  True, "公式Sheet更新非公式列"))
    
    # E12-E15: 数值边界
    tests.append(("E12: 数值边界-正常int查询",
                  "SELECT * FROM NumBoundary WHERE ID <= 3",
                  True, "正常范围整数"))
    
    tests.append(("E13: 数值边界-MAX浮点查询",
                  "SELECT ID, FloatVal FROM NumBoundary WHERE FloatVal > 1e308",
                  True, "极大浮点比较"))
    
    tests.append(("E14: 数值边界-UPDATE极大值",
                  "UPDATE NumBoundary SET IntVal = 999999999 WHERE ID = 1",
                  True, "更新为大整数(R27 P0修复回归)"))
    
    tests.append(("E15: 数值边界-inf/nan处理",
                  "SELECT * FROM NumBoundary WHERE FloatVal = float('inf') OR MixedVal < 0",
                  True, "无穷值和负数查询"))
    
    # E16: Emoji Sheet名(如果存在)
    tests.append(("E16: Emoji Sheet名查询",
                  "SELECT * FROM `Data🔥Test`",
                  True, "Emoji Sheet名(可能不支持)"))
    
    passed = 0
    for name, sql, should_succeed, reason in tests:
        is_update = sql.strip().upper().startswith('UPDATE')
        is_delete = sql.strip().upper().startswith('DELETE')
        is_insert = sql.strip().upper().startswith('INSERT')
        
        try:
            if is_update:
                result = execute_advanced_update_query(file_path, sql)
            elif is_delete:
                result = execute_advanced_delete_query(file_path, sql)
            elif is_insert:
                result = execute_advanced_insert_query(file_path, sql)
            else:
                result = execute_advanced_sql_query(file_path, sql)
            
            is_success = result.get('success', False)
            
            if should_succeed:
                if is_success:
                    record(name, sql, "成功执行", "✅ 成功", True)
                    passed += 1
                else:
                    msg = result.get('message', str(result))[:120]
                    # 对已知问题标记为预期失败
                    if 'Column not found' in msg or '公式' in msg or 'emoji' in name.lower() and '🔥' in sql:
                        record(name, sql, "成功执行", f"已知限制: {msg[:80]}", False, f"已知问题: {msg[:100]}")
                    else:
                        record(name, sql, "成功执行", f"失败: {msg}", False, msg)
            else:
                if is_success:
                    record(name, sql, "应拒绝", "未被拦截!", False, "安全漏洞!")
                else:
                    record(name, sql, "应拒绝", "正确拒绝", True)
                    passed += 1
        except Exception as e:
            err_str = str(e)
            # emoji sheet名可能抛出异常是预期的
            if '🔥' in sql or 'emoji' in name.lower():
                record(name, sql, "成功执行(允许失败)", f"预期异常: {err_str[:60]}", True, "Emoji可能不被支持")
                passed += 1
            elif should_succeed:
                record(name, sql, "成功执行", f"异常: {err_str[:80]}", False, err_str[:100])
            else:
                record(name, sql, "应拒绝", f"异常(被拦截): {err_str[:60]}", True)
                passed += 1
    
    print(f"\n  E组结果: {passed}/{len(tests)} 通过")
    return passed, len(tests)


# ============================================================
# F组: 高级注入绕过技术
# ============================================================
def test_group_f_advanced_bypass(file_path):
    """高级注入绕过技术测试"""
    print("\n" + "="*70)
    print("F组: 高级注入绕过技术测试")
    print("="*70)
    
    tests = []
    
    # F1: 大小写混合关键字
    tests.append(("F1: SeLeCT大小写混合",
                  "SeLeCT * FrOm ShEeT1; DrOp TaBlE ShEeT1",
                  False, "大小写混合不影响SQL解析"))
    
    # F2: 内嵌tab/space
    tests.append(("F2: 关键字内嵌空白",
                  "SEL\tECT * FROM\t Sheet1 ;\tDROP TABLE Sheet1",
                  False, "空白分隔仍是有效SQL"))
    
    # F3: 括号包裹
    tests.append(("F3: 括号包裹注入",
                  "SELECT * FROM (Sheet1); DROP TABLE Sheet1",
                  False, "括号后分号仍为多语句"))
    
    # F4: BETWEEN注入
    tests.append(("F4: BETWEEN中注入",
                  "SELECT * FROM Sheet1 WHERE ID BETWEEN 1 AND 2; DROP TABLE Sheet1",
                  False, "BETWEEN表达式后分号"))
    
    # F5: IN注入
    tests.append(("F5: IN列表中注入",
                  "SELECT * FROM Sheet1 WHERE ID IN (1,2); DROP TABLE Sheet1",
                  False, "IN列表后分号"))
    
    # F6: ORDER BY注入
    tests.append(("F6: ORDER BY后注入",
                  "SELECT * FROM Sheet1 ORDER BY ID; DROP TABLE Sheet1",
                  False, "ORDER BY后分号"))
    
    # F7: LIMIT注入
    tests.append(("F7: LIMIT后注入",
                  "SELECT * FROM Sheet1 LIMIT 10; DROP TABLE Sheet1",
                  False, "LIMIT后分号"))
    
    # F8: GROUP BY + HAVING注入
    tests.append(("F8: HAVING子句注入",
                  "SELECT Rarity, COUNT(*) FROM Sheet1 GROUP BY Rarity HAVING COUNT(*) > 0; DROP TABLE Sheet1",
                  False, "HAVING后分号"))
    
    # F9: 字符串截断技巧
    tests.append(("F9: 字符串值闭合注入",
                  "SELECT * FROM Sheet1 WHERE Name = ''; DROP TABLE Sheet1 --'",
                  False, "字符串闭合后注入"))
    
    # F10: EXISTS子查询注入
    tests.append(("F10: EXISTS子查询注入",
                  "SELECT * FROM Sheet1 WHERE EXISTS(SELECT 1); DROP TABLE Sheet1",
                  False, "EXISTS后分号"))
    
    passed = 0
    for name, sql, should_succeed, reason in tests:
        try:
            result = execute_advanced_sql_query(file_path, sql)
            is_success = result.get('success', False)
            
            if should_succeed:
                if is_success:
                    record(name, sql, "成功执行", "✅", True)
                    passed += 1
                else:
                    msg = result.get('message', '')[:100]
                    record(name, sql, "成功执行", f"失败: {msg}", False, msg)
            else:
                if is_success:
                    data_str = str(result.get('data', ''))
                    record(name, sql, "应拒绝", f"未被拦截! result={str(result)[:80]}", False,
                           "安全漏洞: 未被拦截!")
                else:
                    msg = result.get('message', '被拒绝')[:100]
                    record(name, sql, "应拒绝", f"正确拒绝: {msg}", True)
                    passed += 1
        except Exception as e:
            if not should_succeed:
                record(name, sql, "应拒绝(异常)", f"被拦截: {str(e)[:60]}", True)
                passed += 1
            else:
                record(name, sql, "成功执行", f"异常: {str(e)[:80]}", False, str(e)[:100])
    
    print(f"\n  F组结果: {passed}/{len(tests)} 通过")
    return passed, len(tests)


# ============================================================
# G组: 已知P0/P1全回归
# ============================================================
def test_group_g_known_issues_regression(file_path):
    """已知问题回归测试"""
    print("\n" + "="*70)
    print("G组: 已知P0/P1问题全回归验证")
    print("="*70)
    
    tests = []
    
    # G1: P0-2 SELECT分号多语句 (R25→R29连续确认)
    tests.append(("G1[P0-2]: SELECT分号多语句回归",
                  "SELECT COUNT(*) FROM Sheet1; SELECT COUNT(*) FROM Sheet1",
                  False, "P0-2: 必须仍被确认或已修复"))
    
    # G2: P0-4 UPDATE分号多语句 (R29发现)
    tests.append(("G2[P0-4]: UPDATE分号多语句回归",
                  "UPDATE Sheet1 SET Price = 0 WHERE ID = 1; DROP TABLE Sheet1",
                  False, "P0-4: 必须仍被确认或已修复"))
    
    # G3: P0-5 INSERT分号多语句 (R29发现)
    tests.append(("G3[P0-5]: INSERT分号多语句回归",
                  "INSERT INTO Sheet1(ID,Name,Price,Rarity) VALUES(90,'T',0,'X'); DROP TABLE Sheet1",
                  False, "P0-5: 必须仍被确认或已修复"))
    
    # G4: P0-6 DELETE分号多语句 (R29发现)
    tests.append(("G4[P0-6]: DELETE分号多语句回归",
                  "DELETE FROM Sheet1 WHERE ID = 999; DROP TABLE Sheet1",
                  False, "P0-6: 必须仍被确认或已修复"))
    
    # G5: P0-7 UPDATE注释符注入 (R29发现)
    tests.append(("G5[P0-7]: UPDATE注释符注入回归",
                  "UPDATE Sheet1 SET Price = 0 -- 注释 WHERE ID = 1",
                  False, "P0-7: 全表篡改漏洞"))
    
    # G6: P0-3 uint8溢出 (R27修复, R28/R29验证通过)
    tests.append(("G6[P0-3-FIXED]: uint8溢出修复回归",
                  "UPDATE NumBoundary SET MixedVal = 999 WHERE ID = 1",
                  True, "P0-3已修复: 999不应被截断为231"))
    
    # G7: P1-3 CTE表别名前缀污染
    tests.append(("G7[P1-3]: CTE表别名前缀污染回归",
                  "WITH t AS (SELECT s.ID FROM Sheet1 s) SELECT * FROM t",
                  True, "P1-3: CTE别名可能仍有前缀问题"))
    
    # G8: P2-1 UPDATE || 拼接
    tests.append(("G8[P2-1]: UPDATE ||字符串拼接",
                  "UPDATE Sheet1 SET Name = 'Prefix-' || Name WHERE ID = 1",
                  True, "P2-1: 已知不支持||拼接"))
    
    # G9: P2-2 CASE WHEN算术混合
    tests.append(("G9[P2-2]: CASE WHEN算术操作数",
                  "SELECT Name, Price * CASE WHEN Rarity='Epic' THEN 2 ELSE 1 END AS Adj FROM Sheet1",
                  True, "P2-2: 已知CASE不能做算术操作数"))
    
    # G10: P2-3 公式列聚合
    tests.append(("G10[P2-3]: 公式列SUM聚合",
                  "SELECT SUM(C) FROM FormulaSheet",
                  True, "P2-3: 公式列聚合可能找不到列"))
    
    # G11: NULL字节绕过 (R29 E1发现)
    tests.append(("G11[NULL字节]: NULL字节分隔符回归",
                  f"SELECT * FROM Sheet1\x00DROP TABLE Sheet1",
                  False, "NULL字节作为语句分隔符"))
    
    passed = 0
    for name, sql, should_succeed, reason in tests:
        is_update = sql.strip().upper().startswith('UPDATE')
        is_delete = sql.strip().upper().startswith('DELETE')
        is_insert = sql.strip().upper().startswith('INSERT')
        
        try:
            if is_update:
                result = execute_advanced_update_query(file_path, sql)
            elif is_delete:
                result = execute_advanced_delete_query(file_path, sql)
            elif is_insert:
                result = execute_advanced_insert_query(file_path, sql)
            else:
                result = execute_advanced_sql_query(file_path, sql)
            
            is_success = result.get('success', False)
            
            if should_succeed:
                if is_success:
                    # 对于P0-3回归，需要验证值确实没被截断
                    if 'P0-3' in name or 'uint8' in name.lower():
                        # 需要额外检查实际值
                        verify_result = execute_advanced_sql_query(file_path, 
                            "SELECT MixedVal FROM NumBoundary WHERE ID = 1")
                        if verify_result.get('success'):
                            val = verify_result.get('data', [[], []])
                            if len(val) >= 2 and str(val[1]) == ['999'] or (len(val[1]) > 0 and str(val[1][0]) == '999'):
                                record(name, sql, "值=999(正确)", "✅ 未截断", True)
                                passed += 1
                            else:
                                record(name, sql, "值=999(正确)", f"值={val} 可能被截断!", False, 
                                       f"P0-3回归失败! 值={val}")
                        else:
                            record(name, sql, "值=999(正确)", "验证查询失败", False, "无法验证")
                    else:
                        record(name, sql, "成功执行", "✅", True)
                        passed += 1
                else:
                    msg = result.get('message', '')[:120]
                    # 对已知P2问题，失败是预期的
                    if any(p in name for p in ['P2-1', 'P2-2', 'P2-3', 'P1-3']):
                        record(name, sql, "成功执行(已知限制)", f"已知问题: {msg[:80]}", False, 
                               f"已知P1/P2限制: {msg[:100]}")
                    else:
                        record(name, sql, "成功执行", f"失败: {msg}", False, msg)
            else:
                # 应该失败的用例（安全测试）
                if is_success:
                    record(name, sql, "应拒绝(安全)", f"⚠️ 未被拦截! 结果={str(result)[:60]}", False,
                           "🚨 安全漏洞仍未修复!")
                else:
                    record(name, sql, "应拒绝(安全)", "✅ 正确拦截/拒绝", True)
                    passed += 1
        except Exception as e:
            err_str = str(e)
            if not should_succeed:
                record(name, sql, "应拒绝(异常拦截)", f"被拦截: {err_str[:60]}", True)
                passed += 1
            else:
                if any(p in name for p in ['P2-1', 'P2-2', 'P2-3', 'P1-3']):
                    record(name, sql, "成功执行(已知限制)", f"异常(已知): {err_str[:60]}", False, 
                           f"已知限制异常: {err_str[:100]}")
                else:
                    record(name, sql, "成功执行", f"异常: {err_str[:80]}", False, err_str[:100])
    
    print(f"\n  G组结果: {passed}/{len(tests)} 通过")
    return passed, len(tests)


# ============================================================
# H组: 数据完整性交叉验证
# ============================================================
def test_group_h_data_integrity(file_path):
    """数据完整性验证 - 确保注入测试没有破坏数据"""
    print("\n" + "="*70)
    print("H组: 数据完整性交叉验证")
    print("="*70)
    
    tests = []
    
    # H1: 验证Sheet1原始数据完整性
    tests.append(("H1: Sheet1数据完整性检查",
                  "SELECT COUNT(*), SUM(Price) FROM Sheet1",
                  True, "检查总行数和Price总和"))
    
    # H2: 特殊Sheet数据完整性
    tests.append(("H2: 特殊Sheet数据完整性",
                  "SELECT COUNT(*) FROM `Sheet-Test_数据`",
                  True, "检查特殊Sheet行数"))
    
    # H3: 中文Sheet完整性
    tests.append(("H3: 中文Sheet数据完整性",
                  "SELECT COUNT(*) FROM `装备配置`",
                  True, "检查中文Sheet行数"))
    
    # H4: 公式Sheet完整性
    tests.append(("H4: 公式Sheet完整性",
                  "SELECT COUNT(*), SUM(A), SUM(B) FROM FormulaSheet",
                  True, "检查公式Sheet数据"))
    
    # H5: 宽列Sheet完整性
    tests.append(("H5: 宽列Sheet完整性",
                  "SELECT COUNT(*) FROM WideColumns",
                  True, "检查宽列Sheet行数"))
    
    passed = 0
    for name, sql, should_succeed, reason in tests:
        try:
            result = execute_advanced_sql_query(file_path, sql)
            if result.get('success'):
                data = result.get('data', '')
                record(name, sql, "数据完整", f"data={str(data)[:80]}", True)
                passed += 1
            else:
                record(name, sql, "数据完整", f"查询失败: {result.get('message','')[:80]}", False,
                       result.get('message', '')[:100])
        except Exception as e:
            record(name, sql, "数据完整", f"异常: {str(e)[:80]}", False, str(e)[:100])
    
    print(f"\n  H组结果: {passed}/{len(tests)} 通过")
    return passed, len(tests)


# ============================================================
# 主函数
# ============================================================
def main():
    print("=" * 70)
    print("Round 30 MCP 接口实测")
    print("主题: SQL注入防护深度回归 + 边界组合 + P0全量验证 + 绕过技术")
    print("=" * 70)
    
    setup_test_env()
    
    # 创建测试文件
    basic_file = os.path.join(TEST_DIR, "basic_test.xlsx")
    special_file = os.path.join(TEST_DIR, "special_test.xlsx")
    create_basic_test_file(basic_file)
    special_file = create_special_sheet_test_file(special_file)
    
    total_passed = 0
    total_tests = 0
    
    # 运行各组测试
    p, t = test_group_a_select_injection(special_file)
    total_passed += p; total_tests += t
    
    # 为B/C/D/G组重新创建basic file（因为前面的测试可能修改了它）
    basic_file2 = os.path.join(TEST_DIR, "basic_test2.xlsx")
    create_basic_test_file(basic_file2)
    
    p, t = test_group_b_update_injection(basic_file2)
    total_passed += p; total_tests += t
    
    basic_file3 = os.path.join(TEST_DIR, "basic_test3.xlsx")
    create_basic_test_file(basic_file3)
    
    p, t = test_group_c_insert_injection(basic_file3)
    total_passed += p; total_tests += t
    
    basic_file4 = os.path.join(TEST_DIR, "basic_test4.xlsx")
    create_basic_test_file(basic_file4)
    
    p, t = test_group_d_delete_injection(basic_file4)
    total_passed += p; total_tests += t
    
    p, t = test_group_e_boundary_combination(special_file)
    total_passed += p; total_tests += t
    
    p, t = test_group_f_advanced_bypass(special_file)
    total_passed += p; total_tests += t
    
    p, t = test_group_g_known_issues_regression(special_file)
    total_passed += p; total_tests += t
    
    p, t = test_group_h_data_integrity(special_file)
    total_passed += p; total_tests += t
    
    # ============================================================
    # 最终报告
    # ============================================================
    print("\n" + "=" * 70)
    print("📊 Round 30 MCP 测试最终报告")
    print("=" * 70)
    print(f"  总测试数: {total_tests}")
    print(f"  通过数量: {total_passed}")
    print(f"  失败数量: {total_tests - total_passed}")
    print(f"  通过率:   {total_passed/total_tests*100:.1f}%")
    
    print("\n  失败详情:")
    fail_count = 0
    for r in TEST_RESULTS:
        if not r['status']:
            fail_count += 1
            print(f"    ❌ {r['name']}")
            if r['detail']:
                print(f"       {r['detail'][:120]}")
    
    # 分类统计
    security_fails = [r for r in TEST_RESULTS if not r['status'] and 
                      any(k in r['name'] for k in ['P0-', '注入', '分号', 'DROP', '注释', 'NULL'])]
    known_limit_fails = [r for r in TEST_RESULTS if not r['status'] and 
                         any(k in r['name'] for k in ['P1-', 'P2-', '已知'])]
    
    print(f"\n  🚨 安全相关失败: {len(security_fails)} 个")
    print(f"  🟡 已知限制失败: {len(known_limit_fails)} 个")
    print(f"  🔴 其他失败: {fail_count - len(security_fails) - len(known_limit_fails)} 个")
    
    return total_passed, total_tests


if __name__ == "__main__":
    main()
