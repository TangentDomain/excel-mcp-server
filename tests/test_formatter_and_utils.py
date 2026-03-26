# -*- coding: utf-8 -*-
"""
Excel 格式化和工具函数测试套件

覆盖 formatter.py 和其他工具函数
"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from src.excel_mcp_server_fastmcp.utils.formatter import format_operation_result
from src.excel_mcp_server_fastmcp.models.types import OperationResult, CellInfo, SheetInfo, RangeInfo, RangeType
from src.excel_mcp_server_fastmcp.utils.parsers import RangeParser


class TestFormatter:
    """格式化工具测试"""

    def test_format_operation_result_success(self):
        """测试成功结果的格式化"""
        result = OperationResult(
            success=True,
            data=[["A", "B"], ["C", "D"]],
            metadata={"rows": 2, "columns": 2}
        )
        
        formatted = format_operation_result(result)
        
        assert formatted['success'] is True

    def test_format_operation_result_failure(self):
        """测试失败结果的格式化"""
        result = OperationResult(
            success=False,
            error="Test error"
        )
        
        formatted = format_operation_result(result)
        
        assert formatted is not None

    def test_format_with_nested_data(self):
        """测试嵌套数据格式化"""
        result = OperationResult(
            success=True,
            data=[[{"key": "value"}], ["test"]]
        )
        
        formatted = format_operation_result(result)
        
        assert formatted is not None

    def test_format_with_none_values(self):
        """测试包含None值的格式化"""
        result = OperationResult(
            success=True,
            data=[[None, "value"], ["test", None]]
        )
        
        formatted = format_operation_result(result)
        
        assert formatted is not None


class TestRangeParser:
    """范围解析器测试"""

    def test_parse_standard_range(self):
        """测试标准范围解析"""
        result = RangeParser.parse_range_expression("Sheet1!A1:C10")
        
        assert result is not None

    def test_parse_row_range(self):
        """测试行范围解析"""
        result = RangeParser.parse_range_expression("Sheet1!1:5")
        
        assert result is not None

    def test_parse_column_range(self):
        """测试列范围解析"""
        result = RangeParser.parse_range_expression("Sheet1!A:C")
        
        assert result is not None

    def test_parse_single_cell(self):
        """测试单单元格解析"""
        result = RangeParser.parse_range_expression("Sheet1!A1")
        
        assert result is not None

    def test_validate_range_syntax_valid(self):
        """测试有效范围验证"""
        result = RangeParser.validate_range_syntax("Sheet1!A1:C10")
        
        assert result is True

    def test_validate_range_syntax_invalid(self):
        """测试无效范围验证"""
        result = RangeParser.validate_range_syntax("InvalidRange")
        
        assert result is False


class TestModels:
    """数据模型测试"""

    def test_range_info_creation(self):
        """测试RangeInfo创建"""
        info = RangeInfo(
            sheet_name="Sheet1",
            range_type=RangeType.CELL_RANGE,
            cell_range="A1:C10"
        )
        
        assert info.sheet_name == "Sheet1"
        assert info.range_type == RangeType.CELL_RANGE

    def test_range_type_enum(self):
        """测试RangeType枚举"""
        assert RangeType.CELL_RANGE.value == "cell_range"
        assert RangeType.ROW_RANGE.value == "row_range"
        assert RangeType.COLUMN_RANGE.value == "column_range"


class TestValidators:
    """验证器测试"""

    def test_range_parser_valid_cases(self):
        """测试有效范围"""
        valid_cases = [
            "Sheet1!A1",
            "Sheet1!A1:C10",
            "Sheet1!1:5",
            "Sheet1!A:C",
        ]
        
        for case in valid_cases:
            result = RangeParser.validate_range_syntax(case)
            assert result is True, f"Expected {case} to be valid"


class TestIntegration:
    """集成测试"""

    def test_read_write_cycle(self, temp_dir):
        """测试读写循环"""
        # 创建文件
        file_path = temp_dir / "integration_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        ws['A1'] = "Header"
        ws['A2'] = "Data"
        wb.save(str(file_path))
        
        # 验证文件存在
        assert file_path.exists()
        
        # 读取数据
        from src.excel_mcp_server_fastmcp.core.excel_reader import ExcelReader
        reader = ExcelReader(str(file_path))
        result = reader.get_range("Test!A1:A2")
        assert result.success is True

    def test_data_consistency(self, temp_dir):
        """测试数据一致性"""
        file_path = temp_dir / "consistency_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # 写入测试数据
        test_data = [
            ["ID", "Name", "Value"],
            [1, "Item1", 100],
            [2, "Item2", 200],
            [3, "Item3", 300],
        ]
        
        for row_idx, row_data in enumerate(test_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(str(file_path))
        
        # 读取并验证
        from src.excel_mcp_server_fastmcp.core.excel_reader import ExcelReader
        reader = ExcelReader(str(file_path))
        result = reader.get_range("Data!A1:C4")
        
        assert result.success is True
        assert len(result.data) == 4

    def test_multiple_operations(self, temp_dir):
        """测试多次操作"""
        file_path = temp_dir / "multi_ops_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Ops"
        ws['A1'] = "Start"
        wb.save(str(file_path))
        
        # 第一次读取
        from src.excel_mcp_server_fastmcp.core.excel_reader import ExcelReader
        reader = ExcelReader(str(file_path))
        result1 = reader.get_range("Ops!A1")
        assert result1.success is True
        
        # 第二次读取
        result2 = reader.get_range("Ops!A1")
        assert result2.success is True

    def test_empty_file_handling(self, temp_dir):
        """测试空文件处理"""
        file_path = temp_dir / "empty_test.xlsx"
        wb = Workbook()
        wb.active.title = "Empty"
        wb.save(str(file_path))
        
        from src.excel_mcp_server_fastmcp.core.excel_reader import ExcelReader
        reader = ExcelReader(str(file_path))
        result = reader.get_range("Empty!A1:A10")
        
        # 空文件应该也能返回结果
        assert result is not None
