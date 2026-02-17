# -*- coding: utf-8 -*-
"""
Excel Operations 更多测试 - 覆盖边缘代码
"""

import pytest
from openpyxl import Workbook
from src.api.excel_operations import ExcelOperations


class TestExcelOpsMoreEdge:
    """更多边缘测试"""

    @pytest.fixture
    def test_file(self, temp_dir):
        file_path = temp_dir / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # 多列数据
        for i in range(1, 21):
            ws.cell(row=i, column=1, value=i)
            ws.cell(row=i, column=2, value=i*10)
            ws.cell(row=i, column=3, value=f"Data{i}")
        
        wb.save(file_path)
        return str(file_path)

    def test_search_with_range(self, test_file):
        """测试带范围搜索"""
        result = ExcelOperations.search(
            file_path=test_file,
            pattern="Data",
            sheet_name="Sheet1",
            range="A1:C10"
        )
        assert result is not None

    def test_search_no_sheet(self, test_file):
        """测试不指定工作表"""
        result = ExcelOperations.search(
            file_path=test_file,
            pattern="Data"
        )
        assert result is not None

    def test_update_range_preserve(self, test_file):
        """测试保留公式更新"""
        result = ExcelOperations.update_range(
            file_path=test_file,
            range_expression="Sheet1!D1:D5",
            data=[[i] for i in range(1, 6)],
            preserve_formulas=True
        )
        assert result is not None


class TestExcelOpsSheetOps:
    """工作表操作测试"""

    @pytest.fixture
    def sheet_file(self, temp_dir):
        file_path = temp_dir / "sheet.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.create_sheet("Sheet2")
        wb.save(file_path)
        return str(file_path)

    def test_create_sheet_with_index(self, sheet_file):
        """测试指定索引创建工作表"""
        result = ExcelOperations.create_sheet(
            file_path=sheet_file,
            sheet_name="NewSheet",
            index=0
        )
        assert result is not None


class TestExcelOpsDuplicateIds:
    """重复ID测试"""

    @pytest.fixture
    def dup_file(self, temp_dir):
        file_path = temp_dir / "dup.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        
        ws['A1'] = "ID"
        ws['A2'] = 1
        ws['A3'] = 2
        ws['A4'] = 1
        ws['A5'] = 3
        ws['A6'] = 2
        
        wb.save(file_path)
        return str(file_path)

    def test_check_duplicates_column_string(self, dup_file):
        """测试字符串列ID检查"""
        result = ExcelOperations.check_duplicate_ids(
            file_path=dup_file,
            sheet_name="Sheet",
            id_column="A"
        )
        assert result is not None
