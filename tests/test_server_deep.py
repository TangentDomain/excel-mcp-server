# -*- coding: utf-8 -*-
"""
Server API 深度测试套件

覆盖 server.py 中的更多API端点
"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from src.api.excel_operations import ExcelOperations


class TestServerFileOperations:
    """文件操作测试"""

    @pytest.fixture
    def test_file(self, temp_dir):
        """创建测试文件"""
        file_path = temp_dir / "test_file.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws['A1'] = "ID"
        ws['A2'] = 1
        ws['A3'] = 2
        
        wb.save(file_path)
        return str(file_path)

    def test_get_sheet_headers(self, test_file):
        """测试获取工作表表头"""
        result = ExcelOperations.get_sheet_headers(test_file)
        
        assert result is not None


class TestServerDataOperations:
    """数据操作测试"""

    @pytest.fixture
    def data_file(self, temp_dir):
        """创建数据文件"""
        file_path = temp_dir / "data_file.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DataSheet"
        
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=i)
        
        wb.save(file_path)
        return str(file_path)

    def test_insert_rows_middle(self, data_file):
        """测试在中间插入行"""
        result = ExcelOperations.insert_rows(
            file_path=data_file,
            sheet_name="DataSheet",
            row_index=5,
            count=2
        )
        
        assert result is not None

    def test_delete_rows_middle(self, data_file):
        """测试在中间删除行"""
        result = ExcelOperations.delete_rows(
            file_path=data_file,
            sheet_name="DataSheet",
            row_index=5,
            count=2
        )
        
        assert result is not None


class TestServerMergeOperations:
    """合并操作测试"""

    @pytest.fixture
    def merge_file1(self, temp_dir):
        """创建合并文件1"""
        file_path = temp_dir / "merge1.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws['A1'] = "Data1"
        ws['A2'] = "Data2"
        
        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def merge_file2(self, temp_dir):
        """创建合并文件2"""
        file_path = temp_dir / "merge2.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet2"
        
        ws['A1'] = "Data3"
        ws['A2'] = "Data4"
        
        wb.save(file_path)
        return str(file_path)

    def test_merge_files(self, merge_file1, merge_file2, temp_dir):
        """测试合并文件"""
        output_path = temp_dir / "merged.xlsx"
        
        result = ExcelOperations.merge_files(
            input_files=[merge_file1, merge_file2],
            output_path=str(output_path)
        )
        
        assert result is not None


class TestServerImportExport:
    """导入导出测试"""

    @pytest.fixture
    def import_file(self, temp_dir):
        """创建导入测试文件"""
        file_path = temp_dir / "import_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws['A1'] = "ID"
        ws['A2'] = 1
        ws['A3'] = 2
        
        wb.save(file_path)
        return str(file_path)

    def test_import_from_csv(self, temp_dir):
        """测试从CSV导入"""
        csv_path = temp_dir / "import.csv"
        
        # 创建CSV文件
        csv_path.write_text("ID,Name\n1,Alice\n2,Bob\n")
        
        result = ExcelOperations.import_from_csv(
            csv_path=str(csv_path),
            output_path=str(temp_dir / "output.xlsx")
        )
        
        assert result is not None


class TestServerCompareAdvanced:
    """高级比较测试"""

    @pytest.fixture
    def compare1(self, temp_dir):
        """创建比较文件1"""
        file_path = temp_dir / "comp1.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        
        ws['A1'] = "ID"
        ws['B1'] = "Value"
        
        for i in range(2, 8):
            ws[f'A{i}'] = i - 1
            ws[f'B{i}'] = i * 10
        
        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def compare2(self, temp_dir):
        """创建比较文件2"""
        file_path = temp_dir / "comp2.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        
        ws['A1'] = "ID"
        ws['B1'] = "Value"
        
        for i in range(2, 8):
            ws[f'A{i}'] = i - 1
            ws[f'B{i}'] = i * 20  # 不同值
        
        wb.save(file_path)
        return str(file_path)

    def test_compare_sheets_detailed_results(self, compare1, compare2):
        """测试详细比较结果"""
        result = ExcelOperations.compare_sheets(
            file1_path=compare1,
            sheet1_name="Sheet",
            file2_path=compare2,
            sheet2_name="Sheet"
        )
        
        assert result is not None
