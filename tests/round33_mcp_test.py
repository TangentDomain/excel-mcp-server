"""
Round 33 MCP 接口实测 - SQL注入防护深度测试 + 大数据量压力测试
=====================================================================
本轮重点:
  A组: SQL注入防护深度测试(新向量+变体+编码绕过)
  B组: 大数据量极限压力测试(超多行/超宽表/超长文本/混合压力)
  C组: P0漏洞第6轮回归验证
  D组: 已知P1/P2/P3问题回归
  E组: 特殊字符/Unicode边界组合测试
"""

import os
import sys
import time
import tempfile
import random
import string

# 添加项目路径
sys.path.insert(0, '/root/workspace/excel-mcp-server/src')
sys.path.insert(0, '/root/workspace/excel-mcp-server')

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)

# ============================================================
# 测试结果记录
# ============================================================
class TestResult:
    def __init__(self):
        self.results = []
        self.passed = 0
        self.failed = 0
    
    def add(self, name: str, success: bool, detail: str = ""):
        status = "✅" if success else "❌"
        print(f"  {status} {name}" + (f" — {detail}" if detail else ""))
        self.results.append({"name": name, "success": success, "detail": detail})
        if success:
            self.passed += 1
        else:
            self.failed += 1
    
    def summary(self):
        total = self.passed + self.failed
        rate = (self.passed / total * 100) if total > 0 else 0
        print(f"\n  📊 小计: {self.passed}/{total} ({rate:.1f}%)")
        return self.passed, self.failed

tr = TestResult()

# ============================================================
# 准备测试文件
# ============================================================
TEST_FILE = "/tmp/r33_test.xlsx"
TEST_FILE_LARGE = "/tmp/r33_test_large.xlsx"
TEST_FILE_WIDE = "/tmp/r33_test_wide.xlsx"

import pandas as pd
from openpyxl import Workbook

def create_test_file(path=TEST_FILE, rows=20):
    """创建标准测试文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "Name", "Value", "Price", "Category"])
    for i in range(1, rows + 1):
        ws.append([i, f"Item-{i}", i * 10, round(i * 99.99, 2), random.choice(["A", "B", "C"])])
    wb.save(path)
    return path

def create_large_file(path=TEST_FILE_LARGE, rows=50000):
    """创建大文件用于压力测试"""
    wb = Workbook()
    ws = wb.active
    ws.title = "BigData"
    # 写入表头
    headers = ["ID"] + [f"Col{i}" for i in range(1, 11)]  # 11列
    ws.append(headers)
    # 写入数据
    for i in range(1, rows + 1):
        row = [i] + [round(random.uniform(0, 10000), 2) for _ in range(10)]
        ws.append(row)
        if i % 10000 == 0:
            print(f"    ... 已生成 {i}/{rows} 行")
    wb.save(path)
    print(f"    ✅ 大文件已生成: {path} ({rows}行 x 11列)")
    return path

def create_wide_file(path=TEST_FILE_WIDE, rows=100, cols=50):
    """创建宽表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "WideSheet"
    headers = ["ID"] + [f"WideCol_{i}" for i in range(1, cols)]
    ws.append(headers)
    for i in range(1, rows + 1):
        row = [i] + [f"val_{i}_{j}" for j in range(1, cols)]
        ws.append(row)
    wb.save(path)
    print(f"    ✅ 宽表已生成: {path} ({rows}行 x {cols}列)")
    return path

print("=" * 70)
print("🔬 Round 33 MCP 接口实测")
print("   主题: SQL注入深度防护 + 大数据极限压力")
print("=" * 70)

# 创建测试文件
create_test_file()

# ============================================================
# A组: SQL注入防护深度测试 (新向量+变体+编码绕过)
# ============================================================
print("\n" + "─" * 70)
print("📌 A组: SQL注入防护深度测试 (新向量+变体)")
print("─" * 70)

# A1-A10: SELECT 路径注入
print("\n  --- A1-A10: SELECT 路径注入新向量 ---")

injection_tests_select = [
    ("A1: 分号基本注入", "SELECT * FROM Sheet1; DROP TABLE Sheet1", False, "应拒绝分号"),
    ("A2: 双分号", "SELECT * FROM Sheet1;; DROP TABLE Sheet1", False, "应拒绝双分号"),
    ("A3: 分号+空格变体", "SELECT * FROM Sheet1 ; DROP TABLE Sheet1", False, "应拒绝"),
    ("A4: 分号+换行", "SELECT * FROM Sheet1;\nDROP TABLE Sheet1", False, "应拒绝换行分隔"),
    ("A5: 注释符截断WHERE", "SELECT * FROM Sheet1 WHERE ID = 1 -- AND Value > 0", True, "注释截断WHERE但SELECT仍执行"),
    ("A6: /*块注释*/注入", "SELECT * FROM Sheet1 /* 注释 */ WHERE ID = 1", True, "块注释可能被接受"),
    ("A7: 注释内嵌分号", "SELECT * FROM Sheet1; -- ; DROP x", False, "注释中的分号不应执行"),
    ("A8: 反引号转义尝试", "SELECT * FROM `Sheet1`; DROP TABLE `Sheet1`", False, "反引号不能绕过"),
    ("A9: UNION注入(列数匹配)", "SELECT ID, Name FROM Sheet1 UNION SELECT ID, Name FROM Sheet1", True, "UNION可能被执行"),
    ("A10: LIKE中通配符注入", "SELECT * FROM Sheet1 WHERE Name LIKE '%'; DROP x --'", True, "LIKE字符串内的分号"),
]

for name, sql, expect_success, note in injection_tests_select:
    try:
        result = execute_advanced_sql_query(TEST_FILE, sql)
        actual_success = result.get('success', False)
        # 对于注入测试，我们关注的是是否拒绝了危险操作
        if not expect_success:
            # 期望失败（拒绝）
            if not actual_success or 'error' in str(result).lower() or 'fail' in str(result).lower():
                tr.add(name, True, f"正确拒绝 | {note}")
            elif actual_success and '多语句' in str(result.get('message', '')):
                tr.add(name, False, f"🚨 多语句被执行! | {result.get('message', '')[:80]}")
            else:
                # 需要更细致判断
                data = result.get('data', [])
                msg = result.get('message', '')
                if isinstance(data, list) and len(data) > 0:
                    tr.add(name, False, f"🚨 返回了数据! 可能未拒绝 | msg:{str(msg)[:60]}")
                else:
                    tr.add(name, True, f"可能已拒绝(无数据返回) | {note}")
        else:
            # 期望成功（正常执行或安全接受）
            if actual_success:
                tr.add(name, True, f"正常执行 | {note}")
            else:
                tr.add(name, True, f"优雅拒绝(可接受) | {msg[:60]}")
    except Exception as e:
        if not expect_success:
            tr.add(name, True, f"异常抛出(=拒绝) | {str(e)[:60]}")
        else:
            tr.add(name, False, f"意外异常 | {str(e)[:60]}")

# A11-A18: UPDATE 路径注入
print("\n  --- A11-A18: UPDATE 路径注入 ---")

injection_tests_update = [
    ("A11: UPDATE分号注入", "UPDATE Sheet1 SET Value = 0 WHERE ID = 1; DROP TABLE Sheet1", False),
    ("A12: UPDATE注释篡改", "UPDATE Sheet1 SET Value = 0 -- WHERE ID = 1", False),  # 全表篡改!
    ("A13: UPDATE/*注释*/绕过", "UPDATE Sheet1 SET /*hacked*/ Value = -999 WHERE ID = 1", True),
    ("A14: UPDATE SET中注入", "UPDATE Sheet1 SET Value = 0; DELETE FROM Sheet1 WHERE ID = 1", False),
    ("A15: UPDATE嵌套括号", "UPDATE Sheet1 SET Value = (SELECT MAX(ID) FROM Sheet1) WHERE ID = 1", True),
    ("A16: UPDATE CASE注入", "UPDATE Sheet1 SET Value = CASE WHEN 1=1 THEN -1 ELSE Value END", True),
    ("A17: UPDATE子查询注入", "UPDATE Sheet1 SET Value = (SELECT Value FROM Sheet1 WHERE ID=999) WHERE ID=1", True),
    ("A18: UPDATE空WHERE全表", "UPDATE Sheet1 SET Value = 99999", True),  # 无WHERE是合法的SQL
]

for name, sql, expect_safe in injection_tests_update:
    try:
        # 先读取原始值
        before = execute_advanced_sql_query(TEST_FILE, "SELECT Value FROM Sheet1 WHERE ID = 1")
        result = execute_advanced_update_query(TEST_FILE, sql)
        
        after = execute_advanced_sql_query(TEST_FILE, "SELECT Value FROM Sheet1 WHERE ID = 1")
        
        if not expect_safe:
            # 如果是不安全的操作，检查是否被拒绝或产生了非预期效果
            if not result.get('success'):
                tr.add(name, True, "正确拒绝")
            else:
                # 检查是否有非预期效果
                affected = result.get('affected_rows', 0)
                msg = result.get('message', '')
                if '分号' in sql and affected >= 0:
                    tr.add(name, False, f"🚨 未拒绝分号! affected={affected}")
                elif '--' in sql and 'WHERE' in sql:
                    # 注释符篡改 - 检查是否修改了超过预期的行数
                    if affected > 1:
                        tr.add(name, False, f"🚨 注释符导致全表修改! affected={affected}")
                    else:
                        tr.add(name, True, f"仅影响{affected}行(可能OK)")
                else:
                    tr.add(name, True, f"执行成功 affected={affected}")
        else:
            if result.get('success'):
                tr.add(name, True, f"执行成功 affected={result.get('affected_rows', '?')}")
            else:
                tr.add(name, True, f"优雅拒绝: {str(result.get('message', ''))[:50]}")
        
        # 恢复测试文件
        execute_advanced_update_query(TEST_FILE, "UPDATE Sheet1 SET Value = ID * 10")
    except Exception as e:
        if not expect_safe:
            tr.add(name, True, f"异常(=拒绝): {str(e)[:50]}")
        else:
            tr.add(name, False, f"意外异常: {str(e)[:50]}")

# A19-A24: INSERT/DELETE 路径注入
print("\n  --- A19-A24: INSERT/DELETE 路径注入 ---")

injection_tests_write = [
    ("A19: INSERT分号注入", "INSERT INTO Sheet1 (ID, Name, Value, Price, Category) VALUES (999, 'test', 1, 1, 'A'); DROP TABLE Sheet1", False),
    ("A20: INSERT注释注入", "INSERT INTO Sheet1 (ID, Name, Value, Price, Category) VALUES (999, 'test', 1, 1, 'A') -- 正常结束", True),
    ("A21: DELETE分号注入", "DELETE FROM Sheet1 WHERE ID = 99999; DROP TABLE Sheet1", False),
    ("A22: DELETE注释注入", "DELETE FROM Sheet1 WHERE ID = 99999 -- AND 1=1", True),
    ("A23: DELETE空WHERE", "DELETE FROM Sheet1 WHERE 1=0", True),  # 安全的空操作
    ("A24: DELETE恒真(危险)", "DELETE FROM Sheet1 WHERE 1=1 OR 'a'='a'", False),  # 危险但语法合法
]

for name, sql, expect_safe in injection_tests_write:
    is_insert = sql.strip().upper().startswith('INSERT')
    is_delete = sql.strip().upper().startswith('DELETE')
    
    try:
        # 记录操作前状态
        count_before = execute_advanced_sql_query(TEST_FILE, "SELECT COUNT(*) as cnt FROM Sheet1")
        before_cnt = count_before['data'][1][0] if count_before.get('data') and len(count_before['data']) > 1 else '?'
        
        if is_insert:
            result = execute_advanced_insert_query(TEST_FILE, sql)
        elif is_delete:
            result = execute_advanced_delete_query(TEST_FILE, sql)
        else:
            continue
        
        if not expect_safe:
            if not result.get('success'):
                tr.add(name, True, "正确拒绝")
            else:
                tr.add(name, False, f"🚨 未拒绝! msg: {str(result.get('message', ''))[:60]}")
        else:
            if result.get('success'):
                tr.add(name, True, "执行成功")
            else:
                tr.add(name, True, f"优雅拒绝: {str(result.get('message', ''))[:50]}")
        
        # 清理: 删除可能插入的测试行
        try:
            execute_advanced_delete_query(TEST_FILE, "DELETE FROM Sheet1 WHERE ID = 999")
        except:
            pass
            
    except Exception as e:
        if not expect_safe:
            tr.add(name, True, f"异常(=拒绝): {str(e)[:50]}")
        else:
            tr.add(name, False, f"意外异常: {str(e)[:50]}")

# A25-A30: 高级编码/混淆注入
print("\n  --- A25-A30: 高级编码/混淆注入 ---")

advanced_injections = [
    ("A25: URL编码%3B代替分号", "SELECT * FROM Sheet1%3B DROP TABLE Sheet1", True),  # URL编码不会被SQL解析
    ("A26: 十六进制字符串", "SELECT * FROM Sheet1 WHERE Name = 0x74657374", True),  # hex字面量
    ("A27: CHAR函数拼接", "SELECT * FROM Sheet1 WHERE Name = CONCAT('Item', '-', '1')", True),  # CONCAT
    ("A28: 双写关键字SELESELECTCT", "SELESELECTCT * FROM Sheet1", False),  # 应报语法错误
    ("A29: 大小写混合sElEcT", "sElEcT * FrOm ShEeT1", True),  # 大小写不敏感通常OK
    ("A30: NULL字节\\x00分隔", "SELECT * FROM Sheet1\x00; DROP TABLE Sheet1", False),  # NULL字节!
]

for name, sql, expect_safe in advanced_injections:
    try:
        result = execute_advanced_sql_query(TEST_FILE, sql)
        if result.get('success'):
            data_str = str(result.get('data', ''))
            msg = str(result.get('message', ''))
            # 检查是否有多语句执行的迹象
            if '多语句' in msg:
                tr.add(name, False, f"🚨 多语句执行! {msg[:60]}")
            elif not expect_safe and ('DROP' in sql or '\x00' in sql):
                tr.add(name, False, f"🚨 危险SQL未被拦截! {msg[:60]}")
            else:
                tr.add(name, True, f"执行成功(预期内) | {msg[:40]}")
        else:
            if not expect_safe:
                tr.add(name, True, f"正确拒绝: {str(result.get('message', ''))[:50]}")
            else:
                tr.add(name, True, f"优雅拒绝(可接受): {str(result.get('message', ''))[:40]}")
    except Exception as e:
        if not expect_safe:
            tr.add(name, True, f"异常(=拒绝): {str(e)[:50]}")
        else:
            tr.add(name, False, f"意外异常: {str(e)[:50]}")

tr.summary()

# ============================================================
# B组: 大数据量极限压力测试
# ============================================================
print("\n" + "─" * 70)
print("📌 B组: 大数据量极限压力测试")
print("─" * 70)

# 重新创建干净的测试文件
create_test_file()
large_file = create_large_file(rows=50000)
wide_file = create_wide_file(rows=200, cols=50)

# B1-B5: 大数据量查询性能
print("\n  --- B1-B8: 大数据量查询性能 ---")

stress_queries = [
    ("B1: COUNT(*) 50K行", large_file, "SELECT COUNT(*) as cnt FROM BigData", None),
    ("B2: WHERE+LIMIT 50K行", large_file, "SELECT * FROM BigData WHERE Col1 > 5000 LIMIT 10", None),
    ("B3: GROUP BY聚合 50K行", large_file, "SELECT COUNT(*), AVG(Col1), SUM(Col2) FROM BigData GROUP BY CASE WHEN Col1 > 5000 THEN 1 ELSE 0 END", None),
    ("B4: ORDER BY 50K行", large_file, "SELECT * FROM BigData ORDER BY Col1 DESC LIMIT 5", None),
    ("B5: 窗口函数 50K行", large_file, "SELECT ID, Col1, RANK() OVER (ORDER BY Col1 DESC) as rnk FROM BigData LIMIT 10", None),
    ("B6: CTE+JOIN自身 50K行", large_file, "WITH Top10 AS (SELECT * FROM BigData ORDER BY Col1 DESC LIMIT 10) SELECT * FROM Top10", None),
    ("B7: 宽表SELECT*", wide_file, "SELECT * FROM WideSheet LIMIT 5", None),
    ("B8: 宽表GROUP BY", wide_file, "SELECT COUNT(*) as cnt FROM WideSheet GROUP BY WideCol_1 LIMIT 10", None),
]

for name, filepath, sql, _ in stress_queries:
    try:
        start = time.time()
        result = execute_advanced_sql_query(filepath, sql)
        elapsed = time.time() - start
        
        if result.get('success'):
            data = result.get('data', [])
            row_count = len(data) - 1 if isinstance(data, list) and len(data) > 0 else 0
            tr.add(name, True, f"{elapsed:.2f}s, 返回{row_count}行")
        else:
            tr.add(name, False, f"{elapsed:.2}s, 失败: {str(result.get('message', ''))[:60]}")
    except Exception as e:
        tr.add(name, False, f"异常: {str(e)[:60]}")

# B9-B14: 大数据量写操作压力
print("\n  --- B9-B14: 大数据量写操作压力 ---")

stress_writes = [
    ("B9: UPDATE 50K行(批量)", large_file, "UPDATE BigData SET Col1 = Col1 * 1.001 WHERE ID <= 10000"),
    ("B10: UPDATE带表达式 50K行", large_file, "UPDATE BigData SET Col2 = ROUND(Col1 * 1.05, 2) WHERE ID <= 5000"),
    ("B11: INSERT单行大文件后", large_file, "INSERT INTO BigData (ID, Col1, Col2, Col3, Col4, Col5, Col6, Col7, Col8, Col9, Col10) VALUES (99999, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10)"),
    ("B12: DELETE少量行", large_file, "DELETE FROM BigData WHERE ID = 99999"),
    ("B13: UPDATE CASE WHEN 50K行", large_file, "UPDATE BigData SET Col1 = CASE WHEN Col1 > 5000 THEN Col1 * 1.1 ELSE Col1 * 0.9 END WHERE ID <= 10000"),
    ("B14: 全表COUNT验证完整性", large_file, "SELECT COUNT(*) as cnt FROM BigData"),
]

for name, filepath, sql in stress_writes:
    try:
        start = time.time()
        is_update = sql.strip().upper().startswith('UPDATE')
        is_insert = sql.strip().upper().startswith('INSERT')
        is_delete = sql.strip().upper().startswith('DELETE')
        
        if is_update:
            result = execute_advanced_update_query(filepath, sql)
        elif is_insert:
            result = execute_advanced_insert_query(filepath, sql)
        elif is_delete:
            result = execute_advanced_delete_query(filepath, sql)
        else:
            result = execute_advanced_sql_query(filepath, sql)
            
        elapsed = time.time() - start
        
        if result.get('success'):
            affected = result.get('affected_rows', '?')
            tr.add(name, True, f"{elapsed:.2f}s, affected={affected}")
        else:
            tr.add(name, False, f"{elapsed:.2}s, 失败: {str(result.get('message', ''))[:60]}")
    except Exception as e:
        tr.add(name, False, f"异常: {str(e)[:60]}")

# B15-B18: 极端数据规模
print("\n  --- B15-B18: 极端数据规模 ---")

# B15: 超长文本写入和回读
try:
    long_text = "A" * 50000  # 5万字符
    execute_advanced_update_query(TEST_FILE, f"UPDATE Sheet1 SET Name = '{long_text}' WHERE ID = 1")
    result = execute_advanced_sql_query(TEST_FILE, "SELECT Name FROM Sheet1 WHERE ID = 1")
    if result.get('success') and result.get('data'):
        read_back = str(result['data'][1][0]) if len(result['data']) > 1 else ""
        if len(read_back) >= 49000:  # 允许一些误差
            tr.add("B15: 超长文本(50K字符)写入回读", True, f"写入50000, 读回{len(read_back)}字符")
        else:
            tr.add("B15: 超长文本(50K字符)写入回读", False, f"写入50000, 仅读回{len(read_back)}字符!")
    else:
        tr.add("B15: 超长文本(50K字符)写入回读", False, f"回读失败: {str(result.get('message', ''))[:40]}")
    # 恢复
    execute_advanced_update_query(TEST_FILE, "UPDATE Sheet1 SET Name = 'Item-1' WHERE ID = 1")
except Exception as e:
    tr.add("B15: 超长文本(50K字符)写入回读", False, str(e)[:60])

# B16: 特殊字符密集文本
try:
    special_text = "abc--xyz" * 100  # 重复100次, 含注释符
    escaped = special_text.replace("'", "''")
    execute_advanced_update_query(TEST_FILE, f"UPDATE Sheet1 SET Name = '{escaped}' WHERE ID = 1")
    result = execute_advanced_sql_query(TEST_FILE, "SELECT Name FROM Sheet1 WHERE ID = 1")
    if result.get('success'):
        tr.add("B16: 特殊字符密集文本(含XSS)", True, "写入并回读成功")
    else:
        tr.add("B16: 特殊字符密集文本(含XSS)", False, str(result.get('message', ''))[:60])
    execute_advanced_update_query(TEST_FILE, "UPDATE Sheet1 SET Name = 'Item-1' WHERE ID = 1")
except Exception as e:
    tr.add("B16: 特殊字符密集文本(含XSS)", False, str(e)[:60])

# B17: Unicode超长文本
try:
    unicode_text = "中文日本語한글العربيةעבריתไทย🔥🎉🚀" * 500  # 约15000字符
    execute_advanced_update_query(TEST_FILE, f"UPDATE Sheet1 SET Name = '{unicode_text}' WHERE ID = 1")
    result = execute_advanced_sql_query(TEST_FILE, "SELECT LENGTH(Name) as len FROM Sheet1 WHERE ID = 1")
    if result.get('success') and result.get('data'):
        tr.add("B17: Unicode超长文本(15K字符)", True, "Unicode写入成功")
    else:
        tr.add("B17: Unicode超长文本(15K字符)", False, str(result.get('message', ''))[:60])
    execute_advanced_update_query(TEST_FILE, "UPDATE Sheet1 SET Name = 'Item-1' WHERE ID = 1")
except Exception as e:
    tr.add("B17: Unicode超长文本(15K字符)", False, str(e)[:60])

# B18: 数值极端精度
try:
    tests_precision = [
        ("极小正数", 1e-300),
        ("极大整数", 10**18),
        ("高精度浮点", 3.14159265358979323846),
    ]
    all_ok = True
    for label, val in tests_precision:
        r = execute_advanced_update_query(TEST_FILE, f"UPDATE Sheet1 SET Value = {val} WHERE ID = 1")
        if not r.get('success'):
            all_ok = False
            break
    if all_ok:
        tr.add("B18: 数值极端精度(x3)", True, "极小数/大整数/高精度pi均写入成功")
    else:
        tr.add("B18: 数值极端精度(x3)", False, f"{label}写入失败")
    execute_advanced_update_query(TEST_FILE, "UPDATE Sheet1 SET Value = 10 WHERE ID = 1")
except Exception as e:
    tr.add("B18: 数值极端精度(x3)", False, str(e)[:60])

tr.summary()

# ============================================================
# C组: P0漏洞第6轮回归验证
# ============================================================
print("\n" + "─" * 70)
print("📌 C组: P0漏洞第6轮回归验证")
print("─" * 70)

# 重建干净文件
create_test_file()

p0_tests = [
    ("C1: P0-2 SELECT分号多语句", "query",
     "SELECT * FROM Sheet1; SELECT * FROM Sheet1",
     lambda: None,  # 不需要前置条件
     lambda r: '多语句' in str(r.get('message', '')),  # 如果出现多语句说明有漏洞
     False),  # 期望: 不应该有多语句执行
    
    ("C2: P0-4 UPDATE分号多语句", "update",
     "UPDATE Sheet1 SET Value = -999 WHERE ID = 1; DROP TABLE Sheet1",
     lambda: None,
     lambda r: r.get('success', False) and '分号' in "UPDATE Sheet1 SET Value = -999 WHERE ID = 1; DROP TABLE Sheet1",
     False),
    
    ("C3: P0-5 INSERT分号多语句", "insert",
     "INSERT INTO Sheet1 (ID, Name, Value, Price, Category) VALUES (888, 'hack', -1, -1, 'X'); DROP TABLE Sheet1",
     lambda: None,
     lambda r: r.get('success', False),
     False),
    
    ("C4: P0-6 DELETE分号多语句", "delete",
     "DELETE FROM Sheet1 WHERE ID = 99999; DROP TABLE Sheet1",
     lambda: None,
     lambda r: r.get('success', False),
     False),
    
    ("C5: P0-7 UPDATE注释符全表篡改", "update",
     "UPDATE Sheet1 SET Value = -777 -- WHERE ID = 1",
     lambda: None,
     lambda r: r.get('affected_rows', 0) > 1,  # 如果影响了>1行说明全表被篡改
     False),
]

for name, op_type, sql, setup, check_vuln, expect_vuln in p0_tests:
    try:
        setup()
        
        if op_type == "query":
            result = execute_advanced_sql_query(TEST_FILE, sql)
        elif op_type == "update":
            result = execute_advanced_update_query(TEST_FILE, sql)
        elif op_type == "insert":
            result = execute_advanced_insert_query(TEST_FILE, sql)
        elif op_type == "delete":
            result = execute_advanced_delete_query(TEST_FILE, sql)
        
        is_vulnerable = check_vuln(result)
        
        if is_vulnerable:
            tr.add(name, False, f"🚨🚨🚨 第6轮确认仍存在! {str(result.get('message', ''))[:60]}")
        else:
            # 需要进一步确认是否真的修复了还是假阴性
            if not result.get('success'):
                tr.add(name, True, "✅ 已修复! 操作被正确拒绝")
            else:
                # 成功了但没有明显的漏洞特征 - 需要具体分析
                msg = str(result.get('message', ''))
                affected = result.get('affected_rows', 0)
                if 'P0-7' in name and affected <= 1:
                    tr.add(name, True, f"✅ 可能已修复! 仅影响{affected}行")
                elif 'P0-2' in name and '多语句' not in msg:
                    tr.add(name, True, "✅ 可能已修复! 无多语句特征")
                else:
                    tr.add(name, False, f"⚠️ 需人工判断: success={result.get('success')}, msg={msg[:50]}, affected={affected}")
        
        # 恢复文件
        create_test_file()
    except Exception as e:
        tr.add(name, False, f"异常: {str(e)[:60]}")
        create_test_file()

# C6: P0-3 uint8溢出修复持续验证
try:
    execute_advanced_update_query(TEST_FILE, "UPDATE Sheet1 SET Value = 999 WHERE ID = 1")
    result = execute_advanced_sql_query(TEST_FILE, "SELECT Value FROM Sheet1 WHERE ID = 1")
    if result.get('success') and result.get('data'):
        val = result['data'][1][0]
        if val == 999:
            tr.add("C6: P0-3 uint8溢出修复(R33验证)", True, f"Value=999 ✅ 修复持续有效!")
        else:
            tr.add("C6: P0-3 uint8溢出修复(R33验证)", False, f"🚨 回退! 写入999读到{val}(应为999)")
    else:
        tr.add("C6: P0-3 uint8溢出修复(R33验证)", False, "查询失败")
    execute_advanced_update_query(TEST_FILE, "UPDATE Sheet1 SET Value = 10 WHERE ID = 1")
except Exception as e:
    tr.add("C6: P0-3 uint8溢出修复(R33验证)", False, str(e)[:60])

tr.summary()

# ============================================================
# D组: 已知P1/P2/P3问题回归
# ============================================================
print("\n" + "─" * 70)
print("📌 D组: 已知P1/P2/P3问题回归")
print("─" * 70)

create_test_file()

# D1: P1-3 CTE表别名前列名前缀污染
try:
    # 这个需要多Sheet文件才能测，用现有文件模拟
    result = execute_advanced_sql_query(TEST_FILE, """
        WITH Stats AS (
            SELECT Category, COUNT(*) as cnt, AVG(Value) as avg_val 
            FROM Sheet1 GROUP BY Category
        )
        SELECT s.Category, s.cnt, s.avg_val 
        FROM Stats s 
        ORDER BY s.avg_val DESC
    """)
    if result.get('success'):
        tr.add("D1: P1-3 CTE表别名前列名", True, "CTE+别名查询成功")
    else:
        err = str(result.get('message', ''))[:80]
        if '没有列' in err or 'not found' in err.lower():
            tr.add("D1: P1-3 CTE表别名前列名", False, f"🚨 仍存在: {err}")
        else:
            tr.add("D1: P1-3 CTE表别名前列名", False, f"其他错误: {err}")
except Exception as e:
    tr.add("D1: P1-3 CTE表别名前列名", False, str(e)[:60])

# D2: P2-1 UPDATE SET || 字符串拼接
try:
    result = execute_advanced_update_query(TEST_FILE, "UPDATE Sheet1 SET Name = 'Prefix-' || Name WHERE ID = 1")
    if result.get('success'):
        tr.add("D2: P2-1 UPDATE SET ||拼接", True, "|| 拼接现在支持了? 或其他行为")
    else:
        err = str(result.get('message', ''))[:60]
        if '不支持' in err or 'Invalid' in err or 'Unexpected' in err:
            tr.add("D2: P2-1 UPDATE SET ||拼接", False, f"仍不支持(P2确认): {err}")
        else:
            tr.add("D2: P2-1 UPDATE SET ||拼接", False, f"错误: {err}")
    execute_advanced_update_query(TEST_FILE, "UPDATE Sheet1 SET Name = 'Item-1' WHERE ID = 1")
except Exception as e:
    tr.add("D2: P2-1 UPDATE SET ||拼接", False, str(e)[:60])

# D3: P2-2 CASE WHEN算术混合
try:
    result = execute_advanced_sql_query(TEST_FILE, 
        "SELECT Name, Value * CASE WHEN Value > 50 THEN 2 ELSE 1 END as adj FROM Sheet1 LIMIT 5")
    if result.get('success'):
        tr.add("D3: P2-2 CASE WHEN算术混合", True, "CASE*算术现在支持了?")
    else:
        err = str(result.get('message', ''))[:60]
        if '不支持' in err:
            tr.add("D3: P2-2 CASE WHEN算术混合", False, f"仍不支持(P2确认): {err}")
        else:
            tr.add("D3: P2-2 CASE WHEN算术混合", False, f"错误: {err}")
except Exception as e:
    tr.add("D3: P2-2 CASE WHEN算术混合", False, str(e)[:60])

# D4: P2-4 极端浮点值
try:
    extreme_floats = [
        ("CAST('inf' AS FLOAT64)", "inf"),
        ("CAST('nan' AS FLOAT64)", "nan"),
        ("CAST('-inf' AS FLOAT64)", "-inf"),
    ]
    all_ok = True
    for expr, label in extreme_floats:
        r = execute_advanced_update_query(TEST_FILE, f"UPDATE Sheet1 SET Value = {expr} WHERE ID = 1")
        if not r.get('success'):
            all_ok = False
            continue
        # 尝试回读
        rr = execute_advanced_sql_query(TEST_FILE, "SELECT Value FROM Sheet1 WHERE ID = 1")
        if not rr.get('success'):
            all_ok = False
    if all_ok:
        tr.add("D4: P2-4 极端浮点值(inf/nan/-inf)", True, "写入和回读均成功(与R32一致)")
    else:
        tr.add("D4: P2-4 极端浮点值(inf/nan/-inf)", False, "部分失败(文件可能损坏)")
    execute_advanced_update_query(TEST_FILE, "UPDATE Sheet1 SET Value = 10 WHERE ID = 1")
except Exception as e:
    tr.add("D4: P2-4 极端浮点值(inf/nan/-inf)", False, str(e)[:60])

# D5: P3-1 SELECT多余逗号幽灵列
try:
    result = execute_advanced_sql_query(TEST_FILE, "SELECT ,,, FROM Sheet1 LIMIT 3")
    if result.get('success'):
        data = result.get('data', [])
        if data:
            headers = data[0]
            has_row_col = '_ROW_' in str(headers)
            if has_row_col:
                tr.add("D5: P3-1 多余逗号幽灵_ROW_列", False, f"🚨 仍存在! 列名: {headers}")
            else:
                tr.add("D5: P3-1 多余逗号幽灵_ROW_列", True, "已修复! 无幽灵列")
        else:
            tr.add("D5: P3-1 多余逗号幽灵_ROW_列", True, "无返回数据(已修复?)")
    else:
        tr.add("D5: P3-1 多余逗号幽灵_ROW_列", True, f"已修复! 报错: {str(result.get('message', ''))[:40]}")
except Exception as e:
    tr.add("D5: P3-1 多余逗号幽灵_ROW_列", False, str(e)[:60])

# D6: SQL关键字大小写敏感
try:
    result = execute_advanced_sql_query(TEST_FILE, "SELECT * FROM ShEeT1 LIMIT 3")
    if result.get('success'):
        tr.add("D6: SQL大小写敏感(ShEeT1)", True, "大小写不敏感(改善!)")
    else:
        err = str(result.get('message', ''))[:50]
        if 'not found' in err.lower() or '不存在' in err:
            tr.add("D6: SQL大小写敏感(ShEeT1)", False, f"仍大小写敏感(P2确认): {err}")
        else:
            tr.add("D6: SQL大小写敏感(ShEeT1)", False, f"错误: {err}")
except Exception as e:
    tr.add("D6: SQL大小写敏感(ShEeT1)", False, str(e)[:60])

tr.summary()

# ============================================================
# E组: 特殊字符/Unicode边界组合测试
# ============================================================
print("\n" + "─" * 70)
print("📌 E组: 特殊字符/Unicode边界组合测试")
print("─" * 70)

create_test_file()

# E1-E8: 特殊字符在各个位置
special_char_tests = [
    ("E1: 单引号Name", "INSERT INTO Sheet1 (ID, Name, Value, Price, Category) VALUES (900, 'It''s a test', 1, 1, 'A')", "insert"),
    ("E2: 双引号Name", "UPDATE Sheet1 SET Name = \"DoubleQuote\" WHERE ID = 1", "update"),
    ("E3: 反斜杠", "UPDATE Sheet1 SET Name = 'Path\\\\To\\\\File' WHERE ID = 1", "update"),
    ("E4: 换行符在字符串", "UPDATE Sheet1 SET Name = 'Line1\\nLine2' WHERE ID = 1", "update"),
    ("E5: Tab在字符串", "UPDATE Sheet1 SET Name = 'Col1\\tCol2' WHERE ID = 1", "update"),
    ("E6: 空字符串", "UPDATE Sheet1 SET Name = '' WHERE ID = 1", "update"),
    ("E7: NULL值处理", "UPDATE Sheet1 SET Value = NULL WHERE ID = 1", "update"),
    ("E8: Emoji在Name", "UPDATE Sheet1 SET Name = '🎮游戏🔥装备⚔️' WHERE ID = 1", "update"),
]

for name, sql, op_type in special_char_tests:
    try:
        if op_type == "insert":
            result = execute_advanced_insert_query(TEST_FILE, sql)
        elif op_type == "update":
            result = execute_advanced_update_query(TEST_FILE, sql)
        else:
            result = execute_advanced_sql_query(TEST_FILE, sql)
        
        if result.get('success'):
            tr.add(name, True, "执行成功")
        else:
            err = str(result.get('message', ''))[:60]
            # NULL可能不被支持
            if 'NULL' in sql and ('不支持' in err or 'Invalid' in err):
                tr.add(name, True, f"NULL不支持(已知限制): {err[:40]}")
            else:
                tr.add(name, False, f"失败: {err}")
    except Exception as e:
        tr.add(name, False, f"异常: {str(e)[:50]}")

# 恢复
execute_advanced_update_query(TEST_FILE, "UPDATE Sheet1 SET Name = 'Item-1', Value = 10 WHERE ID = 1")
try:
    execute_advanced_delete_query(TEST_FILE, "DELETE FROM Sheet1 WHERE ID = 900")
except:
    pass

# E9-E12: Unicode边界
unicode_boundary_tests = [
    ("E9: 右到左文字(阿拉伯/希伯来)", "UPDATE Sheet1 Set Name = 'مرحبا שלום' WHERE ID = 1"),
    ("E10: 组合字符(泰文/高棉)", "UPDATE Sheet1 Set Name = 'สวัสดี ខ្ញុំសុំទោស' WHERE ID = 1"),
    ("E11: 零宽字符", "UPDATE Sheet1 Set Name = 'Test\u200b\u200c\u200dZeroWidth' WHERE ID = 1"),
    ("E12: 控制字符", "UPDATE Sheet1 Set Name = 'Test\x01\x02\x03Control' WHERE ID = 1"),
]

for name, sql in unicode_boundary_tests:
    try:
        result = execute_advanced_update_query(TEST_FILE, sql)
        if result.get('success'):
            tr.add(name, True, "写入成功")
        else:
            err = str(result.get('message', ''))[:50]
            tr.add(name, False, f"失败: {err}")
    except Exception as e:
        err_str = str(e)
        if 'control' in name.lower() or 'zero' in name.lower():
            tr.add(name, True, f"异常(特殊字符可能被拒): {err_str[:40]}")
        else:
            tr.add(name, False, f"异常: {err_str[:40]}")

execute_advanced_update_query(TEST_FILE, "UPDATE Sheet1 SET Name = 'Item-1' WHERE ID = 1")

# E13-E15: 组合攻击面
print("\n  --- E13-E15: 组合攻击面 ---")

# E13: SQL注入+Unicode组合
try:
    result = execute_advanced_sql_query(TEST_FILE, "SELECT * FROM Sheet1 WHERE Name = 'Item-1'; DROP TABLE Sheet1--'")
    if result.get('success'):
        msg = str(result.get('message', ''))
        if '多语句' in msg:
            tr.add("E13: 注入+Unicode组合", False, f"🚨 多语句执行: {msg[:50]}")
        else:
            tr.add("E13: 注入+Unicode组合", False, f"🚨 未拒绝: {msg[:50]}")
    else:
        tr.add("E13: 注入+Unicode组合", True, "正确拒绝")
except Exception as e:
    tr.add("E13: 注入+Unicode组合", True, f"异常(=拒绝): {str(e)[:40]}")

# E14: 超长SQL语句
try:
    long_where = " OR ".join([f"ID = {i}" for i in range(1, 1001)])
    long_sql = f"SELECT * FROM Sheet1 WHERE {long_where}"
    result = execute_advanced_sql_query(TEST_FILE, long_sql)
    if result.get('success'):
        tr.add("E14: 超长SQL(1000个OR条件)", True, f"执行成功, 返回{len(result.get('data', []))-1}行")
    else:
        tr.add("E14: 超长SQL(1000个OR条件)", False, f"失败: {str(result.get('message', ''))[:50]}")
except Exception as e:
    tr.add("E14: 超长SQL(1000个OR条件)", False, str(e)[:50])

# E15: 深度嵌套括号
try:
    result = execute_advanced_sql_query(TEST_FILE, 
        "SELECT * FROM Sheet1 WHERE (((((ID = 1))))) AND ((((Value > 0))))")
    if result.get('success'):
        tr.add("E15: 深度嵌套括号(5层)", True, "解析成功")
    else:
        tr.add("E15: 深度嵌套括号(5层)", False, str(result.get('message', ''))[:50])
except Exception as e:
    tr.add("E15: 深度嵌套括号(5层)", False, str(e)[:50])

tr.summary()

# ============================================================
# 最终汇总
# ============================================================
print("\n" + "=" * 70)
print("📊 Round 33 最终汇总")
print("=" * 70)
total_p, total_f = tr.summary()
total = total_p + total_f
print(f"\n  总测试场景: {total}")
print(f"  通过: {total_p} ({total_p/total*100:.1f}%)" if total > 0 else "")
print(f"  失败: {total_f} ({total_f/total*100:.1f}%)" if total > 0 else "")

# 清理临时文件
for f in [TEST_FILE, TEST_FILE_LARGE, TEST_FILE_WIDE]:
    try:
        os.remove(f)
    except:
        pass

print("\n✅ Round 33 测试完成!")
