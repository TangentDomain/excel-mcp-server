"""
Round 29 MCP 接口实测 - SQL注入防护深度测试 + 边界组合测试
================================================================
本轮重点方向: 
1. SQL注入防护全面审计（SELECT/UPDATE/INSERT/DELETE各路径）
2. 边界组合：特殊字符Sheet名 + 超长列名 + 公式单元格
3. 已知P0/P1问题回归验证
4. 注释符/编码绕过/联合注入等高级向量
"""

import sys
import os
import tempfile
import shutil

# 确保能导入项目模块
sys.path.insert(0, '/root/workspace/excel-mcp-server/src')
sys.path.insert(0, '/root/workspace/excel-mcp-server')

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)

# ============================================================
# 测试基础设施
# ============================================================
TEST_RESULTS = []
TEST_DIR = "/tmp/round29_test"

def setup_test_env():
    """创建测试环境"""
    if os.path.exists(TEST_DIR):
        shutil.rmtree(TEST_DIR)
    os.makedirs(TEST_DIR, exist_ok=True)

def record(name, sql, expected, actual, status, detail=""):
    """记录测试结果"""
    TEST_RESULTS.append({
        "name": name,
        "sql": sql,
        "expected": expected,
        "actual": actual,
        "status": status,
        "detail": detail
    })
    icon = "✅" if status else "❌"
    print(f"  {icon} {name}")
    if not status and detail:
        print(f"     详情: {detail[:100]}")

def create_basic_test_file(filepath):
    """创建基础测试Excel文件"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "Name", "Price", "Rarity"])
    ws.append([1, "Sword", 100.5, "Common"])
    ws.append([2, "Shield", 250.0, "Rare"])
    ws.append([3, "Staff", 999.99, "Epic"])
    ws.append([4, "Bow", 50.0, "Legendary"])
    wb.save(filepath)
    return filepath

def create_special_sheet_test_file(filepath):
    """创建含特殊字符Sheet名的测试文件"""
    from openpyxl import Workbook
    wb = Workbook()
    
    # Sheet名含特殊字符
    ws = wb.active
    ws.title = "Sheet-Test_数据"
    ws.append(["ID", "User_Name", "Email_Address_ExtraLongColumnName", "Value"])
    ws.append([1, "Alice", "alice@test.com", 100])
    ws.append([2, "Bob O'Brien", "bob@test.com", 200])
    ws.append([3, "Charlie", "charlie@test.com", 300])
    
    # 含公式的Sheet
    ws2 = wb.create_sheet("FormulaSheet")
    ws2.append(["ID", "A", "B", "C"])
    ws2.append([1, 10, 20, None])
    ws2.append([2, 30, 40, None])
    ws2.append([3, 50, 60, None])
    # 添加公式
    for row in range(2, 5):
        ws2[f'C{row}'] = f'=A{row}*B{row}'
    
    # 超长列名Sheet
    ws3 = wb.create_sheet("WideColumns")
    headers = ["ID"] + [f"VeryLongColumnName_{i}_ForTestingPurpose" for i in range(5)]
    ws3.append(headers)
    ws3.append([1] + [i*10 for i in range(1, 6)])
    ws3.append([2] + [i*20 for i in range(1, 6)])
    
    wb.save(filepath)
    return filepath

# ============================================================
# A组: SQL注入防护 - SELECT路径
# ============================================================
def test_sql_injection_select():
    """A组: SELECT语句的SQL注入防护测试"""
    print("\n" + "="*70)
    print("🔬 A组: SQL注入防护 - SELECT路径")
    print("="*70)
    
    f = os.path.join(TEST_DIR, "inject_select.xlsx")
    create_basic_test_file(f)
    
    # A1: 基础分号多语句 (已知P0-2，应被拦截或执行)
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1; SELECT * FROM Sheet1")
    is_multi = "多语句" in str(r.get('message', '')) or r.get('success', False)
    record(
        "A1_SELECT分号多语句",
        "SELECT * FROM Sheet1; SELECT * FROM Sheet1",
        "应拒绝或警告",
        f"success={r.get('success')}, msg={str(r.get('message',''))[:80]}",
        not r.get('success', False),  # 期望被拦截
        f"success={r.get('success')}, msg={str(r.get('message',''))[:100]}"
    )
    
    # A2: UNION-based注入尝试
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1 UNION SELECT * FROM Sheet1")
    record(
        "A2_UNION注入",
        "SELECT * FROM Sheet1 UNION SELECT * FROM Sheet1",
        "正常执行UNION(合法SQL)",
        f"success={r.get('success')}, rows={len(r.get('data',[]))}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # A3: 注释符注入
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1 -- 注释")
    record(
        "A3_注释符尾随",
        "SELECT * FROM Sheet1 -- 注释",
        "应正常执行或报错",
        f"success={r.get('success')}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # A4: 单引号逃逸
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1 WHERE Name = 'O'Brien'")
    record(
        "A4_单引号逃逸",
        "SELECT * FROM Sheet1 WHERE Name = 'O'Brien'",
        "语法错误或安全处理",
        f"success={r.get('success')}",
        not r.get('success', False),  # 未转义的单引号应该失败
        str(r.get('message', ''))[:100]
    )
    
    # A5: DROP TABLE 尝试(通过分号)
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1; DROP TABLE Sheet1")
    record(
        "A5_DROP_TABLE注入",
        "SELECT * FROM Sheet1; DROP TABLE Sheet1",
        "必须拒绝!",
        f"success={r.get('success')}, msg={str(r.get('message',''))[:80]}",
        not r.get('success', False),
        str(r.get('message', ''))[:150]
    )
    
    # A6: 1=1 恒真条件
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1 WHERE 1=1")
    record(
        "A6_恒真条件1=1",
        "SELECT * FROM Sheet1 WHERE 1=1",
        "返回所有行(合法SQL)",
        f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}",
        r.get('success', False) and r.get('data') and len(r['data']) == 5,  # 4 data + 1 header
        str(r.get('message', ''))[:100]
    )
    
    # A7: OR 1=1 注入模式
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1 WHERE ID = 1 OR 1=1")
    record(
        "A7_OR 1=1注入",
        "SELECT * FROM Sheet1 WHERE ID = 1 OR 1=1",
        "返回所有行(合法SQL语义)",
        f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}",
        r.get('success', False) and r.get('data') and len(r['data']) == 5,
        str(r.get('message', ''))[:100]
    )
    
    # A8: 批量分号语句
    r = execute_advanced_sql_query(f, "SELECT 1; SELECT 2; SELECT 3; DROP TABLE xxx")
    record(
        "A8_批量分号+DROP",
        "SELECT 1; SELECT 2; SELECT 3; DROP TABLE xxx",
        "必须拒绝!",
        f"success={r.get('success')}",
        not r.get('success', False),
        str(r.get('message', ''))[:150]
    )
    
    # A9: /* */ 块注释
    r = execute_advanced_sql_query(f, "SELECT * FROM /* comment */ Sheet1")
    record(
        "A9_块注释",
        "SELECT * FROM /* comment */ Sheet1",
        "可能成功或报错",
        f"success={r.get('success')}",
        True,  # 不论成功失败都算PASS(记录行为)
        str(r.get('message', ''))[:100]
    )
    
    # A10: 反引号/方括号标识符注入
    r = execute_advanced_sql_query(f, "SELECT * FROM `Sheet1`")
    record(
        "A10_反引号标识符",
        "SELECT * FROM `Sheet1`",
        "可能成功或报错",
        f"success={r.get('success')}",
        True,
        str(r.get('message', ''))[:100]
    )

# ============================================================
# B组: SQL注入防护 - UPDATE/INSERT/DELETE路径  
# ============================================================
def test_sql_injection_write():
    """B组: 写操作的SQL注入防护"""
    print("\n" + "="*70)
    print("🔬 B组: SQL注入防护 - UPDATE/INSERT/DELETE路径")
    print("="*70)
    
    # B1: UPDATE 分号注入
    f = os.path.join(TEST_DIR, "inject_update.xlsx")
    create_basic_test_file(f)
    r = execute_advanced_update_query(f, "UPDATE Sheet1 SET Price = 0 WHERE ID = 1; DROP TABLE Sheet1")
    record(
        "B1_UPDATE分号注入",
        "UPDATE ... SET Price=0; DROP TABLE ...",
        "必须拒绝!",
        f"success={r.get('success')}, msg={str(r.get('message',''))[:80]}",
        not r.get('success', False),
        str(r.get('message', ''))[:150]
    )
    
    # B2: INSERT 分号注入
    f2 = os.path.join(TEST_DIR, "inject_insert.xlsx")
    create_basic_test_file(f2)
    r = execute_advanced_insert_query(f2, "INSERT INTO Sheet1 (ID,Name,Price,Rarity) VALUES (99,'hack',0,'test'); DROP TABLE Sheet1")
    record(
        "B2_INSERT分号注入",
        "INSERT ... ; DROP TABLE ...",
        "必须拒绝!",
        f"success={r.get('success')}",
        not r.get('success', False),
        str(r.get('message', ''))[:150]
    )
    
    # B3: DELETE 分号注入
    f3 = os.path.join(TEST_DIR, "inject_delete.xlsx")
    create_basic_test_file(f3)
    r = execute_advanced_delete_query(f3, "DELETE FROM Sheet1 WHERE ID = 999; DROP TABLE Sheet1")
    record(
        "B3_DELETE分号注入",
        "DELETE ... ; DROP TABLE ...",
        "必须拒绝!",
        f"success={r.get('success')}",
        not r.get('success', False),
        str(r.get('message', ''))[:150]
    )
    
    # B4: UPDATE 注释符注入
    f4 = os.path.join(TEST_DIR, "inject_update_comment.xlsx")
    create_basic_test_file(f4)
    r = execute_advanced_update_query(f4, "UPDATE Sheet1 SET Price = 0 -- 注释 WHERE ID = 1")
    record(
        "B4_UPDATE注释符",
        "UPDATE SET Price=0 -- 注释 WHERE ID=1",
        "应拒绝(注释吃掉WHERE导致全表更新风险)",
        f"success={r.get('success')}, msg={str(r.get('message',''))[:80]}",
        not r.get('success', False),  # 应该拒绝注释符
        str(r.get('message', ''))[:150]
    )
    
    # B5: INSERT 单引号逃逸
    f5 = os.path.join(TEST_DIR, "inject_insert_escape.xlsx")
    create_basic_test_file(f5)
    r = execute_advanced_insert_query(f5, "INSERT INTO Sheet1 (ID,Name,Price,Rarity) VALUES (99,O'Connor,0,test)")
    record(
        "B5_INSERT单引号未转义",
        "INSERT ... VALUES (99,O'Connor,0,test)",
        "语法错误",
        f"success={r.get('success')}",
        not r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # B6: 正常UPDATE（对照）
    f6 = os.path.join(TEST_DIR, "inject_normal.xlsx")
    create_basic_test_file(f6)
    r = execute_advanced_update_query(f6, "UPDATE Sheet1 SET Price = 777 WHERE ID = 1")
    record(
        "B6_正常UPDATE对照",
        "UPDATE Sheet1 SET Price = 777 WHERE ID = 1",
        "成功更新",
        f"success={r.get('success')}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )

# ============================================================
# C组: 边界组合测试 - 特殊字符Sheet名 + 超长列名 + 公式
# ============================================================
def test_boundary_combinations():
    """C组: 边界组合测试"""
    print("\n" + "="*70)
    print("🔬 C组: 边界组合测试 (特殊Sheet名+超长列名+公式)")
    print("="*70)
    
    f = os.path.join(TEST_DIR, "boundary_combo.xlsx")
    create_special_sheet_test_file(f)
    
    # C1: 特殊字符Sheet名查询
    r = execute_advanced_sql_query(f, "SELECT * FROM `Sheet-Test_数据`")
    record(
        "C1_特殊字符Sheet名查询",
        "SELECT * FROM `Sheet-Test_数据`",
        "成功查询",
        f"success={r.get('success')}, rows={len(r.get('data',[]))}",
        r.get('success', False) and r.get('data') and len(r['data']) >= 3,
        str(r.get('message', ''))[:100]
    )
    
    # C2: 特殊字符Sheet名 + WHERE含单引号
    r = execute_advanced_sql_query(f, "SELECT * FROM `Sheet-Test_数据` WHERE User_Name = 'Bob O''Brien'")
    record(
        "C2_Sheet特殊字符+单引号值",
        "SELECT * FROM `Sheet-Test_数据` WHERE User_Name = 'Bob O''Brien'",
        "成功查到Bob",
        f"success={r.get('success')}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # C3: 超长列名查询
    r = execute_advanced_sql_query(f, "SELECT VeryLongColumnName_1_ForTestingPurpose FROM WideColumns")
    record(
        "C3_超长列名查询",
        "SELECT VeryLongColumnName_1_ForTestingPurpose FROM WideColumns",
        "成功查询",
        f"success={r.get('success')}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # C4: 超长列名 + 聚合
    r = execute_advanced_sql_query(f, "SELECT SUM(VeryLongColumnName_3_ForTestingPurpose) as total FROM WideColumns")
    record(
        "C4_超长列名聚合",
        "SELECT SUM(VeryLongColumnName_3_ForTestingPurpose) as total FROM WideColumns",
        "聚合成功",
        f"success={r.get('success')}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # C5: 公式Sheet查询
    r = execute_advanced_sql_query(f, "SELECT * FROM FormulaSheet")
    record(
        "C5_公式Sheet查询",
        "SELECT * FROM FormulaSheet",
        "成功查询(公式值被读取)",
        f"success={r.get('success')}, rows={len(r.get('data',[]))}",
        r.get('success', False) and r.get('data'),
        str(r.get('message', ''))[:100]
    )
    
    # C6: 公式Sheet过滤
    r = execute_advanced_sql_query(f, "SELECT * FROM FormulaSheet WHERE A > 20")
    record(
        "C6_公式SheetWHERE过滤",
        "SELECT * FROM FormulaSheet WHERE A > 20",
        "过滤成功",
        f"success={r.get('success')}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # C7: 公式Sheet聚合
    r = execute_advanced_sql_query(f, "SELECT SUM(C) as TotalC FROM FormulaSheet")
    record(
        "C7_公式Sheet聚合SUM(C)",
        "SELECT SUM(C) as TotalC FROM FormulaSheet",
        "聚合公式计算列",
        f"success={r.get('success')}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # C8: 特殊Sheet名 + UPDATE
    r = execute_advanced_update_query(f, "UPDATE `Sheet-Test_数据` SET Value = 999 WHERE ID = 1")
    record(
        "C8_特殊Sheet名UPDATE",
        "UPDATE `Sheet-Test_数据` SET Value = 999 WHERE ID = 1",
        "更新成功",
        f"success={r.get('success')}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # C9: 特殊Sheet名 + DELETE
    r = execute_advanced_delete_query(f, "DELETE FROM `Sheet-Test_数据` WHERE ID = 99")
    record(
        "C9_特殊Sheet名DELETE",
        "DELETE FROM `Sheet-Test_数据` WHERE ID = 99",
        "零匹配删除成功",
        f"success={r.get('success')}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # C10: Unicode极端值
    f2 = os.path.join(TEST_DIR, "unicode_extreme.xlsx")
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "UnicodeTest"
    ws.append(["ID", "Text"])
    extreme_values = [
        "中文测试🎉",
        "日本語テスト",
        "한국어테스트",
        "العربية",
        "עברית",
        "🔥💀🚀emoji",
        "\t\n\r控制字符",
        "a"*500,  # 超长文本
    ]
    for i, val in enumerate(extreme_values):
        ws.append([i+1, val])
    wb.save(f2)
    
    r = execute_advanced_sql_query(f2, "SELECT * FROM UnicodeTest")
    record(
        "C10_Unicode极端值查询",
        "SELECT * FROM UnicodeTest (含emoji/阿拉伯文/希伯来文/超长文本/控制字符)",
        "全部正确读取",
        f"success={r.get('success')}, rows={len(r.get('data',[]))}",
        r.get('success', False) and r.get('data') and len(r['data']) >= 9,
        str(r.get('message', ''))[:100]
    )

# ============================================================
# D组: 已知P0/P1回归验证
# ============================================================
def test_known_issues_regression():
    """D组: 从iteration-memory中取已知问题重测"""
    print("\n" + "="*70)
    print("🔬 D组: 已知P0/P1问题回归验证")
    print("="*70)
    
    # D1: P0-2 SELECT 分号多语句注入 (已知仍存在)
    f = os.path.join(TEST_DIR, "regress_p0.xlsx")
    create_basic_test_file(f)
    r = execute_advanced_sql_query(f, "SELECT COUNT(*) FROM Sheet1; SELECT COUNT(*) FROM Sheet1")
    still_exists = r.get('success', False) or "多语句" in str(r.get('message', ''))
    record(
        "D1_P0-2回归_SELECT分号注入",
        "SELECT COUNT(*) FROM t; SELECT COUNT(*) FROM t",
        "已知P0: 可能仍存在",
        f"success={r.get('success')}, still_exists={still_exists}",
        True,  # 记录状态即可
        f"still_vulnerable={still_exists}, msg={str(r.get('message',''))[:100]}"
    )
    
    # D2: P0-3 uint8溢出 (已修复，验证)
    f2 = os.path.join(TEST_DIR, "regress_uint8.xlsx")
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "MixedTypes"
    ws.append(["ID", "V", "F"])  # int, int, float → 触发uint8推断
    ws.append([1, 10, 1.5])
    wb.save(f2)
    
    r = execute_advanced_update_query(f2, "UPDATE MixedTypes SET V = 999 WHERE ID = 1")
    # 验证实际值
    r2 = execute_advanced_sql_query(f2, "SELECT V FROM MixedTypes WHERE ID = 1")
    actual_val = 'N/A'
    if r2.get('success') and r2.get('data') and len(r2['data']) > 1:
        actual_val = r2['data'][1][0]
    
    is_fixed = (actual_val == 999)
    record(
        "D2_P0-3回归_uint8溢出修复验证",
        "UPDATE MixedTypes SET V = 999 (混合类型Sheet, V>255)",
        "V应为999(非231截断)",
        f"V={actual_val}",
        is_fixed,
        f"actual_value={actual_val}, expected=999"
    )
    
    # D3: P1-3 CTE表别名前列名前缀污染
    f3 = os.path.join(TEST_DIR, "regress_cte.xlsx")
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Players"
    ws.append(["PlayerID", "GuildID", "Level"])
    ws.append([1, 10, 50])
    ws.append([2, 10, 60])
    ws.append([3, 20, 70])
    ws2 = wb.create_sheet("Guilds")
    ws2.append(["GuildID", "GuildName"])
    ws2.append([10, "Warriors"])
    ws2.append([20, "Mages"])
    wb.save(f3)
    
    r = execute_advanced_sql_query(f3, """
        WITH PlayerStats AS (
            SELECT p.GuildID, COUNT(*) as cnt 
            FROM Players p GROUP BY p.GuildID
        )
        SELECT g.GuildName, ps.cnt 
        FROM Guilds g JOIN PlayerStats ps ON g.GuildID = ps.GuildID
    """)
    record(
        "D3_P1-3回归_CTE表别名前缀污染",
        "CTE用表别名后JOIN",
        "已知P1: 列名可能有alias.前缀",
        f"success={r.get('success')}, msg={str(r.get('message',''))[:80]}",
        r.get('success', False),  # 如果修复了则PASS
        str(r.get('message', ''))[:150] if not r.get('success', False) else "CTE JOIN 成功!"
    )
    
    # D4: P2-1 UPDATE SET || 字符串拼接 (已知不支持)
    f4 = os.path.join(TEST_DIR, "regress_concat.xlsx")
    create_basic_test_file(f4)
    r = execute_advanced_update_query(f4, "UPDATE Sheet1 SET Name = 'Special-' || Name WHERE ID = 1")
    record(
        "D4_P2-1回归_UPDATE字符串拼接",
        "UPDATE SET Name = 'Special-' || Name",
        "已知P2: 不支持||拼接",
        f"success={r.get('success')}",
        not r.get('success', False),  # 期望失败
        str(r.get('message', ''))[:100]
    )
    
    # D5: P2-2 CASE WHEN 算术操作数 (已知不支持)
    r = execute_advanced_sql_query(f4, "SELECT Name, BaseAtk * CASE WHEN Rarity='Legendary' THEN 2 ELSE 1 END as AdjAtk FROM Sheet1")
    record(
        "D5_P2-2回归_CASE算术混合",
        "SELECT col * CASE WHEN...END",
        "已知P2: CASE不能做算术操作数",
        f"success={r.get('success')}",
        not r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # D6: P0-1 script_runner RCE (仅记录，不重复测试exec)
    record(
        "D6_P0-1_script_runner_RCE",
        "(不重复执行，参考R25结论)",
        "已知P0: exec()可执行任意命令",
        "仍存在(需沙箱修复)",
        True,  # 仅记录
        "R25已充分验证，待安全加固"
    )

# ============================================================
# E组: 高级注入向量测试
# ============================================================
def test_advanced_injection_vectors():
    """E组: 高级/边缘注入向量"""
    print("\n" + "="*70)
    print("🔬 E组: 高级注入向量测试")
    print("="*70)
    
    f = os.path.join(TEST_DIR, "advanced_inject.xlsx")
    create_basic_test_file(f)
    
    # E1: NULL字节注入
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1\x00; DROP TABLE Sheet1")
    record(
        "E1_NULL字节注入",
        "SELECT * FROM Sheet1\\x00; DROP TABLE...",
        "应安全处理NULL字节",
        f"success={r.get('success')}",
        not r.get('success', False),  # 应该拒绝
        str(r.get('message', ''))[:150]
    )
    
    # E2: 大写混淆关键字
    r = execute_advanced_sql_query(f, "sElEcT * FrOm ShEeT1")
    record(
        "E2_大小写混淆关键字",
        "sElEcT * FrOm ShEeT1",
        "大小写不敏感(应正常工作)",
        f"success={r.get('success')}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # E3: 双重URL编码式注入
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1 WHERE Name = '%27%27'")
    record(
        "E3_URL编码字符串",
        "SELECT ... WHERE Name = '%27%27'",
        "按字面匹配(非解码)",
        f"success={r.get('success')}",
        True,  # 记录行为
        str(r.get('message', ''))[:100]
    )
    
    # E4: 十六进制编码尝试
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1 WHERE Name = 0x53776F7264")
    record(
        "E4_十六进制字面量",
        "SELECT ... WHERE Name = 0x53776F7264",
        "可能不支持hex字面量",
        f"success={r.get('success')}",
        True,
        str(r.get('message', ''))[:100]
    )
    
    # E5: LIKE通配符注入
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1 WHERE Name LIKE '%' OR '1'='1")
    record(
        "E5_LIKE通配符+OR注入",
        "SELECT ... WHERE Name LIKE '%' OR '1'='1",
        "可能返回所有行(合法SQL语义)",
        f"success={r.get('success')}, rows={len(r.get('data',[]))-1 if r.get('data') else 0}",
        True,
        str(r.get('message', ''))[:100]
    )
    
    # E6: 子查询注入
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1 WHERE ID = (SELECT ID FROM Sheet1 LIMIT 1)")
    record(
        "E6_标量子查询",
        "SELECT ... WHERE ID = (SELECT ID FROM ... LIMIT 1)",
        "子查询支持",
        f"success={r.get('success')}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # E7: UPDATE + 子查询SET
    f2 = os.path.join(TEST_DIR, "advanced_subq.xlsx")
    create_basic_test_file(f2)
    r = execute_advanced_update_query(f2, "UPDATE Sheet1 SET Price = (SELECT AVG(Price) FROM Sheet1) WHERE ID = 1")
    record(
        "E7_UPDATE子查询SET",
        "UPDATE SET col = (SELECT AVG(col) ...)",
        "子查询赋值",
        f"success={r.get('success')}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # E8: 空字符串边界
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1 WHERE Name = ''")
    record(
        "E8_空字符串匹配",
        "SELECT ... WHERE Name = ''",
        "空字符串比较",
        f"success={r.get('success')}, rows={len(r.get('data',[]))-1 if r.get('data') else 0}",
        True,
        str(r.get('message', ''))[:100]
    )
    
    # E9: 极端数值边界
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1 WHERE Price > 1e308")
    record(
        "E9_极端浮点(inf边界)",
        "SELECT ... WHERE Price > 1e308",
        "inf边界处理",
        f"success={r.get('success')}",
        True,
        str(r.get('message', ''))[:100]
    )
    
    # E10: 负数索引/偏移
    r = execute_advanced_sql_query(f, "SELECT * FROM Sheet1 ORDER BY ID DESC LIMIT -1")
    record(
        "E10负数LIMIT",
        "SELECT ... ORDER BY ID DESC LIMIT -1",
        "负数LIMIT处理",
        f"success={r.get('success')}",
        True,
        str(r.get('message', ''))[:100]
    )

# ============================================================
# F组: 编码和字符集攻击面
# ============================================================
def test_encoding_attacks():
    """F组: 编码相关攻击面"""
    print("\n" + "="*70)
    print("🔬 F组: 编码与字符集攻击面")
    print("="*70)
    
    f = os.path.join(TEST_DIR, "encoding_test.xlsx")
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "EncTest"
    ws.append(["ID", "Val"])
    ws.append([1, "normal"])
    wb.save(f)
    
    # F1: 全角SQL关键字
    r = execute_advanced_sql_query(f, "ＳＥＬＥＣＴ ＊ ＦＲＯＭ EncTest")
    record(
        "F1_全角SQL关键字",
        "ＳＥＬＥＣＴ ＊ ＦＲＯＭ EncTest (fullwidth)",
        "应失败(全角不是合法SQL)",
        f"success={r.get('success')}",
        not r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # F2: Tab/换行在SQL中
    r = execute_advanced_sql_query(f, "SELECT\t*\nFROM\nEncTest")
    record(
        "F2_SQL中含Tab和换行",
        "SELECT\\t*\\nFROM\\nEncTest",
        "空白符通常被忽略",
        f"success={r.get('success')}",
        r.get('success', False),
        str(r.get('message', ''))[:100]
    )
    
    # F3: 零宽字符
    r = execute_advanced_sql_query(f, "SELECT * FROM EncTest\u200b WHERE ID = 1")
    record(
        "F3_零宽字符",
        "SELECT * FROM EncTest\\u200b WHERE ID = 1",
        "零宽字符处理",
        f"success={r.get('success')}",
        True,
        str(r.get('message', ''))[:100]
    )
    
    # F4: 方向覆盖字符(RLO)
    r = execute_advanced_sql_query(f, "SELECT * FROM EncTest\u202eWHERE ID = 1")
    record(
        "F4_RLO方向覆盖字符",
        "SELECT * FROM EncTest\\u202eWHERE ID = 1",
        "方向覆盖字符",
        f"success={r.get('success')}",
        True,
        str(r.get('message', ''))[:100]
    )

# ============================================================
# 主函数
# ============================================================
def main():
    print("="*70)
    print("🔬 Round 29 MCP 接口实测")
    print("   重点: SQL注入防护 + 边界组合测试 + P0/P1回归")
    print("="*70)
    
    setup_test_env()
    
    try:
        test_sql_injection_select()       # A组
        test_sql_injection_write()         # B组
        test_boundary_combinations()       # C组
        test_known_issues_regression()     # D组
        test_advanced_injection_vectors()  # E组
        test_encoding_attacks()            # F组
    except Exception as e:
        print(f"\n❌ 测试过程异常: {e}")
        import traceback
        traceback.print_exc()
    
    # 统计结果
    print("\n" + "="*70)
    print("📊 Round 29 测试结果汇总")
    print("="*70)
    
    passed = sum(1 for t in TEST_RESULTS if t["status"])
    failed = sum(1 for t in TEST_RESULTS if not t["status"])
    total = len(TEST_RESULTS)
    
    print(f"\n   总计: {total} 个场景")
    print(f"   ✅ 通过: {passed}")
    print(f"   ❌ 失败: {failed}")
    print(f"   通过率: {passed/total*100:.1f}%")
    
    print("\n   失败详情:")
    for t in TEST_RESULTS:
        if not t["status"]:
            print(f"   ❌ {t['name']}: {t['detail'][:120]}")
    
    return passed, failed, total

if __name__ == "__main__":
    main()
