# -*- coding: utf-8 -*-
"""
Excel Operations API 增强测试套件

覆盖 excel_operations.py 中未被充分测试的功能
"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from src.api.excel_operations import ExcelOperations


class TestExcelOperationsEnhanced:
    """ExcelOperations API 增强功能测试"""

    @pytest.fixture
    def test_excel_file(self, temp_dir):
        """创建测试用的Excel文件"""
        file_path = temp_dir / "test_operations_enhanced.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # 添加测试数据
        test_data = [
            ["ID", "Name", "Value", "Category"],
            [1, "Item1", 100, "A"],
            [2, "Item2", 200, "B"],
            [3, "Item3", 300, "A"],
            [4, "Item4", 400, "C"],
            [5, "Item5", 500, "B"],
        ]
        
        for row_idx, row_data in enumerate(test_data, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def multi_sheet_file(self, temp_dir):
        """创建多工作表Excel文件"""
        file_path = temp_dir / "multi_sheet.xlsx"
        
        wb = Workbook()
        
        # Sheet1
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1['A1'] = "ID"
        ws1['B1'] = "Name"
        ws1['A2'] = 1
        ws1['B2'] = "First"
        
        # Sheet2
        ws2 = wb.create_sheet(title="Sheet2")
        ws2['A1'] = "ID"
        ws2['B1'] = "Value"
        ws2['A2'] = 100
        ws2['B2'] = 200
        
        # Sheet3
        ws3 = wb.create_sheet(title="Sheet3")
        ws3['A1'] = "Name"
        ws3['B1'] = "Data"
        ws3['A2'] = "Test"
        ws3['B2'] = "Sample"
        
        wb.save(file_path)
        return str(file_path)

    # ==================== 搜索功能测试 ====================

    def test_search_basic(self, test_excel_file):
        """测试基础搜索"""
        result = ExcelOperations.search(
            file_path=test_excel_file,
            pattern="Item",
            sheet_name="TestSheet"
        )
        
        assert result['success'] is True

    def test_search_regex(self, test_excel_file):
        """测试正则表达式搜索"""
        result = ExcelOperations.search(
            file_path=test_excel_file,
            pattern=r"Item\d+",
            sheet_name="TestSheet",
            use_regex=True
        )
        
        assert result['success'] is True

    def test_search_case_sensitive(self, test_excel_file):
        """测试大小写敏感搜索"""
        result = ExcelOperations.search(
            file_path=test_excel_file,
            pattern="item",
            sheet_name="TestSheet",
            case_sensitive=True
        )
        
        assert result['success'] is True

    def test_search_whole_word(self, test_excel_file):
        """测试全词匹配搜索"""
        result = ExcelOperations.search(
            file_path=test_excel_file,
            pattern="Item",
            sheet_name="TestSheet",
            whole_word=True
        )
        
        assert result['success'] is True

    def test_search_with_range(self, test_excel_file):
        """测试指定范围搜索"""
        result = ExcelOperations.search(
            file_path=test_excel_file,
            pattern="Item",
            sheet_name="TestSheet",
            range="A1:B3"
        )
        
        assert result['success'] is True

    def test_search_all_sheets(self, multi_sheet_file):
        """测试搜索所有工作表"""
        result = ExcelOperations.search(
            file_path=multi_sheet_file,
            pattern="ID"
        )
        
        assert result['success'] is True

    # ==================== 目录搜索测试 ====================

    def test_search_directory(self, temp_dir_with_excel_files):
        """测试目录搜索"""
        result = ExcelOperations.search_directory(
            directory_path=temp_dir_with_excel_files,
            pattern="test"
        )
        
        assert result['success'] is True

    def test_search_directory_regex(self, temp_dir_with_excel_files):
        """测试目录正则搜索"""
        result = ExcelOperations.search_directory(
            directory_path=temp_dir_with_excel_files,
            pattern=r"\d+",
            use_regex=True
        )
        
        assert result['success'] is True

    def test_search_directory_recursive(self, temp_dir_with_excel_files):
        """测试递归目录搜索"""
        result = ExcelOperations.search_directory(
            directory_path=temp_dir_with_excel_files,
            pattern="test",
            recursive=True
        )
        
        assert result['success'] is True

    def test_search_directory_with_extension_filter(self, temp_dir_with_excel_files):
        """测试扩展名过滤"""
        result = ExcelOperations.search_directory(
            directory_path=temp_dir_with_excel_files,
            pattern="test",
            file_extensions=[".xlsx"]
        )
        
        assert result['success'] is True

    # ==================== 工具函数测试 ====================

    def test_get_sheet_headers(self, test_excel_file):
        """测试获取所有工作表表头"""
        result = ExcelOperations.get_sheet_headers(test_excel_file)
        
        assert result['success'] is True

    def test_get_file_info(self, test_excel_file):
        """测试获取文件信息"""
        result = ExcelOperations.get_file_info(test_excel_file)
        
        assert result['success'] is True

    def test_get_file_info_nonexistent(self):
        """测试获取不存在文件的信息"""
        result = ExcelOperations.get_file_info("/nonexistent/file.xlsx")
        
        assert result['success'] is False

    # ==================== 查找最后一行测试 ====================

    def test_find_last_row(self, test_excel_file):
        """测试查找最后一行"""
        result = ExcelOperations.find_last_row(
            file_path=test_excel_file,
            sheet_name="TestSheet"
        )
        
        assert result['success'] is True

    def test_find_last_row_with_data(self, test_excel_file):
        """测试查找有数据的最后一行"""
        result = ExcelOperations.find_last_row(
            file_path=test_excel_file,
            sheet_name="TestSheet",
            column="A"
        )
        
        assert result['success'] is True

    # ==================== 工作表管理测试 ====================

    def test_create_sheet(self, test_excel_file):
        """测试创建工作表"""
        result = ExcelOperations.create_sheet(
            file_path=test_excel_file,
            sheet_name="NewSheet"
        )
        
        assert result['success'] is True

    def test_delete_sheet(self, multi_sheet_file):
        """测试删除工作表"""
        result = ExcelOperations.delete_sheet(
            file_path=multi_sheet_file,
            sheet_name="Sheet3"
        )
        
        assert result['success'] is True
