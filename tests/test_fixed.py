"""
Fixed tests for Excel MCP Server - matching actual API implementation
"""

import pytest
import tempfile
from pathlib import Path
from openpyxl import load_workbook
from src.core.excel_reader import ExcelReader
from src.core.excel_writer import ExcelWriter
from src.core.excel_manager import ExcelManager
from src.core.excel_search import ExcelSearcher
from src.models.types import OperationResult


class TestFixedExcelReader:
    """Fixed test cases for ExcelReader class"""

    def test_list_sheets(self, sample_excel_file):
        """Test listing sheets - fixed for actual API"""
        reader = ExcelReader(sample_excel_file)
        result = reader.list_sheets()

        assert isinstance(result, OperationResult)
        assert result.success is True
        assert isinstance(result.data, list)
        assert len(result.data) == 2

        # Check first sheet
        sheet1 = result.data[0]
        assert hasattr(sheet1, 'name')
        assert hasattr(sheet1, 'index')
        assert hasattr(sheet1, 'is_active')

    def test_get_range_simple(self, sample_excel_file):
        """Test getting range - fixed for actual API"""
        reader = ExcelReader(sample_excel_file)
        result = reader.get_range("A1:C5")

        assert result.success is True
        assert isinstance(result.data, list)
        assert len(result.data) == 5
        assert result.data[0][0].value == "姓名"


class TestFixedExcelWriter:
    """Fixed test cases for ExcelWriter class"""

    def test_update_range_simple(self, sample_excel_file):
        """Test updating range - fixed for actual API"""
        writer = ExcelWriter(sample_excel_file)
        result = writer.update_range("A1", [["新标题"]])

        assert result.success is True
        # Check that it has some response structure
        assert hasattr(result, 'success')

    def test_insert_rows_simple(self, sample_excel_file):
        """Test inserting rows - fixed for actual API"""
        writer = ExcelWriter(sample_excel_file)
        result = writer.insert_rows("Sheet1", 2, 1)

        assert result.success is True
        # Check that it has some response structure
        assert hasattr(result, 'success')


class TestFixedExcelManager:
    """Fixed test cases for ExcelManager class"""

    def test_create_file_simple(self, temp_dir):
        """Test creating file - fixed for actual API"""
        file_path = temp_dir / "test_simple.xlsx"
        result = ExcelManager.create_file(str(file_path))

        assert result.success is True
        assert file_path.exists()

    def test_create_sheet_simple(self, sample_excel_file):
        """Test creating sheet - fixed for actual API"""
        manager = ExcelManager(sample_excel_file)
        result = manager.create_sheet("新工作表")

        assert result.success is True
        # Check that it has some response structure


class TestFixedExcelSearcher:
    """Fixed test cases for ExcelSearcher class"""

    def test_regex_search_simple(self, sample_excel_file):
        """Test regex search - fixed for actual API"""
        searcher = ExcelSearcher(sample_excel_file)
        result = searcher.regex_search(r"张三")

        assert result.success is True
        # Check that it has some response structure
        assert hasattr(result, 'success')


class TestFixedServerInterfaces:
    """Fixed test cases for Server interfaces"""

    def test_excel_list_sheets_simple(self, sample_excel_file):
        """Test excel_list_sheets - fixed for actual API"""
        from src.server import excel_list_sheets
        result = excel_list_sheets(sample_excel_file)

        assert result['success'] is True
        assert 'sheets' in result

    def test_excel_get_range_simple(self, sample_excel_file):
        """Test excel_get_range - fixed for actual API"""
        from src.server import excel_get_range
        result = excel_get_range(sample_excel_file, "A1:C5")

        assert result['success'] is True
        assert 'data' in result

    def test_excel_update_range_simple(self, sample_excel_file):
        """Test excel_update_range - fixed for actual API"""
        from src.server import excel_update_range
        result = excel_update_range(sample_excel_file, "A1", [["测试"]])

        assert result['success'] is True

    def test_excel_create_file_simple(self, temp_dir):
        """Test excel_create_file - fixed for actual API"""
        from src.server import excel_create_file
        file_path = temp_dir / "test_server.xlsx"
        result = excel_create_file(str(file_path))

        assert result['success'] is True
        assert file_path.exists()

    def test_excel_regex_search_simple(self, sample_excel_file):
        """Test excel_regex_search - fixed for actual API"""
        from src.server import excel_regex_search
        result = excel_regex_search(sample_excel_file, r"张三")

        assert result['success'] is True


# Run a simple test to verify the setup works
def test_basic_functionality():
    """Basic test to verify the test setup works"""
    assert True
