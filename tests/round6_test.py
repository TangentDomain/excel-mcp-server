"""
Round 6 迭代测试 - ExcelMCP 持续测试 (修复版)
==============================================
日期: 2026-04-13
版本: 1.9.3
方向: UNION深度测试 / EXISTS子查询 / 多Sheet操作 / 边缘场景
"""

import os
import sys
import json
import random
import tempfile
import shutil
from datetime import datetime

sys.path.insert(0, '/root/workspace/excel-mcp-server/src')

from openpyxl import Workbook

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)

# ============================================================
# 测试数据准备
# ============================================================

TEST_DIR = tempfile.mkdtemp(prefix='excelmcp_r6_')
TEST_FILE = os.path.join(TEST_DIR, 'game_config.xlsx')

def create_test_data():
    wb = Workbook()
    
    # 装备表
    ws_eq = wb.active
    ws_eq.title = "装备"
    ws_eq.append(["ID", "Name", "Type", "BaseAtk", "Price", "Rarity", "LevelReq"])
    for row in [
        [1, "铁剑", "Weapon", 10, 100, "Common", 1],
        [2, "钢剑", "Weapon", 25, 500, "Rare", 5],
        [3, "精钢剑", "Weapon", 45, 1500, "Epic", 15],
        [4, "龙鳞甲", "Armor", 30, 2000, "Epic", 20],
        [5, "皮甲", "Armor", 5, 80, "Common", 1],
        [6, "魔法杖", "Weapon", 35, 1200, "Rare", 12],
        [7, "精灵弓", "Weapon", 28, 900, "Rare", 10],
        [8, "暗影匕首", "Weapon", 55, 3000, "Legendary", 30],
        [9, "生命护符", "Accessory", 0, 2500, "Legendary", 25],
        [10, "力量戒指", "Accessory", 0, 800, "Epic", 18],
        [11, "测试-零攻", "Weapon", 0, 50, "Common", 1],
        [12, "测试-负价", "Armor", 20, -100, "Common", 1],
        [13, "测试-超贵", "Weapon", 999, 999999, "Legendary", 99],
        [14, "测试-小数价", "Accessory", 0, 9.99, "Rare", 3],
        [15, "测试-超长名-" + "A"*200, "Weapon", 1, 1, "Common", 1],
    ]:
        ws_eq.append(row)
    
    # 技能表
    ws_sk = wb.create_sheet("技能")
    ws_sk.append(["ID", "Name", "Type", "Damage", "ManaCost", "Cooldown", "ClassReq"])
    for row in [
        [101, "火焰斩", "Active", 150, 20, 5, "Warrior"],
        [102, "冰霜箭", "Active", 120, 15, 3, "Archer"],
        [103, "治愈术", "Active", 0, 30, 8, "Mage"],
        [104, "雷霆一击", "Active", 300, 50, 15, "Warrior"],
        [105, "隐身", "Buff", 0, 40, 30, "Assassin"],
        [106, "狂暴", "Buff", 0, 25, 60, "Warrior"],
        [107, "火球术", "Active", 200, 35, 4, "Mage"],
        [108, "毒刃", "Active", 80, 10, 2, "Assassin"],
        [109, "神圣护盾", "Buff", 0, 45, 20, "Paladin"],
        [110, "连射", "Active", 180, 25, 6, "Archer"],
    ]:
        ws_sk.append(row)
    
    # 商店表
    ws_sh = wb.create_sheet("商店")
    ws_sh.append(["ID", "ItemID", "ItemType", "Stock", "Discount", "Currency"])
    for row in [
        [1, 1, "装备", 999, 0, "Gold"],
        [2, 2, "装备", 500, 10, "Gold"],
        [3, 3, "装备", 100, 20, "Gold"],
        [4, 6, "装备", 200, 5, "Gold"],
        [5, 7, "装备", 300, 0, "Gold"],
        [6, 101, "技能书", 50, 0, "Diamond"],
        [7, 102, "技能书", 80, 15, "Diamond"],
        [8, 107, "技能书", 30, 25, "Diamond"],
        [9, 9, "装备", 10, 0, "Gem"],
        [10, 10, "装备", 25, 10, "Gem"],
    ]:
        ws_sh.append(row)
    
    # 掉落表
    ws_dt = wb.create_sheet("掉落")
    ws_dt.append(["MonsterID", "MonsterName", "DropItemID", "DropRate", "MinQty", "MaxQty"])
    for row in [
        [1001, "史莱姆", 1, 50.0, 1, 3],
        [1002, "哥布林", 2, 30.0, 1, 2],
        [1003, "兽人", 3, 10.0, 1, 1],
        [1004, "巨龙", 8, 1.0, 1, 1],
        [1005, "骷髅兵", 5, 40.0, 1, 4],
        [1006, "恶魔", 9, 0.5, 1, 1],
        [1007, "精灵", 7, 15.0, 1, 3],
        [1008, "Boss-暗影王", 8, 100.0, 2, 5],
    ]:
        ws_dt.append(row)
    
    # 活动奖励表
    ws_ev = wb.create_sheet("活动")
    ws_ev.append(["EventID", "EventName", "RewardID", "RewardName", "Quantity", "EventType"])
    for row in [
        [1, "新春活动", 1, "金币包", 1000, "Holiday"],
        [2, "新春活动", 2, "经验药水", 50, "Holiday"],
        [3, "周年庆", 8, "暗影匕首券", 1, "Anniversary"],
        [4, "周年庆", 9, "生命护符券", 1, "Anniversary"],
        [5, "限时签到", 3, "精钢剑券", 5, "Daily"],
        [6, "限时签到", 6, "魔法杖券", 3, "Daily"],
        [7, "竞技场S1", 10, "力量戒指券", 1, "PvP"],
        [8, "竞技场S1", 4, "龙鳞甲券", 2, "PvP"],
    ]:
        ws_ev.append(row)
    
    # 空Sheet
    ws_empty = wb.create_sheet("空表")
    ws_empty.append(["ID", "Name", "Value"])
    
    wb.save(TEST_FILE)
    print(f"✅ 测试文件已创建: {TEST_FILE}")
    print(f"   Sheets: {wb.sheetnames}")
    return TEST_FILE


# ============================================================
# 测试执行器 (修复版)
# ============================================================

class TestResult:
    def __init__(self):
        self.results = []
    
    def add(self, name, category, role, passed, detail="", error=""):
        self.results.append({
            "name": name, "category": category, "role": role,
            "passed": passed, "detail": detail, "error": error,
        })
        status = "✅" if passed else "❌"
        print(f"{status} [{category}] {name}")
        if not passed and error:
            err_preview = error[:200] if len(error) > 200 else error
            print(f"   └─ {err_preview}")
    
    def summary(self):
        total = len(self.results)
        passed = sum(1 for r in self.results if r["passed"])
        by_cat, by_role = {}, {}
        for r in self.results:
            by_cat.setdefault(r["category"], {"total": 0, "passed": 0})
            by_cat[r["category"]]["total"] += 1
            if r["passed"]: by_cat[r["category"]]["passed"] += 1
            by_role.setdefault(r["role"], {"total": 0, "passed": 0})
            by_role[r["role"]]["total"] += 1
            if r["passed"]: by_role[r["role"]]["passed"] += 1
        return {
            "total": total, "passed": passed,
            "rate": f"{passed}/{total} = {passed*100//total}%" if total else "N/A",
            "by_category": by_cat, "by_role": by_role,
            "failures": [r for r in self.results if not r["passed"]],
        }


tr = TestResult()


def run_select(sql):
    try:
        return execute_advanced_sql_query(TEST_FILE, sql)
    except Exception as e:
        return {"success": False, "message": str(e), "data": []}

def run_update(sql):
    try:
        return execute_advanced_update_query(TEST_FILE, sql)
    except Exception as e:
        return {"success": False, "message": str(e), "data": []}

def run_insert(sql):
    try:
        return execute_advanced_insert_query(TEST_FILE, sql)
    except Exception as e:
        return {"success": False, "message": str(e), "data": []}

def run_delete(sql):
    try:
        return execute_advanced_delete_query(TEST_FILE, sql)
    except Exception as e:
        return {"success": False, "message": str(e), "data": []}


def check(result, expect_success=True, verify_func=None):
    """
    验证结果。
    verify_func 接收 data，返回 True(通过) 或 (False, "原因") 或 (True, "")
    """
    if expect_success:
        if not result.get('success'):
            return False, result.get('message', '未知错误')
        if verify_func:
            try:
                ret = verify_func(result.get('data', []))
                if ret is True:
                    return True, ""
                elif ret is False:
                    return False, "验证函数返回False"
                elif isinstance(ret, tuple) and len(ret) == 2:
                    ok, msg = ret
                    if not ok:
                        return False, msg
                else:
                    return False, f"验证函数返回异常类型: {type(ret)}"
            except Exception as e:
                return False, f"验证异常: {e}"
        return True, ""
    else:
        if result.get('success'):
            return False, "预期失败但成功"
        return True, ""


# ============================================================
# Round 6 测试用例
# ============================================================

def run_all_tests():
    print("=" * 70)
    print("🔧 Round 6 迭代测试 (修复版)")
    print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)
    
    # ================================================================
    # Group A: UNION / UNION ALL 深度测试
    # ================================================================
    print("\n📦 Group A: UNION / UNION ALL 深度测试")
    print("-" * 50)
    
    # A1: 基础 UNION ALL 合并装备和技能的ID+名称
    r = run_select("""
        SELECT ID, Name, '装备' as 类别 FROM 装备 
        UNION ALL 
        SELECT ID, Name, '技能' as 类别 FROM 技能
    """)
    ok, err = check(r, True, lambda d: (len(d) == 26, f"期望26行(1header+25data), 实际{len(d)}行"))
    tr.add("A1: UNION ALL合并装备+技能", "SELECT", "策划", ok, "", err)
    
    # A2: UNION 去重
    r = run_select("""
        SELECT ItemType FROM 商店 WHERE ItemType = '装备'
        UNION
        SELECT ItemType FROM 商店 WHERE ItemType = '技能书'
    """)
    ok, err = check(r, True, lambda d: (len(d) == 3, f"期望3行(header+2types), 实际{len(d)}行"))
    tr.add("A2: UNION去重商品类型", "SELECT", "运营", ok, "", err)
    
    # A3: UNION ALL + ORDER BY
    r = run_select("""
        SELECT Name, Price, '装备' as 来源 FROM 装备 WHERE Rarity = 'Legendary'
        UNION ALL
        SELECT EventName as Name, Quantity as Price, '活动' as 来源 FROM 活动 WHERE EventType = 'Anniversary'
        ORDER BY Price DESC
    """)
    ok, err = check(r, True, lambda d: (len(d) >= 4, f"期望>=4行, 实际{len(d)}行"))
    tr.add("A3: UNION ALL + ORDER BY DESC", "SELECT", "数据分析", ok, "", err)
    
    # A4: UNION ALL + LIMIT
    r = run_select("""
        SELECT ID, Name FROM 装备 WHERE Type = 'Weapon'
        UNION ALL
        SELECT ID, Name FROM 装备 WHERE Type = 'Armor'
        LIMIT 5
    """)
    ok, err = check(r, True, lambda d: (len(d) <= 6, f"期望<=6行(header+5), 实际{len(d)}行"))
    tr.add("A4: UNION ALL + LIMIT分页", "SELECT", "客户端", ok, "", err)
    
    # A5: UNION ALL 含表达式列
    r = run_select("""
        SELECT Name, Price * 1.1 as 调整后价格, '装备' as 类型 FROM 装备 WHERE LevelReq <= 5
        UNION ALL
        SELECT EventName, Quantity * 10 as 调整后价格, '活动' as 类型 FROM 活动 WHERE EventType = 'Daily'
    """)
    ok, err = check(r, True, lambda d: (len(d) >= 4, f"期望>=4行, 实际{len(d)}行"))
    tr.add("A5: UNION ALL含计算表达式列", "SELECT", "策划", ok, "", err)
    
    # A6: 三表 UNION ALL
    r = run_select("""
        SELECT Name, '装备' as 来源, CAST(Price as int) as 数值 FROM 装备 WHERE ID <= 3
        UNION ALL
        SELECT Name, '技能' as 来源, Damage as 数值 FROM 技能 WHERE ID <= 103
        UNION ALL
        SELECT MonsterName as Name, '怪物' as 来源, DropRate as 数值 FROM 掉落 WHERE MonsterID <= 1003
    """)
    ok, err = check(r, True, lambda d: (len(d) >= 4, f"期望>=4行, 实际{len(d)}行"))
    tr.add("A6: 三表UNION ALL", "SELECT", "数据分析", ok, "", err)
    
    # A7: UNION 子查询 + GROUP BY
    r = run_select("""
        SELECT Type, COUNT(*) as cnt FROM (
            SELECT Type FROM 装备
            UNION ALL
            SELECT ClassReq as Type FROM 技能
        ) t GROUP BY Type
    """)
    ok, err = check(r, True)
    tr.add("A7: UNION子查询+GROUP BY聚合", "SELECT", "数据分析", ok, "", err)
    
    # ================================================================
    # Group B: EXISTS / NOT EXISTS 子查询
    # ================================================================
    print("\n📦 Group B: EXISTS / NOT EXISTS 子查询")
    print("-" * 50)
    
    # B1: EXISTS 基础
    r = run_select("""
        SELECT * FROM 商店 WHERE EXISTS (
            SELECT 1 FROM 装备 WHERE 装备.ID = 商店.ItemID AND 装备.Rarity = 'Legendary'
        )
    """)
    ok, err = check(r, True, lambda d: (len(d) >= 1, f"期望>=1行, 实际{len(d)}行"))
    tr.add("B1: EXISTS查 Legendary 装备在售", "SELECT", "服务端", ok, "", err)
    
    # B2: NOT EXISTS
    r = run_select("""
        SELECT ID, Name, Price FROM 装备 WHERE NOT EXISTS (
            SELECT 1 FROM 商店 WHERE 商店.ItemID = 装备.ID
        )
    """)
    ok, err = check(r, True, lambda d: (len(d) >= 1, f"期望>=1行, 实际{len(d)}行"))
    tr.add("B2: NOT EXISTS查未上架装备", "SELECT", "运营", ok, "", err)
    
    # B3: EXISTS 多条件
    r = run_select("""
        SELECT * FROM 掉落 WHERE EXISTS (
            SELECT 1 FROM 装备 WHERE 装备.ID = 掉落.DropItemID AND 装备.Rarity IN ('Epic', 'Legendary')
        )
    """)
    ok, err = check(r, True)
    tr.add("B3: EXISTS多条件-稀有掉落源", "SELECT", "QA", ok, "", err)
    
    # B4: 双 NOT EXISTS
    r = run_select("""
        SELECT * FROM 商店 WHERE NOT EXISTS (
            SELECT 1 FROM 装备 WHERE 装备.ID = 商店.ItemID
        ) AND NOT EXISTS (
            SELECT 1 FROM 技能 WHERE 技能.ID = 商店.ItemID
        )
    """)
    ok, err = check(r, True)
    tr.add("B4: 双NOT EXISTS查孤儿记录", "SELECT", "服务端", ok, "", err)
    
    # B5: CASE WHEN EXISTS
    r = run_select("""
        SELECT ID, Name, Price, 
            CASE WHEN EXISTS (SELECT 1 FROM 商店 WHERE 商店.ItemID = 装备.ID) THEN '已上架' ELSE '未上架' END as 状态
        FROM 装备 WHERE ID <= 5
    """)
    ok, err = check(r, True, lambda d: (len(d) == 6, f"期望6行(header+5), 实际{len(d)}行"))
    tr.add("B5: CASE WHEN EXISTS标记状态", "SELECT", "策划", ok, "", err)
    
    # ================================================================
    # Group C: 多 Sheet 跨表操作
    # ================================================================
    print("\n📦 Group C: 多 Sheet 跨表操作")
    print("-" * 50)
    
    # C1: INNER JOIN 跨Sheet
    r = run_select("""
        SELECT 装备.ID, 装备.Name, 装备.Price, 商店.Stock, 商店.Discount 
        FROM 装备 INNER JOIN 商店 ON 装备.ID = 商店.ItemID
        WHERE 商店.Discount > 0
    """)
    ok, err = check(r, True, lambda d: (len(d) >= 2, f"期望>=2行, 实际{len(d)}行"))
    tr.add("C1: 跨Sheet JOIN折扣商品", "SELECT", "策划", ok, "", err)
    
    # C2: 三表 JOIN (注意：中文逗号问题已在C2修正为英文逗号)
    r = run_select("""
        SELECT 掉落.MonsterName, 装备.Name as ItemName, 装备.Rarity, 装备.Price, 商店.Stock
        FROM 掉落 
        INNER JOIN 装备 ON 掉落.DropItemID = 装备.ID
        LEFT JOIN 商店 ON 装备.ID = 商店.ItemID
        WHERE 掉落.DropRate >= 10
    """)
    ok, err = check(r, True)
    tr.add("C2: 三表JOIN高概率掉落", "SELECT", "服务端", ok, "", err)
    
    # C3: LEFT JOIN 找未上架
    r = run_select("""
        SELECT 装备.ID, 装备.Name, 装备.Price, 商店.Stock
        FROM 装备 LEFT JOIN 商店 ON 装备.ID = 商店.ItemID
        WHERE 商店.ID IS NULL
    """)
    ok, err = check(r, True, lambda d: (len(d) >= 2, f"期望>=2行, 实际{len(d)}行"))
    tr.add("C3: LEFT JOIN找未上架装备", "SELECT", "运营", ok, "", err)
    
    # C4: 跨Sheet聚合
    r = run_select("""
        SELECT 商店.Currency, SUM(装备.Price * 商店.Stock) as TotalValue, COUNT(*) as ItemCount
        FROM 商店 INNER JOIN 装备 ON 商店.ItemID = 装备.ID
        GROUP BY 商店.Currency
    """)
    ok, err = check(r, True, lambda d: (len(d) >= 2, f"期望>=2行, 实际{len(d)}行"))
    tr.add("C4: 跨Sheet按币种统计价值", "SELECT", "数据分析", ok, "", err)
    
    # ================================================================
    # Group D: 写入操作
    # ================================================================
    print("\n📦 Group D: 写入操作")
    print("-" * 50)
    
    # D1: UPDATE 批量调价
    r = run_update("""UPDATE 装备 SET Price = Price * 1.2 WHERE Rarity = 'Epic' AND ID <= 4""")
    if r.get('success'):
        rv = run_select("SELECT ID, Name, Price FROM 装备 WHERE Rarity = 'Epic' AND ID <= 4")
        ok, err = check(rv, True, lambda d: (len(d) >= 2, f"回读{len(d)}行"))
    else:
        ok, err = False, r.get('message', '')
    tr.add("D1: UPDATE Epic涨价20%+回读", "UPDATE", "策划", ok, "", err)
    
    # D2: INSERT 新装备
    r = run_insert("""
        INSERT INTO 装备 (ID, Name, Type, BaseAtk, Price, Rarity, LevelReq) 
        VALUES (16, '圣光剑', 'Weapon', 88, 8888, 'Legendary', 40)
    """)
    if r.get('success'):
        rv = run_select("SELECT * FROM 装备 WHERE ID = 16")
        ok, err = check(rv, True, lambda d: (len(d) == 2 and '圣光剑' in str(d), "插入后未找到圣光剑"))
    else:
        ok, err = False, r.get('message', '')
    tr.add("D2: INSERT圣光剑+回读", "INSERT", "策划", ok, "", err)
    
    # D3: DELETE 测试数据
    r = run_delete("""DELETE FROM 装备 WHERE Name LIKE '测试-%'""")
    if r.get('success'):
        rv = run_select("SELECT COUNT(*) as cnt FROM 装备 WHERE Name LIKE '测试-%'")
        ok, err = check(rv, True, lambda d: (len(d) == 2 and d[1][0] == 0, f"删除后仍有测试数据, cnt={d[1][0] if len(d)>=2 else '?'}"))
    else:
        ok, err = False, r.get('message', '')
    tr.add("D3: DELETE测试数据+COUNT验证", "DELETE", "运营", ok, "", err)
    
    # D4: UPDATE ROUND取整
    r = run_update("""UPDATE 装备 SET Price = ROUND(Price, 0) WHERE Type = 'Accessory'""")
    if r.get('success'):
        rv = run_select("SELECT ID, Name, Price FROM 装备 WHERE Type = 'Accessory'")
        ok, err = check(rv, True, lambda d: (len(d) >= 2, f"回读{len(d)}行"))
    else:
        ok, err = False, r.get('message', '')
    tr.add("D4: UPDATE ROUND配件取整+回读", "UPDATE", "策划", ok, "", err)
    
    # D5: INSERT 新技能
    r = run_insert("""
        INSERT INTO 技能 (ID, Name, Type, Damage, ManaCost, Cooldown, ClassReq) 
        VALUES (111, '复活术', 'Active', 0, 100, 120, 'Priest')
    """)
    if r.get('success'):
        rv = run_select("SELECT * FROM 技能 WHERE ID = 111")
        ok, err = check(rv, True, lambda d: (len(d) == 2 and '复活术' in str(d), "未找到复活术"))
    else:
        ok, err = False, r.get('message', '')
    tr.add("D5: INSERT复活术+回读", "INSERT", "服务端", ok, "", err)
    
    # D6: UPDATE 商店折扣
    r = run_update("""UPDATE 商店 SET Discount = 50 WHERE Currency = 'Gem' AND Stock <= 25""")
    if r.get('success'):
        rv = run_select("SELECT * FROM 商店 WHERE Currency = 'Gem'")
        ok, err = check(rv, True)
    else:
        ok, err = False, r.get('message', '')
    tr.add("D6: UPDATE Gem商品折扣50%+回读", "UPDATE", "运营", ok, "", err)
    
    # ================================================================
    # Group E: 边缘场景
    # ================================================================
    print("\n📦 Group E: 边缘场景与错误恢复")
    print("-" * 50)
    
    # E1: 空Sheet查询
    r = run_select("SELECT * FROM 空表")
    ok, err = check(r, True, lambda d: (len(d) == 1, f"空表应仅header, 实际{len(d)}行"))
    tr.add("E1: 空Sheet查询(仅header)", "边缘", "QA", ok, "", err)
    
    # E2: 空Sheet聚合
    r = run_select("SELECT COUNT(*) as cnt, AVG(Value) as avg_val FROM 空表")
    ok, err = check(r, True)
    tr.add("E2: 空Sheet聚合默认值", "边缘", "QA", ok, "", err)
    
    # E3: WHERE 无匹配
    r = run_select("SELECT * FROM 装备 WHERE ID = 99999")
    ok, err = check(r, True, lambda d: (len(d) == 1, f"空结果应仅header, 实际{len(d)}行"))
    tr.add("E3: WHERE无匹配空结果集", "边缘", "QA", ok, "", err)
    
    # E4: NULL/零值混合
    r = run_select("SELECT * FROM 装备 WHERE Price IS NULL OR Price = 0")
    ok, err = check(r, True)
    tr.add("E4: NULL/零值混合条件", "边缘", "QA", ok, "", err)
    
    # E5: 超长文本 LIKE
    r = run_select("SELECT ID, Name FROM 装备 WHERE Name LIKE '%测试-超长名%'")
    ok, err = check(r, True, lambda d: (len(d) >= 2 and len(str(d[1][1])) >= 200, f"超长文本不完整或未找到"))
    tr.add("E5: 超长文本LIKE查询", "边缘", "客户端", ok, "", err)
    
    # E6: ORDER BY NULLS FIRST
    r = run_select("SELECT ID, Name, Price FROM 装备 ORDER BY Price ASC NULLS FIRST")
    ok, err = check(r, True)
    tr.add("E6: ORDER BY NULLS FIRST", "边缘", "数据分析", ok, "", err)
    
    # E7: INSERT 后 UNION 一致性
    r_ins = run_insert("""
        INSERT INTO 活动 (EventID, EventName, RewardID, RewardName, Quantity, EventType) 
        VALUES (9, '测试联盟', 16, '圣光剑券', 1, 'Test')
    """)
    if r_ins.get('success'):
        r = run_select("""
            SELECT EventName, RewardName, Quantity, '活动' as src FROM 活动 WHERE EventType = 'Test'
            UNION ALL
            SELECT Name, '直接' as rn, Price as qty, '装备' as src FROM 装备 WHERE ID = 16
        """)
        ok, err = check(r, True, lambda d: (len(d) >= 3, f"UNION应>=3行, 实际{len(d)}行"))
    else:
        ok, err = False, r_ins.get('message', '')
    tr.add("E7: INSERT后UNION一致性验证", "INSERT+SELECT", "服务端", ok, "", err)
    
    # E8: DELETE 不存在的行
    r = run_delete("DELETE FROM 装备 WHERE ID = 99999")
    ok, err = check(r, True)
    tr.add("E8: DELETE不存在行(幂等)", "DELETE", "QA", ok, "", err)
    
    # 完成
    print("\n" + "=" * 70)
    s = tr.summary()
    return s


if __name__ == "__main__":
    try:
        create_test_data()
        summary = run_all_tests()
        
        print("\n" + "📊 Round 6 测试报告".center(60))
        print("=" * 70)
        print(f"总计: {summary['rate']}\n")
        
        print("按分类:")
        for cat, v in sorted(summary['by_category'].items()):
            pct = f"{v['passed']*100//v['total']}%" if v['total'] else "N/A"
            mark = "✅" if v['passed'] == v['total'] else "⚠️"
            print(f"  {mark} {cat}: {v['passed']}/{v['total']} ({pct})")
        
        print("\n按角色:")
        for role, v in sorted(summary['by_role'].items()):
            pct = f"{v['passed']*100//v['total']}%" if v['total'] else "N/A"
            mark = "✅" if v['passed'] == v['total'] else "⚠️"
            print(f"  {mark} {role}: {v['passed']}/{v['total']} ({pct})")
        
        if summary['failures']:
            print(f"\n❌ 失败用例 ({len(summary['failures'])}):")
            for f in summary['failures']:
                print(f"  • [{f['category']}] {f['name']}")
                print(f"    └─ {f['error'][:150]}")
        
        print("\n---JSON_START---")
        print(json.dumps({
            "round": 6, "date": datetime.now().isoformat(), "version": "1.9.3",
            "total": summary['total'], "passed": summary['passed'],
            "by_category": {k: {'p': v['passed'], 't': v['total']} for k, v in summary['by_category'].items()},
            "by_role": {k: {'p': v['passed'], 't': v['total']} for k, v in summary['by_role'].items()},
            "failures": [{"n": f['name'], "c": f['category'], "e": f['error'][:300]} for f in summary['failures']],
        }, ensure_ascii=False, indent=2))
        print("---JSON_END---")
    
    finally:
        if os.path.exists(TEST_DIR):
            shutil.rmtree(TEST_DIR)
            print(f"\n🧹 已清理: {TEST_DIR}")
