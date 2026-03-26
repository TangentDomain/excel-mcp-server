# -*- coding: utf-8 -*-
"""
Excel Compare 深度测试套件

覆盖 excel_compare.py 中未覆盖的代码路径
"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from src.core.excel_compare import ExcelComparer
from src.models.types import ComparisonOptions


class TestExcelCompareAdvanced:
    """高级比较测试"""

    @pytest.fixture
    def compare_file1(self, temp_dir):
        """创建比较文件1"""
        file_path = temp_dir / "compare1.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        ws['A1'] = "ID"
        ws['B1'] = "Name"
        ws['C1'] = "Value"
        
        for i in range(2, 12):
            ws[f'A{i}'] = i - 1
            ws[f'B{i}'] = f"Name{i}"
            ws[f'C{i}'] = i * 10
        
        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def compare_file2(self, temp_dir):
        """创建比较文件2"""
        file_path = temp_dir / "compare2.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        ws['A1'] = "ID"
        ws['B1'] = "Name"
        ws['C1'] = "Value"
        
        # 有一些差异
        for i in range(2, 12):
            ws[f'A{i}'] = i - 1
            ws[f'B{i}'] = f"Name{i}"
            if i < 8:
                ws[f'C{i}'] = i * 20  # 不同的值
            else:
                ws[f'C{i}'] = i * 10
        
        wb.save(file_path)
        return str(file_path)

    def test_compare_with_ignore_empty(self, compare_file1, compare_file2):
        """测试忽略空单元格的比较"""
        options = ComparisonOptions(
            compare_values=True,
            compare_formulas=False,
            compare_formats=False,
            ignore_empty_cells=True,
            case_sensitive=True,
            structured_comparison=False
        )
        
        comparer = ExcelComparer(options)
        result = comparer.compare_files(compare_file1, compare_file2)
        
        assert result is not None

    def test_compare_with_case_insensitive(self, compare_file1, compare_file2):
        """测试不区分大小写的比较"""
        options = ComparisonOptions(
            compare_values=True,
            compare_formulas=False,
            compare_formats=False,
            ignore_empty_cells=True,
            case_sensitive=False,
            structured_comparison=False
        )
        
        comparer = ExcelComparer(options)
        result = comparer.compare_files(compare_file1, compare_file2)
        
        assert result is not None

    def test_compare_formulas(self, temp_dir):
        """测试公式比较"""
        # 创建文件1
        file1 = temp_dir / "formula1.xlsx"
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "Sheet"
        ws1['A1'] = "=1+1"
        ws1['A2'] = 2
        wb1.save(file1)
        
        # 创建文件2
        file2 = temp_dir / "formula2.xlsx"
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Sheet"
        ws2['A1'] = 2
        ws2['A2'] = "=1+1"
        wb2.save(file2)
        
        options = ComparisonOptions(
            compare_values=False,
            compare_formulas=True,
            compare_formats=False,
            ignore_empty_cells=True,
            case_sensitive=True,
            structured_comparison=False
        )
        
        comparer = ExcelComparer(options)
        result = comparer.compare_files(str(file1), str(file2))
        
        assert result is not None

    def test_compare_formats(self, temp_dir):
        """测试格式比较"""
        from openpyxl.styles import Font
        
        # 创建文件1
        file1 = temp_dir / "format1.xlsx"
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "Sheet"
        ws1['A1'] = "Text"
        ws1['A1'].font = Font(bold=True)
        wb1.save(file1)
        
        # 创建文件2
        file2 = temp_dir / "format2.xlsx"
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Sheet"
        ws2['A1'] = "Text"
        wb2.save(file2)
        
        options = ComparisonOptions(
            compare_values=False,
            compare_formulas=False,
            compare_formats=True,
            ignore_empty_cells=True,
            case_sensitive=True,
            structured_comparison=False
        )
        
        comparer = ExcelComparer(options)
        result = comparer.compare_files(str(file1), str(file2))
        
        assert result is not None


class TestExcelCompareEdgeCases:
    """边界情况测试"""

    def test_compare_empty_sheets(self, temp_dir):
        """测试空工作表比较"""
        # 创建文件1
        file1 = temp_dir / "empty1.xlsx"
        wb1 = Workbook()
        wb1.active.title = "Sheet"
        wb1.save(file1)
        
        # 创建文件2
        file2 = temp_dir / "empty2.xlsx"
        wb2 = Workbook()
        wb2.active.title = "Sheet"
        wb2.save(file2)
        
        options = ComparisonOptions(
            compare_values=True,
            compare_formulas=False,
            compare_formats=False,
            ignore_empty_cells=True,
            case_sensitive=True,
            structured_comparison=False
        )
        
        comparer = ExcelComparer(options)
        result = comparer.compare_files(str(file1), str(file2))
        
        assert result is not None

    def test_compare_different_sheets(self, temp_dir):
        """测试不同工作表名称"""
        # 创建文件1
        file1 = temp_dir / "diff1.xlsx"
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "Sheet1"
        ws1['A1'] = "Data"
        wb1.save(file1)
        
        # 创建文件2
        file2 = temp_dir / "diff2.xlsx"
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Sheet2"
        ws2['A1'] = "Data"
        wb2.save(file2)
        
        options = ComparisonOptions(
            compare_values=True,
            compare_formulas=False,
            compare_formats=False,
            ignore_empty_cells=True,
            case_sensitive=True,
            structured_comparison=False
        )
        
        comparer = ExcelComparer(options)
        result = comparer.compare_files(str(file1), str(file2))
        
        assert result is not None
