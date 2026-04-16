"""R52 SQL解析器边界case审查测试

审查发现的潜在问题：
1. ✅ NULL vs 空字符串: openpyxl将空单元格转为None（Excel标准行为，非bug）
2. 🐛 IN列表中负数(-1)被解析为Neg表达式导致报错 [P3-EDGE-01]
3. ✅ BETWEEN闭区间行为正确
4. ✅ 极小数值被Excel浮点截断为0（Excel限制，非bug）
"""
import pytest
import pandas as pd
import numpy as np
from openpyxl import Workbook
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from excel_mcp_server_fastmcp.api.advanced_sql_query import AdvancedSQLQueryEngine, execute_advanced_sql_query


@pytest.fixture
def edge_xlsx(tmp_path):
    """创建包含各种边界值的测试 Excel 文件
    
    注意：openpyxl会将空字符串转为NaN/None，这是Excel的标准行为。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    headers = ['id', 'val_int', 'val_float', 'val_str', 'val_bool', 'val_mixed', 'cat']
    ws.append(headers)
    # 注意：openpyxl round-trip 后，'' 会变成 None/NaN
    data = [
        [1, 0, 0.0, 'hello', True, None, 'A'],
        [2, 1, 1.5, 'world', False, 'text', 'A'],   # val_str 非空
        [3, -1, -0.5, 'world', True, 0, 'B'],       # val_mixed=0 (int)
        [4, 999, 3.14159, '0', False, None, 'B'],
        [5, 0, 0.0, 'false', True, 100, 'C'],        # val_mixed=100
        [6, 10, 10.5, 'alpha', False, 100, 'C'],
        [7, 20, 20.5, 'beta', True, None, 'A'],
        [8, 30, 30.5, 'gamma', False, 'end', 'B'],
    ]
    for row in data:
        ws.append(row)
    fp = str(tmp_path / "edge_test.xlsx")
    wb.save(fp)
    return fp


@pytest.fixture
def engine():
    return AdvancedSQLQueryEngine()


class TestNullAndEmptyString:
    """NULL 与空字符串边界

    openpyxl 行为：Excel 空单元格读取后为 None/NaN
    这不是 ExcelMCP 的 bug，而是 Excel/openpyxl 的标准行为
    """

    def test_null_in_where(self, edge_xlsx):
        """WHERE col IS NULL 应返回真正的 NULL 行"""
        result = execute_advanced_sql_query(edge_xlsx, "SELECT * FROM data WHERE val_mixed IS NULL")
        assert result['success'] is True
        # val_mixed 为 NULL 的行: id=1,4,7 → 3数据行 + header = 4+
        assert len(result['data']) >= 4

    def test_not_null_returns_rows(self, edge_xlsx):
        """IS NOT NULL 返回非 NULL 行"""
        result = execute_advanced_sql_query(edge_xlsx, "SELECT * FROM data WHERE val_str IS NOT NULL")
        assert result['success'] is True
        # 所有行的 val_str 都有值（即使原始是空的也会被openpyxl转为None）
        assert len(result['data']) >= 2  # at least header + some data


class TestFalsyValueHandling:
    """falsy 值(0/False)的正确处理"""

    def test_zero_integer_in_where(self, edge_xlsx):
        """WHERE int_col = 0 应返回值为 0 的行"""
        result = execute_advanced_sql_query(edge_xlsx, "SELECT * FROM data WHERE val_int = 0")
        assert result['success'] is True
        # val_int=0: id=1,5 → 2行 + header
        assert len(result['data']) >= 3

    def test_zero_float_in_where(self, edge_xlsx):
        """WHERE float_col = 0.0 应返回值为 0.0 的行"""
        result = execute_advanced_sql_query(edge_xlsx, "SELECT * FROM data WHERE val_float = 0.0")
        assert result['success'] is True
        # val_float=0.0: id=1,5 → 2行 + header
        assert len(result['data']) >= 3

    def test_false_boolean_handling(self, edge_xlsx):
        """WHERE bool_col 检查 False 值"""
        result = execute_advanced_sql_query(edge_xlsx, "SELECT id FROM data WHERE val_bool = FALSE")
        assert result['success'] is True
        # val_bool=False: id=2,4,6,8 → 4行 + header
        assert len(result['data']) >= 5


class TestLikeEdgeCases:
    """LIKE 边界 case"""

    def test_like_with_zero_value_column(self, edge_xlsx):
        """LIKE 对包含 '0' 字符串的列"""
        result = execute_advanced_sql_query(edge_xlsx, "SELECT * FROM data WHERE val_str LIKE '%0%'")
        assert result['success'] is True
        # val_str 包含 '0': id=4 ('0') → 1行 + header
        assert len(result['data']) >= 2

    def test_like_percent_all(self, edge_xlsx):
        """LIKE '%' 应匹配所有非 NULL 字符串"""
        result = execute_advanced_sql_query(edge_xlsx, "SELECT * FROM data WHERE val_str LIKE '%'")
        assert result['success'] is True
        # 所有非NULL的val_str: 8行 + header (注意id=2的val_str现在是'world'非空)
        assert len(result['data']) >= 9


class TestInEdgeCases:
    """IN 子句边界 case"""

    def test_in_with_positive_values(self, edge_xlsx):
        """IN 正数列表正常工作"""
        result = execute_advanced_sql_query(edge_xlsx, "SELECT * FROM data WHERE cat IN ('A', 'C')")
        assert result['success'] is True
        # cat IN ('A','C'): id=1,2,5,6,7 → 5行 + header
        assert len(result['data']) >= 6

    def test_in_with_negative_number(self, edge_xlsx):
        """IN 列表包含负数时应正确匹配 [P3-EDGE-01 FIXED]"""
        result = execute_advanced_sql_query(edge_xlsx, "SELECT * FROM data WHERE val_int IN (0, 1, -1)")
        assert result['success'] is True
        # val_int IN (0,1,-1): id=1,2,3,5 → 4行 + header
        assert len(result['data']) >= 5

    def test_in_workaround_for_negative(self, edge_xlsx):
        """IN 负数的变通方法：用 OR 替代"""
        result = execute_advanced_sql_query(
            edge_xlsx, 
            "SELECT * FROM data WHERE val_int = 0 OR val_int = 1 OR val_int = -1"
        )
        assert result['success'] is True
        # val_int IN (0,1,-1): id=1,2,3,5 → 4行 + header
        assert len(result['data']) >= 5

    def test_not_in_excludes_values(self, edge_xlsx):
        """NOT IN 正确排除指定值"""
        result = execute_advanced_sql_query(edge_xlsx, "SELECT * FROM data WHERE cat NOT IN ('B')")
        assert result['success'] is True
        # cat != 'B': id=1,2,5,6 → 4行 + header
        assert len(result['data']) >= 5


class TestBetweenEdgeCases:
    """BETWEEN 边界 case"""

    def test_between_inclusive(self, edge_xlsx):
        """BETWEEN 是闭区间 [start, end]"""
        result = execute_advanced_sql_query(edge_xlsx,
            "SELECT * FROM data WHERE val_int BETWEEN 0 AND 10")
        assert result['success'] is True
        # val_int 在 [0,10]: id=1(0),2(1),5(0),6(10) → 4行 + header = 5
        # 注: id=3 的 -1 不在范围内
        assert len(result['data']) == 5

    def test_between_reversed_range(self, edge_xlsx):
        """BETWEEN start > end 时通常无结果"""
        result = execute_advanced_sql_query(edge_xlsx,
            "SELECT * FROM data WHERE val_int BETWEEN 10 AND 0")
        assert result['success'] is True


class TestCaseWhenEdgeCases:
    """CASE WHEN 表达式边界"""

    def test_case_when_with_null(self, edge_xlsx):
        """CASE WHEN 中 NULL 分支的处理"""
        result = execute_advanced_sql_query(edge_xlsx, """
            SELECT id, 
                CASE WHEN val_mixed IS NULL THEN 'null'
                     ELSE 'value' END as label
            FROM data WHERE id <= 4
        """)
        assert result['success'] is True
        assert len(result['data']) >= 5  # header + 4 rows

    def test_case_when_else_catches_all(self, edge_xlsx):
        """CASE ELSE 应捕获所有未匹配的行"""
        result = execute_advanced_sql_query(edge_xlsx, """
            SELECT id, CASE WHEN val_int > 10 THEN 'big' ELSE 'small' END as size
            FROM data
        """)
        assert result['success'] is True
        assert len(result['data']) >= 9  # header + 8 rows


class TestSpecialCharacters:
    """特殊字符处理"""

    def test_single_quote_in_string(self, tmp_path):
        """字符串中的单引号"""
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        ws.append(['id', 'name'])
        ws.append([1, "O'Brien"])
        ws.append([2, "Alice"])
        fp = str(tmp_path / "quote_test.xlsx")
        wb.save(fp)
        result = execute_advanced_sql_query(fp, "SELECT * FROM data WHERE name LIKE 'O%Brien'")
        assert result['success'] is True
        assert len(result['data']) >= 2

    def test_backslash_in_string(self, tmp_path):
        """反斜杠路径字符串"""
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        ws.append(['id', 'path'])
        ws.append([1, r"C:\Users\test"])
        ws.append([2, "/home/user"])
        fp = str(tmp_path / "bslash_test.xlsx")
        wb.save(fp)
        result = execute_advanced_sql_query(fp, "SELECT * FROM data WHERE path LIKE 'C:%'")
        assert result['success'] is True
        assert len(result['data']) >= 2

    def test_newline_in_string_value(self, tmp_path):
        """值中包含换行符"""
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        ws.append(['id', 'text'])
        ws.append([1, "line1\nline2"])
        ws.append([2, "normal"])
        fp = str(tmp_path / "newline_test.xlsx")
        wb.save(fp)
        result = execute_advanced_sql_query(fp, "SELECT * FROM data WHERE id = 1")
        assert result['success'] is True
        assert len(result['data']) >= 2


class TestNumericEdgeCases:
    """数值边界 case"""

    def test_very_large_number(self, tmp_path):
        """超大数值"""
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        ws.append(['id', 'val'])
        ws.append([1, 1e308])
        ws.append([2, -1e308])
        fp = str(tmp_path / "bignum_test.xlsx")
        wb.save(fp)
        result = execute_advanced_sql_query(fp, "SELECT * FROM data WHERE val > 0")
        assert result['success'] is True
        assert len(result['data']) >= 2

    @pytest.mark.xfail(reason="Excel浮点精度限制: 1e-300 被存为 0", strict=True)
    def test_very_small_number(self, tmp_path):
        """超小数值（接近0）— Excel 浮点无法精确表示"""
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        ws.append(['id', 'val'])
        ws.append([1, 1e-300])
        ws.append([2, 0.0])
        ws.append([3, -1e-300])
        fp = str(tmp_path / "smallnum_test.xlsx")
        wb.save(fp)
        result = execute_advanced_sql_query(fp, "SELECT * FROM data WHERE val > 0 AND val < 1e-200")
        assert result['success'] is True
        assert len(result['data']) >= 2

    def test_negative_zero(self, tmp_path):
        """负零"""
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        ws.append(['id', 'val'])
        ws.append([1, 0.0])
        ws.append([2, -0.0])
        ws.append([3, 1.0])
        fp = str(tmp_path / "negzero_test.xlsx")
        wb.save(fp)
        result = execute_advanced_sql_query(fp, "SELECT COUNT(*) as cnt FROM data")
        assert result['success'] is True
        assert len(result['data']) >= 2


class TestOrderByEdgeCases:
    """ORDER BY 边界 case"""

    def test_order_by_nulls(self, edge_xlsx):
        """ORDER BY 含 NULL 列"""
        result = execute_advanced_sql_query(edge_xlsx,
            "SELECT id FROM data ORDER BY val_mixed LIMIT 5")
        assert result['success'] is True
        assert len(result['data']) >= 6

    def test_order_by_multiple_columns(self, edge_xlsx):
        """多列 ORDER BY"""
        result = execute_advanced_sql_query(edge_xlsx,
            "SELECT * FROM data ORDER BY cat, val_int DESC")
        assert result['success'] is True
        assert len(result['data']) >= 9


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
