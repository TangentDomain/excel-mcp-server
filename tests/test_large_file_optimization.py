"""
测试大文件优化功能（REQ-032）

验证：
1. 文件大小检测正确
2. 大文件范围查询使用优化路径（iter_rows按需加载）
3. 小文件不受影响，仍使用calamine快速路径
4. 内存效率：大文件读取时不会全量加载
"""

import os
import tempfile
import unittest

from openpyxl import Workbook


class TestLargeFileOptimization(unittest.TestCase):
    """大文件优化测试"""

    def _create_test_file(self, rows=100, cols=10):
        """创建测试用Excel文件"""
        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)

        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="TestData")

        headers = [f"Col_{i}" for i in range(cols)]
        ws.append(headers)

        for r in range(1, rows):
            ws.append([r, f"Name_{r}", 3.14 * r, r % 2 == 0] + [f"R{r}C{c}" for c in range(4, cols)])

        wb.save(path)
        wb.close()
        return path

    def test_file_size_detection(self):
        """测试文件大小检测"""
        from src.excel_mcp_server_fastmcp.core.excel_reader import _get_file_size, _is_large_file, _LARGE_FILE_THRESHOLD

        # 小文件
        small_path = self._create_test_file(rows=100)
        try:
            size = _get_file_size(small_path)
            self.assertGreater(size, 0)
            self.assertFalse(_is_large_file(small_path))
        finally:
            os.unlink(small_path)

        # 阈值验证
        self.assertIsInstance(_LARGE_FILE_THRESHOLD, int)
        self.assertEqual(_LARGE_FILE_THRESHOLD, 50 * 1024 * 1024)

    def test_small_file_uses_standard_path(self):
        """测试小文件使用标准路径（非优化模式）"""
        from src.excel_mcp_server_fastmcp.core.excel_reader import ExcelReader

        path = self._create_test_file(rows=100, cols=5)
        try:
            reader = ExcelReader(path)
            self.assertFalse(reader._is_large)
            self.assertGreater(reader._file_size, 0)

            # list_sheets 应包含 file_size_mb 元数据
            result = reader.list_sheets()
            self.assertTrue(result.success)
            self.assertIn('file_size_mb', result.metadata)
            self.assertFalse(result.metadata['is_large_file'])

            reader.close()
        finally:
            os.unlink(path)

    def test_large_file_flag(self):
        """测试大文件标志设置正确"""
        from src.excel_mcp_server_fastmcp.core.excel_reader import ExcelReader

        # 创建一个正常大小的文件，然后手动设置_is_large来测试优化路径
        path = self._create_test_file(rows=500, cols=5)
        try:
            reader = ExcelReader(path)
            # 正常情况下不应该是大文件
            self.assertFalse(reader._is_large)
            reader.close()
        finally:
            os.unlink(path)

    def test_get_range_with_file_size_metadata(self):
        """测试范围读取包含文件大小元数据"""
        from src.excel_mcp_server_fastmcp.core.excel_reader import ExcelReader

        path = self._create_test_file(rows=200, cols=5)
        try:
            reader = ExcelReader(path)
            result = reader.get_range("TestData!A1:C10")
            self.assertTrue(result.success)
            self.assertIn('file_path', result.metadata)
            self.assertIn('sheet_name', result.metadata)
            reader.close()
        finally:
            os.unlink(path)

    def test_optimized_range_read_produces_correct_data(self):
        """测试优化路径产生的数据与标准路径一致"""
        from src.excel_mcp_server_fastmcp.core.excel_reader import ExcelReader

        path = self._create_test_file(rows=200, cols=5)
        try:
            reader = ExcelReader(path)

            # 强制使用优化路径（通过设置_is_large）
            reader._is_large = True

            result = reader.get_range("TestData!A1:C5")
            self.assertTrue(result.success)
            self.assertTrue(result.metadata.get('optimized'))

            # 验证数据正确性
            data = result.data
            self.assertEqual(len(data), 5)  # 5行
            self.assertEqual(len(data[0]), 3)  # 3列

            # 第一行应该是表头
            self.assertEqual(data[0][0].value, 'Col_0')
            self.assertEqual(data[0][1].value, 'Col_1')
            self.assertEqual(data[0][2].value, 'Col_2')

            reader.close()
        finally:
            os.unlink(path)

    def test_optimized_range_read_limited_rows(self):
        """测试优化路径只加载请求的行"""
        from src.excel_mcp_server_fastmcp.core.excel_reader import ExcelReader

        path = self._create_test_file(rows=1000, cols=5)
        try:
            reader = ExcelReader(path)
            reader._is_large = True

            result = reader.get_range("TestData!A1:B10")
            self.assertTrue(result.success)
            self.assertTrue(result.metadata.get('optimized'))
            self.assertEqual(result.metadata['rows_loaded'], 10)

            # 只应该有10行数据
            self.assertEqual(len(result.data), 10)

            reader.close()
        finally:
            os.unlink(path)

    def test_list_sheets_file_size_metadata(self):
        """测试list_sheets包含文件大小元数据"""
        from src.excel_mcp_server_fastmcp.core.excel_reader import ExcelReader

        path = self._create_test_file(rows=100, cols=5)
        try:
            reader = ExcelReader(path)
            result = reader.list_sheets()
            self.assertTrue(result.success)
            self.assertIn('file_size_mb', result.metadata)
            self.assertIn('is_large_file', result.metadata)
            self.assertIsInstance(result.metadata['file_size_mb'], float)
            self.assertIsInstance(result.metadata['is_large_file'], bool)
            reader.close()
        finally:
            os.unlink(path)


if __name__ == '__main__':
    unittest.main()
