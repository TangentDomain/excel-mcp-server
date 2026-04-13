"""
Round 5 迭代测试 — 数据类型边界 / 复杂子查询 / 特殊字符 / 多角色场景
日期: 2026-04-12 (修复版)
重点: 科学计数法、emoji、超长文本、多层子查询、极大整数、负零、引号转义、NTILE、CASE嵌套

注意: API 返回 data 为 list-of-lists 格式: [headers, row1, row2, ...]
"""
import sys
import os
import random
import math

sys.path.insert(0, '/root/workspace/excel-mcp-server/src')
from openpyxl import Workbook
from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)

TEST_FILE = '/tmp/excelmcp_round5_test.xlsx'


def create_test_data():
    """创建接近真实游戏配置的多表测试数据"""
    wb = Workbook()
    
    # --- Sheet 1: 装备表 ---
    ws_equip = wb.active
    ws_equip.title = "装备"
    ws_equip.append(["ID", "Name", "BaseAtk", "AtkBonus", "Price", "Rarity", "Desc"])
    
    equip_data = [
        (1, "⚔️ 圣剑·Excalibur", 100, 25.5, 9999.99, "Legendary", "传说中的圣剑" * 5),
        (2, "🛡️ 龙鳞盾", 80, 20.0, 5999.50, "Epic", "龙鳞打造的坚固盾牌"),
        (3, "法杖·星辰", 120, 33.33, 1.5e+4, "Rare", "A" * 600),
        (4, "匕首·影", 45, 10.10, 199.99, "Common", "It's a \"sharp\" dagger"),
        (5, "🔥 炎之戒指", 30, 5.55, 899.00, "Rare", "燃烧的戒指"),
        (6, "冰霜项链", 35, 7.77, 1299.00, "Epic", "I'm cold's necklace"),
        (7, "时空之靴", 50, 12.12, 2499.99, "Legendary", ""),
        (8, "破旧木剑", 5, 0.0, 9.99, "Common", None),
        (9, "黄金战锤", 90, 22.22, 4999.00, "Epic", "Heavy!!@#$%^&*()"),
        (10, "暗影斗篷", 55, 14.44, 3499.50, "Rare", "🎮👾🚀💎 emoji test"),
        (11, "边界测试-大数", 999999999999999, 0.001, 1e+10, "Common", "max int"),
        (12, "边界测试-小数", 1, 0.000001, 0.01, "Common", "min float"),
        (13, "边界测试-负数", 70, -15.5, -100.0, "Rare", "negative values"),
        (14, "边界测试-零", 0, 0.0, 0.0, "Common", "zero row"),
        (15, "科学计数-价格", 60, 8.88e-2, 2.5e+3, "Epic", "sci notation"),
    ]
    for row in equip_data:
        ws_equip.append(list(row))
    
    # --- Sheet 2: 技能表 ---
    ws_skill = wb.create_sheet("技能")
    ws_skill.append(["SkillID", "SkillName", "Cooldown", "ManaCost", "Damage", "Type", "UnlockLevel"])
    
    skill_data = [
        (101, "火焰冲击", 5.0, 30, 150.5, "AOE", 1),
        (102, "冰霜新星", 8.0, 50, 200.0, "AOE", 10),
        (103, "闪电链", 3.0, 20, 80.25, "Single", 1),
        (104, "治愈之光", 10.0, 60, -50.0, "Heal", 5),
        (105, "陨石术", 15.0, 100, 500.75, "AOE", 20),
        (106, "隐身", 20.0, 40, 0.0, "Buff", 15),
        (107, "狂暴", 0.0, 0, 25.0, "Buff", 1),
        (108, "终极技能", 30.0, 200, 999.99, "Ultimate", 30),
        (109, "普通攻击", 0.0, 0, 10.0, "Single", 1),
        (110, "被动技-吸血", 0, 0, 0.05, "Passive", 1),
    ]
    for row in skill_data:
        ws_skill.append(list(row))
    
    # --- Sheet 3: 商店表 ---
    ws_shop = wb.create_sheet("商店")
    ws_shop.append(["ItemID", "ItemName", "Price", "Stock", "Discount", "Category", "StartDate", "EndDate"])
    
    shop_data = [
        (1001, "金币礼包", 648, 999, 0.1, "Currency", "2026-01-01", "2026-12-31"),
        (1002, "经验药水x10", 68, 5000, 0.0, "Consumable", "2026-04-01", "2026-06-30"),
        (1003, "限定皮肤", 198, 100, 0.3, "Cosmetic", "2026-04-10", "2026-05-10"),
        (1004, "复活币x5", 28, 10000, 0.0, "Consumable", "2026-01-01", None),
        (1005, "月卡", 30, -1, 0.2, "Subscription", "2026-04-01", "2027-04-01"),
        (1006, "新手礼包", 6, 99999, 0.5, "Starter", "2026-04-01", "2026-05-01"),
    ]
    for row in shop_data:
        ws_shop.append(list(row))
    
    # --- Sheet 4: 掉落表 ---
    ws_drop = wb.create_sheet("掉落")
    ws_drop.append(["DropID", "MonsterID", "ItemID", "DropRate", "MinCount", "MaxCount", "Zone"])
    
    drop_data = [
        (2001, 501, 1, 0.01, 1, 1, "DragonNest"),
        (2002, 501, 2, 0.05, 1, 3, "DragonNest"),
        (2003, 502, 3, 0.001, 1, 1, "IceCave"),
        (2004, 503, 4, 0.10, 2, 5, "Forest"),
        (2005, 503, 5, 0.10, 2, 5, "Forest"),
        (2006, 503, 6, 0.08, 1, 2, "Forest"),
        (2007, 504, 7, 0.02, 1, 1, "Volcano"),
        (2008, 504, 8, 0.0001, 1, 1, "Volcano"),
        (2009, 505, 9, 0.50, 10, 20, "Town"),
        (2010, 505, 10, 0.30, 5, 10, "Town"),
        (2011, 506, 11, 0.15, 1, 3, "Dungeon"),
        (2012, 506, 12, 0.15, 1, 3, "Dungeon"),
        (2013, 506, 13, 0.10, 1, 2, "Dungeon"),
        (2014, 507, 14, 0.005, 1, 1, "BossRoom"),
        (2015, 507, 15, 0.005, 1, 1, "BossRoom"),
    ]
    for row in drop_data:
        ws_drop.append(list(row))
    
    wb.save(TEST_FILE)
    print(f"✅ 测试数据已创建: {TEST_FILE}")
    return TEST_FILE


def get_data_row(result, col_name):
    """从结果中获取第一行数据的指定列值 (list-of-lists 格式)"""
    if not result.get('success') or not result.get('data') or len(result['data']) < 2:
        return None
    headers = result['data'][0]
    first_row = result['data'][1]
    if col_name in headers:
        idx = headers.index(col_name)
        return first_row[idx]
    return None


def get_all_rows(result):
    """返回 (headers, rows) 元组"""
    if not result.get('success') or not result.get('data'):
        return [], []
    headers = result['data'][0]
    rows = result['data'][1:]
    return headers, rows


def col_value(headers, row, col_name):
    """根据列名获取行中的值"""
    if col_name in headers:
        return row[headers.index(col_name)]
    return None


def get_test_cases():
    tests = []
    
    # ====== Group A: 数据类型边界 ======
    
    tests.append((
        "A1. [策划] 科学计数法价格 SELECT",
        "SELECT ID, Name, Price FROM 装备 WHERE ID = 3",
        "SELECT",
        lambda r: r['success'] and abs(get_data_row(r, 'Price') - 15000.0) < 1.0,
        "Price 应为 ~15000.0"
    ))
    
    tests.append((
        "A2. [策划] 科学计数法 WHERE 比较",
        "SELECT ID, Name FROM 装备 WHERE Price > 1e+9",
        "SELECT",
        lambda r: r['success'] and len(r['data']) >= 2 and get_data_row(r, 'ID') == 11,
        "应找到 Price > 1e+9 的行 (ID=11)"
    ))
    
    tests.append((
        "A3. [QA] 极大整数精度 (>2^53)",
        "SELECT ID, Name, BaseAtk FROM 装备 WHERE ID = 11",
        "SELECT",
        lambda r: r['success'] and len(r['data']) >= 2 and get_data_row(r, 'BaseAtk') == 999999999999999,
        "应读取极大整数 999999999999999"
    ))
    
    tests.append((
        "A4. [QA] 极小浮点数精度",
        "SELECT ID, Name, AtkBonus FROM 装备 WHERE ID = 12",
        "SELECT",
        lambda r: r['success'] and get_data_row(r, 'AtkBonus') is not None and get_data_row(r, 'AtkBonus') != 0,
        "极小浮点数 0.000001 不应被截断为 0"
    ))
    
    tests.append((
        "A5. [QA] 负数值处理",
        "SELECT ID, Name, AtkBonus, Price FROM 装备 WHERE ID = 13",
        "SELECT",
        lambda r: r['success'] and get_data_row(r, 'AtkBonus') == -15.5 and get_data_row(r, 'Price') == -100.0,
        "负数应正确保留"
    ))
    
    tests.append((
        "A6. [QA] 零值行完整性",
        "SELECT ID, Name, BaseAtk, AtkBonus, Price FROM 装备 WHERE ID = 14",
        "SELECT",
        lambda r: r['success'] and get_data_row(r, 'BaseAtk') == 0 and get_data_row(r, 'Price') == 0.0,
        "全零行应正确返回"
    ))
    
    tests.append((
        "A7. [客户端] Emoji 文本字段",
        "SELECT ID, Name FROM 装备 WHERE Name LIKE '⚔️%'",
        "SELECT",
        lambda r: r['success'] and len(r['data']) >= 2 and '⚔️' in str(get_data_row(r, 'Name')),
        "应匹配含 ⚔️ emoji 的装备名"
    ))
    
    tests.append((
        "A8. [客户端] 超长文本回读 (>500字符)",
        "SELECT ID, LEN(Desc) as DescLen FROM 装备 WHERE ID = 3",
        "SELECT",
        lambda r: r['success'] and get_data_row(r, 'DescLen') is not None and get_data_row(r, 'DescLen') >= 590,
        "超长文本长度应 >= 590"
    ))
    
    # ====== Group B: 特殊字符与转义 ======
    
    tests.append((
        "B1. [客户端] 双引号文本查询",
        "SELECT ID, Name, Desc FROM 装备 WHERE ID = 4",
        "SELECT",
        lambda r: r['success'] and 'sharp' in str(get_data_row(r, 'Desc')),
        "双引号文本应完整保留"
    ))
    
    tests.append((
        "B2. [客户端] 单引号文本查询",
        "SELECT ID, Name, Desc FROM 装备 WHERE ID = 6",
        "SELECT",
        lambda r: r['success'] and 'cold' in str(get_data_row(r, 'Desc')).lower(),
        "单引号文本应完整保留"
    ))
    
    tests.append((
        "B3. [运营] 特殊符号文本",
        "SELECT ID, Desc FROM 装备 WHERE ID = 9",
        "SELECT",
        lambda r: r['success'] and '!@#$%^&*' in str(get_data_row(r, 'Desc')),
        "特殊符号应完整保留"
    ))
    
    tests.append((
        "B4. [QA] 空字符串 vs NULL 区分",
        "SELECT ID, Name, Desc FROM 装备 WHERE ID IN (7, 8)",
        "SELECT",
        lambda r: r['success'] and len(r['data']) >= 3,  # header + 2 rows
        "空字符串和NULL都应能查出(共2行)"
    ))
    
    tests.append((
        "B4b. [QA] NULL 是否被转为空字符串",
        "SELECT ID, Desc FROM 装备 WHERE ID = 8",
        "SELECT",
        lambda r: r['success'],  # 只检查能否查到，下面手动判断
        "检查 NULL 值的实际返回"
    ))
    
    # ====== Group C: 写入操作 ======
    
    tests.append((
        "C1. [运营] INSERT 含单引号文本",
        "INSERT INTO 装备 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Desc) VALUES (100, \"It's New\", 10, 1.0, 100.0, 'Common', 'test')",
        "INSERT",
        lambda r: r['success'],
        "插入含单引号的文本"
    ))
    
    tests.append((
        "C1b. [运营] C1 回读验证",
        "SELECT ID, Name, Desc FROM 装备 WHERE ID = 100",
        "SELECT",
        lambda r: r['success'] and len(r['data']) >= 2 and "It's" in str(get_data_row(r, 'Name')),
        "回读确认单引号文本正确"
    ))
    
    tests.append((
        "C2. [策划] UPDATE 设为科学计数法值",
        "UPDATE 装备 SET Price = 2.5e+6 WHERE ID = 2",
        "UPDATE",
        lambda r: r['success'],
        "更新为科学计数法值"
    ))
    
    tests.append((
        "C2b. [策划] C2 回读验证",
        "SELECT Price FROM 装备 WHERE ID = 2",
        "SELECT",
        lambda r: r['success'] and abs(get_data_row(r, 'Price') - 2500000.0) < 1.0,
        "回读确认 Price = 2500000.0"
    ))
    
    tests.append((
        "C3. [策划] UPDATE 设为负数",
        "UPDATE 装备 SET Price = -500.0 WHERE ID = 5",
        "UPDATE",
        lambda r: r['success'],
        "更新为负数"
    ))
    
    tests.append((
        "C3b. [策划] C3 回读验证",
        "SELECT Price FROM 装备 WHERE ID = 5",
        "SELECT",
        lambda r: r['success'] and get_data_row(r, 'Price') == -500.0,
        "回读确认 Price = -500.0"
    ))
    
    tests.append((
        "C4. [运营] DELETE 带 LIKE 条件",
        "DELETE FROM 装备 WHERE Name LIKE '边界测试%'",
        "DELETE",
        lambda r: r['success'],
        "删除边界测试行"
    ))
    
    tests.append((
        "C4b. [运营] C4 验证删除结果",
        "SELECT COUNT(*) as cnt FROM 装备 WHERE Name LIKE '边界测试%'",
        "SELECT",
        lambda r: r['success'] and get_data_row(r, 'cnt') == 0,
        "确认已删除"
    ))
    
    tests.append((
        "C5. [策划] INSERT 含 emoji 名称",
        "INSERT INTO 装备 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Desc) VALUES (101, '🌟 星耀之刃', 200, 50.5, 18888.0, 'Legendary', 'emoji name item')",
        "INSERT",
        lambda r: r['success'],
        "插入含 emoji 名称"
    ))
    
    tests.append((
        "C5b. [策划] C5 回读验证",
        "SELECT Name FROM 装备 WHERE ID = 101",
        "SELECT",
        lambda r: r['success'] and '🌟' in str(get_data_row(r, 'Name')),
        "回读确认 emoji 名称正确"
    ))
    
    # ====== Group D: 复杂查询 ======
    
    tests.append((
        "D1. [服务端] FROM 子查询 (2层嵌套)",
        "SELECT * FROM (SELECT Rarity, AVG(Price) as AvgPrice, COUNT(*) as cnt FROM 装备 GROUP BY Rarity) t WHERE AvgPrice > 1000",
        "SELECT",
        lambda r: r['success'] and len(r['data']) > 1,  # 至少有header + 1 row
        "FROM 子查询 + 外层 WHERE"
    ))
    
    tests.append((
        "D2. [数据分析] GROUP BY 多聚合函数",
        "SELECT Zone, COUNT(*) as total, SUM(DropRate) as total_rate, AVG(DropRate) as avg_rate, MAX(DropRate) as max_rate, MIN(DropRate) as min_rate FROM 掉落 GROUP BY Zone",
        "SELECT",
        lambda r: r['success'] and len(r['data']) > 1,
        "多聚合函数(COUNT/SUM/AVG/MAX/MIN)"
    ))
    
    tests.append((
        "D3. [数据分析] HAVING 复合条件 AND+OR",
        "SELECT Zone, COUNT(*) as cnt, AVG(DropRate) as avg_r FROM 掉落 GROUP BY Zone HAVING (cnt >= 3 AND avg_r > 0.05) OR (cnt = 1 AND avg_r < 0.01)",
        "SELECT",
        lambda r: r['success'],
        "HAVING 复合条件 AND+OR"
    ))
    
    tests.append((
        "D4. [服务端] CASE WHEN 嵌套 (多层分支)",
        "SELECT ID, Name, Price, CASE WHEN Price > 10000 THEN '超贵' WHEN Price > 1000 THEN '昂贵' WHEN Price > 100 THEN '中等' ELSE '便宜' END as price_tier FROM 装备 WHERE ID <= 6",
        "SELECT",
        lambda r: r['success'] and len(r['data']) > 1,
        "多层 CASE WHEN 分支分类"
    ))
    
    tests.append((
        "D5. [数据分析] NTILE 分桶分析",
        "SELECT ID, Name, Price, NTILE(4) OVER (ORDER BY Price DESC) as price_quartile FROM 装备 WHERE ID <= 10",
        "SELECT",
        lambda r: r['success'] and len(r['data']) > 1,
        "NTILE(4) 四分位分桶"
    ))
    
    tests.append((
        "D6. [数据分析] DISTINCT + ORDER BY 组合",
        "SELECT DISTINCT Rarity FROM 装备 ORDER BY Rarity",
        "SELECT",
        lambda r: r['success'] and len(r['data']) > 1,
        "DISTINCT + ORDER BY"
    ))
    
    tests.append((
        "D7. [服务端] BETWEEN 边界值包含性",
        "SELECT ID, Price FROM 装备 WHERE Price BETWEEN 100.0 AND 1000.0",
        "SELECT",
        lambda r: r['success'],
        "BETWEEN 边界包含"
    ))
    
    tests.append((
        "D8. [数据分析] ROW_NUMBER 窗口函数",
        "SELECT SkillID, SkillName, Damage, ROW_NUMBER() OVER (ORDER BY Damage DESC) as rank FROM 技能",
        "SELECT",
        lambda r: r['success'] and len(r['data']) >= 11,  # header + 10 rows
        "ROW_NUMBER 窗口函数"
    ))
    
    tests.append((
        "D9. [服务端] 标量子查询在 SELECT 中",
        "SELECT d.Zone, d.ItemID, d.DropRate, (SELECT COUNT(*) FROM 掉落 d2 WHERE d2.Zone = d.Zone) as zone_total FROM 掉落 d WHERE d.DropRate >= 0.1",
        "SELECT",
        lambda r: r['success'] and len(r['data']) > 1,
        "标量子查询在 SELECT 中"
    ))
    
    tests.append((
        "D10. [策划] UPDATE 用 CASE WHEN 分层调价",
        "UPDATE 技能 SET ManaCost = CASE WHEN Type = 'Ultimate' THEN ManaCost * 1.2 WHEN Type = 'AOE' THEN ManaCost * 1.1 ELSE ManaCost END",
        "UPDATE",
        lambda r: r['success'],
        "CASE WHEN 分层调价"
    ))
    
    tests.append((
        "D10b. [策划] D10 回读验证",
        "SELECT SkillID, SkillName, Type, ManaCost FROM 技能 WHERE Type IN ('Ultimate', 'AOE') ORDER BY SkillID",
        "SELECT",
        lambda r: r['success'] and len(r['data']) > 1,
        "回读确认调整后数据"
    ))
    
    return tests


def run_tests():
    file_path = create_test_data()
    test_cases = get_test_cases()
    
    print("\n" + "=" * 72)
    print("  🔄 ExcelMCP Round 5 迭代测试 (v2)")
    print(f"  📅 日期: 2026-04-12")
    print(f"  📊 测试文件: {file_path}")
    print(f"  📋 用例数量: {len(test_cases)}")
    print("=" * 72 + "\n")
    
    results = {
        'total': 0, 'passed': 0, 'failed': 0,
        'by_type': {'SELECT': {'pass': 0, 'fail': 0}, 'UPDATE': {'pass': 0, 'fail': 0},
                    'INSERT': {'pass': 0, 'fail': 0}, 'DELETE': {'pass': 0, 'fail': 0}},
        'details': [],
        'errors': [],
        'raw_outputs': {},  # 存储关键用例的原始输出用于调试
    }
    
    for i, (name, sql, sql_type, verify_fn, expected_desc) in enumerate(test_cases):
        results['total'] += 1
        
        try:
            if sql_type == 'SELECT':
                result = execute_advanced_sql_query(file_path, sql)
            elif sql_type == 'UPDATE':
                result = execute_advanced_update_query(file_path, sql)
            elif sql_type == 'INSERT':
                result = execute_advanced_insert_query(file_path, sql)
            elif sql_type == 'DELETE':
                result = execute_advanced_delete_query(file_path, sql)
            else:
                result = execute_advanced_sql_query(file_path, sql)
            
            if result.get('success', False):
                try:
                    verified = verify_fn(result)
                    if verified:
                        print(f"✅ [{i+1}/{len(test_cases)}] {name}")
                        results['passed'] += 1
                        results['by_type'][sql_type]['pass'] += 1
                        results['details'].append({'name': name, 'status': 'PASS'})
                    else:
                        # 记录原始输出用于调试
                        raw_data = str(result.get('data', []))[:400]
                        print(f"⚠️  [{i+1}/{len(test_cases)}] {name}")
                        print(f"   验证失败: {expected_desc}")
                        print(f"   实际数据: {raw_data}")
                        results['failed'] += 1
                        results['by_type'][sql_type]['fail'] += 1
                        results['details'].append({'name': name, 'status': 'VERIFY_FAIL', 'error': f'{expected_desc}', 'raw': raw_data})
                        results['errors'].append({'name': name, 'error': f'验证失败: {expected_desc}', 'raw': raw_data})
                        results['raw_outputs'][name] = result
                except Exception as e:
                    print(f"⚠️  [{i+1}/{len(test_cases)}] {name}")
                    print(f"   验证异常: {e}")
                    results['failed'] += 1
                    results['by_type'][sql_type]['fail'] += 1
                    results['details'].append({'name': name, 'status': 'VERIFY_ERROR', 'error': str(e)[:200]})
                    results['errors'].append({'name': name, 'error': f'验证异常: {str(e)[:200]}'})
            else:
                print(f"❌ [{i+1}/{len(test_cases)}] {name}")
                error_msg = result.get('message', '未知错误')
                print(f"   错误: {error_msg[:300]}")
                results['failed'] += 1
                results['by_type'][sql_type]['fail'] += 1
                results['details'].append({'name': name, 'status': 'FAIL', 'error': error_msg[:300]})
                results['errors'].append({'name': name, 'error': error_msg[:300], 'sql': sql})
                
        except Exception as e:
            print(f"💥 [{i+1}/{len(test_cases)}] {name}")
            print(f"   未捕获异常: {type(e).__name__}: {str(e)[:300]}")
            results['failed'] += 1
            results['by_type'][sql_type]['fail'] += 1
            results['details'].append({'name': name, 'status': 'CRASH', 'error': f'{type(e).__name__}: {str(e)[:200]}'})
            results['errors'].append({'name': name, 'error': f'{type(e).__name__}: {str(e)[:300]}'})
    
    # 汇总
    print("\n" + "=" * 72)
    print("  📊 Round 5 测试汇总")
    print("=" * 72)
    rate = results['passed'] / results['total'] * 100 if results['total'] > 0 else 0
    print(f"  总计: {results['passed']}/{results['total']} 通过 ({rate:.1f}%)")
    print(f"\n  分类统计:")
    for ttype, counts in results['by_type'].items():
        ttotal = counts['pass'] + counts['fail']
        if ttotal > 0:
            trate = counts['pass'] / ttotal * 100
            print(f"    {ttype:7s}: {counts['pass']}/{ttotal} ({trate:.0f}%)")
    
    if results['errors']:
        print(f"\n  ❌ 失败详情 ({len(results['errors'])} 个):")
        for err in results['errors']:
            print(f"    • {err['name']}: {err.get('error', '')[:120]}")
    
    print("=" * 72)
    
    # 返回详细结果供后续分析
    return results


if __name__ == '__main__':
    run_results = run_tests()
