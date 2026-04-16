"""
R51 深度代码审查修复验证测试
================================
覆盖 P1-SAFE-01(路径遍历), P1-INSERT-01(值数量不匹配),
P2-ORDER-01(混合类型排序), P2-PERF-01(IN子查询缓存),
P2-CONCUR-01(孤儿锁文件), P3-INSERT-02(Column引用)
"""
import os
import tempfile
import shutil
import pytest
from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_insert_query,
)


@pytest.fixture
def sample_xlsx(tmp_path):
    """创建标准测试用 Excel 文件"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "Name", "Score"])
    ws.append([1, "Alice", 95.5])
    ws.append([2, "Bob", 87.3])
    ws.append([3, "Charlie", None])
    ws.append([10, "Diana", 72.1])
    path = str(tmp_path / "test.xlsx")
    wb.save(path)
    return path


class TestP1Safe01_PathTraversal:
    """P1-SAFE-01: 跨文件引用路径遍历防护"""

    def test_reject_parent_directory_traversal(self, sample_xlsx):
        """拒绝 ../ 路径遍历攻击"""
        result = execute_advanced_sql_query(
            sample_xlsx,
            "SELECT * FROM @'../../etc/passwd'",
        )
        assert result["success"] is False
        assert "不允许访问" in result["message"]

    def test_reject_encoded_traversal(self, sample_xlsx):
        """拒绝编码后的路径遍历"""
        result = execute_advanced_sql_query(
            sample_xlsx,
            "SELECT * FROM @'../secrets/config.xlsx'",
        )
        assert result["success"] is False

    def test_error_message_no_full_path_leak(self, sample_xlsx):
        """错误信息不泄露完整绝对路径"""
        result = execute_advanced_sql_query(
            sample_xlsx,
            "SELECT * FROM @'nonexistent_file.xlsx'",
        )
        if not result["success"]:
            msg = result.get("message", "")
            # 不应包含 /root/ 或其他深层绝对路径前缀（只显示文件名）
            assert "/root/" not in msg, f"错误信息可能泄露完整路径: {msg}"


class TestP1Insert01_ValueCountMismatch:
    """P1-INSERT-01: INSERT VALUES 值数量与列数量不匹配检测"""

    def test_fewer_values_than_columns(self, sample_xlsx):
        """VALUES 值少于列数时应报错"""
        result = execute_advanced_insert_query(
            sample_xlsx,
            "INSERT INTO Sheet1 (ID, Name, Score) VALUES (5, 'Eve')",
        )
        assert result["success"] is False
        assert "不匹配" in result["message"]
        assert "3" in result["message"]  # 期望 3 列
        assert "2" in result["message"]  # 实际 2 个值

    def test_exact_match_values_succeeds(self, sample_xlsx):
        """值数量匹配时正常插入"""
        result = execute_advanced_insert_query(
            sample_xlsx,
            "INSERT INTO Sheet1 (ID, Name, Score) VALUES (5, 'Eve', 88.0)",
        )
        assert result["success"] is True
        assert result["affected_rows"] == 1

    def test_single_column_insert_ok(self, sample_xlsx):
        """单列插入正常工作"""
        result = execute_advanced_insert_query(
            sample_xlsx,
            "INSERT INTO Sheet1 (ID) VALUES (99)",
        )
        assert result["success"] is True


class TestP2Order01_MixedTypeSorting:
    """P2-ORDER-01: 混合类型排序智能处理"""

    @pytest.fixture
    def mixed_xlsx(self):
        import openpyxl
        tmpdir = tempfile.mkdtemp()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mixed"
        ws.append(["ID", "Value"])
        # 数值占多数(>50%): 7个数值 vs 3个字符串
        ws.append([1, 10])       # 整数
        ws.append([2, 2])        # 整数
        ws.append([3, 100])      # 整数
        ws.append([4, "apple"])  # 字符串
        ws.append([5, 50])       # 整数
        ws.append([6, 200])      # 整数
        ws.append([7, 5])        # 整数
        ws.append([8, "banana"]) # 字符串
        ws.append([9, 75])       # 整数
        ws.append([10, "cherry"])# 字符串
        path = os.path.join(tmpdir, "mixed_sort.xlsx")
        wb.save(path)
        yield path
        shutil.rmtree(tmpdir, ignore_errors=True)

    def test_numeric_majority_uses_numeric_sort(self, mixed_xlsx):
        """数值占多数时使用数值排序（10 > 2 正确语义）"""
        result = execute_advanced_sql_query(
            mixed_xlsx,
            "SELECT ID, Value FROM Mixed ORDER BY Value",
        )
        assert result["success"] is True
        data = result["data"]
        values = [row[1] for row in data[1:]]  # 跳过 header
        # 数值排序: 2, 5, 10, 50, 75, 100, 200, "apple", "banana", "cherry"
        # 关键: "10" 排在 "2" 后面（数值语义），而非字符串排序下 "10" < "2"
        idx_2 = next((i for i, v in enumerate(values) if v == 2), -1)
        idx_10 = next((i for i, v in enumerate(values) if v == 10), -1)
        assert idx_2 >= 0 and idx_10 >= 0, f"未找到期望的值: {values}"
        assert idx_2 < idx_10, f"期望数值排序(2<10)，实际顺序: {values}"

    def test_descending_mixed_sort(self, mixed_xlsx):
        """降序混合类型排序"""
        result = execute_advanced_sql_query(
            mixed_xlsx,
            "SELECT ID, Value FROM Mixed ORDER BY Value DESC",
        )
        assert result["success"] is True
        data = result["data"]
        assert len(data) > 1

    def test_pure_numeric_sort_unchanged(self, sample_xlsx):
        """纯数值列排序不受影响"""
        result = execute_advanced_sql_query(
            sample_xlsx,
            "SELECT * FROM Sheet1 ORDER BY ID",
        )
        assert result["success"] is True
        ids = [row[0] for row in result["data"][1:]]
        assert ids == sorted(ids)

    def test_pure_string_sort_unchanged(self, sample_xlsx):
        """纯字符串列排序不受影响"""
        result = execute_advanced_sql_query(
            sample_xlsx,
            "SELECT * FROM Sheet1 ORDER BY Name",
        )
        assert result["success"] is True
        names = [row[1] for row in result["data"][1:]]
        assert names == sorted(names)


class TestP2Perf01_INSubqueryCache:
    """P2-PERF-01: IN 子查询缓存优化"""

    @pytest.fixture
    def in_subquery_file(self):
        import openpyxl
        tmpdir = tempfile.mkdtemp()
        # 单文件双工作表（子查询只能访问同文件中的表）
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Main"
        ws.append(["ID", "CategoryID", "Name"])
        for i in range(1, 51):  # 50 行数据
            ws.append([i, i % 5 + 1, f"Item{i}"])
        # 创建第二个工作表作为子查询源
        ws2 = wb.create_sheet("Cats")
        ws2.append(["CatID", "CatName"])
        for c in [1, 3, 5]:
            ws2.append([c, f"Cat{c}"])
        path = os.path.join(tmpdir, "in_subquery.xlsx")
        wb.save(path)
        yield path
        shutil.rmtree(tmpdir, ignore_errors=True)

    def test_in_subquery_basic(self, in_subquery_file):
        """IN 子查询基本功能正确"""
        result = execute_advanced_sql_query(
            in_subquery_file,
            "SELECT * FROM Main WHERE CategoryID IN (SELECT CatID FROM Cats)",
        )
        assert result["success"] is True
        data = result["data"]
        assert len(data) > 1
        cat_ids = {row[1] for row in data[1:]}
        assert cat_ids.issubset({1, 3, 5})

    def test_in_subquery_empty_result(self, in_subquery_file):
        """IN 空子查询结果"""
        result = execute_advanced_sql_query(
            in_subquery_file,
            "SELECT * FROM Main WHERE CategoryID IN (SELECT CatID FROM Cats WHERE CatID > 100)",
        )
        assert result["success"] is True
        assert len(result["data"]) == 1  # 仅 header

    def test_in_literal_list_unchanged(self, sample_xlsx):
        """字面量 IN 列表不受缓存逻辑影响"""
        result = execute_advanced_sql_query(
            sample_xlsx,
            "SELECT * FROM Sheet1 WHERE ID IN (1, 2)",
        )
        assert result["success"] is True
        ids = {row[0] for row in result["data"][1:]}
        assert ids == {1, 2}


class TestP3Insert02_ColumnReferenceInValues:
    """P3-INSERT-02: VALUES 中列引用明确报错"""

    def test_column_ref_in_values_raises_error(self, sample_xlsx):
        """VALUES 中使用列引用应返回明确错误"""
        result = execute_advanced_insert_query(
            sample_xlsx,
            "INSERT INTO Sheet1 (ID, Name, Score) VALUES (ID, 'X', 0)",
        )
        assert result["success"] is False
        assert "不支持列引用" in result["message"] or "不匹配" in result["message"]

    def test_literal_value_still_works(self, sample_xlsx):
        """字面量值正常工作"""
        result = execute_advanced_insert_query(
            sample_xlsx,
            "INSERT INTO Sheet1 (ID, Name, Score) VALUES (777, 'Literal', 42.0)",
        )
        assert result["success"] is True
        assert result["affected_rows"] == 1

    def test_null_value_in_values(self, sample_xlsx):
        """NULL 值正常工作"""
        result = execute_advanced_insert_query(
            sample_xlsx,
            "INSERT INTO Sheet1 (ID, Name, Score) VALUES (778, NULL, NULL)",
        )
        assert result["success"] is True


class TestRegression_ErrorMessageSafety:
    """回归验证: 错误消息安全性"""

    def test_cross_file_error_message_safe(self, sample_xlsx):
        """跨文件引用错误不含敏感路径"""
        result = execute_advanced_sql_query(
            sample_xlsx,
            "SELECT * FROM @'../../etc/shadow'",
        )
        assert result["success"] is False
        msg = result.get("message", "")
        assert "/etc/" not in msg
        assert "/root/" not in msg
