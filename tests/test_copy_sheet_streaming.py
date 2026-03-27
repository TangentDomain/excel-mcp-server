"""测试copy_sheet的streaming参数支持"""
import pytest
import os
import tempfile
from openpyxl import Workbook, load_workbook


def _create_test_file(path, rows=100, cols=10):
    """创建测试Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=f"R{r}C{c}")
    wb.save(path)
    wb.close()


def test_copy_sheet_streaming_true():
    """测试streaming=True复制工作表"""
    from excel_mcp_server_fastmcp.core.excel_manager import ExcelManager

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        _create_test_file(path, rows=50, cols=5)

        mgr = ExcelManager(path)
        result = mgr.copy_sheet("Sheet1", "副本", streaming=True)

        assert result.success, f"复制失败: {result.error}"
        assert result.metadata.get('mode') == 'streaming', f"应使用streaming模式: {result.metadata}"
        assert result.metadata.get('copied_rows') == 50, f"行数不匹配: {result.metadata.get('copied_rows')}"
        assert result.metadata.get('copied_columns') == 5, f"列数不匹配: {result.metadata.get('copied_columns')}"

        # 验证复制后的文件有两个工作表
        wb2 = load_workbook(path)
        assert "Sheet1" in wb2.sheetnames, "原工作表应存在"
        assert "副本" in wb2.sheetnames, "副本工作表应存在"
        wb2.close()

    finally:
        os.unlink(path)


def test_copy_sheet_streaming_false():
    """测试streaming=False使用传统openpyxl模式"""
    from excel_mcp_server_fastmcp.core.excel_manager import ExcelManager

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        _create_test_file(path, rows=30, cols=5)

        mgr = ExcelManager(path)
        result = mgr.copy_sheet("Sheet1", "传统副本", streaming=False)

        assert result.success, f"复制失败: {result.error}"
        # streaming=False不设置mode metadata
        assert result.metadata.get('mode') != 'streaming', f"不应使用streaming模式"

        wb2 = load_workbook(path)
        assert "传统副本" in wb2.sheetnames, "副本工作表应存在"
        wb2.close()

    finally:
        os.unlink(path)


def test_copy_sheet_auto_name():
    """测试自动生成副本名称"""
    from excel_mcp_server_fastmcp.core.excel_manager import ExcelManager

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        _create_test_file(path, rows=20, cols=3)

        mgr = ExcelManager(path)
        result = mgr.copy_sheet("Sheet1", streaming=True)

        assert result.success, f"复制失败: {result.error}"
        new_name = result.metadata.get('new_name', '')
        assert new_name == "Sheet1_副本", f"自动名称应为'Sheet1_副本': {new_name}"

    finally:
        os.unlink(path)


def test_copy_sheet_name_conflict():
    """测试名称冲突时自动编号"""
    from excel_mcp_server_fastmcp.core.excel_manager import ExcelManager

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        _create_test_file(path, rows=20, cols=3)

        mgr = ExcelManager(path)
        # 先复制一次
        r1 = mgr.copy_sheet("Sheet1", "副本", streaming=True)
        assert r1.success

        # 再复制同名，应该自动编号
        r2 = mgr.copy_sheet("Sheet1", "副本", streaming=True)
        assert r2.success, f"冲突后复制失败: {r2.error}"
        assert r2.metadata.get('new_name') == "副本_1", f"冲突名称应为'副本_1': {r2.metadata.get('new_name')}"

    finally:
        os.unlink(path)


def test_copy_sheet_data_integrity():
    """验证复制后的数据完整性"""
    from excel_mcp_server_fastmcp.core.excel_manager import ExcelManager

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        # 创建有特殊数据的文件
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.cell(row=1, column=1, value="ID")
        ws.cell(row=1, column=2, value="Name")
        ws.cell(row=1, column=3, value="Value")
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="测试")
        ws.cell(row=2, column=3, value=3.14)
        ws.cell(row=3, column=1, value=2)
        ws.cell(row=3, column=2, value="数据")
        ws.cell(row=3, column=3, value=2.71)
        wb.save(path)
        wb.close()

        mgr = ExcelManager(path)
        result = mgr.copy_sheet("Data", "DataCopy", streaming=True)
        assert result.success

        # 验证复制后的数据
        wb2 = load_workbook(path)
        ws2 = wb2["DataCopy"]
        assert ws2.cell(row=1, column=1).value == "ID"
        assert ws2.cell(row=2, column=1).value == 1
        assert ws2.cell(row=2, column=2).value == "测试"
        assert ws2.cell(row=2, column=3).value == 3.14
        wb2.close()

    finally:
        os.unlink(path)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
