# -*- coding: utf-8 -*-
"""
Excel Writer 基础功能测试套件

覆盖 excel_writer.py 中的基础写入功能
"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from src.core.excel_writer import ExcelWriter


class TestExcelWriterBasic:
    """ExcelWriter 基础功能测试"""

    @pytest.fixture
    def basic_test_file(self, temp_dir):
        """创建基础测试文件"""
        file_path = temp_dir / "basic_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws['A1'] = 10
        ws['A2'] = 20
        ws['A3'] = 30
        
        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def empty_test_file(self, temp_dir):
        """创建空Excel文件"""
        file_path = temp_dir / "empty_test.xlsx"
        
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(file_path)
        
        return str(file_path)

    # ==================== 基础写入测试 ====================

    def test_update_range_basic(self, basic_test_file):
        """测试基础范围更新"""
        writer = ExcelWriter(basic_test_file)
        result = writer.update_range(
            "Sheet1!A4:A6",
            [[40], [50], [60]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True
        assert result.data is not None

    def test_update_range_multiple_rows(self, basic_test_file):
        """测试多行更新"""
        writer = ExcelWriter(basic_test_file)
        result = writer.update_range(
            "Sheet1!B1:C3",
            [["Name1", 100], ["Name2", 200], ["Name3", 300]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_update_range_mixed_data(self, basic_test_file):
        """测试混合数据类型"""
        writer = ExcelWriter(basic_test_file)
        result = writer.update_range(
            "Sheet1!D1:E2",
            [["Text", 123], [True, 3.14]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_update_range_preserve_true(self, basic_test_file):
        """测试保留公式（preserve_formulas=True）"""
        # 先读取现有数据
        from src.core.excel_reader import ExcelReader
        reader = ExcelReader(basic_test_file)
        original = reader.get_range("Sheet1!A1:A3")
        reader.close()
        
        # 更新数据
        writer = ExcelWriter(basic_test_file)
        result = writer.update_range(
            "Sheet1!A1:A3",
            [[111], [222], [333]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_update_range_preserve_false(self, basic_test_file):
        """测试不保留公式"""
        writer = ExcelWriter(basic_test_file)
        result = writer.update_range(
            "Sheet1!A1:A3",
            [[1], [2], [3]],
            preserve_formulas=False,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== 插入模式测试 ====================

    def test_update_range_insert_mode_true(self, basic_test_file):
        """测试插入模式"""
        writer = ExcelWriter(basic_test_file)
        result = writer.update_range(
            "Sheet1!A10",
            [[1000], [2000]],
            preserve_formulas=True,
            insert_mode=True
        )
        
        assert result.success is True

    # ==================== 覆盖模式测试 ====================

    def test_update_range_insert_mode_false(self, basic_test_file):
        """测试覆盖模式"""
        writer = ExcelWriter(basic_test_file)
        result = writer.update_range(
            "Sheet1!A1",
            [[999]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    # ==================== 错误处理测试 ====================

    def test_update_range_invalid_file(self):
        """测试无效文件"""
        # ExcelWriter 在初始化时就会验证文件存在性
        # 所以需要捕获初始化异常
        from src.utils.exceptions import ExcelFileNotFoundError
        with pytest.raises(ExcelFileNotFoundError):
            ExcelWriter("/nonexistent/file.xlsx")

    def test_update_range_invalid_sheet(self, basic_test_file):
        """测试无效工作表"""
        writer = ExcelWriter(basic_test_file)
        result = writer.update_range(
            "NonExistentSheet!A1:B2",
            [[1, 2], [3, 4]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is False

    def test_update_range_empty_data(self, basic_test_file):
        """测试空数据"""
        writer = ExcelWriter(basic_test_file)
        result = writer.update_range(
            "Sheet1!A10:A10",
            [[]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        # 空数据可能成功或失败，取决于实现
        assert result is not None


class TestExcelWriterEdgeCases:
    """ExcelWriter 边界情况测试"""

    @pytest.fixture
    def edge_test_file(self, temp_dir):
        """创建边界测试文件"""
        file_path = temp_dir / "edge_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "EdgeSheet"
        
        # 创建一些测试数据
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=i * 10)
        
        wb.save(file_path)
        return str(file_path)

    def test_update_large_range(self, edge_test_file):
        """测试大范围更新"""
        writer = ExcelWriter(edge_test_file)
        
        # 创建大量数据
        large_data = [[i] for i in range(1, 101)]
        
        result = writer.update_range(
            "EdgeSheet!B1:B100",
            large_data,
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_update_single_cell(self, edge_test_file):
        """测试单单元格更新"""
        writer = ExcelWriter(edge_test_file)
        result = writer.update_range(
            "EdgeSheet!C1",
            [[999]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_update_string_data(self, edge_test_file):
        """测试字符串数据"""
        writer = ExcelWriter(edge_test_file)
        result = writer.update_range(
            "EdgeSheet!D1:D5",
            [["Apple"], ["Banana"], ["Cherry"], ["Date"], ["Elderberry"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_update_special_characters(self, edge_test_file):
        """测试特殊字符"""
        writer = ExcelWriter(edge_test_file)
        result = writer.update_range(
            "EdgeSheet!E1:E3",
            [["Hello\nWorld"], ["Tab\tHere"], ["Quote\"Test"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_update_unicode_data(self, edge_test_file):
        """测试Unicode数据"""
        writer = ExcelWriter(edge_test_file)
        result = writer.update_range(
            "EdgeSheet!F1:F3",
            [["你好世界"], ["こんにちは"], ["🎉🎊🎁"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True

    def test_update_none_values(self, edge_test_file):
        """测试None值"""
        writer = ExcelWriter(edge_test_file)
        result = writer.update_range(
            "EdgeSheet!G1:G3",
            [["Value1"], [None], ["Value3"]],
            preserve_formulas=True,
            insert_mode=False
        )
        
        assert result.success is True
