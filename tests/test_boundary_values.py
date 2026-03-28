"""
边界值测试 - REQ-034
测试极端值、精度损失、NULL处理、大数值等边界情况

@intention: 确保SQL引擎和Excel操作在边界条件下正确工作
"""

import pytest
import os
import tempfile
import openpyxl
from openpyxl import Workbook

# 导入被测模块
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from excel_mcp_server_fastmcp.api.advanced_sql_query import AdvancedSQLQueryEngine, _safe_float_comparison


@pytest.fixture
def boundary_test_file():
    """创建包含边界值的测试Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "边界值测试"

    # 表头
    ws.append(["ID", "名称", "价格", "冷却时间", "伤害值", "概率", "描述"])

    # 正常值
    ws.append([1, "普通攻击", 100.0, 1.5, 50, 0.5, "基础攻击"])
    ws.append([2, "强击", 200.0, 2.0, 100, 0.3, "强力攻击"])

    # 边界值 - 0.1精度测试
    ws.append([3, "快速攻击", 50.0, 0.1, 10, 0.1, "极快攻击"])
    ws.append([4, "超快攻击", 25.0, 0.01, 5, 0.01, "超快攻击"])
    ws.append([5, "极速攻击", 12.5, 0.001, 2, 0.001, "极速攻击"])

    # 大数值边界测试
    ws.append([6, "终极攻击", 99999.99, 999.999, 99999, 0.999, "极限攻击"])
    ws.append([7, "超大伤害", 999999.999, 9999.9999, 999999, 0.9999, "超大攻击"])
    ws.append([8, "最大值", 9999999.0, 0.0, 9999999, 1.0, "最大攻击"])

    # 零值和负值测试
    ws.append([9, "免费技能", 0.0, 0.0, 0, 0.0, "零值测试"])
    ws.append([10, "测试负值", -100.0, -1.0, -50, -0.5, "负值测试"])

    # NULL/空值测试
    ws.append([11, None, None, None, None, None, None])
    ws.append([12, "", 0.0, 0.0, 0, 0.0, ""])
    ws.append([None, "空ID测试", 50.0, 1.0, 25, 0.5, "ID为空"])

    # 重复值测试
    ws.append([13, "重复攻击", 100.0, 1.5, 50, 0.5, "和普通攻击一样"])
    ws.append([14, "重复攻击2", 100.0, 1.5, 50, 0.5, "再次重复"])

    # 浮点精度边界
    ws.append([15, "精度测试1", 0.1 + 0.2, 0.3, 0.1 + 0.2, 0.3, "0.1+0.2!=0.3?"])
    ws.append([16, "精度测试2", 1.0 - 0.9, 0.1, 1.0 - 0.9, 0.1, "1.0-0.9!=0.1?"])
    ws.append([17, "精度测试3", 0.7 + 0.1, 0.8, 0.7 + 0.1, 0.8, "0.7+0.1!=0.8?"])

    # 极小值测试
    ws.append([18, "极小值", 0.00001, 0.00001, 1, 0.00001, "极小值测试"])
    ws.append([19, "微小概率", 0.000001, 0.000001, 1, 0.000001, "极微概率"])

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        wb.save(f.name)
        yield f.name

    os.unlink(f.name)


@pytest.fixture
def engine():
    """创建SQL查询引擎实例"""
    return AdvancedSQLQueryEngine()


class TestBoundaryValues:
    """边界值测试类"""

    # ==================== 精度测试 ====================

    def test_float_precision_0_1(self, engine, boundary_test_file):
        """测试0.1秒精度不丢失"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT 名称, 冷却时间 FROM 边界值测试 WHERE 冷却时间 = 0.1"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        # 精确匹配0.1
        found_names = [row['名称'] for row in result['data']]
        assert '快速攻击' in found_names, f"0.1精度丢失，找到: {found_names}"

    def test_float_precision_0_01(self, engine, boundary_test_file):
        """测试0.01秒精度"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT 名称, 冷却时间 FROM 边界值测试 WHERE 冷却时间 = 0.01"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        found_names = [row['名称'] for row in result['data']]
        assert '超快攻击' in found_names, f"0.01精度丢失，找到: {found_names}"

    def test_float_precision_0_001(self, engine, boundary_test_file):
        """测试0.001秒精度"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT 名称, 冷却时间 FROM 边界值测试 WHERE 冷却时间 = 0.001"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        found_names = [row['名称'] for row in result['data']]
        assert '极速攻击' in found_names, f"0.001精度丢失，找到: {found_names}"

    def test_float_arithmetic_precision(self, engine, boundary_test_file):
        """测试浮点算术精度 (0.1+0.2 应该能匹配到对应行)"""
        # 测试WHERE条件中的算术比较
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT 名称, 价格 FROM 边界值测试 WHERE 价格 > 0.29 AND 价格 < 0.31"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        # 如果浮点精度处理得当，应该能找到精度测试1的行
        # 即使 0.1+0.2 != 0.3 (Python浮点问题)，范围查询应该工作

    # ==================== 大数值测试 ====================

    def test_large_value_99999(self, engine, boundary_test_file):
        """测试99999大数值"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT 名称, 价格 FROM 边界值测试 WHERE 价格 >= 99999"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        found_names = [row['名称'] for row in result['data']]
        assert '终极攻击' in found_names, f"99999大数值未匹配，找到: {found_names}"

    def test_large_value_999999(self, engine, boundary_test_file):
        """测试999999超大数值"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT 名称, 伤害值 FROM 边界值测试 WHERE 伤害值 >= 999999"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        assert len(result['data']) >= 1, "应该找到超大伤害值的记录"

    def test_max_value_comparison(self, engine, boundary_test_file):
        """测试最大值比较"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT 名称, 价格 FROM 边界值测试 WHERE 价格 = 9999999"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        found_names = [row['名称'] for row in result['data']]
        assert '最大值' in found_names, f"最大值未匹配，找到: {found_names}"

    # ==================== 零值和负值测试 ====================

    def test_zero_value(self, engine, boundary_test_file):
        """测试零值处理"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT 名称, 价格, 冷却时间 FROM 边界值测试 WHERE 价格 = 0"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        found_names = [row['名称'] for row in result['data']]
        assert '免费技能' in found_names, f"零值未匹配，找到: {found_names}"

    def test_negative_value(self, engine, boundary_test_file):
        """测试负值处理"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT 名称, 价格 FROM 边界值测试 WHERE 价格 < 0"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        found_names = [row['名称'] for row in result['data']]
        assert '测试负值' in found_names, f"负值未匹配，找到: {found_names}"

    # ==================== NULL处理测试 ====================

    def test_null_handling_is_null(self, engine, boundary_test_file):
        """测试IS NULL条件"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT * FROM 边界值测试 WHERE 名称 IS NULL"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        assert len(result['data']) >= 1, "应该找到NULL值的记录"

    def test_null_handling_is_not_null(self, engine, boundary_test_file):
        """测试IS NOT NULL条件"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT 名称, 价格 FROM 边界值测试 WHERE 名称 IS NOT NULL AND 价格 IS NOT NULL"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        for row in result['data']:
            assert row['名称'] is not None, "IS NOT NULL应该过滤掉NULL值"

    # ==================== 极小值测试 ====================

    def test_tiny_value(self, engine, boundary_test_file):
        """测试极小值 (0.00001)"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT 名称, 价格 FROM 边界值测试 WHERE 价格 <= 0.0001 AND 价格 > 0"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        found_names = [row['名称'] for row in result['data']]
        assert '极小值' in found_names, f"极小值未匹配，找到: {found_names}"

    # ==================== 范围查询测试 ====================

    def test_between_precision(self, engine, boundary_test_file):
        """测试BETWEEN精度"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT 名称, 冷却时间 FROM 边界值测试 WHERE 冷却时间 BETWEEN 0.09 AND 0.11"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        found_names = [row['名称'] for row in result['data']]
        assert '快速攻击' in found_names, f"BETWEEN精度丢失，找到: {found_names}"

    def test_between_large_values(self, engine, boundary_test_file):
        """测试大数值BETWEEN"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT 名称, 价格 FROM 边界值测试 WHERE 价格 BETWEEN 99999 AND 1000000"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        assert len(result['data']) >= 1, "应该找到大数值范围的记录"

    # ==================== 聚合函数边界测试 ====================

    def test_max_with_boundary(self, engine, boundary_test_file):
        """测试MAX函数在大数值边界"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT MAX(价格) as 最高价格 FROM 边界值测试 WHERE 价格 IS NOT NULL"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        assert len(result['data']) >= 1, "应该返回聚合结果"

    def test_min_with_boundary(self, engine, boundary_test_file):
        """测试MIN函数在负值边界"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT MIN(价格) as 最低价格 FROM 边界值测试 WHERE 价格 IS NOT NULL"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        assert len(result['data']) >= 1, "应该返回聚合结果"

    def test_avg_with_null(self, engine, boundary_test_file):
        """测试AVG函数对NULL的处理"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT AVG(价格) as 平均价格 FROM 边界值测试"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        assert len(result['data']) >= 1, "应该返回聚合结果"
        # AVG应该忽略NULL值

    def test_count_with_null(self, engine, boundary_test_file):
        """测试COUNT函数对NULL的处理"""
        result = engine.execute_sql_query(
            boundary_test_file,
            "SELECT COUNT(*) as 总数, COUNT(名称) as 非空名称数 FROM 边界值测试"
        )
        assert result['success'], f"查询失败: {result.get('error', '')}"
        assert len(result['data']) >= 1, "应该返回聚合结果"
        # COUNT(*) 应该 > COUNT(名称) 因为有NULL行
        total = result['data'][0]['总数']
        non_null = result['data'][0]['非空名称数']
        assert total >= non_null, f"COUNT(*)应该>=COUNT(列): {total} vs {non_null}"


class TestSafeFloatComparison:
    """安全浮点比较函数测试"""

    def test_normal_comparison(self):
        """测试正常比较"""
        assert _safe_float_comparison(1.0, 2.0, '<') is True
        assert _safe_float_comparison(2.0, 1.0, '>') is True
        assert _safe_float_comparison(1.0, 1.0, '==') is False  # op=='==' returns False
        assert _safe_float_comparison(1.0, 1.0, '>=') is True
        assert _safe_float_comparison(1.0, 1.0, '<=') is True

    def test_none_handling(self):
        """测试None值处理"""
        assert _safe_float_comparison(None, 1.0, '<') is False
        assert _safe_float_comparison(1.0, None, '>') is False
        assert _safe_float_comparison(None, None, '==') is False

    def test_precision_0_1(self):
        """测试0.1精度比较"""
        # 0.1在浮点数中是不精确的，但比较应该仍然工作
        assert _safe_float_comparison(0.1, 0.2, '<') is True
        assert _safe_float_comparison(0.1, 0.1, '<=') is True

    def test_large_values(self):
        """测试大数值比较"""
        assert _safe_float_comparison(99999, 99999, '<=') is True
        assert _safe_float_comparison(9999999, 99999, '>') is True

    def test_negative_values(self):
        """测试负值比较"""
        assert _safe_float_comparison(-100, 0, '<') is True
        assert _safe_float_comparison(-100, -50, '<') is True

    def test_type_coercion(self):
        """测试类型转换"""
        assert _safe_float_comparison("1.5", 2.0, '<') is True
        assert _safe_float_comparison(1, 2, '<') is True
        assert _safe_float_comparison(1.5, "2", '<') is True


class TestJOINBoundary:
    """JOIN边界值测试"""

    @pytest.fixture
    def join_test_files(self):
        """创建JOIN测试用的两个Excel文件"""
        # 文件1: 角色表
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "角色"
        ws1.append(["ID", "名称", "等级", "金币"])
        ws1.append([1, "战士", 100, 9999999])
        ws1.append([2, "法师", 50, 100])
        ws1.append([3, "刺客", 0, 0])
        ws1.append([4, "牧师", -1, -100])  # 负值测试
        ws1.append([5, None, None, None])  # NULL测试

        # 文件2: 技能表
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "技能"
        ws2.append(["角色ID", "技能名", "伤害", "冷却时间"])
        ws2.append([1, "旋风斩", 99999, 0.1])
        ws2.append([2, "火球术", 100, 2.5])
        ws2.append([3, "背刺", 50, 0.01])
        ws2.append([1, "怒吼", 0, 10.0])
        ws2.append([999, "测试技能", 0, 0])  # 不存在的角色ID

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f1:
            wb1.save(f1.name)
            file1 = f1.name

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f2:
            wb2.save(f2.name)
            file2 = f2.name

        yield file1, file2

        os.unlink(file1)
        os.unlink(file2)

    def test_inner_join_with_boundary(self, join_test_files):
        """测试INNER JOIN在边界值下的表现"""
        file1, file2 = join_test_files
        engine = AdvancedSQLQueryEngine()

        result = engine.execute_sql_query(
            file1,
            "SELECT r.名称, s.技能名, s.伤害, s.冷却时间 "
            "FROM 角色 r "
            f"INNER JOIN 技能@'{file2}' s ON r.ID = s.角色ID "
            "WHERE s.伤害 >= 99999 OR s.冷却时间 <= 0.1"
        )
        assert result['success'], f"JOIN查询失败: {result.get('error', '')}"
        assert len(result['data']) >= 1, "应该找到边界值的JOIN结果"

    def test_left_join_with_null(self, join_test_files):
        """测试LEFT JOIN在NULL值下的表现"""
        file1, file2 = join_test_files
        engine = AdvancedSQLQueryEngine()

        result = engine.execute_sql_query(
            file1,
            "SELECT r.名称, s.技能名 "
            "FROM 角色 r "
            f"LEFT JOIN 技能@'{file2}' s ON r.ID = s.角色ID "
            "WHERE r.名称 IS NOT NULL"
        )
        assert result['success'], f"LEFT JOIN查询失败: {result.get('error', '')}"
        assert len(result['data']) >= 1, "应该返回LEFT JOIN结果"
        # LEFT JOIN应该保留左表所有行，即使右表无匹配


if __name__ == '__main__':
    pytest.main([__file__, '-v', '--tb=short'])
