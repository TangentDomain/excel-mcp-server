"""
Tests for ExcelManager class
"""

import pytest
import tempfile
from pathlib import Path
from openpyxl import load_workbook
from src.core.excel_manager import ExcelManager
from src.models.types import OperationResult
from src.utils.exceptions import SheetNotFoundError, DataValidationError


class TestExcelManager:
    """Test cases for ExcelManager class"""
    
    def test_create_file_default_sheet(self, temp_dir):
        """Test creating file with default sheet"""
        file_path = temp_dir / "test_default.xlsx"
        result = ExcelManager.create_file(str(file_path))
        
        assert result.success is True
        assert result.data.file_path == str(file_path)
        assert len(result.data.sheets) == 1
        assert "Sheet" in result.data.sheets[0]
        
        # Verify file exists and is valid
        assert file_path.exists()
        wb = load_workbook(file_path)
        assert len(wb.sheetnames) == 1
    
    def test_create_file_custom_sheets(self, temp_dir):
        """Test creating file with custom sheets"""
        file_path = temp_dir / "test_custom.xlsx"
        sheet_names = ["数据", "图表", "汇总"]
        result = ExcelManager.create_file(str(file_path), sheet_names)
        
        assert result.success is True
        assert result.data.file_path == str(file_path)
        assert len(result.data.sheets) == 3
        assert "数据" in result.data.sheets
        assert "图表" in result.data.sheets
        assert "汇总" in result.data.sheets
        
        # Verify file exists and has correct sheets
        assert file_path.exists()
        wb = load_workbook(file_path)
        assert len(wb.sheetnames) == 3
        assert "数据" in wb.sheetnames
        assert "图表" in wb.sheetnames
        assert "汇总" in wb.sheetnames
    
    def test_create_file_invalid_extension(self, temp_dir):
        """Test creating file with invalid extension"""
        file_path = temp_dir / "test.txt"
        result = ExcelManager.create_file(str(file_path))
        
        assert result.success is False
        assert "扩展名" in result.error
    
    def test_create_file_overwrite_existing(self, temp_dir):
        """Test creating file overwrites existing"""
        file_path = temp_dir / "test_overwrite.xlsx"
        
        # Create initial file
        wb = load_workbook()
        ws = wb.active
        ws.append(["原始数据"])
        wb.save(file_path)
        
        # Overwrite with new file
        sheet_names = ["新工作表"]
        result = ExcelManager.create_file(str(file_path), sheet_names)
        
        assert result.success is True
        assert len(result.data.sheets) == 1
        assert result.data.sheets[0] == "新工作表"
        
        # Verify file was overwritten
        wb = load_workbook(file_path)
        assert len(wb.sheetnames) == 1
        assert "新工作表" in wb.sheetnames
    
    def test_create_sheet(self, sample_excel_file):
        """Test creating a new sheet"""
        manager = ExcelManager(sample_excel_file)
        result = manager.create_sheet("新工作表")
        
        assert result.success is True
        assert result.data.sheet_name == "新工作表"
        assert result.data.total_sheets == 3  # Original 2 + 1 new
        
        # Verify sheet was created
        wb = load_workbook(sample_excel_file)
        assert "新工作表" in wb.sheetnames
    
    def test_create_sheet_at_position(self, sample_excel_file):
        """Test creating sheet at specific position"""
        manager = ExcelManager(sample_excel_file)
        result = manager.create_sheet("首页", 0)  # Create at position 0
        
        assert result.success is True
        assert result.data.sheet_name == "首页"
        
        # Verify sheet is at correct position
        wb = load_workbook(sample_excel_file)
        assert wb.sheetnames[0] == "首页"
    
    def test_create_sheet_duplicate_name(self, sample_excel_file):
        """Test creating sheet with duplicate name"""
        manager = ExcelManager(sample_excel_file)
        result = manager.create_sheet("Sheet1")  # Already exists
        
        assert result.success is False
        assert "已存在" in result.error
    
    def test_create_sheet_invalid_name(self, sample_excel_file):
        """Test creating sheet with invalid name"""
        manager = ExcelManager(sample_excel_file)
        result = manager.create_sheet("")  # Empty name
        
        assert result.success is False
        assert "名称" in result.error
    
    def test_delete_sheet(self, sample_excel_file):
        """Test deleting a sheet"""
        manager = ExcelManager(sample_excel_file)
        result = manager.delete_sheet("Sheet2")
        
        assert result.success is True
        assert result.data.deleted_sheet == "Sheet2"
        assert len(result.data.remaining_sheets) == 1
        assert "Sheet1" in result.data.remaining_sheets
        
        # Verify sheet was deleted
        wb = load_workbook(sample_excel_file)
        assert "Sheet2" not in wb.sheetnames
    
    def test_delete_sheet_nonexistent(self, sample_excel_file):
        """Test deleting non-existent sheet"""
        manager = ExcelManager(sample_excel_file)
        result = manager.delete_sheet("NonExistentSheet")
        
        assert result.success is False
        assert "工作表" in result.error
    
    def test_delete_last_sheet(self, temp_dir):
        """Test deleting the last sheet"""
        file_path = temp_dir / "test_single_sheet.xlsx"
        
        # Create file with single sheet
        wb = load_workbook()
        wb.save(file_path)
        
        manager = ExcelManager(str(file_path))
        result = manager.delete_sheet("Sheet")
        
        assert result.success is False
        assert "最后一个" in result.error
    
    def test_rename_sheet(self, sample_excel_file):
        """Test renaming a sheet"""
        manager = ExcelManager(sample_excel_file)
        result = manager.rename_sheet("Sheet1", "数据表")
        
        assert result.success is True
        assert result.data.old_name == "Sheet1"
        assert result.data.new_name == "数据表"
        
        # Verify sheet was renamed
        wb = load_workbook(sample_excel_file)
        assert "数据表" in wb.sheetnames
        assert "Sheet1" not in wb.sheetnames
    
    def test_rename_sheet_nonexistent(self, sample_excel_file):
        """Test renaming non-existent sheet"""
        manager = ExcelManager(sample_excel_file)
        result = manager.rename_sheet("NonExistentSheet", "新名称")
        
        assert result.success is False
        assert "工作表" in result.error
    
    def test_rename_sheet_duplicate_name(self, sample_excel_file):
        """Test renaming sheet to existing name"""
        manager = ExcelManager(sample_excel_file)
        result = manager.rename_sheet("Sheet1", "Sheet2")  # Sheet2 already exists
        
        assert result.success is False
        assert "已存在" in result.error
    
    def test_rename_sheet_invalid_new_name(self, sample_excel_file):
        """Test renaming sheet with invalid new name"""
        manager = ExcelManager(sample_excel_file)
        result = manager.rename_sheet("Sheet1", "")  # Empty name
        
        assert result.success is False
        assert "名称" in result.error
    
    def test_manager_init_with_valid_file(self, sample_excel_file):
        """Test ExcelManager initialization with valid file"""
        manager = ExcelManager(sample_excel_file)
        assert manager.file_path == sample_excel_file
    
    def test_manager_init_with_invalid_file(self):
        """Test ExcelManager initialization with invalid file"""
        with pytest.raises(FileNotFoundError):
            ExcelManager("nonexistent_file.xlsx")
    
    def test_create_sheet_special_characters(self, sample_excel_file):
        """Test creating sheet with special characters"""
        manager = ExcelManager(sample_excel_file)
        result = manager.create_sheet("数据_2024")
        
        assert result.success is True
        assert result.data.sheet_name == "数据_2024"
        
        # Verify sheet was created
        wb = load_workbook(sample_excel_file)
        assert "数据_2024" in wb.sheetnames
    
    def test_create_sheet_long_name(self, sample_excel_file):
        """Test creating sheet with long name"""
        manager = ExcelManager(sample_excel_file)
        long_name = "这是一个非常长的工作表名称测试" * 2
        result = manager.create_sheet(long_name)
        
        # Excel has sheet name length limits, should handle gracefully
        assert isinstance(result, OperationResult)
    
    def test_delete_sheet_with_content(self, sample_excel_file):
        """Test deleting sheet that contains data"""
        manager = ExcelManager(sample_excel_file)
        result = manager.delete_sheet("Sheet1")  # Contains data
        
        assert result.success is True
        assert result.data.deleted_sheet == "Sheet1"
        
        # Verify sheet was deleted and file is still valid
        wb = load_workbook(sample_excel_file)
        assert "Sheet1" not in wb.sheetnames
        assert len(wb.sheetnames) == 1  # Only Sheet2 remains