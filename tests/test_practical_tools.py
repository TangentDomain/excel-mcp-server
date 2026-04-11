"""Tests for new practical tools (batch ops, charts, data validation, conditional formatting)."""
import os
import pytest
from openpyxl import Workbook

def _create_workbook(path, sheet_title, headers, rows):
    """Create a test workbook and save to path."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for row in rows:
        ws.append(row)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path

@pytest.fixture
def workbook(tmp_path):
    """Create a test workbook with sample game data."""
    path = str(tmp_path / "practical_tools_test.xlsx")
    _create_workbook(path, "技能配置",
                     ["skill_id", "skill_name", "skill_type", "damage", "cooldown"],
                     [
                         [1001, "火球术", "法师", 150, 3.0],
                         [1002, "冰冻术", "法师", 120, 4.0],
                         [1003, "斩击", "战士", 200, 2.0],
                         [1004, "治疗术", "牧师", 80, 5.0],
                     ])
    return path


class TestBatchOperations:
    """Tests for batch operations."""

    def test_merge_multiple_files_append(self, tmp_path):
        """Merge two files with append mode."""
        from src.excel_mcp_server_fastmcp.server import excel_merge_multiple_files

        file1 = str(tmp_path / "source1.xlsx")
        _create_workbook(file1, "技能表", ["id", "name"], [[1, "火球术"], [2, "冰冻术"]])

        file2 = str(tmp_path / "source2.xlsx")
        _create_workbook(file2, "装备表", ["id", "name"], [[1, "长剑"], [2, "法杖"]])

        target = str(tmp_path / "merged.xlsx")
        result = excel_merge_multiple_files([file1, file2], target, "append")

        assert result["success"] is True
        assert result["data"]["source_files_count"] == 2
        assert result["data"]["merged_sheets_count"] == 2

        from openpyxl import load_workbook
        wb = load_workbook(target)
        assert "技能表" in wb.sheetnames
        assert "装备表" in wb.sheetnames


class TestDataValidation:
    """Tests for data validation set / clear."""

    def test_set_list_validation(self, workbook):
        """Set a list-type data validation."""
        from src.excel_mcp_server_fastmcp.server import excel_set_data_validation

        result = excel_set_data_validation(
            workbook, "技能配置", "C2:C10", "list",
            "法师,战士,牧师,刺客", "选择职业", "请选择正确的职业类型"
        )
        assert result["success"] is True
        assert result["data"]["validation_type"] == "list"

    def test_set_validation_nonexistent_sheet(self, workbook):
        """Reject non-existent sheet."""
        from src.excel_mcp_server_fastmcp.server import excel_set_data_validation

        result = excel_set_data_validation(
            workbook, "不存在", "A1:A10", "list", "test"
        )
        assert result["success"] is False

    def test_clear_validation(self, workbook):
        """Clear all validations on a sheet."""
        from src.excel_mcp_server_fastmcp.server import (
            excel_set_data_validation,
            excel_clear_validation,
        )

        excel_set_data_validation(
            workbook, "技能配置", "C2:C10", "list", "法师,战士,牧师"
        )

        result = excel_clear_validation(workbook, "技能配置")
        assert result["success"] is True
        assert result["data"]["cleared_count"] >= 1

    def test_set_whole_number_validation(self, workbook):
        """Set a whole number data validation."""
        from src.excel_mcp_server_fastmcp.server import excel_set_data_validation

        result = excel_set_data_validation(
            workbook, "技能配置", "E2:E10", "whole_number",
            "between,1,10", "冷却时间", "冷却时间应在1-10秒之间"
        )
        assert result["success"] is True
        assert result["data"]["validation_type"] == "whole_number"
        assert result["data"]["input_title"] == "冷却时间"

    def test_set_decimal_validation(self, workbook):
        """Set a decimal data validation."""
        from src.excel_mcp_server_fastmcp.server import excel_set_data_validation

        result = excel_set_data_validation(
            workbook, "技能配置", "E2:E10", "decimal",
            "greaterThan,0", "伤害倍率", "伤害倍率必须大于0"
        )
        assert result["success"] is True
        assert result["data"]["validation_type"] == "decimal"

    def test_set_date_validation(self, workbook):
        """Set a date data validation."""
        from src.excel_mcp_server_fastmcp.server import excel_set_data_validation

        result = excel_set_data_validation(
            workbook, "技能配置", "F2:F10", "date",
            "greaterThanOrEqual,2024-01-01", "发布日期", "发布日期不能早于2024-01-01"
        )
        assert result["success"] is True
        assert result["data"]["validation_type"] == "date"

    def test_set_text_length_validation(self, workbook):
        """Set a text length data validation."""
        from src.excel_mcp_server_fastmcp.server import excel_set_data_validation

        result = excel_set_data_validation(
            workbook, "技能配置", "B2:B10", "text_length",
            "between,2,10", "技能名称", "技能名称长度应在2-10字符之间"
        )
        assert result["success"] is True
        assert result["data"]["validation_type"] == "text_length"

    def test_set_custom_validation(self, workbook):
        """Set a custom formula data validation."""
        from src.excel_mcp_server_fastmcp.server import excel_set_data_validation

        result = excel_set_data_validation(
            workbook, "技能配置", "D2:D10", "custom",
            "=AND(D2>=50,D2<=300)", "伤害值", "伤害值应在50-300之间"
        )
        assert result["success"] is True
        assert result["data"]["validation_type"] == "custom"

    def test_validation_operators(self, workbook):
        """Test various validation operators."""
        from src.excel_mcp_server_fastmcp.server import excel_set_data_validation

        operators_tests = [
            ("equal,5", "equal"),
            ("notEqual,0", "notEqual"),
            ("greaterThan,100", "greaterThan"),
            ("lessThan,1000", "lessThan"),
            ("greaterThanOrEqual,10", "greaterThanOrEqual"),
            ("lessThanOrEqual,100", "lessThanOrEqual"),
            ("between,1,100", "between"),
            ("notBetween,100,1000", "notBetween"),
        ]

        for criteria, expected_op in operators_tests:
            result = excel_set_data_validation(
                workbook, "技能配置", "E2:E10", "whole_number",
                criteria, "测试", "测试验证"
            )
            assert result["success"] is True

    def test_invalid_validation_type(self, workbook):
        """Reject invalid validation types."""
        from src.excel_mcp_server_fastmcp.server import excel_set_data_validation

        result = excel_set_data_validation(
            workbook, "技能配置", "A1:A10", "invalid_type", "test"
        )
        assert result["success"] is False
        assert "不支持的验证类型" in result["message"]

    def test_invalid_criteria_format(self, workbook):
        """Reject invalid criteria format."""
        from src.excel_mcp_server_fastmcp.server import excel_set_data_validation

        result = excel_set_data_validation(
            workbook, "技能配置", "A1:A10", "whole_number", "invalid_format"
        )
        assert result["success"] is False
        assert "验证条件格式错误" in result["message"]

    def test_clear_validation_by_range(self, workbook):
        """Clear validations for a specific range."""
        from src.excel_mcp_server_fastmcp.server import (
            excel_set_data_validation,
            excel_clear_validation,
        )

        excel_set_data_validation(
            workbook, "技能配置", "C2:C5", "list", "法师,战士"
        )
        excel_set_data_validation(
            workbook, "技能配置", "E2:E5", "whole_number", "between,1,10"
        )

        result = excel_clear_validation(workbook, "技能配置", "C2:C5")
        assert result["success"] is True
        assert result["data"]["cleared_count"] >= 1

class TestConditionalFormatting:
    """Tests for conditional formatting add / clear."""

    def test_add_cell_value_format(self, workbook):
        """Add a cell-value conditional format."""
        from src.excel_mcp_server_fastmcp.server import excel_add_conditional_format

        result = excel_add_conditional_format(
            workbook, "技能配置", "D2:D5", "cellValue",
            ">=150", "lightGreen"
        )
        assert result["success"] is True
        assert result["data"]["format_type"] == "cellValue"

    def test_add_format_nonexistent_sheet(self, workbook):
        """Reject non-existent sheet."""
        from src.excel_mcp_server_fastmcp.server import excel_add_conditional_format

        result = excel_add_conditional_format(
            workbook, "不存在", "A1:A10", "cellValue", ">=0"
        )
        assert result["success"] is False

    def test_clear_conditional_format(self, workbook):
        """Clear conditional formatting on a sheet."""
        from src.excel_mcp_server_fastmcp.server import (
            excel_add_conditional_format,
            excel_clear_conditional_format,
        )

        excel_add_conditional_format(
            workbook, "技能配置", "D2:D5", "cellValue", ">=150"
        )

        result = excel_clear_conditional_format(workbook, "技能配置")
        assert result["success"] is True
        assert result["data"]["cleared_count"] >= 1