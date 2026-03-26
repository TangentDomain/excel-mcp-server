# -*- coding: utf-8 -*-
"""
Excel Operations 错误处理和边界情况测试套件

覆盖 excel_operations.py 中的错误处理路径
"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations


class TestExcelOperationsErrors:
    """ExcelOperations 错误处理测试"""

    @pytest.fixture
    def valid_test_file(self, temp_dir):
        """创建有效测试文件"""
        file_path = temp_dir / "valid_test.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws['A1'] = "ID"
        ws['B1'] = "Name"
        ws['A2'] = 1
        ws['B2'] = "Test"
        
        wb.save(file_path)
        return str(file_path)

    # ==================== 文件不存在错误 ====================

    def test_check_duplicate_ids_file_not_found(self):
        """测试文件不存在时的错误处理"""
        result = ExcelOperations.check_duplicate_ids(
            file_path="/nonexistent/file.xlsx",
            sheet_name="Sheet1",
            id_column=1
        )
        
        assert result['success'] is False
        assert '不存在' in result.get('message', '')

    def test_compare_files_file_not_found(self):
        """测试比较不存在的文件"""
        result = ExcelOperations.compare_files(
            file1_path="/nonexistent/file1.xlsx",
            file2_path="/nonexistent/file2.xlsx"
        )
        
        assert result['success'] is False

    # ==================== 工作表不存在错误 ====================

    def test_check_duplicate_ids_sheet_not_found(self, valid_test_file):
        """测试工作表不存在时的错误处理"""
        result = ExcelOperations.check_duplicate_ids(
            file_path=valid_test_file,
            sheet_name="NonExistentSheet",
            id_column=1
        )
        
        assert result['success'] is False
        assert '工作表不存在' in result.get('message', '')

    # ==================== 空参数测试 ====================

    def test_check_duplicate_ids_empty_file_path(self):
        """测试空文件路径"""
        result = ExcelOperations.check_duplicate_ids(
            file_path="",
            sheet_name="Sheet1",
            id_column=1
        )
        
        assert result['success'] is False

    def test_check_duplicate_ids_empty_sheet_name(self, valid_test_file):
        """测试空工作表名"""
        result = ExcelOperations.check_duplicate_ids(
            file_path=valid_test_file,
            sheet_name="",
            id_column=1
        )
        
        assert result['success'] is False

    # ==================== find_last_row 边界测试 ====================

    def test_find_last_row_empty_sheet(self, valid_test_file):
        """测试空工作表"""
        # 创建空工作表
        wb = Workbook()
        ws = wb.create_sheet("EmptySheet")
        file_path = valid_test_file
        wb.save(file_path)
        
        result = ExcelOperations.find_last_row(
            file_path=file_path,
            sheet_name="EmptySheet"
        )
        
        assert result['success'] is True

    def test_find_last_row_with_column(self, valid_test_file):
        """测试指定列查找最后一行"""
        result = ExcelOperations.find_last_row(
            file_path=valid_test_file,
            sheet_name="Sheet1",
            column="A"
        )
        
        assert result['success'] is True

    def test_find_last_row_with_column_index(self, valid_test_file):
        """测试指定列索引查找最后一行"""
        result = ExcelOperations.find_last_row(
            file_path=valid_test_file,
            sheet_name="Sheet1",
            column=1
        )
        
        assert result['success'] is True

    def test_find_last_row_nonexistent_sheet(self, valid_test_file):
        """测试不存在的工作表"""
        result = ExcelOperations.find_last_row(
            file_path=valid_test_file,
            sheet_name="NonExistentSheet"
        )
        
        assert result['success'] is False


class TestExcelOperationsAdvanced:
    """ExcelOperations 高级功能测试"""

    @pytest.fixture
    def multi_sheet_file(self, temp_dir):
        """创建多工作表测试文件"""
        file_path = temp_dir / "multi_sheet.xlsx"
        
        wb = Workbook()
        
        # 第一个工作表
        ws1 = wb.active
        ws1.title = "Data"
        ws1['A1'] = "ID"
        ws1['A2'] = 1
        ws1['A3'] = 2
        ws1['A4'] = 3
        
        # 第二个工作表
        ws2 = wb.create_sheet("Config")
        ws2['A1'] = "Key"
        ws2['A2'] = "Value"
        
        wb.save(file_path)
        return str(file_path)

    def test_list_sheets(self, multi_sheet_file):
        """测试列出所有工作表"""
        result = ExcelOperations.list_sheets(multi_sheet_file)
        
        assert result['success'] is True
        assert 'Data' in result['sheets']
        assert 'Config' in result['sheets']

    def test_get_file_info(self, multi_sheet_file):
        """测试获取文件信息"""
        result = ExcelOperations.get_file_info(multi_sheet_file)
        
        assert result['success'] is True
        assert 'sheet_count' in result['data']

    def test_get_headers_with_dual_header(self, temp_dir):
        """测试获取双行表头"""
        file_path = temp_dir / "dual_header.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # 设置双行表头
        ws['A1'] = "描述"
        ws['A2'] = "字段"
        ws['B1'] = "描述2"
        ws['B2'] = "字段2"
        
        wb.save(file_path)
        
        result = ExcelOperations.get_headers(
            file_path=str(file_path),
            sheet_name="Sheet1",
            header_row=1
        )
        
        assert result['success'] is True


class TestExcelOperationsSearch:
    """ExcelOperations 搜索功能测试"""

    @pytest.fixture
    def search_file(self, temp_dir):
        """创建搜索测试文件"""
        file_path = temp_dir / "search_file.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "SearchSheet"
        
        # 添加测试数据
        ws['A1'] = "Header"
        ws['A2'] = "Apple"
        ws['A3'] = "Banana"
        ws['A4'] = "APPLE"
        ws['A5'] = "Cherry"
        
        # 添加数字
        ws['B1'] = "Number"
        ws['B2'] = 100
        ws['B3'] = 200
        ws['B4'] = 300
        
        wb.save(file_path)
        return str(file_path)

    def test_search_case_insensitive(self, search_file):
        """测试不区分大小写搜索"""
        result = ExcelOperations.search(
            file_path=search_file,
            pattern="apple",
            sheet_name="SearchSheet",
            case_sensitive=False
        )
        
        assert result['success'] is True

    def test_search_with_column_filter(self, search_file):
        """测试带列过滤的搜索"""
        result = ExcelOperations.search(
            file_path=search_file,
            pattern="Apple",
            sheet_name="SearchSheet"
        )
        
        assert result['success'] is True

    def test_search_range(self, search_file):
        """测试范围搜索"""
        result = ExcelOperations.search(
            file_path=search_file,
            pattern="100",
            sheet_name="SearchSheet",
            range="B1:B5"
        )
        
        assert result['success'] is True


class TestExcelOperationsRange:
    """ExcelOperations 范围操作测试"""

    @pytest.fixture
    def range_file(self, temp_dir):
        """创建范围测试文件"""
        file_path = temp_dir / "range_file.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "RangeSheet"
        
        # 添加测试数据
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=i)
            ws.cell(row=i, column=2, value=f"Data{i}")
        
        wb.save(file_path)
        return str(file_path)

    def test_get_range_partial(self, range_file):
        """测试部分范围读取"""
        result = ExcelOperations.get_range(
            file_path=range_file,
            range_expression="RangeSheet!A1:B5"
        )
        
        assert result['success'] is True
        assert len(result['data']) == 5

    def test_get_range_single_row(self, range_file):
        """测试单行范围"""
        result = ExcelOperations.get_range(
            file_path=range_file,
            range_expression="RangeSheet!1:1"
        )
        
        assert result['success'] is True

    def test_get_range_single_column(self, range_file):
        """测试单列范围"""
        result = ExcelOperations.get_range(
            file_path=range_file,
            range_expression="RangeSheet!A:A"
        )
        
        assert result['success'] is True


class TestExcelOperationsEdgeCases:
    """ExcelOperations 边界情况测试"""

    @pytest.fixture
    def edge_file(self, temp_dir):
        """创建边界测试文件"""
        file_path = temp_dir / "edge_file.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "EdgeSheet"
        
        # 包含特殊值
        ws['A1'] = ""
        ws['A2'] = None
        ws['A3'] = 0
        ws['A4'] = False
        ws['A5'] = "0"
        
        wb.save(file_path)
        return str(file_path)

    def test_search_empty_cell(self, edge_file):
        """测试搜索空单元格"""
        result = ExcelOperations.search(
            file_path=edge_file,
            pattern="",
            sheet_name="EdgeSheet"
        )
        
        assert result['success'] is True

    def test_search_none_value(self, edge_file):
        """测试搜索None值"""
        result = ExcelOperations.search(
            file_path=edge_file,
            pattern="None",
            sheet_name="EdgeSheet"
        )
        
        assert result['success'] is True

    def test_search_zero_value(self, edge_file):
        """测试搜索0值"""
        result = ExcelOperations.search(
            file_path=edge_file,
            pattern="0",
            sheet_name="EdgeSheet"
        )
        
        assert result['success'] is True
