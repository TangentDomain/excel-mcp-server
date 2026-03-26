"""Tests for copy_sheet and rename_column functionality."""
import os
import pytest
from openpyxl import Workbook


FIXTURE_PATH = os.path.join(os.path.dirname(__file__), 'test_data', 'copy_sheet_test.xlsx')


@pytest.fixture
def workbook():
    """Create a test workbook with sample data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "技能配置"

    # 双行表头
    ws.append(["技能名称", "技能类型", "伤害"])
    ws.append(["skill_name", "skill_type", "damage"])
    ws.append(["火球术", "法师", 150])
    ws.append(["冰冻术", "法师", 120])
    ws.append(["斩击", "战士", 200])

    # 第二个工作表
    ws2 = wb.create_sheet("装备配置")
    ws2.append(["装备名称", "品质"])
    ws2.append(["铁剑", "普通"])
    ws2.append(["魔杖", "稀有"])

    os.makedirs(os.path.dirname(FIXTURE_PATH), exist_ok=True)
    wb.save(FIXTURE_PATH)
    yield FIXTURE_PATH
    if os.path.exists(FIXTURE_PATH):
        os.remove(FIXTURE_PATH)


class TestCopySheet:
    """Tests for excel_copy_sheet via ExcelOperations."""

    def test_basic_copy(self, workbook):
        """Basic copy: source sheet duplicated with auto-generated name."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.copy_sheet(workbook, "技能配置")
        assert result['success'] is True
        assert "副本" in result.get('data', {}).get('name', '') or "副本" in result.get('message', '')

    def test_copy_with_custom_name(self, workbook):
        """Copy with custom new name."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.copy_sheet(workbook, "技能配置", "副本技能")
        assert result['success'] is True
        assert result['data']['name'] == "副本技能"

    def test_copy_preserves_data(self, workbook):
        """Copied sheet has same row count as source."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.copy_sheet(workbook, "技能配置", "数据验证")
        assert result['success'] is True
        # Source has 5 rows (2 header + 3 data)
        assert result['data']['max_row'] == 5

    def test_copy_nonexistent_sheet(self, workbook):
        """Error when source sheet doesn't exist."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.copy_sheet(workbook, "不存在的表")
        assert result['success'] is False
        assert "不存在" in result.get('message', '')

    def test_copy_duplicate_name_auto_suffix(self, workbook):
        """Auto-suffix when target name already exists."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        # First copy creates "技能配置_副本"
        r1 = ExcelOperations.copy_sheet(workbook, "技能配置")
        assert r1['success'] is True
        # Second copy should auto-suffix to avoid conflict
        r2 = ExcelOperations.copy_sheet(workbook, "技能配置")
        assert r2['success'] is True
        assert r2['data']['name'] != r1['data']['name']

    def test_copy_second_sheet(self, workbook):
        """Copy the second sheet."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.copy_sheet(workbook, "装备配置", "活动装备")
        assert result['success'] is True
        assert result['data']['name'] == "活动装备"


class TestRenameColumn:
    """Tests for excel_rename_column via ExcelOperations."""

    def test_basic_rename(self, workbook):
        """Basic column rename."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.rename_column(workbook, "技能配置", "技能名称", "技能名")
        assert result['success'] is True
        assert "技能名" in result['message']

    def test_rename_updates_cell(self, workbook):
        """Renamed value actually persisted in file."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        ExcelOperations.rename_column(workbook, "技能配置", "技能名称", "技能名")
        # Verify by reading the file
        from openpyxl import load_workbook
        wb = load_workbook(workbook)
        assert wb["技能配置"].cell(row=1, column=1).value == "技能名"
        wb.close()

    def test_rename_nonexistent_column(self, workbook):
        """Error when column name not found."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.rename_column(workbook, "技能配置", "不存在的列", "新列")
        assert result['success'] is False
        assert "未找到" in result.get('message', '')

    def test_rename_nonexistent_sheet(self, workbook):
        """Error when sheet doesn't exist."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.rename_column(workbook, "不存在的表", "技能名称", "技能名")
        assert result['success'] is False
        assert "不存在" in result.get('message', '')

    def test_rename_same_name(self, workbook):
        """Error when old and new names are the same."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.rename_column(workbook, "技能配置", "技能名称", "技能名称")
        assert result['success'] is False
        assert "相同" in result.get('message', '')

    def test_rename_header_row_2(self, workbook):
        """Rename column in row 2 (dual-header scenario, English field name)."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.rename_column(workbook, "技能配置", "skill_name", "name", header_row=2)
        assert result['success'] is True
        # Verify
        from openpyxl import load_workbook
        wb = load_workbook(workbook)
        assert wb["技能配置"].cell(row=2, column=1).value == "name"
        # Row 1 should be unchanged
        assert wb["技能配置"].cell(row=1, column=1).value == "技能名称"
        wb.close()

    def test_rename_empty_old_header(self, workbook):
        """Error when old_header is empty."""
        from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        result = ExcelOperations.rename_column(workbook, "技能配置", "", "新列")
        assert result['success'] is False
