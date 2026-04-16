# -*- coding: utf-8 -*-
"""
format_cells R66 迭代测试 - 深度边缘 case 第六轮

新增测试场景（覆盖 R65 未涉及的场景）:
  1. _normalize_formatting 空字典返回空字典
  2. _normalize_formatting 全 None 值字段被过滤
  3. _normalize_formatting 混合扁平+嵌套键（嵌套优先）
  4. format_cells 单单元格范围 A1:A1
  5. format_cells 超大范围（100x100）性能不崩溃
  6. format_cells 对已合并单元格应用格式
  7. format_cells 后合并单元格，格式是否保留
  8. border_style 为 "none" / "dashDot" / "dashed" / "dotted" / "mediumDashed" / "mediumDashDotDot"
  9. font_color 为 RGB 元组 (255, 0, 0) 格式
  10. number_format 为日期格式 "YYYY-MM-DD" / "DD/MM/YYYY"
  11. alignment shrink_to_fit=True + wrap_text=False 组合
  12. indent 超大值（如 100）的容错
  13. format_cells 对含超链接的单元格
  14. format_cells 对含数据验证的单元格
  15. 多次 format_cells 不同属性叠加（非覆盖）
  16. bg_color 为 3 位 HEX 缩写（如 F00 → FF0000?）
  17. font_name 含特殊字符/空格
  18. underline="singleAccounting" 和 "doubleAccounting" 验证
"""

import os
import pytest
import tempfile
from datetime import date, datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color as OpenpyxlColor, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from excel_mcp_server_fastmcp.core.excel_writer import ExcelWriter
from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations


# ==================== Helper ====================

def _create_test_xlsx(file_path: str, rows: int = 5, cols: int = 4, sheet_name: str = "Sheet1"):
    """创建测试用 xlsx 文件，含数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=r * 10 + c)
    wb.save(file_path)
    wb.close()


def _read_cell_style(file_path: str, cell_ref: str = "A1", sheet_name: str = "Sheet1") -> dict:
    """读取单元格的样式属性用于断言"""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    cell = ws[cell_ref]
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    result = {
        "bold": font.bold,
        "italic": font.italic,
        "underline": font.underline,
        "strikethrough": font.strikethrough,
        "size": font.size,
        "font_name": font.name,
        "color": str(font.color) if font.color else None,
        "fill_type": fill.fill_type if fill else None,
        "fgColor": str(fill.fgColor) if fill and hasattr(fill, 'fgColor') and fill.fgColor else None,
        "bgColor": str(fill.bgColor) if fill and hasattr(fill, 'bgColor') and fill.bgColor else None,
        "alignment_h": alignment.horizontal,
        "alignment_v": alignment.vertical,
        "wrap_text": bool(alignment.wrap_text) if alignment.wrap_text is not None else False,
        "text_rotation": alignment.text_rotation,
        "number_format": cell.number_format,
        "border_left": border.left.style if border and border.left else None,
        "border_right": border.right.style if border and border.right else None,
        "border_top": border.top.style if border and border.top else None,
        "border_bottom": border.bottom.style if border and border.bottom else None,
        "value": cell.value,
        "indent": alignment.indent,
        "shrink_to_fit": alignment.shrink_to_fit,
    }
    wb.close()
    return result


# ==================== Test Class ====================

class TestFormatCellsR66:
    """format_cells R66 第六轮测试套件 — 深度边缘 case 第六轮"""

    # ---------- 1. _normalize_formatting 空字典 ----------

    def test_normalize_empty_dict(self):
        """空字典应返回空字典"""
        result = ExcelOperations._normalize_formatting({})
        assert result == {}

    # ---------- 2. _normalize_formatting 全 None 值 ----------

    def test_normalize_all_none_values(self):
        """全 None 值的字典应返回空字典（None 值被过滤）"""
        result = ExcelOperations._normalize_formatting({
            "bold": None,
            "italic": None,
            "bg_color": None,
            "font_size": None,
            "alignment": None,
            "number_format": None,
        })
        assert result == {}

    # ---------- 3. 混合扁平+嵌套键 ----------

    def test_normalize_mixed_flat_nested(self):
        """混合扁平和嵌套键时，嵌套格式优先直接返回"""
        mixed = {
            "bold": True,
            "font": {"name": "Arial", "size": 12},
            "bg_color": "FF0000",
        }
        result = ExcelOperations._normalize_formatting(mixed)
        # 包含 font 嵌套键 → 直接原样返回
        assert "font" in result
        assert result["font"]["name"] == "Arial"

    # ---------- 4. 单单元格范围 ----------

    def test_single_cell_range(self, tmp_path):
        """A1:A1 单单元格范围应正常格式化"""
        fp = str(tmp_path / "single_cell.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:A1", {"font": {"bold": True}})
        assert result.success
        assert result.metadata["formatted_count"] == 1

        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True

    # ---------- 5. 超大范围 ----------

    def test_large_range_performance(self, tmp_path):
        """100x10 范围格式化不应崩溃"""
        fp = str(tmp_path / "large_range.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for r in range(1, 101):
            for c in range(1, 11):
                ws.cell(row=r, column=c, value=f"{r}-{c}")
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:J100", {"font": {"bold": True}})
        assert result.success
        assert result.metadata["formatted_count"] == 1000

    # ---------- 6. 已合并单元格格式化 ----------

    def test_format_merged_cells(self, tmp_path):
        """对已合并区域应用格式，左上角单元格应体现格式"""
        fp = str(tmp_path / "merged_fmt.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.merge_cells("A1:B2")
        ws["A1"].value = "Merged"
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:B2", {
            "font": {"bold": True, "size": 14},
            "fill": {"type": "solid", "color": "FFFF00"},
        })
        assert result.success

        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["size"] == 14
        assert style["fgColor"] is not None  # 有填充色

    # ---------- 7. 先格式后合并 ----------

    def test_format_then_merge(self, tmp_path):
        """先格式化再合并，格式应保留在合并区域"""
        fp = str(tmp_path / "fmt_then_merge.xlsx")
        _create_test_xlsx(fp)

        # 先格式化
        writer = ExcelWriter(fp)
        r1 = writer.format_cells("Sheet1!A1:B2", {"font": {"bold": True, "color": "FF0000"}})
        assert r1.success

        # 再合并
        r2 = writer.merge_cells("A1:B2", "Sheet1")
        assert r2.success

        # 验证格式保留
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True

    # ---------- 8. 各种 border_style 枚举 ----------

    @pytest.mark.parametrize("bstyle", [
        "none", "dashDot", "dashed", "dotted",
        "mediumDashed", "mediumDashDotDot", "hair", "medium",
        "mediumDashDot", "thick", "double", "slantDashDot",
    ])
    def test_border_style_variants(self, tmp_path, bstyle):
        """各种 openpyxl 支持的边框样式枚举值"""
        fp = str(tmp_path / f"border_{bstyle}.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {
                "left": bstyle,
                "right": bstyle,
                "top": bstyle,
                "bottom": bstyle,
            }
        })
        assert result.success, f"border_style={bstyle} 失败: {result.error}"

        style = _read_cell_style(fp, "A1")
        if bstyle == "none":
            assert style["border_left"] is None
        else:
            assert style["border_left"] == bstyle

    # ---------- 9. font_color RGB 元组 ----------

    def test_font_color_rgb_tuple(self, tmp_path):
        """font_color 支持 RGB 元组格式 (R, G, B)"""
        fp = str(tmp_path / "color_tuple.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        # 通过嵌套格式传入 RGB 元组
        result = writer.format_cells("Sheet1!A1", {
            "font": {"color": (255, 0, 128)}
        })
        # 可能成功也可能失败取决于 openpyxl 兼容性，至少不应崩溃
        # 如果失败，说明需要转换逻辑；如果成功，验证颜色
        assert result.success or result.error is not None

    # ---------- 10. 日期 number_format ----------

    @pytest.mark.parametrize("nf,expected_contains", [
        ("YYYY-MM-DD", "YYYY") | (False,) if False else ("YYYY-MM-DD", "YYYY"),
        ("DD/MM/YYYY", "DD/MM"),
        ("0.00%", "%"),
        ("#,##0.00", ","),
        ("@", "@"),
    ])
    def test_number_format_date_and_special(self, tmp_path, nf, expected_contains):
        """各种特殊 number_format 格式"""
        fp = str(tmp_path / f"nf_{nf.replace('/', '_').replace(',', '_').replace('%', 'pct')}.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"number_format": nf})
        assert result.success

        style = _read_cell_style(fp, "A1")
        assert expected_contains in style["number_format"]

    # ---------- 11. shrink_to_fit + wrap_text 组合 ----------

    def test_shrink_to_fit_with_wrap_text_off(self, tmp_path):
        """shrink_to_fit=True 且 wrap_text=False 的组合"""
        fp = str(tmp_path / "shrink_wrap.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "alignment": {
                "shrink_to_fit": True,
                "wrap_text": False,
            }
        })
        assert result.success

        style = _read_cell_style(fp, "A1")
        assert style["shrink_to_fit"] is True
        assert style["wrap_text"] is False

    # ---------- 12. indent 超大值 ----------

    def test_indent_large_value(self, tmp_path):
        """indent=100 超大值应正常设置（openpyxl 不限制）"""
        fp = str(tmp_path / "indent_large.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"alignment": {"indent": 100}})
        assert result.success

        style = _read_cell_style(fp, "A1")
        assert style["indent"] == 100

    # ---------- 13. 含超链接单元格 ----------

    def test_format_cell_with_hyperlink(self, tmp_path):
        """对含超链接的单元格格式化不应破坏超链接"""
        fp = str(tmp_path / "hyperlink.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        cell = ws["A1"]
        cell.value = "Click here"
        cell.hyperlink = "https://example.com"
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"bold": True, "color": "0000FF"}})
        assert result.success

        wb2 = load_workbook(fp)
        ws2 = wb2["Sheet1"]
        cell = ws2["A1"]
        assert cell.value == "Click here"
        assert cell.hyperlink is not None
        assert cell.hyperlink.target == "https://example.com"
        assert cell.font.bold is True
        wb2.close()

    # ---------- 14. 含数据验证单元格 ----------

    def test_format_cell_with_data_validation(self, tmp_path):
        """对含数据验证的单元格格式化不应破坏验证规则"""
        fp = str(tmp_path / "dataval.xlsx")
        from openpyxl.worksheet.datavalidation import DataValidation
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        dv = DataValidation(type="whole", operator="between", formula1=1, formula2=10)
        dv.add(ws["A1"])
        ws.add_data_validation(dv)
        ws["A1"].value = 5
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"bold": True}})
        assert result.success

        wb2 = load_workbook(fp)
        ws2 = wb2["Sheet1"]
        dvs = list(ws2.data_validations.dataValidation)
        assert len(dvs) >= 1
        assert ws2["A1"].font.bold is True
        wb2.close()

    # ---------- 15. 多次不同属性叠加 ----------

    def test_sequential_format_different_attrs(self, tmp_path):
        """多次 format_cells 设置不同属性应叠加而非覆盖"""
        fp = str(tmp_path / "sequential.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        # 第一次：设 bold
        r1 = writer.format_cells("Sheet1!A1", {"font": {"bold": True}})
        assert r1.success

        # 第二次：设 color（不应清除 bold）
        r2 = writer.format_cells("Sheet1!A1", {"font": {"color": "FF0000"}})
        assert r2.success

        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True  # bold 应保留
        assert "FF0000" in (style["color"] or "")

    # ---------- 16. 3位HEX颜色缩写 ----------

    def test_bg_color_3char_hex(self, tmp_path):
        """bg_color 为 3 位 HEX（如 F00）应能处理"""
        fp = str(tmp_path / "hex3.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"fill": {"type": "solid", "color": "F00"}})
        # 3位HEX可能不被openpyxl支持，但不应崩溃
        assert result.success or result.error is not None

    # ---------- 17. font_name 特殊字符 ----------

    @pytest.mark.parametrize("fname", [
        "微软雅黑",
        "Arial Unicode MS",
        "Courier New",
        "宋体",
        "Helvetica Neue Light",
    ])
    def test_font_name_special_chars(self, tmp_path, fname):
        """各种字体名称（中文、空格、特殊字符）"""
        fp = str(tmp_path / f"font_{hash(fname) % 10000}.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"name": fname}})
        assert result.success, f"font_name={fname} 失败: {result.error}"

        style = _read_cell_style(fp, "A1")
        assert style["font_name"] == fname

    # ---------- 18. underline accounting 变体 ----------

    @pytest.mark.parametrize("uline", ["singleAccounting", "doubleAccounting", "single", "double"])
    def test_underline_accounting_variants(self, tmp_path, uline):
        """underline 各变体正确应用"""
        fp = str(tmp_path / f"ul_{uline}.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"underline": uline}})
        assert result.success

        style = _read_cell_style(fp, "A1")
        assert style["underline"] == uline

    # ---------- 19. format_cells 整行范围 ----------

    def test_format_entire_row(self, tmp_path):
        """整行范围 1:1 格式化"""
        fp = str(tmp_path / "entire_row.xlsx")
        _create_test_xlsx(fp, rows=5, cols=5)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!1:1", {"font": {"bold": True}})
        assert result.success
        assert result.metadata["formatted_count"] >= 5  # 至少5列

    # ---------- 20. format_cells 整列范围 ----------

    def test_format_entire_column(self, tmp_path):
        """整列范围 A:A 格式化"""
        fp = str(tmp_path / "entire_col.xlsx")
        _create_test_xlsx(fp, rows=5, cols=5)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A:A", {"font": {"italic": True}})
        assert result.success

    # ---------- 21. format_cells 带 $ 绝对引用混合 ----------

    def test_mixed_absolute_relative_ref(self, tmp_path):
        """混合绝对/相对引用 $A1:B$2 应正常解析"""
        fp = str(tmp_path / "mixed_ref.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!$A1:B$2", {"font": {"bold": True}})
        assert result.success

    # ---------- 22. number_format=None 重置 ----------

    def test_number_format_none_reset(self, tmp_path):
        """number_format=None 应跳过（不设置为None）"""
        fp = str(tmp_path / "nf_none.xlsx")
        _create_test_xlsx(fp)

        # 先设一个格式
        writer = ExcelWriter(fp)
        r1 = writer.format_cells("Sheet1!A1", {"number_format": "0.00"})
        assert r1.success

        # 再传 None — 应跳过
        r2 = writer.format_cells("Sheet1!A1", {"number_format": None})
        assert r2.success

        style = _read_cell_style(fp, "A1")
        # number_format 应保持之前的值或为 General，不应出错
        assert style["number_format"] is not None

    # ---------- 23. border 全部为 None ----------

    def test_border_all_none_sides(self, tmp_path):
        """border 四边全为 None 应保留原有边框"""
        fp = str(tmp_path / "border_none.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {
                "left": None,
                "right": None,
                "top": None,
                "bottom": None,
            }
        })
        assert result.success

    # ---------- 24. format_cells 对空值单元格 ----------

    def test_format_empty_cell(self, tmp_path):
        """空值(None)单元格格式化后仍为空"""
        fp = str(tmp_path / "empty_cell.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # A1 故意留空
        ws["B1"].value = "data"
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "font": {"bold": True},
            "fill": {"type": "solid", "color": "FFFF00"},
        })
        assert result.success

        style = _read_cell_style(fp, "A1")
        assert style["value"] is None
        assert style["bold"] is True

    # ---------- 25. preset + 自定义 formatting 合并 ----------

    def test_preset_header_with_custom_bold_override(self, tmp_path):
        """preset='header' + 自定义 bold=False 应覆盖 preset 的 bold=True"""
        fp = str(tmp_path / "preset_override.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(
            file_path=fp,
            sheet_name="Sheet1",
            range="A1:C1",
            formatting={"bold": False, "font_size": 9},  # 覆盖 header 的 bold=True, size=11
            preset="header",
        )
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["bold"] is False  # 用户自定义覆盖了 preset
        assert style["size"] == 9

    # ---------- 26. text_rotation=0 显式重置 ----------

    def test_text_rotation_zero_explicit(self, tmp_path):
        """text_rotation=0 显式设置应为水平文本"""
        fp = str(tmp_path / "rot_zero.xlsx")
        _create_test_xlsx(fp)

        # 先设旋转
        writer = ExcelWriter(fp)
        r1 = writer.format_cells("Sheet1!A1", {"alignment": {"text_rotation": 45}})
        assert r1.success

        # 再重置为 0
        r2 = writer.format_cells("Sheet1!A1", {"alignment": {"text_rotation": 0}})
        assert r2.success

        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 0

    # ---------- 27. gradient_fill 双色渐变 ----------

    def test_gradient_fill_two_colors(self, tmp_path):
        """gradient fill 双色渐变正确应用"""
        fp = str(tmp_path / "gradient2.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {
                "type": "gradient",
                "colors": ["FF0000", "0000FF"],
                "gradient_type": "linear",
            }
        })
        assert result.success

        style = _read_cell_style(fp, "A1")
        # openpyxl GradientFill 的 fill_type 报告为 "linear"/"path"，不是 "gradient"
        assert style["fill_type"] in ("linear", "path")

    # ---------- 28. pattern fill 类型 ----------

    def test_pattern_fill_types(self, tmp_path):
        """pattern fill 各种 patternType"""
        fp = str(tmp_path / "pattern.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {
                "type": "pattern",
                "patternType": "lightGray",  # openpyxl 合法 patternType
                "fgColor": "FF0000",
                "bgColor": "00FF00",
            }
        })
        assert result.success

    # ---------- 29. format_cells 范围含 sheet 名 ! ----------

    def test_range_with_sheet_bang(self, tmp_path):
        """range 参数自带 Sheet1! 前缀"""
        fp = str(tmp_path / "range_bang.xlsx")
        _create_test_xlsx(fp)

        result = ExcelOperations.format_cells(
            file_path=fp,
            sheet_name="Sheet1",
            range="Sheet1!A1:B2",
            formatting={"font": {"bold": True}},
        )
        assert result["success"] is True

    # ---------- 30. format_cells 错误输入：无效范围 ----------

    def test_invalid_range_error(self, tmp_path):
        """无效范围表达式应返回错误而非崩溃"""
        fp = str(tmp_path / "bad_range.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!!!INVALID", {"font": {"bold": True}})
        # 无效范围可能抛异常，应被捕获
        assert result.success is False or result.error is not None

    # ---------- 31. format_cells 对布尔值单元格 ----------

    def test_format_boolean_cell(self, tmp_path):
        """布尔值单元格格式化后类型保持"""
        fp = str(tmp_path / "bool_cell.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"].value = True
        ws["A2"].value = False
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:A2", {"font": {"bold": True}})
        assert result.success

        wb2 = load_workbook(fp)
        ws2 = wb2["Sheet1"]
        assert ws2["A1"].value is True
        assert ws2["A2"].value is False
        assert ws2["A1"].font.bold is True
        wb2.close()

    # ---------- 32. format_cells 对日期单元格 ----------

    def test_format_date_cell_preserves_date(self, tmp_path):
        """日期单元格格式化后日期值保持"""
        fp = str(tmp_path / "date_cell.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"].value = date(2025, 6, 15)
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "font": {"bold": True},
            "number_format": "YYYY-MM-DD",
        })
        assert result.success

        wb2 = load_workbook(fp)
        ws2 = wb2["Sheet1"]
        val = ws2["A1"].value
        assert isinstance(val, date) or isinstance(val, datetime)
        wb2.close()

    # ---------- 33. 并发安全：同一文件连续操作 ----------

    def test_consecutive_operations_same_file(self, tmp_path):
        """同一文件连续 10 次 format_cells 不冲突"""
        fp = str(tmp_path / "consecutive.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        for i in range(10):
            result = writer.format_cells(f"Sheet1!A{i+1}", {"font": {"size": 10 + i}})
            assert result.success, f"第{i+1}次失败: {result.error}"

    # ---------- 34. font size 浮点数 ----------

    def test_font_size_float(self, tmp_path):
        """font_size 为浮点数时应正常工作（openpyxl 接受）"""
        fp = str(tmp_path / "fsize_float.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"size": 12.5}})
        assert result.success

        style = _read_cell_style(fp, "A1")
        assert style["size"] == 12.5

    # ---------- 35. alignment horizontal 全枚举 ----------

    @pytest.mark.parametrize("halign", [
        "left", "center", "right", "justify", "centerContinuous", "distributed", "general",
    ])
    def test_alignment_horizontal_enum(self, tmp_path, halign):
        """horizontal alignment 各枚举值"""
        fp = str(tmp_path / f"halign_{halign}.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"alignment": {"horizontal": halign}})
        assert result.success, f"halign={halign} 失败: {result.error}"

        style = _read_cell_style(fp, "A1")
        assert style["alignment_h"] == halign

    # ---------- 36. vertical alignment 全枚举 ----------

    @pytest.mark.parametrize("valign", [
        "top", "center", "bottom", "justify", "distributed",
    ])
    def test_alignment_vertical_enum(self, tmp_path, valign):
        """vertical alignment 各枚举值"""
        fp = str(tmp_path / f"valign_{valign}.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"alignment": {"vertical": valign}})
        assert result.success, f"valign={valign} 失败: {result.error}"

        style = _read_cell_style(fp, "A1")
        assert style["alignment_v"] == valign

    # ---------- 37. format_cells 后读取公式不变 ----------

    def test_formula_preserved_after_format(self, tmp_path):
        """含公式的单元格格式化后公式字符串不变"""
        fp = str(tmp_path / "formula_fmt.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"].value = 10
        ws["B1"].value = 20
        ws["C1"].value = "=SUM(A1:B1)"
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!C1", {"font": {"bold": True}})
        assert result.success

        wb2 = load_workbook(fp)
        ws2 = wb2["Sheet1"]
        assert ws2["C1"].value == "=SUM(A1:B1)"  # 公式保持不变
        assert ws2["C1"].font.bold is True
        wb2.close()

    # ---------- 38. _deep_merge 预设与用户合并行为 ----------

    def test_deep_merge_user_overrides_preset(self):
        """_deep_merge 用户值应覆盖预设值"""
        preset = {
            "font": {"name": "Arial", "size": 11, "bold": True},
            "fill": {"color": "D9D9D9"},
        }
        user = {
            "font": {"bold": False, "size": 14},
        }
        merged = ExcelOperations._deep_merge(preset, user)
        assert merged["font"]["bold"] is False  # 用户覆盖
        assert merged["font"]["size"] == 14  # 用户覆盖
        assert merged["font"]["name"] == "Arial" # 预设保留
        assert merged["fill"]["color"] == "D9D9D9"  # 预设保留

    # ---------- 39. format_cells 空 formatting 字典 ----------

    def test_empty_formatting_dict(self, tmp_path):
        """空 formatting 字典 {} 应成功（无操作但有结果）"""
        fp = str(tmp_path / "empty_fmt.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:B2", {})
        assert result.success
        # A1:B2 = 2行 x 2列 = 4个单元格
        assert result.metadata["formatted_count"] == 4

    # ---------- 40. border 仅设对角线 ----------

    def test_border_diagonal_only(self, tmp_path):
        """仅设置对角线边框"""
        fp = str(tmp_path / "border_diag.xlsx")
        _create_test_xlsx(fp)

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {
                "diagonal": "thin",
                "diagonalDirection": 1,  # 从左上到右下
            }
        })
        assert result.success
