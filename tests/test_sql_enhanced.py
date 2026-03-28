"""SQL引擎增强功能测试 — 子查询/CASE WHEN/COALESCE/EXISTS/LEFT JOIN NULL"""
import os
import sys
import pytest
import pandas as pd
import numpy as np


@pytest.fixture
def game_config():
    """游戏配置表路径"""
    return os.path.join(os.path.dirname(__file__), 'test_data', 'game_config.xlsx')


@pytest.fixture
def multi_sheet_config(tmp_path):
    """创建多工作表测试文件"""
    from openpyxl import Workbook
    wb = Workbook()

    # 技能表
    ws1 = wb.active
    ws1.title = '技能配置'
    ws1.append(['技能名称', 'skill_name', '技能类型', 'skill_type', '伤害', 'damage', '等级', 'level'])
    ws1.append(['火球术', 'fireball', '法师', 'mage', 200, 200, 5])
    ws1.append(['冰冻术', 'ice', '法师', 'mage', 150, 150, 3])
    ws1.append(['斩击', 'slash', '战士', 'warrior', 100, 100, 1])
    ws1.append(['治疗', 'heal', '辅助', 'support', 0, 0, 2])
    ws1.append(['毒雾', 'poison', '刺客', 'assassin', 180, 180, 4])

    # 装备表（用于JOIN和子查询）
    ws2 = wb.create_sheet('装备配置')
    ws2.append(['装备名称', 'equip_name', '适用职业', 'equip_class', '攻击力', 'atk'])
    ws2.append(['法杖', 'staff', '法师', 'mage', 50, 50])
    ws2.append(['剑', 'sword', '战士', 'warrior', 80, 80])
    ws2.append(['匕首', 'dagger', '刺客', 'assassin', 60, 60])
    ws2.append(['法袍', 'robe', '法师', 'mage', 10, 10])

    # 怪物表（用于EXISTS子查询）
    ws3 = wb.create_sheet('怪物配置')
    ws3.append(['怪物名称', 'monster_name', '怪物类型', 'monster_type', '等级', 'level'])
    ws3.append(['火龙', 'dragon', 'boss', 'boss', 10])
    ws3.append(['史莱姆', 'slime', '普通', 'normal', 1])
    ws3.append(['骷髅兵', 'skeleton', '普通', 'normal', 3])
    ws3.append(['暗影刺客', 'shadow', '精英', 'elite', 7])

    path = str(tmp_path / 'test_multi.xlsx')
    wb.save(path)
    return path


class TestCaseWhen:
    """CASE WHEN 表达式测试"""

    def test_case_when_basic(self, game_config):
        """基础CASE WHEN: 根据伤害分级"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            game_config,
            "SELECT skill_name, damage, CASE WHEN damage > 100 THEN '高伤' WHEN damage >= 50 THEN '中伤' ELSE '低伤' END AS level FROM 技能配置"
        )
        assert result['success'] is True
        data = result['data']
        # 跳过表头行
        rows = [r for r in data if r != data[0]]
        # 火球术200→高伤, 斩击100→中伤, 治疗术0→低伤
        high_rows = [r for r in rows if r[2] == '高伤']
        mid_rows = [r for r in rows if r[2] == '中伤']
        low_rows = [r for r in rows if r[2] == '低伤']
        assert len(high_rows) > 0  # 至少火球术200是高伤
        assert len(low_rows) > 0  # 至少治疗术0是低伤

    def test_case_when_with_else(self, game_config):
        """CASE WHEN with ELSE默认值"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            game_config,
            "SELECT skill_name, CASE WHEN damage > 0 THEN '有伤害' ELSE '无伤害' END AS has_damage FROM 技能配置 LIMIT 5"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        # 检查有无伤害分类
        has_damage = [r for r in rows if r[1] == '有伤害']
        no_damage = [r for r in rows if r[1] == '无伤害']
        assert len(has_damage) + len(no_damage) == len(rows)

    def test_case_when_with_order_by(self, game_config):
        """CASE WHEN + ORDER BY"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            game_config,
            "SELECT skill_name, CASE WHEN skill_type = '法师' THEN 1 WHEN skill_type = '战士' THEN 2 ELSE 3 END AS type_order FROM 技能配置 ORDER BY type_order LIMIT 5"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        # 法师应该在前面
        if len(rows) >= 2:
            assert rows[0][1] == 1  # 法师

    def test_case_when_in_select_star_context(self, game_config):
        """CASE WHEN在SELECT中与其他列共存"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            game_config,
            "SELECT skill_type, COUNT(*) AS cnt, CASE WHEN COUNT(*) > 2 THEN '多' ELSE '少' END AS count_level FROM 技能配置 GROUP BY skill_type"
        )
        assert result['success'] is True


class TestCoalesce:
    """COALESCE/IFNULL 表达式测试"""

    def test_coalesce_basic(self, multi_sheet_config):
        """COALESCE基础: 第一个非NULL值"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        # 装备配置中有法师法袍(攻击力10)和法杖(攻击力50)
        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT equip_name, COALESCE(atk, 0) AS effective_atk FROM 装备配置"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        # 所有行都应该有值（COALESCE将NULL替换为0）
        for row in rows:
            assert row[1] is not None and row[1] != ''

    def test_ifnull_alias(self, multi_sheet_config):
        """IFNULL是COALESCE的别名"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT equip_name, IFNULL(atk, 0) AS safe_atk FROM 装备配置"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        assert len(rows) > 0

    def test_coalesce_with_column(self, multi_sheet_config):
        """COALESCE多参数: 返回第一个非NULL"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT equip_name, COALESCE(atk, 999) AS atk_display FROM 装备配置 WHERE equip_name = 'robe'"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        # 法袍有atk=10，COALESCE应返回10而非999
        assert len(rows) == 1
        assert rows[0][1] == 10


class TestSubquery:
    """子查询测试"""

    def test_in_subquery(self, multi_sheet_config):
        """WHERE col IN (SELECT ...) — 子查询在IN中"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        # 查询装备适用职业在技能类型中存在的装备
        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT equip_name, equip_class FROM 装备配置 WHERE equip_class IN (SELECT skill_type FROM 技能配置)"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        # 法师/战士/刺客都在技能表中，应该都能匹配
        classes = set(r[1] for r in rows)
        assert 'mage' in classes or 'warrior' in classes or 'assassin' in classes

    def test_scalar_subquery_comparison(self, multi_sheet_config):
        """WHERE col > (SELECT AVG(...) FROM ...) — 标量子查询"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT skill_name, damage FROM 技能配置 WHERE damage > (SELECT AVG(damage) FROM 技能配置)"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        # 至少应该有一些高于平均值的技能
        assert len(rows) > 0
        # 验证所有返回的伤害都高于平均值
        all_damages = [200, 150, 100, 0, 180, 250, 90, 120, 80]
        avg = sum(all_damages) / len(all_damages)
        for row in rows:
            try:
                assert float(row[1]) > avg
            except (ValueError, TypeError):
                pass

    def test_not_in_subquery(self, multi_sheet_config):
        """WHERE col NOT IN (SELECT ...)"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT equip_name FROM 装备配置 WHERE equip_class NOT IN (SELECT skill_type FROM 技能配置 WHERE skill_type = 'assassin')"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        # 不应包含刺客装备
        equip_names = [r[0] for r in rows]
        assert 'dagger' not in equip_names


class TestExists:
    """EXISTS 子查询测试"""

    def test_exists_basic(self, multi_sheet_config):
        """EXISTS基础: 子查询有结果时匹配"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT skill_name, skill_type FROM 技能配置 WHERE EXISTS (SELECT 1 FROM 装备配置 WHERE 装备配置.equip_class = 技能配置.skill_type)"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        # 法师/战士/刺客都有对应装备
        types_with_equip = set(r[1] for r in rows)
        assert 'mage' in types_with_equip or 'warrior' in types_with_equip

    def test_exists_no_match(self, multi_sheet_config):
        """EXISTS: 子查询无结果时不匹配"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT skill_name FROM 技能配置 WHERE EXISTS (SELECT 1 FROM 装备配置 WHERE equip_class = 'nonexistent_class')"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        # 不应匹配任何行
        assert len(rows) == 0


    def test_exists_unqualified_correlated(self, tmp_path):
        """EXISTS关联子查询: 无表限定符的列引用（测试re.sub参数顺序bug修复）"""
        from openpyxl import Workbook
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        wb = Workbook()
        # 技能表: 单行表头，避免双行表头干扰
        ws1 = wb.active
        ws1.title = '技能表'
        ws1.append(['skill_name', 'damage', 'cooldown'])
        ws1.append(['fireball', 200, 5])
        ws1.append(['heal', 0, 3])

        # 怪物表: hp列仅怪物表有，用于关联
        ws2 = wb.create_sheet('怪物表')
        ws2.append(['monster_name', 'hp', 'atk'])
        ws2.append(['dragon', 5000, 300])
        ws2.append(['slime', 100, 10])

        path = str(tmp_path / 'test_exists_unqual.xlsx')
        wb.save(path)

        # damage仅技能表有 → 无表限定符时触发关联子查询替换
        result = execute_advanced_sql_query(
            path,
            "SELECT skill_name FROM 技能表 WHERE EXISTS (SELECT 1 FROM 怪物表 WHERE hp > damage)"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        # heal(damage=0): 怪物hp 5000>0, 100>0 → EXISTS True
        # fireball(damage=200): 怪物hp 5000>200 True, 100>200 False → EXISTS True
        assert len(rows) == 2

    def test_exists_unqualified_no_match(self, tmp_path):
        """EXISTS关联子查询无匹配: 无表限定符"""
        from openpyxl import Workbook
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        wb = Workbook()
        ws1 = wb.active
        ws1.title = '技能表'
        ws1.append(['skill_name', 'damage', 'cooldown'])
        ws1.append(['fireball', 200, 5])

        ws2 = wb.create_sheet('怪物表')
        ws2.append(['monster_name', 'hp', 'atk'])
        ws2.append(['dragon', 5000, 300])

        path = str(tmp_path / 'test_exists_nomatch.xlsx')
        wb.save(path)

        # damage=200, 怪物hp=5000 < 99999 → 无匹配
        result = execute_advanced_sql_query(
            path,
            "SELECT skill_name FROM 技能表 WHERE EXISTS (SELECT 1 FROM 怪物表 WHERE hp > 99999)"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        assert len(rows) == 0


class TestLeftJoinNull:
    """LEFT JOIN NULL处理测试"""

    def test_left_join_null_values(self, multi_sheet_config):
        """LEFT JOIN不匹配行应正确显示NULL/空值"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        # 辅助没有对应装备，LEFT JOIN应该产生NULL
        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT 技能配置.skill_name, 装备配置.equip_name FROM 技能配置 LEFT JOIN 装备配置 ON 技能配置.skill_type = 装备配置.equip_class"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        # 辅助(support)没有对应装备，JOIN后equip_name应为空
        support_rows = [r for r in rows if r[0] == 'heal']
        assert len(support_rows) == 1
        # NULL值应该被序列化为空字符串
        assert support_rows[0][1] == '' or support_rows[0][1] is None

    def test_left_join_null_serialization(self, multi_sheet_config):
        """LEFT JOIN NULL值序列化不崩溃"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT skill_name, equip_name FROM 技能配置 LEFT JOIN 装备配置 ON skill_type = equip_class"
        )
        assert result['success'] is True
        # 关键：不应该抛出ValueError(cannot convert float NaN to integer)
        assert 'data' in result
        assert len(result['data']) > 1  # 至少有表头+数据


class TestCombinedExpressions:
    """组合表达式测试"""

    def test_case_when_with_subquery(self, multi_sheet_config):
        """CASE WHEN + 子查询组合"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT skill_name, CASE WHEN damage > (SELECT AVG(damage) FROM 技能配置) THEN 'above_avg' ELSE 'below_avg' END AS damage_level FROM 技能配置"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        assert len(rows) > 0

    def test_coalesce_in_select_with_where(self, multi_sheet_config):
        """COALESCE + WHERE组合"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT equip_name, COALESCE(atk, 0) AS safe_atk FROM 装备配置 WHERE COALESCE(atk, 0) > 30"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        # 所有返回的safe_atk应该>30
        for row in rows:
            try:
                assert float(row[1]) > 30
            except (ValueError, TypeError):
                pass


class TestCTE:
    """CTE (WITH ... AS ...) 测试"""

    @pytest.mark.skipif(
        sys.platform == "darwin" and sys.version_info < (3, 11),
        reason="python-calamine stat issue on macOS + Python 3.10"
    )
    def test_basic_cte(self, game_config):
        """基本CTE查询"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            game_config,
            "WITH high AS (SELECT * FROM 技能配置 WHERE 伤害 > 100) SELECT skill_name, damage FROM high ORDER BY damage DESC"
        )
        assert result['success'] is True, f"CTE query failed: {result.get('message', 'N/A')}"
        data = result['data']
        rows = [r for r in data if r != data[0]]
        assert len(rows) >= 3
        # 第一行应该是最高伤害
        assert float(rows[0][1]) >= float(rows[1][1])

    @pytest.mark.skipif(
        sys.platform == "darwin" and sys.version_info < (3, 11),
        reason="python-calamine stat issue on macOS + Python 3.10"
    )
    def test_multi_cte(self, multi_sheet_config):
        """多CTE链式引用"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "WITH mages AS (SELECT * FROM 技能配置 WHERE skill_type='mage'), strong AS (SELECT * FROM mages WHERE damage > 150) SELECT skill_name FROM strong"
        )
        assert result['success'] is True, f"Multi-CTE query failed: {result.get('message', 'N/A')}"
        data = result['data']
        rows = [r for r in data if r != data[0]]
        assert len(rows) >= 1

    @pytest.mark.skipif(
        sys.platform == "darwin" and sys.version_info < (3, 11),
        reason="python-calamine stat issue on macOS + Python 3.10"
    )
    def test_cte_with_aggregation(self, game_config):
        """CTE + 聚合查询"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            game_config,
            "WITH stats AS (SELECT 技能类型, AVG(伤害) as avg_dmg FROM 技能配置 GROUP BY 技能类型) SELECT * FROM stats ORDER BY avg_dmg DESC"
        )
        assert result['success'] is True, f"CTE aggregation query failed: {result.get('message', 'N/A')}"
        data = result['data']
        rows = [r for r in data if r != data[0]]
        assert len(rows) >= 2  # 至少2种类型


class TestStringFunctions:
    """字符串函数测试"""

    def test_upper_in_select(self, multi_sheet_config):
        """UPPER函数"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT UPPER(skill_name) as name FROM 技能配置 WHERE skill_type='warrior'"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        assert len(rows) >= 1
        assert rows[0][0] == 'SLASH'

    def test_lower_in_select(self, multi_sheet_config):
        """LOWER函数"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT LOWER(skill_type) as low_type FROM 技能配置 LIMIT 1"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        assert len(rows) == 1

    def test_trim_in_select(self, multi_sheet_config):
        """TRIM函数"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        # 创建带空格的数据
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'test'
        ws.append(['name', 'val'])
        ws.append(['  hello  ', 1])
        ws.append(['  world  ', 2])
        import tempfile, os
        path = os.path.join(tempfile.gettempdir(), 'test_trim.xlsx')
        wb.save(path)

        result = execute_advanced_sql_query(path, "SELECT TRIM(name) as clean FROM test")
        assert result['success'] is True
        rows = [r for r in result['data'] if r != result['data'][0]]
        assert rows[0][0] == 'hello'

    def test_length_in_select(self, multi_sheet_config):
        """LENGTH函数"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT skill_name, LENGTH(skill_name) as len FROM 技能配置 WHERE skill_type='warrior'"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        assert len(rows) >= 1
        assert rows[0][1] == len('slash')

    def test_concat_in_select(self, multi_sheet_config):
        """CONCAT函数"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT CONCAT(skill_type, '-', skill_name) as label FROM 技能配置 WHERE skill_type='mage'"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        assert len(rows) >= 1
        assert rows[0][0] == 'mage-fireball'

    def test_replace_in_select(self, multi_sheet_config):
        """REPLACE函数"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT REPLACE(skill_name, 'ball', 'BOMB') as new_name FROM 技能配置 WHERE skill_name='fireball'"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        assert len(rows) == 1
        assert rows[0][0] == 'fireBOMB'

    def test_substring_in_select(self, multi_sheet_config):
        """SUBSTRING函数"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT SUBSTRING(skill_name, 1, 3) as short FROM 技能配置 WHERE skill_name='fireball'"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        assert len(rows) == 1
        assert rows[0][0] == 'fir'  # SQL 1-based: first 3 chars

    def test_upper_in_where(self, multi_sheet_config):
        """WHERE中使用UPPER函数"""
        from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

        result = execute_advanced_sql_query(
            multi_sheet_config,
            "SELECT skill_name FROM 技能配置 WHERE UPPER(skill_type) = 'MAGE'"
        )
        assert result['success'] is True
        data = result['data']
        rows = [r for r in data if r != data[0]]
        assert len(rows) >= 1
        for row in rows:
            assert row[0] in ('fireball', 'ice')
