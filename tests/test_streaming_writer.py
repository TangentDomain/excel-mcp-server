"""StreamingWriter 单元测试 — calamine读取 + write_only写入"""

import os
import shutil
import tempfile
import time

import pytest
from openpyxl import Workbook, load_workbook

from excel_mcp_server_fastmcp.core.streaming_writer import StreamingWriter


@pytest.fixture
def sample_file(tmp_path):
    """创建带数据的测试文件"""
    fp = str(tmp_path / "test.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Skills"

    # 设置一些列宽（测试保留）
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15

    # 写入表头 + 数据
    ws.append(["ID", "Name", "Damage"])
    ws.append([1, "Fireball", 100])
    ws.append([2, "Ice Spike", 80])
    ws.append([3, "Thunder", 120])

    # 第二个工作表
    ws2 = wb.create_sheet("Items")
    ws2.append(["ItemID", "ItemName"])
    ws2.append([101, "Sword"])
    ws2.append([102, "Shield"])

    wb.save(fp)
    wb.close()
    return fp


@pytest.fixture
def multi_sheet_file(tmp_path):
    """创建多工作表测试文件"""
    fp = str(tmp_path / "multi.xlsx")
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["A", "B"])
    ws1.append([1, 2])

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["X", "Y"])
    ws2.append([10, 20])
    ws2.append([30, 40])

    wb.save(fp)
    wb.close()
    return fp


class TestStreamingAvailability:
    def test_calamine_available(self):
        assert StreamingWriter.is_available() is True


class TestBatchInsertStreaming:
    def test_basic_insert(self, sample_file):
        """基本批量插入"""
        data = [{"ID": 4, "Name": "Heal", "Damage": 50}]
        success, msg, meta = StreamingWriter.batch_insert_rows(
            sample_file, "Skills", data
        )
        assert success is True
        assert meta['inserted_count'] == 1
        assert meta['action'] == 'batch_insert_streaming'

        # 验证数据
        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert len(rows) == 5  # 1 header + 3 existing + 1 new
        assert rows[-1] == (4, "Heal", 50)

    def test_multi_row_insert(self, sample_file):
        """多行插入"""
        data = [
            {"ID": 4, "Name": "Heal", "Damage": 50},
            {"ID": 5, "Name": "Poison", "Damage": 30},
        ]
        success, msg, meta = StreamingWriter.batch_insert_rows(
            sample_file, "Skills", data
        )
        assert success is True
        assert meta['inserted_count'] == 2

        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert len(rows) == 6

    def test_other_sheets_preserved(self, sample_file):
        """其他工作表不受影响"""
        data = [{"ID": 4, "Name": "Heal", "Damage": 50}]
        StreamingWriter.batch_insert_rows(sample_file, "Skills", data)

        wb = load_workbook(sample_file, read_only=True)
        ws_items = wb["Items"]
        rows = list(ws_items.iter_rows(values_only=True))
        wb.close()
        assert len(rows) == 3  # 1 header + 2 items

    def test_unknown_columns(self, sample_file):
        """未知列不报错，记录到unknown_columns"""
        data = [{"ID": 4, "Name": "Heal", "Damage": 50, "Unknown": "x"}]
        success, msg, meta = StreamingWriter.batch_insert_rows(
            sample_file, "Skills", data
        )
        assert success is True
        assert "Unknown" in meta.get('unknown_columns', [])

    def test_sheet_not_found(self, sample_file):
        """工作表不存在"""
        data = [{"ID": 1}]
        success, msg, meta = StreamingWriter.batch_insert_rows(
            sample_file, "NotExist", data
        )
        assert success is False
        assert "不存在" in msg

    def test_col_widths_preserved(self, sample_file):
        """列宽保留"""
        data = [{"ID": 4, "Name": "Heal", "Damage": 50}]
        StreamingWriter.batch_insert_rows(sample_file, "Skills", data, preserve_col_widths=True)

        wb = load_workbook(sample_file)
        ws = wb["Skills"]
        assert ws.column_dimensions['A'].width == 20
        assert ws.column_dimensions['B'].width == 30
        assert ws.column_dimensions['C'].width == 15
        wb.close()

    def test_col_widths_skip(self, sample_file):
        """不保留列宽时使用默认"""
        data = [{"ID": 4, "Name": "Heal", "Damage": 50}]
        StreamingWriter.batch_insert_rows(sample_file, "Skills", data, preserve_col_widths=False)

        wb = load_workbook(sample_file)
        ws = wb["Skills"]
        # write_only 模式下，不保留列宽时使用 openpyxl 默认列宽
        # 只验证文件正常可读，不检查具体值（默认值由openpyxl决定）
        assert ws.max_row > 0
        wb.close()

    def test_empty_data(self, sample_file):
        """空数据列表"""
        data = []
        success, msg, meta = StreamingWriter.batch_insert_rows(
            sample_file, "Skills", data
        )
        # 空数据也应该成功（只是没插入）
        assert success is True
        assert meta['inserted_count'] == 0

    def test_partial_columns(self, sample_file):
        """部分列写入"""
        data = [{"ID": 4, "Name": "Steal"}]  # 缺少 Damage
        success, msg, meta = StreamingWriter.batch_insert_rows(
            sample_file, "Skills", data
        )
        assert success is True

        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # 最后一行应该有 ID=4, Name="Steal"
        # 注意：calamine截断尾部None，所以tuple可能只有2个元素
        assert rows[-1][0] == 4
        assert rows[-1][1] == "Steal"


class TestUpsertStreaming:
    def test_update_existing(self, sample_file):
        """更新已有行"""
        success, msg, meta = StreamingWriter.upsert_row(
            sample_file, "Skills", "ID", 2,
            {"Damage": 999, "Name": "Ice Lance"}
        )
        assert success is True
        assert meta['action'] == 'update'
        assert meta['row'] is not None

        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # 第3行（index=2）应该是更新后的数据
        assert rows[2][0] == 2
        assert rows[2][1] == "Ice Lance"
        assert rows[2][2] == 999

    def test_insert_new(self, sample_file):
        """插入新行"""
        success, msg, meta = StreamingWriter.upsert_row(
            sample_file, "Skills", "ID", "99",
            {"ID": 99, "Name": "Meteor", "Damage": 500}
        )
        assert success is True
        assert meta['action'] == 'insert'

        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Skills"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert rows[-1][0] == 99
        assert rows[-1][1] == "Meteor"

    def test_upsert_preserves_other_sheets(self, sample_file):
        """upsert不影响其他工作表"""
        StreamingWriter.upsert_row(
            sample_file, "Skills", "ID", "1",
            {"Damage": 200}
        )
        wb = load_workbook(sample_file, read_only=True)
        ws = wb["Items"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        assert len(rows) == 3
        assert rows[1] == (101, "Sword")

    def test_key_column_not_found(self, sample_file):
        """键列不存在"""
        success, msg, meta = StreamingWriter.upsert_row(
            sample_file, "Skills", "NotExist", "1",
            {"Damage": 200}
        )
        assert success is False
        assert "键列" in msg or "不存在" in msg


class TestPerformanceComparison:
    """性能对比：streaming vs openpyxl"""

    def test_large_file_batch_insert(self, tmp_path):
        """大文件批量插入性能对比"""
        # 创建大文件（10000行）
        fp = str(tmp_path / "large.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["ID", "Name", "Value", "Desc", "Extra"])
        for i in range(1, 10001):
            ws.append([i, f"item_{i}", i * 10, f"description_{i}", f"extra_{i}"])
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        wb.save(fp)
        wb.close()

        # streaming 模式
        data = [{"ID": 10001 + i, "Name": f"new_{i}", "Value": i} for i in range(100)]
        start = time.time()
        success, msg, meta = StreamingWriter.batch_insert_rows(fp, "Data", data)
        streaming_time = time.time() - start
        assert success is True

        # 验证数据完整性
        wb = load_workbook(fp, read_only=True)
        ws = wb["Data"]
        row_count = 0
        for _ in ws.iter_rows(values_only=True):
            row_count += 1
        wb.close()
        assert row_count == 10101  # header + 10000 + 100

        # openpyxl 模式对比（相同数据重建）
        fp2 = str(tmp_path / "large2.xlsx")
        shutil.copy(fp, fp2)
        # 重新创建一个相同的源文件用于openpyxl测试
        # （因为streaming已经修改了fp）
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Data"
        ws2.append(["ID", "Name", "Value", "Desc", "Extra"])
        for i in range(1, 10001):
            ws2.append([i, f"item_{i}", i * 10, f"description_{i}", f"extra_{i}"])
        ws2.column_dimensions['A'].width = 10
        ws2.column_dimensions['B'].width = 20
        wb2.save(fp2)
        wb2.close()

        start = time.time()
        from excel_mcp_server_fastmcp.core.excel_manager import ExcelManager
        mgr = ExcelManager(fp2)
        result = mgr.batch_insert_rows("Data", data, streaming=False)
        openpyxl_time = time.time() - start
        assert result.success is True

        # streaming 应该更快（或至少不慢很多）
        # 只记录，不硬性断言（CI环境差异大）
        print(f"\n📊 性能对比 (10000行文件 + 插入100行):")
        print(f"  Streaming: {streaming_time:.3f}s")
        print(f"  OpenPyXL:  {openpyxl_time:.3f}s")
        print(f"  比率: {openpyxl_time / max(streaming_time, 0.001):.1f}x")
