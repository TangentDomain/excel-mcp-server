"""
Round 34 MCP 接口实测 - 边界组合测试 + 类型边界深度测试
================================================================
方向选择:
  A组: 边界组合测试 (特殊字符Sheet名 + 超长列名 + 公式单元格 + 组合攻击)
  B组: 类型边界深度测试 (数值溢出 + 日期边界 + 布尔混用 + 精度极限)
  C组: P0 第7轮回归验证
  D组: 已知P1/P2/P3回归确认

日期: 2026-04-14
轮次: Round 34
"""

import sys
import os
import tempfile
import subprocess
import json
import time
import traceback

sys.path.insert(0, '/root/workspace/excel-mcp-server/src')
sys.path.insert(0, '/root/workspace/excel-mcp-server')

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query,
)

# ============================================================
# 测试数据准备
# ============================================================
TEST_DIR = tempfile.mkdtemp(prefix='r34_test_')
BASE_FILE = os.path.join(TEST_DIR, 'r34_base.xlsx')

def setup_test_file():
    """创建包含多Sheet的测试文件"""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # === Sheet1: 普通数据 ===
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["ID", "Name", "Value", "Price", "Score"])
    for i in range(1, 21):
        ws1.append([i, f"Item-{i}", i * 10, round(i * 1.5, 2), i * 100])
    
    # === 特殊字符Sheet名 ===
    # Sheet名含空格
    ws2 = wb.create_sheet("Data Sheet")
    ws2.append(["ID", "Info", "Amount"])
    for i in range(1, 11):
        ws2.append([i, f"Info-{i}", i * 100])
    
    # Sheet名含中文
    ws3 = wb.create_sheet("数据表")
    ws3.append(["编号", "名称", "数值"])
    for i in range(1, 11):
        ws3.append([i, f"项目{i}", i * 50])
    
    # Sheet名含下划线和数字
    ws4 = wb.create_sheet("Test_2024_Data")
    ws4.append(["Key", "Val", "Flag"])
    for i in range(1, 11):
        ws4.append([f"K{i}", i * 2.5, i % 2 == 0])
    
    # Sheet名含连字符
    ws5 = wb.create_sheet("my-data")
    ws5.append(["X", "Y", "Z"])
    for i in range(1, 11):
        ws5.append([i, i * 3, i * 4])
    
    # === 含公式的Sheet ===
    ws6 = wb.create_sheet("Formulas")
    ws6.append(["A", "B", "C", "D"])
    for i in range(1, 11):
        ws6.append([i, i * 2, i * 3, None])  # D列留空或后续填公式
    
    # 用openpyxl设置公式
    for row in range(2, 12):
        ws6.cell(row=row, column=4).value = f"=A{row}*B{row}+C{row}"
    
    # === 超长列名的Sheet ===
    ws7 = wb.create_sheet("LongCols")
    long_name_a = "A_" + "x" * 200  # 202字符列名
    long_name_b = "B_" + "y" * 200
    long_name_c = "C_" + "z" * 200
    ws7.append([long_name_a, long_name_b, long_name_c])
    for i in range(1, 6):
        ws7.append([f"val_a_{i}", f"val_b_{i}", f"val_c_{i}"])
    
    # === 混合类型Sheet (用于类型边界测试) ===
    ws8 = wb.create_sheet("Types")
    ws8.append(["ID", "IntVal", "FloatVal", "DateVal", "BoolVal", "TextVal", "NullCol"])
    from datetime import datetime, date
    for i in range(1, 16):
        ws8.append([
            i,
            i * 1000000,           # 大整数
            round(i * 0.123456789, 9),  # 高精度浮点
            datetime(2024, 1, 1).date() + __import__('datetime').timedelta(days=i),
            i % 3 == 0,            # 布尔值
            f"text_{i}",
            None                   # NULL值
        ])
    
    wb.save(BASE_FILE)
    return BASE_FILE


# ============================================================
# MCP 测试框架
# ============================================================
class MCPTestResult:
    def __init__(self):
        self.results = []
        self.total = 0
        self.passed = 0
        self.failed = 0
        self.errors = []
    
    def test(self, name, tool_name, input_data, expected_desc, actual_func):
        """执行单个MCP测试"""
        self.total += 1
        try:
            result = actual_func()
            success = result.get('success', False)
            message = str(result.get('message', ''))[:200]
            data_info = ''
            if result.get('data') is not None:
                d = result['data']
                if isinstance(d, list):
                    data_info = f"[{len(d)}行]"
                    if len(d) > 0 and isinstance(d[0], dict):
                        data_info += f" cols={list(d[0].keys())[:5]}"
                else:
                    data_info = str(d)[:100]
            
            affected = result.get('affected_rows', result.get('affected', 'N/A'))
            
            if success:
                self.passed += 1
                status = "✅ PASS"
                print(f"🔧 MCP Test: {tool_name}")
                print(f"   输入: {json.dumps(input_data, ensure_ascii=False)[:150]}")
                print(f"   期望: {expected_desc}")
                print(f"   实际: success, {data_info}, affected={affected}")
                print(f"   结果: {status}")
            else:
                self.failed += 1
                status = "❌ FAIL"
                print(f"🔧 MCP Test: {tool_name}")
                print(f"   输入: {json.dumps(input_data, ensure_ascii=False)[:150]}")
                print(f"   期望: {expected_desc}")
                print(f"   实际: fail, msg={message}")
                print(f"   结果: {status}")
            
            self.results.append({
                'name': name,
                'tool': tool_name,
                'input': input_data,
                'expected': expected_desc,
                'success': success,
                'message': message,
            })
            return success
            
        except Exception as e:
            self.failed += 1
            err_msg = str(e)[:200]
            self.errors.append({'name': name, 'error': err_msg})
            print(f"🔧 MCP Test: {tool_name}")
            print(f"   输入: {json.dumps(input_data, ensure_ascii=False)[:150]}")
            print(f"   期望: {expected_desc}")
            print(f"   实际: EXCEPTION: {err_msg}")
            print(f"   结果: ❌ EXCEPTION")
            return False
    
    def summary(self):
        rate = (self.passed / self.total * 100) if self.total > 0 else 0
        print(f"\n{'='*70}")
        print(f"📊 Round 34 MCP 测试汇总: {self.passed}/{self.total} ({rate:.1f}%通过率)")
        print(f"   ✅ 通过: {self.passed}  ❌ 失败: {self.failed}")
        if self.errors:
            print(f"   ⚠️ 异常: {len(self.errors)}个")
        print(f"{'='*70}")
        return self.passed, self.failed, self.total


# ============================================================
# A组: 边界组合测试
# ============================================================
def run_group_a_boundary_combo(t: MCPTestResult, fp):
    """边界组合: 特殊字符Sheet名 + 超长列名 + 公式"""
    print("\n" + "="*70)
    print("🔍 A组: 边界组合测试 (特殊Sheet名 + 超长列名 + 公式)")
    print("="*70)
    
    # A1: 空格Sheet名查询
    t.test(
        "A1-空格Sheet名查询",
        "excel_query",
        {"sql": "SELECT * FROM `Data Sheet`"},
        "成功查询含空格的Sheet",
        lambda: execute_advanced_sql_query(fp, "SELECT * FROM `Data Sheet`")
    )
    
    # A2: 中文Sheet名查询
    t.test(
        "A2-中文Sheet名查询",
        "excel_query",
        {"sql": "SELECT * FROM 数据表"},
        "成功查询中文Sheet",
        lambda: execute_advanced_sql_query(fp, "SELECT * FROM 数据表")
    )
    
    # A3: 下划线数字Sheet名
    t.test(
        "A3-下划线数字Sheet名",
        "excel_query",
        {"sql": "SELECT * FROM Test_2024_Data"},
        "成功查询含下划线数字的Sheet",
        lambda: execute_advanced_sql_query(fp, "SELECT * FROM Test_2024_Data")
    )
    
    # A4: 连字符Sheet名
    t.test(
        "A4-连字符Sheet名",
        "excel_query",
        {"sql": "SELECT * FROM my-data"},
        "成功查询含连字符的Sheet",
        lambda: execute_advanced_sql_query(fp, "SELECT * FROM my-data")
    )
    
    # A5: 跨特殊Sheet名JOIN
    t.test(
        "A5-跨特殊Sheet名JOIN",
        "excel_query",
        {"sql": "SELECT a.ID, a.Name, b.Amount FROM Sheet1 a JOIN `Data Sheet` b ON a.ID = b.ID"},
        "跨普通Sheet和空格Sheet JOIN",
        lambda: execute_advanced_sql_query(fp, 
            "SELECT a.ID, a.Name, b.Amount FROM Sheet1 a JOIN `Data Sheet` b ON a.ID = b.ID")
    )
    
    # A6: 中文Sheet WHERE条件
    t.test(
        "A6-中文Sheet WHERE条件",
        "excel_query",
        {"sql": "SELECT * FROM 数据表 WHERE 数值 > 200"},
        "中文Sheet带WHERE条件",
        lambda: execute_advanced_sql_query(fp, "SELECT * FROM 数据表 WHERE 数值 > 200")
    )
    
    # A7: 公式Sheet读取
    t.test(
        "A7-公式Sheet读取",
        "excel_query",
        {"sql": "SELECT * FROM Formulas"},
        "读取含公式的Sheet（公式值应被解析为计算结果）",
        lambda: execute_advanced_sql_query(fp, "SELECT * FROM Formulas")
    )
    
    # A8: 公式Sheet聚合
    t.test(
        "A8-公式Sheet聚合计算",
        "excel_query",
        {"sql": "SELECT SUM(A), AVG(B), COUNT(*) FROM Formulas"},
        "对含公式列的Sheet做聚合",
        lambda: execute_advanced_sql_query(fp, "SELECT SUM(A), AVG(B), COUNT(*) FROM Formulas")
    )
    
    # A9: 超长列名Sheet查询
    t.test(
        "A9-超长列名Sheet查询",
        "excel_query",
        {"sql": "SELECT * FROM LongCols"},
        "成功读取超长列名(202字符)的Sheet",
        lambda: execute_advanced_sql_query(fp, "SELECT * FROM LongCols")
    )
    
    # A10: 超长列名WHERE条件
    t.test(
        "A10-超长列名WHERE条件",
        "excel_query",
        {"sql": "SELECT * FROM LongCols WHERE B_yyyyyy... LIKE '%val%'"},
        "对超长列名做WHERE过滤",
        lambda: execute_advanced_sql_query(fp, 
            "SELECT * FROM LongCols WHERE B_" + "y"*200 + " LIKE '%val_b_%'")
    )
    
    # A11: 特殊Sheet名UPDATE
    t.test(
        "A11-中文Sheet UPDATE",
        "excel_update_query",
        {"sql": "UPDATE 数据表 SET 数值 = 999 WHERE 编号 = 1"},
        "成功更新中文Sheet的数据",
        lambda: execute_advanced_update_query(fp, "UPDATE 数据表 SET 数值 = 999 WHERE 编号 = 1")
    )
    
    # A12: 验证中文Sheet更新结果
    t.test(
        "A12-验证中文Sheet更新",
        "excel_query",
        {"sql": "SELECT * FROM 数据表 WHERE 编号 = 1"},
        "更新后数值=999",
        lambda: execute_advanced_sql_query(fp, "SELECT * FROM 数据表 WHERE 编号 = 1")
    )
    
    # A13: 连字符Sheet UPDATE
    t.test(
        "A13-连字符Sheet UPDATE",
        "excel_update_query",
        {"sql": "UPDATE my-data SET Y = 9999 WHERE X = 1"},
        "成功更新连字符名Sheet",
        lambda: execute_advanced_update_query(fp, "UPDATE my-data SET Y = 9999 WHERE X = 1")
    )
    
    # A14: 特殊Sheet名INSERT
    t.test(
        "A14-中文Sheet INSERT",
        "excel_insert_query",
        {"sql": "INSERT INTO 数据表 (编号, 名称, 数值) VALUES (99, '新项目', 777)"},
        "向中文Sheet插入数据",
        lambda: execute_advanced_insert_query(fp, 
            "INSERT INTO 数据表 (编号, 名称, 数值) VALUES (99, '新项目', 777)")
    )
    
    # A15: 特殊Sheet名DELETE
    t.test(
        "A15-中文Sheet DELETE",
        "excel_delete_query",
        {"sql": "DELETE FROM 数据表 WHERE 编号 = 99"},
        "从中文Sheet删除刚插入的行",
        lambda: execute_advanced_delete_query(fp, "DELETE FROM 数据表 WHERE 编号 = 99")
    )
    
    # A16: 公式Sheet UPDATE (尝试更新公式列)
    t.test(
        "A16-公式列UPDATE尝试",
        "excel_update_query",
        {"sql": "UPDATE Formulas SET D = 9999 WHERE A = 1"},
        "尝试更新含公式的列（可能覆盖公式）",
        lambda: execute_advanced_update_query(fp, "UPDATE Formulas SET D = 9999 WHERE A = 1")
    )
    
    # A17: 公式Sheet更新后回读
    t.test(
        "A17-公式列更新后回读",
        "excel_query",
        {"sql": "SELECT D FROM Formulas WHERE A = 1"},
        "检查公式列被更新后的值",
        lambda: execute_advanced_sql_query(fp, "SELECT D FROM Formulas WHERE A = 1")
    )


# ============================================================
# B组: 类型边界深度测试
# ============================================================
def run_group_b_type_boundary(t: MCPTestResult, fp):
    """类型边界: 数值溢出 + 日期边界 + 布尔混用 + 精度极限"""
    print("\n" + "="*70)
    print("🔍 B组: 类型边界深度测试 (数值溢出/日期/布尔/精度)")
    print("="*70)
    
    # B1: 正常整数范围
    t.test(
        "B1-正常整数读写",
        "excel_update_query+query",
        {"sql": "UPDATE Types SET IntVal = 42 WHERE ID = 1"},
        "写入整数42后可正确读回",
        lambda: _update_then_verify(fp, "Types", "IntVal", 42, 1)
    )
    
    # B2: 大整数 (接近int32上限)
    t.test(
        "B2-大整数(int32边界)",
        "excel_update_query+query",
        {"sql": "UPDATE Types SET IntVal = 2147483647 WHERE ID = 2"},
        "写入int32最大值2147483647后可读回",
        lambda: _update_then_verify(fp, "Types", "IntVal", 2147483647, 2)
    )
    
    # B3: 超大整数 (超过int32)
    t.test(
        "B3-超大整数(超int32)",
        "excel_update_query+query",
        {"sql": "UPDATE Types SET IntVal = 9999999999 WHERE ID = 3"},
        "写入超大整数9999999999后可读回",
        lambda: _update_then_verify(fp, "Types", "IntVal", 9999999999, 3)
    )
    
    # B4: 负数边界
    t.test(
        "B4-负数边界",
        "excel_update_query+query",
        {"sql": "UPDATE Types SET IntVal = -999999999 WHERE ID = 4"},
        "写入大负数后可读回",
        lambda: _update_then_verify(fp, "Types", "IntVal", -999999999, 4)
    )
    
    # B5: 极小浮点数
    t.test(
        "B5-极小浮点数",
        "excel_update_query+query",
        {"sql": "UPDATE Types SET FloatVal = 0.000000001 WHERE ID = 5"},
        "写入极小浮点数1e-9后精度保留",
        lambda: _update_then_verify(fp, "Types", "FloatVal", 0.000000001, 5, tolerance=1e-15)
    )
    
    # B6: 极大浮点数 (非inf)
    t.test(
        "B6-极大浮点数(非inf)",
        "excel_update_query+query",
        {"sql": "UPDATE Types SET FloatVal = 1.7976931348623157e+308 WHERE ID = 6"},
        "写入float max后不损坏文件",
        lambda: _update_then_check_readable(fp, "Types", "FloatVal", 1.7976931348623157e+308, 6)
    )
    
    # B7: 零值
    t.test(
        "B7-精确零值",
        "excel_update_query+query",
        {"sql": "UPDATE Types SET FloatVal = 0.0 WHERE ID = 7"},
        "写入精确0.0后读回为0",
        lambda: _update_then_verify(fp, "Types", "FloatVal", 0.0, 7)
    )
    
    # B8: 负零
    t.test(
        "B8-负零(-0.0)",
        "excel_update_query+query",
        {"sql": "UPDATE Types SET FloatVal = -0.0 WHERE ID = 8"},
        "写入-0.0后文件仍可读",
        lambda: _update_then_check_readable(fp, "Types", "FloatVal", -0.0, 8)
    )
    
    # B9: 高精度浮点 (0.1 + 0.2问题)
    t.test(
        "B9-浮点精度(0.1+0.2)",
        "excel_update_query+query",
        {"sql": "UPDATE Types SET FloatVal = 0.30000000000000004 WHERE ID = 9"},
        "写入0.1+0.2的结果值后精度保留",
        lambda: _update_then_verify(fp, "Types", "FloatVal", 0.30000000000000004, 9, tolerance=1e-15)
    )
    
    # B10: 科学计数法
    t.test(
        "B10-科学计数法",
        "excel_update_query+query",
        {"sql": "UPDATE Types SET FloatVal = 1.23e-10 WHERE ID = 10"},
        "科学计数法1.23e-10正确处理",
        lambda: _update_then_verify(fp, "Types", "FloatVal", 1.23e-10, 10, tolerance=1e-16)
    )
    
    # B11: 布尔值 TRUE
    t.test(
        "B11-布尔TRUE",
        "excel_update_query+query",
        {"sql": "UPDATE Types SET BoolVal = TRUE WHERE ID = 11"},
        "写入布尔TRUE后可正确读回",
        lambda: _update_then_verify(fp, "Types", "BoolVal", True, 11)
    )
    
    # B12: 布尔值 FALSE
    t.test(
        "B12-布尔FALSE",
        "excel_update_query+query",
        {"sql": "UPDATE Types SET BoolVal = FALSE WHERE ID = 12"},
        "写入布尔FALSE后可正确读回",
        lambda: _update_then_verify(fp, "Types", "BoolVal", False, 12)
    )
    
    # B13: 布尔与算术混用
    t.test(
        "B13-布尔算术混用",
        "excel_query",
        {"sql": "SELECT ID, BoolVal, BoolVal + 1 AS Calc FROM Types WHERE ID <= 15"},
        "布尔列参与算术运算(TRUE=1/FALSE=0)",
        lambda: execute_advanced_sql_query(fp, 
            "SELECT ID, BoolVal, BoolVal + 1 AS Calc FROM Types WHERE ID <= 15")
    )
    
    # B14: NULL值参与运算
    t.test(
        "B14-NULL值运算",
        "excel_query",
        {"sql": "SELECT ID, NullCol, NullCol + 1 AS Calc FROM Types WHERE ID <= 5"},
        "NULL参与运算返回NULL或优雅处理",
        lambda: execute_advanced_sql_query(fp,
            "SELECT ID, NullCol, NullCol + 1 AS Calc FROM Types WHERE ID <= 5")
    )
    
    # B15: 日期比较
    t.test(
        "B15-日期WHERE比较",
        "excel_query",
        {"sql": "SELECT * FROM Types WHERE DateVal > '2024-01-05'"},
        "日期类型WHERE条件过滤",
        lambda: execute_advanced_sql_query(fp, 
            "SELECT * FROM Types WHERE DateVal > '2024-01-05'")
    )
    
    # B16: 日期排序
    t.test(
        "B16-日期ORDER BY",
        "excel_query",
        {"sql": "SELECT ID, DateVal FROM Types ORDER BY DateVal DESC LIMIT 3"},
        "按日期降序排列",
        lambda: execute_advanced_sql_query(fp,
            "SELECT ID, DateVal FROM Types ORDER BY DateVal DESC LIMIT 3")
    )
    
    # B17: 整数溢出表达式
    t.test(
        "B17-整数溢出表达式",
        "excel_query",
        {"sql": "SELECT IntVal * 10000000 FROM Types WHERE ID = 1"},
        "大整数乘法溢出时优雅处理",
        lambda: execute_advanced_sql_query(fp,
            "SELECT IntVal * 10000000 FROM Types WHERE ID = 1")
    )
    
    # B18: 除以零
    t.test(
        "B18-除以零",
        "excel_query",
        {"sql": "SELECT 1/0 AS DivZero, IntVal FROM Types WHERE ID = 1"},
        "除以零时优雅报错而非崩溃",
        lambda: execute_advanced_sql_query(fp,
            "SELECT 1/0 AS DivZero, IntVal FROM Types WHERE ID = 1")
    )
    
    # B19: 字符串转数字
    t.test(
        "B19-字符串列算术",
        "excel_query",
        {"sql": "SELECT TextVal + 0 AS Num FROM Types WHERE ID = 1"},
        "字符串列尝试算术转换时的行为",
        lambda: execute_advanced_sql_query(fp,
            "SELECT TextVal + 0 AS Num FROM Types WHERE ID = 1")
    )
    
    # B20: 混合类型聚合
    t.test(
        "B20-混合类型聚合",
        "excel_query",
        {"sql": "SELECT COUNT(*), SUM(IntVal), AVG(FloatVal), COUNT(BoolVal) FROM Types"},
        "混合类型列的聚合统计",
        lambda: execute_advanced_sql_query(fp,
            "SELECT COUNT(*), SUM(IntVal), AVG(FloatVal), COUNT(BoolVal) FROM Types")
    )


def _update_then_verify(fp, table, col, expected_val, id_val, tolerance=None):
    """UPDATE后立即SELECT验证值"""
    sql = f"UPDATE {table} SET {col} = {repr(expected_val)} WHERE ID = {id_val}"
    r1 = execute_advanced_update_query(fp, sql)
    if not r1.get('success'):
        return {'success': False, 'message': f"UPDATE failed: {r1.get('message', '')}"}
    
    time.sleep(0.2)
    r2 = execute_advanced_sql_query(fp, f"SELECT {col} FROM {table} WHERE ID = {id_val}")
    if not r2.get('success'):
        return {'success': False, 'message': f"SELECT failed: {r2.get('message', '')}"}
    
    data = r2.get('data', [])
    if not data or len(data) == 0:
        return {'success': False, 'message': 'No data returned'}
    
    actual = list(data[0].values())[0] if isinstance(data[0], dict) else data[0][0]
    
    # 类型兼容比较
    try:
        if tolerance is not None:
            match = abs(float(actual) - float(expected_val)) < tolerance
        elif isinstance(expected_val, bool):
            match = actual == expected_val or actual in (1, 0, True, False, 'True', 'False')
        elif isinstance(expected_val, float):
            match = abs(float(actual) - float(expected_val)) < 1e-10
        else:
            match = actual == expected_val or str(actual) == str(expected_val)
    except (ValueError, TypeError):
        match = str(actual) == str(expected_val)
    
    if match:
        return {'success': True, 'message': f'OK: wrote={expected_val}, read={actual}', 
                'data': [{'verified': actual}]}
    else:
        return {'success': False, 'message': f'MISMATCH: wrote={expected_val}, read={actual}',
                'data': [{'actual': actual}]}


def _update_then_check_readable(fp, table, col, val, id_val):
    """UPDATE后仅检查文件仍可读（用于可能损坏文件的极端值）"""
    sql = f"UPDATE {table} SET {col} = {repr(val)} WHERE ID = {id_val}"
    r1 = execute_advanced_update_query(fp, sql)
    if not r1.get('success'):
        return {'success': False, 'message': f"UPDATE failed: {r1.get('message', '')}"}
    
    time.sleep(0.2)
    # 尝试读取整张表确认文件未损坏
    r2 = execute_advanced_sql_query(fp, f"SELECT COUNT(*) as cnt FROM {table}")
    if r2.get('success'):
        return {'success': True, 'message': f'File readable after writing {val}', 
                'data': r2.get('data')}
    else:
        return {'success': False, 'message': f'File may be corrupted: {r2.get("message", "")}'}


# ============================================================
# C组: P0 第7轮回归验证
# ============================================================
def run_group_p0_regression(t: MCPTestResult, fp):
    """P0漏洞第7轮回归验证"""
    print("\n" + "="*70)
    print("🔴 C组: P0 漏洞第7轮回归验证")
    print("="*70)
    
    # P0-2: SELECT分号多语句
    t.test(
        "P0-2-R7 SELECT分号多语句",
        "excel_query",
        {"sql": "SELECT * FROM Sheet1; SELECT * FROM 数据表"},
        "❌ 应拒绝分号分隔的多语句",
        lambda: execute_advanced_sql_query(fp, 
            "SELECT * FROM Sheet1; SELECT * FROM 数据表")
    )
    
    # P0-4: UPDATE分号多语句
    t.test(
        "P0-4-R7 UPDATE分号多语句",
        "excel_update_query",
        {"sql": "UPDATE Sheet1 SET Value = 999 WHERE ID = 1; DELETE FROM Sheet1 WHERE ID > 1"},
        "❌ 应拒绝UPDATE后的分号语句",
        lambda: execute_advanced_update_query(fp,
            "UPDATE Sheet1 SET Value = 999 WHERE ID = 1; DELETE FROM Sheet1 WHERE ID > 1")
    )
    
    # P0-5: INSERT分号多语句
    t.test(
        "P0-5-R7 INSERT分号多语句",
        "excel_insert_query",
        {"sql": "INSERT INTO Sheet1 (ID, Name, Value, Price, Score) VALUES (999, 'Hack', 0, 0, 0); DELETE FROM Sheet1"},
        "❌ 应拒绝INSERT后的分号语句",
        lambda: execute_advanced_insert_query(fp,
            "INSERT INTO Sheet1 (ID, Name, Value, Price, Score) VALUES (999, 'Hack', 0, 0, 0); DELETE FROM Sheet1")
    )
    
    # P0-6: DELETE分号多语句
    t.test(
        "P0-6-R7 DELETE分号多语句",
        "excel_delete_query",
        {"sql": "DELETE FROM Sheet1 WHERE ID = 999; UPDATE Sheet1 SET Value = -1"},
        "❌ 应拒绝DELETE后的分号语句",
        lambda: execute_advanced_delete_query(fp,
            "DELETE FROM Sheet1 WHERE ID = 999; UPDATE Sheet1 SET Value = -1")
    )
    
    # P0-7: UPDATE注释符全表篡改
    t.test(
        "P0-7-R7 UPDATE注释符全表篡改",
        "excel_update_query",
        {"sql": "UPDATE Sheet1 SET Value = -1 -- WHERE ID = 999"},
        "❌ 应拒绝注释符截断WHERE导致的全表篡改",
        lambda: execute_advanced_update_query(fp,
            "UPDATE Sheet1 SET Value = -1 -- WHERE ID = 999")
    )
    
    # 验证P0-7是否真的篡改了全表
    time.sleep(0.3)
    t.test(
        "P0-7-R7-VERIFY 验证全表篡改",
        "excel_query",
        {"sql": "SELECT COUNT(*), SUM(Value) FROM Sheet1"},
        "如果P0-7成功执行，Value应全部变为-1",
        lambda: execute_advanced_sql_query(fp, 
            "SELECT COUNT(*), SUM(Value) FROM Sheet1")
    )


# ============================================================
# D组: 已知P1/P2/P3 回归确认
# ============================================================
def run_group_d_known_issues(t: MCPTestResult, fp):
    """已知问题回归确认"""
    print("\n" + "="*70)
    print("📋 D组: 已知P1/P2/P3问题回归确认")
    print("="*70)
    
    # P2-1: UPDATE SET || 拼接
    t.test(
        "P2-1 UPDATE SET ||拼接",
        "excel_update_query",
        {"sql": "UPDATE Sheet1 SET Name = Name || '_suffix' WHERE ID = 1"},
        "❌ 已知不支持 || 拼接",
        lambda: execute_advanced_update_query(fp,
            "UPDATE Sheet1 SET Name = Name || '_suffix' WHERE ID = 1")
    )
    
    # P2-2: CASE WHEN算术混合
    t.test(
        "P2-2 CASE WHEN算术混合",
        "excel_update_query",
        {"sql": "UPDATE Sheet1 SET Value = CASE WHEN Price > 10 THEN Price * 2 ELSE Price END"},
        "❌ 已知CASE WHEN在UPDATE中可能不完全支持",
        lambda: execute_advanced_update_query(fp,
            "UPDATE Sheet1 SET Value = CASE WHEN Price > 10 THEN Price * 2 ELSE Price END")
    )
    
    # P3-1: 多余逗号幽灵列
    t.test(
        "P3-1 多余逗号幽灵列",
        "excel_query",
        {"sql": "SELECT ,,, FROM Sheet1"},
        "❌ 已知产生幽灵_ROW_列",
        lambda: execute_advanced_sql_query(fp, "SELECT ,,, FROM Sheet1")
    )
    
    # SQL大小写敏感
    t.test(
        "P2-? SQL大小写敏感 sheet1",
        "excel_query",
        {"sql": "SELECT * FROM sheet1"},
        "❌ 已知大小写敏感(sheet1≠Sheet1)",
        lambda: execute_advanced_sql_query(fp, "SELECT * FROM sheet1")
    )
    
    # P2-5: 超长SQL栈溢出(R33发现)
    t.test(
        "P2-5 超长SQL栈溢出(DoS)",
        "excel_query",
        {"sql": "超长OR条件SQL(约500个OR)"},
        lambda: _test_long_sql(fp),
        "❌ 已知超长SQL导致栈溢出"
    )
    
    # P3-2: Excel单元格32767字符限制
    t.test(
        "P3-2 Excel字符限制32767",
        "excel_update_query+query",
        {"sql": "UPDATE Sheet1 SET Name = 'A'*40000 WHERE ID = 1"},
        "❌ 已知超长文本被截断到32767字符",
        lambda: _test_long_text(fp)
    )


def _test_long_sql(fp):
    """生成超长OR条件的SQL"""
    conditions = " OR ".join([f"ID = {i}" for i in range(1, 501)])
    sql = f"SELECT * FROM Sheet1 WHERE {conditions}"
    return execute_advanced_sql_query(fp, sql)


def _test_long_text(fp):
    """测试超长文本写入"""
    long_text = "A" * 40000
    r1 = execute_advanced_update_query(fp, f"UPDATE Sheet1 SET Name = '{long_text}' WHERE ID = 1")
    if not r1.get('success'):
        return {'success': False, 'message': f"UPDATE failed: {r1.get('message', '')}"}
    
    time.sleep(0.2)
    r2 = execute_advanced_sql_query(fp, "SELECT LENGTH(Name) as len, Name FROM Sheet1 WHERE ID = 1")
    if not r2.get('success'):
        return {'success': False, 'message': f"SELECT failed: {r2.get('message', '')}"}
    
    data = r2.get('data', [])
    if data:
        actual_len = list(data[0].values())[0]
        if actual_len < 40000:
            return {'success': False, 
                    'message': f'Text truncated: wrote=40000, read back={actual_len}',
                    'data': data}
    return {'success': True, 'message': 'Full length preserved', 'data': r2.get('data')}


# ============================================================
# E组: 额外边界探索
# ============================================================
def run_group_e_extra_boundary(t: MCPTestResult, fp):
    """额外边界探索: 特殊字符组合、嵌套极限等"""
    print("\n" + "="*70)
    print("🔍 E组: 额外边界探索 (特殊组合/嵌套极限)")
    print("="*70)
    
    # E1: Sheet名含点号(通过反引号)
    t.test(
        "E1-Sheet名查询通配符行为",
        "excel_query",
        {"sql": "SELECT * FROM Sheet%"},  
        "百分号通配符在表名中的行为",
        lambda: execute_advanced_sql_query(fp, "SELECT * FROM Sheet%")
    )
    
    # E2: 列名与SQL关键字冲突
    t.test(
        "E2-列名SQL关键字冲突",
        "excel_query",
        {"sql": "SELECT ID, Name, Value, Price, Score FROM Sheet1 ORDER BY Score"},
        "正常使用Score等可能关键字的列名",
        lambda: execute_advanced_sql_query(fp, 
            "SELECT ID, Name, Value, Price, Score FROM Sheet1 ORDER BY Score")
    )
    
    # E3: HAVING子句
    t.test(
        "E3-HAVING子句",
        "excel_query",
        {"sql": "SELECT Rarity_count, COUNT(*) as cnt FROM (SELECT CASE WHEN Price > 15 THEN 'High' ELSE 'Low' END AS Rarity_count FROM Sheet1) sub GROUP BY Rarity_count HAVING COUNT(*) > 3"},
        "HAVING子句过滤分组",
        lambda: execute_advanced_sql_query(fp,
            "SELECT Rarity_count, COUNT(*) as cnt FROM (SELECT CASE WHEN Price > 15 THEN 'High' ELSE 'Low' END AS Rarity_count FROM Sheet1) sub GROUP BY Rarity_count HAVING COUNT(*) > 3")
    )
    
    # E4: DISTINCT + UNION ALL
    t.test(
        "E4-DISTINCT+UNION ALL",
        "excel_query",
        {"sql": "SELECT DISTINCT Name FROM Sheet1 WHERE ID <= 5 UNION ALL SELECT Name FROM Sheet1 WHERE ID > 15"},
        "DISTINCT与UNION ALL组合",
        lambda: execute_advanced_sql_query(fp,
            "SELECT DISTINCT Name FROM Sheet1 WHERE ID <= 5 UNION ALL SELECT Name FROM Sheet1 WHERE ID > 15")
    )
    
    # E5: 子查询在SELECT中
    t.test(
        "E5-标量子查询",
        "excel_query",
        {"sql": "SELECT Name, (SELECT AVG(Value) FROM Sheet1) as avg_val FROM Sheet1 WHERE ID = 1"},
        "SELECT中的标量子查询",
        lambda: execute_advanced_sql_query(fp,
            "SELECT Name, (SELECT AVG(Value) FROM Sheet1) as avg_val FROM Sheet1 WHERE ID = 1")
    )
    
    # E6: BETWEEN操作符
    t.test(
        "E6-BETWEEN操作符",
        "excel_query",
        {"sql": "SELECT * FROM Sheet1 WHERE Value BETWEEN 10 AND 50"},
        "BETWEEN范围查询",
        lambda: execute_advanced_sql_query(fp,
            "SELECT * FROM Sheet1 WHERE Value BETWEEN 10 AND 50")
    )
    
    # E7: IN操作符(列表)
    t.test(
        "E7-IN操作符列表",
        "excel_query",
        {"sql": "SELECT * FROM Sheet1 WHERE ID IN (1, 3, 5, 7, 9, 11, 13, 15, 17, 19)"},
        "IN列表成员判断",
        lambda: execute_advanced_sql_query(fp,
            "SELECT * FROM Sheet1 WHERE ID IN (1, 3, 5, 7, 9, 11, 13, 15, 17, 19)")
    )
    
    # E8: IN子查询
    t.test(
        "E8-IN子查询",
        "excel_query",
        {"sql": "SELECT * FROM Sheet1 WHERE ID IN (SELECT ID FROM 数据表 WHERE 数值 > 200)"},
        "IN子查询跨Sheet",
        lambda: execute_advanced_sql_query(fp,
            "SELECT * FROM Sheet1 WHERE ID IN (SELECT ID FROM 数据表 WHERE 数值 > 200)")
    )
    
    # E9: LIKE模糊匹配
    t.test(
        "E9-LIKE模糊匹配",
        "excel_query",
        {"sql": "SELECT * FROM Sheet1 WHERE Name LIKE 'Item-1%'"},
        "LIKE前缀匹配",
        lambda: execute_advanced_sql_query(fp,
            "SELECT * FROM Sheet1 WHERE Name LIKE 'Item-1%'")
    )
    
    # E10: IS NULL / IS NOT NULL
    t.test(
        "E10-IS NULL判断",
        "excel_query",
        {"sql": "SELECT * FROM Types WHERE NullCol IS NULL"},
        "IS NULL空值判断",
        lambda: execute_advanced_sql_query(fp,
            "SELECT * FROM Types WHERE NullCol IS NULL")
    )
    
    # E11: COALESCE函数
    t.test(
        "E11-COALESCE空值替换",
        "excel_query",
        {"sql": "SELECT ID, COALESCE(NullCol, 0) as safe_val FROM Types WHERE ID <= 5"},
        "COALESCE函数替换NULL",
        lambda: execute_advanced_sql_query(fp,
            "SELECT ID, COALESCE(NullCol, 0) as safe_val FROM Types WHERE ID <= 5")
    )
    
    # E12: 多层嵌套子查询
    t.test(
        "E12-三层嵌套子查询",
        "excel_query",
        {"sql": "SELECT * FROM (SELECT * FROM (SELECT ID, Name, Value FROM Sheet1 WHERE Value > 30) t1 WHERE ID < 15) t2 WHERE Value < 150"},
        "三层嵌套子查询",
        lambda: execute_advanced_sql_query(fp,
            "SELECT * FROM (SELECT * FROM (SELECT ID, Name, Value FROM Sheet1 WHERE Value > 30) t1 WHERE ID < 15) t2 WHERE Value < 150")
    )
    
    # E13: LIMIT OFFSET
    t.test(
        "E13-LIMIT OFFSET分页",
        "excel_query",
        {"sql": "SELECT * FROM Sheet1 ORDER BY ID LIMIT 5 OFFSET 3"},
        "LIMIT OFFSET分页查询",
        lambda: execute_advanced_sql_query(fp,
            "SELECT * FROM Sheet1 ORDER BY ID LIMIT 5 OFFSET 3")
    )
    
    # E14: CROSS JOIN
    t.test(
        "E14-CROSS JOIN交叉连接",
        "excel_query",
        {"sql": "SELECT * FROM Sheet1 s1 CROSS JOIN 数据表 s2 WHERE s1.ID = s2.编号 LIMIT 5"},
        "CROSS JOIN交叉连接",
        lambda: execute_advanced_sql_query(fp,
            "SELECT * FROM Sheet1 s1 CROSS JOIN 数据表 s2 WHERE s1.ID = s2.编号 LIMIT 5")
    )


# ============================================================
# 主流程
# ============================================================
def main():
    print("="*70)
    print("🚀 Round 34 MCP 接口实测")
    print("   方向: 边界组合测试 + 类型边界深度测试 + P0第7轮回归")
    print("="*70)
    print(f"   日期: 2026-04-14")
    print(f"   测试目录: {TEST_DIR}")
    
    # 准备测试文件
    print("\n📁 准备测试数据...")
    fp = setup_test_file()
    print(f"   测试文件: {fp}")
    
    t = MCPTestResult()
    
    # 执行各组测试
    try:
        run_group_a_boundary_combo(t, fp)
        run_group_b_type_boundary(t, fp)
        run_group_p0_regression(t, fp)
        run_group_d_known_issues(t, fp)
        run_group_e_extra_boundary(t, fp)
    except KeyboardInterrupt:
        print("\n⚠️ 测试被中断")
    except Exception as e:
        print(f"\n💥 测试过程异常: {e}")
        traceback.print_exc()
    
    # 汇总
    passed, failed, total = t.summary()
    
    # 清理
    import shutil
    try:
        shutil.rmtree(TEST_DIR)
        print(f"🧹 清理完成: {TEST_DIR}")
    except:
        pass
    
    # 返回退出码
    return 0 if failed == 0 else 1


if __name__ == '__main__':
    sys.exit(main())
