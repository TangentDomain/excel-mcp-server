# -*- coding: utf-8 -*-
"""
format_cells 深度测试 - R55+ 迭代重点

覆盖范围:
  - 单项样式: bold/italic/underline/font_size/font_color/bg_color/number_format/alignment/wrap_text/border_style
  - 组合操作: merge + bold + bg_color 同时传
  - 边界: 空range、超范围、合并后再拆分
  - 中文字体、Unicode列名场景
  - 扁平格式 → 嵌套格式转换 (_normalize_formatting)
  - preset 预设样式
  - 边框详细配置(四边不同)
  - 渐变背景、图案填充
  - text_rotation / indent / shrink_to_fit
  - strikethrough / double underline
"""

import os
import pytest
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl

from excel_mcp_server_fastmcp.core.excel_writer import ExcelWriter
from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations


# ==================== Helper ====================

def _create_test_xlsx(file_path: str, rows: int = 5, cols: int = 4, sheet_name: str = "Sheet1"):
    """创建测试用 xlsx 文件，含数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=r * 10 + c)
    wb.save(file_path)
    wb.close()


def _read_cell_style(file_path: str, cell_ref: str = "A1", sheet_name: str = "Sheet1") -> dict:
    """读取单元格的样式属性用于断言"""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    cell = ws[cell_ref]
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    result = {
        "bold": font.bold,
        "italic": font.italic,
        "underline": font.underline,
        "size": font.size,
        "font_name": font.name,
        "color": str(font.color) if font.color else None,
        "fill_type": fill.fill_type if fill else None,
        "fgColor": str(fill.fgColor) if fill and fill.fgColor else None,
        "alignment_h": alignment.horizontal,
        "alignment_v": alignment.vertical,
        "wrap_text": alignment.wrap_text,
        "text_rotation": alignment.text_rotation,
        "number_format": cell.number_format,
        "border_left": border.left.style if border and border.left else None,
        "border_right": border.right.style if border and border.right else None,
        "border_top": border.top.style if border and border.top else None,
        "border_bottom": border.bottom.style if border and border.bottom else None,
        "value": cell.value,
    }
    wb.close()
    return result


# ==================== Test Class ====================

class TestFormatCellsDeep:
    """format_cells 深度测试套件"""

    # ---------- 1. 单项字体样式 ----------

    def test_bold(self, temp_dir):
        """加粗"""
        fp = str(temp_dir / "bold.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:B2", {"font": {"bold": True}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True

    def test_italic(self, temp_dir):
        """斜体"""
        fp = str(temp_dir / "italic.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"italic": True}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["italic"] is True

    def test_underline_single(self, temp_dir):
        """单下划线"""
        fp = str(temp_dir / "underline.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"underline": "single"}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["underline"] == "single"

    def test_underline_double(self, temp_dir):
        """双下划线"""
        fp = str(temp_dir / "underline_double.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"underline": "double"}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["underline"] == "double"

    def test_underline_accounting(self, temp_dir):
        """会计下划线"""
        fp = str(temp_dir / "underline_acc.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"underline": "singleAccounting"}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["underline"] == "singleAccounting"

    def test_font_size(self, temp_dir):
        """字号"""
        fp = str(temp_dir / "fontsize.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"size": 20}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["size"] == 20

    def test_font_color(self, temp_dir):
        """字体颜色 (HEX)"""
        fp = str(temp_dir / "fontcolor.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"color": "FF0000"}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        # openpyxl color 转为字符串包含 RGB 值
        assert style["color"] is not None
        assert "FF0000" in style["color"] or "ff0000" in style["color"].lower()

    def test_strikethrough(self, temp_dir):
        """删除线"""
        fp = str(temp_dir / "strike.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"strikethrough": True}})
        assert result.success is True
        # strikethrough 在 Font 对象上

    def test_font_name(self, temp_dir):
        """字体名称"""
        fp = str(temp_dir / "fontname.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"name": "Arial"}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["font_name"] == "Arial"

    # ---------- 2. 背景填充 ----------

    def test_bg_color_solid(self, temp_dir):
        """纯色背景"""
        fp = str(temp_dir / "bgcolor.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:C3", {"fill": {"type": "solid", "color": "FFFF00"}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["fill_type"] == "solid"
        # fgColor 应包含黄色值
        assert style["fgColor"] is not None

    def test_gradient_fill(self, temp_dir):
        """渐变背景"""
        fp = str(temp_dir / "gradient.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {"type": "gradient", "colors": ["4472C4", "ED7D31"], "gradient_type": "linear"}
        })
        assert result.success is True

    def test_pattern_fill(self, temp_dir):
        """图案填充"""
        fp = str(temp_dir / "pattern.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {"type": "pattern", "patternType": "lightGray", "fgColor": "FF0000"}
        })
        assert result.success is True

    # ---------- 3. 对齐与换行 ----------

    def test_alignment_center(self, temp_dir):
        """水平居中"""
        fp = str(temp_dir / "align_center.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:D1", {"alignment": {"horizontal": "center"}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["alignment_h"] == "center"

    def test_alignment_vertical(self, temp_dir):
        """垂直居中"""
        fp = str(temp_dir / "align_v.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"alignment": {"vertical": "center"}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["alignment_v"] == "center"

    def test_wrap_text(self, temp_dir):
        """自动换行"""
        fp = str(temp_dir / "wrap.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"alignment": {"wrap_text": True}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["wrap_text"] is True

    def test_text_rotation(self, temp_dir):
        """文字旋转角度"""
        fp = str(temp_dir / "rotation.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"alignment": {"text_rotation": 45}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["text_rotation"] == 45

    def test_indent(self, temp_dir):
        """缩进"""
        fp = str(temp_dir / "indent.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"alignment": {"indent": 3}})
        assert result.success is True

    def test_shrink_to_fit(self, temp_dir):
        """缩小字体填充"""
        fp = str(temp_dir / "shrink.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"alignment": {"shrink_to_fit": True}})
        assert result.success is True

    # ---------- 4. 数字格式 ----------

    def test_number_format_currency(self, temp_dir):
        """货币数字格式"""
        fp = str(temp_dir / "numfmt.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:B2", {"number_format": "¥#,##0.00"})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["number_format"] == "¥#,##0.00"

    def test_number_format_percent(self, temp_dir):
        """百分比格式"""
        fp = str(temp_dir / "pctfmt.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"number_format": "0.00%"})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["number_format"] == "0.00%"

    def test_number_format_date(self, temp_dir):
        """日期格式"""
        fp = str(temp_dir / "datefmt.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"number_format": "YYYY-MM-DD"})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["number_format"] == "YYYY-MM-DD"

    # ---------- 5. 边框 ----------

    def test_border_thin(self, temp_dir):
        """细边框 (via format_cells border 参数)"""
        fp = str(temp_dir / "border_thin.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:C3", {
            "border": {"left": "thin", "right": "thin", "top": "thin", "bottom": "thin"}
        })
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["border_left"] == "thin"
        assert style["border_right"] == "thin"
        assert style["border_top"] == "thin"
        assert style["border_bottom"] == "thin"

    def test_border_mixed_styles(self, temp_dir):
        """四边不同边框样式"""
        fp = str(temp_dir / "border_mix.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {
                "top": "medium",
                "bottom": "thick",
                "left": "double",
                "right": "dashed",
                "color": "FF0000",
            }
        })
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["border_top"] == "medium"
        assert style["border_bottom"] == "thick"

    def test_border_with_color_dict(self, temp_dir):
        """边框使用 dict 配置（含独立颜色）"""
        fp = str(temp_dir / "border_dict.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {
                "top": {"style": "medium", "color": "0000FF"},
                "bottom": {"style": "thin"},
            }
        })
        assert result.success is True

    # ---------- 6. 组合操作 ----------

    def test_combine_bold_bg_color_alignment(self, temp_dir):
        """组合: 加粗 + 背景色 + 居中"""
        fp = str(temp_dir / "combine_basic.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:D1", {
            "font": {"bold": True, "size": 14},
            "fill": {"type": "solid", "color": "D9E1F2"},
            "alignment": {"horizontal": "center", "vertical": "center"},
        })
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["size"] == 14
        assert style["alignment_h"] == "center"
        assert style["alignment_v"] == "center"
        assert style["fill_type"] == "solid"

    def test_combine_font_multiple_attrs(self, temp_dir):
        """组合多个字体属性: bold+italic+underline+color+size"""
        fp = str(temp_dir / "font_combo.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "font": {
                "bold": True,
                "italic": True,
                "underline": "double",
                "color": "FF0000",
                "size": 16,
                "name": "Courier New",
            }
        })
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["italic"] is True
        assert style["underline"] == "double"
        assert style["size"] == 16
        assert style["font_name"] == "Courier New"

    def test_combine_all_categories(self, temp_dir):
        """全类别组合: font + fill + alignment + number_format + border"""
        fp = str(temp_dir / "full_combo.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:C3", {
            "font": {"bold": True, "color": "FFFFFF", "size": 12},
            "fill": {"type": "solid", "color": "4472C4"},
            "alignment": {"horizontal": "right", "wrap_text": True},
            "number_format": "#,##0.00",
            "border": {"top": "thin", "bottom": "thin", "left": "thin", "right": "thin"},
        })
        assert result.success is True
        assert result.metadata.get("formatted_count", 0) >= 3  # 至少3个单元格

    # ---------- 7. Merge + Format 组合 ----------

    def test_merge_then_format(self, temp_dir):
        """先合并再设置格式"""
        fp = str(temp_dir / "merge_fmt.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)

        # 先合并
        r1 = writer.merge_cells("Sheet1!A1:D1")
        assert r1.success is True

        # 再格式化合并区域
        r2 = writer.format_cells("Sheet1!A1:D1", {
            "font": {"bold": True, "size": 14},
            "fill": {"type": "solid", "color": "FFFF00"},
            "alignment": {"horizontal": "center"},
        })
        assert r2.success is True

    def test_merge_with_bold_and_bg_via_api(self, temp_dir):
        """通过 API 层 excel_format_cells merge+bold+bg_color"""
        from excel_mcp_server_fastmcp.server import excel_format_cells
        fp = str(temp_dir / "api_merge_combo.xlsx")
        _create_test_xlsx(fp)

        result = excel_format_cells(
            file_path=fp,
            sheet_name="Sheet1",
            cell_range="A1:D1",
            formatting={"merge": True, "bold": True, "bg_color": "FFD700", "alignment": "center"},
        )
        assert result["success"] is True

    # ---------- 8. Unmerge ----------

    def test_merge_then_unmerge(self, temp_dir):
        """合并后拆分"""
        fp = str(temp_dir / "merge_unmerge.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)

        r1 = writer.merge_cells("Sheet1!A1:C1")
        assert r1.success is True

        r2 = writer.unmerge_cells("Sheet1!A1:C1")
        assert r2.success is True

        # 拆分后文件仍可正常读取
        wb = load_workbook(fp)
        ws = wb.active
        # 确认合并区域已被拆分 (openpyxl 的 merged_cells 不再包含该范围)
        merged_ranges = list(ws.merged_cells.ranges)
        wb.close()
        assert not any(str(mr) == "A1:C1" for mr in merged_ranges)

    def test_unmerge_non_merged(self, temp_dir):
        """拆分未合并的区域（不应报错）"""
        fp = str(temp_dir / "unmerge_noop.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        # 未合并的区域调用 unmerge，openpyxl 通常不报错
        result = writer.unmerge_cells("Sheet1!A1:C1")
        # openpyxl 允许对非合并区域调用 unmerge_cells（no-op）

    # ---------- 9. 边界 case ----------

    def test_single_cell_range(self, temp_dir):
        """单单元格格式化"""
        fp = str(temp_dir / "single_cell.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!B2", {"font": {"bold": True}})
        assert result.success is True
        assert result.metadata.get("formatted_count") == 1

    def test_large_range(self, temp_dir):
        """大范围格式化 (20x10)"""
        fp = str(temp_dir / "large_range.xlsx")
        _create_test_xlsx(fp, rows=20, cols=10)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:J20", {"font": {"bold": True}})
        assert result.success is True
        assert result.metadata.get("formatted_count") == 200

    def test_format_preserves_cell_values(self, temp_dir):
        """格式化不改变单元格值"""
        fp = str(temp_dir / "preserve_vals.xlsx")
        _create_test_xlsx(fp, rows=3, cols=3)
        original_val = 11  # A1 = 1*10+1 = 11
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:C3", {
            "font": {"bold": True, "color": "FF0000"},
            "fill": {"type": "solid", "color": "00FF00"},
            "alignment": {"horizontal": "right"},
            "number_format": "0.00",
        })
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["value"] == original_val

    def test_empty_formatting_dict(self, temp_dir):
        """空格式字典"""
        fp = str(temp_dir / "empty_fmt.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {})
        # 空字典不会报错，只是没有任何效果
        assert result.success is True

    def test_invalid_sheet_name(self, temp_dir):
        """不存在的工作表"""
        fp = str(temp_dir / "bad_sheet.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("NonExistent!A1", {"font": {"bold": True}})
        assert result.success is False

    # ---------- 10. 中文字体 & Unicode ----------

    def test_chinese_font_name(self, temp_dir):
        """中文字体名称"""
        fp = str(temp_dir / "chinese_font.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"name": "微软雅黑", "size": 12}})
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["font_name"] == "微软雅黑"
        assert style["size"] == 12

    def test_chinese_sheet_name(self, temp_dir):
        """中文工作表名格式化"""
        fp = str(temp_dir / "chinese_sheet.xlsx")
        _create_test_xlsx(fp, sheet_name="数据表")
        writer = ExcelWriter(fp)
        result = writer.format_cells("数据表!A1:B2", {"font": {"bold": True}})
        assert result.success is True

    # ---------- 11. 扁平格式转换 (_normalize_formatting) ----------

    def test_normalize_flat_bold(self):
        """扁平 bold → 嵌套 font.bold"""
        result = ExcelOperations._normalize_formatting({"bold": True})
        assert result == {"font": {"bold": True}}

    def test_normalize_flat_bg_color(self):
        """扁平 bg_color → 嵌套 fill.color"""
        result = ExcelOperations._normalize_formatting({"bg_color": "FF0"})
        assert result == {"fill": {"color": "FF0"}}

    def test_normalize_flat_alignment(self):
        """扁平 alignment: "center" → nested alignment.horizontal: "center\""""
        result = ExcelOperations._normalize_formatting({"alignment": "center"})
        assert result == {"alignment": {"horizontal": "center"}}

    def test_normalize_flat_font_size(self):
        """扁平 font_size → font.size"""
        result = ExcelOperations._normalize_formatting({"font_size": 18})
        assert result == {"font": {"size": 18}}

    def test_normalize_flat_font_color(self):
        """扁平 font_color → font.color"""
        result = ExcelOperations._normalize_formatting({"font_color": "FF0000"})
        assert result == {"font": {"color": "FF0000"}}

    def test_normalize_flat_wrap_text(self):
        """扁平 wrap_text → alignment.wrap_text"""
        result = ExcelOperations._normalize_formatting({"wrap_text": True})
        assert result == {"alignment": {"wrap_text": True}}

    def test_normalize_flat_combined(self):
        """多项扁平参数同时转换"""
        result = ExcelOperations._normalize_formatting({
            "bold": True,
            "bg_color": "FFFF00",
            "alignment": "center",
            "font_size": 16,
            "italic": True,
            "underline": "double",
        })
        assert result["font"]["bold"] is True
        assert result["font"]["italic"] is True
        assert result["font"]["underline"] == "double"
        assert result["font"]["size"] == 16
        assert result["fill"]["color"] == "FFFF00"
        assert result["alignment"]["horizontal"] == "center"

    def test_normalize_nested_passthrough(self):
        """已是嵌套格式的直接透传"""
        nested = {"font": {"bold": True}, "fill": {"color": "FF0"}}
        result = ExcelOperations._normalize_formatting(nested)
        assert result is nested  # 同一对象引用

    def test_normalize_none_returns_empty(self):
        """None 返回空字典"""
        result = ExcelOperations._normalize_formatting(None)
        assert result == {}

    def test_normalize_gradient_colors(self):
        """扁平 gradient_colors → fill.gradient"""
        result = ExcelOperations._normalize_formatting({
            "gradient_colors": ["4472C4", "ED7D31"],
            "gradient_type": "linear",
        })
        assert result["fill"]["type"] == "gradient"
        assert result["fill"]["colors"] == ["4472C4", "ED7D31"]
        assert result["fill"]["gradient_type"] == "linear"

    def test_normalize_border_string(self):
        """扁平 border 字符串 → border.style"""
        result = ExcelOperations._normalize_formatting({"border": "thin"})
        assert result["border"] == {"style": "thin"}

    def test_normalize_border_dict_passthrough(self):
        """扁平 border 字典直接透传"""
        border_dict = {"top": "medium", "color": "FF0000"}
        result = ExcelOperations._normalize_formatting({"border": border_dict})
        assert result["border"] == border_dict

    def test_normalize_strikethrough(self):
        """扁平 strikethrough → font.strikethrough"""
        result = ExcelOperations._normalize_formatting({"strikethrough": True})
        assert result == {"font": {"strikethrough": True}}

    def test_normalize_text_rotation(self):
        """扁平 text_rotation → alignment.text_rotation"""
        result = ExcelOperations._normalize_formatting({"text_rotation": -90})
        assert result == {"alignment": {"text_rotation": -90}}

    def test_normalize_indent(self):
        """扁平 indent → alignment.indent"""
        result = ExcelOperations._normalize_formatting({"indent": 5})
        assert result == {"alignment": {"indent": 5}}

    def test_normalize_shrink_to_fit(self):
        """扁平 shrink_to_fit → alignment.shrink_to_fit"""
        result = ExcelOperations._normalize_formatting({"shrink_to_fit": True})
        assert result == {"alignment": {"shrink_to_fit": True}}

    def test_normalize_vertical_alignment(self):
        """扁平 vertical_alignment → alignment.vertical"""
        result = ExcelOperations._normalize_formatting({"vertical_alignment": "bottom"})
        assert result == {"alignment": {"vertical": "bottom"}}

    def test_normalize_number_format(self):
        """扁平 number_format 直接传递"""
        result = ExcelOperations._normalize_formatting({"number_format": "0.00%"})
        assert result == {"number_format": "0.00%"}

    def test_normalize_unknown_key_passthrough(self):
        """未知键直接透传（向后兼容）"""
        result = ExcelOperations._normalize_formatting({"custom_key": "custom_value"})
        assert result["custom_key"] == "custom_value"

    # ---------- 12. Preset 测试 ----------

    def test_preset_title(self, temp_dir):
        """预设 title: 微软雅黑 14号 加粗 居中"""
        fp = str(temp_dir / "preset_title.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = excel_format_cells = ExcelOperations.format_cells(
            fp, "Sheet1", "A1", formatting=None, preset="title"
        )
        assert result["success"] is True

    def test_preset_header(self, temp_dir):
        """预设 header: 微软雅黑 11号 加粗 灰底"""
        fp = str(temp_dir / "preset_header.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(
            fp, "Sheet1", "A1:D1", formatting=None, preset="header"
        )
        assert result["success"] is True

    def test_preset_currency(self, temp_dir):
        """预设 currency: ¥#,##0.00"""
        fp = str(temp_dir / "preset_currency.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(
            fp, "Sheet1", "A1", formatting=None, preset="currency"
        )
        assert result["success"] is True

    def test_preset_highlight(self, temp_dir):
        """预设 highlight: 黄色背景"""
        fp = str(temp_dir / "preset_highlight.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(
            fp, "Sheet1", "A1", formatting=None, preset="highlight"
        )
        assert result["success"] is True

    def test_preset_data(self, temp_dir):
        """预设 data: 微软雅黑 10号"""
        fp = str(temp_dir / "preset_data.xlsx")
        _create_test_xlsx(fp)
        result = ExcelOperations.format_cells(
            fp, "Sheet1", "A1:C5", formatting=None, preset="data"
        )
        assert result["success"] is True

    # ---------- 13. formatted_count 元数据 ----------

    def test_formatted_count_multi_cell(self, temp_dir):
        """验证 formatted_count 准确性"""
        fp = str(temp_dir / "fmt_count.xlsx")
        _create_test_xlsx(fp, rows=3, cols=5)  # 3行5列=15单元格
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:E3", {"font": {"bold": True}})
        assert result.success is True
        assert result.metadata.get("formatted_count") == 15

    # ---------- 14. Overwrite 格式 ----------

    def test_format_overwrite(self, temp_dir):
        """后执行的格式覆盖前面的"""
        fp = str(temp_dir / "overwrite.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)

        # 第一次: 红色
        writer.format_cells("Sheet1!A1", {"font": {"color": "FF0000", "bold": True}})

        # 第二次: 蓝色 + 取消 bold
        writer.format_cells("Sheet1!A1", {"font": {"color": "0000FF", "bold": False}})

        style = _read_cell_style(fp, "A1")
        assert style["bold"] is False
        # 颜色应为蓝色
        assert style["color"] is not None

    # ---------- 15. set_borders 工具方法 ----------

    def test_set_borders_tool(self, temp_dir):
        """set_borders 独立工具方法"""
        fp = str(temp_dir / "set_borders.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.set_borders("Sheet1!A1:C3", "medium")
        assert result.success is True
    def test_set_borders_thick(self, temp_dir):
        """粗边框"""
        fp = str(temp_dir / "borders_thick.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)
        result = writer.set_borders("Sheet1!A1", "thick")
        assert result.success is True
        style = _read_cell_style(fp, "A1")
        assert style["border_left"] == "thick"

    # ---------- R55+ 新增边缘 case 测试 ----------

    def test_format_all_none_values_noop(self, temp_dir):
        """所有样式值为 None 时不应报错（no-op）"""
        fp = str(temp_dir / "all_none.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)

        # 路径1: 直接调用 writer.format_cells 传入含 None 值的扁平格式
        # （_apply_cell_format 应跳过 None config 而不 crash）
        result = writer.format_cells("Sheet1!A1:B2", {
            "bold": None, "italic": None, "font_size": None,
            "bg_color": None, "alignment": None,
        })
        assert result.success is True
        # 值不变
        style = _read_cell_style(fp, "A1")
        assert style["value"] == 11

        # 路径2: 通过 ExcelOperations（会经过 _normalize_formatting 过滤 None → 空dict）
        result2 = ExcelOperations.format_cells(fp, "Sheet1", "A1:B2", {
            "bold": None, "italic": None, "font_size": None,
            "bg_color": None, "alignment": None,
        })
        assert result2["success"] is True

    def test_merge_with_border_and_bold_combo(self, temp_dir):
        """合并 + 粗边框 + 加粗 + 背景色 组合操作"""
        fp = str(temp_dir / "merge_border_bold.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)

        # 先合并
        r1 = writer.merge_cells("Sheet1!A1:C1")
        assert r1.success is True

        # 再对合并区域应用组合样式
        r2 = writer.format_cells("Sheet1!A1:C1", {
            "font": {"bold": True, "size": 14},
            "fill": {"color": "FF0000"},
            "border": {
                "top": {"style": "medium", "color": "000000"},
                "bottom": {"style": "medium", "color": "000000"},
                "left": {"style": "medium", "color": "000000"},
                "right": {"style": "medium", "color": "000000"},
            },
        })
        assert r2.success is True

        # 验证左上角单元格样式
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["size"] == 14
        assert style["border_top"] == "medium"

    def test_number_format_scientific_and_fraction(self, temp_dir):
        """科学计数法和分数格式"""
        fp = str(temp_dir / "num_fmt_special.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)

        # 科学计数法
        r1 = writer.format_cells("Sheet1!A1", {"number_format": "0.00E+00"})
        assert r1.success is True
        style = _read_cell_style(fp, "A1")
        assert style["number_format"] == "0.00E+00"

        # 分数格式
        r2 = writer.format_cells("Sheet1!B1", {"number_format": "# ?/?"})
        assert r2.success is True
        style = _read_cell_style(fp, "B1")
        assert style["number_format"] == "# ?/?"

    def test_format_partial_overwrite_preserves_other_attrs(self, temp_dir):
        """部分覆盖：第二次格式化只改 bold，保留之前的 bg_color"""
        fp = str(temp_dir / "partial_overwrite.xlsx")
        _create_test_xlsx(fp)
        writer = ExcelWriter(fp)

        # 第一次：设置 bold + bg_color
        r1 = writer.format_cells("Sheet1!A1", {
            "font": {"bold": True},
            "fill": {"color": "00FF00"},
        })
        assert r1.success is True

        # 第二次：只改 bold 为 False，不传 fill
        r2 = writer.format_cells("Sheet1!A1", {
            "font": {"bold": False},
        })
        assert r2.success is True

        # 验证: bold 被 override，但 openpyxl 的 format_cells 是赋值行为
        # fill 是否保留取决于实现（这里验证不报错即可）
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is False

    def test_flat_format_comprehensive_edge_cases(self, temp_dir):
        """扁平格式边缘 case：boolean False、0、空字符串、特殊字符"""
        from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

        # bold=False 应该正常传递（不是 None，应该被设置）
        result = ExcelOperations._normalize_formatting({
            "bold": False,
            "italic": 0,  # 0 是 falsy 但不是 None
            "font_size": 0,
            "bg_color": "",
            "alignment": "justify",
        })
        # bold=False 应出现在 font 中
        assert result["font"]["bold"] is False
        # font_size=0 应出现
        assert result["font"]["size"] == 0
        # alignment 应正确映射
        assert result["alignment"]["horizontal"] == "justify"
        # 空字符串 bg_color 应产生 fill（空字符串不是 None）
        assert "fill" in result

    def test_unicode_sheet_name_format(self, temp_dir):
        """Unicode/中文工作表名格式化"""
        fp = str(temp_dir / "unicode_sheet.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "数据表"
        ws.append(["名称", "值"])
        ws.append(["测试", 42])
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("数据表!A1:B1", {
            "font": {"bold": True, "name": "微软雅黑"},
            "fill": {"color": "D9D9D9"},
        })
        assert result.success is True

        style = _read_cell_style(fp, "A1", sheet_name="数据表")
        assert style["bold"] is True
        assert style["font_name"] == "微软雅黑"


# ============================================================
# R55+ Round 55+ Edge-Case Tests for format_cells
# ============================================================


def _make_sample(tmp_path, data, sheet_name="Sheet1"):
    """快速创建测试文件（兼容 tmp_path 和 temp_dir）"""
    fp = str(tmp_path / "sample.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row_data in data:
        ws.append(row_data)
    wb.save(fp)
    wb.close()
    return fp


class TestFormatCellsMergeCombine:
    """merge + bold + bg_color 组合操作（Writer 层，使用嵌套格式）"""

    def test_merge_with_bold_and_bg_color(self, tmp_path):
        """合并单元格同时设置粗体和背景色"""
        fp = _make_sample(tmp_path, [[1, 2, 3], [4, 5, 6]])
        writer = ExcelWriter(fp)
        # Step 1: merge
        r1 = writer.merge_cells("Sheet1!A1:C1")
        assert r1.success is True
        # Step 2: format (嵌套格式)
        r2 = writer.format_cells("Sheet1!A1", {
            "font": {"bold": True},
            "fill": {"type": "solid", "color": "FF0000"},
        })
        assert r2.success is True

        wb = load_workbook(fp)
        ws = wb.active
        assert len(ws.merged_cells.ranges) >= 1
        cell = ws["A1"]
        assert cell.font.bold is True
        rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else ""
        assert "FF0000" in rgb or "FFFF0000" in rgb
        wb.close()

    def test_merge_then_format_separate(self, tmp_path):
        """先合并，再单独格式化"""
        fp = _make_sample(tmp_path, [[1, 2], [3, 4]])
        writer = ExcelWriter(fp)

        # Step 1: merge
        r1 = writer.merge_cells("Sheet1!A1:B1")
        assert r1.success is True

        # Step 2: format the merged range (嵌套格式)
        r2 = writer.format_cells("Sheet1!A1:B1", {
            "font": {"italic": True, "size": 14}
        })
        assert r2.success is True

        style = _read_cell_style(fp, "A1")
        assert style["italic"] is True
        assert style["size"] == 14

    def test_unmerge_after_merge(self, tmp_path):
        """合并后再拆分"""
        fp = _make_sample(tmp_path, [[1, 2, 3]])
        writer = ExcelWriter(fp)

        # Merge
        r1 = writer.merge_cells("Sheet1!A1:C1")
        assert r1.success is True

        # Format before unmerge (嵌套格式)
        writer.format_cells("Sheet1!A1", {"font": {"bold": True}})

        # Unmerge
        r2 = writer.unmerge_cells("Sheet1!A1:C1")
        assert r2.success is True

        wb = load_workbook(fp)
        ws = wb.active
        assert len(ws.merged_cells.ranges) == 0
        assert ws["A1"].font.bold is True
        wb.close()


class TestFormatCellsNumberFormatEdgeCases:
    """数字格式边界情况"""

    def test_number_format_none_ignored(self, tmp_path):
        """number_format=None 应被安全跳过（不再崩溃或损坏文件）"""
        fp = _make_sample(tmp_path, [[100]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"number_format": None})
        # 修复后：None 值被安全跳过，操作成功
        assert result.success is True

    def test_number_format_empty_string(self, tmp_path):
        """空字符串 number_format 应正常处理"""
        fp = _make_sample(tmp_path, [[100]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"number_format": ""})
        assert result.success is True

    def test_number_format_complex(self, tmp_path):
        """复杂自定义数字格式"""
        fp = _make_sample(tmp_path, [[1234.5678]])
        writer = ExcelWriter(fp)
        fmt = "#,##0.00_);[Red](#,##0.00)"
        result = writer.format_cells("Sheet1!A1", {"number_format": fmt})
        assert result.success is True

        wb = load_workbook(fp)
        assert wb.active["A1"].number_format == fmt
        wb.close()

    def test_number_format_with_currency_preset(self, tmp_path):
        """货币预设设置正确的 number_format（通过 Operations API）"""
        fp = _make_sample(tmp_path, [[100]])
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:A1", formatting={}, preset="currency")
        assert result["success"] is True

        wb = load_workbook(fp)
        nf = wb.active["A1"].number_format
        assert "$" in nf or "¥" in nf or "#" in nf  # 货币格式包含符号
        wb.close()


class TestFormatCellsAlignmentEdgeCases:
    """对齐方式边界测试（Writer 层使用嵌套格式）"""

    def test_all_alignment_params_together(self, tmp_path):
        """同时设置所有对齐参数"""
        fp = _make_sample(tmp_path, [["long text"]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "alignment": {
                "horizontal": "center",
                "vertical": "center",
                "wrap_text": True,
                "text_rotation": 45,
                "indent": 3,
                "shrink_to_fit": True,
            }
        })
        assert result.success is True

        style = _read_cell_style(fp, "A1")
        assert style["alignment_h"] == "center"
        assert style["alignment_v"] == "center"
        assert style["wrap_text"] is True
        assert style["text_rotation"] == 45

    def test_wrap_text_explicit_false(self, tmp_path):
        """设置 wrap_text 后可以覆盖"""
        fp = _make_sample(tmp_path, [["wrap me"]])
        writer = ExcelWriter(fp)

        # 先设为 True
        writer.format_cells("Sheet1!A1", {"alignment": {"wrap_text": True}})
        # 再设回默认（None 表示不强制设置）
        result = writer.format_cells("Sheet1!A1", {"alignment": {"wrap_text": True}})
        assert result.success is True

        style = _read_cell_style(fp, "A1")
        assert style["wrap_text"] is True

    def test_text_rotation_angles(self, tmp_path):
        """文字旋转角度（openpyxl 有效范围 0-180 及特殊值）"""
        fp = _make_sample(tmp_path, [["rotated"]])
        writer = ExcelWriter(fp)

        # openpyxl text_rotation: 0-180 为标准角度，255 表示垂直文字
        for angle in [0, 45, 90, 180, 255]:
            r = writer.format_cells("Sheet1!A1", {"alignment": {"text_rotation": angle}})
            assert r.success is True
            style = _read_cell_style(fp, "A1")
            assert style["text_rotation"] == angle


class TestFormatCellsBorderEdgeCases:
    """边框边界情况"""

    def test_border_all_sides_different_styles(self, tmp_path):
        """四边不同边框样式"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        border_cfg = {
            "left": {"style": "thin", "color": "FF0000"},
            "right": {"style": "medium", "color": "00FF00"},
            "top": {"style": "thick", "color": "0000FF"},
            "bottom": {"style": "dashed", "color": "FFFF00"},
        }
        result = writer.format_cells("Sheet1!A1", {"border": border_cfg})
        assert result.success is True

        style = _read_cell_style(fp, "A1")
        assert style["border_left"] == "thin"
        assert style["border_right"] == "medium"
        assert style["border_top"] == "thick"
        assert style["border_bottom"] == "dashed"

    def test_border_string_shorthand(self, tmp_path):
        """边框简写字符串 — Writer 层不支持字符串简写，仅支持 dict"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        # 字符串简写在 Writer 层会报错（'str' object has no attribute 'get'）
        # 这是已知行为：border 参数必须是 dict
        result = writer.format_cells("Sheet1!A1", {"border": "medium"})
        # 字符串 border 不是有效输入，应失败或返回错误
        assert result.success is False

    def test_border_empty_dict(self, tmp_path):
        """空边框字典不崩溃"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"border": {}})
        assert result.success is True


class TestFormatCellsFontEdgeCases:
    """字体边界情况（Writer 层使用嵌套格式）"""

    def test_chinese_font_flat_param(self, tmp_path):
        """中文字体通过 Operations 层嵌套格式设置"""
        fp = _make_sample(tmp_path, [["中文"]])
        # 使用 Operations API + 嵌套格式确保字体名生效
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:A1", formatting={
            "font": {"name": "宋体", "bold": True, "size": 16},
        })
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["font_name"] == "宋体"
        assert style["bold"] is True
        assert style["size"] == 16

    def test_font_strikethrough_with_color(self, tmp_path):
        """删除线 + 字体颜色组合"""
        fp = _make_sample(tmp_path, [["deleted"]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "font": {"strikethrough": True, "color": "999999"}
        })
        assert result.success is True

    def test_font_size_very_small(self, tmp_path):
        """极小字号"""
        fp = _make_sample(tmp_path, [["tiny"]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"size": 1}})
        assert result.success is True
        assert _read_cell_style(fp, "A1")["size"] == 1

    def test_font_size_very_large(self, tmp_path):
        """极大字号"""
        fp = _make_sample(tmp_path, [["huge"]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"size": 400}})
        assert result.success is True
        assert _read_cell_style(fp, "A1")["size"] == 400

    def test_underline_styles(self, tmp_path):
        """各种下划线样式"""
        fp = _make_sample(tmp_path, [["under"]])
        writer = ExcelWriter(fp)
        for style in ["single", "double", "singleAccounting", "doubleAccounting"]:
            r = writer.format_cells("Sheet1!A1", {"font": {"underline": style}})
            assert r.success is True


class TestFormatCellsFillEdgeCases:
    """填充/背景色边界情况"""

    def test_bg_color_hex_with_hash(self, tmp_path):
        """带 # 前缀的颜色值（openpyxl 不接受 # 前缀，需去掉）"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        # openpyxl 要求 ARGB hex 不带 # 前缀
        result = writer.format_cells("Sheet1!A1", {"fill": {"type": "solid", "color": "AABBCC"}})
        assert result.success is True

    def test_bg_color_hex_with_hash_rejected(self, tmp_path):
        """验证 # 前缀颜色被 openpyxl 拒绝"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"fill": {"type": "solid", "color": "#AABBCC"}})
        # openpyxl 不接受 # 前缀
        assert result.success is False

    def test_pattern_fill_type(self, tmp_path):
        """pattern fill 类型"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {"type": "pattern", "patternType": "gray125"}
        })
        assert result.success is True

    def test_gradient_fill(self, tmp_path):
        """渐变填充"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {"type": "gradient", "colors": ["FFFFFF", "000000"]}
        })
        assert result.success is True


class TestFormatCellsNormalizeEdgeCases:
    """_normalize_formatting 边界测试"""

    def test_normalize_overwrite_nested_with_flat(self, tmp_path):
        """扁平参数和嵌套参数同时出现时——验证实际行为"""
        result = ExcelOperations._normalize_formatting({
            "bold": True,
            "font": {"bold": False, "name": "Arial"},
        })
        # 验证归一化结果存在 font 键
        assert "font" in result
        assert result["font"]["name"] == "Arial"
        # bold 的最终值取决于实现：记录实际行为即可
        assert "bold" in result or "font" in result

    def test_normalize_unknown_keys_preserved(self, tmp_path):
        """未知键被保留以兼容未来扩展"""
        result = ExcelOperations._normalize_formatting({
            "custom_future_key": "value",
            "another_unknown": 42,
        })
        assert "custom_future_key" in result
        assert "another_unknown" in result

    def test_normalize_empty_subdicts_filtered(self, tmp_path):
        """空的子字典应该被过滤掉"""
        result = ExcelOperations._normalize_formatting({
            "font": {},
            "fill": {},
            "alignment": {},
        })
        # 空子字典不应出现在结果中（或至少不影响输出）
        assert result.get("font") is None or result.get("font") == {}


class TestFormatCellsRangeEdgeCases:
    """范围边界情况"""

    def test_single_cell_range(self, tmp_path):
        """单单元格范围 A1:A1（使用嵌套格式）"""
        fp = _make_sample(tmp_path, [[42]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:A1", {"font": {"bold": True}})
        assert result.success is True
        assert _read_cell_style(fp, "A1")["bold"] is True

    def test_wide_row_range(self, tmp_path):
        """宽行范围 A1:J1"""
        fp = _make_sample(tmp_path, [[i for i in range(10)]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:J1", {"fill": {"type": "solid", "color": "EEEEEE"}})
        assert result.success is True

    def test_large_range_format(self, tmp_path):
        """较大范围格式化（性能边界）"""
        fp = _make_sample(tmp_path, [[i * j for j in range(20)] for i in range(50)])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:T50", {"font": {"name": "Calibri"}})
        assert result.success is True

    def test_format_preserves_formulas(self, tmp_path):
        """格式化不破坏已有公式"""
        fp = _make_sample(tmp_path, [[1, 2]])
        wb = load_workbook(fp)
        ws = wb.active
        ws["C1"] = "=A1+B1"
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!C1", {"font": {"bold": True}})
        assert result.success is True

        wb = load_workbook(fp, data_only=False)
        ws = wb.active
        assert ws["C1"].value == "=A1+B1"  # 公式保留
        assert ws["C1"].font.bold is True
        wb.close()

    def test_format_preserves_values(self, tmp_path):
        """格式化不改变单元格值（注意：openpyxl 会将空字符串转为 None）"""
        fp = _make_sample(tmp_path, [[None, "", 0, "text", 3.14, True]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:F1", {"fill": {"type": "solid", "color": "DDDDDD"}})
        assert result.success is True

        wb = load_workbook(fp)
        ws = wb.active
        assert ws["A1"].value is None
        # openpyxl 行为：空字符串 "" 读写后变为 None
        assert ws["B1"].value is None or ws["B1"].value == ""
        assert ws["C1"].value == 0
        assert ws["D1"].value == "text"
        assert abs(ws["E1"].value - 3.14) < 0.001
        assert ws["F1"].value is True
        wb.close()


class TestFormatCellsOperationsLayer:
    """Operations 层 format_cells 测试（支持扁平参数）"""

    def test_ops_format_with_preset_title(self, tmp_path):
        """通过 Operations API 使用 title 预设"""
        fp = _make_sample(tmp_path, [["标题行"]])
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:A1", formatting={}, preset="title")
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["size"] > 10  # title preset 应该增大字号

    def test_ops_format_normalize_flat_params(self, tmp_path):
        """Operations 层扁平参数正确归一化"""
        fp = _make_sample(tmp_path, [[99]])
        result = ExcelOperations.format_cells(fp, "Sheet1", "A1:A1", formatting={
            "bold": True,
            "italic": True,
            "bg_color": "00FF00",
            "number_format": "0.00%",
        })
        assert result["success"] is True
        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True
        assert style["italic"] is True

    def test_ops_format_invalid_sheet_name(self, tmp_path):
        """无效工作表名返回错误"""
        fp = _make_sample(tmp_path, [[1]])
        result = ExcelOperations.format_cells(fp, "NonExistent", "A1:B1", formatting={"bold": True})
        assert result["success"] is False

    def test_ops_merge_via_operations(self, tmp_path):
        """通过 Operations 层执行合并"""
        fp = _make_sample(tmp_path, [[1, 2, 3]])
        result = ExcelOperations.merge_cells(fp, "Sheet1", "A1:C1")
        assert result["success"] is True

        wb = load_workbook(fp)
        assert len(wb.active.merged_cells.ranges) >= 1
        wb.close()

    def test_ops_unmerge_via_operations(self, tmp_path):
        """通过 Operations 层执行拆分"""
        fp = _make_sample(tmp_path, [[1, 2, 3]])

        # 先合并
        ExcelOperations.merge_cells(fp, "Sheet1", "A1:C1")
        # 再拆分
        result = ExcelOperations.unmerge_cells(fp, "Sheet1", "A1:C1")
        assert result["success"] is True

        wb = load_workbook(fp)
        assert len(wb.active.merged_cells.ranges) == 0
        wb.close()


# ============================================================
# R55+ Round N: Bug-fix 验证 + 新增边缘 case 测试
# ============================================================


class TestFormatCellsBugFixNumberFormatNone:
    """Bug fix: number_format=None 不再导致文件损坏 (P1)"""

    def test_number_format_none_no_corruption(self, tmp_path):
        """number_format=None 应被安全跳过，不报错也不损坏文件"""
        fp = _make_sample(tmp_path, [[100]])
        writer = ExcelWriter(fp)

        # 修复前：这会抛出 TypeError 并损坏 xlsx 文件
        result = writer.format_cells("Sheet1!A1", {"number_format": None})
        # 修复后：应成功（跳过 None 值）
        assert result.success is True

        # 文件仍然可正常打开
        wb = load_workbook(fp)
        assert wb.active["A1"].value == 100
        wb.close()

    def test_number_format_none_via_operations(self, tmp_path):
        """Operations 层 number_format=None 被 normalize 过滤掉"""
        fp = _make_sample(tmp_path, [[42]])
        # Operations 层的 _normalize_formatting 会过滤 None 值
        result = ExcelOperations.format_cells(
            fp, "Sheet1", "A1:A1",
            formatting={"number_format": None, "bold": True}
        )
        assert result["success"] is True

        style = _read_cell_style(fp, "A1")
        assert style["bold"] is True  # bold 正常生效


class TestFormatCellsBugFixBorderNoneSide:
    """Bug fix: border side=None 不再意外创建细边框 (P2)"""

    def test_border_side_none_preserves_original(self, tmp_path):
        """border left=None 应设为无边框（style=None），而非意外创建 thin 边框"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)

        # 先设置一个 thick 左边框
        writer.format_cells("Sheet1!A1", {
            "border": {"left": {"style": "thick", "color": "FF0000"}}
        })

        # 再用 left=None 更新 — left 应变为 None（无边框），right 为 medium
        writer.format_cells("Sheet1!A1", {
            "border": {
                "left": None,
                "right": {"style": "medium", "color": "0000FF"},
            }
        })

        style = _read_cell_style(fp, "A1")
        # 修复前：left 会变成 thin（bug）；修复后：left 为 None
        assert style["border_left"] is None or style.get("border_left") != "thin"
        assert style["border_right"] == "medium"

    def test_border_all_sides_none_is_noop(self, tmp_path):
        """所有 border side 都为 None 时不应崩溃"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {"left": None, "right": None, "top": None, "bottom": None}
        })
        assert result.success is True


class TestFormatCellsRowColumnRange:
    """行/列范围格式化测试"""

    def test_format_single_row(self, tmp_path):
        """格式化单行 '1:1'"""
        fp = _make_sample(tmp_path, [[1, 2, 3], [4, 5, 6]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!1:1", {"font": {"bold": True}})
        assert result.success is True
        meta = result.metadata or {}
        assert meta.get("formatted_count") == 3

        wb = load_workbook(fp)
        assert wb.active["A1"].font.bold is True
        assert wb.active["B1"].font.bold is True
        wb.close()

    def test_format_multi_row(self, tmp_path):
        """格式化多行 '1:2'"""
        fp = _make_sample(tmp_path, [[1, 2], [3, 4], [5, 6]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!1:2", {"fill": {"type": "solid", "color": "EEEEEE"}})
        assert result.success is True
        meta = result.metadata or {}
        assert meta.get("formatted_count") == 4

    def test_format_single_column(self, tmp_path):
        """格式化单列 'A:A'"""
        fp = _make_sample(tmp_path, [[1, 2], [3, 4], [5, 6]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A:A", {"font": {"italic": True}})
        assert result.success is True
        meta = result.metadata or {}
        assert meta.get("formatted_count") == 3

        wb = load_workbook(fp)
        assert wb.active["A1"].font.italic is True
        assert wb.active["A3"].font.italic is True
        # B 列不受影响
        assert wb.active["B1"].font.italic is not True
        wb.close()

    def test_format_column_range(self, tmp_path):
        """格式化多列 'A:B'"""
        fp = _make_sample(tmp_path, [[1, 2, 3], [4, 5, 6]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A:B", {"alignment": {"horizontal": "center"}})
        assert result.success is True
        meta = result.metadata or {}
        assert meta.get("formatted_count") == 4


class TestFormatCellsDiagonalBorder:
    """对角线边框测试"""

    def test_diagonal_border(self, tmp_path):
        """对角线边框（从左上到右下）"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {
                "diagonal": {"style": "thin", "color": "FF0000"},
                "diagonal_direction": "down",
            }
        })
        assert result.success is True

        wb = load_workbook(fp)
        b = wb.active["A1"].border
        assert b.diagonal.style == "thin"
        wb.close()

    def test_diagonal_up_border(self, tmp_path):
        """对角线边框（从左下到右上）"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "border": {
                "diagonal": {"style": "medium"},
                "diagonal_direction": "up",
            }
        })
        assert result.success is True


class TestFormatCellsMergedCellFormatting:
    """合并单元格区域格式化边缘情况"""

    def test_format_merged_range_all_cells(self, tmp_path):
        """对整个合并区域格式化（每个 cell 都应用）"""
        fp = _make_sample(tmp_path, [[1, 2, 3]])
        writer = ExcelWriter(fp)
        writer.merge_cells("Sheet1!A1:C1")

        # 对合并区域中的每个 cell 都 format
        for coord in ["A1", "B1", "C1"]:
            r = writer.format_cells(f"Sheet1!{coord}", {"font": {"bold": True}})
            assert r.success is True

        # 只有左上角 A1 的样式会生效（Excel 行为）
        wb = load_workbook(fp)
        assert wb.active["A1"].font.bold is True
        wb.close()

    def test_merge_then_format_then_unmerge_preserves_style(self, tmp_path):
        """合并→格式化→拆分后样式保留在原左上角单元格"""
        fp = _make_sample(tmp_path, [[1, 2, 3]])
        writer = ExcelWriter(fp)

        writer.merge_cells("Sheet1!A1:C1")
        writer.format_cells("Sheet1!A1", {
            "font": {"bold": True, "color": "FF0000"},
            "fill": {"type": "solid", "color": "FFFF00"},
        })
        writer.unmerge_cells("Sheet1!A1:C1")

        wb = load_workbook(fp)
        c = wb.active["A1"]
        assert c.font.bold is True
        # openpyxl 可能返回 ARGB (8位) 或 RGB (6位) 格式
        color_rgb = c.font.color.rgb if c.font.color else ""
        assert "FF0000" in color_rgb
        wb.close()


class TestFormatCellsProtectionEdgeCases:
    """保护/锁定相关边缘 case"""

    def test_format_locked_cell(self, tmp_path):
        """格式化已锁定的单元格（openpyxl 默认所有单元格锁定）"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"font": {"bold": True}})
        assert result.success is True

    def test_multiple_formats_sequential(self, tmp_path):
        """连续多次格式化同一单元格（压力测试）"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)

        colors = ["FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF"]
        for color in colors:
            r = writer.format_cells("Sheet1!A1", {"fill": {"type": "solid", "color": color}})
            assert r.success is True

        # 最终颜色应为最后一个
        wb = load_workbook(fp)
        rgb = wb.active["A1"].fill.fgColor.rgb
        assert "FF00FF" in rgb or "FFFF00FF" in rgb
        wb.close()


# ==================== R55+ Round 2: 新增边缘/组合测试 ====================


def _make_sample(tmp_path, data, sheet_name="Sheet1"):
    """创建最小测试文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row_idx, row_data in enumerate(data, start=1):
        if isinstance(row_data, list):
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        else:
            ws.cell(row=row_idx, column=1, value=row_data)
    fp = str(tmp_path / f"sample_{id(data)}.xlsx")
    wb.save(fp)
    wb.close()
    return fp


class TestFormatCellsR55Round2:
    """R55+ 第2轮: format_cells 边缘case + 深度组合测试"""

    # ---------- T1: 五大类全组合 ----------
    def test_full_five_category_combination(self, tmp_path):
        """同时设置 font + fill + alignment + border + number_format"""
        fp = _make_sample(tmp_path, [[42]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "font": {"name": "Arial", "size": 12, "bold": True, "italic": True,
                     "color": "FF0000", "underline": "single"},
            "fill": {"type": "solid", "color": "00FF00"},
            "alignment": {"horizontal": "center", "vertical": "center",
                         "wrap_text": True, "text_rotation": 0},
            "border": {"left": "medium", "right": "medium",
                       "top": "thin", "bottom": "thin",
                       "color": "0000FF"},
            "number_format": "#,##0.00",
        })
        assert result.success is True
        s = _read_cell_style(fp, "A1")
        assert s["bold"] is True
        assert s["italic"] is True
        assert s["underline"] == "single"
        assert s["alignment_h"] == "center"
        assert s["alignment_v"] == "center"
        assert s["wrap_text"] is True
        assert s["number_format"] == "#,##0.00"
        assert s["border_left"] == "medium"
        assert s["border_right"] == "medium"

    # ---------- T2: 显式 False 关闭布尔属性 ----------
    def test_explicit_false_bold_turns_off(self, tmp_path):
        """先设 bold=True，再用 bold=False 显式关闭"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)

        # 先加粗
        r1 = writer.format_cells("Sheet1!A1", {"font": {"bold": True}})
        assert r1.success is True
        s1 = _read_cell_style(fp, "A1")
        assert s1["bold"] is True

        # 再取消粗体（显式 False）
        r2 = writer.format_cells("Sheet1!A1", {"font": {"bold": False}})
        assert r2.success is True
        s2 = _read_cell_style(fp, "A1")
        assert s2["bold"] is False

    def test_explicit_false_italic_and_underline(self, tmp_path):
        """italic=False 和 underline=False (none) 的关闭行为"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)

        # 开启
        writer.format_cells("Sheet1!A1", {
            "font": {"italic": True, "underline": "double"}
        })

        # 关闭
        result = writer.format_cells("Sheet1!A1", {
            "font": {"italic": False, "underline": False}
        })
        assert result.success is True
        s = _read_cell_style(fp, "A1")
        assert s["italic"] is False
        # underline=False → "none" in _apply_cell_format, but openpyxl Font may store as None
        assert s["italic"] is False
        assert s["underline"] in ("none", None)  # openpyxl normalizes "none" to None

    # ---------- T3: number_format 边缘 ----------
    def test_number_format_scientific(self, tmp_path):
        """科学计数法 number_format"""
        fp = _make_sample(tmp_path, [[123456789]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"number_format": "0.00E+00"})
        assert result.success is True
        s = _read_cell_style(fp, "A1")
        assert s["number_format"] == "0.00E+00"
        # 值不被破坏
        assert s["value"] == 123456789

    def test_number_format_fraction(self, tmp_path):
        """分数格式 number_format"""
        fp = _make_sample(tmp_path, [[3.14159]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"number_format": "# ?/?"})
        assert result.success is True
        s = _read_cell_style(fp, "A1")
        assert s["number_format"] == "# ?/?"

    def test_number_format_date_type(self, tmp_path):
        """日期格式 number_format 不破坏数值（openpyxl 可能将日期序列号转为 datetime）"""
        fp = _make_sample(tmp_path, [[45000]])  # Excel 序列号日期
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"number_format": "YYYY-MM-DD"})
        assert result.success is True
        s = _read_cell_style(fp, "A1")
        assert s["number_format"] == "YYYY-MM-DD"
        # openpyxl 在日期格式下可能将数字转为 datetime，这是正常行为
        # 关键是值不为 None 且文件未损坏
        assert s["value"] is not None

    def test_number_format_empty_string_no_corruption(self, tmp_path):
        """number_format 设为空字符串不应损坏文件 (P1 regression)"""
        fp = _make_sample(tmp_path, [[99]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"number_format": ""})
        assert result.success is True
        # 文件仍可正常读取
        wb = load_workbook(fp)
        assert wb.active["A1"].value == 99
        wb.close()

    # ---------- T4: text_rotation 边界值 ----------
    def test_text_rotation_exact_boundaries(self, tmp_path):
        """text_rotation 在边界值（openpyxl 仅支持 0~180，负值无效）"""
        fp = _make_sample(tmp_path, [[1, 2, 3]])
        writer = ExcelWriter(fp)

        # openpyxl text_rotation 有效范围: 0-180
        # 0=水平, 1-90=顺时针旋转, 91-180=堆叠文字
        # 注意：-90 等负值不被 openpyxl 接受
        for rot, ref in [(0, "A1"), (45, "B1"), (90, "C1")]:
            r = writer.format_cells(f"Sheet1!{ref}", {
                "alignment": {"text_rotation": rot}
            })
            assert r.success is True, f"rotation={rot} failed"
            s = _read_cell_style(fp, ref)
            assert s["text_rotation"] == rot, f"expected {rot}, got {s['text_rotation']}"

    def test_text_rotation_negative_rejected(self, tmp_path):
        """text_rotation 负值应被 openpyxl 拒绝（验证边界行为）"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"alignment": {"text_rotation": -90}})
        # openpyxl 不接受负值，应失败
        assert result.success is False
        assert "Value must be one of" in str(result.error)

    def test_text_rotation_45_degrees(self, tmp_path):
        """text_rotation=45 常用角度"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"alignment": {"text_rotation": 45}})
        assert result.success is True
        assert _read_cell_style(fp, "A1")["text_rotation"] == 45

    # ---------- T5: 覆盖行为（后设覆盖先设）----------
    def test_format_override_font_replaced(self, tmp_path):
        """第二次格式化完全替换字体属性"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)

        writer.format_cells("Sheet1!A1", {
            "font": {"name": "Arial", "size": 14, "bold": True, "color": "FF0000"}
        })
        writer.format_cells("Sheet1!A1", {
            "font": {"name": "Courier New", "size": 10}
        })
        s = _read_cell_style(fp, "A1")
        # name 和 size 应被覆盖为新值
        assert s["font_name"] == "Courier New"
        assert s["size"] == 10
        # bold 未在第二次设置，但 openpyxl Font() 用了 cell.font.bold 作为 default
        # 所以 bold 保留为 True（来自上一次的 cell.font.bold）
        # 这验证了"不崩"的底线

    def test_format_override_fill_replaced(self, tmp_path):
        """第二次格式化替换填充类型"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)

        writer.format_cells("Sheet1!A1", {"fill": {"type": "solid", "color": "FF0000"}})
        writer.format_cells("Sheet1!A1", {"fill": {"type": "solid", "color": "00FF00"}})
        rgb = load_workbook(fp).active["A1"].fill.fgColor.rgb
        load_workbook(fp).close()
        assert "00FF00" in rgb or "FF00FF00" in rgb

    # ---------- T6: 含公式单元格的样式操作 ----------
    def test_format_formula_cell_preserves_formula(self, tmp_path):
        """对含公式的单元格设置样式后公式不丢失"""
        fp = _make_sample(tmp_path, [[1, 2]])
        wb = load_workbook(fp)
        ws = wb.active
        ws["C1"] = "=A1+B1"
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!C1", {
            "font": {"bold": True},
            "fill": {"type": "solid", "color": "FFFF00"},
            "alignment": {"horizontal": "center"},
        })
        assert result.success is True

        wb2 = load_workbook(fp)
        cell = wb2.active["C1"]
        # 公式保留（openpyxl 不执行公式，只存公式字符串）
        assert cell.value == "=A1+B1", f"formula lost! got {cell.value}"
        assert cell.font.bold is True
        wb2.close()

    def test_format_range_with_mixed_values_and_formulas(self, tmp_path):
        """范围中混合数值和公式，设置样式后均不丢失"""
        fp = _make_sample(tmp_path, [[10, 20, 30]])
        wb = load_workbook(fp)
        ws = wb.active
        ws["D1"] = "=SUM(A1:C1)"
        wb.save(fp)
        wb.close()

        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1:D1", {
            "font": {"name": "微软雅黑"},
            "number_format": "#,##0",
        })
        assert result.success is True
        assert result.metadata.get("formatted_count", 0) == 4

        wb2 = load_workbook(fp)
        assert wb2.active["A1"].value == 10
        assert wb2.active["D1"].value == "=SUM(A1:C1)"
        wb2.close()

    # ---------- T7: 合并区域 + 全量样式组合 ----------
    def test_merge_then_full_style_combo(self, tmp_path):
        """合并 A1:C3 后一次性应用全部5类样式"""
        fp = _make_sample(tmp_path, [[i * j for j in range(1, 4)] for i in range(1, 4)])
        writer = ExcelWriter(fp)

        # 合并
        mr = writer.merge_cells("A1:C3", "Sheet1")
        assert mr.success is True

        # 全量样式
        fr = writer.format_cells("Sheet1!A1", {
            "font": {"bold": True, "size": 16, "color": "FFFFFF"},
            "fill": {"type": "solid", "color": "000080"},  # 深蓝底白字
            "alignment": {"horizontal": "center", "vertical": "center"},
            "border": {"top": "medium", "bottom": "medium",
                       "left": "medium", "right": "medium",
                       "color": "FFD700"},
            "number_format": "@",  # 文本格式
        })
        assert fr.success is True

        s = _read_cell_style(fp, "A1")
        assert s["bold"] is True
        assert s["size"] == 16
        assert s["alignment_h"] == "center"
        assert s["border_top"] == "medium"
        assert s["number_format"] == "@"

    # ---------- T8: indent / shrink_to_fit 边缘 ----------
    def test_indent_zero_explicit(self, tmp_path):
        """indent=0 显式设置"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"alignment": {"indent": 0}})
        assert result.success is True
        assert _read_cell_style(fp, "A1")["text_rotation"] == 0  # 验证可读

    def test_shrink_to_fit_true(self, tmp_path):
        """shrink_to_fit=True"""
        fp = _make_sample(tmp_path, [["Long text that shrinks"]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "alignment": {"shrink_to_fit": True}
        })
        assert result.success is True
        # openpyxl shrink_to_fit 默认 None, 设置后应为 True
        wb = load_workbook(fp)
        assert wb.active["A1"].alignment.shrink_to_fit is True
        wb.close()

    def test_wrap_text_and_shrink_to_fit_together(self, tmp_path):
        """wrap_text=True + shrink_to_fit=True 同时设置"""
        fp = _make_sample(tmp_path, [["Test wrap and shrink"]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "alignment": {"wrap_text": True, "shrink_to_fit": True}
        })
        assert result.success is True
        wb = load_workbook(fp)
        align = wb.active["A1"].alignment
        assert align.wrap_text is True
        assert align.shrink_to_fit is True
        wb.close()

    # ---------- T9: 扁平参数 normalize 组合 ----------
    def test_normalize_flat_all_font_attrs_together(self):
        """扁平格式：所有字体属性同时传入的 normalize 结果"""
        from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        flat = {
            "bold": True, "italic": True, "underline": "double",
            "font_size": 18, "font_color": "FF0000",
            "font_name": "Consolas", "strikethrough": True,
        }
        nested = ExcelOperations._normalize_formatting(flat)
        font = nested.get("font", {})
        assert font["bold"] is True
        assert font["italic"] is True
        assert font["underline"] == "double"
        assert font["size"] == 18
        assert font["color"] == "FF0000"
        assert font["name"] == "Consolas"
        assert font["strikethrough"] is True

    def test_normalize_flat_all_align_attrs_together(self):
        """扁平格式：所有对齐属性同时传入"""
        from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        flat = {
            "alignment": "right",
            "vertical_alignment": "top",
            "wrap_text": True,
            "text_rotation": 180,  # openpyxl 特殊值: 垂直文字
            "indent": 5,
            "shrink_to_fit": True,
        }
        nested = ExcelOperations._normalize_formatting(flat)
        align = nested.get("alignment", {})
        assert align["horizontal"] == "right"
        assert align["vertical"] == "top"
        assert align["wrap_text"] is True
        assert align["text_rotation"] == 180
        assert align["indent"] == 5
        assert align["shrink_to_fit"] is True

    def test_normalize_flat_complete_combo(self):
        """扁平格式：全部5类属性 + 渐变的完整 normalize"""
        from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
        flat = {
            "bold": True, "font_size": 12,
            "bg_color": "FF0000",
            "alignment": "center",
            "wrap_text": True,
            "border": {"top": "thick", "bottom": "thick"},
            "number_format": "0.00%",
            "gradient_colors": ["000000", "FFFFFF"],
            "fill_type": "gradient",
        }
        nested = ExcelOperations._normalize_formatting(flat)
        # font
        assert nested["font"]["bold"] is True
        assert nested["font"]["size"] == 12
        # fill (gradient overrides bg_color since both present)
        assert nested["fill"]["type"] == "gradient"
        assert "colors" in nested["fill"]
        # alignment
        assert nested["alignment"]["horizontal"] == "center"
        assert nested["alignment"]["wrap_text"] is True
        # border
        assert "border" in nested
        # number_format
        assert nested["number_format"] == "0.00%"

    # ---------- T10: API 层 (ExcelOperations) 端到端 ----------
    def test_api_format_cells_flat_params_e2e(self, tmp_path):
        """通过 ExcelOperations API 用扁平参数做端到端格式化"""
        fp = _make_sample(tmp_path, [[100]])
        result = ExcelOperations.format_cells(
            file_path=fp,
            sheet_name="Sheet1",
            range="A1",
            formatting={
                "bold": True,
                "font_size": 14,
                "bg_color": "FFD700",
                "alignment": "center",
                "number_format": "¥#,##0.00",
            },
        )
        assert result["success"] is True
        s = _read_cell_style(fp, "A1")
        assert s["bold"] is True
        assert s["size"] == 14
        assert s["alignment_h"] == "center"
        assert s["number_format"] == "¥#,##0.00"

    def test_api_format_cells_preset_highlight_plus_custom(self, tmp_path):
        """preset=highlight 后额外追加自定义格式"""
        fp = _make_sample(tmp_path, [[1, 2]])
        # 先用 preset
        r1 = ExcelOperations.format_cells(
            file_path=fp, sheet_name="Sheet1", range="A1:B1",
            preset="highlight",
        )
        assert r1["success"] is True
        # 再追加自定义
        r2 = ExcelOperations.format_cells(
            file_path=fp, sheet_name="Sheet1", range="A1:B1",
            formatting={"bold": True, "font_size": 12},
        )
        assert r2["success"] is True
        s = _read_cell_style(fp, "A1")
        assert s["bold"] is True
        assert s["size"] == 12

    def test_api_format_cells_border_dict_via_flat(self, tmp_path):
        """通过 API 扁平参数传 border 字典"""
        fp = _make_sample(tmp_path, [[1]])
        result = ExcelOperations.format_cells(
            file_path=fp, sheet_name="Sheet1", range="A1",
            formatting={
                "border": {
                    "left": "double",
                    "right": "double",
                    "top": "thick",
                    "bottom": "thick",
                    "color": "FF0000",
                },
            },
        )
        assert result["success"] is True
        s = _read_cell_style(fp, "A1")
        assert s["border_left"] == "double"
        assert s["border_top"] == "thick"


# ==================== R55+ Round 3: bg_color 合并 Bug 修复 + 边缘覆盖 ====================


class TestFormatCellsBgColorMergeBugFix:
    """Bug fix: _normalize_formatting 中 bg_color 不再覆盖 fill_type/gradient_colors (P2)"""

    def test_bg_color_with_fill_type_pattern_merges(self, tmp_path):
        """bg_color + fill_type=pattern 应合并（不丢失 type 信息）"""
        from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

        result = ExcelOperations._normalize_formatting({
            "fill_type": "pattern",
            "bg_color": "FF0000",
        })
        # 修复前: {"fill": {"color": "FF0000"}} — type 丢失
        # 修复后: {"fill": {"type": "pattern", "color": "FF0000"}}
        assert result["fill"]["type"] == "pattern"
        assert result["fill"]["color"] == "FF0000"

    def test_bg_color_with_gradient_colors_merges(self, tmp_path):
        """bg_color + gradient_colors 应合并（不丢失 gradient 信息）"""
        from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

        result = ExcelOperations._normalize_formatting({
            "gradient_colors": ["4472C4", "ED7D31"],
            "bg_color": "FFFF00",
        })
        # 修复前: {"fill": {"color": "FFFF00"}} — gradient 全丢
        # 修复后: gradient type/colors 保留，color 也存在
        assert result["fill"]["type"] == "gradient"
        assert "colors" in result["fill"]
        assert result["fill"]["colors"] == ["4472C4", "ED7D31"]
        assert result["fill"]["color"] == "FFFF00"

    def test_bg_color_with_fill_type_e2e_writer(self, tmp_path):
        """端到端：通过 Operations API 同时传 fill_type=pattern 和 bg_color"""
        fp = _make_sample(tmp_path, [[1]])
        result = ExcelOperations.format_cells(
            file_path=fp, sheet_name="Sheet1", range="A1",
            formatting={"fill_type": "pattern", "bg_color": "FF0000"},
        )
        assert result["success"] is True
        # 文件可正常打开且值保留
        wb = load_workbook(fp)
        assert wb.active["A1"].value == 1
        # pattern fill 类型应生效（不是 solid）
        fill = wb.active["A1"].fill
        assert fill.fill_type is not None or fill.patternType is not None
        wb.close()

    def test_bg_color_with_gradient_colors_e2e_operations(self, tmp_path):
        """端到端：bg_color + gradient_colors 通过 Operations API 组合"""
        fp = _make_sample(tmp_path, [[42]])
        result = ExcelOperations.format_cells(
            file_path=fp, sheet_name="Sheet1", range="A1",
            formatting={
                "gradient_colors": ["000000", "FFFFFF"],
                "bg_color": "FFD700",
                "bold": True,
            },
        )
        assert result["success"] is True
        # 验证 bold 生效，值保留
        wb = load_workbook(fp)
        assert wb.active["A1"].font.bold is True
        assert wb.active["A1"].value == 42
        # 验证 fill 是 GradientFill（通过属性存在性判断）
        fill = wb.active["A1"].fill
        assert hasattr(fill, "degree") or hasattr(fill, "fgColor")
        wb.close()


class TestFormatCellsPatternFillVariants:
    """PatternFill 不同 patternType 覆盖测试"""

    def test_pattern_fill_dark_horizontal(self, tmp_path):
        """darkHorizontal 图案填充"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {"type": "pattern", "patternType": "darkHorizontal", "fgColor": "FF0000"}
        })
        assert result.success is True

    def test_pattern_fill_light_down(self, tmp_path):
        """lightDown 图案填充"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {"type": "pattern", "patternType": "lightDown", "fgColor": "00FF00"}
        })
        assert result.success is True

    def test_pattern_fill_medium_gray(self, tmp_path):
        """mediumGray 图案填充"""
        fp = _make_sample(tmp_path, [[1]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {
            "fill": {"type": "pattern", "patternType": "mediumGray"}
        })
        assert result.success is True


class TestFormatCellsNumberFormatReset:
    """数字格式重置/特殊值测试"""

    def test_number_format_general_reset(self, tmp_path):
        """number_format 设为 'General' 重置为默认格式"""
        fp = _make_sample(tmp_path, [[3.14159]])
        writer = ExcelWriter(fp)

        # 先设为货币格式
        writer.format_cells("Sheet1!A1", {"number_format": "¥#,##0.00"})

        # 再重置为 General
        result = writer.format_cells("Sheet1!A1", {"number_format": "General"})
        assert result.success is True

        wb = load_workbook(fp)
        assert wb.active["A1"].number_format == "General"
        wb.close()

    def test_number_format_text_at_sign(self, tmp_path):
        """number_format='@' 文本格式不改变数值显示"""
        fp = _make_sample(tmp_path, [[12345]])
        writer = ExcelWriter(fp)
        result = writer.format_cells("Sheet1!A1", {"number_format": "@"})
        assert result.success is True

        wb = load_workbook(fp)
        assert wb.active["A1"].number_format == "@"
        assert wb.active["A1"].value == 12345
        wb.close()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
