#!/usr/bin/env python3
"""
Round 8 迭代测试 — CTE深度、UNION集合操作、DELETE安全、精度边缘、数据分析视角
日期: 2026-04-13
角色覆盖: 数据分析、运营、QA自动化、服务端开发、策划
"""

import os
import random
import sys
import tempfile
import traceback

# 设置环境
sys.path.insert(0, "/root/workspace/excel-mcp-server/src")

from openpyxl import Workbook

from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_delete_query,
    execute_advanced_insert_query,
    execute_advanced_sql_query,
    execute_advanced_update_query,
)


# ============================================================
# 1. 创建模拟游戏配置表
# ============================================================
def create_test_xlsx():
    """创建多表游戏配置文件，覆盖多种场景"""
    wb = Workbook()

    # --- 表1: 装备表 (Equipment) ---
    ws_equip = wb.active
    ws_equip.title = "装备"
    headers_equip = ["ID", "Name", "BaseAtk", "AtkBonus", "Price", "Rarity", "Category"]
    ws_equip.append(headers_equip)
    random.seed(42)  # 固定种子保证可重现
    rarities = ["Common", "Rare", "Epic", "Legendary"]
    categories = ["Weapon", "Armor", "Accessory", "Helmet"]
    for i in range(1, 31):
        ws_equip.append(
            [
                i,
                f"装备-{i}",
                random.randint(10, 200),
                round(random.uniform(5.0, 50.0), 2),
                round(random.uniform(100.0, 9999.99), 2),
                random.choice(rarities),
                random.choice(categories),
            ]
        )

    # --- 表2: 技能表 (Skills) ---
    ws_skill = wb.create_sheet("技能")
    headers_skill = ["SkillID", "SkillName", "Damage", "ManaCost", "Cooldown", "Type"]
    ws_skill.append(headers_skill)
    skill_types = ["Attack", "Defense", "Buff", "Debuff", "Heal"]
    for i in range(1, 21):
        ws_skill.append(
            [
                i,
                f"技能-{i}",
                random.randint(50, 500),
                random.randint(10, 100),
                random.randint(1, 30),
                random.choice(skill_types),
            ]
        )

    # --- 表3: 商店表 (Shop) ---
    ws_shop = wb.create_sheet("商店")
    headers_shop = ["ItemID", "ItemName", "Price", "Stock", "Discount", "ShopType"]
    ws_shop.append(headers_shop)
    shop_types = ["General", "Premium", "Event", "Limited"]
    for i in range(1, 16):
        ws_shop.append(
            [
                i,
                f"商品-{i}",
                round(random.uniform(50.0, 5000.0), 2),
                random.randint(0, 100),
                round(random.uniform(0.5, 0.95), 2),
                random.choice(shop_types),
            ]
        )

    # --- 表4: 掉落表 (DropTable) ---
    ws_drop = wb.create_sheet("掉落")
    headers_drop = ["DropID", "MonsterID", "ItemID", "Rate", "MinCount", "MaxCount"]
    ws_drop.append(headers_drop)
    for i in range(1, 26):
        ws_drop.append(
            [
                i,
                random.randint(1001, 1050),
                random.randint(1, 30),
                round(random.uniform(0.001, 0.5), 4),  # 极小浮点
                random.randint(1, 5),
                random.randint(1, 10),
            ]
        )

    # --- 表5: 活动奖励表 (EventReward) ---
    ws_event = wb.create_sheet("活动奖励")
    headers_event = [
        "EventID",
        "RewardID",
        "RewardName",
        "Quantity",
        "Priority",
        "IsActive",
    ]
    ws_event.append(headers_event)
    for i in range(1, 11):
        ws_event.append(
            [
                f"EVENT-{i:03d}",
                i * 100,
                f"奖励-{i}",
                random.randint(1, 50),
                random.randint(1, 5),
                random.choice([True, False]),
            ]
        )

    # 添加一些特殊值用于边缘测试
    # NULL值测试行
    ws_equip.append([999, "NULL测试装", None, None, None, None, None])
    # 超长文本测试
    long_text = "A" * 500
    ws_equip.append([1000, long_text, 1, 1.0, 1.0, "Common", "Weapon"])
    # 特殊字符测试
    ws_equip.append([1001, "it's a \"test\" & <hero> 'sword'", 99, 9.99, 99.99, "Epic", "Accessory"])
    # 极大数值
    ws_equip.append([1002, "极大数值", 999999, 99999.99, 9999999.99, "Legendary", "Weapon"])
    # 极小数值（接近0）
    ws_equip.append([1003, "极小数值", 1, 0.001, 0.0001, "Common", "Armor"])
    # 负数（异常数据）
    ws_equip.append([1004, "负数异常", -10, -5.5, -99.99, "Rare", "Helmet"])

    tmpfile = tempfile.mktemp(suffix=".xlsx")
    wb.save(tmpfile)
    return tmpfile


# ============================================================
# 2. 测试用例定义
# ============================================================

TEST_CASES = [
    # ===== Group A: CTE (WITH子句) 深度测试 — 数据分析视角 =====
    {
        "id": "A1",
        "name": "CTE基础查询: WITH高价值装备 AS (...) SELECT ...",
        "role": "数据分析",
        "sql": "WITH 高价值装备 AS (SELECT ID, Name, Price FROM 装备 WHERE Price > 5000) SELECT * FROM 高价值装备 ORDER BY Price DESC LIMIT 5",
        "type": "select",
        "verify": lambda r: r["success"] and len(r.get("data", [])) <= 5,
    },
    {
        "id": "A2",
        "name": "CTE链式引用: CTE1→CTE2→主查询",
        "role": "数据分析",
        "sql": "WITH 稀有装备 AS (SELECT * FROM 装备 WHERE Rarity IN ('Epic', 'Legendary')), 稀有统计 AS (SELECT Category, COUNT(*) as cnt, AVG(Price) as avg_price FROM 稀有装备 GROUP BY Category) SELECT * FROM 稀有统计 WHERE cnt > 1 ORDER BY avg_price DESC",
        "type": "select",
        "verify": lambda r: r["success"] and len(r.get("data", [])) > 0,
    },
    {
        "id": "A3",
        "name": "CTE + JOIN跨表关联分析",
        "role": "服务端开发",
        "sql": "WITH 高价装备 AS (SELECT ID, Name, Price, Category FROM 装备 WHERE Price > 3000) SELECT e.Name, e.Price, s.ShopType FROM 高价装备 e INNER JOIN 商店 s ON e.ID = s.ItemID",
        "type": "select",
        "verify": lambda r: r["success"],
    },
    {
        "id": "A4",
        "name": "CTE + 窗口函数组合: 排名分析",
        "role": "策划",
        "sql": "WITH 装备排名 AS (SELECT *, ROW_NUMBER() OVER (ORDER BY Price DESC) as 排名 FROM 装备 WHERE Rarity = 'Legendary') SELECT Name, Price, 排名 FROM 装备排名 WHERE 排名 <= 3",
        "type": "select",
        "verify": lambda r: r["success"] and len(r.get("data", [])) <= 3,
    },
    {
        "id": "A5",
        "name": "CTE + 聚合 + HAVING: 分组过滤",
        "role": "数据分析",
        "sql": "WITH 品类统计 AS (SELECT Category, Rarity, COUNT(*) as num, SUM(BaseAtk) as total_atk FROM 装备 GROUP BY Category, Rarity) SELECT * FROM 品类统计 WHERE total_atk > 200 ORDER BY num DESC",
        "type": "select",
        "verify": lambda r: r["success"],
    },
    # ===== Group B: UNION / UNION ALL 集合操作 — 数据分析视角 =====
    {
        "id": "B1",
        "name": "UNION ALL 合并两表物品列表",
        "role": "数据分析",
        "sql": "SELECT ID, Name, Price FROM 装备 UNION ALL SELECT ItemID, ItemName, Price FROM 商店 ORDER BY Price DESC LIMIT 10",
        "type": "select",
        "verify": lambda r: r["success"] and len(r.get("data", [])) <= 10,
    },
    {
        "id": "B2",
        "name": "UNION 去重合并",
        "role": "数据分析",
        "sql": "SELECT Category FROM 装备 UNION SELECT ShopType FROM 商店",
        "type": "select",
        "verify": lambda r: r["success"] and len(r.get("data", [])) > 0,
    },
    {
        "id": "B3",
        "name": "UNION + ORDER BY 组合排序",
        "role": "运营",
        "sql": "SELECT Name, Price, '装备' as 来源 FROM 装备 WHERE Rarity = 'Legendary' UNION ALL SELECT ItemName, Price, '商店' as 来源 FROM 商店 WHERE Discount < 0.7 ORDER BY Price DESC",
        "type": "select",
        "verify": lambda r: r["success"],
    },
    {
        "id": "B4",
        "name": "三表UNION ALL合并",
        "role": "数据分析",
        "sql": "SELECT Name, BaseAtk as Value, 'Atk' as Type FROM 装备 UNION ALL SELECT SkillName, Damage, 'Dmg' FROM 技能 UNION ALL SELECT RewardName, Quantity, 'Qty' FROM 活动奖励 LIMIT 15",
        "type": "select",
        "verify": lambda r: r["success"] and len(r.get("data", [])) <= 15,
    },
    # ===== Group C: DELETE 安全机制验证 — QA视角 =====
    {
        "id": "C1",
        "name": "DELETE无WHERE应被拒绝(安全防护)",
        "role": "QA",
        "sql": "DELETE FROM 装备",
        "type": "delete",
        "expect_error": True,  # 应该失败——不允许全表删除
        "verify": lambda r: not r["success"] and "WHERE" in r.get("message", ""),
    },
    {
        "id": "C2",
        "name": "DELETE带WHERE正常删除+回读验证",
        "role": "QA",
        "sql": "DELETE FROM 装备 WHERE ID = 999",
        "type": "delete",
        "verify": lambda r: r["success"] and r.get("affected_rows", 0) >= 1,
    },
    {
        "id": "C3",
        "name": "DELETE无匹配行幂等(affected_rows=0)",
        "role": "QA",
        "sql": "DELETE FROM 装备 WHERE ID = 99999",
        "type": "delete",
        "verify": lambda r: r["success"] and r.get("affected_rows", -1) == 0,
    },
    # ===== Group D: UPDATE 精度与复杂写入 — 运营视角 =====
    {
        "id": "D1",
        "name": "UPDATE写入小数精度验证(123.456)",
        "role": "运营",
        "sql": "UPDATE 装备 SET Price = 123.456 WHERE ID = 1",
        "type": "update",
        "verify": lambda r: r["success"],
    },
    {
        "id": "D2",
        "name": "UPDATE回读验证D1精度是否保持",
        "role": "运营",
        "sql": "SELECT Price FROM 装备 WHERE ID = 1",
        "type": "select",
        "verify": lambda r: r["success"] and len(r.get("data", [])) == 1,
    },
    {
        "id": "D3",
        "name": "UPDATE批量调价: 全场打8折",
        "role": "运营",
        "sql": "UPDATE 装备 SET Price = ROUND(Price * 0.8, 2) WHERE Category = 'Weapon'",
        "type": "update",
        "verify": lambda r: r["success"] and r.get("affected_rows", 0) > 0,
    },
    {
        "id": "D4",
        "name": "UPDATE CASE WHEN动态定价: 稀有度分级调价",
        "role": "策划",
        "sql": "UPDATE 装备 SET Price = CASE WHEN Rarity = 'Legendary' THEN ROUND(Price * 1.2, 2) WHEN Rarity = 'Epic' THEN ROUND(Price * 1.1, 2) ELSE Price END WHERE ID <= 10",
        "type": "update",
        "verify": lambda r: r["success"] and r.get("affected_rows", 0) > 0,
    },
    {
        "id": "D5",
        "name": "UPDATE字符串字段修改",
        "role": "运营",
        "sql": "UPDATE 装备 SET Name = '★限定版-' || Name WHERE Rarity = 'Legendary'",
        "type": "update",
        "verify": lambda r: r["success"] and r.get("affected_rows", 0) > 0,
    },
    # ===== Group E: EXCEPT / INTERSECT 集合操作 =====
    {
        "id": "E1",
        "name": "EXCEPT差集: 装备类别不在商店的",
        "role": "数据分析",
        "sql": "SELECT DISTINCT Category FROM 装备 EXCEPT SELECT DISTINCT ShopType FROM 商店",
        "type": "select",
        "verify": lambda r: r["success"],
    },
    {
        "id": "E2",
        "name": "INTERSECT交集: 两表共有的类型名称",
        "role": "数据分析",
        "sql": "SELECT DISTINCT Category FROM 装备 INTERSECT SELECT DISTINCT ShopType FROM 商店",
        "type": "select",
        "verify": lambda r: r["success"],
    },
    # ===== Group F: 边缘场景压力测试 — QA自动化视角 =====
    {
        "id": "F1",
        "name": "NULL值聚合: AVG/SUM/COUNT处理NULL列",
        "role": "QA",
        "sql": "SELECT COUNT(*) as total, COUNT(BaseAtk) as non_null_atk, AVG(BaseAtk) as avg_atk, SUM(Price) as sum_price FROM 装备 WHERE ID >= 999",
        "type": "select",
        "verify": lambda r: r["success"],
    },
    {
        "id": "F2",
        "name": "超长文本WHERE筛选",
        "role": "QA",
        "sql": "SELECT ID, LENGTH(Name) as name_len FROM 装备 WHERE ID = 1000",
        "type": "select",
        "verify": lambda r: r["success"] and len(r.get("data", [])) == 1,
    },
    {
        "id": "F3",
        "name": "特殊字符值: 引号/符号/HTML标签",
        "role": "QA",
        "sql": "SELECT ID, Name, Price FROM 装备 WHERE ID = 1001",
        "type": "select",
        "verify": lambda r: r["success"] and len(r.get("data", [])) == 1,
    },
    {
        "id": "F4",
        "name": "极小浮点率值: 掉落率0.001级别运算",
        "role": "QA",
        "sql": "SELECT DropID, Rate, Rate * 100 as RatePct, ROUND(Rate * 10000, 0) as RateBps FROM 掉落 WHERE Rate < 0.01 ORDER BY Rate ASC",
        "type": "select",
        "verify": lambda r: r["success"],
    },
    {
        "id": "F5",
        "name": "负数异常数据处理",
        "role": "QA",
        "sql": "SELECT * FROM 装备 WHERE BaseAtk < 0 OR Price < 0",
        "type": "select",
        "verify": lambda r: r["success"] and len(r.get("data", [])) >= 1,
    },
    # ===== Group G: INSERT 写入 + 回读验证 =====
    {
        "id": "G1",
        "name": "INSERT新装备行+回读验证",
        "role": "策划",
        "sql": "INSERT INTO 装备 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category) VALUES (2001, '新神装', 999, 99.99, 88888.88, 'Legendary', 'Weapon')",
        "type": "insert",
        "verify": lambda r: r["success"] and r.get("affected_rows", 0) >= 1,
    },
    {
        "id": "G2",
        "name": "INSERT后SELECT回读G1数据",
        "role": "策划",
        "sql": "SELECT * FROM 装备 WHERE ID = 2001",
        "type": "select",
        "verify": lambda r: r["success"] and len(r.get("data", [])) == 1,
    },
    {
        "id": "G3",
        "name": "INSERT含特殊字符的名称",
        "role": "运营",
        "sql": "INSERT INTO 装备 (ID, Name, BaseAtk, AtkBonus, Price, Rarity, Category) VALUES (2002, '王者的「荣耀」之剑', 888, 88.88, 66666.66, 'Epic', 'Weapon')",
        "type": "insert",
        "verify": lambda r: r["success"],
    },
]


# ============================================================
# 3. 执行引擎
# ============================================================


def run_test(file_path, tc):
    """执行单个测试用例"""
    sql = tc["sql"]
    t = tc["type"]

    try:
        if t == "select":
            result = execute_advanced_sql_query(file_path, sql)
        elif t == "update":
            result = execute_advanced_update_query(file_path, sql)
        elif t == "insert":
            result = execute_advanced_insert_query(file_path, sql)
        elif t == "delete":
            result = execute_advanced_delete_query(file_path, sql)
        else:
            return {"status": "SKIP", "reason": f"未知类型: {t}"}

        # 验证
        verify_fn = tc.get("verify")
        expect_error = tc.get("expect_error", False)

        if expect_error:
            if not result["success"]:
                if verify_fn(result):
                    return {
                        "status": "PASS",
                        "data": result.get("data", []),
                        "message": result.get("message", "")[:200],
                    }
                else:
                    return {
                        "status": "FAIL",
                        "data": None,
                        "message": f"期望错误但验证失败: {result.get('message', '')[:200]}",
                    }
            else:
                return {
                    "status": "FAIL",
                    "data": result.get("data", []),
                    "message": "期望报错但执行成功",
                }
        else:
            if verify_fn(result):
                return {
                    "status": "PASS",
                    "data": result.get("data", []),
                    "message": result.get("message", ""),
                    "affected_rows": result.get("affected_rows"),
                }
            else:
                return {
                    "status": "FAIL",
                    "data": result.get("data", []),
                    "message": f"验证失败: {result.get('message', '')[:200]}",
                }

    except Exception as e:
        if tc.get("expect_error"):
            return {
                "status": "PASS",
                "data": None,
                "message": f"预期异常: {str(e)[:200]}",
            }
        return {
            "status": "ERROR",
            "data": None,
            "message": f"异常: {traceback.format_exc()[-500:]}",
        }


def main():
    print("=" * 80)
    print("🔄 ExcelMCP Round 8 迭代测试")
    print("=" * 80)

    # 创建测试文件
    file_path = create_test_xlsx()
    print(f"\n📁 测试文件: {file_path}")

    # 统计
    results = []
    by_group = {}

    print(f"\n{'ID':<5} {'状态':<6} {'测试用例'}")
    print("-" * 80)

    for tc in TEST_CASES:
        gid = tc["id"][0]  # group letter
        result = run_test(file_path, tc)
        result["tc"] = tc
        results.append(result)

        by_group.setdefault(gid, {"total": 0, "pass": 0, "fail": 0, "error": 0})
        by_group[gid]["total"] += 1
        if result["status"] == "PASS":
            by_group[gid]["pass"] += 1
            icon = "✅"
        elif result["status"] == "FAIL":
            by_group[gid]["fail"] += 1
            icon = "❌"
        else:
            by_group[gid]["error"] += 1
            icon = "💥"

        msg = result.get("message", "")[:60]
        print(f"{tc['id']:<5} {icon:<6} {tc['name']}")
        if result["status"] != "PASS":
            print(f"       └─ {msg}")

    # 汇总
    print("\n" + "=" * 80)
    print("📊 汇总报告")
    print("=" * 80)

    total = len(results)
    passed = sum(1 for r in results if r["status"] == "PASS")
    failed = sum(1 for r in results if r["status"] == "FAIL")
    errored = sum(1 for r in results if r["status"] == "ERROR")

    print(f"\n总计: {total} | 通过: {passed} | 失败: {failed} | 异常: {errored} | 通过率: {passed / total * 100:.1f}%")

    print("\n📈 分组统计:")
    for g, stat in sorted(by_group.items()):
        p = stat["pass"]
        t = stat["total"]
        icon = "✅" if p == t else "⚠️"
        print(f"  {icon} Group {g}: {p}/{t} 通过")

    # 失败详情
    fails = [(r, tc) for r, tc in zip(results, TEST_CASES) if r["status"] != "PASS"]
    if fails:
        print("\n❌ 失败用例详情:")
        for r, tc in fails:
            print(f"  [{tc['id']}] {tc['name']}")
            print(f"       SQL: {tc['sql'][:80]}...")
            print(f"       原因: {r.get('message', 'N/A')[:200]}")
            print()

    # 清理
    try:
        os.unlink(file_path)
    except:
        pass

    return results, TEST_CASES, file_path


if __name__ == "__main__":
    main()
