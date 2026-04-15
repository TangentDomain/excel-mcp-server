#!/usr/bin/env python3
"""Debug script for Round 26 failures analysis"""
import sys
sys.path.insert(0, '/root/workspace/excel-mcp-server/src')
import tempfile, os, random, shutil
from openpyxl import Workbook
from excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

TMPDIR = tempfile.mkdtemp(prefix="r26_debug_")

# --- Debug A2: ORDER BY LIMIT ---
std_xlsx = os.path.join(TMPDIR, "std.xlsx")
wb = Workbook()
ws = wb.active; ws.title = "Sheet1"
ws.append(["ID","Name","Value","Price","Category"])
for i in range(1,51):
    ws.append([i,f"Item_{i}", round(random.uniform(10,1000),2), round(random.uniform(1,9999.99),2), random.choice(["A","B","C","D"])])
wb.save(std_xlsx)

print("=== A2 Debug: ORDER BY + LIMIT ===")
r = execute_advanced_sql_query(std_xlsx, "SELECT * FROM Sheet1 ORDER BY ID DESC LIMIT 3")
print(f"success={r['success']}")
d = r.get('data')
print(f"data type={type(d).__name__}")
if isinstance(d, list):
    print(f"len={len(d)}")
    for i, row in enumerate(d[:5]):
        print(f"  row[{i}]={row}")
else:
    print(f"data={d}")
print(f"columns={r.get('columns')}")
print(f"message: {str(r.get('message',''))[:200]}")
print()

# --- Debug B1: COUNT(*) ---
print("=== B1 Debug: COUNT(*) ===")
large_path = os.path.join(TMPDIR, "large.xlsx")
wb2 = Workbook(); ws2 = wb2.active; ws2.title = "BigTable"
ws2.append(["ID","Name","Value","Price","Category"])
for i in range(1,10001):
    ws2.append([i,f"I_{i}", round(random.uniform(10,1000),2), round(random.uniform(1,9999.99),2), random.choice(["A","B","C","D"])])
wb2.save(large_path)

r = execute_advanced_sql_query(large_path, "SELECT COUNT(*) as cnt FROM BigTable")
print(f"success={r['success']}")
d = r.get('data')
print(f"data type={type(d).__name__}, data={d}")
print(f"columns={r.get('columns')}")
print(f"message snippet: {str(r.get('message',''))[:200]}")
print()

# --- Debug A5: Empty table ---
print("=== A5 Debug: Empty Table ===")
empty_path = os.path.join(TMPDIR, "empty.xlsx")
wb3 = Workbook(); ws3 = wb3.active; ws3.title = "EmptySheet"
ws3.append(["ID","Name","Value"])
wb3.save(empty_path)

r = execute_advanced_sql_query(empty_path, "SELECT * FROM EmptySheet")
d = r.get('data')
print(f"SELECT * success={r['success']}, data_type={type(d).__name__}, data={d}, len={len(d) if isinstance(d,list) else 'N/A'}")

r2 = execute_advanced_sql_query(empty_path, "SELECT COUNT(*) as cnt FROM EmptySheet")
d2 = r2.get('data')
print(f"COUNT(*) success={r2['success']}, data={d2}, columns={r2.get('columns')}")
print()

# --- Debug C1: Multi-statement ---
print("=== C1 Debug: SELECT multi-statement ===")
r = execute_advanced_sql_query(std_xlsx, "SELECT COUNT(*) FROM Sheet1; SELECT COUNT(*) FROM Sheet1")
print(f"success={r['success']}")
d = r.get('data')
print(f"data type={type(d).__name__}, data={d}")
print(f"message: {str(r.get('message',''))[:400]}")
print()

# --- Debug E3: CTE cross-sheet ---
print("=== E3 Debug: CTE cross-sheet ===")
multi_path = os.path.join(TMPDIR, "multi.xlsx")
wb4 = Workbook()
ws_p = wb4.active; ws_p.title = "Players"
ws_p.append(["PlayerID","Name","Level","GuildID"])
for i in range(1,21): ws_p.append([i,f"P_{i}", random.randint(1,99), random.randint(1,5)])
ws_g = wb4.create_sheet("Guilds")
ws_g.append(["GuildID","GuildName","MasterID"])
for i in range(1,6): ws_g.append([i,f"G_{i}", i*4])
wb4.save(multi_path)

r = execute_advanced_sql_query(multi_path, """
WITH PlayerStats AS (
    SELECT p.GuildID, COUNT(*) as cnt, AVG(p.Level) as avg_lvl 
    FROM Players p 
    GROUP BY p.GuildID
)
SELECT g.GuildName, ps.cnt, ps.avg_lvl 
FROM Guilds g 
JOIN PlayerStats ps ON g.GuildID = ps.GuildID
""")
print(f"success={r['success']}")
print(f"message: {str(r.get('message',''))[:400]}")
if r.get('data'):
    d = r.get('data')
    print(f"data type={type(d).__name__}, data={d}")

shutil.rmtree(TMPDIR, ignore_errors=True)
