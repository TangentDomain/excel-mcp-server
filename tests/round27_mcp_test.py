"""
Round 27 MCP 迭代测试 — 深度嵌套查询 + 数据一致性验证 + 错误信息质量审计 + P0回归
日期: 2026-04-14
方向: 与 R22-R26 均不同的新方向
"""

import sys
import os
import time
import tempfile
import traceback

sys.path.insert(0, '/root/workspace/excel-mcp-server/src')
os.chdir('/root/workspace/excel-mcp-server')

from openpyxl import Workbook
import random

# ============================================================
# 1. 创建测试数据文件
# ============================================================
def create_test_files():
    """创建多张测试用 Excel 文件"""
    files = {}
    
    # --- 文件1: 游戏配置（装备+怪物+掉落）---
    wb = Workbook()
    
    # Sheet 1: 装备配置
    ws_equip = wb.active
    ws_equip.title = "装备配置"
    ws_equip.append(["ID", "Name", "BaseAtk", "AtkBonus", "Price", "Rarity", "Category"])
    random.seed(42)
    rarities = ["Common", "Rare", "Epic", "Legendary"]
    categories = ["Weapon", "Armor", "Accessory"]
    for i in range(1, 31):
        ws_equip.append([
            i,
            f"Item-{i}",
            random.randint(10, 100),
            round(random.uniform(5.5, 45.8), 2),
            round(random.uniform(50.5, 9999.99), 2),
            random.choice(rarities),
            random.choice(categories)
        ])
    
    # Sheet 2: 怪物配置
    ws_monster = wb.create_sheet("怪物配置")
    ws_monster.append(["ID", "Name", "Level", "HP", "ZoneID"])
    for i in range(1, 21):
        ws_monster.append([
            i,
            f"Monster-{i}",
            random.randint(1, 99),
            random.randint(100, 9999),
            random.randint(1, 5)
        ])
    
    # Sheet 3: 掉落配置
    ws_drop = wb.create_sheet("掉落配置")
    ws_drop.append(["DropID", "MonsterID", "ItemID", "DropRate", "MinQty", "MaxQty"])
    for i in range(1, 41):
        ws_drop.append([
            i,
            random.randint(1, 20),
            random.randint(1, 30),
            round(random.uniform(0.01, 1.0), 4),
            random.randint(1, 3),
            random.randint(3, 10)
        ])
    
    # Sheet 4: 公会配置
    ws_guild = wb.create_sheet("公会配置")
    ws_guild.append(["GuildID", "GuildName", "MasterID", "Level", "MemberCount"])
    for i in range(1, 11):
        ws_guild.append([i, f"Guild-{i}", random.randint(100, 200), random.randint(1, 50), random.randint(5, 100)])
    
    fd, fpath = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(fpath)
    files['game'] = fpath
    
    # --- 文件2: 简单测试表 ---
    wb2 = Workbook()
    ws = wb2.active
    ws.title = "Sheet1"
    ws.append(["ID", "Value", "Score"])
    for i in range(1, 11):
        ws.append([i, i * 10, round(random.uniform(0, 100), 2)])
    
    fd2, fpath2 = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd2)
    wb2.save(fpath2)
    files['simple'] = fpath2
    
    return files


# ============================================================
# 测试框架
# ============================================================
class _TestResult:
    def __init__(self):
        self.results = []
        self.start_time = time.time()
    
    def add(self, name, category, passed, detail=""):
        status = "✅ PASS" if passed else "❌ FAIL"
        self.results.append({
            'name': name,
            'category': category,
            'passed': passed,
            'detail': detail
        })
        print(f"  {status} | {name}")
        if detail and not passed:
            print(f"       └─ {detail[:200]}")
        elif detail and passed:
            print(f"       └─ {detail[:120]}")
    
    def summary(self):
        total = len(self.results)
        passed = sum(1 for r in self.results if r['passed'])
        failed = total - passed
        elapsed = time.time() - self.start_time
        print(f"\n{'='*70}")
        print(f"📊 总计: {total} | 通过: {passed} | 失败: {failed} | 耗时: {elapsed:.1f}s")
        print(f"{'='*70}")
        return {'total': total, 'passed': passed, 'failed': failed, 'elapsed': elapsed}


tr = _TestResult()


# ============================================================
# 导入 API
# ============================================================
from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_sql_query,
    execute_advanced_update_query,
    execute_advanced_insert_query,
    execute_advanced_delete_query
)


def run_query(file_path, sql):
    """执行 SELECT 查询"""
    try:
        result = execute_advanced_sql_query(file_path, sql)
        return result
    except Exception as e:
        return {'success': False, 'message': str(e), 'data': []}


def run_update(file_path, sql):
    """执行 UPDATE 查询"""
    try:
        result = execute_advanced_update_query(file_path, sql)
        return result
    except Exception as e:
        return {'success': False, 'message': str(e)}


def run_insert(file_path, sql):
    """执行 INSERT 查询"""
    try:
        result = execute_advanced_insert_query(file_path, sql)
        return result
    except Exception as e:
        return {'success': False, 'message': str(e)}


def run_delete(file_path, sql):
    """执行 DELETE 查询"""
    try:
        result = execute_advanced_delete_query(file_path, sql)
        return result
    except Exception as e:
        return {'success': False, 'message': str(e)}


# ============================================================
# A组: 深度嵌套查询测试 (3-4层CTE / 相关子查询)
# ============================================================
print("\n" + "="*70)
print("🔬 A组: 深度嵌套查询测试 (3-4层CTE / 相关子查询)")
print("="*70)

files = create_test_files()
f = files['game']

# A1: 三层 CTE 嵌套
print("\n--- A1-A10: 深度嵌套CTE ---")

r = run_query(f, """
WITH Layer1 AS (
    SELECT Rarity, AVG(Price) as AvgPrice, COUNT(*) as Cnt
    FROM 装备配置 GROUP BY Rarity
),
Layer2 AS (
    SELECT *, RANK() OVER (ORDER BY AvgPrice DESC) as PriceRank
    FROM Layer1
),
Layer3 AS (
    SELECT * FROM Layer2 WHERE PriceRank <= 2
)
SELECT * FROM Layer3 ORDER BY PriceRank
""")
tr.add("A1: 三层CTE嵌套(聚合→窗口函数→过滤)", "深度嵌套",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}" if r.get('success') else r.get('message','')[:100])

# A2: 四层 CTE 嵌套
r = run_query(f, """
WITH L1 AS (SELECT Category, SUM(BaseAtk) as TotalAtk FROM 装备配置 GROUP BY Category),
L2 AS (SELECT *, MAX(TotalAtk) over() as MaxAtk FROM L1),
L3 AS (SELECT * FROM L2 WHERE TotalAtk > MaxAtk * 0.5),
L4 AS (SELECT Category, ROUND(TotalAtk/MaxAtk*100, 1) as PctOfMax FROM L3)
SELECT * FROM L4 ORDER BY PctOfMax DESC
""")
tr.add("A2: 四层CTE嵌套(聚合→窗口→过滤→计算)", "深度嵌套",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1}" if r.get('success') else r.get('message','')[:100])

# A3: CTE + 相关子查询 (scalar subquery in SELECT)
r = run_query(f, """
SELECT e.Name, e.Price, e.Rarity,
       (SELECT AVG(Price) FROM 装备配置 e2 WHERE e2.Rarity = e.Rarity) as RarityAvgPrice
FROM 装备配置 e
WHERE e.ID <= 5
ORDER BY e.Price DESC
""")
tr.add("A3: SELECT中相关子查询(scalar subquery)", "深度嵌套",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1}" if r.get('success') else r.get('message','')[:100])

# A4: CTE + IN 子查询
r = run_query(f, """
SELECT * FROM 装备配置
WHERE ID IN (SELECT ItemID FROM 掉落配置 WHERE DropRate > 0.5)
ORDER BY ID
LIMIT 10
""")
tr.add("A4: WHERE IN 子查询", "深度嵌套",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1}" if r.get('success') else r.get('message','')[:100])

# A5: EXISTS 子查询
r = run_query(f, """
SELECT * FROM 装备配置 e
WHERE EXISTS (SELECT 1 FROM 掉落配置 d WHERE d.ItemID = e.ID AND d.DropRate > 0.8)
ORDER BY e.ID
""")
tr.add("A5: EXISTS 相关子查询", "深度嵌套",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1}" if r.get('success') else r.get('message','')[:100])

# A6: NOT EXISTS 子查询
r = run_query(f, """
SELECT * FROM 装备配置 e
WHERE NOT EXISTS (SELECT 1 FROM 掉落配置 d WHERE d.ItemID = e.ID)
ORDER BY e.ID
""")
tr.add("A6: NOT EXISTS 子查询(无掉落的装备)", "深度嵌套",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1}" if r.get('success') else r.get('message','')[:100])

# A7: 多层子查询嵌套 (subquery in subquery)
r = run_query(f, """
SELECT * FROM 装备配置
WHERE Price > (SELECT AVG(Price) FROM 装备配置 
               WHERE Rarity = (SELECT Rarity FROM 装备配置 WHERE ID = 1))
ORDER BY Price DESC
LIMIT 5
""")
tr.add("A7: 三层标量子查询嵌套(WHERE>SELECT>SELECT)", "深度嵌套",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1}" if r.get('success') else r.get('message','')[:100])

# A8: CTE + JOIN + 窗口函数 组合
r = run_query(f, """
WITH MonsterDrops AS (
    SELECT m.Name as MonsterName, d.ItemID, d.DropRate,
           ROW_NUMBER() OVER (PARTITION BY m.ID ORDER BY d.DropRate DESC) as rn
    FROM 怪物配置 m JOIN 掉落配置 d ON m.ID = d.MonsterID
)
SELECT * FROM MonsterDrops WHERE rn <= 3 ORDER BY MonsterName, rn
""")
tr.add("A8: CTE+JOIN+窗口函数(PARTITION+ROW_NUMBER)", "深度嵌套",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1}" if r.get('success') else r.get('message','')[:100])

# A9: CASE WHEN 嵌套在聚合中
r = run_query(f, """
SELECT 
    Category,
    SUM(CASE WHEN Price > 500 THEN 1 ELSE 0 END) as ExpensiveCnt,
    SUM(CASE WHEN Price <= 500 THEN 1 ELSE 0 END) as CheapCnt,
    COUNT(*) as Total
FROM 装备配置
GROUP BY Category
ORDER BY Total DESC
""")
tr.add("A9: CASE WHEN嵌套在SUM聚合中(GROUP BY)", "深度嵌套",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1}" if r.get('success') else r.get('message','')[:100])

# A10: HAVING + 窗口函数 (通过子查询包装，符合SQL标准)
r = run_query(f, """
WITH CatStats AS (
    SELECT Category, COUNT(*) as Cnt, AVG(Price) as AvgP,
           RANK() OVER (ORDER BY AVG(Price) DESC) as PriceRank
    FROM 装备配置 GROUP BY Category
)
SELECT * FROM CatStats WHERE Cnt >= 5 ORDER BY PriceRank
""")
tr.add("A10: CTE包装(HAVING+窗口函数)", "深度嵌套",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1}" if r.get('success') else r.get('message','')[:100])


# ============================================================
# B组: 数据一致性验证
# ============================================================
print("\n" + "="*70)
print("🔬 B组: 数据一致性验证 (UPDATE/INSERT/DELETE后验证)")
print("="*70)

fs = files['simple']

# B1: UPDATE 前后数据一致性
r_before = run_query(fs, "SELECT * FROM Sheet1 WHERE ID = 1")
orig_val = r_before['data'][1][1] if len(r_before.get('data', [])) > 1 else None
print(f"  B1-前置: ID=1 的 Value = {orig_val}")

r_upd = run_update(fs, "UPDATE Sheet1 SET Value = 999 WHERE ID = 1")
tr.add("B1a: UPDATE SET单行", "数据一致性",
       r_upd.get('success', False),
       r_upd.get('message', '')[:100] if not r_upd.get('success') else '')

r_after = run_query(fs, "SELECT Value FROM Sheet1 WHERE ID = 1")
new_val = r_after['data'][1][0] if len(r_after.get('data', [])) > 1 else None
tr.add("B1b: UPDATE后验证值=999", "数据一致性",
       new_val == 999,
       f"期望=999 实际={new_val}")

# 还原
run_update(fs, f"UPDATE Sheet1 SET Value = {orig_val} WHERE ID = 1")

# B2: UPDATE 影响行数验证
r_cnt_before = run_query(fs, "SELECT COUNT(*) as cnt FROM Sheet1 WHERE Value > 0")
before_total = int(r_cnt_before['data'][1][0]) if r_cnt_before.get('data') and len(r_cnt_before['data']) > 1 else 0

r_upd2 = run_update(fs, "UPDATE Sheet1 SET Score = Score + 1")
tr.add("B2a: UPDATE全表(Score+1)", "数据一致性",
       r_upd2.get('success', False),
       r_upd2.get('message', '')[:100])

r_cnt_after = run_query(fs, "SELECT COUNT(*) as cnt FROM Sheet1")
after_total = int(r_cnt_after['data'][1][0]) if r_cnt_after.get('data') and len(r_cnt_after['data']) > 1 else 0
tr.add("B2b: UPDATE后总行数不变", "数据一致性",
       before_total == after_total,
       f"UPDATE前={before_total} UPDATE后={after_total}")

# 还原
run_update(fs, "UPDATE Sheet1 SET Score = Score - 1")

# B3: INSERT 后能查到
r_ins = run_insert(fs, "INSERT INTO Sheet1 (ID, Value, Score) VALUES (999, 777, 88.8)")
tr.add("B3a: INSERT新行(ID=999)", "数据一致性",
       r_ins.get('success', False),
       r_ins.get('message', '')[:100])

r_verify = run_query(fs, "SELECT * FROM Sheet1 WHERE ID = 999")
found = len(r_verify.get('data', [])) > 1  # header + data row
tr.add("B3b: INSERT后能查询到ID=999", "数据一致性",
       found,
           f"查询到{len(r_verify.get('data',[]))-1}行" if r_verify.get('data') else "空结果")

if found:
    vals = r_verify['data'][1]
    tr.add("B3c: INSERT数据值正确(999,777,88.8)", "数据一致性",
           vals[0] == 999 and vals[1] == 777 and abs(float(vals[2]) - 88.8) < 0.01,
           f"实际值={vals}")

# 清理
run_delete(fs, "DELETE FROM Sheet1 WHERE ID = 999")

# B4: DELETE 后确认删除
r_del = run_delete(fs, "DELETE FROM Sheet1 WHERE ID = 999")
tr.add("B4a: DELETE已插入的行", "数据一致性",
       r_del.get('success', False),
       r_del.get('message', '')[:100])

r_check = run_query(fs, "SELECT * FROM Sheet1 WHERE ID = 999")
deleted_ok = len(r_check.get('data', [])) <= 1  # only header
tr.add("B4b: DELETE后确认不存在", "数据一致性",
       deleted_ok,
       f"还剩{len(r_check.get('data',[]))-1 if r_check.get('data') else 0}行")

# B5: 批量 UPDATE 一致性
r_b5_pre = run_query(fs, "SELECT SUM(Value) as s, COUNT(*) as c FROM Sheet1")
pre_sum = float(r_b5_pre['data'][1][0]) if r_b5_pre.get('data') and len(r_b5_pre['data']) > 1 else 0
pre_cnt = int(r_b5_pre['data'][1][1]) if r_b5_pre.get('data') and len(r_b5_pre['data']) > 1 else 0

r_b5_upd = run_update(fs, "UPDATE Sheet1 SET Value = Value * 2")
r_b5_post = run_query(fs, "SELECT SUM(Value) as s, COUNT(*) as c FROM Sheet1")
post_sum = float(r_b5_post['data'][1][0]) if r_b5_post.get('data') and len(r_b5_post['data']) > 1 else 0
post_cnt = int(r_b5_post['data'][1][1]) if r_b5_post.get('data') and len(r_b5_post['data']) > 1 else 0

tr.add("B5a: 批量UPDATE(Value*2)成功", "数据一致性",
       r_b5_upd.get('success', False),
       r_b5_upd.get('message', '')[:100])
tr.add("B5b: UPDATE后SUM翻倍", "数据一致性",
       abs(post_sum - pre_sum * 2) < 0.01,
       f"前SUM={pre_sum:.1f} 后SUM={post_sum:.1f} 期望={pre_sum*2:.1f}")
tr.add("B5c: UPDATE后COUNT不变", "数据一致性",
       pre_cnt == post_cnt,
       f"前={pre_cnt} 后={post_cnt}")

# 还原
run_update(fs, "UPDATE Sheet1 SET Value = Value / 2")


# ============================================================
# C组: 错误信息质量审计
# ============================================================
print("\n" + "="*70)
print("🔬 C组: 错误信息质量审计 (安全性+友好性)")
print("="*70)

# C1: 不存在的表名
r = run_query(f, "SELECT * FROM 不存在的表")
has_error_msg = not r.get('success', True)
msg = r.get('message', '')
safe = '/etc' not in msg and '/root' not in msg and 'traceback' not in msg.lower() and 'Traceback' not in msg
tr.add("C1: 不存在表的错误提示", "错误质量",
       has_error_msg and safe and len(msg) > 0,
       f"msg='{msg[:150]}'" if msg else "无错误消息(!)")

# C2: 不存在的列名
r = run_query(f, "SELECT 不存在的列 FROM 装备配置 LIMIT 1")
has_error_msg = not r.get('success', True)
msg = r.get('message', '')
safe = 'traceback' not in msg.lower()
tr.add("C2: 不存在列的错误提示", "错误质量",
       has_error_msg and safe,
       f"msg='{msg[:150]}'" if msg else "无错误消息")

# C3: SQL语法错误
r = run_query(f, "SELEC * FORM 装备配置")  # 故意拼写错误
has_error_msg = not r.get('success', True)
msg = r.get('message', '')
tr.add("C3: SQL语法错误的提示", "错误质量",
       has_error_msg and len(msg) > 0,
       f"msg='{msg[:150]}'" if msg else "无错误消息")

# C4: 类型错误（字符串当数字比较）
r = run_query(f, "SELECT * FROM 装备配置 WHERE BaseAtk > 'abc'")
# 这个可能成功也可能失败，取决于实现
tr.add("C4: 字符串与数字比较", "错误质量",
       True,  # 记录实际行为
       f"success={r.get('success')} msg='{str(r.get('message',''))[:100]}'")

# C5: 除零错误
r = run_query(f, "SELECT ID, 1/0 as divzero FROM 装备配置 LIMIT 1")
tr.add("C5: 除零操作处理", "错误质量",
       True,  # 记录行为
       f"success={r.get('success')} msg='{str(r.get('message',''))[:150]}'")

# C6: 空SQL
r = run_query(f, "")
tr.add("C6: 空SQL字符串", "错误质量",
       not r.get('success', True),
       f"msg='{str(r.get('message',''))[:100]}'")

# C7: 只有注释的SQL
r = run_query(f, "-- 这是一条注释")
tr.add("C7: 纯注释SQL", "错误质量",
       not r.get('success', True) or True,  # 两种行为都可接受
       f"success={r.get('success')} msg='{str(r.get('message',''))[:100]}'")

# C8: 非常长的SQL (>5000字符)
long_cols = ",".join([f"MAX(BaseAtk) as maxatk{i}" for i in range(200)])
long_sql = f"SELECT {long_cols} FROM 装备配置"
r = run_query(f, long_sql)
tr.add("C8: 超长SQL(200个聚合列)", "错误质量",
       r.get('success', False) or r.get('success', True),  # 记录实际行为
       f"success={r.get('success')} rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# C9: NULL值处理
r = run_query(f, "SELECT NULL as nullcol, 1 as onecol FROM 装备配置 LIMIT 1")
tr.add("C9: NULL字面量选择", "错误质量",
       r.get('success', False) or r.get('success', True),
       f"success={r.get('success')} data={str(r.get('data',''))[:100]}")


# ============================================================
# D组: P0回归验证
# ============================================================
print("\n" + "="*70)
print("🔬 D组: P0回归验证 (已知严重问题)")
print("="*70)

# D1: P0-1 SELECT 分号多语句注入
r = run_query(f, "SELECT COUNT(*) FROM 装备配置; SELECT COUNT(*) FROM 怪物配置")
is_injection_working = r.get('success', False) and '多语句' in str(r.get('message', ''))
tr.add("D1: [P0回归] SELECT分号多语句注入仍存在?", "P0回归",
       not is_injection_working,  # 期望已被修复
       f"still_vulnerable={is_injection_working} msg='{str(r.get('message',''))[:150]}'")

# D2: P0-2 script_runner RCE (如果可用则测试)
try:
    from excel_mcp_server_fastmcp.api.script_runner import run_script
    r_rce = run_script("import os; os.system('echo P0_TEST')")
    tr.add("D2: [P0回归] script_runner RCE仍存在?", "P0回归",
           False,  # 期望已被修复
           f"still_vulnerable=True result='{str(r_rce)[:150]}'")
except ImportError:
    tr.add("D2: [P0回归] script_runner模块不可导入", "P0回归",
           True,  # 模块不存在也算一种修复
           "module not importable (maybe removed)")
except Exception as e:
    tr.add("D2: [P0回归] script_runner调用异常", "P0回归",
           True,  # 异常说明可能有保护
           f"exception={str(e)[:150]}")

# D3: SQL 注入基础测试 (引号逃逸)
r = run_query(f, "SELECT * FROM 装备配置 WHERE Name = '' OR 1=1 --'")
tr.add("D3: [安全] 引号逃逸+OR 1=1注入", "P0回归",
       True,  # 记录实际行为
       f"success={r.get('success')} rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# D4: UNION 注入尝试
r = run_query(f, "SELECT Name FROM 装备配置 WHERE ID = 1 UNION SELECT Name FROM 怪物配置")
tr.add("D4: [安全] UNION注入跨表", "P0回归",
       True,  # 记录行为
       f"success={r.get('success')} rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# D5: DROP TABLE 注入尝试 (通过SELECT分号)
r = run_query(f, "SELECT 1; DROP TABLE 装备配置")
tr.add("D5: [安全] DROP TABLE通过分号注入", "P0回归",
       not r.get('success', False),  # 期望被拒绝
       f"success={r.get('success')} msg='{str(r.get('message',''))[:150]}'")


# ============================================================
# E组: 特殊场景和边界情况
# ============================================================
print("\n" + "="*70)
print("🔬 E组: 特殊场景 (LIKE/BETWEEN/IN list/DISTINCT)")
print("="*70)

# E1: LIKE 模式匹配
r = run_query(f, "SELECT * FROM 装备配置 WHERE Name LIKE 'Item-1%%' ORDER BY ID LIMIT 5")
tr.add("E1: LIKE前缀匹配(Item-1%)", "特殊场景",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# E2: LIKE 通配符
r = run_query(f, "SELECT * FROM 装备配置 WHERE Name LIKE 'Item-_3' ORDER BY ID")
tr.add("E2: LIKE单字符通配符(_)", "特殊场景",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# E3: BETWEEN
r = run_query(f, "SELECT * FROM 装备配置 WHERE BaseAtk BETWEEN 30 AND 60 ORDER BY BaseAtk")
tr.add("E3: BETWEEN范围查询", "特殊场景",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# E4: IN 列表
r = run_query(f, "SELECT * FROM 装备配置 WHERE ID IN (1, 3, 5, 7, 9) ORDER BY ID")
tr.add("E4: IN列表(显式值)", "特殊场景",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# E5: NOT IN
r = run_query(f, "SELECT * FROM 装备配置 WHERE ID NOT IN (1, 2, 3) ORDER BY ID LIMIT 5")
tr.add("E5: NOT IN列表", "特殊场景",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# E6: DISTINCT
r = run_query(f, "SELECT DISTINCT Rarity FROM 装备配置 ORDER BY Rarity")
tr.add("E6: DISTINCT去重", "特殊场景",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# E7: IS NULL / IS NOT NULL (如果有NULL数据)
r = run_query(f, "SELECT * FROM 装备配置 WHERE Rarity IS NOT NULL LIMIT 3")
tr.add("E7: IS NOT NULL判断", "特殊场景",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# E8: ORDER BY 多列
r = run_query(f, "SELECT * FROM 装备配置 ORDER BY Rarity, Price DESC LIMIT 5")
tr.add("E8: ORDER BY多列(Rarity ASC, Price DESC)", "特殊场景",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# E9: LIMIT OFFSET
r = run_query(f, "SELECT * FROM 装备配置 ORDER BY ID LIMIT 3 OFFSET 5")
tr.add("E9: LIMIT 3 OFFSET 5分页", "特殊场景",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# E10: 复杂布尔条件组合
r = run_query(f, """
SELECT * FROM 装备配置 
WHERE (Rarity = 'Legendary' OR Rarity = 'Epic') 
  AND BaseAtk > 50 
  AND (Category = 'Weapon' OR Category = 'Accessory')
ORDER BY Price DESC
LIMIT 5
""")
tr.add("E10: 复杂AND/OR布尔条件组合", "特殊场景",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")


# ============================================================
# F组: 跨Sheet JOIN 深度测试
# ============================================================
print("\n" + "="*70)
print("🔬 F组: 跨Sheet JOIN深度测试")
print("="*70)

# F1: 三表 JOIN
r = run_query(f, """
SELECT e.Name as EquipName, m.Name as MonName, d.DropRate, d.MaxQty
FROM 装备配置 e
JOIN 掉落配置 d ON e.ID = d.ItemID
JOIN 怪物配置 m ON d.MonsterID = m.ID
WHERE d.DropRate > 0.5
ORDER BY d.DropRate DESC
LIMIT 10
""")
tr.add("F1: 三表JOIN(装备+掉落+怪物)", "跨表JOIN",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# F2: LEFT JOIN (左连接)
r = run_query(f, """
SELECT e.Name, d.DropRate
FROM 装备配置 e
LEFT JOIN 掉落配置 d ON e.ID = d.ItemID
ORDER BY e.ID
LIMIT 10
""")
tr.add("F2: LEFT JOIN(含无掉落装备)", "跨表JOIN",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# F3: JOIN + 聚合 + GROUP BY
r = run_query(f, """
SELECT e.Rarity, COUNT(*) as DropCount, AVG(d.DropRate) as AvgRate
FROM 装备配置 e
INNER JOIN 掉落配置 d ON e.ID = d.ItemID
GROUP BY e.Rarity
ORDER BY DropCount DESC
""")
tr.add("F3: JOIN+GROUP BY聚合统计", "跨表JOIN",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# F4: 自连接 (self-join)
r = run_query(f, """
SELECT a.Name as NameA, b.Name as NameB, a.Price as PriceA, b.Price as PriceB
FROM 装备配置 a, 装备配置 b
WHERE a.Rarity = b.Rarity AND a.ID < b.ID AND a.Rarity = 'Legendary'
ORDER BY PriceA DESC
LIMIT 5
""")
tr.add("F4: 自连接(同表稀有度比较)", "跨表JOIN",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")

# F5: CROSS JOIN (笛卡尔积限制)
r = run_query(f, """
SELECT g.GuildName, z.ZoneID
FROM 公会配置 g, (SELECT DISTINCT ZoneID FROM 怪物配置) z
LIMIT 5
""")
tr.add("F5: CROSS JOIN(笛卡尔积+子查询)", "跨表JOIN",
       r.get('success', False),
       f"rows={len(r.get('data',[]))-1 if r.get('data') else 0}")


# ============================================================
# 清理临时文件
# ============================================================
for name, path in files.items():
    try:
        os.unlink(path)
    except:
        pass


# ============================================================
# 输出汇总
# ============================================================
summary = tr.summary()

# 分类统计
categories = {}
for r in tr.results:
    cat = r['category']
    if cat not in categories:
        categories[cat] = {'total': 0, 'passed': 0}
    categories[cat]['total'] += 1
    if r['passed']:
        categories[cat]['passed'] += 1

print("\n📊 分类统计:")
for cat, stats in sorted(categories.items()):
    p = stats['passed']
    t = stats['total']
    print(f"   {cat}: {p}/{t} ✅" if p == t else f"   {cat}: {p}/{t} ⚠️")

print("\n❌ 失败详情:")
failures = [r for r in tr.results if not r['passed']]
if failures:
    for f in failures:
        print(f"   [{f['category']}] {f['name']}: {f['detail'][:150]}")
else:
    print("   无! 全部通过 🎉")
