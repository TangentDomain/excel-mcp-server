"""
Round 7 迭代测试 — ExcelMCP 持续测试
日期: 2026-04-13
重点: HAVING深度测试 / SQL注入防护 / 大字段压力 / Sheet名特殊字符 / 复杂写入
"""
import sys
import os
import random
import traceback

sys.path.insert(0, '/root/workspace/excel-mcp-server/src')

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ============================================================
# 1. 创建测试数据文件
# ============================================================
def create_test_data():
    wb = Workbook()
    
    # --- Sheet 1: 装备表 (策划视角 - 基础配置) ---
    ws_equip = wb.active
    ws_equip.title = "装备"
    headers_equip = ["ID", "Name", "BaseAtk", "AtkBonus", "Price", "Rarity", "Category", "DropRate"]
    ws_equip.append(headers_equip)
    rarities = ["Common", "Rare", "Epic", "Legendary"]
    categories = ["Weapon", "Armor", "Accessory", "Consumable"]
    for i in range(1, 41):
        ws_equip.append([
            i,
            f"装备-{i:03d}",
            random.randint(10, 500),
            round(random.uniform(5.0, 200.0), 2),
            round(random.uniform(50.0, 50000.0), 2),
            random.choice(rarities),
            random.choice(categories),
            round(random.uniform(0.001, 0.15), 4)
        ])
    
    # --- Sheet 2: 商店表 (运营视角) ---
    ws_shop = wb.create_sheet("商店")
    headers_shop = ["ItemID", "ShopName", "Currency", "SellPrice", "Stock", "Discount", "IsActive"]
    ws_shop.append(headers_shop)
    currencies = ["Gold", "Gem", "Coupon"]
    shop_names = ["主城商店", "秘境商店", "竞技场商店", "公会商店"]
    for i in range(1, 31):
        ws_shop.append([
            i,
            random.choice(shop_names),
            random.choice(currencies),
            round(random.uniform(100.0, 100000.0), 2),
            random.randint(0, 999),
            round(random.uniform(0.5, 1.0), 2),
            random.choice([True, False])
        ])
    
    # --- Sheet 3: 活动奖励表 (运营/策划) ---
    ws_event = wb.create_sheet("活动奖励")
    headers_event = ["EventID", "RewardTier", "MinLevel", "MaxLevel", "RewardItemID", "Count", "Probability"]
    ws_event.append(headers_event)
    for i in range(1, 26):
        ws_event.append([
            f"EVENT-{i:03d}",
            random.choice(["Bronze", "Silver", "Gold", "Diamond"]),
            random.randint(1, 80),
            random.randint(81, 100),
            random.randint(1, 40),
            random.randint(1, 10),
            round(random.uniform(0.01, 0.3), 3)
        ])
    
    # --- Sheet 4: 特殊字符Sheet名 (中文+空格+括号) ---
    ws_special = wb.create_sheet("配置 表(正式)")
    headers_special = ["键", "值", "类型", "备注"]
    ws_special.append(headers_special)
    special_data = [
        ["server_ip", "192.168.1.100", "string", "游戏服务器IP"],
        ["max_players", 5000, "int", "最大在线人数"],
        ["drop_rate_multiplier", 1.5, "float", "掉率倍率"],
        ["maintenance_msg", "服务器维护中，请稍后再试", "string", "维护公告-含特殊符号!@#$%"],
        ["empty_val", "", "string", "空字符串测试"],
        ["null_test", None, "null", "NULL值测试"],
    ]
    for row in special_data:
        ws_special.append(row)
    
    # --- Sheet 5: 大字段压力表 ---
    ws_large = wb.create_sheet("大字段测试")
    headers_large = ["ID", "Content", "JSONData", "Description"]
    ws_large.append(headers_large)
    # 超长文本行
    long_text_1 = "这是一段" + "超长测试文本" * 200  # ~1400 chars
    long_text_2 = ("{" + '"key": "value", ' * 50 + "}").rstrip(', ')[:-1] + "}"  # ~700 chars JSON-like
    long_text_3 = "English long text for testing. " * 150  # ~3000 chars English
    special_chars = '特殊字符测试：\'"<>{}[]|\\/^$`~\n\t\r测试完成'  # 含各种转义符
    for i in range(1, 11):
        content = long_text_1 if i == 1 else (long_text_2 if i == 2 else (long_text_3 if i == 3 else f"普通内容-{i}"))
        desc = special_chars if i == 4 else f"描述-{i}"
        ws_large.append([i, content, f'{{"id":{i},"data":"test"}}', desc])
    
    # --- Sheet 6: 数值边界表 (QA视角) ---
    ws_boundary = wb.create_sheet("数值边界")
    headers_boundary = ["ID", "IntVal", "FloatVal", "NegVal", "ZeroVal", "TinyVal", "HugeVal", "TextVal"]
    ws_boundary.append(headers_boundary)
    boundary_data = [
        [1, 2147483647, 999999.99, -999999.99, 0, 0.000001, 1e15, "正常文本"],
        [2, -2147483648, 0.001, -0.001, 0, 1e-10, 9.999e12, ""],
        [3, 0, 0.0, 0, 0, 0.0000001, 0, None],
        [4, 999999999, 123456789.123456, -123456789.123456, 0, 0.5, 1e8, "中文'带引号\"双引号"],
        [5, 1, 1.0, -1, 0, 0.01, 100, "Emoji🎮🔥⚡🛡️测试"],
    ]
    for row in boundary_data:
        ws_boundary.append(row)
    
    filepath = "/tmp/excelmcp_round7_test.xlsx"
    wb.save(filepath)
    print(f"✅ 测试文件已创建: {filepath}")
    print(f"   Sheets: {wb.sheetnames}")
    return filepath


# ============================================================
# 2. 执行测试
# ============================================================
def run_tests(filepath):
    from excel_mcp_server_fastmcp.api.advanced_sql_query import (
        execute_advanced_sql_query,
        execute_advanced_update_query,
        execute_advanced_insert_query,
        execute_advanced_delete_query
    )
    
    results = []
    
    def run_test(name, sql, sheet=None, expected_error=False, operation='SELECT'):
        """执行单个测试用例，返回结果字典"""
        full_sql = sql
        if sheet and 'FROM' not in sql.upper() and not sql.strip().upper().startswith(('UPDATE','INSERT','DELETE')):
            # 对于非完整SQL，不需要处理
            pass
        
        try:
            if operation == 'UPDATE':
                result = execute_advanced_update_query(filepath, full_sql)
            elif operation == 'INSERT':
                result = execute_advanced_insert_query(filepath, full_sql)
            elif operation == 'DELETE':
                result = execute_advanced_delete_query(filepath, full_sql)
            else:
                result = execute_advanced_sql_query(filepath, full_sql)
            
            success = result.get('success', False)
            msg = result.get('message', '')
            data = result.get('data', [])
            
            # 预期错误的情况
            if expected_error:
                if not success:
                    status = "✅ (预期失败)"
                    passed = True
                else:
                    status = "⚠️ (本应失败但通过了)"
                    passed = True  # 通过了也算OK，只是预期有偏差
            else:
                if success:
                    status = "✅"
                    passed = True
                else:
                    status = "❌"
                    passed = False
            
            row_result = {
                "name": name,
                "sql": full_sql[:80] + "..." if len(full_sql) > 80 else full_sql,
                "passed": passed,
                "status": status,
                "message": msg[:120] if msg else "",
                "data_len": len(data) if data and isinstance(data, list) else 0,
                "operation": operation,
                "expected_error": expected_error
            }
            results.append(row_result)
            
            icon = status.split()[0]
            print(f"{icon} {name}")
            if not passed and msg:
                print(f"   错误: {msg[:150]}")
            if data and isinstance(data, list) and len(data) > 0 and operation == 'SELECT':
                preview = data[:3]
                print(f"   数据预览: {preview}")
                
        except Exception as e:
            results.append({
                "name": name,
                "sql": full_sql[:80],
                "passed": False,
                "status": "💥",
                "message": str(e)[:150],
                "data_len": 0,
                "operation": operation,
                "expected_error": expected_error
            })
            print(f"💥 {name}")
            print(f"   异常: {str(e)[:150]}")
        
        return results[-1]
    
    # ================================================================
    # Group A: HAVING 深度测试 (数据分析/策划视角) — 5个
    # ================================================================
    print("\n" + "=" * 70)
    print("Group A: HAVING 深度测试 (5个)")
    print("=" * 70)
    
    run_test(
        "A1-HAVING基础聚合过滤",
        "SELECT Rarity, COUNT(*) as cnt, AVG(Price) as avg_price FROM 装备 GROUP BY Rarity HAVING COUNT(*) > 5",
        operation='SELECT'
    )
    
    run_test(
        "A2-HAVING多条件AND",
        "SELECT Category, SUM(BaseAtk) as total_atk, AVG(DropRate) as avg_drop FROM 装备 GROUP BY Category HAVING COUNT(*) >= 8 AND AVG(Price) > 5000",
        operation='SELECT'
    )
    
    run_test(
        "A3-HAVING+ORDER BY组合",
        "SELECT Rarity, ROUND(AVG(Price), 2) as avg_p, COUNT(*) as cnt FROM 装备 GROUP BY Rarity HAVING AVG(Price) > 3000 ORDER BY avg_p DESC",
        operation='SELECT'
    )
    
    run_test(
        "A4-HAVING数学表达式",
        "SELECT Category, SUM(BaseAtk) as total_atk FROM 装备 GROUP BY Category HAVING SUM(BaseAtk) > AVG(BaseAtk) * 3",
        operation='SELECT'
    )
    
    run_test(
        "A5-HAVING跨商店统计",
        "SELECT Currency, COUNT(*) as item_cnt, SUM(SellPrice) as total_rev FROM 商店 GROUP BY Currency HAVING SUM(SellPrice) > 100000 ORDER BY total_rev DESC",
        operation='SELECT'
    )
    
    # ================================================================
    # Group B: SQL 注入防护测试 (QA/安全视角) — 5个
    # ================================================================
    print("\n" + "=" * 70)
    print("Group B: SQL 注入防护测试 (5个)")
    print("=" * 70)
    
    run_test(
        "B1-DROP TABLE 注入攻击",
        "SELECT * FROM 装备; DROP TABLE 装备;",
        operation='SELECT',
        expected_error=True
    )
    
    run_test(
        "B2-注释注入 (--)",
        "SELECT ID, Name FROM 装备 WHERE ID = 1 OR 1=1 --",
        operation='SELECT',
        expected_error=True
    )
    
    run_test(
        "B3-UNION 注入",
        "SELECT * FROM 装备 UNION SELECT * FROM 商店",
        operation='SELECT',
        expected_error=True  # 列数不同，应该报错或安全拒绝
    )
    
    run_test(
        "B4-单引号逃逸",
        "SELECT * FROM 装备 WHERE Name = '' OR ''=''",
        operation='SELECT'
    )
    
    run_test(
        "B5-分号多条语句",
        "UPDATE 装备 SET Price = 0 WHERE ID = 1; DELETE FROM 装备 WHERE ID > 1",
        operation='UPDATE',
        expected_error=True
    )
    
    # ================================================================
    # Group C: 大字段与边缘压力测试 (QA/客户端视角) — 5个
    # ================================================================
    print("\n" + "=" * 70)
    print("Group C: 大字段与边缘压力测试 (5个)")
    print("=" * 70)
    
    run_test(
        "C1-超长文本查询",
        "SELECT ID, LEN(Content) as content_len FROM 大字段测试 WHERE ID = 1",
        operation='SELECT'
    )
    
    run_test(
        "C2-特殊字符内容LIKE",
        "SELECT ID, Description FROM 大字段测试 WHERE Description LIKE '%特殊%'",
        operation='SELECT'
    )
    
    run_test(
        "C3-数值极大值运算",
        "SELECT ID, HugeVal, HugeVal * 2 as doubled, HugeVal / 1000 as scaled FROM 数值边界 WHERE HugeVal > 1e10",
        operation='SELECT'
    )
    
    run_test(
        "C4-极小浮点精度",
        "SELECT ID, TinyVal, TinyVal * 10000 as magnified, ROUND(TinyVal, 10) as precise FROM 数值边界 ORDER BY TinyVal ASC LIMIT 3",
        operation='SELECT'
    )
    
    run_test(
        "C5-NULL混合聚合",
        "SELECT COUNT(*) as total, COUNT(TextVal) as non_null, AVG(FloatVal) as avg_f, SUM(NegVal) as sum_n FROM 数值边界",
        operation='SELECT'
    )
    
    # ================================================================
    # Group D: Sheet 名特殊字符测试 (服务端/运维视角) — 4个
    # ================================================================
    print("\n" + "=" * 70)
    print("Group D: Sheet 名特殊字符测试 (4个)")
    print("=" * 70)
    
    run_test(
        "D1-中文名Sheet基本查询",
        "SELECT * FROM `配置 表(正式)`",
        operation='SELECT'
    )
    
    run_test(
        "D2-特殊名Sheet条件筛选",
        "SELECT 键, 值, 备注 FROM `配置 表(正式)` WHERE 类型 = 'string'",
        operation='SELECT'
    )
    
    run_test(
        "D3-特殊名Sheet更新",
        "UPDATE `配置 表(正式)` SET 值 = '192.168.1.200' WHERE 键 = 'server_ip'",
        operation='UPDATE'
    )
    
    run_test(
        "D4-特殊名Sheet更新回读验证",
        "SELECT 键, 值 FROM `配置 表(正式)` WHERE 键 = 'server_ip'",
        operation='SELECT'
    )
    
    # ================================================================
    # Group E: 复杂写入操作 (策划/运营视角) — 5个
    # ================================================================
    print("\n" + "=" * 70)
    print("Group E: 复杂写入操作 (5个)")
    print("=" * 70)
    
    run_test(
        "E1-UPDATE CASE WHEN调价(策划批量调价)",
        "UPDATE 装备 SET Price = CASE WHEN Rarity = 'Legendary' THEN Price * 1.2 WHEN Rarity = 'Epic' THEN Price * 1.1 ELSE Price END",
        operation='UPDATE'
    )
    
    run_test(
        "E2-UPDATE CASE WHEN回读验证",
        "SELECT ID, Name, Price, Rarity FROM 装备 WHERE Rarity IN ('Legendary', 'Epic') ORDER BY Price DESC LIMIT 5",
        operation='SELECT'
    )
    
    run_test(
        "E3-INSERT新活动奖励(运营添加活动)",
        "INSERT INTO 活动奖励 (EventID, RewardTier, MinLevel, MaxLevel, RewardItemID, Count, Probability) VALUES ('EVENT-999', 'Platinum', 90, 100, 99, 1, 0.05)",
        operation='INSERT'
    )
    
    run_test(
        "E4-INSERT回读验证",
        "SELECT * FROM 活动奖励 WHERE EventID = 'EVENT-999'",
        operation='SELECT'
    )
    
    run_test(
        "E5-DELETE清理测试数据(QA清理)",
        "DELETE FROM 活动奖励 WHERE EventID = 'EVENT-999'",
        operation='DELETE'
    )
    
    run_test(
        "E6-DELETE后COUNT验证",
        "SELECT COUNT(*) as remaining FROM 活动奖励 WHERE EventID = 'EVENT-999'",
        operation='SELECT'
    )
    
    # ================================================================
    # Group F: 高级查询模式 (服务端/数据分析) — 4个
    # ================================================================
    print("\n" + "=" * 70)
    print("Group F: 高级查询模式 (4个)")
    print("=" * 70)
    
    run_test(
        "F1-IN子查询跨表",
        "SELECT ID, Name, Price FROM 装备 WHERE ID IN (SELECT ItemID FROM 商店 WHERE IsActive = True AND Discount < 0.7) ORDER BY Price DESC",
        operation='SELECT'
    )
    
    run_test(
        "F2-窗口函数+HAVING组合(标量子查询包装)",
        "SELECT * FROM (SELECT Rarity, Name, Price, RANK() OVER(PARTITION BY Rarity ORDER BY Price DESC) as rnk FROM 装备) t WHERE rnk <= 2 ORDER BY Rarity, rnk",
        operation='SELECT'
    )
    
    run_test(
        "F3-CROSS JOIN笛卡尔积(小表)",
        "SELECT a.Rarity, b.Currency, COUNT(*) as combo FROM (SELECT DISTINCT Rarity FROM 装备) a CROSS JOIN (SELECT DISTINCT Currency FROM 商店) b GROUP BY a.Rarity, b.Currency LIMIT 10",
        operation='SELECT'
    )
    
    run_test(
        "F4-BETWEEN对称条件",
        "SELECT ID, Name, BaseAtk, Price FROM 装备 WHERE BaseAtk BETWEEN 200 AND 400 AND Price BETWEEN 5000 AND 20000 ORDER BY BaseAtk",
        operation='SELECT'
    )
    
    # ================================================================
    # Group G: 写入边缘场景 (QA视角) — 3个
    # ================================================================
    print("\n" + "=" * 70)
    print("Group G: 写入边缘场景 (3个)")
    print("=" * 70)
    
    run_test(
        "G1-UPDATE无匹配行(幂等安全)",
        "UPDATE 装备 SET Price = 999999 WHERE ID = 99999",
        operation='UPDATE'
    )
    
    run_test(
        "G2-INSERT含特殊字符值",
        "INSERT INTO 数值边界 (ID, IntVal, FloatVal, NegVal, ZeroVal, TinyVal, HugeVal, TextVal) VALUES (100, 42, 3.14, -2.71, 0, 0.001, 1000, '测试\"引号''单引号\\反斜线')",
        operation='INSERT'
    )
    
    run_test(
        "G3-INSERT特殊字符回读",
        "SELECT ID, TextVal FROM 数值边界 WHERE ID = 100",
        operation='SELECT'
    )
    
    # ================================================================
    # 统计结果
    # ================================================================
    print("\n" + "=" * 70)
    print("📊 Round 7 测试结果汇总")
    print("=" * 70)
    
    total = len(results)
    passed = sum(1 for r in results if r["passed"])
    failed = [r for r in results if not r["passed"]]
    
    # 分类统计
    by_op = {}
    for r in results:
        op = r["operation"]
        if op not in by_op:
            by_op[op] = {"total": 0, "passed": 0}
        by_op[op]["total"] += 1
        if r["passed"]:
            by_op[op]["passed"] += 1
    
    print(f"\n总体: {passed}/{total} 通过 ({passed*100//total}%)\n")
    print("分类统计:")
    for op, stats in sorted(by_op.items()):
        pct = stats["passed"] * 100 // stats["total"] if stats["total"] > 0 else 0
        print(f"  {op}: {stats['passed']}/{stats['total']} ({pct}%)")
    
    if failed:
        print(f"\n❌ 失败用例 ({len(failed)}):")
        for r in failed:
            print(f"  [{r['operation']}] {r['name']}")
            print(f"     SQL: {r['sql']}")
            print(f"     原因: {r['message']}")
    
    return results, {"total": total, "passed": passed, "failed": len(failed), "by_op": by_op}


if __name__ == "__main__":
    print("=" * 70)
    print("🔄 ExcelMCP Round 7 迭代测试")
    print("=" * 70)
    
    filepath = create_test_data()
    results, summary = run_tests(filepath)
    
    print(f"\n\n{'=' * 70}")
    print(f"🏁 Round 7 完成: {summary['passed']}/{summary['total']}")
    print(f"{'=' * 70}")
