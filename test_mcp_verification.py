#!/usr/bin/env python3
"""
MCP验证测试 - 8项游戏开发场景
"""

import subprocess
import json
import tempfile
import os
from pathlib import Path

def create_test_excel():
    """创建测试Excel文件"""
    test_file = "/tmp/test_game_data.xlsx"
    
    # 使用openpyxl创建包含游戏数据的Excel文件
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "角色表"
    
    # 设置表头
    headers = ["角色ID", "角色名", "等级", "职业", "生命值", "法力值", "攻击力", "防御力"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    # 添加测试数据
    characters = [
        [1, "战士", 10, "战士", 1000, 50, 85, 60],
        [2, "法师", 8, "法师", 400, 200, 45, 30],
        [3, "射手", 12, "射手", 600, 100, 70, 40],
        [4, "刺客", 15, "刺客", 500, 150, 90, 35],
        [5, "牧师", 7, "牧师", 450, 180, 30, 50]
    ]
    
    for row_idx, char in enumerate(characters, 2):
        for col_idx, value in enumerate(char, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 创建装备表
    ws2 = wb.create_sheet("装备表")
    equipment_headers = ["装备ID", "装备名", "类型", "角色ID", "攻击力", "防御力"]
    for col, header in enumerate(equipment_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    equipment = [
        [101, "钢铁剑", "武器", 1, 30, 5],
        [102, "魔法杖", "武器", 2, 25, 3],
        [103, "皮甲", "防具", 1, 0, 15],
        [104, "法袍", "防具", 2, 0, 12],
        [105, "弓", "武器", 3, 28, 4]
    ]
    
    for row_idx, equip in enumerate(equipment, 2):
        for col_idx, value in enumerate(equip, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(test_file)
    return test_file

def run_mcp_verification():
    """运行MCP验证测试"""
    test_file = create_test_excel()
    
    tests = [
        {
            "name": "游戏角色查询",
            "command": f"""
python3 -c "
import sys
sys.path.insert(0, 'src')
from excel_mcp_server_fastmcp.server import excel_query
result = excel_query('{test_file}', 'SELECT * FROM 角色表 WHERE 职业 = \"战士\"')
print('角色查询结果:', len(result) if result else 0)
"
""",
            "expected": "角色查询结果: 1"
        },
        {
            "name": "游戏装备统计",
            "command": f"""
python3 -c "
import sys
sys.path.insert(0, 'src')
from excel_mcp_server_fastmcp.server import excel_query
result = excel_query('{test_file}', 'SELECT 类型, COUNT(*) as 数量 FROM 装备表 GROUP BY 类型')
print('装备统计成功:', len(result) if result else 0)
"
""",
            "expected": "装备统计成功: 3"
        },
        {
            "name": "游戏数据导出",
            "command": f"""
python3 -c "
import sys
sys.path.insert(0, 'src')
from excel_mcp_server_fastmcp.server import excel_export_to_csv
excel_export_to_csv('{test_file}', '角色表', '/tmp/characters.csv')
import os
print('导出成功:', os.path.exists('/tmp/characters.csv'))
"
""",
            "expected": "导出成功: True"
        },
        {
            "name": "游戏角色插入",
            "command": f"""
python3 -c "
import sys
sys.path.insert(0, 'src')
from excel_mcp_server_fastmcp.server import excel_upsert_row
excel_upsert_row('{test_file}', '角色表', [6, '德鲁伊', 9, '德鲁伊', 550, 160, 40, 45])
result = excel_query('{test_file}', 'SELECT COUNT(*) FROM 角色表')
print('角色总数:', result[0][0] if result else 0)
"
""",
            "expected": "角色总数: 6"
        },
        {
            "name": "游戏属性更新",
            "command": f"""
python3 -c "
import sys
sys.path.insert(0, 'src')
from excel_mcp_server_fastmcp.server import excel_update_query
excel_update_query('{test_file}', 'UPDATE 角色表 SET 攻击力 = 攻击力 + 10 WHERE 职业 = \"战士\"')
result = excel_query('{test_file}', 'SELECT 攻击力 FROM 角色表 WHERE 职业 = \"战士\"')
print('战士攻击力:', result[0][0] if result else 0)
"
""",
            "expected": "战士攻击力: 95"
        },
        {
            "name": "游戏装备关联查询",
            "command": f"""
python3 -c "
import sys
sys.path.insert(0, 'src')
from excel_mcp_server_fastmcp.server import excel_query
result = excel_query('{test_file}', '''
SELECT r.角色名, e.装备名 
FROM 角色表 r 
JOIN 装备表 e ON r.角色ID = e.角色ID 
WHERE e.类型 = \"武器\"
''')
print('关联查询结果:', len(result) if result else 0)
"
""",
            "expected": "关联查询结果: 3"
        },
        {
            "name": "游戏数据校验",
            "command": f"""
python3 -c "
import sys
sys.path.insert(0, 'src')
from excel_mcp_server_fastmcp.server import excel_check_duplicate_ids
result = excel_check_duplicate_ids('{test_file}', '角色表', '角色ID')
print('重复ID检查:', '无重复' if result else '有重复')
"
""",
            "expected": "重复ID检查: 无重复"
        },
        {
            "name": "游戏数据处理",
            "command": f"""
python3 -c "
import sys
sys.path.insert(0, 'src')
from excel_mcp_server_fastmcp.server import excel_merge_files
excel_merge_files('{test_file}', '{test_file}', '/tmp/merged_game.xlsx')
import os
print('合并成功:', os.path.exists('/tmp/merged_game.xlsx'))
"
""",
            "expected": "合并成功: True"
        }
    ]
    
    print("🧪 MCP游戏场景验证测试")
    print("=" * 50)
    
    results = []
    for i, test in enumerate(tests, 1):
        print(f"测试{i}/8: {test['name']}")
        
        try:
            result = subprocess.run(
                test['command'], 
                shell=True, 
                capture_output=True, 
                text=True, 
                timeout=30
            )
            
            if result.returncode == 0 and test['expected'] in result.stdout:
                print(f"  ✅ 通过")
                results.append(True)
            else:
                print(f"  ❌ 失败")
                print(f"    预期: {test['expected']}")
                print(f"    实际: {result.stdout}")
                results.append(False)
        except Exception as e:
            print(f"  ❌ 异常: {e}")
            results.append(False)
    
    # 清理
    if os.path.exists(test_file):
        os.remove(test_file)
    
    success_count = sum(results)
    print(f"\n📊 MCP验证结果: {success_count}/8 项测试通过")
    
    if success_count == 8:
        print("🎉 所有游戏场景验证通过！")
        return True
    else:
        print("⚠️ 部分游戏场景验证失败")
        return False

if __name__ == "__main__":
    success = run_mcp_verification()
    exit(0 if success else 1)