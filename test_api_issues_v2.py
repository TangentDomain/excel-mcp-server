#!/usr/bin/env python3
"""
复现监工报告中的5个API问题 - 直接测试MCP工具函数
"""

import json
import sys
import os
import tempfile
from pathlib import Path
import openpyxl

# 添加项目路径
sys.path.insert(0, str(Path(__file__).parent / "src"))

from excel_mcp_server_fastmcp.server import (
    excel_get_range,
    excel_format_cells,
    excel_set_formula,
    excel_search
)

def create_test_excel():
    """创建测试Excel文件"""
    test_file = "/tmp/test_api_problems.xlsx"

    # 创建新的Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 写入测试数据
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["C1"] = "City"
    ws["A2"] = "Alice"
    ws["B2"] = 25
    ws["C2"] = "Shenzhen"
    ws["A3"] = "Bob"
    ws["B3"] = 30
    ws["C3"] = "Beijing"
    ws["A4"] = "Charlie"
    ws["B4"] = 35
    ws["C4"] = "Shanghai"

    # 保存文件
    wb.save(test_file)
    print(f"✅ 创建测试文件: {test_file}")
    return test_file

def test_1_range_query():
    """Test 1: read_data_from_excel range query - parameter order issue"""
    print("\n1️⃣ 测试 read_data_from_excel 范围查询参数顺序...")

    test_file = create_test_excel()

    try:
        # 测试方法1：使用 range 参数（标准方式）
        result1 = excel_get_range(
            file_path=test_file,
            range="Sheet1!A1:C4"
        )
        print(f"✅ 使用 range 参数结果: {json.dumps(result1, indent=2, ensure_ascii=False)[:200]}...")

        # 测试方法2：使用 start_cell + end_cell（替代方式）
        result2 = excel_get_range(
            file_path=test_file,
            range="A1:C4",  # 只有范围，没有工作表名
            sheet_name="Sheet1",
            start_cell="A1",
            end_cell="C4"
        )
        print(f"✅ 使用 start_cell+end_cell 结果: {json.dumps(result2, indent=2, ensure_ascii=False)[:200]}...")

        # 检查参数顺序是否有问题
        import inspect
        sig = inspect.signature(excel_get_range)
        print(f"📝 函数签名: {sig}")
        print(f"📝 参数顺序: {list(sig.parameters.keys())}")

        # 问题分析：range 是必需参数（第2个），但位置可能被误认为sheet_name
        if result1.get('success') and result2.get('success'):
            print("✅ 两种方式都成功")
        else:
            print(f"❌ 测试失败:")
            if not result1.get('success'):
                print(f"   range 方式: {result1.get('message')}")
            if not result2.get('success'):
                print(f"   start_cell+end_cell 方式: {result2.get('message')}")

    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)
            print(f"🧹 清理测试文件: {test_file}")

def test_2_format_range_missing_params():
    """Test 2: format_range missing required parameters"""
    print("\n2️⃣ 测试 format_range 缺少必要参数...")

    test_file = create_test_excel()

    try:
        # 测试方法1：提供 range 参数，但不提供任何格式参数
        result1 = excel_format_cells(
            file_path=test_file,
            sheet_name="Sheet1",
            range="A1:C1"
            # 缺少 formatting 和 preset 参数
        )
        print(f"✅ 缺少格式参数结果: {json.dumps(result1, indent=2, ensure_ascii=False)[:200]}...")

        # 测试方法2：提供 start_cell + end_cell
        result2 = excel_format_cells(
            file_path=test_file,
            sheet_name="Sheet1",
            range="A1:C1",
            start_cell="A1",
            end_cell="C1"
        )
        print(f"✅ 使用 start_cell+end_cell 结果: {json.dumps(result2, indent=2, ensure_ascii=False)[:200]}...")

        # 检查函数签名
        import inspect
        sig = inspect.signature(excel_format_cells)
        print(f"📝 函数签名: {sig}")
        print(f"📝 参数顺序: {list(sig.parameters.keys())}")

        if result1.get('success'):
            print("✅ 缺少格式参数测试通过（应用默认值或无操作）")
        else:
            print(f"❌ 缺少格式参数测试失败: {result1.get('message')}")

    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)
            print(f"🧹 清理测试文件: {test_file}")

def test_3_apply_formula_missing_params():
    """Test 3: apply_formula missing required parameters"""
    print("\n3️⃣ 测试 apply_formula 缺少 formula 参数...")

    test_file = create_test_excel()

    try:
        # 测试缺少 formula 参数
        # 由于 formula 是必需参数，直接调用会抛出异常
        print("📝 尝试直接调用缺少 formula 参数（预期会失败）...")
        try:
            result = excel_set_formula(
                file_path=test_file,
                sheet_name="Sheet1",
                cell_address="B5"
                # 缺少 formula 参数
            )
            print(f"❌ 意外成功（不应该发生）: {result}")
        except TypeError as e:
            print(f"✅ 预期的 TypeError: {e}")

        # 测试正常调用
        print("\n📝 测试正常调用...")
        result = excel_set_formula(
            file_path=test_file,
            sheet_name="Sheet1",
            cell_address="B5",
            formula="=A2*2"
        )
        print(f"✅ 正常调用结果: {json.dumps(result, indent=2, ensure_ascii=False)[:200]}...")

        # 检查函数签名
        import inspect
        sig = inspect.signature(excel_set_formula)
        print(f"\n📝 函数签名: {sig}")
        print(f"📝 参数顺序: {list(sig.parameters.keys())}")

        if result.get('success'):
            print("✅ 正常调用测试通过")
        else:
            print(f"❌ 正常调用测试失败: {result.get('message')}")

    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)
            print(f"🧹 清理测试文件: {test_file}")

def test_4_search_logic():
    """Test 4: read_data_from_excel search logic issue"""
    print("\n4️⃣ 测试 read_data_from_excel 搜索逻辑...")

    test_file = create_test_excel()

    try:
        # 测试搜索功能
        result = excel_search(
            file_path=test_file,
            pattern="Alice",
            sheet_name="Sheet1",
            case_sensitive=False
        )
        print(f"✅ 搜索结果: {json.dumps(result, indent=2, ensure_ascii=False)[:200]}...")

        if result.get('success'):
            print("✅ 搜索功能测试通过")
        else:
            print(f"❌ 搜索功能测试失败: {result.get('message')}")

        # 检查函数签名
        import inspect
        sig = inspect.signature(excel_search)
        print(f"📝 函数签名: {sig}")

    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)
            print(f"🧹 清理测试文件: {test_file}")

def test_5_write_data_format():
    """Test 5: write_data_to_excel data format mismatch"""
    print("\n5️⃣ 测试 write_data_to_excel 数据格式不匹配...")

    test_file = create_test_excel()

    try:
        from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

        # 测试正常数据格式（二维数组）
        print("📝 测试正常数据格式（二维数组）...")
        result1 = ExcelOperations.update_range(
            file_path=test_file,
            range_expression="Sheet1!A5:B5",
            data=[["Test", 100]]  # List of lists（正确格式）
        )
        print(f"✅ 正常数据格式结果: {json.dumps(result1, indent=2, ensure_ascii=False)[:200]}...")

        # 测试错误数据格式（一维数组）
        print("\n📝 测试错误数据格式（一维数组）...")
        result2 = ExcelOperations.update_range(
            file_path=test_file,
            range_expression="Sheet1!A6:B6",
            data=["Wrong", "Format"]  # 一维列表（错误格式）
        )
        print(f"❌ 错误数据格式结果: {json.dumps(result2, indent=2, ensure_ascii=False)[:200]}...")

        # 检查函数签名
        import inspect
        sig = inspect.signature(ExcelOperations.update_range)
        print(f"\n📝 函数签名: {sig}")
        print(f"📝 参数顺序: {list(sig.parameters.keys())}")

        if result1.get('success') and not result2.get('success'):
            print("✅ 数据格式验证测试通过（正确格式成功，错误格式失败）")
        elif result1.get('success') and result2.get('success'):
            print("⚠️ 两种格式都成功（可能格式验证不够严格）")
        else:
            print(f"❌ 测试结果异常:")
            print(f"   正常格式: {'✅ 成功' if result1.get('success') else '❌ 失败'}")
            print(f"   错误格式: {'✅ 成功（不应该）' if result2.get('success') else '❌ 失败（正确）'}")

    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)
            print(f"🧹 清理测试文件: {test_file}")

def main():
    """运行所有测试"""
    print("🔍 开始复现监工报告中的5个API问题...")
    print("=" * 60)

    test_1_range_query()
    test_2_format_range_missing_params()
    test_3_apply_formula_missing_params()
    test_4_search_logic()
    test_5_write_data_format()

    print("\n" + "=" * 60)
    print("✅ 所有测试完成")

if __name__ == "__main__":
    main()
