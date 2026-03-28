#!/usr/bin/env python3
"""
创建测试用的Excel文件用于MCP验证
"""

import os
import sys
import openpyxl
from datetime import datetime

def create_test_excel():
    # 创建测试目录
    test_dir = "test_mcp_verification"
    os.makedirs(test_dir, exist_ok=True)
    
    # 创建测试Excel文件
    wb = openpyxl.Workbook()
    
    # 删除默认工作表
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 创建技能表
    skills_sheet = wb.create_sheet("技能配置", 0)
    headers = ["skill_id", "名称", "类型", "伤害", "冷却时间", "消耗法力", "描述"]
    skills_sheet.append(headers)
    skills_data = [
        [1001, "火球术", "攻击", 100, 3.0, 50, "基础攻击技能"],
        [1002, "冰冻术", "控制", 80, 5.0, 40, "冻结目标"],
        [1003, "治疗术", "治疗", 0, 8.0, 80, "恢复生命值"],
        [1004, "雷电术", "攻击", 150, 4.0, 70, "雷电攻击"],
        [1005, "护盾术", "防御", 0, 6.0, 60, "增加防御力"]
    ]
    for row in skills_data:
        skills_sheet.append(row)
    
    # 创建装备表
    equip_sheet = wb.create_sheet("装备库", 1)
    equip_headers = ["item_id", "名称", "类型", "品质", "攻击力", "防御力", "价格"]
    equip_sheet.append(equip_headers)
    equip_data = [
        [2001, "新手剑", "武器", "普通", 50, 10, 1000],
        [2002, "皮甲", "防具", "普通", 20, 30, 800],
        [2003, "魔法杖", "武器", "稀有", 80, 15, 3000],
        [2004, "盾牌", "防具", "稀有", 10, 50, 2500],
        [2005, "法袍", "防具", "史诗", 40, 80, 8000]
    ]
    for row in equip_data:
        equip_sheet.append(row)
    
    # 创建怪物表
    monster_sheet = wb.create_sheet("怪物数据", 2)
    monster_headers = ["monster_id", "名称", "类型", "生命值", "攻击力", "防御力", "经验值", "掉落率"]
    monster_sheet.append(monster_headers)
    monster_data = [
        [3001, "史莱姆", "小怪", 100, 20, 5, 10, 0.8],
        [3002, "哥布林", "小怪", 150, 30, 10, 20, 0.6],
        [3003, "龙", "BOSS", 1000, 100, 50, 500, 0.3],
        [3004, "巫师", "精英", 300, 80, 20, 100, 0.5],
        [3005, "骑士", "精英", 500, 70, 40, 200, 0.4]
    ]
    for row in monster_data:
        monster_sheet.append(row)
    
    # 保存文件
    test_file = os.path.join(test_dir, "test_game_config.xlsx")
    wb.save(test_file)
    print(f"创建测试文件: {test_file}")
    
    # 创建一个大型文件用于性能测试
    large_wb = openpyxl.Workbook()
    large_sheet = large_wb.active
    large_sheet.title = "大型数据表"
    
    # 添加表头
    large_headers = ["id", "name", "value", "category", "status", "created_time"]
    large_sheet.append(large_headers)
    
    # 添加1000行数据
    for i in range(1000):
        large_sheet.append([i+1, f"项目_{i+1}", (i+1)*100, f"类别_{(i%10)+1}", "正常", datetime.now()])
    
    large_test_file = os.path.join(test_dir, "large_test_data.xlsx")
    large_wb.save(large_test_file)
    print(f"创建大型测试文件: {large_test_file}")
    
    return test_file, large_test_file

if __name__ == "__main__":
    test_file, large_test_file = create_test_excel()
    print(f"测试文件已创建完成:")
    print(f"1. 配置文件: {test_file}")
    print(f"2. 大型文件: {large_test_file}")