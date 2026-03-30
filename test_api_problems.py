#!/usr/bin/env python3
"""
复现监工报告中的5个API问题
"""
import json
import os
import sys
import tempfile
from pathlib import Path
import openpyxl

# 添加项目路径
sys.path.insert(0, str(Path(__file__).parent / "src"))

from excel_mcp_server_fastmcp.server import main

def create_test_excel():
    """创建测试Excel文件"""
    test_file = "test_api_problems.xlsx"
    
    # 创建新的Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 写入测试数据
    ws["A1"] = "测试数据"
    ws["B1"] = 100
    ws["C1"] = "2026-03-30"
    ws["A2"] = "游戏配置"
    ws["B2"] = "技能"
    ws["C2"] = "等级1"
    
    # 保存文件
    wb.save(test_file)
    print(f"✅ 创建测试文件: {test_file}")
    return test_file

def test_api_problems():
    """测试监工报告中的5个API问题"""
    print("🔍 复现监工报告中的5个API问题...")
    
    # 创建测试Excel文件
    test_file = create_test_excel()
    
    try:
        # 测试1: read_data_from_excel 范围查询参数顺序
        print("\n1️⃣ 测试 read_data_from_excel 范围查询参数顺序...")
        result1 = main([
            "read_data_from_excel",
            "--filepath", test_file,
            "--sheet_name", "Sheet1",
            "--start_cell", "B2",
            "--end_cell", "D4"
        ])
        print(f"结果: {json.dumps(result1, indent=2, ensure_ascii=False)}")
        
        # 测试2: format_range 缺少必要参数
        print("\n2️⃣ 测试 format_range 缺少必要参数...")
        result2 = main([
            "format_range",
            "--filepath", test_file,
            "--sheet_name", "Sheet1",
            "--start_cell", "A1",
        ])
        print(f"结果: {json.dumps(result2, indent=2, ensure_ascii=False)}")
        
        # 测试3: apply_formula 缺少 formula 参数
        print("\n3️⃣ 测试 apply_formula 缺少 formula 参数...")
        result3 = main([
            "apply_formula",
            "--filepath", test_file,
            "--sheet_name", "Sheet1",
            "--cell", "A1",
        ])
        print(f"结果: {json.dumps(result3, indent=2, ensure_ascii=False)}")
        
        # 测试4: read_data_from_excel 搜索逻辑
        print("\n4️⃣ 测试 read_data_from_excel 搜索逻辑...")
        result4 = main([
            "read_data_from_excel",
            "--filepath", test_file,
            "--sheet_name", "Sheet1",
            "--start_cell", "A1",
            "--end_cell", "C10"
        ])
        print(f"结果: {json.dumps(result4, indent=2, ensure_ascii=False)}")
        
        # 测试5: write_data_to_excel 数据格式不匹配
        print("\n5️⃣ 测试 write_data_to_excel 数据格式不匹配...")
        # 故意传入错误的数据格式
        result5 = main([
            "write_data_to_excel",
            "--filepath", test_file,
            "--sheet_name", "Sheet1",
            "--data", 'invalid_json_data',
            "--start_cell", "A1"
        ])
        print(f"结果: {json.dumps(result5, indent=2, ensure_ascii=False)}")
        
    finally:
        # 清理测试文件
        if os.path.exists(test_file):
            os.unlink(test_file)
            print(f"🧹 清理测试文件: {test_file}")

if __name__ == "__main__":
    test_api_problems()