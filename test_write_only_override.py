#!/usr/bin/env python3
"""
MCP验证测试 - 第182轮 write_only覆盖修改功能验证

验证新添加的excel_write_only_override工具及其相关功能的实际可用性
"""

import os
import tempfile
import json
from pathlib import Path

# 测试数据
TEST_SKILLS_DATA = [
    ["火球术", 120, 5.0, "火系", 50],
    ["冰箭", 100, 3.5, "冰系", 30],
    ["雷电", 150, 7.2, "雷系", 80],
    ["治疗术", 80, 2.0, "光系", 40],
    ["护盾", 60, 4.0, "土系", 35]
]

TEST_EQUIPMENT_DATA = [
    ["传说剑", 500, 10, "史诗", "武器"],
    ["魔法盾", 300, 8, "稀有", "防具"],
    ["速度靴", 200, 6, "普通", "饰品"]
]

def create_test_excel():
    """创建测试用的Excel文件"""
    try:
        from openpyxl import Workbook
        import tempfile
        
        # 创建临时文件
        fd, temp_path = tempfile.mkstemp(suffix='.xlsx', dir='/tmp')
        os.close(fd)
        
        # 创建测试数据
        wb = Workbook()
        
        # 技能工作表
        ws_skills = wb.active
        ws_skills.title = "技能配置"
        headers = ["技能名称", "伤害值", "冷却时间", "属性", "消耗法力"]
        ws_skills.append(headers)
        for row in TEST_SKILLS_DATA:
            ws_skills.append(row)
        
        # 装备工作表
        ws_equipment = wb.create_sheet("装备列表")
        eq_headers = ["装备名称", "价格", "强化等级", "品质", "类型"]
        ws_equipment.append(eq_headers)
        for row in TEST_EQUIPMENT_DATA:
            ws_equipment.append(row)
        
        # 保存文件
        wb.save(temp_path)
        wb.close()
        
        return temp_path
    except Exception as e:
        print(f"创建测试Excel失败: {e}")
        return None

def test_write_only_override():
    """测试write_only_override功能"""
    print("🧪 开始测试write_only_override功能...")
    
    # 创建测试文件
    test_file = create_test_excel()
    if not test_file:
        print("❌ 测试文件创建失败")
        return False
    
    try:
        # 导入工具
        import sys
        sys.path.insert(0, '/root/.openclaw/workspace/wt-write-only-override/src')
        from excel_mcp_server_fastmcp.server import excel_write_only_override
        
        # 测试1: 更新技能配置
        print("📝 测试1: 更新技能配置...")
        result1 = excel_write_only_override(
            file_path=test_file,
            sheet_name="技能配置",
            range_spec="B2:D6",
            data=[
                [150, 6.0, 60],  # 火球术
                [120, 4.0, 40],  # 冰箭  
                [180, 8.0, 100], # 雷电
                [100, 2.5, 50],  # 治疗术
                [80, 5.0, 45]    # 护盾
            ],
            preserve_formulas=False
        )
        
        if not result1.get('success'):
            print(f"❌ 技能配置更新失败: {result1.get('message')}")
            return False
        
        print(f"✅ 技能配置更新成功: {result1.get('message')}")
        
        # 测试2: 更新装备属性（保留列宽）
        print("📝 测试2: 更新装备属性（保留列宽）...")
        result2 = excel_write_only_override(
            file_path=test_file,
            sheet_name="装备列表", 
            range_spec="B2:D5",
            data=[
                [600, 12, "传说"],  # 传说剑
                [400, 10, "史诗"],  # 魔法盾
                [300, 8, "稀有"]    # 速度靴
            ],
            preserve_col_widths=True
        )
        
        if not result2.get('success'):
            print(f"❌ 装备属性更新失败: {result2.get('message')}")
            return False
            
        print(f"✅ 装备属性更新成功: {result2.get('message')}")
        
        # 测试3: 验证数据写入结果
        print("📊 测试3: 验证数据写入结果...")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(test_file, data_only=True)
            
            # 检查技能配置
            ws_skills = wb["技能配置"]
            # 第2行：火球术
            assert ws_skills.cell(2, 2).value == 150  # 伤害值
            assert ws_skills.cell(2, 3).value == 6.0  # 冷却时间
            assert ws_skills.cell(2, 4).value == 60  # 属性
            
            # 第3行：冰箭
            assert ws_skills.cell(3, 2).value == 120
            assert ws_skills.cell(3, 3).value == 4.0
            
            # 检查装备列表
            ws_equipment = wb["装备列表"]
            # 第2行：传说剑
            assert ws_equipment.cell(2, 2).value == 600  # 价格
            assert ws_equipment.cell(2, 3).value == 12   # 强化等级
            assert ws_equipment.cell(2, 4).value == "传说"  # 品质
            
            wb.close()
            print("✅ 数据验证成功：所有写入数据正确")
            
        except Exception as verify_err:
            print(f"❌ 数据验证失败: {verify_err}")
            return False
        
        # 清理
        os.unlink(test_file)
        
        return True
        
    except Exception as e:
        print(f"❌ 测试过程异常: {e}")
        if os.path.exists(test_file):
            os.unlink(test_file)
        return False

def test_mcp_tools_count():
    """验证工具数量是否正确"""
    try:
        import sys
        sys.path.insert(0, '/root/.openclaw/workspace/wt-write-only-override/src')
        from excel_mcp_server_fastmcp.server import server
        
        # 计算工具数量
        tool_count = len([attr for attr in dir(server) if attr.startswith('excel_') and callable(getattr(server, attr))])
        print(f"🔢 当前工具数量: {tool_count}")
        
        if tool_count != 53:  # 原有52个 + 新增1个
            print(f"⚠️  工具数量不符合预期（预期53个，实际{tool_count}个）")
            return False
            
        print("✅ 工具数量验证通过")
        return True
        
    except Exception as e:
        print(f"❌ 工具数量验证失败: {e}")
        return False

def main():
    """主测试函数"""
    print("🚀 开始第182轮MCP验证 - write_only覆盖修改功能")
    print("=" * 60)
    
    results = []
    
    # 测试1: 工具数量验证
    print("\n📊 测试1: 工具数量验证")
    results.append(("工具数量", test_mcp_tools_count()))
    
    # 测试2: write_only_override功能测试
    print("\n🔧 测试2: write_only_override功能测试")
    results.append(("write_only_override", test_write_only_override()))
    
    # 汇总结果
    print("\n" + "=" * 60)
    print("📋 测试结果汇总:")
    passed = 0
    total = len(results)
    
    for test_name, result in results:
        status = "✅ 通过" if result else "❌ 失败"
        print(f"  {test_name}: {status}")
        if result:
            passed += 1
    
    print(f"\n📊 总体结果: {passed}/{total} 测试通过")
    
    if passed == total:
        print("🎉 所有测试通过！write_only覆盖修改功能验证成功")
        return True
    else:
        print("❌ 部分测试失败，需要修复")
        return False

if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)