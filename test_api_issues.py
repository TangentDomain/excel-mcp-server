#!/usr/bin/env python3
"""
Simple test script to reproduce the 5 API problems using direct MCP calls.
"""

import json
import sys
import os
import tempfile
from pathlib import Path

# Add src to path
sys.path.insert(0, 'src')

def create_test_excel():
    """Create a simple test Excel file"""
    from openpyxl import Workbook
    import shutil
    
    # Create temp file
    test_file = "/tmp/test_api.xlsx"
    if os.path.exists(test_file):
        os.remove(test_file)
    
    # Create workbook with test data
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add headers and data
    headers = ["Name", "Age", "City", "Salary"]
    data = [
        ["Alice", 25, "Shenzhen", 10000],
        ["Bob", 30, "Beijing", 12000],
        ["Charlie", 35, "Shanghai", 15000]
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    wb.save(test_file)
    return test_file

def test_1_range_query():
    """Test 1: excel_get_range parameter order and validation"""
    print("=== Test 1: excel_get_range validation ===")
    test_file = create_test_excel()
    
    try:
        from excel_mcp_server_fastmcp.server import excel_get_range
        
        # Test with valid parameters
        result = excel_get_range(
            file_path=test_file,
            sheet_name="Sheet1", 
            start_cell="A1",
            end_cell="C3"
        )
        print(f"Result: {json.dumps(result, indent=2, ensure_ascii=False)}")
        if result.get('success'):
            print("✅ Range query passed")
        else:
            print(f"❌ Range query failed: {result.get('message')}")
            
    except Exception as e:
        print(f"❌ Range query exception: {e}")
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)

def test_2_format_cells_validation():
    """Test 2: excel_format_cells parameter validation"""
    print("\n=== Test 2: excel_format_cells parameter validation ===")
    test_file = create_test_excel()
    
    try:
        from excel_mcp_server_fastmcp.server import excel_format_cells
        
        # Test without required formatting parameters
        result = excel_format_cells(
            file_path=test_file,
            sheet_name="Sheet1",
            start_cell="A1"
            # Missing bold, italic, etc.
        )
        print(f"Result: {json.dumps(result, indent=2, ensure_ascii=False)}")
        if result.get('success'):
            print("✅ Format cells passed")
        else:
            print(f"❌ Format cells failed: {result.get('message')}")
            
    except Exception as e:
        print(f"❌ Format cells exception: {e}")
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)

def test_3_set_formula_validation():
    """Test 3: excel_set_formula parameter validation"""
    print("\n=== Test 3: excel_set_formula validation ===")
    test_file = create_test_excel()
    
    try:
        from excel_mcp_server_fastmcp.server import excel_set_formula
        
        # Test without formula parameter
        result = excel_set_formula(
            file_path=test_file,
            sheet_name="Sheet1",
            cell_address="D1"
            # Missing formula parameter
        )
        print(f"Result: {json.dumps(result, indent=2, ensure_ascii=False)}")
        if result.get('success'):
            print("✅ Set formula passed")
        else:
            print(f"❌ Set formula failed: {result.get('message')}")
            
    except Exception as e:
        print(f"❌ Set formula exception: {e}")
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)

def test_4_search_confusion():
    """Test 4: excel_search parameter confusion"""
    print("\n=== Test 4: excel_search parameter confusion ===")
    test_file = create_test_excel()
    
    try:
        from excel_mcp_server_fastmcp.server import excel_search
        
        # Test search with various parameter combinations
        result = excel_search(
            file_path=test_file,
            pattern="Alice",
            sheet_name="Sheet1",
            case_sensitive=False,
            whole_word=True
        )
        print(f"Result: {json.dumps(result, indent=2, ensure_ascii=False)}")
        if result.get('success'):
            print("✅ Search passed")
        else:
            print(f"❌ Search failed: {result.get('message')}")
            
    except Exception as e:
        print(f"❌ Search exception: {e}")
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)

def test_5_write_data_format():
    """Test 5: excel_update_range data format handling"""
    print("\n=== Test 5: excel_update_range data format ===")
    test_file = create_test_excel()
    
    try:
        from excel_mcp_server_fastmcp.server import excel_update_range
        
        # Test with different data formats
        test_data = [["David", 40, "Hangzhou", 18000]]  # Single row list of lists
        
        result = excel_update_range(
            file_path=test_file,
            sheet_name="Sheet1",
            data=test_data,
            start_cell="A4",
            # Missing insert_mode parameter
        )
        print(f"Result: {json.dumps(result, indent=2, ensure_ascii=False)}")
        if result.get('success'):
            print("✅ Update range passed")
        else:
            print(f"❌ Update range failed: {result.get('message')}")
            
    except Exception as e:
        print(f"❌ Update range exception: {e}")
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)

if __name__ == "__main__":
    test_1_range_query()
    test_2_format_cells_validation()
    test_3_set_formula_validation()
    test_4_search_confusion()
    test_5_write_data_format()