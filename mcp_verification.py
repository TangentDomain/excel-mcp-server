#!/usr/bin/env python3
"""
MCP真实验证 - 修正版
"""

import sys
import os
sys.path.insert(0, '.')

from src.excel_mcp_server_fastmcp.server import (
    excel_list_sheets,
    excel_get_range,
    excel_query,
    excel_get_headers,
    excel_find_last_row,
    excel_batch_insert_rows,
    excel_delete_rows,
    excel_describe_table,
    excel_search,
    excel_get_file_info,
    excel_update_range,
)
import tempfile

def test_mcp_verification():
    results = {}
    
    # 创建测试文件
    tmpdir = tempfile.mkdtemp()
    test_file = os.path.join(tmpdir, "mcp_test.xlsx")
    
    import openpyxl
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    
    # 创建技能表（用英文名避免编码问题）
    ws1 = wb.create_sheet("Skills")
    ws1.append(["skill_id", "name", "type", "damage", "cooldown"])
    ws1.append([1001, "Fireball", "attack", 100, 3.0])
    ws1.append([1002, "Ice", "control", 80, 5.0])
    ws1.append([1003, "Heal", "heal", 0, 8.0])
    ws1.append([1004, "Lightning", "attack", 150, 4.0])
    ws1.append([1005, "Shield", "defense", 0, 6.0])
    
    # 创建装备表
    ws2 = wb.create_sheet("Equipment")
    ws2.append(["item_id", "name", "quality", "attack"])
    ws2.append([2001, "Sword", "normal", 50])
    ws2.append([2002, "Armor", "normal", 20])
    ws2.append([2003, "Staff", "rare", 80])
    
    wb.save(test_file)
    wb.close()
    
    print(f"Test file: {test_file}")
    print("=" * 60)
    
    # 1. list_sheets
    print("\n[1/12] excel_list_sheets")
    try:
        r = excel_list_sheets(test_file)
        assert r['success'], f"Failed: {r.get('message')}"
        assert len(r['data']['sheets']) == 2
        print(f"  PASS: {r['data']['sheets']}")
        results['list_sheets'] = True
    except Exception as e:
        print(f"  FAIL: {e}")
        results['list_sheets'] = False
    
    # 2. get_range (use sheet!range format)
    print("\n[2/12] excel_get_range")
    try:
        r = excel_get_range(test_file, "Skills!A1:E6")
        assert r['success'], f"Failed: {r.get('message')}"
        assert len(r['data']['data']) == 6
        print(f"  PASS: {len(r['data']['data'])} rows")
        results['get_range'] = True
    except Exception as e:
        print(f"  FAIL: {e}")
        results['get_range'] = False
    
    # 3. query WHERE
    print("\n[3/12] excel_query WHERE")
    try:
        r = excel_query(test_file, "SELECT * FROM Skills WHERE type = 'attack'")
        assert r['success'], f"Failed: {r.get('message')}"
        assert len(r['data']['data']) == 2
        print(f"  PASS: {len(r['data']['data'])} rows")
        results['query_where'] = True
    except Exception as e:
        print(f"  FAIL: {e}")
        results['query_where'] = False
    
    # 4. query JOIN
    print("\n[4/12] excel_query JOIN")
    try:
        r = excel_query(test_file, "SELECT a.name, b.name FROM Skills a JOIN Equipment b ON a.skill_id = b.item_id")
        assert r['success'], f"Failed: {r.get('message')}"
        print(f"  PASS: JOIN returned {len(r['data'].get('data', []))} rows")
        results['query_join'] = True
    except Exception as e:
        print(f"  FAIL: {e}")
        results['query_join'] = False
    
    # 5. query GROUP BY
    print("\n[5/12] excel_query GROUP BY")
    try:
        r = excel_query(test_file, "SELECT type, COUNT(*) as cnt FROM Skills GROUP BY type")
        assert r['success'], f"Failed: {r.get('message')}"
        assert len(r['data']['data']) > 0
        print(f"  PASS: {len(r['data']['data'])} groups")
        results['query_group_by'] = True
    except Exception as e:
        print(f"  FAIL: {e}")
        results['query_group_by'] = False
    
    # 6. query 子查询
    print("\n[6/12] excel_query subquery")
    try:
        r = excel_query(test_file, "SELECT * FROM Skills WHERE damage > (SELECT AVG(damage) FROM Skills)")
        assert r['success'], f"Failed: {r.get('message')}"
        print(f"  PASS: {len(r['data']['data'])} rows")
        results['query_subquery'] = True
    except Exception as e:
        print(f"  FAIL: {e}")
        results['query_subquery'] = False
    
    # 7. query FROM子查询
    print("\n[7/12] excel_query FROM subquery")
    try:
        r = excel_query(test_file, "SELECT * FROM (SELECT * FROM Skills WHERE type = 'attack') WHERE damage > 100")
        assert r['success'], f"Failed: {r.get('message')}"
        print(f"  PASS: FROM subquery returned {len(r['data']['data'])} rows")
        results['query_from_subquery'] = True
    except Exception as e:
        print(f"  FAIL: {e}")
        results['query_from_subquery'] = False
    
    # 8. get_headers
    print("\n[8/12] excel_get_headers")
    try:
        r = excel_get_headers(test_file, "Skills")
        assert r['success'], f"Failed: {r.get('message')}"
        print(f"  PASS: {r['data']['headers']}")
        results['get_headers'] = True
    except Exception as e:
        print(f"  FAIL: {e}")
        results['get_headers'] = False
    
    # 9. find_last_row
    print("\n[9/12] excel_find_last_row")
    try:
        r = excel_find_last_row(test_file, "Skills")
        assert r['success'], f"Failed: {r.get('message')}"
        assert r['data']['last_row'] == 6
        print(f"  PASS: last_row = {r['data']['last_row']}")
        results['find_last_row'] = True
    except Exception as e:
        print(f"  FAIL: {e}")
        results['find_last_row'] = False
    
    # 10. batch_insert_rows (use dict format)
    print("\n[10/12] excel_batch_insert_rows")
    try:
        r = excel_batch_insert_rows(test_file, "Skills", [{"skill_id": 1006, "name": "Test", "type": "test", "damage": 50, "cooldown": 2.0}])
        assert r['success'], f"Failed: {r.get('message')}"
        print(f"  PASS: inserted")
        results['batch_insert_rows'] = True
    except Exception as e:
        print(f"  FAIL: {e}")
        results['batch_insert_rows'] = False
    
    # 11. delete_rows (check correct parameter names)
    print("\n[11/12] excel_delete_rows")
    try:
        import inspect
        sig = inspect.signature(excel_delete_rows)
        print(f"  Parameters: {list(sig.parameters.keys())}")
        r = excel_delete_rows(test_file, "Skills", 1, 7, 7)
        assert r['success'], f"Failed: {r.get('message')}"
        print(f"  PASS: deleted")
        results['delete_rows'] = True
    except Exception as e:
        print(f"  FAIL: {e}")
        results['delete_rows'] = False
    
    # 12. describe_table
    print("\n[12/12] excel_describe_table")
    try:
        r = excel_describe_table(test_file, "Skills")
        assert r['success'], f"Failed: {r.get('message')}"
        assert len(r['data']['columns']) == 5
        print(f"  PASS: {len(r['data']['columns'])} columns")
        results['describe_table'] = True
    except Exception as e:
        print(f"  FAIL: {e}")
        results['describe_table'] = False
    
    # 清理
    os.remove(test_file)
    os.rmdir(tmpdir)
    
    # 汇总
    print("\n" + "=" * 60)
    passed = sum(1 for v in results.values() if v)
    total = len(results)
    print(f"MCP Verification: {passed}/{total} passed")
    for name, ok in results.items():
        print(f"  {'PASS' if ok else 'FAIL'} {name}")
    
    return passed, total

if __name__ == "__main__":
    passed, total = test_mcp_verification()
    sys.exit(0 if passed == total else 1)