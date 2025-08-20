#!/usr/bin/env python3
"""
测试openpyxl的计算引擎
"""

import sys
import tempfile
import os
from pathlib import Path

# 添加src路径
sys.path.insert(0, str(Path(__file__).parent / "src"))

from openpyxl import Workbook, load_workbook

def test_calculation_engine():
    """测试计算引擎"""

    print("🔧 测试Excel计算引擎...")

    # 创建工作簿
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test"

    # 添加数据
    sheet['A1'] = 10
    sheet['A2'] = 20
    sheet['A3'] = 30

    # 启用自动计算
    workbook.calculation.calcMode = 'auto'

    # 添加公式
    sheet['B1'] = "=SUM(A1:A3)"
    print(f"📝 设置公式后 B1: {sheet['B1'].value}")

    # 保存文件
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()

    try:
        workbook.save(temp_file.name)
        workbook.close()

        # 重新打开文件并强制计算
        wb = load_workbook(temp_file.name, data_only=False)
        ws = wb.active

        # 手动触发计算（虽然openpyxl没有直接的计算引擎）
        # 我们可以通过重新设置公式来刷新
        original_formula = ws['B1'].value
        ws['B1'] = original_formula

        wb.save(temp_file.name)
        wb.close()

        # 使用LibreOffice或其他方式计算（如果可用）
        # 但最简单的方法是使用xlwings或python-excel等库

        # 读取结果
        result_wb = load_workbook(temp_file.name, data_only=True)
        result_ws = result_wb.active
        result = result_ws['B1'].value
        print(f"🎯 计算结果: {result}")
        result_wb.close()

        # 尝试使用xlcalculator（如果安装了）
        try:
            from xlcalculator import ModelCompiler, Evaluator
            print("📊 尝试使用xlcalculator...")

            # 编译模型
            compiler = ModelCompiler()
            model = compiler.read_and_parse_archive(temp_file.name)
            evaluator = Evaluator(model)

            # 计算B1
            val = evaluator.evaluate('Test!B1')
            print(f"📊 xlcalculator结果: {val}")

        except ImportError:
            print("⚠️  xlcalculator未安装，无法使用")

            # 手动计算作为回退
            print("📊 使用手动计算...")
            # 简单的SUM函数手动实现
            if original_formula == "=SUM(A1:A3)":
                a1 = result_ws['A1'].value
                a2 = result_ws['A2'].value
                a3 = result_ws['A3'].value
                manual_result = a1 + a2 + a3
                print(f"🔢 手动计算结果: {manual_result}")

    finally:
        try:
            os.unlink(temp_file.name)
        except:
            pass

    print("✅ 测试完成")

if __name__ == "__main__":
    test_calculation_engine()
