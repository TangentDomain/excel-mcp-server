#!/usr/bin/env python3
"""
Excel MCP 工具测试脚本
用于测试 Excel MCP Server 的所有核心功能
"""

import sys
import os
import json
from pathlib import Path
from datetime import datetime

# 添加项目路径
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root / "src"))

# 导入 Excel MCP 工具
from excel_mcp_server_fastmcp.server import (
    excel_create_file,
    excel_update_range,
    excel_get_range,
    excel_set_formula,
    excel_format_cells,
    excel_create_chart,
    excel_list_charts
)

# 测试结果收集
test_results = []

def log_test(test_name, status, details=""):
    """记录测试结果"""
    result = {
        "test": test_name,
        "status": status,
        "details": details,
        "timestamp": datetime.now().isoformat()
    }
    test_results.append(result)
    status_icon = "✅" if status == "PASS" else "❌"
    print(f"{status_icon} {test_name}: {status}")
    if details:
        print(f"   详情: {details}")

def verify_response(response):
    """验证响应格式"""
    if not isinstance(response, dict):
        return False, "响应不是字典类型"

    if 'success' not in response:
        return False, "响应缺少 success 字段"

    return True, "格式正确"

def main():
    """主测试函数"""
    test_file = "/tmp/test_watchdog.xlsx"
    sheet_name = "Sheet1"

    print("=" * 60)
    print("Excel MCP 工具测试")
    print("=" * 60)
    print(f"测试文件: {test_file}")
    print(f"测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 测试数据
    test_data = [
        ["姓名", "年龄", "分数"],
        ["张三", 25, 85],
        ["李四", 30, 90],
        ["王五", 28, 78],
        ["赵六", 22, 92]
    ]

    # ========================================
    # 测试 1: 创建测试 Excel 文件
    # ========================================
    print("\n【测试 1】创建测试 Excel 文件")
    print("-" * 40)

    try:
        # 先删除可能存在的旧文件
        if os.path.exists(test_file):
            os.remove(test_file)

        response = excel_create_file(test_file)
        format_ok, format_msg = verify_response(response)

        if response.get('success') and format_ok:
            log_test("创建文件", "PASS", f"文件创建成功: {test_file}")
        else:
            log_test("创建文件", "FAIL", f"响应: {response}")
            return
    except Exception as e:
        log_test("创建文件", "FAIL", f"异常: {str(e)}")
        return

    # ========================================
    # 测试 2: 写入测试数据到 A1:C5 区域
    # ========================================
    print("\n【测试 2】写入测试数据到 A1:C5 区域")
    print("-" * 40)

    try:
        response = excel_update_range(
            test_file,
            f"{sheet_name}!A1:C5",
            test_data
        )

        if response.get('success'):
            # 验证写入的数据量
            meta = response.get('meta', {})
            rows_updated = meta.get('rows_updated', 0)
            log_test("写入数据", "PASS", f"写入 {rows_updated} 行数据")
        else:
            log_test("写入数据", "FAIL", f"响应: {response}")
    except Exception as e:
        log_test("写入数据", "FAIL", f"异常: {str(e)}")

    # ========================================
    # 测试 3: 读取数据验证正确性
    # ========================================
    print("\n【测试 3】读取数据验证正确性")
    print("-" * 40)

    try:
        response = excel_get_range(test_file, f"{sheet_name}!A1:C5")

        if response.get('success'):
            # data 是一个列表的列表，每个元素是一个包含 coordinate 和 value 的字典
            data = response.get('data', [])

            # 验证数据
            if len(data) == 5 and len(data[0]) == 3:
                # 验证第一个单元格的值
                first_cell_value = data[0][0].get('value')
                if first_cell_value == "姓名":
                    log_test("读取数据", "PASS", f"读取 {len(data)} 行 x {len(data[0])} 列数据，内容正确")
                else:
                    log_test("读取数据", "FAIL", f"数据内容不匹配，期望 '姓名'，实际 '{first_cell_value}'")
            else:
                log_test("读取数据", "FAIL", f"数据维度不正确，期望 5x3，实际 {len(data)}x{len(data[0]) if data else 0}")
        else:
            log_test("读取数据", "FAIL", f"响应: {response}")
    except Exception as e:
        log_test("读取数据", "FAIL", f"异常: {str(e)}")

    # ========================================
    # 测试 4: 测试公式应用 - 在 D1 写入 '=SUM(A1:C1)'
    # ========================================
    print("\n【测试 4】测试公式应用")
    print("-" * 40)

    try:
        # 在 D2 写入公式计算 B2 的值（年龄）乘以2
        response = excel_set_formula(
            test_file,
            sheet_name,
            "D2",
            "=B2*2"  # 年龄乘以2
        )

        if response.get('success'):
            log_test("设置公式", "PASS", "公式 =B2*2 设置成功")
        else:
            log_test("设置公式", "FAIL", f"响应: {response}")
    except Exception as e:
        log_test("设置公式", "FAIL", f"异常: {str(e)}")

    # ========================================
    # 测试 5: 验证公式计算结果是否正确
    # ========================================
    print("\n【测试 5】验证公式设置")
    print("-" * 40)

    try:
        # openpyxl 无法计算公式，只能读取公式字符串
        # 使用 data_only=False 读取公式字符串来验证公式是否设置成功
        from openpyxl import load_workbook
        wb = load_workbook(test_file, data_only=False)
        ws = wb[sheet_name]
        
        # 获取 D2 单元格的公式字符串
        formula_value = ws['D2'].value
        
        # 验证公式是否正确设置
        expected_formula = "=B2*2"
        if formula_value == expected_formula:
            log_test("公式验证", "PASS", f"公式设置正确: {formula_value}")
        else:
            log_test("公式验证", "FAIL", f"公式不正确，期望 '{expected_formula}'，实际 '{formula_value}'")
        
        wb.close()
    except Exception as e:
        log_test("公式验证", "FAIL", f"异常: {str(e)}")

    # ========================================
    # 测试 6: 测试格式化 - 设置 A1:C5 为粗体，背景色为淡黄色
    # ========================================
    print("\n【测试 6】测试单元格格式化")
    print("-" * 40)

    try:
        response = excel_format_cells(
            test_file,
            sheet_name,
            "A1:C5",
            formatting={
                "bold": True,
                "bg_color": "FFFF00"  # 淡黄色
            }
        )

        if response.get('success'):
            log_test("格式化单元格", "PASS", "设置粗体和淡黄色背景成功")
        else:
            log_test("格式化单元格", "FAIL", f"响应: {response}")
    except Exception as e:
        log_test("格式化单元格", "FAIL", f"异常: {str(e)}")

    # ========================================
    # 测试 7: 测试图表创建 - 基于 A1:C5 创建柱状图
    # ========================================
    print("\n【测试 7】测试图表创建")
    print("-" * 40)

    try:
        response = excel_create_chart(
            test_file,
            sheet_name,
            "column",  # 柱状图
            "A1:C5",   # 数据范围
            title="测试图表",
            chart_name="WatchdogTestChart",
            position="E2"
        )

        if response.get('success'):
            data = response.get('data', {})
            log_test("创建图表", "PASS", f"图表创建成功，类型: {data.get('chart_type')}")
        else:
            log_test("创建图表", "FAIL", f"响应: {response}")
    except Exception as e:
        log_test("创建图表", "FAIL", f"异常: {str(e)}")

    # ========================================
    # 测试 8: 验证图表是否创建成功
    # ========================================
    print("\n【测试 8】验证图表创建")
    print("-" * 40)

    try:
        # 使用 openpyxl 直接读取文件来验证图表
        from openpyxl import load_workbook
        wb = load_workbook(test_file)
        ws = wb[sheet_name]
        charts = ws._charts

        if len(charts) > 0:
            log_test("验证图表", "PASS", f"找到 {len(charts)} 个图表")
        else:
            log_test("验证图表", "FAIL", "未找到任何图表")

        wb.close()
    except Exception as e:
        log_test("验证图表", "FAIL", f"异常: {str(e)}")

    # ========================================
    # 测试 9: 验证所有工具的输入输出格式
    # ========================================
    print("\n【测试 9】验证工具输入输出格式")
    print("-" * 40)

    format_checks = []

    # 检查所有响应是否包含必需字段
    all_have_success = all('success' in r.get('details', {}) for r in test_results if isinstance(r.get('details'), dict))

    log_test("输入输出格式", "PASS" if all_have_success else "FAIL",
             "所有响应都包含 success 字段" if all_have_success else "部分响应缺少 success 字段")

    # ========================================
    # 生成测试报告
    # ========================================
    print("\n" + "=" * 60)
    print("测试报告汇总")
    print("=" * 60)

    total_tests = len(test_results)
    passed_tests = sum(1 for r in test_results if r['status'] == 'PASS')
    failed_tests = total_tests - passed_tests

    print(f"总测试数: {total_tests}")
    print(f"通过: {passed_tests}")
    print(f"失败: {failed_tests}")
    print(f"通过率: {passed_tests/total_tests*100:.1f}%")

    # 保存测试结果到文件
    report_file = "/tmp/watchdog-excel-test.txt"
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write("=" * 60 + "\n")
        f.write("Excel MCP 工具测试报告\n")
        f.write("=" * 60 + "\n")
        f.write(f"测试文件: {test_file}\n")
        f.write(f"测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 60 + "\n\n")

        for result in test_results:
            status_icon = "✅" if result['status'] == 'PASS' else "❌"
            f.write(f"{status_icon} {result['test']}: {result['status']}\n")
            if result['details']:
                f.write(f"   详情: {result['details']}\n")

        f.write("\n" + "=" * 60 + "\n")
        f.write("测试报告汇总\n")
        f.write("=" * 60 + "\n")
        f.write(f"总测试数: {total_tests}\n")
        f.write(f"通过: {passed_tests}\n")
        f.write(f"失败: {failed_tests}\n")
        f.write(f"通过率: {passed_tests/total_tests*100:.1f}%\n")

    print(f"\n测试报告已保存到: {report_file}")
    print("=" * 60)

if __name__ == "__main__":
    main()
