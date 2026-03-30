#!/usr/bin/env python3
"""
测试API问题的脚本
"""
import openpyxl
import os
import json
import sys
from pathlib import Path

def create_test_file():
    """创建测试Excel文件"""
    test_file = 'test_api_validation.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws['A1'] = 'ID'
    ws['B1'] = 'Name'
    ws['C1'] = 'Value'
    ws['A2'] = 1
    ws['B2'] = '技能'
    ws['C2'] = 100
    wb.save(test_file)
    print(f'✅ 创建测试文件: {test_file}')
    return test_file

def test_api_problems():
    """测试API问题"""
    test_file = create_test_file()
    
    # 测试案例
    test_cases = [
        {
            'name': 'read_data_from_excel 范围查询',
            'params': ['--filepath', test_file, '--sheet_name', 'Sheet1', '--start_cell', 'B2', '--end_cell', 'D4'],
            'expected_issue': '参数顺序可能颠倒'
        },
        {
            'name': 'format_range 缺少参数', 
            'params': ['--filepath', test_file, '--sheet_name', 'Sheet1', '--start_cell', 'A1'],
            'expected_issue': '缺少bold等必要参数'
        },
        {
            'name': 'apply_formula 缺少formula',
            'params': ['--filepath', test_file, '--sheet_name', 'Sheet1', '--cell', 'A1'],
            'expected_issue': '缺少formula参数'
        },
        {
            'name': 'write_data_to_excel 格式错误',
            'params': ['--filepath', test_file, '--sheet_name', 'Sheet1', '--data', 'invalid_json', '--start_cell', 'A1'],
            'expected_issue': '数据格式不匹配'
        }
    ]
    
    print('🔍 开始测试API问题...')
    for i, test in enumerate(test_cases, 1):
        print(f'\n{i}. 测试 {test["name"]}: 预期问题 - {test["expected_issue"]}')
    
    # 清理
    os.unlink(test_file)
    print('\n✅ 测试完成')

if __name__ == "__main__":
    test_api_problems()