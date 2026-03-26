#!/usr/bin/env python3
"""
ExcelMCP 性能基准测试脚本

独立运行，测量各核心操作的性能并输出结构化报告。
支持历史结果对比，用于性能回归检测。

用法:
    python3 scripts/benchmark.py                  # 运行基准测试
    python3 scripts/benchmark.py --compare        # 与上次结果对比
    python3 scripts/benchmark.py --output report.json  # 指定输出文件
    python3 scripts/benchmark.py --quick          # 快速模式（跳过大表测试）

输出:
    默认保存到 .benchmark-history/benchmark_YYYYMMDD_HHMMSS.json
    --compare 时输出与上次结果的对比表格
"""

import argparse
import json
import os
import statistics
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path

# 确保项目根目录在 Python 路径中
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
from src.excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query


def _query(file_path: str, sql: str):
    """SQL查询的快捷调用（与server.py中excel_query一致）"""
    return execute_advanced_sql_query(file_path, sql)


# ==================== 测试数据生成 ====================

def create_game_config_file(file_path: str, rows: int, sheets: int = 1,
                            dual_header: bool = True) -> float:
    """创建模拟游戏配置表的Excel文件，返回创建耗时（秒）"""
    from openpyxl import Workbook

    wb = Workbook()
    skill_types = ["战士", "法师", "刺客", "辅助", "射手"]
    rarities = ["普通", "精良", "史诗", "传说"]
    elements = ["火", "冰", "雷", "暗", "光"]

    for sheet_idx in range(sheets):
        if sheet_idx == 0:
            ws = wb.active
            ws.title = "技能配置" if dual_header else "Skills"
        else:
            ws = wb.create_sheet(f"配置表{sheet_idx + 1}" if dual_header else f"Sheet{sheet_idx + 1}")

        col_count = 10
        cn_headers = ["技能名称", "技能类型", "伤害", "冷却时间", "消耗法力",
                       "技能等级", "品质", "元素属性", "描述", "是否被动"]
        en_headers = ["skill_name", "skill_type", "damage", "cooldown", "mana_cost",
                       "level", "rarity", "element", "description", "is_passive"]

        if dual_header:
            for col, header in enumerate(cn_headers, start=1):
                ws.cell(row=1, column=col, value=header)
            for col, header in enumerate(en_headers, start=1):
                ws.cell(row=2, column=col, value=header)
            data_start_row = 3
        else:
            for col, header in enumerate(en_headers, start=1):
                ws.cell(row=1, column=col, value=header)
            data_start_row = 2

        import random
        random.seed(42)

        for row_offset in range(rows):
            r = data_start_row + row_offset
            ws.cell(row=r, column=1, value=f"技能_{row_offset + 1}")
            ws.cell(row=r, column=2, value=skill_types[row_offset % len(skill_types)])
            ws.cell(row=r, column=3, value=random.randint(10, 500))
            ws.cell(row=r, column=4, value=round(random.uniform(0.5, 30.0), 1))
            ws.cell(row=r, column=5, value=random.randint(0, 200))
            ws.cell(row=r, column=6, value=random.randint(1, 10))
            ws.cell(row=r, column=7, value=rarities[row_offset % len(rarities)])
            ws.cell(row=r, column=8, value=elements[row_offset % len(elements)])
            ws.cell(row=r, column=9, value=f"第{row_offset + 1}号技能的描述文本")
            ws.cell(row=r, column=10, value=random.choice([True, False]))

    wb.save(file_path)
    return 0


# ==================== 基准测试项 ====================

class BenchmarkResult:
    """单个基准测试的结果"""

    def __init__(self, name: str, category: str):
        self.name = name
        self.category = category
        self.times: list[float] = []
        self.success = True
        self.error: str | None = None
        self.extra: dict = {}

    def add_time(self, t: float):
        self.times.append(t)

    @property
    def avg_ms(self) -> float:
        return statistics.mean(self.times) * 1000 if self.times else 0

    @property
    def min_ms(self) -> float:
        return min(self.times) * 1000 if self.times else 0

    @property
    def max_ms(self) -> float:
        return max(self.times) * 1000 if self.times else 0

    @property
    def median_ms(self) -> float:
        return statistics.median(self.times) * 1000 if self.times else 0

    def to_dict(self) -> dict:
        d = {
            "name": self.name,
            "category": self.category,
            "success": self.success,
            "runs": len(self.times),
            "avg_ms": round(self.avg_ms, 2),
            "min_ms": round(self.min_ms, 2),
            "max_ms": round(self.max_ms, 2),
            "median_ms": round(self.median_ms, 2),
        }
        if self.error:
            d["error"] = self.error
        if self.extra:
            d["extra"] = self.extra
        return d


def measure(name: str, category: str, func, *args, runs: int = 3, warmup: int = 0) -> BenchmarkResult:
    """执行函数多次并收集耗时"""
    result = BenchmarkResult(name, category)

    # 预热
    for _ in range(warmup):
        try:
            func(*args)
        except Exception:
            pass

    for _ in range(runs):
        start = time.perf_counter()
        try:
            func(*args)
            result.add_time(time.perf_counter() - start)
        except Exception as e:
            result.success = False
            result.error = str(e)
            break

    return result


# ==================== 测试函数 ====================


class Benchmarks:
    """所有基准测试项的集合"""

    def __init__(self, small_file: str, medium_file: str, large_file: str, quick: bool):
        self.small_file = small_file
        self.medium_file = medium_file
        self.large_file = large_file
        self.quick = quick
        self.results: list[BenchmarkResult] = []

    def run_all(self):
        """运行所有基准测试"""
        self.results = []

        print("=" * 60, flush=True)
        print("  ExcelMCP Benchmark", flush=True)
        print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", flush=True)
        print("=" * 60, flush=True)

        self._run_sql_benchmarks()
        self._run_read_benchmarks()
        self._run_search_benchmarks()
        self._run_write_benchmarks()

        if not self.quick:
            self._run_large_table_benchmarks()

        print()
        return self.results

    def _run_sql_benchmarks(self):
        """SQL查询性能测试"""
        print("\n📊 SQL查询性能测试")
        print("-" * 40)

        # 1. 简单查询
        r = measure("sql_select_all_limit50", "sql",
                     lambda: execute_advanced_sql_query(self.small_file, "SELECT * FROM 技能配置 LIMIT 50"),
                     runs=5, warmup=1)
        self.results.append(r)
        print(f"  SELECT * LIMIT 50 (小表):     {r.avg_ms:.1f}ms (median {r.median_ms:.1f})")

        # 2. WHERE条件查询
        r = measure("sql_where_and", "sql",
                     lambda: execute_advanced_sql_query(self.medium_file,
                         "SELECT * FROM 技能配置 WHERE 技能类型='法师' AND 伤害>100"),
                     runs=5, warmup=1)
        self.results.append(r)
        print(f"  WHERE + AND (中表):            {r.avg_ms:.1f}ms (median {r.median_ms:.1f})")

        # 3. GROUP BY聚合
        r = measure("sql_group_by", "sql",
                     lambda: execute_advanced_sql_query(self.medium_file,
                         "SELECT 技能类型, AVG(伤害) as avg_dmg, COUNT(*) as cnt "
                         "FROM 技能配置 GROUP BY 技能类型"),
                     runs=5, warmup=1)
        self.results.append(r)
        print(f"  GROUP BY聚合 (中表):           {r.avg_ms:.1f}ms (median {r.median_ms:.1f})")

        # 4. 复杂查询（多条件+排序+聚合）
        r = measure("sql_complex", "sql",
                     lambda: execute_advanced_sql_query(self.medium_file,
                         "SELECT 技能类型, AVG(伤害) as avg_dmg, MAX(冷却时间) as max_cd "
                         "FROM 技能配置 WHERE 伤害>50 AND 技能等级>=3 "
                         "GROUP BY 技能类型 ORDER BY avg_dmg DESC LIMIT 5"),
                     runs=5, warmup=1)
        self.results.append(r)
        print(f"  复杂查询 WHERE+GROUP+ORDER (中表): {r.avg_ms:.1f}ms (median {r.median_ms:.1f})")

        # 5. 中文列名查询
        r = measure("sql_chinese_column", "sql",
                     lambda: execute_advanced_sql_query(self.small_file,
                         "SELECT 技能名称, 伤害 FROM 技能配置 WHERE 元素属性='火'"),
                     runs=5, warmup=1)
        self.results.append(r)
        print(f"  中文列名查询 (小表):           {r.avg_ms:.1f}ms (median {r.median_ms:.1f})")

        # 6. DESCRIBE
        r = measure("describe_table", "sql",
                     lambda: execute_advanced_sql_query(self.small_file, "DESCRIBE 技能配置"),
                     runs=3, warmup=1)
        self.results.append(r)
        print(f"  DESCRIBE (小表):               {r.avg_ms:.1f}ms (median {r.median_ms:.1f})")

    def _run_read_benchmarks(self):
        """数据读取性能测试"""
        print("\n📖 数据读取性能测试")
        print("-" * 40)

        # 小表读取（calamine引擎，多轮测量取平均）
        r = measure("read_50x10", "read",
                     lambda: ExcelOperations.get_range(self.small_file, "技能配置!A1:J50"),
                     runs=5, warmup=2)
        self.results.append(r)
        print(f"  读取 50行×10列 (小表):         {r.avg_ms:.1f}ms")

        # 中表读取（calamine引擎，多轮测量取平均）
        r = measure("read_100x10", "read",
                     lambda: ExcelOperations.get_range(self.medium_file, "技能配置!A1:J100"),
                     runs=5, warmup=2)
        self.results.append(r)
        print(f"  读取 100行×10列 (中表):        {r.avg_ms:.1f}ms")

        # list_sheets
        r = measure("list_sheets", "read",
                     lambda: ExcelOperations.list_sheets(self.small_file),
                     runs=5, warmup=2)
        self.results.append(r)
        print(f"  列出工作表 (小表):             {r.avg_ms:.1f}ms")

        # get_headers
        r = measure("get_headers", "read",
                     lambda: ExcelOperations.get_headers(self.small_file, "技能配置"),
                     runs=5, warmup=2)
        self.results.append(r)
        print(f"  获取表头 (小表):               {r.avg_ms:.1f}ms")

    def _run_search_benchmarks(self):
        """搜索性能测试"""
        print("\n🔍 搜索性能测试")
        print("-" * 40)

        # 精确搜索（calamine引擎，多轮测量）
        r = measure("search_exact", "search",
                     lambda: ExcelOperations.search(self.medium_file, "技能_42", "技能配置"),
                     runs=5, warmup=2)
        self.results.append(r)
        print(f"  精确搜索 (中表):               {r.avg_ms:.1f}ms")

        # 模糊搜索
        r = measure("search_fuzzy", "search",
                     lambda: ExcelOperations.search(self.medium_file, "法师", "技能配置"),
                     runs=1, warmup=0)
        self.results.append(r)
        print(f"  模糊搜索 (中表):               {r.avg_ms:.1f}ms")

        # 正则搜索
        r = measure("search_regex", "search",
                     lambda: ExcelOperations.search(self.medium_file, r"技能_\d{2}$", "技能配置", use_regex=True),
                     runs=1, warmup=0)
        self.results.append(r)
        print(f"  正则搜索 (中表):               {r.avg_ms:.1f}ms")

    def _run_write_benchmarks(self):
        """写入性能测试"""
        print("\n✏️  写入性能测试")
        print("-" * 40)

        write_data = [[f"W_{r}_{c}" for c in range(10)] for r in range(20)]

        # 小批量写入
        r = measure("write_20x10", "write",
                     lambda: ExcelOperations.update_range(self.small_file, "技能配置!A1:J20",
                                                         write_data, preserve_formulas=False),
                     runs=1, warmup=0)
        self.results.append(r)
        print(f"  写入 20行×10列 (小表):         {r.avg_ms:.1f}ms")

        # find_last_row
        r = measure("find_last_row", "write",
                     lambda: ExcelOperations.find_last_row(self.medium_file, "技能配置"),
                     runs=1, warmup=0)
        self.results.append(r)
        print(f"  查找最后行 (中表):             {r.avg_ms:.1f}ms")

    def _run_large_table_benchmarks(self):
        """大表性能测试（仅非快速模式）"""
        print("\n🗄️  大表性能测试 (500行)")
        print("-" * 40)

        # 大表SELECT
        r = measure("sql_large_select", "sql_large",
                     lambda: execute_advanced_sql_query(self.large_file, "SELECT * FROM 技能配置 LIMIT 100"),
                     runs=3, warmup=1)
        self.results.append(r)
        print(f"  SELECT LIMIT 100 (大表):       {r.avg_ms:.1f}ms (median {r.median_ms:.1f})")

        # 大表GROUP BY
        r = measure("sql_large_group_by", "sql_large",
                     lambda: execute_advanced_sql_query(self.large_file,
                         "SELECT 技能类型, AVG(伤害) as avg_dmg, COUNT(*) as cnt "
                         "FROM 技能配置 GROUP BY 技能类型 ORDER BY avg_dmg DESC"),
                     runs=3, warmup=1)
        self.results.append(r)
        print(f"  GROUP BY聚合 (大表):           {r.avg_ms:.1f}ms (median {r.median_ms:.1f})")

        # 大表WHERE
        r = measure("sql_large_where_order", "sql_large",
                     lambda: execute_advanced_sql_query(self.large_file,
                         "SELECT * FROM 技能配置 WHERE 伤害>200 AND 技能等级>=5 "
                         "ORDER BY 伤害 DESC LIMIT 20"),
                     runs=3, warmup=1)
        self.results.append(r)
        print(f"  WHERE+ORDER LIMIT 20 (大表):   {r.avg_ms:.1f}ms (median {r.median_ms:.1f})")

        # 大表搜索
        r = measure("search_large", "search_large",
                     lambda: ExcelOperations.search(self.large_file, "传说", "技能配置"),
                     runs=1, warmup=0)
        self.results.append(r)
        print(f"  搜索 (大表):                   {r.avg_ms:.1f}ms")

        # 大表范围读取
        r = measure("read_large_100x10", "read_large",
                     lambda: ExcelOperations.get_range(self.large_file, "技能配置!A1:J100"),
                     runs=1, warmup=0)
        self.results.append(r)
        print(f"  读取 100行×10列 (大表):        {r.avg_ms:.1f}ms")

        # 缓存加速测试：先触发一次 GROUP BY 加载缓存，再测第二次
        execute_advanced_sql_query(self.large_file,
                              "SELECT 技能类型, AVG(伤害) FROM 技能配置 GROUP BY 技能类型")
        r = measure("sql_large_group_by_cached", "sql_large",
                     lambda: execute_advanced_sql_query(self.large_file,
                         "SELECT 技能类型, AVG(伤害) FROM 技能配置 GROUP BY 技能类型"),
                     runs=2, warmup=0)
        self.results.append(r)
        print(f"  GROUP BY缓存后 (大表):         {r.avg_ms:.1f}ms (median {r.median_ms:.1f})")


# ==================== 报告输出 ====================

def generate_report(results: list[BenchmarkResult]) -> dict:
    """生成结构化报告"""
    report = {
        "timestamp": datetime.now().isoformat(),
        "version": "1.0.0",
        "summary": {
            "total_tests": len(results),
            "passed": sum(1 for r in results if r.success),
            "failed": sum(1 for r in results if not r.success),
        },
        "benchmarks": [r.to_dict() for r in results],
        "categories": {},
    }

    # 按类别汇总
    for r in results:
        cat = r.category
        if cat not in report["categories"]:
            report["categories"][cat] = {"count": 0, "total_avg_ms": 0, "tests": []}
        report["categories"][cat]["count"] += 1
        report["categories"][cat]["total_avg_ms"] += r.avg_ms
        report["categories"][cat]["tests"].append(r.name)

    # 类别平均
    for cat, data in report["categories"].items():
        data["avg_ms"] = round(data["total_avg_ms"] / data["count"], 2)
        del data["total_avg_ms"]

    return report


def compare_with_previous(current: dict, history_dir: Path) -> dict | None:
    """与上次结果对比"""
    # 找到最新的历史文件（排除当前正在写的）
    history_files = sorted(history_dir.glob("benchmark_*.json"), reverse=True)
    if not history_files:
        return None

    try:
        with open(history_files[0]) as f:
            previous = json.load(f)
    except (json.JSONDecodeError, KeyError):
        return None

    comparison = {
        "previous_timestamp": previous.get("timestamp", "unknown"),
        "current_timestamp": current["timestamp"],
        "changes": [],
    }

    prev_map = {b["name"]: b for b in previous.get("benchmarks", [])}
    curr_map = {b["name"]: b for b in current.get("benchmarks", [])}

    for name, curr in curr_map.items():
        if name in prev_map:
            prev = prev_map[name]
            diff_pct = ((curr["avg_ms"] - prev["avg_ms"]) / prev["avg_ms"] * 100
                        if prev["avg_ms"] > 0 else 0)
            change = {
                "name": name,
                "previous_ms": prev["avg_ms"],
                "current_ms": curr["avg_ms"],
                "diff_ms": round(curr["avg_ms"] - prev["avg_ms"], 2),
                "diff_pct": round(diff_pct, 1),
                "status": "OK",
            }
            # 性能退化警告：变慢超过30%
            if diff_pct > 30 and curr["avg_ms"] > 100:  # 排除本身很快的测试
                change["status"] = "⚠️ REGRESSION"
            elif diff_pct < -20:
                change["status"] = "✅ IMPROVED"
            comparison["changes"].append(change)

    # 按退化程度排序
    comparison["changes"].sort(key=lambda x: x["diff_pct"], reverse=True)
    return comparison


def print_comparison(comparison: dict):
    """打印对比表格"""
    print("\n" + "=" * 60)
    print("  📊 性能对比报告")
    print(f"  上次: {comparison['previous_timestamp']}")
    print(f"  本次: {comparison['current_timestamp']}")
    print("=" * 60)
    print(f"  {'测试项':<40} {'上次':>8} {'本次':>8} {'变化':>8}")
    print("-" * 70)

    regressions = []
    improvements = []

    for change in comparison["changes"]:
        status = change["status"]
        if status == "⚠️ REGRESSION":
            regressions.append(change)
        elif status == "✅ IMPROVED":
            improvements.append(change)

        marker = "🔴" if "REGRESSION" in status else ("🟢" if "IMPROVED" in status else "  ")
        print(f"  {marker} {change['name']:<38} {change['previous_ms']:>7.1f}ms "
              f"{change['current_ms']:>7.1f}ms {change['diff_pct']:>+6.1f}%")

    print()
    if regressions:
        print(f"  ⚠️  {len(regressions)} 项性能退化:")
        for r in regressions:
            print(f"      {r['name']}: +{r['diff_pct']:.1f}% ({r['previous_ms']:.1f}→{r['current_ms']:.1f}ms)")
    if improvements:
        print(f"  ✅ {len(improvements)} 项性能提升:")
        for i in improvements:
            print(f"      {i['name']}: {i['diff_pct']:.1f}% ({i['previous_ms']:.1f}→{i['current_ms']:.1f}ms)")
    if not regressions and not improvements:
        print("  ✅ 所有测试项性能稳定")
    print()


# ==================== 主入口 ====================

def main():
    parser = argparse.ArgumentParser(description="ExcelMCP 性能基准测试")
    parser.add_argument("--compare", action="store_true", help="与上次结果对比")
    parser.add_argument("--output", "-o", help="输出JSON文件路径")
    parser.add_argument("--quick", "-q", action="store_true", help="快速模式（跳过大表测试）")
    args = parser.parse_args()

    # 历史目录
    history_dir = PROJECT_ROOT / ".benchmark-history"
    history_dir.mkdir(exist_ok=True)

    # 创建临时测试文件
    with tempfile.TemporaryDirectory() as tmpdir:
        small_file = os.path.join(tmpdir, "small.xlsx")     # 50 rows
        medium_file = os.path.join(tmpdir, "medium.xlsx")   # 100 rows
        large_file = os.path.join(tmpdir, "large.xlsx")     # 500 rows

        print("📦 生成测试数据...")
        t0 = time.perf_counter()
        create_game_config_file(small_file, rows=50, dual_header=True)
        create_game_config_file(medium_file, rows=100, dual_header=True)
        if not args.quick:
            create_game_config_file(large_file, rows=500, dual_header=True)
        print(f"   完成 ({(time.perf_counter() - t0) * 1000:.0f}ms)")

        # 运行基准测试
        bench = Benchmarks(small_file, medium_file, large_file, args.quick)
        results = bench.run_all()

    # 生成报告
    report = generate_report(results)

    # 保存报告
    if args.output:
        output_path = Path(args.output)
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = history_dir / f"benchmark_{ts}.json"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)

    print(f"📁 报告已保存: {output_path}")

    # 对比模式
    if args.compare:
        comparison = compare_with_previous(report, history_dir)
        if comparison:
            print_comparison(comparison)
            # 保存对比结果
            compare_path = history_dir / f"compare_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(compare_path, "w", encoding="utf-8") as f:
                json.dump(comparison, f, indent=2, ensure_ascii=False)
        else:
            print("\n  没有找到历史数据，无法对比。下次运行时可用 --compare 查看变化。")

    # 汇总
    print("=" * 60)
    print(f"  总计: {report['summary']['passed']}/{report['summary']['total_tests']} 项通过")
    if report['summary']['failed'] > 0:
        print(f"  ❌ {report['summary']['failed']} 项失败:")
        for r in results:
            if not r.success:
                print(f"      {r.name}: {r.error}")
    else:
        print("  ✅ 全部通过")
    print("=" * 60)

    return 0 if report['summary']['failed'] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
