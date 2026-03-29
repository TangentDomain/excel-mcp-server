import asyncio
import json
import tempfile
import os
from pathlib import Path
import openpyxl
from typing import Dict, Any

# 创建测试用的Excel文件用于MCP验证
def create_test_excel_files():
    """创建游戏开发测试用的Excel文件"""
    
    # 测试文件1: 技能表.xlsx
    skill_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    skill_file.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "技能表"
    
    # 写入表头
    headers = ["ID", "技能名称", "类型", "消耗MP", "冷却时间", "描述", "稀有度"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 写入技能数据
    skills_data = [
        [1, "火球术", "攻击魔法", 30, 0, "发射火球攻击敌人", "普通"],
        [2, "冰箭", "攻击魔法", 25, 0, "发射冰箭冻结敌人", "普通"],
        [3, "治愈术", "回复魔法", 40, 5, "恢复HP", "稀有"],
        [4, "闪电链", "攻击魔法", 60, 10, "连锁闪电攻击多个敌人", "史诗"],
        [5, "护盾术", "防御魔法", 50, 0, "提供魔法护盾", "稀有"],
        [6, "传送", "特殊魔法", 80, 30, "传送到指定位置", "传说"],
        [7, "复活术", "特殊魔法", 100, 60, "复活死亡队友", "传说"],
        [8, "暴风雪", "攻击魔法", 90, 15, "范围冰系伤害", "史诗"],
    ]
    
    for row_idx, skill in enumerate(skills_data, 2):
        for col_idx, value in enumerate(skill, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(skill_file.name)
    
    # 测试文件2: 角色表.xlsx  
    char_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    char_file.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "角色表"
    
    # 写入表头
    headers = ["ID", "角色名称", "职业", "等级", "HP", "MP", "攻击力", "防御力", "技能ID"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 写入角色数据
    char_data = [
        [1, "艾莉娅", "法师", 25, 2800, 1200, 85, 60, "1,2,3"],
        [2, "加尔", "战士", 30, 3500, 800, 120, 95, "1,4,5"],
        [3, "莉娜", "弓箭手", 28, 2600, 900, 95, 75, "2,6,7"],
        [4, "汤姆", "牧师", 22, 2000, 1500, 70, 65, "3,5,8"],
        [5, "杰克", "刺客", 26, 2200, 1100, 110, 70, "2,7,8"],
    ]
    
    for row_idx, char in enumerate(char_data, 2):
        for col_idx, value in enumerate(char, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(char_file.name)
    
    return skill_file.name, char_file.name

async def test_mcp_core_functions(skill_file: str, char_file: str):
    """测试MCP核心功能"""
    
    # 导入MCP工具
    from excel_mcp_server_fastmcp.server import (
        excel_list_sheets, excel_get_range, excel_get_headers,
        excel_search, excel_find_last_row, excel_describe_table,
        excel_batch_insert_rows, excel_delete_rows,
        excel_update_range, excel_query
    )
    
    print("🧪 开始MCP真实验证...")
    print("=" * 60)
    
    results = {}
    
    # 1. 测试 list_sheets
    try:
        result = excel_list_sheets(skill_file)
        results['list_sheets'] = "✅ 通过" if result.get('success') else "❌ 失败"
        print(f"📋 list_sheets: {results['list_sheets']}")
    except Exception as e:
        results['list_sheets'] = f"❌ 异常: {e}"
        print(f"📋 list_sheets: {results['list_sheets']}")
    
    # 2. 测试 get_headers
    try:
        result = excel_get_headers(skill_file, "技能表")
        results['get_headers'] = "✅ 通过" if result.get('success') else "❌ 失败"
        print(f"📋 get_headers: {results['get_headers']}")
    except Exception as e:
        results['get_headers'] = f"❌ 异常: {e}"
        print(f"📋 get_headers: {results['get_headers']}")
    
    # 3. 测试 get_range
    try:
        result = excel_get_range(skill_file, "技能表!A1:G5")
        results['get_range'] = "✅ 通过" if result.get('success') else "❌ 失败"
        print(f"📋 get_range: {results['get_range']}")
    except Exception as e:
        results['get_range'] = f"❌ 异常: {e}"
        print(f"📋 get_range: {results['get_range']}")
    
    # 4. 测试 find_last_row
    try:
        result = excel_find_last_row(skill_file, "技能表")
        results['find_last_row'] = "✅ 通过" if result.get('success') else "❌ 失败"
        print(f"📋 find_last_row: {results['find_last_row']}")
    except Exception as e:
        results['find_last_row'] = f"❌ 异常: {e}"
        print(f"📋 find_last_row: {results['find_last_row']}")
    
    # 5. 测试 describe_table
    try:
        result = excel_describe_table(skill_file, "技能表")
        results['describe_table'] = "✅ 通过" if result.get('success') else "❌ 失败"
        print(f"📋 describe_table: {results['describe_table']}")
    except Exception as e:
        results['describe_table'] = f"❌ 异常: {e}"
        print(f"📋 describe_table: {results['describe_table']}")
    
    # 6. 测试 search (搜索稀有度为"传说"的技能)
    try:
        result = excel_search(skill_file, "传说", sheet_name="技能表")
        results['search_传说'] = "✅ 通过" if result.get('success') else "❌ 失败"
        print(f"📋 search_传说: {results['search_传说']}")
    except Exception as e:
        results['search_传说'] = f"❌ 异常: {e}"
        print(f"📋 search_传说: {results['search_传说']}")
    
    # 7. 测试 batch_insert_rows
    try:
        # 插入新技能 (需要字典格式)
        new_skill = [{"ID": 9, "技能名称": "治疗之光", "类型": "回复魔法", "消耗MP": 30, "冷却时间": 0, "描述": "群体治疗", "稀有度": "稀有"}]
        result = excel_batch_insert_rows(skill_file, "技能表", new_skill)
        results['batch_insert_rows'] = "✅ 通过" if result.get('success') else "❌ 失败"
        print(f"📋 batch_insert_rows: {results['batch_insert_rows']}")
    except Exception as e:
        results['batch_insert_rows'] = f"❌ 异常: {e}"
        print(f"📋 batch_insert_rows: {results['batch_insert_rows']}")
    
    # 8. 测试 query (查询MP>50的技能)
    try:
        result = excel_query(skill_file, "SELECT * FROM '技能表' WHERE 消耗MP > 50")
        results['query_MP'] = "✅ 通过" if result.get('success') else "❌ 失败"
        print(f"📋 query_MP: {results['query_MP']}")
    except Exception as e:
        results['query_MP'] = f"❌ 异常: {e}"
        print(f"📋 query_MP: {results['query_MP']}")
    
    # 9. 测试 update_range (更新第一个技能的描述)
    try:
        update_data = [["超级火球术"]]
        result = excel_update_range(skill_file, "技能表!G2", update_data)
        results['update_range'] = "✅ 通过" if result.get('success') else "❌ 失败"
        print(f"📋 update_range: {results['update_range']}")
    except Exception as e:
        results['update_range'] = f"❌ 异常: {e}"
        print(f"📋 update_range: {results['update_range']}")
    
    # 10. 测试 delete_rows (删除最后插入的测试技能)
    try:
        result = excel_delete_rows(skill_file, "技能表", 9, 1)
        results['delete_rows'] = "✅ 通过" if result.get('success') else "❌ 失败"
        print(f"📋 delete_rows: {results['delete_rows']}")
    except Exception as e:
        results['delete_rows'] = f"❌ 异常: {e}"
        print(f"📋 delete_rows: {results['delete_rows']}")
    
    print("=" * 60)
    
    # 统计结果
    passed = sum(1 for r in results.values() if r.startswith("✅"))
    total = len(results)
    
    print(f"🎯 MCP验证结果: {passed}/{total} 通过")
    
    if passed == total:
        print("🎉 所有核心功能验证通过！")
    else:
        print(f"⚠️  有 {total - passed} 个功能需要修复")
    
    return results, passed == total

if __name__ == "__main__":
    # 创建测试文件
    skill_file, char_file = create_test_excel_files()
    print(f"📁 创建测试文件:")
    print(f"   技能表: {skill_file}")
    print(f"   角色表: {char_file}")
    
    try:
        # 运行测试
        results, success = asyncio.run(test_mcp_core_functions(skill_file, char_file))
        
        # 输出详细结果
        print("\n📊 详细测试结果:")
        for test_name, result in results.items():
            print(f"   {test_name}: {result}")
        
        # 清理临时文件
        os.unlink(skill_file)
        os.unlink(char_file)
        
    except Exception as e:
        print(f"❌ 测试运行失败: {e}")
        # 清理临时文件
        try:
            os.unlink(skill_file)
            os.unlink(char_file)
        except:
            pass