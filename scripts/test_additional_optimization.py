#!/usr/bin/env python3
"""Token优化额外效果测试"""

import json
import sys
from pathlib import Path

# 添加项目路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root / "src"))

from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

def test_additional_optimizations():
    """测试额外的优化机会"""
    print("🔍 测试额外优化机会...")
    
    test_file = "/tmp/test_additional.xlsx"
    
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Skills"
        
        # 添加包含更多样化数据的测试表
        headers = ['skill_id', 'skill_name', 'skill_type', 'damage', 'cooldown', 'mana_cost', 'description', 'is_active', 'rarity']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        data = [
            [1, '火球术', '攻击', 100, 5.0, 50, '发射火球攻击敌人', True, '普通'],
            [2, '冰霜护盾', '防御', 0, 10.0, 100, '提供冰霜防护', False, '稀有'],
            [3, '雷电术', '攻击', 150, 8.0, 80, '召唤雷电', True, '史诗'],
            [4, '治疗术', '治疗', 0, 0.0, 150, '恢复生命值', True, '普通'],
        ]
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(test_file)
        
        # 测试更多工具的响应
        tools = [
            ('excel_list_sheets', lambda: ExcelOperations.list_sheets(test_file)),
            ('excel_get_headers', lambda: ExcelOperations.get_headers(test_file, "Skills")),
            ('excel_get_range', lambda: ExcelOperations.get_range(test_file, "Skills!A1:C5")),
            ('excel_describe_table', lambda: ExcelOperations.describe_table(test_file, "Skills")),
        ]
        
        print(f"\n📊 各工具响应大小分析:")
        total_size = 0
        for name, func in tools:
            try:
                result = func()
                size = len(json.dumps(result))
                print(f"  {name}: {size} 字符")
                total_size += size
                
                # 分析可以进一步优化的点
                if result.get('success') and result.get('data'):
                    data_obj = result['data']
                    if isinstance(data_obj, dict):
                        # 检查是否有更多的重复字段
                        data_keys = set(data_obj.keys())
                        meta_keys = set(result.get('meta', {}).keys()) if result.get('meta') else set()
                        
                        # 检查是否有不必要的元数据
                        if 'file_path' in meta_keys:
                            print(f"    💡 可以优化: 移除meta中的file_path（通常不必要）")
                        
                        # 检查是否有冗余的元数据
                        redundant_meta = []
                        for key in meta_keys:
                            if key in data_obj and data_obj[key] == result['meta'][key]:
                                redundant_meta.append(key)
                        
                        if redundant_meta:
                            print(f"    💡 可以优化: meta和data中重复的字段: {redundant_meta}")
                            
            except Exception as e:
                print(f"  {name}: 错误 - {e}")
        
        print(f"\n📈 当前总大小: {total_size} 字符")
        
        # 估算如果移除冗余字段后的节省
        estimated_savings = 50  # 估算每个工具可以再节省50字符
        potential_total = total_size - (len(tools) * estimated_savings)
        potential_savings_percent = ((total_size - potential_total) / total_size) * 100
        
        print(f"💡 如果移除冗余字段，预计可以再节省: {len(tools) * estimated_savings} 字符")
        print(f"💡 优化后预计总大小: {potential_total} 字符")
        print(f"💡 预计总节省: {potential_savings_percent:.1f}%")
        
        # 清理
        import os
        os.remove(test_file)
        
        return True
        
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        return False

if __name__ == "__main__":
    test_additional_optimizations()