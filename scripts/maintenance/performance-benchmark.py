#!/usr/bin/env python3
"""
Excel MCP Server - 性能基准测试脚本

测量不同文件大小下的读取/写入性能，生成对比报告。
用于验证性能优化效果（REQ-032）。

用法:
    python3 scripts/performance-benchmark.py [--output-dir ./bench_output]
"""

import argparse
import json
import os
import sys
import tempfile
import time
import traceback
from datetime import datetime

# 确保可以导入项目模块
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

try:
    import psutil
    _HAS_PSUTIL = True
except ImportError:
    _HAS_PSUTIL = False

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    from python_calamine import CalamineWorkbook
    _HAS_CALAMINE = True
except ImportError:
    _HAS_CALAMINE = False

try:
    import pandas as pd
    _HAS_PANDAS = True
except ImportError:
    _HAS_PANDAS = False

# 测试配置
TEST_SIZES = [
    {"rows": 1_000, "cols": 10, "label": "1K行×10列"},
    {"rows": 10_000, "cols": 10, "label": "10K行×10列"},
    {"rows": 50_000, "cols": 10, "label": "50K行×10列"},
    {"rows": 100_000, "cols": 10, "label": "100K行×10列"},
]

# 大文件阈值（字节）
LARGE_FILE_THRESHOLD = 50 * 1024 * 1024  # 50MB


def get_memory_mb():
    """获取当前进程内存占用（MB）"""
    if _HAS_PSUTIL:
        process = psutil.Process(os.getpid())
        return process.memory_info().rss / 1024 / 1024
    return 0


def generate_test_file(file_path, rows, cols):
    """生成测试用Excel文件

    Args:
        file_path: 输出文件路径
        rows: 行数（含表头）
        cols: 列数
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="TestData")

    # 写入表头
    headers = [f"Col_{i}" for i in range(cols)]
    ws.append(headers)

    # 写入数据行
    for r in range(1, rows):
        row_data = [
            r,                          # 整数列
            f"Name_{r}",                # 字符串列
            3.14159 * r,                # 浮点列
            True if r % 2 == 0 else False,  # 布尔列
        ]
        # 其余列填充混合数据
        for c in range(4, cols):
            row_data.append(f"R{r}C{c}")
        ws.append(row_data)

    wb.save(file_path)
    wb.close()


def bench_calamine_read(file_path):
    """测试calamine读取性能"""
    if not _HAS_CALAMINE:
        return None

    start_mem = get_memory_mb()
    start_time = time.perf_counter()

    wb = CalamineWorkbook.from_path(file_path)
    ws = wb.get_sheet_by_name("TestData")
    all_rows = list(ws.to_python())
    row_count = len(all_rows)

    elapsed = time.perf_counter() - start_time
    end_mem = get_memory_mb()

    return {
        "engine": "calamine",
        "rows": row_count,
        "time_s": round(elapsed, 3),
        "memory_mb": round(end_mem - start_mem, 1),
        "throughput_rows_per_s": round(row_count / elapsed, 0) if elapsed > 0 else 0,
    }


def bench_openpyxl_read_only(file_path):
    """测试openpyxl read_only模式读取性能"""
    if not _HAS_OPENPYXL:
        return None

    start_mem = get_memory_mb()
    start_time = time.perf_counter()

    wb = _openpyxl_import().load_workbook(file_path, read_only=True)
    ws = wb["TestData"]
    row_count = 0
    for row in ws.iter_rows(values_only=True):
        row_count += 1
    wb.close()

    elapsed = time.perf_counter() - start_time
    end_mem = get_memory_mb()

    return {
        "engine": "openpyxl_read_only",
        "rows": row_count,
        "time_s": round(elapsed, 3),
        "memory_mb": round(end_mem - start_mem, 1),
        "throughput_rows_per_s": round(row_count / elapsed, 0) if elapsed > 0 else 0,
    }


def bench_openpyxl_range_read(file_path, max_rows=1000):
    """测试openpyxl读取指定范围性能（只读前N行）"""
    if not _HAS_OPENPYXL:
        return None

    start_mem = get_memory_mb()
    start_time = time.perf_counter()

    wb = _openpyxl_import().load_workbook(file_path, read_only=True)
    ws = wb["TestData"]
    row_count = 0
    for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        row_count += 1
    wb.close()

    elapsed = time.perf_counter() - start_time
    end_mem = get_memory_mb()

    return {
        "engine": f"openpyxl_range({max_rows})",
        "rows": row_count,
        "time_s": round(elapsed, 3),
        "memory_mb": round(end_mem - start_mem, 1),
        "throughput_rows_per_s": round(row_count / elapsed, 0) if elapsed > 0 else 0,
    }


def bench_pandas_read(file_path):
    """测试pandas读取性能"""
    if not _HAS_PANDAS or not _HAS_CALAMINE:
        return None

    start_mem = get_memory_mb()
    start_time = time.perf_counter()

    df = pd.read_excel(file_path, sheet_name="TestData", engine="calamine", keep_default_na=False)
    row_count = len(df)

    elapsed = time.perf_counter() - start_time
    end_mem = get_memory_mb()

    return {
        "engine": "pandas_calamine",
        "rows": row_count,
        "time_s": round(elapsed, 3),
        "memory_mb": round(end_mem - start_mem, 1),
        "df_memory_mb": round(df.memory_usage(deep=True).sum() / 1024 / 1024, 2),
        "throughput_rows_per_s": round(row_count / elapsed, 0) if elapsed > 0 else 0,
    }


def bench_pandas_dtype_optimized(file_path):
    """测试pandas读取 + dtype优化后的内存节省"""
    if not _HAS_PANDAS or not _HAS_CALAMINE:
        return None

    start_time = time.perf_counter()

    df = pd.read_excel(file_path, sheet_name="TestData", engine="calamine", keep_default_na=False)
    raw_df_mem = df.memory_usage(deep=True).sum() / 1024 / 1024

    # 模拟 _optimize_dtypes
    for col in df.columns:
        col_type = df[col].dtype
        if col_type == 'object':
            num_unique = df[col].nunique()
            if num_unique > 0 and num_unique / len(df) < 0.3:
                df[col] = df[col].astype('category')
        elif col_type in ['int64', 'int32']:
            col_min = df[col].min()
            col_max = df[col].max()
            if col_min >= 0:
                if col_max < 256:
                    df[col] = df[col].astype('uint8')
                elif col_max < 65536:
                    df[col] = df[col].astype('uint16')
                elif col_max < 4294967296:
                    df[col] = df[col].astype('uint32')
            else:
                if col_min > -128 and col_max < 127:
                    df[col] = df[col].astype('int8')
                elif col_min > -32768 and col_max < 32767:
                    df[col] = df[col].astype('int16')
                elif col_min > -2147483648 and col_max < 2147483647:
                    df[col] = df[col].astype('int32')
        elif col_type == 'float64':
            df[col] = df[col].astype('float32')

    elapsed = time.perf_counter() - start_time
    opt_df_mem = df.memory_usage(deep=True).sum() / 1024 / 1024
    reduction = (1 - opt_df_mem / raw_df_mem) * 100 if raw_df_mem > 0 else 0

    return {
        "engine": "pandas_dtype_optimized",
        "rows": len(df),
        "time_s": round(elapsed, 3),
        "raw_df_memory_mb": round(raw_df_mem, 2),
        "optimized_df_memory_mb": round(opt_df_mem, 2),
        "memory_reduction_pct": round(reduction, 1),
        "throughput_rows_per_s": round(len(df) / elapsed, 0) if elapsed > 0 else 0,
    }


def bench_sql_query(file_path, sql):
    """测试SQL查询性能"""
    try:
        from excel_mcp_server_fastmcp.api.advanced_sql_query import AdvancedSQLQueryEngine

        engine = AdvancedSQLQueryEngine()
        start_time = time.perf_counter()
        result = engine.execute_sql_query(file_path, sql)
        elapsed = time.perf_counter() - start_time

        row_count = len(result.get('data', []))
        engine.clear_cache()

        return {
            "engine": f"sql_query",
            "sql": sql[:60],
            "rows_returned": row_count,
            "time_s": round(elapsed, 3),
            "success": result.get('success', False),
            "throughput_rows_per_s": round(row_count / elapsed, 0) if elapsed > 0 else 0,
        }
    except Exception as e:
        return {"engine": "sql_query", "error": str(e)}


def bench_write_only(file_path, rows, cols):
    """测试openpyxl write_only模式写入性能"""
    if not _HAS_OPENPYXL:
        return None

    start_mem = get_memory_mb()
    start_time = time.perf_counter()

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="TestData")

    headers = [f"Col_{i}" for i in range(cols)]
    ws.append(headers)

    for r in range(1, rows):
        row_data = [r, f"Name_{r}", 3.14159 * r, r % 2 == 0]
        for c in range(4, cols):
            row_data.append(f"R{r}C{c}")
        ws.append(row_data)

    wb.save(file_path)
    wb.close()

    elapsed = time.perf_counter() - start_time
    end_mem = get_memory_mb()

    return {
        "engine": "openpyxl_write_only",
        "rows": rows,
        "time_s": round(elapsed, 3),
        "memory_mb": round(end_mem - start_mem, 1),
        "throughput_rows_per_s": round(rows / elapsed, 0) if elapsed > 0 else 0,
    }


def _openpyxl_import():
    """延迟导入openpyxl"""
    from openpyxl import load_workbook
    return load_workbook


def run_benchmark(output_dir):
    """运行完整的基准测试"""
    results = {
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "environment": {
            "python": sys.version,
            "psutil": _HAS_PSUTIL,
            "calamine": _HAS_CALAMINE,
            "openpyxl": _HAS_OPENPYXL,
            "pandas": _HAS_PANDAS,
        },
        "tests": [],
    }

    os.makedirs(output_dir, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"Excel MCP Server 性能基准测试")
    print(f"{'='*60}")
    print(f"环境: Python {sys.version.split()[0]}")
    print(f"  calamine: {'✓' if _HAS_CALAMINE else '✗'}")
    print(f"  openpyxl: {'✓' if _HAS_OPENPYXL else '✗'}")
    print(f"  pandas:   {'✓' if _HAS_PANDAS else '✗'}")
    print(f"  psutil:   {'✓' if _HAS_PSUTIL else '✗ (内存数据不可用)'}")
    print(f"{'='*60}\n")

    for config in TEST_SIZES:
        rows, cols, label = config["rows"], config["cols"], config["label"]
        test_result = {"label": label, "rows": rows, "cols": cols, "benchmarks": []}

        # 生成测试文件
        test_file = os.path.join(output_dir, f"bench_{rows}x{cols}.xlsx")
        print(f"[生成] {label} ({rows:,}行)...", end=" ", flush=True)
        gen_start = time.perf_counter()
        generate_test_file(test_file, rows, cols)
        gen_time = time.perf_counter() - gen_start
        file_size_mb = os.path.getsize(test_file) / 1024 / 1024
        print(f"完成 ({gen_time:.1f}s, {file_size_mb:.1f}MB)")

        test_result["file_size_mb"] = round(file_size_mb, 2)
        test_result["generate_time_s"] = round(gen_time, 2)

        # 读取测试
        benchmarks = [
            ("calamine全量读取", bench_calamine_read, [test_file]),
            ("openpyxl read_only全量", bench_openpyxl_read_only, [test_file]),
            ("openpyxl 范围读取(1000行)", bench_openpyxl_range_read, [test_file, 1000]),
            ("pandas+calamine读取", bench_pandas_read, [test_file]),
            ("pandas+dtype优化", bench_pandas_dtype_optimized, [test_file]),
        ]

        print(f"  [读取测试]")
        for bench_name, bench_fn, args in benchmarks:
            try:
                result = bench_fn(*args)
                if result:
                    test_result["benchmarks"].append(result)
                    mem_str = f", 内存+{result['memory_mb']}MB" if result['memory_mb'] else ""
                    print(f"    {bench_name:30s} → {result['time_s']:.3f}s "
                          f"({result['throughput_rows_per_s']:.0f} 行/s{mem_str})")
            except Exception as e:
                print(f"    {bench_name:30s} → 失败: {e}")

        # 写入测试
        write_file = os.path.join(output_dir, f"bench_write_{rows}x{cols}.xlsx")
        print(f"  [写入测试]")
        try:
            write_result = bench_write_only(write_file, rows, cols)
            if write_result:
                test_result["benchmarks"].append(write_result)
                mem_str = f", 内存+{write_result['memory_mb']}MB" if write_result['memory_mb'] else ""
                print(f"    {'write_only写入':30s} → {write_result['time_s']:.3f}s "
                      f"({write_result['throughput_rows_per_s']:.0f} 行/s{mem_str})")
        except Exception as e:
            print(f"    {'write_only写入':30s} → 失败: {e}")

        # SQL查询测试
        print(f"  [SQL查询测试]")
        sql_queries = [
            ("SELECT * LIMIT 100", "SELECT * FROM TestData LIMIT 100"),
            ("SELECT WHERE", f"SELECT * FROM TestData WHERE Col_0 > {rows // 2}"),
            ("SELECT GROUP BY", "SELECT Col_3, COUNT(*) as cnt FROM TestData GROUP BY Col_3"),
            ("SELECT AGG", "SELECT AVG(Col_2) as avg_val, MAX(Col_0) as max_id FROM TestData"),
        ]
        for sql_name, sql in sql_queries:
            try:
                sql_result = bench_sql_query(test_file, sql)
                if sql_result and sql_result.get("success"):
                    test_result["benchmarks"].append(sql_result)
                    print(f"    {sql_name:30s} → {sql_result['time_s']:.3f}s "
                          f"({sql_result['rows_returned']}行)")
                elif sql_result and sql_result.get("error"):
                    print(f"    {sql_name:30s} → 失败: {sql_result['error'][:60]}")
            except Exception as e:
                print(f"    {sql_name:30s} → 失败: {e}")

        results["tests"].append(test_result)

        # 清理
        for f in [test_file, write_file]:
            if os.path.exists(f):
                os.unlink(f)

        print()

    # 保存结果
    result_file = os.path.join(output_dir, "benchmark_results.json")
    with open(result_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"结果已保存: {result_file}")

    return results


def main():
    parser = argparse.ArgumentParser(description="Excel MCP Server 性能基准测试")
    parser.add_argument(
        "--output-dir", default="./bench_output",
        help="测试文件输出目录 (默认: ./bench_output)"
    )
    args = parser.parse_args()

    results = run_benchmark(args.output_dir)

    # 打印摘要
    print(f"\n{'='*60}")
    print("摘要")
    print(f"{'='*60}")

    for test in results["tests"]:
        print(f"\n{test['label']} ({test['file_size_mb']}MB):")
        for bm in test["benchmarks"]:
            print(f"  {bm['engine']:35s} {bm['time_s']:>8.3f}s  "
                  f"{bm['throughput_rows_per_s']:>10.0f} 行/s")


if __name__ == "__main__":
    main()
