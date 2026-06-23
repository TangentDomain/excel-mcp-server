#!/usr/bin/env python
"""对抗测试评分脚本 — 多维量化指标体系。

运行所有对抗测试用例，输出 JSON 报告到 data/adversarial-score.jsonl。

量化维度：
1. 准确率（accuracy）：pass/total，按操作类型 + 数据类型分类
2. 功能覆盖率（tool_coverage）：26 个工具中多少个被测试覆盖
3. SQL 特性覆盖率（sql_coverage）：各 SQL 特性的覆盖情况
4. 边界值覆盖（edge_coverage）：各边界值类型的覆盖情况
5. 写操作安全性（write_safety）：affected_rows / 文件完整性 / 无匹配
6. 性能指标（performance）：各操作类型的平均执行时间

Usage:
    python scripts/adversarial-score.py
    python scripts/adversarial-score.py --json       # 只输出 JSON
    python scripts/adversarial-score.py --verbose    # 详细输出每个 case
"""

from __future__ import annotations

import json
import math
import os
import random
import sqlite3
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# 确保 repo root 在 path 中
REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "src"))

from openpyxl import Workbook

import excel_mcp_server_fastmcp.api.advanced_sql_query as _query_module
from excel_mcp_server_fastmcp.api.advanced_sql_query import (
    execute_advanced_delete_query,
    execute_advanced_insert_query,
    execute_advanced_sql_query,
    execute_advanced_update_query,
)

# ============================================================
# 种子数据
# ============================================================

SEED_TABLE = "商品"
SEED_COLUMNS = ["ID", "Name", "Price", "Stock", "Active"]
SEED_DATA = [
    [1, "铁剑", 100.0, 50, "是"],
    [2, "银剑", 200.0, 30, "是"],
    [3, "木盾", 50.0, 100, "否"],
    [4, "铁甲", 180.0, 20, "是"],
    [5, "皮甲", 80.0, 60, "是"],
    [6, "神器", 999.99, 1, "是"],
]


# ============================================================
# 工具函数
# ============================================================


def _make_seed_excel(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SEED_TABLE
    ws.append(SEED_COLUMNS)
    for row in SEED_DATA:
        ws.append(row)
    wb.save(path)


def _make_oracle() -> sqlite3.Connection:
    conn = sqlite3.connect(":memory:")
    col_defs = ", ".join(
        f"[{c}] {'REAL' if c == 'Price' else 'TEXT' if c in ('Name', 'Active') else 'INTEGER'}"
        for c in SEED_COLUMNS
    )
    conn.execute(f"CREATE TABLE [{SEED_TABLE}] ({col_defs})")
    ph = ", ".join(["?"] * len(SEED_COLUMNS))
    for row in SEED_DATA:
        conn.execute(f"INSERT INTO [{SEED_TABLE}] VALUES ({ph})", row)
    conn.commit()
    return conn


def _reset_engine() -> None:
    _query_module._shared_engine = None


def _values_match(a: Any, b: Any, tol: float = 0.01) -> bool:
    # Excel 存储层无法区分空字符串 '' 和 None（openpyxl 往返后 '' → None）
    # 差分测试中将两者视为等价（这是 Excel 的固有限制，非 SQL 引擎 bug）
    if a == "" and b is None:
        return True
    if a is None and b == "":
        return True
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        if isinstance(a, float) or isinstance(b, float):
            return abs(float(a) - float(b)) < tol
        return a == b
    return a == b


def _rows_match(excel_row: list, sqlite_row: tuple, tol: float = 0.01) -> bool:
    if len(excel_row) != len(sqlite_row):
        return False
    return all(_values_match(a, b, tol) for a, b in zip(excel_row, sqlite_row))


def _query_excel(path: str, sql: str) -> dict:
    return execute_advanced_sql_query(path, sql)


def _query_sqlite(conn: sqlite3.Connection, sql: str) -> list[tuple]:
    return conn.execute(sql).fetchall()


def _write_excel(path: str, sql: str, op: str) -> dict:
    if op == "update":
        return execute_advanced_update_query(path, sql)
    elif op == "insert":
        return execute_advanced_insert_query(path, sql)
    elif op == "delete":
        return execute_advanced_delete_query(path, sql)


def _write_sqlite(conn: sqlite3.Connection, sql: str) -> int:
    cur = conn.execute(sql)
    n = cur.rowcount
    conn.commit()
    return n


def _check_query(excel_res: dict, sqlite_rows: list[tuple], sql: str, tol: float = 0.01) -> tuple[bool, str]:
    if not excel_res["success"]:
        return False, f"ExcelMCP failed: {excel_res.get('message', 'N/A')}"
    data = excel_res["data"]
    if len(data) <= 1:
        if len(sqlite_rows) == 0:
            return True, ""
        return False, f"ExcelMCP has 0 data rows, SQLite has {len(sqlite_rows)}"
    excel_rows = data[1:]
    if len(excel_rows) != len(sqlite_rows):
        return False, f"Row count: ExcelMCP={len(excel_rows)} SQLite={len(sqlite_rows)}"
    for i, (er, sr) in enumerate(zip(excel_rows, sqlite_rows)):
        if not _rows_match(er, sr, tol):
            return False, f"Row {i}: ExcelMCP={er} SQLite={sr}"
    return True, ""


def _check_affected(excel_res: dict, sqlite_n: int, sql: str) -> tuple[bool, str]:
    if not excel_res["success"]:
        return False, f"ExcelMCP failed: {excel_res.get('message', 'N/A')}"
    ea = excel_res.get("affected_rows", -1)
    if ea != sqlite_n:
        return False, f"affected_rows: ExcelMCP={ea} SQLite={sqlite_n}"
    return True, ""


# ============================================================
# 对抗用例定义
# ============================================================


def _build_test_cases() -> list[dict]:
    """构建所有对抗测试用例，每个 case 有 category/sub_category/sql/op_type/tol/tags"""

    cases = []

    # ---- UPDATE readback ----
    cases.append({
        "category": "UPDATE", "sub_category": "single_col",
        "sql": "UPDATE 商品 SET Price=999 WHERE ID=1",
        "op_type": "update", "tol": 0.01,
        "tags": ["write_safety", "affected_rows", "readback"],
    })
    cases.append({
        "category": "UPDATE", "sub_category": "multi_col",
        "sql": "UPDATE 商品 SET Price=888, Stock=99 WHERE ID=2",
        "op_type": "update", "tol": 0.01,
        "tags": ["write_safety", "affected_rows", "readback"],
    })
    cases.append({
        "category": "UPDATE", "sub_category": "expression",
        "sql": "UPDATE 商品 SET Price=Price*1.1 WHERE Active='是'",
        "op_type": "update", "tol": 0.1,
        "tags": ["write_safety", "affected_rows", "readback", "sql_expression"],
    })
    cases.append({
        "category": "UPDATE", "sub_category": "no_match",
        "sql": "UPDATE 商品 SET Price=999 WHERE ID=99999",
        "op_type": "update", "tol": 0.01,
        "tags": ["write_safety", "affected_rows", "no_match_safety"],
    })
    cases.append({
        "category": "UPDATE", "sub_category": "all_rows",
        "sql": "UPDATE 商品 SET Stock=0",
        "op_type": "update", "tol": 0.01,
        "tags": ["write_safety", "affected_rows", "readback"],
    })

    # ---- INSERT ----
    cases.append({
        "category": "INSERT", "sub_category": "basic",
        "sql": "INSERT INTO 商品 (ID, Name, Price, Stock, Active) VALUES (7, '圣剑', 500, 10, '是')",
        "op_type": "insert", "tol": 0.01,
        "tags": ["write_safety", "affected_rows", "row_count"],
    })
    cases.append({
        "category": "INSERT", "sub_category": "float_val",
        "sql": "INSERT INTO 商品 (ID, Name, Price, Stock, Active) VALUES (8, '法杖', 333.33, 5, '否')",
        "op_type": "insert", "tol": 0.01,
        "tags": ["write_safety", "affected_rows", "edge_float", "row_count"],
    })

    # ---- DELETE ----
    cases.append({
        "category": "DELETE", "sub_category": "basic",
        "sql": "DELETE FROM 商品 WHERE ID=3",
        "op_type": "delete", "tol": 0.01,
        "tags": ["write_safety", "affected_rows", "row_count"],
    })
    cases.append({
        "category": "DELETE", "sub_category": "no_match",
        "sql": "DELETE FROM 商品 WHERE ID=99999",
        "op_type": "delete", "tol": 0.01,
        "tags": ["write_safety", "affected_rows", "no_match_safety"],
    })

    # ---- WHERE conditions ----
    cases.append({
        "category": "WHERE", "sub_category": "equal",
        "sql": "SELECT * FROM 商品 WHERE ID=2",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_where_equal", "sql_select"],
    })
    cases.append({
        "category": "WHERE", "sub_category": "greater",
        "sql": "SELECT * FROM 商品 WHERE Price > 150",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_where_compare", "sql_select"],
    })
    cases.append({
        "category": "WHERE", "sub_category": "less",
        "sql": "SELECT * FROM 商品 WHERE Stock < 50",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_where_compare", "sql_select"],
    })
    cases.append({
        "category": "WHERE", "sub_category": "in",
        "sql": "SELECT * FROM 商品 WHERE ID IN (1, 3, 5)",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_where_in", "sql_select"],
    })
    cases.append({
        "category": "WHERE", "sub_category": "like",
        "sql": "SELECT * FROM 商品 WHERE Name LIKE '%剑%'",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_where_like", "sql_select"],
    })
    cases.append({
        "category": "WHERE", "sub_category": "and_or",
        "sql": "SELECT * FROM 商品 WHERE (Price > 100 AND Active='是') OR ID=3",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_where_and_or", "sql_select"],
    })
    cases.append({
        "category": "WHERE", "sub_category": "between",
        "sql": "SELECT * FROM 商品 WHERE Price BETWEEN 50 AND 200",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_where_between", "sql_select"],
    })

    # ---- SQL features ----
    cases.append({
        "category": "SQL", "sub_category": "order_by",
        "sql": "SELECT * FROM 商品 ORDER BY Price DESC",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_order_by", "sql_select"],
    })
    cases.append({
        "category": "SQL", "sub_category": "group_by",
        "sql": "SELECT Active, COUNT(*) as cnt, AVG(Price) as avg_price FROM 商品 GROUP BY Active",
        "op_type": "select", "tol": 0.1,
        "tags": ["sql_group_by", "sql_aggregate", "sql_select"],
    })
    cases.append({
        "category": "SQL", "sub_category": "having",
        "sql": "SELECT Active, COUNT(*) as cnt FROM 商品 GROUP BY Active HAVING cnt > 2",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_group_by", "sql_having", "sql_aggregate", "sql_select"],
    })
    cases.append({
        "category": "SQL", "sub_category": "distinct",
        "sql": "SELECT DISTINCT Active FROM 商品",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_distinct", "sql_select"],
    })
    cases.append({
        "category": "SQL", "sub_category": "limit",
        "sql": "SELECT * FROM 商品 LIMIT 3",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_limit", "sql_select"],
    })
    cases.append({
        "category": "SQL", "sub_category": "count_star",
        "sql": "SELECT COUNT(*) FROM 商品",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_aggregate", "sql_count", "sql_select"],
    })
    cases.append({
        "category": "SQL", "sub_category": "sum_avg",
        "sql": "SELECT SUM(Price) as total, AVG(Stock) as avg_stock FROM 商品",
        "op_type": "select", "tol": 0.1,
        "tags": ["sql_aggregate", "sql_sum", "sql_avg", "sql_select"],
    })
    cases.append({
        "category": "SQL", "sub_category": "case_when",
        "sql": "SELECT Name, CASE WHEN Price > 100 THEN '贵' ELSE '便宜' END as tier FROM 商品",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_case_when", "sql_select"],
    })
    cases.append({
        "category": "SQL", "sub_category": "aliased_columns",
        "sql": "SELECT ID as 编号, Name as 名称, Price*2 as 双倍价 FROM 商品 WHERE ID < 4",
        "op_type": "select", "tol": 0.01,
        "tags": ["sql_alias", "sql_expression", "sql_select"],
    })
    cases.append({
        "category": "SQL", "sub_category": "subquery_where",
        "sql": "SELECT * FROM 商品 WHERE Price > (SELECT AVG(Price) FROM 商品)",
        "op_type": "select", "tol": 0.1,
        "tags": ["sql_subquery", "sql_aggregate", "sql_select"],
    })

    # ---- Edge values ----
    cases.append({
        "category": "EDGE", "sub_category": "float_precision",
        "sql": "UPDATE 商品 SET Price=999.99*1.1 WHERE ID=6",
        "op_type": "update", "tol": 0.1,
        "tags": ["edge_float", "write_safety", "affected_rows", "readback"],
    })
    cases.append({
        "category": "EDGE", "sub_category": "negative",
        "sql": "UPDATE 商品 SET Price=-100 WHERE ID=1",
        "op_type": "update", "tol": 0.01,
        "tags": ["edge_negative", "write_safety", "affected_rows", "readback"],
    })
    cases.append({
        "category": "EDGE", "sub_category": "zero",
        "sql": "UPDATE 商品 SET Price=0 WHERE ID=1",
        "op_type": "update", "tol": 0.01,
        "tags": ["edge_zero", "write_safety", "affected_rows", "readback"],
    })
    cases.append({
        "category": "EDGE", "sub_category": "empty_string",
        "sql": "UPDATE 商品 SET Name='' WHERE ID=1",
        "op_type": "update", "tol": 0.01,
        "tags": ["edge_empty_string", "write_safety", "affected_rows", "readback"],
    })
    cases.append({
        "category": "EDGE", "sub_category": "special_char",
        "sql": "UPDATE 商品 SET Name='O''Brien' WHERE ID=1",
        "op_type": "update", "tol": 0.01,
        "tags": ["edge_special_char", "write_safety", "affected_rows", "readback"],
    })
    cases.append({
        "category": "EDGE", "sub_category": "chinese",
        "sql": "UPDATE 商品 SET Name='神剑·破军' WHERE ID=1",
        "op_type": "update", "tol": 0.01,
        "tags": ["edge_chinese", "write_safety", "affected_rows", "readback"],
    })
    cases.append({
        "category": "EDGE", "sub_category": "null_write",
        "sql": "UPDATE 商品 SET Price=NULL WHERE ID=1",
        "op_type": "update", "tol": 0.01,
        "tags": ["edge_null", "write_safety", "affected_rows", "readback"],
    })
    cases.append({
        "category": "EDGE", "sub_category": "large_number",
        "sql": "UPDATE 商品 SET Price=9999999999 WHERE ID=1",
        "op_type": "update", "tol": 0.01,
        "tags": ["edge_large_number", "write_safety", "affected_rows", "readback"],
    })
    cases.append({
        "category": "EDGE", "sub_category": "small_number",
        "sql": "UPDATE 商品 SET Price=0.001 WHERE ID=1",
        "op_type": "update", "tol": 0.001,
        "tags": ["edge_small_number", "write_safety", "affected_rows", "readback"],
    })

    # ---- Random fuzz ----
    cases.append({
        "category": "FUZZ", "sub_category": "random_20",
        "sql": "__RANDOM_20__",
        "op_type": "random",
        "tol": 0.01,
        "tags": ["fuzz"],
    })

    return cases


# ============================================================
# 随机 fuzz 执行器
# ============================================================


def _run_random_fuzz(excel_path: str, conn: sqlite3.Connection, seed: int = 42) -> list[dict]:
    """执行 20 步随机写操作，返回每步结果"""
    random.seed(seed)

    update_ops = [
        ("UPDATE 商品 SET Price={v} WHERE ID={rid}", "update", lambda: {
            "v": random.choice([0, -50, 99.99, 1000, 0.01, 500]),
            "rid": random.randint(1, 6),
        }),
        ("UPDATE 商品 SET Stock={v} WHERE ID={rid}", "update", lambda: {
            "v": random.randint(0, 200),
            "rid": random.randint(1, 6),
        }),
        ("UPDATE 商品 SET Name='{v}' WHERE ID={rid}", "update", lambda: {
            "v": random.choice(["测试", "Test", "O''Brien", "A" * 50]),
            "rid": random.randint(1, 6),
        }),
        ("UPDATE 商品 SET Price=Price*1.1 WHERE Active='是'", "update", lambda: {}),
        ("UPDATE 商品 SET Stock=Stock+10", "update", lambda: {}),
    ]

    insert_ops = [
        ("INSERT INTO 商品 (ID, Name, Price, Stock, Active) VALUES ({rid}, '{name}', {price}, {stock}, '{active}')", "insert", lambda: {
            "rid": random.randint(100, 999),
            "name": random.choice(["随机物品", "RandItem", "空名"]),
            "price": round(random.uniform(1, 1000), 2),
            "stock": random.randint(0, 100),
            "active": random.choice(["是", "否"]),
        }),
    ]

    delete_ops = [
        ("DELETE FROM 商品 WHERE ID={rid}", "delete", lambda: {
            "rid": random.randint(100, 999),
        }),
    ]

    all_ops = update_ops + insert_ops + delete_ops
    results = []

    for i in range(20):
        template, op_type, gen = random.choice(all_ops)
        params = gen()
        sql = template.format(**params)

        t0 = time.perf_counter()
        er = _write_excel(excel_path, sql, op_type)
        t_excel = time.perf_counter() - t0

        t0 = time.perf_counter()
        sr = _write_sqlite(conn, sql)
        t_sqlite = time.perf_counter() - t0

        # affected_rows check
        ok_af, msg_af = _check_affected(er, sr, sql)

        # readback check
        ok_q, msg_q = True, ""
        if ok_af:
            sel = "SELECT * FROM 商品"
            t0 = time.perf_counter()
            qr = _query_excel(excel_path, sel)
            t_excel += time.perf_counter() - t0
            t0 = time.perf_counter()
            qs = _query_sqlite(conn, sel)
            t_sqlite += time.perf_counter() - t0
            ok_q, msg_q = _check_query(qr, qs, sel)

        passed = ok_af and ok_q
        results.append({
            "step": i,
            "sql": sql,
            "op_type": op_type,
            "passed": passed,
            "expected_affected": sr,
            "actual_affected": er.get("affected_rows", -1) if er.get("success") else -1,
            "error": "" if passed else f"{msg_af}; {msg_q}".strip("; "),
            "time_excel_ms": round(t_excel * 1000, 1),
            "time_sqlite_ms": round(t_sqlite * 1000, 1),
        })

    return results


# ============================================================
# 单个 case 执行器
# ============================================================


def _run_single_case(excel_path: str, conn: sqlite3.Connection, case: dict) -> dict:
    """执行单个对抗用例，返回结果"""
    sql = case["sql"]
    op = case["op_type"]
    tol = case["tol"]
    category = case["category"]
    sub = case["sub_category"]

    result = {
        "category": category,
        "sub_category": sub,
        "sql": sql,
        "op_type": op,
        "tags": case.get("tags", []),
        "passed": True,
        "expected": None,
        "actual": None,
        "error": "",
        "time_excel_ms": 0,
        "time_sqlite_ms": 0,
    }

    _reset_engine()

    if op == "random":
        fuzz_results = _run_random_fuzz(excel_path, conn)
        total_t_excel = sum(r["time_excel_ms"] for r in fuzz_results)
        total_t_sqlite = sum(r["time_sqlite_ms"] for r in fuzz_results)
        fuzz_passed = sum(1 for r in fuzz_results if r["passed"])
        result.update({
            "passed": fuzz_passed == 20,
            "expected": 20,
            "actual": fuzz_passed,
            "error": f"{20 - fuzz_passed}/20 steps failed" if fuzz_passed < 20 else "",
            "time_excel_ms": total_t_excel,
            "time_sqlite_ms": total_t_sqlite,
            "fuzz_details": fuzz_results,
        })
        return result

    if op == "select":
        t0 = time.perf_counter()
        er = _query_excel(excel_path, sql)
        t_excel = time.perf_counter() - t0
        t0 = time.perf_counter()
        sr = _query_sqlite(conn, sql)
        t_sqlite = time.perf_counter() - t0

        ok, msg = _check_query(er, sr, sql, tol)
        result.update({
            "passed": ok,
            "expected": len(sr),
            "actual": len(er.get("data", [])) - 1 if er.get("success") and len(er.get("data", [])) > 0 else 0,
            "error": msg,
            "time_excel_ms": round(t_excel * 1000, 1),
            "time_sqlite_ms": round(t_sqlite * 1000, 1),
        })
        return result

    # Write operation (update/insert/delete)
    t0 = time.perf_counter()
    er = _write_excel(excel_path, sql, op)
    t_excel = time.perf_counter() - t0
    t0 = time.perf_counter()
    sr = _write_sqlite(conn, sql)
    t_sqlite = time.perf_counter() - t0

    ok_af, msg_af = _check_affected(er, sr, sql)

    ok_q, msg_q = True, ""
    if ok_af:
        sel = "SELECT * FROM 商品"
        t0 = time.perf_counter()
        qr = _query_excel(excel_path, sel)
        t_excel += time.perf_counter() - t0
        t0 = time.perf_counter()
        qs = _query_sqlite(conn, sel)
        t_sqlite += time.perf_counter() - t0
        ok_q, msg_q = _check_query(qr, qs, sel)

    passed = ok_af and ok_q
    result.update({
        "passed": passed,
        "expected": sr,
        "actual": er.get("affected_rows", -1) if er.get("success") else -1,
        "error": "" if passed else f"{msg_af}; {msg_q}".strip("; "),
        "time_excel_ms": round(t_excel * 1000, 1),
        "time_sqlite_ms": round(t_sqlite * 1000, 1),
    })
    return result


# ============================================================
# 多维量化指标计算
# ============================================================


def _compute_accuracy(results: list[dict]) -> dict:
    """维度 1: 准确率"""
    total = len(results)
    passed = sum(1 for r in results if r["passed"])

    by_category: dict[str, dict] = {}
    for r in results:
        cat = r["category"]
        if cat not in by_category:
            by_category[cat] = {"total": 0, "passed": 0}
        by_category[cat]["total"] += 1
        if r["passed"]:
            by_category[cat]["passed"] += 1

    return {
        "total": total,
        "passed": passed,
        "accuracy_pct": round(passed / total * 100, 1) if total > 0 else 0,
        "by_category": {
            k: {
                "total": v["total"],
                "passed": v["passed"],
                "accuracy_pct": round(v["passed"] / v["total"] * 100, 1) if v["total"] > 0 else 0,
            }
            for k, v in sorted(by_category.items())
        },
    }


def _compute_tool_coverage() -> dict:
    """维度 2: 功能覆盖率（26 个工具）"""
    all_tools = [
        "excel_query", "excel_update_query", "excel_insert_query", "excel_delete_query",
        "excel_get_range", "excel_update_range", "excel_upsert_row",
        "excel_describe_table", "excel_get_headers", "excel_search",
        "excel_search_directory", "excel_create_file", "excel_list_sheets",
        "excel_create_sheet", "excel_delete_sheet", "excel_rename_sheet",
        "excel_copy_sheet", "excel_structure", "excel_set_layout",
        "excel_format_cells", "excel_set_formula", "excel_find_last_row",
        "excel_compare_sheets", "excel_backup", "excel_run_python",
        "excel_get_cell",
    ]
    # 对抗测试当前覆盖的工具（通过 SQL API 间接覆盖）
    covered_tools = {
        "excel_query": True,         # SELECT via execute_advanced_sql_query
        "excel_update_query": True,  # UPDATE via execute_advanced_update_query
        "excel_insert_query": True,  # INSERT via execute_advanced_insert_query
        "excel_delete_query": True,  # DELETE via execute_advanced_delete_query
    }
    covered_count = sum(1 for t in all_tools if covered_tools.get(t))
    return {
        "total_tools": len(all_tools),
        "covered_tools": covered_count,
        "coverage_pct": round(covered_count / len(all_tools) * 100, 1),
        "covered_list": [t for t in all_tools if covered_tools.get(t)],
        "uncovered_list": [t for t in all_tools if not covered_tools.get(t)],
    }


def _compute_sql_coverage(results: list[dict]) -> dict:
    """维度 3: SQL 特性覆盖率"""
    features = {
        "sql_select": "SELECT 查询",
        "sql_where_equal": "WHERE = 等值条件",
        "sql_where_compare": "WHERE > < 比较条件",
        "sql_where_in": "WHERE IN",
        "sql_where_like": "WHERE LIKE",
        "sql_where_and_or": "WHERE AND/OR",
        "sql_where_between": "WHERE BETWEEN",
        "sql_order_by": "ORDER BY",
        "sql_group_by": "GROUP BY",
        "sql_having": "HAVING",
        "sql_aggregate": "聚合函数 (COUNT/SUM/AVG/MAX/MIN)",
        "sql_distinct": "DISTINCT",
        "sql_limit": "LIMIT",
        "sql_case_when": "CASE WHEN",
        "sql_alias": "列别名 (AS)",
        "sql_expression": "表达式 (数学运算)",
        "sql_subquery": "子查询",
        "sql_count": "COUNT(*)",
        "sql_sum": "SUM()",
        "sql_avg": "AVG()",
        "sql_join": "JOIN",
        "sql_union": "UNION / UNION ALL",
        "sql_cte": "CTE (WITH)",
        "sql_window": "窗口函数 (ROW_NUMBER/RANK)",
        "sql_offset": "OFFSET",
    }

    all_tags = set()
    for r in results:
        for tag in r.get("tags", []):
            all_tags.add(tag)

    covered = {}
    for feat_id, feat_name in features.items():
        covered[feat_id] = {
            "name": feat_name,
            "covered": feat_id in all_tags,
        }

    total = len(features)
    covered_count = sum(1 for v in covered.values() if v["covered"])
    return {
        "total_features": total,
        "covered_features": covered_count,
        "coverage_pct": round(covered_count / total * 100, 1),
        "features": covered,
    }


def _compute_edge_coverage(results: list[dict]) -> dict:
    """维度 4: 边界值覆盖"""
    edge_types = {
        "edge_float": "浮点数",
        "edge_negative": "负数",
        "edge_zero": "零",
        "edge_empty_string": "空字符串",
        "edge_special_char": "特殊字符 (单引号等)",
        "edge_chinese": "中文/Unicode",
        "edge_null": "NULL",
        "edge_large_number": "大数",
        "edge_small_number": "极小数",
    }

    all_tags = set()
    for r in results:
        for tag in r.get("tags", []):
            all_tags.add(tag)

    covered = {}
    for edge_id, edge_name in edge_types.items():
        covered[edge_id] = {
            "name": edge_name,
            "covered": edge_id in all_tags,
        }

    total = len(edge_types)
    covered_count = sum(1 for v in covered.values() if v["covered"])
    return {
        "total_types": total,
        "covered_types": covered_count,
        "coverage_pct": round(covered_count / total * 100, 1),
        "types": covered,
    }


def _compute_write_safety(results: list[dict]) -> dict:
    """维度 5: 写操作安全性"""
    write_results = [r for r in results if r["op_type"] in ("update", "insert", "delete")]

    affected_tests = [r for r in write_results if "affected_rows" in r.get("tags", [])]
    readback_tests = [r for r in write_results if "readback" in r.get("tags", [])]
    no_match_tests = [r for r in write_results if "no_match_safety" in r.get("tags", [])]
    row_count_tests = [r for r in write_results if "row_count" in r.get("tags", [])]

    def _pass_count(tests):
        if not tests:
            return {"total": 0, "passed": 0, "pct": 0}
        p = sum(1 for t in tests if t["passed"])
        return {"total": len(tests), "passed": p, "pct": round(p / len(tests) * 100, 1)}

    return {
        "affected_rows_accuracy": _pass_count(affected_tests),
        "readback_consistency": _pass_count(readback_tests),
        "no_match_safety": _pass_count(no_match_tests),
        "row_count_consistency": _pass_count(row_count_tests),
        "total_write_tests": len(write_results),
        "total_write_passed": sum(1 for r in write_results if r["passed"]),
    }


def _compute_performance(results: list[dict]) -> dict:
    """维度 6: 性能指标"""
    by_op: dict[str, list[float]] = {}
    for r in results:
        op = r["op_type"]
        if op == "random":
            continue
        if op not in by_op:
            by_op[op] = []
        by_op[op].append(r.get("time_excel_ms", 0))

    perf = {}
    for op, times in sorted(by_op.items()):
        if times:
            perf[op] = {
                "count": len(times),
                "avg_ms": round(sum(times) / len(times), 1),
                "max_ms": round(max(times), 1),
                "min_ms": round(min(times), 1),
                "total_ms": round(sum(times), 1),
            }

    # Also include random fuzz aggregate
    fuzz = [r for r in results if r["op_type"] == "random"]
    if fuzz:
        perf["random_fuzz"] = {
            "count": 20,
            "avg_ms": round(fuzz[0].get("time_excel_ms", 0) / 20, 1),
            "total_ms": round(fuzz[0].get("time_excel_ms", 0), 1),
        }

    return perf


# ============================================================
# 报告生成
# ============================================================


def generate_report(results: list[dict]) -> dict:
    """生成多维量化报告"""
    return {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "accuracy": _compute_accuracy(results),
        "tool_coverage": _compute_tool_coverage(),
        "sql_coverage": _compute_sql_coverage(results),
        "edge_coverage": _compute_edge_coverage(results),
        "write_safety": _compute_write_safety(results),
        "performance": _compute_performance(results),
        "failed_cases": [
            {
                "category": r["category"],
                "sub_category": r["sub_category"],
                "sql": r["sql"],
                "expected": r["expected"],
                "actual": r["actual"],
                "error": r["error"],
            }
            for r in results if not r["passed"]
        ],
    }


def append_score_jsonl(report: dict, path: str) -> None:
    """追加一行 JSON 到 jsonl 文件"""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(report, ensure_ascii=False) + "\n")


# ============================================================
# 主函数
# ============================================================


def main() -> int:
    import argparse
    parser = argparse.ArgumentParser(description="对抗测试评分")
    parser.add_argument("--json", action="store_true", help="只输出 JSON（无文本）")
    parser.add_argument("--verbose", action="store_true", help="详细输出每个 case")
    parser.add_argument("--output", default=str(REPO_ROOT / "data" / "adversarial-score.jsonl"), help="JSONL 输出路径")
    args = parser.parse_args()

    # 临时文件
    import tempfile
    tmp_dir = tempfile.mkdtemp(prefix="adv_test_")
    excel_path = os.path.join(tmp_dir, "seed.xlsx")

    try:
        _make_seed_excel(excel_path)
        cases = _build_test_cases()
        results = []

        for i, case in enumerate(cases):
            # 每个独立的 write case 需要新建文件和 oracle
            if case["op_type"] != "select":
                # 对写操作，每个 case 用独立的文件和 oracle
                p = os.path.join(tmp_dir, f"case_{i}.xlsx")
                _make_seed_excel(p)
                conn = _make_oracle()
                _reset_engine()
                r = _run_single_case(p, conn, case)
                conn.close()
            else:
                # SELECT 用同一个种子（只读，不会修改）
                conn = _make_oracle()
                _reset_engine()
                r = _run_single_case(excel_path, conn, case)
                conn.close()

            results.append(r)

            if args.verbose:
                status = "PASS" if r["passed"] else "FAIL"
                print(f"  [{status}] {r['category']}/{r['sub_category']}: {r['sql'][:60]}")
                if not r["passed"]:
                    print(f"         expected={r['expected']} actual={r['actual']}")
                    print(f"         error: {r['error']}")

        report = generate_report(results)
        append_score_jsonl(report, args.output)

        if not args.json:
            acc = report["accuracy"]
            tc = report["tool_coverage"]
            sc = report["sql_coverage"]
            ec = report["edge_coverage"]
            ws = report["write_safety"]
            perf = report["performance"]

            print("=" * 60)
            print("对抗测试评分报告")
            print("=" * 60)
            print()
            print(f"  准确率:     {acc['passed']}/{acc['total']} = {acc['accuracy_pct']}%")
            print(f"  工具覆盖:   {tc['covered_tools']}/{tc['total_tools']} = {tc['coverage_pct']}%")
            print(f"  SQL 特性:   {sc['covered_features']}/{sc['total_features']} = {sc['coverage_pct']}%")
            print(f"  边界值:     {ec['covered_types']}/{ec['total_types']} = {ec['coverage_pct']}%")
            print()
            print("  写操作安全性:")
            ws_af = ws["affected_rows_accuracy"]
            ws_rb = ws["readback_consistency"]
            ws_nm = ws["no_match_safety"]
            ws_rc = ws["row_count_consistency"]
            print(f"    affected_rows:   {ws_af['passed']}/{ws_af['total']} = {ws_af['pct']}%")
            print(f"    readback:        {ws_rb['passed']}/{ws_rb['total']} = {ws_rb['pct']}%")
            print(f"    no_match_safety:  {ws_nm['passed']}/{ws_nm['total']} = {ws_nm['pct']}%")
            print(f"    row_count:       {ws_rc['passed']}/{ws_rc['total']} = {ws_rc['pct']}%")
            print()
            print("  按操作类型准确率:")
            for cat, info in acc["by_category"].items():
                print(f"    {cat:10s}: {info['passed']}/{info['total']} = {info['accuracy_pct']}%")
            print()
            print("  性能 (avg ms):")
            for op, info in perf.items():
                print(f"    {op:15s}: avg={info['avg_ms']}ms (n={info['count']})")
            print()

            if report["failed_cases"]:
                print(f"  失败案例 ({len(report['failed_cases'])}):")
                for fc in report["failed_cases"]:
                    print(f"    [{fc['category']}/{fc['sub_category']}] {fc['sql'][:50]}")
                    print(f"      expected={fc['expected']} actual={fc['actual']}")
                    print(f"      {fc['error']}")
            else:
                print("  全部通过!")
            print()
            print(f"  报告已追加到: {args.output}")
            print("=" * 60)

        # 退出码
        if report["failed_cases"]:
            return 1
        return 0

    finally:
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
