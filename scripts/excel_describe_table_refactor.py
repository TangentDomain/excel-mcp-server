"""
Refactored excel_describe_table function - breaking down into smaller functions
"""

def _detect_dual_header(rows) -> tuple:
    """
    检测双行表头模式
    
    Args:
        rows: 工作表前几行数据
        
    Returns:
        tuple: (is_dual_header, header_row_idx, descriptions)
    """
    is_dual_header = False
    header_row_idx = 0
    descriptions = None
    
    if len(rows) >= 2:
        second_row = rows[1]
        if second_row and len(second_row) >= 3:
            all_valid = all(
                isinstance(c, str) and c.strip().startswith(tuple('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ_'))
                for c in second_row if c is not None
            )
            first_row = rows[0]
            any_chinese = any(
                isinstance(c, str) and any('\u4e00' <= ch <= '\u9fff' for ch in c)
                for c in first_row if c is not None
            )
            if all_valid and any_chinese:
                is_dual_header = True
                header_row_idx = 1
                descriptions = first_row
    
    return is_dual_header, header_row_idx, descriptions

def _collect_column_statistics(ws, data_start, num_cols, col_name_list):
    """
    收集列统计信息 - 优化版本，单次遍历
    
    Args:
        ws: 工作表对象
        data_start: 数据开始行
        num_cols: 列数
        col_name_list: 列名列表
        
    Returns:
        dict: 列统计信息
    """
    col_stats = {}
    for col_idx in range(num_cols):
        col_name = col_name_list[col_idx]
        col_stats[col_name] = {'non_null': 0, 'samples': [], 'type_values': []}
    
    # 单次遍历所有行和列，同时统计行数和列数据
    total_rows = 0
    for row in ws.iter_rows(min_row=data_start + 1, values_only=True):
        total_rows += 1
        for col_idx in range(min(len(row), num_cols)):
            val = row[col_idx]
            if val is not None:
                s = col_stats[col_name_list[col_idx]]
                s['non_null'] += 1
                if len(s['samples']) < 3:
                    s['samples'].append(val)
                if len(s['type_values']) < 100:
                    s['type_values'].append(val)
    
    return col_stats, total_rows

def _analyze_data_types(col_stats):
    """
    分析列数据类型
    
    Args:
        col_stats: 列统计信息
        
    Returns:
        list: 列信息列表
    """
    columns = []
    for col_name, stats in col_stats.items():
        samples = stats['samples']
        
        # 分析数据类型
        if not samples:
            col_type = 'empty'
        elif all(isinstance(x, (int, float)) for x in samples if x is not None):
            col_type = 'number'
        elif all(isinstance(x, str) for x in samples if x is not None):
            # 检查是否可能是日期
            date_like = False
            for sample in samples:
                if sample and isinstance(sample, str):
                    # 简单的日期格式检测
                    if any(char in sample for char in ['-', '/', ':']):
                        date_like = True
                        break
            col_type = 'date' if date_like else 'text'
        else:
            col_type = 'mixed'
        
        columns.append({
            'name': col_name,
            'type': col_type,
            'sample_values': samples,
            'non_null_count': stats['non_null']
        })
    
    return columns

def _prepare_describe_result(sheet_name, is_dual_header, columns, row_count, file_path, sheet):
    """
    准备describe结果
    
    Args:
        sheet_name: 工作表名称
        is_dual_header: 是否双行表头
        columns: 列信息
        row_count: 行数
        file_path: 文件路径
        sheet: 工作表对象
        
    Returns:
        dict: 返回结果
    """
    return _ok(f"表 '{sheet_name}': {len(columns)}列, {row_count}行数据, {'双行表头' if is_dual_header else '单行表头'}", data={
        'sheet_name': sheet_name,
        'header_type': 'dual' if is_dual_header else 'single',
        'row_count': row_count,
        'column_count': len(columns),
        'columns': columns
    }, meta={"file_path": file_path, "sheet": sheet_name})

# Now refactor the main excel_describe_table function to use these helper functions
def excel_describe_table_refactored(
    file_path: str,
    sheet_name: str = None
) -> Dict[str, Any]:
    """
    📋 表结构分析器 - 重构版本（更易维护）
    
    **核心功能**: 快速分析Excel工作表结构，返回列名、数据类型、空值统计等元信息
    
    **改进**:
    • 函数分解：216行拆分为多个小函数，每个函数职责单一
    • 性能优化：单次遍历收集统计信息，减少I/O操作
    • 代码复用：检测逻辑可被其他函数复用
    • 易于测试：每个小函数可独立测试
    
    **参数说明**:
    • **file_path**: Excel文件路径
    • **sheet_name**: 工作表名称（可选）
    
    **返回信息**:
    • **columns**: 列信息列表{name, type, sample_values, non_null_count}
    • **table_stats**: 表统计信息
    • **success**: 操作是否成功
    """
    # 文件验证和加载（保持不变）
    _path_err = _validate_path(file_path)
    if _path_err:
        return _path_err
    if not file_path or not file_path.strip():
        return _fail('文件路径不能为空', meta={"error_code": "MISSING_FILE_PATH"})

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        return _fail(f'无法打开文件: {e}', meta={"error_code": "FILE_OPEN_FAILED"})

    try:
        # 选择工作表
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return _fail(f'工作表 "{sheet_name}" 不存在。可用工作表: {wb.sheetnames}', meta={"error_code": "SHEET_NOT_FOUND"})
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[0]
            sheet_name = ws.title

        # 读取前几行来判断表头类型
        rows = list(ws.iter_rows(max_row=4, values_only=True))
        if not rows:
            return _fail('工作表为空', meta={"error_code": "EMPTY_SHEET"})

        # 使用重构的双行表头检测函数
        is_dual_header, header_row_idx, descriptions = _detect_dual_header(rows)
        
        headers = rows[header_row_idx]
        data_start = header_row_idx + 1
        
        # 准备列名列表
        col_name_list = []
        for col_idx in range(len(headers)):
            col_name = headers[col_idx]
            if col_name is None:
                col_name = f"column_{col_idx + 1}"
            col_name = str(col_name).strip()
            if not col_name:
                col_name = f"column_{col_idx + 1}"
            col_name_list.append(col_name)

        # 使用重构的列统计收集函数
        col_stats, row_count = _collect_column_statistics(ws, data_start, len(headers), col_name_list)
        
        # 使用重构的数据类型分析函数
        columns = _analyze_data_types(col_stats)
        
        # 使用重构的结果准备函数
        return _prepare_describe_result(sheet_name, is_dual_header, columns, row_count, file_path, ws)
        
    except Exception as e:
        return _fail(f'查看表结构失败: {e}', meta={"error_code": "DESCRIBE_FAILED"})
    finally:
        wb.close()