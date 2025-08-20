#!/usr/bin/env python3
"""
简化的公式计算调试
"""

import sys
import tempfile
import os
from pathlib import Path

# 添加src路径
sys.path.insert(0, str(Path(__file__).parent / "src"))

from openpyxl import Workbook, load_workbook

def debug_formula_calculation():
    """调试公式计算过程"""

    print("🔧 开始调试公式计算...")

    # 创建简单的工作簿
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test"

    # 添加数据
    sheet['A1'] = 10
    sheet['A2'] = 20
    sheet['A3'] = 30

    print(f"📊 添加数据: A1={sheet['A1'].value}, A2={sheet['A2'].value}, A3={sheet['A3'].value}")

    # 添加公式
    formula = "SUM(A1:A3)"
    sheet['B1'] = f"={formula}"
    print(f"📝 设置公式: B1={sheet['B1'].value}")

    # 保存到临时文件
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()

    try:
        print(f"💾 保存到文件: {temp_file.name}")
        workbook.save(temp_file.name)
        workbook.close()

        print("🔄 重新加载文件...")

        # 读取公式模式
        formula_workbook = load_workbook(temp_file.name, data_only=False)
        formula_sheet = formula_workbook["Test"]
        print(f"📐 公式模式 B1: {formula_sheet['B1'].value}")
        formula_workbook.close()

        # 读取数据模式
        data_workbook = load_workbook(temp_file.name, data_only=True)
        data_sheet = data_workbook["Test"]
        print(f"📊 数据模式 B1: {data_sheet['B1'].value}")
        print(f"📊 数据模式 A1: {data_sheet['A1'].value}")
        print(f"📊 数据模式 A2: {data_sheet['A2'].value}")
        print(f"📊 数据模式 A3: {data_sheet['A3'].value}")
        data_workbook.close()

        # 测试在Z1位置计算
        z_workbook = load_workbook(temp_file.name, data_only=False)
        z_sheet = z_workbook["Test"]
        z_sheet['Z1'] = f"={formula}"
        z_workbook.save(temp_file.name)
        z_workbook.close()

        # 读取Z1结果
        result_workbook = load_workbook(temp_file.name, data_only=True)
        result_sheet = result_workbook["Test"]
        z1_value = result_sheet['Z1'].value
        print(f"🎯 Z1位置计算结果: {z1_value}")
        result_workbook.close()

    finally:
        # 清理
        try:
            os.unlink(temp_file.name)
        except:
            pass

    print("✅ 调试完成")

if __name__ == "__main__":
    debug_formula_calculation()
