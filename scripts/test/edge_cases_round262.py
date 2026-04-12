#!/usr/bin/env python3
"""
Edge case tests for round 262 (T316-T335).
Tests focus on: forbidden sheet name chars, boundary sheet name lengths,
merge/delete interactions, search edge cases, data validation, format edge cases.
"""

import os
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openpyxl import Workbook, load_workbook
from excel_mcp_server_fastmcp.server import (
    excel_create_sheet,
    excel_delete_sheet,
    excel_rename_sheet,
    excel_copy_sheet,
    excel_search,
    excel_get_headers,
    excel_find_last_row,
    excel_insert_rows,
    excel_insert_columns,
    excel_update_range,
    excel_merge_cells,
    excel_unmerge_cells,
    excel_set_formula,
    excel_format_cells,
    excel_set_data_validation,
    excel_query,
    excel_compare_sheets,
    excel_export_to_csv,
    excel_batch_insert_rows,
    excel_list_sheets,
    excel_upsert_row,
    excel_delete_rows,
)


def create_test_file():
    """Create a test xlsx file with sample data."""
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "Beijing"])
    ws.append(["Bob", 25, "Shanghai"])
    ws.append(["Charlie", 35, "Guangzhou"])
    ws.append(["Diana", 28, "Shenzhen"])
    wb.save(path)
    return path


class TestRunner:
    def __init__(self):
        self.results = []

    def run(self, test_id, description, test_func):
        """Run a single test and record result."""
        try:
            msg = test_func()
            self.results.append((test_id, description, "PASS", msg or ""))
            print(f"[PASS] {test_id}: {description}")
            if msg:
                print(f"       {msg}")
        except AssertionError as e:
            self.results.append((test_id, description, "FAIL", str(e)))
            print(f"[FAIL] {test_id}: {description} - {e}")
        except Exception as e:
            self.results.append((test_id, description, "FAIL", f"{type(e).__name__}: {e}"))
            print(f"[FAIL] {test_id}: {description} - {type(e).__name__}: {e}")
        print()

    def info(self, test_id, description, msg):
        """Record an INFO test."""
        self.results.append((test_id, description, "INFO", msg))
        print(f"[INFO] {test_id}: {description}")
        print(f"       {msg}")
        print()


runner = TestRunner()


# ===== Tests =====

def test_T316():
    """Sheet name with backslash (forbidden in Excel)."""
    path = create_test_file()
    try:
        result = excel_create_sheet(path, "Test\\Sheet")
        assert not result.get('success'), f"Should reject backslash: {result}"
        return "Backslash correctly rejected"
    finally:
        os.unlink(path)


def test_T317():
    """Sheet name exactly 31 characters (Excel max)."""
    path = create_test_file()
    try:
        name_31 = "A" * 31
        result = excel_create_sheet(path, name_31)
        assert result.get('success'), f"31-char sheet name failed: {result}"
        list_result = excel_list_sheets(path)
        sheets = list_result.get('sheets', [])
        assert name_31 in sheets, f"31-char sheet not in list: {sheets}"
        return f"Created and listed 31-char sheet name successfully"
    finally:
        os.unlink(path)


def test_T318():
    """Sheet name with 32 characters (over Excel limit)."""
    path = create_test_file()
    try:
        name_32 = "B" * 32
        result = excel_create_sheet(path, name_32)
        assert not result.get('success'), f"Should reject 32-char name: {result}"
        return "32-char name correctly rejected"
    finally:
        os.unlink(path)


def test_T319():
    """Rename sheet to existing name (same name)."""
    path = create_test_file()
    try:
        result = excel_rename_sheet(path, "Test", "Test")
        # Either fail or be a no-op
        assert result is not None, "No response"
        if result.get('success'):
            return "Rename to same name succeeded (no-op behavior)"
        else:
            return "Rename to same name correctly rejected"
    finally:
        os.unlink(path)


def test_T320():
    """Merge cells spanning entire used range."""
    path = create_test_file()
    try:
        result = excel_merge_cells(path, "Test", "A1:C5")
        assert result.get('success'), f"Merge entire range failed: {result}"
        wb = load_workbook(path)
        ws = wb["Test"]
        assert len(ws.merged_cells.ranges) > 0, "No merged cells found"
        return f"Merged {len(ws.merged_cells.ranges)} range(s)"
    finally:
        os.unlink(path)


def test_T321():
    """Search with regex .* pattern."""
    path = create_test_file()
    try:
        result = excel_search(path, ".*", use_regex=True)
        assert result.get('success'), f"Regex search failed: {result}"
        # data is a list of matches directly
        data = result.get('data', [])
        assert isinstance(data, list), f"Expected list, got {type(data)}"
        assert len(data) > 0, "No matches for .* pattern"
        return f"Found {len(data)} matches for .*"
    finally:
        os.unlink(path)


def test_T322():
    """Query with HAVING clause."""
    path = create_test_file()
    try:
        result = excel_query(path, "SELECT City, COUNT(*) as cnt FROM Test GROUP BY City HAVING cnt >= 1")
        assert result.get('success'), f"HAVING query failed: {result}"
        data = result.get('data', [])
        assert len(data) > 0, "No results"
        return f"HAVING query returned {len(data)} rows"
    finally:
        os.unlink(path)


def test_T323():
    """Batch insert rows with empty data list."""
    path = create_test_file()
    try:
        result = excel_batch_insert_rows(path, "Test", [])
        assert not result.get('success'), f"Should reject empty data: {result}"
        return "Empty batch correctly rejected"
    finally:
        os.unlink(path)


def test_T324():
    """Insert columns at position 1 (before all data)."""
    path = create_test_file()
    try:
        result = excel_insert_columns(path, "Test", column_index=1, count=1)
        assert result.get('success'), f"Insert column at pos 1 failed: {result}"
        headers = excel_get_headers(path, "Test")
        header_list = headers.get('data', {}).get('headers', [])
        assert len(header_list) >= 3, f"Expected >=3 headers after insert, got {len(header_list)}: {header_list}"
        return f"Inserted column, now {len(header_list)} headers"
    finally:
        os.unlink(path)


def test_T325():
    """Search for empty string."""
    path = create_test_file()
    try:
        result = excel_search(path, "")
        if result.get('success'):
            return "Empty search string accepted"
        else:
            return "Empty search string correctly rejected"
    finally:
        os.unlink(path)


def test_T326():
    """Get headers on empty sheet."""
    path = create_test_file()
    try:
        excel_create_sheet(path, "Empty")
        result = excel_get_headers(path, "Empty")
        if result.get('success'):
            headers = result.get('data', {}).get('headers', [])
            return f"Headers on empty sheet: {headers}"
        else:
            return f"Headers on empty sheet returned: {result.get('error', result.get('message', 'unknown'))}"
    finally:
        os.unlink(path)


def test_T327():
    """Format cells with custom number format."""
    path = create_test_file()
    try:
        result = excel_format_cells(path, "Test", "B2:B5",
                                    formatting={"number_format": "#,##0.00"})
        assert result.get('success'), f"Custom format failed: {result}"
        wb = load_workbook(path)
        ws = wb["Test"]
        assert ws["B2"].number_format == "#,##0.00", \
            f"Format not applied: {ws['B2'].number_format}"
        return "Custom number_format applied correctly"
    finally:
        os.unlink(path)


def test_T328():
    """Set data validation with list type."""
    path = create_test_file()
    try:
        result = excel_set_data_validation(path, "Test", "A2:A5",
                                          "list", '"Red,Green,Blue"',
                                          "Color", "Choose a color")
        assert result.get('success'), f"Data validation failed: {result}"
        wb = load_workbook(path)
        ws = wb["Test"]
        dv = ws.data_validations.dataValidation if ws.data_validations else []
        assert len(dv) > 0, "No data validation found"
        return f"Data validation set with {len(dv)} rule(s)"
    finally:
        os.unlink(path)


def test_T329():
    """Delete rows then find_last_row."""
    path = create_test_file()
    try:
        result = excel_delete_rows(path, "Test", row_index=2, count=4)
        assert result.get('success'), f"Delete rows failed: {result}"
        last = excel_find_last_row(path, "Test")
        last_row = last.get('data', {}).get('last_row', None)
        assert last_row is not None, "find_last_row returned None"
        return f"After deleting rows 2-5, last_row={last_row}"
    finally:
        os.unlink(path)


def test_T330():
    """Copy sheet and verify data independence."""
    path = create_test_file()
    try:
        result = excel_copy_sheet(path, "Test", "TestCopy")
        assert result.get('success'), f"Copy sheet failed: {result}"
        excel_update_range(path, "A2", [["MODIFIED"]])
        wb = load_workbook(path)
        ws_copy = wb["TestCopy"]
        assert ws_copy["A2"].value == "Alice", \
            f"Copy not independent: expected 'Alice', got '{ws_copy['A2'].value}'"
        return "Copy is data-independent from original"
    finally:
        os.unlink(path)


def test_T331():
    """Query with BETWEEN clause."""
    path = create_test_file()
    try:
        result = excel_query(path, "SELECT * FROM Test WHERE Age BETWEEN 25 AND 30")
        assert result.get('success'), f"BETWEEN query failed: {result}"
        data = result.get('data', [])
        assert isinstance(data, list), f"Expected list, got {type(data)}"
        assert len(data) >= 2, f"Expected >=2 rows, got {len(data)}"
        return f"BETWEEN query returned {len(data)} rows"
    finally:
        os.unlink(path)


def test_T332():
    """Export to CSV and verify content."""
    path = create_test_file()
    csv_path = path.replace('.xlsx', '.csv')
    try:
        result = excel_export_to_csv(path, csv_path)
        assert result.get('success'), f"Export CSV failed: {result}"
        assert os.path.exists(csv_path), "CSV file not created"
        with open(csv_path, 'r') as f:
            content = f.read()
        assert "Alice" in content and "Beijing" in content, "Data missing from CSV"
        return "CSV export verified with correct content"
    finally:
        os.unlink(path)
        if os.path.exists(csv_path):
            os.unlink(csv_path)


def test_T333():
    """Compare sheets with different column orders."""
    path1 = create_test_file()
    path2 = create_test_file()
    try:
        wb2 = load_workbook(path2)
        ws2 = wb2["Test"]
        for row in ws2.iter_rows(min_row=1, max_row=5):
            row[1].value, row[2].value = row[2].value, row[1].value
        wb2.save(path2)

        result = excel_compare_sheets(path1, "Test", path2, "Test")
        assert result.get('success'), f"Compare failed: {result}"
        diff_count = result.get('data', {}).get('difference_count', 0)
        return f"Column swap detected: {diff_count} differences"
    finally:
        os.unlink(path1)
        os.unlink(path2)


def test_T334():
    """Copy sheet with CJK data and special name."""
    path = create_test_file()
    try:
        # Add CJK data
        excel_update_range(path, "A6", [["日本語テスト", " Café Résumé ", "中文测试"]], sheet_name="Test")
        result = excel_copy_sheet(path, "Test", "特殊データ")
        assert result.get('success'), f"Copy sheet failed: {result}"
        wb = load_workbook(path)
        ws = wb["特殊データ"]
        val = ws["A6"].value
        assert val == "日本語テスト", f"CJK data not copied: {val}"
        return "CJK data and special sheet name copied correctly"
    finally:
        os.unlink(path)


def test_T335():
    """Upsert row - update existing."""
    path = create_test_file()
    try:
        result = excel_upsert_row(path, "Test", "Name", "Alice",
                                   {"Age": 31, "City": "Hangzhou"})
        assert result.get('success'), f"Upsert failed: {result}"
        wb = load_workbook(path)
        ws = wb["Test"]
        assert ws["B2"].value == 31, f"Age not updated: {ws['B2'].value}"
        assert ws["C2"].value == "Hangzhou", f"City not updated: {ws['C2'].value}"
        return "Upsert updated existing row correctly"
    finally:
        os.unlink(path)


# ===== Run all tests =====
if __name__ == "__main__":
    print("=" * 60)
    print("Edge Case Tests - Round 262 (T316-T335)")
    print("=" * 60)
    print()

    tests = [
        ("T316", "Sheet name with backslash (forbidden)", test_T316),
        ("T317", "Sheet name exactly 31 chars (Excel max)", test_T317),
        ("T318", "Sheet name with 32 chars (over limit)", test_T318),
        ("T319", "Rename sheet to same name", test_T319),
        ("T320", "Merge cells spanning entire used range", test_T320),
        ("T321", "Search with regex .* pattern", test_T321),
        ("T322", "Query with HAVING clause", test_T322),
        ("T323", "Batch insert rows with empty data", test_T323),
        ("T324", "Insert column at position 1", test_T324),
        ("T325", "Search for empty string", test_T325),
        ("T326", "Get headers on empty sheet", test_T326),
        ("T327", "Format cells with custom number_format", test_T327),
        ("T328", "Set data validation (list type)", test_T328),
        ("T329", "Delete all data rows then find_last_row", test_T329),
        ("T330", "Copy sheet data independence", test_T330),
        ("T331", "Query with BETWEEN clause", test_T331),
        ("T332", "Export to CSV and verify content", test_T332),
        ("T333", "Compare sheets with different column orders", test_T333),
        ("T334", "Copy sheet with CJK data and special name", test_T334),
        ("T335", "Upsert row - update existing", test_T335),
    ]

    for test_id, desc, func in tests:
        runner.run(test_id, desc, func)

    # Summary
    print("=" * 60)
    print("Summary")
    print("=" * 60)
    pass_count = sum(1 for r in runner.results if r[2] == "PASS")
    fail_count = sum(1 for r in runner.results if r[2] == "FAIL")
    info_count = sum(1 for r in runner.results if r[2] == "INFO")
    print(f"Total: {len(runner.results)} | PASS: {pass_count} | FAIL: {fail_count} | INFO: {info_count}")

    if fail_count > 0:
        print("\nFailed tests:")
        for r in runner.results:
            if r[2] == "FAIL":
                print(f"  {r[0]}: {r[1]} - {r[3]}")

    if info_count > 0:
        print("\nInfo tests:")
        for r in runner.results:
            if r[2] == "INFO":
                print(f"  {r[0]}: {r[1]} - {r[3]}")
