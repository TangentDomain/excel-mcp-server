#!/usr/bin/env python3
"""
边缘案例测试脚本 - 第268轮 - T436-T455
聚焦：format_cells边界、find_last_row边界、batch_insert边界、
     describe_table边界、evaluate_formula边界、check_duplicate_ids边界、
     SQL查询功能验证、空文件处理、get_range边界
"""

import subprocess
import json
import tempfile
import os
from openpyxl import Workbook


class EdgeCaseTester:
    def __init__(self):
        self.results = []

    def create_test_file(self, data=None, sheet_names=None):
        """用openpyxl创建测试文件，避免MCP进程间文件不共享的问题"""
        filepath = tempfile.mktemp(suffix=".xlsx")
        wb = Workbook()
        if sheet_names and len(sheet_names) > 1:
            wb.remove(wb.active)
            for name in sheet_names:
                wb.create_sheet(name)
        else:
            ws = wb.active
            if sheet_names:
                ws.title = sheet_names[0]
            if data:
                for row in data:
                    ws.append(row)
        wb.save(filepath)
        return filepath

    def call_mcp(self, method, params):
        """调用MCP工具，返回解析后的工具结果"""
        proc = subprocess.Popen(
            ["uvx", "--from", ".", "excel-mcp-server-fastmcp"],
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )

        init = json.dumps({
            "jsonrpc": "2.0", "id": 0,
            "method": "initialize",
            "params": {
                "protocolVersion": "2024-11-05",
                "capabilities": {},
                "clientInfo": {"name": "test", "version": "1.0"}
            }
        })
        proc.stdin.write(init + "\n")
        proc.stdin.flush()
        proc.stdout.readline()

        request = json.dumps({
            "jsonrpc": "2.0", "id": 1,
            "method": "tools/call",
            "params": {"name": method, "arguments": params}
        })

        try:
            proc.stdin.write(request + "\n")
            proc.stdin.flush()
            response = proc.stdout.readline()
            proc.terminate()

            result = json.loads(response)
            if "error" in result:
                return {"success": False, "message": result["error"].get("message", str(result["error"]))}

            r = result.get("result", {})
            content = r.get("content", [])
            if content and isinstance(content, list):
                text = content[0].get("text", "")
                try:
                    return json.loads(text)
                except json.JSONDecodeError:
                    return {"success": bool(text), "data": text}
            return r
        except Exception as e:
            return {"success": False, "message": str(e)}

    def test(self, num, description, operation, expected_pass=True):
        """执行单个测试"""
        print(f"测试T{num}: {description}")
        try:
            result = operation()
            success = result.get("success", False)
            has_data = result.get("data") is not None

            if expected_pass:
                if success or has_data:
                    status = "PASS"
                else:
                    status = "FAIL"
            else:
                if not success:
                    status = "PASS"
                else:
                    status = "FAIL"

            self.results.append({
                "num": num, "desc": description,
                "status": status, "result": str(result)[:200]
            })
            print(f"  结果: {status} - {str(result)[:120]}")
        except Exception as e:
            self.results.append({
                "num": num, "desc": description,
                "status": "ERROR", "result": str(e)
            })
            print(f"  错误: {e}")

    def run_tests(self):
        """运行所有测试"""
        # 创建有数据的基础文件
        base_file = self.create_test_file(
            data=[
                ["id", "name", "value"],
                ["1", "Alice", 100],
                ["2", "Bob", 200],
                ["3", "Charlie", 300],
                ["4", "Diana", 400],
                ["5", "Eve", 500]
            ],
            sheet_names=["TestData"]
        )

        # 创建空sheet文件
        empty_file = self.create_test_file(sheet_names=["EmptySheet"])
        # 清除空sheet的默认行
        wb = __import__('openpyxl').load_workbook(empty_file)
        ws = wb.active
        ws.delete_rows(1, ws.max_row)
        wb.save(empty_file)

        # T436: format_cells 超大范围 + 正确参数
        self.test(436, "format_cells超大范围(A1:Z1000)+bold", lambda: self.call_mcp(
            "excel_format_cells",
            {"file_path": base_file, "sheet_name": "TestData",
             "range": "A1:Z1000", "formatting": {"bold": True}}
        ), expected_pass=True)

        # T437: format_cells 单单元格
        self.test(437, "format_cells单单元格+highlight", lambda: self.call_mcp(
            "excel_format_cells",
            {"file_path": base_file, "sheet_name": "TestData",
             "range": "B2", "preset": "highlight"}
        ), expected_pass=True)

        # T438: format_cells 无效range格式
        self.test(438, "format_cells无效range", lambda: self.call_mcp(
            "excel_format_cells",
            {"file_path": base_file, "sheet_name": "TestData",
             "range": "INVALID_RANGE", "formatting": {"bold": True}}
        ), expected_pass=False)

        # T439: find_last_row 空sheet
        self.test(439, "find_last_row空sheet", lambda: self.call_mcp(
            "excel_find_last_row",
            {"file_path": empty_file, "sheet_name": "EmptySheet"}
        ), expected_pass=True)

        # T440: find_last_row 不存在的sheet
        self.test(440, "find_last_row不存在的sheet", lambda: self.call_mcp(
            "excel_find_last_row",
            {"file_path": base_file, "sheet_name": "NonExistentSheet"}
        ), expected_pass=False)

        # T441: describe_table 空sheet
        self.test(441, "describe_table空sheet", lambda: self.call_mcp(
            "excel_describe_table",
            {"file_path": empty_file, "sheet_name": "EmptySheet"}
        ), expected_pass=True)

        # T442: evaluate_formula 简单数学表达式
        self.test(442, "evaluate_formula简单算术", lambda: self.call_mcp(
            "excel_evaluate_formula",
            {"formula": "=1+2+3"}
        ), expected_pass=True)

        # T443: evaluate_formula 引用单元格
        self.test(443, "evaluate_formula单元格引用", lambda: self.call_mcp(
            "excel_evaluate_formula",
            {"formula": "=SUM({1,2,3,4,5})"}
        ), expected_pass=True)

        # T444: evaluate_formula 复杂嵌套函数
        self.test(444, "evaluate_formula嵌套IF", lambda: self.call_mcp(
            "excel_evaluate_formula",
            {"formula": '=IF(10>5,IF(3>1,"yes","no"),"zero")'}
        ), expected_pass=True)

        # T445: evaluate_formula 数组公式
        self.test(445, "evaluate_formulaSUMPRODUCT", lambda: self.call_mcp(
            "excel_evaluate_formula",
            {"formula": "=SUMPRODUCT({1,2,3},{4,5,6})"}
        ), expected_pass=True)

        # T446: check_duplicate_ids 无重复
        self.test(446, "check_duplicate_ids无重复", lambda: self.call_mcp(
            "excel_check_duplicate_ids",
            {"file_path": base_file, "sheet_name": "TestData", "id_column": "A"}
        ), expected_pass=True)

        # T447: check_duplicate_ids 不存在的列
        self.test(447, "check_duplicate_ids不存在的列", lambda: self.call_mcp(
            "excel_check_duplicate_ids",
            {"file_path": base_file, "sheet_name": "TestData", "id_column": "Z"}
        ), expected_pass=False)

        # T448: batch_insert_rows 超大数据量
        self.test(448, "batch_insert_rows超大数据(100行)", lambda: self.call_mcp(
            "excel_batch_insert_rows",
            {"file_path": base_file, "sheet_name": "TestData",
             "data": [{"id": str(i), "name": f"user_{i}", "value": i * 10} for i in range(100)]}
        ), expected_pass=True)

        # T449: batch_insert_rows 空数据
        self.test(449, "batch_insert_rows空数据", lambda: self.call_mcp(
            "excel_batch_insert_rows",
            {"file_path": base_file, "sheet_name": "TestData", "data": []}
        ), expected_pass=True)

        # T450: get_range 超出数据范围
        self.test(450, "get_range超出数据范围", lambda: self.call_mcp(
            "excel_get_range",
            {"file_path": base_file, "sheet_name": "TestData", "range": "A100:Z200"}
        ), expected_pass=True)

        # T451: SQL查询 ORDER BY DESC
        self.test(451, "SQL ORDER BY DESC", lambda: self.call_mcp(
            "excel_query",
            {"file_path": base_file, "query_expression": "SELECT * FROM TestData ORDER BY value DESC"}
        ), expected_pass=True)

        # T452: SQL查询 DISTINCT
        self.test(452, "SQL DISTINCT查询", lambda: self.call_mcp(
            "excel_query",
            {"file_path": base_file, "query_expression": "SELECT DISTINCT name FROM TestData"}
        ), expected_pass=True)

        # T453: SQL查询 LIKE模糊匹配
        self.test(453, "SQL LIKE模糊匹配", lambda: self.call_mcp(
            "excel_query",
            {"file_path": base_file, "query_expression": "SELECT * FROM TestData WHERE name LIKE 'A%' OR name LIKE 'B%'"}
        ), expected_pass=True)

        # T454: SQL查询 COUNT聚合
        self.test(454, "SQL COUNT聚合", lambda: self.call_mcp(
            "excel_query",
            {"file_path": base_file, "query_expression": "SELECT COUNT(*) as cnt FROM TestData"}
        ), expected_pass=True)

        # T455: SQL查询 IN子句
        self.test(455, "SQL IN子句", lambda: self.call_mcp(
            "excel_query",
            {"file_path": base_file, "query_expression": "SELECT * FROM TestData WHERE name IN ('Alice', 'Eve')"}
        ), expected_pass=True)

        # 清理
        for f in [base_file, empty_file]:
            try:
                os.unlink(f)
            except:
                pass

    def print_summary(self):
        pass_count = sum(1 for r in self.results if r["status"] == "PASS")
        fail_count = sum(1 for r in self.results if r["status"] == "FAIL")
        error_count = sum(1 for r in self.results if r["status"] == "ERROR")

        print("\n=== 第268轮统计 ===")
        print(f"总计: {len(self.results)}个边缘案例")
        print(f"通过: {pass_count}个")
        print(f"失败: {fail_count}个")
        print(f"错误: {error_count}个")

        for r in self.results:
            if r["status"] in ("FAIL", "ERROR"):
                print(f"  [{r['status']}] T{r['num']}: {r['desc']}")
                print(f"    {r['result']}")

        return {"total": len(self.results), "pass": pass_count, "fail": fail_count, "error": error_count}


if __name__ == "__main__":
    tester = EdgeCaseTester()
    tester.run_tests()
    summary = tester.print_summary()
    print(json.dumps(summary))
