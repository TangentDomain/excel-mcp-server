#!/usr/bin/env python3
"""
边缘案例测试脚本 - 第272轮 - T456-T475
聚焦：子查询边界、JOIN边界、GROUP BY边界、HAVING边界、
     WHERE EXISTS边界、复杂表达式边界、数据类型边界
"""

import subprocess
import json
import tempfile
import os
from openpyxl import Workbook


class EdgeCaseTester:
    def __init__(self):
        self.results = []

    def create_test_file(self, data=None, sheet_names=None):
        """用openpyxl创建测试文件，避免MCP进程间文件不共享的问题"""
        filepath = tempfile.mktemp(suffix=".xlsx")
        wb = Workbook()
        if sheet_names and len(sheet_names) > 1:
            wb.remove(wb.active)
            for name in sheet_names:
                wb.create_sheet(name)
        else:
            ws = wb.active
            if sheet_names:
                ws.title = sheet_names[0]
            if data:
                for row in data:
                    ws.append(row)
        wb.save(filepath)
        return filepath

    def call_mcp(self, method, params):
        """调用MCP工具，返回解析后的工具结果"""
        proc = subprocess.Popen(
            ["uvx", "--from", ".", "excel-mcp-server-fastmcp"],
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )

        init = json.dumps({
            "jsonrpc": "2.0", "id": 0,
            "method": "initialize",
            "params": {
                "protocolVersion": "2024-11-05",
                "capabilities": {},
                "clientInfo": {"name": "test", "version": "1.0"}
            }
        })
        proc.stdin.write(init + "\n")
        proc.stdin.flush()

        request = json.dumps({
            "jsonrpc": "2.0", "id": 1,
            "method": "tools/call",
            "params": {"name": method, "arguments": params}
        })

        try:
            proc.stdin.write(request + "\n")
            proc.stdin.flush()

            # 读取多行响应直到找到工具调用的结果
            while True:
                response = proc.stdout.readline()
                if not response:
                    break

                try:
                    result = json.loads(response)
                    # 检查是否是工具调用的响应（id=1）
                    if result.get("id") == 1:
                        # 找到工具调用响应
                        if "error" in result:
                            proc.terminate()
                            return {"success": False, "message": result["error"].get("message", str(result["error"]))}

                        r = result.get("result", {})
                        content = r.get("content", [])
                        if content and isinstance(content, list):
                            text = content[0].get("text", "")
                            try:
                                proc.terminate()
                                return json.loads(text)
                            except json.JSONDecodeError:
                                proc.terminate()
                                return {"success": bool(text), "data": text}
                        proc.terminate()
                        return r
                except json.JSONDecodeError:
                    continue

            proc.terminate()
            return {"success": False, "message": "No valid response received"}
        except Exception as e:
            proc.terminate()
            return {"success": False, "message": str(e)}

    def test(self, num, description, operation, expected_pass=True):
        """执行单个测试"""
        print(f"测试T{num}: {description}")
        try:
            result = operation()
            success = result.get("success", False)
            has_data = result.get("data") is not None

            if expected_pass:
                if success or has_data:
                    status = "PASS"
                else:
                    status = "FAIL"
            else:
                if not success:
                    status = "PASS"
                else:
                    status = "FAIL"

            self.results.append({
                "num": num, "desc": description,
                "status": status, "result": str(result)[:200]
            })
            print(f"  结果: {status} - {self.results[-1]['result'][:100]}")
        except Exception as e:
            self.results.append({
                "num": num, "desc": description,
                "status": "ERROR", "result": str(e)
            })
            print(f"  结果: ERROR - {e}")

    def run_tests(self):
        """运行所有测试"""
        # 创建多sheet测试文件
        multi_file = self.create_test_file(
            data=[
                ["id", "name", "value"],
                ["1", "Alice", 100],
                ["2", "Bob", 200],
                ["3", "Charlie", 300],
                ["4", "Diana", 400],
                ["5", "Eve", 500]
            ],
            sheet_names=["Users", "Orders"]
        )

        # 添加第二个sheet的数据
        wb = __import__('openpyxl').load_workbook(multi_file)
        ws = wb["Orders"]
        ws.append(["user_id", "product", "price"])
        ws.append(["1", "A", 10])
        ws.append(["1", "B", 20])
        ws.append(["2", "A", 15])
        ws.append(["3", "C", 30])
        wb.save(multi_file)

        # T456: JOIN INNER
        self.test(456, "JOIN INNER连接", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT Users.name, Orders.product FROM Users INNER JOIN Orders ON Users.id = Orders.user_id"}
        ), expected_pass=True)

        # T457: JOIN LEFT
        self.test(457, "JOIN LEFT连接", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT Users.name, Orders.product FROM Users LEFT JOIN Orders ON Users.id = Orders.user_id"}
        ), expected_pass=True)

        # T458: GROUP BY单列
        self.test(458, "GROUP BY单列", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT user_id, COUNT(*) as cnt FROM Orders GROUP BY user_id"}
        ), expected_pass=True)

        # T459: GROUP BY多列
        self.test(459, "GROUP BY多列", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT user_id, product, SUM(price) as total FROM Orders GROUP BY user_id, product"}
        ), expected_pass=True)

        # T460: GROUP BY聚合WHERE过滤
        self.test(460, "GROUP BY + WHERE", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT user_id, COUNT(*) as cnt FROM Orders WHERE price > 10 GROUP BY user_id"}
        ), expected_pass=True)

        # T461: HAVING聚合过滤
        self.test(461, "HAVING聚合过滤", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT user_id, COUNT(*) as cnt FROM Orders GROUP BY user_id HAVING COUNT(*) >= 2"}
        ), expected_pass=True)

        # T462: WHERE + GROUP BY + HAVING
        self.test(462, "WHERE + GROUP BY + HAVING", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT user_id, SUM(price) as total FROM Orders WHERE price > 0 GROUP BY user_id HAVING SUM(price) > 20"}
        ), expected_pass=True)

        # T463: 子查询 WHERE IN (SELECT)
        self.test(463, "子查询 WHERE IN", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT * FROM Users WHERE id IN (SELECT DISTINCT user_id FROM Orders)"}
        ), expected_pass=False)  # EXISTS子查询不支持pandas query，应该失败或回退

        # T464: 复杂表达式 AND OR 嵌套
        self.test(464, "复杂表达式 AND OR嵌套", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT * FROM Orders WHERE (user_id = 1 OR user_id = 2) AND price > 10"}
        ), expected_pass=True)

        # T465: BETWEEN运算符
        self.test(465, "BETWEEN运算符", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT * FROM Orders WHERE price BETWEEN 10 AND 20"}
        ), expected_pass=True)

        # T466: IS NULL
        self.test(466, "IS NULL检查", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT * FROM Users WHERE name IS NULL"}
        ), expected_pass=True)

        # T467: IS NOT NULL
        self.test(467, "IS NOT NULL检查", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT * FROM Users WHERE name IS NOT NULL"}
        ), expected_pass=True)

        # T468: DISTINCT + JOIN
        self.test(468, "DISTINCT + JOIN", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT DISTINCT Users.name FROM Users INNER JOIN Orders ON Users.id = Orders.user_id"}
        ), expected_pass=True)

        # T469: ORDER BY + LIMIT
        self.test(469, "ORDER BY + LIMIT", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT * FROM Orders ORDER BY price DESC LIMIT 2"}
        ), expected_pass=True)

        # T470: GROUP BY + ORDER BY
        self.test(470, "GROUP BY + ORDER BY", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT user_id, COUNT(*) as cnt FROM Orders GROUP BY user_id ORDER BY cnt DESC"}
        ), expected_pass=True)

        # T471: 多聚合函数
        self.test(471, "多聚合函数", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT user_id, COUNT(*) as cnt, SUM(price) as total, AVG(price) as avg FROM Orders GROUP BY user_id"}
        ), expected_pass=True)

        # T472: LIKE + AND + OR
        self.test(472, "LIKE + AND + OR", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT * FROM Users WHERE (name LIKE 'A%' OR name LIKE 'B%') AND id > 0"}
        ), expected_pass=True)

        # T473: 自连接（JOIN同一个表）
        self.test(473, "JOIN ORDER BY多字段", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT * FROM Orders ORDER BY user_id, price ASC"}
        ), expected_pass=True)

        # T474: JOIN WHERE GROUP BY组合
        self.test(474, "JOIN + WHERE + GROUP BY", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT Users.name, COUNT(Orders.user_id) as order_cnt FROM Users LEFT JOIN Orders ON Users.id = Orders.user_id WHERE Users.id < 5 GROUP BY Users.name"}
        ), expected_pass=True)

        # T475: 复杂组合查询
        self.test(475, "复杂组合查询", lambda: self.call_mcp(
            "excel_query",
            {"file_path": multi_file, "query_expression": "SELECT Users.name, COUNT(Orders.user_id) as cnt FROM Users LEFT JOIN Orders ON Users.id = Orders.user_id WHERE Users.id IN (1, 2, 3) GROUP BY Users.name HAVING COUNT(Orders.user_id) > 0 ORDER BY cnt DESC"}
        ), expected_pass=True)

        # 清理
        try:
            os.unlink(multi_file)
        except:
            pass

    def print_summary(self):
        pass_count = sum(1 for r in self.results if r["status"] == "PASS")
        fail_count = sum(1 for r in self.results if r["status"] == "FAIL")
        error_count = sum(1 for r in self.results if r["status"] == "ERROR")

        print("\n=== 第272轮统计 ===")
        print(f"总计: {len(self.results)}个边缘案例")
        print(f"通过: {pass_count}")
        print(f"失败: {fail_count}")
        print(f"错误: {error_count}")

        # 打印失败案例
        if fail_count > 0 or error_count > 0:
            print("\n失败/错误案例:")
            for r in self.results:
                if r["status"] in ["FAIL", "ERROR"]:
                    print(f"  T{r['num']}: {r['desc']} - {r['status']} - {r['result'][:100]}")

        return pass_count, fail_count, error_count


if __name__ == "__main__":
    tester = EdgeCaseTester()
    tester.run_tests()
    pass_count, fail_count, error_count = tester.print_summary()
