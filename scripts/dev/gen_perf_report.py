"""Generate performance report Excel from pytest --durations=0 output."""
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Read raw pytest output
raw_file = Path(r"C:\Users\Administrator\AppData\Local\Temp\perf_raw.txt")
lines = raw_file.read_text(encoding="utf-8").splitlines()

# Parse duration lines: "0.05s call     tests/test_file.py::TestClass::test_method"
duration_pattern = re.compile(r'^(\d+\.\d+)s\s+(setup|call|teardown)\s+(.+)$')
# Parse test result lines
result_pattern = re.compile(r'^(tests/[^:]+\.py::\S+)\s+(PASSED|FAILED|XFAIL|XPASS|ERROR)')

results = {}
for line in lines:
    m = result_pattern.match(line.strip())
    if m:
        test_id, status = m.group(1), m.group(2)
        results.setdefault(test_id, {'status': status, 'durations': {}})

    m = duration_pattern.match(line.strip())
    if m:
        secs, phase, test_id = m.group(1), m.group(2), m.group(3)
        results.setdefault(test_id, {'status': 'PASSED', 'durations': {}})
        results[test_id]['durations'][phase] = float(secs)

# Build rows
rows = []
for test_id, info in sorted(results.items()):
    parts = test_id.split('::')
    file_name = parts[0] if len(parts) >= 1 else ''
    class_name = parts[1] if len(parts) >= 3 else ''
    method_name = parts[-1] if len(parts) >= 2 else ''

    durations = info['durations']
    setup_t = durations.get('setup', 0)
    call_t = durations.get('call', 0)
    teardown_t = durations.get('teardown', 0)
    total_t = setup_t + call_t + teardown_t

    rows.append({
        'file': file_name,
        'class': class_name,
        'test': method_name,
        'status': info['status'],
        'setup_s': round(setup_t, 4),
        'call_s': round(call_t, 4),
        'teardown_s': round(teardown_t, 4),
        'total_s': round(total_t, 4),
    })

rows.sort(key=lambda r: r['total_s'], reverse=True)

# Write Excel
wb = Workbook()
ws = wb.active
ws.title = "Test Performance Report"

headers = ['#', 'File', 'Class', 'Test', 'Status', 'Setup (s)', 'Call (s)', 'Teardown (s)', 'Total (s)']
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

status_colors = {
    'PASSED': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    'FAILED': PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    'XFAIL': PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
}
slow_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
medium_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

for idx, row in enumerate(rows, 1):
    r = idx + 1
    ws.cell(row=r, column=1, value=idx)
    ws.cell(row=r, column=2, value=row['file'])
    ws.cell(row=r, column=3, value=row['class'])
    ws.cell(row=r, column=4, value=row['test'])
    ws.cell(row=r, column=5, value=row['status'])
    ws.cell(row=r, column=6, value=row['setup_s']).number_format = '0.0000'
    ws.cell(row=r, column=7, value=row['call_s']).number_format = '0.0000'
    ws.cell(row=r, column=8, value=row['teardown_s']).number_format = '0.0000'
    ws.cell(row=r, column=9, value=row['total_s']).number_format = '0.0000'

    if row['status'] in status_colors:
        ws.cell(row=r, column=5).fill = status_colors[row['status']]

    total = row['total_s']
    if total >= 1.0:
        ws.cell(row=r, column=9).fill = slow_fill
        ws.cell(row=r, column=9).font = Font(bold=True, color="CC0000")
    elif total >= 0.3:
        ws.cell(row=r, column=9).fill = medium_fill

col_widths = [6, 45, 35, 50, 10, 12, 12, 14, 12]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:I{len(rows)+1}"

# === Summary sheet ===
ws2 = wb.create_sheet("Summary")

file_stats = {}
for row in rows:
    f = row['file']
    file_stats.setdefault(f, {'count': 0, 'passed': 0, 'failed': 0, 'total_s': 0, 'max_s': 0, 'slow_tests': []})
    file_stats[f]['count'] += 1
    file_stats[f]['total_s'] += row['total_s']
    if row['total_s'] > file_stats[f]['max_s']:
        file_stats[f]['max_s'] = row['total_s']
    if row['status'] == 'PASSED':
        file_stats[f]['passed'] += 1
    elif row['status'] == 'FAILED':
        file_stats[f]['failed'] += 1
    if row['total_s'] >= 0.5:
        file_stats[f]['slow_tests'].append(f"{row['test']} ({row['total_s']:.2f}s)")

sorted_files = sorted(file_stats.items(), key=lambda x: x[1]['total_s'], reverse=True)

sum_headers = ['File', 'Tests', 'Passed', 'Failed', 'Total Time (s)', 'Avg (s)', 'Max (s)', 'Slow Tests (>0.5s)']
for col, h in enumerate(sum_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

for idx, (fname, stats) in enumerate(sorted_files, 1):
    r = idx + 1
    ws2.cell(row=r, column=1, value=fname)
    ws2.cell(row=r, column=2, value=stats['count'])
    ws2.cell(row=r, column=3, value=stats['passed'])
    ws2.cell(row=r, column=4, value=stats['failed'])
    ws2.cell(row=r, column=5, value=round(stats['total_s'], 3)).number_format = '0.000'
    ws2.cell(row=r, column=6, value=round(stats['total_s'] / stats['count'], 4)).number_format = '0.0000'
    ws2.cell(row=r, column=7, value=round(stats['max_s'], 3)).number_format = '0.000'
    ws2.cell(row=r, column=8, value='; '.join(stats['slow_tests']) if stats['slow_tests'] else '')

sum_col_widths = [50, 8, 8, 8, 16, 12, 12, 80]
for i, w in enumerate(sum_col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f"A1:H{len(sorted_files)+1}"

total_row = len(sorted_files) + 2
ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
ws2.cell(row=total_row, column=2, value=sum(s['count'] for s in file_stats.values())).font = Font(bold=True)
ws2.cell(row=total_row, column=5, value=round(sum(s['total_s'] for s in file_stats.values()), 3)).font = Font(bold=True)

output_path = Path(r"D:\excel-mcp-server\test_performance_report.xlsx")
wb.save(str(output_path))
print(f"Saved to {output_path}")
print(f"Total tests: {len(rows)}")
print(f"Tests >= 1.0s: {sum(1 for r in rows if r['total_s'] >= 1.0)}")
print(f"Tests 0.3-1.0s: {sum(1 for r in rows if 0.3 <= r['total_s'] < 1.0)}")
print(f"\nTop 10 slowest:")
for r in rows[:10]:
    print(f"  {r['total_s']:.3f}s  {r['file']}::{r['test']}")
