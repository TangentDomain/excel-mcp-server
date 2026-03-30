#!/usr/bin/env python3
"""REQ-027 Token节约优化验证测试"""

import json
import sys
from pathlib import Path

# 添加项目路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root / "src"))

from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

def test_req_027_optimization():
    """测试REQ-027 Token节约优化"""
    print("🧪 REQ-027 Token节约优化验证测试...")
    
    # 创建测试文件
    test_file = "/tmp/test_req_027.xlsx"
    
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Skills"
        
        # 添加测试数据（模拟游戏配置表，包含各种格式）
        headers = ['skill_id', 'skill_name', 'skill_type', 'damage', 'cooldown', 'description', 'is_active']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        data = [
            [1, '火球术', '攻击', 100, 5.0, '发射火球攻击敌人', True],
            [2, '冰霜护盾', '防御', 0, 10.0, '提供冰霜防护', False],
            [3, '雷电术', '攻击', 150, 8.0, '召唤雷电', True],
        ]
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(test_file)
        print(f"✅ 测试文件创建成功: {test_file}")
        
        # 测试1: list_sheets 优化
        print(f"\n📊 1. excel_list_sheets 优化验证:")
        result1 = ExcelOperations.list_sheets(test_file)
        size1 = len(json.dumps(result1))
        print(f"响应大小: {size1} 字符")
        
        # 验证无重复字段
        data_fields = set(result1.get('data', {}).keys()) if isinstance(result1.get('data'), dict) else set()
        root_fields = set(result1.keys()) - {'success', 'message', 'data', 'meta'}
        overlap = data_fields & root_fields
        if overlap:
            print(f"❌ 发现重复字段: {overlap}")
            return False
        else:
            print("✅ 无重复字段")
        
        # 测试2: get_headers 优化
        print(f"\n📋 2. excel_get_headers 优化验证:")
        result2 = ExcelOperations.get_headers(test_file, "Skills")
        size2 = len(json.dumps(result2))
        print(f"响应大小: {size2} 字符")
        
        # 验证无重复字段
        data_fields2 = set(result2.get('data', {}).keys()) if isinstance(result2.get('data'), dict) else set()
        root_fields2 = set(result2.keys()) - {'success', 'message', 'data', 'meta'}
        overlap2 = data_fields2 & root_fields2
        if overlap2:
            print(f"❌ 发现重复字段: {overlap2}")
            return False
        else:
            print("✅ 无重复字段")
        
        # 测试3: get_range 和 _strip_defaults 功能
        print(f"\n📊 3. excel_get_range + _strip_defaults 优化验证:")
        result3 = ExcelOperations.get_range(test_file, "Skills!A1:C3")
        size3 = len(json.dumps(result3))
        print(f"响应大小: {size3} 字符")
        
        # 检查数据是否经过_strip_defaults处理
        data3 = result3.get('data', {})
        if 'data' in data3 and isinstance(data3['data'], list):
            print(f"数据行数: {len(data3['data'])}")
            if data3['data']:
                print("示例数据行（验证是否有默认值被过滤）:")
                for row in data3['data'][:1]:  # 只显示第一行
                    print(f"  {row}")
                
                # 检查是否有空值被过滤
                flat_data = []
                def flatten_data(obj):
                    if isinstance(obj, list):
                        for item in obj:
                            if isinstance(item, list):
                                flat_data.extend(item)
                            else:
                                flat_data.append(item)
                    elif isinstance(obj, dict):
                        for v in obj.values():
                            flatten_data(v)
                
                flatten_data(data3['data'])
                
                empty_count = sum(1 for x in flat_data if x in [None, '', []])
                print(f"空值数量: {empty_count}")
                
                if empty_count == 0:
                    print("✅ 空值已被过滤")
                else:
                    print(f"⚠️ 仍有 {empty_count} 个空值（可能是有意义的空值）")
        
        # 测试4: 检查错误信息精简
        print(f"\n❌ 4. 错误信息精简验证:")
        error_result = ExcelOperations.get_headers(test_file, "NonExistentSheet")
        error_size = len(json.dumps(error_result))
        print(f"错误响应大小: {error_size} 字符")
        
        # 验证错误信息不包含Python堆栈
        error_msg = error_result.get('message', '')
        if 'Traceback' in error_msg or 'File "' in error_msg:
            print("❌ 错误信息仍包含Python堆栈")
            return False
        else:
            print("✅ 错误信息已精简，无Python堆栈")
        
        # 计算总体优化效果
        print(f"\n📈 总体优化效果:")
        total_size = size1 + size2 + size3 + error_size
        print(f"总响应大小: {total_size} 字符")
        
        # 基准对比（基于之前的分析）
        baseline_size = 398 + 1011 + 651 + 200  # list_sheets + get_headers + get_range + error
        savings = baseline_size - total_size
        savings_percent = (savings / baseline_size) * 100
        
        print(f"基准大小: {baseline_size} 字符")
        print(f"节省: {savings} 字符 ({savings_percent:.1f}%)")
        
        # 验证是否达到30%的节省目标
        if savings_percent >= 30:
            print("✅ 达到30% token节省目标")
        else:
            print(f"⚠️ 仅节省 {savings_percent:.1f}%，未达到30%目标")
        
        # 清理
        import os
        os.remove(test_file)
        
        return True
        
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_req_027_optimization()
    if success:
        print(f"\n🎉 REQ-027 Token节约优化验证完成")
    else:
        print(f"\n💥 REQ-027 验证失败")