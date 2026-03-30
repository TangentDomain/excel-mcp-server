#!/usr/bin/env python3
"""Token优化验证脚本"""

import json
import sys
from pathlib import Path

# 添加项目路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root / "src"))

from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

def test_optimization():
    """测试优化效果"""
    print("🧪 Token优化验证测试...")
    
    # 创建测试文件
    test_file = "/tmp/test_optimization.xlsx"
    
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Skills"
        
        # 添加测试数据（模拟游戏配置表）
        headers = ['skill_id', 'skill_name', 'skill_type', 'damage', 'cooldown', 'description']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        data = [
            [1, '火球术', '攻击', 100, 5.0, '发射火球攻击敌人'],
            [2, '冰霜护盾', '防御', 0, 10.0, '提供冰霜防护'],
            [3, '雷电术', '攻击', 150, 8.0, '召唤雷电'],
        ]
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 添加第二个工作表
        ws2 = wb.create_sheet("Items")
        headers2 = ['item_id', 'item_name', 'rarity', 'price', 'stackable']
        for col, header in enumerate(headers2, 1):
            ws2.cell(row=1, column=col, value=header)
        
        items_data = [
            [101, '生命药水', '普通', 50, True],
            [102, '魔法药水', '稀有', 100, True],
            [103, '传说武器', '传说', 5000, False],
        ]
        
        for row_idx, row_data in enumerate(items_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws2.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(test_file)
        print(f"✅ 测试文件创建成功: {test_file}")
        
        # 测试优化后的list_sheets
        print("\n📊 优化后 excel_list_sheets 响应:")
        result = ExcelOperations.list_sheets(test_file)
        print(f"响应大小: {len(json.dumps(result))} 字符")
        print("响应结构:")
        print(json.dumps(result, indent=2, ensure_ascii=False))
        
        # 验证没有重复字段
        data_fields = set(result.get('data', {}).keys()) if isinstance(result.get('data'), dict) else set()
        root_fields = set(result.keys()) - {'success', 'message', 'data', 'meta'}
        
        overlap = data_fields & root_fields
        if overlap:
            print(f"❌ 发现重复字段: {overlap}")
        else:
            print("✅ 无重复字段")
        
        # 测试get_headers
        print(f"\n📋 excel_get_headers 响应:")
        headers_result = ExcelOperations.get_headers(test_file, "Skills")
        print(f"响应大小: {len(json.dumps(headers_result))} 字符")
        
        # 检查是否有可以优化的默认值
        data = headers_result.get('data', {})
        print("数据字段检查:")
        
        def check_for_defaults(obj, path=""):
            if isinstance(obj, dict):
                for k, v in obj.items():
                    current_path = f"{path}.{k}" if path else k
                    # 检查常见的默认值
                    if v in [False, 0, ''] or (isinstance(v, (list, dict)) and len(v) == 0):
                        print(f"  - {current_path}: {repr(v)} (可能是可优化的默认值)")
                    elif isinstance(v, dict):
                        check_for_defaults(v, current_path)
                    elif isinstance(v, list):
                        check_for_defaults(v, current_path)
        
        check_for_defaults(data)
        
        # 测试get_range
        print(f"\n📊 excel_get_range 响应:")
        range_result = ExcelOperations.get_range(test_file, "Skills!A1:C3")
        print(f"响应大小: {len(json.dumps(range_result))} 字符")
        
        # 检查数据的token优化
        range_data = range_result.get('data', {})
        if 'data' in range_data and isinstance(range_data['data'], list):
            print(f"数据行数: {len(range_data['data'])}")
            if range_data['data']:
                print("示例数据行:")
                for row in range_data['data'][:2]:  # 只显示前两行
                    print(f"  {row}")
        
        # 清理
        import os
        os.remove(test_file)
        
        print(f"\n✅ 优化验证完成")
        
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_optimization()