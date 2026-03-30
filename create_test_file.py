#!/usr/bin/env python3
"""
创建测试Excel文件
"""
import openpyxl
import os

def create_test_file():
    """创建测试Excel文件"""
    test_file = "test_api_problems.xlsx"
    
    # 创建新的Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 写入测试数据
    ws["A1"] = "测试数据"
    ws["B1"] = 100
    ws["C1"] = "2026-03-30"
    ws["A2"] = "游戏配置"
    ws["B2"] = "技能"
    ws["C2"] = "等级1"
    
    # 保存文件
    wb.save(test_file)
    print(f"✅ 创建测试文件: {test_file}")
    return test_file

if __name__ == "__main__":
    test_file = create_test_file()
    print(f"文件路径: {os.path.abspath(test_file)}")