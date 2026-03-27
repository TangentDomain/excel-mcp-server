#!/usr/bin/env python3
"""
REQ-029 真实验证测试 - 修复后的版本
"""

import json
import tempfile
import os
from pathlib import Path
import sys

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from src.excel_mcp_server_fastmcp.server import excel_describe_table, excel_query, excel_batch_insert_rows
from src.excel_mcp_server_fastmcp.api.advanced_sql_query import AdvancedSQLQueryEngine

def create_test_excel():
    """创建测试用Excel文件"""
    temp_dir = tempfile.mkdtemp()
    test_file = os.path.join(temp_dir, "test_req029.xlsx")
    
    # 创建角色和技能数据
    char_data = [
        ["ID", "名称", "职业", "等级"],
        [1, "战士", "战士", 10],
        [2, "法师", "法师", 8],
        [3, "弓箭手", "弓箭手", 9]
    ]
    
    skill_data = [
        ["ID", "名称", "职业限制"],
        [1, "剑术", "战士"],
        [2, "火球术", "法师"],
        [3, "箭术", "弓箭手"]
    ]
    
    # 使用openpyxl写入测试文件
    try:
        from openpyxl import Workbook
        
        # 写入文件
        wb = Workbook()
        
        # 角色表
        ws_char = wb.create_sheet("角色")
        for row in char_data:
            ws_char.append(row)
            
        # 技能表
        ws_skill = wb.create_sheet("技能")
        for row in skill_data:
            ws_skill.append(row)
            
        wb.save(test_file)
        return test_file
    except ImportError:
        print("ERROR: openpyxl not installed")
        return None

def test_join_alias():
    """测试JOIN表别名功能"""
    print("\n=== 测试Bug 1: JOIN表别名 ===")
    
    test_file = create_test_excel()
    if not test_file:
        return False
    
    try:
        # 测试JOIN查询，使用表别名
        query = """
        SELECT r.名称, s.名称 as 技能名称
        FROM 角色 r JOIN 技能 s ON r.职业 = s.职业限制
        WHERE r.等级 >= 9
        """
        
        print(f"执行JOIN查询: {query}")
        
        # 使用AdvancedSQLQueryEngine
        engine = AdvancedSQLQueryEngine()
        result = engine.execute_sql_query(test_file, query)
        
        print(f"JOIN结果: {json.dumps(result, indent=2, ensure_ascii=False)}")
        
        # 验证结果是否包含正确的列名和数据
        if result and isinstance(result, dict) and result.get('success'):
            data = result.get('data', [])
            if len(data) > 1:  # 包含标题行和数据
                headers = data[0]  # 第一行是标题
                rows = data[1:]     # 其余是数据行
                
                # 检查列名是否正确（包含SQL别名）
                if "技能名称" in headers:
                    print("✅ JOIN表别名功能正常 - 列别名正确")
                    # 检查数据内容
                    expected_data = [["战士", "剑术"], ["弓箭手", "箭术"]]
                    if rows == expected_data:
                        print("✅ JOIN表别名功能正常 - 数据内容正确")
                        return True
                    else:
                        print(f"❌ 数据内容不匹配，期望: {expected_data}, 实际: {rows}")
                        return False
                else:
                    print(f"❌ JOIN表别名功能异常 - 列别名不正确，实际列名: {headers}")
                    return False
            else:
                print("❌ JOIN表别名功能异常 - 无数据")
                return False
        else:
            print("❌ JOIN表别名功能异常 - 查询失败")
            return False
            
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)

def test_streaming_describe():
    """测试streaming写入后describe_table功能"""
    print("\n=== 测试Bug 2: streaming写入后describe_table ===")
    
    temp_dir = tempfile.mkdtemp()
    test_file = os.path.join(temp_dir, "test_streaming.xlsx")
    
    try:
        # 创建初始文件
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # 写入一些测试数据
        headers = ["ID", "名称", "数值"]
        data = [[1, "测试1", 100], [2, "测试2", 200]]
        
        for header in headers:
            ws.append([header])
        for row in data:
            ws.append(row)
            
        wb.save(test_file)
        
        # 使用streaming模式插入数据
        insert_data = [
            {"ID": 3, "名称": "新增1", "数值": 300},
            {"ID": 4, "名称": "新增2", "数值": 400}
        ]
        
        print("执行streaming插入...")
        excel_batch_insert_rows(
            file_path=test_file,
            sheet_name='Sheet1',
            data=insert_data,
            streaming=True
        )
        
        # 然后调用describe_table
        print("调用describe_table...")
        result = excel_describe_table(test_file, sheet_name='Sheet1')
        
        print(f"describe_table结果: {json.dumps(result, indent=2, ensure_ascii=False)}")
        
        # 检查是否正常返回
        if isinstance(result, dict) and result.get('success'):
            data = result.get('data', {})
            columns = data.get('columns', [])
            
            # 检查列名是否正确
            expected_columns = ['ID', '名称', '数值']
            actual_columns = [col.get('name') for col in columns]
            
            if actual_columns == expected_columns:
                print("✅ streaming写入后describe_table功能正常 - 列名正确")
                return True
            else:
                print(f"❌ streaming写入后describe_table功能异常 - 列名不匹配")
                print(f"期望列名: {expected_columns}")
                print(f"实际列名: {actual_columns}")
                return False
        else:
            print("❌ streaming写入后describe_table功能异常 - 描述失败")
            return False
            
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)
        if os.path.exists(temp_dir):
            os.rmdir(temp_dir)

def main():
    """主测试函数"""
    print("REQ-029 真实验证开始")
    
    test1_passed = test_join_alias()
    test2_passed = test_streaming_describe()
    
    print(f"\n=== 测试结果 ===")
    print(f"JOIN表别名测试: {'✅ 通过' if test1_passed else '❌ 失败'}")
    print(f"streaming describe测试: {'✅ 通过' if test2_passed else '❌ 失败'}")
    
    overall_success = test1_passed and test2_passed
    print(f"整体结果: {'✅ REQ-029已修复' if overall_success else '❌ REQ-029仍有问题'}")
    
    return overall_success

if __name__ == "__main__":
    main()