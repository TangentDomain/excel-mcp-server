#!/usr/bin/env python3
"""
REQ-015 验证测试：streaming写入后读取工具正常性验证

测试步骤：
1. 创建测试Excel文件
2. 使用streaming模式写入大量数据
3. 验证各种读取工具在streaming写入后的表现
4. 检查是否有崩溃或异常行为
"""

import os
import sys
import tempfile
import json
from datetime import datetime

# 添加项目路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations
from excel_mcp_server_fastmcp.api.advanced_sql_query import AdvancedSQLQueryEngine, execute_advanced_sql_query


def create_test_excel(file_path: str):
    """创建包含测试数据的Excel文件"""
    print(f"创建测试文件: {file_path}")
    
    # 先创建基础Excel文件
    from openpyxl import Workbook
    wb = Workbook()
    
    # 删除默认工作表，创建我们需要的
    ws_default = wb.active
    wb.remove(ws_default)
    
    # 创建角色数据工作表
    ws_chars = wb.create_sheet('角色数据')
    # 添加表头
    headers = ['ID', '名称', '职业', '等级', '生命值', '魔法值', '攻击力', '防御力', '创建时间']
    ws_chars.append(headers)
    
    # 添加测试数据
    for i in range(1, 101):
        ws_chars.append([
            i,
            f'角色{i}',
            ['战士', '法师', '射手'][i % 3],
            i * 10,
            100 + i * 5,
            50 + i * 3,
            20 + i * 2,
            10 + i,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    # 创建装备数据工作表
    ws_items = wb.create_sheet('装备数据')
    item_headers = ['ID', '装备名称', '品质', '等级需求', '价格', '描述']
    ws_items.append(item_headers)
    
    # 添加一些初始装备数据
    for i in range(1, 21):
        ws_items.append([
            i,
            f'装备{i}',
            ['普通', '优秀', '精良', '史诗', '传说'][i % 5],
            i * 2,
            i * 100,
            f'这是装备{i}的详细描述'
        ])
    
    wb.save(file_path)
    
    print("✅ 测试文件创建成功")
    return [headers] + [[i, f'角色{i}', ['战士', '法师', '射手'][i % 3], i * 10, 100 + i * 5, 50 + i * 3, 20 + i * 2, 10 + i, datetime.now().strftime('%Y-%m-%d %H:%M:%S')] for i in range(1, 101)]


def test_streaming_write(file_path: str):
    """测试streaming写入"""
    print("\n=== 测试1: Streaming写入大量数据 ===")
    
    # 准备大量数据
    large_data = []
    headers = ['ID', '装备名称', '品质', '等级需求', '价格', '描述']
    large_data.append(headers)
    
    for i in range(1, 1001):  # 1000行数据
        large_data.append([
            i,
            f'装备{i}',
            ['普通', '优秀', '精良', '史诗', '传说'][i % 5],
            i * 2,
            i * 100,
            f'这是装备{i}的详细描述'
        ])
    
    # 使用streaming模式写入
    result = ExcelOperations.update_range(
        file_path,
        '装备数据!A1:F1001',
        large_data,
        streaming=True,
        insert_mode=False
    )
    
    if not result['success']:
        print(f"❌ Streaming写入失败: {result['message']}")
        return False
    
    print("✅ Streaming写入成功")
    return True


def test_reading_tools_after_streaming(file_path: str):
    """测试streaming写入后的各种读取工具"""
    print("\n=== 测试2: Streaming写入后的读取工具验证 ===")
    
    test_cases = [
        {
            'name': 'get_range',
            'func': ExcelOperations.get_range,
            'params': {'file_path': file_path, 'range_expression': '角色数据!A1:I10'}
        },
        {
            'name': 'get_headers', 
            'func': ExcelOperations.get_headers,
            'params': {'file_path': file_path, 'sheet_name': '角色数据', 'header_row': 1}
        },
        {
            'name': 'find_last_row',
            'func': ExcelOperations.find_last_row,
            'params': {'file_path': file_path, 'sheet_name': '角色数据'}
        },
        {
            'name': 'get_file_info',
            'func': ExcelOperations.get_file_info,
            'params': {'file_path': file_path}
        },
        {
            'name': 'list_sheets',
            'func': ExcelOperations.list_sheets,
            'params': {'file_path': file_path}
        }
    ]
    
    results = []
    for test_case in test_cases:
        try:
            print(f"  测试 {test_case['name']}...", end=' ')
            
            result = test_case['func'](**test_case['params'])
            
            if result['success']:
                print("✅ 成功")
                results.append({'name': test_case['name'], 'status': 'success'})
            else:
                print(f"❌ 失败: {result['message']}")
                results.append({'name': test_case['name'], 'status': 'failed', 'error': result['message']})
                
        except Exception as e:
            print(f"❌ 异常: {str(e)}")
            results.append({'name': test_case['name'], 'status': 'error', 'error': str(e)})
    
    return results


def test_sql_query_after_streaming(file_path: str):
    """测试streaming写入后的SQL查询功能"""
    print("\n=== 测试3: Streaming写入后的SQL查询验证 ===")
    
    test_cases = [
        {
            'name': '基础查询',
            'query': 'SELECT * FROM 角色数据 WHERE 等级 >= 50 LIMIT 5'
        },
        {
            'name': '条件查询',
            'query': 'SELECT 名称, 等级, 攻击力 FROM 角色数据 WHERE 职业 = "战士" LIMIT 3'
        },
        {
            'name': 'JOIN查询',
            'query': '''
                SELECT r.名称, r.职业, e.装备名称, e.品质 
                FROM 角色数据 r 
                JOIN 装备数据 e ON r.ID = e.ID 
                LIMIT 3
            '''
        }
    ]
    
    results = []
    for test_case in test_cases:
        try:
            print(f"  测试 {test_case['name']}...", end=' ')
            
            result = execute_advanced_sql_query(
                file_path=file_path,
                sql=test_case['query']
            )
            
            if result['success']:
                print("✅ 成功")
                results.append({'name': test_case['name'], 'status': 'success'})
            else:
                print(f"❌ 失败: {result['message']}")
                results.append({'name': test_case['name'], 'status': 'failed', 'error': result['message']})
                
        except Exception as e:
            print(f"❌ 异常: {str(e)}")
            results.append({'name': test_case['name'], 'status': 'error', 'error': str(e)})
    
    return results


def main():
    """主测试函数"""
    print("🧪 REQ-015 验证测试开始：Streaming写入后读取工具正常性")
    
    # 创建临时测试文件
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        test_file = tmp_file.name
    
    try:
        # 1. 创建测试文件
        original_data = create_test_excel(test_file)
        
        # 2. 测试streaming写入
        if not test_streaming_write(test_file):
            return False
        
        # 3. 测试streaming写入后的读取工具
        reading_results = test_reading_tools_after_streaming(test_file)
        
        # 4. 测试streaming写入后的SQL查询
        sql_results = test_sql_query_after_streaming(test_file)
        
        # 5. 生成测试报告
        print("\n" + "="*60)
        print("📊 测试结果汇总")
        print("="*60)
        
        # 读取工具结果
        reading_success = sum(1 for r in reading_results if r['status'] == 'success')
        reading_failed = sum(1 for r in reading_results if r['status'] in ['failed', 'error'])
        
        print(f"读取工具: {reading_success}/{len(reading_results)} 成功")
        for result in reading_results:
            if result['status'] != 'success':
                print(f"  ❌ {result['name']}: {result.get('error', '未知错误')}")
        
        # SQL查询结果
        sql_success = sum(1 for r in sql_results if r['status'] == 'success')
        sql_failed = sum(1 for r in sql_results if r['status'] in ['failed', 'error'])
        
        print(f"SQL查询: {sql_success}/{len(sql_results)} 成功")
        for result in sql_results:
            if result['status'] != 'success':
                print(f"  ❌ {result['name']}: {result.get('error', '未知错误')}")
        
        # 总体评估
        total_success = reading_success + sql_success
        total_tests = len(reading_results) + len(sql_results)
        
        print(f"\n🎯 总体评估: {total_success}/{total_tests} 测试通过")
        
        if total_success == total_tests:
            print("✅ 所有测试通过！Streaming写入后读取工具工作正常")
            return True
        else:
            print("❌ 部分测试失败，发现streaming写入后的问题")
            return False
            
    finally:
        # 清理测试文件
        if os.path.exists(test_file):
            os.unlink(test_file)


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)