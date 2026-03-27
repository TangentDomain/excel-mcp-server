#!/usr/bin/env python3
"""
调试JOIN表别名问题 - 修复版
"""

import json
import tempfile
import os
from pathlib import Path
import sys

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from src.excel_mcp_server_fastmcp.api.advanced_sql_query import AdvancedSQLQueryEngine

def debug_join_fixed():
    """调试JOIN表别名功能 - 修复版"""
    temp_dir = tempfile.mkdtemp()
    test_file = os.path.join(temp_dir, "debug_join.xlsx")
    
    # 创建角色和技能数据
    char_data = [
        ["ID", "名称", "职业", "等级"],
        [1, "战士", "战士", 10],
        [2, "法师", "法师", 8],
        [3, "弓箭手", "弓箭手", 9]
    ]
    
    skill_data = [
        ["ID", "名称", "职业限制"],
        [1, "剑术", "战士"],
        [2, "火球术", "法师"],
        [3, "箭术", "弓箭手"]
    ]
    
    # 使用openpyxl写入测试文件
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        
        # 角色表
        ws_char = wb.create_sheet("角色")
        for row in char_data:
            ws_char.append(row)
            
        # 技能表
        ws_skill = wb.create_sheet("技能")
        for row in skill_data:
            ws_skill.append(row)
            
        wb.save(test_file)
        
        # 创建引擎并执行JOIN查询
        engine = AdvancedSQLQueryEngine()
        
        # 先加载数据
        worksheets_data = engine._load_data_with_cache(test_file)
        print(f"角色表列名: {list(worksheets_data['角色'].columns)}")
        print(f"技能表列名: {list(worksheets_data['技能'].columns)}")
        
        # 执行JOIN查询
        query = """
        SELECT r.名称, s.名称 as 技能名称
        FROM 角色 r JOIN 技能 s ON r.职业 = s.职业限制
        WHERE r.等级 >= 9
        """
        
        print(f"执行JOIN查询: {query}")
        
        # 模拟JOIN过程，查看中间结果
        try:
            # 1. 解析SQL
            import sqlglot
            parsed_sql = sqlglot.parse_one(query, dialect="mysql")
            print(f"解析后的SQL: {parsed_sql}")
            
            # 2. 执行JOIN并查看中间结果
            result_df = engine._execute_query(parsed_sql, worksheets_data)
            
            print(f"\nJOIN后的DataFrame列名: {list(result_df.columns)}")
            print(f"JOIN后的DataFrame形状: {result_df.shape}")
            print(f"JOIN后的DataFrame内容:")
            print(result_df.to_string())
            
            # 3. 检查JOIN列映射
            print(f"\nJOIN列映射: {getattr(engine, '_join_column_mapping', '无')}")
            print(f"表别名映射: {getattr(engine, '_table_aliases', '无')}")
            
        except Exception as e:
            print(f"JOIN执行错误: {e}")
            import traceback
            traceback.print_exc()
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)
        if os.path.exists(temp_dir):
            os.rmdir(temp_dir)

if __name__ == "__main__":
    debug_join_fixed()