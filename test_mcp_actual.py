#!/usr/bin/env python3
"""
Actual MCP tool testing to reproduce the 5 API issues
"""

import subprocess
import tempfile
import os
import json

def create_test_excel():
    """Create a test Excel file with some data"""
    tmp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_path = tmp_file.name
    tmp_file.close()
    
    # Create a simple Excel file with openpyxl
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add some test data
        ws['A1'] = 'Name'
        ws['B1'] = 'Age'
        ws['C1'] = 'City'
        ws['A2'] = 'Alice'
        ws['B2'] = '25'
        ws['C2'] = 'New York'
        ws['A3'] = 'Bob'
        ws['B3'] = '30'
        ws['C3'] = 'Boston'
        
        wb.save(tmp_path)
        return tmp_path
    except ImportError:
        print("openpyxl not available, creating minimal file")
        # Create a minimal file that can be read
        with open(tmp_path, 'w') as f:
            f.write("test")
        return tmp_path

def test_mcp_tools():
    """Test the actual MCP tools to reproduce issues"""
    
    test_file = create_test_excel()
    
    try:
        print(f"Using test file: {test_file}")
        
        # Test 1: read_data_from_excel range query
        print("\n=== Test 1: read_data_from_excel range query ===")
        cmd = f"echo '{{\"jsonrpc\":\"2.0\",\"id\":1,\"method\":\"tools/call\",\"params\":{{\"name\":\"read_data_from_excel\",\"arguments\":{{\"file_path\":\"{test_file}\",\"sheet_name\":\"Sheet1\",\"start_cell\":\"A1\",\"end_cell\":\"C3\"}}}}}}' | nc localhost 8000"
        print(f"Command: {cmd}")
        # Note: This won't work directly, need proper MCP client
        
        # Test 2: format_range missing parameters  
        print("\n=== Test 2: format_range missing bold parameter ===")
        cmd = f"echo '{{\"jsonrpc\":\"2.0\",\"id\":2,\"method\":\"tools/call\",\"params\":{{\"name\":\"format_range\",\"arguments\":{{\"file_path\":\"{test_file}\",\"sheet_name\":\"Sheet1\",\"start_cell\":\"A1\",\"end_cell\":\"A1\"}}}}}}' | nc localhost 8000"
        print(f"Command: {cmd}")
        
        # Test 3: apply_formula missing formula
        print("\n=== Test 3: apply_formula missing formula parameter ===")
        cmd = f"echo '{{\"jsonrpc\":\"2.0\",\"id\":3,\"method\":\"tools/call\",\"params\":{{\"name\":\"apply_formula\",\"arguments\":{{\"file_path\":\"{test_file}\",\"sheet_name\":\"Sheet1\",\"cell\":\"A1\"}}}}}}' | nc localhost 8000"
        print(f"Command: {cmd}")
        
        # Test 4: read_data_from_excel with sheet_name only
        print("\n=== Test 4: read_data_from_excel search logic ===")
        cmd = f"echo '{{\"jsonrpc\":\"2.0\",\"id\":4,\"method\":\"tools/call\",\"params\":{{\"name\":\"read_data_from_excel\",\"arguments\":{{\"file_path\":\"{test_file}\",\"sheet_name\":\"Sheet1\"}}}}}}' | nc localhost 8000"
        print(f"Command: {cmd}")
        
        # Test 5: write_data_to_excel data format
        print("\n=== Test 5: write_data_to_excel data format ===")
        data = json.dumps([["Name", "Age"], ["Alice", "25"]])
        cmd = f"echo '{{\"jsonrpc\":\"2.0\",\"id\":5,\"method\":\"tools/call\",\"params\":{{\"name\":\"write_data_to_excel\",\"arguments\":{{\"file_path\":\"{test_file}\",\"sheet_name\":\"Sheet1\",\"start_cell\":\"A1\",\"data\":{data}}}}}}}}' | nc localhost 8000"
        print(f"Command: {cmd}")
        
        print("\nNote: These are the MCP tool calls that would reproduce the issues.")
        print("Actual testing requires MCP server connection or using MCP client tools.")
        
    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)

if __name__ == "__main__":
    test_mcp_tools()