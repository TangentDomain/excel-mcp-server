#!/usr/bin/env python3
"""
调试脚本 - 测试max_row=None问题
"""

import os
import sys
import tempfile
from pathlib import Path

# 添加项目路径到Python路径
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root / "src"))

from excel_mcp_server_fastmcp.server import excel_describe_table
import logging

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def test_max_row_issue():
    """测试max_row问题"""
    temp_dir = tempfile.mkdtemp(prefix="excel_max_row_test_")
    test_file = os.path.join(temp_dir, "test_max_row.xlsx")
    
    try:
        # 1. 创建初始数据
        initial_data = {
            'id': [1, 2, 3],
            'name': ['Test1', 'Test2', 'Test3'],
            'value': [10.5, 20.3, 30.8]
        }
        
        # 使用openpyxl直接创建
        from openpyxl import Workbook, load_workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # 写入表头
        ws.append(['id', 'name', 'value'])
        
        # 写入初始数据
        for row in initial_data.values():
            ws.append(row)
        
        # 保存
        wb.save(test_file)
        wb.close()
        
        logger.info("创建初始数据，行数应该为4（含表头）")
        
        # 2. 检查初始max_row
        wb_check = load_workbook(test_file, read_only=True)
        ws_check = wb_check['Sheet1']
        logger.info(f"初始max_row: {ws_check.max_row}")
        wb_check.close()
        
        # 3. 追加数据
        wb_append = load_workbook(test_file)
        ws_append = wb_append['Sheet1']
        
        # 追加1000行数据
        for i in range(1000):
            ws_append.append([i+4, f"Test{i+4}", float(i * 1.1)])
        
        # 保存
        wb_append.save(test_file)
        wb_append.close()
        
        logger.info("追加1000行数据")
        
        # 4. 再次检查max_row
        wb_final = load_workbook(test_file, read_only=True)
        ws_final = wb_final['Sheet1']
        logger.info(f"追加后max_row: {ws_final.max_row}")
        wb_final.close()
        
        # 5. 测试describe_table
        result = excel_describe_table(test_file, 'Sheet1')
        logger.info(f"describe_table结果: {result}")
        
    except Exception as e:
        logger.error(f"测试异常: {e}")
    finally:
        # 清理
        if os.path.exists(test_file):
            os.remove(test_file)
        os.rmdir(temp_dir)

if __name__ == "__main__":
    test_max_row_issue()