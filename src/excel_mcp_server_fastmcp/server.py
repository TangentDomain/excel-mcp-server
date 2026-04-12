#!/usr/bin/env python3
"""Excel MCP Server - 基于 FastMCP 和 openpyxl 的游戏开发Excel配置表工具。

统一返回格式: {success: bool, message: str, data: Any, meta: dict}
- 成功: success=true, data含实际数据
- 失败: success=false, message含错误描述+💡修复建议
"""

# 标准库导入
import functools
import glob
import json
import logging
import os
import re
import shutil
import sys
import tempfile
import threading
import time
from collections import defaultdict
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Optional, List, Dict, Any, Union

# 配置logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

try:
    from mcp.server.fastmcp import FastMCP
except ImportError as e:
    logger.error(f"缺少必要的依赖包: {e}")
    logger.error("请运行: pip install mcp openpyxl")
    exit(1)

# 导入API模块
from .api.excel_operations import ExcelOperations
from .utils.validators import ExcelValidator, DataValidationError
from .utils import extract_rich_text
from .utils.config import (
    MAX_SEARCH_FILES,
    MAX_FILE_SIZE_MB,
    DATA_DENSITY_THRESHOLD,
    LARGE_OPERATION_CELL_THRESHOLD,
)

# ==================== 操作日志系统 ====================
class OperationLogger:
    """操作日志记录器，用于跟踪所有Excel操作"""

    def __init__(self):
        self.log_file = None
        self.current_session = []

    def start_session(self, file_path: str):
        """开始新的操作会话
        
        创建一个新的操作会话，初始化日志记录系统。
        
        Args:
            file_path (str): 要操作的Excel文件路径
            
        Returns:
            None
        """
        self.log_file = os.path.join(
            os.path.dirname(file_path),
            ".excel_mcp_logs",
            f"operations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        )

        os.makedirs(os.path.dirname(self.log_file), exist_ok=True)

        self.current_session = [{
            'session_id': datetime.now().isoformat(),
            'file_path': file_path,
            'operations': []
        }]

        self._save_log()

    def log_operation(self, operation: str, details: Dict[str, Any]):
        """记录操作
        
        将操作记录添加到当前会话中，并保存日志。
        
        Args:
            operation (str): 操作名称（如 'read_cell', 'write_range'）
            details (Dict[str, Any]): 操作详情，包含参数和返回结果
            
        Returns:
            None
        """
        if not self.current_session:
            return

        operation_record = {
            'timestamp': datetime.now().isoformat(),
            'operation': operation,
            'details': details
        }

        self.current_session[0]['operations'].append(operation_record)
        self._save_log()

    def _save_log(self):
        """保存日志到文件
        
        将当前会话的日志记录保存到JSON文件。
        
        Args:
            None
            
        Returns:
            None
            
        Raises:
            IOError: 当文件写入失败时记录错误日志
        """
        if not self.log_file:
            return

        try:
            with open(self.log_file, 'w', encoding='utf-8') as f:
                json.dump(self.current_session, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logger.error(f"保存操作日志失败: {e}")

    def get_recent_operations(self, limit: int = 10) -> List[Dict[str, Any]]:
        """获取最近的操作记录
        
        返回当前会话中最近的操作记录。
        
        Args:
            limit (int): 返回的最大记录数，默认为10
            
        Returns:
            List[Dict[str, Any]]: 最近的操作记录列表，每条记录包含timestamp, operation, details
            
        Raises:
            ValueError: 当limit为负数时
        """
        if not self.current_session:
            return []

        operations = self.current_session[0]['operations']
        return operations[-limit:] if len(operations) > limit else operations

# ==================== 工具调用追踪器 ====================
class ToolCallTracker:
    """工具调用追踪器，记录每个工具的调用次数、耗时、错误率和错误分类"""

    # 错误分类规则：按优先级匹配（前缀/关键词 → 错误类型）
    _ERROR_RULES = [
        ('🔒', 'security'),
        ('文件不存在', 'file_not_found'),
        ('无法加载文件', 'file_load'),
        ('文件路径', 'validation'),
        ('工作表不存在', 'sheet_not_found'),
        ('文件格式', 'file_format'),
        ('文件太大', 'file_too_large'),
        ('无效的', 'validation'),
        ('不支持', 'unsupported'),
        ('列名', 'column'),
        ('SQL语法', 'sql_syntax'),
        ('语法错误', 'sql_syntax'),
        ('engine_error', 'engine'),
        ('execution_error', 'execution'),
        ('syntax_error', 'sql_syntax'),
        ('unsupported_sql', 'unsupported'),
        ('unsupported_feature', 'unsupported'),
        ('file_not_found', 'file_not_found'),
        ('data_load_failed', 'file_load'),
    ]

    def __init__(self):
        self._stats: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
            'call_count': 0,
            'total_time_ms': 0.0,
            'error_count': 0,
            'error_types': defaultdict(int),
            'last_called': None,
            'min_time_ms': float('inf'),
            'max_time_ms': 0.0,
        })
        self._lock = threading.Lock()
        self._start_time = datetime.now()

    def record(self, tool_name: str, duration_ms: float, success: bool = True,
               error_type: str = None):
        """记录一次工具调用，可选指定错误类型"""
        with self._lock:
            s = self._stats[tool_name]
            s['call_count'] += 1
            s['total_time_ms'] += duration_ms
            s['last_called'] = datetime.now().isoformat()
            if duration_ms < s['min_time_ms']:
                s['min_time_ms'] = duration_ms
            if duration_ms > s['max_time_ms']:
                s['max_time_ms'] = duration_ms
            if not success:
                s['error_count'] += 1
                classified = error_type or 'unknown'
                s['error_types'][classified] += 1

    @staticmethod
    def classify_error(message: str) -> str:
        """根据错误消息内容自动分类错误类型"""
        if not message:
            return 'unknown'
        msg_lower = message.lower()
        for keyword, error_type in ToolCallTracker._ERROR_RULES:
            if keyword.lower() in msg_lower:
                return error_type
        return 'unknown'

    def get_stats(self) -> Dict[str, Any]:
        """获取所有工具的调用统计，按调用次数降序排列"""
        with self._lock:
            tools = {}
            global_error_types: Dict[str, int] = defaultdict(int)
            for name, s in sorted(self._stats.items(), key=lambda x: -x[1]['call_count']):
                tools[name] = {
                    'call_count': s['call_count'],
                    'avg_time_ms': round(s['total_time_ms'] / s['call_count'], 1) if s['call_count'] else 0,
                    'min_time_ms': round(s['min_time_ms'], 1) if s['min_time_ms'] != float('inf') else 0,
                    'max_time_ms': round(s['max_time_ms'], 1),
                    'error_count': s['error_count'],
                    'error_types': dict(sorted(s['error_types'].items())),
                    'last_called': s['last_called'],
                }
                for et, count in s['error_types'].items():
                    global_error_types[et] += count
            return {
                'uptime_seconds': round((datetime.now() - self._start_time).total_seconds(), 0),
                'total_calls': sum(s['call_count'] for s in self._stats.values()),
                'total_errors': sum(s['error_count'] for s in self._stats.values()),
                'error_types': dict(sorted(global_error_types.items())),
                'tools': tools,
            }

    def reset(self):
        """重置所有统计"""
        with self._lock:
            self._stats.clear()
            self._start_time = datetime.now()


_tracker = ToolCallTracker()


def _track_call(func):
    """工具调用追踪装饰器，记录每次调用的耗时和结果，自动检测返回值中的错误"""
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        """包装函数，记录工具调用的耗时和结果信息
        
        Args:
            *args: 传递给被装饰函数的位置参数
            **kwargs: 传递给被装饰函数的关键字参数
            
        Returns:
            被装饰函数的执行结果
            
        Note:
            自动检测函数返回值中的错误状态，记录到追踪器中
            如果返回值是dict且success=False，记录为失败
            异常情况也会被捕获并记录
        """
        start = time.perf_counter()
        try:
            result = func(*args, **kwargs)
            duration_ms = (time.perf_counter() - start) * 1000
            # 自动检测返回值中的错误（大多数工具返回dict而非抛异常）
            if isinstance(result, dict) and result.get('success') is False:
                error_msg = result.get('message', '')
                error_type = ToolCallTracker.classify_error(error_msg)
                # 优先使用SQL引擎的错误类型
                qi = result.get('query_info') or {}
                if isinstance(qi, dict) and 'error_type' in qi:
                    error_type = qi['error_type']
                _tracker.record(func.__name__, duration_ms, success=False,
                                error_type=error_type)
                logger.debug(f"[TOOL] {func.__name__}: {duration_ms:.1f}ms ERROR [{error_type}]",
                             extra={'tool': func.__name__, 'duration_ms': round(duration_ms, 1),
                                    'error': error_msg})
            else:
                _tracker.record(func.__name__, duration_ms, success=True)
                logger.debug(f"[TOOL] {func.__name__}: {duration_ms:.1f}ms",
                             extra={'tool': func.__name__, 'duration_ms': round(duration_ms, 1)})
            return result
        except Exception as e:
            duration_ms = (time.perf_counter() - start) * 1000
            error_type = ToolCallTracker.classify_error(str(e))
            _tracker.record(func.__name__, duration_ms, success=False,
                            error_type=error_type)
            logger.debug(f"[TOOL] {func.__name__}: {duration_ms:.1f}ms ERROR [{error_type}]: {e}",
                         extra={'tool': func.__name__, 'duration_ms': round(duration_ms, 1),
                                'error': str(e)})
            raise
    return wrapper


# ==================== 安全验证模块 ====================
class SecurityValidator:
    """文件路径安全验证，防止路径穿越和资源耗尽"""

    # 允许的文件扩展名
    ALLOWED_EXTENSIONS = {'.xlsx', '.xlsm', '.xls', '.csv', '.json', '.bak'}
    # 文件大小上限（50MB）
    MAX_FILE_SIZE = MAX_FILE_SIZE_MB * 1024 * 1024
    # 危险公式模式（DDE攻击、命令执行等）
    DANGEROUS_FORMULA_PATTERNS = [
        re.compile(r'\bDDE\b', re.IGNORECASE),
        re.compile(r'\bCMD\b', re.IGNORECASE),
        re.compile(r'\bSHELL\b', re.IGNORECASE),
        re.compile(r'\bREGISTER\b', re.IGNORECASE),
        re.compile(r'\|.*\|', re.IGNORECASE),  # DDE链接格式
        re.compile(r'@SUM', re.IGNORECASE),  # Lotus 1-2-3 兼容性攻击
    ]

    @classmethod
    def validate_file_path(cls, file_path: str) -> Dict[str, Any]:
        """验证文件路径安全性，返回 {'valid': bool, 'error': str|None}"""
        if not file_path:
            return {'valid': False, 'error': '文件路径不能为空'}

        # 解析为绝对路径并规范化
        try:
            resolved = Path(file_path).resolve()
        except (OSError, ValueError) as e:
            return {'valid': False, 'error': f'无效的文件路径: {e}'}

        # 路径穿越检测：规范化后路径必须以 / 开头（绝对路径）且不含 ..
        if '..' in Path(file_path).parts:
            return {'valid': False, 'error': '文件路径不允许包含 ".." 路径穿越'}

        # 拒绝符号链接（用原始路径检查，resolve会解析符号链接）
        try:
            if Path(file_path).is_symlink():
                return {'valid': False, 'error': '不允许通过符号链接访问文件'}
        except OSError:
            pass

        # 拒绝隐藏文件（以 . 开头）
        basename = os.path.basename(file_path)
        if basename.startswith('.') and not basename.endswith('.bak'):
            return {'valid': False, 'error': f'不允许访问隐藏文件: {basename}'}

        # 扩展名检查（对有扩展名的文件检查）
        _, ext = os.path.splitext(file_path)
        if ext and ext.lower() not in cls.ALLOWED_EXTENSIONS:
            return {'valid': False, 'error': f'不支持的文件格式: {ext}（允许: {", ".join(sorted(cls.ALLOWED_EXTENSIONS))}）'}

        return {'valid': True, 'error': None}

    @classmethod
    def validate_file_size(cls, file_path: str) -> Dict[str, Any]:
        """验证文件大小是否在允许范围内"""
        try:
            size = os.path.getsize(file_path)
            if size > cls.MAX_FILE_SIZE:
                return {'valid': False, 'error': f'文件过大: {size / 1024 / 1024:.1f}MB（上限{cls.MAX_FILE_SIZE / 1024 / 1024:.0f}MB）'}
            return {'valid': True, 'error': None}
        except OSError as e:
            return {'valid': False, 'error': f'无法获取文件大小: {e}'}

    @classmethod
    def validate_formula(cls, formula: str) -> Dict[str, Any]:
        """验证公式安全性，检测危险模式"""
        for pattern in cls.DANGEROUS_FORMULA_PATTERNS:
            if pattern.search(formula):
                return {'valid': False, 'error': f'检测到危险公式模式，公式中不允许包含 {pattern.pattern}'}
        return {'valid': True, 'error': None}

    @classmethod
    def cleanup_orphan_temp_files(cls, temp_dir: str = None) -> int:
        """清理孤儿临时文件（.xlsx.bak），返回清理数量"""
        target = temp_dir or tempfile.gettempdir()
        pattern = os.path.join(target, '*.xlsx.bak')
        cleaned = 0
        for f in glob.glob(pattern):
            try:
                # 超过1小时的临时备份文件自动清理
                age = time.time() - os.path.getmtime(f)
                if age > 3600:
                    os.remove(f)
                    cleaned += 1
            except OSError:
                pass
        return cleaned


def _validate_path(file_path: str) -> Optional[Dict[str, Any]]:
    """统一的文件路径安全验证入口，失败返回错误dict，通过返回None"""
    result = SecurityValidator.validate_file_path(file_path)
    if not result['valid']:
        return _fail(f'🔒 安全验证失败: {result["error"]}', meta={"error_code": "PATH_VALIDATION_FAILED"})
    # 文件存在时额外检查大小
    if os.path.exists(file_path):
        size_result = SecurityValidator.validate_file_size(file_path)
        if not size_result['valid']:
            return _fail(f'🔒 安全验证失败: {size_result["error"]}', meta={"error_code": "FILE_SIZE_EXCEEDED"})
    return None


def _validate_file_path(param='file_path'):
    """路径安全验证装饰器，自动验证指定参数的文件路径。

    Args:
        param: 要验证的参数名（str）或参数名列表（list[str]），
               默认为 'file_path'。列表时依次验证所有参数。
               仅当参数值不为None时验证（支持Optional参数）。

    Returns:
        装饰器函数

    Example:
        @_validate_file_path()          # 验证 file_path 参数
        @_validate_file_path('backup_path')  # 验证 backup_path 参数
        @_validate_file_path(['csv_path', 'output_path'])  # 验证多个参数
    """
    def decorator(func):
        """创建验证装饰器。

        Args:
            func: 被装饰的函数。
        """
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            """包装函数，在调用前验证文件路径参数
            
            Args:
                *args: 传递给被装饰函数的位置参数
                **kwargs: 传递给被装饰函数的关键字参数，包含需要验证的路径参数
                
            Returns:
                如果验证通过，返回被装饰函数的执行结果
                如果验证失败，返回错误信息
                
            Note:
                根据@param_list指定的参数名，从kwargs中获取对应的路径值
                对每个路径值调用_validate_path进行验证
                参数值为None时跳过验证
            """
            params = [param] if isinstance(param, str) else param
            for p_name in params:
                p_value = kwargs.get(p_name)
                if p_value is None:
                    continue
                err = _validate_path(p_value)
                if err:
                    return err
            return func(*args, **kwargs)
        return wrapper
    return decorator


# 全局操作日志器
operation_logger = OperationLogger()

# ==================== 结构化JSON日志 ====================
class JsonFormatter(logging.Formatter):
    """结构化JSON日志格式化器，每条日志输出为单行JSON对象
    
    输出字段: ts(ISO时间戳), level, module, message
    可选字段: tool(工具名), duration_ms(耗时), error(错误信息), file_path, operation
    激活方式: EXCEL_MCP_JSON_LOG=1
    """
    def format(self, record):
        """格式化日志记录为JSON格式
        
        Args:
            record: 日志记录对象，包含levelname, module, message等属性
            
        Returns:
            str: 格式化后的JSON字符串，包含时间戳、级别、模块、消息等字段
            
        Note:
            根据记录中的可选字段(tool, duration_ms, error等)动态添加到JSON中
            使用ensure_ascii=False确保中文字符正常输出
        """
        log_entry = {
            'ts': datetime.now().isoformat(),
            'level': record.levelname,
            'module': record.module,
            'message': record.getMessage(),
        }
        for field in ('tool', 'duration_ms', 'error', 'file_path', 'operation'):
            value = getattr(record, field, None)
            if value is not None:
                log_entry[field] = value
        return json.dumps(log_entry, ensure_ascii=False)


# ==================== 配置和初始化 ====================
# 日志级别: 默认WARNING，设置EXCEL_MCP_DEBUG=1开启DEBUG
_log_level = logging.DEBUG if os.environ.get('EXCEL_MCP_DEBUG') else logging.WARNING
_json_log = os.environ.get('EXCEL_MCP_JSON_LOG')
if _json_log:
    _handler = logging.StreamHandler()
    _handler.setFormatter(JsonFormatter())
    logging.basicConfig(level=_log_level, handlers=[_handler], force=True)
else:
    logging.basicConfig(
        level=_log_level,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(),
        ],
        force=True,
    )
logger = logging.getLogger(__name__)

# 启动时清理孤儿临时文件
_cleaned = SecurityValidator.cleanup_orphan_temp_files()
if _cleaned:
    logger.info(f"启动清理: 删除 {_cleaned} 个孤儿临时文件")

# 创建FastMCP服务器实例，开启调试模式和详细日志
mcp = FastMCP(
    name="excel-mcp",
    instructions=r"""🎮 游戏开发Excel配置表管理专家 — 34个工具

## 🔥 核心原则：SQL优先

**优先使用 `excel_query`** - 所有数据查询分析任务
- 复杂条件筛选 ✅ WHERE, LIKE, IN, BETWEEN, 子查询
- 聚合统计分析 ✅ COUNT, SUM, AVG, MAX, MIN, GROUP BY, HAVING
- 高级查询 ✅ CASE WHEN, CTE(WITH), EXISTS, JOIN(5种)
- 排序限制 ✅ ORDER BY, LIMIT, OFFSET
- 字符串函数 ✅ UPPER, LOWER, TRIM, LENGTH, CONCAT, REPLACE, SUBSTRING
- 结果合并 ✅ UNION, UNION ALL
- 窗口函数 ✅ ROW_NUMBER, RANK, DENSE_RANK

## 📊 工具选择决策树
```
═══ 读数据 ═══
🔥 首选：所有数据查询/分析任务 → excel_query（SQL引擎，批量分析首选）
│   ✅ 复杂条件筛选 → WHERE, LIKE, IN, BETWEEN, 子查询
│   ✅ 聚合统计 → COUNT, SUM, AVG, MAX, MIN, GROUP BY, HAVING
│   ✅ 多表关联 → 5种JOIN类型，支持跨文件查询
│   ✅ 窗口函数 → ROW_NUMBER, RANK, DENSE_RANK
│   ✅ 字符串函数 → UPPER, LOWER, TRIM, LENGTH, CONCAT, REPLACE, SUBSTRING
│
├─ 📍 已知精确坐标（如A1:C10）────────────→ excel_get_range
│
├─ 📊 快速了解表结构（列名+类型+样本值）───→ excel_describe_table
│
├─ 📋 只需表头信息（中文+英文）───────────→ excel_get_headers（更轻量）
│
├─ 🔍 定位文本位置───────────────────────→ excel_search（返回row/column）
│
└─ 🌐 跨文件搜索─────────────────────────→ excel_search_directory

═══ 写数据（重要！选对工具） ═══
┌─ 🔥 批量修改多行？（改10行以上/按条件改）──→ excel_update_query（SQL UPDATE）
│   例: UPDATE 怪物表 SET 血量=血量*2 WHERE 等级>5
│   ✅ 需要计算表达式 → SET 血量=血量*2
│   ✅ 需要预览变更 → dry_run=True
│   ✅ 条件复杂 → WHERE 等级>5 AND 稀有度='传说'
│
├─ 📍 精确坐标写入（知道具体A1:C10）───────→ excel_update_range
│   ⚠️⚠️⚠️ 默认覆盖模式！insert_mode=True 才是插入！
│   🟢 安全追加数据？→ 先 find_last_row → 再 update_range(..., insert_mode=True)
│   🔴 直接覆盖数据？→ update_range(..., insert_mode=False) 【默认，危险】
│
├─ 👍 按ID改单行（知道 key_column + key_value）→ excel_upsert_row
│   例: upsert_row(file, sheet, "ID", 3, {"血量": 900, "攻击力": 70})
│   ✅ 优点：不会误改其他行、行不存在自动插入、参数自文档化
│   ✅ 只改2-3个字段、dict传参方便、幂等安全
│
├─ ➕ 批量插入新行────────────────────────→ excel_insert_query（SQL INSERT）
│   例: INSERT INTO 怪物表 (ID,名称,血量) VALUES (6,'Boss',9999)
│
└─ 🗑️ 删行？
    按条件删？────────────────────────────→ excel_delete_query（SQL DELETE，必须WHERE）
    按行号删？────────────────────────────→ excel_delete_rows（row_index从1开始）

═══ 结构操作 ═══
文件？                      → excel_create_file / excel_export_to_csv / excel_import_from_csv
工作表？                    → excel_create_sheet / excel_delete_sheet / excel_rename_sheet / excel_copy_sheet
行？                        → excel_insert_rows / excel_delete_rows / excel_set_row_height
列？                        → excel_insert_columns / excel_delete_columns / excel_rename_column / excel_set_column_width
样式/合并/边框？            → excel_format_cells（字体+合并+边框+预设样式，一个工具全搞定）
公式？                      → excel_set_formula（以=开头）
末行定位？                  → excel_find_last_row（追加数据前必用）

═══ 对比 & 备份 ═══
按ID对比两表差异？           → excel_compare_sheets（对象级: 新增/删除/修改）
逐单元格对比差异？           → excel_compare_files（单元格级）
备份恢复？                   → excel_create_backup / excel_restore_backup / excel_list_backups
```

## 🔥 LLM 防错自查清单（调用前速查）

> 选完工具后、调用前，花3秒检查这5项，避免90%的错误：

| # | 自查问题 | 常见错误 | 正确做法 |
|---|---------|---------|---------|
| 1 | 我要追加还是覆盖？ | 用 update_range 追加数据却忘了 insert_mode=True | 追加→`insert_mode=True`+先`find_last_row`；覆盖→默认即可 |
| 2 | 我改的是一行还是批量？ | 改单行却用了 update_query（大材小用） | 单行2-3字段→`upsert_row`；需计算/预览/复杂条件→`update_query` |
| 3 | 范围含工作表名吗？ | 多表文件中 cell_range="A1:C10" 报错 | 优先用 `"Sheet名!A1:C10"` 格式 |
| 4 | SQL语句类型对吗？ | 把 SELECT 传给 update_query 或 UPDATE 传给 query | 查询→query / 改→update_query / 增→insert_query / 删→delete_query |
| 5 | 写入后验证了吗？ | 写完就结束，没确认实际效果 | 安全链路：写入 → query验证 → 有备份可恢复 |

**错误信号速认**：返回值含 `覆盖模式` → 你可能忘了 insert_mode=True；含 `💡` → 按提示修复即可

## ⚠️ 双行表头列名注意事项

当 Excel 有双行表头（第1行中文 + 第2行英文）时：
- **SQL工具**（query/update_query/insert_query/delete_query）：中英文名都 ✅ 能用
- **describe_table 返回的列名**：是第2行英文名（如 skill_id, equip_name）
- **upsert_row 的 key_column**：✅ 中英文名都能用（自动检测双行表头，自动兼容）
- **建议**：直接用 describe_table 返回的英文名即可，无需额外确认原始列名

## ⚡ SQL vs Direct 选择原则
```
✅ 用 SQL (query/update_query/insert_query/delete_query)：
   - 需要条件筛选(WHERE)、聚合(GROUP BY)、JOIN、排序(ORDER BY)
   - 批量操作（影响10行以上）
   - 需要预览效果(dry_run)

✅ 用 Direct (get_range/update_range/upsert_row)：
   - 精确坐标操作（知道具体是A1还是C5）
   - 单行或少量行的精确读写
   - 需要保持格式/公式的精确区域写入
```

## ✅ SQL已支持功能
基础: SELECT, DISTINCT, 别名(AS), 数学表达式(+-*/%)
条件: WHERE, 比较运算(=/></≤/≥/≠), LIKE, NOT LIKE, IN, NOT IN, BETWEEN, AND, OR, IS NULL, IS NOT NULL, NOT
高级: WHERE子查询, FROM子查询(FROM (SELECT ...) AS alias), CASE WHEN, COALESCE, EXISTS, CTE(WITH ... AS ...)
聚合: COUNT(*), COUNT(col), COUNT(DISTINCT), SUM, AVG, MAX, MIN, GROUP BY, HAVING, TOTAL行
排序: ORDER BY DESC/ASC, LIMIT, OFFSET
合并: UNION(去重), UNION ALL(不去重)
关联: INNER JOIN, LEFT JOIN, RIGHT JOIN, FULL JOIN, CROSS JOIN（同文件内工作表）
字符串: UPPER, LOWER, TRIM, LENGTH, CONCAT, REPLACE, SUBSTRING, LEFT, RIGHT
窗口: ROW_NUMBER, RANK, DENSE_RANK, LAG, LEAD, FIRST_VALUE, LAST_VALUE, NTILE, PERCENT_RANK, CUME_DIST, AVG/SUM/COUNT/MIN/MAX OVER
数据修改: INSERT INTO ... VALUES (...), DELETE FROM ... WHERE ..., UPDATE ... SET ... WHERE ...

## ❌ SQL不支持
NATURAL JOIN, WITH RECURSIVE, LATERAL JOIN(请改用子查询或CTE), 跨文件JOIN需使用@'path'语法

## ✅ 跨文件JOIN
使用@'path'语法引用其他Excel文件的工作表：
```sql
SELECT a.*, b.掉落物品 FROM 技能表@'./data/技能配置.xlsx' a JOIN 掉落表@'./data/掉落配置.xlsx' b ON a.id = b.skill_id
```

## ✅ FROM子查询
支持多层嵌套FROM子查询：`FROM (SELECT ...) AS alias`。
```sql
SELECT * FROM (SELECT skill_name, damage FROM 技能配置 WHERE damage > 100) AS 高伤技能
```

## ✅ UNION / UNION ALL
合并多个SELECT查询结果。支持ORDER BY和LIMIT。
```sql
SELECT name FROM 技能配置 WHERE 类型='法师' UNION ALL SELECT name FROM 技能配置 WHERE 类型='战士' ORDER BY name LIMIT 10
```

## ⚠️ 重要原则
- 双行表头: 第1行中文描述+第2行英文字段名，中英文列名均可查询
- **统一1-based索引: 第1行=1, 第1列=1（所有工具一致，含insert_rows/delete_rows）**
- **范围格式(cell_range参数): 优先使用 "工作表名!A1:C10"；不含!时单工作表文件自动推断**
- **🔴🔴🔴 update_range 默认覆盖模式（insert_mode=False）！目标区域数据会被直接替换！**
  - ✅ 追加/插入数据：必须设置 `insert_mode=True` + 先用 `find_last_row` 定位末行
  - ✅ 确实要覆盖：不用改参数，但返回消息会标注 `[覆盖模式]`
  - 🟢 安全链路：`find_last_row` → `update_range(..., insert_mode=True)` → `query` 验证

## 🎮 典型用法示例
```sql
-- 聚合统计
SELECT 技能类型, AVG(伤害), COUNT(*) FROM 技能表 GROUP BY 技能类型

-- 高级筛选+子查询
SELECT * FROM 技能表 WHERE 伤害 > (SELECT AVG(伤害) FROM 技能表)

-- 跨文件JOIN
SELECT a.*, b.掉落物品 FROM 技能表@'./data/技能配置.xlsx' a JOIN 掉落表@'./data/掉落配置.xlsx' b ON a.id = b.skill_id
```

## 📦 统一返回格式
所有工具返回统一JSON结构：`{success, message, data, meta}`
- 成功时 `success=true`，`data` 为实际数据
- 失败时 `success=false`，`meta.error_code` 为机器可读错误码
- 所有错误均包含💡修复提示

## 🔄 常用工作流
```
了解结构 → describe_table → 查询分析 → query → 修改数据 → update_query/update_range → 验证结果 → compare_sheets
```
""",
    debug=bool(os.environ.get('EXCEL_MCP_DEBUG')),
    log_level="DEBUG" if os.environ.get('EXCEL_MCP_DEBUG') else "WARNING"
)



# ==================== 统一响应格式 ====================

def _ok(message: str = "", data=None, meta: dict = None) -> dict:
    """统一成功响应生成器
    
    生成标准化的成功响应格式，用于所有MCP工具函数。
    
    Args:
        message (str, optional): 成功消息描述，默认为空字符串
        data (any, optional): 成功时的数据内容，默认为None
        meta (dict, optional): 元信息字典，包含执行时间、缓存状态等，默认为None
        
    Returns:
        dict: 统一的成功响应格式 {success: true, message: str, data: any, meta: dict}
    """
    r: dict = {"success": True}
    if message:
        r["message"] = message
    if data is not None:
        r["data"] = data
    if meta:
        r["meta"] = meta
    return r


# 集中式非SQL错误提示映射 — 让AI在收到错误后知道如何修复
# 格式: error_code → hint字符串
_ERROR_HINTS = {
    'FILE_NOT_FOUND': '请检查文件路径是否正确，或用excel_list_sheets确认文件存在。路径支持绝对路径和相对于工作目录的路径。',
    'SHEET_NOT_FOUND': '请用excel_list_sheets查看该文件有哪些工作表，确认表名拼写正确。',
    'EMPTY_SHEET': '工作表没有数据行。请确认文件内容或检查表头行数设置（默认前1-2行为表头）。',
    'FILE_OPEN_FAILED': '请确认文件是有效的.xlsx格式且未被其他程序锁定。如文件损坏，尝试用Excel打开后另存为新文件。',
    'FILE_SIZE_EXCEEDED': '文件过大，超出安全限制。如需处理大文件，请联系管理员调整限制。',
    'PATH_VALIDATION_FAILED': '文件路径不合法。请使用绝对路径或相对于当前工作目录的路径，不要使用".."路径穿越。',
    'MISSING_FILE_PATH': '请提供文件路径参数。例如：excel_query("配置表.xlsx", "SELECT * FROM 技能表")',
    'MISSING_QUERY': '请提供SQL查询语句。例如：excel_query("配置表.xlsx", "SELECT * FROM 技能表 WHERE 伤害 > 100")',
    'INVALID_FORMAT': '请检查参数值是否在允许范围内。查看工具描述了解支持的选项。',
    'PARAMETER_ORDER_ERROR': '参数顺序或类型错误。cell_range应该是单元格范围（如A1:E10），不是工作表名。正确示例: excel_get_range("文件.xlsx", "Sheet1!A1:E10")',
    'VALIDATION_FAILED': '请检查参数格式是否正确。常见原因：范围表达式格式错误（应为"工作表名!A1:Z100"）、参数类型不匹配。',
    'OPERATION_FAILED': '写入操作失败，文件可能被锁定或磁盘空间不足。请关闭其他正在使用该文件的程序后重试。',
    'PREVIEW_FAILED': '无法预览操作影响。请先用excel_get_range查看当前数据，再重新尝试操作。',
    'ASSESSMENT_FAILED': '数据影响评估失败。请检查文件路径和工作表名是否正确。',
    'DEPENDENCY_MISSING': '需要安装额外依赖。按提示运行pip install命令后重试。',
    'BACKUP_FAILED': '备份创建失败。请检查磁盘空间和文件权限。',
    'BACKUP_NOT_FOUND': '备份文件不存在。请用excel_list_backups查看可用的备份列表。',
    'RESTORE_FAILED': '恢复失败。备份文件可能已损坏。请用excel_list_backups选择其他备份。',
    'LIST_BACKUPS_FAILED': '获取备份列表失败。请检查备份目录是否存在。',
    'DELETE_SHEET_FAILED': '删除工作表失败。文件可能被锁定或工作表名不正确。',
    'DELETE_ROWS_FAILED': '删除行失败。请确认行号范围有效（不能删除表头行）。',
    'DELETE_COLUMNS_FAILED': '删除列失败。请确认列号范围有效。',
    'FORMULA_SECURITY_FAILED': '公式包含不允许的模式（如外部链接）。请移除危险部分后重试。',
    'HISTORY_RETRIEVAL_FAILED': '获取操作历史失败。历史记录文件可能不存在或已损坏。',
    'DESCRIBE_FAILED': '查看表结构失败。请确认文件路径和工作表名正确。',
    'SQL_EXECUTION_FAILED': 'SQL查询执行失败。请检查语法、表名和列名。建议先用excel_describe_table了解表结构。',
    'UPDATE_EXECUTION_FAILED': 'UPDATE执行失败。请检查SQL语法，确保SET和WHERE子句正确。可用dry_run=True预览。',
    'UNSUPPORTED_SQL': '不支持该SQL语法。查询请用excel_query，修改请用excel_update_query。',
}


def _fail(message: str, meta: dict = None) -> dict:
    """统一错误响应生成器
    
    生成标准化的错误响应格式，自动附加错误代码对应的修复提示。
    
    Args:
        message (str): 错误消息描述，通常包含具体的错误信息
        meta (dict, optional): 元信息字典，包含error_code、file_path、suggested_fix等，默认为None
        
    Returns:
        dict: 统一的错误响应格式 {success: false, message: str, meta: dict}
        
    Note:
        当meta中包含error_code时，会自动添加对应的💡修复提示到message中
    """
    r: dict = {"success": False, "message": message}
    if meta:
        r["meta"] = meta
    # 自动附加集中式错误提示（仅当message中还没有💡提示时）
    error_code = (meta or {}).get('error_code', '')
    hint = _ERROR_HINTS.get(error_code, '')
    if hint and '💡' not in message:
        r["message"] = message + f'\n💡 {hint}'
    return r


def _strip_defaults(obj: Any, depth: int = 0) -> Any:
    """递归移除默认值和空值以减少token消耗。

    ⚠️ LLM注意：以下字段若缺失表示其值为默认值（False/0），并非字段不存在：
    bold, italic, underline, wrap_text, border_top/bottom/left/right,
    horizontal_alignment, vertical_alignment, text_rotation, merge_cells

    移除规则：
    - None, 空字符串, 空列表, 空字典
    - 常见Excel默认值：False, 0, None（基于字段名判断）
    - 保留有意义的值
    """
    if depth > 5 or not isinstance(obj, dict):
        return obj
    
    # 常见Excel默认值字段名（这些字段通常不需要返回False/0/None）
    excel_default_fields = {
        'bold', 'italic', 'underline', 'wrap_text', 'auto_filter',
        'border_top', 'border_bottom', 'border_left', 'border_right',
        'horizontal_alignment', 'vertical_alignment', 'text_rotation',
        'indent', 'shrink_to_fit', 'merge_cells'
    }
    
    # 有语义的空列表字段名，即使为空也不应移除
    semantic_list_fields = {
        'headers', 'sheets', 'sheets_with_headers', 'field_names',
        'descriptions', 'data', 'columns', 'rows'
    }
    
    # 单元格语义字段：即使为None也必须保留（LLM需要区分「空单元格」和「字段缺失」）
    # 当dict同时包含coordinate时，说明这是CellInfo结构，value=None代表空单元格
    cell_semantic_fields = {'value'}
    _is_cell_info = 'coordinate' in obj
    
    cleaned = {}
    for k, v in obj.items():
        # 单元格value字段：保留None和空字符串（空单元格有语义意义，calamine返回''而openpyxl返回None）
        if k in cell_semantic_fields and _is_cell_info:
            cleaned[k] = v  # 保留原值（包括None和''），让JSON序列化为null
            continue
        
        # 移除空值（CellInfo的value已由上面处理）
        if v is None or v == '':
            continue
        
        # 移除空容器（但有语义的列表字段除外）
        if isinstance(v, (list, dict)) and len(v) == 0 and k not in semantic_list_fields:
            continue
        
        # 基于字段名的默认值处理
        if k.lower() in excel_default_fields and v in [False, 0, None]:
            continue
        
        # 递归处理嵌套对象
        if isinstance(v, dict):
            cleaned[k] = _strip_defaults(v, depth + 1)
        elif isinstance(v, list):
            cleaned[k] = [_strip_defaults(i, depth + 1) if isinstance(i, dict) else i for i in v]
        else:
            cleaned[k] = v
    
    return cleaned


def _ensure_dict(result) -> dict:
    """将OperationResult等dataclass转换为dict，已是dict则原样返回。"""
    if isinstance(result, dict):
        return result
    if hasattr(result, '__dataclass_fields__'):
        from dataclasses import asdict
        return asdict(result)
    return result


def _wrap(result, meta: dict = None) -> dict:
    """包装Operations层返回，metadata→meta，添加上下文meta，统一success字段。
    
    统一返回格式：{success, message, data, meta}
    - 成功时：若缺少message则自动补充默认message
    - metadata→meta：Operations层的metadata自动映射到meta
    - error→message：确保AI只需检查message键
    """
    result = _ensure_dict(result)
    if not isinstance(result, dict):
        return result
    # 统一error→message，确保AI只需检查message键
    err_val = result.get('error')
    if isinstance(err_val, str) and not result.get('message'):
        result['message'] = result.pop('error')
    if "success" not in result:
        result["success"] = True
    # 成功时若缺少message，自动补充默认message
    if result.get('success') is True and 'message' not in result:
        result['message'] = '操作成功'
    if "metadata" in result:
        m = result.pop("metadata")
        if isinstance(m, dict) and m:
            merged = {**m, **(meta or {})}
            result["meta"] = merged
            meta = None  # 已合并，不再重复设置
    if meta and "meta" not in result:
        result["meta"] = meta

    # Token优化：过滤默认值和空值，减少响应体积
    if result.get('success') is True and 'data' in result and isinstance(result['data'], dict):
        result['data'] = _strip_defaults(result['data'])

    # 对Operations层返回的错误也附加集中式提示（Operations层无error_code）
    if result.get('success') is False:
        msg = result.get('message', '')
        existing_code = (result.get('meta') or {}).get('error_code', '')
        if not existing_code and '💡' not in msg:
            # 根据消息内容推断error_code并附加提示
            inferred = _infer_error_code(msg)
            if inferred:
                hint = _ERROR_HINTS.get(inferred, '')
                if hint:
                    result['message'] = msg + f'\n💡 {hint}'
                    if 'meta' not in result:
                        result['meta'] = {}
                    result['meta']['error_code'] = inferred

    return result


def _infer_error_code(message: str) -> str:
    """根据错误消息内容推断error_code，用于Operations层错误提示"""
    msg = message
    if '文件不存在' in msg or 'No such file' in msg.lower():
        return 'FILE_NOT_FOUND'
    if '工作表' in msg and ('不存在' in msg or '未找到' in msg):
        return 'SHEET_NOT_FOUND'
    if '无法打开' in msg or 'cannot open' in msg.lower() or 'Permission denied' in msg:
        return 'FILE_OPEN_FAILED'
    if '工作表为空' in msg or '没有数据' in msg:
        return 'EMPTY_SHEET'
    if '路径' in msg and ('不合法' in msg or '不允许' in msg or '穿越' in msg):
        return 'PATH_VALIDATION_FAILED'
    return ''

# ==================== MCP 工具定义 ====================

@mcp.tool()
@_validate_file_path()
@_track_call
def excel_list_sheets(file_path: str) -> Dict[str, Any]:
    """列出Excel文件中的所有工作表名称。查询前先用此工具确认工作表存在。

    Args:
        file_path: Excel文件路径
        
    Returns:
        Dict[str, Any]: 包含工作表列表的字典，结构为 {"sheets": ["sheet1", "sheet2"], "success": bool}
    """
    return _wrap(ExcelOperations.list_sheets(file_path))


@mcp.tool()
@_validate_file_path()
@_track_call
def excel_search(
    file_path: str,
    pattern: str,
    sheet_name: Optional[str] = None,
    *,
    case_sensitive: bool = False,
    whole_word: bool = False,
    use_regex: Optional[bool] = None,
    cell_range: Optional[str] = None
) -> Dict[str, Any]:
    """在Excel中搜索匹配pattern的单元格。

    Args:
        file_path: Excel文件路径
        pattern: 搜索模式（文本或正则，use_regex=None时自动检测）
        sheet_name: 工作表名称，默认搜索全部
        case_sensitive: 区分大小写，默认False
        whole_word: 全词匹配，默认False
        use_regex: None=自动检测(含特殊字符时启用)，True/False=强制
        cell_range: 搜索范围（如 "A1:C10" 或 "Sheet1!A1:C10"，不含!时配合sheet_name使用）
    """
    # 统一范围格式：如果cell_range不含!但传了sheet_name，自动拼接
    _range_arg = cell_range
    if _range_arg and '!' not in _range_arg and sheet_name:
        _range_arg = f"{sheet_name}!{_range_arg}"

    if use_regex is None:
        use_regex = bool(re.match(r'.*[\[\](){}*+?|^$\\.]', pattern))
    return _wrap(ExcelOperations.search(
        file_path, pattern, sheet_name, case_sensitive, whole_word,
        use_regex, include_values=True, include_formulas=False, range=_range_arg))


@mcp.tool()
@_track_call
def excel_search_directory(
    directory_path: str,
    pattern: str,
    *,
    recursive: bool = True,
    file_extensions: Optional[List[str]] = None,
    max_files: int = MAX_SEARCH_FILES,
    case_sensitive: bool = False,
    whole_word: bool = False,
    use_regex: Optional[bool] = None
) -> Dict[str, Any]:
    """在目录下所有Excel文件中搜索内容。

    Args:
        directory_path: 搜索目录路径
        pattern: 搜索模式（文本或正则）
        recursive: 递归子目录，默认True
        file_extensions: 扩展名过滤，如 [".xlsx", ".xls"]
        max_files: 最大搜索文件数
        case_sensitive: 区分大小写，默认False
        whole_word: 全词匹配，默认False
        use_regex: None=自动检测，True/False=强制
    """
    _path_err = _validate_path(directory_path)
    if _path_err:
        return _path_err
    if use_regex is None:
        use_regex = bool(re.match(r'.*[\[\](){}*+?|^$\\.]', pattern))
    return _wrap(ExcelOperations.search_directory(
        directory_path, pattern, case_sensitive, whole_word,
        use_regex, True, False,
        recursive=recursive, file_extensions=file_extensions,
        file_pattern=None, max_files=max_files))


@mcp.tool()
@_validate_file_path()
@_track_call
def excel_get_range(
    file_path: str,
    cell_range: str,
    include_formatting: bool = False,
    sheet_name: Optional[str] = None
) -> Dict[str, Any]:
    """📍 **精确读取**：已知单元格坐标时使用（如A1:C10）。

    💡 **使用场景**：
    • 已知精确坐标 → excel_get_range（精确读取，如 "Sheet1!A1:C10"）
    • 需要单元格格式信息 → include_formatting=True

    ⚠️ **不推荐用于**：
    • 需要筛选/聚合/JOIN/排序 → 使用 excel_query（SQL引擎）
    • 快速了解表结构 → 使用 excel_describe_table（列名+类型+样本值）
    • 只需表头信息 → 使用 excel_get_headers

    返回二维CellInfo数组[[{coordinate,value},...],...]。
    每个单元格返回 {coordinate: "A1", value: ...}，空单元格value为null。
    支持include_formatting获取样式信息（额外返回font/fill等字段）。

    Args:
        file_path: Excel文件路径
        cell_range: 单元格范围（如 "A1:C10" 或 "Sheet1!A1:C10"，不含!时配合sheet_name使用）
        include_formatting: 是否包含格式信息，默认False
        sheet_name: 工作表名称（可选，cell_range不含!时自动拼接为 "sheet_name!cell_range"）
    """
    
    # 如果range不包含工作表名但指定了sheet_name，添加工作表名前缀
    if sheet_name and '!' not in cell_range:
        cell_range = f"{sheet_name}!{cell_range}"
    
    # 参数顺序问题检测：cell_range看起来像误传的sheet_name
    if cell_range and '!' not in cell_range and not cell_range.startswith(('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J')) and len(cell_range) <= 10:
        return _fail(
            f"参数顺序可能错误：cell_range参数值 '{cell_range}' 看起来不像有效的单元格范围。\n"
            f"请检查参数顺序：第二个参数应该是范围表达式（如 'A1:C10' 或 'Sheet1!A1:C10'），\n"
            f"如果要以默认工作表读取，请明确指定: excel_get_range('{file_path}', 'A1:C10')\n"
            f"💡 正确示例: excel_get_range('文件.xlsx', 'Sheet1!A1:E10') 或 excel_get_range('文件.xlsx', 'A1:E10', sheet_name='Sheet1')",
            meta={
                "error_code": "PARAMETER_ORDER_ERROR",
                "received_range": cell_range,
                "hint": "cell_range参数应该是单元格范围（如A1:E10），不是工作表名"
            }
        )
    
    # 原有的参数验证逻辑

    try:
        # 验证范围表达式格式
        # 如果没有工作表前缀，尝试自动推断（单工作表文件）
        if '!' not in cell_range:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                if len(sheet_names) == 1:
                    cell_range = f"{sheet_names[0]}!{cell_range}"
                else:
                    return _fail(
                        f"范围表达式缺少工作表名，且文件有多个工作表({', '.join(sheet_names)})。"
                        f"请使用格式: '工作表名!A1:C10'",
                        meta={"error_code": "VALIDATION_FAILED", "available_sheets": sheet_names}
                    )
            except Exception:
                return _fail(
                    f"范围表达式缺少工作表名。正确格式: 'Sheet1!A1:C10'，当前: '{cell_range}'",
                    meta={"error_code": "VALIDATION_FAILED"}
                )

        range_validation = ExcelValidator.validate_range_expression(cell_range)

        # 验证操作规模
        scale_validation = ExcelValidator.validate_operation_scale(range_validation['range_info'])

        # 记录验证成功到调试日志
        logger.debug(f"范围验证成功: {range_validation['normalized_range']}")

    except DataValidationError as e:
        # 记录验证失败
        logger.error(f"范围验证失败: {str(e)}")

        return _fail(f"范围表达式验证失败: {str(e)}", meta={"error_code": "VALIDATION_FAILED"})

    # 调用原始函数
    result = ExcelOperations.get_range(file_path, cell_range, include_formatting)
    result = _ensure_dict(result)

    # 如果成功，添加验证信息到结果中
    if result.get('success'):
        validation_info = {
            'normalized_range': range_validation['normalized_range'],
            'sheet_name': range_validation['sheet_name'],
            'range_type': range_validation['range_info']['type'],
            'scale_assessment': scale_validation
        }
        # 向后兼容：保留顶层validation_info
        result['validation_info'] = validation_info
        # 新格式：同时合并到meta中
        if 'meta' not in result:
            result['meta'] = {}
        result['meta']['validation_info'] = validation_info

    return _wrap(result)


@mcp.tool()
@_validate_file_path()
@_track_call
def excel_get_headers(
    file_path: str,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    max_columns: Optional[int] = None
) -> Dict[str, Any]:
    """📋 **表头信息**：轻量级获取列名（中文+英文）。

    💡 **使用场景**：
    • 只需表头信息 → excel_get_headers（更轻量）
    • 需要双行表头（中文+英文）→ 自动识别
    • 获取所有表的表头 → 不传sheet_name参数

    ⚠️ **不推荐用于**：
    • 需要数据类型/样本值 → 使用 excel_describe_table（完整分析）
    • 需要筛选/聚合 → 使用 excel_query（SQL引擎）

    提取工作表表头信息。支持双行表头（中文描述+英文字段名）。
    不传sheet_name获取所有表的表头。

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称，默认为None表示获取所有表的表头
        header_row: 表头行号，默认为1
        max_columns: 最大列数限制，默认为None
    """
    if sheet_name is None:
        return _wrap(ExcelOperations.get_all_headers(file_path, header_row, max_columns))
    return _wrap(ExcelOperations.get_headers(file_path, sheet_name, header_row, max_columns))


@mcp.tool()
@_validate_file_path()
@_track_call
def excel_update_range(
    file_path: str,
    cell_range: str,
    data: List[List[Any]],
    preserve_formulas: bool = True,
    insert_mode: bool = False,
    sheet_name: Optional[str] = None
) -> Dict[str, Any]:
    """📍 **精确坐标写入**：知道具体单元格范围时使用。

    💡 **使用场景**：
    • 🔴 精确覆盖指定区域（知道具体A1:C10）→ excel_update_range
    • 🟢 安全追加数据 → 先 find_last_row → 再 update_range(..., insert_mode=True)
    • 🔴 直接覆盖数据 → update_range(..., insert_mode=False)【默认，危险】

    ⚠️ **不推荐用于**：
    • 批量修改多行（改10行以上/按条件改）→ 使用 excel_update_query（SQL UPDATE）
    • 按ID改单行（知道 key_column + key_value）→ 使用 excel_upsert_row

    ⚠️ **重要**：默认为覆盖模式(insert_mode=False)，会直接替换目标区域数据！
       如需保留原有数据并插入新行，必须显式设置 insert_mode=True

    Args:
        file_path: Excel文件路径
        cell_range: 单元格范围（如 "A1:C10" 或 "Sheet1!A1:C10"，不含!时配合sheet_name使用）
        data: 要写入的数据，二维数组格式 [[row1], [row2], ...]
        preserve_formulas: 是否保留已有公式不被覆盖，默认True
        insert_mode: False=覆盖模式(默认，直接替换) | True=插入模式(原有数据下移)
        sheet_name: 工作表名称（可选，cell_range不含!时自动拼接为 "sheet_name!cell_range"）
    """
    
    # 如果range不包含工作表名但指定了sheet_name，添加工作表名前缀
    if sheet_name and '!' not in cell_range:
        cell_range = f"{sheet_name}!{cell_range}"

    try:
        # 验证范围表达式格式
        range_validation = ExcelValidator.validate_range_expression(cell_range)

        # 验证操作规模
        scale_validation = ExcelValidator.validate_operation_scale(range_validation['range_info'])

        # 如果有警告信息，记录到操作日志
        if scale_validation.get('warning'):
            logger.warning(f"操作规模警告: {scale_validation['warning']}")

    except DataValidationError as e:
        # 记录验证失败
        operation_logger.start_session(file_path)
        operation_logger.log_operation("validation_failed", {
            "operation": "update_range",
            "cell_range": cell_range,
            "error": str(e)
        })

        return _fail(f"参数验证失败: {str(e)}", meta={"error_code": "VALIDATION_FAILED"})

    # 开始操作会话
    operation_logger.start_session(file_path)

    # 记录操作日志
    operation_logger.log_operation("update_range", {
        "cell_range": cell_range,
        "validated_range": range_validation['normalized_range'],
        "data_rows": len(data),
        "insert_mode": insert_mode,
        "preserve_formulas": preserve_formulas,
        "scale_info": scale_validation
    })

    try:
        # 扩展流式路径：支持覆盖模式和插入模式
        use_streaming = True and not preserve_formulas
        # 插入模式会自动使用流式插入，覆盖模式使用流式覆盖
        result = ExcelOperations.update_range(file_path, cell_range, data, preserve_formulas, insert_mode, use_streaming)
        result = _ensure_dict(result)

        # 记录操作结果
        operation_logger.log_operation("operation_result", {
            "success": result.get('success', False),
            "updated_cells": result.get('updated_cells', 0),
            "message": result.get('message', '')
        })

        result = _wrap(result)
        # 🔴 覆盖模式安全警告：在返回值中明确标注操作模式，防止LLM误判
        if insert_mode is False and result.get('success'):
            _mode_note = "⚠️ [覆盖模式] 目标区域原有数据已被替换。如需保留原数据并插入新行，请设置 insert_mode=True"
            result['message'] = f"{result.get('message', '')} | {_mode_note}"
            if 'meta' not in result:
                result['meta'] = {}
            result['meta']['write_mode'] = 'overwrite'
            result['meta']['safe_alternative'] = 'Set insert_mode=True for insert mode'
        elif insert_mode is True and result.get('success'):
            if 'meta' not in result:
                result['meta'] = {}
            result['meta']['write_mode'] = 'insert'

        return result

    except Exception as e:
        # 记录错误
        operation_logger.log_operation("operation_error", {
            "error": str(e),
            "message": f"更新操作失败: {str(e)}"
        })

        return _fail(f"更新操作失败: {str(e)}", meta={"error_code": "OPERATION_FAILED"})



@mcp.tool()
@_track_call
def excel_create_backup(
    file_path: str,
    backup_dir: Optional[str] = None
) -> Dict[str, Any]:
    """为Excel文件创建备份。备份存放在同级backup目录。

    Args:
        file_path: Excel文件路径
        backup_dir: 备份目录路径，默认为None表示同级backup目录
    """
    if not os.path.exists(file_path):
        return _fail(f"源文件不存在: {file_path}", meta={"error_code": "FILE_NOT_FOUND"})

    try:
        # 创建备份目录
        if backup_dir is None:
            base_dir = os.path.dirname(file_path)
            backup_dir = os.path.join(base_dir, ".excel_mcp_backups")

        os.makedirs(backup_dir, exist_ok=True)

        # 生成备份文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.basename(file_path)
        name, ext = os.path.splitext(filename)
        backup_filename = f"{name}_backup_{timestamp}{ext}"
        backup_path = os.path.join(backup_dir, backup_filename)

        # 创建备份
        shutil.copy2(file_path, backup_path)

        # 检查备份大小
        original_size = os.path.getsize(file_path)
        backup_size = os.path.getsize(backup_path)

        return _ok(f"备份创建成功: {backup_filename}", data={'backup_file': backup_path, 'backup_directory': backup_dir, 'file_size': {'original': original_size, 'backup': backup_size}, 'timestamp': timestamp}, meta={"file_path": file_path})

    except Exception as e:
        return _fail(f"备份创建失败: {str(e)}", meta={"error_code": "BACKUP_FAILED"})


@mcp.tool()
@_track_call
def excel_restore_backup(
    backup_path: str,
    target_path: Optional[str] = None
) -> Dict[str, Any]:
    """从备份文件恢复Excel。target_path不传则覆盖原文件。

    Args:
        backup_path: 备份文件路径
        target_path: 目标文件路径，默认为None表示覆盖原文件
    """
    _path_err = _validate_path(backup_path)
    if _path_err:
        return _path_err

    if not os.path.exists(backup_path):
        return _fail(f"备份文件不存在: {backup_path}", meta={"error_code": "BACKUP_NOT_FOUND"})

    try:
        # 确定目标路径
        if target_path is None:
            # 尝试从备份文件名推断原始文件名
            filename = os.path.basename(backup_path)
            if "_backup_" in filename:
                # 移除备份时间戳
                parts = filename.split("_backup_")
                target_path = parts[0] + os.path.splitext(backup_path)[1]
            else:
                target_path = filename.replace("_backup_", ".")

        # 创建目标目录
        target_dir = os.path.dirname(target_path)
        if target_dir:
            os.makedirs(target_dir, exist_ok=True)

        # 检查目标文件是否存在
        target_exists = os.path.exists(target_path)

        # 执行恢复
        shutil.copy2(backup_path, target_path)

        return _ok(f"文件恢复成功: {os.path.basename(target_path)}", data={'backup_file': backup_path, 'target_file': target_path, 'target_existed': target_exists}, meta={"file_path": backup_path})

    except Exception as e:
        return _fail(f"恢复失败: {str(e)}", meta={"error_code": "RESTORE_FAILED"})


@mcp.tool()
@_track_call
def excel_list_backups(
    file_path: str,
    backup_dir: Optional[str] = None
) -> Dict[str, Any]:
    """列出文件的所有备份版本及时间。

    Args:
        file_path: Excel文件路径
        backup_dir: 备份目录路径，默认为None
    """
    try:
        # 确定备份目录
        if backup_dir is None:
            base_dir = os.path.dirname(file_path)
            backup_dir = os.path.join(base_dir, ".excel_mcp_backups")

        if not os.path.exists(backup_dir):
            return _ok("备份目录不存在", data={'backups': []}, meta={"file_path": file_path})

        # 获取文件名
        filename = os.path.basename(file_path)
        name, ext = os.path.splitext(filename)
        backup_pattern = f"{name}_backup_*{ext}"

        # 查找备份文件
        backup_files = []
        for file in os.listdir(backup_dir):
            if file.startswith(f"{name}_backup_") and file.endswith(ext):
                full_path = os.path.join(backup_dir, file)
                stat = os.stat(full_path)
                backup_files.append({
                    'filename': file,
                    'path': full_path,
                    'size': stat.st_size,
                    'created_time': datetime.fromtimestamp(stat.st_ctime),
                    'modified_time': datetime.fromtimestamp(stat.st_mtime)
                })

        # 按时间排序
        backup_files.sort(key=lambda x: x['created_time'], reverse=True)

        return _ok(f"找到 {len(backup_files)} 个备份", data={'backups': backup_files, 'backup_directory': backup_dir, 'total_backups': len(backup_files)}, meta={"file_path": file_path})

    except Exception as e:
        return _fail(f"列出备份失败: {str(e)}", meta={"error_code": "LIST_BACKUPS_FAILED"})


@mcp.tool()
@_track_call
def excel_insert_rows(
    file_path: str,
    sheet_name: str,
    row_index: int,
    count: int = 1
) -> Dict[str, Any]:
    """在指定位置插入空行。row_index从1开始（第1行前插入传1）。

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        row_index: 插入位置的行索引（从1开始，在该行上方插入）
        count: 插入的行数，默认为1    """
    return _wrap(ExcelOperations.insert_rows(file_path, sheet_name, row_index, count, True))


@mcp.tool()
@_track_call
def excel_insert_columns(
    file_path: str,
    sheet_name: str,
    column_index: int,
    count: int = 1
) -> Dict[str, Any]:
    """在指定位置插入空列。column_index从1开始。

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        column_index: 插入位置的列索引（从1开始）
        count: 插入的列数，默认为1    """
    return _wrap(ExcelOperations.insert_columns(file_path, sheet_name, column_index, count, True))


@mcp.tool()
@_track_call
def excel_find_last_row(
    file_path: str,
    sheet_name: str,
    column: Optional[Union[str, int]] = None
) -> Dict[str, Any]:
    """查找工作表最后一行。可指定列来找该列最后一个有值的行。追加数据前必用。

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        column: 列名或列索引，默认为None
    """
    return _wrap(ExcelOperations.find_last_row(file_path, sheet_name, column))


@mcp.tool()
@_validate_file_path()
@_track_call
def excel_create_file(
    file_path: str,
    sheet_names: Optional[List[str]] = None
) -> Dict[str, Any]:
    """创建新Excel文件。可指定初始工作表名称列表。

    Args:
        file_path: Excel文件路径
        sheet_names: 初始工作表名称列表，默认为None
    """
    return _wrap(ExcelOperations.create_file(file_path, sheet_names))


@mcp.tool()
@_validate_file_path(['file_path', 'output_path'])
@_track_call
def excel_export_to_csv(
    file_path: str,
    output_path: str,
    sheet_name: Optional[str] = None,
    encoding: str = "utf-8"
) -> Dict[str, Any]:
    """将工作表导出为CSV文件。

    Args:
        file_path: Excel文件路径
        output_path: CSV输出路径
        sheet_name: 工作表名称，默认为None
        encoding: 编码格式，默认为"utf-8"
    """
    return _wrap(ExcelOperations.export_to_csv(file_path, output_path, sheet_name, encoding))


@mcp.tool()
@_validate_file_path(['csv_path', 'output_path'])
@_track_call
def excel_import_from_csv(
    csv_path: str,
    output_path: str,
    sheet_name: str = "Sheet1",
    encoding: str = "utf-8",
    has_header: bool = True
) -> Dict[str, Any]:
    """从CSV文件创建Excel工作表。

    Args:
        csv_path: CSV文件路径
        output_path: Excel输出路径
        sheet_name: 工作表名称，默认为"Sheet1"
        encoding: 编码格式，默认为"utf-8"
        has_header: CSV是否有表头，默认为True
    """
    for _p in [csv_path, output_path]:
        _err = _validate_path(_p)
        if _err:
            return _err

    return _wrap(ExcelOperations.import_from_csv(csv_path, output_path, sheet_name, encoding, has_header))





@mcp.tool()
@_track_call
def excel_create_sheet(
    file_path: str,
    sheet_name: str,
    index: Optional[int] = None
) -> Dict[str, Any]:
    """创建新工作表。可指定插入位置index（从0开始，0=最前面）。

    Args:
        file_path: Excel文件路径
        sheet_name: 新工作表名称
        index: 插入位置索引（从0开始，默认为None表示追加到末尾）
    """
    return _wrap(ExcelOperations.create_sheet(file_path, sheet_name, index))


@mcp.tool()
@_validate_file_path()
@_track_call
def excel_create_chart(
    file_path: str,
    sheet_name: str,
    chart_type: str,
    data_range: str,
    title: str = "",
    chart_name: str = "",
    position: str = "B15"
) -> Dict[str, Any]:
    """在工作表中创建图表。chart_type: line/bar/column/pie/scatter/area等。支持'column'作为'bar'的别名。

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        chart_type: 图表类型
        data_range: 数据范围
        title: 图表标题，默认为空字符串
        chart_name: 图表名称，默认为空字符串
        position: 图表位置，默认为"B15"
    """
    return _wrap(ExcelOperations.create_chart(
        file_path, sheet_name, chart_type, data_range,
        title=title, chart_name=chart_name, position=position
    ))


@mcp.tool()
@_track_call
def excel_delete_sheet(
    file_path: str,
    sheet_name: str
) -> Dict[str, Any]:
    """删除指定工作表。

    Args:
        file_path: Excel文件路径
        sheet_name: 要删除的工作表名称
    """
    # 开始操作会话
    operation_logger.start_session(file_path)

    # 记录删除工作表操作日志
    operation_logger.log_operation("delete_sheet", {
        "sheet_name": sheet_name
    })

    try:
        result = ExcelOperations.delete_sheet(file_path, sheet_name)
        result = _ensure_dict(result)
        operation_logger.log_operation("operation_result", {
            "success": result.get('success', False),
            "deleted_sheet": result.get('deleted_sheet', ''),
            "remaining_sheets": result.get('remaining_sheets', 0),
            "message": result.get('message', '')
        })

        return _wrap(result)

    except Exception as e:
        # 记录错误
        operation_logger.log_operation("operation_error", {
            "error": str(e),
            "message": f"删除工作表操作失败: {str(e)}"
        })

        return _fail(f"删除工作表操作失败: {str(e)}", meta={"error_code": "DELETE_SHEET_FAILED"})


@mcp.tool()
@_track_call
def excel_rename_sheet(
    file_path: str,
    old_name: str,
    new_name: str
) -> Dict[str, Any]:
    """重命名工作表。

    Args:
        file_path: Excel文件路径
        old_name: 原工作表名称
        new_name: 新工作表名称
    """
    return _wrap(ExcelOperations.rename_sheet(file_path, old_name, new_name))


@mcp.tool()
@_track_call
def excel_copy_sheet(
    file_path: str,
    source_name: str,
    new_name: Optional[str] = None,
    index: Optional[int] = None
) -> Dict[str, Any]:
    """复制工作表（含数据和格式）。新工作表在同一文件内创建。

    Args:
        file_path: Excel文件路径
        source_name: 源工作表名称
        new_name: 新工作表名称，默认为None（自动命名为"源名_副本"）
        index: 插入位置索引（从0开始），默认为None（追加到末尾）    """
    return _wrap(ExcelOperations.copy_sheet(file_path, source_name, new_name, index, True))


@mcp.tool()
@_track_call
def excel_rename_column(
    file_path: str,
    sheet_name: str,
    old_header: str,
    new_header: str,
    header_row: int = 1
) -> Dict[str, Any]:
    """修改表头（列名）。只改header_row指定的行。

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        old_header: 原列名
        new_header: 新列名
        header_row: 表头行号，默认为1
    """
    return _wrap(ExcelOperations.rename_column(file_path, sheet_name, old_header, new_header, header_row))


@mcp.tool()
@_track_call
def excel_upsert_row(
    file_path: str,
    sheet_name: str,
    key_column: str,
    key_value: Any,
    updates: Dict[str, Any],
    header_row: int = 1
) -> Dict[str, Any]:
    """👍 **按ID更新单行首选**：知道key_column+key_value时使用。

    💡 **使用场景**：
    • 按ID改单行（知道 key_column + key_value）→ excel_upsert_row
    • 只改2-3个字段、dict传参方便 → 更安全
    • 需要幂等操作（行不存在自动插入）

    ⚠️ **不推荐用于**：
    • 批量修改多行 → 使用 excel_update_query（SQL UPDATE）
    • 精确覆盖指定区域 → 使用 excel_update_range

    ✅ **优点**：不会误改其他行、行不存在自动插入、参数自文档化、双行表头兼容

    按key_column+key_value查找行，存在则更新，不存在则插入。

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        key_column: 键列名
        key_value: 键值
        updates: 要更新的字段字典（如 {"伤害": 200, "冷却": 5}）
        header_row: 表头行号，默认为1
    """
    return _wrap(ExcelOperations.upsert_row(file_path, sheet_name, key_column, key_value, updates, header_row, True))


@mcp.tool()
@_track_call
def excel_delete_rows(
    file_path: str,
    sheet_name: str,
    row_index: int,
    count: int = 1
) -> Dict[str, Any]:
    """按行号删除行。row_index从1开始（第1行前删除传1）。

    ⚠️ 按条件删除请用 excel_delete_query（SQL DELETE FROM ... WHERE ...）

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        row_index: 要删除的起始行索引（从1开始，在该行位置删除）
        count: 删除的行数，默认为1
    """

    # 开始操作会话
    operation_logger.start_session(file_path)

    # 记录删除操作日志
    operation_logger.log_operation("delete_rows", {
        "sheet_name": sheet_name,
        "row_index": row_index,
        "count": count
    })

    try:
        result = _ensure_dict(ExcelOperations.delete_rows(file_path, sheet_name, row_index, count, True))

        # 记录操作结果
        operation_logger.log_operation("operation_result", {
            "success": result.get('success', False),
            "deleted_rows": result.get('deleted_rows', 0),
            "message": result.get('message', '')
        })

        return _wrap(result)

    except Exception as e:
        # 记录错误
        operation_logger.log_operation("operation_error", {
            "error": str(e),
            "message": f"删除行操作失败: {str(e)}"
        })

        return _fail(f"删除行操作失败: {str(e)}", meta={"error_code": "DELETE_ROWS_FAILED"})


@mcp.tool()
@_track_call
def excel_delete_columns(
    file_path: str,
    sheet_name: str,
    column_index: int,
    count: int = 1
) -> Dict[str, Any]:
    """删除指定位置开始的列。column_index从1开始。

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        column_index: 起始列索引（从1开始）
        count: 删除的列数，默认为1    """
    # 开始操作会话
    operation_logger.start_session(file_path)

    # 记录删除列操作日志
    operation_logger.log_operation("delete_columns", {
        "sheet_name": sheet_name,
        "column_index": column_index,
        "count": count
    })

    try:
        result = _ensure_dict(ExcelOperations.delete_columns(file_path, sheet_name, column_index, count, True))

        # 记录操作结果
        operation_logger.log_operation("operation_result", {
            "success": result.get('success', False),
            "deleted_columns": result.get('deleted_columns', 0),
            "message": result.get('message', '')
        })

        return _wrap(result)

    except Exception as e:
        # 记录错误
        operation_logger.log_operation("operation_error", {
            "error": str(e),
            "message": f"删除列操作失败: {str(e)}"
        })

        return _fail(f"删除列操作失败: {str(e)}", meta={"error_code": "DELETE_COLUMNS_FAILED"})

@mcp.tool()
@_track_call
def excel_set_formula(
    file_path: str,
    sheet_name: str,
    cell_address: str,
    formula: str
) -> Dict[str, Any]:
    """在单元格写入Excel公式。

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        cell_address: 单元格地址，如"A1"
        formula: Excel公式，以等号开头
    """
    # 参数验证：formula 不能为空
    if not formula or not formula.strip():
        return _fail(
            '📝 公式参数缺失：excel_set_formula() 需要一个有效的Excel公式。\n'
            '✅ 正确用法：excel_set_formula("file.xlsx", "Sheet1", "A1", "=SUM(B1:C1)")\n'
            '🔧 支持的公式类型：\n'
            '  - 数学公式: "=A1+B1", "=SUM(A1:A10)"\n'
            '  - 引用公式: "=B2", "=Sheet2!A1"\n'
            '  - 函数公式: "=VLOOKUP(A1, Sheet2!A:B, 2, FALSE)"\n'
            '❌ 不要只传单元格地址，公式必须以等号 "=" 开头',
            meta={
                "error_code": "MISSING_FORMULA",
                "received_formula": formula,
                "expected_format": "Excel公式（以=开头）",
                "example": "=SUM(A1:C1)"
            }
        )

    _formula_err = SecurityValidator.validate_formula(formula)
    if not _formula_err['valid']:
        return _fail(f'🔒 安全验证失败: {_formula_err["error"]}', meta={"error_code": "FORMULA_SECURITY_FAILED"})
    return _wrap(ExcelOperations.set_formula(file_path, sheet_name, cell_address, formula))


@mcp.tool()
@_track_call
def excel_query(
    file_path: str,
    query_expression: str,
    include_headers: bool = True,
    output_format: str = "table"
) -> Dict[str, Any]:
    """🔥 **首选工具**：所有数据查询/分析任务，优先使用此工具。

    💡 **使用场景**：
    • 需要筛选/聚合/JOIN/排序 → excel_query（SQL引擎，批量分析首选）
    • 复杂条件筛选 → WHERE, LIKE, IN, BETWEEN, 子查询
    • 聚合统计 → COUNT, SUM, AVG, MAX, MIN, GROUP BY, HAVING
    • 多表关联 → 5种JOIN类型，支持跨文件查询
    • 窗口函数 → ROW_NUMBER, RANK, DENSE_RANK
    • 字符串函数 → UPPER, LOWER, TRIM, LENGTH, CONCAT, REPLACE, SUBSTRING

    ⚠️ **不推荐用于**：
    • 已知精确坐标（如A1:C10）→ 使用 excel_get_range
    • 只需表头信息 → 使用 excel_get_headers 或 excel_describe_table

    query_expression: SELECT * FROM 技能表 WHERE 伤害>100 | GROUP BY | JOIN ON

    Args:
        file_path: Excel文件路径
        query_expression: SQL查询语句
        include_headers: 是否包含表头，默认为True
        output_format: 输出格式，默认为"table"
    """
    # 参数验证
    if not file_path or not file_path.strip():
        return _fail('文件路径不能为空', meta={"error_code": "MISSING_FILE_PATH"})

    if not query_expression or not query_expression.strip():
        return _fail('SQL查询语句不能为空', meta={"error_code": "MISSING_QUERY"})

    # 验证 output_format
    valid_formats = ('table', 'json', 'csv')
    if output_format and output_format not in valid_formats:
        return _fail(f'不支持的输出格式: {output_format}。可选: {", ".join(valid_formats)}', meta={"error_code": "INVALID_FORMAT"})

    # 使用高级SQL查询引擎
    try:
        from .api.advanced_sql_query import execute_advanced_sql_query
        result = _wrap(execute_advanced_sql_query(
            file_path=file_path,
            sql=query_expression,
            sheet_name=None,  # 统一使用SQL FROM子句中的表名
            limit=None,  # 统一使用SQL中的LIMIT
            include_headers=include_headers,
            output_format=output_format or 'table'
        ))
        # 确保meta字段存在（query_info保留在顶层向后兼容）
        if 'meta' not in result:
            result['meta'] = {'file_path': file_path}
        return result

    except ImportError:
        return _fail('SQLGlot未安装，无法使用高级SQL功能。请运行: pip install sqlglot\n\n💡 智能降级建议：\n• 对于简单数据读取：尝试使用 excel_get_range("文件路径", "工作表名!A1:Z100")\n• 对于文本搜索：尝试使用 excel_search("文件路径", "关键词", "工作表名")\n• 对于表头信息：尝试使用 excel_get_headers("文件路径", "工作表名")', meta={"error_code": "DEPENDENCY_MISSING"})
    except Exception as e:
        # SQL引擎已处理大部分错误并返回结构化响应，此处仅捕获未预期的异常
        return _fail(f'SQL查询失败: {str(e)}', meta={"error_code": "SQL_EXECUTION_FAILED"})


@mcp.tool()
@_track_call
def excel_update_query(
    file_path: str,
    update_expression: str,
    dry_run: bool = False
) -> Dict[str, Any]:
    """🔥 **批量修改首选**：按条件批量修改多行数据。

    💡 **使用场景**：
    • 批量修改多行（改10行以上/按条件改）→ excel_update_query（SQL UPDATE）
    • 需要计算表达式 → SET 血量=血量*2
    • 需要预览变更 → dry_run=True
    • 条件复杂 → WHERE 等级>5 AND 稀有度='传说'

    ⚠️ **不推荐用于**：
    • 🔴 精确覆盖指定区域（知道具体A1:C10）→ 使用 excel_update_range
    • 按ID改单行（只改2-3个字段）→ 使用 excel_upsert_row（更安全）

    SQL批量修改。dry_run=True预览变更不实际写入。

    示例::
        UPDATE 技能表 SET 伤害=200 WHERE 等级>=5
        UPDATE LootList SET PropType='主武器' WHERE _ROW_NUMBER_ IN (11,21,36)
        UPDATE 数据表 SET 状态='已处理' WHERE _ROW_NUMBER_ BETWEEN 10 AND 50

    行号支持(_ROW_NUMBER_):
        在UPDATE的WHERE条件中可使用 _ROW_NUMBER_ 虚拟列，基于Excel数据行号(不含表头)精确定位行。
        适用于：有重复记录无法用字段值唯一确定行的场景。
        注意：_ROW_NUMBER_ 仅在UPDATE中可用，SELECT查询暂不支持。
        不允许对 _ROW_NUMBER_ 本身执行SET操作。

    Args:
        file_path: Excel文件路径
        update_expression: UPDATE语句
        dry_run: 是否仅预览不实际写入，默认为False
    """

    if not file_path or not file_path.strip():
        return _fail('文件路径不能为空', meta={"error_code": "MISSING_FILE_PATH"})

    if not update_expression or not update_expression.strip():
        return _fail('UPDATE语句不能为空', meta={"error_code": "MISSING_QUERY"})

    # 安全检查：只允许UPDATE语句
    stripped = update_expression.strip().upper()
    if not stripped.startswith('UPDATE'):
        return _fail('只支持UPDATE语句。查询请使用 excel_query', meta={"error_code": "UNSUPPORTED_SQL"})

    try:
        from .api.advanced_sql_query import execute_advanced_update_query
        return _wrap(execute_advanced_update_query(
            file_path=file_path,
            sql=update_expression,
            dry_run=dry_run
        ))
    except ImportError:
        return _fail('SQLGlot未安装，无法使用UPDATE功能', meta={"error_code": "DEPENDENCY_MISSING"})
    except Exception as e:
        return _fail(f'UPDATE执行失败: {str(e)}', meta={"error_code": "UPDATE_EXECUTION_FAILED"})


@mcp.tool()
@_track_call
def excel_insert_query(
    file_path: str,
    insert_expression: str,
    dry_run: bool = False
) -> Dict[str, Any]:
    """SQL插入数据。支持单行/多行INSERT。

    示例::
        INSERT INTO 技能表 (技能名称, 伤害, 冷却) VALUES ('火球术', 300, 6)
        INSERT INTO Raids (RID, CID, Score) VALUES (6, 105, 7000), (7, 106, 8000)

    Args:
        file_path: Excel文件路径
        insert_expression: INSERT语句
        dry_run: 是否仅预览不实际写入，默认为False
    """

    if not file_path or not file_path.strip():
        return _fail('文件路径不能为空', meta={"error_code": "MISSING_FILE_PATH"})

    if not insert_expression or not insert_expression.strip():
        return _fail('INSERT语句不能为空', meta={"error_code": "MISSING_QUERY"})

    stripped = insert_expression.strip().upper()
    if not stripped.startswith('INSERT'):
        return _fail('只支持INSERT语句。查询请使用 excel_query', meta={"error_code": "UNSUPPORTED_SQL"})

    try:
        from .api.advanced_sql_query import execute_advanced_insert_query
        return _wrap(execute_advanced_insert_query(
            file_path=file_path,
            sql=insert_expression,
            dry_run=dry_run
        ))
    except ImportError:
        return _fail('SQLGlot未安装，无法使用INSERT功能', meta={"error_code": "DEPENDENCY_MISSING"})
    except Exception as e:
        return _fail(f'INSERT执行失败: {str(e)}', meta={"error_code": "INSERT_EXECUTION_FAILED"})


@mcp.tool()
@_track_call
def excel_delete_query(
    file_path: str,
    delete_expression: str,
    dry_run: bool = False
) -> Dict[str, Any]:
    """SQL删除数据。必须指定WHERE条件。

    示例::
        DELETE FROM Raids WHERE Score < 8000
        DELETE FROM Raids WHERE _ROW_NUMBER_ IN (3, 5, 7)

    Args:
        file_path: Excel文件路径
        delete_expression: DELETE语句
        dry_run: 是否仅预览不实际删除，默认为False
    """

    if not file_path or not file_path.strip():
        return _fail('文件路径不能为空', meta={"error_code": "MISSING_FILE_PATH"})

    if not delete_expression or not delete_expression.strip():
        return _fail('DELETE语句不能为空', meta={"error_code": "MISSING_QUERY"})

    stripped = delete_expression.strip().upper()
    if not stripped.startswith('DELETE'):
        return _fail('只支持DELETE语句。查询请使用 excel_query', meta={"error_code": "UNSUPPORTED_SQL"})

    try:
        from .api.advanced_sql_query import execute_advanced_delete_query
        return _wrap(execute_advanced_delete_query(
            file_path=file_path,
            sql=delete_expression,
            dry_run=dry_run
        ))
    except ImportError:
        return _fail('SQLGlot未安装，无法使用DELETE功能', meta={"error_code": "DEPENDENCY_MISSING"})
    except Exception as e:
        return _fail(f'DELETE执行失败: {str(e)}', meta={"error_code": "DELETE_EXECUTION_FAILED"})



def _collect_column_statistics(ws, data_start, num_cols, col_name_list):
    """
    收集列统计信息 - 优化版本，单次遍历
    
    Args:
        ws: 工作表对象
        data_start: 数据开始行
        num_cols: 列数
        col_name_list: 列名列表
        
    Returns:
        dict: 列统计信息
    """
    col_stats = {}
    for col_idx in range(num_cols):
        col_name = col_name_list[col_idx]
        col_stats[col_name] = {'non_null': 0, 'samples': [], 'type_values': []}
    
    # 单次遍历所有行和列，同时统计行数和列数据
    total_rows = 0
    for row in ws.iter_rows(min_row=data_start + 1, values_only=True):
        total_rows += 1
        for col_idx in range(min(len(row), num_cols)):
            val = row[col_idx]
            if val is not None:
                s = col_stats[col_name_list[col_idx]]
                s['non_null'] += 1
                if len(s['samples']) < 3:
                    s['samples'].append(val)
                if len(s['type_values']) < 100:
                    s['type_values'].append(val)
    
    return col_stats, total_rows

def _build_describe_columns(col_stats, col_name_list, is_dual_header, descriptions):
    """
    分析列数据类型并构建最终列信息列表（保留原始完整行为）
    
    Args:
        col_stats: 列统计信息
        col_name_list: 列名列表
        is_dual_header: 是否双行表头
        descriptions: 双行表头描述列表
        
    Returns:
        list: 列信息列表
    """
    # 推断类型并构建最终结果
    for col_idx, col_name in enumerate(col_name_list):
        s = col_stats[col_name]
        tv = s['type_values']
        if not tv:
            col_type = "empty"
        elif all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in tv):
            col_type = "number"
        elif all(isinstance(v, str) for v in tv):
            col_type = "text"
        else:
            col_type = "mixed"

        description = str(descriptions[col_idx]).strip() if is_dual_header and descriptions and col_idx < len(descriptions) and descriptions[col_idx] else None
        col_stats[col_name] = {
            'name': col_name, 'type': col_type, 'description': description,
            'non_null': s['non_null'], 'sample_values': s['samples']
        }

    return list(col_stats.values())

def _resolve_row_count(ws, data_start, total_rows):
    """
    统计行数 — 优先使用max_row，为None则使用total_rows（已统计避免重复计算）
    
    Args:
        ws: 工作表对象
        data_start: 数据开始行
        total_rows: 已统计的总行数
        
    Returns:
        int: 行数
    """
    try:
        # 修复：streaming写入后max_row可能为None的问题
        if hasattr(ws, 'max_row') and ws.max_row is not None and ws.max_row > data_start:
            row_count = ws.max_row - data_start
        else:
            # streaming写入后max_row可能为None，直接使用已统计的total_rows
            # total_rows已经在上面的循环中统计完成，避免重复计算
            row_count = total_rows if total_rows > 0 else 0
            # 安全检查：如果total_rows为0，回退到iter_rows统计
            if row_count == 0:
                try:
                    for idx, row in enumerate(ws.iter_rows(min_row=data_start + 1, values_only=True), start=1):
                        # 只要行中有一个非空单元格，就计数为有效数据行
                        if row and any(cell is not None for cell in row):
                            row_count += 1
                except Exception as e:
                    logger.warning(f"iter_rows统计行数失败: {e}，使用默认值0")
                    row_count = 0
    except Exception as e:
        # 极端情况处理：如果max_row访问失败，使用已统计的total_rows
        logger.warning(f"访问ws.max_row失败: {e}，使用已统计的total_rows")
        row_count = total_rows if total_rows > 0 else 0
        # 确保total_rows无效时，使用iter_rows重新统计
        if row_count == 0:
            try:
                for idx, row in enumerate(ws.iter_rows(min_row=data_start + 1, values_only=True), start=1):
                    if row and any(cell is not None for cell in row):
                        row_count += 1
            except Exception as e2:
                logger.warning(f"iter_rows回退统计也失败: {e2}")
                row_count = 0
    return row_count

def _prepare_describe_result(sheet_name, is_dual_header, columns, row_count, file_path, sheet):
    """
    准备describe结果
    
    Args:
        sheet_name: 工作表名称
        is_dual_header: 是否双行表头
        columns: 列信息
        row_count: 行数
        file_path: 文件路径
        sheet: 工作表对象
        
    Returns:
        dict: 返回结果
    """
    return _ok(f"表 '{sheet_name}': {len(columns)}列, {row_count}行数据, {'双行表头' if is_dual_header else '单行表头'}", data={
        'sheet_name': sheet_name,
        'header_type': 'dual' if is_dual_header else 'single',
        'row_count': row_count,
        'column_count': len(columns),
        'columns': columns
    }, meta={"file_path": file_path, "sheet": sheet_name})


def _detect_dual_header(rows: list) -> tuple:
    """检测双行表头模式（第1行中文描述 + 第2行英文字段名）。
    
    Args:
        rows: 工作表前几行数据（values_only=True的列表）
    
    Returns:
        tuple: (is_dual_header, header_row_idx, descriptions)
    """
    if len(rows) < 2:
        return False, 0, []
    
    first_row = [_cell_str(c) for c in rows[0]]
    second_row = [_cell_str(c) for c in rows[1]]
    
    # 第二行是否全是有效英文字段名
    second_row_strs = [v for v in second_row if v is not None]
    all_valid_names = (
        len(second_row_strs) >= 2
        and all(isinstance(v, str) and v.strip() and v.strip()[0].isalpha() and v.strip()[0].isascii()
               for v in second_row_strs)
    )
    
    # 第一行是否有中文
    first_row_strs = [v for v in first_row if v is not None]
    any_chinese = any(
        isinstance(v, str) and any('\u4e00' <= ch <= '\u9fff' for ch in v)
        for v in first_row_strs
    )
    
    is_dual_header = all_valid_names and any_chinese
    header_row_idx = 1 if is_dual_header else 0  # 双行时数据从第3行开始(idx=2)，表头用第2行(idx=1)
    descriptions = first_row if is_dual_header else []
    
    return is_dual_header, header_row_idx, descriptions


def _cell_str(c):
    """统一单元格值转字符串"""
    if c is None:
        return None
    if hasattr(c, 'value'):
        return c.value
    return str(c)


@mcp.tool()
@_validate_file_path()
@_track_call
def excel_describe_table(
    file_path: str,
    sheet_name: str = None
) -> Dict[str, Any]:
    """📊 **表结构分析**：快速了解表的列名、类型、样本值。

    💡 **使用场景**：
    • 快速了解表结构 → excel_describe_table（列名+类型+样本值+行数）
    • 数据操作前了解表结构 → 避免列名错误
    • 需要数据类型和非空统计 → 自动推断类型、非空比例

    ⚠️ **不推荐用于**：
    • 只需表头信息 → 使用 excel_get_headers（更轻量）
    • 已知精确坐标读取数据 → 使用 excel_get_range
    • 需要筛选/聚合 → 使用 excel_query（SQL引擎）

    分析Excel工作表(sheet)结构：列名、数据类型、非空统计、样本数据。
    任何数据操作前应先调用此工具了解表结构，避免列名错误。
    自动识别双行表头（中文描述+英文字段名）。

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称，默认为None表示第一个工作表

    Returns:
        data.sheet_name: 工作表名称
        data.header_type: 表头类型 ("dual"=双行 / "single"=单行)
        data.row_count: 数据行数
        data.column_count: 列数
        data.columns: 列信息列表 [{name, type, description, non_null, sample_values}]
    """
    # 文件验证和加载
    if not file_path or not file_path.strip():
        return _fail('文件路径不能为空', meta={"error_code": "MISSING_FILE_PATH"})

    try:
        wb, ws = ExcelValidator.get_workbook_and_sheet(file_path, sheet_name, read_only=True, data_only=True)
        if not sheet_name:
            sheet_name = ws.title
    except DataValidationError as e:
        return _fail(e.message, meta={"error_code": "SHEET_NOT_FOUND", "hint": e.hint, "suggested_fix": e.suggested_fix})
    except Exception as e:
        return _fail(f'无法打开文件: {e}', meta={"error_code": "FILE_OPEN_FAILED"})

    try:

        # 使用 HeaderAnalyzer 统一检测双表头
        from src.excel_mcp_server_fastmcp.api.header_analyzer import HeaderAnalyzer
        HeaderAnalyzer.invalidate(file_path)  # 确保读取最新数据（read_only 模式下可能需要）
        info = HeaderAnalyzer.analyze(file_path, sheet_name)
        
        is_dual_header = info.is_dual
        header_row_idx = 1 if is_dual_header else 0  # 0-based index into rows list
        descriptions = info.descriptions
        
        # 从 openpyxl 重新读取数据（HeaderAnalyzer 用 calamine，这里需要 openpyxl 取类型信息）
        rows = list(ws.iter_rows(max_row=4 if info.is_dual else 3, values_only=True))
        if not rows:
            return _fail('工作表为空', meta={"error_code": "EMPTY_SHEET"})
        
        headers = rows[header_row_idx] if header_row_idx < len(rows) else rows[0]
        data_start = info.data_start_row

        # 准备列名列表
        num_cols = len(headers)
        col_name_list = []
        for col_idx in range(num_cols):
            col_name = headers[col_idx]
            if col_name is None:
                col_name = f"column_{col_idx + 1}"
            col_name = str(col_name).strip()
            if not col_name:
                col_name = f"column_{col_idx + 1}"
            col_name_list.append(col_name)

        # 单次遍历收集统计信息（提取为独立函数）
        col_stats, total_rows = _collect_column_statistics(ws, data_start, num_cols, col_name_list)

        # 推断类型并构建最终结果（提取为独立函数）
        columns = _build_describe_columns(col_stats, col_name_list, is_dual_header, descriptions)

        # 统计行数 — 优先使用max_row，为None则使用total_rows
        row_count = _resolve_row_count(ws, data_start, total_rows)

        return _prepare_describe_result(sheet_name, is_dual_header, columns, row_count, file_path, ws)
    except Exception as e:
        return _fail(f'查看表结构失败: {e}', meta={"error_code": "DESCRIBE_FAILED"})
    finally:
        wb.close()


def _check_merge_data_loss(file_path: str, sheet_name: str, cell_range: str) -> Optional[str]:
    """检查合并操作是否会导致数据丢失。返回警告信息或None。
    
    Excel合并单元格时，只有左上角单元格保留数据，其余单元格的值会丢失。
    此函数在合并前检测目标范围内是否有非空的非左上角单元格，如有则返回警告。
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        
        # 解析范围
        if '!' in cell_range:
            range_part = cell_range.split('!')[1]
        else:
            range_part = cell_range
        
        # 获取范围内所有非空单元格（排除第一个/左上角）
        non_empty_cells = []
        first_cell = True
        for row in ws[range_part]:
            for cell in row:
                if first_cell:
                    first_cell = False  # 跳过左上角
                    continue
                if cell.value is not None:
                    non_empty_cells.append(f"{cell.coordinate}={cell.value!r}")
        wb.close()
        
        if non_empty_cells:
            sample = non_empty_cells[:3]
            more = f" 等{len(non_empty_cells)}个" if len(non_empty_cells) > 3 else ""
            return (f"⚠️ 合并将清除以下单元格的数据: {', '.join(sample)}{more}。"
                    f"Excel合并后只有左上角单元格保留值。如需保留数据请先备份或改用加粗+背景色代替合并。")
        return None
    except Exception:
        return None  # 检查失败不阻塞合并操作


@mcp.tool()
@_track_call
def excel_format_cells(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    formatting: Optional[Dict[str, Any]] = None,
    preset: Optional[str] = None
) -> Dict[str, Any]:
    """单元格样式统一入口：字体样式 + 合并/拆分 + 边框，一个工具完成所有外观操作。

    支持的操作类别（可在单次调用中组合使用）：
      📝 字体: bold, italic, underline('single'/'double'/'singleAccounting'), strikethrough,
              font_size, font_color, font_name
      🎨 单元格: bg_color(背景), fill_type(solid/gradient/pattern), gradient_colors, alignment,
              wrap_text, text_rotation, indent, shrink_to_fit, number_format
      🔗 结构: merge(True=合并), unmerge(True=取消合并)
      📦 边框: border_style(thin/thick/double/dotted/dashed) 或 border{left/right/top/bottom+color}

    常用示例:
      加粗表头:           {"bold": True}
      蓝底白字+双下划线:   {"bg_color": "0000FF", "font_color": "FFFFFF", "underline": "double", "bold": True}
      合并+加粗居中:       {"merge": True, "bold": True, "alignment": "center"}
      删除线+红色背景:     {"strikethrough": True, "bg_color": "FFCCCC"}
      渐变背景色:         {"gradient_colors": ["4472C4", "ED7D31"], "gradient_type": "linear"}
      边框（四边不同）:    {"border": {"top": "medium", "bottom": "thin", "color": "000000"}}
      仅合并:             {"merge": True}
      边框:               {"border_style": "thin"}
      合并+边框+背景色:     {"merge": True, "border_style": "thin", "bg_color": "FFFF00"}
      预设样式:           preset="header"（等价于 bold + center + bg_color）

    ⚠️ 合并警告：merge=True 会清除合并区域内非左上角单元格的值！Excel合并后只有左上角单元格保留数据。
       例如合并 A1:E1 后，B1~E1 的值会丢失。如需保留数据，请先复制到其他位置或改用加粗+背景色代替合并。

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        cell_range: 单元格范围（如 "A1:C10" 或 "Sheet1!A1:C10"，不含!时自动拼接sheet_name）
        formatting: 样式配置字典（可同时指定多项，按 merge/unmerge → format → border 顺序执行）:
            bold/italic/strikethrough: bool
            underline: 'single'(默认) | 'double' | 'singleAccounting' | 'doubleAccounting'
            font_size: int | font_color/bg_color: str (HEX)
            fill_type: 'solid'(默认) | 'gradient' | 'pattern'
            gradient_colors: list[str] 渐变色数组（如 ["4472C4", "ED7D31"]）
            number_format: str | alignment: left/center/right/top/bottom
            wrap_text: bool | text_rotation: int(-90~90度) | indent: int
            shrink_to_fit: bool
            border: dict{left/right/top/bottom/diagonal: style|dict, color: str}
            merge: bool (合并单元格) | unmerge: bool (取消合并)
        preset: 预设样式名（bold/italic/highlight/header/currency/title/data）
        start_cell/end_cell: 可选，替代 cell_range 使用（如 start_cell="A1", end_cell="E1"）
    """
    if formatting is None and preset is None:
        return _fail(
            '未提供样式参数。示例: {"bold": True} 或 {"merge": True} 或 {"border_style": "thin"}',
            meta={"error_code": "MISSING_FORMATTING_PARAMS"})
    if not sheet_name or not sheet_name.strip():
        return _fail('缺少必需参数 sheet_name',
                    meta={"error_code": "MISSING_REQUIRED_PARAM", "param": "sheet_name"})

    if '!' not in cell_range:
        cell_range = f"{sheet_name}!{cell_range}"

    # 提取特殊操作（从formatting中pop，避免传给底层format_cells引起混淆）
    _do_merge = formatting.pop('merge', None) if formatting else None
    _do_unmerge = formatting.pop('unmerge', None) if formatting else None
    _border_style = formatting.pop('border_style', None) if formatting else None

    # 如果提取后formatting为空且没有preset且无特殊操作，报错
    if not formatting and not preset and _do_merge is None and _do_unmerge is None and _border_style is None:
        return _fail(
            '未提供样式参数。示例: {"bold": True} 或 {"merge": True} 或 {"border_style": "thin"}',
            meta={"error_code": "MISSING_FORMATTING_PARAMS"})

    _ops_result: list[tuple[str, bool, str]] = []
    _merge_warning: Optional[str] = None

    # Step 1: 合并/取消合并（结构操作先执行）
    if _do_merge:
        # 检查合并是否会导致数据丢失
        _merge_warning = _check_merge_data_loss(file_path, sheet_name, cell_range)
        r = _ensure_dict(ExcelOperations.merge_cells(file_path, sheet_name, cell_range))
        _ops_result.append(('merge', r.get('success', False), r.get('message', '')))
    elif _do_unmerge:
        r = _ensure_dict(ExcelOperations.unmerge_cells(file_path, sheet_name, cell_range))
        _ops_result.append(('unmerge', r.get('success', False), r.get('message', '')))

    # Step 2: 字体/单元格样式（仅当有剩余格式参数或preset时）
    if formatting or preset:
        r = _ensure_dict(ExcelOperations.format_cells(file_path, sheet_name, cell_range, formatting, preset))
        ok = r.get('success', False)
        _ops_result.append(('format', ok, r.get('message', '')))
        if ok and r.get('data') is None:
            r['message'] += ' (注意：没有单元格需要格式化)'

    # Step 3: 边框
    if _border_style:
        try:
            r = _ensure_dict(ExcelOperations.set_borders(file_path, sheet_name, cell_range, _border_style))
            _ops_result.append(('border', r.get('success', False), r.get('message', '')))
        except Exception as e:
            _ops_result.append(('border', False, str(e)))

    # 汇总结果
    _ops_ok = [name for name, ok, _ in _ops_result if ok]
    _ops_fail = [name for name, ok, _ in _ops_result if not ok]

    if _ops_fail and not _ops_ok:
        return _fail(f"格式化操作全部失败: {', '.join(_ops_fail)}",
                     meta={"error_code": "FORMAT_FAILED", "operations": _ops_result})

    _msg = f"已完成格式化: {', '.join(_ops_ok)}"
    if _ops_fail:
        _msg += f" | 部分失败: {', '.join(_ops_fail)}"
    if _merge_warning:
        _msg += f" | {_merge_warning}"
    return _ok(_msg, data={'operations': [(n, o) for n, o, _ in _ops_result]})



@mcp.tool()
@_validate_file_path()
@_track_call
def excel_set_row_height(file_path: str, sheet_name: str, row_index: int,
                          height: float, count: int = 1) -> Dict[str, Any]:
    """设置行高（磅值）。"""
    if row_index < 1:
        return _fail("row_index 必须大于 0", meta={"error_code": "INVALID_PARAMETER"})
    if height <= 0:
        return _fail("height 必须大于 0", meta={"error_code": "INVALID_PARAMETER"})
    return _wrap(ExcelOperations.set_row_height(file_path, sheet_name, row_index, height, count))


@mcp.tool()
@_validate_file_path()
@_track_call
def excel_set_column_width(file_path: str, sheet_name: str, column_index: int,
                             width: float, count: int = 1) -> Dict[str, Any]:
    """设置列宽（字符单位）。"""
    if column_index < 1:
        return _fail("column_index 必须大于 0", meta={"error_code": "INVALID_PARAMETER"})
    if width <= 0:
        return _fail("width 必须大于 0", meta={"error_code": "INVALID_PARAMETER"})
    return _wrap(ExcelOperations.set_column_width(file_path, sheet_name, column_index, width, count))


# ==================== Excel比较功能 ====================

@mcp.tool()
@_validate_file_path(['file1_path', 'file2_path'])
@_track_call
def excel_compare_files(
    file1_path: str,
    file2_path: str
) -> Dict[str, Any]:
    """逐单元格比较两个Excel文件的所有工作表差异（单元格级对比）。

    与 excel_compare_sheets 的区别：本工具逐个单元格比对值，返回每个不同单元格的位置和前后值。
    适用于：精确找出哪些单元格被修改了。

    Args:
        file1_path: 第一个文件路径（基准/旧版本）
        file2_path: 第二个文件路径（对比/新版本）
    """
    for _p in [file1_path, file2_path]:
        _err = _validate_path(_p)
        if _err:
            return _err
    return _wrap(ExcelOperations.compare_files(file1_path, file2_path))



@mcp.tool()
@_track_call
def excel_compare_sheets(
    file1_path: str,
    sheet1_name: str,
    file2_path: str,
    sheet2_name: str,
    id_column: Union[int, str] = 1,
    header_row: int = 1
) -> Dict[str, Any]:
    """按ID列比较两个工作表的行级差异（对象级对比：新增/删除/修改）。

    与 excel_compare_files 的区别：本工具按ID匹配行，返回行级别的变更摘要（新增/删除/修改了哪些行）。
    适用于：比较两个版本的配置表，了解数据变动概况。

    Args:
        file1_path: 第一个文件路径（基准/旧版本）
        sheet1_name: 第一个工作表名称
        file2_path: 第二个文件路径（对比/新版本）
        sheet2_name: 第二个工作表名称
        id_column: ID列名（字符串）或列索引（从1开始的整数），默认为1表示第1列
        header_row: 表头行号，默认为1
    """
    for _p in [file1_path, file2_path]:
        _err = _validate_path(_p)
        if _err:
            return _err
    return _wrap(ExcelOperations.compare_sheets(file1_path, sheet1_name, file2_path, sheet2_name, id_column, header_row))


# ==================== 主程序 ====================
def main():
    """Entry point for excel-mcp-server-fastmcp.

    Args:
        --stdio: 标准输入输出模式（默认），本地使用，uvx/claude/cursor
        --sse: Server-Sent Events远程模式
        --streamable-http: Streamable HTTP远程模式，推荐用于团队共享
        --mount-path=<path>: HTTP模式挂载路径
        --version, -v: 显示版本号
    """
    if len(sys.argv) > 1 and sys.argv[1] in ('--version', '-v'):
        from excel_mcp_server_fastmcp import __version__
        logger.info(f"excel-mcp-server-fastmcp {__version__}")
        sys.exit(0)

    transport = 'stdio'
    mount_path = None
    for arg in sys.argv[1:]:
        if arg in ('--stdio', '--sse', '--streamable-http'):
            transport = arg[2:]  # remove '--'
        elif arg.startswith('--mount-path='):
            mount_path = arg.split('=', 1)[1]

    mcp.run(transport=transport, mount_path=mount_path)

if __name__ == "__main__":
    main()
