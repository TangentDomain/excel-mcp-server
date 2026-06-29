#!/usr/bin/env python3
"""ExcelMCP CLI — 包内模块入口。

pip install 后通过 `excel-cli` 命令调用。
用法: excel-cli <command> [options]
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import urllib.request
from datetime import datetime

VERSION = "1.0.0"
GITHUB_REPO = "TangentDomain/excel-mcp-server"
GITHUB_RAW = f"https://raw.githubusercontent.com/{GITHUB_REPO}/main/scripts/excel-cli.py"
GITHUB_RELEASE = f"https://github.com/{GITHUB_REPO}/releases/latest/download"
GITHUB_API_COMMITS = f"https://api.github.com/repos/{GITHUB_REPO}/commits/main"
_IS_FROZEN = getattr(sys, "frozen", False)


def detect_platform() -> tuple[str, str]:
    """检测当前平台和架构，返回 (platform, arch) 标识符。

    platform: windows | macos | linux
    arch:     amd64  | arm64  | x86_64（linux 用 x86_64）

    后续 PyInstaller 打包时，self-update 会用此拼接二进制下载 URL:
        {GITHUB_RELEASE}/excel-cli-{platform}-{arch}{ext}
    当前源码阶段仍拉 .py 文件。
    """
    s = sys.platform.lower()
    if s.startswith("win"):
        platform = "windows"
    elif s.startswith("darwin"):
        platform = "macos"
    else:
        platform = "linux"

    machine = platform.machine().lower()
    if machine in ("x86_64", "amd64"):
        arch = "amd64"
    elif machine in ("aarch64", "arm64"):
        arch = "arm64"
    else:
        arch = machine  # fallback

    return platform, arch


def _self_update_url() -> str:
    """返回 self-update 下载 URL。

    源码模式（非冻结）: 拉 GitHub raw .py 文件
    二进制模式（PyInstaller）: 拉对应平台的预编译二进制
    """
    platform, arch = detect_platform()
    if _IS_FROZEN:
        ext = ".exe" if platform == "windows" else ""
        return f"{GITHUB_RELEASE}/excel-cli-{platform}-{arch}{ext}"
    else:
        return GITHUB_RAW


# ==================== 统一响应格式（复现 server.py 的 _wrap/_ok/_fail） ====================


def _ok(message: str = "", data=None, meta: dict = None) -> dict:
    r = {"success": True}
    if message:
        r["message"] = message
    if data is not None:
        r["data"] = data
    if meta:
        r["meta"] = meta
    return r


def _fail(message: str, meta: dict = None) -> dict:
    r = {"success": False, "message": message}
    if meta:
        r["meta"] = meta
    return r


def _ensure_dict(result) -> dict:
    if isinstance(result, dict):
        return result
    if hasattr(result, "__dataclass_fields__"):
        from dataclasses import asdict

        return asdict(result)
    return result


def _strip_defaults(obj, depth=0):
    if depth > 5 or not isinstance(obj, dict):
        return obj
    excel_default_fields = {
        "bold",
        "italic",
        "underline",
        "wrap_text",
        "auto_filter",
        "border_top",
        "border_bottom",
        "border_left",
        "border_right",
        "horizontal_alignment",
        "vertical_alignment",
        "text_rotation",
        "indent",
        "shrink_to_fit",
        "merge_cells",
    }
    semantic_list_fields = {
        "headers",
        "sheets",
        "sheets_with_headers",
        "field_names",
        "descriptions",
        "data",
        "columns",
        "rows",
    }
    cell_semantic_fields = {"value"}
    _is_cell_info = "coordinate" in obj
    cleaned = {}
    for k, v in obj.items():
        if k in cell_semantic_fields and _is_cell_info:
            cleaned[k] = v
            continue
        if v is None or v == "":
            continue
        if isinstance(v, (list, dict)) and len(v) == 0 and k not in semantic_list_fields:
            continue
        if k.lower() in excel_default_fields and v in [False, 0, None]:
            continue
        if isinstance(v, dict):
            cleaned[k] = _strip_defaults(v, depth + 1)
        elif isinstance(v, list):
            cleaned[k] = [_strip_defaults(i, depth + 1) if isinstance(i, dict) else i for i in v]
        else:
            cleaned[k] = v
    return cleaned


def wrap_result(result, meta=None):
    """统一返回格式 {success, data, message, meta} — 复现 server.py _wrap 逻辑"""
    result = _ensure_dict(result)
    if not isinstance(result, dict):
        return result
    err_val = result.get("error")
    if isinstance(err_val, str) and not result.get("message"):
        result["message"] = result.pop("error")
    if "success" not in result:
        result["success"] = True
    if result.get("success") is True and "message" not in result:
        result["message"] = "操作成功"
    if "metadata" in result:
        m = result.pop("metadata")
        if isinstance(m, dict) and m:
            merged = {**m, **(meta or {})}
            result["meta"] = merged
            meta = None
    if meta and "meta" not in result:
        result["meta"] = meta
    if result.get("success") is True and "data" in result and isinstance(result["data"], dict):
        result["data"] = _strip_defaults(result["data"])
    return result


def output(result):
    """输出 JSON 到 stdout，返回退出码"""
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result.get("success") else 1


# ==================== JSON 参数解析辅助 ====================


def parse_json(s, default=None):
    """安全解析 JSON 字符串"""
    if s is None:
        return default
    try:
        return json.loads(s)
    except (json.JSONDecodeError, TypeError):
        return default


# ==================== 查询类子命令 (9) ====================


def cmd_list_sheets(args):
    """列出工作表。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        result = ExcelOperations.list_sheets(args.file)
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"列出工作表失败: {e}"))


def cmd_get_headers(args):
    """获取表头。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        result = ExcelOperations.get_headers(
            args.file,
            args.sheet,
            args.header_row,
            args.max_columns,
        )
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"获取表头失败: {e}"))


def cmd_get_range(args):
    """读取单元格范围数据。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        cell_range = args.range
        if args.sheet and "!" not in cell_range:
            cell_range = f"{args.sheet}!{cell_range}"
        result = ExcelOperations.get_range(args.file, cell_range, args.formatting)
        result = _ensure_dict(result)
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"获取范围失败: {e}"))


def cmd_describe_table(args):
    """查看表结构（列名+类型+样本值），支持双行表头自动检测。"""
    try:
        from openpyxl import load_workbook

        from excel_mcp_server_fastmcp.api.header_analyzer import _cell_str, detect_from_rows

        wb = load_workbook(args.file, read_only=True, data_only=True)
        sheet_name = args.sheet or wb.sheetnames[0]
        ws = wb[sheet_name]

        # 读取前几行来推断结构
        all_rows = list(ws.iter_rows(max_row=min(ws.max_row or 10, 50), values_only=True))
        if not all_rows:
            return output(_fail("工作表为空"))

        # 双行表头检测
        raw_first = list(all_rows[0]) if len(all_rows) > 0 else []
        raw_second = list(all_rows[1]) if len(all_rows) > 1 else []
        is_dual, header_row_idx, descriptions = detect_from_rows([raw_first, raw_second] if len(all_rows) >= 2 else [raw_first])

        # 根据检测结果取表头和数据行
        if is_dual:
            headers = [_cell_str(c) or "" for c in raw_second]
            sample_row = all_rows[2] if len(all_rows) > 2 else []
            header_type = "dual"
            data_start = 3  # Excel 行号
        else:
            headers = [_cell_str(c) or "" for c in raw_first]
            sample_row = list(all_rows[1]) if len(all_rows) > 1 else []
            header_type = "single"
            data_start = 2

        columns = []
        for i, h in enumerate(headers):
            sample = sample_row[i] if i < len(sample_row) else None
            col_type = type(sample).__name__ if sample is not None else "null"
            columns.append(
                {
                    "name": h,
                    "type": col_type,
                    "sample": sample,
                }
            )

        data = {
            "sheet_name": sheet_name,
            "header_type": header_type,
            "row_count": (ws.max_row or 0) - (data_start - 1),
            "column_count": ws.max_column or 0,
            "columns": columns,
        }
        wb.close()
        return output(_ok("表结构分析完成", data=data))
    except Exception as e:
        return output(_fail(f"表结构分析失败: {e}"))


def cmd_search(args):
    """搜索单元格文本。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        _range_arg = args.range
        if _range_arg and "!" not in _range_arg and args.sheet:
            _range_arg = f"{args.sheet}!{_range_arg}"
        use_regex = args.regex
        if use_regex is None:
            use_regex = bool(re.match(r".*[\[\](){}*+?|^$\\.]", args.pattern))
        result = ExcelOperations.search(
            args.file,
            args.pattern,
            args.sheet,
            args.case_sensitive,
            args.whole_word,
            use_regex,
            include_values=True,
            include_formulas=False,
            range=_range_arg,
        )
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"搜索失败: {e}"))


def cmd_search_directory(args):
    """跨文件搜索。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        use_regex = args.regex
        if use_regex is None:
            use_regex = bool(re.match(r".*[\[\](){}*+?|^$\\.]", args.pattern))
        extensions = parse_json(args.extensions, None)
        result = ExcelOperations.search_directory(
            args.dir,
            args.pattern,
            args.case_sensitive,
            args.whole_word,
            use_regex,
            True,
            False,
            recursive=args.recursive,
            file_extensions=extensions,
            file_pattern=None,
            max_files=args.max_files,
        )
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"目录搜索失败: {e}"))


def cmd_find_last_row(args):
    """定位数据末行。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        result = ExcelOperations.find_last_row(args.file, args.sheet, args.column)
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"查找末行失败: {e}"))


def cmd_query(args):
    """执行 SQL 查询。"""
    from excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_sql_query

    try:
        result = execute_advanced_sql_query(
            file_path=args.file,
            sql=args.sql,
            sheet_name=getattr(args, "sheet", None),
            limit=None,
            include_headers=not args.no_headers,
            output_format=args.format or "table",
        )
        wrapped = wrap_result(result)
        if "meta" not in wrapped:
            wrapped["meta"] = {"file_path": args.file}
        return output(wrapped)
    except Exception as e:
        return output(_fail(f"SQL查询失败: {e}"))


def cmd_compare_sheets(args):
    """对比两个工作表差异。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        result = ExcelOperations.compare_sheets(
            args.file1,
            args.sheet1,
            args.file2,
            args.sheet2,
            args.id_column,
            args.header_row,
        )
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"比较工作表失败: {e}"))


# ==================== 写入类子命令 (7) ====================


def cmd_update_range(args):
    """写入单元格范围数据。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        cell_range = args.range
        if args.sheet and "!" not in cell_range:
            cell_range = f"{args.sheet}!{cell_range}"
        data = parse_json(args.data, [])
        result = ExcelOperations.update_range(
            args.file,
            cell_range,
            data,
            args.preserve_formulas,
            args.insert_mode,
        )
        result = _ensure_dict(result)
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"更新范围失败: {e}"))


def cmd_upsert_row(args):
    """按主键插入或更新行。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        updates = parse_json(args.updates, {})
        result = ExcelOperations.upsert_row(
            args.file,
            args.sheet,
            args.key_column,
            args.key_value,
            updates,
            args.header_row,
            True,
        )
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"插入/更新行失败: {e}"))


def cmd_set_formula(args):
    """设置单元格公式。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        result = ExcelOperations.set_formula(
            args.file,
            args.sheet,
            args.cell,
            args.formula,
        )
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"设置公式失败: {e}"))


def cmd_update_query(args):
    """执行 SQL UPDATE。"""
    from excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_update_query

    try:
        result = execute_advanced_update_query(
            file_path=args.file,
            sql=args.sql,
            dry_run=args.dry_run,
        )
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"UPDATE执行失败: {e}"))


def cmd_insert_query(args):
    """执行 SQL INSERT。"""
    from excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_insert_query

    try:
        result = execute_advanced_insert_query(
            file_path=args.file,
            sql=args.sql,
            dry_run=args.dry_run,
        )
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"INSERT执行失败: {e}"))


def cmd_delete_query(args):
    """执行 SQL DELETE。"""
    from excel_mcp_server_fastmcp.api.advanced_sql_query import execute_advanced_delete_query

    try:
        result = execute_advanced_delete_query(
            file_path=args.file,
            sql=args.sql,
            dry_run=args.dry_run,
        )
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"DELETE执行失败: {e}"))


def cmd_run_python(args):
    """执行 Python 脚本。"""
    from excel_mcp_server_fastmcp.api.script_runner import execute_python_script

    try:
        result = execute_python_script(args.file, args.code, args.sheet, args.timeout)
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"脚本执行失败: {e}"))


# ==================== 结构操作类子命令 (6) ====================


def cmd_create_sheet(args):
    """创建工作表。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        result = ExcelOperations.create_sheet(args.file, args.name, args.index)
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"创建工作表失败: {e}"))


def cmd_delete_sheet(args):
    """删除工作表。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        result = ExcelOperations.delete_sheet(args.file, args.name)
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"删除工作表失败: {e}"))


def cmd_rename_sheet(args):
    """重命名工作表。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        result = ExcelOperations.rename_sheet(args.file, args.old_name, args.new_name)
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"重命名工作表失败: {e}"))


def cmd_copy_sheet(args):
    """复制工作表。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        result = ExcelOperations.copy_sheet(
            args.file,
            args.source,
            args.new_name,
            args.index,
            True,
        )
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"复制工作表失败: {e}"))


def cmd_structure(args):
    """增删行列结构。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        ops_map = {
            "insert_rows": "insert_rows",
            "insert_columns": "insert_columns",
            "insert_cols": "insert_columns",
            "delete_rows": "delete_rows",
            "delete_columns": "delete_columns",
            "delete_cols": "delete_columns",
        }
        method_name = ops_map.get(args.operation)
        if not method_name:
            return output(
                _fail(
                    f"不支持的operation: {args.operation}。可选: {', '.join(ops_map.keys())}",
                    meta={"error_code": "INVALID_OPERATION"},
                )
            )
        method = getattr(ExcelOperations, method_name)
        result = method(args.file, args.sheet, args.index, args.count)
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"结构操作失败: {e}"))


def cmd_rename_column(args):
    """重命名列。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        result = ExcelOperations.rename_column(
            args.file,
            args.sheet,
            args.old_header,
            args.new_header,
            args.header_row,
        )
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"重命名列失败: {e}"))


# ==================== 格式化类子命令 (2) ====================


def cmd_format_cells(args):
    """格式化单元格。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        cell_range = args.range
        if "!" not in cell_range:
            cell_range = f"{args.sheet}!{cell_range}"

        formatting = parse_json(args.formatting, None)
        preset = args.preset

        # 复现 server.py excel_format_cells 的组合操作逻辑
        _do_merge = formatting.pop("merge", None) if formatting else None
        _do_unmerge = formatting.pop("unmerge", None) if formatting else None
        _border_style = formatting.pop("border_style", None) if formatting else None

        if not formatting and not preset and _do_merge is None and _do_unmerge is None and _border_style is None:
            return output(
                _fail(
                    '未提供样式参数。示例: {"bold": True} 或 {"merge": True} 或 {"border_style": "thin"}',
                    meta={"error_code": "MISSING_FORMATTING_PARAMS"},
                )
            )

        _ops_result = []
        _merge_warning = None

        # Step 1: 合并/取消合并
        if _do_merge:
            r = _ensure_dict(ExcelOperations.merge_cells(args.file, args.sheet, cell_range))
            _ops_result.append(("merge", r.get("success", False), r.get("message", "")))
        elif _do_unmerge:
            r = _ensure_dict(ExcelOperations.unmerge_cells(args.file, args.sheet, cell_range))
            _ops_result.append(("unmerge", r.get("success", False), r.get("message", "")))

        # Step 2: 样式
        if formatting or preset:
            r = _ensure_dict(ExcelOperations.format_cells(args.file, args.sheet, cell_range, formatting, preset))
            ok = r.get("success", False)
            _ops_result.append(("format", ok, r.get("message", "")))

        # Step 3: 边框
        if _border_style:
            try:
                r = _ensure_dict(ExcelOperations.set_borders(args.file, args.sheet, cell_range, _border_style))
                _ops_result.append(("border", r.get("success", False), r.get("message", "")))
            except Exception as e:
                _ops_result.append(("border", False, str(e)))

        _ops_ok = [name for name, ok, _ in _ops_result if ok]
        _ops_fail = [name for name, ok, _ in _ops_result if not ok]

        if _ops_fail and not _ops_ok:
            return output(
                _fail(
                    f"格式化操作全部失败: {', '.join(_ops_fail)}",
                    meta={"error_code": "FORMAT_FAILED", "operations": _ops_result},
                )
            )

        _msg = f"已完成格式化: {', '.join(_ops_ok)}"
        if _ops_fail:
            _msg += f" | 部分失败: {', '.join(_ops_fail)}"
        if _merge_warning:
            _msg += f" | {_merge_warning}"
        return output(_ok(_msg, data={"operations": [(n, o) for n, o, _ in _ops_result]}))
    except Exception as e:
        return output(_fail(f"格式化失败: {e}"))


def cmd_set_layout(args):
    """设置行高列宽。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        op_map = {
            "set_row_height": "set_row_height",
            "row_height": "set_row_height",
            "set_column_width": "set_column_width",
            "column_width": "set_column_width",
        }
        method_name = op_map.get(args.operation)
        if not method_name:
            return output(
                _fail(
                    f"不支持的operation: {args.operation}。可选: set_row_height, set_column_width",
                    meta={"error_code": "INVALID_OPERATION"},
                )
            )
        if args.index < 1:
            return output(_fail("index 必须大于 0", meta={"error_code": "INVALID_PARAMETER"}))
        if args.value <= 0:
            return output(_fail("value 必须大于 0", meta={"error_code": "INVALID_PARAMETER"}))
        method = getattr(ExcelOperations, method_name)
        result = method(args.file, args.sheet, args.index, args.value, args.count)
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"布局设置失败: {e}"))


# ==================== 文件操作类子命令 (2) ====================


def cmd_create_file(args):
    """创建 Excel 文件。"""
    from excel_mcp_server_fastmcp.api.excel_operations import ExcelOperations

    try:
        sheets = parse_json(args.sheets, None)
        result = ExcelOperations.create_file(args.file, sheets)
        return output(wrap_result(result))
    except Exception as e:
        return output(_fail(f"创建文件失败: {e}"))


def cmd_backup(args):
    """备份或恢复文件。"""
    try:
        if args.operation == "create":
            if not os.path.exists(args.file):
                return output(_fail(f"源文件不存在: {args.file}", meta={"error_code": "FILE_NOT_FOUND"}))
            _bd = args.backup_dir or os.path.join(os.path.dirname(args.file), ".excel_mcp_backups")
            os.makedirs(_bd, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.basename(args.file)
            name, ext = os.path.splitext(filename)
            backup_filename = f"{name}_backup_{timestamp}{ext}"
            backup_path = os.path.join(_bd, backup_filename)
            shutil.copy2(args.file, backup_path)
            _ts = datetime.now().isoformat()
            return output(
                _ok(
                    f"备份创建成功: {backup_filename}",
                    data={"backup_file": backup_path, "backup_directory": _bd, "timestamp": _ts},
                    meta={"file_path": args.file},
                )
            )

        elif args.operation == "list":
            _bd = args.backup_dir or os.path.join(os.path.dirname(args.file), ".excel_mcp_backups")
            if not os.path.exists(_bd):
                return output(_ok("备份目录不存在", data={"backups": []}, meta={"file_path": args.file}))
            filename = os.path.basename(args.file)
            name, ext = os.path.splitext(filename)
            backup_files = []
            for fn in os.listdir(_bd):
                if fn.startswith(f"{name}_backup_") and fn.endswith(ext):
                    fp = os.path.join(_bd, fn)
                    st = os.stat(fp)
                    backup_files.append(
                        {
                            "filename": fn,
                            "path": fp,
                            "size": st.st_size,
                            "created_time": datetime.fromtimestamp(st.st_ctime).isoformat(),
                        }
                    )
            backup_files.sort(key=lambda x: x["created_time"], reverse=True)
            return output(
                _ok(
                    f"找到 {len(backup_files)} 个备份",
                    data={"backups": backup_files, "backup_directory": _bd},
                    meta={"file_path": args.file},
                )
            )

        elif args.operation == "restore":
            if not args.backup_id:
                return output(_fail("restore 操作需要 --backup-id 参数", meta={"error_code": "MISSING_PARAM"}))
            backup_path = args.backup_id
            if not os.path.exists(backup_path):
                return output(_fail(f"备份文件不存在: {backup_path}", meta={"error_code": "BACKUP_NOT_FOUND"}))
            filename = os.path.basename(backup_path)
            if "_backup_" in filename:
                parts = filename.split("_backup_")
                target_path = os.path.join(os.path.dirname(backup_path), parts[0] + os.path.splitext(backup_path)[1])
            else:
                target_path = args.file
            shutil.copy2(backup_path, target_path)
            return output(
                _ok(
                    f"文件恢复成功: {os.path.basename(target_path)}",
                    data={"backup_file": backup_path, "target_file": target_path},
                    meta={"file_path": backup_path},
                )
            )
        else:
            return output(
                _fail(
                    f"不支持的operation: {args.operation}。可选: create, list, restore",
                    meta={"error_code": "INVALID_OPERATION"},
                )
            )
    except Exception as e:
        return output(_fail(f"备份操作失败: {e}"))


# ==================== self-update 子命令 ====================


def cmd_self_update(args):
    """检查/更新 CLI 到最新版本。"""
    try:
        check_result = _check_update()

        if args.check:
            return check_result

        # check_result 的 data 里有 need_update, local_version, remote_version, local_sha, remote_sha
        data = json.loads(check_result)
        if not data.get("success"):
            return check_result

        if not data["data"].get("need_update"):
            return check_result

        return _do_update()
    except Exception as e:
        return output(_fail(f"self-update 失败: {e}"))


def _check_update():
    """对比本地版本和 GitHub 最新 commit。"""
    try:
        # 本地 SHA（从 .last-update 文件读取）
        script_dir = os.path.dirname(os.path.abspath(__file__))
        version_file = os.path.join(script_dir, ".last-update")
        local_sha = ""
        if os.path.exists(version_file):
            with open(version_file) as f:
                local_sha = f.read().strip()

        # 请求 GitHub API
        req = urllib.request.Request(
            f"https://api.github.com/repos/{GITHUB_REPO}/commits/main",
            headers={"User-Agent": f"excel-cli/{VERSION}", "Accept": "application/vnd.github.v3+json"},
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        remote_sha = data.get("sha", "")[:12]
        commit_message = data.get("commit", {}).get("message", "").split("\n")[0]
        commit_date = data.get("commit", {}).get("committer", {}).get("date", "")[:10]

        need_update = not local_sha or (remote_sha and remote_sha != local_sha)

        return output(
            _ok(
                "有新版本" if need_update else f"已是最新版 ({VERSION})",
                data={
                    "local_version": VERSION,
                    "need_update": need_update,
                    "local_sha": local_sha or "(首次检查)",
                    "remote_sha": remote_sha,
                    "commit_message": commit_message,
                    "commit_date": commit_date,
                },
            )
        )
    except urllib.error.HTTPError as e:
        return output(_fail(f"检查更新失败: HTTP {e.code}", meta={"error": "HTTP_ERROR"}))
    except Exception as e:
        return output(_fail(f"检查更新失败: {e}", meta={"error": "NETWORK_ERROR"}))


def _do_update():
    """执行更新。优先用 uv/venv 模式，回退到传统下载替换。"""
    is_venv = hasattr(sys, "real_prefix") or (hasattr(sys, "base_prefix") and sys.base_prefix != sys.prefix)

    if is_venv:
        # venv 模式：pip install --upgrade
        import subprocess

        try:
            result = subprocess.run(
                [sys.executable, "-m", "pip", "install", "--upgrade", f"git+https://github.com/{GITHUB_REPO}"],
                capture_output=True,
                text=True,
                timeout=120,
            )
            if result.returncode != 0:
                return output(_fail(f"pip 安装失败: {result.stderr}", meta={"error": "PIP_INSTALL_FAILED"}))

            # 更新成功后写 .last-update
            try:
                req = urllib.request.Request(
                    f"https://api.github.com/repos/{GITHUB_REPO}/commits/main",
                    headers={"User-Agent": f"excel-cli/{VERSION}", "Accept": "application/vnd.github.v3+json"},
                )
                with urllib.request.urlopen(req, timeout=15) as resp:
                    remote_data = json.loads(resp.read().decode("utf-8"))
                script_dir = os.path.dirname(os.path.abspath(__file__))
                with open(os.path.join(script_dir, ".last-update"), "w") as f:
                    f.write(remote_data.get("sha", "")[:12])
            except Exception:
                pass

            return output(_ok(f"更新成功: {VERSION} -> 最新版", data={"status": "updated", "mode": "pip"}))
        except subprocess.TimeoutExpired:
            return output(_fail("pip 安装超时", meta={"error": "TIMEOUT"}))
    else:
        # 传统模式：下载替换
        return _legacy_update()


def _legacy_update():
    """传统下载替换更新（源码或 PyInstaller 二进制）。"""
    try:
        platform, arch = detect_platform()
        self_path = os.path.abspath(__file__)
        if _IS_FROZEN:
            self_path = os.path.abspath(sys.executable)
        download_url = _self_update_url()

        req = urllib.request.Request(
            download_url,
            headers={"User-Agent": f"excel-cli/{VERSION} ({platform}-{arch})"},
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            if resp.status != 200:
                return output(
                    _fail(
                        f"下载失败: HTTP {resp.status} {resp.reason}",
                        meta={"error_code": "UPDATE_CHECK_FAILED", "url": download_url},
                    )
                )
            new_content = resp.read()

        # 检测新版本号（源码模式从文件内容提取）
        new_version = VERSION
        if not _IS_FROZEN:
            try:
                new_text = new_content.decode("utf-8")
                import ast

                for line in new_text.splitlines():
                    if line.startswith("VERSION = "):
                        new_version = ast.literal_eval(line.split("=", 1)[1].strip())
                        break
            except Exception:
                pass

        # 原子替换：Windows 先写 .new 避免文件锁
        new_path = self_path + ".new"
        with open(new_path, "wb") as f:
            f.write(new_content)
        try:
            os.replace(new_path, self_path)
        except PermissionError:
            return output(
                _ok(
                    f"文件被占用，已写入: {new_path}。请手动替换 {self_path} 后重启。",
                    data={
                        "current_version": VERSION,
                        "new_version": new_version,
                        "temp_file": new_path,
                        "platform": platform,
                        "arch": arch,
                        "mode": "binary" if _IS_FROZEN else "source",
                        "status": "file_locked",
                    },
                )
            )

        # 更新成功后写 .last-update
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            version_file = os.path.join(script_dir, ".last-update")
            req2 = urllib.request.Request(
                f"https://api.github.com/repos/{GITHUB_REPO}/commits/main",
                headers={"User-Agent": f"excel-cli/{VERSION}", "Accept": "application/vnd.github.v3+json"},
            )
            with urllib.request.urlopen(req2, timeout=15) as resp2:
                sha_data = json.loads(resp2.read().decode("utf-8"))
            with open(version_file, "w") as f:
                f.write(sha_data.get("sha", "")[:12])
        except Exception:
            pass

        return output(
            _ok(
                f"更新成功: {VERSION} -> {new_version}",
                data={
                    "current_version": VERSION,
                    "new_version": new_version,
                    "platform": platform,
                    "arch": arch,
                    "mode": "binary" if _IS_FROZEN else "source",
                    "status": "updated",
                },
            )
        )
    except Exception as e:
        return output(_fail(f"legacy update 失败: {e}"))


# ==================== 版本子命令 ====================


def cmd_version(args):

    print(
        json.dumps(
            {
                "success": True,
                "data": {"version": VERSION, "python": sys.version.split()[0]},
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


# ==================== argparse 定义 ====================


def build_parser():
    """构建 argparse 命令行解析器。"""
    parser = argparse.ArgumentParser(
        prog="excel-cli",
        description="ExcelMCP CLI — 将 26 个 MCP 工具转为 CLI 子命令，输出 JSON。",
        epilog="示例:\n"
        "  excel-cli list-sheets --file data.xlsx\n"
        '  excel-cli query --file data.xlsx --sql "SELECT * FROM Sheet1 LIMIT 5"\n'
        '  excel-cli update-query --file data.xlsx --sql "UPDATE Sheet1 SET Col=1 WHERE ID>5"\n'
        "  excel-cli self-update",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--version", action="version", version=f"%(prog)s {VERSION}")
    subparsers = parser.add_subparsers(dest="command", help="子命令")

    # ---- 查询类 (9) ----
    # list-sheets
    p = subparsers.add_parser("list-sheets", help="列出所有工作表名称")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.set_defaults(func=cmd_list_sheets)

    # get-headers
    p = subparsers.add_parser("get-headers", help="获取表头信息（中文+英文）")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sheet", default=None, help="工作表名称")
    p.add_argument("--header-row", type=int, default=1, help="表头行号，默认1")
    p.add_argument("--max-columns", type=int, default=None, help="最大列数")
    p.set_defaults(func=cmd_get_headers)

    # get-range
    p = subparsers.add_parser("get-range", help="获取指定单元格范围的数据")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--range", required=True, help="单元格范围（如 A1:C10 或 Sheet1!A1:C10）")
    p.add_argument("--sheet", default=None, help="工作表名称")
    p.add_argument("--formatting", action="store_true", help="包含格式信息")
    p.set_defaults(func=cmd_get_range)

    # describe-table
    p = subparsers.add_parser("describe-table", help="查看表结构（列名+类型+样本值）")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sheet", default=None, help="工作表名称")
    p.set_defaults(func=cmd_describe_table)

    # search
    p = subparsers.add_parser("search", help="在 Excel 中搜索单元格")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--pattern", required=True, help="搜索模式")
    p.add_argument("--sheet", default=None, help="工作表名称")
    p.add_argument("--case-sensitive", action="store_true", help="区分大小写")
    p.add_argument("--whole-word", action="store_true", help="全词匹配")
    p.add_argument("--regex", default=None, nargs="?", const=True, type=lambda x: x.lower() in ("true", "1", "yes"), help="使用正则")
    p.add_argument("--range", default=None, help="搜索范围")
    p.set_defaults(func=cmd_search)

    # search-directory
    p = subparsers.add_parser("search-directory", help="在目录下搜索 Excel 文件")
    p.add_argument("--dir", required=True, help="搜索目录")
    p.add_argument("--pattern", required=True, help="搜索模式")
    p.add_argument("--case-sensitive", action="store_true", help="区分大小写")
    p.add_argument("--whole-word", action="store_true", help="全字匹配")
    p.add_argument("--regex", default=None, help="正则模式 (None=自动检测)")
    p.add_argument("--recursive", type=lambda x: x.lower() in ("true", "1", "yes"), default=True, help="递归子目录")
    p.add_argument("--extensions", default=None, help='扩展名过滤 JSON（如 [".xlsx"]）')
    p.add_argument("--max-files", type=int, default=100, help="最大搜索文件数")
    p.set_defaults(func=cmd_search_directory)

    # find-last-row
    p = subparsers.add_parser("find-last-row", help="查找最后一行")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sheet", required=True, help="工作表名称")
    p.add_argument("--column", default=None, help="列名或列号")
    p = subparsers.add_parser("query", help="SQL 查询（优先使用）")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sql", required=True, help="SQL 查询语句")
    p.add_argument("--sheet", default=None, help="工作表名称（可选，多表文件时指定）")
    p.add_argument("--no-headers", action="store_true", help="不包含表头")
    p.add_argument("--format", default=None, choices=["table", "json", "csv"], help="输出格式")
    p.set_defaults(func=cmd_query)

    # compare-sheets
    p = subparsers.add_parser("compare-sheets", help="按 ID 列比较两个工作表差异")
    p.add_argument("--file1", required=True, help="基准文件路径")
    p.add_argument("--sheet1", required=True, help="基准工作表")
    p.add_argument("--file2", required=True, help="对比文件路径")
    p.add_argument("--sheet2", required=True, help="对比工作表")
    p.add_argument("--id-column", default=1, help="ID 列名或列索引（从1开始），默认1")
    p.add_argument("--header-row", type=int, default=1, help="表头行号")
    p.set_defaults(func=cmd_compare_sheets)

    # ---- 写入类 (7) ----
    # update-range
    p = subparsers.add_parser("update-range", help="精确坐标写入数据")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--range", required=True, help="单元格范围")
    p.add_argument("--data", required=True, help="数据 JSON 二维数组 [[row1], [row2], ...]")
    p.add_argument("--preserve-formulas", action="store_true", default=True, help="保留公式（默认开启）")
    p.add_argument("--no-preserve-formulas", action="store_false", dest="preserve_formulas", help="不保留公式")
    p.add_argument("--insert-mode", action="store_true", help="插入模式（原有数据下移）")
    p.add_argument("--sheet", default=None, help="工作表名称")
    p.set_defaults(func=cmd_update_range)

    # upsert-row
    p = subparsers.add_parser("upsert-row", help="按 key_column+key_value 插入或更新行")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sheet", required=True, help="工作表名称")
    p.add_argument("--key-column", required=True, help="键列名")
    p.add_argument("--key-value", required=True, help="键值")
    p.add_argument("--updates", required=True, help='更新字段 JSON（如 {"伤害": 200}）')
    p.add_argument("--header-row", type=int, default=1, help="表头行号")
    p.set_defaults(func=cmd_upsert_row)

    # set-formula
    p = subparsers.add_parser("set-formula", help="写入 Excel 公式")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sheet", required=True, help="工作表名称")
    p.add_argument("--cell", required=True, help="单元格地址（如 A1）")
    p.add_argument("--formula", required=True, help="公式（以 = 开头）")
    p.set_defaults(func=cmd_set_formula)

    # update-query
    p = subparsers.add_parser("update-query", help="SQL UPDATE 批量修改")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sql", required=True, help="UPDATE 语句")
    p.add_argument("--dry-run", action="store_true", help="仅预览不实际写入")
    p.set_defaults(func=cmd_update_query)

    # insert-query
    p = subparsers.add_parser("insert-query", help="SQL INSERT 插入数据")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sql", required=True, help="INSERT 语句")
    p.add_argument("--dry-run", action="store_true", help="仅预览不实际写入")
    p.set_defaults(func=cmd_insert_query)

    # delete-query
    p = subparsers.add_parser("delete-query", help="SQL DELETE 删除数据")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sql", required=True, help="DELETE 语句")
    p.add_argument("--dry-run", action="store_true", help="仅预览不实际删除")
    p.set_defaults(func=cmd_delete_query)

    # run-python
    p = subparsers.add_parser("run-python", help="直接执行 Python 代码操作 Excel")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--code", required=True, help="Python 代码")
    p.add_argument("--sheet", default=None, help="工作表名称")
    p.add_argument("--timeout", type=int, default=30, help="超时秒数（默认30，最大120）")
    p.set_defaults(func=cmd_run_python)

    # ---- 结构操作类 (6) ----
    # create-sheet
    p = subparsers.add_parser("create-sheet", help="创建新工作表")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--name", required=True, help="工作表名称")
    p.add_argument("--index", type=int, default=None, help="插入位置（从0开始）")
    p.set_defaults(func=cmd_create_sheet)

    # delete-sheet
    p = subparsers.add_parser("delete-sheet", help="删除工作表")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--name", required=True, help="工作表名称")
    p.set_defaults(func=cmd_delete_sheet)

    # rename-sheet
    p = subparsers.add_parser("rename-sheet", help="重命名工作表")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--old-name", required=True, help="原工作表名")
    p.add_argument("--new-name", required=True, help="新工作表名")
    p.set_defaults(func=cmd_rename_sheet)

    # copy-sheet
    p = subparsers.add_parser("copy-sheet", help="复制工作表")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--source", required=True, help="源工作表名")
    p.add_argument("--new-name", default=None, help="新工作表名（默认: 源名_副本）")
    p.add_argument("--index", type=int, default=None, help="插入位置")
    p.set_defaults(func=cmd_copy_sheet)

    # structure
    p = subparsers.add_parser("structure", help="插入或删除行和列")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sheet", required=True, help="工作表名称")
    p.add_argument("--operation", required=True, choices=["insert_rows", "delete_rows", "insert_columns", "insert_cols", "delete_columns", "delete_cols"], help="操作类型")
    p.add_argument("--index", type=int, required=True, help="行/列位置（从1开始）")
    p.add_argument("--count", type=int, default=1, help="数量（默认1）")
    p.set_defaults(func=cmd_structure)

    # rename-column
    p = subparsers.add_parser("rename-column", help="修改列名（表头）")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sheet", required=True, help="工作表名称")
    p.add_argument("--old-header", required=True, help="原列名")
    p.add_argument("--new-header", required=True, help="新列名")
    p.add_argument("--header-row", type=int, default=1, help="表头行号")
    p.set_defaults(func=cmd_rename_column)

    # ---- 格式化类 (2) ----
    # format-cells
    p = subparsers.add_parser("format-cells", help="设置单元格样式")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sheet", required=True, help="工作表名称")
    p.add_argument("--range", required=True, help="单元格范围")
    p.add_argument("--formatting", default=None, help='样式 JSON（如 {"bold": true}）')
    p.add_argument("--preset", default=None, help="预设样式（header/title/data 等）")
    p.set_defaults(func=cmd_format_cells)

    # set-layout
    p = subparsers.add_parser("set-layout", help="设置行高或列宽")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--sheet", required=True, help="工作表名称")
    p.add_argument("--operation", required=True, choices=["set_row_height", "set_column_width", "row_height", "column_width"], help="操作类型")
    p.add_argument("--index", type=int, required=True, help="行号或列号（从1开始）")
    p.add_argument("--value", type=float, required=True, help="行高（磅）或列宽（字符）")
    p.add_argument("--count", type=int, default=1, help="连续影响数量")
    p.set_defaults(func=cmd_set_layout)

    # ---- 文件操作类 (2) ----
    # create-file
    p = subparsers.add_parser("create-file", help="创建新 Excel 文件")
    p.add_argument("--file", required=True, help="文件路径")
    p.add_argument("--sheets", default=None, help='初始工作表名称列表 JSON（如 ["Sheet1","Sheet2"]）')
    p.set_defaults(func=cmd_create_file)

    # backup
    p = subparsers.add_parser("backup", help="备份与恢复")
    p.add_argument("--file", required=True, help="Excel 文件路径")
    p.add_argument("--operation", required=True, choices=["create", "list", "restore"], help="操作类型")
    p.add_argument("--backup-dir", default=None, help="备份目录")
    p.add_argument("--backup-id", default=None, help="restore 时指定备份文件路径")
    p.set_defaults(func=cmd_backup)

    # ---- self-update ----
    p = subparsers.add_parser("self-update", help="检查/更新 CLI 到最新版本")
    p.add_argument("--check", action="store_true", help="只检查版本，不更新")
    p.set_defaults(func=cmd_self_update)

    return parser


def main():
    """CLI 入口：解析参数并分发到对应子命令。"""
    parser = build_parser()
    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        return 1
    return args.func(args)
