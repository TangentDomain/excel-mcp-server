"""
Excel MCP Server - Excel比较模块

提供Excel文件和工作表比较功能
"""

import logging
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..models.types import (
    OperationResult, ComparisonResult, SheetComparison,
    CellDifference, ComparisonOptions, DifferenceType,
    RowDifference, StructuredSheetComparison, FieldDifference
)
from ..utils.validators import ExcelValidator
from ..utils.exceptions import SheetNotFoundError

logger = logging.getLogger(__name__)


class ExcelComparer:
    """Excel文件比较器"""

    def __init__(self, comparison_options: Optional[ComparisonOptions] = None):
        """
        初始化Excel比较器

        Args:
            comparison_options: 比较选项配置
        """
        self.options = comparison_options or ComparisonOptions()

    def compare_files(
        self,
        file1_path: str,
        file2_path: str,
        options: Optional[ComparisonOptions] = None
    ) -> OperationResult:
        """
        比较两个Excel文件

        Args:
            file1_path: 第一个Excel文件路径
            file2_path: 第二个Excel文件路径
            options: 比较选项，如果未提供则使用初始化时的选项

        Returns:
            OperationResult: 包含比较结果的操作结果
        """
        try:
            # 验证文件路径
            file1_path = ExcelValidator.validate_file_path(file1_path)
            file2_path = ExcelValidator.validate_file_path(file2_path)

            # 使用提供的选项或默认选项
            compare_options = options or self.options

            logger.info(f"开始比较文件: {file1_path} vs {file2_path}")

            # 加载两个Excel文件 - 比较操作使用只读模式
            workbook1 = load_workbook(file1_path, data_only=not compare_options.compare_formulas, read_only=True)
            workbook2 = load_workbook(file2_path, data_only=not compare_options.compare_formulas, read_only=True)

            # 比较文件结构
            structural_differences = self._compare_file_structure(workbook1, workbook2)

            # 比较工作表
            sheet_comparisons = []
            all_sheet_names = set(workbook1.sheetnames + workbook2.sheetnames)

            total_differences = 0
            for sheet_name in all_sheet_names:
                sheet_comparison = self._compare_sheets(
                    workbook1, workbook2, sheet_name, compare_options
                )
                sheet_comparisons.append(sheet_comparison)
                total_differences += sheet_comparison.total_differences

            # 判断文件是否完全相同
            identical = (total_differences == 0 and len(structural_differences) == 0)

            # 生成摘要
            summary = self._generate_comparison_summary(
                sheet_comparisons, structural_differences, total_differences
            )

            comparison_result = ComparisonResult(
                file1_path=file1_path,
                file2_path=file2_path,
                identical=identical,
                total_differences=total_differences,
                sheet_comparisons=sheet_comparisons,
                structural_differences=structural_differences,
                summary=summary
            )

            logger.info(f"文件比较完成，共发现 {total_differences} 处差异")

            return OperationResult(
                success=True,
                data=comparison_result,
                message=f"成功比较两个Excel文件，发现 {total_differences} 处差异",
                metadata={
                    'total_differences': total_differences,
                    'sheets_compared': len(sheet_comparisons),
                    'identical': identical
                }
            )

        except Exception as e:
            logger.error(f"文件比较失败: {e}")
            return OperationResult(
                success=False,
                error=f"文件比较失败: {str(e)}"
            )

    def compare_sheets(
        self,
        file1_path: str,
        sheet1_name: str,
        file2_path: str,
        sheet2_name: str,
        options: Optional[ComparisonOptions] = None
    ) -> OperationResult:
        """
        比较两个工作表

        Args:
            file1_path: 第一个Excel文件路径
            sheet1_name: 第一个工作表名称
            file2_path: 第二个Excel文件路径
            sheet2_name: 第二个工作表名称
            options: 比较选项

        Returns:
            OperationResult: 包含比较结果的操作结果
        """
        try:
            # 验证文件路径
            file1_path = ExcelValidator.validate_file_path(file1_path)
            file2_path = ExcelValidator.validate_file_path(file2_path)

            # 使用提供的选项或默认选项
            compare_options = options or self.options

            logger.info(f"开始比较工作表: {file1_path}[{sheet1_name}] vs {file2_path}[{sheet2_name}]")

            # 加载Excel文件 - 比较操作使用只读模式
            workbook1 = load_workbook(file1_path, data_only=not compare_options.compare_formulas, read_only=True)
            workbook2 = load_workbook(file2_path, data_only=not compare_options.compare_formulas, read_only=True)

            # 检查工作表是否存在
            if sheet1_name not in workbook1.sheetnames:
                raise SheetNotFoundError(f"工作表 '{sheet1_name}' 在文件 '{file1_path}' 中不存在")

            if sheet2_name not in workbook2.sheetnames:
                raise SheetNotFoundError(f"工作表 '{sheet2_name}' 在文件 '{file2_path}' 中不存在")

            # 获取工作表
            sheet1 = workbook1[sheet1_name]
            sheet2 = workbook2[sheet2_name]

            # 根据比较选项选择比较方式
            if compare_options.structured_comparison and compare_options.header_row:
                # 结构化数据比较
                row_differences = self._compare_structured_data(sheet1, sheet2, compare_options)
                structural_changes = self._get_sheet_structural_changes(sheet1, sheet2)

                # 创建StructuredDataComparison对象
                from ..models.types import StructuredDataComparison
                result_data = StructuredDataComparison(
                    sheet_name=f"{sheet1_name} vs {sheet2_name}",
                    exists_in_file1=True,
                    exists_in_file2=True,
                    row_differences=row_differences,
                    total_differences=len(row_differences),
                    structural_changes=structural_changes
                )
                message = f"成功比较两个工作表的结构化数据，发现 {len(row_differences)} 处差异"
            else:
                # 传统单元格级比较
                differences = self._compare_worksheet_data(sheet1, sheet2, compare_options)
                structural_changes = self._get_sheet_structural_changes(sheet1, sheet2)

                result_data = SheetComparison(
                    sheet_name=f"{sheet1_name} vs {sheet2_name}",
                    exists_in_file1=True,
                    exists_in_file2=True,
                    differences=differences,
                    total_differences=len(differences),
                    structural_changes=structural_changes
                )
                message = f"成功比较两个工作表，发现 {len(differences)} 处差异"

            logger.info(f"工作表比较完成，共发现 {result_data.total_differences} 处差异")

            return OperationResult(
                success=True,
                data=result_data,
                message=message,
                metadata={
                    'file1': file1_path,
                    'sheet1': sheet1_name,
                    'file2': file2_path,
                    'sheet2': sheet2_name,
                    'total_differences': result_data.total_differences,
                    'comparison_type': 'structured' if compare_options.structured_comparison else 'cell-by-cell'
                }
            )

        except Exception as e:
            logger.error(f"工作表比较失败: {e}")
            return OperationResult(
                success=False,
                error=f"工作表比较失败: {str(e)}"
            )

    # ==================== 分支 ====================
    # --- 文件结构比较 ---
    def _compare_file_structure(self, workbook1, workbook2) -> Dict[str, Any]:
        """比较文件结构差异"""
        structural_differences = {}

        # 比较工作表数量
        sheet_count1 = len(workbook1.sheetnames)
        sheet_count2 = len(workbook2.sheetnames)

        if sheet_count1 != sheet_count2:
            structural_differences['sheet_count'] = {
                'file1': sheet_count1,
                'file2': sheet_count2,
                'difference': sheet_count2 - sheet_count1
            }

        # 比较工作表名称
        sheets1 = set(workbook1.sheetnames)
        sheets2 = set(workbook2.sheetnames)

        added_sheets = sheets2 - sheets1
        removed_sheets = sheets1 - sheets2

        if added_sheets:
            structural_differences['added_sheets'] = list(added_sheets)

        if removed_sheets:
            structural_differences['removed_sheets'] = list(removed_sheets)

        return structural_differences

    def _compare_sheets(
        self,
        workbook1,
        workbook2,
        sheet_name: str,
        options: ComparisonOptions
    ) -> SheetComparison:
        """比较单个工作表"""
        exists_in_file1 = sheet_name in workbook1.sheetnames
        exists_in_file2 = sheet_name in workbook2.sheetnames

        differences = []
        structural_changes = {}

        if exists_in_file1 and exists_in_file2:
            # 两个文件都有这个工作表，进行详细比较
            sheet1 = workbook1[sheet_name]
            sheet2 = workbook2[sheet_name]
            differences = self._compare_worksheet_data(sheet1, sheet2, options)
            structural_changes = self._get_sheet_structural_changes(sheet1, sheet2)

        elif exists_in_file1 and not exists_in_file2:
            # 第二个文件中没有这个工作表
            differences.append(CellDifference(
                coordinate="SHEET",
                difference_type=DifferenceType.SHEET_REMOVED,
                sheet_name=sheet_name
            ))

        elif not exists_in_file1 and exists_in_file2:
            # 第一个文件中没有这个工作表
            differences.append(CellDifference(
                coordinate="SHEET",
                difference_type=DifferenceType.SHEET_ADDED,
                sheet_name=sheet_name
            ))

        return SheetComparison(
            sheet_name=sheet_name,
            exists_in_file1=exists_in_file1,
            exists_in_file2=exists_in_file2,
            differences=differences,
            total_differences=len(differences) if isinstance(differences, list) else 0,
            structural_changes=structural_changes
        )

    def _compare_worksheet_data(
        self,
        sheet1,
        sheet2,
        options: ComparisonOptions
    ) -> List[CellDifference]:
        """比较工作表数据"""
        # 如果启用结构化比较，使用ID-based比较
        if options.structured_comparison:
            return self._compare_structured_data(sheet1, sheet2, options)

        # 否则使用传统的单元格级比较
        return self._compare_cell_by_cell(sheet1, sheet2, options)

    def _compare_cell_by_cell(
        self,
        sheet1,
        sheet2,
        options: ComparisonOptions
    ) -> List[CellDifference]:
        """传统的逐单元格比较"""
        differences = []

        # 获取两个工作表的实际范围（空工作表max_row/max_column可能为None）
        max_row = max(sheet1.max_row or 0, sheet2.max_row or 0)
        max_col = max(sheet1.max_column or 0, sheet2.max_column or 0)

        # 遍历所有单元格
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                coord = f"{get_column_letter(col)}{row}"

                # 获取单元格
                cell1 = sheet1.cell(row=row, column=col)
                cell2 = sheet2.cell(row=row, column=col)

                # 比较单元格值
                if options.compare_values:
                    diff = self._compare_cell_values(cell1, cell2, coord, options)
                    if diff:
                        differences.append(diff)

                # 比较单元格格式
                if options.compare_formats:
                    format_diff = self._compare_cell_formats(cell1, cell2, coord)
                    if format_diff:
                        differences.append(format_diff)

        return differences

    def _compare_structured_data(
        self,
        sheet1,
        sheet2,
        options: ComparisonOptions
    ) -> List[RowDifference]:
        """结构化数据比较，返回行级差异"""
        # 获取表头
        headers1 = self._extract_headers(sheet1, options.header_row)
        headers2 = self._extract_headers(sheet2, options.header_row)

        # 解析数据行
        data_rows1 = self._extract_data_rows(sheet1, options.header_row, headers1, options.id_column)
        data_rows2 = self._extract_data_rows(sheet2, options.header_row, headers2, options.id_column)

        # 比较数据行，传入工作表名称
        sheet_name = getattr(sheet1, 'title', 'Unknown')
        differences = self._compare_data_rows(
            data_rows1, data_rows2, headers1, headers2, options, sheet_name
        )

        return differences

    def _compare_cell_values(
        self,
        cell1,
        cell2,
        coordinate: str,
        options: ComparisonOptions
    ) -> Optional[CellDifference]:
        """比较单元格值"""
        value1 = cell1.value
        value2 = cell2.value

        # 处理空单元格
        if options.ignore_empty_cells:
            if value1 is None and value2 is None:
                return None
            if value1 is None:
                value1 = ""
            if value2 is None:
                value2 = ""

        # 处理大小写敏感
        if isinstance(value1, str) and isinstance(value2, str) and not options.case_sensitive:
            value1 = value1.lower()
            value2 = value2.lower()

        # 比较值
        if value1 != value2:
            return CellDifference(
                coordinate=coordinate,
                difference_type=DifferenceType.VALUE_CHANGED,
                old_value=cell1.value,
                new_value=cell2.value
            )

        return None

    def _compare_cell_formats(self, cell1, cell2, coordinate: str) -> Optional[CellDifference]:
        """比较单元格格式"""
        # 简化的格式比较，可以根据需要扩展
        format1 = str(cell1.number_format) if cell1.number_format else ""
        format2 = str(cell2.number_format) if cell2.number_format else ""

        if format1 != format2:
            return CellDifference(
                coordinate=coordinate,
                difference_type=DifferenceType.FORMAT_CHANGED,
                old_format=format1,
                new_format=format2
            )

        return None

    def _get_sheet_structural_changes(self, sheet1, sheet2) -> Dict[str, Any]:
        """获取工作表结构变化"""
        structural_changes = {}

        # 比较行数和列数
        if sheet1.max_row != sheet2.max_row:
            structural_changes['max_row'] = {
                'sheet1': sheet1.max_row,
                'sheet2': sheet2.max_row,
                'difference': sheet2.max_row - sheet1.max_row
            }

        if sheet1.max_column != sheet2.max_column:
            structural_changes['max_column'] = {
                'sheet1': sheet1.max_column,
                'sheet2': sheet2.max_column,
                'difference': sheet2.max_column - sheet1.max_column
            }

        return structural_changes

    def _generate_comparison_summary(
        self,
        sheet_comparisons: List[SheetComparison],
        structural_differences: Dict[str, Any],
        total_differences: int
    ) -> str:
        """生成比较结果摘要"""
        if total_differences == 0 and len(structural_differences) == 0:
            return "两个Excel文件完全相同"

        summary_parts = []

        if total_differences > 0:
            summary_parts.append(f"发现 {total_differences} 处数据差异")

        if 'sheet_count' in structural_differences:
            sheet_diff = structural_differences['sheet_count']['difference']
            if sheet_diff > 0:
                summary_parts.append(f"增加了 {sheet_diff} 个工作表")
            else:
                summary_parts.append(f"减少了 {abs(sheet_diff)} 个工作表")

        if 'added_sheets' in structural_differences:
            added = len(structural_differences['added_sheets'])
            summary_parts.append(f"新增 {added} 个工作表")

        if 'removed_sheets' in structural_differences:
            removed = len(structural_differences['removed_sheets'])
            summary_parts.append(f"删除 {removed} 个工作表")

        # 统计有差异的工作表
        sheets_with_diff = sum(1 for sc in sheet_comparisons if sc.total_differences > 0)
        if sheets_with_diff > 0:
            summary_parts.append(f"{sheets_with_diff} 个工作表存在差异")

        return "；".join(summary_parts)

    # --- 结构化数据比较（旧版） ---
    def _compare_structured_data_old(
        self,
        sheet1,
        sheet2,
        options: ComparisonOptions
    ) -> StructuredSheetComparison:
        """比较结构化数据（表格化数据）"""
        # 提取表头
        headers1 = self._extract_headers(sheet1, options.header_row)
        headers2 = self._extract_headers(sheet2, options.header_row)

        # 比较表头差异
        header_differences = self._compare_headers(headers1, headers2)

        # 提取数据行
        data_rows1 = self._extract_data_rows(sheet1, options.header_row, headers1, options.id_column)
        data_rows2 = self._extract_data_rows(sheet2, options.header_row, headers2, options.id_column)

        # 比较数据行，传入工作表名称
        sheet_name = getattr(sheet1, 'title', 'Unknown')
        row_differences = self._compare_data_rows(data_rows1, data_rows2, headers1, headers2, options, sheet_name)

        # 统计差异（不再预计算统计信息，减少冗余）
        total_differences = len(row_differences) + len(header_differences)

        return StructuredSheetComparison(
            sheet_name=f"{sheet1.title} vs {sheet2.title}",
            exists_in_file1=True,
            exists_in_file2=True,
            headers1=headers1,
            headers2=headers2,
            header_differences=header_differences,
            row_differences=row_differences,
            total_differences=total_differences
            # 已移除冗余统计字段：identical_rows, modified_rows, added_rows, removed_rows
            # 客户端可以从 row_differences 按需计算这些统计信息
        )

    def _extract_headers(self, sheet, header_row: int) -> List[str]:
        """从工作表中提取表头"""
        headers = []
        for col in range(1, (sheet.max_column or 0) + 1):
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value is not None:
                headers.append(str(cell_value))
            else:
                headers.append(f"Column{col}")  # 为空表头生成默认名称
        return headers

    def _extract_data_rows(self, sheet, header_row: int, headers: List[str], id_column) -> Dict[Any, Dict]:
        """从工作表中提取数据行，以ID为键"""
        data_rows = {}

        # 确定ID列的索引
        id_col_index = self._get_id_column_index(id_column, headers)

        # 从表头行的下一行开始提取数据
        for row_num in range(header_row + 1, (sheet.max_row or 0) + 1):
            row_data = {}
            row_id = None

            # 提取行数据
            for col_index, header in enumerate(headers, 1):
                cell_value = sheet.cell(row=row_num, column=col_index).value
                row_data[header] = cell_value

                # 获取行ID
                if col_index == id_col_index:
                    row_id = cell_value

            # 如果没有ID列或ID为空，使用行号作为ID
            if row_id is None:
                row_id = f"Row{row_num}"

            # 检查是否为空行
            if not self._is_empty_row(row_data):
                data_rows[row_id] = {
                    'data': row_data,
                    'row_index': row_num
                }

        return data_rows

    def _get_id_column_index(self, id_column, headers: List[str]) -> Optional[int]:
        """获取ID列的索引"""
        if id_column is None:
            return None

        if isinstance(id_column, int):
            return id_column

        if isinstance(id_column, str):
            try:
                return headers.index(id_column) + 1  # 转换为1-based索引
            except ValueError:
                logger.warning(f"指定的ID列 '{id_column}' 在表头中不存在")
                return None

        return None

    def _compare_headers(self, headers1: List[str], headers2: List[str]) -> List[str]:
        """比较表头差异"""
        differences = []

        # 检查长度差异
        if len(headers1) != len(headers2):
            differences.append(f"表头数量不同: {len(headers1)} vs {len(headers2)}")

        # 检查表头内容差异
        max_len = max(len(headers1), len(headers2))
        for i in range(max_len):
            header1 = headers1[i] if i < len(headers1) else None
            header2 = headers2[i] if i < len(headers2) else None

            if header1 != header2:
                differences.append(f"列{i+1}: '{header1}' vs '{header2}'")

        return differences

    def _compare_data_rows(
        self,
        data_rows1: Dict,
        data_rows2: Dict,
        headers1: List[str],
        headers2: List[str],
        options: ComparisonOptions,
        sheet_name: str
    ) -> List[RowDifference]:
        """比较数据行（ID对象变化优化版）"""
        differences = []

        # 获取所有行ID
        all_ids = set(data_rows1.keys()) | set(data_rows2.keys())

        for row_id in all_ids:
            row1 = data_rows1.get(row_id)
            row2 = data_rows2.get(row_id)

            if row1 and row2:
                # 两个文件都有这一行，比较内容
                changed_differences, unchanged_differences = self._compare_row_data_detailed(
                    row1['data'], row2['data'], headers1, headers2, options
                )

                if changed_differences:
                    differences.append(RowDifference(
                        row_id=row_id,
                        difference_type=DifferenceType.ROW_MODIFIED,
                        row_index1=row1['row_index'],
                        row_index2=row2['row_index'],
                        sheet_name=sheet_name,
                        detailed_field_differences=changed_differences,
                        unchanged_field_differences=unchanged_differences
                    ))

            elif row1 and not row2:
                # 第二个文件中没有这一行
                differences.append(RowDifference(
                    row_id=row_id,
                    difference_type=DifferenceType.ROW_REMOVED,
                    row_index1=row1['row_index'],
                    row_index2=0,  # 不存在于第二个文件，使用0表示
                    sheet_name=sheet_name
                ))

            elif not row1 and row2:
                # 第一个文件中没有这一行
                differences.append(RowDifference(
                    row_id=row_id,
                    difference_type=DifferenceType.ROW_ADDED,
                    row_index1=0,  # 不存在于第一个文件，使用0表示
                    row_index2=row2['row_index'],
                    sheet_name=sheet_name
                ))

        return differences

    def _compare_row_data_detailed(
        self,
        row_data1: Dict,
        row_data2: Dict,
        headers1: List[str],
        headers2: List[str],
        options: ComparisonOptions
    ) -> tuple[List[FieldDifference], List[FieldDifference]]:
        """比较行数据，返回详细格式的差异和未变化字段（精简版 - 去除字符串格式冗余）"""
        changed_differences = []
        unchanged_differences = []

        # 获取所有字段名
        all_fields = set(headers1) | set(headers2)

        for field in all_fields:
            value1 = row_data1.get(field)
            value2 = row_data2.get(field)

            # 处理空值
            if options.ignore_empty_cells:
                if value1 is None:
                    value1 = ""
                if value2 is None:
                    value2 = ""

            # 处理大小写 - 用于比较但保持原始值用于展示
            compare_value1 = value1
            compare_value2 = value2
            if isinstance(value1, str) and isinstance(value2, str) and not options.case_sensitive:
                compare_value1 = value1.lower()
                compare_value2 = value2.lower()

            # 比较值
            if compare_value1 != compare_value2:
                # 生成变化字段的详细差异对象
                changed_diff = self._create_field_difference_simplified(field, value1, value2, options)
                changed_differences.append(changed_diff)
            else:
                # 生成未变化字段的信息（使用相同值作为old_value和new_value）
                unchanged_diff = FieldDifference(
                    field_name=field,
                    old_value=value1,
                    new_value=value1,  # 未变化，所以新旧值相同
                    change_type="unchanged"
                )
                unchanged_differences.append(unchanged_diff)

        return changed_differences, unchanged_differences

    def _create_field_difference_simplified(
        self,
        field_name: str,
        old_value: Any,
        new_value: Any,
        options: ComparisonOptions
    ) -> FieldDifference:
        """创建精简的字段差异对象（去除计算冗余）"""
        # 确定变化类型
        old_num = self._try_parse_number(old_value)
        new_num = self._try_parse_number(new_value)

        if old_num is not None and new_num is not None:
            change_type = "numeric_change"
        elif self._is_game_config_field(field_name):
            change_type = "config_change"
        else:
            change_type = "text_change"

        return FieldDifference(
            field_name=field_name,
            old_value=old_value,
            new_value=new_value,
            change_type=change_type
        )

    def _create_field_difference(
        self,
        field_name: str,
        old_value: Any,
        new_value: Any,
        options: ComparisonOptions
    ) -> FieldDifference:
        """创建详细的字段差异对象（已弃用 - 使用精简版）"""
        # 重定向到精简版本
        return self._create_field_difference_simplified(field_name, old_value, new_value, options)

    def _compare_row_data(
        self,
        row_data1: Dict,
        row_data2: Dict,
        headers1: List[str],
        headers2: List[str],
        options: ComparisonOptions
    ) -> List[str]:
        """比较单行数据的字段差异（游戏开发友好版，简化返回）- 已弃用，保留兼容性"""
        # 获取详细差异并转换为字符串格式（为了保持向后兼容性）
        detailed_differences = self._compare_row_data_detailed(row_data1, row_data2, headers1, headers2, options)

        # 从详细差异生成简化字符串列表
        field_differences = []
        for diff in detailed_differences:
            if options.show_numeric_changes and options.game_friendly_format:
                formatted = self._format_field_difference_for_summary(diff)
            else:
                formatted = f"{diff.field_name}: '{diff.old_value}' -> '{diff.new_value}'"
            field_differences.append(formatted)

        return field_differences

    def _format_game_friendly_difference(self, field: str, old_value: Any, new_value: Any) -> str:
        """格式化游戏开发友好的差异显示"""
        # 尝试解析为数字进行数值分析
        old_num = self._try_parse_number(old_value)
        new_num = self._try_parse_number(new_value)

        if old_num is not None and new_num is not None and old_num != 0:
            # 数值类型，显示变化量和百分比
            change = new_num - old_num
            change_percent = (change / old_num) * 100

            if change > 0:
                return f"🔺 {field}: {old_num} → {new_num} (+{change}, +{change_percent:.1f}%)"
            else:
                return f"🔻 {field}: {old_num} → {new_num} ({change}, {change_percent:.1f}%)"
        else:
            # 非数值类型或特殊情况，使用标准格式
            if self._is_game_config_field(field):
                return f"🔄 {field}: '{old_value}' → '{new_value}'"
            else:
                return f"{field}: '{old_value}' → '{new_value}'"

    def _try_parse_number(self, value: Any) -> Optional[float]:
        """尝试将值解析为数字"""
        if isinstance(value, (int, float)):
            return float(value)

        if isinstance(value, str):
            try:
                # 移除可能的单位符号和空格
                clean_value = value.strip().replace('%', '').replace(',', '')
                return float(clean_value)
            except ValueError:
                pass

        return None

    def _is_game_config_field(self, field: str) -> bool:
        """判断是否是常见的游戏配置字段"""
        game_fields = {
            '名称', 'name', '技能名', '装备名', '道具名', '怪物名',
            '描述', 'description', 'desc', '说明',
            '品质', 'quality', '等级', 'level', 'lv',
            '类型', 'type', '分类', 'category'
        }
        return field.lower() in [f.lower() for f in game_fields]

    def _extract_object_name(self, row_data: Dict[str, Any], headers: List[str]) -> str:
        """从行数据中提取对象名称"""
        # 特殊处理：如果是表头行（ID行），返回专门的标识
        first_col_value = row_data.get(headers[0]) if headers else None
        if first_col_value and str(first_col_value).strip().upper() == "ID":
            return "表头定义"

        # 常见的名称字段
        name_fields = ['名称', 'name', '技能名', '装备名', '道具名', '怪物名', '称号', 'title', '备注']

        # 优先查找专门的名称字段
        for header in headers:
            if header and header.lower() in [f.lower() for f in name_fields]:
                name_value = row_data.get(header)
                if name_value and str(name_value).strip():
                    return str(name_value).strip()

        # 如果没有找到专门的名称字段，尝试使用第2列（通常是名称列）
        if len(headers) >= 2:
            second_col_value = row_data.get(headers[1])
            if second_col_value and str(second_col_value).strip():
                return str(second_col_value).strip()

        return "未知对象"

    def _generate_id_based_summary_from_detailed(
        self,
        row_id: Any,
        object_name: str,
        detailed_differences: List[FieldDifference],
        game_friendly: bool
    ) -> str:
        """从详细差异生成ID对象的变化摘要（精简版）"""
        if not game_friendly:
            return f"ID {row_id} modified: {len(detailed_differences)} fields changed"

        # 游戏开发友好格式 - 从详细差异对象生成格式化显示
        if len(detailed_differences) == 1:
            diff = detailed_differences[0]
            formatted_diff = self._format_field_difference_for_summary(diff)
            return f"🔧 ID {row_id} ({object_name}): {formatted_diff}"
        else:
            # 多个属性变化
            summary = f"🔧 ID {row_id} ({object_name}): {len(detailed_differences)}个属性变化"

            # 按行显示变化，每行最多2个
            for i in range(0, len(detailed_differences), 2):
                line_diffs = detailed_differences[i:i+2]
                formatted_diffs = [self._format_field_difference_for_summary(d) for d in line_diffs]
                summary += f"\n   • {' | '.join(formatted_diffs)}"

            return summary

    def _format_field_difference_for_summary(self, diff: FieldDifference) -> str:
        """为摘要格式化单个字段差异"""
        field = diff.field_name
        old_val = diff.old_value
        new_val = diff.new_value

        # 尝试数值格式化
        if diff.change_type == "numeric_change":
            old_num = self._try_parse_number(old_val)
            new_num = self._try_parse_number(new_val)
            if old_num is not None and new_num is not None and old_num != 0:
                change = new_num - old_num
                change_percent = (change / old_num) * 100
                if change > 0:
                    return f"🔺 {field}: {old_val} → {new_val} (+{change}, +{change_percent:.1f}%)"
                else:
                    return f"🔻 {field}: {old_val} → {new_val} ({change}, {change_percent:.1f}%)"

        # 配置变化或文本变化
        if self._is_game_config_field(field):
            return f"⚙️ {field}: {old_val} → {new_val}"
        else:
            return f"{field}: '{old_val}' → '{new_val}'"

    def _generate_id_based_summary(
        self,
        row_id: Any,
        object_name: str,
        field_differences: List[str],
        game_friendly: bool
    ) -> str:
        """生成ID对象的变化摘要（已弃用 - 保留兼容性）"""
        if not game_friendly:
            return f"ID {row_id} modified: {len(field_differences)} fields changed"

        # 游戏开发友好格式 - 显示所有属性变化
        if len(field_differences) == 1:
            return f"🔧 ID {row_id} ({object_name}): {field_differences[0]}"
        else:
            # 显示所有属性变化，但分行显示以便阅读
            summary = f"🔧 ID {row_id} ({object_name}): {len(field_differences)}个属性变化"

            # 将所有变化按行显示，每行最多显示2个变化以保持可读性
            for i in range(0, len(field_differences), 2):
                line_changes = field_differences[i:i+2]
                summary += f"\n   • {' | '.join(line_changes)}"

            return summary

    def _is_empty_row(self, row_data: Dict) -> bool:
        """检查行是否为空"""
        return all(value is None or value == "" for value in row_data.values())
