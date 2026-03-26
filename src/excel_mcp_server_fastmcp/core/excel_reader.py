"""
Excel MCP Server - Excel读取模块

提供Excel文件读取功能
使用python-calamine（Rust引擎）加速纯数据读取，openpyxl作为格式化读取的后备方案
"""

import logging
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, column_index_from_string, get_column_letter

from ..models.types import (
    RangeInfo, RangeType, CellInfo, SheetInfo,
    ExcelData, ExcelDimensions, OperationResult
)
from ..utils.validators import ExcelValidator
from ..utils.parsers import RangeParser
from ..utils.exceptions import SheetNotFoundError

logger = logging.getLogger(__name__)

# 尝试导入calamine，不可用时降级到openpyxl
try:
    from python_calamine import CalamineWorkbook
    _HAS_CALAMINE = True
except ImportError:
    _HAS_CALAMINE = False
    logger.debug("python-calamine未安装，读取性能将受影响")


class ExcelReader:
    """Excel文件读取器（calamine加速 + openpyxl后备）"""

    def __init__(self, file_path: str):
        """
        初始化Excel读取器

        Args:
            file_path: Excel文件路径
        """
        self.file_path = ExcelValidator.validate_file_path(file_path)
        self._workbook_cache = {}  # 缓存不同参数的openpyxl工作簿
        self._calamine_wb = None   # calamine工作簿缓存（只缓存一个，因为不支持参数变体）

    def _get_calamine_workbook(self):
        """
        获取缓存的calamine工作簿或加载新工作簿

        Returns:
            CalamineWorkbook: calamine工作簿对象
        """
        if self._calamine_wb is None and _HAS_CALAMINE:
            self._calamine_wb = CalamineWorkbook.from_path(self.file_path)
        return self._calamine_wb

    def _get_workbook(self, read_only: bool = True, data_only: bool = False):
        """
        获取缓存的工作簿或加载新工作簿（openpyxl）

        Args:
            read_only: 是否以只读模式打开
            data_only: 是否只读取值（不包含公式）

        Returns:
            Workbook: openpyxl工作簿对象
        """
        cache_key = (read_only, data_only)
        if cache_key not in self._workbook_cache:
            self._workbook_cache[cache_key] = load_workbook(
                self.file_path,
                read_only=read_only,
                data_only=data_only
            )
        return self._workbook_cache[cache_key]

    def close(self):
        """关闭所有缓存的工作簿"""
        for workbook in self._workbook_cache.values():
            if workbook is not None:
                workbook.close()
        self._workbook_cache.clear()
        self._calamine_wb = None

    def list_sheets(self) -> OperationResult:
        """
        获取Excel文件中所有工作表的信息

        Returns:
            OperationResult: 包含所有工作表信息的结果
        """
        # calamine快速路径
        if _HAS_CALAMINE:
            try:
                wb = self._get_calamine_workbook()
                sheets_info = []
                for i, name in enumerate(wb.sheet_names):
                    ws = wb.get_sheet_by_name(name)
                    sheet_info = SheetInfo(
                        index=i,
                        name=name,
                        max_row=ws.height,
                        max_column=ws.width,
                        max_column_letter=get_column_letter(ws.width)
                    )
                    sheets_info.append(sheet_info)

                return OperationResult(
                    success=True,
                    data=sheets_info,
                    metadata={
                        'file_path': self.file_path,
                        'total_sheets': len(sheets_info)
                    }
                )
            except Exception as e:
                logger.debug(f"calamine list_sheets失败，回退openpyxl: {e}")

        # openpyxl后备路径
        try:
            workbook = self._get_workbook(read_only=True)

            sheets_info = []
            for i, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]

                sheet_info = SheetInfo(
                    index=i,
                    name=sheet_name,
                    max_row=sheet.max_row,
                    max_column=sheet.max_column,
                    max_column_letter=get_column_letter(sheet.max_column)
                )
                sheets_info.append(sheet_info)

            return OperationResult(
                success=True,
                data=sheets_info,
                metadata={
                    'file_path': self.file_path,
                    'total_sheets': len(sheets_info)
                }
            )

        except Exception as e:
            logger.error(f"获取工作表列表失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def get_range(
        self,
        range_expression: str,
        include_formatting: bool = False
    ) -> OperationResult:
        """
        获取Excel文件中指定范围的数据

        Args:
            range_expression: 范围表达式
            include_formatting: 是否包含格式信息

        Returns:
            OperationResult: 包含范围数据的结果
        """
        # 解析范围表达式
        range_info = RangeParser.parse_range_expression(range_expression)

        # calamine快速路径（仅限纯数据读取）
        if not include_formatting and _HAS_CALAMINE:
            try:
                return self._get_range_calamine(range_info, range_expression)
            except Exception as e:
                logger.debug(f"calamine get_range失败，回退openpyxl: {e}")

        # openpyxl路径（格式化读取或calamine不可用）
        return self._get_range_openpyxl(range_info, range_expression, include_formatting)

    def _get_range_calamine(self, range_info: RangeInfo, range_expression: str) -> OperationResult:
        """使用calamine快速读取范围数据（纯值，无格式）"""
        wb = self._get_calamine_workbook()
        sheet_name = range_info.sheet_name

        if not sheet_name or not sheet_name.strip():
            raise SheetNotFoundError("工作表名称不能为空，必须明确指定工作表")

        if sheet_name not in wb.sheet_names:
            raise SheetNotFoundError(
                f"工作表不存在: {sheet_name}，可用工作表: {', '.join(wb.sheet_names)}"
            )

        ws = wb.get_sheet_by_name(sheet_name)
        all_rows = list(ws.to_python())

        data, dimensions = self._extract_range_from_rows(
            all_rows, range_info
        )

        return OperationResult(
            success=True,
            data=data,
            metadata={
                'file_path': self.file_path,
                'range': range_expression,
                'range_type': range_info.range_type.value,
                'sheet_name': sheet_name,
                'dimensions': dimensions.__dict__
            }
        )

    def _extract_range_from_rows(
        self,
        all_rows: List[List],
        range_info: RangeInfo
    ) -> tuple[ExcelData, ExcelDimensions]:
        """从已加载的行数据中提取指定范围（0-based索引）"""
        data = []

        if range_info.range_type in [RangeType.ROW_RANGE, RangeType.SINGLE_ROW]:
            row_parts = range_info.cell_range.split(':')
            start_row = int(row_parts[0]) - 1  # 转为0-based
            end_row = int(row_parts[1]) - 1

            for r in range(start_row, end_row + 1):
                row_values = all_rows[r] if r < len(all_rows) else []
                row_data = []
                for c, val in enumerate(row_values):
                    col_letter = get_column_letter(c + 1)
                    cell_info = CellInfo(
                        coordinate=f"{col_letter}{r + 1}",
                        value=self._normalize_calamine_value(val)
                    )
                    row_data.append(cell_info)
                data.append(row_data)

            max_col = max((len(row) for row in data), default=0)
            dimensions = ExcelDimensions(
                rows=end_row - start_row + 1,
                columns=max_col,
                start_row=start_row + 1,
                start_column=1
            )

        elif range_info.range_type in [RangeType.COLUMN_RANGE, RangeType.SINGLE_COLUMN]:
            col_parts = range_info.cell_range.split(':')
            start_col = column_index_from_string(col_parts[0]) - 1  # 转为0-based
            end_col = column_index_from_string(col_parts[1]) - 1

            for r, row_values in enumerate(all_rows):
                row_data = []
                for c in range(start_col, end_col + 1):
                    col_letter = get_column_letter(c + 1)
                    val = row_values[c] if c < len(row_values) else None
                    cell_info = CellInfo(
                        coordinate=f"{col_letter}{r + 1}",
                        value=self._normalize_calamine_value(val)
                    )
                    row_data.append(cell_info)
                data.append(row_data)

            dimensions = ExcelDimensions(
                rows=len(all_rows),
                columns=end_col - start_col + 1,
                start_row=1,
                start_column=start_col + 1
            )

        else:
            # CELL_RANGE
            min_col, min_row, max_col, max_row = range_boundaries(range_info.cell_range)

            for r in range(min_row - 1, max_row):  # 转为0-based
                row_values = all_rows[r] if r < len(all_rows) else []
                row_data = []
                for c in range(min_col - 1, max_col):  # 转为0-based
                    col_letter = get_column_letter(c + 1)
                    val = row_values[c] if c < len(row_values) else None
                    cell_info = CellInfo(
                        coordinate=f"{col_letter}{r + 1}",
                        value=self._normalize_calamine_value(val)
                    )
                    row_data.append(cell_info)
                data.append(row_data)

            dimensions = ExcelDimensions(
                rows=max_row - min_row + 1,
                columns=max_col - min_col + 1,
                start_row=min_row,
                start_column=min_col
            )

        return data, dimensions

    @staticmethod
    def _normalize_calamine_value(val):
        """归一化calamine值：整型浮点数转整数（25.0→25），与openpyxl行为一致"""
        if isinstance(val, float) and val == int(val) and not (val != val):  # 排除NaN
            return int(val)
        return val

    def _get_range_openpyxl(self, range_info: RangeInfo, range_expression: str, include_formatting: bool) -> OperationResult:
        """使用openpyxl读取范围数据（支持格式化）"""
        try:
            workbook = self._get_workbook(read_only=True, data_only=True)
            sheet = self._get_worksheet_openpyxl(workbook, range_info.sheet_name)

            data, dimensions = self._get_range_data_openpyxl(
                sheet, range_info, include_formatting
            )

            return OperationResult(
                success=True,
                data=data,
                metadata={
                    'file_path': self.file_path,
                    'range': range_expression,
                    'range_type': range_info.range_type.value,
                    'sheet_name': sheet.title,
                    'dimensions': dimensions.__dict__
                }
            )

        except Exception as e:
            logger.error(f"获取范围数据失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def _get_worksheet_openpyxl(self, workbook, sheet_name: Optional[str]):
        """获取openpyxl工作表"""
        if not sheet_name or not sheet_name.strip():
            raise SheetNotFoundError(f"工作表名称不能为空，必须明确指定工作表")

        if not workbook.sheetnames:
            raise SheetNotFoundError(f"Excel文件中没有任何工作表")

        if sheet_name not in workbook.sheetnames:
            raise SheetNotFoundError(f"工作表不存在: {sheet_name}，可用工作表: {', '.join(workbook.sheetnames)}")

        return workbook[sheet_name]

    # 向后兼容别名（excel_operations.py直接调用此方法）
    _get_worksheet = _get_worksheet_openpyxl

    def _get_range_data_openpyxl(
        self,
        sheet,
        range_info: RangeInfo,
        include_formatting: bool
    ) -> tuple[ExcelData, ExcelDimensions]:
        """根据范围类型获取数据（openpyxl）"""

        if range_info.range_type in [RangeType.ROW_RANGE, RangeType.SINGLE_ROW]:
            return self._get_row_data(sheet, range_info, include_formatting)
        elif range_info.range_type in [RangeType.COLUMN_RANGE, RangeType.SINGLE_COLUMN]:
            return self._get_column_data(sheet, range_info, include_formatting)
        else:
            return self._get_cell_range_data(sheet, range_info, include_formatting)

    def _get_row_data(
        self,
        sheet,
        range_info: RangeInfo,
        include_formatting: bool
    ) -> tuple[ExcelData, ExcelDimensions]:
        """获取行范围数据"""
        row_parts = range_info.cell_range.split(':')
        start_row = int(row_parts[0])
        end_row = int(row_parts[1])

        # 优化：使用iter_rows而不是手动遍历所有列
        data = []
        max_col_found = 0

        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=False):
            row_data = []
            last_data_col = 0

            for col_idx, cell in enumerate(row, 1):
                if cell.value is not None:
                    last_data_col = col_idx

            if last_data_col == 0:
                last_data_col = 1

            for col_idx in range(1, last_data_col + 1):
                if col_idx - 1 < len(row):
                    cell = row[col_idx - 1]
                else:
                    cell = sheet.cell(row=start_row + len(data), column=col_idx)
                cell_info = self._create_cell_info(cell, include_formatting)
                row_data.append(cell_info)

            data.append(row_data)
            max_col_found = max(max_col_found, last_data_col)

        dimensions = ExcelDimensions(
            rows=end_row - start_row + 1,
            columns=max_col_found,
            start_row=start_row,
            start_column=1
        )

        return data, dimensions

    def _get_column_data(
        self,
        sheet,
        range_info: RangeInfo,
        include_formatting: bool
    ) -> tuple[ExcelData, ExcelDimensions]:
        """获取列范围数据"""
        col_parts = range_info.cell_range.split(':')
        start_col = column_index_from_string(col_parts[0])
        end_col = column_index_from_string(col_parts[1])
        max_row = sheet.max_row

        data = []
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(start_col, end_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell_info = self._create_cell_info(cell, include_formatting)
                row_data.append(cell_info)
            data.append(row_data)

        dimensions = ExcelDimensions(
            rows=max_row,
            columns=end_col - start_col + 1,
            start_row=1,
            start_column=start_col
        )

        return data, dimensions

    def _get_cell_range_data(
        self,
        sheet,
        range_info: RangeInfo,
        include_formatting: bool
    ) -> tuple[ExcelData, ExcelDimensions]:
        """获取单元格范围数据"""
        min_col, min_row, max_col, max_row = range_boundaries(range_info.cell_range)

        data = []
        for row_idx in range(min_row, max_row + 1):
            row_data = []
            for col_idx in range(min_col, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell_info = self._create_cell_info(cell, include_formatting)
                row_data.append(cell_info)
            data.append(row_data)

        dimensions = ExcelDimensions(
            rows=max_row - min_row + 1,
            columns=max_col - min_col + 1,
            start_row=min_row,
            start_column=min_col
        )

        return data, dimensions

    def _create_cell_info(self, cell, include_formatting: bool) -> CellInfo:
        """创建单元格信息对象"""
        if hasattr(cell, 'coordinate'):
            coordinate = cell.coordinate
        else:
            coordinate = f"{get_column_letter(cell.column)}{cell.row}" if hasattr(cell, 'row') and hasattr(cell, 'column') else "A1"

        cell_info = CellInfo(
            coordinate=coordinate,
            value=cell.value
        )

        if include_formatting and hasattr(cell, 'data_type'):
            cell_info.data_type = cell.data_type
            cell_info.number_format = getattr(cell, 'number_format', None)
            cell_info.font = str(cell.font) if hasattr(cell, 'font') and cell.font else None
            cell_info.fill = str(cell.fill) if hasattr(cell, 'fill') and cell.fill else None

        return cell_info
