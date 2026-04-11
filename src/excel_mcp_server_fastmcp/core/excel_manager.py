"""
Excel MCP Server - Excel管理模块

提供Excel文件和工作表管理功能
"""

import logging
import os
from typing import List, Optional
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from ..models.types import SheetInfo, OperationResult
from ..utils.validators import ExcelValidator
from ..utils.exceptions import SheetNotFoundError, DataValidationError

# 设置日志级别为INFO以便调试
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)


class ExcelManager:
    """Excel文件和工作表管理器"""

    @classmethod
    def create_file(
        cls,
        file_path: str,
        sheet_names: Optional[List[str]] = None
    ) -> OperationResult:
        """
        创建新的Excel文件

        Args:
            file_path: 要创建的Excel文件路径
            sheet_names: 工作表名称列表

        Returns:
            OperationResult: 创建操作的结果
        """
        try:
            # 验证路径，默认允许覆盖已存在的文件
            validated_path = ExcelValidator.validate_file_for_creation(file_path, overwrite=True)

            # 确保目录存在
            Path(validated_path).parent.mkdir(parents=True, exist_ok=True)

            # 使用 write_only 模式创建工作簿（减少内存占用）
            # write_only 模式：不维护单元格内存模型，直接流式写入磁盘
            workbook = Workbook(write_only=True)

            # 处理工作表
            if sheet_names:
                # write_only 模式不创建默认工作表，直接创建指定的工作表
                created_sheets = []
                for i, sheet_name in enumerate(sheet_names):
                    if not sheet_name or not sheet_name.strip():
                        raise DataValidationError(
                            f"工作表名称不能为空: 索引 {i}",
                            "工作表名称不能为空白字符串",
                            "请为所有工作表提供有效的名称"
                        )

                    sheet = workbook.create_sheet(title=sheet_name.strip())
                    created_sheets.append(SheetInfo(
                        index=i,
                        name=sheet.title,
                        max_row=1,
                        max_column=1,
                        max_column_letter='A'
                    ))
            else:
                # write_only 模式需要手动创建默认工作表
                sheet = workbook.create_sheet(title='Sheet1')
                created_sheets = [SheetInfo(
                    index=0,
                    name='Sheet1',
                    max_row=1,
                    max_column=1,
                    max_column_letter='A'
                )]

            # 保存文件
            workbook.save(validated_path)

            return OperationResult(
                success=True,
                data=created_sheets,
                message=f"成功创建Excel文件，包含{len(created_sheets)}个工作表",
                metadata={
                    'file_path': validated_path,
                    'total_sheets': len(created_sheets)
                }
            )

        except Exception as e:
            logger.error(f"创建Excel文件失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def __init__(self, file_path: str):
        """
        初始化Excel管理器

        Args:
            file_path: Excel文件路径
        """
        self.file_path = ExcelValidator.validate_file_path(file_path)

    def list_sheets(self) -> OperationResult:
        """
        列出所有工作表

        Returns:
            OperationResult: 包含工作表信息列表的结果
        """
        from .excel_reader import ExcelReader
        reader = ExcelReader(self.file_path)
        return reader.list_sheets()

    def create_sheet(
        self,
        sheet_name: str,
        index: Optional[int] = None
    ) -> OperationResult:
        """
        在Excel文件中创建新工作表，支持中文字符

        Args:
            sheet_name: 新工作表名称
            index: 插入位置索引

        Returns:
            OperationResult: 创建操作的结果
        """
        try:
            # 验证工作表名称
            if not sheet_name or not sheet_name.strip():
                raise DataValidationError("工作表名称不能为空")

            # 验证工作表名称合法性
            sheet_name = sheet_name.strip()
            self._validate_sheet_name(sheet_name)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 检查工作表名称是否已存在
            if sheet_name in workbook.sheetnames:
                raise DataValidationError(f"工作表名称已存在: {sheet_name}")

            # 验证索引
            total_sheets = len(workbook.sheetnames)
            if index is not None:
                if index < 0 or index > total_sheets:
                    raise DataValidationError(f"索引超出范围: {index}，应在 0-{total_sheets} 之间")

            # 创建新工作表，使用安全的编码处理
            try:
                new_sheet = workbook.create_sheet(title=sheet_name, index=index)
            except Exception as sheet_error:
                # 如果直接创建失败，尝试使用ASCII兼容的名称
                logger.warning(f"创建工作表失败，尝试备用方法: {sheet_error}")
                fallback_name = self._create_fallback_name(sheet_name, workbook.sheetnames)
                new_sheet = workbook.create_sheet(title=fallback_name, index=index)
                logger.info(f"使用备用名称创建工作表: {fallback_name}")
                sheet_name = fallback_name

            # 保存文件
            workbook.save(self.file_path)

            # 获取新工作表信息
            sheet_info = SheetInfo(
                index=workbook.sheetnames.index(sheet_name),
                name=new_sheet.title,
                max_row=1,
                max_column=1,
                max_column_letter='A'
            )

            return OperationResult(
                success=True,
                data=sheet_info,
                message=f"成功创建工作表: {sheet_name}",
                metadata={
                    'file_path': self.file_path,
                    'total_sheets': len(workbook.sheetnames),
                    'all_sheets': workbook.sheetnames
                }
            )

        except Exception as e:
            logger.error(f"创建工作表失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def _validate_sheet_name(self, name: str) -> None:
        """
        验证工作表名称是否合法，不合法则抛出 DataValidationError

        Excel工作表名称限制：
        1. 不能超过31个字符
        2. 不能包含: / \\ ? * [ ] :
        3. 不能为空

        Args:
            name: 待验证的工作表名称

        Raises:
            DataValidationError: 名称不合法时抛出，包含具体原因
        """
        import re

        if not name or not name.strip():
            raise DataValidationError("工作表名称不能为空")

        # 检查非法字符
        invalid_chars = r'[/\\?*\[\]:]'
        illegal = re.findall(invalid_chars, name)
        if illegal:
            chars_str = ', '.join(sorted(set(illegal)))
            raise DataValidationError(
                f"工作表名称包含非法字符: {chars_str}。"
                f"Excel工作表名称不能包含以下字符: / \\ ? * [ ] :"
            )

        # 检查长度限制（31字符是Excel硬限制）
        if len(name) > 31:
            raise DataValidationError(
                f"工作表名称过长: {len(name)}个字符，"
                f"Excel限制最多31个字符。请缩短名称或使用缩写。"
            )

    def _sanitize_sheet_name(self, name: str) -> str:
        """
        静默清理工作表名称，用于系统自动生成的名称（如复制工作表时的默认名）。
        会替换非法字符、截断超长名称，不抛出异常。

        Args:
            name: 原始工作表名称

        Returns:
            str: 清理后的名称
        """
        import re

        # 替换非法字符
        invalid_chars = r'[/\\?*\[\]:]'
        name = re.sub(invalid_chars, '_', name)

        # 移除首尾空白
        name = name.strip()

        # 截断超长名称（留3字符给省略号）
        if len(name) > 31:
            name = name[:28] + "..."

        # 确保名称不为空
        if not name:
            name = "Sheet"

        return name

    def _create_fallback_name(self, original_name: str, existing_names: list) -> str:
        """
        创建备用工作表名称

        Args:
            original_name: 原始名称
            existing_names: 已存在的名称列表

        Returns:
            str: 备用名称
        """
        import re

        # 尝试创建ASCII兼容的名称
        fallback_base = "Sheet"

        # 尝试从原始名称中提取英文字符
        ascii_chars = re.findall(r'[a-zA-Z0-9]', original_name)
        if ascii_chars:
            fallback_base = ''.join(ascii_chars)[:10]  # 最多取10个字符

        # 确保名称唯一
        counter = 1
        fallback_name = fallback_base
        while fallback_name in existing_names:
            fallback_name = f"{fallback_base}_{counter}"
            counter += 1

        return fallback_name

    def delete_sheet(self, sheet_name: str) -> OperationResult:
        """
        删除Excel文件中的工作表

        Args:
            sheet_name: 要删除的工作表名称

        Returns:
            OperationResult: 删除操作的结果
        """
        try:
            # 验证工作表名称
            if not sheet_name or not sheet_name.strip():
                raise DataValidationError("工作表名称不能为空")

            sheet_name = sheet_name.strip()

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 检查工作表是否存在
            if sheet_name not in workbook.sheetnames:
                raise SheetNotFoundError(f"工作表不存在: {sheet_name}")

            # 检查是否为最后一个工作表
            if len(workbook.sheetnames) <= 1:
                raise DataValidationError("无法删除最后一个工作表，Excel文件至少需要一个工作表")

            # 记录删除前的信息
            deleted_sheet_index = workbook.sheetnames.index(sheet_name)
            deleted_sheet_name = sheet_name  # 保存要删除的工作表名称

            logger.info(f"准备删除工作表: {deleted_sheet_name}, 索引: {deleted_sheet_index}")
            logger.info(f"删除前工作表列表: {workbook.sheetnames}")

            # 删除工作表
            workbook.remove(workbook[sheet_name])

            logger.info(f"删除后工作表列表: {workbook.sheetnames}")

            # 如果删除的是第一个工作表，设置新的第一个为活动工作表
            if deleted_sheet_index == 0 and workbook.sheetnames:
                workbook.active = 0
            elif deleted_sheet_index < len(workbook.sheetnames):
                workbook.active = deleted_sheet_index
            else:
                workbook.active = deleted_sheet_index - 1

            # 保存文件
            workbook.save(self.file_path)

            # 返回更新后的工作表信息
            remaining_sheet_infos = []
            for i, sheet_name_iter in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name_iter]
                remaining_sheet_infos.append(SheetInfo(
                    index=i,
                    name=sheet_name_iter,
                    max_row=sheet.max_row,
                    max_column=sheet.max_column,
                    max_column_letter=get_column_letter(sheet.max_column)
                ))

            return OperationResult(
                success=True,
                data=remaining_sheet_infos,  # 添加data字段
                message=f"成功删除工作表: {deleted_sheet_name}",
                metadata={
                    'file_path': self.file_path,
                    'deleted_sheet': deleted_sheet_name,
                    'deleted_index': deleted_sheet_index,
                    'new_active_sheet': workbook.active.title,
                    'remaining_sheets': workbook.sheetnames,
                    'total_sheets': len(workbook.sheetnames)
                }
            )

        except Exception as e:
            logger.error(f"删除工作表失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def rename_sheet(self, old_name: str, new_name: str) -> OperationResult:
        """
        重命名Excel文件中的工作表

        Args:
            old_name: 原工作表名称
            new_name: 新工作表名称

        Returns:
            OperationResult: 重命名操作的结果
        """
        try:
            # 验证工作表名称
            if not old_name or not old_name.strip():
                raise DataValidationError("原工作表名称不能为空")
            if not new_name or not new_name.strip():
                raise DataValidationError("新工作表名称不能为空")

            old_name = old_name.strip()
            new_name = new_name.strip()

            # 验证新名称合法性
            self._validate_sheet_name(new_name)

            if old_name == new_name:
                raise DataValidationError("新名称与原名称相同，无需重命名")

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 检查原工作表是否存在
            if old_name not in workbook.sheetnames:
                raise SheetNotFoundError(f"原工作表不存在: {old_name}")

            # 检查新名称是否已存在
            if new_name in workbook.sheetnames:
                raise DataValidationError(f"新工作表名称已存在: {new_name}")

            # 获取工作表
            sheet = workbook[old_name]
            old_index = workbook.sheetnames.index(old_name)

            # 重命名工作表
            sheet.title = new_name

            # 保存文件
            workbook.save(self.file_path)

            # 构建重命名后的工作表信息
            renamed_sheet_info = SheetInfo(
                index=old_index,
                name=new_name,
                max_row=sheet.max_row,
                max_column=sheet.max_column,
                max_column_letter=get_column_letter(sheet.max_column) if sheet.max_column > 0 else 'A'
            )

            return OperationResult(
                success=True,
                message=f"成功将工作表 '{old_name}' 重命名为 '{new_name}'",
                data=renamed_sheet_info,
                metadata={
                    'file_path': self.file_path,
                    'old_name': old_name,
                    'new_name': new_name,
                    'sheet_index': old_index,
                    'all_sheets': workbook.sheetnames
                }
            )

        except Exception as e:
            logger.error(f"重命名工作表失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def copy_sheet(
        self,
        source_name: str,
        new_name: Optional[str] = None,
        index: Optional[int] = None,
        streaming: bool = True
    ) -> OperationResult:
        """
        复制工作表到同一文件中（含数据和格式）

        Args:
            source_name: 源工作表名称
            new_name: 新工作表名称（为空则自动生成 "源表名_副本"）
            index: 插入位置索引（None表示追加到末尾）
            streaming: 是否使用流式复制（默认True，大文件性能更好）

        Returns:
            OperationResult: 复制操作的结果
        """
        try:
            if not source_name or not source_name.strip():
                raise DataValidationError("源工作表名称不能为空")

            source_name = source_name.strip()

            # 流式复制路径：读取数据后重建工作表，大文件性能更好
            if streaming:
                from .streaming_writer import StreamingWriter
                if StreamingWriter.is_available():
                    try:
                        return self._copy_sheet_streaming(source_name, new_name, index)
                    except Exception as streaming_err:
                        logger.warning(f"流式复制工作表失败，降级到openpyxl: {streaming_err}")

            workbook = load_workbook(self.file_path)

            if source_name not in workbook.sheetnames:
                available = ', '.join(workbook.sheetnames)
                raise SheetNotFoundError(f"工作表不存在: {source_name}（可用: {available}）")

            source_sheet = workbook[source_name]

            # 自动生成新名称
            if not new_name or not new_name.strip():
                new_name = f"{source_name}_副本"
            else:
                new_name = new_name.strip()

            # 处理名称冲突
            base_name = new_name
            counter = 1
            while new_name in workbook.sheetnames:
                new_name = f"{base_name}_{counter}"
                counter += 1
                if counter > 100:
                    raise DataValidationError(f"无法生成唯一工作表名称: {base_name}")

            # 规范化名称（自动生成的名称允许静默清理）
            new_name = self._sanitize_sheet_name(new_name)

            # openpyxl copy_worksheet 创建副本
            target = workbook.copy_worksheet(source_sheet)
            target.title = new_name

            # 调整位置（copy_worksheet 默认追加到末尾）
            if index is not None:
                total = len(workbook.sheetnames)
                if index < 0 or index > total:
                    raise DataValidationError(f"索引超出范围: {index}，应在 0-{total} 之间")
                workbook.move_sheet(new_name, offset=index - (total - 1))

            workbook.save(self.file_path)

            new_index = workbook.sheetnames.index(new_name)
            sheet_info = SheetInfo(
                index=new_index,
                name=new_name,
                max_row=target.max_row,
                max_column=target.max_column,
                max_column_letter=get_column_letter(target.max_column) if target.max_column > 0 else 'A'
            )

            return OperationResult(
                success=True,
                data=sheet_info,
                message=f"成功复制工作表 '{source_name}' 为 '{new_name}'（{target.max_row}行 × {target.max_column}列）",
                metadata={
                    'file_path': self.file_path,
                    'source_name': source_name,
                    'new_name': new_name,
                    'copied_rows': target.max_row,
                    'copied_columns': target.max_column,
                    'new_index': new_index,
                    'total_sheets': len(workbook.sheetnames),
                    'all_sheets': workbook.sheetnames
                }
            )

        except Exception as e:
            logger.error(f"复制工作表失败: {e}")
            return OperationResult(success=False, error=str(e))

    def _copy_sheet_streaming(
        self,
        source_name: str,
        new_name: Optional[str] = None,
        index: Optional[int] = None
    ) -> OperationResult:
        """
        流式复制工作表：读取源数据后用calamine+write_only重建，大文件性能更好

        Args:
            source_name: 源工作表名称
            new_name: 新工作表名称（为空则自动生成 "源表名_副本"）
            index: 插入位置索引（None表示追加到末尾）

        Returns:
            OperationResult: 复制操作的结果
        """
        import tempfile
        import shutil

        # 自动生成新名称
        if not new_name or not new_name.strip():
            new_name = f"{source_name}_副本"
        else:
            new_name = new_name.strip()

        # 使用calamine读取源工作表数据
        try:
            from python_calamine import CalamineWorkbook
            cal_wb = CalamineWorkbook.from_path(self.file_path)
        except Exception as e:
            logger.warning(f"calamine读取失败，降级到openpyxl: {e}")
            return self.copy_sheet(source_name, new_name, index, streaming=False)

        if source_name not in cal_wb.sheet_names:
            available = ', '.join(cal_wb.sheet_names)
            raise SheetNotFoundError(f"工作表不存在: {source_name}（可用: {available}）")

        source_rows = cal_wb.get_sheet_by_name(source_name).to_python()
        if not source_rows:
            return OperationResult(
                success=False,
                error=f"源工作表 '{source_name}' 为空或读取失败"
            )

        # 处理名称冲突（只需检查sheet_names列表，无需读取数据）
        existing_names = set(cal_wb.sheet_names)
        base_name = new_name
        counter = 1
        new_name = self._sanitize_sheet_name(new_name)
        while new_name in existing_names:
            new_name = f"{base_name}_{counter}"
            counter += 1
            if counter > 100:
                raise DataValidationError(f"无法生成唯一工作表名称: {base_name}")

        # 只在openpyxl中用copy_worksheet复制，保留格式和列宽
        # 对于大文件，先用calamine读取源数据，再用openpyxl直接添加新sheet
        wb = load_workbook(self.file_path)
        try:
            source_ws = wb[source_name]
            new_ws = wb.copy_worksheet(source_ws)
            new_ws.title = new_name

            if index is not None and 0 <= index < len(wb.sheetnames):
                # 移动到指定位置
                sheet_list = wb.sheetnames
                current_idx = sheet_list.index(new_name)
                wb.move_sheet(new_name, offset=index - current_idx)

            wb.save(self.file_path)
        finally:
            wb.close()

        total_rows = len(source_rows)
        total_cols = max(len(row) for row in source_rows) if source_rows else 0

        return OperationResult(
            success=True,
            data={
                'index': len(cal_wb.sheet_names),
                'name': new_name,
                'max_row': total_rows,
                'max_column': total_cols,
                'max_column_letter': get_column_letter(total_cols) if total_cols > 0 else 'A'
            },
            message=f"成功复制工作表 '{source_name}' 为 '{new_name}'（{total_rows}行 × {total_cols}列，流式模式）",
            metadata={
                'file_path': self.file_path,
                'source_name': source_name,
                'new_name': new_name,
                'copied_rows': total_rows,
                'copied_columns': total_cols,
                'mode': 'streaming',
                'col_widths_preserved': True
            }
        )

    def rename_column(
        self,
        sheet_name: str,
        old_header: str,
        new_header: str,
        header_row: int = 1
    ) -> OperationResult:
        """
        重命名指定工作表的列（修改表头单元格值）

        Args:
            sheet_name: 工作表名称
            old_header: 当前列名（必须精确匹配表头单元格值）
            new_header: 新列名
            header_row: 表头所在行号（默认1，双行表头场景可设为2）

        Returns:
            OperationResult: 操作结果
        """
        try:
            if not sheet_name or not sheet_name.strip():
                raise DataValidationError("工作表名称不能为空")
            if not old_header or not old_header.strip():
                raise DataValidationError("当前列名不能为空")
            if not new_header or not new_header.strip():
                raise DataValidationError("新列名不能为空")

            old_header = old_header.strip()
            new_header = new_header.strip()

            if old_header == new_header:
                raise DataValidationError("新列名与当前列名相同，无需修改")

            workbook = load_workbook(self.file_path)

            if sheet_name not in workbook.sheetnames:
                available = ', '.join(workbook.sheetnames)
                raise SheetNotFoundError(f"工作表不存在: {sheet_name}（可用: {available}）")

            sheet = workbook[sheet_name]

            if header_row < 1 or header_row > sheet.max_row:
                raise DataValidationError(f"表头行号 {header_row} 超出范围（1-{sheet.max_row}）")

            # 查找匹配的表头单元格
            col_idx = None
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row, column=col).value
                if cell_value is not None and str(cell_value).strip() == old_header:
                    col_idx = col
                    break

            if col_idx is None:
                # 收集实际列名用于提示
                actual_headers = []
                for col in range(1, sheet.max_column + 1):
                    v = sheet.cell(row=header_row, column=col).value
                    if v is not None:
                        actual_headers.append(str(v).strip())
                raise DataValidationError(
                    f"在行 {header_row} 中未找到列名 '{old_header}'（实际列名: {', '.join(actual_headers[:10])}）"
                )

            # 修改列名
            old_value = sheet.cell(row=header_row, column=col_idx).value
            sheet.cell(row=header_row, column=col_idx).value = new_header
            col_letter = get_column_letter(col_idx)

            workbook.save(self.file_path)

            return OperationResult(
                success=True,
                message=f"成功将列 '{old_header}' 重命名为 '{new_header}'（{col_letter}{header_row}）",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet_name,
                    'old_header': old_header,
                    'new_header': new_header,
                    'cell': f"{col_letter}{header_row}",
                    'header_row': header_row
                }
            )

        except Exception as e:
            logger.error(f"重命名列失败: {e}")
            return OperationResult(success=False, error=str(e))

    @staticmethod
    def get_file_info(file_path: str) -> OperationResult:
        """
        获取Excel文件的详细信息

        Args:
            file_path: Excel文件路径

        Returns:
            OperationResult: 文件信息结果
        """
        try:
            import os
            from datetime import datetime

            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")

            # 获取文件系统信息
            stat_info = os.stat(file_path)
            file_size = stat_info.st_size
            created_time = datetime.fromtimestamp(stat_info.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
            modified_time = datetime.fromtimestamp(stat_info.st_mtime).strftime('%Y-%m-%d %H:%M:%S')

            # 获取文件格式
            file_format = Path(file_path).suffix.lower().lstrip('.')

            # 加载工作簿获取详细信息（非read_only模式以获取准确数据维度）
            workbook = load_workbook(file_path)
            sheet_count = len(workbook.sheetnames)
            sheet_names = workbook.sheetnames
            has_macros = file_format == 'xlsm'

            # 统计第一个工作表的行列范围
            total_rows = 0
            total_cols = 0
            formatted_rows = 0
            formatted_cols = 0
            if workbook.worksheets:
                ws = workbook.worksheets[0]
                # 格式化范围（包含仅有格式无数据的单元格）
                if ws.max_row and ws.max_column:
                    formatted_rows = ws.max_row
                    formatted_cols = ws.max_column
                # 实际数据范围（仅含非空值单元格）
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            total_rows = max(total_rows, cell.row)
                            total_cols = max(total_cols, cell.column)

            workbook.close()

            # 构建返回数据
            file_data = {
                'file_path': file_path,
                'file_name': Path(file_path).name,
                'file_size': file_size,
                'file_size_mb': round(file_size / 1024 / 1024, 2),
                'created_time': created_time,
                'modified_time': modified_time,
                'format': file_format,
                'sheet_count': sheet_count,
                'sheet_names': sheet_names,
                'has_macros': has_macros,
                'total_rows': total_rows,
                'total_cols': total_cols
            }
            # 仅当格式化范围与数据范围不同时才报告格式化维度
            if formatted_rows != total_rows or formatted_cols != total_cols:
                file_data['formatted_rows'] = formatted_rows
                file_data['formatted_cols'] = formatted_cols

            return OperationResult(
                success=True,
                message=f"成功获取文件信息: {file_path}",
                data=file_data,
                metadata={
                    'file_format': file_format,
                    'is_macro_enabled': has_macros,
                    'sheet_summary': {name: f"工作表{i+1}" for i, name in enumerate(sheet_names)}
                }
            )

        except Exception as e:
            logger.error(f"获取文件信息失败: {e}")
            return OperationResult(
                success=False,
                error=f"获取文件信息失败: {str(e)}"
            )

    def upsert_row(
        self,
        sheet_name: str,
        key_column: str,
        key_value,
        updates: dict,
        header_row: int = 1,
        streaming: bool = True
    ) -> OperationResult:
        """
        Upsert行：按键列查找，存在则更新，不存在则插入新行。

        Args:
            sheet_name: 工作表名称
            key_column: 用于匹配的列名
            key_value: 用于匹配的值
            updates: 要写入的列值字典（含key_column对应的值）
            header_row: 表头所在行号（默认1）
            streaming: 是否使用流式写入（默认True）

        Returns:
            OperationResult: 操作结果，含action(update/insert)、行号等信息
        """
        try:
            if not sheet_name or not sheet_name.strip():
                raise DataValidationError("工作表名称不能为空")
            if not key_column or not key_column.strip():
                raise DataValidationError("键列名不能为空")
            if key_value is None:
                raise DataValidationError("键值不能为None")
            if not updates or not isinstance(updates, dict):
                raise DataValidationError("更新数据不能为空，需提供列值字典")

            key_column = key_column.strip()
            sheet_name = sheet_name.strip()

            # 流式写入路径
            if streaming:
                from .streaming_writer import StreamingWriter
                if StreamingWriter.is_available():
                    success, message, meta = StreamingWriter.upsert_row(
                        self.file_path, sheet_name, key_column, key_value,
                        updates, header_row
                    )
                    if success:
                        return OperationResult(
                            success=True,
                            data=meta,
                            message=message,
                            metadata={
                                'file_path': self.file_path,
                                'sheet_name': sheet_name,
                                'key_column': key_column,
                                'key_value': str(key_value),
                                **meta
                            }
                        )
                    else:
                        logger.warning(f"流式upsert失败，降级到openpyxl: {message}")

            # openpyxl 传统路径
            workbook = load_workbook(self.file_path)

            if sheet_name not in workbook.sheetnames:
                available = ', '.join(workbook.sheetnames)
                raise SheetNotFoundError(f"工作表不存在: {sheet_name}（可用: {available}）")

            sheet = workbook[sheet_name]

            if header_row < 1 or header_row > sheet.max_row:
                raise DataValidationError(f"表头行号 {header_row} 超出范围（1-{sheet.max_row}）")

            # 构建列名→列索引映射（支持双行表头：同时注册中英文列名）
            col_map = {}
            for col in range(1, sheet.max_column + 1):
                cell_val = sheet.cell(row=header_row, column=col).value
                if cell_val is not None:
                    col_map[str(cell_val).strip()] = col

            # 双行表头自动检测：如果下一行全是英文字段名，也注册到col_map
            # 这样LLM用describe_table返回的英文名也能匹配到
            _next_row = header_row + 1
            if _next_row <= sheet.max_row:
                _next_vals = [sheet.cell(row=_next_row, column=c).value for c in range(1, sheet.max_column + 1)]
                _next_strs = [v for v in _next_vals if v is not None]
                _is_english_row = (
                    len(_next_strs) >= 2
                    and all(isinstance(v, str) and v.strip() and v.strip()[0].isalpha() and v.strip()[0].isascii()
                           for v in _next_strs)
                )
                if _is_english_row:
                    for col in range(1, min(sheet.max_column + 1, len(_next_vals) + 1)):
                        val = _next_vals[col - 1]
                        if val is not None:
                            name = str(val).strip()
                            if name and name not in col_map:  # 不覆盖已有映射
                                col_map[name] = col

            if key_column not in col_map:
                actual = list(col_map.keys())[:10]
                raise DataValidationError(
                    f"键列 '{key_column}' 不存在。实际列名: {', '.join(actual)}"
                )

            key_col_idx = col_map[key_column]

            # 查找匹配行
            target_row = None
            for row in range(header_row + 1, sheet.max_row + 1):
                cell_val = sheet.cell(row=row, column=key_col_idx).value
                if cell_val is not None and str(cell_val).strip() == str(key_value).strip():
                    target_row = row
                    break

            # 执行upsert
            if target_row is not None:
                # UPDATE: 更新已有行
                updated_cols = []
                for col_name, value in updates.items():
                    col_name_stripped = col_name.strip()
                    if col_name_stripped in col_map:
                        col_idx = col_map[col_name_stripped]
                        sheet.cell(row=target_row, column=col_idx, value=value)
                        updated_cols.append(col_name_stripped)

                workbook.save(self.file_path)
                workbook.close()

                return OperationResult(
                    success=True,
                    data={
                        'action': 'update',
                        'row': target_row,
                        'updated_columns': updated_cols,
                        'updated_count': len(updated_cols)
                    },
                    message=f"更新行 {target_row}（键列 '{key_column}'='{key_value}'），修改了 {len(updated_cols)} 列",
                    metadata={
                        'file_path': self.file_path,
                        'sheet_name': sheet_name,
                        'key_column': key_column,
                        'key_value': str(key_value),
                        'action': 'update',
                        'target_row': target_row,
                        'updated_columns': updated_cols
                    }
                )
            else:
                # INSERT: 在末尾追加新行
                last_row = sheet.max_row
                new_row = last_row + 1

                # 确保所有列都写入（按表头顺序）
                inserted_cols = []
                for col_name, col_idx in sorted(col_map.items(), key=lambda x: x[1]):
                    if col_name in updates:
                        sheet.cell(row=new_row, column=col_idx, value=updates[col_name])
                        inserted_cols.append(col_name)
                    elif col_name == key_column:
                        # 确保key_column的值被写入（即使updates中没有）
                        sheet.cell(row=new_row, column=col_idx, value=key_value)
                        inserted_cols.append(col_name)

                workbook.save(self.file_path)
                workbook.close()

                return OperationResult(
                    success=True,
                    data={
                        'action': 'insert',
                        'row': new_row,
                        'inserted_columns': inserted_cols,
                        'inserted_count': len(inserted_cols)
                    },
                    message=f"插入新行 {new_row}（键列 '{key_column}'='{key_value}'），写入了 {len(inserted_cols)} 列",
                    metadata={
                        'file_path': self.file_path,
                        'sheet_name': sheet_name,
                        'key_column': key_column,
                        'key_value': str(key_value),
                        'action': 'insert',
                        'new_row': new_row,
                        'inserted_columns': inserted_cols
                    }
                )

        except Exception as e:
            logger.error(f"Upsert行失败: {e}")
            return OperationResult(success=False, error=str(e))

    def batch_insert_rows(
        self,
        sheet_name: str,
        data: list,
        header_row: int = 1,
        streaming: bool = True
    ) -> OperationResult:
        """
        批量插入多行数据到工作表末尾。

        Args:
            sheet_name: 工作表名称
            data: 行数据列表，每行为{列名: 值}字典
            header_row: 表头所在行号（默认1）
            streaming: 是否使用流式写入（calamine+write_only，内存更低）
                       默认True，大幅减少大文件操作的内存和时间
                       注意：流式模式不保留单元格格式（字体/填充/边框/合并），
                       但保留列宽、行高、数据值

        Returns:
            OperationResult: 操作结果
        """
        try:
            if not sheet_name or not sheet_name.strip():
                raise DataValidationError("工作表名称不能为空")
            if not data:
                raise DataValidationError("数据不能为空，需提供行数据列表")
            # 接受list和tuple（MCP可能传递tuple）
            if isinstance(data, dict):
                data = [data]  # 单个字典自动包装为列表
            if not isinstance(data, (list, tuple)):
                raise DataValidationError("数据必须是列表或元组，每项为字典格式的行数据")
            if len(data) > 10000:
                raise DataValidationError(f"单次最多插入10000行，当前{len(data)}行")

            sheet_name = sheet_name.strip()

            # 流式写入路径（calamine读取 + write_only写入）
            if streaming:
                from .streaming_writer import StreamingWriter
                if StreamingWriter.is_available():
                    success, message, meta = StreamingWriter.batch_insert_rows(
                        self.file_path, sheet_name, data, header_row
                    )
                    if success:
                        return OperationResult(
                            success=True,
                            data=meta,
                            message=message,
                            metadata={
                                'file_path': self.file_path,
                                'sheet_name': sheet_name,
                                **meta
                            }
                        )
                    else:
                        logger.warning(f"流式写入失败，降级到openpyxl: {message}")

            # openpyxl 传统路径（降级或 streaming=False）
            # 优先用 calamine 读取表头（比 openpyxl load_workbook 快 10x+）
            col_map = {}
            try:
                from python_calamine import CalamineWorkbook
                cal_wb = CalamineWorkbook.from_path(self.file_path)
                for sn in cal_wb.sheet_names:
                    if sn == sheet_name:
                        rows = cal_wb.get_sheet_by_name(sn).to_python()
                        if rows and header_row <= len(rows):
                            for col_idx, cell_val in enumerate(rows[header_row - 1], 1):
                                if cell_val is not None:
                                    col_map[str(cell_val).strip()] = col_idx
                        break
            except ImportError:
                pass

            # calamine 未获取到表头时，降级到 openpyxl
            workbook = load_workbook(self.file_path)
            if sheet_name not in workbook.sheetnames:
                available = ', '.join(workbook.sheetnames)
                workbook.close()
                raise SheetNotFoundError(f"工作表不存在: {sheet_name}（可用: {available}）")

            sheet = workbook[sheet_name]

            if not col_map:
                if header_row < 1 or header_row > sheet.max_row:
                    workbook.close()
                    raise DataValidationError(f"表头行号 {header_row} 超出范围（1-{sheet.max_row}）")
                for col in range(1, sheet.max_column + 1):
                    cell_val = sheet.cell(row=header_row, column=col).value
                    if cell_val is not None:
                        col_map[str(cell_val).strip()] = col

            if not col_map:
                raise DataValidationError("未找到表头列名")

            # 从末尾行开始追加
            start_row = sheet.max_row + 1
            unknown_cols = set()

            for i, row_data in enumerate(data):
                if not isinstance(row_data, dict):
                    raise DataValidationError(f"第{i + 1}行数据必须是字典，实际类型: {type(row_data).__name__}")
                row_num = start_row + i
                for col_name, value in row_data.items():
                    col_name_stripped = col_name.strip()
                    if col_name_stripped in col_map:
                        sheet.cell(row=row_num, column=col_map[col_name_stripped], value=value)
                    else:
                        unknown_cols.add(col_name_stripped)

            workbook.save(self.file_path)
            workbook.close()

            inserted_count = len(data)
            unknown_list = sorted(unknown_cols)[:5] if unknown_cols else []

            return OperationResult(
                success=True,
                data={
                    'action': 'batch_insert',
                    'start_row': start_row,
                    'end_row': start_row + inserted_count - 1,
                    'inserted_count': inserted_count,
                    'unknown_columns': unknown_list
                },
                message=f"批量插入 {inserted_count} 行（第{start_row}-{start_row + inserted_count - 1}行）",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet_name,
                    'action': 'batch_insert',
                    'start_row': start_row,
                    'inserted_count': inserted_count,
                    'unknown_columns': unknown_list
                }
            )

        except Exception as e:
            logger.error(f"批量插入行失败: {e}")
            return OperationResult(success=False, error=str(e))
