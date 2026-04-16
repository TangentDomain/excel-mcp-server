"""
Excel MCP Server - Excel写入模块

提供Excel文件写入和修改功能
"""

import ast
import json
import logging
import math
import os
import re
import time
from collections import Counter
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, GradientFill, PatternFill, Side
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import coordinate_from_string

from ..models.types import ModifiedCell, OperationResult, RangeType
from ..utils.exceptions import DataValidationError, SheetNotFoundError
from ..utils.formula_cache import get_formula_cache
from ..utils.parsers import RangeParser
from ..utils.temp_file_manager import TempFileManager
from ..utils.validators import ExcelValidator

logger = logging.getLogger(__name__)


# Fix: P2-4 极端浮点值导致文件损坏 — 单元格值清理函数
def _sanitize_cell_value(value: Any) -> Any:
    """清理单元格值,防止极端浮点值导致xlsx文件损坏.

    处理以下情况:
    - NaN → None(Excel空值)
    - Inf/-Inf → None(Excel空值)
    - 超出 ±1e308 的浮点值 → 截断到边界值
    - 其他类型 → 原样返回

    Args:
        value: 待写入的单元格值

    Returns:
        清理后的安全值
    """
    if value is None:
        return None

    # 仅处理浮点类型(含numpy浮点)
    if isinstance(value, (float,)):
        import numpy as np
        if np.isnan(value) or math.isinf(value):
            logger.warning(f"P2-4: 清理极端浮点值 {value} → None (NaN/Inf)")
            return None
        # 超出IEEE 754 double范围的值
        if abs(value) > 1e308:
            clamped = 1e308 if value > 0 else -1e308
            logger.warning(f"P2-4: 截断超范围浮点值 {value} → {clamped}")
            return clamped

    return value


class ExcelWriter:
    """Excel文件写入器"""

    def __init__(self, file_path: str):
        """
        初始化Excel写入器

        Args:
            file_path: Excel文件路径（允许空字符串，用于临时计算场景）
        """
        if file_path:
            self.file_path = ExcelValidator.validate_file_path(file_path)
        else:
            self.file_path = ""

    def _check_file_lock(self) -> None:
        """检查文件是否被锁定

        Raises:
            PermissionError: 文件被锁定或无法访问
        """
        try:
            # 尝试以写入模式打开文件来检测锁定
            test_file = open(self.file_path, "rb+")
            test_file.close()
        except PermissionError as e:
            raise PermissionError(f"文件被锁定或权限不足: {self.file_path}") from e
        except OSError as e:
            if e.errno == 13:  # Permission denied
                raise PermissionError(f"文件被锁定: {self.file_path}") from e
            raise

    def _safe_save_workbook(self, workbook, operation_name: str = "保存文件") -> None:
        """安全保存工作簿，包含文件锁定检测和错误处理

        Args:
            workbook: openpyxl工作簿对象
            operation_name: 操作名称，用于错误消息

        Raises:
            PermissionError: 文件被锁定或权限不足
            IOError: 保存文件时发生IO错误
            Exception: 其他保存错误
        """
        # 检查文件锁定
        self._check_file_lock()

        try:
            workbook.save(self.file_path)
            logger.info(f"{operation_name}成功: {self.file_path}")
        except PermissionError as e:
            logger.error(f"{operation_name}失败 - 权限错误: {e}")
            raise PermissionError(f"文件保存失败，权限不足或文件被锁定: {self.file_path}") from e
        except OSError as e:
            logger.error(f"{operation_name}失败 - IO错误: {e}")
            raise OSError(f"文件保存失败，IO错误: {e}") from e
        except Exception as e:
            logger.error(f"{operation_name}失败 - 未知错误: {e}")
            raise Exception(f"文件保存失败: {str(e)}") from e

    def update_range(
        self,
        range_expression: str,
        data: list[list[Any]],
        preserve_formulas: bool = True,
        insert_mode: bool = False,
    ) -> OperationResult:
        """
        修改Excel文件中指定范围的数据

        Args:
            range_expression: 范围表达式
            data: 要写入的二维数据数组
            preserve_formulas: 是否保留现有的公式
            insert_mode: 数据写入模式 (默认值: False)
                - True: 插入模式，在指定位置插入新行然后写入数据（更安全）
                - False: 覆盖模式，直接覆盖目标范围的现有数据（默认行为）

        Returns:
            OperationResult: 修改操作的结果
        """
        try:
            # 解析范围表达式
            range_info = RangeParser.parse_range_expression(range_expression)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, range_info.sheet_name)

            # 根据范围类型处理不同的范围格式
            cell_range_for_boundaries = self._convert_to_cell_range(range_info.cell_range, range_info.range_type, sheet, data)

            # 获取范围边界
            min_col, min_row, max_col, max_row = range_boundaries(cell_range_for_boundaries)

            # 处理插入模式
            is_smart_append = False
            if insert_mode:
                rows_to_insert = len(data)
                if rows_to_insert > 0:
                    # 智能追加：目标行在数据末尾之后时，跳过 insert_rows（O(n)行移动）
                    # 追加场景无需遍历公式，因为新行位于所有现有数据之后
                    current_max_row = sheet.max_row
                    if min_row > current_max_row:
                        is_smart_append = True
                        logger.info(f"智能追加：目标行{min_row} > 数据末尾{current_max_row}，跳过行插入")
                    else:
                        # 插入模式：在指定位置插入足够的行数
                        # 保存公式位置信息，以便后续调整
                        formula_positions = {}
                        if preserve_formulas:
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.data_type == "f":
                                        formula_positions[cell.coordinate] = cell.value

                        sheet.insert_rows(min_row, rows_to_insert)
                        logger.info(f"插入模式：在第{min_row}行插入了{rows_to_insert}行")

            # 写入数据（追加模式下新行无公式，跳过公式保留检查）
            modified_cells = self._write_data(sheet, data, min_row, min_col, preserve_formulas and not is_smart_append)

            # 保存文件
            self._safe_save_workbook(workbook, "更新范围数据")
            workbook.close()

            mode_description = "插入模式" if insert_mode else "覆盖模式"
            if is_smart_append:
                mode_description = "智能追加模式"

            return OperationResult(
                success=True,
                data=modified_cells,
                metadata={
                    "file_path": self.file_path,
                    "range": range_expression,
                    "sheet_name": sheet.title,
                    "modified_cells_count": len(modified_cells),
                    "insert_mode": insert_mode,
                    "mode_description": mode_description,
                    "rows_inserted": len(data) if insert_mode else 0,
                    "smart_append": is_smart_append,
                },
            )

        except Exception as e:
            logger.error(f"更新范围数据失败: {e}")
            return OperationResult(success=False, error=str(e))

    def insert_rows(self, sheet_name: str | None = None, row_index: int = 1, count: int = 1) -> OperationResult:
        """
        在Excel文件中插入空白行

        Args:
            sheet_name: 工作表名称
            row_index: 插入行的位置（1-based）
            count: 要插入的行数

        Returns:
            OperationResult: 插入操作的结果
        """
        try:
            # 验证参数
            ExcelValidator.validate_row_operations(row_index, count)
            ExcelValidator.validate_sheet_name(sheet_name)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, sheet_name)

            # 记录操作前的信息
            original_max_row = sheet.max_row

            # 插入行
            sheet.insert_rows(row_index, count)

            # 保存文件
            self._safe_save_workbook(workbook, "插入行")
            workbook.close()

            return OperationResult(
                success=True,
                message=f"成功在第{row_index}行前插入{count}行",
                metadata={
                    "file_path": self.file_path,
                    "sheet_name": sheet.title,
                    "inserted_at_row": row_index,
                    "inserted_count": count,
                    "original_max_row": original_max_row,
                    "new_max_row": sheet.max_row,
                },
            )

        except Exception as e:
            logger.error(f"插入行失败: {e}")
            return OperationResult(success=False, error=str(e))

    def insert_columns(self, sheet_name: str | None = None, column_index: int = 1, count: int = 1) -> OperationResult:
        """
        在Excel文件中插入空白列

        Args:
            sheet_name: 工作表名称
            column_index: 插入列的位置（1-based）
            count: 要插入的列数

        Returns:
            OperationResult: 插入操作的结果
        """
        try:
            # 验证参数
            ExcelValidator.validate_column_operations(column_index, count)
            ExcelValidator.validate_sheet_name(sheet_name)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, sheet_name)

            # 记录操作前的信息
            original_max_column = sheet.max_column

            # 插入列
            sheet.insert_cols(column_index, count)

            # 记录插入后的最大列数（在关闭前获取）
            new_max_column = sheet.max_column

            # 保存文件
            self._safe_save_workbook(workbook, "插入列")
            workbook.close()

            # 验证：重新加载文件确认列已插入
            verification_workbook = load_workbook(self.file_path)
            try:
                verification_sheet = verification_workbook[sheet.title]
                actual_max_column = verification_sheet.max_column
                expected_max_column = original_max_column + count

                if actual_max_column != expected_max_column:
                    raise Exception(f"插入列验证失败：期望最大列数为 {expected_max_column}，实际为 {actual_max_column}")
            finally:
                verification_workbook.close()

            return OperationResult(
                success=True,
                message=f"成功插入{count}列",
                metadata={
                    "file_path": self.file_path,
                    "sheet_name": sheet.title,
                    "inserted_at_column": column_index,
                    "inserted_count": count,
                    "original_max_column": original_max_column,
                    "new_max_column": new_max_column,
                },
            )

        except Exception as e:
            logger.error(f"插入列失败: {e}")
            return OperationResult(success=False, error=str(e))

    def delete_rows(self, sheet_name: str | None = None, start_row: int = 1, count: int = 1) -> OperationResult:
        """
        删除Excel文件中的行

        Args:
            sheet_name: 工作表名称
            start_row: 开始删除的行号
            count: 要删除的行数

        Returns:
            OperationResult: 删除操作的结果
        """
        try:
            # 验证参数
            ExcelValidator.validate_row_operations(start_row, count)
            ExcelValidator.validate_sheet_name(sheet_name)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, sheet_name)

            # 记录操作前的信息
            original_max_row = sheet.max_row

            # 验证删除范围
            if start_row > original_max_row:
                raise DataValidationError(
                    f"起始行号({start_row})超过工作表最大行数({original_max_row})",
                    f"工作表最大行数为{original_max_row}",
                    f"请使用不超过{original_max_row}的行号",
                )

            # 计算实际删除的行数
            actual_count = min(count, original_max_row - start_row + 1)

            # 删除行
            sheet.delete_rows(start_row, actual_count)

            # 保存文件
            self._safe_save_workbook(workbook, "删除行")

            workbook.close()
            return OperationResult(
                success=True,
                message=f"成功从第{start_row}行开始删除{actual_count}行",
                metadata={
                    "file_path": self.file_path,
                    "sheet_name": sheet.title,
                    "deleted_start_row": start_row,
                    "actual_deleted_count": actual_count,
                    "original_max_row": original_max_row,
                    "new_max_row": sheet.max_row,
                },
            )

        except Exception as e:
            logger.error(f"删除行失败: {e}")
            return OperationResult(success=False, error=str(e))

    def delete_columns(self, sheet_name: str | None = None, start_column: int = 1, count: int = 1) -> OperationResult:
        """
        删除Excel文件中的列

        Args:
            sheet_name: 工作表名称
            start_column: 开始删除的列号
            count: 要删除的列数

        Returns:
            OperationResult: 删除操作的结果
        """
        try:
            # 验证参数
            ExcelValidator.validate_column_operations(start_column, count)
            ExcelValidator.validate_sheet_name(sheet_name)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, sheet_name)

            # 记录操作前的信息
            original_max_column = sheet.max_column

            # 验证删除范围
            if start_column > original_max_column:
                raise DataValidationError(f"起始列号({start_column})超过工作表最大列数({original_max_column})")

            # 计算实际删除的列数
            actual_count = min(count, original_max_column - start_column + 1)

            # 删除列
            sheet.delete_cols(start_column, actual_count)

            # 保存文件
            self._safe_save_workbook(workbook, "删除列")

            workbook.close()
            return OperationResult(
                success=True,
                message=f"成功删除{actual_count}列",
                metadata={
                    "file_path": self.file_path,
                    "sheet_name": sheet.title,
                    "deleted_start_column": start_column,
                    "actual_deleted_count": actual_count,
                    "original_max_column": original_max_column,
                    "new_max_column": sheet.max_column,
                },
            )

        except Exception as e:
            logger.error(f"删除列失败: {e}")
            return OperationResult(success=False, error=str(e))

    def _convert_to_cell_range(self, cell_range: str, range_type: RangeType, sheet, data: list[list[Any]]) -> str:
        """
        检查范围表达式格式的合法性

        Args:
            cell_range: 原始范围表达式
            range_type: 范围类型
            sheet: 工作表对象
            data: 数据数组

        Returns:
            str: 标准的单元格范围表达式

        Raises:
            ValueError: 当使用不支持的范围格式时
        """
        if range_type == RangeType.ROW_RANGE:
            # 对于纯行范围格式，抛出明确错误并提供建议
            raise ValueError(
                f'不支持纯行范围格式 "{cell_range}"。请使用以下格式之一：\n'
                f'- 标准格式: "A{cell_range.split(":")[0]}:Z{cell_range.split(":")[-1]}" (明确指定列范围)\n'
                f'- 单列格式: "A{cell_range.split(":")[0]}:A{cell_range.split(":")[-1]}" (单列数据)\n'
                f'- 指定范围: "B{cell_range.split(":")[0]}:E{cell_range.split(":")[-1]}" (B到E列)'
            )

        elif range_type == RangeType.COLUMN_RANGE:
            # 对于列范围，直接返回不做自动扩展
            return cell_range

        elif range_type == RangeType.SINGLE_ROW:
            # 对于单行范围，抛出明确错误
            row_num = int(cell_range.split(":")[0])  # 规范化后是 "1:1" 格式
            raise ValueError(f'不支持单行范围格式 "{row_num}"。请使用以下格式之一：\n- 标准格式: "A{row_num}:Z{row_num}" (明确指定列范围)\n- 单列格式: "A{row_num}:A{row_num}" (单列数据)')

        elif range_type == RangeType.SINGLE_COLUMN:
            # 对于单列范围，直接返回不做自动扩展
            return cell_range

        else:
            # 其他情况（CELL_RANGE）直接返回
            return cell_range

    def _get_worksheet(self, workbook, sheet_name: str | None):
        """获取工作表 - 强制要求指定工作表名称

        Args:
            workbook: openpyxl工作簿对象
            sheet_name: 工作表名称

        Returns:
            Worksheet: openpyxl工作表对象

        Raises:
            SheetNotFoundError: 工作表不存在或为空时抛出
        """
        if not sheet_name or not sheet_name.strip():
            raise SheetNotFoundError("工作表名称不能为空，必须明确指定工作表")

        if not workbook.sheetnames:
            raise SheetNotFoundError("Excel文件中没有任何工作表")

        if sheet_name not in workbook.sheetnames:
            raise SheetNotFoundError(f"工作表不存在: {sheet_name}，可用工作表: {', '.join(workbook.sheetnames)}")

        return workbook[sheet_name]

    def _write_data(
        self,
        sheet,
        data: list[list[Any]],
        start_row: int,
        start_col: int,
        preserve_formulas: bool,
    ) -> list[ModifiedCell]:
        """写入数据到工作表

        Args:
            sheet: openpyxl工作表对象
            data: 要写入的二维数据数组
            start_row: 起始行号（1-based）
            start_col: 起始列号（1-based）
            preserve_formulas: 是否保留公式

        Returns:
            List[ModifiedCell]: 修改的单元格列表
        """
        modified_cells = []

        for row_offset, row_data in enumerate(data):
            for col_offset, value in enumerate(row_data):
                row_idx = start_row + row_offset
                col_idx = start_col + col_offset
                cell = sheet.cell(row=row_idx, column=col_idx)

                # 保留公式检查
                if preserve_formulas and cell.data_type == "f":
                    continue

                old_value = cell.value

                # Fix: P2-4 极端浮点值导致文件损坏 — 写入前校验并清理NaN/Inf/超范围值
                value = _sanitize_cell_value(value)

                # 处理复杂数据类型
                try:
                    # 尝试直接设置值
                    cell.value = value
                except (ValueError, TypeError) as e:
                    # 如果直接设置失败，尝试转换为字符串
                    logger.warning(f"无法直接设置值 {value} ({type(value).__name__})，转换为字符串: {e}")
                    try:
                        if isinstance(value, (list, dict, tuple)):
                            # 复杂数据类型转换为JSON字符串
                            cell.value = json.dumps(value, ensure_ascii=False)
                        elif hasattr(value, "__str__"):
                            # 有字符串表示的对象
                            cell.value = str(value)
                        else:
                            # 最后尝试转换为字符串
                            cell.value = repr(value)
                    except Exception as conversion_error:
                        logger.error(f"无法转换值 {value}: {conversion_error}")
                        # 设置为空字符串作为最后手段
                        cell.value = ""

                modified_cells.append(ModifiedCell(coordinate=cell.coordinate, old_value=old_value, new_value=value))

        return modified_cells

    def set_formula(self, cell_address: str, formula: str, sheet_name: str | None = None) -> OperationResult:
        """
        设置单元格公式

        Args:
            cell_address: 目标单元格地址（如"A1"）
            formula: Excel公式（不包含等号）
            sheet_name: 目标工作表名称

        Returns:
            OperationResult: 公式设置结果
        """
        try:
            # 验证公式格式（简单验证）
            if not formula.strip():
                return OperationResult(success=False, error="公式不能为空")

            # 确保公式不以等号开头（openpyxl会自动添加）
            if formula.startswith("="):
                formula = formula[1:]

            # 验证单元格地址格式
            try:
                coordinate_from_string(cell_address)
            except ValueError:
                return OperationResult(success=False, error=f"单元格地址格式错误: {cell_address}")

            # 加载工作簿并设置公式
            workbook = load_workbook(self.file_path)
            sheet = self._get_worksheet(workbook, sheet_name)

            # 设置公式
            cell = sheet[cell_address]
            old_value = cell.value
            old_formula = cell.formula if hasattr(cell, "formula") else None

            cell.value = f"={formula}"

            # 保存文件
            self._safe_save_workbook(workbook, "设置公式")
            workbook.close()

            # 重新读取以获取计算值 - 使用只读模式
            workbook_read = load_workbook(self.file_path, data_only=True, read_only=True)
            sheet_read = self._get_worksheet(workbook_read, sheet_name)
            calculated_value = sheet_read[cell_address].value
            workbook_read.close()

            logger.info(f"成功设置公式: {cell_address} = {formula}")

            return OperationResult(
                success=True,
                message="公式设置成功",
                metadata={
                    "file_path": self.file_path,
                    "sheet_name": sheet.title,
                    "cell_address": cell_address,
                    "formula": formula,
                    "calculated_value": calculated_value,
                    "old_value": old_value,
                    "old_formula": old_formula,
                },
            )

        except Exception as e:
            logger.error(f"设置公式失败: {e}")
            return OperationResult(success=False, error=str(e))

    def format_cells(self, range_expression: str, formatting: dict, sheet_name: str | None = None) -> OperationResult:
        """
        设置单元格格式

        Args:
            range_expression: 目标范围
            formatting: 格式配置字典
            sheet_name: 目标工作表名

        Returns:
            OperationResult: 格式应用结果
        """
        try:
            # 解析范围表达式
            range_info = RangeParser.parse_range_expression(range_expression)

            # 加载工作簿
            workbook = load_workbook(self.file_path)
            sheet = self._get_worksheet(workbook, sheet_name or range_info.sheet_name)

            # 获取范围边界
            if range_info.range_type in [
                RangeType.COLUMN_RANGE,
                RangeType.SINGLE_COLUMN,
                RangeType.ROW_RANGE,
                RangeType.SINGLE_ROW,
            ]:
                # 处理整行或整列
                cells_range = sheet[range_expression.replace(f"{sheet.title}!", "")]
            else:
                min_col, min_row, max_col, max_row = range_boundaries(range_info.cell_range)
                cells_range = sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)

            formatted_count = 0

            # 应用格式
            for row in cells_range:
                if isinstance(row, tuple):
                    for cell in row:
                        self._apply_cell_format(cell, formatting)
                        formatted_count += 1
                else:
                    self._apply_cell_format(row, formatting)
                    formatted_count += 1

            # 保存文件
            self._safe_save_workbook(workbook, "格式化单元格")
            workbook.close()

            logger.info(f"成功格式化{formatted_count}个单元格")

            return OperationResult(
                success=True,
                message=f"成功格式化{formatted_count}个单元格",
                metadata={
                    "file_path": self.file_path,
                    "sheet_name": sheet.title,
                    "range": range_expression,
                    "formatted_count": formatted_count,
                    "formatting_applied": formatting,
                },
            )

        except Exception as e:
            logger.error(f"格式化失败: {e}")
            return OperationResult(success=False, error=str(e))

    def evaluate_formula(self, formula: str, context_sheet: str | None = None) -> OperationResult:
        """
        临时执行Excel公式并返回计算结果，不修改文件
        使用缓存机制提升性能

        Args:
            formula: Excel公式（不包含等号）
            context_sheet: 公式执行的上下文工作表

        Returns:
            OperationResult: 公式执行结果
        """
        try:
            start_time = time.time()

            # 确保公式不以等号开头
            if formula.startswith("="):
                formula = formula[1:]

            # 验证公式格式
            if not formula.strip():
                return OperationResult(success=False, error="公式不能为空")

            # 尝试从缓存获取结果
            cache = get_formula_cache()
            cached_result = cache.get(self.file_path, formula, context_sheet)

            if cached_result is not None:
                execution_time = round((time.time() - start_time) * 1000, 2)
                logger.info(f"缓存命中，公式: {formula} = {cached_result}")

                # 确定结果类型
                result_type = self._get_result_type(cached_result)

                return OperationResult(
                    success=True,
                    message="公式执行成功（缓存）",
                    data=cached_result,
                    metadata={
                        "formula": formula,
                        "result": cached_result,
                        "result_type": result_type,
                        "execution_time_ms": execution_time,
                        "context_sheet": context_sheet or "default",
                        "cached": True,
                        "cache_stats": cache.get_stats(),
                    },
                )

            # 缓存未命中
            # 快速路径：先判断是否纯算术公式（无单元格引用）
            # 纯算术公式直接用 Python 计算，完全跳过临时工作簿创建和磁盘I/O
            if not re.search(r"[A-Za-z]+\d+", formula):
                logger.debug(f"检测到纯算术公式，使用快速计算（跳过工作簿创建）: {formula}")
                calculated_value = self._fast_calculate(formula)
            else:
                # 含单元格引用的公式需要工作簿上下文
                cached_workbook_data = cache.get_cached_workbook(self.file_path)

                if cached_workbook_data:
                    temp_workbook, temp_file_path = cached_workbook_data
                    logger.debug("使用缓存的工作簿进行计算")
                else:
                    # 创建新的临时工作簿（传入公式用于判断是否需要加载数据）
                    temp_workbook, temp_file_path = self._create_temp_workbook(context_sheet, cache, formula)

                try:
                    # 使用xlcalculator计算公式
                    calculated_value = self._calculate_with_xlcalculator(temp_file_path, formula, temp_workbook)

                except ImportError:
                    return OperationResult(
                        success=False,
                        error="需要安装xlcalculator库来支持公式计算: pip install xlcalculator",
                    )
                except Exception as calc_error:
                    # 如果xlcalculator失败，尝试基础的手动解析
                    logger.warning(f"xlcalculator计算失败，尝试基础解析: {calc_error}")
                    calculated_value = self._fallback_calculation(temp_file_path, formula)

            # 缓存计算结果
            cache.put(self.file_path, formula, calculated_value, context_sheet)

            # 确定结果类型
            result_type = self._get_result_type(calculated_value)

            execution_time = round((time.time() - start_time) * 1000, 2)
            logger.info(f"成功计算公式: {formula} = {calculated_value}")

            return OperationResult(
                success=True,
                message="公式执行成功",
                data=calculated_value,
                metadata={
                    "formula": formula,
                    "result": calculated_value,
                    "result_type": result_type,
                    "execution_time_ms": execution_time,
                    "context_sheet": context_sheet or "default",
                    "cached": False,
                    "cache_stats": cache.get_stats(),
                },
            )

        except Exception as e:
            logger.error(f"公式执行失败: {e}")
            return OperationResult(success=False, error=f"公式执行失败: {str(e)}")

    def _detect_file_format(self) -> str:
        """检测Excel文件格式

        根据文件扩展名检测文件格式。支持的格式：
        - xlsx: Excel 工作簿（默认格式）
        - xls: 旧版 Excel 工作簿（仅读取）
        - xlsm: 启用宏的工作簿
        - xltx: Excel 模板
        - xltm: 启用宏的模板
        - xlsb: 二进制工作簿

        注意：使用 os.path.splitext 获取扩展名后，会转换为小写并去掉前导点。
        如果扩展名不在支持列表中，将返回默认格式 'xlsx'。

        Returns:
            str: 文件格式字符串，默认返回 'xlsx'
        """
        if not self.file_path or not os.path.exists(self.file_path):
            logger.debug("无文件路径或文件不存在，使用默认格式 xlsx")
            return "xlsx"

        # 从文件扩展名提取格式
        _, ext = os.path.splitext(self.file_path)
        ext = ext.lower().lstrip(".")

        # 验证格式是否受支持
        supported_formats = {"xlsx", "xls", "xlsm", "xltx", "xltm", "xlsb"}
        if ext in supported_formats:
            logger.debug(f"检测到文件格式: {ext}")
            return ext

        logger.warning(f"不支持的文件格式 '{ext}'，使用默认格式 xlsx")
        return "xlsx"

    def _create_temp_workbook(self, context_sheet: str | None, cache, formula: str = None) -> tuple:
        """创建临时工作簿用于计算，检测并保留原始文件格式

        Args:
            context_sheet: 上下文工作表名称
            cache: 公式缓存对象
            formula: 待计算的公式（用于判断是否需要加载原始数据）

        Returns:
            tuple: (temp_workbook, temp_file_path) 元组，包含临时工作簿和文件路径
        """
        # 检测公式是否包含单元格引用（字母+数字模式如A1, B2, SUM(A1:A10)）
        # formula=None 表示直接调用（非evaluate_formula路径），默认需要加载数据
        has_cell_ref = formula is None or bool(re.search(r"[A-Za-z]+\d+", formula or ""))

        # 创建临时工作簿进行计算
        temp_workbook = Workbook()
        temp_sheet = temp_workbook.active
        temp_sheet.title = "Calculation"

        # 只在公式包含单元格引用时才加载原始数据
        if has_cell_ref and self.file_path and os.path.exists(self.file_path):
            try:
                original_workbook = load_workbook(self.file_path, data_only=False, read_only=True)

                if context_sheet and context_sheet in original_workbook.sheetnames:
                    source_sheet = original_workbook[context_sheet]
                else:
                    source_sheet = original_workbook.active

                # 使用values_only批量读取，避免逐cell访问
                max_rows = min(source_sheet.max_row or 1000, 1000)
                max_cols = min(source_sheet.max_column or 100, 100)
                for row_idx, row_data in enumerate(
                    source_sheet.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True),
                    start=1,
                ):
                    for col_idx, value in enumerate(row_data, start=1):
                        if value is not None:
                            temp_sheet.cell(row=row_idx, column=col_idx, value=value)

                original_workbook.close()
            except Exception as e:
                logger.warning(f"无法加载原始工作簿，使用空工作簿: {e}")
        else:
            logger.debug("无文件路径或公式不含单元格引用，使用空工作簿进行计算")

        # 保存到临时文件
        temp_file_path = TempFileManager.create_temp_excel_file(suffix=".xlsx")
        temp_workbook.save(temp_file_path)

        # 缓存工作簿
        cache.cache_workbook(self.file_path or "temp", temp_workbook, temp_file_path)

        return temp_workbook, temp_file_path

    def _calculate_with_xlcalculator(self, temp_file_path: str, formula: str, temp_workbook) -> any:
        """使用xlcalculator进行计算

        Args:
            temp_file_path: 临时文件路径
            formula: Excel公式
            temp_workbook: 临时工作簿对象

        Returns:
            any: 计算结果
        """
        from xlcalculator import Evaluator, ModelCompiler

        # 在临时单元格中设置要计算的公式
        temp_sheet = temp_workbook.active
        calc_cell = temp_sheet["Z1"]
        calc_cell.value = f"={formula}"

        # 一次性保存（包含数据和公式）
        temp_workbook.save(temp_file_path)

        # 编译模型
        compiler = ModelCompiler()
        model = compiler.read_and_parse_archive(temp_file_path)
        evaluator = Evaluator(model)

        # 计算Z1位置的公式
        calculated_value = evaluator.evaluate("Calculation!Z1")
        return calculated_value

    def _fast_calculate(self, formula: str) -> any:
        """快速计算纯算术公式（无单元格引用），使用 Python 内置 eval。

        支持: SUM, AVG, MIN, MAX, COUNT, ABS, ROUND, INT, MOD, POWER, SQRT,
              IF, AND, OR, NOT, CONCATENATE, LEFT, RIGHT, MID, LEN, UPPER, LOWER,
              TRIM, TEXT, DATE, TODAY, NOW 等常用函数，以及 + - * / ^ () 运算。

        Args:
            formula: Excel 公式（不含等号）

        Returns:
            计算结果
        """
        from datetime import date, datetime

        # Excel 函数名 → Python 映射
        safe_names = {
            # 数学函数
            "SUM": lambda *args: sum(args) if args else 0,
            "AVG": lambda *args: sum(args) / len(args) if args else 0,
            "MIN": min,
            "MAX": max,
            "COUNT": lambda *args: sum(1 for a in args if a is not None),
            "ABS": abs,
            "ROUND": round,
            "INT": int,
            "MOD": lambda a, b: a % b,
            "POWER": pow,
            "SQRT": math.sqrt,
            "FLOOR": math.floor,
            "CEILING": math.ceil,
            "LN": math.log,
            "LOG": math.log10,
            "LOG10": math.log10,
            "EXP": math.exp,
            "PI": math.pi,
            "E": math.e,
            "SIN": math.sin,
            "COS": math.cos,
            "TAN": math.tan,
            "DEGREES": math.degrees,
            "RADIANS": math.radians,
            # 逻辑函数
            "IF": lambda cond, t, f: t if cond else f,
            "AND": lambda *args: all(args),
            "OR": lambda *args: any(args),
            "NOT": lambda x: not x,
            "TRUE": True,
            "FALSE": False,
            "NA": None,
            # 文本函数
            "CONCATENATE": lambda *args: "".join(str(a) for a in args),
            "LEFT": lambda s, n=1: str(s)[: int(n)],
            "RIGHT": lambda s, n=1: str(s)[-int(n) :] if int(n) > 0 else "",
            "MID": lambda s, start, length: str(s)[int(start) - 1 : int(start) - 1 + int(length)],
            "LEN": lambda s: len(str(s)),
            "UPPER": lambda s: str(s).upper(),
            "LOWER": lambda s: str(s).lower(),
            "TRIM": lambda s: str(s).strip(),
            "TEXT": lambda v, fmt=None: str(v),
            "VALUE": lambda v: float(v) if "." in str(v) else int(v),
            # 日期函数
            "TODAY": date.today(),
            "NOW": datetime.now(),
            "DATE": lambda y, m, d: date(int(y), int(m), int(d)),
            "YEAR": lambda d: d.year if hasattr(d, "year") else 0,
            "MONTH": lambda d: d.month if hasattr(d, "month") else 0,
            "DAY": lambda d: d.day if hasattr(d, "day") else 0,
            # 统计
            "MEDIAN": lambda *args: sorted(args)[len(args) // 2] if args else 0,
            "MODE": lambda *args: max(set(args), key=args.count) if args else 0,
        }

        try:
            # 将 Excel 风格公式转为 Python 表达式
            py_expr = formula

            # 处理 Excel 比较运算符（Python 已支持）
            # 处理 & 连接符 → +
            py_expr = re.sub(r"&", "+", py_expr)

            # 处理 != (Excel 用 <>)
            py_expr = re.sub(r"<>", "!=", py_expr)

            result = eval(py_expr, {"__builtins__": {}}, safe_names)
            return result
        except Exception as e:
            logger.warning(f"快速计算失败，回退到 xlcalculator: {e}")
            # 回退：创建临时工作簿用 xlcalculator 计算
            temp_workbook = Workbook()
            temp_sheet = temp_workbook.active
            temp_sheet.title = "Calculation"
            temp_file_path = TempFileManager.create_temp_excel_file(suffix=".xlsx")
            calc_cell = temp_sheet["Z1"]
            calc_cell.value = f"={formula}"
            temp_workbook.save(temp_file_path)

            from xlcalculator import Evaluator, ModelCompiler

            compiler = ModelCompiler()
            model = compiler.read_and_parse_archive(temp_file_path)
            evaluator = Evaluator(model)
            return evaluator.evaluate("Calculation!Z1")

    def _fallback_calculation(self, temp_file_path: str, formula: str) -> any:
        """备用计算方法

        Args:
            temp_file_path: 临时文件路径
            formula: Excel公式

        Returns:
            any: 计算结果
        """
        # 重新加载工作簿获取数据 - 使用只读模式
        data_workbook = load_workbook(temp_file_path, data_only=True, read_only=True)
        data_sheet = data_workbook["Calculation"]

        # 尝试基础的公式解析
        calculated_value = self._basic_formula_parse(formula, data_sheet)
        data_workbook.close()

        return calculated_value

    def _get_result_type(self, value) -> str:
        """确定结果类型

        Args:
            value: 待判断的值

        Returns:
            str: 结果类型 ("null", "boolean", "number", "text", "date", "unknown")
        """
        if value is None:
            return "null"
        elif isinstance(value, bool):
            return "boolean"  # 布尔值要在数值之前检查，因为bool是int的子类
        elif isinstance(value, (int, float)):
            return "number"
        elif isinstance(value, str):
            return "text"
        else:
            try:
                # 检查是否是xlcalculator的数字类型
                try:
                    from xlcalculator.xlfunctions.func_xltypes import Number

                    if isinstance(value, Number):
                        return "number"
                except ImportError:
                    pass

                # 检查是否是日期
                if isinstance(value, (datetime, date)):
                    return "date"

                # 如果是xlcalculator的类型，尝试获取实际值
                if hasattr(value, "value"):
                    actual_value = value.value
                    if isinstance(actual_value, (int, float)):
                        return "number"
                    elif isinstance(actual_value, str):
                        return "text"
                    elif isinstance(actual_value, bool):
                        return "boolean"

            except Exception:
                pass
            return "unknown"

    def _safe_eval_expr(self, expr: str):
        """安全求值简单数学表达式（仅支持数字和+-*/运算符）

        Args:
            expr: 数学表达式字符串

        Returns:
            Any: 计算结果，如果表达式不合法则返回 None
        """
        try:
            tree = ast.parse(expr, mode="eval")
            # 只允许数字常量和算术运算符
            for node in ast.walk(tree):
                if isinstance(node, (ast.Expression, ast.BinOp, ast.UnaryOp, ast.operator)):
                    pass
                elif isinstance(node, ast.Constant):
                    if not isinstance(node.value, (int, float)):
                        return None
                else:
                    return None
            return eval(compile(tree, "<expr>", "eval"))
        except (SyntaxError, TypeError, ValueError, ZeroDivisionError):
            return None

    # 范围统计函数分发表: (正则模式, 处理函数)
    # 模式组: (start_cell, end_cell) 或 (start_cell, end_cell, extra_param)
    _RANGE_FORMULAS = [
        (r"SUM\(([A-Z]+\d+):([A-Z]+\d+)\)", "_formula_range_sum"),
        (r"AVERAGE\(([A-Z]+\d+):([A-Z]+\d+)\)", "_formula_range_average"),
        (r"COUNT\(([A-Z]+\d+):([A-Z]+\d+)\)", "_formula_range_count"),
        (r"MIN\(([A-Z]+\d+):([A-Z]+\d+)\)", "_formula_range_min"),
        (r"MAX\(([A-Z]+\d+):([A-Z]+\d+)\)", "_formula_range_max"),
        (r"MEDIAN\(([A-Z]+\d+):([A-Z]+\d+)\)", "_formula_range_median"),
        (r"STDEV(?:\.S)?\(([A-Z]+\d+):([A-Z]+\d+)\)", "_formula_range_stdev"),
        (r"VAR(?:\.S)?\(([A-Z]+\d+):([A-Z]+\d+)\)", "_formula_range_var"),
        (
            r"PERCENTILE\(([A-Z]+\d+):([A-Z]+\d+),\s*([0-9.]+)\)",
            "_formula_range_percentile",
        ),
        (r"QUARTILE\(([A-Z]+\d+):([A-Z]+\d+),\s*([0-3])\)", "_formula_range_quartile"),
        (r"MODE(?:\.SNGL)?\(([A-Z]+\d+):([A-Z]+\d+)\)", "_formula_range_mode"),
        (r"SKEW\(([A-Z]+\d+):([A-Z]+\d+)\)", "_formula_range_skew"),
        (r"KURT\(([A-Z]+\d+):([A-Z]+\d+)\)", "_formula_range_kurt"),
        (r"GEOMEAN\(([A-Z]+\d+):([A-Z]+\d+)\)", "_formula_range_geomean"),
        (r"HARMEAN\(([A-Z]+\d+):([A-Z]+\d+)\)", "_formula_range_harmean"),
        # 条件统计函数（第三组是条件表达式）
        (
            r'COUNTIF\(([A-Z]+\d+):([A-Z]+\d+),\s*"?([^"]+)"?\)',
            "_formula_range_countif",
        ),
        (r'SUMIF\(([A-Z]+\d+):([A-Z]+\d+),\s*"?([^"]+)"?\)', "_formula_range_sumif"),
        (
            r'AVERAGEIF\(([A-Z]+\d+):([A-Z]+\d+),\s*"?([^"]+)"?\)',
            "_formula_range_averageif",
        ),
    ]

    def _basic_formula_parse(self, formula: str, sheet) -> any:
        """增强的基础公式解析器 - 支持numpy统计函数"""
        formula = formula.strip()

        # 简单数学表达式（优先处理，不依赖工作表）
        if re.match(r"^[\d\+\-\*\/\s\(\)\.]+$", formula):
            result = self._safe_eval_expr(formula)
            if result is not None:
                return result

        # 数字列表统计函数（不依赖工作表）
        list_match = re.match(r"(SUM|AVERAGE)\(([\d\s\,]+)\)", formula, re.IGNORECASE)
        if list_match:
            func_name, numbers_str = list_match.groups()
            try:
                numbers = [float(n.strip()) for n in numbers_str.split(",") if n.strip()]
                if not numbers:
                    return 0
                if func_name.upper() == "SUM":
                    return sum(numbers)
                return sum(numbers) / len(numbers)
            except Exception as e:
                logger.warning(f"{func_name}函数解析失败: {numbers_str}, 错误: {e}")

        # 范围统计函数分发
        for pattern, handler_name in self._RANGE_FORMULAS:
            match = re.match(pattern, formula, re.IGNORECASE)
            if match:
                handler = getattr(self, handler_name)
                return handler(sheet, *match.groups())

        # IF函数
        if_match = re.match(r'IF\((.+),\s*"?([^,"]+)"?,\s*"?([^,"]+)"?\)', formula, re.IGNORECASE)
        if if_match:
            return self._formula_if(if_match.group(1), if_match.group(2), if_match.group(3))

        # CONCATENATE函数
        concat_match = re.match(r"CONCATENATE\((.+)\)", formula, re.IGNORECASE)
        if concat_match:
            args = concat_match.group(1).split(",")
            return "".join(arg.strip().strip('"') for arg in args)

        return None

    # ---- 范围统计函数分发处理方法 ----

    def _formula_range_sum(self, sheet, start, end):
        return sum(self._get_range_values(sheet, start, end))

    def _formula_range_average(self, sheet, start, end):
        return self._numpy_average(self._get_range_values(sheet, start, end))

    def _formula_range_count(self, sheet, start, end):
        return len(self._get_range_values(sheet, start, end))

    def _formula_range_min(self, sheet, start, end):
        return self._numpy_min(self._get_range_values(sheet, start, end))

    def _formula_range_max(self, sheet, start, end):
        return self._numpy_max(self._get_range_values(sheet, start, end))

    def _formula_range_median(self, sheet, start, end):
        return self._numpy_median(self._get_range_values(sheet, start, end))

    def _formula_range_stdev(self, sheet, start, end):
        return self._numpy_stdev(self._get_range_values(sheet, start, end))

    def _formula_range_var(self, sheet, start, end):
        return self._numpy_var(self._get_range_values(sheet, start, end))

    def _formula_range_percentile(self, sheet, start, end, p):
        return self._numpy_percentile(self._get_range_values(sheet, start, end), float(p))

    def _formula_range_quartile(self, sheet, start, end, q):
        return self._numpy_quartile(self._get_range_values(sheet, start, end), int(q))

    def _formula_range_mode(self, sheet, start, end):
        return self._numpy_mode(self._get_range_values(sheet, start, end))

    def _formula_range_skew(self, sheet, start, end):
        return self._numpy_skew(self._get_range_values(sheet, start, end))

    def _formula_range_kurt(self, sheet, start, end):
        return self._numpy_kurtosis(self._get_range_values(sheet, start, end))

    def _formula_range_geomean(self, sheet, start, end):
        return self._numpy_geomean(self._get_range_values(sheet, start, end))

    def _formula_range_harmean(self, sheet, start, end):
        return self._numpy_harmean(self._get_range_values(sheet, start, end))

    def _formula_range_countif(self, sheet, start, end, condition):
        return self._apply_condition(self._get_range_values(sheet, start, end), condition, "count")

    def _formula_range_sumif(self, sheet, start, end, condition):
        return self._apply_condition(self._get_range_values(sheet, start, end), condition, "sum")

    def _formula_range_averageif(self, sheet, start, end, condition):
        return self._apply_condition(self._get_range_values(sheet, start, end), condition, "average")

    def _formula_if(self, condition: str, true_val: str, false_val: str):
        """IF函数简单实现（支持>和<比较）"""
        for op_str, op_fn in [(">", float.__gt__), ("<", float.__lt__)]:
            if op_str in condition:
                parts = condition.split(op_str)
                if len(parts) == 2:
                    try:
                        if op_fn(float(parts[0].strip()), float(parts[1].strip())):
                            return true_val
                        return false_val
                    except (ValueError, TypeError):
                        pass
        return None

    @staticmethod
    def _apply_condition(values: list, condition: str, mode: str):
        """通用条件筛选 — 支持 COUNTIF/SUMIF/AVERAGEIF

        Args:
            values: 数值列表
            condition: 条件表达式（如 ">50", "<=100", "=25", "50"）
            mode: 'count' | 'sum' | 'average'

        Returns:
            float or int: 筛选后的聚合结果
        """
        # 解析条件
        if condition.startswith(">="):
            op, threshold = ">=", float(condition[2:])
        elif condition.startswith("<="):
            op, threshold = "<=", float(condition[2:])
        elif condition.startswith(">"):
            op, threshold = ">", float(condition[1:])
        elif condition.startswith("<"):
            op, threshold = "<", float(condition[1:])
        elif condition.startswith("="):
            op, threshold = "=", float(condition[1:])
        else:
            op, threshold = "=", float(condition)

        # 筛选
        if op == ">":
            filtered = [v for v in values if v > threshold]
        elif op == "<":
            filtered = [v for v in values if v < threshold]
        elif op == ">=":
            filtered = [v for v in values if v >= threshold]
        elif op == "<=":
            filtered = [v for v in values if v <= threshold]
        else:
            filtered = [v for v in values if v == threshold]

        # 聚合
        if mode == "count":
            return len(filtered)
        elif mode == "sum":
            return sum(filtered)
        else:  # average
            return sum(filtered) / len(filtered) if filtered else 0

    def _get_range_values(self, sheet, start_cell: str, end_cell: str) -> list:
        """获取范围内的数值列表

        Args:
            sheet: openpyxl工作表对象
            start_cell: 起始单元格地址（如 "A1"）
            end_cell: 结束单元格地址（如 "C10"）

        Returns:
            list: 数值列表
        """

        min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
        values = []

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    values.append(float(cell.value))

        return values

    # ==================== Numpy统计函数实现 ====================

    def _numpy_op(
        self,
        values: list,
        np_func_name,
        fallback_func,
        min_values: int = 0,
        numpy_kwargs: dict = None,
    ) -> float:
        """统一的numpy计算+Python降级模式，消除15个numpy方法的重复try/except

        Args:
            values: 数值列表
            np_func_name: numpy函数名（如'min', 'max', 'median'）或callable
            fallback_func: numpy不可用时的Python降级函数
            min_values: 最小值数量（不足直接返回0）
            numpy_kwargs: 传递给numpy函数的额外参数

        Returns:
            float: 计算结果
        """
        try:
            import numpy as np

            if len(values) < min_values:
                return 0
            kwargs = numpy_kwargs or {}
            if callable(np_func_name):
                np_func = np_func_name
            else:
                np_func = getattr(np, np_func_name)
            return float(np_func(values, **kwargs))
        except Exception:
            return fallback_func(values) if values else 0

    @staticmethod
    def _python_median(values: list) -> float:
        """Python原生中位数计算"""
        sorted_values = sorted(values)
        n = len(sorted_values)
        if n % 2 == 1:
            return sorted_values[n // 2]
        return (sorted_values[n // 2 - 1] + sorted_values[n // 2]) / 2

    @staticmethod
    def _python_stdev(values: list) -> float:
        """Python原生样本标准差"""
        mean = sum(values) / len(values)
        variance = sum((x - mean) ** 2 for x in values) / (len(values) - 1)
        return variance**0.5

    @staticmethod
    def _python_var(values: list) -> float:
        """Python原生样本方差"""
        mean = sum(values) / len(values)
        return sum((x - mean) ** 2 for x in values) / (len(values) - 1)

    @staticmethod
    def _python_percentile(values: list, p: float) -> float:
        """Python原生百分位数"""
        sorted_values = sorted(values)
        k = p * (len(sorted_values) - 1)
        f = int(k)
        c = k - f
        if f + 1 < len(sorted_values):
            return sorted_values[f] * (1 - c) + sorted_values[f + 1] * c
        return sorted_values[f]

    def _numpy_average(self, values: list) -> float:
        """计算平均值"""
        return self._numpy_op(values, "mean", lambda v: sum(v) / len(v) if v else 0)

    def _numpy_min(self, values: list) -> float:
        """计算最小值"""
        return self._numpy_op(values, "min", min)

    def _numpy_max(self, values: list) -> float:
        """计算最大值"""
        return self._numpy_op(values, "max", max)

    def _numpy_median(self, values: list) -> float:
        """计算中位数"""
        return self._numpy_op(values, "median", self._python_median)

    def _numpy_stdev(self, values: list) -> float:
        """计算样本标准差"""
        return self._numpy_op(values, "std", self._python_stdev, min_values=2, numpy_kwargs={"ddof": 1})

    def _numpy_var(self, values: list) -> float:
        """计算样本方差"""
        return self._numpy_op(values, "var", self._python_var, min_values=2, numpy_kwargs={"ddof": 1})

    def _numpy_percentile(self, values: list, percentile: float) -> float:
        """计算百分位数"""
        return self._numpy_op(
            values,
            "percentile",
            lambda v: self._python_percentile(v, percentile),
            numpy_kwargs={"q": percentile * 100},
        )

    def _numpy_quartile(self, values: list, quartile: int) -> float:
        """计算四分位数"""
        quartile_map = {0: 0, 1: 0.25, 2: 0.5, 3: 0.75}
        return self._numpy_percentile(values, quartile_map.get(quartile, 0.5))

    def _numpy_mode(self, values: list) -> float:
        """计算众数"""
        try:
            from scipy import stats

            if not values:
                return 0
            mode_result = stats.mode(values, keepdims=True)
            return float(mode_result[0][0])
        except Exception:
            # 简单实现：返回最频繁出现的值
            if not values:
                return 0
            counts = Counter(values)
            return float(counts.most_common(1)[0][0])

    def _numpy_skew(self, values: list) -> float:
        """计算偏度"""
        try:
            from scipy import stats

            if len(values) < 3:
                return 0
            return float(stats.skew(values))
        except Exception:
            return 0

    def _numpy_kurtosis(self, values: list) -> float:
        """计算峰度"""
        try:
            from scipy import stats

            if len(values) < 4:
                return 0
            return float(stats.kurtosis(values))
        except Exception:
            return 0

    def _numpy_geomean(self, values: list) -> float:
        """计算几何平均数"""
        try:
            from scipy import stats

            if not values or any(v <= 0 for v in values):
                return 0
            return float(stats.gmean(values))
        except Exception:
            if not values or any(v <= 0 for v in values):
                return 0
            product = 1
            for v in values:
                product *= v
            return product ** (1.0 / len(values))

    def _numpy_harmean(self, values: list) -> float:
        """计算调和平均数"""
        try:
            from scipy import stats

            if not values or any(v <= 0 for v in values):
                return 0
            return float(stats.hmean(values))
        except Exception:
            if not values or any(v <= 0 for v in values):
                return 0
            return len(values) / sum(1.0 / v for v in values)

    def _apply_cell_format(self, cell, formatting: dict):
        """应用单元格格式（v1.9.3+ 增强版）

        Args:
            cell: openpyxl单元格对象
            formatting: 格式配置字典，支持以下键：

                font: {
                    name(str), size(int), bold(bool), italic(bool),
                    color(str/RGB), underline('single'|'double'|'singleAccounting'|'doubleAccounting'),
                    strikethrough(bool)
                }
                fill: {
                    type('solid'|'gradient'|'pattern'),
                    color(str),           # solid 模式的填充色
                    colors(list[str]),    # gradient 模式的渐变色数组
                    gradient_type, patternType, fgColor, bgColor
                }
                alignment: {
                    horizontal, vertical,
                    wrap_text(bool), text_rotation(int, -90~90),
                    indent(int), shrink_to_fit(bool)
                }
                number_format: str
                border: {
                    left/right/top/bottom/diagonal: str|dict(style, color),
                    color(str), diagonalDirection, outline, start, end
                }

        Returns:
            None
        """
        # 字体格式
        if "font" in formatting and formatting["font"] is not None:
            font_config = formatting["font"]
            # 处理 underline 值：支持 'single'/'double'/'singleAccounting'/'doubleAccounting'/True/False
            _underline = font_config.get("underline", None)
            if _underline is True:
                _underline = "single"
            elif _underline is False:
                _underline = "none"

            cell.font = Font(
                name=font_config.get("name", cell.font.name),
                size=font_config.get("size", cell.font.size),
                bold=font_config.get("bold", cell.font.bold),
                italic=font_config.get("italic", cell.font.italic),
                color=font_config.get("color", cell.font.color),
                underline=_underline if _underline is not None else cell.font.underline,
                strikethrough=font_config.get("strikethrough", cell.font.strikethrough),
            )

        # 背景颜色 / 填充
        if "fill" in formatting and formatting["fill"] is not None:
            fill_config = formatting["fill"]
            _fill_type = fill_config.get("type", "solid").lower()
            if _fill_type == "solid":
                cell.fill = PatternFill(
                    start_color=fill_config.get("color", "FFFFFF"),
                    end_color=fill_config.get("color", "FFFFFF"),
                    fill_type="solid",
                )
            elif _fill_type == "gradient":
                from openpyxl.styles.colors import Color as Clr
                from openpyxl.styles.fills import Stop

                _colors = fill_config.get("colors", ["FFFFFF", "D9D9D9"])
                _n = len(_colors)
                cell.fill = GradientFill(
                    type=fill_config.get("gradient_type", "linear"),
                    degree=fill_config.get("degree", 0),
                    stop=[Stop(color=Clr(c), position=i / (_n - 1) if _n > 1 else 0) for i, c in enumerate(_colors)],
                )
            elif _fill_type == "pattern":
                cell.fill = PatternFill(
                    patternType=fill_config.get("patternType", "lightGray"),
                    fgColor=fill_config.get("fgColor", "00000000"),
                    bgColor=fill_config.get("bgColor", "00000000"),
                )

        # 对齐方式（含换行、旋转、缩进、自动换行）
        if "alignment" in formatting and formatting["alignment"] is not None:
            align_config = formatting["alignment"]
            cell.alignment = Alignment(
                horizontal=align_config.get("horizontal", cell.alignment.horizontal),
                vertical=align_config.get("vertical", cell.alignment.vertical),
                wrap_text=align_config.get("wrap_text", cell.alignment.wrap_text),
                text_rotation=align_config.get("text_rotation", cell.alignment.text_rotation),
                indent=align_config.get("indent", cell.alignment.indent),
                shrink_to_fit=align_config.get("shrink_to_fit", cell.alignment.shrink_to_fit),
            )

        # 数字格式（None 值跳过，避免 openpyxl 抛 TypeError 导致文件损坏）
        if "number_format" in formatting and formatting["number_format"] is not None:
            cell.number_format = formatting["number_format"]

        # 行内边框（可选，与 set_borders 工具互补）
        if "border" in formatting and formatting["border"] is not None:
            border_config = formatting["border"]
            from openpyxl.styles import Border as Bdr
            from openpyxl.styles import Side

            def _make_side(cfg):
                if cfg is None:
                    # None 表示不设置该边（保留原值），返回无样式 Side
                    return Side(style=None)
                if isinstance(cfg, str):
                    return Side(style=cfg, color=border_config.get("color", "000000"))
                return Side(
                    style=cfg.get("style", "thin"),
                    color=cfg.get("color", border_config.get("color", "000000")),
                )

            cell.border = Bdr(
                left=_make_side(border_config.get("left")),
                right=_make_side(border_config.get("right")),
                top=_make_side(border_config.get("top")),
                bottom=_make_side(border_config.get("bottom")),
                diagonal=_make_side(border_config.get("diagonal")),
                diagonal_direction=border_config.get("diagonal_direction"),
                outline=border_config.get("outline", True),
                start=border_config.get("start"),
                end=border_config.get("end"),
            )

    def _parse_and_resolve_sheet(self, workbook, range_expression: str, sheet_name: str | None = None):
        """解析范围表达式并获取工作表，消除merge/unmerge/borders的重复逻辑

        Args:
            workbook: openpyxl Workbook实例
            range_expression: 范围表达式
            sheet_name: 工作表名称（可选）

        Returns:
            tuple: (range_info, worksheet)

        Raises:
            SheetNotFoundError: 工作表不存在
        """
        if sheet_name and "!" not in range_expression:
            full_range = f"{sheet_name}!{range_expression}"
        else:
            full_range = range_expression

        range_info = RangeParser.parse_range_expression(full_range)

        if range_info.sheet_name not in workbook.sheetnames:
            raise SheetNotFoundError(f"工作表 '{range_info.sheet_name}' 不存在")

        return range_info, workbook[range_info.sheet_name]

    def _get_worksheet_or_raise(self, workbook, sheet_name: str | None = None):
        """获取工作表（支持可选sheet_name，不存在则用活动表）

        Args:
            workbook: openpyxl Workbook实例
            sheet_name: 工作表名称（可选）

        Returns:
            tuple: (worksheet, resolved_sheet_name)

        Raises:
            SheetNotFoundError: 指定的工作表不存在
        """
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise SheetNotFoundError(f"工作表 '{sheet_name}' 不存在")
            return workbook[sheet_name], sheet_name
        else:
            worksheet = workbook.active
            return worksheet, worksheet.title

    def merge_cells(self, range_expression: str, sheet_name: str | None = None) -> OperationResult:
        """
        合并单元格

        Args:
            range_expression: 范围表达式，如 "A1:C3" 或 "Sheet1!A1:C3"
            sheet_name: 工作表名称（可选，如果range_expression包含工作表名则忽略）

        Returns:
            OperationResult: 操作结果
        """
        try:
            workbook = load_workbook(self.file_path)
            range_info, worksheet = self._parse_and_resolve_sheet(workbook, range_expression, sheet_name)

            worksheet.merge_cells(range_info.cell_range)

            self._safe_save_workbook(workbook, "合并单元格")
            workbook.close()

            return OperationResult(
                success=True,
                message=f"成功合并单元格范围: {range_info.cell_range}",
                data={
                    "merged_range": range_info.cell_range,
                    "sheet_name": range_info.sheet_name,
                },
                metadata={"operation": "merge_cells", "file_path": self.file_path},
            )

        except Exception as e:
            logger.error(f"合并单元格失败: {e}")
            return OperationResult(success=False, error=f"合并单元格失败: {str(e)}")

    def unmerge_cells(self, range_expression: str, sheet_name: str | None = None) -> OperationResult:
        """
        取消合并单元格

        Args:
            range_expression: 范围表达式，如 "A1:C3" 或 "Sheet1!A1:C3"
            sheet_name: 工作表名称（可选，如果range_expression包含工作表名则忽略）

        Returns:
            OperationResult: 操作结果
        """
        try:
            workbook = load_workbook(self.file_path)
            range_info, worksheet = self._parse_and_resolve_sheet(workbook, range_expression, sheet_name)

            worksheet.unmerge_cells(range_info.cell_range)

            self._safe_save_workbook(workbook, "取消合并单元格")
            workbook.close()

            return OperationResult(
                success=True,
                message=f"成功取消合并单元格范围: {range_info.cell_range}",
                data={
                    "unmerged_range": range_info.cell_range,
                    "sheet_name": range_info.sheet_name,
                },
                metadata={"operation": "unmerge_cells", "file_path": self.file_path},
            )

        except Exception as e:
            logger.error(f"取消合并单元格失败: {e}")
            return OperationResult(success=False, error=f"取消合并单元格失败: {str(e)}")

    def set_borders(
        self,
        range_expression: str,
        border_style: str = "thin",
        sheet_name: str | None = None,
    ) -> OperationResult:
        """
        设置单元格边框

        Args:
            range_expression: 范围表达式，如 "A1:C3" 或 "Sheet1!A1:C3"
            border_style: 边框样式 ("thin", "thick", "double", "dashed", "dotted")
            sheet_name: 工作表名称（可选）

        Returns:
            OperationResult: 操作结果
        """
        try:
            workbook = load_workbook(self.file_path)
            range_info, worksheet = self._parse_and_resolve_sheet(workbook, range_expression, sheet_name)

            # 创建边框样式
            side = Side(style=border_style)
            border = Border(left=side, right=side, top=side, bottom=side)

            # 应用边框到指定范围
            cell_count = 0
            try:
                # 尝试直接使用范围
                for row in worksheet[range_info.cell_range]:
                    if hasattr(row, "__iter__"):  # 确保row是可迭代的
                        for cell in row:
                            cell.border = border
                            cell_count += 1
                    else:  # 如果是单个单元格
                        row.border = border
                        cell_count += 1
            except TypeError:
                # 如果cell_range不是预期的格式，尝试其他方法
                min_col, min_row, max_col, max_row = range_boundaries(range_info.cell_range)
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = border
                        cell_count += 1

            # 保存文件
            self._safe_save_workbook(workbook, "设置边框")
            workbook.close()

            return OperationResult(
                success=True,
                message=f"成功设置 {cell_count} 个单元格的边框",
                data={
                    "range": range_info.cell_range,
                    "border_style": border_style,
                    "cell_count": cell_count,
                    "sheet_name": range_info.sheet_name,
                },
                metadata={"operation": "set_borders", "file_path": self.file_path},
            )

        except Exception as e:
            logger.error(f"设置边框失败: {e}")
            return OperationResult(success=False, error=f"设置边框失败: {str(e)}")

    def set_row_height(self, row_number: int, height: float, sheet_name: str | None = None) -> OperationResult:
        """
        设置行高

        Args:
            row_number: 行号（从1开始）
            height: 行高（磅值）
            sheet_name: 工作表名称（可选，使用活动工作表）

        Returns:
            OperationResult: 操作结果
        """
        try:
            workbook = load_workbook(self.file_path)
            worksheet, sheet_name = self._get_worksheet_or_raise(workbook, sheet_name)

            # 设置行高
            worksheet.row_dimensions[row_number].height = height

            # 保存文件
            self._safe_save_workbook(workbook, "设置行高")
            workbook.close()

            return OperationResult(
                success=True,
                message=f"成功设置第 {row_number} 行的高度为 {height} 磅",
                data={
                    "row_number": row_number,
                    "height": height,
                    "sheet_name": sheet_name,
                },
                metadata={"operation": "set_row_height", "file_path": self.file_path},
            )

        except Exception as e:
            logger.error(f"设置行高失败: {e}")
            return OperationResult(success=False, error=f"设置行高失败: {str(e)}")

    def set_column_width(self, column: str, width: float, sheet_name: str | None = None) -> OperationResult:
        """
        设置列宽

        Args:
            column: 列标识符，如 "A", "B", "C"
            width: 列宽（字符单位）
            sheet_name: 工作表名称（可选，使用活动工作表）

        Returns:
            OperationResult: 操作结果
        """
        try:
            workbook = load_workbook(self.file_path)
            worksheet, sheet_name = self._get_worksheet_or_raise(workbook, sheet_name)

            # 设置列宽
            worksheet.column_dimensions[column.upper()].width = width

            # 保存文件
            self._safe_save_workbook(workbook, "设置列宽")
            workbook.close()

            return OperationResult(
                success=True,
                message=f"成功设置列 {column.upper()} 的宽度为 {width} 字符",
                data={
                    "column": column.upper(),
                    "width": width,
                    "sheet_name": sheet_name,
                },
                metadata={"operation": "set_column_width", "file_path": self.file_path},
            )

        except Exception as e:
            logger.error(f"设置列宽失败: {e}")
            return OperationResult(success=False, error=f"设置列宽失败: {str(e)}")
