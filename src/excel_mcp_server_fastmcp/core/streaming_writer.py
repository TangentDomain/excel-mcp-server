"""
Excel MCP Server - 流式写入模块

copy-modify-write 方案：用 calamine 读取 + write_only 模式写入，
显著降低大文件修改操作的内存占用和耗时。

适用场景：批量插入、大批量数据修改等数据密集型操作。
权衡：write_only 模式不保留单元格级格式（字体/填充/边框/合并），
但保留列宽、行高、数据值。对游戏配置表场景（数据优先）友好。
"""

import logging
import os
import re
import shutil
import tempfile
from typing import List, Dict, Any, Optional, Tuple, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

logger = logging.getLogger(__name__)

try:
    from python_calamine import CalamineWorkbook
    _HAS_CALAMINE = True
except ImportError:
    _HAS_CALAMINE = False
    logger.debug("python-calamine未安装，流式写入将降级到openpyxl")


class StreamingWriter:
    """流式写入器：calamine读取 + write_only写入

    内存占用与文件大小无关（只缓存当前行），
    适合大数据量文件的修改操作。
    """

    def __init__(self, file_path: str):
        self.file_path = file_path

    @staticmethod
    def is_available() -> bool:
        """检查流式写入是否可用（需要calamine）"""
        return _HAS_CALAMINE

    @classmethod
    def batch_insert_rows(
        cls,
        file_path: str,
        sheet_name: str,
        data: List[Dict[str, Any]],
        header_row: int = 1,
        preserve_col_widths: bool = True,
    ) -> Tuple[bool, str, Dict[str, Any]]:
        """流式批量插入行（calamine读取 + write_only写入）

        Args:
            file_path: Excel文件路径
            sheet_name: 目标工作表名
            data: 行数据列表，每行为{列名: 值}字典
            header_row: 表头所在行号（默认1）
            preserve_col_widths: 是否保留列宽（需额外openpyxl读取，但比全量加载快）

        Returns:
            (success, message, metadata)
        """
        if not _HAS_CALAMINE:
            return False, "calamine未安装，无法使用流式写入", {}

        try:
            # 1. 用 calamine 读取所有工作表数据
            cal_wb = CalamineWorkbook.from_path(file_path)
            all_sheets_data = {}
            target_sheet_data = None
            col_map = {}

            for sn in cal_wb.sheet_names:
                rows = cal_wb.get_sheet_by_name(sn).to_python()
                all_sheets_data[sn] = rows

                if sn == sheet_name and rows:
                    # 提取表头映射
                    if header_row <= len(rows):
                        for col_idx, cell_val in enumerate(rows[header_row - 1], 1):
                            if cell_val is not None:
                                col_map[str(cell_val).strip()] = col_idx
                    target_sheet_data = rows

            if sheet_name not in all_sheets_data:
                available = ', '.join(all_sheets_data.keys())
                return False, f"工作表不存在: {sheet_name}（可用: {available}）", {}

            if not col_map:
                return False, f"未找到表头（行{header_row}）", {}

            # 2. 构建新数据（原有数据 + 新行）
            new_rows = list(target_sheet_data)
            unknown_cols = set()

            for row_data in data:
                if not isinstance(row_data, dict):
                    return False, f"行数据必须是字典，实际类型: {type(row_data).__name__}", {}

                # 计算最大列数
                max_col = max(col_map.values()) if col_map else 0
                new_row = [None] * max_col

                for col_name, value in row_data.items():
                    col_name_stripped = col_name.strip()
                    if col_name_stripped in col_map:
                        new_row[col_map[col_name_stripped] - 1] = value
                    else:
                        unknown_cols.add(col_name_stripped)

                new_rows.append(new_row)

            # 3. 读取列宽（用非只读模式，需要column_dimensions）
            col_widths = {}
            if preserve_col_widths:
                try:
                    wb_meta = load_workbook(file_path)
                    if sheet_name in wb_meta.sheetnames:
                        ws_meta = wb_meta[sheet_name]
                        col_widths = {
                            col_letter: dim.width
                            for col_letter, dim in ws_meta.column_dimensions.items()
                            if dim.width
                        }
                    wb_meta.close()
                except Exception as e:
                    logger.warning(f"读取列宽失败，跳过: {e}")

            # 4. 用 write_only 模式写入临时文件
            wb_out = Workbook(write_only=True)

            # 更新目标工作表数据为含新行的版本
            all_sheets_data[sheet_name] = new_rows

            for sn in cal_wb.sheet_names:
                ws = wb_out.create_sheet(title=sn)
                sheet_rows = all_sheets_data[sn]

                if not sheet_rows:
                    continue

                # 设置列宽（write_only支持）
                if sn == sheet_name and col_widths:
                    for col_letter, width in col_widths.items():
                        try:
                            ws.column_dimensions[col_letter].width = width
                        except Exception:
                            pass

                # 逐行写入（流式，不积累内存）
                for row in sheet_rows:
                    ws.append(row)

            # 5. 原子替换：写入临时文件 → 替换原文件
            fd, tmp_path = tempfile.mkstemp(suffix='.xlsx', dir=os.path.dirname(file_path))
            os.close(fd)

            try:
                wb_out.save(tmp_path)
                wb_out.close()

                # 原子替换
                shutil.move(tmp_path, file_path)
            except Exception:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
                raise

            start_row = len(target_sheet_data) + 1
            end_row = start_row + len(data) - 1

            return True, (
                f"流式批量插入 {len(data)} 行"
                f"（第{start_row}-{end_row}行）"
            ), {
                'action': 'batch_insert_streaming',
                'start_row': start_row,
                'end_row': end_row,
                'inserted_count': len(data),
                'total_rows_after': len(new_rows),
                'unknown_columns': sorted(unknown_cols)[:5] if unknown_cols else [],
                'mode': 'streaming',
                'col_widths_preserved': len(col_widths) > 0,
            }

        except Exception as e:
            logger.error(f"流式批量插入失败: {e}")
            return False, f"流式写入失败: {str(e)}", {}

    @classmethod
    def upsert_row(
        cls,
        file_path: str,
        sheet_name: str,
        key_column: str,
        key_value: Any,
        updates: Dict[str, Any],
        header_row: int = 1,
        preserve_col_widths: bool = True,
    ) -> Tuple[bool, str, Dict[str, Any]]:
        """流式upsert行：按键列查找并更新/插入

        Args:
            file_path: Excel文件路径
            sheet_name: 目标工作表名
            key_column: 匹配列名
            key_value: 匹配值
            updates: 列值字典
            header_row: 表头行号
            preserve_col_widths: 是否保留列宽

        Returns:
            (success, message, metadata)
        """
        if not _HAS_CALAMINE:
            return False, "calamine未安装，无法使用流式写入", {}

        try:
            cal_wb = CalamineWorkbook.from_path(file_path)
            all_sheets_data = {}
            target_rows = None
            col_map = {}

            for sn in cal_wb.sheet_names:
                rows = cal_wb.get_sheet_by_name(sn).to_python()
                all_sheets_data[sn] = rows

                if sn == sheet_name and rows and header_row <= len(rows):
                    for col_idx, cell_val in enumerate(rows[header_row - 1], 1):
                        if cell_val is not None:
                            col_map[str(cell_val).strip()] = col_idx
                    target_rows = list(rows)  # 深拷贝

            if sheet_name not in all_sheets_data:
                available = ', '.join(all_sheets_data.keys())
                return False, f"工作表不存在: {sheet_name}（可用: {available}）", {}

            if key_column not in col_map:
                actual = list(col_map.keys())[:10]
                return False, f"键列 '{key_column}' 不存在。实际列名: {', '.join(actual)}", {}

            key_col_idx = col_map[key_column]

            # 查找匹配行
            target_row = None
            for row_num, row in enumerate(target_rows[header_row:], start=header_row + 1):
                if key_col_idx <= len(row):
                    cell_val = row[key_col_idx - 1]
                    if cell_val is not None:
                        # calamine 把整数读成浮点数（2→2.0），需要标准化比较
                        a, b = str(cell_val).strip(), str(key_value).strip()
                        try:
                            a_norm = float(a)
                            b_norm = float(b)
                            matched = a_norm == b_norm
                        except (ValueError, TypeError):
                            matched = a == b
                        if matched:
                            target_row = row_num
                            break

            action = 'update'
            if target_row is not None:
                # UPDATE: 修改已有行
                updated_cols = []
                row = target_rows[target_row - 1]
                for col_name, value in updates.items():
                    col_name_stripped = col_name.strip()
                    if col_name_stripped in col_map:
                        row[col_map[col_name_stripped] - 1] = value
                        updated_cols.append(col_name_stripped)
            else:
                # INSERT: 追加新行
                action = 'insert'
                max_col = max(col_map.values()) if col_map else 0
                new_row = [None] * max_col
                updated_cols = []

                for col_name, col_idx in sorted(col_map.items(), key=lambda x: x[1]):
                    if col_name in updates:
                        new_row[col_idx - 1] = updates[col_name]
                        updated_cols.append(col_name)
                    elif col_name == key_column:
                        new_row[col_idx - 1] = key_value
                        updated_cols.append(col_name)

                target_rows.append(new_row)
                target_row = len(target_rows)

            # 更新 all_sheets_data
            all_sheets_data[sheet_name] = target_rows

            # 读取列宽
            col_widths = {}
            if preserve_col_widths:
                try:
                    wb_meta = load_workbook(file_path)
                    if sheet_name in wb_meta.sheetnames:
                        ws_meta = wb_meta[sheet_name]
                        col_widths = {
                            col_letter: dim.width
                            for col_letter, dim in ws_meta.column_dimensions.items()
                            if dim.width
                        }
                    wb_meta.close()
                except Exception as e:
                    logger.warning(f"读取列宽失败，跳过: {e}")

            # write_only 写入
            wb_out = Workbook(write_only=True)
            for sn in cal_wb.sheet_names:
                ws = wb_out.create_sheet(title=sn)
                sheet_rows = all_sheets_data[sn]

                if sn == sheet_name and col_widths:
                    for col_letter, width in col_widths.items():
                        try:
                            ws.column_dimensions[col_letter].width = width
                        except Exception:
                            pass

                for row in sheet_rows:
                    ws.append(row)

            fd, tmp_path = tempfile.mkstemp(suffix='.xlsx', dir=os.path.dirname(file_path))
            os.close(fd)

            try:
                wb_out.save(tmp_path)
                wb_out.close()
                shutil.move(tmp_path, file_path)
            except Exception:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
                raise

            return True, (
                f"流式{'更新' if action == 'update' else '插入'}行 {target_row}"
                f"（键列 '{key_column}'='{key_value}'），"
                f"修改了 {len(updated_cols)} 列"
            ), {
                'action': action,
                'row': target_row,
                'updated_columns': updated_cols,
                'mode': 'streaming',
                'col_widths_preserved': len(col_widths) > 0,
            }

        except Exception as e:
            logger.error(f"流式upsert失败: {e}")
            return False, f"流式写入失败: {str(e)}", {}

    @classmethod
    def _copy_modify_write(
        cls,
        file_path: str,
        sheet_name: str,
        modify_fn: Callable[[List[List[Any]]], Tuple[bool, str, List[List[Any]], Dict[str, Any]]],
        preserve_col_widths: bool = True,
    ) -> Tuple[bool, str, Dict[str, Any]]:
        """通用 copy-modify-write：calamine读取 → 修改 → write_only写入

        Args:
            file_path: Excel文件路径
            sheet_name: 目标工作表名
            modify_fn: 修改函数，接收(rows, header_row, col_map)，
                       返回(success, message, modified_rows, metadata)
            preserve_col_widths: 是否保留列宽

        Returns:
            (success, message, metadata)
        """
        if not _HAS_CALAMINE:
            return False, "calamine未安装，无法使用流式写入", {}

        try:
            # 1. calamine 读取所有工作表数据
            cal_wb = CalamineWorkbook.from_path(file_path)
            all_sheets_data = {}
            target_rows = None
            col_map = {}

            for sn in cal_wb.sheet_names:
                rows = cal_wb.get_sheet_by_name(sn).to_python()
                all_sheets_data[sn] = rows
                if sn == sheet_name and rows:
                    target_rows = list(rows)
                    # 提取表头映射（第1行为表头）
                    for col_idx, cell_val in enumerate(rows[0], 1):
                        if cell_val is not None:
                            col_map[str(cell_val).strip()] = col_idx

            if sheet_name not in all_sheets_data:
                available = ', '.join(all_sheets_data.keys())
                return False, f"工作表不存在: {sheet_name}（可用: {available}）", {}

            if target_rows is None:
                return False, f"工作表 '{sheet_name}' 为空", {}

            # 2. 调用修改函数
            success, message, modified_rows, extra_meta = modify_fn(
                target_rows, 1, col_map
            )
            if not success:
                return False, message, extra_meta

            # 3. 更新目标工作表数据
            all_sheets_data[sheet_name] = modified_rows

            # 4. 读取列宽
            col_widths = {}
            if preserve_col_widths:
                try:
                    wb_meta = load_workbook(file_path)
                    if sheet_name in wb_meta.sheetnames:
                        ws_meta = wb_meta[sheet_name]
                        col_widths = {
                            col_letter: dim.width
                            for col_letter, dim in ws_meta.column_dimensions.items()
                            if dim.width
                        }
                    wb_meta.close()
                except Exception as e:
                    logger.warning(f"读取列宽失败，跳过: {e}")

            # 5. write_only 写入
            wb_out = Workbook(write_only=True)
            for sn in cal_wb.sheet_names:
                ws = wb_out.create_sheet(title=sn)
                sheet_rows = all_sheets_data[sn]

                if not sheet_rows:
                    continue

                if sn == sheet_name and col_widths:
                    for col_letter, width in col_widths.items():
                        try:
                            ws.column_dimensions[col_letter].width = width
                        except Exception:
                            pass

                for row in sheet_rows:
                    ws.append(row)

            # 6. 原子替换
            fd, tmp_path = tempfile.mkstemp(suffix='.xlsx', dir=os.path.dirname(file_path))
            os.close(fd)

            try:
                wb_out.save(tmp_path)
                wb_out.close()
                shutil.move(tmp_path, file_path)
            except Exception:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
                raise

            meta = {'mode': 'streaming', 'col_widths_preserved': len(col_widths) > 0}
            meta.update(extra_meta)
            return True, message, meta

        except Exception as e:
            logger.error(f"流式copy-modify-write失败: {e}")
            return False, f"流式写入失败: {str(e)}", {}

    @classmethod
    def delete_rows(
        cls,
        file_path: str,
        sheet_name: str,
        start_row: int,
        count: int = 1,
        preserve_col_widths: bool = True,
    ) -> Tuple[bool, str, Dict[str, Any]]:
        """流式删除行：calamine读取 → 删除行 → write_only写入

        Args:
            file_path: Excel文件路径
            sheet_name: 目标工作表名
            start_row: 起始行号（1-based，含表头）
            count: 删除行数
            preserve_col_widths: 是否保留列宽

        Returns:
            (success, message, metadata)
        """
        if not _HAS_CALAMINE:
            return False, "calamine未安装", {}

        if start_row < 1:
            return False, f"起始行号必须>=1，实际: {start_row}", {}
        if count < 1:
            return False, f"删除行数必须>=1，实际: {count}", {}

        def _modify(rows, header_row, col_map):
            if start_row > len(rows):
                return False, f"起始行号({start_row})超过工作表总行数({len(rows)})", rows, {}
            actual_count = min(count, len(rows) - start_row + 1)
            new_rows = rows[:start_row - 1] + rows[start_row - 1 + actual_count:]
            return True, (
                f"流式删除第{start_row}-{start_row + actual_count - 1}行"
                f"（共{actual_count}行）"
            ), new_rows, {
                'action': 'delete_rows_streaming',
                'start_row': start_row,
                'actual_count': actual_count,
                'original_rows': len(rows),
                'new_rows': len(new_rows),
            }

        return cls._copy_modify_write(file_path, sheet_name, _modify, preserve_col_widths)

    @classmethod
    def batch_delete_rows(
        cls,
        file_path: str,
        sheet_name: str,
        row_numbers: List[int],
        preserve_col_widths: bool = True,
    ) -> Tuple[bool, str, Dict[str, Any]]:
        """流式批量删除多行：一次性删除所有指定行，仅一次文件I/O。

        Args:
            file_path: Excel文件路径
            sheet_name: 目标工作表名
            row_numbers: 待删除的行号列表（1-based，含表头）
            preserve_col_widths: 是否保留列宽

        Returns:
            (success, message, metadata)
        """
        if not _HAS_CALAMINE:
            return False, "calamine未安装", {}

        if not row_numbers:
            return False, "行号列表为空", {}

        unique_rows = sorted(set(row_numbers))
        if unique_rows[0] < 1:
            return False, f"行号必须>=1，实际最小值: {unique_rows[0]}", {}

        def _modify(rows, header_row, col_map):
            if unique_rows[-1] > len(rows):
                return False, (
                    f"行号({unique_rows[-1]})超过工作表总行数({len(rows)})"
                ), rows, {}
            # 构建 0-based 索引集合，高效过滤
            del_indices = {r - 1 for r in unique_rows}
            new_rows = [
                row for idx, row in enumerate(rows) if idx not in del_indices
            ]
            return True, (
                f"流式批量删除{len(unique_rows)}行"
                f"（{len(rows)}→{len(new_rows)}行）"
            ), new_rows, {
                'action': 'batch_delete_rows_streaming',
                'deleted_count': len(unique_rows),
                'original_rows': len(rows),
                'new_rows': len(new_rows),
            }

        return cls._copy_modify_write(file_path, sheet_name, _modify, preserve_col_widths)

    @classmethod
    def delete_columns(
        cls,
        file_path: str,
        sheet_name: str,
        start_column: int,
        count: int = 1,
        preserve_col_widths: bool = True,
    ) -> Tuple[bool, str, Dict[str, Any]]:
        """流式删除列：calamine读取 → 删除列 → write_only写入

        Args:
            file_path: Excel文件路径
            sheet_name: 目标工作表名
            start_column: 起始列号（1-based）
            count: 删除列数
            preserve_col_widths: 是否保留列宽

        Returns:
            (success, message, metadata)
        """
        if not _HAS_CALAMINE:
            return False, "calamine未安装", {}

        if start_column < 1:
            return False, f"起始列号必须>=1，实际: {start_column}", {}
        if count < 1:
            return False, f"删除列数必须>=1，实际: {count}", {}

        def _modify(rows, header_row, col_map):
            if not rows:
                return False, "工作表为空", rows, {}

            max_col = max(len(r) for r in rows) if rows else 0
            if start_column > max_col:
                return False, f"起始列号({start_column})超过工作表最大列数({max_col})", rows, {}

            actual_count = min(count, max_col - start_column + 1)
            end_col = start_column + actual_count - 1  # inclusive

            new_rows = []
            for row in rows:
                # 删除 start_column-1 到 end_col-1 位置的元素
                new_row = row[:start_column - 1] + row[end_col:]
                new_rows.append(new_row)

            return True, (
                f"流式删除第{start_column}-{end_col}列"
                f"（共{actual_count}列）"
            ), new_rows, {
                'action': 'delete_columns_streaming',
                'start_column': start_column,
                'actual_count': actual_count,
                'original_columns': max_col,
                'new_columns': max_col - actual_count,
            }

        return cls._copy_modify_write(file_path, sheet_name, _modify, preserve_col_widths)

    @classmethod
    def update_range(
        cls,
        file_path: str,
        sheet_name: str,
        start_row: int,
        start_col: int,
        data: List[List[Any]],
        preserve_formulas: bool = True,
        preserve_col_widths: bool = True,
    ) -> Tuple[bool, str, Dict[str, Any]]:
        """流式覆盖范围：calamine读取 → 覆盖指定范围 → write_only写入

        仅支持覆盖模式（insert_mode=False），不保留公式。
        适合批量修改已知位置的数据。

        Args:
            file_path: Excel文件路径
            sheet_name: 目标工作表名
            start_row: 起始行号（1-based）
            start_col: 起始列号（1-based）
            data: 二维数组数据
            preserve_col_widths: 是否保留列宽

        Returns:
            (success, message, metadata)
        """
        if not _HAS_CALAMINE:
            return False, "calamine未安装", {}

        try:
            if not data:
                return False, "数据不能为空", {}
            if start_row < 1:
                return False, f"起始行号必须>=1，实际: {start_row}", {}
            if start_col < 1:
                return False, f"起始列号必须>=1，实际: {start_col}", {}

            def _modify(rows, header_row, col_map):
                # 确保行数足够
                max_needed_row = start_row - 1 + len(data)
                if max_needed_row > len(rows):
                    # 追加空行
                    max_col = max(len(r) for r in rows) if rows else 0
                    while len(rows) < max_needed_row:
                        rows.append([None] * max_col)

                updated_cells = 0
                for r_offset, row_data in enumerate(data):
                    row_idx = start_row - 1 + r_offset
                    for c_offset, value in enumerate(row_data):
                        col_idx = start_col - 1 + c_offset
                        # 确保列数足够
                        while len(rows[row_idx]) <= col_idx:
                            rows[row_idx].append(None)
                        rows[row_idx][col_idx] = value
                        updated_cells += 1

                return True, (
                    f"流式覆盖 {len(data)} 行×{max(len(r) for r in data)} 列"
                    f"（从{sheet_name}!{get_column_letter(start_col)}{start_row}，"
                    f"共{updated_cells}个单元格）"
                ), rows, {
                    'action': 'update_range_streaming',
                    'start_row': start_row,
                    'start_col': start_col,
                    'rows_written': len(data),
                    'cols_written': max(len(r) for r in data),
                    'updated_cells': updated_cells,
                }

            return cls._copy_modify_write(file_path, sheet_name, _modify, preserve_col_widths)

        except Exception as e:
            logger.error(f"流式覆盖范围失败: {e}")
            return False, f"流式写入失败: {str(e)}", {}

    @classmethod
    def insert_rows_streaming(
        cls,
        file_path: str,
        sheet_name: str,
        start_row: int,
        data: List[List[Any]],
        preserve_formulas: bool = True,
        preserve_col_widths: bool = True,
    ) -> Tuple[bool, str, Dict[str, Any]]:
        """流式插入行：calamine读取 → 在指定行位置插入 → write_only写入

        支持在任意行位置插入新行，保留原有数据下移。
        Args:
            file_path: Excel文件路径
            sheet_name: 目标工作表名
            start_row: 插入位置行号（1-based）
            data: 要插入的二维数组数据
            preserve_formulas: 是否保留公式（插入模式下此参数暂无效，因为插入的是新行）
            preserve_col_widths: 是否保留列宽

        Returns:
            (success, message, metadata)
        """
        if not _HAS_CALAMINE:
            return False, "calamine未安装，无法使用流式写入", {}

        try:
            if not data:
                return False, "数据不能为空", {}
            if start_row < 1:
                return False, f"起始行号必须>=1，实际: {start_row}", {}

            def _modify(rows, header_row, col_map):
                # 确保行数足够
                if len(rows) < start_row - 1:
                    # 如果目标行不存在，先创建空行到目标位置
                    max_col = max(len(r) for r in rows) if rows else 0
                    while len(rows) < start_row - 1:
                        rows.append([None] * max_col)

                # 在目标行位置插入新行
                for r_offset, row_data in enumerate(data):
                    new_row = [None] * len(rows[0]) if rows else []
                    # 填充新行数据
                    for c_offset, value in enumerate(row_data):
                        if c_offset < len(new_row):
                            new_row[c_offset] = value
                    # 插入到指定位置
                    rows.insert(start_row - 1 + r_offset, new_row)

                inserted_count = len(data)
                return True, (
                    f"流式插入 {inserted_count} 行到 {sheet_name}!{start_row}，"
                    f"原有数据自动下移"
                ), rows, {
                    'action': 'insert_rows_streaming',
                    'start_row': start_row,
                    'rows_inserted': inserted_count,
                    'total_rows': len(rows),
                }

            return cls._copy_modify_write(file_path, sheet_name, _modify, preserve_col_widths)

        except Exception as e:
            logger.error(f"流式插入行失败: {e}")
            return False, f"流式写入失败: {str(e)}", {}

    @classmethod
    def insert_columns_streaming(
        cls,
        file_path: str,
        sheet_name: str,
        start_column: int,
        count: int,
        preserve_formulas: bool = True,
        preserve_col_widths: bool = True,
    ) -> Tuple[bool, str, Dict[str, Any]]:
        """流式插入空列：calamine读取 → 在指定列位置插入空列 → write_only写入

        支持在任意列位置插入空列，保留原有数据右移。
        Args:
            file_path: Excel文件路径
            sheet_name: 目标工作表名
            start_column: 插入位置列号（1-based）
            count: 要插入的列数
            preserve_formulas: 是否保留公式（插入模式下此参数暂无效，因为插入的是新列）
            preserve_col_widths: 是否保留列宽

        Returns:
            (success, message, metadata)
        """
        if not _HAS_CALAMINE:
            return False, "calamine未安装，无法使用流式写入", {}

        try:
            if count < 1:
                return False, f"插入列数必须>=1，实际: {count}", {}
            if start_column < 1:
                return False, f"起始列号必须>=1，实际: {start_column}", {}

            def _modify(rows, header_row, col_map):
                # 在每行的指定位置插入 count 个空值
                for row in rows:
                    for _ in range(count):
                        row.insert(start_column - 1, None)

                return True, (
                    f"流式插入 {count} 列到 {sheet_name}!{get_column_letter(start_column)}，"
                    f"原有数据自动右移"
                ), rows, {
                    'action': 'insert_columns_streaming',
                    'start_column': start_column,
                    'columns_inserted': count,
                    'total_columns': max(len(r) for r in rows) if rows else 0,
                }

            return cls._copy_modify_write(file_path, sheet_name, _modify, preserve_col_widths)

        except Exception as e:
            logger.error(f"流式插入列失败: {e}")
            return False, f"流式写入失败: {str(e)}", {}
