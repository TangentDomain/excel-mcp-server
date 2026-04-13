"""
Excel MCP Server - Excel搜索模块

提供Excel文件搜索功能
使用python-calamine（Rust引擎）加速纯值搜索，openpyxl作为公式搜索的后备方案
"""

import logging
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..models.types import MatchType, OperationResult, RangeType, SearchMatch
from ..utils.config import MAX_SEARCH_FILES
from ..utils.parsers import RangeParser
from ..utils.validators import ExcelValidator

logger = logging.getLogger(__name__)

# 尝试导入calamine
try:
    from python_calamine import CalamineWorkbook

    _HAS_CALAMINE = True
except ImportError:
    _HAS_CALAMINE = False


class ExcelSearcher:
    """Excel文件搜索器"""

    def __init__(self, file_path: str):
        """
        初始化Excel搜索器

        Args:
            file_path: Excel文件路径
        """
        self.file_path = ExcelValidator.validate_file_path(file_path)

    def regex_search(
        self,
        pattern: str,
        flags: str = "",
        search_values: bool = True,
        search_formulas: bool = False,
        sheet_name: str | None = None,
        range_expression: str | None = None,
    ) -> OperationResult:
        """
        在Excel文件中使用正则表达式搜索单元格内容

        Args:
            pattern: 正则表达式模式
            flags: 正则表达式标志 (i=忽略大小写, m=多行, s=点匹配换行)
            search_values: 是否搜索单元格的显示值
            search_formulas: 是否搜索单元格的公式
            sheet_name: 工作表名称 (可选，不指定时搜索所有工作表)
            range_expression: 搜索范围表达式 (如"A1:C10"或"Sheet1!A1:C10")

        Returns:
            OperationResult: 包含搜索结果的结果对象
        """
        try:
            # 构建正则表达式标志
            regex_flags = self._build_regex_flags(flags)

            # 编译正则表达式
            try:
                regex = re.compile(pattern, regex_flags)
            except re.error as e:
                raise ValueError(f"无效的正则表达式: {e}")

            # calamine快速路径（仅限纯值搜索，不搜公式）
            if not search_formulas and _HAS_CALAMINE:
                try:
                    matches = self._search_calamine(regex, sheet_name, range_expression)
                    return OperationResult(
                        success=True,
                        data=matches,
                        metadata={
                            "file_path": self.file_path,
                            "pattern": pattern,
                            "total_matches": len(matches),
                            "search_values": search_values,
                            "search_formulas": search_formulas,
                            "range_expression": range_expression,
                        },
                    )
                except Exception as e:
                    logger.debug(f"calamine搜索失败，回退openpyxl: {e}")

            # openpyxl路径（公式搜索或calamine不可用）
            try:
                workbook = load_workbook(
                    self.file_path,
                    data_only=not search_formulas,
                    keep_vba=False,  # 禁用VBA以避免兼容性问题
                    read_only=True,  # 搜索操作使用只读模式
                )
            except Exception as e:
                # 如果加载失败，尝试更保守的加载方式
                logger.warning(f"使用标准方式加载失败，尝试兼容模式: {e}")
                try:
                    workbook = load_workbook(
                        self.file_path,
                        data_only=True,  # 强制只读取数据
                        keep_vba=False,
                        read_only=True,  # 只读模式
                    )
                except Exception as e2:
                    raise ValueError(f"无法加载Excel文件: {e2}")

            # 执行搜索
            matches = self._search_workbook(
                workbook,
                regex,
                search_values,
                search_formulas,
                sheet_name,
                range_expression,
            )

            return OperationResult(
                success=True,
                data=matches,
                metadata={
                    "file_path": self.file_path,
                    "pattern": pattern,
                    "total_matches": len(matches),
                    "search_values": search_values,
                    "search_formulas": search_formulas,
                    "range_expression": range_expression,
                },
            )

        except Exception as e:
            logger.error(f"正则搜索失败: {e}")
            return OperationResult(success=False, error=str(e))

    def _build_regex_flags(self, flags: str) -> int:
        """构建正则表达式标志

        Args:
            flags: 正则表达式标志字符串 (i=忽略大小写, m=多行, s=点匹配换行)

        Returns:
            int: 组合后的正则表达式标志位
        """
        regex_flags = 0
        if "i" in flags.lower():
            regex_flags |= re.IGNORECASE
        if "m" in flags.lower():
            regex_flags |= re.MULTILINE
        if "s" in flags.lower():
            regex_flags |= re.DOTALL
        return regex_flags

    def _search_calamine(
        self,
        regex: re.Pattern,
        sheet_name: str | None = None,
        range_expression: str | None = None,
    ) -> list[SearchMatch]:
        """使用calamine快速搜索纯值（不搜公式）

        Args:
            regex: 编译后的正则表达式对象
            sheet_name: 工作表名称（可选）
            range_expression: 搜索范围表达式（可选）

        Returns:
            List[SearchMatch]: 匹配结果列表
        """
        wb = CalamineWorkbook.from_path(self.file_path)
        matches = []

        # 解析范围表达式
        range_info = None
        target_sheet_name = sheet_name
        if range_expression:
            range_info = RangeParser.parse_range_expression(range_expression)
            if range_info.sheet_name:
                target_sheet_name = range_info.sheet_name

        # 确定要搜索的工作表
        if target_sheet_name:
            if target_sheet_name not in wb.sheet_names:
                raise ValueError(f"工作表 '{target_sheet_name}' 不存在")
            sheets_to_search = [target_sheet_name]
        else:
            sheets_to_search = wb.sheet_names

        for s_name in sheets_to_search:
            ws = wb.get_sheet_by_name(s_name)
            # 跳过空工作表（calamine iter_rows在空表上会panic）
            if ws.height == 0:
                continue
            all_rows = list(ws.iter_rows())

            for r_idx, row in enumerate(all_rows):
                for c_idx, val in enumerate(row):
                    if val is None:
                        continue

                    # 范围过滤
                    if range_info:
                        r1, c1 = r_idx + 1, c_idx + 1  # 转为1-based
                        if not self._in_range(range_info, r1, c1):
                            continue

                    cell_str = str(val)
                    for match in regex.finditer(cell_str):
                        coord = f"{get_column_letter(c_idx + 1)}{r_idx + 1}"
                        matches.append(
                            SearchMatch(
                                sheet=s_name,
                                cell=coord,
                                value=cell_str,
                                match=match.group(),
                                match_start=match.start(),
                                match_end=match.end(),
                                match_type=MatchType.VALUE,
                            )
                        )

        return matches

    def _in_range(self, range_info, row: int, col: int) -> bool:
        """检查单元格(row,col)是否在范围内（1-based）

        Args:
            range_info: 范围信息对象
            row: 行号（1-based）
            col: 列号（1-based）

        Returns:
            bool: 单元格是否在指定范围内
        """
        rt = range_info.range_type
        if rt in [RangeType.ROW_RANGE, RangeType.SINGLE_ROW]:
            parts = range_info.cell_range.split(":")
            min_r = int(parts[0])
            max_r = int(parts[1]) if len(parts) > 1 else min_r
            return min_r <= row <= max_r
        elif rt in [RangeType.COLUMN_RANGE, RangeType.SINGLE_COLUMN]:
            from openpyxl.utils import column_index_from_string

            parts = range_info.cell_range.split(":")
            min_c = column_index_from_string(parts[0])
            max_c = column_index_from_string(parts[1]) if len(parts) > 1 else min_c
            return min_c <= col <= max_c
        else:
            from openpyxl.utils import range_boundaries

            min_c, min_r, max_c, max_r = range_boundaries(range_info.cell_range)
            return min_r <= row <= max_r and min_c <= col <= max_c

    def _search_workbook(
        self,
        workbook,
        regex: re.Pattern,
        search_values: bool,
        search_formulas: bool,
        sheet_name: str | None = None,
        range_expression: str | None = None,
    ) -> list[SearchMatch]:
        """在工作簿中搜索

        Args:
            workbook: openpyxl工作簿对象
            regex: 编译后的正则表达式对象
            search_values: 是否搜索单元格值
            search_formulas: 是否搜索单元格公式
            sheet_name: 工作表名称（可选）
            range_expression: 搜索范围表达式（可选）

        Returns:
            List[SearchMatch]: 匹配结果列表
        """
        matches = []

        # 解析范围表达式（如果提供）
        range_info = None
        target_sheet_name = sheet_name

        if range_expression:
            range_info = RangeParser.parse_range_expression(range_expression)
            # 如果范围表达式包含工作表名，使用它
            if range_info.sheet_name:
                target_sheet_name = range_info.sheet_name

        # 确定要搜索的工作表
        if target_sheet_name:
            # 搜索指定工作表
            if target_sheet_name not in workbook.sheetnames:
                raise ValueError(f"工作表 '{target_sheet_name}' 不存在")
            sheet_names = [target_sheet_name]
        else:
            # 搜索所有工作表（但如果有范围表达式，这种情况不应该发生）
            sheet_names = workbook.sheetnames

        # 遍历指定的工作表
        for current_sheet_name in sheet_names:
            sheet = workbook[current_sheet_name]

            # 如果指定了范围，则使用范围内的单元格；否则使用所有单元格
            if range_info:
                matches.extend(
                    self._search_in_range(
                        sheet,
                        current_sheet_name,
                        regex,
                        search_values,
                        search_formulas,
                        range_info,
                    )
                )
            else:
                matches.extend(self._search_entire_sheet(sheet, current_sheet_name, regex, search_values, search_formulas))

        return matches

    def _search_entire_sheet(
        self,
        sheet,
        sheet_name: str,
        regex: re.Pattern,
        search_values: bool,
        search_formulas: bool,
    ) -> list[SearchMatch]:
        """在整个工作表中搜索

        Args:
            sheet: openpyxl工作表对象
            sheet_name: 工作表名称
            regex: 编译后的正则表达式对象
            search_values: 是否搜索单元格值
            search_formulas: 是否搜索单元格公式

        Returns:
            List[SearchMatch]: 匹配结果列表
        """
        matches = []

        # 遍历所有单元格
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                # 搜索单元格值
                if search_values:
                    cell_value = str(cell.value)
                    for match in regex.finditer(cell_value):
                        matches.append(
                            SearchMatch(
                                sheet=sheet_name,
                                cell=cell.coordinate,
                                value=cell_value,
                                match=match.group(),
                                match_start=match.start(),
                                match_end=match.end(),
                                match_type=MatchType.VALUE,
                            )
                        )

                # 搜索单元格公式
                if search_formulas and hasattr(cell, "formula") and cell.formula:
                    formula = str(cell.formula)
                    for match in regex.finditer(formula):
                        matches.append(
                            SearchMatch(
                                sheet=sheet_name,
                                cell=cell.coordinate,
                                formula=formula,
                                match=match.group(),
                                match_start=match.start(),
                                match_end=match.end(),
                                match_type=MatchType.FORMULA,
                            )
                        )

        return matches

    def _search_in_range(
        self,
        sheet,
        sheet_name: str,
        regex: re.Pattern,
        search_values: bool,
        search_formulas: bool,
        range_info,
    ) -> list[SearchMatch]:
        """在指定范围内搜索

        Args:
            sheet: openpyxl工作表对象
            sheet_name: 工作表名称
            regex: 编译后的正则表达式对象
            search_values: 是否搜索单元格值
            search_formulas: 是否搜索单元格公式
            range_info: 范围信息对象

        Returns:
            List[SearchMatch]: 匹配结果列表
        """
        from openpyxl.utils import column_index_from_string, range_boundaries

        matches = []

        # 根据范围类型确定搜索边界
        if range_info.range_type.value in ["row_range", "single_row"]:
            # 行范围搜索 (如 "3:5" 或 "3")
            min_col, max_col = 1, sheet.max_column
            if range_info.range_type.value == "single_row":
                # 单行 (如 "3")
                row_num = int(range_info.cell_range.split(":")[0])
                min_row = max_row = row_num
            else:
                # 行范围 (如 "3:5")
                start_row, end_row = map(int, range_info.cell_range.split(":"))
                min_row, max_row = start_row, end_row

        elif range_info.range_type.value in ["column_range", "single_column"]:
            # 列范围搜索 (如 "B:D" 或 "B")
            min_row, max_row = 1, sheet.max_row
            if range_info.range_type.value == "single_column":
                # 单列 (如 "B")
                col_letter = range_info.cell_range.split(":")[0]
                min_col = max_col = column_index_from_string(col_letter)
            else:
                # 列范围 (如 "B:D")
                start_col, end_col = range_info.cell_range.split(":")
                min_col = column_index_from_string(start_col)
                max_col = column_index_from_string(end_col)

        else:
            # 单元格范围搜索 (如 "A1:C10")
            min_col, min_row, max_col, max_row = range_boundaries(range_info.cell_range)

        # 遍历范围内的单元格
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value is None:
                    continue

                # 搜索单元格值
                if search_values:
                    cell_value = str(cell.value)
                    for match in regex.finditer(cell_value):
                        matches.append(
                            SearchMatch(
                                sheet=sheet_name,
                                cell=cell.coordinate,
                                value=cell_value,
                                match=match.group(),
                                match_start=match.start(),
                                match_end=match.end(),
                                match_type=MatchType.VALUE,
                            )
                        )

                # 搜索单元格公式
                if search_formulas and hasattr(cell, "formula") and cell.formula:
                    formula = str(cell.formula)
                    for match in regex.finditer(formula):
                        matches.append(
                            SearchMatch(
                                sheet=sheet_name,
                                cell=cell.coordinate,
                                formula=formula,
                                match=match.group(),
                                match_start=match.start(),
                                match_end=match.end(),
                                match_type=MatchType.FORMULA,
                            )
                        )

        return matches

    def regex_search_directory(
        self,
        directory_path: str,
        pattern: str,
        flags: str = "",
        search_values: bool = True,
        search_formulas: bool = False,
        recursive: bool = True,
        file_extensions: list[str] | None = None,
        file_pattern: str | None = None,
        max_files: int = MAX_SEARCH_FILES,
    ) -> OperationResult:
        """
        在目录下的所有Excel文件中使用正则表达式搜索单元格内容

        Args:
            directory_path: 目录路径
            pattern: 正则表达式模式
            flags: 正则表达式标志 (i=忽略大小写, m=多行, s=点匹配换行)
            search_values: 是否搜索单元格的显示值
            search_formulas: 是否搜索单元格的公式
            recursive: 是否递归搜索子目录
            file_extensions: 文件扩展名过滤，如['.xlsx', '.xlsm']
            file_pattern: 文件名正则模式过滤
            max_files: 最大搜索文件数限制

        Returns:
            OperationResult: 包含聚合搜索结果的结果对象
        """
        try:
            # 验证目录路径
            directory_path = Path(directory_path)
            if not directory_path.exists():
                raise ValueError(f"目录不存在: {directory_path}")
            if not directory_path.is_dir():
                raise ValueError(f"路径不是目录: {directory_path}")

            # 设置默认文件扩展名
            if file_extensions is None:
                file_extensions = [".xlsx", ".xlsm"]

            # 构建正则表达式标志
            regex_flags = self._build_regex_flags(flags)

            # 编译正则表达式
            try:
                re.compile(pattern, regex_flags)
            except re.error as e:
                raise ValueError(f"无效的正则表达式: {e}")

            # 编译文件名过滤正则（如果提供）
            file_regex = None
            if file_pattern:
                try:
                    file_regex = re.compile(file_pattern)
                except re.error as e:
                    raise ValueError(f"无效的文件名正则表达式: {e}")

            # 查找Excel文件
            excel_files = self._find_excel_files(directory_path, file_extensions, file_regex, recursive, max_files)

            # 执行搜索
            all_matches = []
            searched_files = []
            skipped_files = []
            file_errors = []

            for file_path in excel_files:
                try:
                    # 临时创建搜索器实例（使用当前文件路径）
                    temp_searcher = ExcelSearcher(str(file_path))
                    result = temp_searcher.regex_search(pattern, flags, search_values, search_formulas)

                    if result.success and result.data:
                        # 为每个匹配添加文件路径信息
                        for match in result.data:
                            match_dict = match.__dict__ if hasattr(match, "__dict__") else match
                            if isinstance(match_dict, dict):
                                match_dict["file_path"] = str(file_path)
                            all_matches.append(match_dict)
                        searched_files.append(str(file_path))
                    elif result.success:
                        # 没有匹配但搜索成功
                        searched_files.append(str(file_path))
                    else:
                        # 搜索失败
                        file_errors.append({"file_path": str(file_path), "error": result.error})
                        skipped_files.append(str(file_path))

                except Exception as e:
                    logger.warning(f"搜索文件 {file_path} 时发生错误: {e}")
                    file_errors.append({"file_path": str(file_path), "error": str(e)})
                    skipped_files.append(str(file_path))

            return OperationResult(
                success=True,
                data=all_matches,
                metadata={
                    "directory_path": str(directory_path),
                    "pattern": pattern,
                    "total_matches": len(all_matches),
                    "total_files_found": len(excel_files),
                    "searched_files": searched_files,
                    "skipped_files": skipped_files,
                    "file_errors": file_errors,
                    "search_values": search_values,
                    "search_formulas": search_formulas,
                    "recursive": recursive,
                    "file_extensions": file_extensions,
                },
            )

        except Exception as e:
            logger.error(f"目录正则搜索失败: {e}")
            return OperationResult(success=False, error=str(e))

    def _find_excel_files(
        self,
        directory: Path,
        extensions: list[str],
        file_regex: re.Pattern | None,
        recursive: bool,
        max_files: int,
    ) -> list[Path]:
        """查找目录中的Excel文件

        Args:
            directory: 目录路径对象
            extensions: 文件扩展名列表
            file_regex: 文件名正则表达式（可选）
            recursive: 是否递归搜索子目录
            max_files: 最大文件数量限制

        Returns:
            List[Path]: 找到的Excel文件路径列表
        """
        excel_files = []

        # 构建搜索模式
        search_patterns = []
        for ext in extensions:
            if not ext.startswith("."):
                ext = f".{ext}"
            search_patterns.append(f"*{ext}")

        try:
            if recursive:
                # 递归搜索
                for pattern in search_patterns:
                    for file_path in directory.rglob(pattern):
                        if len(excel_files) >= max_files:
                            break
                        if self._should_include_file(file_path, file_regex):
                            excel_files.append(file_path)
            else:
                # 仅搜索当前目录
                for pattern in search_patterns:
                    for file_path in directory.glob(pattern):
                        if len(excel_files) >= max_files:
                            break
                        if self._should_include_file(file_path, file_regex):
                            excel_files.append(file_path)

        except Exception as e:
            logger.error(f"查找Excel文件时发生错误: {e}")

        return excel_files[:max_files]  # 确保不超过最大限制

    def _should_include_file(self, file_path: Path, file_regex: re.Pattern | None) -> bool:
        """判断是否应该包含该文件

        Args:
            file_path: 文件路径对象
            file_regex: 文件名正则表达式（可选）

        Returns:
            bool: 是否应该包含该文件
        """
        # 检查文件是否存在且是文件
        if not file_path.is_file():
            return False

        # 如果提供了文件名正则模式，进行匹配
        if file_regex and not file_regex.search(file_path.name):
            return False

        # 排除临时文件
        if file_path.name.startswith("~") or file_path.name.startswith("."):
            return False

        return True

    @staticmethod
    def search_directory_static(
        directory_path: str,
        pattern: str,
        flags: str = "",
        search_values: bool = True,
        search_formulas: bool = False,
        recursive: bool = True,
        file_extensions: list[str] | None = None,
        file_pattern: str | None = None,
        max_files: int = MAX_SEARCH_FILES,
    ) -> OperationResult:
        """静态方法：在目录下的所有Excel文件中使用正则表达式搜索单元格内容

        这是一个静态方法，不需要创建ExcelSearcher实例

        Args:
            directory_path: 目录路径
            pattern: 正则表达式模式
            flags: 正则表达式标志 (i=忽略大小写, m=多行, s=点匹配换行)
            search_values: 是否搜索单元格的显示值
            search_formulas: 是否搜索单元格的公式
            recursive: 是否递归搜索子目录
            file_extensions: 文件扩展名过滤，如['.xlsx', '.xlsm']
            file_pattern: 文件名正则模式过滤
            max_files: 最大搜索文件数限制

        Returns:
            OperationResult: 包含聚合搜索结果的结果对象
        """
        try:
            # 验证目录路径
            directory_path = Path(directory_path)
            if not directory_path.exists():
                raise ValueError(f"目录不存在: {directory_path}")
            if not directory_path.is_dir():
                raise ValueError(f"路径不是目录: {directory_path}")

            # 设置默认文件扩展名
            if file_extensions is None:
                file_extensions = [".xlsx", ".xlsm"]

            # 构建正则表达式标志
            regex_flags = ExcelSearcher._build_regex_flags_static(flags)

            # 编译正则表达式
            try:
                re.compile(pattern, regex_flags)
            except re.error as e:
                raise ValueError(f"无效的正则表达式: {e}")

            # 编译文件名过滤正则（如果提供）
            file_regex = None
            if file_pattern:
                try:
                    file_regex = re.compile(file_pattern)
                except re.error as e:
                    raise ValueError(f"无效的文件名正则表达式: {e}")

            # 查找Excel文件
            excel_files = ExcelSearcher._find_excel_files_static(directory_path, file_extensions, file_regex, recursive, max_files)

            # 执行搜索
            all_matches = []
            searched_files = []
            skipped_files = []
            file_errors = []

            for file_path in excel_files:
                try:
                    # 临时创建搜索器实例（使用当前文件路径）
                    temp_searcher = ExcelSearcher(str(file_path))
                    result = temp_searcher.regex_search(pattern, flags, search_values, search_formulas)

                    if result.success and result.data:
                        # 为每个匹配添加文件路径信息
                        for match in result.data:
                            match_dict = match.__dict__ if hasattr(match, "__dict__") else match
                            if isinstance(match_dict, dict):
                                match_dict["file_path"] = str(file_path)
                            all_matches.append(match_dict)
                        searched_files.append(str(file_path))
                    elif result.success:
                        # 没有匹配但搜索成功
                        searched_files.append(str(file_path))
                    else:
                        # 搜索失败
                        file_errors.append({"file_path": str(file_path), "error": result.error})
                        skipped_files.append(str(file_path))

                except Exception as e:
                    logger.warning(f"搜索文件 {file_path} 时发生错误: {e}")
                    file_errors.append({"file_path": str(file_path), "error": str(e)})
                    skipped_files.append(str(file_path))

            return OperationResult(
                success=True,
                data=all_matches,
                metadata={
                    "directory_path": str(directory_path),
                    "pattern": pattern,
                    "total_matches": len(all_matches),
                    "total_files_found": len(excel_files),
                    "searched_files": searched_files,
                    "skipped_files": skipped_files,
                    "file_errors": file_errors,
                    "search_values": search_values,
                    "search_formulas": search_formulas,
                    "recursive": recursive,
                    "file_extensions": file_extensions,
                },
            )

        except Exception as e:
            logger.error(f"目录正则搜索失败: {e}")
            return OperationResult(success=False, error=str(e))

    @staticmethod
    def _build_regex_flags_static(flags: str) -> int:
        """静态方法：构建正则表达式标志

        Args:
            flags: 正则表达式标志字符串 (i=忽略大小写, m=多行, s=点匹配换行)

        Returns:
            int: 组合后的正则表达式标志位
        """
        regex_flags = 0
        if "i" in flags.lower():
            regex_flags |= re.IGNORECASE
        if "m" in flags.lower():
            regex_flags |= re.MULTILINE
        if "s" in flags.lower():
            regex_flags |= re.DOTALL
        return regex_flags

    @staticmethod
    def _find_excel_files_static(
        directory: Path,
        extensions: list[str],
        file_regex: re.Pattern | None,
        recursive: bool,
        max_files: int,
    ) -> list[Path]:
        """静态方法：查找目录中的Excel文件

        Args:
            directory: 目录路径对象
            extensions: 文件扩展名列表
            file_regex: 文件名正则表达式（可选）
            recursive: 是否递归搜索子目录
            max_files: 最大文件数量限制

        Returns:
            List[Path]: 找到的Excel文件路径列表
        """
        excel_files = []

        # 构建搜索模式
        search_patterns = []
        for ext in extensions:
            if not ext.startswith("."):
                ext = f".{ext}"
            search_patterns.append(f"*{ext}")

        try:
            if recursive:
                # 递归搜索
                for pattern in search_patterns:
                    for file_path in directory.rglob(pattern):
                        if len(excel_files) >= max_files:
                            break
                        if ExcelSearcher._should_include_file_static(file_path, file_regex):
                            excel_files.append(file_path)
            else:
                # 仅搜索当前目录
                for pattern in search_patterns:
                    for file_path in directory.glob(pattern):
                        if len(excel_files) >= max_files:
                            break
                        if ExcelSearcher._should_include_file_static(file_path, file_regex):
                            excel_files.append(file_path)

        except Exception as e:
            logger.error(f"查找Excel文件时发生错误: {e}")

        return excel_files[:max_files]  # 确保不超过最大限制

    @staticmethod
    def _should_include_file_static(file_path: Path, file_regex: re.Pattern | None) -> bool:
        """静态方法：判断是否应该包含该文件

        Args:
            file_path: 文件路径对象
            file_regex: 文件名正则表达式（可选）

        Returns:
            bool: 是否应该包含该文件
        """
        # 检查文件是否存在且是文件
        if not file_path.is_file():
            return False

        # 如果提供了文件名正则模式，进行匹配
        if file_regex and not file_regex.search(file_path.name):
            return False

        # 排除临时文件
        if file_path.name.startswith("~") or file_path.name.startswith("."):
            return False

        return True
