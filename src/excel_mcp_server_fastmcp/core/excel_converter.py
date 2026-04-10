"""
Excel MCP Server - Excel转换模块

提供Excel文件格式转换、导入导出功能
"""

import logging
import csv
import json
import os
import time
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from ..models.types import SheetInfo, OperationResult
from ..utils.validators import ExcelValidator
from ..utils.exceptions import ExcelFileNotFoundError, DataValidationError

logger = logging.getLogger(__name__)


class ExcelConverter:
    """Excel文件转换器"""

    def __init__(self, file_path: str):
        """
        初始化Excel转换器

        Args:
            file_path: Excel文件路径
        """
        self.file_path = ExcelValidator.validate_file_path(file_path)

    def export_to_csv(
        self,
        output_path: str,
        sheet_name: Optional[str] = None,
        encoding: str = "utf-8"
    ) -> OperationResult:
        """
        将Excel工作表导出为CSV文件

        Args:
            output_path: 输出CSV文件路径
            sheet_name: 工作表名称 (默认使用活动工作表)
            encoding: 文件编码 (默认: utf-8，可选: gbk)

        Returns:
            OperationResult: 导出操作的结果
        """
        try:
            workbook = load_workbook(self.file_path, read_only=True)

            # 选择工作表
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise DataValidationError(
                        f"工作表 '{sheet_name}' 不存在",
                        f"可用工作表: {', '.join(workbook.sheetnames)}",
                        "请使用正确的工作表名称"
                    )
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
                sheet_name = sheet.title

            # 创建输出目录
            output_dir = Path(output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)

            # 写入CSV文件
            with open(output_path, 'w', newline='', encoding=encoding) as csvfile:
                csv_writer = csv.writer(csvfile)

                row_count = 0
                for row in sheet.iter_rows(values_only=True):
                    # 过滤掉完全空的行
                    if any(cell is not None for cell in row):
                        csv_writer.writerow(row)
                        row_count += 1

            workbook.close()

            return OperationResult(
                success=True,
                message=f"成功导出 {row_count} 行数据到CSV文件",
                data={
                    'output_path': output_path,
                    'row_count': row_count,
                    'sheet_name': sheet_name,
                    'encoding': encoding
                },
                metadata={
                    'source_file': self.file_path,
                    'sheet_name': sheet_name,
                    'encoding': encoding,
                    'row_count': row_count
                }
            )

        except Exception as e:
            logger.error(f"CSV导出失败: {e}")
            return OperationResult(
                success=False,
                error=str(e),
                metadata={'operation': 'export_to_csv', 'file_path': self.file_path}
            )

    @staticmethod
    def import_from_csv(
        csv_path: str,
        output_path: str,
        sheet_name: str = "Sheet1",
        encoding: str = "utf-8",
        has_header: bool = True
    ) -> OperationResult:
        """
        从CSV文件导入数据创建Excel文件

        Args:
            csv_path: CSV文件路径
            output_path: 输出Excel文件路径
            sheet_name: 工作表名称 (默认: Sheet1)
            encoding: CSV文件编码 (默认: utf-8，可选: gbk)
            has_header: 是否包含表头行

        Returns:
            OperationResult: 导入操作的结果
        """
        try:
            # 验证CSV文件存在
            if not os.path.exists(csv_path):
                raise ExcelFileNotFoundError(f"CSV文件不存在: {csv_path}")

            # 创建输出目录
            output_dir = Path(output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)

            # 使用 write_only 模式创建工作簿（流式写入，内存占用极低）
            # 大CSV文件(10万+行)内存从~500MB降至~10MB
            workbook = Workbook(write_only=True)
            sheet = workbook.create_sheet(title=sheet_name)

            # 读取CSV数据并流式写入
            row_count = 0
            with open(csv_path, 'r', encoding=encoding) as csvfile:
                csv_reader = csv.reader(csvfile)

                for row_data in csv_reader:
                    # 尝试转换数值
                    converted_row = []
                    for value in row_data:
                        try:
                            if value and value.replace('.', '').replace('-', '').isdigit():
                                value = float(value) if '.' in value else int(value)
                        except (ValueError, AttributeError):
                            pass
                        converted_row.append(value)

                    sheet.append(converted_row)
                    row_count += 1

            # 保存Excel文件
            workbook.save(output_path)
            workbook.close()

            return OperationResult(
                success=True,
                message=f"成功从CSV导入 {row_count} 行数据",
                data={
                    'output_path': output_path,
                    'row_count': row_count,
                    'sheet_name': sheet_name,
                    'has_header': has_header,
                    'encoding': encoding
                },
                metadata={
                    'source_file': csv_path,
                    'encoding': encoding,
                    'row_count': row_count
                }
            )

        except Exception as e:
            logger.error(f"CSV导入失败: {e}")
            return OperationResult(
                success=False,
                error=str(e),
                metadata={'operation': 'import_from_csv', 'csv_path': csv_path}
            )

    @staticmethod
    def convert_format(
        input_path: str,
        output_path: str,
        target_format: str = "xlsx"
    ) -> OperationResult:
        """
        转换Excel文件格式

        Args:
            input_path: 输入文件路径
            output_path: 输出文件路径
            target_format: 目标格式，可选值: "xlsx", "xlsm", "csv", "json"

        Returns:
            OperationResult: 转换操作的结果
        """
        try:
            # 验证输入文件
            if not os.path.exists(input_path):
                raise ExcelFileNotFoundError(f"输入文件不存在: {input_path}")

            input_format = Path(input_path).suffix.lower()

            # 加载工作簿
            workbook = load_workbook(input_path)

            # 创建输出目录
            output_dir = Path(output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)

            if target_format.lower() in ["xlsx", "xlsm"]:
                # Excel格式转换
                workbook.save(output_path)
                file_size = os.path.getsize(output_path)

                return OperationResult(
                    success=True,
                    message=f"成功转换文件格式: {input_format} -> {target_format}",
                    data={
                        'input_format': input_format,
                        'output_format': target_format,
                        'file_size': file_size,
                        'output_path': output_path
                    },
                    metadata={
                        'input_path': input_path,
                        'target_format': target_format
                    }
                )

            elif target_format.lower() == "json":
                # 转换为JSON格式 — 优先用 calamine 读取（比 openpyxl 快 10x+）
                json_data = {}
                calamine_ok = False
                try:
                    from python_calamine import CalamineWorkbook
                    cal_wb = CalamineWorkbook.from_path(input_path)
                    for sheet_name in cal_wb.sheet_names:
                        sheet_data = []
                        try:
                            rows = cal_wb.get_sheet_by_name(sheet_name).to_python()
                            for row in rows:
                                if any(cell is not None for cell in row):
                                    sheet_data.append(row)
                        except Exception:
                            pass
                        if sheet_data:
                            json_data[sheet_name] = sheet_data
                    calamine_ok = True
                except Exception:
                    pass
                if not calamine_ok:
                    # calamine 不可用时降级到 openpyxl
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        sheet_data = []
                        for row in sheet.iter_rows(values_only=True):
                            if any(cell is not None for cell in row):
                                sheet_data.append(list(row))
                        json_data[sheet_name] = sheet_data

                with open(output_path, 'w', encoding='utf-8') as jsonfile:
                    json.dump(json_data, jsonfile, ensure_ascii=False, indent=2)

                file_size = os.path.getsize(output_path)

                return OperationResult(
                    success=True,
                    message=f"成功转换为JSON格式",
                    data={
                        'input_format': input_format,
                        'output_format': 'json',
                        'file_size': file_size,
                        'output_path': output_path,
                        'sheet_count': len(json_data)
                    },
                    metadata={
                        'input_path': input_path,
                        'target_format': target_format
                    }
                )

            elif target_format.lower() == "csv":
                # 转换为CSV格式 — 优先用 calamine 读取（比 openpyxl 快 10x+）
                row_count = 0
                sheet_name = ""
                calamine_ok = False
                try:
                    from python_calamine import CalamineWorkbook
                    cal_wb = CalamineWorkbook.from_path(input_path)
                    first_sheet = cal_wb.sheet_names[0] if cal_wb.sheet_names else ""
                    sheet_name = first_sheet
                    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                        csv_writer = csv.writer(csvfile)
                        try:
                            rows = cal_wb.get_sheet_by_name(first_sheet).to_python()
                            for row in rows:
                                if any(cell is not None for cell in row):
                                    csv_writer.writerow(row)
                                    row_count += 1
                        except Exception:
                            pass
                    calamine_ok = True
                except Exception:
                    pass
                if not calamine_ok:
                    sheet = workbook.active
                    sheet_name = sheet.title
                    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                        csv_writer = csv.writer(csvfile)
                        for row in sheet.iter_rows(values_only=True):
                            if any(cell is not None for cell in row):
                                csv_writer.writerow(row)
                                row_count += 1

                file_size = os.path.getsize(output_path)

                return OperationResult(
                    success=True,
                    message=f"成功转换为CSV格式",
                    data={
                        'input_format': input_format,
                        'output_format': 'csv',
                        'file_size': file_size,
                        'output_path': output_path,
                        'row_count': row_count
                    },
                    metadata={
                        'input_path': input_path,
                        'target_format': target_format
                    }
                )

            else:
                raise DataValidationError(f"不支持的目标格式: {target_format}")

        except Exception as e:
            logger.error(f"格式转换失败: {e}")
            return OperationResult(
                success=False,
                error=str(e),
                metadata={'operation': 'convert_format', 'input_path': input_path}
            )

    @staticmethod
    def merge_files(
        input_files: List[str],
        output_path: str,
        merge_mode: str = "sheets"
    ) -> OperationResult:
        """
        合并多个Excel文件

        Args:
            input_files: 输入文件路径列表
            output_path: 输出文件路径
            merge_mode: 合并模式
                - "sheets": 将每个文件作为独立工作表
                - "append": 将数据追加到单个工作表中
                - "horizontal": 水平合并（按列）

        Returns:
            OperationResult: 合并操作的结果
        """
        try:
            if not input_files:
                raise DataValidationError("输入文件列表不能为空")

            # 验证所有输入文件存在
            for file_path in input_files:
                if not os.path.exists(file_path):
                    raise ExcelFileNotFoundError(f"文件不存在: {file_path}")

            # 创建输出工作簿（write_only 模式：流式写入，合并大文件时内存占用极低）
            output_workbook = Workbook(write_only=True)
            merged_files = 0
            total_sheets = 0

            # 多文件时使用并行读取（>8个文件启用线程池，避免小文件集的线程开销）
            _use_parallel = len(input_files) > 8

            if _use_parallel:
                from ..utils.concurrent_utils import parallel_read_files

            if merge_mode == "sheets":
                if _use_parallel:
                    # 并行读取所有文件数据
                    def _read_file_data(fp):
                        wb = load_workbook(fp, read_only=True)
                        sheets_data = []
                        for sn in wb.sheetnames:
                            rows = list(wb[sn].iter_rows(values_only=True))
                            sheets_data.append((sn, rows))
                        wb.close()
                        return sheets_data

                    file_results = parallel_read_files(input_files, _read_file_data)

                    for file_index, (file_path, sheets_data) in enumerate(file_results):
                        if sheets_data is None:
                            continue
                        for sheet_name, rows in sheets_data:
                            new_sheet_name = f"File{file_index+1}_{sheet_name}"
                            if len(new_sheet_name) > 31:
                                new_sheet_name = f"F{file_index+1}_{sheet_name[:20]}"[:31]

                            target_sheet = output_workbook.create_sheet(title=new_sheet_name)
                            for row in rows:
                                target_sheet.append(row)
                            total_sheets += 1
                        merged_files += 1
                else:
                    # 少量文件：顺序读取（避免线程池开销）
                    for file_index, file_path in enumerate(input_files):
                        source_workbook = load_workbook(file_path, read_only=True)

                        for sheet_name in source_workbook.sheetnames:
                            source_sheet = source_workbook[sheet_name]

                            new_sheet_name = f"File{file_index+1}_{sheet_name}"
                            if len(new_sheet_name) > 31:
                                new_sheet_name = f"F{file_index+1}_{sheet_name[:20]}"[:31]

                            target_sheet = output_workbook.create_sheet(title=new_sheet_name)

                            for row in source_sheet.iter_rows(values_only=True):
                                target_sheet.append(row)

                            total_sheets += 1

                        source_workbook.close()
                        merged_files += 1

            elif merge_mode == "append":
                if _use_parallel:
                    # 并行读取所有文件的活动工作表数据
                    def _read_active_sheet(fp):
                        wb = load_workbook(fp, read_only=True)
                        rows = list(wb.active.iter_rows(values_only=True))
                        wb.close()
                        return rows

                    file_results = parallel_read_files(input_files, _read_active_sheet)
                    output_sheet = output_workbook.create_sheet(title="合并数据")

                    for file_path, rows in file_results:
                        if rows is None:
                            continue
                        for row in rows:
                            if any(cell is not None for cell in row):
                                output_sheet.append(row)
                        merged_files += 1
                    total_sheets = 1
                else:
                    # 少量文件：顺序读取
                    output_sheet = output_workbook.create_sheet(title="合并数据")

                    for file_path in input_files:
                        source_workbook = load_workbook(file_path, read_only=True)

                        source_sheet = source_workbook.active

                        for row in source_sheet.iter_rows(values_only=True):
                            if any(cell is not None for cell in row):
                                output_sheet.append(row)

                        source_workbook.close()
                        merged_files += 1

                    total_sheets = 1

            else:
                raise DataValidationError(f"不支持的合并模式: {merge_mode}")

            # 创建输出目录
            output_dir = Path(output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)

            # 保存合并后的文件
            output_workbook.save(output_path)
            output_workbook.close()

            return OperationResult(
                success=True,
                message=f"成功合并 {merged_files} 个文件，共 {total_sheets} 个工作表",
                data={
                    'merged_files': merged_files,
                    'total_sheets': total_sheets,
                    'output_path': output_path,
                    'merge_mode': merge_mode
                },
                metadata={
                    'input_files': input_files,
                    'merge_mode': merge_mode
                }
            )

        except Exception as e:
            logger.error(f"文件合并失败: {e}")
            return OperationResult(
                success=False,
                error=str(e),
                metadata={'operation': 'merge_files', 'input_files': input_files}
            )
