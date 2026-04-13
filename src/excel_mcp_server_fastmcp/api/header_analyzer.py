"""
表头分析器 - 集中式双表头/单表头自动检测

核心职责：
1. 自动检测工作表是单行表头还是双行表头（中文描述 + 英文字段名）
2. 返回标准化的表头信息：表头行、数据起始行、列映射
3. 提供缓存机制，避免重复检测

使用方式：
    from src.excel_mcp_server_fastmcp.api.header_analyzer import HeaderAnalyzer

    info = HeaderAnalyzer.analyze(file_path, sheet_name)
    # info.is_dual → bool
    # info.header_rows → [1] 或 [1, 2]
    # info.data_start_row → 2 或 3
    # info.column_names → ['col1', 'col2', ...]
    # info.column_map → {'中文名': 'english_name', ...}
"""

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


@dataclass
class HeaderInfo:
    """表头分析结果"""

    is_dual: bool = False  # 是否双行表头
    header_rows: list[int] = field(default_factory=list)  # 表头所在行号列表 (1-based)
    data_start_row: int = 2  # 数据起始行号 (1-based)
    column_names: list[str] = field(default_factory=list)  # 有效列名（英文/实际字段名）
    descriptions: list[str] = field(default_factory=list)  # 中文描述（仅双表头时有值）
    column_map: dict[str, str] = field(default_factory=dict)  # {中文描述: 英文字段名}
    name_to_index: dict[str, int] = field(default_factory=dict)  # {列名: 列索引(0-based)}
    total_columns: int = 0  # 总列数
    raw_first_row: list[Any] = field(default_factory=list)  # 第1行原始数据
    raw_second_row: list[Any] = field(default_factory=list)  # 第2行原始数据（可能为空）

    def effective_header_row(self) -> int:
        """返回用于匹配列名的表头行号（双表头时返回第2行）"""
        return self.header_rows[-1] if self.header_rows else 1

    def resolve_column(self, key: str) -> int | None:
        """
        解析列名/中文描述 → 列索引 (0-based)
        支持英文字段名、中文描述、大小写不敏感
        """
        if not key:
            return None
        key_str = str(key).strip()

        # 精确匹配
        if key_str in self.name_to_index:
            return self.name_to_index[key_str]

        # 大小写不敏感
        lower_key = key_str.lower()
        for name, idx in self.name_to_index.items():
            if name.lower() == lower_key:
                return idx

        # 通过 column_map 匹配中文描述
        if self.is_dual:
            for desc, eng in self.column_map.items():
                if desc == key_str or desc.lower() == lower_key:
                    return self.name_to_index.get(eng)

        return None


# 模块级缓存：{file_path: {sheet_name: HeaderInfo}}
_cache: dict[str, dict[str, HeaderInfo]] = {}


def _cell_str(c: Any) -> str | None:
    """统一单元格值转字符串"""
    if c is None:
        return None
    if hasattr(c, "value"):
        v = c.value
    else:
        v = c
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _is_valid_field_name(s: str) -> bool:
    """判断字符串是否像有效的英文字段名（字母/下划线开头，含字母数字下划线）"""
    if not s or len(s) < 1:
        return False
    return bool(re.match(r"^[a-zA-Z_][a-zA-Z0-9_.#]*$", s))


def _contains_chinese(s: str) -> bool:
    """判断字符串是否包含中文"""
    return bool(any("\u4e00" <= ch <= "\u9fff" for ch in s))


def detect_from_rows(rows: list[list]) -> tuple[bool, int, list[str]]:
    """
    从原始行数据检测是否为双行表头。

    Args:
        rows: 工作表前几行数据（至少需要2行）

    Returns:
        tuple: (is_dual_header, header_row_idx, descriptions)
               - header_row_idx: 用于取列名的行索引(0-based)，单表头=0，双表头=1
               - descriptions: 第一行的中文描述列表（仅双表头时有值）
    """
    if not rows or len(rows) < 2:
        return False, 0, []

    first_row = [_cell_str(c) for c in rows[0]]
    second_row = [_cell_str(c) for c in rows[1]]

    # 第二行非空值
    second_non_empty = [v for v in second_row if v is not None]
    first_non_empty = [v for v in first_row if v is not None]

    # 至少需要2个非空列才能判断
    if len(second_non_empty) < 2:
        return False, 0, []

    # 第二行是否全是有效英文字段名
    all_valid_names = all(_is_valid_field_name(v) for v in second_non_empty)

    # 第一行是否有中文（或者第一行不是纯英文字段名格式）
    any_chinese = any(_contains_chinese(v) for v in first_non_empty)

    # 如果第二行全是有效英文名且第一行有中文 → 双表头
    # 额外条件：第一行不能也全是有效英文名（否则可能是两层都是英文名）
    first_all_valid = all(_is_valid_field_name(v) for v in first_non_empty) if first_non_empty else False

    is_dual = all_valid_names and any_chinese and not first_all_valid

    header_row_idx = 1 if is_dual else 0
    descriptions = first_row if is_dual else []

    return is_dual, header_row_idx, descriptions


class HeaderAnalyzer:
    """
    集中式表头分析器

    用法：
        info = HeaderAnalyzer.analyze(file_path, 'Sheet1')
        if info.is_dual:
            print(f"双表头，数据从第{info.data_start_row}行开始")
        col_idx = info.resolve_column('技能名称')  # 支持中英文
    """

    @classmethod
    def analyze(cls, file_path: str, sheet_name: str, force_refresh: bool = False) -> HeaderInfo:
        """
        分析指定工作表的表头结构（带缓存）。

        Args:
            file_path: Excel 文件路径
            sheet_name: 工作表名称
            force_refresh: 强制刷新缓存

        Returns:
            HeaderInfo: 表头分析结果
        """
        cache_key = str(Path(file_path).resolve())

        # 缓存命中
        if not force_refresh and cache_key in _cache:
            sheet_cache = _cache[cache_key]
            if sheet_name in sheet_cache:
                return sheet_cache[sheet_name]

        # 执行分析
        info = cls._do_analyze(file_path, sheet_name)

        # 写入缓存
        if cache_key not in _cache:
            _cache[cache_key] = {}
        _cache[cache_key][sheet_name] = info

        return info

    @classmethod
    def _do_analyze(cls, file_path: str, sheet_name: str) -> HeaderInfo:
        """执行实际的表头分析（内部方法）"""
        from openpyxl import load_workbook

        info = HeaderInfo()

        try:
            # 优先用 calamine 快速读取前两行（Rust 引擎，毫秒级）
            from python_calamine import CalamineWorkbook

            cal_wb = CalamineWorkbook.from_path(file_path)
            cal_ws = cal_wb.get_sheet_by_name(sheet_name)

            # 检查空工作表（calamine 对空表 iter_rows 会 panic）
            if cal_ws.height == 0 or not hasattr(cal_ws, "iter_rows"):
                return info

            rows_iter = cal_ws.iter_rows()
            first_raw = list(next(rows_iter, []))
            second_raw = list(next(rows_iter, []))

            info.raw_first_row = first_raw
            info.raw_second_row = second_raw

        except BaseException:
            # 捕获所有异常包括 pyo3_runtime.PanicException
            # fallback 到 openpyxl
            try:
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb[sheet_name]

                row_iter = ws.iter_rows(max_row=2, values_only=True)
                first_raw = list(next(row_iter, []))
                second_raw = list(next(row_iter, []))

                info.raw_first_row = first_raw
                info.raw_second_row = second_raw

                wb.close()
            except Exception:
                # 无法读取，返回默认值
                return info

        # 检测双表头
        rows_for_detect = [info.raw_first_row, info.raw_second_row]
        is_dual, header_row_idx, descriptions = detect_from_rows(rows_for_detect)

        info.is_dual = is_dual

        if is_dual:
            info.header_rows = [1, 2]
            info.data_start_row = 3
            info.descriptions = [_cell_str(c) or "" for c in info.raw_first_row]
            info.column_names = [_cell_str(c) or "" for c in info.raw_second_row]
        else:
            info.header_rows = [1]
            info.data_start_row = 2
            info.descriptions = []
            info.column_names = [_cell_str(c) or "" for c in info.raw_first_row]

        # 计算总列数（取第一个非空行末尾）
        non_empty_cols = [i for i, v in enumerate(info.column_names) if v]
        info.total_columns = max(non_empty_cols) + 1 if non_empty_cols else 0

        # 构建列名→索引映射
        for i, name in enumerate(info.column_names):
            if name:
                info.name_to_index[name] = i

        # 构建 中文名→英文名 映射（双表头）
        if is_dual:
            for i, (desc, eng) in enumerate(zip(info.descriptions, info.column_names)):
                if desc and eng:
                    info.column_map[desc] = eng

        return info

    @classmethod
    def invalidate(cls, file_path: str | None = None):
        """
        使缓存失效。

        Args:
            file_path: 指定文件路径则只清除该文件的缓存；None 表示全部清除
        """
        global _cache
        if file_path:
            cache_key = str(Path(file_path).resolve())
            _cache.pop(cache_key, None)
        else:
            _cache.clear()

    @classmethod
    def get_effective_header_row(cls, file_path: str, sheet_name: str, explicit_header_row: int | None = None) -> int:
        """
        获取有效的表头行号。

        如果调用方显式传了 header_row，优先使用显式值；
        否则自动检测后返回正确的表头行。

        Args:
            file_path: 文件路径
            sheet_name: 工作表名
            explicit_header_row: 显式指定的表头行号（用户传入的参数）

        Returns:
            int: 有效的表头行号 (1-based)
        """
        if explicit_header_row is not None and explicit_header_row > 0:
            return explicit_header_row

        info = cls.analyze(file_path, sheet_name)
        return info.effective_header_row()

    @classmethod
    def get_data_start_row(cls, file_path: str, sheet_name: str, explicit_header_row: int | None = None) -> int:
        """
        获取数据起始行号。

        Args:
            file_path: 文件路径
            sheet_name: 工作表名
            explicit_header_row: 显式指定的表头行号

        Returns:
            int: 数据起始行号 (1-based)
        """
        if explicit_header_row is not None and explicit_header_row > 0:
            # 用户显式指定了表头行，数据从表头下一行开始
            info = cls.analyze(file_path, sheet_name)
            # 如果用户指定的是第1行但实际是双表头，数据应该从第3行开始
            if info.is_dual and explicit_header_row == 1:
                return 3  # 双表头时即使指定 header_row=1，数据也从第3行开始
            return explicit_header_row + 1

        info = cls.analyze(file_path, sheet_name)
        return info.data_start_row

    @classmethod
    def resolve_key_to_column(
        cls,
        file_path: str,
        sheet_name: str,
        key: str,
        explicit_header_row: int | None = None,
    ) -> int | None:
        """
        将用户传入的键（可能是中文或英文）解析为列索引。

        这是所有接收 dict 数据的操作（batch_insert_rows、upsert_row 等）的核心入口。

        Args:
            file_path: 文件路径
            sheet_name: 工作表名
            key: 列名（中文描述 或 英文字段名）
            explicit_header_row: 显式指定的表头行号

        Returns:
            int: 列索引 (0-based)，找不到返回 None
        """
        info = cls.analyze(file_path, sheet_name)
        return info.resolve_column(key)

    @classmethod
    def summary_text(cls, info: HeaderInfo) -> str:
        """生成人类可读的表头摘要"""
        if info.is_dual:
            return f"双行表头(第{info.header_rows[0]}-{info.header_rows[1]}行), 数据从第{info.data_start_row}行开始, 共{info.total_columns}列"
        else:
            return f"单行表头(第{info.header_rows[0]}行), 数据从第{info.data_start_row}行开始, 共{info.total_columns}列"
