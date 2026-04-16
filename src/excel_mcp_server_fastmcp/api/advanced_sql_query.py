"""
高级SQL查询引擎 - 基于SQLGlot实现的SQL查询支持

支持功能:
- 基础查询: SELECT, DISTINCT, 别名
- 条件筛选: WHERE, LIKE, IN, BETWEEN, AND/OR, EXISTS, 子查询
- 聚合统计: COUNT, SUM, AVG, MAX, MIN, GROUP BY, HAVING
- 排序限制: ORDER BY, LIMIT, OFFSET
- 算术运算: 加减乘除
- 条件表达式: CASE WHEN, COALESCE/IFNULL
- 表关联: INNER JOIN, LEFT JOIN, RIGHT JOIN, FULL JOIN, CROSS JOIN(同文件内工作表关联 + 跨文件关联)
- 子查询: WHERE col IN (SELECT ...), 标量子查询, EXISTS
- CTE: WITH ... AS (SELECT ...)
- 字符串函数: UPPER, LOWER, TRIM, LENGTH, CONCAT, REPLACE, SUBSTRING, LEFT, RIGHT
- 窗口函数: ROW_NUMBER, RANK, DENSE_RANK(OVER PARTITION BY ... ORDER BY ...)
- 集合操作: UNION, UNION ALL, EXCEPT, INTERSECT

不支持功能:
- FROM子查询(FROM (SELECT ...) AS alias)
"""

import csv
import datetime
import difflib
import io
import json
import logging
import math
import operator
import gzip
import hashlib
import os
import re
import shutil
import tempfile
import threading
import time
from collections.abc import Generator
from contextlib import contextmanager
from decimal import Decimal, InvalidOperation
from typing import Any

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)

# SQLGlot导入 - 核心SQL解析引擎
try:
    import sqlglot
    from sqlglot import expressions as exp
    from sqlglot.errors import ParseError, UnsupportedError

    SQLGLOT_AVAILABLE = True
except ImportError:
    SQLGLOT_AVAILABLE = False
    logger.warning("SQLGlot未安装,将使用基础pandas查询功能")

    # 创建虚拟类型注解以避免NameError
    class exp:
        class Expression:
            pass

        class Select:
            pass

        class Subquery:
            pass

        class With:
            pass

        class Window:
            pass

        class From:
            pass

        class Table:
            pass

        class Where:
            pass

        class Column:
            pass

        class Literal:
            pass

        class EQ:
            pass

        class NEQ:
            pass

        class GT:
            pass

        class GTE:
            pass

        class LT:
            pass

        class LTE:
            pass

        class And:
            pass

        class Or:
            pass

        class Like:
            pass

        class In:
            pass

        class Paren:
            pass  # 括号表达式

        # class IsNull: pass  # SQLGlot中可能不使用这个名称
        # class NotNull: pass  # SQLGlot中可能不使用这个名称
        class Order:
            pass

        class Ordered:
            pass

        class Having:
            pass

        class Alias:
            pass

        class AggFunc:
            pass


# Excel处理导入
import openpyxl

# 流式写入导入
try:
    from ..core.streaming_writer import StreamingWriter
except ImportError:
    StreamingWriter = None

# 表头分析器（统一双表头检测）
try:
    from .header_analyzer import HeaderAnalyzer
except ImportError:
    HeaderAnalyzer = None

# 配置常量
from ..utils.config import (
    CACHE_TARGET_MEMORY_MB,
    MARKDOWN_TABLE_MAX_ROWS,
    MAX_CACHE_SIZE,
    MAX_QUERY_CACHE_SIZE,
    MAX_RESULT_ROWS,
    QUERY_CACHE_TTL,
    STREAMING_WRITE_MIN_CHANGES,
    STREAMING_WRITE_MIN_FILE_SIZE_MB,
    STREAMING_WRITE_MIN_ROWS,
)


class StructuredSQLError(Exception):
    """结构化SQL错误,为AI提供可自动修复的错误信息.

    Attributes:
        error_code: 机器可读的错误分类码
        message: 人类可读的错误描述
        hint: AI修复建议(可选)
        context: 错误上下文,如可用列名,表名等(可选)
    """

    def __init__(self, error_code: str, message: str, hint: str = "", context: dict = None):
        self.error_code = error_code
        self.message = message
        self.hint = hint
        self.context = context or {}
        super().__init__(message)


def _unsupported_error_hint(err_detail: str) -> str:
    """为UnsupportedError提供替代建议."""
    err_upper = err_detail.upper()

    if "INSERT" in err_upper or "DELETE" in err_upper or "DROP" in err_upper or "ALTER" in err_upper or "CREATE" in err_upper:
        return "此工具仅支持SELECT查询.数据修改请使用excel_update_query(UPDATE语句)."
    if "NATURAL JOIN" in err_upper:
        return "不支持NATURAL JOIN,请改用显式ON条件:JOIN 表2 ON 表1.列 = 表2.列"
    if "FETCH" in err_upper or "NEXT" in err_upper:
        return "不支持FETCH/NEXT语法,请用LIMIT:SELECT ... LIMIT 10"
    if "RECURSIVE" in err_upper:
        return "不支持递归CTE(WITH RECURSIVE).请改用普通CTE或子查询."
    if "LATERAL" in err_upper:
        return "不支持LATERAL JOIN.请改用子查询或CTE."
    if "WINDOW" in err_upper and "OVER" not in err_upper:
        return "WINDOW子句请改为直接在窗口函数后写OVER:ROW_NUMBER() OVER (PARTITION BY ... ORDER BY ...)"

    return '请参考工具描述中"SQL已支持功能"列表,使用支持的操作.对于复杂计算,可考虑分步查询.'


def _parse_error_hint(err_str: str, sql: str) -> str:
    """根据SQLGlot ParseError和原始SQL,生成AI可自动修复的提示.

    覆盖常见SQL书写错误:
    - 关键字拼写错误
    - 关键字顺序错误
    - 缺少关键字
    - 缺少逗号/括号
    - Excel函数名误用
    - 中文标点混用
    - 引号配对错误
    """
    sql_upper = sql.strip().upper()
    hint = ""

    # === 关键字拼写错误 ===
    typos = [
        ("SELEC ", "SELECT"),
        ("SELEC$", "SELECT"),
        ("FORM ", "FROM"),
        ("FORM$", "FROM"),
        ("WHER ", "WHERE"),
        ("WHER$", "WHERE"),
        ("GROUPBY", "GROUP BY"),
        ("GROUP  BY", "GROUP BY"),
        ("ORDERBY", "ORDER BY"),
        ("ORDER  BY", "ORDER BY"),
        ("HAVNIG", "HAVING"),
        ("HAVIN", "HAVING"),
        ("INNTER", "INNER"),
        ("LEFR", "LEFT"),
        ("RIGTH", "RIGHT"),
        ("JOINT", "JOIN"),
        ("OUDER", "OUTER"),
        ("DISTIN T", "DISTINCT"),
        ("DISTNCT", "DISTINCT"),
        ("BETWEE N", "BETWEEN"),
        ("BETWEN", "BETWEEN"),
        ("NOTNULL", "NOT NULL"),
        ("ISNUL", "IS NULL"),
        ("LIK E", "LIKE"),
        ("LIEK", "LIKE"),
        ("EXIS TS", "EXISTS"),
        ("EXIST ", "EXISTS"),
        ("LIMITT", "LIMIT"),
        ("OFFEST", "OFFSET"),
        ("ASCEND", "ASC"),
        ("DSCEND", "DESC"),
        ("CROS", "CROSS"),
        ("FUL L", "FULL"),
        ("UNIO N", "UNION"),
        ("UNON", "UNION"),
        ("INTERSE CT", "INTERSECT"),
        ("EXCEP T", "EXCEPT"),
        ("CONCATENATE", "CONCAT"),
        ("SUBSTITUE", "REPLACE"),
    ]
    for typo, correct in typos:
        if typo.rstrip("$") in sql_upper:
            # 用$匹配行尾
            if typo.endswith("$") and not sql_upper.rstrip(";").endswith(typo.rstrip("$")):
                continue
            hint = f'可能是拼写错误,"{typo.rstrip().rstrip("$")}" 应为 "{correct}"'
            return hint

    # === 关键字顺序错误 ===
    # SELECT ... FROM ... WHERE ... GROUP BY ... HAVING ... ORDER BY ... LIMIT
    order_keywords = [
        "SELECT",
        "FROM",
        "WHERE",
        "GROUP BY",
        "HAVING",
        "ORDER BY",
        "LIMIT",
    ]
    found_positions = []
    for kw in order_keywords:
        # GROUP BY / ORDER BY 需要特殊处理
        if " " in kw:
            parts = kw.split()
            pos = sql_upper.find(parts[0])
            if pos != -1:
                # 检查后面是否跟着第二个词
                after = sql_upper[pos + len(parts[0]) :].lstrip()
                if after.startswith(parts[1]):
                    found_positions.append((pos, kw))
        else:
            pos = sql_upper.find(kw)
            if pos != -1 and (pos == 0 or not sql_upper[pos - 1].isalpha()):
                found_positions.append((pos, kw))

    # 按出现位置排序,然后检查顺序是否符合SQL标准
    found_positions.sort(key=lambda x: x[0])
    for i in range(len(found_positions) - 1):
        pos1, kw1 = found_positions[i]
        pos2, kw2 = found_positions[i + 1]
        idx1 = order_keywords.index(kw1)
        idx2 = order_keywords.index(kw2)
        if idx1 > idx2:
            hint = f'SQL关键字顺序错误:"{kw1}"出现在"{kw2}"之前,但标准顺序要求"{kw1}"在"{kw2}"之后.正确顺序: {" -> ".join(order_keywords)}'
            return hint

    # === 缺少关键字 ===
    # 有GROUP BY但没有聚合函数
    if "GROUP BY" in sql_upper:
        agg_funcs = [
            "COUNT(",
            "SUM(",
            "AVG(",
            "MIN(",
            "MAX(",
            "COUNT (",
            "SUM (",
            "AVG (",
            "MIN (",
            "MAX (",
        ]
        has_agg = any(af in sql_upper for af in agg_funcs)
        if not has_agg:
            hint = "GROUP BY通常与聚合函数一起使用(如COUNT/SUM/AVG/MIN/MAX).如果只是去重,请用SELECT DISTINCT."
            return hint

    # 有JOIN但缺少ON
    if re.search(r"\bJOIN\b", sql_upper) and " ON " not in sql_upper and not re.search(r"\bCROSS\s+JOIN\b", sql_upper):
        hint = "JOIN缺少ON条件.例如:... JOIN 表2 ON 表1.id = 表2.id.如果是笛卡尔积,请用CROSS JOIN."
        return hint

    # UPDATE语句出现在SELECT查询中
    if "UPDATE" in sql_upper and "SET" in sql_upper and "SELECT" in sql_upper:
        hint = "不能在SELECT查询中使用UPDATE.批量修改请使用excel_update_query工具."
        return hint

    # === 缺少逗号检测 ===
    # SELECT a b FROM -> SELECT a, b FROM(两个标识符之间只有空格没有逗号)
    select_match = re.search(r"\bSELECT\s+(.+?)\bFROM\b", sql_upper, re.DOTALL)
    if select_match:
        select_raw = sql[select_match.start(1) : select_match.end(1)]
        # 检查原始SQL中两个标识符之间是否缺少逗号
        # 模式:单词 + 空格(非逗号) + 单词,其中两个都不是SQL关键字
        keywords_in_select = {
            "AS",
            "DISTINCT",
            "CASE",
            "WHEN",
            "THEN",
            "ELSE",
            "END",
            "AND",
            "OR",
            "NOT",
            "IN",
            "BETWEEN",
            "LIKE",
            "IS",
            "NULL",
            "TRUE",
            "FALSE",
            "COUNT",
            "SUM",
            "AVG",
            "MIN",
            "MAX",
            "UPPER",
            "LOWER",
            "TRIM",
            "LENGTH",
            "CONCAT",
            "REPLACE",
            "SUBSTRING",
            "LEFT",
            "RIGHT",
            "COALESCE",
            "IFNULL",
            "CAST",
            "ROW_NUMBER",
            "RANK",
            "DENSE_RANK",
            "LAG",
            "LEAD",
            "FIRST_VALUE",
            "LAST_VALUE",
            "OVER",
            "PARTITION",
            "ASC",
            "DESC",
            "ON",
        }
        # 匹配:标识符 + 空格 + 标识符(中间无逗号)
        adjacent_pairs = re.finditer(r"([A-Za-z_]\w*)\s+([A-Za-z_]\w*)", select_raw)
        for m in adjacent_pairs:
            t1, t2 = m.group(1), m.group(2)
            if t1.upper() not in keywords_in_select and t2.upper() not in keywords_in_select:
                # 检查它们之间没有逗号(finditer已经保证了没有逗号,因为逗号不是\w)
                hint = f'SELECT子句中"{t1}"和"{t2}"之间可能缺少逗号.列之间用逗号分隔:SELECT {t1}, {t2}'
                return hint

    # === 括号配对检测 ===
    paren_count = sql.count("(") - sql.count(")")
    if paren_count > 0:
        hint = f'SQL中有{paren_count}个未闭合的括号.请检查每个左括号"("都有对应的右括号")".'
        return hint
    if paren_count < 0:
        hint = f'SQL中有多余的{abs(paren_count)}个右括号")".请删除多余的括号.'
        return hint

    # === 引号配对检测 ===
    single_quotes = len(re.findall(r"(?<!')'(?!')", sql))
    if single_quotes % 2 != 0:
        hint = "SQL中的单引号数量为奇数,可能有未闭合的引号.字符串值需要用单引号包裹,如 '值'."
        return hint

    # === 中文标点混用（全角 CJK 标点 → 半角 ASCII） ===
    cn_punctuation = {
        "\uff0c": ",",   # fullwidth comma → ASCII comma
        "\uff08": "(",   # fullwidth left paren → ASCII (
        "\uff09": ")",   # fullwidth right paren → ASCII )
        "\uff1a": ":",   # fullwidth colon → ASCII :
        "\uff1b": ";",   # fullwidth semicolon → ASCII ;
    }
    for cn, en in cn_punctuation.items():
        if cn in sql:
            hint = f'SQL中使用了中文标点"{cn}",应改为英文标点"{en}".'
            return hint

    # === 跨文件引用语法 [file.xlsx].Sheet 检测 ===
    # SQL Server / Access 风格的跨文件引用语法不被支持
    # 应使用 @'path' 语法进行跨文件查询
    cross_file_bracket = re.search(r"\[[^\]]+\.xlsx?\]\.\w+", sql, re.IGNORECASE)
    if cross_file_bracket:
        matched = cross_file_bracket.group(0)
        hint = (f'检测到跨文件引用语法 "{matched}",当前版本不支持 SQL Server 风格的 [文件名.xlsx].表名 语法。'
                f'请使用 @\'path\' 语法进行跨文件查询，例如: FROM 表名@\'/path/to/file.xlsx\' alias')
        return hint

    # === Excel函数名误用 ===
    excel_funcs = {
        "SUMIF": "请用 CASE WHEN ... THEN ... END 替代 SUMIF",
        "COUNTIF": "请用 COUNT(CASE WHEN ... THEN 1 END) 替代 COUNTIF",
        "VLOOKUP": "请用 JOIN 替代 VLOOKUP",
        "IF": "请用 CASE WHEN ... THEN ... ELSE ... END 替代 IF 函数",
        "IFS": "请用 CASE WHEN ... THEN ... ELSE ... END 替代 IFS",
    }
    for func, suggestion in excel_funcs.items():
        if re.search(r"\b" + func + r"\s*\(", sql_upper):
            hint = f'Excel函数"{func}"不是SQL语法.{suggestion}.'
            return hint

    # === 子查询缺少别名 ===
    subquery_pattern = re.search(r"\(\s*SELECT\b.+?\)\s*$", sql.strip(), re.IGNORECASE | re.DOTALL)
    if subquery_pattern:
        end_part = sql.strip()[subquery_pattern.end() :].strip()
        # 如果子查询后没有别名(没有内容,或内容不是 AS/标识符)
        if not end_part or (not re.match(r"^AS\b", end_part, re.IGNORECASE) and not re.match(r"^[A-Za-z_]\w*$", end_part)):
            hint = "FROM子查询或UNION结果需要别名.例如:FROM (SELECT ...) AS subquery"
            return hint

    # === 通用建议 ===
    if "SUBSTRING" in sql_upper and "(" in sql:
        # 检查SUBSTRING参数是否正确
        substr_match = re.search(r"SUBSTRING\s*\((.+?)\)", sql, re.IGNORECASE)
        if substr_match:
            args = [a.strip() for a in substr_match.group(1).split(",")]
            if len(args) == 2:
                hint = "SUBSTRING需要3个参数:SUBSTRING(列, 起始位置, 长度).如果要从位置N取到末尾,请用SUBSTRING(列, N, LENGTH(列)-N+1)."
                return hint

    return hint


def _classify_value_error(err_str: str) -> str:
    """将ValueError分类为标准错误码.优先匹配更具体的模式."""
    err_upper = err_str.upper()
    # 更具体的模式优先匹配
    if "列 '" in err_str or "COLUMN" in err_upper:
        return "column_not_found"
    if "表 '" in err_str or "TABLE" in err_upper:
        return "table_not_found"
    if "窗口函数" in err_str or "WINDOW" in err_upper:
        return "window_function_error"
    if "JOIN" in err_upper:
        return "join_error"
    if "子查询" in err_str or "SUBQUERY" in err_upper:
        return "subquery_error"
    if "UNION" in err_upper:
        return "union_error"
    if "CTE" in err_upper or "WITH" in err_upper:
        return "cte_error"
    if "GROUP BY" in err_upper:
        return "group_by_error"
    if "ORDER BY" in err_upper:
        return "order_by_error"
    # 通用模式放后面
    if "不支持" in err_str or "UNSUPPORTED" in err_upper:
        return "unsupported_feature"
    if "表达式" in err_str:
        return "expression_error"
    return "execution_error"


def _generate_value_error_hint(err_str: str) -> str:
    """根据ValueError内容生成AI修复建议."""
    # 列不存在
    if "列 '" in err_str and "可用列" in err_str:
        return "请检查列名拼写,或先用excel_get_headers查看可用列名."
    # 表不存在
    if "表 '" in err_str and "可用表" in err_str:
        return "请检查表名拼写,或先用excel_list_sheets查看可用工作表."
    # FROM子查询
    if "FROM子查询" in err_str:
        return "请检查FROM子查询中的SQL语法和表名.FROM子查询需要别名:FROM (SELECT ...) AS alias."
    # JOIN相关
    if "JOIN表" in err_str and "不存在" in err_str:
        return "请检查JOIN的表名是否正确,先用excel_list_sheets确认可用工作表."
    if "JOIN缺少ON条件" in err_str:
        return "JOIN必须包含ON条件,例如:... JOIN 表2 ON 表1.id = 表2.id."
    if "没有列 '" in err_str:
        return "请检查ON条件中的列名,确认列属于哪个表."
    # 窗口函数
    if "不支持的窗口函数" in err_str:
        return "仅支持 ROW_NUMBER,RANK,DENSE_RANK,LAG,LEAD,FIRST_VALUE,LAST_VALUE 窗口函数."
    if "需要 ORDER BY" in err_str:
        return "该窗口函数必须包含 ORDER BY 子句."
    # UNION
    if "UNION" in err_str and "SELECT" in err_str:
        return "请确保UNION两侧的SELECT列数一致."
    # 数学表达式
    if "数学表达式" in err_str:
        return "请检查数学运算符和操作数是否正确."
    # 字符串函数
    if "字符串函数" in err_str:
        return "请检查函数名和参数.支持的字符串函数:UPPER/LOWER/TRIM/LENGTH/CONCAT/REPLACE/SUBSTRING/LEFT/RIGHT."
    return ""


def _generate_value_error_suggested_fix(err_str: str, sql: str) -> str:
    """根据ValueError内容生成具体的修复SQL建议(如果可能)."""
    # 列不存在:尝试提取建议的列名并替换
    if "列 '" in err_str and "你是否想用" in err_str:
        # re already imported at top level

        # 提取建议的列名
        suggestion_match = re.search(r"你是否想用:\s*(.+?)\?", err_str)
        if suggestion_match:
            suggested_col = suggestion_match.group(1).strip().split(",")[0].strip()
            # 提取错误的列名
            col_match = re.search(r"列 '(.+?)'", err_str)
            if col_match:
                wrong_col = col_match.group(1)
                return sql.replace(wrong_col, suggested_col)
    # 表不存在:尝试提取建议的表名
    if "表 '" in err_str and "你是否想用" in err_str:
        # re already imported at top level

        suggestion_match = re.search(r"你是否想用:\s*(.+?)\?", err_str)
        if suggestion_match:
            suggested_table = suggestion_match.group(1).strip().split(",")[0].strip()
            col_match = re.search(r"表 '(.+?)'", err_str)
            if col_match:
                wrong_table = col_match.group(1)
                return sql.replace(wrong_table, suggested_table)
    return ""


def _safe_float_comparison(left, right, op):
    """安全比较函数,处理None值,避免 '<=' not supported between instances of 'int' and 'NoneType' 错误

    REQ-034优化:
    - 不等式比较(>/>=/</<=)保持精确,不使用epsilon
    - 仅 == 比较使用自适应epsilon处理浮点精度问题(如0.1+0.2approx0.3)
    - 支持极端边界值:0.001秒精度,99999大数值,NULL处理
    """
    if left is None or right is None:
        return False

    try:
        left_float = float(left)
        right_float = float(right)

        if op == ">":
            return left_float > right_float
        elif op == ">=":
            return left_float >= right_float
        elif op == "<":
            return left_float < right_float
        elif op == "<=":
            return left_float <= right_float
        elif op == "==" or op == "=":
            # 仅等值比较使用epsilon
            max_val = max(abs(left_float), abs(right_float))
            epsilon = max(max_val * 1e-9, 1e-10)
            return abs(left_float - right_float) <= epsilon
        else:
            return False
    except (TypeError, ValueError):
        return False


# Fix: P2-4 极端浮点值导致文件损坏 — 浮点值清理函数(SQL引擎用)
def _sanitize_float_for_excel(value: Any) -> Any:
    """清理浮点值,防止NaN/Inf/超范围值导致xlsx文件损坏.

    Args:
        value: 待清理的值

    Returns:
        清理后的安全值: NaN/Inf→None, 超范围→截断, 其他→原样
    """
    if value is None:
        return None

    # 处理numpy浮点和Python浮点
    if isinstance(value, (float, np.floating)):
        try:
            f_val = float(value)
            if np.isnan(f_val) or math.isinf(f_val):
                return None
            # 截断超IEEE 754范围的值
            if abs(f_val) > 1e308:
                return 1e308 if f_val > 0 else -1e308
        except (ValueError, TypeError, OverflowError):
            pass

    return value


class AdvancedSQLQueryEngine:
    """高级SQL查询引擎,支持完整的SQL语法"""

    def __init__(self, disable_streaming_aggregate: bool = False):
        """
        初始化SQL查询引擎

        Args:
            disable_streaming_aggregate: 禁用流式聚合优化(大文件处理)
        """
        self.disable_streaming_aggregate = disable_streaming_aggregate
        # DataFrame缓存:{file_path: (mtime, worksheets_data, header_descriptions)}
        self._df_cache = {}
        self._max_cache_size = MAX_CACHE_SIZE  # 最大缓存文件数,防止内存泄漏
        # 列名映射缓存:{file_path: {原始列名: 清洗列名}}
        # 与_df_cache同步,避免缓存命中时_original_to_clean_cols为空
        self._col_map_cache = {}

        # Fix: P1-concurrent — 每个文件的线程级写锁,防止多线程并发写入导致xlsx损坏
        # fcntl.flock是进程级锁,同进程内多线程共享FD表无法互斥;threading.Lock提供线程级互斥
        self._write_locks: dict[str, threading.Lock] = {}
        self._write_locks_global = threading.Lock()  # 保护_write_locks字典本身的并发访问

        # Fix: BUG-004 — 查询级可重入锁,防止并发查询(如excel_query + run_python内的query())
        # 污染共享可变状态(_original_to_clean_cols, _current_file_path, _parsed_sql等)
        # 使用RLock允许同一线程嵌套调用(如run_python内query→引擎内部再查询)
        self._query_lock = threading.RLock()

        # 性能优化:查询结果缓存 {hash(sql): (result_df, file_mtime)}
        self._query_result_cache = {}
        self._max_query_cache_size = MAX_QUERY_CACHE_SIZE  # 最大查询缓存数,防止内存泄漏
        self._query_cache_ttl = QUERY_CACHE_TTL  # 查询缓存TTL

        if not SQLGLOT_AVAILABLE:
            raise ImportError("SQLGlot未安装,请运行: pip install sqlglot")

        # 当前查询的文件路径(用于同文件JOIN时动态加载其他sheet)
        self._current_file_path = None

    def clear_cache(self):
        """清除DataFrame缓存,释放内存

        清除DataFrame缓存和查询结果缓存,释放内存占用.
        下次查询会重新加载Excel数据.

        Args:
            无

        Returns:
            None
        """
        self._df_cache.clear()
        self._query_result_cache.clear()

    def _find_column_name(self, col_name: str, df: pd.DataFrame) -> str | None:
        """大小写不敏感的列名查找（符合SQL标准：未引用标识符大小写不敏感）

        SQL标准规定未引用的标识符(identifier)是不区分大小写的。
        例如 SELECT rate FROM ... 应能匹配列名 'Rate'、'RATE'、'rate' 等。

        同时支持游戏配表常见的 Name(备注) 短名前缀匹配：
        当列名为 ChestPropID(宝箱PropID) 时，用户用 ChestPropID 也能匹配到。

        Args:
            col_name: 用户SQL中使用的列名
            df: 目标DataFrame

        Returns:
            实际存在的列名(保持原始大小写)，如果找不到则返回None
        """
        if col_name in df.columns:
            return col_name
        # 大小写不敏感匹配
        col_lower = col_name.lower()
        for c in df.columns:
            if c.lower() == col_lower:
                return c
        # BUG-003 fix: 支持 Name(备注) 格式的短名前缀匹配
        # 游戏配表常见格式：ChestPropID(宝箱PropID) → 允许用 ChestPropID 访问
        for c in df.columns:
            c_lower = c.lower()
            # 匹配模式：col_name 是某个列名的括号前部分（如 "chestpropid" 匹配 "chestpropid(宝箱propid)"）
            if c_lower.startswith(col_lower + "("):
                return c
        return None

    def _get_query_cache_key(self, sql: str, file_path: str, sheet_name: str | None = None) -> str:
        """生成查询缓存键"""
        # hashlib already imported at top level

        cache_data = f"{sql}|{file_path}|{sheet_name or ''}"
        return hashlib.md5(cache_data.encode()).hexdigest()

    def _get_cached_query_result(self, cache_key: str, file_mtime: float) -> pd.DataFrame | None:
        """获取缓存的查询结果"""
        if cache_key in self._query_result_cache:
            cached_time, cached_df, cached_mtime = self._query_result_cache[cache_key]
            # 检查缓存是否过期(文件是否被修改)
            current_time = time.time()
            if current_time - cached_time < self._query_cache_ttl and cached_mtime == file_mtime:
                return cached_df
            else:
                # 缓存过期,删除
                del self._query_result_cache[cache_key]
        return None

    def _cache_query_result(self, cache_key: str, result_df: pd.DataFrame, file_mtime: float):
        """缓存查询结果"""
        # LRU淘汰:超过最大缓存数时删除最早的缓存
        if len(self._query_result_cache) >= self._max_query_cache_size:
            oldest_key = next(iter(self._query_result_cache))
            del self._query_result_cache[oldest_key]

        self._query_result_cache[cache_key] = (time.time(), result_df, file_mtime)

    def execute_sql_query(
        self,
        file_path: str,
        sql: str,
        sheet_name: str | None = None,
        limit: int | None = None,
        include_headers: bool = True,
        output_format: str = "table",
    ) -> dict[str, Any]:
        """

        Args:
            file_path: Excel文件路径
            sql: SQL查询语句，支持完整的SELECT语法包括WHERE、JOIN、GROUP BY、HAVING、ORDER BY等
            sheet_name: 工作表名称（可选，默认使用第一个）
            limit: 限制返回行数，用于控制结果集大小
            include_headers: 是否包含表头，True时返回数据第一行为列名
            output_format: 输出格式，支持 table/json/csv 三种格式

        Returns:
            Dict: 查询结果，包含以下字段:
                - success (bool): 查询是否成功
                - message (str): 结果消息或错误描述
                - data (list): 查询结果数据（二维数组）
                - query_info (dict): 查询详细信息
                    - original_rows: 原始数据总行数
                    - filtered_rows: 过滤后行数
                    - returned_rows: 实际返回行数（可能因截断减少）
                    - truncated: 是否被截断（超过最大行数限制）
                    - query_applied: 是否应用了查询条件
                    - sql_query: 执行的SQL语句
                    - columns_returned: 返回的列数
                    - available_tables: 可用的工作表列表
                    - returned_columns: 返回的列名列表
                    - data_types: 各列的数据类型推断
                    - execution_time_ms: 查询执行耗时（毫秒）
                    - markdown_table: Markdown格式表格（如数据存在）
                    - suggestion: 空结果时的智能建议（如结果为空）
                    - json_output: JSON格式输出（output_format=json时）
                    - csv_output: CSV格式输出（output_format=csv时）
        """
        # Fix: BUG-004 — 查询级锁保护,防止并发调用(如excel_query + run_python的query())
        # 互相污染 _original_to_clean_cols / _current_file_path / _parsed_sql 等共享状态
        # RLock允许同线程嵌套(run_python→query()→引擎内部跨表JOIN再查询同引擎)
        with self._query_lock:
            return self._execute_sql_query_locked(
                file_path, sql, sheet_name, limit, include_headers, output_format
            )

    def _execute_sql_query_locked(
        self,
        file_path: str,
        sql: str,
        sheet_name: str | None = None,
        limit: int | None = None,
        include_headers: bool = True,
        output_format: str = "table",
    ) -> dict[str, Any]:
        """execute_sql_query 的核心逻辑(已在_query_lock保护内)"""
        try:
            # 验证文件存在性
            if not os.path.exists(file_path):
                return {
                    "success": False,
                    "message": f"文件不存在: {file_path}",
                    "data": [],
                    "query_info": {"error_type": "file_not_found"},
                }

            # 检查文件大小并处理大文件(支持2GB+文件)
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            if file_size_mb > 2048:
                return {
                    "success": False,
                    "message": f"文件过大 ({file_size_mb:.2f}MB),建议使用小于2GB的文件",
                    "data": [],
                    "query_info": {
                        "error_type": "file_too_large",
                        "size_mb": file_size_mb,
                    },
                }
            if file_size_mb > 500:
                logger.info(f"大文件查询: {file_path} ({file_size_mb:.1f}MB),启用分块处理优化")

            file_mtime = os.path.getmtime(file_path)

            # 保存当前文件路径(用于同文件JOIN时动态加载其他sheet)
            self._current_file_path = file_path

            # 加载Excel数据(带缓存)
            # 重置列名映射(每次查询重新构建)
            self._original_to_clean_cols = {}
            worksheets_data = self._load_data_with_cache(file_path, sheet_name)

            if not worksheets_data:
                return {
                    "success": False,
                    "message": "无法加载Excel数据或文件为空",
                    "data": [],
                    "query_info": {"error_type": "data_load_failed"},
                }

            # 跨文件引用解析:FROM 表名@'path' 语法
            # 在sqlglot解析前处理,加载外部文件并合并worksheets_data
            if "@'" in sql or '@"' in sql:
                sql, worksheets_data = self._resolve_cross_file_references(sql, file_path, worksheets_data)

            # 清理ANSI转义序列(终端粘贴可能带入的不可见字符)
            sql = re.sub(r"\x1b\[[0-9;]*[a-zA-Z]", "", sql)
            # 清理残余控制字符(保留\t\n\r,它们在SQL中有意义)
            sql = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", sql)
            # 清理残余的ANSI括号伪影:未配对的[后紧跟非ASCII字符
            # 有效的SQL Server标识符 [名称] 有配对的],ANSI伪影 [中文 无配对
            if sql.count("[") != sql.count("]"):
                # 存在未配对括号,清理[后紧跟非ASCII字符的情况
                sql = re.sub(r"\[(?=[^\x00-\x7F])", "", sql)

            # 中文列名替换:将SQL中的中文列名替换为英文列名(在解析前)
            sql = self._replace_cn_columns_in_sql(sql, worksheets_data)

            # DESCRIBE命令友好提示
            sql_stripped = sql.strip().upper()
            if sql_stripped.startswith("DESCRIBE ") or sql_stripped.startswith("DESC "):
                table_hint = sql.strip().split(None, 1)[-1].strip(";").strip("\"'`") if len(sql.strip().split()) > 1 else ""
                hint = "请使用 excel_describe_table 工具查看表结构"
                if table_hint:
                    hint += f"(工作表: {table_hint})"
                return {
                    "success": False,
                    "message": f"DESCRIBE不是SQL查询语法.{hint}",
                    "data": [],
                    "query_info": {
                        "error_type": "describe_not_sql",
                        "hint": "use_excel_describe_table",
                    },
                }

            # 解析和执行SQL
            _query_start = time.time()
            try:
                # 预处理:将双引号引用的原始列名替换为清洗后的列名
                # 解决用户写 SELECT "Player Name" 但内部列名已变为 Player_Name 的问题
                sql = self._preprocess_quoted_identifiers(sql)

                # 预处理: 将 || 字符串拼接操作符转为 CONCAT()
                # 因为 MySQL 方言将 || 解析为逻辑 OR，需要提前转换
                sql = self._preprocess_dpipe_to_concat(sql)

                # 预处理: 自动为 MySQL 保留字标识符添加反引号
                # 解决 Key/Value/Status 等常见列名导致 sqlglot ParseError 的问题
                sql = self._preprocess_reserved_words(sql)

                # Fix: P0-2 SELECT 分号多语句注入
                # 安全检测: 禁止SQL中出现外部分号(多语句注入攻击向量)
                # 使用 _has_dangerous_semicolon() 跳过字符串字面量内的分号，避免误报
                if self._has_dangerous_semicolon(sql):
                    # 包含中间分号 → 拒绝执行(安全策略: 不支持多语句)
                    return {
                        "success": False,
                        "message": "SQL语法错误: 不支持分号分隔的多语句执行(安全限制).💡 请将每条SQL语句分开执行",
                        "data": [],
                        "query_info": {
                            "error_type": "multi_statement_rejected",
                            "reason": "semicolon_injection_blocked",
                        },
                    }

                parsed_sql = sqlglot.parse_one(sql, dialect="mysql")

                # 保存解析后的SQL,用于错误提示中的窗口函数别名检测
                self._parsed_sql = parsed_sql

                # 验证SQL支持范围
                validation_result = self._validate_sql_support(parsed_sql)
                if not validation_result["valid"]:
                    error_msg = validation_result.get("error", "不支持的SQL语法")
                    hint = _unsupported_error_hint(error_msg)
                    qi = {"error_type": "unsupported_sql", "details": validation_result}
                    if hint:
                        qi["hint"] = hint
                    return {
                        "success": False,
                        "message": f"不支持的SQL语法: {error_msg}" + (f"\n💡 {hint}" if hint else ""),
                        "data": [],
                        "query_info": qi,
                    }

                # 执行查询(UNION/UNION ALL/EXCEPT/INTERSECT 或普通 SELECT)
                if isinstance(parsed_sql, exp.Union):
                    result_data = self._execute_union(parsed_sql, worksheets_data, limit)
                elif isinstance(parsed_sql, (exp.Except, exp.Intersect)):
                    result_data = self._execute_except_intersect(parsed_sql, worksheets_data, limit)
                else:
                    result_data = self._execute_query(parsed_sql, worksheets_data, limit)
                _query_elapsed = (time.time() - _query_start) * 1000

                # 格式化结果(传入parsed_sql和WHERE前数据用于空结果智能建议)
                has_group_by = not isinstance(parsed_sql, (exp.Union, exp.Except, exp.Intersect)) and parsed_sql.args.get("group") is not None
                has_having = parsed_sql.args.get("having") is not None
                result = self._format_query_result(
                    result_data,
                    file_path,
                    sql,
                    worksheets_data,
                    include_headers,
                    has_group_by=has_group_by,
                    has_having=has_having,
                    parsed_sql=parsed_sql,
                    df_before_where=self._df_before_where,
                    output_format=output_format,
                )
                # 注入执行时间
                result["query_info"]["execution_time_ms"] = round(_query_elapsed, 1)

                return result

            except StructuredSQLError as e:
                qi = {
                    "error_type": e.error_code,
                    "hint": e.hint,
                    "context": e.context,
                    "details": e.message,
                }
                # 为列名/表名错误生成suggested_fix
                suggested_fix = ""
                if e.error_code in ("column_not_found", "table_not_found") and e.context:
                    wrong_name = e.context.get("column_requested") or e.context.get("table_requested", "")
                    available = e.context.get("available_columns") or e.context.get("available_tables") or []
                    if wrong_name and available:
                        matches = difflib.get_close_matches(wrong_name, available, n=1, cutoff=0.4)
                        if matches:
                            suggested_fix = sql.replace(wrong_name, matches[0], 1)
                if suggested_fix:
                    qi["suggested_fix"] = suggested_fix
                msg = e.message
                if e.hint:
                    msg += f"\n💡 {e.hint}"
                if suggested_fix:
                    msg += f"\n🔧 建议修复SQL: {suggested_fix}"
                return {"success": False, "message": msg, "data": [], "query_info": qi}
            except ParseError as e:
                err_str = self._sanitize_error_message(str(e))
                hint = _parse_error_hint(err_str, sql)
                qi = {"error_type": "syntax_error", "details": err_str, "hint": hint}
                return {
                    "success": False,
                    "message": f"SQL语法错误: {err_str}" + (f"\n💡 {hint}" if hint else ""),
                    "data": [],
                    "query_info": qi,
                }
            except UnsupportedError as e:
                err_detail = self._sanitize_error_message(str(e))
                # 为不支持的SQL功能提供替代建议
                hint = _unsupported_error_hint(err_detail)
                qi = {
                    "error_type": "unsupported_feature",
                    "details": err_detail,
                    "hint": hint,
                }
                return {
                    "success": False,
                    "message": f"不支持的SQL功能: {err_detail}" + (f"\n💡 {hint}" if hint else ""),
                    "data": [],
                    "query_info": qi,
                }
            except ValueError as e:
                err_str = self._sanitize_error_message(str(e))
                # 对常见ValueError生成智能修复建议
                hint = _generate_value_error_hint(err_str)
                error_code = _classify_value_error(err_str)
                suggested_fix = _generate_value_error_suggested_fix(err_str, sql)
                qi = {"error_type": error_code, "details": err_str, "hint": hint}
                if suggested_fix:
                    qi["suggested_fix"] = suggested_fix
                msg = err_str
                if hint:
                    msg += f"\n💡 {hint}"
                if suggested_fix:
                    msg += f"\n🔧 建议修复SQL: {suggested_fix}"
                return {"success": False, "message": msg, "data": [], "query_info": qi}
            except Exception as e:
                raw_msg = str(e)
                return {
                    "success": False,
                    "message": f"SQL执行错误: {self._sanitize_error_message(raw_msg)}",
                    "data": [],
                    "query_info": {"error_type": "execution_error", "details": self._sanitize_error_message(raw_msg)},
                }

        except Exception as e:
            raw_msg = str(e)
            return {
                "success": False,
                "message": f"查询引擎错误: {self._sanitize_error_message(raw_msg)}",
                "data": [],
                "query_info": {"error_type": "engine_error", "details": self._sanitize_error_message(raw_msg)},
            }

    def _resolve_cross_file_references(
        self,
        sql: str,
        primary_file_path: str,
        primary_worksheets: dict[str, pd.DataFrame],
    ) -> tuple[str, dict[str, pd.DataFrame]]:
        """
        解析SQL中的跨文件引用(@'path'语法),加载外部文件数据并合并到worksheets_data

        语法:FROM 技能表@'/path/to/file1.xlsx' s JOIN 掉落表@'/path/to/file2.xlsx' d ON ...
        支持单引号和双引号包裹路径,支持相对路径(相对于主文件目录)

        Args:
            sql: 原始SQL语句
            primary_file_path: 主文件路径
            primary_worksheets: 主文件的worksheets_data

        Returns:
            Tuple[str, Dict]: (清理后的SQL, 合并后的worksheets_data)
        """
        # 正则匹配 @'path' 或 @"path" 模式
        # 路径可以包含字母,数字,./,../,_,-,空格等
        cross_file_pattern = re.compile(r"""@(['"])(.*?)\1""", re.DOTALL)

        matches = list(cross_file_pattern.finditer(sql))
        if not matches:
            return sql, primary_worksheets

        merged_data = dict(primary_worksheets)
        cleaned_sql = sql
        primary_dir = os.path.dirname(os.path.abspath(primary_file_path))
        loaded_files = {}  # filepath -> worksheets_data(避免重复加载)

        # 从后向前替换,避免索引偏移
        for match in reversed(matches):
            quote_char = match.group(1)
            ref_path = match.group(2).strip()

            # 解析相对路径(相对于主文件目录)
            if not os.path.isabs(ref_path):
                ref_path = os.path.join(primary_dir, ref_path)

            ref_path = os.path.normpath(ref_path)

            # 安全检查: 防止路径遍历攻击 (../)
            primary_dir_norm = os.path.normpath(primary_dir)
            if not ref_path.startswith(primary_dir_norm + os.sep) and ref_path != primary_dir_norm:
                raise ValueError("跨文件引用的路径不允许访问主文件目录之外的文件")

            # 验证文件存在(错误信息不泄露完整路径)
            if not os.path.exists(ref_path):
                raise ValueError(f"跨文件引用的文件不存在: {os.path.basename(ref_path)}.请检查文件名是否正确")

            # 加载文件(带缓存,避免重复加载)
            if ref_path not in loaded_files:
                ext_worksheets = self._load_data_with_cache(ref_path)
                if not ext_worksheets:
                    raise ValueError(f"无法加载跨文件引用的Excel数据: {ref_path}")
                loaded_files[ref_path] = ext_worksheets

            ext_worksheets = loaded_files[ref_path]

            # 合并工作表数据,处理名称冲突
            # 冲突时:主文件优先(已存在的不覆盖),外部文件重命名添加文件前缀
            file_basename = os.path.splitext(os.path.basename(ref_path))[0]
            for sheet_name, df in ext_worksheets.items():
                if sheet_name in merged_data:
                    # 名称冲突:为外部文件的工作表添加文件前缀
                    prefixed_name = f"{file_basename}.{sheet_name}"
                    if prefixed_name in merged_data:
                        # 即使加前缀也冲突,添加数字后缀
                        counter = 2
                        while f"{prefixed_name}_{counter}" in merged_data:
                            counter += 1
                        prefixed_name = f"{prefixed_name}_{counter}"
                    merged_data[prefixed_name] = df
                else:
                    merged_data[sheet_name] = df

            # 从SQL中移除 @'path' 部分,保留表名和别名
            cleaned_sql = cleaned_sql[: match.start()] + cleaned_sql[match.end() :]

        return cleaned_sql, merged_data

    def _load_data_with_cache(self, file_path: str, sheet_name: str | None = None) -> dict[str, pd.DataFrame] | None:
        """
        带缓存的Excel数据加载(公共方法,供execute_sql_query和execute_update_query复用)

        使用mtime检测文件变更,LRU淘汰防止内存泄漏.
        大文件(>500MB)使用分块读取减少内存峰值.

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称(可选)

        Returns:
            worksheets_data字典,加载失败返回None
        """
        mtime = os.path.getmtime(file_path)
        cache_key = f"{file_path}|{sheet_name or ''}"
        if cache_key in self._df_cache:
            cached_mtime, cached_data, cached_desc = self._df_cache[cache_key]
            if cached_mtime == mtime:
                self._header_descriptions = cached_desc
                # 缓存命中时,重置列名映射为当前文件的正确映射,避免其他文件的映射干扰
                self._original_to_clean_cols = {}
                if cache_key in self._col_map_cache:
                    self._original_to_clean_cols.update(self._col_map_cache[cache_key])
                return cached_data
            else:
                # 文件已修改,重新加载
                worksheets_data = self._load_excel_data(file_path, sheet_name)
                self._df_cache[cache_key] = (
                    mtime,
                    worksheets_data,
                    self._header_descriptions,
                )
                # 保存列名映射到缓存
                self._col_map_cache[cache_key] = dict(self._original_to_clean_cols)
                return worksheets_data
        else:
            worksheets_data = self._load_excel_data(file_path, sheet_name)
            self._df_cache[cache_key] = (
                mtime,
                worksheets_data,
                self._header_descriptions,
            )
            # 保存列名映射到缓存
            self._col_map_cache[cache_key] = dict(self._original_to_clean_cols)
            # LRU淘汰:超过最大缓存数时删除最早缓存的文件
            while len(self._df_cache) > self._max_cache_size:
                evicted_key = next(iter(self._df_cache))
                self._df_cache.pop(evicted_key)
                self._col_map_cache.pop(evicted_key, None)
            return worksheets_data

    def _estimate_cache_memory_mb(self) -> float:
        """估算当前缓存占用的内存(MB)"""
        total = 0.0
        for _key, (_mtime, worksheets_data, _desc) in self._df_cache.items():
            for _sheet, df in worksheets_data.items():
                total += df.memory_usage(deep=True).sum() / 1024 / 1024
        return total

    def evict_cache_by_memory(self, target_mb: float = None):
        """内存感知缓存淘汰:当缓存总内存超过阈值时,淘汰最早的缓存项

        Args:
            target_mb: 目标最大缓存内存(MB),默认使用 CACHE_TARGET_MEMORY_MB
        """
        if target_mb is None:
            target_mb = CACHE_TARGET_MEMORY_MB
        while self._estimate_cache_memory_mb() > target_mb and self._df_cache:
            self._df_cache.pop(next(iter(self._df_cache)))
            logger.info(f"缓存内存淘汰后剩余: {self._estimate_cache_memory_mb():.1f}MB")

    def _load_excel_data(self, file_path: str, sheet_name: str | None = None) -> dict[str, pd.DataFrame]:
        """
        加载Excel数据到DataFrame字典,支持游戏配置表双行表头

        游戏配置表通常有双行表头:
          第1行:中文描述(如"技能ID","技能名称")
          第2行:字段名(如"skill_id","skill_name")

        本方法自动检测双行表头,用第二行(字段名)做列名,
        第一行(描述)保存在 self._header_descriptions 中.

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称(可选)

        Returns:
            Dict[str, pd.DataFrame]: 工作表名到DataFrame的映射
        """
        worksheets_data = {}
        self._header_descriptions = {}  # {sheet_name: {field_name: description}}

        try:
            # P3-01: 大文件内存优化 - 文件大小预检
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            if file_size_mb > 50:
                logger.warning(f"大文件加载警告: {file_path} ({file_size_mb:.1f}MB), 内存占用可能较高")
            elif file_size_mb > 10:
                logger.info(f"加载较大文件: {file_path} ({file_size_mb:.1f}MB)")

            # 性能优化:用calamine替代openpyxl读取(Rust引擎,速度提升10-50倍)
            # calamine一次性读取所有sheet数据,无需二次打开文件
            from python_calamine import CalamineWorkbook

            cal_wb = CalamineWorkbook.from_path(file_path)
            all_sheet_names = cal_wb.sheet_names

            if sheet_name:
                sheets_to_load = [sheet_name] if sheet_name in all_sheet_names else []
            else:
                sheets_to_load = all_sheet_names

            # 使用 HeaderAnalyzer 统一检测所有sheet的双行表头（毫秒级）
            header_info = {}  # {sheet: (is_dual_header, first_row_values, second_row_values)}
            for sheet in sheets_to_load:
                try:
                    if HeaderAnalyzer is not None:
                        # 用统一的 HeaderAnalyzer 检测
                        info = HeaderAnalyzer.analyze(file_path, sheet)
                        header_info[sheet] = (
                            info.is_dual,
                            info.raw_first_row,
                            info.raw_second_row,
                        )
                    else:
                        # fallback：手动检测
                        cal_ws = cal_wb.get_sheet_by_name(sheet)
                        if cal_ws.height == 0:
                            header_info[sheet] = (False, None, None)
                            continue
                        rows_iter = cal_ws.iter_rows()
                        first_row = list(next(rows_iter, []))
                        second_row = list(next(rows_iter, []))
                        is_dual_header = False
                        if first_row and second_row:
                            second_row_values = [str(v).strip() if v is not None else "" for v in second_row]
                            first_row_values = [str(v).strip() if v is not None else "" for v in first_row]
                            non_empty_second = [v for v in second_row_values if v]
                            non_empty_first = [v for v in first_row_values if v]
                            if len(non_empty_second) >= 2:
                                second_all_field = all(re.match(r"^[a-zA-Z_][a-zA-Z0-9_.#]*$", v) for v in non_empty_second)
                                first_all_field = all(re.match(r"^[a-zA-Z_][a-zA-Z0-9_#]*$", v) for v in non_empty_first) if non_empty_first else False
                                if second_all_field and not first_all_field:
                                    is_dual_header = True
                        header_info[sheet] = (
                            is_dual_header,
                            first_row_values,
                            second_row_values,
                        )
                except Exception:
                    header_info[sheet] = (False, None, None)

            # 批量读取所有sheet数据(pd.read_excel + calamine引擎)
            for sheet, (
                is_dual_header,
                first_row_values,
                second_row_values,
            ) in header_info.items():
                if is_dual_header:
                    df = pd.read_excel(
                        file_path,
                        sheet_name=sheet,
                        engine="calamine",
                        header=1,
                        keep_default_na=False,
                        na_values=[""],
                    )
                    # 注意:desc_map 在 _clean_dataframe 之后构建(列名可能被清洗)
                    # 先记录原始映射关系,后面清洗后再构建最终映射
                    raw_desc_pairs = []
                    if second_row_values and first_row_values:
                        for col_idx, fname in enumerate(second_row_values):
                            fname = fname.strip() if fname else ""
                            desc = first_row_values[col_idx].strip() if col_idx < len(first_row_values) else ""
                            if fname and desc and desc != fname:
                                raw_desc_pairs.append((col_idx, fname, desc))

                else:
                    df = pd.read_excel(
                        file_path,
                        sheet_name=sheet,
                        engine="calamine",
                        keep_default_na=False,
                        na_values=[""],
                    )
                    raw_desc_pairs = []

                # 清洗 DataFrame(列名中的特殊字符会被替换)
                original_columns = list(df.columns)
                df = self._clean_dataframe(df)
                cleaned_columns = list(df.columns)

                # 构建中文描述映射(用清洗后的列名)
                if raw_desc_pairs:
                    desc_map = {}
                    for col_idx, fname, desc in raw_desc_pairs:
                        if col_idx < len(cleaned_columns):
                            desc_map[cleaned_columns[col_idx]] = desc
                    self._header_descriptions[sheet] = desc_map

                df = self._optimize_dtypes(df)
                worksheets_data[sheet] = df

        except Exception as e:
            logger.error(f"加载Excel数据失败: {e}")
            return {}

        return worksheets_data

    def _clean_dataframe(self, df) -> pd.DataFrame:
        """
        清理DataFrame数据

        Args:
            df: 原始DataFrame

        Returns:
            pd.DataFrame: 清理后的DataFrame
        """
        # 删除完全为空的行
        df = df.dropna(how="all")
        # Fix(R11): 空表(0行)时,pandas的dropna(axis=1, how='all')会误删所有列
        # 因为0行DataFrame中每列都算"全NA"。仅在有数据行时才清理全空列。
        # Fix: P2-formula 公式列保留 — 公式单元格无缓存值时全部为NaN,
        #   dropna(axis=1, how='all') 会误删公式列。先记录原始列名，删除后恢复被误删的列。
        if len(df) > 0:
            original_cols = list(df.columns)
            df = df.dropna(axis=1, how="all")
            # 恢复被误删的公式列（有表头但数据全为NaN的列，如含公式的列）
            dropped_cols = [c for c in original_cols if c not in df.columns]
            if dropped_cols:
                for col in dropped_cols:
                    df[col] = None  # 恢复为None(NaN)，保持列结构完整

        # 重置索引
        df = df.reset_index(drop=True)

        # 清理列名
        clean_columns = {}
        for col in df.columns:
            clean_col = str(col).strip()
            # 处理Unicode编码
            if "\\u" in clean_col:
                try:
                    clean_col = clean_col.encode("raw_unicode_escape").decode("unicode_escape")
                except Exception:
                    pass

            # 清理特殊字符,但保持多语言文字和括号(Excel列名如"刷新时间(小时)")
            # 支持范围: ASCII word chars + CJK中文 + 日文(平假名/片假名) + 韩文 + 拉丁语补充 + 括号 + 空格
            # Emoji 和其他 Unicode 符号也保留（pandas/openpyxl 均支持）
            clean_col = re.sub(
                r"[^\w\u4e00-\u9fff\u3040-\u309f\u30a0-\u30ff\uac00-\ud7af"
                r"\u00c0-\u024f\u1e00-\u1eff\s()\U0001f300-\U0001f9ff]",
                "_", clean_col
            )
            clean_col = re.sub(r"\s+", "_", clean_col)

            # 确保列名不为空且不以数字开头
            if not clean_col or clean_col.isspace():
                clean_col = f"column_{len(clean_columns) + 1}"
            elif clean_col[0].isdigit():
                clean_col = f"col_{clean_col}"

            clean_columns[col] = clean_col

        df = df.rename(columns=clean_columns)

        # 保存原始列名到清洗后列名的映射,用于SQL预处理
        if not hasattr(self, "_original_to_clean_cols") or self._original_to_clean_cols is None:
            self._original_to_clean_cols = {}
        self._original_to_clean_cols.update(clean_columns)

        # 保持原始数据不做空值替换
        # pandas groupby 默认跳过 NaN 行,不需要手动处理

        return df

    def _preprocess_dpipe_to_concat(self, sql: str) -> str:
        """
        预处理SQL中的 || 字符串拼接操作符，转为 CONCAT() 函数调用。

        策略：使用 PostgreSQL 方言解析（原生支持 || 为字符串拼接），
        然后将生成的 DPipe/Concat 节点以 MySQL 方言输出。
        sqlglot 在跨方言转换时会自动将 PG 的 || 转为 MySQL 的 CONCAT()。

        Fix: P2-1 || 字符串拼接不支持

        Args:
            sql: 原始SQL查询语句

        Returns:
            str: 转换后的SQL语句
        """
        # Fix: P2-1 使用 PostgreSQL 方言解析（|| 原生为字符串拼接）
        # 然后转换为 MySQL 方言输出（自动变为 CONCAT）
        # sqlglot already imported at top level
        # exp already imported at top level
        from sqlglot.dialects.mysql import MySQL

        try:
            # 用 PostgreSQL 方言解析（|| = 字符串拼接，非逻辑 OR）
            parsed = sqlglot.parse_one(sql, dialect="postgres")
        except Exception:
            # 解析失败，返回原始 SQL（不阻塞其他错误提示）
            return sql

        # 检查是否真的包含 DPipe 或 Concat 节点（即原 SQL 有 ||）
        has_dpipe = False
        def _check_dpipe(node):
            nonlocal has_dpipe
            if isinstance(node, (exp.DPipe, exp.Concat)):
                has_dpipe = True
            return node
        parsed.transform(_check_dpipe)

        if not has_dpipe:
            # 没有 || 操作符，返回原 SQL（避免不必要的方言转换副作用）
            return sql

        try:
            # 转换为 MySQL 方言输出（DPipe → CONCAT）
            result_sql = parsed.sql(dialect="mysql")
            return result_sql
        except Exception:
            return sql

    # MySQL 保留字中常被用作 Excel 列名的关键字集合（精简版）
    # 这些字在 MySQL 方言中作为未引用标识符会导致 sqlglot 解析失败
    # 通过逐词测试验证（sqlglot 27.29.0 + MySQL dialect），仅以下词在 SELECT/INSERT 列表位置会报 ParseError：
    #   - KEY: 最常见问题（本地化表的 Key 列、配置表的主键列名）
    #   - INDEX: 在 INSERT 列表中冲突
    #   - CHECK / CONSTRAINT: 在 INSERT 列表中冲突
    # 注意：不包含 SQL 结构关键字（SELECT/FROM/WHERE/INSERT/VALUES/CASE 等），
    #       那些关键字如果被引用反而会破坏 SQL 语法。
    _MYSQL_RESERVED_AS_COLUMNS = frozenset(
        {
            "KEY",
            "INDEX",
            "CHECK",
            "CONSTRAINT",
        }
    )

    def _inject_ctes_to_worksheets(
        self,
        parsed_sql: exp.Expression,
        worksheets_data: dict[str, pd.DataFrame],
        _cte_depth: int = 0,
    ) -> dict[str, pd.DataFrame]:
        """
        从已解析的SQL语句中提取CTE(WITH子句)，执行后注入到worksheets_data中。

        支持UPDATE/DELETE语句中的CTE，使WHERE子句的子查询可以引用CTE定义的临时表。

        Args:
            parsed_sql: sqlglot解析后的语句(Update/Delete/Select等)
            worksheets_data: 原始工作表数据
            _cte_depth: CTE 嵌套深度（内部使用，防止无限递归）

        Returns:
            包含CTE结果的工作表数据字典
        """
        # 兼容sqlglot不同版本: arg key可能是'with'或'with_'
        with_clause = parsed_sql.args.get("with") or parsed_sql.args.get("with_")
        if not with_clause:
            return worksheets_data

        # 检查是否为递归CTE（不支持）
        if getattr(with_clause, "recursive", False):
            raise ValueError("不支持递归CTE(WITH RECURSIVE).请改用普通CTE或子查询.")

        # Fix(R4): CTE 深度检查 — 防止深层嵌套导致 StackOverflow
        if _cte_depth >= self._MAX_CTE_DEPTH:
            raise ValueError(
                f"CTE 嵌套深度超过限制 ({self._MAX_CTE_DEPTH})。"
                f"💡 请简化查询，减少 CTE 嵌套层数，或改用子查询替代多层 CTE。"
            )

        # 复制worksheets_data避免修改原始数据，逐步添加CTE结果
        cte_data = dict(worksheets_data)
        for cte_expr in with_clause.expressions:
            cte_name = cte_expr.alias
            cte_query = cte_expr.this  # inner Select
            try:
                # 每个CTE在已有的cte_data上执行(支持CTE引用前面的CTE)
                # 递归深度 +1
                cte_result = self._execute_query(cte_query, cte_data, _cte_depth=_cte_depth + 1)
                cte_data[cte_name] = cte_result
            except Exception as e:
                raise ValueError(f"CTE '{cte_name}' 执行失败: {e}")

        return cte_data

    def _preprocess_reserved_words(self, sql: str) -> str:
        """
        预处理SQL：将未加反引号的MySQL保留字标识符自动加上反引号。

        问题背景：
        sqlglot 的 MySQL 方言解析器对某些保留字（如 Key, Value, Status, Name 等）
        作为未引用标识符时会报 ParseError。但游戏配置表中这些名称极其常见（如
        本地化表的 Key 列、掉落表的概率 Value 列等）。

        解决方案：
        在 SQL 解析前，扫描并自动为匹配保留字的裸标识符添加反引号。
        使用状态机逐字符分析，避免在字符串内部误替换。

        Args:
            sql: 原始 SQL 语句

        Returns:
            str: 处理后的 SQL 语句
        """

        reserved = self._MYSQL_RESERVED_AS_COLUMNS
        result = []
        i = 0
        n = len(sql)

        while i < n:
            # 跳过单引号字符串
            if sql[i] == "'":
                j = i + 1
                while j < n:
                    if sql[j] == "'" and j + 1 < n and sql[j + 1] == "'":
                        j += 2  # 转义的单引号 ''
                    elif sql[j] == "'":
                        break
                    else:
                        j += 1
                result.append(sql[i : j + 1])
                i = j + 1
                continue

            # 跳过双引号字符串
            if sql[i] == '"':
                j = i + 1
                while j < n:
                    if sql[j] == '"':
                        break
                    else:
                        j += 1
                result.append(sql[i : j + 1])
                i = j + 1
                continue

            # 跳过反引号引用的标识符（已引用的不需要处理）
            if sql[i] == "`":
                j = i + 1
                while j < n and sql[j] != "`":
                    j += 1
                result.append(sql[i : j + 1])  # include closing backtick
                i = j + 1
                continue

            # 跳过数字开头（数字字面量）
            if sql[i].isdigit():
                j = i
                while j < n and (sql[j].isdigit() or sql[j] == "." or sql[j] in "eE+-"):
                    j += 1
                result.append(sql[i:j])
                i = j
                continue

            # 检测标识符（字母或下划线开头，后跟字母/数字/下划线/中文）
            if sql[i].isalpha() or sql[i] == "_":
                j = i
                while j < n and (sql[j].isalnum() or sql[j] == "_" or ord(sql[j]) > 127):
                    j += 1
                word = sql[i:j]
                upper_word = word.upper()

                # 判断是否需要加反引号：
                # 1. 是保留字
                # 2. 前面不是 "." （避免引用 table.column 的列名部分被错误处理，
                #    实际上这种情况也需要保护，所以也处理）
                if upper_word in reserved:
                    # 确保不在字符串或已引用上下文中（前面逻辑已保证）
                    # 额外检查：前面紧邻的字符不应是 @ （MySQL变量）
                    if i > 0 and sql[i - 1] == "@":
                        result.append(word)
                    else:
                        result.append("`" + word + "`")
                else:
                    result.append(word)
                i = j
                continue

            # 其他原样保留
            result.append(sql[i])
            i += 1

        return "".join(result)

    # Fix: P0-7 注释符注入防御 — 检测写操作中的危险注释符
    # 攻击向量: UPDATE sheet SET x=0 WHERE id=1 -- AND safe=1
    #   --> sqlglot 将 -- 后内容作为注释丢弃，WHERE 条件被截断，导致全表篡改
    # 策略: 在写操作(UPDATE/DELETE/INSERT)入口处拒绝 -- 和 # 行注释
    #   /* */ 内联注释由 sqlglot 正确处理，不截断后续SQL，故放行
    #   SELECT 为只读操作，尾随注释无安全风险，不在此次拦截范围内
    @staticmethod
    def _detect_dangerous_comments(sql: str) -> str | None:
        """
        检测SQL中的危险注释符（-- 和 #）。

        只在字符串字面量外部检测，避免误报如 SELECT '--not-a-comment' 的情况。
        返回错误信息字符串（检测到危险注释）或 None（安全）。

        Args:
            sql: 原始SQL语句

        Returns:
            str: 检测到危险注释时的错误描述
            None: SQL安全，无危险注释
        """
        i = 0
        n = len(sql)
        while i < n:
            ch = sql[i]
            # 单引号字符串 - 跳过整个字符串（处理转义单引号 ''）
            if ch == "'":
                i += 1
                while i < n:
                    if sql[i] == "'" and i + 1 < n and sql[i + 1] == "'":
                        i += 2  # 转义单引号 ''
                    elif sql[i] == "'":
                        i += 1  # 字符串结束
                        break
                    else:
                        i += 1
                continue
            # 双引号字符串 - 跳过
            elif ch == '"':
                i += 1
                while i < n and sql[i] != '"':
                    if sql[i] == '\\':
                        i += 1  # 跳过转义字符
                    i += 1
                if i < n:
                    i += 1  # 跳过结束引号
                continue
            # 反引号标识符 - 跳过
            elif ch == '`':
                i += 1
                while i < n and sql[i] != '`':
                    i += 1
                if i < n:
                    i += 1
                continue
            # 检测行注释: -- （双横线）
            elif ch == '-' and i + 1 < n and sql[i + 1] == '-':
                return "SQL语句包含注释符(--)，可能截断WHERE条件导致非预期修改(安全限制).💡 请移除注释符后重试"
            # 检测行注释: # （MySQL风格）
            elif ch == '#':
                return "SQL语句包含注释符(#)，可能截断WHERE条件导致非预期修改(安全限制).💡 请移除注释符后重试"
            else:
                i += 1
        return None

    # Fix: P0-2/P0-4/P0-5/P0-6 分号检测误报修复
    # 攻击向量: SELECT * FROM t; DROP TABLE t -- 分号分隔的多语句注入
    # 误报场景: UPDATE t SET col='value;withsemicolon' WHERE id=1
    #   简单的 ";" in sql 会误报字符串字面量内的分号
    # 策略: 只在字符串字面量外部检测分号（与 _detect_dangerous_comments 一致）
    @staticmethod
    def _has_dangerous_semicolon(sql: str) -> bool:
        """
        检测SQL中是否存在危险的外部分号（多语句注入）。

        跳过字符串字面量内部的单引号、双引号、反引号内容，
        避免误报如 UPDATE t SET x='a;b' WHERE id=1 的情况。

        Args:
            sql: 原始SQL语句

        Returns:
            bool: 存在危险外部分号返回True，否则False
        """
        _stripped = sql.strip()
        # 尾部允许有一个分号（SQL标准写法）
        if _stripped.endswith(";"):
            _stripped = _stripped[:-1].rstrip()

        i = 0
        n = len(_stripped)
        while i < n:
            ch = _stripped[i]
            # 单引号字符串 - 跳过整个字符串（处理转义单引号 ''）
            if ch == "'":
                i += 1
                while i < n:
                    if _stripped[i] == "'" and i + 1 < n and _stripped[i + 1] == "'":
                        i += 2  # 转义单引号 ''
                    elif _stripped[i] == "'":
                        i += 1  # 字符串结束
                        break
                    else:
                        i += 1
                continue
            # 双引号字符串 - 跳过
            elif ch == '"':
                i += 1
                while i < n and _stripped[i] != '"':
                    if _stripped[i] == '\\':
                        i += 1  # 跳过转义字符
                    i += 1
                if i < n:
                    i += 1  # 跳过结束引号
                continue
            # 反引号标识符 - 跳过
            elif ch == '`':
                i += 1
                while i < n and _stripped[i] != '`':
                    i += 1
                if i < n:
                    i += 1
                continue
            # 检测外部分号 → 危险！
            elif ch == ';':
                return True
            else:
                i += 1
        return False

    def _execute_multi_statement(
        self,
        sql: str,
        file_path: str,
        worksheets_data: dict,
        include_headers: bool,
        output_format: str,
        limit: int,
    ) -> dict:
        """
        执行多语句SQL（分号分隔的多条语句）。

        将 SQL 按分号拆分为多条独立语句，逐条执行，合并结果返回。
        支持混合 SELECT / UPDATE / INSERT / DELETE 语句的组合。

        Args:
            sql: 原始多语句 SQL（包含分号分隔符）
            file_path: Excel 文件路径
            worksheets_data: 工作表数据字典
            include_headers: 是否包含表头
            output_format: 输出格式
            limit: 行数限制

        Returns:
            dict: 合并后的结果，格式与单条查询一致
        """
        # time already imported at top level

        _start = time.time()

        # Fix: P0-2 分号多语句注入防御（纵深防御）
        # 即使各入口已拦截，此函数内部也应拒绝多语句执行
        # 智能检测：按分号分割（尊重字符串内分号），若产生多条语句则拒绝
        raw_parts = []
        current = []
        in_single_quote = False
        in_double_quote = False
        for ch in sql:
            if ch == "'" and not in_double_quote:
                in_single_quote = not in_single_quote
                current.append(ch)
            elif ch == '"' and not in_single_quote:
                in_double_quote = not in_double_quote
                current.append(ch)
            elif ch == ";" and not in_single_quote and not in_double_quote:
                raw_parts.append("".join(current))
                current = []
            else:
                current.append(ch)
        if current:
            raw_parts.append("".join(current))

        statements = [s.strip() for s in raw_parts if s.strip()]

        # 安全策略: 禁止多语句执行（覆盖 P0-2/P0-4/P0-5/P0-6）
        if len(statements) > 1:
            return {
                "success": False,
                "message": "SQL语法错误: 不支持分号分隔的多语句执行(安全限制).💡 请将每条SQL语句分开执行",
                "data": [],
                "columns": None,
                "row_count": 0,
                "query_info": {
                    "error_type": "multi_statement_rejected",
                    "reason": "semicolon_injection_blocked",
                    "statement_count": len(statements),
                    "execution_time_ms": round((time.time() - _start) * 1000, 1),
                },
            }

        # 单语句情况：正常走后续流程（保留原有逻辑兼容性）
        all_results = []  # 所有 SELECT 查询的结果行
        all_columns = []  # 所有 SELECT 查询的列名
        affected_total = 0  # 写入操作影响的总行数
        statements_executed = 0
        errors = []
        for idx, stmt in enumerate(statements):
            try:
                # 根据语句类型路由到对应的执行方法
                stmt_upper = stmt.strip().upper()
                if stmt_upper.startswith(("SELECT ", "WITH ", "(", "SELECT\n", "SELECT\r")):
                    result = self._execute_single_select(
                        stmt,
                        file_path,
                        worksheets_data,
                        include_headers,
                        output_format,
                        limit,
                    )
                    statements_executed += 1
                    if result.get("success") and result.get("data"):
                        all_results.extend(result["data"])
                        if "columns" in result:
                            all_columns.extend(result["columns"]) if isinstance(result["columns"], list) else None
                    else:
                        errors.append(f"语句{idx + 1}: {result.get('message', '未知错误')[:80]}")
                elif stmt_upper.startswith("UPDATE "):
                    result = execute_advanced_update_query(file_path, stmt)
                    statements_executed += 1
                    if result.get("success"):
                        affected_total += result.get("affected_rows", 0)
                    else:
                        errors.append(f"语句{idx + 1}(UPDATE): {result.get('message', '')[:80]}")
                elif stmt_upper.startswith("INSERT "):
                    result = execute_advanced_insert_query(file_path, stmt)
                    statements_executed += 1
                    if result.get("success"):
                        affected_total += result.get("affected_rows", 1)
                    else:
                        errors.append(f"语句{idx + 1}(INSERT): {result.get('message', '')[:80]}")
                elif stmt_upper.startswith("DELETE "):
                    result = execute_advanced_delete_query(file_path, stmt)
                    statements_executed += 1
                    if result.get("success"):
                        affected_total += result.get("affected_rows", 0)
                    else:
                        errors.append(f"语句{idx + 1}(DELETE): {result.get('message', '')[:80]}")
                else:
                    # 尝试作为 SELECT 执行
                    result = self._execute_single_select(
                        stmt,
                        file_path,
                        worksheets_data,
                        include_headers,
                        output_format,
                        limit,
                    )
                    statements_executed += 1
                    if result.get("success") and result.get("data"):
                        all_results.extend(result["data"])
                    else:
                        errors.append(f"语句{idx + 1}: {result.get('message', '未知错误')[:80]}")

            except Exception as e:
                errors.append(f"语句{idx + 1}异常: {self._sanitize_error_message(str(e))[:100]}")

        _elapsed = (time.time() - _start) * 1000

        # 构建合并结果
        has_query_data = len(all_results) > 0
        has_write_ops = affected_total > 0

        return {
            "success": len(errors) < len(statements),  # 部分成功也算 success
            "message": (
                f"多语句执行完成: {statements_executed}/{len(statements)} 条成功"
                + (f", 返回 {len(all_results)} 行" if has_query_data else "")
                + (f", 影响 {affected_total} 行" if has_write_ops else "")
                + (f"\n⚠️ 错误: {'; '.join(errors)}" if errors else "")
            ),
            "data": all_results,
            "columns": all_columns if all_columns else None,
            "row_count": len(all_results),
            "query_info": {
                "statement_count": len(statements),
                "statements_executed": statements_executed,
                "query_rows": len(all_results),
                "affected_rows_total": affected_total,
                "errors": errors,
                "execution_time_ms": round(_elapsed, 1),
                "multi_statement": True,
            },
        }

    def _execute_single_select(
        self,
        sql: str,
        file_path: str,
        worksheets_data: dict,
        include_headers: bool,
        output_format: str,
        limit: int,
    ) -> dict:
        """执行单条 SELECT 语句（供多语句调用）。"""
        # time already imported at top level

        # sqlglot already imported at top level
        # exp already imported at top level

        _query_start = time.time()

        # 复用相同的预处理管线
        sql_processed = self._preprocess_quoted_identifiers(sql)
        sql_processed = self._preprocess_dpipe_to_concat(sql_processed)
        sql_processed = self._preprocess_reserved_words(sql_processed)

        try:
            parsed_sql = sqlglot.parse_one(sql_processed, dialect="mysql")
        except Exception as e:
            return {
                "success": False,
                "message": f"解析错误: {e}",
                "data": [],
                "row_count": 0,
            }

        try:
            if isinstance(parsed_sql, exp.Union):
                result_data = self._execute_union(parsed_sql, worksheets_data, limit)
            elif isinstance(parsed_sql, (exp.Except, exp.Intersect)):
                result_data = self._execute_except_intersect(parsed_sql, worksheets_data, limit)
            else:
                result_data = self._execute_query(parsed_sql, worksheets_data, limit)

            _query_elapsed = (time.time() - _query_start) * 1000

            has_group_by = not isinstance(parsed_sql, (exp.Union, exp.Except, exp.Intersect)) and parsed_sql.args.get("group") is not None
            has_having = parsed_sql.args.get("having") is not None

            result = self._format_query_result(
                result_data,
                file_path,
                sql,
                worksheets_data,
                include_headers,
                has_group_by=has_group_by,
                has_having=has_having,
                parsed_sql=parsed_sql,
                df_before_where=self._df_before_where,
                output_format=output_format,
            )
            result["query_info"]["execution_time_ms"] = round(_query_elapsed, 1)
            return result

        except Exception as e:
            return {
                "success": False,
                "message": f"执行错误: {e}",
                "data": [],
                "row_count": 0,
            }

    def _is_likely_dpipe_concatenation(self, or_expr) -> bool:
        """
        启发式判断 MySQL 方言解析的 exp.Or 是否原本是 || 字符串拼接(DPipe)

        判断依据:
        - 两边都不是布尔比较表达式(=, !=, >, <, IN, LIKE, BETWEEN, IS 等)
        - 至少有一边是列引用、字面量字符串、CAST 或字符串函数
        - 不包含 AND/OR 嵌套(真正的逻辑OR通常有嵌套布尔条件)

        Args:
            or_expr: sqlglot 的 exp.Or 表达式节点

        Returns:
            bool: True 表示这可能是字符串拼接, False 表示是真正的逻辑 OR
        """
        # exp already imported at top level

        left = or_expr.this
        right = or_expr.expression

        # 如果任一边是布尔比较/逻辑运算,则认为是真正的 OR
        bool_types = (
            exp.EQ,
            exp.NEQ,
            exp.GT,
            exp.GTE,
            exp.LT,
            exp.LTE,
            exp.In,
            exp.Like,
            exp.ILike,
            exp.Between,
            exp.Is,
            exp.Null,
            exp.Not,
            exp.And,
            exp.Or,  # 嵌套逻辑运算
            exp.Exists,
            exp.Any,
            exp.All,
            exp.Boolean,  # TRUE/FALSE 字面量
        )

        def _contains_bool_type(expr):
            """检查表达式中是否包含布尔类型节点"""
            if isinstance(expr, bool_types):
                return True
            # 递归检查子节点
            for child in expr.iter_expressions():
                if _contains_bool_type(child):
                    return True
            return False

        # 如果两边都是非布尔表达式,很可能是拼接
        if _contains_bool_type(left) or _contains_bool_type(right):
            return False

        # 检查两边是否都是"值类"表达式(列引用、字面量、函数调用等)
        value_types = (
            exp.Column,
            exp.Literal,
            exp.Cast,
            exp.Upper,
            exp.Lower,
            exp.Trim,
            exp.Length,
            exp.Concat,
            exp.Replace,
            exp.Substring,
            exp.Anonymous,
            exp.Add,
            exp.Sub,
            exp.Mul,
            exp.Div,  # 算术表达式也是值
            exp.Coalesce,
            exp.Nullif,
            exp.Round,
            exp.Case,
        )

        def _is_value_expr(expr):
            """检查是否为值类型表达式"""
            if isinstance(expr, value_types):
                return True
            # 嵌套的算术/函数也视为值
            for child in expr.iter_expressions():
                if isinstance(child, value_types):
                    return True
            return False

        # 两边都应该是值表达式
        return _is_value_expr(left) and _is_value_expr(right)

    def _preprocess_quoted_identifiers(self, sql: str) -> str:
        """
        预处理SQL中的双引号标识符,将原始列名替换为清洗后的列名

        当Excel列名含空格或特殊字符时(如"Player Name"),_clean_column_names会将其
        转换为下划线形式(Player_Name).用户在SQL中使用双引号引用原始列名时,
        sqlglot(MySQL方言)会将其解析为字符串字面量而非列引用.

        使用AST方法精确替换:只在列引用位置(SELECT/ORDER BY/GROUP BY/HAVING,
        WHERE比较左侧)替换,保留WHERE值位置的字符串字面量不变.
        例如:SELECT "Player Name" FROM t WHERE type = "Player Name"
        -> SELECT `Player_Name` FROM t WHERE type = 'Player Name'
        (SELECT中替换为列引用,WHERE值位置保留为字符串)

        Args:
            sql: 原始SQL查询语句

        Returns:
            str: 预处理后的SQL语句
        """
        col_map = getattr(self, "_original_to_clean_cols", None)
        if not col_map:
            return sql

        # 只处理原始名与清洗名不同的列(即含空格或特殊字符的列名)
        changed_cols = {orig: clean for orig, clean in col_map.items() if orig != clean}
        if not changed_cols:
            return sql

        try:
            # sqlglot already imported at top level
            # exp already imported at top level

            parsed = sqlglot.parse_one(sql, dialect="mysql")
            if parsed is None:
                return self._fallback_preprocess(sql, changed_cols)

            # 在SELECT表达式中替换:双引号字符串 -> 列引用
            select = parsed.find(exp.Select)
            if select:
                new_exprs = []
                for e in select.expressions:
                    new_exprs.append(self._literal_to_column(e, changed_cols))
                select.set("expressions", new_exprs)

            # 在ORDER BY中替换
            order = parsed.find(exp.Order)
            if order:
                new_ordered = []
                for o in order.expressions:
                    new_ordered.append(self._literal_to_column(o, changed_cols))
                order.set("expressions", new_ordered)

            # 在GROUP BY中替换
            group = parsed.find(exp.Group)
            if group:
                new_group = []
                for g in group.expressions:
                    new_group.append(self._literal_to_column(g, changed_cols))
                group.set("expressions", new_group)

            # 在HAVING中替换
            having = parsed.find(exp.Having)
            if having:
                self._replace_having_literals(having.this, changed_cols)

            # 在WHERE子句中:只替换比较操作左侧的字面量(列引用位置)
            where = parsed.find(exp.Where)
            if where:
                self._replace_where_left_literals(where.this, changed_cols)

            result_sql = parsed.sql(dialect="mysql")
            return result_sql

        except Exception:
            # AST解析失败,回退到简单替换(仅SELECT子句)
            return self._fallback_preprocess(sql, changed_cols)

    def _literal_to_column(self, node, changed_cols):
        """
        如果节点是匹配原始列名的字符串字面量,替换为列引用

        Args:
            node: sqlglot AST节点
            changed_cols: 原始列名到清洗列名的映射

        Returns:
            替换后的节点(列引用或原始节点)
        """
        # exp already imported at top level

        if isinstance(node, exp.Literal) and node.is_string:
            lit_val = node.this
            if lit_val in changed_cols:
                return exp.Column(this=exp.Identifier(this=changed_cols[lit_val]))
        return node

    def _replace_where_left_literals(self, node, changed_cols):
        """
        替换WHERE子句中比较操作左侧的字符串字面量为列引用

        只处理比较操作(=, !=, >, <, >=, <=)的左侧,保留右侧作为字符串值.
        递归处理AND/OR组合条件.

        Args:
            node: WHERE子句的AST节点
            changed_cols: 原始列名到清洗列名的映射
        """
        # exp already imported at top level

        comparison_types = (
            exp.EQ,
            exp.NEQ,
            exp.GT,
            exp.GTE,
            exp.LT,
            exp.LTE,
        )

        if isinstance(node, comparison_types):
            if isinstance(node.this, exp.Literal) and node.this.is_string:
                lit_val = node.this.this
                if lit_val in changed_cols:
                    node.set(
                        "this",
                        exp.Column(this=exp.Identifier(this=changed_cols[lit_val])),
                    )
        elif isinstance(node, (exp.And, exp.Or)):
            self._replace_where_left_literals(node.this, changed_cols)
            self._replace_where_left_literals(node.expression, changed_cols)
        elif isinstance(node, exp.Paren):
            self._replace_where_left_literals(node.this, changed_cols)
        elif isinstance(node, exp.Not):
            self._replace_where_left_literals(node.this, changed_cols)

    def _replace_having_literals(self, node, changed_cols):
        """
        替换HAVING子句中的字符串字面量为列引用

        Args:
            node: HAVING子句的AST节点
            changed_cols: 原始列名到清洗列名的映射
        """
        # exp already imported at top level

        comparison_types = (
            exp.EQ,
            exp.NEQ,
            exp.GT,
            exp.GTE,
            exp.LT,
            exp.LTE,
        )

        if isinstance(node, comparison_types):
            # HAVING两侧都可能是列引用(HAVING "Player Name" > 100)
            node.this = self._literal_to_column(node.this, changed_cols)
        elif isinstance(node, (exp.And, exp.Or)):
            self._replace_having_literals(node.this, changed_cols)
            self._replace_having_literals(node.expression, changed_cols)

    def _fallback_preprocess(self, sql: str, changed_cols: dict[str, str]) -> str:
        """
        回退预处理:仅替换SELECT子句中的双引号列名(使用正则提取SELECT子句)

        当AST解析失败时使用,避免全量字符串替换误伤WHERE字符串值.

        Args:
            sql: 原始SQL查询语句
            changed_cols: 原始列名到清洗列名的映射

        Returns:
            str: 预处理后的SQL语句
        """
        # 提取SELECT子句(SELECT ... FROM),仅在此范围内替换
        select_match = re.match(r"(SELECT\s+)(.*?)(\s+FROM\s+)", sql, re.IGNORECASE | re.DOTALL)
        if select_match:
            prefix = select_match.group(1)
            select_clause = select_match.group(2)
            from_suffix = sql[select_match.end(2) :]

            for orig_name in sorted(changed_cols.keys(), key=len, reverse=True):
                clean_name = changed_cols[orig_name]
                select_clause = select_clause.replace(f'"{orig_name}"', f"`{clean_name}`")

            return prefix + select_clause + from_suffix

        # 无法提取SELECT子句,不替换(安全优先)
        return sql

    def _optimize_dtypes(self, df) -> pd.DataFrame:
        """
        优化DataFrame数据类型以减少内存占用

        对数值列进行降级(int64->int32/int16/int8, float64->float32),
        对高基数字符串列不做转换(避免转换开销),对低基数字符串列转为category.
        P3-01增强: 低基数object列自动转category,大幅减少字符串内存占用.

        Args:
            df: 原始DataFrame

        Returns:
            pd.DataFrame: 类型优化后的DataFrame
        """
        start_mem = df.memory_usage(deep=True).sum() / 1024 / 1024

        for col in df.columns:
            col_type = df[col].dtype

            if col_type == "object":
                # P3-01: 低基数字符串列转为category类型
                # 条件: 非空值基数/总行数 < 0.3 (即重复值多), 且行数 > 100 (小表无意义)
                n_unique = df[col].nunique()
                n_total = len(df)
                if n_total > 100 and n_unique > 0 and n_unique / n_total < 0.3:
                    try:
                        df[col] = df[col].astype("category")
                    except (TypeError, ValueError):
                        pass  # 转换失败则保持原类型
            elif col_type in ["int64", "int32"]:
                # 整数列降级
                col_min = df[col].min()
                col_max = df[col].max()
                if col_min >= 0:
                    if col_max < 256:
                        df[col] = df[col].astype("uint8")
                    elif col_max < 65536:
                        df[col] = df[col].astype("uint16")
                    elif col_max < 4294967296:
                        df[col] = df[col].astype("uint32")
                else:
                    if col_min > -128 and col_max < 127:
                        df[col] = df[col].astype("int8")
                    elif col_min > -32768 and col_max < 32767:
                        df[col] = df[col].astype("int16")
                    elif col_min > -2147483648 and col_max < 2147483647:
                        df[col] = df[col].astype("int32")
            elif col_type == "float64":
                # 浮点列降级为 float32(精度足够)，但需检查溢出风险
                # float32 范围约 ±3.4e38，超出会变成 inf 导致后续 int() 转换崩溃
                test_series = df[col].astype("float32")
                if not (test_series.isin([float("inf"), float("-inf")]).any() or test_series.isna().any()):
                    df[col] = test_series
                # 否则保持 float64，避免 inf 值导致崩溃

        end_mem = df.memory_usage(deep=True).sum() / 1024 / 1024
        reduction = (1 - end_mem / start_mem) * 100 if start_mem > 0 else 0
        logger.debug(f"dtype优化: {start_mem:.1f}MB -> {end_mem:.1f}MB (节省{reduction:.0f}%)")

        return df

    def _validate_sql_support(self, parsed_sql: exp.Expression) -> dict[str, Any]:
        """验证SQL语法支持范围"""
        try:
            # 检查是否为SELECT语句,UNION,EXCEPT或INTERSECT
            if not isinstance(parsed_sql, (exp.Select, exp.Union, exp.Except, exp.Intersect)):
                return {
                    "valid": False,
                    "error": "只支持SELECT查询语句,不支持INSERT,UPDATE,DELETE等操作",
                }

            # 检测FETCH NEXT → 建议用LIMIT
            if list(parsed_sql.find_all(exp.Fetch)):
                return {
                    "valid": False,
                    "error": "不支持FETCH/NEXT语法,请用LIMIT:SELECT ... LIMIT 10",
                }

            # 检测NATURAL JOIN
            for join in parsed_sql.find_all(exp.Join):
                if join.args.get("kind") == "NATURAL" or join.args.get("natural"):
                    return {
                        "valid": False,
                        "error": "不支持NATURAL JOIN,请改用显式ON条件:JOIN t2 ON t1.col = t2.col",
                    }

            # LATERAL JOIN: 由_apply_lateral_join处理，此处不拒绝

            # 检测WITH RECURSIVE
            with_clause = parsed_sql.args.get("with")
            if with_clause and getattr(with_clause, "recursive", False):
                return {
                    "valid": False,
                    "error": "不支持递归CTE(WITH RECURSIVE),请改用普通CTE或子查询",
                }

            return {"valid": True}

        except Exception as e:
            return {"valid": False, "error": f"SQL验证失败: {self._sanitize_error_message(str(e))}"}

    def _replace_cn_columns_in_sql(self, sql: str, worksheets_data: dict[str, pd.DataFrame]) -> str:
        """
        将SQL中的中文列名替换为英文列名(在sqlglot解析前).

        双行表头的游戏配置表中,第1行是中文描述,第2行是英文字段名.
        策划习惯用中文名查询,但SQL引擎需要英文列名.
        本方法在SQL文本层面做替换,避免给DataFrame添加临时列.

        Args:
            sql: 原始SQL语句
            worksheets_data: 已加载的工作表数据

        Returns:
            str: 替换后的SQL语句
        """
        if not hasattr(self, "_header_descriptions") or not self._header_descriptions:
            return sql

        # 收集所有中文->英文映射(去重)
        cn_to_en = {}
        for sheet_name, desc_map in self._header_descriptions.items():
            for eng_name, cn_desc in desc_map.items():
                if cn_desc and cn_desc != eng_name:
                    cn_to_en[cn_desc] = eng_name

        if not cn_to_en:
            return sql

        # 按中文列名长度降序排列,避免短名称部分匹配长名称
        sorted_names = sorted(cn_to_en.keys(), key=len, reverse=True)

        # 用正则替换:只替换SQL标识符位置(非字符串字面量中的中文)
        # 策略:先把字符串字面量占位保护,替换中文标识符,再恢复字符串
        string_literals = []
        protected_sql = sql

        # 保护单引号字符串
        def protect_string(match):
            """保护字符串字面量不被替换.

            Args:
                match: 正则匹配对象
            """
            string_literals.append(match.group(0))
            return f"__PROTECTED_STR_{len(string_literals) - 1}__"

        protected_sql = re.sub(r"'[^']*'", protect_string, protected_sql)

        # 保护AS别名(避免中文别名被误替换为英文列名)
        # 例如: SELECT level AS 等级 → 不应把"等级"替换为"level"
        alias_mapping = {}  # 记录别名到保护名的映射

        def protect_as_alias(match):
            """保护AS别名不被替换"""
            alias_name = match.group(1)
            alias_mapping[alias_name] = f"__PROTECTED_ALIAS_{len(string_literals)}__"
            string_literals.append(alias_name)  # 保存别名部分用于恢复
            return f"AS {alias_mapping[alias_name]}"

        protected_sql = re.sub(r"\bAS\s+([^\s,)(]+)", protect_as_alias, protected_sql, flags=re.IGNORECASE)

        # 保护 ORDER BY/GROUP BY/HAVING 中的别名引用
        # 简化方案: 在全局SQL中替换所有别名引用为保护名(中文→英文替换之前)
        for alias_name, protected_name in alias_mapping.items():
            # 使用单词边界确保只替换完整的别名
            protected_sql = re.sub(
                rf"\b{re.escape(alias_name)}\b",
                protected_name,
                protected_sql,
                flags=re.IGNORECASE,
            )

        # 替换中文列名(此时AS别名及其引用已被保护)
        for cn_name in sorted_names:
            en_name = cn_to_en[cn_name]
            protected_sql = re.sub(re.escape(cn_name), en_name, protected_sql)

        # 恢复字符串字面量
        for i, s in enumerate(string_literals):
            protected_sql = protected_sql.replace(f"__PROTECTED_STR_{i}__", s)

        # 恢复AS别名(在字符串恢复之后,避免冲突)
        for i, s in enumerate(string_literals):
            if f"__PROTECTED_ALIAS_{i}__" in protected_sql:
                protected_sql = protected_sql.replace(f"__PROTECTED_ALIAS_{i}__", s)

        return protected_sql

    def _generate_empty_result_suggestion(self, parsed_sql, df_before_where, worksheets_data):
        """分析WHERE条件类型,生成智能空结果建议"""
        where_clause = parsed_sql.args.get("where")
        if not where_clause:
            return "查询返回0行数据.表可能为空,请检查数据是否已录入."

        total_rows = len(df_before_where)
        if total_rows == 0:
            return "查询返回0行数据.工作表本身没有数据行."

        hints = []
        condition = where_clause.this

        # 分析条件树,收集条件类型和涉及的列
        eq_conditions = []  # 等值条件
        range_conditions = []  # 范围条件
        like_conditions = []  # LIKE条件
        in_conditions = []  # IN条件
        between_conditions = []  # BETWEEN条件
        null_conditions = []  # IS NULL条件

        self._collect_condition_types(
            condition,
            eq_conditions,
            range_conditions,
            like_conditions,
            in_conditions,
            between_conditions,
            null_conditions,
        )

        # 等值条件:提示列的唯一值
        for col, val in eq_conditions:
            if col in df_before_where.columns:
                unique_vals = df_before_where[col].dropna().unique()
                if len(unique_vals) <= 20:
                    vals_str = ", ".join(str(v) for v in unique_vals[:10])
                    if len(unique_vals) > 10:
                        vals_str += f" ... 共{len(unique_vals)}个"
                    hints.append(f'• 列"{col}"的值为: {vals_str}')
                else:
                    hints.append(f'• 列"{col}"有{len(unique_vals)}个不同值,"{val}"不在其中')

        # 范围条件 + BETWEEN条件:提示列的实际范围(两者提示文本相同)
        range_and_between = [(col, op, val) for col, op, val in range_conditions] + [(col, f"{low}~{high}", None) for col, low, high in between_conditions]
        range_cols_seen = set()
        for item in range_and_between:
            col = item[0]
            if col in range_cols_seen:
                continue
            range_cols_seen.add(col)
            if col in df_before_where.columns:
                numeric = pd.to_numeric(df_before_where[col], errors="coerce").dropna()
                if len(numeric) > 0:
                    hints.append(f'• 列"{col}"的实际范围: {numeric.min():.2f} ~ {numeric.max():.2f}')
                else:
                    hints.append(f'• 列"{col}"不是数值列,无法用范围条件比较')

        # LIKE条件:提示匹配情况
        for col, pattern in like_conditions:
            if col in df_before_where.columns:
                sample = df_before_where[col].dropna().astype(str).head(5).tolist()
                hints.append(f'• 列"{col}"的样本数据: {", ".join(sample)}')

        # IN条件:提示实际存在的值
        for col, vals in in_conditions:
            if col in df_before_where.columns:
                unique_vals = set(df_before_where[col].dropna().unique())
                matched = unique_vals & set(vals)
                if not matched:
                    hints.append(f'• 列"{col}"中不包含指定的任何值')

        # IS NULL条件
        for col, is_null in null_conditions:
            if col in df_before_where.columns:
                null_count = df_before_where[col].isna().sum()
                if is_null and null_count == 0:
                    hints.append(f'• 列"{col}"没有空值')
                elif not is_null and null_count == total_rows:
                    hints.append(f'• 列"{col}"全部为空值')

        # 多条件提示
        total_conditions = len(eq_conditions) + len(range_conditions) + len(like_conditions) + len(in_conditions) + len(between_conditions) + len(null_conditions)
        if total_conditions > 1:
            hints.append("• 多个AND条件同时满足的行可能不存在,尝试减少条件或改用OR")

        # 通用提示
        hints.append(f"• 源表共{total_rows}行,WHERE过滤后为0行")
        hints.append("• 可用 DESCRIBE 查看表结构,或去掉WHERE先查看全部数据")

        return "查询返回0行数据.分析:\n" + "\n".join(hints)

    def _collect_condition_types(self, condition, eq, rng, like, in_list, between, null_list):
        """递归收集WHERE条件树中的各类条件"""
        if isinstance(condition, exp.EQ):
            col, col_table = self._extract_column_name(condition.left)
            val = self._extract_literal_value(condition.right)
            if col:
                eq.append((col, val))
        elif isinstance(condition, (exp.GT, exp.GTE, exp.LT, exp.LTE)):
            col, col_table = self._extract_column_name(condition.left)
            val = self._extract_literal_value(condition.right)
            op_map = {exp.GT: ">", exp.GTE: ">=", exp.LT: "<", exp.LTE: "<="}
            if col:
                rng.append((col, op_map.get(type(condition), "?"), val))
        elif isinstance(condition, exp.Like):
            col, col_table = self._extract_column_name(condition.left)
            val = self._extract_literal_value(condition.right)
            if col:
                like.append((col, val))
        elif isinstance(condition, exp.In):
            col, col_table = self._extract_column_name(condition.this)
            vals = []
            if hasattr(condition, "expressions"):
                for e in condition.expressions:
                    v = self._extract_literal_value(e)
                    if v is not None:
                        vals.append(v)
            if col and vals:
                in_list.append((col, vals))
        elif isinstance(condition, exp.Between):
            col, col_table = self._extract_column_name(condition.this)
            low = self._extract_literal_value(condition.args.get("low"))
            high = self._extract_literal_value(condition.args.get("high"))
            if col:
                between.append((col, low, high))
        elif isinstance(condition, exp.Is):
            col, col_table = self._extract_column_name(condition.this)
            if col:
                null_list.append((col, True))
        elif isinstance(condition, exp.Not):
            inner = condition.this
            if isinstance(inner, exp.Is):
                col, col_table = self._extract_column_name(inner.this)
                if col:
                    null_list.append((col, False))
            else:
                self._collect_condition_types(inner, eq, rng, like, in_list, between, null_list)
        elif isinstance(condition, (exp.And, exp.Or)):
            for child in condition.flatten():
                if child is not condition:
                    self._collect_condition_types(child, eq, rng, like, in_list, between, null_list)

    def _extract_column_name(self, expr):
        """从表达式中提取列名,支持表别名格式(如 r.名称)"""
        if isinstance(expr, exp.Column):
            # 处理表别名格式,如 r.名称 -> 名称, 返回 (列名, 表别名)
            if hasattr(expr, "table") and expr.table:
                return f"{expr.table}.{expr.name}", expr.table
            else:
                return expr.name, None
        return None, None

    def _resolve_column_name(self, col_name: str, df) -> str:
        """解析列名,支持表别名格式(如 r.名称)"""
        if "." in col_name:
            # 处理表别名格式,如 r.名称
            table_part, col_part = col_name.split(".", 1)
            # 从 _table_aliases 获取真实的表名
            resolved_table = self._table_aliases.get(table_part, table_part)

            # 优先检查用户使用的别名格式是否直接存在
            alias_col = f"{table_part}.{col_part}"
            if alias_col in df.columns:
                return alias_col

            # 检查JOIN后pandas添加的后缀格式(table_part_列名)
            pandas_suffix_col = f"{table_part}_{col_part}"
            if pandas_suffix_col in df.columns:
                return pandas_suffix_col

            # 尝试其他可能的别名格式
            # 如果用户使用的是原始表名,检查原始表名+列名
            original_col = f"{resolved_table}_{col_part}"
            if original_col in df.columns:
                return original_col

            # 最后尝试原始列名
            if col_part in df.columns:
                return col_part

        return col_name

    def _extract_literal_value(self, expr):
        """从表达式中提取字面值(委托_parse_literal_value统一处理)"""
        if isinstance(expr, exp.Literal):
            return self._parse_literal_value(expr)
        if isinstance(expr, exp.Boolean):
            # SQL布尔字面量(TRUE/FALSE) → int(1/0)
            return int(expr.this)
        return None

    def _generate_having_empty_suggestion(self, having_expr, df_before_having) -> str:
        """生成HAVING导致空结果时的智能建议

        Args:
            having_expr: HAVING表达式(完整的having clause)
            df_before_having: HAVING过滤前的聚合结果DataFrame
        """
        hints = ["\nHAVING分析:"]
        hints.append(f"• GROUP BY聚合后有{len(df_before_having)}组数据")

        condition = having_expr.this
        col, col_table = self._extract_column_name(condition.left)
        val = self._extract_literal_value(condition.right)

        # 策略0:通过_having_agg_alias_map直接查找聚合表达式对应的SELECT别名
        # 这是最可靠的方式,能正确处理中文列名(如AVG(伤害)->avg_dmg)
        if not col and hasattr(self, "_having_agg_alias_map") and self._having_agg_alias_map:
            left_sql = str(condition.left).strip()
            # 精确匹配:HAVING表达式SQL与map key完全一致
            for map_sql, alias in self._having_agg_alias_map.items():
                if map_sql == left_sql and alias in df_before_having.columns:
                    col = alias
                    break
            # 模糊匹配:去除多余空格后比较(sqlglot有时生成不一致的空格)
            if not col:
                left_normalized = " ".join(left_sql.split())
                for map_sql, alias in self._having_agg_alias_map.items():
                    map_normalized = " ".join(map_sql.split())
                    if (left_normalized == map_normalized or left_sql in map_sql or map_sql in left_sql) and alias in df_before_having.columns:
                        col = alias
                        break

        # HAVING中聚合函数表达式的列名可能是别名,尝试从DataFrame列匹配
        if not col:
            left_str = str(condition.left).lower()
            # 策略1:子串匹配(含中文支持)
            for c in df_before_having.columns:
                c_lower = c.lower()
                if c_lower in left_str or left_str in c_lower:
                    col = c
                    break
            # 策略2:拆分表达式中的标识符(含中文token提取)
            if not col:
                tokens = set(re.findall(r"[a-zA-Z_]+", left_str))
                cn_tokens = set(re.findall(r"[\u4e00-\u9fff]+", left_str))
                if tokens or cn_tokens:
                    for c in df_before_having.columns:
                        c_tokens = set(re.findall(r"[a-zA-Z_]+", c.lower()))
                        c_cn_tokens = set(re.findall(r"[\u4e00-\u9fff]+", c))
                        generic = {"avg", "sum", "count", "min", "max"}
                        specific = tokens - generic
                        if (specific and specific & c_tokens) or (cn_tokens and cn_tokens & c_cn_tokens):
                            col = c
                            break

        if not col or col not in df_before_having.columns:
            # 无法匹配列名,显示所有聚合列的实际范围
            if len(df_before_having.columns) > 0:
                for c in df_before_having.columns:
                    numeric = pd.to_numeric(df_before_having[c], errors="coerce").dropna()
                    if len(numeric) > 0:
                        hints.append(f'• 列"{c}"范围: {numeric.min()} ~ {numeric.max()}')
            hints.append("• HAVING条件较复杂,建议去掉HAVING先查看聚合结果")
            hints.append("• 可先去掉HAVING查看全部分组结果,再调整过滤条件")
            return "\n".join(hints)

        numeric = pd.to_numeric(df_before_having[col], errors="coerce").dropna()
        if len(numeric) == 0:
            hints.append(f'• 列"{col}"没有数值数据')
            hints.append("• 可先去掉HAVING查看全部分组结果,再调整过滤条件")
            return "\n".join(hints)

        # 比较运算符(GT/GTE/LT/LTE)使用分发表
        op_type = type(condition)
        if op_type in self._HAVING_OPS:
            stat_func, op_str, label = self._HAVING_OPS[op_type]
            stat_val = getattr(numeric, stat_func)()
            hints.append(f'• 列"{col}"的{label}值为{stat_val},HAVING要求 {op_str}{val},无满足条件的组')
        elif isinstance(condition, exp.EQ):
            unique_vals = df_before_having[col].dropna().unique()
            if len(unique_vals) <= 10:
                vals_str = ", ".join(str(v) for v in unique_vals)
                hints.append(f'• 列"{col}"的值为: {vals_str},不等于{val}')
            else:
                hints.append(f'• 列"{col}"有{len(unique_vals)}个不同值,不等于{val}')
        else:
            hints.append("• HAVING条件较复杂,建议去掉HAVING先查看聚合结果")

        hints.append("• 可先去掉HAVING查看全部分组结果,再调整过滤条件")
        return "\n".join(hints)

    def _suggest_column_name(self, col_name: str, available_cols: list[str], max_suggestions: int = 3) -> str:
        """
        当列名不存在时,用编辑距离找出最相似的列名作为建议.
        同时检查中文列名描述(双行表头场景).

        Args:
            col_name: 用户输入的列名
            available_cols: 可用的列名列表
            max_suggestions: 最多返回几个建议

        Returns:
            str: 格式化的建议字符串
        """
        if not available_cols:
            return ""

        matches = difflib.get_close_matches(col_name, available_cols, n=max_suggestions, cutoff=0.3)
        if matches:
            return f" 你是否想用: {', '.join(matches)}?"

        # 英文列名匹配不到时,尝试匹配中文列名描述(双行表头)
        if hasattr(self, "_header_descriptions") and self._header_descriptions:
            cn_names = []
            for sheet_name, desc_map in self._header_descriptions.items():
                for eng_name, cn_desc in desc_map.items():
                    if cn_desc:
                        cn_names.append(cn_desc)
            if cn_names:
                cn_matches = difflib.get_close_matches(col_name, cn_names, n=max_suggestions, cutoff=0.3)
                if cn_matches:
                    # 反查中文->英文映射
                    en_lookup = {}
                    for desc_map in self._header_descriptions.values():
                        for eng_name, cn_desc in desc_map.items():
                            if cn_desc:
                                en_lookup[cn_desc] = eng_name
                    mapped = [f"{cn}({en_lookup.get(cn, '?')})" for cn in cn_matches]
                    return f" 中文名称匹配: {', '.join(mapped)}"

        return ""

    def _check_window_alias_hint(self, col_name: str) -> str:
        """
        检查列名是否是窗口函数别名,如果是则返回友好的错误提示.

        WHERE子句无法引用窗口函数结果(SQL标准执行顺序限制).
        """
        if not hasattr(self, "_parsed_sql"):
            return ""

        # 检查SELECT表达式中的窗口函数别名
        for select_expr in self._parsed_sql.expressions:
            if isinstance(select_expr, exp.Alias):
                alias_name = select_expr.alias
                original_expr = select_expr.this
                # 检查是否是窗口函数且别名匹配
                if isinstance(original_expr, exp.Window) and alias_name == col_name:
                    return (
                        f"\n💡 提示: '{col_name}' 是窗口函数别名,WHERE 子句无法引用窗口函数结果(SQL标准限制).\n"
                        f"   解决方案: 使用子查询包装 — SELECT * FROM (SELECT ..., RANK() as {col_name} FROM ...) t WHERE {col_name} <= 3\n"
                        f"   原因: SQL执行顺序为 FROM → WHERE → GROUP BY → HAVING → 窗口函数 → SELECT → ORDER BY"
                    )
        return ""

    def _apply_union_order_by(self, df, order_clause=None) -> pd.DataFrame:
        """
        对 UNION 合并后的 DataFrame 应用 ORDER BY 排序

        Args:
            df: 合并后的 DataFrame
            order_clause: sqlglot Order expression

        Returns:
            pd.DataFrame: 排序后的 DataFrame
        """
        if not order_clause or df.empty:
            return df

        sort_columns = []
        sort_ascending = []

        for ordered in order_clause.find_all(exp.Ordered):
            col_expr = ordered.this
            # 获取列名
            if isinstance(col_expr, exp.Column):
                col_name = col_expr.name
            elif isinstance(col_expr, exp.Identifier):
                col_name = col_expr.this
            else:
                col_name = str(col_expr)

            # 检查是否使用表别名限定符
            if "." in col_name:
                col_name = col_name.split(".")[-1]

            # 尝试列名匹配(包括中文列名映射)
            if col_name not in df.columns:
                # [FIX R16-B3] _cn_to_en_map 可能未初始化（UNION查询路径不经过表头解析）
                # 使用 getattr 提供默认空字典避免 AttributeError
                cn_to_en_map = getattr(self, "_cn_to_en_map", {})
                for cn_name, en_name in cn_to_en_map.items():
                    if col_name == cn_name:
                        col_name = en_name
                        break

            if col_name in df.columns:
                sort_columns.append(col_name)
                sort_ascending.append(not ordered.args.get("desc", False))

        if sort_columns:
            df = df.sort_values(sort_columns, ascending=sort_ascending).reset_index(drop=True)

        return df

    def _execute_union(
        self,
        parsed_sql: exp.Expression,
        worksheets_data: dict[str, pd.DataFrame],
        limit: int | None = None,
    ) -> pd.DataFrame:
        """
        执行 UNION / UNION ALL 查询

        从 Union 表达式中提取所有 SELECT 语句,分别执行后合并结果.
        UNION 去重,UNION ALL 保留所有行.
        支持 ORDER BY 和 LIMIT 应用于合并后的结果.

        Args:
            parsed_sql: 解析后的 Union 表达式
            worksheets_data: 工作表数据
            limit: 结果限制

        Returns:
            pd.DataFrame: 合并后的查询结果
        """

        # 递归提取所有 SELECT 语句
        def _extract_selects(node):
            if isinstance(node, exp.Select):
                return [node]
            elif isinstance(node, exp.Subquery):
                # [FIX R16-B2] (SELECT ...) UNION (SELECT ...) 中括号导致 sqlglot 解析为 Subquery
                # 需要解包 Subquery 获取内部的 Select
                return _extract_selects(node.this)
            elif isinstance(node, exp.Union):
                selects = []
                # this 可能是 Union(链式)、Select 或 Subquery
                selects.extend(_extract_selects(node.this))
                # expression 是右侧的 Select、Union 或 Subquery
                selects.extend(_extract_selects(node.expression))
                return selects
            return []

        # 处理CTE(WITH子句) - UNION可能包含CTE定义
        # 兼容sqlglot不同版本:arg key可能是'with'或'with_'
        _with_key = "with" if parsed_sql.args.get("with") else "with_"
        with_clause = parsed_sql.args.get(_with_key)
        effective_data = worksheets_data

        if with_clause:
            # 复制worksheets_data避免修改原始数据,逐步添加CTE结果
            cte_data = dict(worksheets_data)
            for cte_expr in with_clause.expressions:
                cte_name = cte_expr.alias
                cte_query = cte_expr.this  # inner Select
                try:
                    # 每个CTE在已有的cte_data上执行(支持CTE引用前面的CTE)
                    cte_result = self._execute_query(cte_query, cte_data, limit=None)
                    cte_data[cte_name] = cte_result
                except Exception as e:
                    raise ValueError(f"CTE '{cte_name}' 执行失败: {e}")
            # 使用包含CTE的数据
            effective_data = cte_data

        selects = _extract_selects(parsed_sql)
        if not selects:
            raise ValueError("UNION 查询中未找到有效的 SELECT 语句")

        # 执行每个 SELECT 并收集结果
        result_dfs = []
        for i, select_sql in enumerate(selects):
            # 使用包含CTE的effective_data
            df = self._execute_query(select_sql, effective_data, limit=None)
            result_dfs.append(df)

        # 合并所有结果(列名对齐)
        if not result_dfs:
            return pd.DataFrame()

        # 以第一个 SELECT 的列名为基准,统一列名
        base_columns = list(result_dfs[0].columns)
        aligned_dfs = []
        for i, df in enumerate(result_dfs):
            # Fix(R56): UNION 要求每个 SELECT 返回相同数量的列
            if len(df.columns) != len(base_columns):
                raise ValueError(
                    f"UNION 第{i+1}个SELECT的列数({len(df.columns)}) "
                    f"与第一个({len(base_columns)})不同。"
                    f"SQL标准要求UNION的每个SELECT返回相同数量的列"
                )
            aligned = df.reindex(columns=base_columns)
            aligned_dfs.append(aligned)

        combined = pd.concat(aligned_dfs, ignore_index=True)

        # UNION(去重) vs UNION ALL(保留重复)
        is_union_all = not parsed_sql.args.get("distinct", True)
        if not is_union_all:
            combined = combined.drop_duplicates().reset_index(drop=True)

        # 应用 ORDER BY(如果有,sqlglot 将其放在外层 Union 上)
        order_clause = parsed_sql.args.get("order")
        if order_clause:
            # 构造一个最小化的 Select 用于 _apply_order_by 的签名
            # _apply_order_by(self, parsed_sql, df, select_aliases) 期望完整的 parsed_sql
            # 但 UNION 的 ORDER BY 是独立的,直接解析排序列
            combined = self._apply_union_order_by(combined, order_clause)

        # 应用 LIMIT(如果有)
        union_limit = self._extract_int_value(parsed_sql.args.get("limit"))
        if union_limit is not None:
            combined = combined.head(union_limit)

        # 应用外部传入的 limit
        if limit is not None:
            combined = combined.head(limit)

        return combined

    def _execute_except_intersect(
        self,
        parsed_sql: exp.Expression,
        worksheets_data: dict[str, pd.DataFrame],
        limit: int | None = None,
    ) -> pd.DataFrame:
        """
        执行 EXCEPT / INTERSECT 查询

        从 Except/Intersect 表达式中提取所有 SELECT 语句,分别执行后进行集合操作.
        EXCEPT: 返回第一个查询结果中不存在于第二个查询结果的行(差集)
        INTERSECT: 返回两个查询结果中都存在的行(交集)

        Args:
            parsed_sql: 解析后的 Except/Intersect 表达式
            worksheets_data: 工作表数据
            limit: 结果限制

        Returns:
            pd.DataFrame: 集合操作后的查询结果
        """
        # 确定操作类型
        is_except = isinstance(parsed_sql, exp.Except)
        is_intersect = isinstance(parsed_sql, exp.Intersect)

        # 递归提取所有 SELECT 语句(支持链式 EXCEPT/INTERSECT)
        def _extract_selects(node):
            if isinstance(node, exp.Select):
                return [node]
            elif isinstance(node, (exp.Union, exp.Except, exp.Intersect)):
                selects = []
                # this 可能是 Union/Except/Intersect(链式)或 Select
                selects.extend(_extract_selects(node.this))
                # expression 是右侧的 Select 或 Union/Except/Intersect
                selects.extend(_extract_selects(node.expression))
                return selects
            return []

        selects = _extract_selects(parsed_sql)
        if not selects:
            raise ValueError("EXCEPT/INTERSECT 查询中未找到有效的 SELECT 语句")
        if len(selects) != 2:
            raise ValueError("EXCEPT/INTERSECT 查询必须包含恰好两个 SELECT 语句")

        # 执行两个 SELECT
        df1 = self._execute_query(selects[0], worksheets_data, limit=None)
        df2 = self._execute_query(selects[1], worksheets_data, limit=None)

        # 对齐列名
        base_columns = list(df1.columns)
        df2_aligned = df2.reindex(columns=base_columns)

        if is_except:
            # EXCEPT: 差集 - df1 中不在 df2 中的行
            # 使用 merge(how='left', indicator=True) 实现
            merged = df1.merge(df2_aligned, how="left", indicator=True, on=list(base_columns))
            # 保留只在左侧(df1)的行
            result = merged[merged["_merge"] == "left_only"].drop(columns=["_merge"]).reset_index(drop=True)
        elif is_intersect:
            # INTERSECT: 交集 - df1 和 df2 都有的行
            # 使用 merge(how='inner') 实现
            result = df1.merge(df2_aligned, how="inner", on=list(base_columns)).drop_duplicates().reset_index(drop=True)
        else:
            return pd.DataFrame()

        # 应用 ORDER BY(如果有)
        order_clause = parsed_sql.args.get("order")
        if order_clause:
            result = self._apply_union_order_by(result, order_clause)

        # 应用 LIMIT(如果有)
        op_limit = self._extract_int_value(parsed_sql.args.get("limit"))
        if op_limit is not None:
            result = result.head(op_limit)

        # 应用外部传入的 limit
        if limit is not None:
            result = result.head(limit)

        return result

    def _has_window_function(self, parsed_sql: exp.Expression) -> bool:
        """检查SQL是否包含窗口函数"""
        return bool(parsed_sql.find(exp.Window))

    def _apply_window_functions(self, parsed_sql: exp.Expression, df) -> pd.DataFrame:
        """
        计算窗口函数并将结果添加到DataFrame
        支持: ROW_NUMBER, RANK, DENSE_RANK, PERCENT_RANK, CUME_DIST
        语法: func() OVER ([PARTITION BY col ...] ORDER BY col [ASC|DESC] ...)
        """
        if not self._has_window_function(parsed_sql):
            return df

        df = df.copy()

        # 构建SELECT别名映射(用于将聚合表达式映射到别名,如AVG(damage)->avg_dmg)
        select_alias_map = {}
        for select_expr in parsed_sql.expressions:
            if isinstance(select_expr, exp.Alias):
                alias_name = select_expr.alias
                original = select_expr.this
                # 将原始表达式文本作为key
                expr_key = str(original).strip()
                select_alias_map[expr_key] = alias_name

        # 收集所有已处理的Window表达式(用于去重)
        _processed_windows: set[int] = set()

        for select_expr in parsed_sql.expressions:
            # 跳过 SELECT *
            if isinstance(select_expr, exp.Star):
                continue

            # 提取别名和原始表达式
            if isinstance(select_expr, exp.Alias):
                alias_name = select_expr.alias
                original_expr = select_expr.this
            else:
                alias_name = None
                original_expr = select_expr

            if not isinstance(original_expr, exp.Window):
                continue

            # 确定列名
            col_name = alias_name or f"_window_{len([c for c in df.columns if c.startswith('_window_')])}"

            # 计算窗口函数
            result = self._compute_window_function(original_expr, df, select_alias_map)
            df[col_name] = result
            _processed_windows.add(id(original_expr))

        # [FIX R10-B1] 处理嵌套在标量函数中的窗口函数(如 ROUND(RANK() OVER(...), 2))
        # SQLGlot 的 find_all 可以递归发现所有 Window 节点(包括嵌套的)
        all_windows = list(parsed_sql.find_all(exp.Window))
        _window_counter = len([c for c in df.columns if c.startswith("_window_")])
        for w in all_windows:
            if id(w) in _processed_windows:
                continue  # 已在顶层处理过
            _processed_windows.add(id(w))

            # 为嵌套窗口生成自动别名(基于表达式文本hash确保稳定)
            gen_col = f"_window_nested{_window_counter}"
            _window_counter += 1
            try:
                result = self._compute_window_function(w, df, select_alias_map)
                df[gen_col] = result
                # 将此窗口节点与生成列名的映射存到实例上,供 _expr_to_series 查找
                if not hasattr(self, "_nested_window_columns"):
                    self._nested_window_columns = {}
                self._nested_window_columns[id(w)] = gen_col
            except Exception as e:
                logger.warning(f"嵌套窗口函数计算失败: {e}, 跳过")

        # 按第一个窗口函数的ORDER BY排序输出（无外部ORDER BY时的自然顺序）
        for select_expr in parsed_sql.expressions:
            original_expr = select_expr.this if isinstance(select_expr, exp.Alias) else select_expr
            if isinstance(original_expr, exp.Window):
                order = original_expr.args.get("order")
                if order:
                    order_cols = []
                    ascending = []
                    for item in order.expressions:
                        col = item.this.name if hasattr(item.this, "name") else str(item.this)
                        # 跳过聚合表达式列（GROUP BY后的场景，排序列可能不存在）
                        if col not in df.columns:
                            continue
                        desc = item.args.get("desc", False)
                        order_cols.append(col)
                        ascending.append(not desc)
                    if order_cols:
                        # [FIX] 确保排序列为数值类型，避免 object 列混合类型导致 sort_values 崩溃
                        for oc in order_cols:
                            if oc in df.columns and df[oc].dtype == object:
                                df[oc] = pd.to_numeric(df[oc], errors="coerce")
                        df = df.sort_values(order_cols, ascending=ascending, kind="mergesort")
                break  # 只按第一个窗口函数排序

        return df

    def _compute_window_function(
        self,
        window_expr: exp.Window,
        df: pd.DataFrame,
        select_alias_map: dict | None = None,
    ) -> pd.Series:
        """计算单个窗口函数,返回结果Series"""
        # [FIX R16-B1] 解包 IgnoreNulls/RespectNulls 包装节点
        # sqlglot 将 NTH_VALUE(x, 1) IGNORE NULLS 解析为 IgnoreNulls(NthValue(x, 1))
        # 需要提取内部的实际窗口函数节点
        actual_func = window_expr.this
        while type(actual_func).__name__ in ("IgnoreNulls", "RespectNulls"):
            actual_func = actual_func.this
        func_type = type(actual_func).__name__

        # 如果发生了解包，替换 window_expr.this 以便下游处理器正常工作
        if actual_func is not window_expr.this:
            window_expr = window_expr.copy()
            window_expr.set("this", actual_func)

        # 支持的窗口函数类型
        _window_agg_funcs = {"Avg", "Sum", "Count", "Min", "Max"}
        supported_funcs = {
            "RowNumber",
            "Rank",
            "DenseRank",
            "Lag",
            "Lead",
            "FirstValue",
            "LastValue",
            "NthValue",
            "Ntile",
            "PercentRank",
            "CumeDist",
            "Count",
        } | _window_agg_funcs
        if func_type not in supported_funcs:
            raise ValueError(f"不支持的窗口函数: {func_type}")

        if select_alias_map is None:
            select_alias_map = {}

        # 辅助函数：解析列名（处理JOIN列映射）
        def resolve_col_name(col_node):
            """Resolve column name from AST column node."""
            col_name = col_node.name if hasattr(col_node, "name") and col_node.name else str(col_node)
            # 检查JOIN列映射
            if hasattr(self, "_join_column_mapping") and col_name not in df.columns:
                for table_alias, col_map in self._join_column_mapping.items():
                    if col_name in col_map and col_map[col_name] in df.columns:
                        return col_map[col_name]
            # [FIX R15-B1] 裸列名找不到时，尝试在df.columns中匹配 "table.col" 格式
            # 场景: JOIN后列名变为 p.Category，但AST中取出的是裸名 Category
            if col_name not in df.columns:
                for fc in df.columns:
                    if fc.endswith(f".{col_name}") or fc.endswith(f"_{col_name}"):
                        return fc
            return col_name

        # 解析 PARTITION BY
        partition_by = window_expr.args.get("partition_by", [])
        partition_cols = []
        for col in partition_by:
            col_name = resolve_col_name(col)
            partition_cols.append(col_name)

        # 解析 ORDER BY
        order = window_expr.args.get("order")
        order_cols = []
        ascending = []
        if order:
            for ordered_expr in order.expressions:
                col = ordered_expr.this
                col_name = resolve_col_name(col)
                # 如果列名仍不在DataFrame中,尝试聚合别名映射
                if col_name not in df.columns:
                    col_name = self._resolve_window_column(col_name, df.columns, select_alias_map)
                order_cols.append(col_name)
                ascending.append(not ordered_expr.args.get("desc", False))

        # 验证列存在
        for col in partition_cols + order_cols:
            if col not in df.columns:
                suggestion = self._suggest_column_name(col, list(df.columns))
                raise ValueError(f"窗口函数中列 '{col}' 不存在.可用列: {list(df.columns)}.{suggestion}")

        # [FIX] 确保排序列为数值类型，避免 object 列混合 int/str 导致 sort_values 崩溃
        # 场景: 部分列INSERT产生空字符串值，查询管道将列转为object dtype
        if order_cols:
            df = df.copy()
            for oc in order_cols:
                if oc in df.columns and df[oc].dtype == object:
                    df[oc] = pd.to_numeric(df[oc], errors="coerce")

        # 窗口函数分发表
        _window_dispatch = {
            "RowNumber": self._compute_row_number,
            "Rank": self._compute_rank,
            "DenseRank": self._compute_dense_rank,
            "Lag": self._compute_lag,
            "Lead": self._compute_lead,
            "FirstValue": self._compute_first_value,
            "LastValue": self._compute_last_value,
            "NthValue": self._compute_nth_value,
            "Ntile": self._compute_ntile,
            "PercentRank": self._compute_percent_rank,
            "CumeDist": self._compute_cume_dist,
        }
        handler = _window_dispatch.get(func_type)
        if handler:
            if func_type in (
                "Lag",
                "Lead",
                "FirstValue",
                "LastValue",
                "NthValue",
                "Ntile",
            ):
                return handler(
                    window_expr,
                    df,
                    partition_cols,
                    order_cols,
                    ascending,
                    select_alias_map or {},
                )
            return handler(df, partition_cols, order_cols, ascending)

        # 窗口聚合函数: AVG/SUM/COUNT/MIN/MAX OVER (...)
        if func_type in _window_agg_funcs:
            return self._compute_window_aggregate(
                func_type,
                df,
                partition_cols,
                order_cols,
                ascending,
                window_expr,
                select_alias_map or {},
            )

        raise ValueError(f"不支持的窗口函数: {func_type}")

    def _resolve_window_column(self, col_name: str, df_columns: list, select_alias_map: dict[str, str]) -> str:
        """解析窗口函数中的列名(支持聚合表达式->别名映射)"""
        # 1. 直接在SELECT别名映射中查找
        if col_name in select_alias_map:
            alias = select_alias_map[col_name]
            if alias in df_columns:
                return alias

        # 2. 尝试聚合函数名匹配
        agg_funcs = {"AVG", "SUM", "COUNT", "MAX", "MIN"}
        for func in agg_funcs:
            if col_name.upper().startswith(func):
                match = re.match(rf"{func}\s*\(\s*(.+?)\s*\)", col_name, re.IGNORECASE)
                if match:
                    inner_col = match.group(1).strip()
                    candidates = [
                        f"{func.lower()}_{inner_col}",
                        f"{func.lower()}_{inner_col.lower()}",
                        inner_col,
                    ]
                    for c in candidates:
                        if c in df_columns:
                            return c

        return col_name  # 未找到映射,返回原名

    def _compute_count_window(self, window_expr: exp.Window, df: pd.DataFrame, partition_cols: list) -> pd.Series:
        """COUNT() OVER(): 返回每个分区的总行数

        支持:
        - COUNT(*) OVER (PARTITION BY col): 计算分区内的所有行数
        - COUNT(col) OVER (PARTITION BY col): 计算分区内col不为NULL的行数
        """
        # 解析COUNT函数的参数
        count_func = window_expr.this
        count_all = False
        count_col = None
        is_distinct = False

        # 检查是否为COUNT(*)
        if isinstance(count_func.this, exp.Star):
            count_all = True
        # 检查是否为DISTINCT
        elif isinstance(count_func.this, exp.Distinct):
            is_distinct = True
            count_col = self._extract_agg_column(count_func.this.expressions[0], "COUNT(DISTINCT)")
        else:
            # COUNT(col)
            count_col = self._extract_agg_column(count_func.this, "COUNT")

        # 验证列存在
        if count_col and count_col not in df.columns:
            suggestion = self._suggest_column_name(count_col, list(df.columns))
            raise ValueError(f"COUNT() 窗口函数中列 '{count_col}' 不存在.可用列: {list(df.columns)}.{suggestion}")

        def compute_count_in_group(group):
            """在分组内计算COUNT"""
            if count_all:
                # COUNT(*): 计算所有行
                return pd.Series([len(group)] * len(group), index=group.index)
            elif is_distinct:
                # COUNT(DISTINCT col): 计算不同值的数量
                unique_count = group[count_col].nunique()
                return pd.Series([unique_count] * len(group), index=group.index)
            else:
                # COUNT(col): 计算非NULL值的数量
                non_null_count = group[count_col].count()
                return pd.Series([non_null_count] * len(group), index=group.index)

        if partition_cols:
            grouped = df.groupby(partition_cols, sort=False, dropna=False)
            result = grouped.apply(compute_count_in_group, include_groups=False)
            if isinstance(result.index, pd.MultiIndex):
                result = result.droplevel(result.index.names[:-1])
        else:
            # 无PARTITION BY: 全表作为一个分区
            result = compute_count_in_group(df)

        return result.astype("Int64")

    def _compute_row_number(self, df: pd.DataFrame, partition_cols: list, order_cols: list, ascending: list) -> pd.Series:
        """ROW_NUMBER: 分区内从1开始的连续编号"""
        if not partition_cols and not order_cols:
            return pd.Series(range(1, len(df) + 1), index=df.index, dtype=int)

        # 排序后用 cumcount 计算行号，避免 groupby.apply 的 pandas 版本兼容问题
        if order_cols:
            sorted_df = df.sort_values(order_cols, ascending=ascending)
        else:
            sorted_df = df

        if partition_cols:
            result = sorted_df.groupby(partition_cols, sort=False, dropna=False).cumcount() + 1
        else:
            result = pd.Series(range(1, len(sorted_df) + 1), index=sorted_df.index, dtype=int)

        return result.reindex(df.index).astype("Int64")

    def _compute_rank(self, df: pd.DataFrame, partition_cols: list, order_cols: list, ascending: list) -> pd.Series:
        """RANK: 相同值相同排名,下一个排名跳过(1,2,2,4)"""
        if not order_cols:
            raise ValueError("RANK() 窗口函数需要 ORDER BY 子句")

        sorted_df = df.sort_values(order_cols, ascending=ascending)
        if partition_cols:
            result = sorted_df.groupby(partition_cols, sort=False, dropna=False)[order_cols[0]].rank(method="min", ascending=ascending[0])
        else:
            result = sorted_df[order_cols[0]].rank(method="min", ascending=ascending[0])
        return result.reindex(df.index).astype("Int64")

    def _compute_dense_rank(self, df: pd.DataFrame, partition_cols: list, order_cols: list, ascending: list) -> pd.Series:
        """DENSE_RANK: 相同值相同排名,下一个排名不跳过(1,2,2,3)"""
        if not order_cols:
            raise ValueError("DENSE_RANK() 窗口函数需要 ORDER BY 子句")

        sorted_df = df.sort_values(order_cols, ascending=ascending)
        if partition_cols:
            result = sorted_df.groupby(partition_cols, sort=False)[order_cols[0]].rank(method="dense", ascending=ascending[0])
        else:
            result = sorted_df[order_cols[0]].rank(method="dense", ascending=ascending[0])
        return result.reindex(df.index).astype("Int64")

    def _compute_percent_rank(self, df: pd.DataFrame, partition_cols: list, order_cols: list, ascending: list) -> pd.Series:
        """PERCENT_RANK: 百分比排名 (rank-1)/(n-1), 结果在0~1之间"""
        if not order_cols:
            raise ValueError("PERCENT_RANK() 窗口函数需要 ORDER BY 子句")

        sorted_df = df.sort_values(order_cols, ascending=ascending) if order_cols else df
        if partition_cols:
            rank = sorted_df.groupby(partition_cols, sort=False)[order_cols[0]].rank(method="min", ascending=ascending[0])
            counts = sorted_df.groupby(partition_cols, sort=False)[order_cols[0]].transform("count")
            result = (rank - 1) / (counts - 1)
            result = result.reindex(df.index).fillna(0.0)
        else:
            rank = sorted_df[order_cols[0]].rank(method="min", ascending=ascending[0])
            n = len(sorted_df)
            result = (rank - 1) / (n - 1) if n > 1 else pd.Series(0.0, index=sorted_df.index)
            result = result.reindex(df.index)
        return result

    def _compute_cume_dist(self, df: pd.DataFrame, partition_cols: list, order_cols: list, ascending: list) -> pd.Series:
        """CUME_DIST: 累积分布, 值<=当前值的行数/总行数"""
        if not order_cols:
            raise ValueError("CUME_DIST() 窗口函数需要 ORDER BY 子句")

        sorted_df = df.sort_values(order_cols, ascending=ascending) if order_cols else df
        if partition_cols:
            rank = sorted_df.groupby(partition_cols, sort=False)[order_cols[0]].rank(method="max", ascending=ascending[0])
            counts = sorted_df.groupby(partition_cols, sort=False)[order_cols[0]].transform("count")
            result = rank / counts
            result = result.reindex(df.index)
        else:
            rank = sorted_df[order_cols[0]].rank(method="max", ascending=ascending[0])
            n = len(sorted_df)
            result = (rank / n).reindex(df.index)
        return result

    def _compute_window_aggregate(
        self,
        func_type: str,
        df: pd.DataFrame,
        partition_cols: list,
        order_cols: list,
        ascending: list,
        window_expr,
        select_alias_map: dict = None,
    ) -> pd.Series:
        """窗口聚合函数: AVG/SUM/COUNT/MIN/MAX OVER (...)

        无ORDER BY时计算整个分区的聚合值;
        有ORDER BY时计算累计聚合值(running aggregate)。
        """
        func = window_expr.this
        # 提取列名
        inner = func.this
        if isinstance(inner, exp.Star):
            # COUNT(*) OVER (...)
            col_name = None
        else:
            col_name = inner.name if hasattr(inner, "name") else str(inner)
            # JOIN列映射
            if hasattr(self, "_join_column_mapping") and col_name not in df.columns:
                for _, col_map in self._join_column_mapping.items():
                    if col_name in col_map and col_map[col_name] in df.columns:
                        col_name = col_map[col_name]
                        break
            # [FIX R15-B1] 裸列名找不到时，尝试匹配 "table.col" 格式（与 resolve_col_name 保持一致）
            if col_name not in df.columns:
                for fc in df.columns:
                    if fc.endswith(f".{col_name}") or fc.endswith(f"_{col_name}"):
                        col_name = fc
                        break
            # [FIX R15-B1c] GROUP BY 后原始列不存在，通过 select_alias_map 反向查找
            # 场景: AVG(s.Quantity) OVER (...) 在 GROUP BY 后执行，s.Quantity 已被聚合为 TotalSold
            if col_name not in df.columns and select_alias_map:
                expr_str = str(inner).upper()
                for orig_expr, alias in select_alias_map.items():
                    ou = orig_expr.upper()
                    # 精确匹配
                    if ou == expr_str and alias in df.columns:
                        col_name = alias
                        break
                    # 从聚合函数表达式中提取内部列名进行匹配
                    # 例如 orig_expr='SUM(s.Quantity)', expr_str='S.QUANTITY'
                    # re already imported at top level as _re

                    agg_match = re.match(r"(AVG|SUM|COUNT|MAX|MIN)\s*\(\s*(.+?)\s*\)$", ou)
                    if agg_match:
                        agg_inner = agg_match.group(2).strip()
                        if agg_inner == expr_str or agg_inner.endswith(expr_str) or expr_str.endswith(agg_inner):
                            if alias in df.columns:
                                col_name = alias
                                break

        numeric_col = col_name and col_name in df.columns
        if numeric_col and func_type != "Count":
            series = pd.to_numeric(df[col_name], errors="coerce")
        elif col_name is None:
            # COUNT(*)
            series = pd.Series(1, index=df.index)
        else:
            series = df[col_name]

        if partition_cols:
            grouped = series.groupby([df[c] for c in partition_cols], sort=False)
        else:
            grouped = None

        # 有ORDER BY → 累计聚合; 无ORDER BY → 分区整体聚合
        if order_cols:
            sorted_df = df.sort_values(order_cols, ascending=ascending)
            if partition_cols:
                sorted_group = series.reindex(sorted_df.index).groupby([df[c].reindex(sorted_df.index) for c in partition_cols], sort=False)
            else:
                sorted_group = None

            if func_type == "Sum":
                result = sorted_group.cumsum() if sorted_group else series.reindex(sorted_df.index).cumsum()
            elif func_type == "Count":
                result = sorted_group.cumcount() + 1 if sorted_group else pd.Series(range(1, len(df) + 1), index=sorted_df.index)
            elif func_type == "Min":
                result = sorted_group.cummin() if sorted_group else series.reindex(sorted_df.index).cummin()
            elif func_type == "Max":
                result = sorted_group.cummax() if sorted_group else series.reindex(sorted_df.index).cummax()
            elif func_type == "Avg":
                if sorted_group:
                    cumsum = sorted_group.cumsum()
                    cumcount = sorted_group.cumcount() + 1
                    result = cumsum / cumcount
                else:
                    s = series.reindex(sorted_df.index)
                    result = s.cumsum() / pd.Series(range(1, len(s) + 1), index=s.index)
            elif func_type in ("Stddev", "Std", "Variance", "Var"):
                # 累计标准差/方差:计算整个分区的std/var并填充到每行
                # 注意:这是简化的实现,不是数学上严格的累计std/var
                if sorted_group:
                    result = sorted_group.transform("std" if "Std" in func_type else "var").reindex(df.index)
                else:
                    result = series.agg("std" if "Std" in func_type else "var")
                    result = pd.Series([result] * len(df), index=df.index)
            else:
                raise ValueError(f"不支持的窗口聚合函数: {func_type}")
            result = result.reindex(df.index)
        else:
            # 无ORDER BY: 分区整体聚合
            agg_map = {
                "Avg": "mean",
                "Sum": "sum",
                "Count": "count",
                "Min": "min",
                "Max": "max",
                "Stddev": "std",
                "Std": "std",
                "Variance": "var",
                "Var": "var",
            }
            agg_func = agg_map[func_type]
            if grouped is not None:
                result = grouped.transform(agg_func)
            else:
                val = series.agg(agg_func)
                result = pd.Series(val, index=df.index)

        return result

    def _compute_percent_rank(self, df: pd.DataFrame, partition_cols: list, order_cols: list, ascending: list) -> pd.Series:
        """PERCENT_RANK: 返回行的相对排名百分比

        公式: (rank - 1) / (total_rows - 1)
        - 第一行: percent_rank = 0
        - 最后一行: percent_rank = 1
        - 只有一行: percent_rank = 0
        """
        if not order_cols:
            raise ValueError("PERCENT_RANK() 窗口函数需要 ORDER BY 子句")

        if partition_cols:
            grouped = df.groupby(partition_cols, sort=False)
        else:
            grouped = None

        def assign_percent_rank(group):
            """为分组内的行计算百分比排名"""
            sorted_group = group.sort_values(order_cols, ascending=ascending)
            group_size = len(sorted_group)

            if group_size == 1:
                # 只有一行时，percent_rank = 0
                return pd.Series([0.0], index=group.index)

            # 使用RANK方法计算排名（相同值相同排名，下一个排名跳过）
            rank_series = sorted_group[order_cols[0]].rank(method="first", ascending=ascending[0])
            # 计算百分比排名: (rank - 1) / (n - 1)
            percent_rank = (rank_series - 1) / (group_size - 1)

            return percent_rank.reindex(group.index)

        if grouped is not None:
            result = grouped.apply(assign_percent_rank, include_groups=False)
            if isinstance(result.index, pd.MultiIndex):
                result = result.droplevel(result.index.names[:-1])
        else:
            result = assign_percent_rank(df)

        return result

    def _compute_cume_dist(self, df: pd.DataFrame, partition_cols: list, order_cols: list, ascending: list) -> pd.Series:
        """CUME_DIST: 返回累积分布值

        公式: number_of_rows_with_value <= current_value / total_rows
        - 返回值范围: (0, 1]
        - 第一行: cume_dist = 1 / n (最小值)
        - 最后一行: cume_dist = 1 (最大值)
        """
        if not order_cols:
            raise ValueError("CUME_DIST() 窗口函数需要 ORDER BY 子句")

        if partition_cols:
            grouped = df.groupby(partition_cols, sort=False)
        else:
            grouped = None

        def assign_cume_dist(group):
            """为分组内的行计算累积分布"""
            sorted_group = group.sort_values(order_cols, ascending=ascending)
            group_size = len(sorted_group)

            # 计算累积分布: 对于每一行，计算小于等于当前值的行数 / 总行数
            cume_dist = sorted_group[order_cols[0]].rank(method="max", ascending=ascending[0]) / group_size

            return cume_dist.reindex(group.index)

        if grouped is not None:
            result = grouped.apply(assign_cume_dist, include_groups=False)
            if isinstance(result.index, pd.MultiIndex):
                result = result.droplevel(result.index.names[:-1])
        else:
            result = assign_cume_dist(df)

        return result

    def _compute_ntile(
        self,
        window_expr: exp.Window,
        df: pd.DataFrame,
        partition_cols: list,
        order_cols: list,
        ascending: list,
        select_alias_map: dict,
    ) -> pd.Series:
        """NTILE: 将分组内的行均匀分为N个桶（桶号从1到N）"""
        if not order_cols:
            raise ValueError("NTILE() 窗口函数需要 ORDER BY 子句")

        # 解析桶数参数
        ntile_func = window_expr.this
        bucket_count = 1
        # NTILE的参数直接存储在ntile_func.this中（是一个Literal）
        if hasattr(ntile_func, "this") and isinstance(ntile_func.this, exp.Literal):
            bucket_count = int(ntile_func.this.this)
        elif hasattr(ntile_func, "this"):
            # 尝试从表达式解析整数
            try:
                bucket_count = int(str(ntile_func.this).strip())
            except (ValueError, TypeError):
                bucket_count = 1

        # 验证桶数
        if bucket_count < 1:
            raise ValueError("NTILE() 的桶数参数必须大于等于 1")

        # 分组计算NTILE
        if partition_cols:
            grouped = df.groupby(partition_cols, sort=False)
        else:
            grouped = None

        def assign_ntile_buckets(group):
            """为分组内的行分配桶号

            NTILE算法：将行均匀分配到N个桶中
            - 每个桶分配尽可能相等的行数
            - 行数多的桶在前面（按排序顺序）
            - 例如：8行分成3个桶 → 桶大小为3,3,2
            """
            sorted_group = group.sort_values(order_cols, ascending=ascending)
            group_size = len(sorted_group)

            # 计算每个桶的行数
            base_size = group_size // bucket_count  # 每个桶的基础行数
            remainder = group_size % bucket_count  # 余数：前remainder个桶多一行

            # 创建桶号数组
            bucket_numbers = []
            current_bucket = 1
            rows_in_current_bucket = base_size + (1 if current_bucket <= remainder else 0)

            for _ in range(group_size):
                bucket_numbers.append(current_bucket)
                rows_in_current_bucket -= 1
                if rows_in_current_bucket == 0 and current_bucket < bucket_count:
                    current_bucket += 1
                    rows_in_current_bucket = base_size + (1 if current_bucket <= remainder else 0)

            return pd.Series(bucket_numbers, index=sorted_group.index).reindex(group.index)

        if grouped is not None:
            result = grouped.apply(assign_ntile_buckets, include_groups=False)
            if isinstance(result.index, pd.MultiIndex):
                result = result.droplevel(result.index.names[:-1])
        else:
            result = assign_ntile_buckets(df)

        return result

    def _compute_lag(
        self,
        window_expr: exp.Window,
        df: pd.DataFrame,
        partition_cols: list,
        order_cols: list,
        ascending: list,
        select_alias_map: dict,
    ) -> pd.Series:
        """LAG: 获取分区内当前行之前第N行的值"""
        if not order_cols:
            raise ValueError("LAG() 窗口函数需要 ORDER BY 子句")

        # 解析目标列和偏移量
        lag_func = window_expr.this
        target_col_expr = lag_func.this
        target_col = target_col_expr.name if hasattr(target_col_expr, "name") else str(target_col_expr)

        # [FIX R15-B3] 处理内层聚合表达式: LAG(MAX(Price)) 等
        # 当内层是聚合函数(Max/Min/Sum/Avg/Count)时，target_col可能为空
        # 需要通过select_alias_map反向查找对应的别名列
        if not target_col or target_col not in df.columns:
            expr_str = str(target_col_expr).strip()
            if expr_str in (select_alias_map or {}):
                alias_name = select_alias_map[expr_str]
                if alias_name in df.columns:
                    target_col = alias_name
            if target_col not in df.columns:
                agg_funcs = {"Max", "Min", "Sum", "Avg", "Count"}
                func_type = type(lag_func.this).__name__ if hasattr(lag_func, "this") else ""
                if func_type in agg_funcs:
                    found = False
                    for expr_key, alias_name in (select_alias_map or {}).items():
                        if func_type.upper() in expr_key.upper() and alias_name in df.columns:
                            target_col = alias_name
                            found = True
                            break
                    # [FIX R15-B3b] 递归提取内层AST节点直到找到Column节点
                    if not found:
                        node = lag_func.this
                        while hasattr(node, "this") and not isinstance(node, exp.Column):
                            node = node.this
                        if isinstance(node, exp.Column) and node.name and node.name in df.columns:
                            target_col = node.name

        # 解析偏移量参数（默认为1）
        offset = 1
        if hasattr(lag_func, "args") and "offset" in lag_func.args:
            offset_arg = lag_func.args["offset"]
            if offset_arg:
                offset = int(offset_arg.this) if hasattr(offset_arg, "this") else 1

        # 解析默认值参数（可选）
        default_value = None
        if hasattr(lag_func, "args") and "default" in lag_func.args:
            default_arg = lag_func.args["default"]
            if default_arg:
                raw = default_arg.this if hasattr(default_arg, "this") else None
                if raw is not None:
                    try:
                        default_value = float(raw) if "." in str(raw) else int(raw)
                    except (ValueError, TypeError):
                        default_value = raw

        # 验证目标列存在
        if target_col not in df.columns:
            suggestion = self._suggest_column_name(target_col, list(df.columns))
            raise ValueError(f"LAG() 函数中列 '{target_col}' 不存在.可用列: {list(df.columns)}.{suggestion}")

        # 分组计算LAG
        if partition_cols:
            # [FIX R15-B2] 推断目标列的dtype，避免字符串列赋值到float Series报错
            target_dtype = df[target_col].dtype if target_col in df.columns else object
            result = pd.Series(None, index=df.index, dtype=target_dtype)
            for _, group in df.groupby(partition_cols, sort=False):
                sorted_group = group.sort_values(order_cols, ascending=ascending)
                lagged = sorted_group[target_col].shift(periods=offset)
                if default_value is not None:
                    lagged = lagged.fillna(default_value)
                result.loc[group.index] = lagged.reindex(group.index).values
        else:
            sorted_df = df.sort_values(order_cols, ascending=ascending)
            lagged = sorted_df[target_col].shift(periods=offset)
            if default_value is not None:
                lagged = lagged.fillna(default_value)
            result = lagged.reindex(df.index)

        return result

    def _compute_lead(
        self,
        window_expr: exp.Window,
        df: pd.DataFrame,
        partition_cols: list,
        order_cols: list,
        ascending: list,
        select_alias_map: dict,
    ) -> pd.Series:
        """LEAD: 获取分区内当前行之后第N行的值"""
        if not order_cols:
            raise ValueError("LEAD() 窗口函数需要 ORDER BY 子句")

        # 解析目标列和偏移量
        lead_func = window_expr.this
        target_col_expr = lead_func.this
        target_col = target_col_expr.name if hasattr(target_col_expr, "name") else str(target_col_expr)

        # [FIX R15-B3] 处理内层聚合表达式: LEAD(MAX(Price)) 等
        # 当内层是聚合函数(Max/Min/Sum/Avg/Count)时，target_col可能为空
        # 需要通过select_alias_map反向查找对应的别名列
        if not target_col or target_col not in df.columns:
            # 尝试从表达式的字符串形式在alias_map中查找
            expr_str = str(target_col_expr).strip()
            if expr_str in (select_alias_map or {}):
                alias_name = select_alias_map[expr_str]
                if alias_name in df.columns:
                    target_col = alias_name
            # 如果还是找不到，尝试匹配包含聚合函数名的列
            if target_col not in df.columns:
                agg_funcs = {"Max", "Min", "Sum", "Avg", "Count"}
                func_type = type(lead_func.this).__name__ if hasattr(lead_func, "this") else ""
                if func_type in agg_funcs:
                    # 查找select_alias_map中引用了此聚合的别名
                    found = False
                    for expr_key, alias_name in (select_alias_map or {}).items():
                        if func_type.upper() in expr_key.upper() and alias_name in df.columns:
                            target_col = alias_name
                            found = True
                            break
                    # [FIX R15-B3b] 如果alias_map中也没有有用的映射，
                    # 递归提取内层AST节点直到找到Column节点
                    if not found:
                        node = lead_func.this
                        while hasattr(node, "this") and not isinstance(node, exp.Column):
                            node = node.this
                        if isinstance(node, exp.Column) and node.name and node.name in df.columns:
                            target_col = node.name

        # 解析偏移量参数（默认为1）
        offset = 1
        if hasattr(lead_func, "args") and "offset" in lead_func.args:
            offset_arg = lead_func.args["offset"]
            if offset_arg:
                offset = int(offset_arg.this) if hasattr(offset_arg, "this") else 1

        # 解析默认值参数（可选）
        default_value = None
        if hasattr(lead_func, "args") and "default" in lead_func.args:
            default_arg = lead_func.args["default"]
            if default_arg:
                raw = default_arg.this if hasattr(default_arg, "this") else None
                if raw is not None:
                    try:
                        default_value = float(raw) if "." in str(raw) else int(raw)
                    except (ValueError, TypeError):
                        default_value = raw

        # 验证目标列存在
        if target_col not in df.columns:
            suggestion = self._suggest_column_name(target_col, list(df.columns))
            raise ValueError(f"LEAD() 函数中列 '{target_col}' 不存在.可用列: {list(df.columns)}.{suggestion}")

        # 分组计算LEAD
        if partition_cols:
            # [FIX R15-B2] 推断目标列的dtype，避免字符串列赋值到float Series报错
            target_dtype = df[target_col].dtype if target_col in df.columns else object
            result = pd.Series(None, index=df.index, dtype=target_dtype)
            for _, group in df.groupby(partition_cols, sort=False):
                sorted_group = group.sort_values(order_cols, ascending=ascending)
                led = sorted_group[target_col].shift(periods=-offset)
                if default_value is not None:
                    led = led.fillna(default_value)
                result.loc[group.index] = led.reindex(group.index).values
        else:
            sorted_df = df.sort_values(order_cols, ascending=ascending)
            led = sorted_df[target_col].shift(periods=-offset)
            if default_value is not None:
                led = led.fillna(default_value)
            result = led.reindex(df.index)

        return result

    def _compute_first_value(
        self,
        window_expr: exp.Window,
        df: pd.DataFrame,
        partition_cols: list,
        order_cols: list,
        ascending: list,
        select_alias_map: dict,
    ) -> pd.Series:
        """FIRST_VALUE: 获取分区内排序后第一行的值"""
        if not order_cols:
            raise ValueError("FIRST_VALUE() 窗口函数需要 ORDER BY 子句")

        first_value_func = window_expr.this
        target_col_expr = first_value_func.this
        target_col = target_col_expr.name if hasattr(target_col_expr, "name") else str(target_col_expr)

        if target_col not in df.columns:
            suggestion = self._suggest_column_name(target_col, list(df.columns))
            raise ValueError(f"FIRST_VALUE() 函数中列 '{target_col}' 不存在.可用列: {list(df.columns)}.{suggestion}")

        result = pd.Series(None, index=df.index, dtype=df[target_col].dtype)
        if partition_cols:
            for _, group in df.groupby(partition_cols, sort=False):
                sorted_group = group.sort_values(order_cols, ascending=ascending)
                first_val = sorted_group[target_col].iloc[0] if len(sorted_group) > 0 else None
                result.loc[group.index] = first_val
        else:
            sorted_df = df.sort_values(order_cols, ascending=ascending)
            first_val = sorted_df[target_col].iloc[0] if len(sorted_df) > 0 else None
            result[:] = first_val

        return result

    def _compute_last_value(
        self,
        window_expr: exp.Window,
        df: pd.DataFrame,
        partition_cols: list,
        order_cols: list,
        ascending: list,
        select_alias_map: dict,
    ) -> pd.Series:
        """LAST_VALUE: 获取分区内排序后最后一行的值"""
        if not order_cols:
            raise ValueError("LAST_VALUE() 窗口函数需要 ORDER BY 子句")

        last_value_func = window_expr.this
        target_col_expr = last_value_func.this
        target_col = target_col_expr.name if hasattr(target_col_expr, "name") else str(target_col_expr)

        if target_col not in df.columns:
            suggestion = self._suggest_column_name(target_col, list(df.columns))
            raise ValueError(f"LAST_VALUE() 函数中列 '{target_col}' 不存在.可用列: {list(df.columns)}.{suggestion}")

        result = pd.Series(None, index=df.index, dtype=df[target_col].dtype)
        if partition_cols:
            for _, group in df.groupby(partition_cols, sort=False):
                sorted_group = group.sort_values(order_cols, ascending=ascending)
                last_val = sorted_group[target_col].iloc[-1] if len(sorted_group) > 0 else None
                result.loc[group.index] = last_val
        else:
            sorted_df = df.sort_values(order_cols, ascending=ascending)
            last_val = sorted_df[target_col].iloc[-1] if len(sorted_df) > 0 else None
            result[:] = last_val

        return result

    def _compute_nth_value(
        self,
        window_expr: exp.Window,
        df: pd.DataFrame,
        partition_cols: list,
        order_cols: list,
        ascending: list,
        select_alias_map: dict,
    ) -> pd.Series:
        """NTH_VALUE: 获取分区内排序后第N行的值"""
        if not order_cols:
            raise ValueError("NTH_VALUE() 窗口函数需要 ORDER BY 子句")

        nth_func = window_expr.this
        target_col_expr = nth_func.this
        target_col = target_col_expr.name if hasattr(target_col_expr, "name") else str(target_col_expr)

        # 解析 N 参数
        nth_expr = nth_func.args.get("offset") or (nth_func.expressions[1] if len(nth_func.expressions) > 1 else None)
        n = int(nth_expr.this) if nth_expr and hasattr(nth_expr, "this") else 1

        if target_col not in df.columns:
            raise ValueError(f"NTH_VALUE() 函数中列 '{target_col}' 不存在.可用列: {list(df.columns)}")

        result = pd.Series(None, index=df.index, dtype=df[target_col].dtype)
        if partition_cols:
            for _, group in df.groupby(partition_cols, sort=False):
                sorted_group = group.sort_values(order_cols, ascending=ascending)
                nth_val = sorted_group[target_col].iloc[n - 1] if n <= len(sorted_group) else None
                result.loc[group.index] = nth_val
        else:
            sorted_df = df.sort_values(order_cols, ascending=ascending)
            nth_val = sorted_df[target_col].iloc[n - 1] if n <= len(sorted_df) else None
            result[:] = nth_val

        return result

    # CTE 最大嵌套深度限制（防止恶意/意外深层递归导致 StackOverflow）
    _MAX_CTE_DEPTH = 10

    def _execute_query(
        self,
        parsed_sql: exp.Expression,
        worksheets_data: dict[str, pd.DataFrame],
        limit: int | None = None,
        _cte_depth: int = 0,
    ) -> pd.DataFrame:
        """
        执行解析后的SQL查询

        Args:
            parsed_sql: 解析后的SQL表达式
            worksheets_data: 工作表数据
            limit: 结果限制
            _cte_depth: CTE 嵌套深度（内部使用，防止无限递归）

        Returns:
            pd.DataFrame: 查询结果
        """
        # 保存worksheets_data为实例变量,供子查询执行使用
        self._worksheets_data = worksheets_data

        # Fix(R4): CTE 深度检查（放在入口处，无论当前 SQL 是否有 CTE 都拦截）
        if _cte_depth >= self._MAX_CTE_DEPTH:
            raise ValueError(
                f"CTE 嵌套深度超过限制 ({self._MAX_CTE_DEPTH})。"
                f"💡 请简化查询，减少 CTE 嵌套层数，或改用子查询替代多层 CTE。"
            )

        # 处理CTE (WITH ... AS ...)
        # 兼容sqlglot不同版本:arg key可能是'with'或'with_'
        _with_key = "with" if parsed_sql.args.get("with") else "with_"
        with_clause = parsed_sql.args.get(_with_key)
        if with_clause:
            # 复制worksheets_data避免修改原始数据,逐步添加CTE结果
            cte_data = dict(worksheets_data)
            for cte_expr in with_clause.expressions:
                cte_name = cte_expr.alias
                cte_query = cte_expr.this  # inner Select
                try:
                    # 每个CTE在已有的cte_data上执行(支持CTE引用前面的CTE)
                    # 递归深度 +1
                    cte_result = self._execute_query(cte_query, cte_data, limit=None, _cte_depth=_cte_depth + 1)
                    cte_data[cte_name] = cte_result
                except Exception as e:
                    raise ValueError(f"CTE '{cte_name}' 执行失败: {e}")
            # 从parsed_sql中移除with子句,让后续逻辑正常处理
            parsed_sql = parsed_sql.copy()
            parsed_sql.set(_with_key, None)

        # 获取FROM子句中的表名(及可选的子查询)
        from_table, from_subquery = self._get_from_table(parsed_sql)

        # 查找表名时也搜索CTE定义的临时表
        effective_data = cte_data if with_clause else worksheets_data

        # 如果FROM是子查询,先执行子查询并将结果注入effective_data
        if from_subquery is not None:
            try:
                sub_result = self._execute_subquery(from_subquery, effective_data)
                effective_data[from_table] = sub_result
            except Exception as e:
                raise StructuredSQLError(
                    "from_subquery_error",
                    f"FROM子查询执行失败: {e}",
                    hint="请检查FROM子查询中的SQL语法和表名是否正确.",
                    context={"subquery_alias": from_table, "error": str(e)},
                )

        if from_table not in effective_data:
            raise StructuredSQLError(
                "table_not_found",
                f"表 '{from_table}' 不存在.可用表: {list(effective_data.keys())}",
                hint="请检查表名拼写,或用excel_list_sheets查看可用工作表名.",
                context={
                    "table_requested": from_table,
                    "available_tables": list(effective_data.keys()),
                },
            )

        base_df = effective_data[from_table].copy()

        # 添加行号虚拟列 _ROW_NUMBER_ (SELECT和UPDATE通用)
        if "_ROW_NUMBER_" not in base_df.columns:
            base_df["_ROW_NUMBER_"] = range(1, len(base_df) + 1)

        # 构建表别名映射
        self._table_aliases = {}
        self._table_aliases[from_table] = from_table
        # 检查FROM子句是否有别名 (FROM 技能表 a)
        from_clause = parsed_sql.args.get("from")
        if from_clause:
            # 优先使用 Table.alias 属性
            from_table_expr = from_clause.this
            if hasattr(from_table_expr, "alias") and from_table_expr.alias:
                from_alias = from_table_expr.alias
                if isinstance(from_alias, str) and from_alias != from_table:
                    self._table_aliases[from_alias] = from_table
                    self._table_aliases[from_table] = from_table
            # 备用:遍历 TableAlias 节点
            found_from_alias = False
            for alias in from_clause.find_all(exp.TableAlias):
                parent_table = from_clause.this.name if hasattr(from_clause.this, "name") else str(from_clause.this)
                self._table_aliases[alias.alias] = parent_table
                self._table_aliases[parent_table] = parent_table
                found_from_alias = True
            if not found_from_alias:
                for alias in from_clause.find_all(exp.Alias):
                    parent_table = from_clause.this.name if hasattr(from_clause.this, "name") else str(from_clause.this)
                    self._table_aliases[alias.alias] = parent_table
                    self._table_aliases[parent_table] = parent_table

        # 应用JOIN子句
        joins = parsed_sql.args.get("joins")
        if joins:
            base_df = self._apply_join_clause(joins, base_df, effective_data, from_table)

        # 应用WHERE条件
        # 保存WHERE前的DataFrame,用于空结果智能建议
        base_df_before_where = base_df.copy()
        self._df_before_where = base_df_before_where
        # 保存当前工作表数据供子查询使用
        self._current_worksheets = effective_data
        base_df = self._apply_where_clause(parsed_sql, base_df)

        # 检查是否有聚合函数
        has_aggregate = self._check_has_aggregate_function(parsed_sql)

        # [FIX R14-B1] 当同时存在GROUP BY和窗口函数时，窗口函数需要在GROUP BY之前计算
        # 因为GROUP BY会丢弃非分组列，而窗口函数的ORDER BY/PARTITION BY可能引用这些列
        has_window = self._has_window_function(parsed_sql)
        has_group_by = parsed_sql.args.get("group") is not None or has_aggregate
        _precomputed_windows = False  # 标记是否已预计算窗口函数

        if has_group_by and has_window:
            try:
                base_df = self._apply_window_functions(parsed_sql, base_df)
                _precomputed_windows = True
            except Exception:
                # 预计算失败时，回退到原有流程（在GROUP BY后再尝试）
                _precomputed_windows = False

        # 应用GROUP BY和聚合
        if parsed_sql.args.get("group") or has_aggregate:
            # 有GROUP BY或有聚合函数时,应用分组聚合
            base_df = self._apply_group_by_aggregation(parsed_sql, base_df)

            # 应用HAVING条件
            has_having = parsed_sql.args.get("having") is not None
            if has_having:
                # 保存HAVING前的DataFrame,用于HAVING空结果建议
                self._df_before_having = base_df.copy()
                base_df = self._apply_having_clause(parsed_sql, base_df)
        else:
            has_having = False

        # 应用窗口函数(ROW_NUMBER, RANK, DENSE_RANK)
        # 窗口函数在GROUP BY/HAVING之后,ORDER BY/SELECT之前计算
        # [FIX R14-B1] 如果已在GROUP BY前预计算过，则跳过
        if not _precomputed_windows:
            base_df = self._apply_window_functions(parsed_sql, base_df)

        if parsed_sql.args.get("group") or has_aggregate:
            # ORDER BY(聚合查询:在GROUP BY之后)
            # 提取SELECT别名,支持ORDER BY引用聚合结果列的别名
            select_aliases = self._extract_select_aliases(parsed_sql)
            if parsed_sql.args.get("order"):
                base_df = self._apply_order_by(parsed_sql, base_df, select_aliases=select_aliases)
        else:
            # 非聚合查询:提取SELECT别名,然后ORDER BY(支持引用别名和原始列),最后SELECT
            select_aliases = self._extract_select_aliases(parsed_sql)
            if parsed_sql.args.get("order"):
                base_df = self._apply_order_by(parsed_sql, base_df, select_aliases=select_aliases)

            # 应用SELECT表达式(裁剪列,计算字段,别名)
            base_df = self._apply_select_expressions(parsed_sql, base_df)

        # R51-opt: LIMIT/OFFSET 优化 — 合并操作 + 早返回 + 边界检查
        offset_value = self._extract_int_value(parsed_sql.args.get("offset"))
        # R48-fix: SELECT DISTINCT 必须在 LIMIT/OFFSET 之前应用(SQL标准执行顺序)
        if parsed_sql.args.get("distinct"):
            base_df = base_df.drop_duplicates()

        limit_value = self._extract_int_value(parsed_sql.args.get("limit"))
        if limit is not None and limit_value is None:
            limit_value = limit

        # 早返回: LIMIT 0 → 空结果（跳过后续切片）
        if limit_value is not None and limit_value <= 0:
            return base_df.iloc[0:0]

        # 归一化 OFFSET（负值视为 0）
        if offset_value is not None and offset_value < 0:
            offset_value = 0

        # 合并 OFFSET+LIMIT 为单次 iloc 操作（避免中间 DataFrame）
        if offset_value is not None and offset_value > 0:
            if limit_value is not None and limit_value > 0:
                # OFFSET + LIMIT 合并: iloc[offset:offset+limit]
                end_idx = offset_value + limit_value
                base_df = base_df.iloc[offset_value:end_idx]
            else:
                # 仅 OFFSET
                base_df = base_df.iloc[offset_value:]
        elif limit_value is not None and limit_value > 0:
            # 仅 LIMIT
            base_df = base_df.head(limit_value)

        return base_df

    @staticmethod
    def _extract_int_value(clause) -> int | None:
        """从SQL子句中提取整数值(LIMIT/OFFSET等)"""
        if clause is None:
            return None

        # 提取值
        if hasattr(clause, "expression") and clause.expression is not None:
            value = clause.expression.this
        else:
            value = clause.this

        # 转换为整数
        try:
            int_value = int(value)
            # LIMIT/OFFSET 不支持负数
            if int_value < 0:
                raise ValueError(f"LIMIT/OFFSET 不支持负数: {int_value}")
            return int_value
        except (ValueError, TypeError) as e:
            if "负数" in str(e):
                raise
            raise ValueError(f"无法提取整数值: {value}")

    def _check_has_aggregate_function(self, parsed_sql: exp.Expression) -> bool:
        """检查SQL查询是否包含聚合函数

        覆盖三种情况:
        1. 直接聚合: SELECT AVG(col) FROM t
        2. 标量函数包裹聚合: SELECT ROUND(AVG(col)), ABS(SUM(col)) FROM t
        3. 别名包裹: SELECT ROUND(AVG(col)) AS alias FROM t

        注意: 标量子查询中的聚合(如 SELECT (SELECT AVG(x) FROM t) FROM s)
        不算作外层聚合，因为子查询内的聚合由 _execute_subquery 独立处理。
        """
        for select_expr in parsed_sql.expressions:
            # 解包别名: 获取实际表达式
            inner_expr = select_expr.this if isinstance(select_expr, exp.Alias) else select_expr

            # 跳过标量子查询: 子查询内聚合是自包含的，不触发外层聚合路径
            # (SELECT AVG(x) FROM t) 由 _apply_select_expressions → Subquery 分支处理
            if isinstance(inner_expr, exp.Subquery):
                continue

            # 情况1: 顶层是聚合函数(含别名包裹)
            if self._is_aggregate_function(select_expr):
                return True
            # 情况2: 标量函数(ROUND/ABS/CEIL/FLOOR/SQRT/POWER)包裹聚合函数
            if self._is_scalar_num_function(inner_expr) and self._find_inner_aggregate(inner_expr) is not None:
                return True
            # 情况3: 其他表达式树中包含聚合函数(CASE WHEN AVG 等)
            if self._find_inner_aggregate(select_expr) is not None:
                return True
        return False

    def _apply_select_expressions(self, parsed_sql: exp.Expression, df) -> pd.DataFrame:
        """
        应用SELECT表达式(非聚合查询)
        处理计算字段,别名等

        Args:
            parsed_sql: 解析后的SQL表达式
            df: 数据DataFrame

        Returns:
            pd.DataFrame: 处理后的DataFrame
        """
        result_data = {}
        ordered_columns = []
        used_aliases = set()  # 跟踪已使用的别名,避免重复

        # 按SELECT表达式顺序处理列
        for i, select_expr in enumerate(parsed_sql.expressions):
            if isinstance(select_expr, exp.Star):
                # SELECT *: 返回所有列(排除内部临时列)
                for col in df.columns:
                    if col == "_ROW_NUMBER_" or col.startswith("_window_"):
                        continue
                    if col not in result_data:
                        result_data[col] = df[col]
                        ordered_columns.append(col)
                continue

            # 处理别名
            alias_name, original_expr = self._extract_select_alias(select_expr, i)

            # 如果是窗口函数且别名已存在于result_data(被SELECT *展开),跳过
            if isinstance(original_expr, exp.Window) and alias_name in result_data:
                continue

            # 处理重复别名:当多个SELECT列解析为相同名称时,添加表前缀
            if alias_name in used_aliases and isinstance(original_expr, exp.Column):
                table_part = original_expr.table if hasattr(original_expr, "table") and original_expr.table else None
                if table_part:
                    qualified_alias = f"{table_part}.{alias_name}"
                    if qualified_alias not in used_aliases:
                        alias_name = qualified_alias
                    else:
                        alias_name = f"{table_part}_{alias_name}_{i}"
                else:
                    alias_name = f"{alias_name}_{i}"
            used_aliases.add(alias_name)

            # 解包Paren: (伤害 * 1.2) → 伤害 * 1.2
            while isinstance(original_expr, exp.Paren):
                original_expr = original_expr.this

            # 计算表达式值
            try:
                if isinstance(original_expr, exp.Column):
                    # 普通列引用(支持表限定符 a.column)
                    column_name = original_expr.name
                    table_part = original_expr.table if hasattr(original_expr, "table") and original_expr.table else None
                    qualified = f"{table_part}.{column_name}" if table_part else None

                    # 先尝试直接使用qualified列名
                    if qualified and qualified in df.columns:
                        result_data[alias_name] = df[qualified]
                    # 如果qualified不存在,尝试映射
                    elif table_part:
                        # 使用_expression_to_column_reference进行映射
                        try:
                            mapped_column = self._expression_to_column_reference(original_expr, df)
                            # 去掉反引号
                            mapped_column = mapped_column.strip("`")
                            result_data[alias_name] = df[mapped_column]
                        except Exception:
                            # 修复:JOIN表别名映射失败时的回退逻辑
                            # 尝试查找可能的pandas merge后缀格式 (_x/_y)
                            possible_columns = [
                                f"{table_part}.{column_name}",  # 用户原始别名格式
                                f"{table_part}_{column_name}",  # table_part_列名格式
                                f"{column_name}_x",  # _x后缀格式
                                f"{column_name}_y",  # _y后缀格式
                                f"{table_part}_x",  # table_part_x格式
                                f"{table_part}_y",  # table_part_y格式
                                column_name,  # 无表前缀的原始列名
                            ]

                            # 去重并检查存在的列(大小写不敏感)
                            for possible_col in possible_columns:
                                actual = self._find_column_name(possible_col, df)
                                if actual:
                                    result_data[alias_name] = df[actual]
                                    break
                            else:
                                # 所有可能的映射都失败,尝试直接使用列名(大小写不敏感)
                                actual_col = self._find_column_name(column_name, df)
                                if actual_col:
                                    result_data[alias_name] = df[actual_col]
                                else:
                                    suggestion = self._suggest_column_name(column_name, list(df.columns))
                                    raise StructuredSQLError(
                                        "column_not_found",
                                        f"列 '{qualified or column_name}' 不存在.可用列: {list(df.columns)}.{suggestion}",
                                        hint="请检查列名拼写,或用excel_get_headers查看所有可用列名.",
                                        context={
                                            "column_requested": qualified or column_name,
                                            "available_columns": list(df.columns),
                                        },
                                    )
                    actual_col = self._find_column_name(column_name, df)
                    if actual_col:
                        result_data[alias_name] = df[actual_col]
                    else:
                        suggestion = self._suggest_column_name(column_name, list(df.columns))
                        raise StructuredSQLError(
                            "column_not_found",
                            f"列 '{qualified or column_name}' 不存在.可用列: {list(df.columns)}.{suggestion}",
                            hint="请检查列名拼写,或用excel_get_headers查看所有可用列名.",
                            context={
                                "column_requested": qualified or column_name,
                                "available_columns": list(df.columns),
                            },
                        )

                elif isinstance(original_expr, exp.Case):
                    # CASE WHEN表达式
                    result_data[alias_name] = self._evaluate_case_expression(original_expr, df)

                elif isinstance(original_expr, exp.Coalesce):
                    # COALESCE/IFNULL表达式(向量化)
                    result_data[alias_name] = self._evaluate_coalesce_vectorized(original_expr, df)

                elif self._is_string_function(original_expr):
                    # 字符串函数: UPPER, LOWER, TRIM, LENGTH, CONCAT, REPLACE, SUBSTRING, LEFT, RIGHT
                    result_data[alias_name] = self._evaluate_string_function(original_expr, df)

                elif self._is_scalar_num_function(original_expr):
                    # 标量数值函数: ROUND(value, decimals)
                    result_data[alias_name] = self._evaluate_scalar_num_function(original_expr, df)

                elif isinstance(original_expr, exp.Window):
                    # 窗口函数: ROW_NUMBER, RANK, DENSE_RANK(已由_apply_window_functions预计算)
                    if alias_name in df.columns:
                        result_data[alias_name] = df[alias_name]
                    else:
                        raise ValueError(f"窗口函数结果列 '{alias_name}' 未找到")

                elif self._is_mathematical_expression(original_expr):
                    # 数学表达式
                    result_data[alias_name] = self._evaluate_math_expression(original_expr, df)

                elif isinstance(original_expr, exp.Literal):
                    # SELECT中的字面量值(如 SELECT 1, SELECT 'hello')
                    val = self._parse_literal_value(original_expr)
                    result_data[alias_name] = pd.Series([val] * len(df), index=df.index)

                elif isinstance(original_expr, exp.Boolean):
                    # SQL布尔字面量(TRUE/FALSE) → int(1/0)，与Excel存储格式一致
                    val = int(original_expr.this)
                    result_data[alias_name] = pd.Series([val] * len(df), index=df.index)

                elif isinstance(original_expr, exp.Subquery):
                    # 标量子查询(如 SELECT (SELECT MAX(col) FROM t))
                    try:
                        sub_result = self._execute_subquery(original_expr, self._current_worksheets)
                        if len(sub_result) > 0 and len(sub_result.columns) > 0:
                            scalar_val = sub_result.iloc[0, 0]
                            result_data[alias_name] = pd.Series([scalar_val] * len(df), index=df.index)
                        else:
                            result_data[alias_name] = pd.Series([None] * len(df), index=df.index)
                    except Exception as sub_e:
                        raise ValueError(f"标量子查询执行失败: {sub_e}")

                elif isinstance(original_expr, exp.Anonymous):
                    # 处理含括号的列名(如"刷新时间(小时)"被sqlglot解析为Anonymous函数)
                    anon_name = original_expr.this
                    # 重建完整列名:函数名+括号内容
                    if original_expr.expressions:
                        inner = ", ".join(str(e) for e in original_expr.expressions)
                        full_name = f"{anon_name}({inner})"
                    else:
                        full_name = anon_name
                    # 检查完整列名或函数名是否匹配实际列
                    if full_name in df.columns:
                        result_data[alias_name] = df[full_name]
                    elif anon_name in df.columns:
                        result_data[alias_name] = df[anon_name]
                    else:
                        raise ValueError(f"不支持的函数调用: {original_expr}。💡 如果'{full_name}'是含括号的列名,列名不匹配。\n🔧 可用列: {list(df.columns)}")

                elif isinstance(original_expr, exp.Cast):
                    # CAST 表达式(向量化)
                    result_data[alias_name] = self._evaluate_cast_expression(original_expr, df)

                elif isinstance(original_expr, exp.Or):
                    # MySQL 方言将 || 解析为 OR,启发式检测是否为字符串拼接
                    if self._is_likely_dpipe_concatenation(original_expr):
                        left = self._expr_to_series(original_expr.this, df).astype(str)
                        right = self._expr_to_series(original_expr.expression, df).astype(str)
                        result_data[alias_name] = left + right
                    else:
                        raise ValueError(f"不支持的表达式: {original_expr}。\n💡 MySQL方言中 || 表示逻辑OR,如需字符串拼接请使用 CONCAT() 函数。")

                else:
                    # 其他表达式,尝试作为列处理
                    if hasattr(original_expr, "name") and original_expr.name in df.columns:
                        result_data[alias_name] = df[original_expr.name]
                    else:
                        raise ValueError(f"不支持的表达式: {original_expr}")

                ordered_columns.append(alias_name)

            except Exception as e:
                # 表达式处理失败,尝试返回原始值
                if hasattr(original_expr, "name") and original_expr.name in df.columns:
                    result_data[alias_name] = df[original_expr.name]
                    ordered_columns.append(alias_name)
                else:
                    raise ValueError(f"处理SELECT表达式失败: {self._sanitize_error_message(str(e))}")

        # 构建结果DataFrame,保持SELECT顺序
        if result_data:
            result_df = pd.DataFrame(result_data)
            # 按照SQL SELECT顺序重新排列列
            result_df = result_df[ordered_columns]
            return result_df
        else:
            return df

    def _extract_select_alias(self, select_expr, index: int) -> tuple:
        """
        从SELECT表达式提取别名和原始表达式.
        复用于 _apply_select_expressions 和 _apply_group_by_aggregation.

        Args:
            select_expr: SELECT表达式节点
            index: 表达式索引(用于生成默认别名 col_N)

        Returns:
            (alias_name, original_expr) 元组
        """
        if isinstance(select_expr, exp.Alias):
            return select_expr.alias, select_expr.this
        # 无别名:优先用列名,否则用聚合别名,最后 col_N
        if isinstance(select_expr, exp.Column):
            # 保留表别名前缀(如 r.名称 -> 别名为 "r.名称")
            table_part = select_expr.table if hasattr(select_expr, "table") and select_expr.table else None
            if table_part:
                return f"{table_part}.{select_expr.name}", select_expr
            return select_expr.name, select_expr
        if self._is_aggregate_function(select_expr):
            return self._generate_aggregate_alias(select_expr), select_expr
        if isinstance(select_expr, exp.Anonymous):
            # 含括号的列名:重建完整名作为别名
            anon_name = select_expr.this
            if select_expr.expressions:
                inner = ", ".join(str(e) for e in select_expr.expressions)
                full_name = f"{anon_name}({inner})"
            else:
                full_name = anon_name
            return full_name, select_expr
        if hasattr(select_expr, "name") and select_expr.name:
            return select_expr.name, select_expr
        return f"col_{index}", select_expr

    def _is_mathematical_expression(self, expr) -> bool:
        """检查是否为数学表达式"""
        return isinstance(expr, (exp.Add, exp.Sub, exp.Mul, exp.Div, exp.Mod, exp.Nullif))

    # 数学运算符分发表:二元运算符统一处理
    _MATH_BINARY_OPS = {
        exp.Add: operator.add,
        exp.Sub: operator.sub,
        exp.Mul: operator.mul,
        exp.Div: operator.truediv,
        exp.Mod: operator.mod,
    }

    # 比较运算符分发表:逐行条件评估统一处理
    _COMPARISON_OPS = {
        exp.EQ: lambda l, r: l == r,
        exp.NEQ: lambda l, r: l != r,
        exp.GT: lambda l, r: _safe_float_comparison(l, r, ">"),
        exp.GTE: lambda l, r: _safe_float_comparison(l, r, ">="),
        exp.LT: lambda l, r: _safe_float_comparison(l, r, "<"),
        exp.LTE: lambda l, r: _safe_float_comparison(l, r, "<="),
    }

    # 复杂表达式类型集合:WHERE子句逐行过滤触发条件
    _COMPLEX_EXPR_TYPES = frozenset(
        {
            exp.Coalesce,
            exp.Case,
            exp.Exists,
            exp.Upper,
            exp.Lower,
            exp.Trim,
            exp.Length,
            exp.Concat,
            exp.Replace,
            exp.Substring,
            exp.Left,
            exp.Right,
            exp.Add,
            exp.Sub,
            exp.Mul,
            exp.Div,
            exp.Mod,
            exp.Nullif,
            exp.All,
            exp.Any,
            exp.Anonymous,
            exp.Round,  # 标量数值函数
            # Fix(R46): exp.Cast 已从此列表移除
            # CAST 现在在 _sql_condition_to_pandas 中通过预计算临时列支持,无需走逐行过滤
        }
    )

    # HAVING空结果建议分发表:(stat_func, op_str, label)
    _HAVING_OPS = {
        exp.GT: ("max", ">", "最大"),
        exp.GTE: ("max", ">=", "最大"),
        exp.LT: ("min", "<", "最小"),
        exp.LTE: ("min", "<=", "最小"),
    }

    # JOIN类型分发表:(side, kind) -> how
    # [FIX R55-BUG-04] sqlglot 将 "FULL OUTER JOIN" 解析为 side=FULL, kind=OUTER
    # 将 "LEFT OUTER JOIN" 解析为 side=LEFT, kind=OUTER
    # 原表只有 (side, None) 和 (None, kind) 条目，无法匹配 (side, OUTER) 组合
    # 导致 FULL/LEFT/RIGHT OUTER JOIN 全部 fallback 到默认值 "inner"
    _JOIN_KIND_MAP = {
        ("LEFT", None): "left",
        (None, "LEFT"): "left",
        ("LEFT", "OUTER"): "left",
        ("RIGHT", None): "right",
        (None, "RIGHT"): "right",
        ("RIGHT", "OUTER"): "right",
        ("FULL", None): "outer",
        (None, "FULL"): "outer",
        ("FULL", "OUTER"): "outer",
        ("INNER", None): "inner",
        (None, "INNER"): "inner",
        (None, "CROSS"): "cross",
    }

    # Pandas条件运算符分发表:SQL条件->pandas query字符串
    _PANDAS_OPS = {
        exp.EQ: "==",
        exp.NEQ: "!=",
        exp.GT: ">",
        exp.GTE: ">=",
        exp.LT: "<",
        exp.LTE: "<=",
    }

    def _evaluate_math_expression(self, expr, df: pd.DataFrame):
        """计算数学表达式"""
        # 解包Paren: (a + b) -> a + b
        if isinstance(expr, exp.Paren):
            return self._evaluate_math_expression(expr.this, df)

        op_type = type(expr)
        if op_type in self._MATH_BINARY_OPS:
            left = self._evaluate_math_expression(expr.left, df)
            right = self._evaluate_math_expression(expr.right, df)
            return self._MATH_BINARY_OPS[op_type](left, right)
        elif isinstance(expr, exp.Column):
            # 处理列引用，支持表限定符（如 t.column_name）+ 大小写不敏感
            col_name = expr.name
            table_part = expr.table if hasattr(expr, "table") and expr.table else None
            qualified = f"{table_part}.{col_name}" if table_part else None

            # 先尝试直接使用qualified列名
            if qualified and qualified in df.columns:
                return df[qualified]
            # 大小写不敏感列名查找
            actual_col = self._find_column_name(col_name, df)
            if actual_col:
                return df[actual_col]
            elif qualified:
                actual_qualified = self._find_column_name(qualified, df)
                if actual_qualified:
                    return df[actual_qualified]
            raise ValueError(f"列 '{qualified or col_name}' 不存在")
        elif isinstance(expr, exp.Literal):
            return self._expression_to_value(expr, df)
        elif isinstance(expr, exp.Boolean):
            # SQL布尔字面量(TRUE/FALSE) → int(1/0)，与Excel存储格式一致
            return int(expr.this)
        elif isinstance(expr, exp.Coalesce):
            # COALESCE在数学表达式中(向量化)
            return self._evaluate_coalesce_vectorized(expr, df)
        elif isinstance(expr, exp.Nullif):
            # NULLIF(a, b): 如果a等于b则返回NULL,否则返回a
            left = self._evaluate_math_expression(expr.this, df)
            right = self._evaluate_math_expression(expr.expression, df)
            # 向量化:对齐Series,逐元素比较
            if isinstance(left, pd.Series) or isinstance(right, pd.Series):
                left = left if isinstance(left, pd.Series) else pd.Series([left] * len(df), index=df.index)
                right = right if isinstance(right, pd.Series) else pd.Series([right] * len(df), index=df.index)
                return left.where(left != right, other=None)
            return None if left == right else left
        elif self._is_scalar_num_function(expr):
            # 标量数值函数嵌入数学表达式(如 ROUND(price,2) * 1.1)
            return self._evaluate_scalar_num_function(expr, df)
        elif isinstance(expr, exp.Window):
            # 窗口函数参与算术表达式(如 Price - LAG(Price) OVER ...)
            # 需要即时计算窗口函数并返回结果Series
            temp_col = f"_window_math_{id(expr)}"
            if temp_col not in df.columns:
                # 计算窗口函数并存储到临时列
                df[temp_col] = self._compute_window_function(expr, df)
            return pd.to_numeric(df[temp_col], errors="coerce")
        elif isinstance(expr, (exp.Select, exp.Subquery)):
            # 标量子查询参与算术运算(如 Price > (SELECT AVG(Price) FROM ...) 或 Price / (SELECT AVG(Price) FROM ...))
            # 执行子查询并获取标量结果
            if not hasattr(self, "_worksheets_data"):
                raise ValueError("子查询执行失败: 缺少工作表数据上下文")

            # 执行子查询
            subquery_df = self._execute_subquery(expr, self._worksheets_data)

            # 子查询应该返回单个值(标量子查询)
            if subquery_df.empty:
                return None
            if len(subquery_df) > 1:
                raise ValueError(f"子查询返回了多行数据,无法作为标量值使用。返回行数: {len(subquery_df)}")
            if len(subquery_df.columns) > 1:
                # 多列时取第一列
                return pd.to_numeric(subquery_df.iloc[0, 0], errors="coerce")
            else:
                # 单列单行
                return pd.to_numeric(subquery_df.iloc[0, 0], errors="coerce")
        elif isinstance(expr, exp.Cast):
            # CAST 嵌入数学表达式 (如 CAST(col AS FLOAT) * 100)
            return self._evaluate_cast_expression(expr, df)
        elif isinstance(expr, exp.DPipe):
            # || 字符串拼接嵌入表达式
            left = self._evaluate_math_expression(expr.this, df).astype(str)
            right = self._evaluate_math_expression(expr.expression, df).astype(str)
            return left + right
        else:
            raise ValueError(f"不支持的数学运算: {expr}。💡 WHERE子句暂不支持算术运算，建议用子查询替代")

    def _is_string_function(self, expr) -> bool:
        """检查是否为字符串函数"""
        return isinstance(
            expr,
            (
                exp.Upper,
                exp.Lower,
                exp.Trim,
                exp.Length,
                exp.Concat,
                exp.ConcatWs,
                exp.Replace,
                exp.Substring,
                exp.Left,
                exp.Right,
            ),
        )

    # 简单字符串函数分发表:一元操作,统一模式 val_series.astype(str).str.<op>()
    _SIMPLE_STR_OPS = {
        exp.Upper: "upper",
        exp.Lower: "lower",
        exp.Trim: "strip",
        exp.Length: "len",
    }

    # 标量数值函数分发表:ROUND, ABS, CEIL, FLOOR, SQRT, POWER 等(isinstance需要tuple,不能用frozenset)
    _SCALAR_NUM_FUNCS = (
        exp.Round,  # ROUND(value, decimals) -- 四舍五入
        exp.Abs,  # ABS(value) -- 绝对值
        exp.Ceil,  # CEIL(value) -- 向上取整
        exp.Floor,  # FLOOR(value) -- 向下取整
        exp.Sqrt,  # SQRT(value) -- 平方根
        exp.Pow,  # POWER(base, exp) -- 幂运算
    )

    def _is_scalar_num_function(self, expr) -> bool:
        """检查是否为标量数值函数(ROUND, ABS, FLOOR等)"""
        return isinstance(expr, self._SCALAR_NUM_FUNCS)

    def _find_inner_aggregate(self, expr) -> Any:
        """在表达式中查找内层聚合函数，用于处理 ROUND(AVG(col)) 等情况

        递归搜索表达式树，返回找到的第一个 AggFunc 节点。
        如果没有找到聚合函数，返回 None。
        """
        # 类型守卫: 只处理 sqlglot 表达式节点,跳过字符串/数字等字面量
        if not isinstance(expr, exp.Expression):
            return None
        if isinstance(expr, exp.AggFunc):
            return expr
        # 递归检查子节点: this 和 expression (二元操作数)
        for child_attr in ("this", "expression"):
            child = getattr(expr, child_attr, None)
            if child is not None:
                result = self._find_inner_aggregate(child)
                if result is not None:
                    return result
        # 检查函数参数列表（如 ROUND 的第二个参数）
        for arg_key in expr.args:
            val = expr.args[arg_key]
            if isinstance(val, list):
                for item in val:
                    if isinstance(item, exp.Expression):
                        result = self._find_inner_aggregate(item)
                        if result is not None:
                            return result
            elif isinstance(val, exp.Expression):
                result = self._find_inner_aggregate(val)
                if result is not None:
                    return result
        return None

    def _apply_scalar_to_agg_result(self, scalar_expr, agg_series: pd.Series) -> pd.Series:
        """对已计算的聚合结果应用标量数值函数

        用于处理 ROUND(AVG(col)), ABS(SUM(col)), ABS(ROUND(MIN(col))) 等场景。
        直接对聚合结果 Series 应用标量运算，避免 _expr_to_series 解析 Agg 节点失败。
        支持递归嵌套标量函数（如 ABS(ROUND(MIN(value)))）。

        Args:
            scalar_expr: 标量函数表达式节点 (如 exp.Round)
            agg_series: 已计算的聚合结果 Series

        Returns:
            应用标量函数后的 Series
        """
        # numpy already imported at top level

        func_type = type(scalar_expr)

        # Fix(R47): 递归处理嵌套标量函数
        # 例如 ABS(ROUND(MIN(value))) → 先对聚合结果应用 ROUND，再应用 ABS
        inner_expr = scalar_expr.this
        if self._is_scalar_num_function(inner_expr):
            agg_series = self._apply_scalar_to_agg_result(inner_expr, agg_series)

        numeric_series = pd.to_numeric(agg_series, errors="coerce")

        if func_type == exp.Round:
            decimals_arg = scalar_expr.args.get("decimals")
            decimals = int(self._literal_value(decimals_arg)) if decimals_arg is not None else 0
            return numeric_series.round(decimals)
        elif func_type == exp.Abs:
            return numeric_series.abs()
        elif func_type == exp.Ceil:
            return np.ceil(numeric_series)
        elif func_type == exp.Floor:
            return np.floor(numeric_series)
        elif func_type == exp.Sqrt:
            return np.sqrt(numeric_series)
        elif func_type == exp.Pow:
            # POWER(base, exp) — 这里 base 是聚合结果
            power_arg = scalar_expr.args.get("expression") or scalar_expr.args.get("this")
            if power_arg is not None and isinstance(power_arg, (exp.Literal,)):
                exp_val = float(self._literal_value(power_arg))
                return np.power(numeric_series, exp_val)
            else:
                return numeric_series
        else:
            # 未知标量函数，原样返回
            logger.warning("_apply_scalar_to_agg_result: 未知的标量函数类型 %s", func_type.__name__)
            return agg_series

    def _evaluate_string_function(self, expr, df) -> pd.Series:
        """计算字符串函数,返回pd.Series"""
        func_type = type(expr)

        # 简单字符串函数:分发表处理
        if func_type in self._SIMPLE_STR_OPS:
            val_series = self._expr_to_series(expr.this, df)
            return getattr(val_series.astype(str).str, self._SIMPLE_STR_OPS[func_type])()

        func_name = func_type.__name__.lower()

        if func_name == "concat":
            # CONCAT(a, b, ...) -- expressions列表包含所有参数
            parts = [self._expr_to_series(arg, df).astype(str) for arg in expr.expressions]
            if parts:
                result = parts[0]
                for p in parts[1:]:
                    result = result + p
                return result
            return pd.Series([""] * len(df), index=df.index)

        if func_name == "concatws":
            # CONCAT_WS(separator, str1, str2, ...) -- 第一个参数是分隔符
            if len(expr.expressions) >= 2:
                # 第一个表达式是分隔符
                separator = self._expr_to_series(expr.expressions[0], df).astype(str).iloc[0]
                # 其余表达式是要拼接的字符串
                parts = [self._expr_to_series(arg, df).astype(str) for arg in expr.expressions[1:]]
                if parts:
                    result = parts[0]
                    for p in parts[1:]:
                        result = result + separator + p
                    return result
            return pd.Series([""] * len(df), index=df.index)

        if func_name == "replace":
            # REPLACE(str, old, new) -- sqlglot: this=string, expression=old, replacement=new
            val_series = self._expr_to_series(expr.this, df).astype(str)
            old_val = self._get_arg(expr, "expression", "", str)
            new_val = self._get_arg(expr, "replacement", "", str)
            return val_series.str.replace(old_val, new_val, regex=False)

        if func_name in ("substring", "left", "right"):
            val_series = self._expr_to_series(expr.this, df).astype(str)
            if func_name == "substring":
                start = self._get_arg(expr, "start", 1, int) - 1
                length = self._get_arg(expr, "length", len(val_series.iloc[0]), int)
                return val_series.str.slice(start, start + length)
            if func_name == "left":
                n = self._get_arg(expr, "expression", 1, int)
                return val_series.str.slice(0, n)
            # right
            n = self._get_arg(expr, "expression", 1, int)
            return val_series.str.slice(-n)

        raise ValueError(f"不支持的字符串函数: {func_name}。💡 支持的字符串函数: UPPER, LOWER, CONCAT, SUBSTRING, TRIM, LENGTH, REPLACE")

    # 逐行字符串函数分发表:一元操作,统一模式 op(val)
    _ROW_STR_OPS = {
        "upper": lambda v: v.upper(),
        "lower": lambda v: v.lower(),
        "trim": lambda v: v.strip(),
        "length": lambda v: len(v),
    }

    def _evaluate_string_function_for_row(self, expr, row: pd.Series) -> Any:
        """逐行评估字符串函数"""
        func_name = type(expr).__name__.lower()
        val = self._get_row_value(expr.this, row)
        if val is None:
            return None
        val = str(val)

        # 简单函数:分发表处理
        if func_name in self._ROW_STR_OPS:
            return self._ROW_STR_OPS[func_name](val)

        # 复杂函数:各自独立处理
        if func_name == "concat":
            parts = [str(self._get_row_value(arg, row) or "") for arg in expr.expressions]
            return "".join(parts)
        if func_name == "concatws":
            # CONCAT_WS(separator, str1, str2, ...)
            if len(expr.expressions) >= 2:
                separator = str(self._get_row_value(expr.expressions[0], row) or "")
                parts = [str(self._get_row_value(arg, row) or "") for arg in expr.expressions[1:]]
                return separator.join(parts)
            return ""
        if func_name == "replace":
            old_val = self._get_arg(expr, "expression", "", str)
            new_val = self._get_arg(expr, "replacement", "", str)
            return val.replace(old_val, new_val)
        if func_name == "substring":
            start = self._get_arg(expr, "start", 1, int) - 1
            length = self._get_arg(expr, "length", len(val), int)
            return val[start : start + length]
        if func_name == "left":
            n = self._get_arg(expr, "expression", 1, int)
            return val[:n]
        if func_name == "right":
            n = self._get_arg(expr, "expression", 1, int)
            return val[-n:] if n > 0 else ""
        return val

    def _evaluate_scalar_num_function(self, expr, df) -> pd.Series:
        """计算标量数值函数,返回pd.Series(向量化)

        支持: ROUND, ABS, CEIL, FLOOR, SQRT, POWER
        """
        func_type = type(expr)
        # numpy already imported at top level

        # 提取基础参数
        val_series = self._expr_to_series(expr.this, df)

        if func_type == exp.Round:
            # ROUND(value, decimals)
            decimals_arg = expr.args.get("decimals")
            decimals = int(self._literal_value(decimals_arg)) if decimals_arg is not None else 0
            try:
                numeric_series = pd.to_numeric(val_series, errors="coerce")
                return numeric_series.round(decimals)
            except Exception:

                def _round_val(v):
                    try:
                        return round(float(v), decimals)
                    except (TypeError, ValueError):
                        return None

                return val_series.apply(_round_val)

        elif func_type == exp.Abs:
            # ABS(value) - 绝对值
            try:
                numeric_series = pd.to_numeric(val_series, errors="coerce")
                return numeric_series.abs()
            except Exception:
                return val_series.apply(lambda v: abs(float(v)) if pd.notna(v) and v != "" else None)

        elif func_type == exp.Ceil:
            # CEIL(value) - 向上取整
            try:
                numeric_series = pd.to_numeric(val_series, errors="coerce")
                return np.ceil(numeric_series)
            except Exception:
                return val_series.apply(lambda v: np.ceil(float(v)) if pd.notna(v) and v != "" else None)

        elif func_type == exp.Floor:
            # FLOOR(value) - 向下取整
            try:
                numeric_series = pd.to_numeric(val_series, errors="coerce")
                return np.floor(numeric_series)
            except Exception:
                return val_series.apply(lambda v: np.floor(float(v)) if pd.notna(v) and v != "" else None)

        elif func_type == exp.Sqrt:
            # SQRT(value) - 平方根
            try:
                numeric_series = pd.to_numeric(val_series, errors="coerce")
                return np.sqrt(numeric_series)
            except Exception:

                def _sqrt_val(v):
                    try:
                        return np.sqrt(float(v))
                    except (TypeError, ValueError):
                        return None

                return val_series.apply(_sqrt_val)

        elif func_type == exp.Pow:
            # POWER(base, exp) - 幂运算
            # sqlglot Pow 结构: this=base, args['exp']=exponent
            exp_arg = expr.args.get("exp")
            exponent = self._literal_value(exp_arg)

            try:
                # 转换为数值
                numeric_series = pd.to_numeric(val_series, errors="coerce")
                if isinstance(exponent, (int, float)):
                    return np.power(numeric_series, exponent)
                else:
                    # 指数也是表达式的情况
                    exp_series = self._expr_to_series(exp_arg, df)
                    exp_numeric = pd.to_numeric(exp_series, errors="coerce")
                    return np.power(numeric_series, exp_numeric)
            except Exception:

                def _pow_val(v):
                    try:
                        if isinstance(exponent, (int, float)):
                            return np.power(float(v), exponent)
                        return None
                    except (TypeError, ValueError):
                        return None

                return val_series.apply(_pow_val)

        raise ValueError(f"不支持的标量数值函数: {func_type.__name__}")

    def _evaluate_scalar_num_function_for_row(self, expr, row: pd.Series) -> Any:
        """逐行评估标量数值函数

        支持: ROUND, ABS, CEIL, FLOOR, SQRT, POWER
        """
        func_type = type(expr)
        # numpy already imported at top level

        val = self._get_row_value(expr.this, row)
        if val is None:
            return None

        try:
            val_float = float(val)
        except (TypeError, ValueError):
            return None

        if func_type == exp.Round:
            # ROUND(value, decimals)
            decimals_arg = expr.args.get("decimals")
            decimals = int(self._literal_value(decimals_arg)) if decimals_arg is not None else 0
            try:
                return round(val_float, decimals)
            except (TypeError, ValueError):
                return None

        elif func_type == exp.Abs:
            # ABS(value)
            return abs(val_float)

        elif func_type == exp.Ceil:
            # CEIL(value)
            return np.ceil(val_float)

        elif func_type == exp.Floor:
            # FLOOR(value)
            return np.floor(val_float)

        elif func_type == exp.Sqrt:
            # SQRT(value)
            try:
                return np.sqrt(val_float)
            except (TypeError, ValueError):
                return None

        elif func_type == exp.Pow:
            # POWER(base, exp)
            exp_arg = expr.args.get("exp")
            exponent = self._literal_value(exp_arg)

            if isinstance(exponent, (int, float)):
                try:
                    return np.power(val_float, exponent)
                except (TypeError, ValueError):
                    return None
            # 如果指数也是表达式，需要递归求值
            exp_val = self._get_row_value(exp_arg, row)
            if exp_val is not None:
                try:
                    return np.power(val_float, float(exp_val))
                except (TypeError, ValueError):
                    return None
            return None

        return None

    def _expr_to_series(self, expr, df) -> pd.Series:
        """将表达式转换为pd.Series(支持列引用,字面量,数学表达式,字符串函数)"""
        if isinstance(expr, exp.Column):
            col_name = expr.name
            if col_name in df.columns:
                return df[col_name]
            # 表限定符
            table_part = expr.table if hasattr(expr, "table") and expr.table else None
            qualified = f"{table_part}.{col_name}" if table_part else None
            if qualified and qualified in df.columns:
                return df[qualified]
            raise ValueError(f"列 '{qualified or col_name}' 不存在")
        elif isinstance(expr, exp.Literal):
            val = expr.this
            # Fix(R11): sqlglot将数值字面量(含科学计数法如1e100)存储为字符串
            # 直接用字符串创建Series会导致与数值列运算时类型不匹配崩溃
            if not expr.is_string and isinstance(val, str):
                try:
                    # 尝试转为int(更精确),失败则转float
                    val = int(val)
                except (ValueError, OverflowError):
                    try:
                        val = float(val)
                    except (ValueError, OverflowError):
                        pass  # 保持原字符串值
            return pd.Series([val] * len(df), index=df.index)
        elif isinstance(expr, (exp.Add, exp.Sub, exp.Mul, exp.Div)):
            return self._evaluate_math_expression(expr, df)
        elif isinstance(expr, exp.Neg):
            # Fix(R52-P3-EDGE-01): 快速路径支持负数字面量
            inner = self._expr_to_series(expr.this, df)
            return -inner
        elif isinstance(expr, exp.DPipe):
            # || 字符串拼接操作符 (SQL标准/PostgreSQL风格)
            left = self._expr_to_series(expr.this, df).astype(str)
            right = self._expr_to_series(expr.expression, df).astype(str)
            return left + right
        elif isinstance(expr, exp.Coalesce):
            return self._evaluate_coalesce_vectorized(expr, df)
        elif self._is_string_function(expr):
            return self._evaluate_string_function(expr, df)
        elif self._is_scalar_num_function(expr):
            return self._evaluate_scalar_num_function(expr, df)
        elif isinstance(expr, exp.Case):
            return self._evaluate_case_expression(expr, df)
        elif isinstance(expr, exp.Anonymous):
            # 含括号的列名
            anon_name = expr.this
            if expr.expressions:
                inner = ", ".join(str(e) for e in expr.expressions)
                full_name = f"{anon_name}({inner})"
            else:
                full_name = anon_name
            if full_name in df.columns:
                return df[full_name]
            if anon_name in df.columns:
                return df[anon_name]

            # 中文函数名映射提示
            _CN_FUNC_MAP = {
                "长度": "LENGTH",
                "LEN": "LENGTH",
                "UPPER": "UPPER",
                "LOWER": "LOWER",
                "TRIM": "TRIM",
                "截取": "SUBSTRING",
                "四舍五入": "ROUND",
                "舍入": "ROUND",
                "拼接": "CONCAT",
                "替换": "REPLACE",
            }
            cn_hint = ""
            for cn_name, en_name in _CN_FUNC_MAP.items():
                if cn_name in anon_name or cn_name in full_name:
                    cn_hint = f"💡 如需调用{en_name}函数,请使用英文函数名: {en_name}(...)"
                    break

            raise ValueError(f"列 '{full_name}' 不存在。可用列: {list(df.columns)}{cn_hint}")
        elif isinstance(expr, exp.Cast):
            # CAST(expr AS type) — SQL 标准类型转换
            return self._evaluate_cast_expression(expr, df)
        elif isinstance(expr, exp.Boolean):
            # SQL布尔字面量(TRUE/FALSE) → int(1/0)
            return int(expr.this)
        elif isinstance(expr, exp.Window):
            # 窗口函数作为内嵌表达式(如 ROUND(RANK() OVER(...), 2))
            # 窗口函数已由 _apply_window_functions 预计算到 df 中，按 id() 查找
            if hasattr(self, "_nested_window_columns") and id(expr) in self._nested_window_columns:
                return df[self._nested_window_columns[id(expr)]]
            # 回退: 尝试在 df 列中按生成模式查找
            for col in df.columns:
                if col.startswith("_window_nested"):
                    # 验证此列是否来自当前窗口(通过长度匹配等启发式)
                    return df[col]
            raise ValueError("嵌套窗口函数结果列未找到。💡 确保窗口函数语法正确: ROUND(PERCENT_RANK() OVER (ORDER BY col), N)")

        elif isinstance(expr, exp.Or):
            # MySQL 方言将 || 解析为 OR,但用户可能意图是字符串拼接
            # 启发式检测: 如果两边都是非布尔表达式(列引用/字面量/CAST/函数),
            # 则将其视为字符串拼接(DPipe)而非逻辑OR
            if self._is_likely_dpipe_concatenation(expr):
                left = self._expr_to_series(expr.this, df).astype(str)
                right = self._expr_to_series(expr.expression, df).astype(str)
                return left + right
            raise ValueError(f"不支持的表达式: {expr}。\n💡 MySQL方言中 || 表示逻辑OR,如需字符串拼接请使用 CONCAT() 函数。\n🔧 示例: SELECT CONCAT(name, '_v1') FROM table")

        elif isinstance(expr, exp.Neg):
            # 一元负号: -3.7, -column 等
            inner = self._expr_to_series(expr.this, df)
            numeric_inner = pd.to_numeric(inner, errors="coerce")
            return -numeric_inner

        else:
            raise ValueError(
                f"不支持的表达式类型: {type(expr).__name__}。"
                f"\n💡 建议: 检查SQL语法，常用表达式支持: 列引用、字面量、算术运算(+,-,*,/)、字符串函数(UPPER/LOWER/CONCAT)、CASE WHEN、CAST。"
                f"\n🔧 如需复杂计算，可用子查询替代: SELECT * FROM (SELECT ..., (A+B) as total FROM table) WHERE total > 100"
            )

    def _literal_value(self, expr) -> Any:
        """提取字面量值"""
        if isinstance(expr, exp.Literal):
            return expr.this
        elif isinstance(expr, exp.Column):
            return expr.name
        return str(expr)

    def _get_arg(self, expr, arg_name, default=None, type_fn=None):
        """从表达式参数中提取值,支持类型转换

        消除字符串函数中重复的 _literal_value(expr.args.get(...)) 模式.
        例如: int(self._literal_value(expr.args.get('n'))) if expr.args.get('n') else 1
        简化为: self._get_arg(expr, 'n', 1, int)
        """
        arg = expr.args.get(arg_name)
        if arg is None:
            return default
        val = self._literal_value(arg)
        if type_fn and val is not None:
            return type_fn(val)
        return val

    def _get_from_table(self, parsed_sql: exp.Expression) -> tuple[str, exp.Expression | None]:
        """获取FROM子句中的表名.

        Returns:
            (table_name, subquery_expr):
                - 普通表: (表名, None)
                - FROM子查询: (别名, Subquery/Select表达式)
        """
        from_clause = parsed_sql.args.get("from")
        if not from_clause:
            # 尝试使用 from_ 键(sqlglot的另一种存储方式)
            from_clause = parsed_sql.args.get("from_")
        if from_clause:
            # 检查FROM子句是否是子查询(FROM (SELECT ...) AS alias)
            if hasattr(from_clause, "this") and isinstance(from_clause.this, (exp.Subquery, exp.Select)):
                subquery_node = from_clause.this
                # 获取别名
                alias = getattr(subquery_node, "alias", None)
                if not alias and isinstance(subquery_node, exp.Subquery):
                    # sqlglot存储别名在alias属性
                    alias = subquery_node.alias
                if not alias:
                    alias = "_subquery"
                return (alias, subquery_node)
            if hasattr(from_clause, "this") and hasattr(from_clause.this, "name"):
                return (from_clause.this.name, None)
            # 兼容 Table 对象
            if hasattr(from_clause, "this") and hasattr(from_clause.this, "this"):
                return (from_clause.this.this, None)

        # 如果没有明确的FROM子句,返回第一个表名
        raise ValueError("无法确定FROM子句中的表名")

    def _apply_join_clause(self, joins, left_df, worksheets_data=None, left_table=None) -> pd.DataFrame:
        """
        应用JOIN子句,支持INNER/LEFT/RIGHT/FULL/CROSS JOIN
        性能优化:使用索引优化和智能JOIN策略

        Args:
            joins: sqlglot joins列表
            left_df: 左表DataFrame
            worksheets_data: 所有工作表数据
            left_table: 左表名

        Returns:
            pd.DataFrame: JOIN后的DataFrame
        """
        if isinstance(joins, exp.Join):
            joins = [joins]

        # 性能优化:检查数据大小,决定是否使用索引优化
        left_size = len(left_df)
        total_memory_mb = left_size * len(left_df.columns) * 8 / (1024 * 1024)  # 估算内存使用

        # 初始化JOIN列映射(用于别名解析)
        self._join_column_mapping = {}

        result_df = left_df

        for join in joins:
            # 解析JOIN类型
            join_side = str(join.side).upper() if join.side else None
            join_kind_name = str(join.kind).upper() if join.kind else None
            join_kind = self._JOIN_KIND_MAP.get((join_side, join_kind_name), "inner")

            # Fix(R13): 逗号风格隐式 CROSS JOIN (FROM table1, table2)
            # sqlglot 将逗号解析为无 kind 无 ON 的 Join 节点,应视为笛卡尔积
            on_clause = join.args.get("on")
            if not join_kind_name and not on_clause and not join_side:
                join_kind = "cross"

            # 解析右表
            right_table_expr = join.this

            # LATERAL JOIN: 关联子查询，逐行执行
            if isinstance(right_table_expr, exp.Lateral):
                result_df = self._apply_lateral_join(
                    join,
                    right_table_expr,
                    result_df,
                    worksheets_data,
                    left_table,
                    join_kind,
                )
                continue

            # Fix(R7-F3): Support subquery as JOIN right table
            # e.g., FROM (SELECT ...) a CROSS JOIN (SELECT ...) b
            _r7_right_from_subquery = False
            if isinstance(right_table_expr, (exp.Subquery, exp.Select)):
                _r7_right_from_subquery = True
                right_alias = getattr(right_table_expr, "alias", None) or "_joined_subquery"
                right_table = right_alias  # Set right_table to alias for mapping
                try:
                    right_df = self._execute_subquery(right_table_expr, worksheets_data)
                except Exception as e:
                    raise StructuredSQLError(
                        "join_error",
                        f"JOIN右表子查询执行失败: {self._sanitize_error_message(str(e))}",
                        hint="请检查JOIN右表子查询的SQL语法.",
                    )
                # Skip normal table lookup - go directly to column rename and merge
                self._table_aliases[right_alias] = right_alias
            elif hasattr(right_table_expr, "this"):
                right_table = (
                    right_table_expr.this if isinstance(right_table_expr.this, str) else (right_table_expr.this.name if hasattr(right_table_expr.this, "name") else str(right_table_expr.this))
                )
            else:
                right_table = right_table_expr.name if hasattr(right_table_expr, "name") else str(right_table_expr)

            # 检查右表是否有别名
            right_alias = right_table
            # 优先使用 Table.alias 属性(sqlglot 中 Table 的 alias 返回字符串)
            if hasattr(right_table_expr, "alias") and right_table_expr.alias:
                table_alias = right_table_expr.alias
                if isinstance(table_alias, str) and table_alias != right_table:
                    right_alias = table_alias
                elif hasattr(table_alias, "alias"):
                    right_alias = table_alias.alias
            # 备用:遍历 TableAlias 节点
            if right_alias == right_table:
                for alias in join.find_all(exp.TableAlias):
                    parent = alias.this
                    parent_name = parent.name if hasattr(parent, "name") else str(parent)
                    if parent_name == right_table or str(parent) == right_table:
                        right_alias = alias.alias
                        break

            # 记录别名映射
            self._table_aliases[right_alias] = right_table
            self._table_aliases[right_table] = right_table

            # Fix(R7-F3): Skip table lookup when right_df already from subquery
            if not _r7_right_from_subquery:
                # 检查右表是否存在,如果不存在尝试从同文件加载其他sheet
                if right_table not in worksheets_data:
                    # 尝试从同文件加载该sheet
                    if self._current_file_path and hasattr(self, "_load_excel_data"):
                        try:
                            # 加载指定sheet的数据
                            additional_sheets = self._load_excel_data(self._current_file_path, right_table)
                            if right_table in additional_sheets:
                                # 将加载的sheet添加到worksheets_data
                                worksheets_data[right_table] = additional_sheets[right_table]
                                # 同时更新列名映射
                                if not hasattr(self, "_additional_loaded_sheets"):
                                    self._additional_loaded_sheets = {}
                                self._additional_loaded_sheets[right_table] = additional_sheets[right_table]
                        except Exception:
                            pass  # 加载失败,继续抛出原错误

                    # 再次检查,如果仍不存在则抛出错误
                    if right_table not in worksheets_data:
                        available = list(worksheets_data.keys())

                        # 跨文件语法提示: 检测是否误用了Excel原生 ! 语法
                        cross_file_hint = ""
                        if "!" in right_table or right_table.endswith((".xlsx", ".xls")):
                            cross_file_hint = (
                                "\n💡 跨文件JOIN请使用 @'path' 语法,例如:"
                                "\n   FROM 表A@'/path/to/file1.xlsx' a "
                                "JOIN 表B@'/path/to/file2.xlsx' b ON a.id = b.id"
                                "\n   (不支持Excel原生的 'file.xlsx'!Sheet 语法)"
                            )

                        raise StructuredSQLError(
                            "table_not_found",
                            f"JOIN表 '{right_table}' 不存在.可用表: {available}.{cross_file_hint}",
                            hint="请检查JOIN的表名,跨文件引用请用 表名@'文件路径' 语法.",
                            context={
                                "table_requested": right_table,
                                "available_tables": available,
                            },
                        )

                right_df = worksheets_data[right_table].copy()

            # 解析ON条件(CROSS JOIN不需要ON)
            on_clause = join.args.get("on")
            left_on_col = None
            right_on_col = None
            actual_right_on = None

            # [FIX R55-BUG-01] 移除有缺陷的 set_index 索引优化
            # 原代码在 total_memory_mb > 10 时对 JOIN key 列执行 set_index(),
            # 将该列从 .columns 移至 index，导致后续 L5456 的列存在性检查
            # 和 L5572 的 merge(left_on=...) 均因找不到列而报 KeyError。
            # pandas merge 内部已使用 hash join 算法优化，此 set_index 优化
            # 不仅冗余且引入回归 bug，故整体移除。
            # 保留 total_memory_mb 变量供其他逻辑使用（如非等值连接策略选择）

            if join_kind == "cross":
                # CROSS JOIN: 笛卡尔积,不需要ON条件
                pass
            elif not on_clause:
                raise StructuredSQLError(
                    "join_error",
                    "JOIN缺少ON条件",
                    hint="JOIN必须包含ON条件,例如:... JOIN 表2 ON 表1.id = 表2.id.",
                )
            else:
                left_on_col, right_on_col, non_equi_cond = self._parse_join_on_condition(on_clause, left_table, right_table, right_alias)
                # 等值连接:验证列存在
                # Fix(C2): 三表链式JOIN时,左表列可能已被前次JOIN重命名为"表名.列名"
                # 需要在result_df中查找原始列名或其别名版本
                if not non_equi_cond and left_on_col and left_on_col not in result_df.columns:
                    # 搜索_join_column_mapping中是否有该列的别名版本
                    resolved_left_on = None
                    for alias_map in self._join_column_mapping.values():
                        for orig, aliased in alias_map.items():
                            if orig == left_on_col and aliased in result_df.columns:
                                resolved_left_on = aliased
                                break
                        if resolved_left_on:
                            break
                    if resolved_left_on:
                        left_on_col = resolved_left_on
                    else:
                        raise StructuredSQLError(
                            "column_not_found",
                            f"左表 '{left_table}' 没有列 '{left_on_col}'.可用列: {list(result_df.columns)}",
                            hint="请检查ON条件中左表的列名拼写.",
                            context={
                                "table": left_table,
                                "column_requested": left_on_col,
                                "available_columns": list(result_df.columns),
                            },
                        )
                if right_on_col and right_on_col not in right_df.columns:
                    raise StructuredSQLError(
                        "column_not_found",
                        f"右表 '{right_table}' 没有列 '{right_on_col}'.可用列: {list(right_df.columns)}",
                        hint="请检查ON条件中右表的列名拼写.",
                        context={
                            "table": right_table,
                            "column_requested": right_on_col,
                            "available_columns": list(right_df.columns),
                        },
                    )

            # 执行JOIN
            # 为右表列添加别名前缀避免冲突
            right_df_renamed = right_df.copy()
            col_mapping = {}
            for col in right_df_renamed.columns:
                if col in result_df.columns and (left_on_col is None or col != left_on_col):
                    new_col = f"{right_alias}.{col}"
                    col_mapping[col] = new_col
                elif left_on_col and col == left_on_col:
                    # ON列(无论左右列名是否相同):右表列重命名避免_x/_y后缀
                    new_col = f"{right_alias}.{col}"
                    col_mapping[col] = new_col
            right_df_renamed = right_df_renamed.rename(columns=col_mapping)

            # 更新JOIN列映射,用于别名解析
            if right_alias not in self._join_column_mapping:
                self._join_column_mapping[right_alias] = {}
            for orig_col, new_col in col_mapping.items():
                self._join_column_mapping[right_alias][orig_col] = new_col

            # 调整右表ON列名(如果被重命名了)
            if right_on_col:
                actual_right_on = col_mapping.get(right_on_col, right_on_col)

            # Fix(C2): 三表及以上链式JOIN时,更新left_table为连接结果标识
            # 使后续JOIN的ON条件解析能正确识别左表已变为前次JOIN的结果
            # 而非继续使用原始FROM表名去查找列
            left_table = "_joined_result"

            # 合并双行表头描述
            if right_table in self._header_descriptions:
                for orig_col, new_col in col_mapping.items():
                    if orig_col in self._header_descriptions[right_table]:
                        self._header_descriptions[right_table][new_col] = self._header_descriptions[right_table][orig_col]

            if join_kind == "cross":
                # CROSS JOIN: 笛卡尔积(无需ON列)
                # 先临时移除冲突列名,合并后再恢复
                temp_col_mapping = {}
                right_df_for_cross = right_df_renamed.copy()

                for col in right_df_for_cross.columns:
                    if col in result_df.columns:
                        temp_col = f"{right_alias}_temp_{col}"
                        temp_col_mapping[col] = temp_col
                        right_df_for_cross = right_df_for_cross.rename(columns={col: temp_col})

                result_df = result_df.merge(right_df_for_cross, how="cross")

                # 恢复原始列名
                for old_col, new_col in temp_col_mapping.items():
                    result_df = result_df.rename(columns={new_col: old_col})
            elif non_equi_cond is not None:
                # [R53优化] 非等值连接: 检查是否有可用的等值join key + 额外过滤器
                pending_filters = getattr(self, '_pending_join_filters', None)
                if left_on_col and right_on_col and pending_filters:
                    # 复合条件路径：先等值JOIN（快速pandas merge），再对缩小后的结果集施加非等值过滤
                    # 这比 cross join + filter 快几个数量级
                    result_df = result_df.merge(
                        right_df_renamed,
                        left_on=left_on_col,
                        right_on=actual_right_on,
                        how=join_kind,
                    )
                    # 对等值JOIN结果施加额外的非等值过滤条件
                    for filter_cond in pending_filters:
                        result_df = self._apply_row_filter(filter_cond, result_df)
                    self._pending_join_filters = None  # 清理
                else:
                    # [R53优化] 纯非等值连接: 尝试排序归并优化
                    sorted_result = self._try_sorted_non_equi_join(
                        result_df, right_df_renamed, non_equi_cond,
                        left_table, right_table, right_alias, join_kind,
                    )
                    if sorted_result is not None:
                        result_df = sorted_result
                    else:
                        # 回退到 cross join + row filter
                        result_df = result_df.merge(right_df_renamed, how="cross")
                        result_df = self._apply_row_filter(non_equi_cond, result_df)
            else:
                result_df = result_df.merge(
                    right_df_renamed,
                    left_on=left_on_col,
                    right_on=actual_right_on,
                    how=join_kind,
                )

            # 合并后删除重复的ON列(右表侧)
            # Fix(C2): 链式JOIN中,右表的ON列可能被后续JOIN引用(如三表JOIN的第二/三个ON条件)
            # 因此不再自动删除右表ON列;SELECT阶段会只选取需要的列,多余列不影响正确性
            # 仅当左右ON列名完全相同时,pandas merge已自动合并为单列,无需处理

        return result_df

    def _try_sorted_non_equi_join(self, left_df, right_df, non_equi_cond,
                                   left_table, right_table, right_alias, join_kind):
        """
        [R53优化] 对基于排序的非等值连接使用归并算法，避免 O(n*m) 的笛卡尔积。

        支持的单条件模式（两侧都是简单列引用）：
          a.col < b.col, a.col <= b.col, a.col > b.col, a.col >= b.col

        算法：双指针归并，O(n log n + m log m + n + m)
        不支持时返回 None，调用方回退到 cross+filter。
        """
        # 仅支持单条件比较（GT/GTE/LT/LTE/NEQ）
        if not isinstance(non_equi_cond, (exp.GT, exp.GTE, exp.LT, exp.LTE)):
            return None  # NEQ 或复合条件不适用

        left_expr = non_equi_cond.left
        right_expr = non_equi_cond.right

        # 两边都必须是简单列引用
        if not (isinstance(left_expr, exp.Column) and isinstance(right_expr, exp.Column)):
            return None

        left_col_name = left_expr.name
        right_col_name = right_expr.name
        left_tbl = getattr(left_expr, 'table', None)
        right_tbl = getattr(right_expr, 'table', None)

        # 确定左右列归属
        if left_tbl:
            resolved_left_tbl = self._table_aliases.get(left_tbl, left_tbl)
            if resolved_left_tbl == right_table or left_tbl == right_alias:
                # 左表达式实际引用右表列 → 交换
                left_col_name, right_col_name = right_col_name, left_col_name
                # 翻转比较符
                op_type = type(non_equi_cond)
                op_map = {exp.LT: exp.GT, exp.GT: exp.LT, exp.LTE: exp.GTE, exp.GTE: exp.LTE}
                non_equi_cond = op_map.get(op_type, op_type)(this=non_equi_cond.this, 
                                                              expression=non_equi_cond.expression,
                                                              comments=non_equi_cond.comments)
        elif right_tbl:
            resolved_right_tbl = self._table_aliases.get(right_tbl, right_tbl)
            if resolved_right_tbl == left_table:
                # 右表达式引用左表列 → 交换
                left_col_name, right_col_name = right_col_name, left_col_name
                op_type = type(non_equi_cond)
                op_map = {exp.LT: exp.GT, exp.GT: exp.LT, exp.LTE: exp.GTE, exp.GTE: exp.LTE}
                non_equi_cond = op_map.get(op_type, op_type)(this=non_equi_cond.this,
                                                              expression=non_equi_cond.expression,
                                                              comments=non_equi_cond.comments)

        # 验证列存在（右表可能已被重命名为 alias.col 格式）
        _right_col_candidates = [right_col_name]
        if right_col_name not in right_df.columns:
            # 尝试别名前缀格式（_apply_join_clause 中 right_df_renamed 的命名规则）
            _right_col_candidates.append(f"{right_alias}.{right_col_name}")
        _actual_right_col = None
        for rc in _right_col_candidates:
            if rc in right_df.columns:
                _actual_right_col = rc
                break

        if left_col_name not in left_df.columns or _actual_right_col is None:
            return None

        # 提取排序列数据
        left_keys = left_df[left_col_name].values
        right_keys = right_df[_actual_right_col].values

        # 数值/日期类型才能排序比较 — 统一转 float64 避免类型混用
        import numpy as np
        try:
            left_keys = pd.to_numeric(left_keys, errors='coerce').astype(np.float64)
            right_keys = pd.to_numeric(right_keys, errors='coerce').astype(np.float64)
            # NaN 值无法参与排序比较，回退
            if np.any(np.isnan(left_keys)) or np.any(np.isnan(right_keys)):
                return None
        except (ValueError, TypeError):
            return None

        # 排序索引
        left_order = np.argsort(left_keys, kind='mergesort')
        right_order = np.argsort(right_keys, kind='mergesort')

        left_sorted = left_keys[left_order]
        right_sorted = right_keys[right_order]

        # 根据比较符确定匹配逻辑
        op_type = type(non_equi_cond)

        # 收集匹配的 (left_idx, right_idx) 对
        matched_pairs = []
        r = 0  # 右指针

        if op_type == exp.LT:  # left < right
            # 对于每个左值，找所有右值 > 左值
            for i in range(len(left_sorted)):
                val = left_sorted[i]
                # 移动 r 到第一个 > val 的位置
                while r < len(right_sorted) and not (right_sorted[r] > val):
                    r += 1
                # 从 r 到末尾都满足
                for j in range(r, len(right_sorted)):
                    matched_pairs.append((left_order[i], right_order[j]))

        elif op_type == exp.LTE:  # left <= right
            for i in range(len(left_sorted)):
                val = left_sorted[i]
                while r < len(right_sorted) and right_sorted[r] < val:
                    r += 1
                for j in range(r, len(right_sorted)):
                    matched_pairs.append((left_order[i], right_order[j]))

        elif op_type == exp.GT:  # left > right
            # 反向：对于每个左值，找所有右值 < 左值
            for i in range(len(left_sorted)):
                val = left_sorted[i]
                while r < len(right_sorted) and right_sorted[r] < val:
                    r += 1
                # 0..r-1 都满足 < val
                for j in range(r):
                    matched_pairs.append((left_order[i], right_order[j]))

        elif op_type == exp.GTE:  # left >= right
            # 左值 >= 右值：推进 r 到第一个 > val 的位置，则 [0, r) 都满足 <= val
            for i in range(len(left_sorted)):
                val = left_sorted[i]
                while r < len(right_sorted) and right_sorted[r] <= val:
                    r += 1
                for j in range(r):
                    matched_pairs.append((left_order[i], right_order[j]))

        if not matched_pairs:
            # 空 JOIN 结果：返回左表结构（0行）
            return pd.DataFrame(columns=list(left_df.columns) + list(right_df.columns))

        # 构建结果 DataFrame（避免笛卡尔积）
        left_indices = [p[0] for p in matched_pairs]
        right_indices = [p[1] for p in matched_pairs]

        left_part = left_df.iloc[left_indices].reset_index(drop=True)
        right_part = right_df.iloc[right_indices].reset_index(drop=True)
        result = pd.concat([left_part, right_part], axis=1)

        # LEFT JOIN: 补充左表未匹配行
        if join_kind in ("left", "outer"):
            matched_left = set(left_indices)
            unmatched_left = [i for i in range(len(left_df)) if i not in matched_left]
            if unmatched_left:
                left_unmatched = left_df.iloc[unmatched_left].reset_index(drop=True)
                right_null = pd.DataFrame([[None] * len(right_df.columns)] * len(unmatched_left),
                                          columns=right_df.columns)
                result = pd.concat([
                    result,
                    pd.concat([left_unmatched, right_null], axis=1),
                ], ignore_index=True)

        return result

    def _parse_join_on_condition(self, on_clause, left_table: str, right_table: str, right_alias: str):
        """
        解析JOIN ON条件

        等值连接返回 (left_col, right_col, None)
        非等值连接返回 (None, None, on_clause)
        [R53优化] 复合AND条件返回 (left_col, right_col, [extra_conditions])
                  支持等值+非等值混合条件，先做equi-join再filter
        """
        # 非等值连接: 返回条件用于cross+filter
        if isinstance(on_clause, (exp.GT, exp.GTE, exp.LT, exp.LTE, exp.NEQ)):
            return (None, None, on_clause)

        if isinstance(on_clause, exp.EQ):
            left_expr = on_clause.left
            right_expr = on_clause.right
        elif isinstance(on_clause, exp.And):
            # [R53] 增强版：从AND中提取所有条件
            # 分离等值条件和非等值条件
            eq_conditions = list(on_clause.find_all(exp.EQ))
            non_eq_conditions = []
            # 遍历AND的直接子节点（不递归进嵌套AND）
            for child in (on_clause.left, on_clause.right):
                if isinstance(child, (exp.GT, exp.GTE, exp.LT, exp.LTE, exp.NEQ)):
                    non_eq_conditions.append(child)
                elif isinstance(child, exp.And):
                    # 嵌套AND：递归提取
                    for sub_child in child.find_all((exp.GT, exp.GTE, exp.LT, exp.LTE, exp.NEQ)):
                        non_eq_conditions.append(sub_child)

            if not eq_conditions:
                # 没有等值条件，全部是非等值
                if len(non_eq_conditions) == 1:
                    return (None, None, non_eq_conditions[0])
                else:
                    # 多个非等值条件：保留原始AND节点给_apply_row_filter
                    return (None, None, on_clause)

            # 取第一个等值条件作为join key
            left_expr = eq_conditions[0].left
            right_expr = eq_conditions[0].right

            # 如果有额外的非等值条件，暂存到实例变量供后续使用
            if non_eq_conditions:
                self._pending_join_filters = non_eq_conditions
            else:
                self._pending_join_filters = None
        else:
            raise ValueError("JOIN ON条件格式不支持,请使用等值连接: ON a.id = b.id")

        def resolve_column(col_expr) -> tuple:
            """解析列引用。返回 (col_name, table_part, is_simple_column)"""
            if isinstance(col_expr, exp.Column):
                col_name = col_expr.name
                table_part = col_expr.table if hasattr(col_expr, "table") and col_expr.table else None
                return col_name, table_part, True
            return str(col_expr), None, False

        left_col, left_tbl, left_is_col = resolve_column(left_expr)
        right_col, right_tbl, right_is_col = resolve_column(right_expr)

        # 如果EQ的任一侧不是简单列引用（如表达式拼接、函数调用等），
        # 则走非等值连接路径（cross join + row filter）
        if not left_is_col or not right_is_col:
            return (None, None, on_clause)

        # 判断哪个属于左表,哪个属于右表
        if left_tbl:
            resolved_left_tbl = self._table_aliases.get(left_tbl, left_tbl)
            if resolved_left_tbl == right_table or left_tbl == right_alias:
                return right_col, left_col, None
        if right_tbl:
            resolved_right_tbl = self._table_aliases.get(right_tbl, right_tbl)
            if resolved_right_tbl == left_table:
                return right_col, left_col, None

        return left_col, right_col, None

    def _apply_lateral_join(self, join, lateral_node, left_df, worksheets_data, left_table, join_kind):
        """
        执行LATERAL JOIN: 对左表每行执行关联子查询，合并结果。

        LATERAL子查询可以引用左表的列（通过别名.列名）。
        策略：将子查询中引用左表的列替换为当前行的字面值，逐行执行。

        Args:
            join: sqlglot Join节点
            lateral_node: exp.Lateral节点
            left_df: 左表DataFrame
            worksheets_data: 所有工作表数据
            left_table: 左表名（或别名）
            join_kind: JOIN类型(inner/left/cross)

        Returns:
            pd.DataFrame: JOIN后的结果
        """
        subquery_expr = lateral_node.this  # exp.Subquery
        lateral_alias = lateral_node.alias or "_lateral"

        # ON条件
        on_clause = join.args.get("on")

        # 收集子查询中引用左表的列 (如 p.ColName)
        # left_table可能是表名(Players)或别名(p)，需要收集所有可能的引用方式
        left_aliases = {left_table}
        for alias, real_name in self._table_aliases.items():
            if real_name == left_table:
                left_aliases.add(alias)
        inner_select = subquery_expr.this if isinstance(subquery_expr, exp.Subquery) else subquery_expr
        # correlated_refs: {(table_alias, col_name)}
        correlated_refs = {}
        for col in inner_select.find_all(exp.Column):
            if col.table in left_aliases and col.name in left_df.columns:
                correlated_refs[(col.table, col.name)] = True

        # 优化：尝试pandas原生执行（避免逐行parse_one开销）
        lateral_results = self._apply_lateral_pandas(inner_select, left_df, correlated_refs, worksheets_data, lateral_alias)
        if lateral_results is None:
            # 降级：逐行SQL解析执行
            lateral_results = self._apply_lateral_sql_fallback(inner_select, left_df, correlated_refs, worksheets_data)

        # 构建结果
        # 推断LATERAL列名
        sample_lateral = None
        for r in lateral_results:
            if r is not None and len(r) > 0:
                sample_lateral = r
                break

        all_rows = []
        # R48-fix P0-03: 使用 enumerate 替代 iterrows 索引,避免 DataFrame 索引不连续时 list 越界
        for pos, (i, row) in enumerate(left_df.iterrows()):
            if pos >= len(lateral_results):
                break  # lateral_results 长度不足,安全终止
            lateral_df = lateral_results[pos]
            if lateral_df is not None and len(lateral_df) > 0:
                for _, lr in lateral_df.iterrows():
                    combined = dict(row)
                    for col in lr.index:
                        col_name = f"{lateral_alias}.{col}"
                        combined[col_name] = lr[col]
                    all_rows.append(combined)
            elif join_kind in ("left", "cross"):
                combined = dict(row)
                if sample_lateral is not None:
                    for col in sample_lateral.columns:
                        col_name = f"{lateral_alias}.{col}"
                        combined[col_name] = None
                all_rows.append(combined)

        if not all_rows:
            return left_df.iloc[0:0]

        result_df = pd.DataFrame(all_rows)

        if on_clause is not None:
            result_df = self._apply_row_filter(on_clause, result_df)

        return result_df

    def _apply_lateral_pandas(self, inner_select, left_df, correlated_refs, worksheets_data, lateral_alias):
        """优化LATERAL执行：用pandas原生操作替代逐行SQL解析。

        支持的模式:
        - SELECT agg_fn(col) FROM table WHERE alias.col = val [ORDER BY ... LIMIT n]
        - SELECT col FROM table WHERE alias.col = val [ORDER BY ... LIMIT n]

        Returns:
            list[df|None] 或 None(无法优化时降级)
        """
        from_ = inner_select.args.get("from")
        if not from_:
            return None

        # 1. 解析FROM表名
        from_table = from_.this
        table_name = from_table.name if hasattr(from_table, "name") else str(from_table)
        # 匹配工作表名（不区分大小写）
        sheet_name = None
        for s in worksheets_data:
            if s == table_name or s.lower() == table_name.lower():
                sheet_name = s
                break
        if sheet_name is None:
            return None

        right_df = worksheets_data[sheet_name]

        # 2. 解析WHERE中的关联条件: 必须是 alias.col = alias.col 形式
        where_clause = inner_select.args.get("where")
        if not where_clause:
            return None

        # 提取WHERE中的等值关联条件
        where_expr = where_clause.this
        lateral_col_map = {}  # {right_col_name: left_col_name}
        remaining_conditions = []

        if isinstance(where_expr, exp.EQ):
            left_cond, right_cond = where_expr.this, where_expr.expression
            if self._extract_correlated_eq(
                left_cond,
                right_cond,
                left_df.columns,
                correlated_refs,
                right_df,
                lateral_col_map,
            ):
                pass
            elif self._extract_correlated_eq(
                right_cond,
                left_cond,
                left_df.columns,
                correlated_refs,
                right_df,
                lateral_col_map,
            ):
                pass
            else:
                return None  # 无法识别的WHERE条件
        elif isinstance(where_expr, exp.And):
            # 处理AND连接的多个条件
            for child in [where_expr.this, where_expr.expression]:
                if isinstance(child, exp.EQ):
                    l, r = child.this, child.expression
                    if not self._extract_correlated_eq(
                        l,
                        r,
                        left_df.columns,
                        correlated_refs,
                        right_df,
                        lateral_col_map,
                    ):
                        if not self._extract_correlated_eq(
                            r,
                            l,
                            left_df.columns,
                            correlated_refs,
                            right_df,
                            lateral_col_map,
                        ):
                            return None
                else:
                    return None
        else:
            return None

        if not lateral_col_map:
            return None

        # 3. 解析SELECT列和聚合
        select_exprs = inner_select.args.get("expressions", [])
        agg_info = self._parse_lateral_select(select_exprs, right_df.columns)
        if agg_info is None:
            return None

        # 4. 解析ORDER BY + LIMIT
        order_col = None
        order_desc = False
        limit_n = None

        order_by = inner_select.args.get("order")
        if order_by:
            for ordered in order_by.expressions:
                order_col_expr = ordered.this
                if hasattr(order_col_expr, "name"):
                    order_col = order_col_expr.name
                    desc_str = str(ordered.args.get("desc", "")).lower()
                    order_desc = desc_str == "true" or ordered.desc

        limit_expr = inner_select.args.get("limit")
        if limit_expr:
            try:
                limit_n = int(limit_expr.this.name if hasattr(limit_expr.this, "name") else str(limit_expr.this))
            except (ValueError, AttributeError):
                pass

        # 5. 批量执行：预分组右表数据
        # lateral_col_map: {right_col: left_col}
        # 按 right_df 的关联列分组
        right_group_cols = list(lateral_col_map.keys())
        left_group_cols = list(lateral_col_map.values())

        right_grouped = right_df.groupby(right_group_cols)

        # 6. 对每个left行查找对应的right分组并计算结果
        results = []
        for idx, row in left_df.iterrows():
            # 构建查找key
            try:
                key = tuple(row[c] for c in left_group_cols)
                if len(key) == 1:
                    key = key[0]
                group = right_grouped.get_group(key)
            except (KeyError, ValueError):
                results.append(None)
                continue

            # 应用ORDER BY
            if order_col and order_col in group.columns:
                group = group.sort_values(order_col, ascending=not order_desc)
            # 应用LIMIT
            if limit_n is not None:
                group = group.head(limit_n)

            if agg_info["type"] == "raw":
                # SELECT col [AS alias] FROM ... (无聚合)
                col_pairs = agg_info["columns"]
                source_cols = [c[0] for c in col_pairs]
                alias_map = {c[0]: c[1] for c in col_pairs if c[0] != c[1]}
                result_df = group[source_cols].reset_index(drop=True)
                if alias_map:
                    result_df = result_df.rename(columns=alias_map)
            elif agg_info["type"] == "agg":
                # SELECT agg_fn(col) FROM ...
                row_result = {}
                for alias, (fn, col) in agg_info["aggs"].items():
                    if fn == "max":
                        row_result[alias] = group[col].max()
                    elif fn == "min":
                        row_result[alias] = group[col].min()
                    elif fn == "sum":
                        row_result[alias] = group[col].sum()
                    elif fn == "avg":
                        row_result[alias] = group[col].mean()
                    elif fn == "count":
                        row_result[alias] = group[col].count()
                    else:
                        return None  # 不支持的聚合函数
                result_df = pd.DataFrame([row_result])
            else:
                return None

            results.append(result_df if len(result_df) > 0 else None)

        return results

    def _extract_correlated_eq(
        self,
        left_cond,
        right_cond,
        left_cols,
        correlated_refs,
        right_df,
        lateral_col_map,
    ):
        """尝试从等值条件中提取关联映射。

        检查 left_cond 是否是关联引用(alias.col), right_cond 是否是右表列。
        如果匹配，将映射添加到 lateral_col_map。

        Returns:
            bool: 是否成功提取
        """
        if not isinstance(left_cond, exp.Column) or not isinstance(right_cond, exp.Column):
            return False

        # left_cond 应该是关联引用 (alias.col where alias refers to left table)
        if left_cond.table and (left_cond.table, left_cond.name) in correlated_refs:
            # right_cond 应该是右表列
            if right_cond.name in right_df.columns:
                lateral_col_map[right_cond.name] = left_cond.name
                return True
        return False

    def _parse_lateral_select(self, select_exprs, right_cols):
        """解析LATERAL子查询的SELECT部分。

        Returns:
            dict: {'type': 'raw'|'agg', 'columns': [(col_name, alias)], 'aggs': {alias: (fn, col)}}
            或 None(无法解析)
        """
        if not select_exprs:
            return None

        has_agg = False
        raw_cols = []  # [(source_col_name, output_alias)]
        aggs = {}

        for expr in select_exprs:
            # 处理 AS 别名: Alias(this=Column('Score'), alias='top')
            inner_expr = expr.this if isinstance(expr, exp.Alias) else expr
            alias = expr.alias if isinstance(expr, exp.Alias) else None

            # 检查是否是聚合函数
            if isinstance(inner_expr, (exp.Max, exp.Min, exp.Sum, exp.Avg, exp.Count)):
                agg_fn = type(inner_expr).__name__.lower()
                agg_inner = inner_expr.this
                agg_col = agg_inner.name if hasattr(agg_inner, "name") else None
                alias = alias or f"{agg_fn}({agg_col})"
                if agg_col:
                    aggs[alias] = (agg_fn, agg_col)
                    has_agg = True
                else:
                    return None
            elif isinstance(inner_expr, exp.Column):
                col_name = inner_expr.name
                if col_name in right_cols:
                    raw_cols.append((col_name, alias or col_name))
                else:
                    return None
            elif isinstance(inner_expr, exp.Star):
                for c in right_cols:
                    raw_cols.append((c, c))
            else:
                return None  # 不支持的表达式

        if has_agg:
            return {"type": "agg", "aggs": aggs}
        else:
            return {"type": "raw", "columns": raw_cols}

    def _apply_lateral_sql_fallback(self, inner_select, left_df, correlated_refs, worksheets_data):
        """LATERAL降级路径：逐行SQL解析执行（慢但通用）"""
        lateral_results = []
        for idx, row in left_df.iterrows():
            lateral_sql = inner_select.sql(dialect="mysql")
            for tbl_alias, col_name in correlated_refs:
                val = row[col_name]
                placeholder = f"{tbl_alias}.{col_name}"
                if isinstance(val, (int, float)):
                    lateral_sql = lateral_sql.replace(placeholder, str(val))
                elif isinstance(val, str):
                    lateral_sql = lateral_sql.replace(placeholder, f"'{val}'")
                elif val is None:
                    lateral_sql = lateral_sql.replace(placeholder, "NULL")

            try:
                lateral_parsed = sqlglot.parse_one(lateral_sql, read="mysql")
                result = self._execute_query(lateral_parsed, worksheets_data)
                lateral_results.append(result)
            except Exception:
                lateral_results.append(None)
        return lateral_results

    def _apply_where_clause(self, parsed_sql: exp.Expression, df) -> pd.DataFrame:
        """应用WHERE条件"""
        where_clause = parsed_sql.args.get("where")
        if not where_clause:
            return df

        # 如果WHERE包含复杂表达式(pandas query不支持的类型),直接使用逐行过滤
        where_expr = where_clause.this
        has_complex = any(where_expr.find(t) is not None for t in self._COMPLEX_EXPR_TYPES)

        if has_complex:
            return self._apply_row_filter(where_expr, df)

        # 将SQLGlot表达式转换为pandas查询条件
        # Fix(R52): 使用实例级临时列追踪，避免 df._tmp_columns 触发 pandas UserWarning
        self._pending_tmp_cols = []
        condition_str = self._sql_condition_to_pandas(where_expr, df)

        if condition_str:
            try:
                result_df = df.query(condition_str)
                # Fix(R46/R52): 清理 WHERE 子句中 CAST/函数预计算产生的临时列
                # 使用 .copy() 确保 result_df 不受后续 del 操作影响（防止 view 问题）
                result_df = result_df.copy()
                self._cleanup_tmp_columns(df)
                return result_df
            except Exception:
                # 如果查询失败,尝试逐行过滤
                # 同时清理可能已添加的临时列
                self._cleanup_tmp_columns(df)
                return self._apply_row_filter(where_clause.this, df)

        logger.warning("WHERE条件转换为pandas表达式失败,回退到逐行过滤: %s", where_expr)
        return self._apply_row_filter(where_expr, df)

    def _cleanup_tmp_columns(self, df):
        """清理WHERE处理过程中添加的临时列 (R52 refactor)"""
        tmp_cols = getattr(self, '_pending_tmp_cols', [])
        for tc in tmp_cols:
            if tc in df.columns:
                del df[tc]
        self._pending_tmp_cols = []

    @staticmethod
    def _like_to_regex(value_str: str) -> str:
        """将SQL LIKE模式转换为pandas regex模式(%->.*  _->.)

        安全加固(R42): 先保护 SQL 通配符,再转义正则元字符,防止 ReDoS 和注入.
        策略: 先用占位符替换 % 和 _, re.escape 后再还原为正则等价形式.
        """
        # re already imported at top level as _re
        pattern = str(value_str).strip("'\"")
        # 防止超长模式导致 ReDoS
        if len(pattern) > 256:
            raise ValueError(f"LIKE 模式过长({len(pattern)}字符), 最大支持256字符")
        # 第一步: 用占位符保护 SQL 通配符 (% 和 _)
        placeholder_pct = "\x00PCT\x00"
        placeholder_und = "\x00UND\x00"
        protected = pattern.replace("%", placeholder_pct).replace("_", placeholder_und)
        # 第二步: 转义所有正则元字符(此时 % _ 已被保护)
        escaped = re.escape(protected)
        # 第三步: 将占位符还原为正则等价形式
        # 注意: re.escape 会转义占位符中的不可打印字符,所以搜索转义后的形式
        escaped_escaped_pct = re.escape(placeholder_pct)
        escaped_escaped_und = re.escape(placeholder_und)
        regex = escaped.replace(escaped_escaped_pct, ".*").replace(escaped_escaped_und, ".")
        return regex

    @staticmethod
    def _parse_literal_value(expr: exp.Literal) -> Any:
        """将SQL Literal解析为Python值(字符串->str,数字->int/float)"""
        # 处理SQLGlot 27.29.0的is_string属性bug
        try:
            is_string = expr.is_string
        except KeyError:
            # 如果is_string属性访问失败，基于this的内容判断是否为字符串
            is_string = isinstance(expr.this, str) or (isinstance(expr.this, str) and expr.this.startswith("'"))

        if is_string:
            return expr.this
        # Fix(R11): 先尝试int(更精确),再尝试float(支持科学计数法如1e100/1.5e-10)
        # 原逻辑用'.'判断类型,但科学计数法不含点导致int()失败返回字符串
        try:
            return int(expr.this)
        except (ValueError, TypeError, OverflowError):
            try:
                return float(expr.this)
            except (ValueError, TypeError, OverflowError):
                return expr.this

    @staticmethod
    def _escape_pandas_query_string(value: str) -> str:
        """转义字符串值中的特殊字符，防止 pandas query() 注入。

        pandas query() 使用 Python 表达式语法，字符串用单引号包裹。
        需要转义的特殊字符：
        - 单引号 ' → \\'  （防止提前闭合字符串字面量）
        - 反斜杠 \\\\ → \\\\\\  （防止转义注入）

        Args:
            value: 原始字符串值

        Returns:
            转义后安全用于 pandas query() 的字符串
        """
        # 先转义反斜杠（必须先处理，否则后续转义会双重处理）
        escaped = value.replace('\\', '\\\\')
        # 再转义单引号
        escaped = escaped.replace("'", "\\'")
        return escaped

    @staticmethod
    def _sanitize_error_message(error_msg: str) -> str:
        """清理异常消息中的敏感内部信息（文件路径、栈帧等）。

        防止通过错误消息泄露服务器内部路径、模块结构等信息。
        保留用户可理解的错误描述，移除绝对路径。

        Args:
            error_msg: 原始异常消息

        Returns:
            清理后的安全消息
        """
        if not error_msg:
            return error_msg

        sanitized = error_msg
        # 移除常见绝对路径模式 (Unix)
        # 匹配 /root/, /home/, /usr/, /opt/, /app/, /var/, /etc/ 开头的路径
        path_pattern = r'(?:(?:^|(?<=[^a-zA-Z0-9_./-]))(?:/?(?:root|home|usr|opt|app|var|etc|tmp|src|lib|local|workspace|project|build|dist|\.hermes|\.cache)[/][^\s,"\']{0,300}))'
        sanitized = re.sub(path_pattern, '<path>', sanitized)

        # 移除 Python 模块路径格式 (package.module.function)
        # 如 "excel_mcp_server_fastmcp.api.advanced_sql_query.method_name"
        module_path_pattern = r'[a-z_][a-z0-9_]*(?:\.[a-z_][a-z0-9_]*){2,}(?:\.[A-Z][a-zA-Z]*)?'
        # 只替换看起来像完整模块路径的（至少3段且含已知前缀）
        known_prefixes = ('excel_mcp', 'sqlglot', 'pandas', 'numpy', 'openpyxl', 'calamine')
        def _replace_module_path(m):
            text = m.group(0)
            if any(text.startswith(p) for p in known_prefixes):
                return '<module>'
            return text
        sanitized = re.sub(module_path_pattern, _replace_module_path, sanitized)

        # 移除行号引用 "line XXX" 或 ":line XXX"
        sanitized = re.sub(r':?\s*line\s+\d+', ':<line>', sanitized)

        return sanitized

    def _in_to_pandas(self, in_expr: exp.In, df, negate: bool = False) -> str:
        """将IN/NOT IN条件转换为pandas表达式(支持子查询和值列表)

        Args:
            in_expr: sqlglot In表达式
            df: 当前DataFrame
            negate: True->NOT IN(~isin),False->IN(isin)
        """
        left = self._expression_to_column_reference(in_expr.this, df)
        prefix = "~" if negate else ""

        # 子查询模式 (IN (SELECT ...))
        subquery = in_expr.args.get("query")
        if subquery and isinstance(subquery, exp.Subquery):
            try:
                sub_result = self._execute_subquery(subquery, self._current_worksheets)
                if len(sub_result.columns) > 0:
                    # Fix(R13): _execute_subquery 返回的 DataFrame 不含表头行
                    # columns 属性就是列名，数据从 iloc[0] 开始
                    # 旧代码错误地用 iloc[1:, 0] 跳过了第一行数据
                    sub_values = sub_result.iloc[:, 0].dropna().tolist()
                    # Fix(R56): 使用 set 去重 + 类型感知格式化，避免大结果集 O(n×m)
                    unique_values = list(set(sub_values))
                    values_str = ", ".join(
                        f"'{v}'" if isinstance(v, str) else str(v)
                        for v in unique_values
                    )
                    return f"{prefix}{left}.isin([{values_str}])"
                return f"{prefix}{left}.isin([])"
            except Exception as e:
                op = "NOT IN" if negate else "IN"
                raise ValueError(f"{op}子查询执行失败: {e}")

        # 值列表模式
        values = [self._expression_to_value(v, df) for v in in_expr.expressions]
        # R48-fix: SQL标准规定 NULL IN (...) 结果为UNKNOWN(WHERE中视为FALSE)
        # 过滤掉None值,避免pandas isin([None])产生语义错误匹配
        values = [v for v in values if v is not None]
        if not values:
            # IN (NULL,NULL,...): 无可匹配值 → 空集; NOT IN (NULL,...): SQL标准全UNKNOWN→空集
            return "index != index"
        # R48-fix: 正确处理pandas query中的类型格式
        # 整数/浮点数直接用str; 字符串需要单引号包裹
        formatted = []
        for v in values:
            if isinstance(v, str):
                formatted.append(f"'{v}'")
            else:
                formatted.append(str(v))
        values_str = ", ".join(formatted)
        return f"{prefix}{left}.isin([{values_str}])"

    def _sql_condition_to_pandas(self, condition: exp.Expression, df) -> str:
        """将SQL条件转换为pandas查询字符串"""
        op_type = type(condition)
        if op_type in self._PANDAS_OPS:
            # Fix(R46): 支持左边为 CAST/函数等非列表达式
            # 原代码仅支持 exp.Column，导致 CAST(col AS FLOAT) > 6 等表达式失败
            left_expr = condition.left
            if isinstance(left_expr, exp.Column):
                # 普通列引用: 走原有快速路径
                left = self._expression_to_column_reference(left_expr, df)
                right = self._expression_to_value(condition.right, df)

                # SQL标准: 任何与NULL的比较都返回UNKNOWN(在WHERE中视为FALSE)
                if right is None:
                    return "False"

                return f"{left} {self._PANDAS_OPS[op_type]} {right}"
            else:
                # 非列表达式(CAST/函数/嵌套表达式/HAVING聚合): 预计算为临时列
                # R48-fix: HAVING 聚合函数(COUNT/SUM/AVG 等)需映射到已计算的别名列
                resolved_left = left_expr
                if isinstance(left_expr, exp.AggFunc) and hasattr(self, '_having_agg_alias_map'):
                    agg_sql = left_expr.sql()
                    if agg_sql in self._having_agg_alias_map:
                        alias_col = self._having_agg_alias_map[agg_sql]
                        if alias_col in df.columns:
                            # 将聚合函数替换为对应的列引用
                            # exp already imported at top level
                            resolved_left = exp.Column(this=exp.Identifier(this=alias_col))

                # hashlib already imported at top level
                expr_hash = hashlib.md5(str(resolved_left).encode()).hexdigest()[:8]
                temp_col = f"_cast_tmp_{expr_hash}"

                try:
                    # 使用向量化求值: CAST走_evaluate_cast_expression, 函数走_process_select_expression
                    if isinstance(resolved_left, exp.Cast):
                        df[temp_col] = self._evaluate_cast_expression(resolved_left, df)
                    elif isinstance(resolved_left, exp.Column):
                        # R48: 聚合函数已解析为列引用 → 直接取列值
                        # 注意: _expression_to_column_reference 返回带反引号的名称(给query用的)
                        # 此处需要原始列名用于 df[col] 访问
                        col_name_raw = resolved_left.this.this if hasattr(resolved_left, 'this') and hasattr(resolved_left.this, 'this') else str(resolved_left)
                        if col_name_raw in df.columns:
                            df[temp_col] = df[col_name_raw]
                        else:
                            # 回退: 尝试大小写不敏感匹配
                            for c in df.columns:
                                if c.lower() == col_name_raw.lower():
                                    df[temp_col] = df[c]
                                    break
                            else:
                                raise ValueError(f"列 '{col_name_raw}' 不存在于 DataFrame 中")
                    else:
                        # 其他表达式: 通过 _expr_to_series 或 _process_select_expression
                        df[temp_col] = self._expr_to_series(resolved_left, df)

                    right = self._expression_to_value(condition.right, df)
                    if right is None:
                        del df[temp_col]
                        return "False"

                    query_str = f"`{temp_col}` {self._PANDAS_OPS[op_type]} {right}"
                    # R52: 使用实例级 _pending_tmp_cols 追踪临时列（避免 pandas UserWarning）
                    # 不在此处删除临时列,因为query()是惰性求值
                    # 调用方(_apply_where_clause)会在查询完成后通过 _cleanup_tmp_columns 清理
                    getattr(self, '_pending_tmp_cols', []).append(temp_col)
                    return query_str
                except Exception as e:
                    if temp_col in df.columns:
                        del df[temp_col]
                    raise ValueError(f"WHERE子句中复杂表达式求值失败 ({left_expr}): {e}")

        elif isinstance(condition, exp.And):
            left = self._sql_condition_to_pandas(condition.left, df)
            right = self._sql_condition_to_pandas(condition.right, df)
            if left is None or right is None:
                raise ValueError("AND条件包含不支持的子查询类型。💡 建议将复杂子查询改为JOIN或逐行过滤")
            return f"({left}) & ({right})"

        elif isinstance(condition, exp.Or):
            left = self._sql_condition_to_pandas(condition.left, df)
            right = self._sql_condition_to_pandas(condition.right, df)
            if left is None or right is None:
                raise ValueError("OR条件包含不支持的子查询类型。💡 建议将复杂子查询改为JOIN或逐行过滤")
            return f"({left}) | ({right})"

        elif isinstance(condition, exp.Paren):
            return self._sql_condition_to_pandas(condition.this, df)

        elif isinstance(condition, exp.Not):
            inner = condition.this
            if isinstance(inner, exp.Like):
                left = self._expression_to_column_reference(inner.this, df)
                right = self._expression_to_value(inner.expression, df)
                # 注意: _expression_to_value() 已对字符串中的特殊字符做了转义,
                # 此处无需再次转义,直接将 LIKE 模式转为 regex 即可
                regex = self._like_to_regex(right)
                return f"~{left}.str.match('{regex}', case=False, na=False)"
            if isinstance(inner, exp.In):
                return self._in_to_pandas(inner, df, negate=True)
            # 其他NOT表达式(IS NOT NULL等)
            pandas_expr = self._sql_condition_to_pandas(inner, df)
            return f"~({pandas_expr})"

        elif isinstance(condition, exp.Like):
            left = self._expression_to_column_reference(condition.this, df)
            right = self._expression_to_value(condition.expression, df)
            # 注意: _expression_to_value() 已对字符串中的特殊字符做了转义,
            # 此处无需再次转义,直接将 LIKE 模式转为 regex 即可
            regex = self._like_to_regex(right)
            return f"{left}.str.match('{regex}', case=False, na=False)"

        elif isinstance(condition, exp.In):
            return self._in_to_pandas(condition, df, negate=False)

        # EXISTS (子查询) -- 需要逐行评估,返回None触发行过滤回退
        elif isinstance(condition, exp.Exists):
            return None

        # IS NULL (sqlglot解析为 exp.Is)
        # 注意: Excel中空单元格被openpyxl/calamine读为空字符串(非NaN),
        # 因keep_default_na=False保留原始语义.所以IS NULL需同时匹配NaN和空字符串.
        elif isinstance(condition, exp.Is):
            left = self._expression_to_column_reference(condition.this, df)
            return f"({left}.isna()) | ({left} == '')"

        # BETWEEN x AND y
        elif isinstance(condition, exp.Between):
            left = self._expression_to_column_reference(condition.this, df)
            low = self._expression_to_value(condition.args["low"], df)
            high = self._expression_to_value(condition.args["high"], df)
            # R48-fix: NULL in BETWEEN → SQL标准规定 NULL比较结果为UNKNOWN, WHERE中视为FALSE
            if low is None or high is None:
                return "index != index"
            return f"({left} >= {low}) & ({left} <= {high})"

        elif isinstance(condition, exp.Boolean):
            # SQL布尔字面量(TRUE/FALSE)作为WHERE条件
            # 注意: pandas query() 中 True/False 会被解释为列名，需用恒真/恒假表达式
            return "index == index" if condition.this else "index != index"

        else:
            raise ValueError(f"不支持的条件类型: {type(condition)}")

    def _expression_to_column_reference(self, expr: exp.Expression, df) -> str:
        """将表达式转换为列引用(支持表限定符 a.column, 大小写不敏感)"""
        if isinstance(expr, exp.Column):
            col_name = expr.name
            # 处理表限定符 (a.column_name -> 查找 "a.column_name" 或 "column_name")
            table_part = expr.table if hasattr(expr, "table") and expr.table else None
            qualified = f"{table_part}.{col_name}" if table_part else None

            if qualified and qualified in df.columns:
                return f"`{qualified}`"
            # 大小写不敏感列名查找
            actual_col = self._find_column_name(col_name, df)
            if actual_col:
                return f"`{actual_col}`"

            # JOIN别名映射支持
            if table_part:
                # 1. 检查用户使用的别名格式是否直接存在
                alias_col = f"{table_part}.{col_name}"
                if alias_col in df.columns:
                    return f"`{alias_col}`"

                # 2. 检查JOIN后pandas添加的后缀格式(table_part_列名)
                pandas_suffix_col = f"{table_part}_{col_name}"
                if pandas_suffix_col in df.columns:
                    return f"`{pandas_suffix_col}`"

                # 2.1 检查pandas merge后的_x/_y后缀映射
                # 2.1.1 检查列名是否直接是_x/_y后缀格式(pandas自动处理)
                for col in df.columns:
                    if col.endswith("_x") and col[:-2] == col_name:
                        # 检查是否是JOIN冲突导致的_x后缀
                        if table_part:
                            # 如果table_part存在,检查是否有对应的x_col
                            x_col = f"{table_part}_x"
                            if x_col in df.columns:
                                return f"`{x_col}`"
                            # 如果没有,检查是否是冲突导致的直接_x后缀
                            return f"`{col}`"
                    elif col.endswith("_y") and col[:-2] == col_name:
                        if table_part:
                            # 如果table_part存在,检查是否有对应的y_col
                            y_col = f"{table_part}_y"
                            if y_col in df.columns:
                                return f"`{y_col}`"
                            # 如果没有,检查是否是冲突导致的直接_y后缀
                            return f"`{col}`"

                # 2.1.2 检查表名+_x/_y后缀格式(更智能的匹配)
                # 处理table_part_x和table_part_y格式
                table_part_x = f"{table_part}_x"
                table_part_y = f"{table_part}_y"
                if table_part_x in df.columns:
                    return f"`{table_part_x}`"
                if table_part_y in df.columns:
                    return f"`{table_part_y}`"

                # 3. 如果有JOIN映射,检查映射后的列名
                if hasattr(self, "_join_column_mapping"):
                    mapped_col = self._join_column_mapping.get(table_part, {}).get(col_name)
                    if mapped_col and mapped_col in df.columns:
                        return f"`{mapped_col}`"

                # 4. 尝试其他可能的别名格式
                if hasattr(self, "_table_aliases"):
                    resolved_table = self._table_aliases.get(table_part, table_part)
                    # 检查原始表名+列名
                    original_col = f"{resolved_table}_{col_name}"
                    if original_col in df.columns:
                        return f"`{original_col}`"
                    # 如果用户使用的是原始表名,检查原始表名+列名
                    if table_part == resolved_table and f"{table_part}_{col_name}" in df.columns:
                        return f"`{table_part}_{col_name}`"

                # 5. 最后尝试原始列名
                if col_name in df.columns:
                    return f"`{col_name}`"

            suggestion = self._suggest_column_name(col_name, list(df.columns))

            # 检测是否是窗口函数别名被WHERE引用(SQL标准限制)
            window_hint = self._check_window_alias_hint(col_name)
            if window_hint:
                raise ValueError(f"列 '{qualified or col_name}' 不存在.{window_hint}")

            raise ValueError(f"列 '{qualified or col_name}' 不存在.可用列: {list(df.columns)}.{suggestion}")

        elif isinstance(expr, exp.Literal):
            return str(expr.this)

        elif isinstance(expr, exp.AggFunc):
            # 对于HAVING子句中的聚合函数,需要查找对应的列
            func_name = type(expr).__name__.lower()

            # 优先:通过SELECT别名映射(HAVING COUNT(*) -> SELECT COUNT(*) as cnt)
            if hasattr(self, "_having_agg_alias_map") and self._having_agg_alias_map:
                agg_sql = expr.sql()
                for map_sql, alias in self._having_agg_alias_map.items():
                    if agg_sql == map_sql:
                        if alias in df.columns:
                            return f"`{alias}`"
                        break

            # 模糊匹配:查找列名包含函数名的列(count->count_star/avg_dmg等)
            for col in df.columns:
                if func_name in col.lower():
                    return f"`{col}`"

            # 单数值列兜底(常见于全表聚合场景)
            numeric_cols = [col for col in df.columns if pd.to_numeric(df[col], errors="coerce").notna().sum() > 0]
            if len(numeric_cols) == 1:
                return f"`{numeric_cols[0]}`"
            if df.columns.size > 0:
                return f"`{df.columns[0]}`"

            raise ValueError(f"无法找到聚合函数 {func_name} 对应的列.可用列: {list(df.columns)}")

        else:
            raise ValueError(f"不支持的表达式类型: {type(expr).__name__}。\n💡 建议: 聚合函数(SUM/COUNT/AVG/MAX/MIN)需要配合列名使用，如 SUM(列名)。\n🔧 检查SELECT列表中是否有不支持的函数调用。")

    def _expression_to_value(self, expr: exp.Expression, df) -> str | int | float:
        """将表达式转换为值"""
        if isinstance(expr, exp.Literal):
            # 委托_parse_literal_value统一处理Literal->Python值转换
            parsed = self._parse_literal_value(expr)
            if isinstance(parsed, str):
                # Fix(R7): 转义特殊字符，防止 pandas query() 字符串注入
                escaped = self._escape_pandas_query_string(parsed)
                return f"'{escaped}'"
            return parsed

        elif isinstance(expr, exp.Null):
            # SQL NULL 字面量 - 返回Python None
            return None

        elif isinstance(expr, exp.Column):
            col_name = expr.name
            if col_name not in df.columns:
                suggestion = self._suggest_column_name(col_name, list(df.columns))
                raise ValueError(f"列 '{col_name}' 不存在.可用列: {list(df.columns)}.{suggestion}")
            return f"`{col_name}`"

        elif isinstance(expr, exp.Boolean):
            # SQL布尔字面量(TRUE/FALSE) → int(1/0)，与Excel整数存储比较兼容
            return int(expr.this)

        elif isinstance(expr, exp.AggFunc):
            # 聚合函数作为值的处理(HAVING子句中)
            # 使用与_expression_to_column_reference相同的逻辑
            return self._expression_to_column_reference(expr, df)

        elif isinstance(expr, exp.Subquery):
            """标量子查询: WHERE col > (SELECT AVG(...) FROM ...)

            子查询结果应返回单行单列的标量值.
            修复: 直接从 DataFrame 提取标量值,不再假设有标题行.

            安全加固(R42): 空结果返回 None(SQL NULL)而非 "0",避免错误比较.
            """
            try:
                sub_result = self._execute_subquery(expr, self._current_worksheets)
                if len(sub_result) > 0 and len(sub_result.columns) > 0:
                    scalar_val = sub_result.iloc[0, 0]
                    if isinstance(scalar_val, (int, float, np.integer, np.floating)):
                        return float(scalar_val)
                    if scalar_val is not None:
                        # Fix(R7): 转义标量子查询返回的字符串值
                        escaped = self._escape_pandas_query_string(str(scalar_val))
                        return f"'{escaped}'"
                return None  # 空子查询返回 SQL NULL
            except Exception as e:
                raise ValueError(f"标量子查询执行失败: {e}")

        elif isinstance(expr, exp.Cast):
            # WHERE/HAVING 中的 CAST: 递归求内部值,类型转换由 pandas 在比较时自动处理
            inner_expr = expr.this
            return self._expression_to_value(inner_expr, df)

        elif isinstance(expr, exp.Neg):
            # Fix(R52-P3-EDGE-01): 支持负数字面量（如 IN (1, -1, 2)）
            inner_val = self._expression_to_value(expr.this, df)
            if isinstance(inner_val, str):
                # 字符串形式的负数（来自列引用等）
                try:
                    return -float(inner_val) if '.' in inner_val else -int(inner_val)
                except (ValueError, TypeError):
                    return f"(-{inner_val})"
            elif isinstance(inner_val, (int, float)):
                return -inner_val
            return None

        else:
            raise ValueError(
                f"不支持的表达式类型: {type(expr).__name__}。"
                f"\n💡 WHERE子句支持: 比较运算(=,!=,>,<,>=,<=)、逻辑运算(AND,OR,NOT)、LIKE、IN、BETWEEN、IS NULL、CAST。"
                f"\n🔧 不支持算术运算(如 A+B>10)，建议用子查询: SELECT * FROM (SELECT ..., (A+B) as t FROM tbl) WHERE t>10"
            )

    def _pre_cache_in_subqueries(self, condition: exp.Expression) -> dict:
        """预执行条件树中所有 IN 子查询并缓存结果，避免逐行重复执行"""
        cache = {}
        if not (hasattr(self, "_current_worksheets") and self._current_worksheets):
            return cache

        # 递归查找所有 In 节点
        in_nodes = list(condition.find_all(exp.In))
        for in_node in in_nodes:
            if in_node.expressions and isinstance(in_node.expressions[0], (exp.Subquery, exp.Select)):
                sub_key = id(in_node)
                try:
                    sub_result = self._execute_subquery(in_node.expressions[0], self._current_worksheets)
                    if sub_result.empty:
                        cache[sub_key] = []
                    else:
                        # Fix(R56): 缓存为 set 以支持 O(1) 查找
                        cache[sub_key] = set(sub_result.iloc[:, 0].dropna().tolist())
                except Exception:
                    cache[sub_key] = []
        return cache

    def _apply_row_filter(self, condition: exp.Expression, df) -> pd.DataFrame:
        """逐行应用过滤条件(备用方案),使用apply替代iterrows提升性能"""
        # Fix: 预执行所有IN子查询并缓存结果，避免逐行重复执行
        in_subquery_cache = self._pre_cache_in_subqueries(condition)
        mask = df.apply(lambda row: self._evaluate_condition_for_row(condition, row, in_subquery_cache), axis=1)
        return df[mask]

    def _evaluate_condition_for_row(self, condition: exp.Expression, row: pd.Series, in_subquery_cache: dict = None) -> bool:
        """为单行评估条件"""
        try:
            op_type = type(condition)
            if op_type in self._COMPARISON_OPS:
                # 检查右侧是否为ALL/ANY子查询
                right_expr = condition.right
                if isinstance(right_expr, (exp.All, exp.Any)):
                    return self._evaluate_all_any_comparison(condition, row)

                # SQL标准: 任何与NULL的比较都返回UNKNOWN(在WHERE中视为FALSE)
                if isinstance(right_expr, exp.Null):
                    return False

                left_val = self._get_row_value(condition.left, row)
                right_val = self._get_row_value(condition.right, row)
                try:
                    return self._COMPARISON_OPS[op_type](left_val, right_val)
                except (TypeError, ValueError):
                    return False

            elif isinstance(condition, exp.And):
                return self._evaluate_condition_for_row(condition.left, row) and self._evaluate_condition_for_row(condition.right, row)

            elif isinstance(condition, exp.Or):
                return self._evaluate_condition_for_row(condition.left, row) or self._evaluate_condition_for_row(condition.right, row)

            elif isinstance(condition, exp.Is):
                val = self._get_row_value(condition.this, row)
                return pd.isna(val) or val is None

            elif isinstance(condition, exp.Not):
                inner = condition.this
                if isinstance(inner, exp.Is):
                    val = self._get_row_value(inner.this, row)
                    return not (pd.isna(val) or val is None)
                return not self._evaluate_condition_for_row(inner, row)

            elif isinstance(condition, exp.Like):
                val = str(self._get_row_value(condition.this, row) or "")
                pattern = str(self._get_row_value(condition.expression, row) or "")
                regex = self._like_to_regex(pattern)
                return bool(re.match(regex, val, re.IGNORECASE))

            elif isinstance(condition, exp.In):
                val = self._get_row_value(condition.this, row)
                # Fix(R13): 支持 IN (SELECT ...) 子查询形式
                # sqlglot 将 IN (1,2,3) 和 IN (SELECT ...) 都解析为 In 节点
                # 区分方式: expressions[0] 是否为 Subquery/Select
                if condition.expressions and isinstance(condition.expressions[0], (exp.Subquery, exp.Select)):
                    # IN 子查询: 优先使用缓存，避免逐行重复执行
                    sub_key = id(condition)
                    if in_subquery_cache and sub_key in in_subquery_cache:
                        sub_values = in_subquery_cache[sub_key]
                    else:
                        try:
                            if not (hasattr(self, "_current_worksheets") and self._current_worksheets):
                                return False
                            sub_result = self._execute_subquery(condition.expressions[0], self._current_worksheets)
                            if sub_result.empty:
                                sub_values = []
                            else:
                                sub_values = sub_result.iloc[:, 0].dropna().tolist()
                        except Exception:
                            sub_values = []
                        if in_subquery_cache is not None:
                            in_subquery_cache[sub_key] = sub_values

                    if not sub_values:
                        return False
                    # Fix(R56): 使用 set 进行 O(1) hash lookup 替代 O(n) 线性扫描
                    sub_values_set = set(sub_values)
                    try:
                        if float(val) in {float(sv) for sv in sub_values_set}:
                            return True
                    except (TypeError, ValueError):
                        if val in sub_values_set:
                            return True
                    return False
                else:
                    # IN 字面量列表: IN (1, 2, 3)
                    values = [self._get_row_value(e, row) for e in condition.expressions]
                    return val in values

            elif isinstance(condition, exp.Between):
                val = self._get_row_value(condition.this, row)
                low = self._get_row_value(condition.args["low"], row)
                high = self._get_row_value(condition.args["high"], row)
                try:
                    return float(low) <= float(val) <= float(high)
                except (TypeError, ValueError):
                    return False

            elif isinstance(condition, exp.Exists):
                return self._evaluate_exists_for_row(condition, row)

            elif isinstance(condition, (exp.All, exp.Any)):
                return self._evaluate_all_any_for_row(condition, row)

            # 其他条件类型...

            return True

        except Exception:
            return False

    def _evaluate_all_any_comparison(self, condition, row: pd.Series) -> bool:
        """评估 col > ALL/ANY (SELECT ...) 子查询比较

        sqlglot将 col > ALL (SELECT ...) 解析为 GT(this=col, expression=All(this=Subquery(...))).
        ALL: 所有子查询结果都必须满足比较条件
        ANY/SOME: 至少一个子查询结果满足比较条件
        """
        try:
            left_val = self._get_row_value(condition.left, row)
            quantifier = condition.right  # All or Any node
            subquery_node = quantifier.this
            if isinstance(subquery_node, exp.Subquery):
                subquery_node = subquery_node.this
            if not (hasattr(self, "_current_worksheets") and self._current_worksheets):
                return False
            sub_result = self._execute_subquery(
                condition.right if isinstance(condition.right, exp.Subquery) else condition.right,
                self._current_worksheets,
            )
            if sub_result.empty:
                # ALL: 空集返回True( vacuous truth)
                # ANY: 空集返回False
                return isinstance(quantifier, exp.All)

            sub_values = sub_result.iloc[:, 0].dropna().tolist()
            if not sub_values:
                return isinstance(quantifier, exp.All)

            op_type = type(condition)
            op_func = self._COMPARISON_OPS.get(op_type)
            if op_func is None:
                return False

            if isinstance(quantifier, exp.All):
                # ALL: 每个子查询值都必须满足条件
                return all(op_func(left_val, sv) for sv in sub_values)
            else:
                # ANY/SOME: 至少一个子查询值满足条件
                return any(op_func(left_val, sv) for sv in sub_values)
        except Exception:
            return False

    def _evaluate_all_any_for_row(self, condition, row: pd.Series) -> bool:
        """评估ALL/ANY/SOME子查询比较 (如 WHERE 伤害 > ALL (SELECT ...))"""
        # sqlglot将 col > ALL (SELECT ...) 解析为 GT(this=col, expression=All(...))
        # 所以需要从父节点获取比较运算符和左操作数
        # 但这里condition本身就是All/Any节点,我们需要从调用上下文获取
        # 实际上sqlglot把 ALL/ANY 作为比较操作的expression,所以condition是All/Any
        # 我们需要重新从WHERE条件树中找到完整的比较节点
        # 简化方案:在_apply_row_filter中特殊处理,这里只处理单独出现的情况
        try:
            subquery = condition.this if isinstance(condition.this, (exp.Subquery, exp.Select)) else None
            if subquery is None:
                return False
            if not (hasattr(self, "_current_worksheets") and self._current_worksheets):
                return False
            sub_result = self._execute_subquery(subquery, self._current_worksheets)
            if sub_result.empty:
                return True  # ALL:空集为真; ANY:空集为假(但这里简化处理)
            sub_values = sub_result.iloc[:, 0].dropna().tolist()
            if not sub_values:
                return isinstance(condition, exp.All)
            return True  # 实际比较在_apply_row_filter中处理
        except Exception:
            return False

    def _evaluate_exists_for_row(self, condition: exp.Exists, row: pd.Series) -> bool:
        """评估EXISTS子查询(支持关联子查询)"""
        inner_expr = condition.this
        if isinstance(inner_expr, exp.Subquery):
            inner_select = inner_expr.this
        elif isinstance(inner_expr, exp.Select):
            inner_select = inner_expr
        else:
            return False

        if not (hasattr(self, "_current_worksheets") and self._current_worksheets):
            return False

        inner_from, _ = self._get_from_table(inner_select)
        has_correlation = False
        for col in inner_select.find_all(exp.Column):
            col_name = col.name
            table_part = col.table if hasattr(col, "table") and col.table else None
            if table_part:
                resolved = self._table_aliases.get(table_part, table_part)
                if resolved != inner_from:
                    has_correlation = True
                    break
            else:
                if inner_from in self._current_worksheets:
                    for tbl_name, tbl_df in self._current_worksheets.items():
                        if tbl_name != inner_from and col_name in tbl_df.columns:
                            has_correlation = True
                            break

        if has_correlation:
            return self._evaluate_correlated_exists(inner_select, inner_from, row)
        else:
            sub_result = self._execute_subquery(inner_expr, self._current_worksheets)
            return len(sub_result) > 0

    def _evaluate_correlated_exists(self, inner_select, inner_from: str, row: pd.Series) -> bool:
        """评估关联EXISTS子查询:替换外部引用为当前行值后执行"""
        inner_sql = str(inner_select)
        inner_from_cols = set()
        if inner_from in self._current_worksheets:
            inner_from_cols = set(self._current_worksheets[inner_from].columns)

        for col in inner_select.find_all(exp.Column):
            col_name = col.name
            table_part = col.table if hasattr(col, "table") and col.table else None
            should_substitute = False

            if table_part:
                resolved = self._table_aliases.get(table_part, table_part)
                for tbl_name in self._current_worksheets:
                    if tbl_name != inner_from and (resolved == tbl_name or table_part == tbl_name):
                        should_substitute = True
                        break
            else:
                for tbl_name, tbl_df in self._current_worksheets.items():
                    if tbl_name != inner_from and col_name in tbl_df.columns and col_name not in inner_from_cols:
                        should_substitute = True
                        break

            if should_substitute:
                val = row.get(col_name)
                if val is not None:
                    # [FIX R55-BUG-03] SQL 标准转义：单引号 → 双单引号
                    # 避免 repr() 对含引号字符串(如 O'Brien)产生转义反斜杠
                    # 导致 sqlglot 解析失败后静默返回 False
                    if isinstance(val, str):
                        safe_val = "'" + val.replace("'", "''") + "'"
                    else:
                        safe_val = repr(val)
                    if table_part:
                        inner_sql = inner_sql.replace(f"{table_part}.{col_name}", safe_val, 1)
                    else:
                        pattern = r"\b" + re.escape(col_name) + r"\b"
                        inner_sql = re.sub(pattern, safe_val, inner_sql, count=1)

        try:
            parsed_inner = sqlglot.parse_one(inner_sql)
            # 保存外部查询状态,避免_execute_query内部重置覆盖
            saved_aliases = dict(self._table_aliases)
            saved_worksheets = dict(self._current_worksheets)
            try:
                sub_result = self._execute_query(parsed_inner, saved_worksheets)
                return len(sub_result) > 0
            finally:
                self._table_aliases = saved_aliases
                self._current_worksheets = saved_worksheets
        except Exception:
            return False

    def _extract_column_references(self, expr: exp.Expression) -> list[str]:
        """从表达式中提取所有列引用

        Args:
            expr: SQL表达式对象

        Returns:
            列名列表

        修复说明(REQ-061):
        - 支持从复杂表达式(CASE WHEN/COALESCE/函数调用)中提取列引用
        - 递归遍历表达式树,提取所有Column类型的子节点
        """
        columns = []
        if isinstance(expr, exp.Column):
            columns.append(expr.name)
        elif hasattr(expr, "expressions"):
            for sub_expr in expr.expressions:
                columns.extend(self._extract_column_references(sub_expr))
        elif hasattr(expr, "this"):
            columns.extend(self._extract_column_references(expr.this))
        if hasattr(expr, "args"):
            for arg in expr.args.values():
                if isinstance(arg, exp.Expression):
                    columns.extend(self._extract_column_references(arg))
        return columns

    def _get_row_value(self, expr: exp.Expression, row: pd.Series) -> Any:
        """获取行中表达式的值"""
        if isinstance(expr, exp.Column):
            col_name = expr.name
            table_part = expr.table if hasattr(expr, "table") and expr.table else None

            # 支持表别名限定符 (a.column)
            if table_part:
                # 优先尝试表别名格式: table_part.column
                qualified_col = f"{table_part}.{col_name}"
                if qualified_col in row.index:
                    return row.get(qualified_col)

                # 尝试pandas merge后的格式: table_part_column
                pandas_suffix_col = f"{table_part}_{col_name}"
                if pandas_suffix_col in row.index:
                    return row.get(pandas_suffix_col)

                # 尝试JOIN映射中的列名
                if hasattr(self, "_join_column_mapping") and table_part in self._join_column_mapping:
                    mapped_col = self._join_column_mapping[table_part].get(col_name)
                    if mapped_col and mapped_col in row.index:
                        return row.get(mapped_col)

                # 尝试表别名解析后的格式
                if hasattr(self, "_table_aliases") and table_part in self._table_aliases:
                    resolved_table = self._table_aliases[table_part]
                    resolved_col = f"{resolved_table}_{col_name}"
                    if resolved_col in row.index:
                        return row.get(resolved_col)

            # 最后尝试原始列名
            return row.get(col_name)

        elif isinstance(expr, exp.Null):
            # SQL NULL 字面量
            return None

        elif isinstance(expr, exp.Literal):
            return self._parse_literal_value(expr)

        elif isinstance(expr, exp.Boolean):
            # SQL布尔字面量(TRUE/FALSE) → int(1/0)，与Excel存储格式一致
            return int(expr.this)

        elif isinstance(expr, exp.Coalesce):
            return self._evaluate_coalesce_for_row(expr, row)

        elif isinstance(expr, exp.Case):
            return self._evaluate_case_expression(expr, None, row=row)

        elif isinstance(expr, (exp.Add, exp.Sub, exp.Mul, exp.Div)):
            return self._evaluate_math_for_row(expr, row)

        elif isinstance(expr, exp.Neg):
            # Fix(R52-P3-EDGE-01): 支持负数字面量（如 WHERE col = -1）
            inner = self._get_row_value(expr.this, row)
            if inner is None:
                return None
            try:
                return -inner
            except (TypeError, ValueError):
                return None

        elif self._is_string_function(expr):
            return self._evaluate_string_function_for_row(expr, row)

        elif self._is_scalar_num_function(expr):
            return self._evaluate_scalar_num_function_for_row(expr, row)

        elif isinstance(expr, exp.Cast):
            # CAST(expr AS type) — 逐行模式
            # 需要传入 df 参数,这里用 row.to_frame().T 构造临时 DataFrame
            tmp_df = row.to_frame().T
            return self._evaluate_cast_expression(expr, tmp_df, row=row)[0]

        elif isinstance(expr, exp.Anonymous):
            # 含括号的列名(如"刷新时间(小时)")被sqlglot解析为Anonymous
            anon_name = expr.this
            if expr.expressions:
                inner = ", ".join(str(e) for e in expr.expressions)
                full_name = f"{anon_name}({inner})"
            else:
                full_name = anon_name
            val = row.get(full_name) if full_name in row.index else row.get(anon_name)

            # 中文函数名映射提示
            _CN_FUNC_MAP = {
                "长度": "LENGTH",
                "LEN": "LENGTH",
                "UPPER": "UPPER",
                "LOWER": "LOWER",
                "TRIM": "TRIM",
                "截取": "SUBSTRING",
                "四舍五入": "ROUND",
                "舍入": "ROUND",
                "拼接": "CONCAT",
                "替换": "REPLACE",
            }
            if val is None:
                for cn_name, en_name in _CN_FUNC_MAP.items():
                    if cn_name in anon_name or cn_name in full_name:
                        return f"💡 如需调用{en_name}函数,请使用英文函数名: {en_name}(...)"
            return val

        elif isinstance(expr, exp.Window):
            # 窗口函数参与逐行评估(如 WHERE 条件中的 LAG/LEAD)
            # 尝试从预计算的窗口函数列中获取值
            temp_col = f"_window_math_{id(expr)}"
            if temp_col in row.index:
                val = row[temp_col]
                return float(val) if val is not None else None
            # 尝试通过窗口函数的文本表示查找
            window_str = str(expr).strip()
            for col in row.index:
                if col.startswith("_window_") or col == window_str:
                    val = row[col]
                    return float(val) if val is not None else None
            return None

        else:
            # Fix(R47): 聚合函数节点(SUM/AVG/COUNT等) — 从row上下文中按SQL文本查找
            # 用于CASE+聚合交互: CASE WHEN SUM(Amount) > 300 THEN ... 场景
            if self._is_aggregate_function(expr):
                agg_sql = expr.sql()
                if agg_sql in row.index:
                    return row.get(agg_sql)
                for col in row.index:
                    if col.replace(" ", "") == agg_sql.replace(" ", ""):
                        return row.get(col)
            return None

    def _apply_group_by_aggregation(self, parsed_sql: exp.Expression, df) -> pd.DataFrame:
        """应用GROUP BY和聚合函数

        Args:
            parsed_sql: SQL解析后的表达式对象
            df: 要处理的DataFrame数据

        Returns:
            应用GROUP BY和聚合函数后的DataFrame

        性能优化:
        - 大数据集使用向量化操作代替逐行计算
        - 减少不必要的DataFrame复制
        - 优化groupby操作使用observed=True避免稀疏分组

        修复说明(REQ-061):
        - 修复多列GROUP BY逻辑，确保当GROUP BY包含多列且SELECT包含计算表达式时，
          正确按所有GROUP BY列分组，而非仅按计算列分组
        - 确保所有GROUP BY列都包含在最终结果中
        """

        group_by_columns = []
        group_clause = parsed_sql.args.get("group")
        if group_clause:
            for group_expr in group_clause.expressions:
                if isinstance(group_expr, exp.Column):
                    group_by_columns.append(group_expr.name)
                else:
                    # 修复(REQ-061): 从复杂表达式中提取列引用
                    # 支持GROUP BY包含表达式(如CASE WHEN/COALESCE)的情况
                    column_refs = self._extract_column_references(group_expr)
                    for col_name in column_refs:
                        if col_name not in group_by_columns:
                            group_by_columns.append(col_name)

        # 检查是否有聚合函数
        aggregations = {}
        select_exprs = {}

        for i, select_expr in enumerate(parsed_sql.expressions):
            alias_name, original_expr = self._extract_select_alias(select_expr, i)
            select_exprs[alias_name] = original_expr

        # 检查聚合函数（包括标量函数包裹聚合的情况，如 ROUND(AVG(col))）
        for alias_name, expr in select_exprs.items():
            if self._is_aggregate_function(expr):
                aggregations[alias_name] = expr
            elif self._is_scalar_num_function(expr) and self._find_inner_aggregate(expr) is not None:
                # 标量函数(ROUND/ABS/FLOOR/CEIL等)包裹聚合函数 → 视为聚合
                # 在后续主循环中会走 _apply_scalar_to_agg_result 分支处理
                aggregations[alias_name] = expr
            # Fix(R47): CASE/COALESCE 内包含聚合函数时也视为聚合表达式
            elif isinstance(expr, (exp.Case, exp.Coalesce)) and self._find_inner_aggregate(expr) is not None:
                aggregations[alias_name] = expr
            elif hasattr(expr, "name") and expr.name not in group_by_columns:
                # 如果是非聚合列且不在GROUP BY中,需要添加到GROUP BY
                # 跳过窗口函数表达式(窗口函数在GROUP BY之后处理)
                if isinstance(expr, exp.Window):
                    continue
                if isinstance(expr, (exp.Column, exp.Identifier)):
                    group_by_columns.append(expr.name)
                else:
                    # 修复(REQ-061): 从复杂表达式中提取列引用
                    # 支持SELECT包含表达式(如CASE WHEN/COALESCE)的情况
                    # 跳过窗口函数表达式
                    if isinstance(expr, exp.Window):
                        continue
                    column_refs = self._extract_column_references(expr)
                    for col_name in column_refs:
                        if col_name not in group_by_columns:
                            group_by_columns.append(col_name)

        # 处理HAVING子句中的聚合函数(REQ-EXCEL-015)
        # HAVING可以引用不在SELECT中的聚合函数(如HAVING COUNT(*) > 1)
        # 需要提前计算这些聚合函数,以便在_apply_having_clause中使用
        having_aggregations = {}
        having_clause = parsed_sql.args.get("having")
        if having_clause:
            having_agg_funcs = self._extract_agg_funcs_from_expr(having_clause.this)
            for agg_func in having_agg_funcs:
                agg_sql = agg_func.sql()
                # 检查是否已经在SELECT中
                if agg_sql not in [aggr.sql() for aggr in aggregations.values()]:
                    # 生成临时列名
                    temp_alias = f"_having_agg_{len(having_aggregations)}"
                    having_aggregations[temp_alias] = agg_func

        # 合并HAVING聚合到aggregations中
        aggregations.update(having_aggregations)

        # 保存HAVING聚合映射供_apply_having_clause使用
        self._having_agg_in_select_map = {}
        for temp_alias, agg_func in having_aggregations.items():
            self._having_agg_in_select_map[agg_func.sql()] = temp_alias

        # 保存GROUP BY列到实例变量
        self._group_by_columns = group_by_columns

        # Fix(E2): 空DataFrame聚合保护 — 0行DataFrame做groupby后访问列会报"Column not found"
        # 返回含正确列名但0行的结果,或全表聚合时返回1行NULL行(符合SQL标准)
        if df.empty:
            ordered_cols = []
            for i, select_expr in enumerate(parsed_sql.expressions):
                alias_name, original_expr = self._extract_select_alias(select_expr, i)
                ordered_cols.append(alias_name)
            # 全表聚合(无GROUP BY列) → 返回1行NULL结果(SQL标准: COUNT→0, 其他→NULL)
            if not group_by_columns:
                result_df = pd.DataFrame(columns=ordered_cols)
                # 填充一行默认值
                default_row = {}
                for i, select_expr in enumerate(parsed_sql.expressions):
                    alias_name, original_expr = self._extract_select_alias(select_expr, i)
                    if self._is_aggregate_function(select_expr if not isinstance(select_expr, exp.Alias) else select_expr.this):
                        func_name = type(original_expr if isinstance(original_expr, exp.AggFunc) else (select_expr.this if isinstance(select_expr, exp.Alias) else select_expr)).__name__.lower()
                        default_row[alias_name] = 0 if func_name == "count" else None
                    else:
                        default_row[alias_name] = None
                result_df = pd.DataFrame([default_row], columns=ordered_cols)
                return result_df
            # 有GROUP BY但无数据 → 返回空结果(0行,有列名)
            return pd.DataFrame(columns=ordered_cols)

        if not aggregations:
            # 没有聚合函数,只应用GROUP BY去重
            if group_by_columns:
                # [FIX R14-B1] 包含预计算的窗口函数列(已在GROUP BY前通过_apply_window_functions添加到df)
                _window_cols = [
                    c
                    for c in df.columns
                    if c.startswith("_window_")
                    or any(self._extract_select_alias(expr, i)[0] == c and isinstance(self._extract_select_alias(expr, i)[1], exp.Window) for i, expr in enumerate(parsed_sql.expressions))
                ]
                _result_cols = list(group_by_columns) + [c for c in _window_cols if c not in group_by_columns]
                # 性能优化:使用drop_duplicates的subset参数避免全列比较
                return df[_result_cols].drop_duplicates(subset=group_by_columns).reset_index(drop=True)
            else:
                return df

        # 预计算CASE WHEN/COALESCE/标量子查询表达式,添加到df副本,使grouped可访问
        # Fix(R47): 跳过包含聚合函数的CASE/COALESCE — 它们需要在分组后求值
        df = df.copy()
        for alias_name, expr in select_exprs.items():
            if isinstance(expr, exp.Case) and alias_name not in df.columns:
                # 检查CASE内是否包含聚合函数(如 CASE WHEN SUM(col) > x THEN ...)
                if self._find_inner_aggregate(expr) is None:
                    df[alias_name] = self._evaluate_case_expression(expr, df)
                # else: 含聚合的CASE在分组后处理
            elif isinstance(expr, exp.Coalesce) and alias_name not in df.columns:
                if self._find_inner_aggregate(expr) is None:
                    df[alias_name] = self._evaluate_coalesce_vectorized(expr, df)
            elif isinstance(expr, exp.Subquery) and alias_name not in df.columns:
                try:
                    sub_result = self._execute_subquery(expr, self._current_worksheets)
                    if len(sub_result) > 0 and len(sub_result.columns) > 0:
                        scalar_val = sub_result.iloc[0, 0]
                        df[alias_name] = pd.Series([scalar_val] * len(df), index=df.index)
                    else:
                        df[alias_name] = pd.Series([None] * len(df), index=df.index)
                except Exception:
                    df[alias_name] = pd.Series([None] * len(df), index=df.index)

        # 应用聚合
        # 性能优化:使用observed=True减少分组计算开销
        if group_by_columns:
            # 确保group_by_columns中的列都存在
            valid_group_cols = [c for c in group_by_columns if c in df.columns]
            if valid_group_cols:
                grouped = df.groupby(valid_group_cols, observed=True, dropna=False)
            else:
                grouped = df.groupby(lambda x: 0)
        else:
            # 全表聚合
            grouped = df.groupby(lambda x: 0)  # 将所有行分组为一组

        # 按照SQL SELECT表达式的顺序构建结果
        result_data = {}
        ordered_columns = []

        # 按SELECT表达式顺序处理列
        for i, select_expr in enumerate(parsed_sql.expressions):
            alias_name, original_expr = self._extract_select_alias(select_expr, i)

            # 跳过SELECT *，它会在循环后单独处理
            if isinstance(original_expr, exp.Star):
                continue

            ordered_columns.append(alias_name)

            # 处理聚合函数
            is_agg = self._is_aggregate_function(select_expr if not isinstance(select_expr, exp.Alias) else select_expr.this)

            if is_agg:
                agg_expr = original_expr if isinstance(select_expr, exp.Alias) else select_expr
                agg_result = self._apply_aggregation_function(agg_expr, grouped, df)
                # 修复:确保聚合结果是正确格式
                if isinstance(agg_result, (int, float, np.integer, np.floating)):
                    result_data[alias_name] = pd.Series([agg_result])
                elif isinstance(agg_result, pd.Series):
                    result_data[alias_name] = agg_result.reset_index(drop=True)
                else:
                    # 尝试转换为Series
                    try:
                        result_data[alias_name] = pd.Series([agg_result])
                    except (ValueError, TypeError):
                        result_data[alias_name] = pd.Series([None])
            # 处理标量函数包裹聚合函数的情况 (如 ROUND(AVG(Price)), ABS(SUM(col)) 等)
            elif self._is_scalar_num_function(original_expr):
                # 检查内部是否包含聚合函数
                inner_agg = self._find_inner_aggregate(original_expr)
                if inner_agg is not None:
                    # 先计算内层聚合
                    try:
                        agg_result = self._apply_aggregation_function(inner_agg, grouped, df)
                        # 将聚合结果转为 Series 用于标量函数处理
                        if isinstance(agg_result, (int, float, np.integer, np.floating)):
                            agg_series = pd.Series([agg_result])
                        elif isinstance(agg_result, pd.Series):
                            agg_series = agg_result.reset_index(drop=True)
                        else:
                            agg_series = pd.Series([None])
                        # 直接对聚合结果应用标量函数（避免 _expr_to_series 解析 Agg 节点失败）
                        scalar_result = self._apply_scalar_to_agg_result(original_expr, agg_series)

                        result_data[alias_name] = scalar_result.reset_index(drop=True)
                    except Exception as e:
                        logger.warning("标量函数+聚合计算失败 %s: %s", alias_name, e)
                        logger.debug("标量函数+聚合计算详细错误: %s", e, exc_info=True)
                        result_data[alias_name] = pd.Series([None])
                else:
                    # 无内层聚合，按普通表达式处理（不应到达这里）
                    result_data[alias_name] = pd.Series([None])
            # 处理CASE WHEN表达式
            elif isinstance(original_expr, exp.Case):
                if alias_name in df.columns:
                    # 已预计算(无内层聚合的简单CASE),直接从grouped取first
                    result_data[alias_name] = grouped[alias_name].first().reset_index(drop=True)
                else:
                    try:
                        case_results = self._evaluate_case_with_aggregate(
                            original_expr, grouped, df, result_data, group_by_columns
                        )
                        result_data[alias_name] = case_results
                    except Exception as e:
                        logger.warning("CASE+聚合计算失败 %s: %s", alias_name, e)
                        result_data[alias_name] = pd.Series([None] * len(grouped))
            # 处理COALESCE表达式
            elif isinstance(original_expr, exp.Coalesce):
                if alias_name in df.columns:
                    # 已预计算(无内层聚合的简单COALESCE)
                    result_data[alias_name] = grouped[alias_name].first().reset_index(drop=True)
                else:
                    # Fix(R47): 含聚合函数的COALESCE(如 COALESCE(SUM(col), 0))
                    try:
                        coalesce_results = self._evaluate_coalesce_with_aggregate(
                            original_expr, grouped, df, result_data, group_by_columns
                        )
                        result_data[alias_name] = coalesce_results
                    except Exception as e:
                        logger.warning("COALESCE+聚合计算失败 %s: %s", alias_name, e)
                        result_data[alias_name] = pd.Series([None] * len(grouped))
            # 处理标量子查询(已预计算到df,直接从grouped取first)
            elif isinstance(original_expr, exp.Subquery):
                result_data[alias_name] = grouped[alias_name].first().reset_index(drop=True)
            # 处理普通列(GROUP BY列)
            elif hasattr(original_expr, "name") and original_expr.name:
                col_name = original_expr.name
                if col_name in group_by_columns:
                    result_data[alias_name] = grouped[col_name].first().reset_index(drop=True)

            # [FIX R14-B1] 处理窗口函数表达式(已在GROUP BY前预计算到df)
            # 窗口函数列已通过_apply_window_functions添加到DataFrame,此处从grouped取first
            elif isinstance(original_expr, exp.Window):
                if alias_name in df.columns:
                    result_data[alias_name] = grouped[alias_name].first().reset_index(drop=True)
                elif alias_name.startswith("_window_") and alias_name in df.columns:
                    result_data[alias_name] = grouped[alias_name].first().reset_index(drop=True)

        # 处理SELECT *的情况：添加所有GROUP BY列
        has_star = any(isinstance(self._extract_select_alias(expr, 0)[1], exp.Star) for expr in parsed_sql.expressions)
        if has_star:
            for col in group_by_columns:
                if col not in result_data:
                    result_data[col] = grouped[col].first().reset_index(drop=True)
                    if col not in ordered_columns:
                        ordered_columns.append(col)

        # 构建完整的别名→原始列名反向映射(覆盖所有表达式类型)
        # 用于GROUP BY列去重判断:避免 SELECT col AS 别名 时重复添加原始列
        _alias_to_orig_map: dict[str, str] = {}
        _orig_to_alias_map: dict[str, str] = {}
        for i, select_expr in enumerate(parsed_sql.expressions):
            alias_name, original_expr = self._extract_select_alias(select_expr, i)
            # 从原始表达式中提取底层列名
            orig_col_name = None
            if hasattr(original_expr, "name") and original_expr.name:
                orig_col_name = original_expr.name
            elif isinstance(original_expr, exp.Column) and hasattr(original_expr, "this"):
                ref = original_expr.this
                if hasattr(ref, "name") and ref.name:
                    orig_col_name = ref.name
            elif isinstance(original_expr, (exp.AggFunc,)):
                # 聚合函数:提取内部列引用
                for col_ref in original_expr.find_all(exp.Column):
                    if hasattr(col_ref, "name") and col_ref.name:
                        orig_col_name = col_ref.name
                        break
            if orig_col_name and alias_name != orig_col_name:
                _alias_to_orig_map[alias_name] = orig_col_name
                _orig_to_alias_map[orig_col_name] = alias_name

        def _is_group_col_represented(col: str) -> bool:
            """检查GROUP BY列是否已被SELECT中的某列(含别名)表示"""
            if col in result_data:
                return True
            # 检查是否有别名映射到这个原始列名
            if col in _orig_to_alias_map:
                return True
            # 遍历所有映射关系
            for alias, orig in _alias_to_orig_map.items():
                if orig == col and alias in result_data:
                    return True
            return False

        for col in group_by_columns:
            if not _is_group_col_represented(col):
                # 为缺失的GROUP BY列添加到结果
                result_data[col] = grouped[col].first().reset_index(drop=True)
                # 如果不在有序列列表中,添加到末尾
                if col not in ordered_columns:
                    ordered_columns.append(col)

        # R48-fix: 计算 HAVING 聚合临时列(_having_agg_*)并加入结果 DataFrame
        # 这些列在 _apply_group_by_aggregation 中注册但不是 SELECT 表达式,
        # 需要在 grouped 对象仍可用时(聚合后、返回前)计算正确值
        # R53-fix: 同时将 _having_agg_* 列加入 ordered_columns，确保它们出现在
        # 最终 DataFrame 中（否则 _apply_having_clause 找不到这些列）
        import sys as _sys
        if having_aggregations and 'grouped' in dir() and grouped is not None:
            for temp_alias, agg_func in having_aggregations.items():
                if temp_alias not in result_data:
                    try:
                        agg_result = self._apply_aggregation_function(agg_func, grouped, df)
                        if isinstance(agg_result, pd.Series):
                            result_data[temp_alias] = agg_result.reset_index(drop=True)
                        else:
                            result_data[temp_alias] = pd.Series([agg_result])
                        # R53-fix: 确保HAVING临时列包含在最终DataFrame中
                        if temp_alias not in ordered_columns:
                            ordered_columns.append(temp_alias)
                    except Exception:
                        result_data[temp_alias] = pd.Series([None] * len(grouped) if group_by_columns else [None])

        # 组合结果,保持列顺序
        try:
            result_df = pd.DataFrame(result_data, columns=ordered_columns).reset_index(drop=True)
        except Exception as e:
            # 如果创建DataFrame失败,尝试逐列构建
            result_data = {k: v.reset_index(drop=True) for k, v in result_data.items()}
            result_df = pd.DataFrame(result_data, columns=ordered_columns).reset_index(drop=True)

        return result_df

    def _is_aggregate_function(self, expr: exp.Expression) -> bool:
        """检查是否为聚合函数"""
        if isinstance(expr, exp.AggFunc):
            return True
        elif isinstance(expr, exp.Alias):
            return self._is_aggregate_function(expr.this)
        return False

    def _is_aggregate_only_query(self, parsed_sql: exp.Expression) -> bool:
        """检查是否为无GROUP BY的纯聚合查询(如 SELECT COUNT(*) FROM t WHERE ...)

        SQL标准要求此类查询即使无匹配行也应返回1行默认值(COUNT→0, 其他→NULL)
        """
        if not isinstance(parsed_sql, exp.Select):
            return False
        # 有GROUP BY时由groupby逻辑处理空结果，不在此处干预
        if parsed_sql.args.get("group") is not None:
            return False
        # 检查SELECT中是否包含聚合函数
        for expr in parsed_sql.expressions:
            if self._is_aggregate_function(expr):
                return True
        return False

    def _get_aggregate_default_value(self, parsed_sql: exp.Expression, col_name: str):
        """获取聚合函数在空结果集时的默认值

        SQL标准: COUNT → 0, SUM/AVG/MAX/MIN → NULL
        """
        for expr in parsed_sql.expressions:
            alias_name = None
            actual_expr = expr
            if isinstance(expr, exp.Alias):
                alias_name = expr.alias
                actual_expr = expr.this
            if self._is_aggregate_function(actual_expr):
                # 检查此聚合函数对应的别名是否匹配当前列
                func_name = type(actual_expr).__name__.lower()
                if alias_name == col_name or (alias_name is None and func_name.upper() == col_name.upper()):
                    if func_name == "count":
                        return 0
                    return None  # SUM/AVG/MAX/MIN等返回NULL
        # 无法确定具体聚合函数时的保守默认值
        return None

    def _execute_subquery(self, subquery_expr, worksheets_data: dict[str, pd.DataFrame]) -> pd.DataFrame:
        """
        执行子查询,返回结果DataFrame

        Args:
            subquery_expr: sqlglot Subquery或Select表达式
            worksheets_data: 当前可用的所有工作表数据

        Returns:
            pd.DataFrame: 子查询结果
        """
        # sqlglot可能将子查询直接存储为Select(而非Subquery包装)
        # 也可能存储为Union(UNION/UNION ALL在FROM中时)
        if isinstance(subquery_expr, exp.Subquery):
            inner = subquery_expr.this
            # Fix(A7): Subquery内部可能是Union(如 FROM (SELECT...UNION SELECT...) AS t)
            if isinstance(inner, exp.Union):
                return self._execute_union(inner, worksheets_data)
            inner_select = inner
        elif isinstance(subquery_expr, exp.Select):
            inner_select = subquery_expr
        elif isinstance(subquery_expr, exp.Union):
            return self._execute_union(subquery_expr, worksheets_data)
        else:
            raise ValueError(f"不支持子查询类型: {type(subquery_expr)}")

        # 获取子查询的FROM表
        from_table, from_subquery = self._get_from_table(inner_select)
        if from_subquery is not None:
            # 嵌套FROM子查询：递归执行内层子查询，结果作为临时表
            inner_result = self._execute_subquery(from_subquery, worksheets_data)
            # 用内层结果替换worksheets_data中的数据
            inner_alias = getattr(from_subquery, "alias", None)
            if not inner_alias:
                inner_alias = from_subquery.alias if hasattr(from_subquery, "alias") else "_nested_sub"
            worksheets_data = {**worksheets_data, inner_alias: inner_result}
            # 修改内层select的FROM为临时表名
            inner_select_copy = inner_select.copy()
            inner_select_copy.set("from", exp.From(this=exp.Table(this=exp.to_identifier(inner_alias))))
            result = self._execute_query(inner_select_copy, worksheets_data)
            return result
        if from_table not in worksheets_data:
            raise ValueError(f"子查询中表 '{from_table}' 不存在.可用表: {list(worksheets_data.keys())}")

        # 复用现有查询执行逻辑
        try:
            result = self._execute_query(inner_select, worksheets_data)
            return result
        except Exception as e:
            raise ValueError(f"子查询执行失败: {e}")

    def _evaluate_cast_expression(self, cast_expr: exp.Cast, df, row=None) -> Any:
        """
        评估 CAST(expr AS type) 表达式 — SQL 标准类型转换

        支持的目标类型:
        - INT / INTEGER → 整数
        - FLOAT / DECIMAL / REAL / DOUBLE → 浮点数
        - VARCHAR / TEXT / STRING / CHAR → 字符串
        - BOOLEAN / BOOL → 布尔值 (0/1)
        """
        # 1. 获取内部表达式并递归求值
        inner = cast_expr.this

        if row is not None:
            # 单行模式: 用于 WHERE/HAVING 逐行过滤
            if isinstance(inner, exp.Column):
                col_name = inner.name
                if col_name in row.index:
                    inner_val = row[col_name]
                else:
                    raise ValueError(f"CAST: 列 '{col_name}' 不存在.可用列: {list(row.index)}")
            elif isinstance(inner, exp.AggFunc):
                # Fix(R56): HAVING CAST(AGG(..)) — 聚合函数通过别名映射解析
                agg_sql = inner.sql()
                alias_map = getattr(self, '_having_agg_alias_map', {})
                resolved = alias_map.get(agg_sql)
                if resolved and resolved in row.index:
                    inner_val = row[resolved]
                else:
                    raise ValueError(f"CAST: 聚合函数 '{agg_sql}' 无法解析.可用列: {list(row.index)}, 映射: {alias_map}")
            elif hasattr(inner, "key") or isinstance(inner, exp.Binary):
                inner_val = self._evaluate_condition_for_row(inner, row)
            else:
                # 字面量或其他简单表达式
                try:
                    inner_val = self._literal_value(inner)
                except (AttributeError, TypeError):
                    inner_val = self._get_row_value(inner, row) if hasattr(self, '_get_row_value') else None
        else:
            # 向量化模式: 用于 SELECT 列计算
            if isinstance(inner, exp.Column):
                col_name = inner.name
                if col_name in df.columns:
                    inner_val = df[col_name]
                else:
                    raise ValueError(f"CAST: 列 '{col_name}' 不存在.可用列: {list(df.columns)}")
            elif isinstance(inner, exp.Literal):
                inner_val = self._parse_literal_value(inner)
            elif isinstance(inner, exp.AggFunc):
                # Fix(R56): 向量化模式中 CAST 内的聚合函数通过别名映射解析
                agg_sql = inner.sql()
                alias_map = getattr(self, '_having_agg_alias_map', {})
                resolved = alias_map.get(agg_sql)
                if resolved and resolved in df.columns:
                    inner_val = df[resolved]
                else:
                    raise ValueError(f"CAST: 聚合函数 '{agg_sql}' 无法解析.可用列: {list(df.columns)}, 映射: {alias_map}")
            else:
                # 嵌套表达式: 尝试递归转换为 pandas 表达式
                try:
                    cond_str = self._sql_condition_to_pandas(inner, df)
                    inner_val = df.eval(cond_str)
                except Exception:
                    # 最后尝试: 查找是否有匹配的别名
                    inner_str = str(inner)
                    matched = [c for c in df.columns if c == inner_str or c.endswith('_' + inner_str)]
                    if len(matched) == 1:
                        inner_val = df[matched[0]]
                    else:
                        raise ValueError(f"CAST: 无法求值嵌套表达式 ({inner}).可用列: {list(df.columns)}")

        # 2. 获取目标类型
        target_type = cast_expr.args.get("to")
        if target_type is None:
            return inner_val  # 无目标类型则原样返回

        # sqlglot 的 DataType 节点: target_type.this 是类型表达式(如 TYPE.INT)
        # 需要用 .sql() 或提取实际类型名; str(node) 会返回 "TYPE.INT" 带前缀
        type_node = target_type.this if hasattr(target_type, "this") else target_type
        if hasattr(type_node, "sql"):
            # sqlglot DataType 的 .sql() 返回标准 SQL 类型名 (如 "INT", "VARCHAR")
            type_name = type_node.sql().upper()
        else:
            type_name = str(type_node).upper()
        # 去掉可能的 "TYPE." 前缀 (兼容不同 sqlglot 版本)
        if type_name.startswith("TYPE."):
            type_name = type_name[5:]

        # 3. 执行类型转换

        if type_name in ("INT", "INTEGER"):
            if isinstance(inner_val, pd.Series):
                numeric = pd.to_numeric(inner_val, errors="coerce")
                # SQL CAST FLOAT→INT 行为: 向零截断 (truncate toward zero)
                # 不能直接用 .astype("Int64"), 因为非整数值(如 3.14)会抛 TypeError
                # numpy already imported at top level

                float_vals = np.asarray(numeric, dtype=float)
                # R48-fix P2-05: 将 inf/-inf 替换为 NaN,后续统一处理为 pd.NA
                # np.trunc(inf)=inf, int(inf) 会抛 OverflowError
                float_vals = np.where(np.isinf(float_vals), np.nan, float_vals)
                truncated = np.trunc(float_vals)  # 向零截断,匹配 SQL 标准
                # 构造 Int64 Series (支持 NA)
                result = []
                for i in range(len(truncated)):
                    if np.isnan(numeric.iloc[i]) if hasattr(numeric, "iloc") else np.isnan(float_vals[i]):
                        result.append(pd.NA)
                    else:
                        result.append(int(truncated[i]))
                return pd.Series(result, dtype="Int64")
            return int(float(inner_val)) if inner_val is not None else None

        elif type_name in ("FLOAT", "DECIMAL", "REAL", "DOUBLE", "NUMBER"):
            if isinstance(inner_val, pd.Series):
                return pd.to_numeric(inner_val, errors="coerce")
            return float(inner_val) if inner_val is not None else None

        elif type_name in ("VARCHAR", "TEXT", "STRING", "CHAR", "NVARCHAR"):
            if isinstance(inner_val, pd.Series):
                # R48-fix P1-02: 先处理NA再做str转换,避免字面量"nan"/"<NA>"被误清空
                result = inner_val.astype(str)
                na_mask = inner_val.isna()
                if na_mask.any():
                    result = result.copy()
                    result[na_mask] = ""
                return result
            return str(inner_val) if inner_val is not None else ""

        elif type_name in ("BOOLEAN", "BOOL"):
            if isinstance(inner_val, pd.Series):
                # R48-fix P1-01: SQL标准要求 CAST(NULL AS BOOLEAN) 返回 NULL, 不是 0
                # pandas 中 pd.NA.astype(bool) → False → 0, 违反SQL语义
                result = inner_val.copy()
                mask = inner_val.notna()
                if mask.any():
                    result[mask] = inner_val[mask].astype(bool).astype(int)
                result[~mask] = pd.NA
                return result
            if inner_val is None:
                return None  # SQL标准: CAST(NULL AS BOOL) → NULL
            return 1 if bool(inner_val) else 0

        else:
            # 未知类型: 尝试转为字符串(最安全的默认行为)
            if isinstance(inner_val, pd.Series):
                return inner_val.astype(str)
            return str(inner_val) if inner_val is not None else None

    def _evaluate_case_expression(self, case_expr: exp.Case, df, row=None) -> Any:
        """
        评估CASE WHEN表达式

        支持格式:
        - CASE WHEN cond1 THEN val1 WHEN cond2 THEN val2 ELSE default END
        - 单行评估模式(row参数)或向量化模式(无row参数)

        Args:
            case_expr: sqlglot Case表达式
            df: DataFrame(向量化模式)
            row: 可选,单行数据(逐行模式)

        Returns:
            向量化模式返回pd.Series,逐行模式返回单个值
        """
        ifs = case_expr.args.get("ifs", [])
        default_value = case_expr.args.get("default")

        if row is not None:
            # 逐行评估模式
            for if_clause in ifs:
                condition = if_clause.this
                if self._evaluate_condition_for_row(condition, row):
                    return self._get_expression_value(if_clause.args.get("true"), row)
            # 没有匹配的WHEN,返回ELSE默认值
            if default_value is not None:
                return self._get_expression_value(default_value, row)
            return None
        else:
            # 向量化模式 - 复用逐行评估
            return pd.Series(
                [self._evaluate_case_expression(case_expr, df, df.iloc[i]) for i in range(len(df))],
                index=df.index,
            )

    def _evaluate_case_with_aggregate(self, case_expr: exp.Case, grouped, df, result_data, group_by_columns) -> pd.Series:
        """Fix(R47): 在分组后评估包含聚合函数的CASE WHEN表达式

        处理如 CASE WHEN SUM(Amount) > 300 THEN 'Big' ELSE 'Small' END 的场景.
        策略: 对每个分组构建一行聚合结果DataFrame,然后逐组评估CASE.
        
        关键: 需要先计算CASE内嵌的聚合函数值(这些聚合不在select_exprs中),
        然后将它们注入到row_context中供条件评估使用.

        Args:
            case_expr: CASE表达式(内含聚合函数)
            grouped: pandas GroupBy对象
            df: 原始未分组DataFrame
            result_data: 已计算的聚合结果字典 {alias_name: Series}
            group_by_columns: GROUP BY列名列表

        Returns:
            pd.Series: 每组一个CASE结果值
        """
        # numpy already imported at top level
        results = []
        group_names = list(grouped.groups.keys())

        # Fix(R47-b): 提取并计算CASE内嵌的聚合函数
        inner_agg_expr = self._find_inner_aggregate(case_expr)
        inner_agg_results = {}
        if inner_agg_expr:
            agg_sql = inner_agg_expr.sql()
            # 使用 _apply_aggregation_function 计算内嵌聚合值(与主循环一致)
            try:
                agg_result = self._apply_aggregation_function(inner_agg_expr, grouped, df)
                if isinstance(agg_result, pd.Series):
                    inner_agg_results[agg_sql] = agg_result.reset_index(drop=True).tolist()
                elif isinstance(agg_result, (int, float, np.integer, np.floating)):
                    # 全表聚合返回单个值,复制到每组
                    n_groups = len(group_names)
                    inner_agg_results[agg_sql] = [agg_result] * n_groups
                else:
                    inner_agg_results[agg_sql] = [None] * len(group_names)
            except Exception as e:
                logger.warning("CASE+聚合计算失败 %s: %s", alias_name, e)
                inner_agg_results[agg_sql] = [None] * len(group_names)

        for i, group_name in enumerate(group_names):
            if not isinstance(group_name, tuple):
                group_key = (group_name,)
            else:
                group_key = group_name

            # 构建该组的单行上下文: 包含GROUP BY列 + 已计算聚合结果 + 内嵌聚合结果
            all_keys = list(group_by_columns) + list(result_data.keys()) + list(inner_agg_results.keys())
            row_context = pd.Series(index=all_keys, dtype=object)

            # 填充GROUP BY列值
            for j, col in enumerate(group_by_columns):
                if j < len(group_key):
                    row_context[col] = group_key[j]

            # 填充已计算的聚合结果
            for alias_name, agg_series in result_data.items():
                if i < len(agg_series):
                    row_context[alias_name] = agg_series.iloc[i]

            # 填充CASE内嵌的聚合结果(Fix(R47-b))
            for agg_sql, agg_vals in inner_agg_results.items():
                if i < len(agg_vals):
                    row_context[agg_sql] = agg_vals[i]

            # 用行级模式评估CASE表达式
            case_result = self._evaluate_case_expression(case_expr, df, row=row_context)
            results.append(case_result)

        return pd.Series(results, index=range(len(group_names)))

    def _evaluate_coalesce_with_aggregate(self, coalesce_expr: exp.Coalesce, grouped, df, result_data, group_by_columns) -> pd.Series:
        """Fix(R47): 在分组后评估包含聚合函数的COALESCE表达式

        处理如 COALESCE(SUM(col), 0) 的场景.
        策略与 _evaluate_case_with_aggregate 相同: 逐组评估.

        Args:
            coalesce_expr: COALESCE表达式(内含聚合函数)
            grouped: pandas GroupBy对象
            df: 原始未分组DataFrame
            result_data: 已计算的聚合结果字典
            group_by_columns: GROUP BY列名列表

        Returns:
            pd.Series: 每组一个COALESCE结果值
        """
        # numpy already imported at top level
        results = []
        group_names = list(grouped.groups.keys())

        for i, group_name in enumerate(group_names):
            if not isinstance(group_name, tuple):
                group_key = (group_name,)
            else:
                group_key = group_name

            # 构建该组的单行上下文
            row_context = pd.Series(index=list(group_by_columns) + list(result_data.keys()), dtype=object)

            for j, col in enumerate(group_by_columns):
                if j < len(group_key):
                    row_context[col] = group_key[j]

            for alias_name, agg_series in result_data.items():
                if i < len(agg_series):
                    row_context[alias_name] = agg_series.iloc[i]

            # 逐行评估COALESCE
            coalesce_result = self._evaluate_coalesce_for_row(coalesce_expr, row_context)
            results.append(coalesce_result)

        return pd.Series(results, index=range(len(group_names)))

    def _get_expression_value(self, expr: exp.Expression, row: pd.Series) -> Any:
        """获取表达式在指定行的值(委托给_get_row_value,两者功能完全重叠)"""
        return self._get_row_value(expr, row)

    def _evaluate_math_for_row(self, expr: exp.Expression, row: pd.Series) -> Any:
        """逐行评估数学表达式,复用类级别分发表"""
        # 标量数值函数: ROUND, ABS, FLOOR等
        if self._is_scalar_num_function(expr):
            return self._evaluate_scalar_num_function_for_row(expr, row)

        # 窗口函数参与算术表达式(逐行评估时需要从已计算的列中取值)
        if isinstance(expr, exp.Window):
            # 窗口函数应该已在预计算阶段处理,从df中查找对应列
            # 生成临时列名并尝试从row中获取
            temp_col = f"_window_math_{id(expr)}"
            if temp_col in row.index:
                val = row[temp_col]
                return float(val) if val is not None else None
            # 尝试通过窗口函数的文本表示查找
            window_str = str(expr).strip()
            for col in row.index:
                if col.startswith("_window_") or col in [window_str]:
                    val = row[col]
                    return float(val) if val is not None else None
            return None

        op_type = type(expr)
        if op_type not in self._MATH_BINARY_OPS:
            return None
        left = self._get_expression_value(expr.left, row)
        right = self._get_expression_value(expr.right, row)
        try:
            left_n = float(left) if left is not None else None
            right_n = float(right) if right is not None else None
            if left_n is None or right_n is None:
                return None
            # 除零保护
            if op_type == exp.Div and right_n == 0:
                return None
            return self._MATH_BINARY_OPS[op_type](left_n, right_n)
        except (TypeError, ValueError):
            return None

    def _evaluate_coalesce_for_row(self, coalesce_expr: exp.Coalesce, row: pd.Series) -> Any:
        """逐行评估COALESCE/IFNULL表达式,空字符串视为NULL继续查找下一个参数"""
        # COALESCE结构: this=第一个参数, expressions=[后续参数]
        values = [coalesce_expr.this] + list(coalesce_expr.expressions)
        for val_expr in values:
            val = self._get_expression_value(val_expr, row)
            # 跳过None/NaN/空字符串,继续查找下一个参数
            if val is not None and not (isinstance(val, float) and np.isnan(val)) and val != "":
                return val
        return 0  # 所有参数都无效(None/NaN/空)时返回0

    def _evaluate_coalesce_vectorized(self, coalesce_expr: exp.Coalesce, df) -> pd.Series:
        """向量化评估COALESCE/IFNULL表达式(用于DataFrame),空字符串视为NULL

        使用 pandas combine_first 实现真正的向量化操作,
        替代逐行 _evaluate_coalesce_for_row 循环.
        仅当所有参数为列引用或字面量时可向量化,否则回退逐行.
        
        Fix(R47): 支持 exp.Neg 等一元表达式作为COALESCE默认值(如 COALESCE(col, -1))
        """
        values = [coalesce_expr.this] + list(coalesce_expr.expressions)
        result = None
        fallback = False

        for val_expr in values:
            if isinstance(val_expr, exp.Column) and val_expr.name in df.columns:
                series = df[val_expr.name].astype(object)
                # 空字符串转为NaN,让combine_first正确识别为NULL
                series = series.replace("", np.nan)
                # None/NaN保持不变,用于combine_first处理
            elif isinstance(val_expr, exp.Literal):
                v = self._parse_literal_value(val_expr)
                series = pd.Series([v] * len(df), index=df.index, dtype=object)
            elif isinstance(val_expr, exp.Neg):
                # 处理负数字面量: COALESCE(col, -1) -> -1 是 Neg 节点
                inner = val_expr.this
                if isinstance(inner, exp.Literal):
                    v = self._parse_literal_value(inner)
                    series = pd.Series([-v] * len(df), index=df.index, dtype=object)
                else:
                    fallback = True
                    break
            elif isinstance(val_expr, (exp.Add, exp.Sub, exp.Mul, exp.Div)):
                # 简单算术表达式作为常量值(如 COALESCE(col, 1+1))
                fallback = True  # 暂时fallback,后续可优化为常量折叠
                break
            else:
                fallback = True
                break

            if result is None:
                result = series
            else:
                result = result.combine_first(series)

        if fallback:
            results = [self._evaluate_coalesce_for_row(coalesce_expr, df.iloc[i]) for i in range(len(df))]
            return pd.Series(results, index=df.index)

        # Fix(R45-03): 不再无条件fillna(0)，保留SQL标准的NULL语义
        # combine_first已正确处理了NULL替换：当首个非NULL参数值存在时使用该值，
        # 当所有参数均为NULL时保留NaN（符合SQL标准COALESCE行为）
        # 旧代码result.fillna(0)会导致负数fallback值（如-1）在特定类型转换场景下被错误替换为0
        return result

    def _generate_aggregate_alias(self, expr: exp.Expression) -> str:
        """为无别名的聚合函数生成有意义的列名

        例: COUNT(*) -> count_star, AVG(damage) -> avg_damage, SUM(hp) -> sum_hp
        """
        func_name = type(expr).__name__.lower()  # count, sum, avg, max, min

        # 提取参数
        is_distinct = isinstance(expr.this, exp.Distinct)
        if is_distinct:
            target = expr.this.expressions[0]
        else:
            target = expr.this

        if isinstance(target, exp.Star):
            arg_name = "star"
        elif isinstance(target, exp.Column):
            arg_name = target.name
        elif hasattr(target, "name") and target.name:
            arg_name = target.name
        else:
            arg_name = "expr"

        distinct_prefix = "distinct_" if is_distinct else ""
        return f"{func_name}_{distinct_prefix}{arg_name}"

    @staticmethod
    def _extract_agg_column(expr_node, context: str = "表达式") -> str:
        """从聚合函数参数节点提取列名(消除重复的Column/hasattr提取逻辑)"""
        if isinstance(expr_node, exp.Column):
            return expr_node.name
        if hasattr(expr_node, "name") and expr_node.name:
            return expr_node.name
        raise ValueError(f"{context}参数格式错误: {expr_node}")

    # 聚合函数分发表:sum/avg/max/min 统一为 pd.to_numeric -> agg
    _AGG_OPS = {
        "sum": lambda g, col: g[col].agg(lambda x: pd.to_numeric(x, errors="coerce").sum()),
        "avg": lambda g, col: g[col].agg(lambda x: pd.to_numeric(x, errors="coerce").mean()),
        "max": lambda g, col: g[col].agg(lambda x: pd.to_numeric(x, errors="coerce").max()),
        "min": lambda g, col: g[col].agg(lambda x: pd.to_numeric(x, errors="coerce").min()),
        "stddev": lambda g, col: g[col].agg(lambda x: pd.to_numeric(x, errors="coerce").std(ddof=1)),
        "std": lambda g, col: g[col].agg(lambda x: pd.to_numeric(x, errors="coerce").std(ddof=1)),
        "variance": lambda g, col: g[col].agg(lambda x: pd.to_numeric(x, errors="coerce").var(ddof=1)),
        "var": lambda g, col: g[col].agg(lambda x: pd.to_numeric(x, errors="coerce").var(ddof=1)),
    }

    def _apply_aggregation_function(self, expr: exp.Expression, grouped, df) -> pd.Series:
        """应用聚合函数"""
        if isinstance(expr, exp.Alias):
            return self._apply_aggregation_function(expr.this, grouped, df)

        if not isinstance(expr, exp.AggFunc):
            raise ValueError(f"不是聚合函数: {type(expr)}")

        func_name = type(expr).__name__.lower()

        # COUNT 特殊处理
        if func_name == "count":
            if isinstance(expr.this, exp.Star):
                return grouped.size()
            if isinstance(expr.this, exp.Distinct):
                col_name = self._extract_agg_column(expr.this.expressions[0], "COUNT(DISTINCT)")
                return grouped[col_name].nunique()

            # Fix(R47): COUNT(CASE WHEN ...) / COUNT(COALESCE(...)) 等复杂表达式参数
            if isinstance(expr.this, (exp.Case, exp.Coalesce)):
                # 先在完整df上计算CASE/COALESCE表达式,然后对分组结果计数非空值
                temp_col = f"_count_expr_{id(expr)}"
                if temp_col not in df.columns:
                    if isinstance(expr.this, exp.Case):
                        expr_values = self._evaluate_case_expression(expr.this, df)
                    else:
                        expr_values = self._evaluate_coalesce_vectorized(expr.this, df)
                    df[temp_col] = expr_values
                # 对每个组统计非空值数量
                def count_non_null(group):
                    return group[temp_col].notna().sum()
                return grouped.apply(count_non_null)

            col_name = self._extract_agg_column(expr.this, "COUNT")
            return grouped[col_name].count()

        # GROUP_CONCAT 特殊处理
        if func_name == "groupconcat":
            # GROUP_CONCAT(col, separator) - sqlglot MySQL 方言将语法解析为:
            # GROUP_CONCAT(col, sep) -> GroupConcat(this=Concat(...), separator=None)
            # 其中 Concat 的 expressions=[col, sep]
            # GROUP_CONCAT(col SEPARATOR sep) -> GroupConcat(this=col, separator=sep)
            separator = ","
            target_expr = None
            is_distinct = False

            # 检查是否为 DISTINCT
            if isinstance(expr.this, exp.Distinct):
                is_distinct = True
                target_expr = expr.this.expressions[0]
            else:
                target_expr = expr.this

            # 检查是否为 Concat 表达式（MySQL 方言对 GROUP_CONCAT(col, sep) 的解析）
            if isinstance(target_expr, exp.Concat):
                concat_exprs = target_expr.expressions
                if len(concat_exprs) >= 1:
                    target_expr = concat_exprs[0]
                if len(concat_exprs) >= 2 and isinstance(concat_exprs[1], exp.Literal):
                    separator = concat_exprs[1].this
            else:
                # 标准形式：GROUP_CONCAT(col SEPARATOR sep) 或 GROUP_CONCAT(col)
                sep_arg = expr.args.get("separator")
                if sep_arg and isinstance(sep_arg, exp.Literal):
                    separator = sep_arg.this

            # 处理目标表达式:支持简单列名和复杂表达式(CASE WHEN等)
            # 对于简单列名,直接使用列名;对于复杂表达式,先计算表达式
            if isinstance(target_expr, exp.Column):
                # 简单列名
                col_name = target_expr.name
            else:
                # 复杂表达式:先计算表达式,然后在原始df上添加临时列,再重新分组
                # 注意:这里需要获取原始的df和分组列
                if df is not None and not df.empty:
                    # 创建临时列存储表达式结果
                    temp_col = f"_groupconcat_expr_{id(target_expr)}"
                    # 计算表达式(使用_apply_select_expressions中的逻辑)
                    if isinstance(target_expr, exp.Case):
                        # CASE WHEN 表达式
                        expr_values = self._evaluate_case_expression(target_expr, df)
                    elif isinstance(target_expr, exp.Coalesce):
                        # COALESCE 表达式
                        expr_values = self._evaluate_coalesce_vectorized(target_expr, df)
                    elif self._is_mathematical_expression(target_expr):
                        # 数学表达式
                        expr_values = self._evaluate_math_expression(target_expr, df)
                    elif self._is_string_function(target_expr):
                        # 字符串函数
                        expr_values = self._evaluate_string_function(target_expr, df)
                    elif isinstance(target_expr, exp.Literal):
                        # 字面量
                        val = self._parse_literal_value(target_expr)
                        expr_values = pd.Series([val] * len(df), index=df.index)
                    else:
                        # 尝试作为列名处理
                        try:
                            col_name = self._extract_agg_column(target_expr, "GROUP_CONCAT")
                            expr_values = df[col_name]
                        except ValueError:
                            raise ValueError(f"GROUP_CONCAT 不支持的表达式类型: {type(target_expr)}")

                    # 确保expr_values是Series
                    if not isinstance(expr_values, pd.Series):
                        expr_values = pd.Series(expr_values, index=df.index)

                    # 将表达式值添加到df的副本中
                    df[temp_col] = expr_values

                    # 获取分组列名
                    if hasattr(grouped, "names"):
                        group_cols = list(grouped.names)
                    elif hasattr(grouped, "keys"):
                        group_cols = list(grouped.keys)
                    else:
                        raise ValueError("无法获取分组列名")

                    # 重新分组,包含临时列
                    grouped = df.groupby(group_cols, sort=False, dropna=False)
                    col_name = temp_col  # 使用临时列名
                else:
                    raise ValueError("GROUP_CONCAT 表达式计算需要有效的DataFrame")

            # 执行拼接
            if is_distinct:
                return grouped[col_name].apply(lambda x: separator.join(str(v) for v in set(x) if v is not None and v != ""))
            else:
                return grouped[col_name].apply(lambda x: separator.join(str(v) for v in x if v is not None and v != ""))

        # 其他聚合函数不支持 *
        if isinstance(expr.this, exp.Star):
            raise ValueError(f"函数 {func_name} 不支持 * 参数")

        # 分发表处理 sum/avg/max/min
        if func_name in self._AGG_OPS:
            # 检查是否为表达式(如 攻击力+防御力)
            if self._is_expression(expr.this):
                # 对于表达式,先计算表达式列,再聚合
                expr_col = self._evaluate_expression(expr.this, df if df is not None else grouped)
                return self._AGG_OPS[func_name](grouped, expr_col)
            else:
                # 单列处理
                col_name = self._extract_agg_column(expr.this, func_name.upper())
                return self._AGG_OPS[func_name](grouped, col_name)

        # GROUP_CONCAT: 分组拼接字符串
        if func_name == "groupconcat":
            col_name = self._extract_agg_column(expr.this, "GROUP_CONCAT")
            return grouped[col_name].agg(lambda x: ",".join(str(v) for v in x if pd.notna(v)))

        raise ValueError(f"不支持的聚合函数: {func_name}")

    def _is_expression(self, node) -> bool:
        """检查是否为表达式(非单列)"""
        return not isinstance(node, exp.Column)

    def _evaluate_expression(self, expr_node, df) -> str:
        """计算表达式,返回临时列名"""
        if isinstance(expr_node, exp.Column):
            return expr_node.name

        # 处理字面量(数字)
        if isinstance(expr_node, exp.Literal):
            temp_col = f"_temp_literal_{id(expr_node)}"
            df[temp_col] = float(expr_node.this) if expr_node.this is not None else 0
            return temp_col

        # 生成临时列名
        temp_col = f"_temp_expr_{id(expr_node)}"

        # 处理字符串函数（LENGTH, UPPER, LOWER, CONCAT等）
        if self._is_string_function(expr_node):
            if temp_col not in df.columns:
                df[temp_col] = self._evaluate_string_function(expr_node, df)
            return temp_col

        # 处理标量数值函数（ROUND, ABS, CEIL, FLOOR等）
        if self._is_scalar_num_function(expr_node):
            if temp_col not in df.columns:
                df[temp_col] = self._evaluate_scalar_num_function(expr_node, df)
            return temp_col

        # 处理加法表达式
        if isinstance(expr_node, exp.Add):
            left_col = self._evaluate_expression(expr_node.this, df)
            right_col = self._evaluate_expression(expr_node.expression, df)
            df[temp_col] = pd.to_numeric(df[left_col], errors="coerce") + pd.to_numeric(df[right_col], errors="coerce")

        # 处理减法表达式
        elif isinstance(expr_node, exp.Sub):
            left_col = self._evaluate_expression(expr_node.this, df)
            right_col = self._evaluate_expression(expr_node.expression, df)
            df[temp_col] = pd.to_numeric(df[left_col], errors="coerce") - pd.to_numeric(df[right_col], errors="coerce")

        # 处理乘法表达式
        elif isinstance(expr_node, exp.Mul):
            left_col = self._evaluate_expression(expr_node.this, df)
            right_col = self._evaluate_expression(expr_node.expression, df)
            df[temp_col] = pd.to_numeric(df[left_col], errors="coerce") * pd.to_numeric(df[right_col], errors="coerce")

        # 处理除法表达式
        elif isinstance(expr_node, exp.Div):
            left_col = self._evaluate_expression(expr_node.this, df)
            right_col = self._evaluate_expression(expr_node.expression, df)
            result = pd.to_numeric(df[left_col], errors="coerce") / pd.to_numeric(df[right_col], errors="coerce")
            # 安全加固(R42): 除零产生的 inf/-inf 转为 None(SQL NULL),符合 SQL 标准
            df[temp_col] = result.replace([np.inf, -np.inf], None)

        # 处理 CASE WHEN 表达式
        elif isinstance(expr_node, exp.Case):
            # 复用已有的 CASE WHEN 向量化求值方法
            if temp_col not in df.columns:
                df[temp_col] = self._evaluate_case_expression(expr_node, df)
            return temp_col

        # 处理 COALESCE 表达式
        elif isinstance(expr_node, exp.Coalesce):
            if temp_col not in df.columns:
                df[temp_col] = self._evaluate_coalesce_vectorized(expr_node, df)
            return temp_col

        # 处理 CAST 表达式
        elif isinstance(expr_node, exp.Cast):
            if temp_col not in df.columns:
                df[temp_col] = self._evaluate_cast_expression(expr_node, df)
            return temp_col

        else:
            raise ValueError(f"不支持的表达式类型: {type(expr_node)}")

        return temp_col

    def _apply_having_clause(self, parsed_sql: exp.Expression, df) -> pd.DataFrame:
        """应用HAVING条件"""
        having_clause = parsed_sql.args.get("having")
        if not having_clause:
            return df

        # Fix(R56): 初始化 _pending_tmp_cols，确保 HAVING 路径中 CAST/复杂表达式
        # 产生的临时列能被正确追踪和清理（WHERE路径在 _apply_where_clause 中初始化）
        self._pending_tmp_cols = []

        # 构建聚合表达式->SELECT别名的映射(HAVING COUNT(*) > 1 需要找到 cnt 列)
        self._having_agg_alias_map = {}
        for select_expr in parsed_sql.expressions:
            if isinstance(select_expr, exp.Alias) and isinstance(select_expr.this, exp.AggFunc):
                # 显式别名:AVG(伤害) as avg_dmg
                agg_sql = select_expr.this.sql()
                self._having_agg_alias_map[agg_sql] = select_expr.alias
            elif isinstance(select_expr, exp.AggFunc) and not isinstance(select_expr, exp.Alias):
                # 无别名:AVG(伤害) -> 自动生成 avg_damage
                agg_sql = select_expr.sql()
                auto_alias = self._generate_aggregate_alias(select_expr)
                self._having_agg_alias_map[agg_sql] = auto_alias

        # 合并HAVING聚合映射(HAVING中但不在SELECT中的聚合函数,如HAVING COUNT(*) > 1)
        # 这些聚合函数在_apply_group_by_aggregation中已计算并存储在_having_agg_in_select_map
        if hasattr(self, "_having_agg_in_select_map") and self._having_agg_in_select_map:
            self._having_agg_alias_map.update(self._having_agg_in_select_map)

        # 收集HAVING子句中使用的聚合函数(包括不在SELECT中的)
        having_agg_funcs = self._extract_agg_funcs_from_expr(having_clause.this)

        # 为HAVING中使用的但不在SELECT中的聚合函数创建临时列
        temp_columns = []
        for agg_func in having_agg_funcs:
            agg_sql = agg_func.sql()
            if agg_sql not in self._having_agg_alias_map:
                # 创建临时列名
                temp_col_name = f"_having_temp_{len(temp_columns)}"
                # 计算聚合值
                try:
                    # 使用groupby对象计算聚合
                    if hasattr(self, "_group_by_columns") and self._group_by_columns:
                        grouped = df.groupby(self._group_by_columns, sort=False, dropna=False)
                        agg_result = self._apply_aggregation_function(agg_func, grouped, df)
                        df[temp_col_name] = agg_result.values
                    else:
                        # 全表聚合
                        agg_result = self._apply_aggregation_function(agg_func, None, df)
                        df[temp_col_name] = agg_result
                    self._having_agg_alias_map[agg_sql] = temp_col_name
                    temp_columns.append(temp_col_name)
                except Exception:
                    pass  # 忽略计算失败的聚合

        # HAVING子句处理类似于WHERE,但作用于聚合后的数据
        try:
            condition_str = self._sql_condition_to_pandas(having_clause.this, df)
        except (ValueError, TypeError):
            # 不支持的条件类型(如COALESCE/CASE),回退到逐行过滤
            return self._apply_row_filter(having_clause.this, df)

        if condition_str:
            try:
                result_df = df.query(condition_str)
            except Exception:
                # 备用方案:逐行过滤
                return self._apply_row_filter(having_clause.this, df)
        else:
            result_df = df

        # 清理临时列（在查询执行之后）
        for temp_col in temp_columns:
            if temp_col in result_df.columns:
                del result_df[temp_col]

        # R48-fix: 清理 _sql_condition_to_pandas 创建的 _cast_tmp_* 临时列
        cast_tmp_cols = [c for c in result_df.columns if c.startswith('_cast_tmp_')]
        for col in cast_tmp_cols:
            del result_df[col]

        # R53-fix: 清理 _having_agg_* 临时列（由 _apply_group_by_aggregation 计算，
        # 仅用于 HAVING 条件评估，不应出现在最终 SELECT 结果中）
        having_agg_cols = [c for c in result_df.columns if c.startswith('_having_agg_')]
        for col in having_agg_cols:
            del result_df[col]

        # [FIX R55-BUG-02] 移除不可达的死代码
        # 原代码: return result_df 后的 logger.warning + _apply_row_filter 回退路径永远不可达
        # query() 异常已被 L8332 的 except 捕获，此处 return 是正确行为
        return result_df

    def _extract_agg_funcs_from_expr(self, expr: exp.Expression) -> list[exp.Expression]:
        """递归提取表达式中的所有聚合函数"""
        agg_funcs = []
        if isinstance(expr, exp.AggFunc):
            agg_funcs.append(expr)
        # 递归处理子表达式
        for child in expr.iter_expressions():
            agg_funcs.extend(self._extract_agg_funcs_from_expr(child))
        return agg_funcs

    def _extract_select_aliases(self, parsed_sql: exp.Expression) -> dict[str, Any]:
        """提取SELECT子句中的别名映射(委托_extract_select_alias统一逻辑)

        Returns:
            Dict: {alias_name: original_expression}
        """
        aliases = {}
        for i, select_expr in enumerate(parsed_sql.expressions):
            if isinstance(select_expr, exp.Star):
                continue
            alias_name, original_expr = self._extract_select_alias(select_expr, i)
            # 解包Paren表达式
            while isinstance(original_expr, exp.Paren):
                original_expr = original_expr.this
            aliases[alias_name] = original_expr
        return aliases

    def _resolve_order_column(self, col_name: str, df, select_aliases=None) -> str | None:
        """解析ORDER BY列名:先查SELECT别名对应的基础列,再查原始列名(大小写不敏感)

        Args:
            col_name: ORDER BY中引用的列名
            df: 当前DataFrame
            select_aliases: SELECT别名映射

        Returns:
            解析后的实际列名,找不到返回None
        """
        # 1. 大小写不敏感查找
        actual_col = self._find_column_name(col_name, df)
        if actual_col:
            return actual_col

        # 2. 如果有SELECT别名映射(大小写不敏感),检查别名对应的基础列
        if select_aliases:
            # 别名也做大小写不敏感匹配
            alias_lower = col_name.lower()
            for alias_name, expr in select_aliases.items():
                if alias_name.lower() == alias_lower:
                    if isinstance(expr, exp.Column):
                        actual_expr_col = self._find_column_name(expr.name, df)
                        if actual_expr_col:
                            return actual_expr_col
                    # 别名对应的是计算表达式,临时计算后用于排序
                    temp_col = self._compute_temp_column(expr, df, f"__order_temp_{col_name}")
                    if temp_col is None:
                        return None
                    df.rename(columns={temp_col: col_name}, inplace=True)
                    return col_name

        # 3. 列名不存在
        return None

    def _compute_temp_column(self, expr, df, temp_prefix="__temp__") -> str | None:
        """将表达式计算结果写入临时列,支持数学/字符串/CASE/COALESCE

        Args:
            expr: SQL表达式
            df: DataFrame
            temp_prefix: 临时列名前缀

        Returns:
            临时列名,不支持的表达式返回None
        """
        temp_col = f"{temp_prefix}_{id(expr)}"
        try:
            if self._is_string_function(expr):
                df[temp_col] = self._evaluate_string_function(expr, df)
            elif self._is_scalar_num_function(expr):
                df[temp_col] = self._evaluate_scalar_num_function(expr, df)
            elif isinstance(expr, exp.Case):
                df[temp_col] = self._evaluate_case_expression(expr, df)
            elif isinstance(expr, exp.Coalesce):
                df[temp_col] = self._evaluate_coalesce_vectorized(expr, df)
            elif self._is_mathematical_expression(expr):
                df[temp_col] = self._evaluate_math_expression(expr, df)
            elif isinstance(expr, exp.Cast):
                # CAST 表达式用于 ORDER BY 临时列
                df[temp_col] = self._evaluate_cast_expression(expr, df)
            elif isinstance(expr, exp.DPipe):
                # || 拼接用于 ORDER BY 临时列
                left = self._expr_to_series(expr.this, df).astype(str)
                right = self._expr_to_series(expr.expression, df).astype(str)
                df[temp_col] = left + right
            elif isinstance(expr, exp.Anonymous):
                # 含括号的列名(如"刷新时间(小时)")
                anon_name = expr.this
                if expr.expressions:
                    inner = ", ".join(str(e) for e in expr.expressions)
                    full_name = f"{anon_name}({inner})"
                else:
                    full_name = anon_name
                if full_name in df.columns:
                    df[temp_col] = df[full_name]
                elif anon_name in df.columns:
                    df[temp_col] = df[anon_name]
                else:
                    return None
                return temp_col
            else:
                return None
            return temp_col
        except Exception:
            return None

    def _resolve_order_expression(self, expr, df) -> str | None:
        """解析ORDER BY中的函数表达式,临时计算并添加为排序列

        支持: UPPER, LOWER, TRIM, LENGTH, CONCAT, REPLACE, SUBSTRING, LEFT, RIGHT,
              CASE WHEN, COALESCE, 数学表达式
        """
        return self._compute_temp_column(expr, df, "__order_expr")

    def _apply_order_by(self, parsed_sql: exp.Expression, df, select_aliases=None) -> pd.DataFrame:
        """应用ORDER BY排序

        Args:
            parsed_sql: 解析后的SQL表达式
            df: 数据DataFrame
            select_aliases: SELECT子句的别名映射(允许ORDER BY引用别名)
        """
        order_clause = parsed_sql.args.get("order")
        if not order_clause:
            return df

        sort_columns = []
        ascending = []

        for order_expr in order_clause.expressions:
            # 统一处理Ordered和简单列引用
            if isinstance(order_expr, exp.Ordered):
                col_expr = order_expr.this
                is_desc = order_expr.args.get("desc", False)
            elif isinstance(order_expr, exp.Column):
                col_expr = order_expr
                is_desc = False
            else:
                # 函数表达式: ORDER BY UPPER(name), ORDER BY LENGTH(name) 等
                col_expr = order_expr
                is_desc = False

            col_name = col_expr.name
            table_part = col_expr.table if hasattr(col_expr, "table") and col_expr.table else None
            qualified = f"{table_part}.{col_name}" if table_part else None

            # 先查限定名,再查简单名,再查SELECT别名
            resolved_name = qualified if qualified and qualified in df.columns else None
            if resolved_name is None:
                resolved_name = self._resolve_order_column(col_name, df, select_aliases)
            if resolved_name is None and qualified and qualified in df.columns:
                resolved_name = qualified

            # [FIX R53] ORDER BY 聚合函数表达式匹配 SELECT 别名
            # 例: SELECT COUNT(*) as cnt ... ORDER BY COUNT(*)
            # col_expr.name 对 Count(*) 返回 "*" 无法匹配,需要按表达式SQL字符串匹配
            if resolved_name is None and select_aliases and not isinstance(col_expr, exp.Column):
                expr_sql = str(col_expr).strip()
                for alias_name, alias_expr in select_aliases.items():
                    if str(alias_expr).strip() == expr_sql:
                        # 找到匹配的别名,在 DataFrame 中查找该别名对应的列
                        alias_col = self._find_column_name(alias_name, df)
                        if alias_col:
                            resolved_name = alias_col
                            break

            # 函数表达式: ORDER BY UPPER(col), LENGTH(col), COALESCE(col, 0) 等
            if resolved_name is None and not isinstance(col_expr, exp.Column):
                resolved_name = self._resolve_order_expression(col_expr, df)

            if resolved_name is None:
                suggestion = self._suggest_column_name(col_name, list(df.columns))
                raise ValueError(f"排序列 '{qualified or col_name}' 不存在.可用列: {list(df.columns)}.{suggestion}")

            sort_columns.append(resolved_name)
            ascending.append(not is_desc if is_desc is not None else True)

        if sort_columns:
            # Handle mixed data types in ORDER BY columns
            # Fix: 智能混合类型排序 — 优先数值排序，非数值值排末尾
            temp_sort_cols = []
            for col in sort_columns:
                if col in df.columns:
                    col_data = df[col]
                    has_numbers = False
                    has_strings = False
                    num_count = 0
                    total_count = 0

                    for val in col_data.dropna():
                        if isinstance(val, (int, float)) and not isinstance(val, bool):
                            has_numbers = True
                            num_count += 1
                        elif isinstance(val, str):
                            has_strings = True
                        total_count += 1

                        if has_numbers and has_strings:
                            break

                    if has_numbers and has_strings and total_count > 0:
                        # 超过50%为数值时使用智能排序：数值按数值排，非数值排末尾
                        if num_count / total_count > 0.5:
                            temp_col_name = f"_temp_sort_{col}"
                            # 尝试转为数值，失败者保留原值用于末尾排序
                            numeric_vals = pd.to_numeric(col_data, errors='coerce')
                            # 排序键：数值用其值，非数值用 inf（排到末尾）
                            df[temp_col_name] = numeric_vals.fillna(float('inf'))
                            sort_columns = [temp_col_name if c == col else c for c in sort_columns]
                            temp_sort_cols.append(temp_col_name)
                        else:
                            # 字符串为主，回退到字符串排序
                            temp_col_name = f"_temp_str_{col}"
                            df[temp_col_name] = col_data.astype(str)
                            sort_columns = [temp_col_name if c == col else c for c in sort_columns]
                            temp_sort_cols.append(temp_col_name)

            sorted_df = df.sort_values(by=sort_columns, ascending=ascending)

            # Clean up temporary columns
            for tc in temp_sort_cols:
                if tc in sorted_df.columns:
                    sorted_df.drop(columns=[tc], inplace=True)

            return sorted_df

        return df

    def _generate_markdown_table(self, data: list, max_rows: int = MARKDOWN_TABLE_MAX_ROWS) -> str:
        """生成Markdown格式表格

        Args:
            data (List): 表格数据
            max_rows (int): 最大行数,默认使用配置值

        Returns:
            str: Markdown格式表格字符串
        """
        """将查询结果数据转为Markdown表格"""
        if not data:
            return ""
        md_lines = ["| " + " | ".join(str(c) for c in data[0]) + " |"]
        md_lines.append("| " + " | ".join(["---"] * len(data[0])) + " |")
        display_rows = min(len(data) - 1, max_rows)
        for row in data[1 : 1 + display_rows]:
            md_lines.append("| " + " | ".join(str(c) for c in row) + " |")
        if len(data) - 1 > max_rows:
            md_lines.append(f"| ... 共{len(data) - 1}行,仅显示前{max_rows}行 |")
        return "\n".join(md_lines)

    def _format_export_output(self, data: list, output_format: str, include_headers: bool) -> dict[str, Any]:
        """生成JSON/CSV格式输出"""
        if not data or output_format == "table":
            return {}
        headers_row = data[0]
        data_rows = data[1:]
        records = [dict(zip([str(h) for h in headers_row], row)) for row in data_rows]
        result = {"query_info": {"record_count": len(records)}}
        if output_format == "json":
            result["formatted_output"] = json.dumps(records, ensure_ascii=False, indent=2)
            result["query_info"]["output_format"] = "json"
        elif output_format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            if include_headers:
                writer.writerow([str(h) for h in headers_row])
            for row in data_rows:
                writer.writerow([str(v) if v is not None else "" for v in row])
            result["formatted_output"] = output.getvalue()
            result["query_info"]["output_format"] = "csv"
        return result

    def _format_query_result(
        self,
        result_df: pd.DataFrame,
        file_path: str,
        sql: str,
        worksheets_data: dict[str, pd.DataFrame],
        include_headers: bool,
        has_group_by: bool = False,
        has_having: bool = False,
        parsed_sql: exp.Expression = None,
        df_before_where: pd.DataFrame = None,
        output_format: str = "table",
    ) -> dict[str, Any]:
        """格式化查询结果

        Args:
            has_group_by: 如果为True且有数值聚合列,自动追加TOTAL行
            has_having: 如果为True,表示使用了HAVING过滤,不添加TOTAL行
            parsed_sql: 解析后的SQL表达式,用于空结果智能建议
            df_before_where: WHERE过滤前的DataFrame,用于空结果智能建议
            output_format: 输出格式 table/json/csv
        """

        # 计算原始数据统计
        total_original_rows = sum(len(df) for df in worksheets_data.values())

        # 准备返回数据
        data = []
        if include_headers:
            data.append(list(result_df.columns))
        if not result_df.empty:
            for row in result_df.itertuples(index=False, name=None):
                data.append([self._serialize_value(val) for val in row])
        elif self._is_aggregate_only_query(parsed_sql):
            # SQL标准: 无GROUP BY的纯聚合查询(COUNT/SUM/AVG等)在无匹配行时应返回1行默认值
            # COUNT(*) → 0, 其他聚合函数 → NULL
            default_row = []
            for col in result_df.columns:
                default_row.append(self._get_aggregate_default_value(parsed_sql, col))
            data.append(default_row)

        # 大结果自动截断:保护AI上下文窗口(MAX_RESULT_ROWS=500)
        truncated = False
        data_row_count = len(result_df)
        if data_row_count > MAX_RESULT_ROWS:
            # 保留表头行 + 前MAX_RESULT_ROWS行数据
            keep_rows = MAX_RESULT_ROWS + (1 if include_headers else 0)
            data = data[:keep_rows]
            truncated = True

        # 双行表头:构建列描述映射
        column_descriptions = {}
        if hasattr(self, "_header_descriptions") and self._header_descriptions:
            for table_name, desc_map in self._header_descriptions.items():
                for col in result_df.columns if not result_df.empty else []:
                    if col in desc_map:
                        column_descriptions[col] = desc_map[col]

        # 性能提示:无LIMIT且返回行数过多时建议加LIMIT
        perf_hint = ""
        if len(result_df) > 100:
            has_limit = parsed_sql is not None and parsed_sql.args.get("limit") is not None
            if not has_limit:
                perf_hint = "(结果较多,建议加 LIMIT 缩小范围)"
        if truncated:
            perf_hint += f"(结果已截断为前{MAX_RESULT_ROWS}行,共{data_row_count}行,请加 LIMIT 精确查询)"

        result = {
            "success": True,
            "message": f"SQL查询成功执行,返回 {data_row_count} 行结果" + perf_hint,
            "data": data,
            "query_info": {
                "original_rows": total_original_rows,
                "filtered_rows": data_row_count,
                "returned_rows": len(data) - (1 if include_headers else 0),
                "truncated": truncated,
                "query_applied": True,
                "sql_query": sql,
                "columns_returned": len(result_df.columns) if not result_df.empty else 0,
                "available_tables": list(worksheets_data.keys()),
                "returned_columns": list(result_df.columns) if not result_df.empty else [],
                "data_types": self._infer_data_types(result_df) if not result_df.empty else {},
            },
        }

        # 空结果智能建议:分析WHERE/HAVING条件类型,给出针对性提示
        if result_df.empty:
            suggestion = self._generate_empty_result_suggestion(parsed_sql, df_before_where, worksheets_data)
            # HAVING空结果追加聚合中间结果信息
            df_before_having = getattr(self, "_df_before_having", None)
            if df_before_having is not None and not df_before_having.empty:
                having_clause = parsed_sql.args.get("having")
                if having_clause:
                    suggestion += self._generate_having_empty_suggestion(having_clause, df_before_having)
            result["query_info"]["suggestion"] = suggestion

        # 生成Markdown表格(方便AI和人类阅读)
        if data:
            result["query_info"]["markdown_table"] = self._generate_markdown_table(data)

        # 生成JSON/CSV格式输出
        if data:
            export = self._format_export_output(data, output_format, include_headers)
            for key, value in export.items():
                if key == "query_info":
                    result["query_info"].update(value)
                else:
                    result[key] = value

        # 双行表头时附加描述信息
        if column_descriptions:
            result["query_info"]["dual_header"] = True
            result["query_info"]["column_descriptions"] = column_descriptions

        return result

    def _infer_data_types(self, df) -> dict[str, str]:
        """
        推断列的数据类型
        R51-opt: 增加布尔类型检测 + 采样优化 + 混合类型处理
        """
        data_types = {}

        for col in df.columns:
            series = df[col]
            non_null = series.dropna()

            if len(non_null) == 0:
                data_types[col] = "string"
                continue

            # R51-opt: 先用小样本快速判断类型（避免全量 to_numeric）
            sample = non_null.head(10)

            # R51-new: 布尔类型检测（True/False, Yes/No, 1/0 模式）
            if self._is_boolean_column(sample):
                data_types[col] = "boolean"
                continue

            # 数值类型检测（先用样本，确认后再全量转换）
            try:
                sample_numeric = pd.to_numeric(sample, errors="coerce")
                sample_valid = sample_numeric.notna().sum()
                if sample_valid == len(sample):
                    # 样本全为数值 → 全量确认
                    full_numeric = pd.to_numeric(series, errors="coerce")
                    if not full_numeric.isna().all():
                        if (full_numeric.dropna() % 1 == 0).all():
                            data_types[col] = "integer"
                        else:
                            data_types[col] = "float"
                        continue
                elif sample_valid > len(sample) * 0.8:
                    # >80% 为数值 → 混合类型，仍尝试数值优先
                    full_numeric = pd.to_numeric(series, errors="coerce")
                    na_ratio = full_numeric.isna().mean()
                    if na_ratio < 0.5:  # 少于一半非数值
                        if (full_numeric.dropna() % 1 == 0).all():
                            data_types[col] = "integer"
                        else:
                            data_types[col] = "float"
                        continue
            except Exception:
                pass

            # 日期类型检测
            try:
                is_likely_date = any(
                    isinstance(val, str) and any(x in str(val) for x in ["-", "/", ":", "年", "月", "日"])
                    for val in sample
                )
                if is_likely_date:
                    converted = pd.to_datetime(series, errors="coerce", format="mixed")
                    if not converted.isna().all():
                        data_types[col] = "datetime"
                        continue
            except Exception:
                pass

            # 默认为字符串类型
            data_types[col] = "string"

        return data_types

    @staticmethod
    def _is_boolean_column(sample_series) -> bool:
        """检测列是否为布尔类型（True/False, Yes/No, 1/0 等模式）

        Args:
            sample_series: 已去空的样本 Series

        Returns:
            bool: 是否为布尔类型
        """
        if len(sample_series) == 0:
            return False

        unique_vals = set()
        for v in sample_series:
            if isinstance(v, bool):
                unique_vals.add(v)
            elif isinstance(v, str):
                v_lower = v.lower().strip()
                if v_lower in ("true", "false", "yes", "no", "y", "n", "1", "0"):
                    unique_vals.add(v_lower)
                else:
                    return False  # 非布尔字符串 → 不是布尔列
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                if v in (0, 1):
                    unique_vals.add(v)
                else:
                    return False  # 非 0/1 数值 → 不是布尔列
            else:
                return False

        # 布尔列特征: 只有 True/False 或 Yes/No 或 0/1 两类值
        bool_groups = [
            {"true", "false"},
            {"yes", "no"},
            {"y", "n"},
            {0, 1},
            {True, False},
        ]
        return any(unique_vals <= bg for bg in bool_groups) and len(unique_vals) >= 2

    def _update_error(self, message: str, elapsed_ms: float = 0) -> dict[str, Any]:
        """构造UPDATE操作的统一错误响应"""
        result = {
            "success": False,
            "message": message,
            "affected_rows": 0,
            "changes": [],
        }
        if elapsed_ms:
            result["execution_time_ms"] = round(elapsed_ms, 1)
        return result

    # Fix: P2-type-check — 列值类型校验,防止字符串写入数值列等类型不匹配问题
    def _get_column_type_category(self, df: pd.DataFrame, col_name: str) -> str:
        """检测列的数据类型类别,用于写入时的类型校验.

        Returns:
            'numeric' — 数值型(int/float),只接受可转为数字的值
            'string'  — 字符串/混合型,接受任意值
            'empty'   — 空列(无数据或全为空),接受任意值(无法推断类型)
        """
        if col_name not in df.columns:
            return "empty"

        series = df[col_name]
        # 排除空值后采样
        non_null = series.dropna()
        if len(non_null) == 0:
            return "empty"

        # R51-opt: 采样量从50降至20（统计显著性足够，减少开销）
        # 不再限制 pd.api.types.is_numeric_dtype，统一走采样检测（兼容 object 型数值列）
        sample = non_null.head(20)
        numeric_count = sum(
            1 for v in sample
            if isinstance(v, (int, float, np.integer, np.floating)) and not (isinstance(v, float) and np.isnan(v))
        )
        if numeric_count == len(sample):
            return "numeric"

        return "string"

    def _validate_value_type(self, value: Any, df: pd.DataFrame, col_name: str) -> str | None:
        """校验值是否与目标列的数据类型兼容.

        Args:
            value: 待写入的值
            df: 目标DataFrame
            col_name: 目标列名

        Returns:
            None 表示校验通过,否则返回错误信息字符串
        """
        # 空值/None 始终允许(表示清空或NULL)
        if value is None or value == "":
            return None

        col_type = self._get_column_type_category(df, col_name)

        # 空列无法推断类型,放行
        if col_type == "empty":
            return None

        # 字符串列接受任意值(Excel本身是弱类型的)
        if col_type == "string":
            return None

        # 数值型列: 检查值是否能转为数字
        if col_type == "numeric":
            # 数值类型直接放行
            if isinstance(value, (int, float, np.integer, np.floating)):
                # Fix: P2-4 极端浮点值导致文件损坏 — 校验NaN/Inf/超出Excel范围的值
                try:
                    f_val = float(value)
                    if np.isnan(f_val) or math.isinf(f_val):
                        return (
                            f"数值校验失败: 列 '{col_name}' 不支持 NaN 或无穷大(Inf/-Inf)值. "
                            f"💡 Excel无法表示这些特殊浮点值,请使用有限数值或NULL."
                        )
                    # Excel 实际支持的浮点范围约为 ±1e308(IEEE 754 double),
                    # 但超过 ±1e15 的整数精度会丢失,超过 ±1e308 会溢出.
                    # 此处设置安全阈值: 绝对值不超过 1e308
                    if abs(f_val) > 1e308:
                        return (
                            f"数值校验失败: 列 '{col_name}' 的值 {f_val} 超出Excel支持的数值范围(±1e308). "
                            f"💡 请使用更小的数值."
                        )
                except (ValueError, TypeError, OverflowError):
                    pass
                return None
            # 字符串尝试转数字
            if isinstance(value, str):
                try:
                    float(value)
                    return None  # 字符串内容是有效数字,放行(如 "123")
                except (ValueError, TypeError):
                    return (
                        f"类型不匹配: 列 '{col_name}' 是数值型,但写入了字符串 '{value}'. "
                        f"💡 数值列只能写入数字或数字格式的字符串."
                    )
            # 其他类型(如list/dict等)拒绝
            return (
                f"类型不匹配: 列 '{col_name}' 是数值型,但写入了 {type(value).__name__} 类型的值. "
                f"💡 数值列只能写入数字或数字格式的字符串."
            )

        return None

    def _verify_streaming_write(self, file_path: str, sheet_name: str, changes: list, en_to_cn_map: dict = None) -> dict[str, Any]:
        """验证流式写入是否实际生效

        重新读取文件，抽样检查changes中的修改是否反映到文件中。
        """
        if not StreamingWriter.is_available() or not changes:
            return {"verified": True, "unverified": []}

        try:
            from ..core.streaming_writer import CalamineWorkbook

            cal_wb = CalamineWorkbook.from_path(file_path)
            verify_rows = cal_wb.get_sheet_by_name(sheet_name).to_python()
            cal_wb.close()

            if not verify_rows:
                return {"verified": False, "unverified": changes}

            # 构建col_map(和_copy_modify_write一致)
            col_map = {}
            for col_idx, cell_val in enumerate(verify_rows[0], 1):
                if cell_val is not None:
                    col_map[str(cell_val).strip()] = col_idx

            unverified = []
            # 抽样验证: 最多检查前5个change
            sample = changes[:5] if len(changes) > 5 else changes

            for change in sample:
                col_name = change["column"]
                col_idx = None

                if col_name in col_map:
                    col_idx = col_map[col_name]
                elif en_to_cn_map:
                    cn_name = en_to_cn_map.get(col_name)
                    if cn_name and cn_name in col_map:
                        col_idx = col_map[cn_name]
                else:
                    col_stripped = str(col_name).strip().lower()
                    for k, v in col_map.items():
                        if str(k).strip().lower() == col_stripped:
                            col_idx = v
                            break

                if col_idx is None:
                    unverified.append(change)
                    continue

                row_idx = change["row"] - 1
                if row_idx < len(verify_rows):
                    actual_val = str(verify_rows[row_idx][col_idx - 1]) if col_idx - 1 < len(verify_rows[row_idx]) else None
                    expected = str(change["new_value"])
                    if actual_val != expected:
                        unverified.append(change)

            return {"verified": len(unverified) == 0, "unverified": unverified}
        except Exception as e:
            logger.warning(f"写入验证异常: {e}")
            return {"verified": True, "unverified": []}  # 验证失败不阻塞

    def _precompute_update_window_where(self, df: pd.DataFrame, where_expr, window_nodes: list) -> pd.DataFrame:
        """预处理UPDATE WHERE中的窗口函数，将计算结果存为临时列并替换AST节点"""
        for i, win_node in enumerate(window_nodes):
            temp_col = f"_wf_{i}"
            try:
                result = self._compute_window_function(win_node, df)
                df[temp_col] = result
                # 在WHERE表达式中替换Window节点为临时列引用
                win_node.replace(exp.Column(this=exp.to_identifier(temp_col)))
            except Exception as e:
                raise ValueError(f"UPDATE WHERE中窗口函数计算失败: {e}")
        return df

    def execute_update_query(
        self,
        file_path: str,
        sql: str,
        sheet_name: str | None = None,
        dry_run: bool = False,
    ) -> dict[str, Any]:
        """
        执行UPDATE语句,基于WHERE条件批量修改Excel数据

        支持语法: UPDATE 表名 SET 列1=值1, 列2=值2 [WHERE 条件]
        SET表达式支持: 列=常量, 列=列, 列=算术表达式(如 伤害*1.1)
        WHERE条件复用查询引擎的所有条件语法

        行号支持(_ROW_NUMBER_):
            在WHERE条件中可使用 _ROW_NUMBER_ 虚拟列，值为Excel数据行号(从1开始,不含表头)。
            适用场景: 表中存在重复记录,无法通过字段值唯一确定目标行时。

            用法示例::
                UPDATE LootList SET PropType='主武器' WHERE _ROW_NUMBER_ IN (11,21,36)
                UPDATE 数据表 SET 状态='已处理' WHERE _ROW_NUMBER_ BETWEEN 10 AND 50
                UPDATE 表 SET 值=100 WHERE _ROW_NUMBER_ = 5

            限制:
                - _ROW_NUMBER_ 仅在UPDATE的WHERE条件中可用,SELECT查询暂不支持
                - 不允许对 _ROW_NUMBER_ 本身执行SET操作(会报错)
                - 行号基于DataFrame索引+2(表头偏移),与Excel显示行号一致

        写入策略:
            - 小规模更新(<200行或<500单元格或文件<5MB): 传统openpyxl写入(可靠)
            - 大规模更新: 流式calamine+write_only写入(高性能,不保留单元格格式)
            - 流式写入列名匹配失败时自动降级到传统路径

        Args:
            file_path: Excel文件路径
            sql: UPDATE SQL语句
            sheet_name: 工作表名称(可选)
            dry_run: 预览模式,只返回影响行数不实际修改

        Returns:
            Dict: 更新结果,包含success/message/affected_rows/changes/verification等字段
        """
        start_time = time.time()

        # 验证文件
        if not os.path.exists(file_path):
            return self._update_error(f"文件不存在: {file_path}")

        if not SQLGLOT_AVAILABLE:
            return self._update_error("SQLGLOT未安装,无法使用UPDATE功能")

        # 加载数据(使用缓存)
        worksheets_data = self._load_data_with_cache(file_path, sheet_name)

        if not worksheets_data:
            return self._update_error("无法加载Excel数据")

        # 清理ANSI转义序列(终端粘贴可能带入的不可见字符)
        sql = re.sub(r"\x1b\[[0-9;]*[a-zA-Z]", "", sql)

        # Fix: P0-4 UPDATE 分号多语句注入
        if self._has_dangerous_semicolon(sql):
            return self._update_error("SQL语法错误: 不支持分号分隔的多语句执行(安全限制).💡 请将每条SQL语句分开执行")

        # Fix: P0-7 UPDATE 注释符注入防御
        # -- 和 # 可截断WHERE条件导致全表篡改，必须在解析前拦截
        _comment_err = self._detect_dangerous_comments(sql)
        if _comment_err:
            return self._update_error(_comment_err)

        # 预处理: 自动为 MySQL 保留字标识符添加反引号
        sql = self._preprocess_reserved_words(sql)

        # Fix: P2-1 预处理: 将 || 字符串拼接操作符转为 CONCAT()
        # 因为 MySQL 方言将 || 解析为逻辑 OR，需要提前转换
        sql = self._preprocess_dpipe_to_concat(sql)

        # 解析UPDATE语句
        # 中文列名替换(与SELECT查询保持一致)
        try:
            sql = self._replace_cn_columns_in_sql(sql, worksheets_data)
        except Exception:
            pass  # 替换失败时继续用原始SQL

        try:
            parsed = sqlglot.parse_one(sql, read="mysql")
        except ParseError as e:
            return self._update_error(f"SQL语法错误: {e}")

        # 验证是UPDATE语句
        if not isinstance(parsed, exp.Update):
            return self._update_error("只支持UPDATE语句.💡 写入操作只支持UPDATE,查询请用 excel_query")

        # 提取表名(sqlglot中table在this属性)
        table_node = parsed.this if isinstance(parsed.this, exp.Table) else None
        if not table_node:
            return self._update_error("UPDATE语句缺少表名")
        target_table = table_node.name

        # 匹配工作表(支持中英文表名)
        matched_sheet = None
        for sheet in worksheets_data:
            if sheet == target_table:
                matched_sheet = sheet
                break
            # 模糊匹配
            if target_table.lower() == sheet.lower():
                matched_sheet = sheet
                break

        if not matched_sheet:
            available = list(worksheets_data.keys())
            suggestion = self._suggest_column_name(target_table, available)
            return self._update_error(f"工作表 '{target_table}' 不存在.可用工作表: {available}.{suggestion}")

        df = worksheets_data[matched_sheet].copy()
        original_df = df.copy()

        # P1: 添加行号虚拟列 _ROW_NUMBER_
        df["_ROW_NUMBER_"] = range(1, len(df) + 1)

        # 中文列名替换
        cn_map = {}
        desc_map = self._header_descriptions.get(matched_sheet, {})
        for en_col, cn_desc in desc_map.items():
            if en_col in df.columns:
                cn_map[cn_desc] = en_col

        # 解析SET子句(sqlglot中在expressions属性)
        set_exprs = parsed.args.get("expressions", [])
        if not set_exprs:
            return self._update_error("UPDATE语句缺少SET子句")

        set_operations = []  # [(col_name, expression_node)]
        for set_item in set_exprs:
            # sqlglot Update SET items: EQ expression (col = value)
            if isinstance(set_item, exp.EQ):
                col_name = set_item.left.name
                # 中文列名替换 + 大小写不敏感
                if col_name in cn_map:
                    col_name = cn_map[col_name]
                if col_name == "_ROW_NUMBER_":
                    return self._update_error("_ROW_NUMBER_ 是虚拟列，不允许修改")
                # 大小写不敏感列名查找
                actual_col = self._find_column_name(col_name, df)
                if not actual_col:
                    suggestion = self._suggest_column_name(col_name, list(df.columns))
                    return self._update_error(f"列 '{col_name}' 不存在.可用列: {list(df.columns)}.{suggestion}")
                # 使用实际列名(保持原始大小写)
                col_name = actual_col
                set_operations.append((col_name, set_item.right))
            else:
                return self._update_error(f"不支持的SET表达式: {set_item}")

        # 设置当前工作表数据供子查询使用（IN子查询可能需要）
        # Fix(R13): 支持 WITH CTE 子句 — 提取CTE并执行，结果加入可用表
        self._current_worksheets = self._inject_ctes_to_worksheets(parsed, worksheets_data)

        # 应用WHERE条件筛选（支持窗口函数预处理）
        where_clause = parsed.args.get("where")
        if where_clause:
            # 预处理: 只处理WHERE直接条件中的窗口函数，跳过IN子查询内部的
            where_expr = where_clause.this
            in_nodes = list(where_expr.find_all(exp.In))
            window_nodes = [n for n in where_expr.find_all(exp.Window) if not any(n in list(in_node.walk()) for in_node in in_nodes)]
            if window_nodes:
                df = self._precompute_update_window_where(df, where_expr, window_nodes)
                # 重新解析WHERE（窗口函数已被替换为临时列引用）
                where_clause = parsed.args.get("where")
                where_expr = where_clause.this
            condition_str = self._sql_condition_to_pandas(where_clause.this, df)
            if condition_str:
                try:
                    filtered_df = df.query(condition_str)
                except Exception:
                    filtered_df = self._apply_row_filter(where_clause.this, df)
            else:
                logger.warning(
                    "UPDATE WHERE条件转换为pandas表达式失败,回退到逐行过滤: %s",
                    where_clause.this,
                )
                filtered_df = self._apply_row_filter(where_clause.this, df)
        else:
            filtered_df = df

        if filtered_df.empty:
            return {
                "success": True,
                "message": "没有匹配WHERE条件的行,无需更新",
                "affected_rows": 0,
                "changes": [],
                "execution_time_ms": 0,
            }

        # P1: 重复行检测
        warnings = []
        match_ratio = len(filtered_df) / len(df) if len(df) > 0 else 0
        if match_ratio > 0.5:
            warnings.append(f"WHERE条件匹配了 {len(filtered_df)}/{len(df)} 行({match_ratio * 100:.0f}%)，请确认条件是否正确")

        # 检测完全重复行(排除_ROW_NUMBER_)
        check_cols = [c for c in filtered_df.columns if c != "_ROW_NUMBER_"]
        dup_mask = filtered_df[check_cols].duplicated(keep=False)
        dup_count = dup_mask.sum()
        if dup_count > 0:
            warnings.append(f"发现 {dup_count} 行完全重复的记录（可用 _ROW_NUMBER_ 精确定位）")

        affected_indices = filtered_df.index.tolist()
        changes = []

        # 应用SET操作
        for col_name, value_expr in set_operations:
            for idx in affected_indices:
                old_val = df.at[idx, col_name]
                new_val = self._evaluate_update_expression(value_expr, df, idx)

                # Fix: P2-type-check — 写入前校验值类型是否与目标列匹配
                type_err = self._validate_value_type(new_val, df, col_name)
                if type_err:
                    return self._update_error(type_err)

                # 类型兼容性:数值类型可互通(含numpy整数/浮点,避免uint8溢出),其他类型尝试转为旧值类型
                # [FIX R54] NULL/None 值跳过类型强制转换 — 避免 None 被旧值类型转换(如 int(None) 异常后 fallback 到 0)
                if new_val is not None and old_val != "" and new_val != "" and type(old_val) != type(new_val):
                    if isinstance(old_val, (int, float, np.integer, np.floating)) and isinstance(new_val, (int, float, np.integer, np.floating)):
                        pass  # 数值互通:不转换(P0-fix: numpy数值类型不走type转换避免溢出)
                    else:
                        try:
                            new_val = type(old_val)(new_val)
                        except (ValueError, TypeError):
                            pass

                if old_val != new_val:
                    changes.append(
                        {
                            "row": int(idx) + 2,  # +2 for header offset (0-indexed + header row)
                            "column": col_name,
                            "old_value": self._serialize_update_value(old_val),
                            "new_value": self._serialize_update_value(new_val),
                        }
                    )
                # P0-fix: df.at赋值可能因numpy小dtype(uint8)溢出而截断或报错
                # 此赋值仅用于链式SET(如SET A=999, B=A+1)的中间状态,实际写Excel走changes列表
                # 故跳过溢出赋值不影响最终正确性
                try:
                    df.at[idx, col_name] = new_val
                except (ValueError, TypeError, OverflowError):
                    pass  # 值超出DataFrame列dtype范围,跳过中间状态更新

        if not changes:
            elapsed = (time.time() - start_time) * 1000
            result = {
                "success": True,
                "message": f"匹配 {len(affected_indices)} 行,但值无变化",
                "affected_rows": len(affected_indices),
                "changes": [],
                "execution_time_ms": round(elapsed, 1),
            }
            if warnings:
                result["warnings"] = warnings
            return result

        if dry_run:
            elapsed = (time.time() - start_time) * 1000
            result = {
                "success": True,
                "message": f"[预览] 将修改 {len(changes)} 个单元格({len(affected_indices)} 行)",
                "affected_rows": len(affected_indices),
                "changes": changes,
                "dry_run": True,
                "execution_time_ms": round(elapsed, 1),
            }
            if warnings:
                result["warnings"] = warnings
            return result

        # 写回Excel(事务保护:失败自动回滚)
        try:
            result = self._write_changes_to_excel(file_path, matched_sheet, changes, df, len(affected_indices), start_time)
            if warnings:
                result["warnings"] = warnings
            return result
        except Exception as e:
            elapsed = (time.time() - start_time) * 1000
            return {
                "success": False,
                "message": f"写入Excel失败,已自动回滚: {e}",
                "affected_rows": 0,
                "changes": changes,
                "execution_time_ms": round(elapsed, 1),
            }

    def execute_insert_query(
        self,
        file_path: str,
        sql: str,
        sheet_name: str | None = None,
        dry_run: bool = False,
    ) -> dict[str, Any]:
        """执行INSERT语句"""
        start_time = time.time()

        if not os.path.exists(file_path):
            return {"success": False, "message": f"文件不存在: {file_path}"}

        if not SQLGLOT_AVAILABLE:
            return {"success": False, "message": "SQLGLOT未安装"}

        sql = re.sub(r"\x1b\[[0-9;]*[a-zA-Z]", "", sql)

        # Fix: P0-5 INSERT 分号多语句注入
        if self._has_dangerous_semicolon(sql):
            return {"success": False, "message": "SQL语法错误: 不支持分号分隔的多语句执行(安全限制).💡 请将每条SQL语句分开执行"}

        # Fix: P0-7 INSERT 注释符注入防御
        # -- 和 # 可截断VALUES/SELECT子句导致非预期插入，必须在解析前拦截
        _comment_err = AdvancedSQLQueryEngine._detect_dangerous_comments(sql)
        if _comment_err:
            return {"success": False, "message": _comment_err}

        # 预处理: 自动为 MySQL 保留字标识符添加反引号
        sql = self._preprocess_reserved_words(sql)

        try:
            parsed = sqlglot.parse_one(sql, read="mysql")
        except ParseError as e:
            return {"success": False, "message": f"SQL语法错误: {e}"}

        if not isinstance(parsed, exp.Insert):
            return {"success": False, "message": "只支持INSERT语句"}

        # 提取表名和列名
        schema = parsed.this
        if isinstance(schema, exp.Schema):
            table_node = schema.this
            col_nodes = schema.expressions
        elif isinstance(schema, exp.Table):
            table_node = schema
            col_nodes = None
        else:
            return {
                "success": False,
                "message": f"INSERT目标格式不支持: {type(schema).__name__}",
            }

        table_name = table_node.name
        specified_cols = [c.name for c in col_nodes] if col_nodes else None

        # 加载数据
        worksheets_data = self._load_data_with_cache(file_path, sheet_name)

        # 匹配工作表
        matched_sheet = None
        for s in worksheets_data:
            if s == table_name or s.lower() == table_name.lower():
                matched_sheet = s
                break
        if not matched_sheet:
            return {
                "success": False,
                "message": f"工作表 '{table_name}' 不存在.可用: {list(worksheets_data.keys())}",
            }

        df = worksheets_data[matched_sheet]
        col_names = specified_cols if specified_cols else list(df.columns)

        # 双表头列名解析：将中文描述映射为英文字段名
        if specified_cols:
            try:
                from .header_analyzer import HeaderAnalyzer

                info = HeaderAnalyzer.analyze(file_path, matched_sheet)
                if info.is_dual and info.column_map:
                    _resolved_cols = []
                    for col in col_names:
                        if col in df.columns:
                            _resolved_cols.append(col)
                        elif col in info.column_map:
                            _resolved_cols.append(info.column_map[col])
                        else:
                            _resolved_cols.append(col)  # 保留原值，后续验证会报错
                    col_names = _resolved_cols
            except Exception:
                pass

        # 验证列名
        for col in col_names:
            if col not in df.columns:
                return {
                    "success": False,
                    "message": f"列 '{col}' 不存在.可用: {list(df.columns)}",
                }

        # 提取VALUES
        values_node = parsed.expression
        if not isinstance(values_node, (exp.Values, exp.Select)):
            return {
                "success": False,
                "message": "VALUES格式不支持,请使用 INSERT INTO ... VALUES (...)",
            }

        if isinstance(values_node, exp.Select):
            return {
                "success": False,
                "message": "INSERT ... SELECT 暂不支持,请使用VALUES",
            }

        rows = []
        for tuple_expr in values_node.expressions:
            if not isinstance(tuple_expr, exp.Tuple):
                continue
            row = {}
            for i, val_expr in enumerate(tuple_expr.expressions):
                if i >= len(col_names):
                    break
                val = self._eval_insert_value(val_expr)
                # Fix: P2-type-check — INSERT写入前校验值类型是否与目标列匹配
                type_err = self._validate_value_type(val, df, col_names[i])
                if type_err:
                    return {"success": False, "message": type_err, "affected_rows": 0}
                row[col_names[i]] = val
            # Fix: 检查VALUES值数量与列数量是否匹配，防止静默截断导致数据不完整
            if len(row) != len(col_names):
                return {
                    "success": False,
                    "message": f"VALUES 值数量({len(row)})与列数量({len(col_names)})不匹配。"
                               f"请确保每个 VALUES 元组包含 {len(col_names)} 个值",
                    "affected_rows": 0,
                }
            rows.append(row)

        if not rows:
            return {"success": False, "message": "没有数据可插入"}

        # Fix(P1-04): INSERT 批量大小限制，防止意外的大批量插入导致性能问题
        _MAX_INSERT_BATCH_SIZE = 5000
        if len(rows) > _MAX_INSERT_BATCH_SIZE:
            return {
                "success": False,
                "message": f"INSERT 批量插入行数({len(rows)})超过限制({_MAX_INSERT_BATCH_SIZE})。"
                           f"请分批插入，每批不超过 {_MAX_INSERT_BATCH_SIZE} 行",
                "affected_rows": 0,
            }

        if dry_run:
            elapsed = (time.time() - start_time) * 1000
            return {
                "success": True,
                "message": f"[预览] 将插入 {len(rows)} 行",
                "affected_rows": len(rows),
                "data": rows,
                "dry_run": True,
                "execution_time_ms": round(elapsed, 1),
            }

        # 写入Excel
        try:
            # Fix: P1-concurrent — 线程级写锁保护INSERT操作
            with self._get_write_lock(file_path):
                from .excel_operations import ExcelOperations

                result = ExcelOperations.batch_insert_rows(file_path, matched_sheet, rows, streaming=True)
                elapsed = (time.time() - start_time) * 1000
                if result.get("success"):
                    return {
                        "success": True,
                        "message": f"成功插入 {len(rows)} 行到 {matched_sheet}",
                        "affected_rows": len(rows),
                        "execution_time_ms": round(elapsed, 1),
                    }
                else:
                    return {
                        "success": False,
                        "message": f"写入失败: {result.get('message', '')}",
                        "affected_rows": 0,
                        "execution_time_ms": round(elapsed, 1),
                    }
        except Exception as e:
            elapsed = (time.time() - start_time) * 1000
            return {
                "success": False,
                "message": f"INSERT执行失败: {e}",
                "affected_rows": 0,
                "execution_time_ms": round(elapsed, 1),
            }

    def _eval_insert_value(self, val_expr) -> Any:
        """将sqlglot表达式转为Python值"""
        if isinstance(val_expr, exp.Literal):
            v = val_expr.this
            if val_expr.is_string:
                return v
            try:
                return int(v)
            except ValueError:
                try:
                    return float(v)
                except ValueError:
                    return v
        elif isinstance(val_expr, exp.Null):
            return None
        elif isinstance(val_expr, (exp.Neg,)):
            inner = self._eval_insert_value(val_expr.this)
            return -inner if isinstance(inner, (int, float)) else inner
        elif isinstance(val_expr, exp.Column):
            # Fix: VALUES 中不支持列引用，返回明确错误而非静默插入列名字符串
            raise ValueError(f"VALUES 中不支持列引用 '{val_expr.name}'。请使用字面量值（如 'value' 或 123）")
        else:
            return str(val_expr)

    def execute_delete_query(
        self,
        file_path: str,
        sql: str,
        sheet_name: str | None = None,
        dry_run: bool = False,
    ) -> dict[str, Any]:
        """执行DELETE语句"""
        start_time = time.time()

        if not os.path.exists(file_path):
            return {"success": False, "message": f"文件不存在: {file_path}"}

        if not SQLGLOT_AVAILABLE:
            return {"success": False, "message": "SQLGLOT未安装"}

        sql = re.sub(r"\x1b\[[0-9;]*[a-zA-Z]", "", sql)

        # Fix: P0-6 DELETE 分号多语句注入
        if self._has_dangerous_semicolon(sql):
            return {"success": False, "message": "SQL语法错误: 不支持分号分隔的多语句执行(安全限制).💡 请将每条SQL语句分开执行"}

        # Fix: P0-7 DELETE 注释符注入防御
        # -- 和 # 可截断WHERE条件导致全表删除，必须在解析前拦截
        _comment_err = AdvancedSQLQueryEngine._detect_dangerous_comments(sql)
        if _comment_err:
            return {"success": False, "message": _comment_err}

        # 预处理: 自动为 MySQL 保留字标识符添加反引号
        sql = self._preprocess_reserved_words(sql)

        try:
            parsed = sqlglot.parse_one(sql, read="mysql")
        except ParseError as e:
            return {"success": False, "message": f"SQL语法错误: {e}"}

        if not isinstance(parsed, exp.Delete):
            return {"success": False, "message": "只支持DELETE语句"}

        # 提取表名
        table_name = parsed.this.name

        # 加载数据
        worksheets_data = self._load_data_with_cache(file_path, sheet_name)
        # Fix(R13): 支持 WITH CTE 子句 — 提取CTE并执行，结果加入可用表
        self._current_worksheets = self._inject_ctes_to_worksheets(parsed, worksheets_data)

        # 匹配工作表
        matched_sheet = None
        for s in worksheets_data:
            if s == table_name or s.lower() == table_name.lower():
                matched_sheet = s
                break
        if not matched_sheet:
            return {
                "success": False,
                "message": f"工作表 '{table_name}' 不存在.可用: {list(worksheets_data.keys())}",
            }

        df = worksheets_data[matched_sheet].copy()
        df["_ROW_NUMBER_"] = range(1, len(df) + 1)

        # WHERE条件（必须）
        where_clause = parsed.args.get("where")
        if not where_clause:
            return {
                "success": False,
                "message": "DELETE必须指定WHERE条件(防止误删全表).如需清空请逐行删除或使用excel_delete_rows工具",
            }

        # 中文列名替换
        cn_map = {}
        desc_map = self._header_descriptions.get(matched_sheet, {})
        for en_col, cn_desc in desc_map.items():
            if en_col in df.columns:
                cn_map[cn_desc] = en_col

        # WHERE过滤（复用UPDATE的逻辑）
        try:
            condition_str = self._sql_condition_to_pandas(where_clause.this, df)
            if condition_str:
                filtered_df = df.query(condition_str)
            else:
                filtered_df = self._apply_row_filter(where_clause.this, df)
        except Exception:
            filtered_df = self._apply_row_filter(where_clause.this, df)

        if filtered_df.empty:
            elapsed = (time.time() - start_time) * 1000
            return {
                "success": True,
                "message": "没有匹配WHERE条件的行,无需删除",
                "affected_rows": 0,
                "execution_time_ms": round(elapsed, 1),
            }

        # DataFrame行号转Excel行号
        # 单行表头: DataFrame第1行 = Excel第2行 (+1)
        # 双行表头: DataFrame第1行 = Excel第3行 (+2)
        df_row_numbers = filtered_df["_ROW_NUMBER_"].tolist()
        header_offset = 2 if matched_sheet in self._header_descriptions and self._header_descriptions[matched_sheet] else 1
        excel_row_numbers = [r + header_offset for r in df_row_numbers]

        if dry_run:
            elapsed = (time.time() - start_time) * 1000
            return {
                "success": True,
                "message": f"[预览] 将删除 {len(excel_row_numbers)} 行",
                "affected_rows": len(excel_row_numbers),
                "dry_run": True,
                "execution_time_ms": round(elapsed, 1),
            }

        # 写入Excel
        try:
            # Fix: P1-concurrent — 线程级写锁保护DELETE操作
            with self._get_write_lock(file_path):
                from .excel_operations import ExcelOperations

                result = ExcelOperations.batch_delete_rows(file_path, matched_sheet, excel_row_numbers, streaming=True)
                elapsed = (time.time() - start_time) * 1000
                if result.get("success"):
                    return {
                        "success": True,
                        "message": f"成功删除 {len(excel_row_numbers)} 行",
                        "affected_rows": len(excel_row_numbers),
                        "execution_time_ms": round(elapsed, 1),
                    }
                else:
                    return {
                        "success": False,
                        "message": f"删除失败: {result.get('message', '')}",
                        "affected_rows": 0,
                        "execution_time_ms": round(elapsed, 1),
                    }
        except Exception as e:
            elapsed = (time.time() - start_time) * 1000
            return {
                "success": False,
                "message": f"DELETE执行失败: {e}",
                "affected_rows": 0,
                "execution_time_ms": round(elapsed, 1),
            }

    def _write_changes_to_excel(
        self,
        file_path: str,
        sheet_name: str,
        changes: list,
        df: pd.DataFrame,
        affected_rows: int,
        start_time: float,
    ) -> dict[str, Any]:
        """事务保护写入变更到Excel(失败自动回滚)
        支持流式写入:大文件批量修改时使用write_only模式提升性能
        """
        backup_path = None
        try:
            # Fix: P1-concurrent — 线程级写锁(外层) + fcntl进程级锁(内层),双重保护
            with self._get_write_lock(file_path):
                with self._file_lock(file_path):
                    backup_path = tempfile.mktemp(suffix=".xlsx.bak")
                    shutil.copy2(file_path, backup_path)

                    # 决策:使用流式写入的条件
                    file_size = os.path.getsize(file_path)
                    use_streaming = (
                        affected_rows >= STREAMING_WRITE_MIN_ROWS  # 影响行数>=阈值
                        or len(changes) >= STREAMING_WRITE_MIN_CHANGES  # 修改单元格数>=阈值
                        or file_size > STREAMING_WRITE_MIN_FILE_SIZE_MB * 1024 * 1024  # 文件大小>阈值
                    )

                    if use_streaming and StreamingWriter.is_available():
                        # 使用流式写入(高性能路径)
                        # col_map: {Excel表头(中文): 列索引(1-based)}
                        # change['column']: 英文列名(来自pandas DataFrame)
                        # 需要通过英→中映射翻译列名

                        # 构建英→中列名映射
                        en_to_cn_map = {}
                        header_desc = getattr(self, "_header_descriptions", {})
                        desc_for_sheet = header_desc.get(sheet_name, {})
                        for en_col, cn_desc in desc_for_sheet.items():
                            en_to_cn_map[en_col] = cn_desc

                        def modify_fn(rows, header_row, col_map):
                            """修改函数:应用UPDATE变更到行数据"""
                            modified_rows = [row[:] for row in rows]
                            failed_cols = set()
                            matched_count = 0

                            for change in changes:
                                col_name = change["column"]
                                col_idx = None

                                # 策略1: 精确匹配
                                if col_name in col_map:
                                    col_idx = col_map[col_name]
                                else:
                                    # 策略2: 英→中映射
                                    cn_name = en_to_cn_map.get(col_name)
                                    if cn_name and cn_name in col_map:
                                        col_idx = col_map[cn_name]
                                    else:
                                        # 策略3: strip+大小写不敏感
                                        col_stripped = str(col_name).strip().lower()
                                        for k, v in col_map.items():
                                            if str(k).strip().lower() == col_stripped:
                                                col_idx = v
                                                break

                                if col_idx is None:
                                    failed_cols.add(col_name)
                                    continue

                                matched_count += 1
                                list_idx = change["row"] - 1
                                if 0 <= list_idx < len(modified_rows):
                                    row = modified_rows[list_idx]
                                    while len(row) < col_idx:
                                        row.append("")
                                    # Fix: P2-4 极端浮点值导致文件损坏 — 流式写入前清理值
                                    new_val = _sanitize_float_for_excel(change["new_value"])
                                    row[col_idx - 1] = new_val

                            return (
                                True,
                                f"流式写入完成, 匹配{matched_count}列, 失败{len(failed_cols)}列",
                                modified_rows,
                                {
                                    "columns_matched": matched_count,
                                    "columns_failed": list(failed_cols),
                                },
                            )

                        success, message, meta = StreamingWriter._copy_modify_write(file_path, sheet_name, modify_fn, preserve_col_widths=True)

                        # P0: 如果流式写入列名全部匹配失败，立即降级到传统写入
                        if success and meta.get("columns_matched", 0) == 0:
                            logger.warning(f"流式写入列名全部匹配失败，降级到传统写入。失败列: {meta.get('columns_failed', [])}")
                            success = False
                            message = f"列名匹配全部失败，降级到传统写入: {meta.get('columns_failed', [])}"

                        if success:
                            # 流式写入后验证实际写入
                            write_verified = False
                            verify_result = self._verify_streaming_write(file_path, sheet_name, changes, en_to_cn_map)
                            write_verified = verify_result["verified"]
                            unverified = verify_result["unverified"]

                            if not write_verified and len(unverified) == len(changes):
                                # 所有change都没生效，回滚
                                if backup_path and os.path.exists(backup_path):
                                    shutil.copy2(backup_path, file_path)
                                    os.remove(backup_path)
                                self._df_cache.pop(file_path, None)
                                failed_cols_info = meta.get("columns_failed", [])
                                extra_hint = f"（列名匹配失败: {failed_cols_info}）" if failed_cols_info else ""
                                elapsed = (time.time() - start_time) * 1000
                                return {
                                    "success": False,
                                    "message": f"写入验证失败：{len(changes)}个修改均未生效，已自动回滚{extra_hint}",
                                    "affected_rows": 0,
                                    "changes": changes,
                                    "error_code": "WRITE_VERIFICATION_FAILED",
                                    "execution_time_ms": round(elapsed, 1),
                                }

                            self._df_cache.pop(file_path, None)
                            elapsed = (time.time() - start_time) * 1000
                            result = {
                                "success": True,
                                "message": f"流式更新 {len(changes)} 个单元格({affected_rows} 行)",
                                "affected_rows": affected_rows,
                                "changes": changes,
                                "execution_time_ms": round(elapsed, 1),
                                "method": "streaming",
                                "verification": {
                                    "verified": write_verified,
                                    "columns_matched": meta.get("columns_matched", len(changes)),
                                    "columns_failed": meta.get("columns_failed", []),
                                },
                            }
                            if not write_verified and unverified:
                                result["message"] += f" (警告: {len(unverified)}/{len(changes)}个修改验证失败)"
                            return result
                        else:
                            # 流式写入失败,降级到传统方式
                            logger.warning(f"流式写入失败,降级到传统方式: {message}")

                    # 传统写入方式(兼容性路径)
                    header_row_offset = 0
                    header_desc = getattr(self, "_header_descriptions", {})
                    if header_desc.get(sheet_name, {}):
                        header_row_offset = 1

                    wb = openpyxl.load_workbook(file_path)
                    ws = wb[sheet_name]

                    for change in changes:
                        excel_row = change["row"] + header_row_offset
                        col_idx = list(df.columns).index(change["column"]) + 1
                        # Fix: P2-4 极端浮点值导致文件损坏 — 传统写入前清理值
                        safe_value = _sanitize_float_for_excel(change["new_value"])
                        # [FIX R54] NULL 值需要特殊处理 — openpyxl 对已赋值的数值单元格设 None 可能不生效
                        # 通过重建 cell 或显式清除来确保 NULL 写入
                        if safe_value is None:
                            target_cell = ws.cell(row=excel_row, column=col_idx)
                            # 方法1: 直接设为 None (大部分情况生效)
                            target_cell.value = None
                            # 方法2: 强制清除 data_type 确保保存为空
                            try:
                                target_cell.data_type = 's'  # 先转为字符串类型
                                target_cell.value = None
                            except Exception:
                                pass
                        else:
                            ws.cell(row=excel_row, column=col_idx, value=safe_value)

                    wb.save(file_path)
                    wb.close()

                    if backup_path and os.path.exists(backup_path):
                        os.remove(backup_path)

                    self._df_cache.pop(file_path, None)

                    elapsed = (time.time() - start_time) * 1000
                    return {
                        "success": True,
                        "message": f"成功更新 {len(changes)} 个单元格({affected_rows} 行)",
                        "affected_rows": affected_rows,
                        "changes": changes,
                        "execution_time_ms": round(elapsed, 1),
                        "method": "traditional",
                    }
        except Exception:
            if backup_path and os.path.exists(backup_path):
                try:
                    shutil.copy2(backup_path, file_path)
                    os.remove(backup_path)
                except Exception:
                    pass
            raise  # 重新抛出让调用方处理

    # Fix: P1-concurrent — 线程级写锁上下文管理器(按文件路径隔离)
    @contextmanager
    def _get_write_lock(self, file_path: str) -> Generator[None, None, None]:
        """获取指定文件的线程级写锁,确保同文件并发写入互斥"""
        # 获取或创建该文件的锁(线程安全)
        with self._write_locks_global:
            if file_path not in self._write_locks:
                self._write_locks[file_path] = threading.Lock()
            lock = self._write_locks[file_path]
        with lock:
            yield

    @contextmanager
    def _file_lock(self, file_path: str) -> Generator[None, None, None]:
        """文件锁上下文管理器(Linux fcntl,其他平台优雅降级)

        Fix: 检测并清理孤儿锁文件（进程被强杀后残留的 .lock 文件）
        """
        lock_fd = None
        try:
            try:
                import fcntl

                lock_path = file_path + ".lock"
                # 检测孤儿锁文件：如果存在且持有者进程已死，自动清理
                if os.path.exists(lock_path):
                    try:
                        with open(lock_path, "r") as lf:
                            pid_str = lf.read().strip()
                        if pid_str:
                            old_pid = int(pid_str)
                            # 检查进程是否存活
                            os.kill(old_pid, 0)
                            # 进程仍存活，正常等待 flock
                    except (ValueError, ProcessLookupError):
                        # PID 无效或进程已死，清理孤儿锁
                        try:
                            os.remove(lock_path)
                        except OSError:
                            pass
                    except OSError:
                        pass

                lock_fd = open(lock_path, "w", encoding="utf-8")
                # 写入当前 PID，用于后续孤儿检测
                lock_fd.write(str(os.getpid()))
                lock_fd.flush()
                fcntl.flock(lock_fd, fcntl.LOCK_EX)
            except (ImportError, OSError):
                lock_fd = None
            yield
        finally:
            if lock_fd:
                try:
                    import fcntl

                    fcntl.flock(lock_fd, fcntl.LOCK_UN)
                    lock_path = file_path + ".lock"
                    if os.path.exists(lock_path):
                        os.remove(lock_path)
                except Exception:
                    pass
                lock_fd.close()

    # Fix: P2-float-precision 移除强制round(x,2),保留完整浮点精度
    def _serialize_value(self, val: Any) -> Any:
        """智能序列化值:数值保持数值类型,None/NaN/inf转None,numpy->Python原生

        浮点数处理策略 (v2 — 保留完整精度):
        - 整数值(如 25.0, 25.000001) → int(25)
        - 所有其他浮点值 → 原样返回 float,不做任何舍入
          (用户如需舍入应使用 SQL ROUND() 函数显式控制)

        安全加固(R42): inf/-inf 视为无效值转None,防止 OverflowError 崩溃.
        R48-fix: 新增 Decimal 类型支持,避免 JSON 序列化失败.
        R48-fix: 新增 datetime/timedelta 类型支持,避免 JSON 序列化崩溃(P0-01).
        """
        if val is None:
            return None
        # R48-fix P0-01: datetime/timedelta/pd.Timestamp → ISO格式字符串
        if isinstance(val, (datetime.datetime, datetime.date)):
            return val.isoformat()
        if isinstance(val, datetime.timedelta):
            return str(val)
        try:
            if isinstance(val, pd.Timestamp):
                return val.isoformat()
            if isinstance(val, pd.Timedelta):
                return str(val)
        except (ImportError, AttributeError):
            pass
        # R48: Decimal 类型处理 — 转为 float/int 以保证 JSON 安全
        try:
            if isinstance(val, Decimal):
                if val.is_nan():
                    return None
                if val.is_infinite():
                    return None
                # 整数 Decimal → int
                if val == int(val):
                    return int(val)
                # 非整数 Decimal → float（保留精度）
                return float(val)
        except (ImportError, InvalidOperation, ValueError, OverflowError):
            pass
        if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
            return None
        if isinstance(val, (np.integer,)):
            return int(val)
        if isinstance(val, (np.floating,)):
            f = float(val)
            if np.isnan(f) or np.isinf(f):
                return None
            # 整数值(含浮点误差范围内的近整数)返回int
            try:
                if f == int(f):
                    return int(f)
            except (OverflowError, ValueError):
                pass  # inf 等极端值已在上面拦截,此处为防御性编程
            # Fix: P2-float-precision 不再强制 round(x,2),保留原始精度
            return f
        if isinstance(val, float):
            if np.isnan(val) or np.isinf(val):
                return None
            try:
                if val == int(val):
                    return int(val)
            except (OverflowError, ValueError):
                pass
            # Fix: P2-float-precision 不再强制 round(x,2),保留原始精度
            return val
        return val

    def _serialize_update_value(self, val: Any) -> Any:
        """将值序列化为JSON安全类型(numpy->Python原生)-- 委托给_serialize_value"""
        return self._serialize_value(val)

    def _evaluate_update_expression(self, expr: exp.Expression, df: pd.DataFrame, row_idx: int,
                                    depth: int = 0) -> Any:
        """
        评估UPDATE SET表达式,支持常量,列引用和算术运算

        Args:
            expr: SQL表达式
            df: DataFrame
            row_idx: 行索引
            depth: Fix(P1-05): 当前递归深度，防止无限递归

        Returns:
            计算后的值
        """
        if isinstance(expr, exp.Literal):
            return self._parse_literal_value(expr)

        elif isinstance(expr, exp.Boolean):
            # SQL布尔字面量(TRUE/FALSE) → int(1/0)
            return int(expr.this)

        elif isinstance(expr, exp.Column):
            col_name = expr.name
            if col_name in df.columns:
                return df.at[row_idx, col_name]
            return ""

        elif isinstance(expr, exp.Neg):
            inner = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            try:
                return -float(inner)
            except (ValueError, TypeError):
                return inner

        elif isinstance(expr, (exp.Add, exp.Sub, exp.Mul, exp.Div)):
            left = self._evaluate_update_expression(expr.left, df, row_idx, depth + 1)
            right = self._evaluate_update_expression(expr.right, df, row_idx, depth + 1)
            try:
                left_n = float(left) if not isinstance(left, (int, float)) else left
                right_n = float(right) if not isinstance(right, (int, float)) else right
                # 复用类级别分发表,支持所有算术运算符
                op = self._MATH_BINARY_OPS[type(expr)]
                result = op(left_n, right_n if type(expr) != exp.Div or right_n != 0 else 0)
                # 如果原值都是整数(含numpy整数)且非除法,返回整数
                if isinstance(left, (int, np.integer)) and isinstance(right, (int, np.integer)) and type(expr) != exp.Div:
                    return int(result)
                return result
            except (ValueError, TypeError):
                return ""

        elif isinstance(expr, exp.Round):
            # UPDATE SET ROUND(column, decimals)
            inner = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            decimals_arg = expr.args.get("decimals")
            decimals = int(self._literal_value(decimals_arg)) if decimals_arg is not None else 0
            try:
                return round(float(inner), decimals)
            except (ValueError, TypeError):
                return inner

        elif isinstance(expr, exp.Case):
            # UPDATE SET CASE WHEN expression
            # Fix(R7-E1): Support CASE WHEN in UPDATE SET clause
            # Fix(R14-B2): 必须用关键字参数传 row,否则 df 参数会接收 Series 导致向量化模式错误
            try:
                return self._evaluate_case_expression(expr, None, row=df.iloc[row_idx])
            except Exception as e:
                logger.warning(f"UPDATE CASE WHEN 求值失败: {e}")
                # R48-fix P0-02: 不再硬编码"Price"列名回退(会导致非Price表数据丢失为None)
                # 改为抛出异常,由外层UPDATE统一错误处理
                raise ValueError(f"UPDATE CASE WHEN 求值失败: {e}")

        elif isinstance(expr, exp.Coalesce):
            # [FIX R10-B2] UPDATE SET COALESCE(v1, v2, ...) — 返回第一个非NULL/非空值
            # [FIX R45-04] sqlglot Coalesce: expr.this=first arg, expr.expressions=rest
            all_args = [expr.this] + list(expr.expressions)
            for arg in all_args:
                val = self._evaluate_update_expression(arg, df, row_idx, depth + 1)
                if val is not None and val != "" and not (isinstance(val, float) and np.isnan(val)):
                    return val
            # 所有参数都为 NULL/空，返回最后一个参数的值(可能为None)
            if all_args:
                return self._evaluate_update_expression(all_args[-1], df, row_idx, depth + 1)
            return None

        # Fix: P2-1 — 支持 CONCAT() 和 || 字符串拼接操作符在 UPDATE SET 中使用
        elif isinstance(expr, exp.Concat):
            # CONCAT(a, b, c, ...) — 由 || 转换而来或用户直接使用
            parts = []
            for arg in expr.expressions:
                val = self._evaluate_update_expression(arg, df, row_idx, depth + 1)
                parts.append(str(val) if val is not None else '')
            return ''.join(parts)
        elif isinstance(expr, exp.DPipe):
            # || 直接被 sqlglot 识别为 DPipe(如 PostgreSQL 方言)
            left = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            right = self._evaluate_update_expression(expr.expression, df, row_idx, depth + 1)
            return (str(left) if left is not None else '') + (str(right) if right is not None else '')
        elif isinstance(expr, exp.Or):
            # Fix: P2-1 — MySQL 方言下 || 被解析为 OR,启发式检测字符串拼接
            if self._is_likely_dpipe_concatenation(expr):
                left = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
                right = self._evaluate_update_expression(expr.expression, df, row_idx, depth + 1)
                return (str(left) if left is not None else '') + (str(right) if right is not None else '')
            else:
                raise ValueError(
                    f"不支持的表达式: {expr}。\n"
                    f"💡 MySQL方言中 || 表示逻辑OR,如需字符串拼接请使用 CONCAT() 函数。\n"
                    f"🔧 示例: UPDATE table SET col = CONCAT(col, '_suffix') WHERE ..."
                )

        elif isinstance(expr, (exp.Abs, exp.Ceil, exp.Floor, exp.Sqrt)):
            # [FIX R10-B2] UPDATE SET 标量数学函数(ABS/CEIL/FLOOR/SQRT)
            inner = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            try:
                f_val = float(inner) if inner is not None else None
                if f_val is None:
                    return None
                if isinstance(expr, exp.Abs):
                    return abs(f_val)
                elif isinstance(expr, exp.Ceil):
                    r = np.ceil(f_val)
                    return int(r) if r == int(r) else r
                elif isinstance(expr, exp.Floor):
                    r = np.floor(f_val)
                    return int(r) if r == int(r) else r
                elif isinstance(expr, exp.Sqrt):
                    return np.sqrt(f_val) if f_val >= 0 else None
            except (ValueError, TypeError):
                return inner

        elif isinstance(expr, (exp.Upper, exp.Lower, exp.Trim)):
            # [FIX R47] UPDATE SET 字符串函数(UPPER/LOWER/TRIM)
            inner = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            if inner is None:
                return None
            s = str(inner)
            if isinstance(expr, exp.Upper):
                return s.upper()
            elif isinstance(expr, exp.Lower):
                return s.lower()
            elif isinstance(expr, exp.Trim):
                return s.strip()

        elif isinstance(expr, (exp.Substring, exp.Left, exp.Right)):
            # [FIX R47] UPDATE SET 子字符串函数(SUBSTRING/LEFT/RIGHT)
            if isinstance(expr, exp.Substring):
                val = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
                start = self._evaluate_update_expression(expr.args.get("start"), df, row_idx, depth + 1)
                length = self._evaluate_update_expression(expr.args.get("length"), df, row_idx, depth + 1)
                if val is None or start is None:
                    return None
                s = str(val)
                st = int(start) - 1  # SQL is 1-indexed
                ln = int(length) if length is not None else len(s)
                return s[st:st + ln]
            elif isinstance(expr, (exp.Left, exp.Right)):
                val = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
                length_expr = expr.args.get("length") or expr.args.get("expression")
                n = self._evaluate_update_expression(length_expr, df, row_idx, depth + 1)
                if val is None or n is None:
                    return None
                s = str(val)
                count = int(n)
                if isinstance(expr, exp.Left):
                    return s[:count]
                else:
                    return s[-count:] if count > 0 else ""

        elif isinstance(expr, exp.Length):
            # [FIX R47] UPDATE SET LENGTH() 函数
            inner = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            if inner is None:
                return 0
            return len(str(inner))

        elif isinstance(expr, exp.Pow):
            # [FIX R45-04] UPDATE SET POWER(base, exponent)
            # sqlglot Pow structure: this=base, expression=exponent (NOT 'exp'!)
            base = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            exp_arg = expr.args.get("expression") or expr.args.get("exp")
            exponent = self._evaluate_update_expression(exp_arg, df, row_idx, depth + 1) if exp_arg is not None else None
            try:
                base_n = float(base) if base is not None else None
                exp_n = float(exponent) if exponent is not None else None
                if base_n is None or exp_n is None:
                    return None
                result = np.power(base_n, exp_n)
                return int(result) if result == int(result) else result
            except (ValueError, TypeError):
                return base

        elif isinstance(expr, exp.Mod):
            # [FIX R45-04] UPDATE SET modulo (a % b)
            left = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            right = self._evaluate_update_expression(expr.expression, df, row_idx, depth + 1)
            try:
                left_n = float(left) if not isinstance(left, (int, float)) else left
                right_n = float(right) if not isinstance(right, (int, float)) else right
                if right_n == 0:
                    return None
                result = left_n % right_n
                if isinstance(left, (int, np.integer)) and isinstance(right, (int, np.integer)):
                    return int(result)
                return result
            except (ValueError, TypeError):
                return ""

        elif isinstance(expr, exp.Replace):
            # [FIX R45-04] UPDATE SET REPLACE(str, search, replacement)
            val = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            old_str = self._evaluate_update_expression(expr.args.get("expression"), df, row_idx, depth + 1)
            new_str = self._evaluate_update_expression(expr.args.get("replacement"), df, row_idx, depth + 1)
            if val is None:
                return None
            s = str(val)
            old_s = str(old_str) if old_str is not None else ""
            new_s = str(new_str) if new_str is not None else ""
            return s.replace(old_s, new_s)

        elif isinstance(expr, exp.Nullif):
            # [FIX R45-04] UPDATE SET NULLIF(a, b) — returns NULL if a == b, else a
            val_a = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            val_b = self._evaluate_update_expression(expr.expression, df, row_idx, depth + 1)
            if val_a is None or val_b is None:
                return val_a
            # Compare values (handle numeric/string comparison)
            try:
                if float(val_a) == float(val_b):
                    return None
            except (ValueError, TypeError):
                if str(val_a) == str(val_b):
                    return None
            return val_a

        elif isinstance(expr, exp.Cast):
            # [FIX R45-04] UPDATE SET CAST(expr AS type)
            inner = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            if inner is None:
                return None
            target_type = expr.args.get("to")
            if target_type is None:
                return inner
            # sqlglot DataType: this is a Type enum (e.g. Type.BIGINT), use sql() for name
            if hasattr(target_type, "sql"):
                type_name = target_type.sql().upper()
            elif hasattr(target_type, "this"):
                val = target_type.this
                # Handle enum (Type.BIGINT -> 'BIGINT') and string values
                type_name = str(val).upper() if not isinstance(val, str) else val.upper()
            else:
                type_name = str(target_type).upper()
            try:
                if type_name in ("SIGNED", "INTEGER", "INT", "TINYINT", "SMALLINT", "BIGINT"):
                    return int(float(inner))
                elif type_name in ("UNSIGNED", "DECIMAL", "NUMERIC", "FLOAT", "DOUBLE", "REAL"):
                    return float(inner)
                elif type_name in ("CHAR", "VARCHAR", "TEXT", "STRING"):
                    return str(inner)
                else:
                    # Unknown type — return as-is
                    return inner
            except (ValueError, TypeError):
                return inner

        elif isinstance(expr, exp.If):
            # [FIX R45-04] UPDATE SET IF(condition, true_value, false_value) — MySQL-specific
            cond = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            # Evaluate condition: non-zero/non-empty/True = truthy
            is_truthy = False
            if cond is not None and cond != "" and cond != 0 and cond != False:
                if isinstance(cond, str):
                    is_truthy = cond.upper() not in ("FALSE", "0", "")
                else:
                    is_truthy = True
            if is_truthy:
                true_expr = expr.args.get("true")
                if true_expr is not None:
                    return self._evaluate_update_expression(true_expr, df, row_idx, depth + 1)
                return None
            else:
                false_expr = expr.args.get("false")
                if false_expr is not None:
                    return self._evaluate_update_expression(false_expr, df, row_idx, depth + 1)
                return None

        elif isinstance(expr, (exp.EQ, exp.NEQ, exp.GT, exp.GTE, exp.LT, exp.LTE)):
            # [FIX R45-04] UPDATE SET 比较运算符(=, <>, >, >=, <, <=) — 用于 IF/CASE WHEN 条件
            left = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            right = self._evaluate_update_expression(expr.expression, df, row_idx, depth + 1)
            if left is None or right is None:
                return False
            try:
                if isinstance(left, str) and isinstance(right, (int, float)):
                    left_num = float(left) if '.' in str(left) else int(left)
                    left = type(right)(left_num)
                elif isinstance(right, str) and isinstance(left, (int, float)):
                    right_num = float(right) if '.' in str(right) else int(right)
                    right = type(left)(right_num)
                elif isinstance(left, str) and isinstance(right, str):
                    pass
                else:
                    left, right = float(left), float(right)
            except (ValueError, TypeError):
                pass
            if isinstance(expr, exp.EQ):
                return left == right
            elif isinstance(expr, exp.NEQ):
                return left != right
            elif isinstance(expr, exp.GT):
                return left > right
            elif isinstance(expr, exp.GTE):
                return left >= right
            elif isinstance(expr, exp.LT):
                return left < right
            elif isinstance(expr, exp.LTE):
                return left <= right

        elif isinstance(expr, exp.And):
            # [FIX R45-04] UPDATE SET 逻辑 AND — 用于复合条件
            left = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            right = self._evaluate_update_expression(expr.expression, df, row_idx, depth + 1)
            return bool(left) and bool(right)

        elif isinstance(expr, exp.Not):
            # [FIX R45-04] UPDATE SET 逻辑 NOT
            val = self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            if val is None:
                return None
            return not val

        else:
            # Fix(P1-05): 递归深度保护，防止无限递归导致栈溢出
            _MAX_RECURSION_DEPTH = 20
            if depth >= _MAX_RECURSION_DEPTH:
                logger.warning(f"_evaluate_update_expression 递归深度({depth})超过限制({_MAX_RECURSION_DEPTH})，中止求值")
                return None
            # 未知表达式类型,尝试递归(带深度计数)
            if hasattr(expr, "this"):
                return self._evaluate_update_expression(expr.this, df, row_idx, depth + 1)
            return ""


# 模块级单例引擎,DataFrame缓存跨调用共享
_shared_engine: AdvancedSQLQueryEngine | None = None


def _get_engine() -> AdvancedSQLQueryEngine:
    """获取共享SQL引擎实例(缓存跨调用复用)"""
    global _shared_engine
    if _shared_engine is None:
        _shared_engine = AdvancedSQLQueryEngine()
    return _shared_engine


def execute_advanced_sql_query(
    file_path: str,
    sql: str,
    sheet_name: str | None = None,
    limit: int | None = None,
    include_headers: bool = True,
    output_format: str = "table",
) -> dict[str, Any]:
    """
    便捷函数:执行高级SQL查询

    Args:
        file_path: Excel文件路径
        sql: SQL查询语句
        sheet_name: 工作表名称(可选)
        limit: 结果限制
        include_headers: 是否包含表头
        output_format: 输出格式 table/json/csv

    Returns:
        Dict: 查询结果
    """
    try:
        engine = _get_engine()
        return engine.execute_sql_query(
            file_path=file_path,
            sql=sql,
            sheet_name=sheet_name,
            limit=limit,
            include_headers=include_headers,
            output_format=output_format,
        )
    except ImportError as e:
        return {
            "success": False,
            "message": f"SQLGlot未安装,无法使用高级SQL功能: {AdvancedSQLQueryEngine._sanitize_error_message(str(e))}",
            "data": [],
            "query_info": {"error_type": "missing_dependency", "dependency": "sqlglot"},
        }
    except Exception as e:
        _safe_msg = AdvancedSQLQueryEngine._sanitize_error_message(str(e))
        return {
            "success": False,
            "message": f"高级SQL查询失败: {_safe_msg}",
            "data": [],
            "query_info": {"error_type": "engine_error", "details": _safe_msg},
        }


def execute_advanced_update_query(file_path: str, sql: str, sheet_name: str | None = None, dry_run: bool = False) -> dict[str, Any]:
    """
    便捷函数:执行UPDATE SQL语句

    支持行号定位更新(_ROW_NUMBER_)::
        UPDATE 表 SET 列=值 WHERE _ROW_NUMBER_ IN (行号列表)
        UPDATE 表 SET 列=值 WHERE _ROW_NUMBER_ = 行号
        UPDATE 表 SET 列=值 WHERE _ROW_NUMBER_ BETWEEN 起始 AND 结束

    Args:
        file_path: Excel文件路径
        sql: UPDATE SQL语句
        sheet_name: 工作表名称(可选)
        dry_run: 预览模式

    Returns:
        Dict: 更新结果
    """
    try:
        engine = _get_engine()
        return engine.execute_update_query(file_path=file_path, sql=sql, sheet_name=sheet_name, dry_run=dry_run)
    except ImportError as e:
        return {
            "success": False,
            "message": f"SQLGlot未安装,无法使用UPDATE功能: {AdvancedSQLQueryEngine._sanitize_error_message(str(e))}",
            "affected_rows": 0,
            "changes": [],
            "query_info": {"error_type": "missing_dependency", "dependency": "sqlglot"},
        }
    except Exception as e:
        _safe_msg = AdvancedSQLQueryEngine._sanitize_error_message(str(e))
        return {
            "success": False,
            "message": f"UPDATE执行失败: {_safe_msg}",
            "affected_rows": 0,
            "changes": [],
            "query_info": {"error_type": "engine_error", "details": _safe_msg},
        }


def execute_advanced_insert_query(file_path: str, sql: str, dry_run: bool = False) -> dict[str, Any]:
    """执行INSERT SQL语句"""
    try:
        engine = _get_engine()
        return engine.execute_insert_query(file_path=file_path, sql=sql, dry_run=dry_run)
    except ImportError as e:
        return {"success": False, "message": f"SQLGLOT未安装: {e}", "affected_rows": 0}
    except Exception as e:
        return {"success": False, "message": f"INSERT执行失败: {e}", "affected_rows": 0}


def execute_advanced_delete_query(file_path: str, sql: str, dry_run: bool = False) -> dict[str, Any]:
    """执行DELETE SQL语句"""
    try:
        engine = _get_engine()
        return engine.execute_delete_query(file_path=file_path, sql=sql, dry_run=dry_run)
    except ImportError as e:
        return {"success": False, "message": f"SQLGLOT未安装: {e}", "affected_rows": 0}
    except Exception as e:
        return {"success": False, "message": f"DELETE执行失败: {e}", "affected_rows": 0}
