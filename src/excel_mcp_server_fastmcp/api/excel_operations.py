"""
Excel MCP Server - Excel操作API模块

提供高内聚的Excel业务操作功能，包含完整的参数验证、业务逻辑、错误处理和结果格式化

@intention: 将Excel操作的具体实现从server.py中分离，提高代码内聚性和可维护性
"""

import logging
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from ..core.excel_compare import ExcelComparer
from ..core.excel_converter import ExcelConverter
from ..core.excel_manager import ExcelManager
from ..core.excel_reader import ExcelReader
from ..core.excel_search import ExcelSearcher
from ..core.excel_writer import ExcelWriter
from ..models.types import ComparisonOptions
from ..utils.config import MAX_SEARCH_FILES
from ..utils.formatter import format_operation_result
from .header_analyzer import HeaderAnalyzer

logger = logging.getLogger(__name__)


class ExcelOperations:
    """
    @class ExcelOperations
    @brief Excel业务操作的高内聚封装
    @intention 提供完整的Excel操作功能，包含参数验证、错误处理、结果格式化
    """

    # ==================== 日志系统 ====================
    DEBUG_LOG_ENABLED: bool = False
    _LOG_PREFIX = "[API][ExcelOperations]"

    # ==================== 主干API ====================

    @classmethod
    def query(cls, file_path: str, sql: str, include_headers: bool = True) -> dict[str, Any]:
        """
        @intention 执行SQL查询Excel数据

        Args:
            file_path: Excel文件路径
            sql: SQL查询语句
            include_headers: 是否包含表头

        Returns:
            Dict: 查询结果
        """
        try:
            from .advanced_sql_query import AdvancedSQLQueryEngine

            engine = AdvancedSQLQueryEngine()
            result = engine.execute_sql_query(file_path, sql, include_headers=include_headers)
            return result
        except Exception as e:
            error_msg = f"SQL查询失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def get_range(cls, file_path: str, range_expression: str, include_formatting: bool = False) -> dict[str, Any]:
        """
        @intention 获取Excel文件中指定范围的数据，提供完整的业务逻辑处理

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            range_expression: 范围表达式，必须包含工作表名
            include_formatting: 是否包含格式信息

        Returns:
            Dict: 标准化的操作结果

        Example:
            result = ExcelOperations.get_range("data.xlsx", "Sheet1!A1:C10")
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始获取范围数据: {range_expression}")

        try:
            # 步骤1: 验证参数格式
            validation_result = cls._validate_range_format(range_expression)
            if not validation_result["valid"]:
                return cls._format_error_result(validation_result["error"])

            # 步骤2: 执行数据读取
            reader = ExcelReader(file_path)
            result = reader.get_range(range_expression, include_formatting)
            reader.close()

            # 步骤3: 格式化结果
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"获取范围数据失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def update_range(
        cls,
        file_path: str,
        range_expression: str,
        data: list[list[Any]],
        preserve_formulas: bool = True,
        insert_mode: bool = False,
        streaming: bool = True,
    ) -> dict[str, Any]:
        """
        @intention 更新Excel文件中指定范围的数据，支持插入和覆盖模式

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            range_expression: 范围表达式，必须包含工作表名
            data: 二维数组数据 [[row1], [row2], ...]
            preserve_formulas: 是否保留现有公式
            insert_mode: 数据写入模式 (默认值: False)
                - True: 插入模式，在指定位置插入新行然后写入数据（更安全）
                - False: 覆盖模式，直接覆盖目标范围的现有数据（默认行为）
            streaming: 是否使用流式写入（仅覆盖模式有效）

        Returns:
            Dict: 标准化的操作结果
        """
        # 新增：数据格式验证
        if not data:
            return cls._format_error_result("数据不能为空")

        if not isinstance(data, list):
            return cls._format_error_result(f"数据格式错误：data 应该是二维数组 [[row1], [row2], ...]，实际收到类型: {type(data).__name__}")

        # 验证每一行是否都是列表
        for i, row in enumerate(data):
            if not isinstance(row, list):
                return cls._format_error_result(f"数据格式错误：第 {i + 1} 行应该是列表，实际收到类型: {type(row).__name__}。data 应该是二维数组 [[row1], [row2], ...]")

        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始更新范围数据: {range_expression}, 模式: {'插入' if insert_mode else '覆盖'}")

        try:
            # 步骤1: 验证参数格式
            validation_result = cls._validate_range_format(range_expression)
            if not validation_result["valid"]:
                return cls._format_error_result(validation_result["error"])

            # 步骤2: 扩展流式写入路径（支持覆盖+插入模式）
            if streaming:
                from excel_mcp_server_fastmcp.core.streaming_writer import (
                    StreamingWriter,
                )

                if StreamingWriter.is_available():
                    try:
                        from excel_mcp_server_fastmcp.utils.parsers import RangeParser

                        range_info_parsed = RangeParser.parse_range_expression(range_expression)
                        sheet_name = range_info_parsed.sheet_name

                        if sheet_name:
                            # 解析起始行列
                            from openpyxl.utils import range_boundaries

                            min_col, min_row, max_col, max_row = range_boundaries(range_info_parsed.cell_range)

                            # 根据模式选择不同的流式写入方法
                            if not insert_mode:
                                # 覆盖模式 - 直接使用流式写入
                                success, message, meta = StreamingWriter.update_range(
                                    file_path,
                                    sheet_name,
                                    min_row,
                                    min_col,
                                    data,
                                    preserve_formulas=preserve_formulas,
                                )
                            else:
                                # 插入模式 - 使用流式插入
                                success, message, meta = StreamingWriter.insert_rows_streaming(
                                    file_path,
                                    sheet_name,
                                    min_row,
                                    data,
                                    preserve_formulas=preserve_formulas,
                                )

                            if success:
                                return {
                                    "success": True,
                                    "message": message,
                                    "data": meta,
                                    "metadata": {
                                        "file_path": file_path,
                                        "sheet_name": sheet_name,
                                        "range": range_expression,
                                        "streaming_mode": "insert" if insert_mode else "overwrite",
                                        **meta,
                                    },
                                }
                            else:
                                logger.warning(f"流式{'插入' if insert_mode else '覆盖'}操作失败，降级到openpyxl: {message}")
                    except Exception as parse_err:
                        logger.warning(f"流式update_range范围解析失败，降级到openpyxl: {parse_err}")

            # 步骤3: 传统openpyxl路径
            writer = ExcelWriter(file_path)
            result = writer.update_range(range_expression, data, preserve_formulas, insert_mode)

            # 步骤3: 格式化结果
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"更新范围数据失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def list_sheets(cls, file_path: str) -> dict[str, Any]:
        """
        @intention 获取Excel文件中所有工作表基本信息（Token优化版本）

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)

        Returns:
            Dict: 仅包含工作表名称、行数、列数的简化信息

        Example:
            result = ExcelOperations.list_sheets("data.xlsx")
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始获取工作表列表: {file_path}")

        reader = None
        try:
            # 步骤1: 读取工作表信息
            reader = ExcelReader(file_path)
            result = reader.list_sheets()

            # 步骤2: 提取和格式化数据（Token优化：只返回核心信息）
            sheets_info = []
            sheet_names = []  # 向后兼容：简单的工作表名称列表
            if result.data:
                for sheet in result.data:
                    # 获取工作表的基本信息
                    sheet_info = {
                        "name": sheet.name,
                        "rows": 0,  # 默认值，后续更新
                        "cols": 0,  # 默认值，后续更新
                        "state": sheet.sheet_state,
                    }
                    sheets_info.append(sheet_info)
                    sheet_names.append(sheet.name)  # 向后兼容：简单名称列表

            total_sheets = result.metadata.get("total_sheets", len(sheets_info)) if result.metadata else len(sheets_info)

            # Token优化：简化响应结构，移除重复字段，只保留核心信息
            # 保持向后兼容：提供两种格式，旧格式直接在顶层，新格式在data中
            response = {
                "success": True,
                "message": f"获取到 {len(sheets_info)} 个工作表",
                "sheets": sheet_names,  # 向后兼容：简单的工作表名称列表
                "total_sheets": total_sheets,  # 向后兼容：直接提供total_sheets字段
                "data": {
                    "sheets": sheets_info,  # Token优化：详细信息列表
                    "total_sheets": total_sheets,
                },
                "meta": {"total_sheets": total_sheets, "file_path": file_path},
            }

            return response

        except Exception as e:
            error_msg = f"获取工作表列表失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)
        finally:
            # 步骤3: 确保清理资源，即使在异常情况下
            if reader is not None:
                try:
                    reader.close()
                except Exception as cleanup_error:
                    logger.warning(f"{cls._LOG_PREFIX} 清理ExcelReader资源时发生错误: {cleanup_error}")

    @classmethod
    def get_headers(
        cls,
        file_path: str,
        sheet_name: str,
        header_row: int = 1,
        max_columns: int | None = None,
    ) -> dict[str, Any]:
        """
        @intention 获取指定工作表的表头信息，支持游戏开发双行模式（字段描述+字段名）

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            sheet_name: 工作表名称
            header_row: 表头起始行号 (1-based，默认从第1行开始获取两行)
            max_columns: 最大列数限制，None表示自动截取到空列

        Returns:
            Dict: 包含双行表头信息
            {
                'success': bool,
                'data': List[str],  # 字段名列表（兼容性）
                'headers': List[str],  # 字段名列表（兼容性）
                'descriptions': List[str],  # 字段描述列表（第1行）
                'field_names': List[str],   # 字段名列表（第2行）
                'header_count': int,
                'sheet_name': str,
                'header_row': int,
                'message': str
            }

        Example:
            result = ExcelOperations.get_headers("data.xlsx", "Sheet1")
            # 第1行：['技能ID描述', '技能名称描述', '技能类型描述']
            # 第2行：['skill_id', 'skill_name', 'skill_type']
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始获取双行表头: {sheet_name}")

        reader = None
        try:
            # 步骤1: 构建双行范围表达式
            range_expression = cls._build_header_range(sheet_name, header_row, max_columns, dual_row=True)

            # 步骤2: 读取表头数据（两行）
            reader = ExcelReader(file_path)
            result = reader.get_range(range_expression)

            if not result.success:
                return cls._format_error_result(f"无法读取表头数据: {result.message}")

            # 步骤3: 解析双行表头信息
            header_info = cls._parse_dual_header_data(result.data, max_columns)
            is_dual_header = header_info.get("is_dual_header", False)

            # 如果是双行表头，field_names用第2行；否则用第1行
            effective_headers = header_info["field_names"] if is_dual_header else header_info["descriptions"]

            return {
                "success": True,
                # 向后兼容：直接提供核心字段在顶层
                "field_names": header_info["field_names"],
                "descriptions": header_info["descriptions"],
                "headers": effective_headers,
                "header_count": len(effective_headers),
                "sheet_name": sheet_name,
                "is_dual_header": is_dual_header,
                # 统一data字段，集中核心数据
                "data": {
                    "field_names": header_info["field_names"],
                    "descriptions": header_info["descriptions"],
                    "headers": effective_headers,
                    "is_dual_header": is_dual_header,
                },
                "meta": {
                    "sheet_name": sheet_name,
                    "header_row": header_row,
                    "header_count": len(effective_headers),
                    "max_columns": max_columns,
                    "dual_row_mode": is_dual_header,
                    "header_type": "dual" if is_dual_header else "single",
                },
                "message": f"成功获取{len(effective_headers)}个表头字段（{'双行表头：描述+字段名' if is_dual_header else '单行表头'}）",
            }

        except Exception as e:
            error_msg = f"获取表头失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)
        finally:
            # 确保清理资源
            if reader is not None:
                try:
                    reader.close()
                except Exception as cleanup_error:
                    logger.warning(f"{cls._LOG_PREFIX} 清理ExcelReader资源时发生错误: {cleanup_error}")

    @classmethod
    def create_file(cls, file_path: str, sheet_names: list[str] | None = None) -> dict[str, Any]:
        """
        @intention 创建新的Excel文件，支持自定义工作表配置

        Args:
            file_path: 新文件路径 (必须以.xlsx或.xlsm结尾)
            sheet_names: 工作表名称列表，None表示默认工作表

        Returns:
            Dict: 包含创建结果和文件信息

        Example:
            result = ExcelOperations.create_file("new_file.xlsx", ["数据", "分析"])
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始创建文件: {file_path}")

        try:
            # 步骤1: 执行文件创建
            result = ExcelManager.create_file(file_path, sheet_names)

            # 步骤2: 格式化结果
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"创建文件失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    # ==================== 分支实现 ====================

    # --- 参数验证 ---
    @classmethod
    def _validate_range_format(cls, range_expression: str) -> dict[str, Any]:
        """
        @intention 验证范围表达式格式

        Args:
            range_expression: 范围表达式，如 'Sheet1!A1:C10'

        Returns:
            Dict: 验证结果，包含以下字段:
                - valid (bool): 验证是否通过
                - error (str, optional): 错误信息，仅在 valid=False 时存在
        """
        if not range_expression or not range_expression.strip():
            return {"valid": False, "error": "range参数不能为空"}

        if "!" not in range_expression:
            return {
                "valid": False,
                "error": f"range必须包含工作表名。当前格式: '{range_expression}'，正确格式示例: 'Sheet1!A1:B2'",
            }

        return {"valid": True}

    @classmethod
    def _build_header_range(
        cls,
        sheet_name: str,
        header_row: int,
        max_columns: int | None,
        dual_row: bool = False,
    ) -> str:
        """
        @intention 构建表头范围表达式，支持单行或双行模式

        Args:
            sheet_name: 工作表名称
            header_row: 表头起始行号 (1-based)
            max_columns: 最大列数限制，None表示使用默认值（100列）
            dual_row: 是否使用双行模式（True时读取两行，False时读取一行）

        Returns:
            str: 范围表达式，如 'Sheet1!A1:CV2' 或 'Sheet1!A1:Z1'
        """
        if max_columns:
            # 如果指定了最大列数，使用具体范围
            end_column = get_column_letter(max_columns)
            if dual_row:
                # 双行模式：获取连续两行
                end_row = header_row + 1
                return f"{sheet_name}!A{header_row}:{end_column}{end_row}"
            else:
                # 单行模式（保持兼容性）
                return f"{sheet_name}!A{header_row}:{end_column}{header_row}"
        else:
            # 否则使用一个合理的默认范围（读取前100列）
            if dual_row:
                # 双行模式：获取连续两行
                end_row = header_row + 1
                return f"{sheet_name}!A{header_row}:CV{end_row}"  # CV = 第100列
            else:
                # 单行模式（保持兼容性）
                return f"{sheet_name}!A{header_row}:CV{header_row}"  # CV = 第100列

    @classmethod
    def _parse_header_data(cls, data: list[list], max_columns: int | None) -> list[str]:
        """
        @intention 解析表头数据

        Args:
            data: 单元格数据列表，每个元素可能是 CellInfo 对象或普通值
            max_columns: 最大列数限制，None表示自动截取到空列

        Returns:
            List[str]: 解析后的表头字符串列表
        """
        headers = []
        if data and len(data) > 0:
            first_row = data[0]
            for i, cell_info in enumerate(first_row):
                # 处理CellInfo对象和普通值
                cell_value = getattr(cell_info, "value", cell_info) if hasattr(cell_info, "value") else cell_info

                # 转换为字符串并清理
                if cell_value is not None:
                    str_value = str(cell_value).strip()
                    if str_value != "":
                        headers.append(str_value)
                    else:
                        # 空字符串的处理
                        if max_columns:
                            headers.append("")  # 指定max_columns时保留空字符串
                        else:
                            break  # 否则停止
                else:
                    # None值的处理
                    if max_columns:
                        headers.append("")  # 指定max_columns时将None转为空字符串
                    else:
                        break  # 否则停止

                # 如果指定了max_columns，检查是否已达到限制
                if max_columns and len(headers) >= max_columns:
                    break

        return headers

    @classmethod
    def _parse_dual_header_data(cls, data: list[list], max_columns: int | None) -> dict[str, list[str]]:
        """
        @intention 解析双行表头数据（字段描述 + 字段名），支持空值fallback机制

        Args:
            data: 二维数组数据，第一行为字段描述，第二行为字段名
            max_columns: 最大列数限制，None表示自动截取到空列

        Returns:
            Dict[str, List[str]]: 包含以下字段的字典:
                - descriptions (List[str]): 字段描述列表（第1行）
                - field_names (List[str]): 字段名列表（第2行）
        """
        descriptions = []
        field_names = []

        if not data or len(data) < 2:
            return {
                "descriptions": descriptions,
                "field_names": field_names,
                "is_dual_header": False,
            }

        # 解析第一行（字段描述）
        first_row = data[0] if len(data) > 0 else []
        # 解析第二行（字段名）
        second_row = data[1] if len(data) > 1 else []

        # 确定实际处理的列数
        max_cols = max(len(first_row), len(second_row))  # 改为取最大值，不遗漏任何列
        if max_columns:
            max_cols = min(max_cols, max_columns)

        # 检测是否为双行表头（与server.py的_detect_dual_header逻辑一致）
        # 注意：data中的元素可能是CellInfo对象，需要通过.value访问实际值
        def _cell_str(c):
            if c is None:
                return None
            if hasattr(c, "value"):
                return c.value
            return c

        is_dual_header = False
        # 需要至少2列才能判断
        if len(first_row) >= 2 and len(second_row) >= 2:
            # 第二行是否全是有效英文字段名
            second_row_strs = [_cell_str(c) for c in second_row if _cell_str(c) is not None]
            all_valid_names = all(isinstance(v, str) and v.strip() and v.strip()[0].isalpha() and v.strip()[0].isascii() for v in second_row_strs)
            # 第一行是否有中文
            first_row_strs = [_cell_str(c) for c in first_row if _cell_str(c) is not None]
            any_chinese = any(isinstance(v, str) and any("\u4e00" <= ch <= "\u9fff" for ch in v) for v in first_row_strs)
            is_dual_header = all_valid_names and any_chinese and len(second_row_strs) >= 2

        for i in range(max_cols):
            # 处理字段描述（第1行）
            desc_cell = first_row[i] if i < len(first_row) else None
            desc_value = getattr(desc_cell, "value", desc_cell) if hasattr(desc_cell, "value") else desc_cell
            desc_str = str(desc_value).strip() if desc_value is not None and str(desc_value).strip() else ""

            # 处理字段名（第2行）
            name_cell = second_row[i] if i < len(second_row) else None
            name_value = getattr(name_cell, "value", name_cell) if hasattr(name_cell, "value") else name_cell
            name_str = str(name_value).strip() if name_value is not None and str(name_value).strip() else ""

            # 🆕 智能Fallback机制
            column_letter = get_column_letter(i + 1)  # 1-based列名：A, B, C...

            if is_dual_header:
                # 双行模式：第1行描述，第2行字段名
                if not desc_str:
                    desc_str = f"列{column_letter}"
                if not name_str:
                    name_str = column_letter.lower()
            else:
                # 单行模式：第1行就是字段名，描述为空
                if not desc_str and name_str:
                    desc_str = name_str
                    name_str = ""
                elif not desc_str:
                    desc_str = f"列{column_letter}"
                if not name_str:
                    name_str = ""  # 单行模式不需要fallback字段名

            # 🆕 检查是否应该停止（简化的停止条件）
            # 只有在没有指定max_columns时才进行智能停止
            if not max_columns:
                # 检查原始数据是否为完全空（描述和字段名都是原始空值）
                desc_is_empty = desc_cell is None or (hasattr(desc_cell, "value") and desc_cell.value is None) or (not hasattr(desc_cell, "value") and desc_cell is None)
                name_is_empty = name_cell is None or (hasattr(name_cell, "value") and name_cell.value is None) or (not hasattr(name_cell, "value") and name_cell is None)

                # 如果当前列完全为空，检查接下来连续3列是否也为空
                if desc_is_empty and name_is_empty:
                    consecutive_empty = 0
                    for j in range(i, min(i + 3, max_cols)):  # 检查当前及后续2列
                        check_desc = first_row[j] if j < len(first_row) else None
                        check_name = second_row[j] if j < len(second_row) else None

                        desc_empty = check_desc is None or (hasattr(check_desc, "value") and check_desc.value is None) or (not hasattr(check_desc, "value") and check_desc is None)
                        name_empty = check_name is None or (hasattr(check_name, "value") and check_name.value is None) or (not hasattr(check_name, "value") and check_name is None)

                        if desc_empty and name_empty:
                            consecutive_empty += 1
                        else:
                            break

                    # 如果连续3列都为空，则停止
                    if consecutive_empty >= 3:
                        break

            descriptions.append(desc_str)
            field_names.append(name_str)

            # 如果指定了max_columns，检查是否已达到限制
            if max_columns and len(field_names) >= max_columns:
                break

        return {
            "descriptions": descriptions,
            "field_names": field_names,
            "is_dual_header": is_dual_header,
        }

    @classmethod
    def search(
        cls,
        file_path: str,
        pattern: str,
        sheet_name: str | None = None,
        case_sensitive: bool = False,
        whole_word: bool = False,
        use_regex: bool = False,
        include_values: bool = True,
        include_formulas: bool = False,
        range: str | None = None,
    ) -> dict[str, Any]:
        """
        @intention 在Excel文件中搜索单元格内容（VSCode风格搜索选项）

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            pattern: 搜索模式（正则表达式或字面字符串）
            sheet_name: 工作表名称 (可选)
            case_sensitive: 大小写敏感
            whole_word: 全词匹配
            use_regex: 启用正则表达式
            include_values: 是否搜索单元格值
            include_formulas: 是否搜索公式内容
            range: 搜索范围表达式

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            search_type = "正则" if use_regex else ("全词" if whole_word else "字符串")
            case_info = "大小写敏感" if case_sensitive else "忽略大小写"
            logger.info(f"{cls._LOG_PREFIX} 开始{search_type}搜索({case_info}): {pattern}")

        try:
            searcher = ExcelSearcher(file_path)

            # 构建正则表达式模式
            if use_regex:
                # 直接使用用户提供的正则表达式
                regex_pattern = pattern
            else:
                # 将字面字符串转义为正则表达式
                escaped_pattern = re.escape(pattern)

                # 如果是全词匹配，添加单词边界
                if whole_word:
                    regex_pattern = r"\b" + escaped_pattern + r"\b"
                else:
                    regex_pattern = escaped_pattern

            # 构建正则表达式标志
            regex_flags = "" if case_sensitive else "i"

            result = searcher.regex_search(
                regex_pattern,
                regex_flags,
                include_values,
                include_formulas,
                sheet_name,
                range,
            )
            return format_operation_result(result)

        except Exception as e:
            search_type = "正则" if use_regex else ("全词" if whole_word else "字符串")
            error_msg = f"{search_type}搜索失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def search_directory(
        cls,
        directory_path: str,
        pattern: str,
        case_sensitive: bool = False,
        whole_word: bool = False,
        use_regex: bool = False,
        include_values: bool = True,
        include_formulas: bool = False,
        recursive: bool = True,
        file_extensions: list[str] | None = None,
        file_pattern: str | None = None,
        max_files: int = MAX_SEARCH_FILES,
    ) -> dict[str, Any]:
        """
        @intention 在目录下的所有Excel文件中搜索内容（VSCode风格搜索选项）

        Args:
            directory_path: 目录路径
            pattern: 搜索模式（正则表达式或字面字符串）
            case_sensitive: 大小写敏感
            whole_word: 全词匹配
            use_regex: 启用正则表达式
            其他参数同search方法

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            search_type = "正则" if use_regex else ("全词" if whole_word else "字符串")
            case_info = "大小写敏感" if case_sensitive else "忽略大小写"
            logger.info(f"{cls._LOG_PREFIX} 开始目录{search_type}搜索({case_info}): {directory_path}")

        try:
            # 构建正则表达式模式
            if use_regex:
                # 直接使用用户提供的正则表达式
                regex_pattern = pattern
            else:
                # 将字面字符串转义为正则表达式
                escaped_pattern = re.escape(pattern)

                # 如果是全词匹配，添加单词边界
                if whole_word:
                    regex_pattern = r"\b" + escaped_pattern + r"\b"
                else:
                    regex_pattern = escaped_pattern

            # 构建正则表达式标志
            regex_flags = "" if case_sensitive else "i"

            result = ExcelSearcher.search_directory_static(
                directory_path,
                regex_pattern,
                regex_flags,
                include_values,
                include_formulas,
                recursive,
                file_extensions,
                file_pattern,
                max_files,
            )
            return format_operation_result(result)

        except Exception as e:
            search_type = "正则" if use_regex else ("全词" if whole_word else "字符串")
            error_msg = f"目录{search_type}搜索失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def get_all_headers(cls, file_path: str, header_row: int = 1, max_columns: int | None = None) -> dict[str, Any]:
        """
        @intention 获取Excel文件中所有工作表的双行表头信息（字段描述+字段名）

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            header_row: 表头起始行号 (1-based，默认1)
            max_columns: 最大列数限制，None表示自动截取到空列

        Returns:
            Dict: 包含所有工作表的双行表头信息
            {
                'success': bool,
                'sheets_with_headers': [
                    {
                        'name': str,
                        'headers': List[str],       # 字段名（兼容性）
                        'descriptions': List[str],  # 字段描述（第1行）
                        'field_names': List[str],   # 字段名（第2行）
                        'header_count': int
                    }
                ],
                'file_path': str,
                'total_sheets': int
            }
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始获取所有工作表双行表头: {file_path}")

        try:
            # 步骤1: 获取所有工作表列表
            sheets_result = cls.list_sheets(file_path)
            if not sheets_result.get("success"):
                return sheets_result

            # 步骤2: 获取每个工作表的双行表头
            sheets_with_headers = []
            # Operations层内部调用，从data中取sheets列表
            sheets_data = sheets_result.get("data", {}).get("sheets", sheets_result.get("sheets", []))
            sheets = [s["name"] if isinstance(s, dict) else s for s in sheets_data]

            for sheet_name in sheets:
                try:
                    header_result = cls.get_headers(
                        file_path,
                        sheet_name,
                        header_row=header_row,
                        max_columns=max_columns,
                    )

                    if header_result.get("success"):
                        # Operations层内部调用，优先从data取
                        hdr_data = header_result.get("data", header_result)
                        headers = hdr_data.get("headers", [])
                        descriptions = hdr_data.get("descriptions", [])
                        field_names = hdr_data.get("field_names", [])

                        # 如果没有获取到field_names，使用headers作为fallback
                        if not field_names and headers:
                            field_names = headers

                        sheets_with_headers.append(
                            {
                                "name": sheet_name,
                                "headers": field_names,  # 兼容性字段，使用字段名
                                "descriptions": descriptions,  # 字段描述（第1行）
                                "field_names": field_names,  # 字段名（第2行）
                                "header_count": len(field_names),
                            }
                        )
                    else:
                        sheets_with_headers.append(
                            {
                                "name": sheet_name,
                                "headers": [],
                                "descriptions": [],
                                "field_names": [],
                                "header_count": 0,
                                "error": header_result.get("error", "未知错误"),
                            }
                        )

                except Exception as e:
                    sheets_with_headers.append(
                        {
                            "name": sheet_name,
                            "headers": [],
                            "descriptions": [],
                            "field_names": [],
                            "header_count": 0,
                            "error": str(e),
                        }
                    )

            return format_operation_result(
                {
                    "success": True,
                    "data": {
                        "sheets_with_headers": sheets_with_headers,
                        "total_sheets": len(sheets),
                    },
                    # 保持顶层兼容性
                    "sheets_with_headers": sheets_with_headers,
                    "file_path": file_path,
                    "total_sheets": len(sheets),
                }
            )

        except Exception as e:
            error_msg = f"获取工作表表头失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def insert_rows(
        cls,
        file_path: str,
        sheet_name: str,
        row_index: int,
        count: int = 1,
        streaming: bool = True,
    ) -> dict[str, Any]:
        """
        @intention 在指定位置插入空行

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            sheet_name: 工作表名称
            row_index: 插入位置 (1-based)
            count: 插入行数
            streaming: 是否使用流式写入（默认True，大文件性能更好）

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始插入行: {sheet_name} 第{row_index}行")

        try:
            # 流式写入路径
            if streaming:
                from excel_mcp_server_fastmcp.core.streaming_writer import (
                    StreamingWriter,
                )

                if StreamingWriter.is_available():
                    # 需要先知道列数来创建空行
                    reader = ExcelReader(file_path)
                    range_expr = f"{sheet_name}!A1"
                    header_result = reader.get_range(range_expr)
                    reader.close()
                    # 获取工作表列数
                    col_count = 0
                    if header_result.success and header_result.data:
                        col_count = max(len(row) for row in header_result.data)
                    if col_count == 0:
                        col_count = 1  # fallback
                    # 创建空行数据用于插入
                    empty_rows = [[None] * col_count for _ in range(count)]
                    success, message, meta = StreamingWriter.insert_rows_streaming(file_path, sheet_name, row_index, empty_rows)
                    if success:
                        return {
                            "success": True,
                            "message": message,
                            "data": meta,
                            "metadata": {
                                "file_path": file_path,
                                "sheet_name": sheet_name,
                                "row_index": row_index,
                                "count": count,
                                "mode": "streaming",
                                **meta,
                            },
                        }
                    else:
                        logger.warning(f"流式插入行失败，降级到openpyxl: {message}")

            # 传统openpyxl路径
            writer = ExcelWriter(file_path)
            result = writer.insert_rows(sheet_name, row_index, count)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"插入行操作失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def insert_columns(
        cls,
        file_path: str,
        sheet_name: str,
        column_index: int,
        count: int = 1,
        streaming: bool = True,
    ) -> dict[str, Any]:
        """
        @intention 在指定位置插入空列

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            sheet_name: 工作表名称
            column_index: 插入位置 (1-based)
            count: 插入列数
            streaming: 是否使用流式写入（默认True，大文件性能更好）

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始插入列: {sheet_name} 第{column_index}列")

        try:
            # 流式写入路径
            if streaming:
                from excel_mcp_server_fastmcp.core.streaming_writer import (
                    StreamingWriter,
                )

                if StreamingWriter.is_available():
                    success, message, meta = StreamingWriter.insert_columns_streaming(file_path, sheet_name, column_index, count)
                    if success:
                        return {
                            "success": True,
                            "message": message,
                            "data": meta,
                            "metadata": {
                                "file_path": file_path,
                                "sheet_name": sheet_name,
                                "column_index": column_index,
                                "count": count,
                                "mode": "streaming",
                                **meta,
                            },
                        }
                    else:
                        logger.warning(f"流式插入列失败，降级到openpyxl: {message}")

            # 传统openpyxl路径
            writer = ExcelWriter(file_path)
            result = writer.insert_columns(sheet_name, column_index, count)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"插入列操作失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def export_to_csv(
        cls,
        file_path: str,
        output_path: str,
        sheet_name: str | None = None,
        encoding: str = "utf-8",
    ) -> dict[str, Any]:
        """
        @intention 将Excel工作表导出为CSV文件

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            output_path: 输出CSV文件路径
            sheet_name: 工作表名称 (默认使用活动工作表)
            encoding: 文件编码 (默认: utf-8)

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始导出为CSV: {output_path}")

        try:
            converter = ExcelConverter(file_path)
            result = converter.export_to_csv(output_path, sheet_name, encoding)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"导出为CSV失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def import_from_csv(
        cls,
        csv_path: str,
        output_path: str,
        sheet_name: str = "Sheet1",
        encoding: str = "utf-8",
        has_header: bool = True,
    ) -> dict[str, Any]:
        """
        @intention 从CSV文件导入数据创建Excel文件

        Args:
            csv_path: CSV文件路径
            output_path: 输出Excel文件路径
            sheet_name: 工作表名称
            encoding: CSV文件编码
            has_header: 是否包含表头行

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始从CSV导入: {csv_path}")

        try:
            result = ExcelConverter.import_from_csv(csv_path, output_path, sheet_name, encoding, has_header)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"从CSV导入失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def convert_format(cls, input_path: str, output_path: str, target_format: str = "xlsx") -> dict[str, Any]:
        """
        @intention 转换Excel文件格式

        Args:
            input_path: 输入文件路径
            output_path: 输出文件路径
            target_format: 目标格式

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始格式转换: {input_path} -> {output_path}")

        try:
            result = ExcelConverter.convert_format(input_path, output_path, target_format)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"文件格式转换失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def merge_files(cls, input_files: list[str], output_path: str, merge_mode: str = "sheets") -> dict[str, Any]:
        """
        @intention 合并多个Excel文件

        Args:
            input_files: 输入文件路径列表
            output_path: 输出文件路径
            merge_mode: 合并模式

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始合并文件: {len(input_files)}个文件")

        try:
            result = ExcelConverter.merge_files(input_files, output_path, merge_mode)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"合并Excel文件失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def get_file_info(cls, file_path: str) -> dict[str, Any]:
        """
        @intention 获取Excel文件的详细信息

        Args:
            file_path: Excel文件路径

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始获取文件信息: {file_path}")

        try:
            result = ExcelManager.get_file_info(file_path)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"获取文件信息失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def create_sheet(cls, file_path: str, sheet_name: str, index: int | None = None) -> dict[str, Any]:
        """
        @intention 在文件中创建新工作表

        Args:
            file_path: Excel文件路径
            sheet_name: 新工作表名称
            index: 插入位置

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始创建工作表: {sheet_name}")

        try:
            manager = ExcelManager(file_path)
            result = manager.create_sheet(sheet_name, index)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"创建工作表失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def delete_sheet(cls, file_path: str, sheet_name: str) -> dict[str, Any]:
        """
        @intention 删除指定工作表

        Args:
            file_path: Excel文件路径
            sheet_name: 要删除的工作表名称

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始删除工作表: {sheet_name}")

        try:
            manager = ExcelManager(file_path)
            result = manager.delete_sheet(sheet_name)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"删除工作表失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def rename_sheet(cls, file_path: str, old_name: str, new_name: str) -> dict[str, Any]:
        """
        @intention 重命名工作表

        Args:
            file_path: Excel文件路径
            old_name: 当前工作表名称
            new_name: 新工作表名称

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始重命名工作表: {old_name} -> {new_name}")

        try:
            manager = ExcelManager(file_path)
            result = manager.rename_sheet(old_name, new_name)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"重命名工作表失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def copy_sheet(
        cls,
        file_path: str,
        source_name: str,
        new_name: str | None = None,
        index: int | None = None,
        streaming: bool = True,
    ) -> dict[str, Any]:
        """
        @intention 复制工作表（含数据和格式）

        Args:
            file_path: Excel文件路径
            source_name: 源工作表名称
            new_name: 新工作表名称（为空自动生成 "源表名_副本"）
            index: 插入位置（None追加到末尾）
            streaming: 是否使用流式复制（默认True，大文件性能更好）

        Returns:
            Dict: 标准化的操作结果
        """
        try:
            manager = ExcelManager(file_path)
            result = manager.copy_sheet(source_name, new_name, index, streaming)
            return format_operation_result(result)
        except Exception as e:
            error_msg = f"复制工作表失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def rename_column(
        cls,
        file_path: str,
        sheet_name: str,
        old_header: str,
        new_header: str,
        header_row: int = 1,
    ) -> dict[str, Any]:
        """
        @intention 重命名列（修改表头单元格值）

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            old_header: 当前列名（支持中文/英文）
            new_header: 新列名
            header_row: 表头所在行号（默认1=自动检测）

        Returns:
            Dict: 标准化的操作结果
        """
        try:
            # 自动检测双表头：确定实际表头行
            _effective_header_row = header_row
            if header_row == 1:
                try:
                    info = HeaderAnalyzer.analyze(file_path, sheet_name)
                    if info.is_dual:
                        _effective_header_row = 2  # 双表头时重命名第2行（英文字段名）
                except Exception:
                    pass

            manager = ExcelManager(file_path)
            result = manager.rename_column(sheet_name, old_header, new_header, _effective_header_row)
            return format_operation_result(result)
        except Exception as e:
            error_msg = f"重命名列失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def delete_rows(
        cls,
        file_path: str,
        sheet_name: str,
        row_index: int,
        count: int = 1,
        streaming: bool = True,
    ) -> dict[str, Any]:
        """
        @intention 删除指定行

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            row_index: 起始行号 (1-based)
            count: 删除行数
            streaming: 是否使用流式写入（默认True，大文件性能更好）

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始删除行: {sheet_name} 第{row_index}行")

        try:
            # 流式写入路径
            if streaming:
                from excel_mcp_server_fastmcp.core.streaming_writer import (
                    StreamingWriter,
                )

                if StreamingWriter.is_available():
                    success, message, meta = StreamingWriter.delete_rows(file_path, sheet_name, row_index, count)
                    if success:
                        return {
                            "success": True,
                            "message": message,
                            "data": meta,
                            "metadata": {
                                "file_path": file_path,
                                "sheet_name": sheet_name,
                                **meta,
                            },
                        }
                    else:
                        logger.warning(f"流式删除行失败，降级到openpyxl: {message}")

            writer = ExcelWriter(file_path)
            result = writer.delete_rows(sheet_name, row_index, count)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"删除行操作失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def batch_delete_rows(
        cls,
        file_path: str,
        sheet_name: str,
        row_numbers: list[int],
        streaming: bool = True,
    ) -> dict[str, Any]:
        """批量删除多个行号对应的行，仅一次文件I/O。

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            row_numbers: 待删除的行号列表（1-based，调用方应确保从大到小排序或使用streaming）
            streaming: 是否使用流式写入（默认True，推荐）

        Returns:
            Dict: 标准化的操作结果，包含 deleted_count
        """
        if not row_numbers:
            return {"success": False, "message": "行号列表为空", "data": None}

        try:
            if streaming:
                from excel_mcp_server_fastmcp.core.streaming_writer import (
                    StreamingWriter,
                )

                if StreamingWriter.is_available():
                    success, message, meta = StreamingWriter.batch_delete_rows(file_path, sheet_name, row_numbers)
                    if success:
                        return {
                            "success": True,
                            "message": message,
                            "data": meta,
                            "metadata": {
                                "file_path": file_path,
                                "sheet_name": sheet_name,
                                **meta,
                            },
                        }
                    else:
                        logger.warning(f"流式批量删除行失败，降级到openpyxl: {message}")

            # openpyxl 降级路径：合并相邻行号为连续区间，逐区间删除
            # Fix(P1-03): 必须从大到小(逆序)处理区间，避免先删小行号导致后续行号偏移
            from excel_mcp_server_fastmcp.core.excel_writer import ExcelWriter

            writer = ExcelWriter(file_path)
            sorted_rows = sorted(set(row_numbers))

            # 合并为连续区间以减少操作次数
            ranges = []
            start = sorted_rows[0]
            prev = start
            for r in sorted_rows[1:]:
                if r == prev + 1:
                    prev = r
                else:
                    ranges.append((start, prev - start + 1))
                    start = r
                    prev = r
            ranges.append((start, prev - start + 1))

            # 逆序排列：从最大行号开始删除，避免行号偏移
            ranges.sort(reverse=True)

            total_deleted = 0
            for range_start, range_count in ranges:
                result = writer.delete_rows(sheet_name, range_start, range_count)
                if result.success:
                    total_deleted += range_count
                else:
                    logger.warning(f"删除行 {range_start}+{range_count} 失败: {result.message}")

            return {
                "success": True,
                "message": f"批量删除了{total_deleted}行（{len(ranges)}个区间）",
                "data": {"deleted_rows": total_deleted, "ranges": len(ranges)},
                "metadata": {
                    "file_path": file_path,
                    "sheet_name": sheet_name,
                    "deleted_rows": total_deleted,
                },
            }

        except Exception as e:
            error_msg = f"批量删除行失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def delete_columns(
        cls,
        file_path: str,
        sheet_name: str,
        column_index: int,
        count: int = 1,
        streaming: bool = True,
    ) -> dict[str, Any]:
        """
        @intention 删除指定列

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            column_index: 起始列号 (1-based)
            count: 删除列数
            streaming: 是否使用流式写入（默认True，大文件性能更好）

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始删除列: {sheet_name} 第{column_index}列")

        try:
            # 流式写入路径
            if streaming:
                from excel_mcp_server_fastmcp.core.streaming_writer import (
                    StreamingWriter,
                )

                if StreamingWriter.is_available():
                    success, message, meta = StreamingWriter.delete_columns(file_path, sheet_name, column_index, count)
                    if success:
                        return {
                            "success": True,
                            "message": message,
                            "data": meta,
                            "metadata": {
                                "file_path": file_path,
                                "sheet_name": sheet_name,
                                **meta,
                            },
                        }
                    else:
                        logger.warning(f"流式删除列失败，降级到openpyxl: {message}")

            writer = ExcelWriter(file_path)
            result = writer.delete_columns(sheet_name, column_index, count)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"删除列操作失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @staticmethod
    def _normalize_formatting(formatting: dict | None) -> dict:
        """将LLM友好的扁平格式转换为Writer层需要的嵌套格式。

        扁平格式（API层/LLM输入）:
            {"bold": True, "alignment": "center", "bg_color": "FF0", "font_size": 14,
             "underline": "double", "strikethrough": True,
             "gradient_colors": ["4472C4", "ED7D31"],
             "border": {"top": "medium", "color": "000000"},
             "text_rotation": 45, "indent": 2}
        嵌套格式（Writer层/_apply_cell_format）:
            {"font": {"bold": True, "size": 14, "underline": "double"},
             "alignment": {"horizontal": "center", "text_rotation": 45},
             "fill": {"type": "gradient", "colors": ["4472C4", "ED7D31"]},
             "border": {"top": {"style": "medium"}, "color": "000000"}}

        已是嵌套格式的（如preset产生的）直接原样返回，不做二次转换。

        支持的扁平字段映射（v1.9.3+ 完整列表）：
            字体: bold, italic, underline(single/double/singleAccounting/doubleAccounting),
                  strikethrough, font_size(size), font_color(color), font_name(name)
            填充: bg_color → fill.color, fill_type(solid/gradient/pattern),
                  gradient_colors(list), gradient_type(linear/radial/path)
            对齐: alignment(horizontal), vertical_alignment(vertical),
                  wrap_text, text_rotation(-90~90), indent, shrink_to_fit
            边框: border(dict) — 支持 {left/right/top/bottom/diagonal: style|dict, color}
            数字: number_format
        """
        if not formatting or not isinstance(formatting, dict):
            return formatting or {}

        # 检测是否已经是嵌套格式（包含 font/fill/alignment 等嵌套键）
        nested_keys = {"font", "fill", "alignment"}
        if any(k in formatting and isinstance(formatting.get(k), dict) for k in nested_keys):
            return formatting  # 已经是嵌套格式，直接返回

        # 扁平 → 嵌套 转换
        nested: dict[str, Any] = {}
        font_attrs: dict[str, Any] = {}
        align_attrs: dict[str, Any] = {}

        flat_to_font = {
            "bold": "bold",
            "italic": "italic",
            "underline": "underline",
            "font_size": "size",
            "font_color": "color",
            "font_name": "name",
            "strikethrough": "strikethrough",
        }
        flat_to_align = {
            "alignment": "horizontal",
            "vertical_alignment": "vertical",
            "wrap_text": "wrap_text",
            "text_rotation": "text_rotation",
            "indent": "indent",
            "shrink_to_fit": "shrink_to_fit",
        }

        for key, value in formatting.items():
            if key in flat_to_font and value is not None:
                font_attrs[flat_to_font[key]] = value
            elif key in flat_to_align and value is not None:
                align_attrs[flat_to_align[key]] = value
            elif key == "bg_color" and value is not None:
                # 合并到已有 fill 配置（而非覆盖 fill_type/gradient_colors 等字段）
                if "fill" not in nested:
                    nested["fill"] = {}
                nested["fill"]["color"] = str(value)
            elif key == "number_format" and value is not None:
                nested["number_format"] = str(value)
            elif key == "fill_type" and value is not None:
                # 渐变/图案填充类型，与 bg_color 合并到 fill
                if "fill" not in nested:
                    nested["fill"] = {}
                nested["fill"]["type"] = str(value)
            elif key == "gradient_colors" and value is not None:
                if "fill" not in nested:
                    nested["fill"] = {}
                nested["fill"]["type"] = "gradient"
                nested["fill"]["colors"] = list(value) if isinstance(value, (list, tuple)) else [str(value)]
            elif key == "gradient_type" and value is not None:
                if "fill" not in nested:
                    nested["fill"] = {}
                nested["fill"]["gradient_type"] = str(value)
            elif key == "border" and value is not None:
                # 行内边框配置：支持简写字符串或详细字典
                if isinstance(value, str):
                    nested["border"] = {"style": value}
                else:
                    nested["border"] = value
            else:
                # 未知键直接透传（保持向后兼容）
                nested[key] = value

        if font_attrs:
            nested["font"] = font_attrs
        if align_attrs:
            nested["alignment"] = align_attrs

        return nested

    @classmethod
    def format_cells(
        cls,
        file_path: str,
        sheet_name: str,
        range: str,
        formatting: dict[str, Any] | None = None,
        preset: str | None = None,
    ) -> dict[str, Any]:
        """
        @intention 设置单元格格式

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            range: 目标范围
            formatting: 自定义格式配置
            preset: 预设样式

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始格式化单元格: {range}")

        try:
            writer = ExcelWriter(file_path)
            # 处理预设格式
            if preset:
                preset_formats = {
                    "title": {
                        "font": {"name": "微软雅黑", "size": 14, "bold": True},
                        "alignment": {"horizontal": "center"},
                    },
                    "header": {
                        "font": {"name": "微软雅黑", "size": 11, "bold": True},
                        "fill": {"color": "D9D9D9"},
                    },
                    "data": {"font": {"name": "微软雅黑", "size": 10}},
                    "highlight": {"fill": {"color": "FFFF00"}},
                    "currency": {"number_format": "¥#,##0.00"},
                }
                formatting = preset_formats.get(preset, formatting or {})

            # 扁平格式 → 嵌套格式转换（LLM友好API用扁平格式，Writer层需要嵌套格式）
            # 扁平: {"bold": True, "alignment": "center", "bg_color": "FF0"}
            # 嵌套: {"font": {"bold": True}, "alignment": {"horizontal": "center"}, "fill": {"color": "FF0"}}
            formatting = cls._normalize_formatting(formatting)

            # 构建完整的range表达式
            if "!" not in range:
                range_expression = f"{sheet_name}!{range}"
            else:
                range_expression = range

            result = writer.format_cells(range_expression, formatting or {})
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"单元格格式化失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    # --- 错误处理 ---
    @classmethod
    def _format_error_result(cls, error_message: str) -> dict[str, Any]:
        """
        @intention 创建标准化的错误响应

        Args:
            error_message: 错误消息字符串

        Returns:
            Dict[str, Any]: 标准化的错误结果，包含以下字段:
                - success (bool): 固定为 False
                - message (str): 错误消息
                - data (None): 固定为 None
        """
        return {"success": False, "message": error_message, "data": None}

    # --- 单元格操作扩展 ---
    @classmethod
    def merge_cells(cls, file_path: str, sheet_name: str, range: str) -> dict[str, Any]:
        """
        @intention 合并指定范围的单元格

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            range: 要合并的单元格范围（如 A1:C3）

        Returns:
            Dict: 操作结果，成功或失败信息
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始合并单元格: {range}")

        try:
            writer = ExcelWriter(file_path)
            result = writer.merge_cells(range, sheet_name)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"合并单元格失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def unmerge_cells(cls, file_path: str, sheet_name: str, range: str) -> dict[str, Any]:
        """
        @intention 取消合并指定范围的单元格

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            range: 要取消合并的单元格范围（如 A1:C3）

        Returns:
            Dict: 操作结果，成功或失败信息
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始取消合并单元格: {range}")

        try:
            writer = ExcelWriter(file_path)
            result = writer.unmerge_cells(range, sheet_name)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"取消合并单元格失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def set_borders(cls, file_path: str, sheet_name: str, range: str, border_style: str = "thin") -> dict[str, Any]:
        """
        @intention 为指定范围设置边框样式

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            range: 要设置边框的单元格范围（如 A1:C3）
            border_style: 边框样式，如 thin、medium、thick（默认：thin）

        Returns:
            Dict: 操作结果，成功或失败信息
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始设置边框: {range}, 样式: {border_style}")

        try:
            writer = ExcelWriter(file_path)
            result = writer.set_borders(range, border_style, sheet_name)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"设置边框失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def set_row_height(
        cls,
        file_path: str,
        sheet_name: str,
        row_index: int,
        height: float,
        count: int = 1,
    ) -> dict[str, Any]:
        """
        @intention 调整指定行的高度

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            row_index: 起始行号（从1开始）
            height: 行高（磅）
            count: 要设置的连续行数（默认：1）

        Returns:
            Dict: 操作结果，成功或失败信息
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始调整行高: 行{row_index}, 高度{height}, 数量{count}")

        try:
            writer = ExcelWriter(file_path)

            # ExcelWriter.set_row_height(row_number, height, sheet_name)
            for i in range(count):
                row_num = row_index + i
                result = writer.set_row_height(row_num, height, sheet_name)
                if not result.success:
                    break

            return format_operation_result(result)

        except Exception as e:
            error_msg = f"调整行高失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def set_column_width(
        cls,
        file_path: str,
        sheet_name: str,
        column_index: int,
        width: float,
        count: int = 1,
    ) -> dict[str, Any]:
        """
        @intention 调整指定列的宽度

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            column_index: 起始列号（从1开始）
            width: 列宽（字符）
            count: 要设置的连续列数（默认：1）

        Returns:
            Dict: 操作结果，成功或失败信息
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始调整列宽: 列{column_index}, 宽度{width}, 数量{count}")

        try:
            writer = ExcelWriter(file_path)

            # ExcelWriter.set_column_width(column, width, sheet_name)
            for i in range(count):
                col_idx = column_index + i
                column_letter = get_column_letter(col_idx)
                result = writer.set_column_width(column_letter, width, sheet_name)
                if not result.success:
                    break

            return format_operation_result(result)

        except Exception as e:
            error_msg = f"调整列宽失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def compare_sheets(
        cls,
        file1_path: str,
        sheet1_name: str,
        file2_path: str,
        sheet2_name: str,
        id_column: int | str = 1,
        header_row: int = 1,
    ) -> dict[str, Any]:
        """
        @intention 比较两个Excel工作表，识别ID对象的新增、删除、修改

        Args:
            file1_path: 第一个Excel文件路径
            sheet1_name: 第一个工作表名称
            file2_path: 第二个Excel文件路径
            sheet2_name: 第二个工作表名称
            id_column: ID列的索引或列名（默认：1，即第一列）
            header_row: 表头行号（默认：1）

        Returns:
            Dict: 比较结果，包含新增、删除、修改的记录
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始比较工作表: {file1_path}:{sheet1_name} vs {file2_path}:{sheet2_name}")

        try:
            # 创建比较选项
            options = ComparisonOptions()
            comparer = ExcelComparer(options)

            # 执行比较 - 使用正确的参数顺序
            result = comparer.compare_sheets(file1_path, sheet1_name, file2_path, sheet2_name, options)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"比较工作表失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    # --- 公式操作扩展 ---
    @classmethod
    def set_formula(cls, file_path: str, sheet_name: str, cell_range: str, formula: str) -> dict[str, Any]:
        """
        @intention 设置指定单元格或区域的公式

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            cell_range: 单元格或区域（如 A1 或 A1:C3）
            formula: 公式表达式（以 = 开头）

        Returns:
            Dict: 操作结果，成功或失败信息
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始设置公式: {cell_range} = {formula}")

        try:
            writer = ExcelWriter(file_path)
            result = writer.set_formula(cell_range, formula, sheet_name)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"设置公式失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def evaluate_formula(cls, formula: str, context_sheet: str | None = None) -> dict[str, Any]:
        """
        @intention 计算公式的值，不修改文件

        Args:
            formula: 要计算的公式表达式（可以不包含 =）
            context_sheet: 上下文工作表（可选），用于引用单元格值

        Returns:
            Dict: 计算结果，成功或失败信息
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始计算公式: {formula}")

        try:
            writer = ExcelWriter("")  # 临时实例，不需要文件
            result = writer.evaluate_formula(formula, context_sheet)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"公式计算失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def compare_files(cls, file1_path: str, file2_path: str) -> dict[str, Any]:
        """
        @intention 比较两个Excel文件的所有工作表

        Args:
            file1_path: 第一个Excel文件路径
            file2_path: 第二个Excel文件路径

        Returns:
            Dict: 比较结果，包含各工作表的差异统计
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始比较文件: {file1_path} vs {file2_path}")

        try:
            # 标准文件比较配置
            options = ComparisonOptions(
                compare_values=True,
                compare_formulas=False,
                compare_formats=False,
                ignore_empty_cells=True,
                case_sensitive=True,
                structured_comparison=False,
            )

            comparer = ExcelComparer(options)
            result = comparer.compare_files(file1_path, file2_path)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"比较文件失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def find_last_row(cls, file_path: str, sheet_name: str, column: str | int | None = None) -> dict[str, Any]:
        """
        @intention 查找表格中最后一行有数据的位置

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            sheet_name: 工作表名称
            column: 指定列来查找最后一行（可选）
                - None: 查找整个工作表的最后一行
                - 整数: 列索引 (1-based，1=A列)
                - 字符串: 列名（先在表头行中匹配，匹配不到再按列字母解释如A/B/C）

        Returns:
            Dict: 包含 success、last_row、message 等信息

        Example:
            # 查找整个工作表的最后一行
            result = ExcelOperations.find_last_row("data.xlsx", "Sheet1")
            # 查找A列的最后一行有数据的位置
            result = ExcelOperations.find_last_row("data.xlsx", "Sheet1", "A")
            # 查找第3列的最后一行有数据的位置
            result = ExcelOperations.find_last_row("data.xlsx", "Sheet1", 3)
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始查找最后一行: {sheet_name}")

        try:
            # 自动检测双表头：列名解析需要支持中英文
            reader = ExcelReader(file_path)

            # 获取工作簿和工作表
            workbook = reader._get_workbook(read_only=True, data_only=True)
            sheet = reader._get_worksheet(workbook, sheet_name)

            last_row = 0
            search_info = ""

            if column is None:
                # 查找整个工作表的最后一行
                if sheet.max_row is not None and sheet.max_column is not None:
                    # 正常路径：有 dimension 元数据
                    for row_num in range(sheet.max_row, 0, -1):
                        has_data = False
                        for col_num in range(1, sheet.max_column + 1):
                            cell_value = sheet.cell(row=row_num, column=col_num).value
                            if cell_value is not None and str(cell_value).strip():
                                has_data = True
                                break
                        if has_data:
                            last_row = row_num
                            break
                else:
                    # 降级路径：write_only 写入的文件缺少 dimension 元数据，遍历所有行
                    # read_only 模式下迭代行是高效的（不加载全部到内存）
                    for row_num, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                        if any(cell.value is not None and str(cell.value).strip() for cell in row):
                            last_row = row_num
                search_info = "整个工作表"
            else:
                # 转换列参数为列索引（支持双表头：中文名/英文名都能匹配）
                if isinstance(column, str):
                    col_index = None

                    # 优先用 HeaderAnalyzer 解析（支持双表头中英文）
                    try:
                        info = HeaderAnalyzer.analyze(file_path, sheet_name)
                        col_index = info.resolve_column(column)
                        if col_index is not None:
                            col_index += 1  # 转为 1-based
                    except Exception:
                        pass

                    # 回退：手动搜索第1行
                    if col_index is None:
                        try:
                            max_col = sheet.max_column or 1000
                            for row in sheet.iter_rows(min_row=1, max_row=1, max_col=max_col, values_only=False):
                                for idx, cell in enumerate(row, start=1):
                                    if cell.value is not None and str(cell.value).strip() == column.strip():
                                        col_index = idx
                                        break
                                if col_index is not None:
                                    break
                        except Exception:
                            pass

                    if col_index is None:
                        # 最后回退：按列字母解释（如 'A'→1, 'B'→2）
                        try:
                            col_index = column_index_from_string(column.upper())
                        except ValueError:
                            reader.close()
                            return cls._format_error_result(f"列名 '{column}' 未在表头行中找到，也不是有效的列字母")
                elif isinstance(column, int):
                    if column < 1:
                        reader.close()
                        return cls._format_error_result("列索引必须大于等于1")
                    col_index = column
                else:
                    reader.close()
                    return cls._format_error_result("列参数必须是字符串或整数")

                # 查找指定列的最后一行有数据
                if sheet.max_row is not None:
                    for row_num in range(sheet.max_row, 0, -1):
                        cell_value = sheet.cell(row=row_num, column=col_index).value
                        if cell_value is not None and str(cell_value).strip():
                            last_row = row_num
                            break
                else:
                    # 降级路径：遍历所有行
                    for row_num, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                        try:
                            # 安全地获取单元格值
                            if hasattr(row, "__len__") and col_index <= len(row):
                                cell_value = row[col_index - 1].value
                            else:
                                cell_value = None
                        except (IndexError, AttributeError, TypeError):
                            cell_value = None

                        if cell_value is not None and str(cell_value).strip():
                            last_row = row_num

                col_letter = get_column_letter(col_index)
                search_info = f"{col_letter}列"

            reader.close()

            return {
                "success": True,
                "data": {
                    "last_row": last_row,
                    "sheet_name": sheet_name,
                    "column": column,
                    "search_scope": search_info,
                },
                "last_row": last_row,  # 兼容性字段
                "message": f"成功查找{search_info}最后一行: 第{last_row}行" if last_row > 0 else f"{search_info}没有数据",
            }

        except Exception as e:
            error_msg = f"查找最后一行失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def check_duplicate_ids(
        cls,
        file_path: str,
        sheet_name: str,
        id_column: int | str = 1,
        header_row: int = 1,
    ) -> dict[str, Any]:
        """
        检查Excel工作表中的ID重复情况

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            id_column: ID列位置 (1-based数字或列名)
            header_row: 表头行号 (1-based)

        Returns:
            Dict: 包含success、has_duplicates、duplicate_count、total_ids、unique_ids、duplicates、message
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始检查ID重复: {sheet_name}")

        try:
            # 参数验证
            if not file_path or not sheet_name:
                return {
                    "success": False,
                    "message": "文件路径和工作表名不能为空",
                    "has_duplicates": False,
                    "duplicate_count": 0,
                    "total_ids": 0,
                    "unique_ids": 0,
                    "duplicates": [],
                }

            # 自动检测双表头：确定实际表头行和数据起始行
            _effective_header_row = header_row
            _data_start_row = header_row + 1  # 默认数据从表头下一行开始
            if header_row == 1:
                try:
                    info = HeaderAnalyzer.analyze(file_path, sheet_name)
                    if info.is_dual:
                        _effective_header_row = 2  # 双表头时用第2行（英文字段名）
                        _data_start_row = info.data_start_row  # 数据从第3行开始
                except Exception:
                    pass

            # 加载工作簿
            try:
                wb = load_workbook(file_path, read_only=True)
            except FileNotFoundError:
                return {
                    "success": False,
                    "message": f"文件不存在: {file_path}",
                    "has_duplicates": False,
                    "duplicate_count": 0,
                    "total_ids": 0,
                    "unique_ids": 0,
                    "duplicates": [],
                }
            except Exception as e:
                return {
                    "success": False,
                    "message": f"无法加载文件: {str(e)}",
                    "has_duplicates": False,
                    "duplicate_count": 0,
                    "total_ids": 0,
                    "unique_ids": 0,
                    "duplicates": [],
                }

            # 检查工作表是否存在
            if sheet_name not in wb.sheetnames:
                return {
                    "success": False,
                    "message": f"工作表不存在: {sheet_name}",
                    "has_duplicates": False,
                    "duplicate_count": 0,
                    "total_ids": 0,
                    "unique_ids": 0,
                    "duplicates": [],
                }

            ws = wb[sheet_name]

            # 处理列索引（支持双表头：中文名/英文名都能匹配）
            if isinstance(id_column, str):
                col_idx = None

                # 优先用 HeaderAnalyzer 解析（支持双表头中英文）
                try:
                    info = HeaderAnalyzer.analyze(file_path, sheet_name)
                    resolved = info.resolve_column(id_column)
                    if resolved is not None:
                        col_idx = resolved + 1  # 转为 1-based
                except Exception:
                    pass

                # 回退：手动搜索 header_row 行
                if col_idx is None:
                    try:
                        for row in ws.iter_rows(
                            min_row=header_row,
                            max_row=header_row,
                            max_col=ws.max_column or 1000,
                        ):
                            for idx, cell in enumerate(row, start=1):
                                if cell.value is not None and str(cell.value).strip() == id_column.strip():
                                    col_idx = idx
                                    break
                            if col_idx is not None:
                                break
                    except Exception:
                        pass
                if col_idx is None:
                    # 回退：按列字母解释（如 'A'→1, 'B'→2）
                    try:
                        col_idx = column_index_from_string(id_column)
                    except Exception:
                        return {
                            "success": False,
                            "message": f'列名 "{id_column}" 未在表头行中找到，也不是有效的列字母',
                            "has_duplicates": False,
                            "duplicate_count": 0,
                            "total_ids": 0,
                            "unique_ids": 0,
                            "duplicates": [],
                        }
            else:
                col_idx = id_column

            # 检查表头行是否存在（streaming写入后max_row可能为None，跳过检查）
            if header_row < 1 or (ws.max_row is not None and header_row > ws.max_row):
                return {
                    "success": False,
                    "message": f"表头行不存在: {header_row}",
                    "has_duplicates": False,
                    "duplicate_count": 0,
                    "total_ids": 0,
                    "unique_ids": 0,
                    "duplicates": [],
                }

            # 检查列是否存在（streaming写入后max_column可能为None，跳过检查）
            if col_idx < 1 or (ws.max_column is not None and col_idx > ws.max_column):
                return {
                    "success": False,
                    "message": f"列不存在或索引超出范围: {col_idx}",
                    "has_duplicates": False,
                    "duplicate_count": 0,
                    "total_ids": 0,
                    "unique_ids": 0,
                    "duplicates": [],
                }

            # 单次遍历收集ID并构建行号映射
            id_to_rows = {}  # id_value -> [row_numbers]
            total_ids = 0

            for row_idx, row_data in enumerate(
                ws.iter_rows(
                    min_row=_data_start_row,
                    min_col=col_idx,
                    max_col=col_idx,
                    values_only=True,
                ),
                start=_data_start_row,
            ):
                cell_value = row_data[0] if row_data else None
                if cell_value is not None:
                    total_ids += 1
                    if cell_value not in id_to_rows:
                        id_to_rows[cell_value] = []
                    id_to_rows[cell_value].append(row_idx)

            # 直接从映射构建重复列表（无需额外遍历）
            duplicates = []
            duplicate_count = 0
            for id_value, rows in id_to_rows.items():
                if len(rows) > 1:
                    duplicate_count += 1
                    duplicates.append({"id_value": id_value, "count": len(rows), "rows": rows})

            unique_ids = len(id_to_rows)

            has_duplicates = duplicate_count > 0

            # 构建返回结果
            message = f"共检查{total_ids}个ID，发现{duplicate_count}个重复ID" if has_duplicates else f"共检查{total_ids}个ID，无重复ID"

            return {
                "success": True,
                "has_duplicates": has_duplicates,
                "duplicate_count": duplicate_count,
                "total_ids": total_ids,
                "unique_ids": unique_ids,
                "duplicates": duplicates,
                "message": message,
            }

        except Exception as e:
            error_msg = f"检查ID重复时发生错误: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def upsert_row(
        cls,
        file_path: str,
        sheet_name: str,
        key_column: str,
        key_value,
        updates: dict,
        header_row: int = 1,
        streaming: bool = True,
    ) -> dict[str, Any]:
        """
        @intention Upsert行：按键列查找，存在则更新，不存在则插入新行

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            key_column: 用于匹配的列名（支持中文/英文）
            key_value: 用于匹配的值
            updates: 要写入的列值字典（key 支持中文/英文）
            header_row: 表头所在行号（默认1=自动检测）
            streaming: 是否使用流式写入（默认True）

        Returns:
            Dict: 标准化的操作结果
        """
        try:
            # 自动检测双表头 + 中文key解析
            _effective_header_row = header_row
            _resolved_key = key_column
            _resolved_updates = updates.copy() if updates else {}

            if header_row == 1:
                try:
                    info = HeaderAnalyzer.analyze(file_path, sheet_name)
                    if info.is_dual:
                        _effective_header_row = 2
                        # 解析中文 key → 英文列名
                        _resolved_col = info.resolve_column(key_column)
                        if _resolved_col is not None:
                            # 用英文名替换（取 column_names 中对应索引的值）
                            _resolved_key = info.column_names[_resolved_col]
                        # 解析 updates 中的中文 key
                        _new_updates = {}
                        for k, v in _resolved_updates.items():
                            _rc = info.resolve_column(k)
                            if _rc is not None and k != info.column_names[_rc]:
                                _new_updates[info.column_names[_rc]] = v
                            else:
                                _new_updates[k] = v
                        _resolved_updates = _new_updates
                except Exception:
                    pass

            manager = ExcelManager(file_path)
            result = manager.upsert_row(
                sheet_name,
                _resolved_key,
                key_value,
                _resolved_updates,
                _effective_header_row,
                streaming,
            )
            return format_operation_result(result)
        except Exception as e:
            error_msg = f"Upsert行失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def batch_insert_rows(
        cls,
        file_path: str,
        sheet_name: str,
        data: list,
        header_row: int = 1,
        streaming: bool = True,
    ) -> dict[str, Any]:
        """
        @intention 批量插入多行数据到工作表末尾

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            data: 行数据列表，每行为{列名: 值}字典
            header_row: 表头所在行号（默认1=自动检测）
            streaming: 是否使用流式写入（默认True）

        Returns:
            Dict: 标准化的操作结果

        📌 双行表头支持（v1.9.3+）：
            自动检测第1行中文描述 + 第2行英文字段名的双行表头模式。
            data 字典中无论用中文还是英文列名作为 key 都能正确匹配。
            header_row 默认为 1（自动检测），显式传入时可覆盖。
        """
        try:
            # 自动检测双表头：batch_insert_rows 的 streaming_writer 已内置双表头检测
            # 这里保持 header_row=1 让它自行处理；仅做记录和缓存预热
            _effective_header_row = header_row
            if header_row == 1:
                try:
                    info = HeaderAnalyzer.analyze(file_path, sheet_name)
                    # 不修改 header_row！streaming_writer 内部会正确检测 [row1, row2]
                except Exception:
                    pass

            manager = ExcelManager(file_path)
            result = manager.batch_insert_rows(sheet_name, data, _effective_header_row, streaming)
            return format_operation_result(result)
        except Exception as e:
            error_msg = f"批量插入行失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def batch_insert_rows_at(
        cls,
        file_path: str,
        sheet_name: str,
        data: list,
        target_row: int,
        header_row: int = 1,
        streaming: bool = True,
    ) -> dict[str, Any]:
        """
        @intention 在指定行位置前批量插入多行数据

        策略：先插入空行，再逐行写入数据

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            data: 行数据列表，每行为{列名: 值}字典
            target_row: 在此行号前插入（1-based）
            header_row: 表头所在行号（默认1）
            streaming: 是否使用流式写入（默认True）

        Returns:
            Dict: 标准化的操作结果
        """
        try:
            count = len(data)
            if count == 0:
                return {"success": False, "message": "没有数据需要插入", "data": None}

            # 自动检测双表头：确定实际表头行
            _effective_header_row = header_row
            if header_row == 1:
                try:
                    info = HeaderAnalyzer.analyze(file_path, sheet_name)
                    if info.is_dual:
                        _effective_header_row = 2
                except Exception:
                    pass

            # 步骤1：在目标位置插入空行
            insert_result = cls.insert_rows(file_path, sheet_name, target_row, count, streaming)
            if not insert_result.get("success", False):
                return insert_result

            # 步骤2：获取表头列名
            reader = ExcelReader(file_path)
            range_expr = f"{sheet_name}!A{_effective_header_row}"
            header_result = reader.get_range(range_expr)
            reader.close()

            if not header_result.success or not header_result.data:
                return {
                    "success": False,
                    "message": f"无法读取表头: 第{_effective_header_row}行",
                    "data": None,
                }

            headers = header_result.data[0] if header_result.data else []
            # 从CellInfo对象中提取value（get_range可能返回CellInfo而非原始值）
            headers = [h.value if hasattr(h, "value") else h for h in headers]

            # 步骤3：构建写入数据并使用update_range一次性写入
            from excel_mcp_server_fastmcp.core.excel_writer import ExcelWriter

            writer = ExcelWriter(file_path)

            num_cols = len(headers)
            write_data = []
            for row_data in data:
                if not isinstance(row_data, dict):
                    continue
                row_vals = [row_data.get(col_name, "") for col_name in headers]
                write_data.append(row_vals)

            if write_data:
                start_cell = f"A{target_row}"
                end_col = get_column_letter(num_cols)
                end_row = target_row + len(write_data) - 1
                range_expr = f"{sheet_name}!{start_cell}:{end_col}{end_row}"
                writer.update_range(range_expr, write_data)

            return {
                "success": True,
                "message": f"在行{target_row}前插入了{count}行数据",
                "data": {
                    "inserted_count": count,
                    "start_row": target_row,
                    "end_row": target_row + count - 1,
                    "mode": "streaming" if streaming else "standard",
                },
            }
        except Exception as e:
            error_msg = f"指定位置批量插入行失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def create_chart(
        cls,
        file_path: str,
        sheet_name: str,
        chart_type: str,
        data_range: str,
        title: str = "",
        chart_name: str = "",
        position: str = "B15",
    ) -> dict[str, Any]:
        """在工作表中创建图表。

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            chart_type: 图表类型 (line/bar/column/pie/scatter/area等)
            data_range: 数据范围 (如 "Sheet1!A1:B10")
            title: 图表标题
            chart_name: 图表名称
            position: 图表位置(单元格引用, 如 "B15")
        """
        import openpyxl
        from openpyxl.chart import (
            AreaChart,
            BarChart,
            LineChart,
            PieChart,
            Reference,
            ScatterChart,
        )

        # 图表类型映射
        type_map = {
            "bar": BarChart,
            "column": BarChart,  # openpyxl中column也是BarChart(不同方向)
            "line": LineChart,
            "pie": PieChart,
            "scatter": ScatterChart,
            "area": AreaChart,
        }

        chart_type_lower = chart_type.lower().strip()
        if chart_type_lower not in type_map:
            return cls._format_error_result(f"不支持的图表类型: {chart_type}。支持的类型: {', '.join(type_map.keys())}")

        try:
            wb = openpyxl.load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return cls._format_error_result(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 解析数据范围
            if "!" in data_range:
                range_sheet, range_ref = data_range.split("!", 1)
                if range_sheet != sheet_name:
                    wb.close()
                    return cls._format_error_result(f"数据范围的工作表 '{range_sheet}' 与目标工作表 '{sheet_name}' 不匹配")
            else:
                range_ref = data_range

            # 解析数据范围边界 (兼容 openpyxl 3.1+: ws[range] 返回 tuple 而非 CellRange)
            from openpyxl.utils import range_boundaries

            try:
                min_col, min_row, max_col, max_row = range_boundaries(range_ref)
                if max_col < min_col or max_row < min_row:
                    wb.close()
                    return cls._format_error_result(f"无效的数据范围: {data_range}")
            except Exception:
                wb.close()
                return cls._format_error_result(f"无效的数据范围: {data_range}")

            # 验证数据区域有内容 (至少2行2列: 表头+数据, 类别+数值)
            if max_row <= min_row or max_col <= min_col:
                wb.close()
                return cls._format_error_result("数据范围为空或只有一行/一列，无法创建图表")

            # 创建图表
            ChartClass = type_map[chart_type_lower]
            chart = ChartClass()

            # 设置标题
            if title:
                chart.title = title
            elif chart_name:
                chart.title = chart_name

            # 设置数据源（第一列作为类别，后续列作为数据系列）
            categories = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
            chart.set_categories(categories)

            # 添加数据系列（跳过第一列作为标签）
            for col_idx in range(min_col + 1, max_col + 1):
                values = Reference(ws, min_col=col_idx, min_row=min_row, max_row=max_row)
                chart.add_data(values, titles_from_data=True)

            # 设置图表样式
            chart.style = 10
            chart.shape = 4

            # column 类型改为垂直方向
            if chart_type_lower == "column":
                chart.type = "col"

            # 解析位置
            try:
                pos_cell = ws[position]
                chart.anchor = position
            except Exception:
                # 如果位置无效，使用默认位置
                chart.anchor = "B15"

            # 添加图表到工作表
            ws.add_chart(chart, chart_name or None)

            # 保存文件
            wb.save(file_path)
            wb.close()

            return {
                "success": True,
                "message": "图表创建成功",
                "data": {
                    "chart_type": chart_type,
                    "chart_title": title or chart_name or "",
                    "data_range": data_range,
                    "position": position,
                    "sheet_name": sheet_name,
                },
            }
        except Exception as e:
            error_msg = f"图表创建失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)
