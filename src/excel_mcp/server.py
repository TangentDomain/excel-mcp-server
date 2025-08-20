#!/usr/bin/env python3
"""
Excel MCP Server - 基于 FastMCP 和 openpyxl 实现

主要功能：
1. 正则搜索：在Excel文件中搜索符合正则表达式的单元格
2. 范围获取：读取指定范围的Excel数据
3. 范围修改：修改指定范围的Excel数据

技术栈：
- FastMCP: 用于MCP服务器框架
- openpyxl: 用于Excel文件操作
"""

import os
import re
import logging
from typing import Optional, List, Dict, Any, Union
from pathlib import Path

try:
    from mcp.server.fastmcp import FastMCP
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter, range_boundaries, column_index_from_string
    from openpyxl.cell import Cell
except ImportError as e:
    print(f"Error: 缺少必要的依赖包: {e}")
    print("请运行: pip install fastmcp openpyxl")
    exit(1)

# ==================== 配置和初始化 ====================
logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger(__name__)

# 创建FastMCP服务器实例
mcp = FastMCP("excel-mcp-server")

class ExcelHandler:
    """Excel文件操作处理器"""

    @staticmethod
    def validate_file_path(file_path: str) -> str:
        """验证并规范化文件路径"""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")
        if path.suffix.lower() not in ['.xlsx', '.xlsm', '.xls']:
            raise ValueError(f"不支持的文件格式: {path.suffix}")
        return str(path.absolute())

    @staticmethod
    def parse_range_expression(range_expr: str) -> Dict[str, Any]:
        """解析范围表达式 (如 'Sheet1!A1:C10' 或 'A1:C10' 或 '1:1' 或 'A:A')"""
        if '!' in range_expr:
            sheet_name, cell_range = range_expr.split('!', 1)
        else:
            sheet_name = None
            cell_range = range_expr

        # 检测是否为行列访问模式
        range_type = 'cell_range'  # 默认为单元格范围

        # 检测整行模式 (如 "1:1", "3:5")
        if re.match(r'^\d+:\d+$', cell_range):
            range_type = 'row_range'
        # 检测整列模式 (如 "A:A", "B:D")
        elif re.match(r'^[A-Z]+:[A-Z]+$', cell_range):
            range_type = 'column_range'
        # 检测单行模式 (如 "1", 只读取第1行)
        elif re.match(r'^\d+$', cell_range):
            range_type = 'single_row'
            cell_range = f"{cell_range}:{cell_range}"
        # 检测单列模式 (如 "A", 只读取A列)
        elif re.match(r'^[A-Z]+$', cell_range):
            range_type = 'single_column'
            cell_range = f"{cell_range}:{cell_range}"

        return {
            'sheet_name': sheet_name,
            'cell_range': cell_range,
            'range_type': range_type
        }

# ==================== MCP 工具定义 ====================

@mcp.tool()
def excel_list_sheets(file_path: str) -> Dict[str, Any]:
    """
    获取Excel文件中所有工作表的名称列表

    Args:
        file_path: Excel文件路径

    Returns:
        包含所有工作表信息的字典
    """
    try:
        # 验证文件路径
        validated_path = ExcelHandler.validate_file_path(file_path)

        # 加载Excel文件
        workbook = load_workbook(validated_path, read_only=True)

        # 获取所有工作表信息
        sheets_info = []
        for i, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]

            # 获取工作表基本信息
            sheet_info = {
                'index': i,
                'name': sheet_name,
                'is_active': sheet == workbook.active,
                'max_row': sheet.max_row,
                'max_column': sheet.max_column,
                'max_column_letter': get_column_letter(sheet.max_column)
            }
            sheets_info.append(sheet_info)

        return {
            'success': True,
            'file_path': validated_path,
            'total_sheets': len(sheets_info),
            'active_sheet': workbook.active.title,
            'sheets': sheets_info
        }

    except Exception as e:
        logger.error(f"Excel工作表列表获取错误: {e}")
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path
        }

@mcp.tool()
def excel_regex_search(
    file_path: str,
    pattern: str,
    flags: str = "",
    search_values: bool = True,
    search_formulas: bool = False
) -> Dict[str, Any]:
    """
    在Excel文件中使用正则表达式搜索单元格内容

    Args:
        file_path: Excel文件路径
        pattern: 正则表达式模式
        flags: 正则表达式标志 (i=忽略大小写, m=多行, s=点匹配换行)
        search_values: 是否搜索单元格的显示值
        search_formulas: 是否搜索单元格的公式

    Returns:
        包含搜索结果的字典，包含匹配的单元格信息
    """
    try:
        # 验证文件路径
        validated_path = ExcelHandler.validate_file_path(file_path)

        # 构建正则表达式标志
        regex_flags = 0
        if 'i' in flags.lower():
            regex_flags |= re.IGNORECASE
        if 'm' in flags.lower():
            regex_flags |= re.MULTILINE
        if 's' in flags.lower():
            regex_flags |= re.DOTALL

        # 编译正则表达式
        regex = re.compile(pattern, regex_flags)

        # 加载Excel文件
        workbook = load_workbook(validated_path, data_only=not search_formulas)
        results = []

        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # 遍历所有单元格
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue

                    # 搜索单元格值
                    if search_values:
                        cell_value = str(cell.value)
                        matches = regex.finditer(cell_value)
                        for match in matches:
                            results.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'value': cell_value,
                                'match': match.group(),
                                'match_start': match.start(),
                                'match_end': match.end(),
                                'type': 'value'
                            })

                    # 搜索单元格公式
                    if search_formulas and hasattr(cell, 'formula') and cell.formula:
                        formula = str(cell.formula)
                        matches = regex.finditer(formula)
                        for match in matches:
                            results.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'formula': formula,
                                'match': match.group(),
                                'match_start': match.start(),
                                'match_end': match.end(),
                                'type': 'formula'
                            })

        return {
            'success': True,
            'file_path': validated_path,
            'pattern': pattern,
            'total_matches': len(results),
            'matches': results
        }

    except Exception as e:
        logger.error(f"Excel正则搜索错误: {e}")
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path,
            'pattern': pattern
        }

@mcp.tool()
def excel_get_range(
    file_path: str,
    range_expression: str,
    include_formatting: bool = False
) -> Dict[str, Any]:
    """
    获取Excel文件中指定范围的数据
    支持多种范围模式：
    - 单元格范围: 'A1:C10' 或 'Sheet1!A1:C10'
    - 整行访问: '1:1' (第1行), '3:5' (第3-5行), '2' (仅第2行)
    - 整列访问: 'A:A' (A列), 'B:D' (B-D列), 'C' (仅C列)

    Args:
        file_path: Excel文件路径
        range_expression: 范围表达式
        include_formatting: 是否包含格式信息

    Returns:
        包含范围数据的字典
    """
    try:
        # 验证文件路径
        validated_path = ExcelHandler.validate_file_path(file_path)

        # 解析范围表达式
        range_info = ExcelHandler.parse_range_expression(range_expression)

        # 加载Excel文件
        workbook = load_workbook(validated_path, data_only=True)

        # 确定工作表
        if range_info['sheet_name']:
            if range_info['sheet_name'] not in workbook.sheetnames:
                raise ValueError(f"工作表不存在: {range_info['sheet_name']}")
            sheet = workbook[range_info['sheet_name']]
        else:
            sheet = workbook.active

        # 根据范围类型处理不同的访问模式
        range_type = range_info.get('range_type', 'cell_range')

        if range_type in ['row_range', 'single_row']:
            # 行访问模式：读取整行数据
            row_parts = range_info['cell_range'].split(':')
            start_row = int(row_parts[0])
            end_row = int(row_parts[1])

            # 获取工作表的实际数据范围
            max_col = sheet.max_column

            data = []
            for row_idx in range(start_row, end_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell_info = {
                        'coordinate': cell.coordinate,
                        'value': cell.value
                    }

                    if include_formatting:
                        cell_info.update({
                            'data_type': cell.data_type,
                            'number_format': cell.number_format,
                            'font': str(cell.font) if cell.font else None,
                            'fill': str(cell.fill) if cell.fill else None
                        })

                    row_data.append(cell_info)
                data.append(row_data)

            dimensions = {
                'rows': end_row - start_row + 1,
                'columns': max_col,
                'start_row': start_row,
                'start_column': 1
            }

        elif range_type in ['column_range', 'single_column']:
            # 列访问模式：读取整列数据
            col_parts = range_info['cell_range'].split(':')
            start_col_letter = col_parts[0]
            end_col_letter = col_parts[1]

            # 转换列字母为数字
            start_col = column_index_from_string(start_col_letter)
            end_col = column_index_from_string(end_col_letter)

            # 获取工作表的实际数据范围
            max_row = sheet.max_row

            data = []
            for row_idx in range(1, max_row + 1):
                row_data = []
                for col_idx in range(start_col, end_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell_info = {
                        'coordinate': cell.coordinate,
                        'value': cell.value
                    }

                    if include_formatting:
                        cell_info.update({
                            'data_type': cell.data_type,
                            'number_format': cell.number_format,
                            'font': str(cell.font) if cell.font else None,
                            'fill': str(cell.fill) if cell.fill else None
                        })

                    row_data.append(cell_info)
                data.append(row_data)

            dimensions = {
                'rows': max_row,
                'columns': end_col - start_col + 1,
                'start_row': 1,
                'start_column': start_col
            }

        else:
            # 标准单元格范围模式
            # 获取范围边界
            min_col, min_row, max_col, max_row = range_boundaries(range_info['cell_range'])

            # 读取数据
            data = []
            for row_idx in range(min_row, max_row + 1):
                row_data = []
                for col_idx in range(min_col, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell_info = {
                        'coordinate': cell.coordinate,
                        'value': cell.value
                    }

                    if include_formatting:
                        cell_info.update({
                            'data_type': cell.data_type,
                            'number_format': cell.number_format,
                            'font': str(cell.font) if cell.font else None,
                            'fill': str(cell.fill) if cell.fill else None
                        })

                    row_data.append(cell_info)
                data.append(row_data)

            dimensions = {
                'rows': max_row - min_row + 1,
                'columns': max_col - min_col + 1,
                'start_row': min_row,
                'start_column': min_col
            }

        return {
            'success': True,
            'file_path': validated_path,
            'range': range_expression,
            'range_type': range_type,
            'sheet_name': sheet.title,
            'dimensions': dimensions,
            'data': data
        }

    except Exception as e:
        logger.error(f"Excel范围获取错误: {e}")
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path,
            'range': range_expression
        }

@mcp.tool()
def excel_update_range(
    file_path: str,
    range_expression: str,
    data: List[List[Any]],
    preserve_formulas: bool = True
) -> Dict[str, Any]:
    """
    修改Excel文件中指定范围的数据

    Args:
        file_path: Excel文件路径
        range_expression: 范围表达式 (如 'A1:C10' 或 'Sheet1!A1:C10')
        data: 要写入的二维数据数组
        preserve_formulas: 是否保留现有的公式

    Returns:
        修改操作的结果信息
    """
    try:
        # 验证文件路径
        validated_path = ExcelHandler.validate_file_path(file_path)

        # 解析范围表达式
        range_info = ExcelHandler.parse_range_expression(range_expression)

        # 加载Excel文件
        workbook = load_workbook(validated_path)

        # 确定工作表
        if range_info['sheet_name']:
            if range_info['sheet_name'] not in workbook.sheetnames:
                raise ValueError(f"工作表不存在: {range_info['sheet_name']}")
            sheet = workbook[range_info['sheet_name']]
        else:
            sheet = workbook.active

        # 获取范围边界
        min_col, min_row, max_col, max_row = range_boundaries(range_info['cell_range'])

        # 验证数据维度
        range_rows = max_row - min_row + 1
        range_cols = max_col - min_col + 1

        if len(data) > range_rows:
            raise ValueError(f"数据行数({len(data)})超过范围行数({range_rows})")

        # 写入数据
        modified_cells = []
        for row_offset, row_data in enumerate(data):
            if len(row_data) > range_cols:
                raise ValueError(f"第{row_offset + 1}行数据列数({len(row_data)})超过范围列数({range_cols})")

            for col_offset, value in enumerate(row_data):
                row_idx = min_row + row_offset
                col_idx = min_col + col_offset
                cell = sheet.cell(row=row_idx, column=col_idx)

                # 保留公式检查
                if preserve_formulas and cell.data_type == 'f':
                    # 如果是公式单元格且要求保留公式，跳过修改
                    continue

                old_value = cell.value
                cell.value = value

                modified_cells.append({
                    'coordinate': cell.coordinate,
                    'old_value': old_value,
                    'new_value': value
                })

        # 保存文件
        workbook.save(validated_path)

        return {
            'success': True,
            'file_path': validated_path,
            'range': range_expression,
            'sheet_name': sheet.title,
            'modified_cells_count': len(modified_cells),
            'modified_cells': modified_cells
        }

    except Exception as e:
        logger.error(f"Excel范围修改错误: {e}")
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path,
            'range': range_expression
        }

@mcp.tool()
def excel_insert_rows(
    file_path: str,
    sheet_name: Optional[str] = None,
    row_index: int = 1,
    count: int = 1
) -> Dict[str, Any]:
    """
    在Excel文件中插入空白行

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称 (可选，默认使用活动工作表)
        row_index: 插入行的位置（1-based），新行将插入到此位置之前
        count: 要插入的行数

    Returns:
        插入操作的结果信息
    """
    try:
        # 验证文件路径
        validated_path = ExcelHandler.validate_file_path(file_path)

        # 验证参数
        if row_index < 1:
            raise ValueError("行索引必须大于等于1")
        if count < 1:
            raise ValueError("插入行数必须大于等于1")
        if count > 1000:
            raise ValueError("一次最多插入1000行")

        # 加载Excel文件
        workbook = load_workbook(validated_path)

        # 确定工作表
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"工作表不存在: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        # 记录操作前的信息
        original_max_row = sheet.max_row

        # 插入行
        sheet.insert_rows(row_index, count)

        # 保存文件
        workbook.save(validated_path)

        return {
            'success': True,
            'file_path': validated_path,
            'sheet_name': sheet.title,
            'inserted_at_row': row_index,
            'inserted_count': count,
            'original_max_row': original_max_row,
            'new_max_row': sheet.max_row,
            'message': f"成功在第{row_index}行前插入{count}行"
        }

    except Exception as e:
        logger.error(f"Excel行插入错误: {e}")
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path,
            'sheet_name': sheet_name,
            'row_index': row_index,
            'count': count
        }

@mcp.tool()
def excel_insert_columns(
    file_path: str,
    sheet_name: Optional[str] = None,
    column_index: int = 1,
    count: int = 1
) -> Dict[str, Any]:
    """
    在Excel文件中插入空白列

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称 (可选，默认使用活动工作表)
        column_index: 插入列的位置（1-based），新列将插入到此位置之前
        count: 要插入的列数

    Returns:
        插入操作的结果信息
    """
    try:
        # 验证文件路径
        validated_path = ExcelHandler.validate_file_path(file_path)

        # 验证参数
        if column_index < 1:
            raise ValueError("列索引必须大于等于1")
        if count < 1:
            raise ValueError("插入列数必须大于等于1")
        if count > 100:
            raise ValueError("一次最多插入100列")

        # 加载Excel文件
        workbook = load_workbook(validated_path)

        # 确定工作表
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"工作表不存在: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        # 记录操作前的信息
        original_max_column = sheet.max_column
        original_max_column_letter = get_column_letter(original_max_column)

        # 插入列
        sheet.insert_cols(column_index, count)

        # 保存文件
        workbook.save(validated_path)

        new_max_column_letter = get_column_letter(sheet.max_column)

        return {
            'success': True,
            'file_path': validated_path,
            'sheet_name': sheet.title,
            'inserted_at_column': column_index,
            'inserted_at_column_letter': get_column_letter(column_index),
            'inserted_count': count,
            'original_max_column': original_max_column,
            'original_max_column_letter': original_max_column_letter,
            'new_max_column': sheet.max_column,
            'new_max_column_letter': new_max_column_letter,
            'message': f"成功在第{get_column_letter(column_index)}列前插入{count}列"
        }

    except Exception as e:
        logger.error(f"Excel列插入错误: {e}")
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path,
            'sheet_name': sheet_name,
            'column_index': column_index,
            'count': count
        }

@mcp.tool()
def excel_create_file(
    file_path: str,
    sheet_names: Optional[List[str]] = None
) -> Dict[str, Any]:
    """
    创建新的Excel文件

    Args:
        file_path: 要创建的Excel文件路径
        sheet_names: 工作表名称列表 (可选，默认创建"Sheet1")

    Returns:
        创建操作的结果信息
    """
    try:
        # 验证路径合法性
        path = Path(file_path)
        if path.exists():
            raise FileExistsError(f"文件已存在: {file_path}")
        if path.suffix.lower() not in ['.xlsx', '.xlsm']:
            raise ValueError(f"不支持的文件格式: {path.suffix}，请使用 .xlsx 或 .xlsm")

        # 创建工作簿
        workbook = Workbook()

        # 删除默认工作表（如果需要自定义工作表）
        if sheet_names:
            # 删除默认的工作表
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            # 创建指定的工作表
            created_sheets = []
            for i, sheet_name in enumerate(sheet_names):
                if not sheet_name or not sheet_name.strip():
                    raise ValueError(f"工作表名称不能为空: 索引 {i}")
                sheet = workbook.create_sheet(title=sheet_name.strip())
                created_sheets.append({
                    'index': i,
                    'name': sheet.title,
                    'is_active': i == 0
                })

            # 设置第一个工作表为活动工作表
            if created_sheets:
                workbook.active = workbook[created_sheets[0]['name']]
        else:
            # 使用默认工作表
            created_sheets = [{
                'index': 0,
                'name': 'Sheet1',
                'is_active': True
            }]

        # 确保目录存在
        path.parent.mkdir(parents=True, exist_ok=True)

        # 保存文件
        workbook.save(file_path)

        return {
            'success': True,
            'file_path': str(path.absolute()),
            'total_sheets': len(created_sheets),
            'sheets': created_sheets,
            'message': f"成功创建Excel文件，包含{len(created_sheets)}个工作表"
        }

    except Exception as e:
        logger.error(f"Excel文件创建错误: {e}")
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path,
            'sheet_names': sheet_names
        }

@mcp.tool()
def excel_create_sheet(
    file_path: str,
    sheet_name: str,
    index: Optional[int] = None
) -> Dict[str, Any]:
    """
    在Excel文件中创建新工作表

    Args:
        file_path: Excel文件路径
        sheet_name: 新工作表名称
        index: 插入位置索引 (可选，默认在最后添加)

    Returns:
        创建操作的结果信息
    """
    try:
        # 验证文件路径
        validated_path = ExcelHandler.validate_file_path(file_path)

        # 验证工作表名称
        if not sheet_name or not sheet_name.strip():
            raise ValueError("工作表名称不能为空")

        sheet_name = sheet_name.strip()

        # 加载Excel文件
        workbook = load_workbook(validated_path)

        # 检查工作表名称是否已存在
        if sheet_name in workbook.sheetnames:
            raise ValueError(f"工作表名称已存在: {sheet_name}")

        # 验证索引
        total_sheets = len(workbook.sheetnames)
        if index is not None:
            if index < 0 or index > total_sheets:
                raise ValueError(f"索引超出范围: {index}，应在 0-{total_sheets} 之间")

        # 创建新工作表
        new_sheet = workbook.create_sheet(title=sheet_name, index=index)

        # 保存文件
        workbook.save(validated_path)

        # 获取新工作表信息
        sheet_info = {
            'name': new_sheet.title,
            'index': workbook.sheetnames.index(sheet_name),
            'is_active': new_sheet == workbook.active
        }

        return {
            'success': True,
            'file_path': validated_path,
            'sheet_info': sheet_info,
            'total_sheets': len(workbook.sheetnames),
            'all_sheets': workbook.sheetnames,
            'message': f"成功创建工作表: {sheet_name}"
        }

    except Exception as e:
        logger.error(f"Excel工作表创建错误: {e}")
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path,
            'sheet_name': sheet_name,
            'index': index
        }

@mcp.tool()
def excel_delete_sheet(
    file_path: str,
    sheet_name: str
) -> Dict[str, Any]:
    """
    删除Excel文件中的工作表

    Args:
        file_path: Excel文件路径
        sheet_name: 要删除的工作表名称

    Returns:
        删除操作的结果信息
    """
    try:
        # 验证文件路径
        validated_path = ExcelHandler.validate_file_path(file_path)

        # 验证工作表名称
        if not sheet_name or not sheet_name.strip():
            raise ValueError("工作表名称不能为空")

        sheet_name = sheet_name.strip()

        # 加载Excel文件
        workbook = load_workbook(validated_path)

        # 检查工作表是否存在
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表不存在: {sheet_name}")

        # 检查是否为最后一个工作表
        if len(workbook.sheetnames) <= 1:
            raise ValueError("无法删除最后一个工作表，Excel文件至少需要一个工作表")

        # 记录删除前的信息
        deleted_sheet_index = workbook.sheetnames.index(sheet_name)
        was_active = workbook[sheet_name] == workbook.active

        # 删除工作表
        workbook.remove(workbook[sheet_name])

        # 如果删除的是活动工作表，设置新的活动工作表
        if was_active:
            # 选择下一个工作表，如果是最后一个则选择前一个
            if deleted_sheet_index < len(workbook.sheetnames):
                workbook.active = deleted_sheet_index
            else:
                workbook.active = deleted_sheet_index - 1

        # 保存文件
        workbook.save(validated_path)

        return {
            'success': True,
            'file_path': validated_path,
            'deleted_sheet': sheet_name,
            'deleted_index': deleted_sheet_index,
            'was_active': was_active,
            'new_active_sheet': workbook.active.title,
            'remaining_sheets': workbook.sheetnames,
            'total_sheets': len(workbook.sheetnames),
            'message': f"成功删除工作表: {sheet_name}"
        }

    except Exception as e:
        logger.error(f"Excel工作表删除错误: {e}")
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path,
            'sheet_name': sheet_name
        }

@mcp.tool()
def excel_rename_sheet(
    file_path: str,
    old_name: str,
    new_name: str
) -> Dict[str, Any]:
    """
    重命名Excel文件中的工作表

    Args:
        file_path: Excel文件路径
        old_name: 原工作表名称
        new_name: 新工作表名称

    Returns:
        重命名操作的结果信息
    """
    try:
        # 验证文件路径
        validated_path = ExcelHandler.validate_file_path(file_path)

        # 验证工作表名称
        if not old_name or not old_name.strip():
            raise ValueError("原工作表名称不能为空")
        if not new_name or not new_name.strip():
            raise ValueError("新工作表名称不能为空")

        old_name = old_name.strip()
        new_name = new_name.strip()

        if old_name == new_name:
            raise ValueError("新名称与原名称相同，无需重命名")

        # 加载Excel文件
        workbook = load_workbook(validated_path)

        # 检查原工作表是否存在
        if old_name not in workbook.sheetnames:
            raise ValueError(f"原工作表不存在: {old_name}")

        # 检查新名称是否已存在
        if new_name in workbook.sheetnames:
            raise ValueError(f"新工作表名称已存在: {new_name}")

        # 获取工作表
        sheet = workbook[old_name]
        old_index = workbook.sheetnames.index(old_name)
        was_active = sheet == workbook.active

        # 重命名工作表
        sheet.title = new_name

        # 保存文件
        workbook.save(validated_path)

        return {
            'success': True,
            'file_path': validated_path,
            'old_name': old_name,
            'new_name': new_name,
            'sheet_index': old_index,
            'is_active': was_active,
            'all_sheets': workbook.sheetnames,
            'message': f"成功将工作表 '{old_name}' 重命名为 '{new_name}'"
        }

    except Exception as e:
        logger.error(f"Excel工作表重命名错误: {e}")
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path,
            'old_name': old_name,
            'new_name': new_name
        }

@mcp.tool()
def excel_delete_rows(
    file_path: str,
    sheet_name: Optional[str] = None,
    start_row: int = 1,
    count: int = 1
) -> Dict[str, Any]:
    """
    在Excel文件中删除行

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称 (可选，默认使用活动工作表)
        start_row: 开始删除的行号（1-based）
        count: 要删除的行数

    Returns:
        删除操作的结果信息
    """
    try:
        # 验证文件路径
        validated_path = ExcelHandler.validate_file_path(file_path)

        # 验证参数
        if start_row < 1:
            raise ValueError("起始行号必须大于等于1")
        if count < 1:
            raise ValueError("删除行数必须大于等于1")
        if count > 1000:
            raise ValueError("一次最多删除1000行")

        # 加载Excel文件
        workbook = load_workbook(validated_path)

        # 确定工作表
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"工作表不存在: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        # 记录操作前的信息
        original_max_row = sheet.max_row

        # 验证删除范围
        if start_row > original_max_row:
            raise ValueError(f"起始行号({start_row})超过工作表最大行数({original_max_row})")

        # 计算实际删除的行数（不超过剩余行数）
        actual_count = min(count, original_max_row - start_row + 1)

        # 删除行
        sheet.delete_rows(start_row, actual_count)

        # 保存文件
        workbook.save(validated_path)

        return {
            'success': True,
            'file_path': validated_path,
            'sheet_name': sheet.title,
            'deleted_start_row': start_row,
            'requested_count': count,
            'actual_deleted_count': actual_count,
            'original_max_row': original_max_row,
            'new_max_row': sheet.max_row,
            'message': f"成功从第{start_row}行开始删除{actual_count}行"
        }

    except Exception as e:
        logger.error(f"Excel行删除错误: {e}")
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path,
            'sheet_name': sheet_name,
            'start_row': start_row,
            'count': count
        }

@mcp.tool()
def excel_delete_columns(
    file_path: str,
    sheet_name: Optional[str] = None,
    start_column: int = 1,
    count: int = 1
) -> Dict[str, Any]:
    """
    在Excel文件中删除列

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称 (可选，默认使用活动工作表)
        start_column: 开始删除的列号（1-based）
        count: 要删除的列数

    Returns:
        删除操作的结果信息
    """
    try:
        # 验证文件路径
        validated_path = ExcelHandler.validate_file_path(file_path)

        # 验证参数
        if start_column < 1:
            raise ValueError("起始列号必须大于等于1")
        if count < 1:
            raise ValueError("删除列数必须大于等于1")
        if count > 100:
            raise ValueError("一次最多删除100列")

        # 加载Excel文件
        workbook = load_workbook(validated_path)

        # 确定工作表
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"工作表不存在: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        # 记录操作前的信息
        original_max_column = sheet.max_column
        original_max_column_letter = get_column_letter(original_max_column)

        # 验证删除范围
        if start_column > original_max_column:
            raise ValueError(f"起始列号({start_column})超过工作表最大列数({original_max_column})")

        # 计算实际删除的列数（不超过剩余列数）
        actual_count = min(count, original_max_column - start_column + 1)

        # 删除列
        sheet.delete_cols(start_column, actual_count)

        # 保存文件
        workbook.save(validated_path)

        new_max_column_letter = get_column_letter(sheet.max_column) if sheet.max_column > 0 else "A"

        return {
            'success': True,
            'file_path': validated_path,
            'sheet_name': sheet.title,
            'deleted_start_column': start_column,
            'deleted_start_column_letter': get_column_letter(start_column),
            'requested_count': count,
            'actual_deleted_count': actual_count,
            'original_max_column': original_max_column,
            'original_max_column_letter': original_max_column_letter,
            'new_max_column': sheet.max_column,
            'new_max_column_letter': new_max_column_letter,
            'message': f"成功从第{get_column_letter(start_column)}列开始删除{actual_count}列"
        }

    except Exception as e:
        logger.error(f"Excel列删除错误: {e}")
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path,
            'sheet_name': sheet_name,
            'start_column': start_column,
            'count': count
        }

# ==================== 主程序 ====================
if __name__ == "__main__":
    # 运行FastMCP服务器
    mcp.run()
