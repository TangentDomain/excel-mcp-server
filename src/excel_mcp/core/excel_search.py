"""
Excel MCP Server - Excel搜索模块

提供Excel文件搜索功能
"""

import re
import logging
from typing import List, Optional
from openpyxl import load_workbook

from ..models.types import SearchMatch, MatchType, OperationResult
from ..utils.validators import ExcelValidator

logger = logging.getLogger(__name__)


class ExcelSearcher:
    """Excel文件搜索器"""

    def __init__(self, file_path: str):
        """
        初始化Excel搜索器

        Args:
            file_path: Excel文件路径
        """
        self.file_path = ExcelValidator.validate_file_path(file_path)

    def regex_search(
        self,
        pattern: str,
        flags: str = "",
        search_values: bool = True,
        search_formulas: bool = False
    ) -> OperationResult:
        """
        在Excel文件中使用正则表达式搜索单元格内容

        Args:
            pattern: 正则表达式模式
            flags: 正则表达式标志 (i=忽略大小写, m=多行, s=点匹配换行)
            search_values: 是否搜索单元格的显示值
            search_formulas: 是否搜索单元格的公式

        Returns:
            OperationResult: 包含搜索结果的结果对象
        """
        try:
            # 构建正则表达式标志
            regex_flags = self._build_regex_flags(flags)

            # 编译正则表达式
            try:
                regex = re.compile(pattern, regex_flags)
            except re.error as e:
                raise ValueError(f"无效的正则表达式: {e}")

            # 加载Excel文件
            workbook = load_workbook(self.file_path, data_only=not search_formulas)

            # 执行搜索
            matches = self._search_workbook(
                workbook, regex, search_values, search_formulas
            )

            return OperationResult(
                success=True,
                data=matches,
                metadata={
                    'file_path': self.file_path,
                    'pattern': pattern,
                    'total_matches': len(matches),
                    'search_values': search_values,
                    'search_formulas': search_formulas
                }
            )

        except Exception as e:
            logger.error(f"正则搜索失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def _build_regex_flags(self, flags: str) -> int:
        """构建正则表达式标志"""
        regex_flags = 0
        if 'i' in flags.lower():
            regex_flags |= re.IGNORECASE
        if 'm' in flags.lower():
            regex_flags |= re.MULTILINE
        if 's' in flags.lower():
            regex_flags |= re.DOTALL
        return regex_flags

    def _search_workbook(
        self,
        workbook,
        regex: re.Pattern,
        search_values: bool,
        search_formulas: bool
    ) -> List[SearchMatch]:
        """在工作簿中搜索"""
        matches = []

        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # 遍历所有单元格
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue

                    # 搜索单元格值
                    if search_values:
                        cell_value = str(cell.value)
                        for match in regex.finditer(cell_value):
                            matches.append(SearchMatch(
                                sheet=sheet_name,
                                cell=cell.coordinate,
                                value=cell_value,
                                match=match.group(),
                                match_start=match.start(),
                                match_end=match.end(),
                                match_type=MatchType.VALUE
                            ))

                    # 搜索单元格公式
                    if search_formulas and hasattr(cell, 'formula') and cell.formula:
                        formula = str(cell.formula)
                        for match in regex.finditer(formula):
                            matches.append(SearchMatch(
                                sheet=sheet_name,
                                cell=cell.coordinate,
                                formula=formula,
                                match=match.group(),
                                match_start=match.start(),
                                match_end=match.end(),
                                match_type=MatchType.FORMULA
                            ))

        return matches
