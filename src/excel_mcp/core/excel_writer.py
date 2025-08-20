"""
Excel MCP Server - Excel写入模块

提供Excel文件写入和修改功能
"""

import logging
from typing import List, Any, Optional
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from ..models.types import RangeInfo, ModifiedCell, OperationResult, RangeType
from ..utils.validators import ExcelValidator
from ..utils.parsers import RangeParser
from ..utils.exceptions import SheetNotFoundError, DataValidationError

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Excel文件写入器"""

    def __init__(self, file_path: str):
        """
        初始化Excel写入器

        Args:
            file_path: Excel文件路径
        """
        self.file_path = ExcelValidator.validate_file_path(file_path)

    def update_range(
        self,
        range_expression: str,
        data: List[List[Any]],
        preserve_formulas: bool = True
    ) -> OperationResult:
        """
        修改Excel文件中指定范围的数据

        Args:
            range_expression: 范围表达式
            data: 要写入的二维数据数组
            preserve_formulas: 是否保留现有的公式

        Returns:
            OperationResult: 修改操作的结果
        """
        try:
            # 解析范围表达式
            range_info = RangeParser.parse_range_expression(range_expression)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, range_info.sheet_name)

            # 获取范围边界
            min_col, min_row, max_col, max_row = range_boundaries(range_info.cell_range)

            # 验证数据维度
            range_rows = max_row - min_row + 1
            range_cols = max_col - min_col + 1
            ExcelValidator.validate_range_data(data, range_rows, range_cols)

            # 写入数据
            modified_cells = self._write_data(
                sheet, data, min_row, min_col, preserve_formulas
            )

            # 保存文件
            workbook.save(self.file_path)

            return OperationResult(
                success=True,
                data=modified_cells,
                metadata={
                    'file_path': self.file_path,
                    'range': range_expression,
                    'sheet_name': sheet.title,
                    'modified_cells_count': len(modified_cells)
                }
            )

        except Exception as e:
            logger.error(f"更新范围数据失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def insert_rows(
        self,
        sheet_name: Optional[str] = None,
        row_index: int = 1,
        count: int = 1
    ) -> OperationResult:
        """
        在Excel文件中插入空白行

        Args:
            sheet_name: 工作表名称
            row_index: 插入行的位置（1-based）
            count: 要插入的行数

        Returns:
            OperationResult: 插入操作的结果
        """
        try:
            # 验证参数
            ExcelValidator.validate_row_operations(row_index, count)
            ExcelValidator.validate_sheet_name(sheet_name)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, sheet_name)

            # 记录操作前的信息
            original_max_row = sheet.max_row

            # 插入行
            sheet.insert_rows(row_index, count)

            # 保存文件
            workbook.save(self.file_path)

            return OperationResult(
                success=True,
                message=f"成功在第{row_index}行前插入{count}行",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet.title,
                    'inserted_at_row': row_index,
                    'inserted_count': count,
                    'original_max_row': original_max_row,
                    'new_max_row': sheet.max_row
                }
            )

        except Exception as e:
            logger.error(f"插入行失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def insert_columns(
        self,
        sheet_name: Optional[str] = None,
        column_index: int = 1,
        count: int = 1
    ) -> OperationResult:
        """
        在Excel文件中插入空白列

        Args:
            sheet_name: 工作表名称
            column_index: 插入列的位置（1-based）
            count: 要插入的列数

        Returns:
            OperationResult: 插入操作的结果
        """
        try:
            # 验证参数
            ExcelValidator.validate_column_operations(column_index, count)
            ExcelValidator.validate_sheet_name(sheet_name)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, sheet_name)

            # 记录操作前的信息
            original_max_column = sheet.max_column

            # 插入列
            sheet.insert_cols(column_index, count)

            # 保存文件
            workbook.save(self.file_path)

            return OperationResult(
                success=True,
                message=f"成功插入{count}列",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet.title,
                    'inserted_at_column': column_index,
                    'inserted_count': count,
                    'original_max_column': original_max_column,
                    'new_max_column': sheet.max_column
                }
            )

        except Exception as e:
            logger.error(f"插入列失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def delete_rows(
        self,
        sheet_name: Optional[str] = None,
        start_row: int = 1,
        count: int = 1
    ) -> OperationResult:
        """
        删除Excel文件中的行

        Args:
            sheet_name: 工作表名称
            start_row: 开始删除的行号
            count: 要删除的行数

        Returns:
            OperationResult: 删除操作的结果
        """
        try:
            # 验证参数
            ExcelValidator.validate_row_operations(start_row, count)
            ExcelValidator.validate_sheet_name(sheet_name)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, sheet_name)

            # 记录操作前的信息
            original_max_row = sheet.max_row

            # 验证删除范围
            if start_row > original_max_row:
                raise DataValidationError(
                    f"起始行号({start_row})超过工作表最大行数({original_max_row})"
                )

            # 计算实际删除的行数
            actual_count = min(count, original_max_row - start_row + 1)

            # 删除行
            sheet.delete_rows(start_row, actual_count)

            # 保存文件
            workbook.save(self.file_path)

            return OperationResult(
                success=True,
                message=f"成功从第{start_row}行开始删除{actual_count}行",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet.title,
                    'deleted_start_row': start_row,
                    'actual_deleted_count': actual_count,
                    'original_max_row': original_max_row,
                    'new_max_row': sheet.max_row
                }
            )

        except Exception as e:
            logger.error(f"删除行失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def delete_columns(
        self,
        sheet_name: Optional[str] = None,
        start_column: int = 1,
        count: int = 1
    ) -> OperationResult:
        """
        删除Excel文件中的列

        Args:
            sheet_name: 工作表名称
            start_column: 开始删除的列号
            count: 要删除的列数

        Returns:
            OperationResult: 删除操作的结果
        """
        try:
            # 验证参数
            ExcelValidator.validate_column_operations(start_column, count)
            ExcelValidator.validate_sheet_name(sheet_name)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, sheet_name)

            # 记录操作前的信息
            original_max_column = sheet.max_column

            # 验证删除范围
            if start_column > original_max_column:
                raise DataValidationError(
                    f"起始列号({start_column})超过工作表最大列数({original_max_column})"
                )

            # 计算实际删除的列数
            actual_count = min(count, original_max_column - start_column + 1)

            # 删除列
            sheet.delete_cols(start_column, actual_count)

            # 保存文件
            workbook.save(self.file_path)

            return OperationResult(
                success=True,
                message=f"成功删除{actual_count}列",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet.title,
                    'deleted_start_column': start_column,
                    'actual_deleted_count': actual_count,
                    'original_max_column': original_max_column,
                    'new_max_column': sheet.max_column
                }
            )

        except Exception as e:
            logger.error(f"删除列失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def _get_worksheet(self, workbook, sheet_name: Optional[str]):
        """获取工作表"""
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise SheetNotFoundError(f"工作表不存在: {sheet_name}")
            return workbook[sheet_name]
        else:
            return workbook.active

    def _write_data(
        self,
        sheet,
        data: List[List[Any]],
        start_row: int,
        start_col: int,
        preserve_formulas: bool
    ) -> List[ModifiedCell]:
        """写入数据到工作表"""
        modified_cells = []

        for row_offset, row_data in enumerate(data):
            for col_offset, value in enumerate(row_data):
                row_idx = start_row + row_offset
                col_idx = start_col + col_offset
                cell = sheet.cell(row=row_idx, column=col_idx)

                # 保留公式检查
                if preserve_formulas and cell.data_type == 'f':
                    continue

                old_value = cell.value
                cell.value = value

                modified_cells.append(ModifiedCell(
                    coordinate=cell.coordinate,
                    old_value=old_value,
                    new_value=value
                ))

        return modified_cells

    def set_formula(
        self,
        cell_address: str,
        formula: str,
        sheet_name: Optional[str] = None
    ) -> OperationResult:
        """
        设置单元格公式

        Args:
            cell_address: 目标单元格地址（如"A1"）
            formula: Excel公式（不包含等号）
            sheet_name: 目标工作表名称

        Returns:
            OperationResult: 公式设置结果
        """
        try:
            # 验证公式格式（简单验证）
            if not formula.strip():
                return OperationResult(
                    success=False,
                    error="公式不能为空"
                )

            # 确保公式不以等号开头（openpyxl会自动添加）
            if formula.startswith('='):
                formula = formula[1:]

            # 验证单元格地址格式
            from openpyxl.utils.cell import coordinate_from_string
            try:
                coordinate_from_string(cell_address)
            except ValueError as e:
                return OperationResult(
                    success=False,
                    error=f"单元格地址格式错误: {cell_address}"
                )

            # 加载工作簿并设置公式
            workbook = load_workbook(self.file_path)
            sheet = self._get_worksheet(workbook, sheet_name)

            # 设置公式
            cell = sheet[cell_address]
            old_value = cell.value
            old_formula = cell.formula if hasattr(cell, 'formula') else None

            cell.value = f"={formula}"

            # 保存文件
            workbook.save(self.file_path)
            workbook.close()

            # 重新读取以获取计算值
            workbook_read = load_workbook(self.file_path, data_only=True)
            sheet_read = self._get_worksheet(workbook_read, sheet_name)
            calculated_value = sheet_read[cell_address].value
            workbook_read.close()

            logger.info(f"成功设置公式: {cell_address} = {formula}")

            return OperationResult(
                success=True,
                message=f"公式设置成功",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet.title,
                    'cell_address': cell_address,
                    'formula': formula,
                    'calculated_value': calculated_value,
                    'old_value': old_value,
                    'old_formula': old_formula
                }
            )

        except Exception as e:
            logger.error(f"设置公式失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def format_cells(
        self,
        range_expression: str,
        formatting: dict,
        sheet_name: Optional[str] = None
    ) -> OperationResult:
        """
        设置单元格格式

        Args:
            range_expression: 目标范围
            formatting: 格式配置字典
            sheet_name: 目标工作表名

        Returns:
            OperationResult: 格式应用结果
        """
        try:
            from openpyxl.styles import Font, PatternFill, Border, Alignment

            # 解析范围表达式
            range_info = RangeParser.parse_range_expression(range_expression)

            # 加载工作簿
            workbook = load_workbook(self.file_path)
            sheet = self._get_worksheet(workbook, sheet_name or range_info.sheet_name)

            # 获取范围边界
            if range_info.range_type in [RangeType.COLUMN_RANGE, RangeType.SINGLE_COLUMN, RangeType.ROW_RANGE, RangeType.SINGLE_ROW]:
                # 处理整行或整列
                cells_range = sheet[range_expression.replace(f"{sheet.title}!", "")]
            else:
                min_col, min_row, max_col, max_row = range_boundaries(range_info.cell_range)
                cells_range = sheet.iter_rows(
                    min_row=min_row, max_row=max_row,
                    min_col=min_col, max_col=max_col
                )

            formatted_count = 0

            # 应用格式
            for row in cells_range:
                if isinstance(row, tuple):
                    for cell in row:
                        self._apply_cell_format(cell, formatting)
                        formatted_count += 1
                else:
                    self._apply_cell_format(row, formatting)
                    formatted_count += 1

            # 保存文件
            workbook.save(self.file_path)
            workbook.close()

            logger.info(f"成功格式化{formatted_count}个单元格")

            return OperationResult(
                success=True,
                message=f"成功格式化{formatted_count}个单元格",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet.title,
                    'range': range_expression,
                    'formatted_count': formatted_count,
                    'formatting_applied': formatting
                }
            )

        except Exception as e:
            logger.error(f"格式化失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def evaluate_formula(
        self,
        formula: str,
        context_sheet: Optional[str] = None
    ) -> OperationResult:
        """
        临时执行Excel公式并返回计算结果，不修改文件

        Args:
            formula: Excel公式（不包含等号）
            context_sheet: 公式执行的上下文工作表

        Returns:
            OperationResult: 公式执行结果
        """
        try:
            import tempfile
            import os
            from openpyxl import Workbook
            import time

            start_time = time.time()

            # 确保公式不以等号开头
            if formula.startswith('='):
                formula = formula[1:]

            # 验证公式格式
            if not formula.strip():
                return OperationResult(
                    success=False,
                    error="公式不能为空"
                )

            # 加载原始工作簿（用于提供数据上下文）
            original_workbook = load_workbook(self.file_path, data_only=False)

            # 创建临时工作簿进行计算
            temp_workbook = Workbook()
            temp_sheet = temp_workbook.active
            temp_sheet.title = "Calculation"

            # 选择要复制的源工作表
            if context_sheet and context_sheet in original_workbook.sheetnames:
                source_sheet = original_workbook[context_sheet]
            else:
                # 使用活动工作表或第一个工作表
                source_sheet = original_workbook.active

            # 复制数据到临时工作表
            for row in source_sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        target_cell = temp_sheet.cell(
                            row=cell.row,
                            column=cell.column
                        )
                        target_cell.value = cell.value

            # 在临时单元格中设置要计算的公式
            calc_cell = temp_sheet['Z1']  # 使用Z1作为计算单元格
            calc_cell.value = f"={formula}"

            # 保存到临时文件
            temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            temp_file.close()

            try:
                # 保存工作簿
                temp_workbook.save(temp_file.name)
                temp_workbook.close()
                original_workbook.close()

                # 使用xlcalculator计算公式
                try:
                    from xlcalculator import ModelCompiler, Evaluator

                    # 编译模型
                    compiler = ModelCompiler()
                    model = compiler.read_and_parse_archive(temp_file.name)
                    evaluator = Evaluator(model)

                    # 计算Z1位置的公式
                    calculated_value = evaluator.evaluate('Calculation!Z1')

                except ImportError:
                    return OperationResult(
                        success=False,
                        error="需要安装xlcalculator库来支持公式计算: pip install xlcalculator"
                    )
                except Exception as calc_error:
                    # 如果xlcalculator失败，尝试基础的手动解析
                    logger.warning(f"xlcalculator计算失败，尝试基础解析: {calc_error}")

                    # 重新加载工作簿获取数据
                    data_workbook = load_workbook(temp_file.name, data_only=True)
                    data_sheet = data_workbook["Calculation"]

                    # 尝试基础的公式解析
                    calculated_value = self._basic_formula_parse(formula, data_sheet)
                    data_workbook.close()

                # 确定结果类型
                result_type = "unknown"
                if calculated_value is None:
                    result_type = "null"
                elif isinstance(calculated_value, (int, float)):
                    result_type = "number"
                elif isinstance(calculated_value, str):
                    result_type = "text"
                elif isinstance(calculated_value, bool):
                    result_type = "boolean"
                else:
                    try:
                        # 检查是否是日期
                        from datetime import datetime, date
                        if isinstance(calculated_value, (datetime, date)):
                            result_type = "date"
                    except:
                        pass

                execution_time = round((time.time() - start_time) * 1000, 2)

                logger.info(f"成功计算公式: {formula} = {calculated_value}")

                return OperationResult(
                    success=True,
                    message="公式执行成功",
                    metadata={
                        'formula': formula,
                        'result': calculated_value,
                        'result_type': result_type,
                        'execution_time_ms': execution_time,
                        'context_sheet': context_sheet or "default"
                    }
                )

            finally:
                # 清理临时文件
                try:
                    os.unlink(temp_file.name)
                except:
                    pass

        except Exception as e:
            logger.error(f"公式执行失败: {e}")
            return OperationResult(
                success=False,
                error=f"公式执行失败: {str(e)}"
            )

    def _basic_formula_parse(self, formula: str, sheet) -> any:
        """基础公式解析器 - 支持简单的函数"""
        import re

        formula = formula.strip()

        # SUM函数
        sum_match = re.match(r'SUM\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if sum_match:
            start_cell, end_cell = sum_match.groups()
            return self._calculate_range_sum(sheet, start_cell, end_cell)

        # AVERAGE函数
        avg_match = re.match(r'AVERAGE\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if avg_match:
            start_cell, end_cell = avg_match.groups()
            total = self._calculate_range_sum(sheet, start_cell, end_cell)
            count = self._calculate_range_count(sheet, start_cell, end_cell)
            return total / count if count > 0 else 0

        # COUNT函数
        count_match = re.match(r'COUNT\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if count_match:
            start_cell, end_cell = count_match.groups()
            return self._calculate_range_count(sheet, start_cell, end_cell)

        # 简单的数学表达式
        if re.match(r'^[\d\+\-\*\/\s\(\)\.]+$', formula):
            try:
                return eval(formula)  # 注意：这在生产环境中需要更安全的实现
            except:
                pass

        # IF函数简单实现
        if_match = re.match(r'IF\((.+),\s*"?([^,"]+)"?,\s*"?([^,"]+)"?\)', formula, re.IGNORECASE)
        if if_match:
            condition, true_val, false_val = if_match.groups()
            # 简单条件判断
            if '>' in condition:
                parts = condition.split('>')
                if len(parts) == 2:
                    left = float(parts[0].strip())
                    right = float(parts[1].strip())
                    return true_val if left > right else false_val
            elif '<' in condition:
                parts = condition.split('<')
                if len(parts) == 2:
                    left = float(parts[0].strip())
                    right = float(parts[1].strip())
                    return true_val if left < right else false_val

        # CONCATENATE函数
        concat_match = re.match(r'CONCATENATE\((.+)\)', formula, re.IGNORECASE)
        if concat_match:
            args = concat_match.group(1).split(',')
            result = ""
            for arg in args:
                arg = arg.strip().strip('"')
                result += arg
            return result

        return None

    def _calculate_range_sum(self, sheet, start_cell: str, end_cell: str) -> float:
        """计算范围求和"""
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
        total = 0

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    total += cell.value

        return total

    def _calculate_range_count(self, sheet, start_cell: str, end_cell: str) -> int:
        """计算范围内数值个数"""
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
        count = 0

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    count += 1

        return count

    def _apply_cell_format(self, cell, formatting: dict):
        """应用单元格格式"""
        from openpyxl.styles import Font, PatternFill, Border, Alignment

        # 字体格式
        if 'font' in formatting:
            font_config = formatting['font']
            cell.font = Font(
                name=font_config.get('name', cell.font.name),
                size=font_config.get('size', cell.font.size),
                bold=font_config.get('bold', cell.font.bold),
                italic=font_config.get('italic', cell.font.italic),
                color=font_config.get('color', cell.font.color)
            )

        # 背景颜色
        if 'fill' in formatting:
            fill_config = formatting['fill']
            cell.fill = PatternFill(
                start_color=fill_config.get('color', 'FFFFFF'),
                end_color=fill_config.get('color', 'FFFFFF'),
                fill_type='solid'
            )

        # 对齐方式
        if 'alignment' in formatting:
            align_config = formatting['alignment']
            cell.alignment = Alignment(
                horizontal=align_config.get('horizontal', cell.alignment.horizontal),
                vertical=align_config.get('vertical', cell.alignment.vertical)
            )
