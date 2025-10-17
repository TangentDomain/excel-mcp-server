"""
Excel MCP Server - Excel操作API模块

提供高内聚的Excel业务操作功能，包含完整的参数验证、业务逻辑、错误处理和结果格式化

@intention: 将Excel操作的具体实现从server.py中分离，提高代码内聚性和可维护性
"""

import logging
import os
import threading
import time
import uuid
from pathlib import Path
from typing import Dict, Any, List, Optional, Union

from ..core.excel_reader import ExcelReader
from ..core.excel_writer import ExcelWriter
from ..core.excel_manager import ExcelManager
from ..utils.formatter import format_operation_result

logger = logging.getLogger(__name__)


class OperationManager:
    """
    @class OperationManager
    @brief 管理正在进行的Excel操作，提供取消和状态跟踪功能
    @intention 实现操作的异步管理，支持取消、进度跟踪和状态查询
    """

    _instance = None
    _lock = threading.Lock()

    def __new__(cls):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
                    cls._instance._operations = {}
                    cls._instance._history = []
        return cls._instance

    def __init__(self):
        if not hasattr(self, '_initialized'):
            self._operations = {}
            self._history = []
            self._initialized = True

    def start_operation(
        self,
        operation_type: str,
        file_path: str,
        range_expression: str,
        user_id: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> str:
        """开始一个新操作，返回操作ID"""
        operation_id = str(uuid.uuid4())

        operation = {
            'id': operation_id,
            'type': operation_type,
            'file_path': file_path,
            'range_expression': range_expression,
            'user_id': user_id,
            'status': 'pending',
            'progress': 0,
            'start_time': time.time(),
            'end_time': None,
            'cancelled': False,
            'metadata': metadata or {},
            'error': None,
            'result': None,
            'cancellable': True
        }

        self._operations[operation_id] = operation
        logger.info(f"{self._LOG_PREFIX} 开始操作: {operation_id} - {operation_type}")

        return operation_id

    def update_progress(self, operation_id: str, progress: int, message: Optional[str] = None):
        """更新操作进度"""
        if operation_id in self._operations:
            self._operations[operation_id]['progress'] = max(0, min(100, progress))
            if message:
                self._operations[operation_id]['current_message'] = message

    def check_cancelled(self, operation_id: str) -> bool:
        """检查操作是否已被取消"""
        if operation_id in self._operations:
            return self._operations[operation_id].get('cancelled', False)
        return False

    def cancel_operation(self, operation_id: str, reason: str = "用户取消") -> Dict[str, Any]:
        """取消操作"""
        if operation_id not in self._operations:
            return {
                'success': False,
                'error': f'操作不存在: {operation_id}',
                'message': '无法取消不存在的操作'
            }

        operation = self._operations[operation_id]

        if not operation.get('cancellable', True):
            return {
                'success': False,
                'error': 'OPERATION_NOT_CANCELLABLE',
                'message': '此操作不允许取消'
            }

        if operation['status'] == 'completed':
            return {
                'success': False,
                'error': 'OPERATION_ALREADY_COMPLETED',
                'message': '操作已完成，无法取消'
            }

        if operation['status'] == 'cancelled':
            return {
                'success': False,
                'error': 'OPERATION_ALREADY_CANCELLED',
                'message': '操作已被取消'
            }

        # 标记为取消
        operation['cancelled'] = True
        operation['status'] = 'cancelled'
        operation['end_time'] = time.time()
        operation['cancel_reason'] = reason

        logger.info(f"{self._LOG_PREFIX} 操作已取消: {operation_id} - {reason}")

        # 移动到历史记录
        self._move_to_history(operation_id)

        return {
            'success': True,
            'operation_id': operation_id,
            'message': f'操作已成功取消: {reason}',
            'cancelled_at': operation['end_time']
        }

    def complete_operation(self, operation_id: str, result: Optional[Dict[str, Any]] = None):
        """标记操作完成"""
        if operation_id in self._operations:
            operation = self._operations[operation_id]
            operation['status'] = 'completed'
            operation['progress'] = 100
            operation['end_time'] = time.time()
            operation['result'] = result

            # 移动到历史记录
            self._move_to_history(operation_id)

    def fail_operation(self, operation_id: str, error: str):
        """标记操作失败"""
        if operation_id in self._operations:
            operation = self._operations[operation_id]
            operation['status'] = 'failed'
            operation['end_time'] = time.time()
            operation['error'] = error

            # 移动到历史记录
            self._move_to_history(operation_id)

    def get_operation_status(self, operation_id: str) -> Dict[str, Any]:
        """获取操作状态"""
        if operation_id in self._operations:
            return self._operations[operation_id].copy()
        else:
            # 在历史记录中查找
            for operation in self._history:
                if operation['id'] == operation_id:
                    return operation.copy()

        return {
            'success': False,
            'error': f'操作不存在: {operation_id}',
            'status': 'not_found'
        }

    def list_active_operations(self, user_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """列出活跃操作"""
        operations = list(self._operations.values())

        if user_id:
            operations = [op for op in operations if op.get('user_id') == user_id]

        return operations

    def list_operation_history(self, limit: int = 50, user_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """列出操作历史"""
        history = self._history.copy()

        if user_id:
            history = [op for op in history if op.get('user_id') == user_id]

        # 按时间倒序排列
        history.sort(key=lambda x: x.get('start_time', 0), reverse=True)

        return history[:limit]

    def _move_to_history(self, operation_id: str):
        """将操作移动到历史记录"""
        if operation_id in self._operations:
            operation = self._operations.pop(operation_id)
            self._history.append(operation)

            # 限制历史记录数量
            if len(self._history) > 1000:
                self._history = self._history[-500:]  # 保留最近500条

    @property
    def _LOG_PREFIX(self):
        return '[OperationManager]'


class ExcelOperations:
    """
    @class ExcelOperations
    @brief Excel业务操作的高内聚封装
    @intention 提供完整的Excel操作功能，包含参数验证、错误处理、结果格式化
    """

    # ==================== 日志系统 ====================
    DEBUG_LOG_ENABLED: bool = False
    _LOG_PREFIX = '[API][ExcelOperations]'

    # ==================== 主干API ====================

    @classmethod
    def get_range(
        cls,
        file_path: str,
        range_expression: str,
        include_formatting: bool = False
    ) -> Dict[str, Any]:
        """
        @intention 获取Excel文件中指定范围的数据，提供完整的业务逻辑处理

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            range_expression: 范围表达式，必须包含工作表名
            include_formatting: 是否包含格式信息

        Returns:
            Dict: 标准化的操作结果

        Example:
            result = ExcelOperations.get_range("data.xlsx", "Sheet1!A1:C10")
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始获取范围数据: {range_expression}")

        try:
            # 步骤1: 验证参数格式
            validation_result = cls._validate_range_format(range_expression)
            if not validation_result['valid']:
                return cls._format_error_result(validation_result['error'])

            # 步骤2: 执行数据读取
            reader = ExcelReader(file_path)
            result = reader.get_range(range_expression, include_formatting)
            reader.close()

            # 步骤3: 格式化结果
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"获取范围数据失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def update_range(
        cls,
        file_path: str,
        range_expression: str,
        data: List[List[Any]],
        preserve_formulas: bool = True,
        insert_mode: bool = True,
        require_confirmation: bool = False,
        skip_safety_checks: bool = False
    ) -> Dict[str, Any]:
        """
        @intention 更新Excel文件中指定范围的数据，支持插入和覆盖模式

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            range_expression: 范围表达式，必须包含工作表名
            data: 二维数组数据 [[row1], [row2], ...]
            preserve_formulas: 是否保留现有公式
            insert_mode: 数据写入模式 (默认值: True)
                - True: 插入模式，在指定位置插入新行然后写入数据（更安全）
                - False: 覆盖模式，直接覆盖目标范围的现有数据
            require_confirmation: 是否需要用户确认（基于风险评估自动触发）
            skip_safety_checks: 跳过安全检查（仅用于系统维护）

        Returns:
            Dict: 标准化的操作结果，包含安全警告和影响分析

        Example:
            data = [["姓名", "年龄"], ["张三", 25]]
            # 插入模式（默认，更安全）
            result = ExcelOperations.update_range("test.xlsx", "Sheet1!A1:B2", data)
            # 覆盖模式（显式指定）
            result = ExcelOperations.update_range("test.xlsx", "Sheet1!A1:B2", data, insert_mode=False)
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始更新范围数据: {range_expression}, 模式: {'插入' if insert_mode else '覆盖'}")

        try:
            # 步骤1: 验证参数格式
            validation_result = cls._validate_range_format(range_expression)
            if not validation_result['valid']:
                return cls._format_error_result(validation_result['error'])

            # 步骤2: 安全检查和影响评估（除非明确跳过）
            if not skip_safety_checks:
                # 子步骤2.1: 文件状态检查
                file_status_check = cls.check_file_status(file_path)
                if not file_status_check['success']:
                    return cls._format_error_result(f"文件状态检查失败: {file_status_check.get('error', '未知错误')}")

                file_status = file_status_check['file_status']
                if file_status.get('locked', False):
                    return {
                        'success': False,
                        'error': 'FILE_LOCKED',
                        'message': f"文件被锁定: {file_status.get('locked_by', '未知程序')}",
                        'file_status': file_status,
                        'security_warnings': file_status_check.get('security_warnings', []),
                        'safety_recommendations': file_status_check.get('safety_recommendations', [])
                    }

                if not file_status.get('writable', False):
                    return {
                        'success': False,
                        'error': 'FILE_NOT_WRITABLE',
                        'message': "文件不可写，请检查权限或关闭占用程序",
                        'file_status': file_status,
                        'security_warnings': file_status_check.get('security_warnings', []),
                        'safety_recommendations': file_status_check.get('safety_recommendations', [])
                    }

                # 子步骤2.2: 操作影响评估
                impact_assessment = cls.assess_operation_impact(
                    file_path, range_expression, "update", data
                )

                if not impact_assessment['success']:
                    return cls._format_error_result(f"安全检查失败: {impact_assessment.get('error', '未知错误')}")

                impact_analysis = impact_assessment['impact_analysis']
                risk_level = impact_analysis['operation_risk_level']

                # 步骤3: 生成安全警告
                safety_warnings = cls._generate_safety_warnings(
                    "update", range_expression, impact_analysis, data, insert_mode
                )

                # 步骤4: 根据风险等级决定是否需要确认
                auto_require_confirmation = risk_level in ['high', 'critical']
                final_require_confirmation = require_confirmation or auto_require_require_confirmation

                # 步骤5: 高风险操作自动创建备份
                backup_result = None
                if risk_level in ['high', 'critical']:
                    import time
                    backup_result = cls.create_auto_backup(
                        file_path,
                        backup_name=f"pre_{operation_type}_{time.strftime('%H%M%S')}",
                        backup_reason=f"高风险{operation_type}操作前自动备份"
                    )
                    if not backup_result['success']:
                        return {
                            'success': False,
                            'error': 'BACKUP_CREATION_FAILED',
                            'message': f'高风险操作前创建备份失败: {backup_result.get("error", "未知错误")}',
                            'risk_level': risk_level,
                            'backup_error': backup_result.get('error')
                        }

                # 如果需要确认但未获得确认，返回警告信息
                if final_require_confirmation:
                    return {
                        'success': False,
                        'error': 'OPERATION_REQUIRES_CONFIRMATION',
                        'message': '此操作存在数据安全风险，需要用户确认后才能执行',
                        'risk_level': risk_level,
                        'impact_analysis': impact_analysis,
                        'safety_warnings': safety_warnings,
                        'preview_data': impact_assessment['preview_data'],
                        'safe_execution_plan': impact_assessment['safe_execution_plan'],
                        'confirmation_required': True,
                        'backup_available': backup_result.get('success', False) if backup_result else False
                    }

            # 步骤5: 执行数据写入
            writer = ExcelWriter(file_path)
            result = writer.update_range(range_expression, data, preserve_formulas, insert_mode)

            # 步骤6: 格式化结果，包含安全信息
            formatted_result = format_operation_result(result)

            # 添加安全信息到结果中
            if not skip_safety_checks:
                formatted_result['safety_info'] = {
                    'risk_level': risk_level,
                    'warnings': safety_warnings,
                    'impact_analysis': impact_analysis
                }

            return formatted_result

        except Exception as e:
            error_msg = f"更新范围数据失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def list_sheets(cls, file_path: str) -> Dict[str, Any]:
        """
        @intention 获取Excel文件中所有工作表信息，提供完整的文件结构概览

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)

        Returns:
            Dict: 包含工作表列表、总数量、活动工作表等信息

        Example:
            result = ExcelOperations.list_sheets("data.xlsx")
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始获取工作表列表: {file_path}")

        try:
            # 步骤1: 读取工作表信息
            reader = ExcelReader(file_path)
            result = reader.list_sheets()

            # 步骤2: 提取和格式化数据
            sheets = [sheet.name for sheet in result.data] if result.data else []

            response = {
                'success': True,
                'sheets': sheets,
                'file_path': file_path,
                'total_sheets': result.metadata.get('total_sheets', len(sheets)) if result.metadata else len(sheets)
            }

            # 步骤3: 清理资源
            reader.close()

            return response

        except Exception as e:
            error_msg = f"获取工作表列表失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def get_headers(
        cls,
        file_path: str,
        sheet_name: str,
        header_row: int = 1,
        max_columns: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        @intention 获取指定工作表的表头信息，支持游戏开发双行模式（字段描述+字段名）

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            sheet_name: 工作表名称
            header_row: 表头起始行号 (1-based，默认从第1行开始获取两行)
            max_columns: 最大列数限制，None表示自动截取到空列

        Returns:
            Dict: 包含双行表头信息
            {
                'success': bool,
                'data': List[str],  # 字段名列表（兼容性）
                'headers': List[str],  # 字段名列表（兼容性）
                'descriptions': List[str],  # 字段描述列表（第1行）
                'field_names': List[str],   # 字段名列表（第2行）
                'header_count': int,
                'sheet_name': str,
                'header_row': int,
                'message': str
            }

        Example:
            result = ExcelOperations.get_headers("data.xlsx", "Sheet1")
            # 第1行：['技能ID描述', '技能名称描述', '技能类型描述']
            # 第2行：['skill_id', 'skill_name', 'skill_type']
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始获取双行表头: {sheet_name}")

        try:
            # 步骤1: 构建双行范围表达式
            range_expression = cls._build_header_range(sheet_name, header_row, max_columns, dual_row=True)

            # 步骤2: 读取表头数据（两行）
            reader = ExcelReader(file_path)
            result = reader.get_range(range_expression)
            reader.close()

            if not result.success:
                return cls._format_error_result(f"无法读取表头数据: {result.message}")

            # 步骤3: 解析双行表头信息
            header_info = cls._parse_dual_header_data(result.data, max_columns)

            return {
                'success': True,
                'data': header_info['field_names'],  # 兼容性字段，返回字段名
                'headers': header_info['field_names'],  # 兼容性字段，返回字段名
                'descriptions': header_info['descriptions'],  # 字段描述（第1行）
                'field_names': header_info['field_names'],    # 字段名（第2行）
                'header_count': len(header_info['field_names']),
                'sheet_name': sheet_name,
                'header_row': header_row,
                'message': f"成功获取{len(header_info['field_names'])}个表头字段（描述+字段名）"
            }

        except Exception as e:
            error_msg = f"获取表头失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def create_file(
        cls,
        file_path: str,
        sheet_names: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        @intention 创建新的Excel文件，支持自定义工作表配置

        Args:
            file_path: 新文件路径 (必须以.xlsx或.xlsm结尾)
            sheet_names: 工作表名称列表，None表示默认工作表

        Returns:
            Dict: 包含创建结果和文件信息

        Example:
            result = ExcelOperations.create_file("new_file.xlsx", ["数据", "分析"])
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始创建文件: {file_path}")

        try:
            # 步骤1: 执行文件创建
            result = ExcelManager.create_file(file_path, sheet_names)

            # 步骤2: 格式化结果
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"创建文件失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    # ==================== 分支实现 ====================

    # --- 参数验证 ---
    @classmethod
    def _validate_range_format(cls, range_expression: str) -> Dict[str, Any]:
        """验证范围表达式格式"""
        if not range_expression or not range_expression.strip():
            return {'valid': False, 'error': 'range参数不能为空'}

        if '!' not in range_expression:
            return {
                'valid': False,
                'error': f"range必须包含工作表名。当前格式: '{range_expression}'，正确格式示例: 'Sheet1!A1:B2'"
            }

        return {'valid': True}

    @classmethod
    def _build_header_range(cls, sheet_name: str, header_row: int, max_columns: Optional[int], dual_row: bool = False) -> str:
        """构建表头范围表达式，支持单行或双行模式"""
        if max_columns:
            # 如果指定了最大列数，使用具体范围
            from openpyxl.utils import get_column_letter
            end_column = get_column_letter(max_columns)
            if dual_row:
                # 双行模式：获取连续两行
                end_row = header_row + 1
                return f"{sheet_name}!A{header_row}:{end_column}{end_row}"
            else:
                # 单行模式（保持兼容性）
                return f"{sheet_name}!A{header_row}:{end_column}{header_row}"
        else:
            # 否则使用一个合理的默认范围（读取前100列）
            if dual_row:
                # 双行模式：获取连续两行
                end_row = header_row + 1
                return f"{sheet_name}!A{header_row}:CV{end_row}"  # CV = 第100列
            else:
                # 单行模式（保持兼容性）
                return f"{sheet_name}!A{header_row}:CV{header_row}"  # CV = 第100列

    @classmethod
    def _parse_header_data(cls, data: List[List], max_columns: Optional[int]) -> List[str]:
        """解析表头数据"""
        headers = []
        if data and len(data) > 0:
            first_row = data[0]
            for i, cell_info in enumerate(first_row):
                # 处理CellInfo对象和普通值
                cell_value = getattr(cell_info, 'value', cell_info) if hasattr(cell_info, 'value') else cell_info

                # 转换为字符串并清理
                if cell_value is not None:
                    str_value = str(cell_value).strip()
                    if str_value != "":
                        headers.append(str_value)
                    else:
                        # 空字符串的处理
                        if max_columns:
                            headers.append("")  # 指定max_columns时保留空字符串
                        else:
                            break  # 否则停止
                else:
                    # None值的处理
                    if max_columns:
                        headers.append("")  # 指定max_columns时将None转为空字符串
                    else:
                        break  # 否则停止

                # 如果指定了max_columns，检查是否已达到限制
                if max_columns and len(headers) >= max_columns:
                    break

        return headers

    @classmethod
    def _parse_dual_header_data(cls, data: List[List], max_columns: Optional[int]) -> Dict[str, List[str]]:
        """解析双行表头数据（字段描述 + 字段名），支持空值fallback机制"""
        descriptions = []
        field_names = []

        if not data or len(data) < 2:
            # 如果数据不足两行，返回空结果
            return {
                'descriptions': descriptions,
                'field_names': field_names
            }

        # 解析第一行（字段描述）
        first_row = data[0] if len(data) > 0 else []
        # 解析第二行（字段名）
        second_row = data[1] if len(data) > 1 else []

        # 确定实际处理的列数
        max_cols = max(len(first_row), len(second_row))  # 改为取最大值，不遗漏任何列
        if max_columns:
            max_cols = min(max_cols, max_columns)

        # 导入列名转换工具
        from openpyxl.utils import get_column_letter

        for i in range(max_cols):
            # 处理字段描述（第1行）
            desc_cell = first_row[i] if i < len(first_row) else None
            desc_value = getattr(desc_cell, 'value', desc_cell) if hasattr(desc_cell, 'value') else desc_cell
            desc_str = str(desc_value).strip() if desc_value is not None and str(desc_value).strip() else ""

            # 处理字段名（第2行）
            name_cell = second_row[i] if i < len(second_row) else None
            name_value = getattr(name_cell, 'value', name_cell) if hasattr(name_cell, 'value') else name_cell
            name_str = str(name_value).strip() if name_value is not None and str(name_value).strip() else ""

            # 🆕 智能Fallback机制
            column_letter = get_column_letter(i + 1)  # 1-based列名：A, B, C...

            # 描述为空时使用列标识作为fallback
            if not desc_str:
                desc_str = f"列{column_letter}"  # 中文：列A, 列B, 列C...

            # 字段名为空时使用列名作为fallback
            if not name_str:
                name_str = column_letter.lower()  # 小写：a, b, c...

            # 🆕 检查是否应该停止（简化的停止条件）
            # 只有在没有指定max_columns时才进行智能停止
            if not max_columns:
                # 检查原始数据是否为完全空（描述和字段名都是原始空值）
                desc_is_empty = (desc_cell is None or
                               (hasattr(desc_cell, 'value') and desc_cell.value is None) or
                               (not hasattr(desc_cell, 'value') and desc_cell is None))
                name_is_empty = (name_cell is None or
                               (hasattr(name_cell, 'value') and name_cell.value is None) or
                               (not hasattr(name_cell, 'value') and name_cell is None))

                # 如果当前列完全为空，检查接下来连续3列是否也为空
                if desc_is_empty and name_is_empty:
                    consecutive_empty = 0
                    for j in range(i, min(i + 3, max_cols)):  # 检查当前及后续2列
                        check_desc = first_row[j] if j < len(first_row) else None
                        check_name = second_row[j] if j < len(second_row) else None

                        desc_empty = (check_desc is None or
                                    (hasattr(check_desc, 'value') and check_desc.value is None) or
                                    (not hasattr(check_desc, 'value') and check_desc is None))
                        name_empty = (check_name is None or
                                    (hasattr(check_name, 'value') and check_name.value is None) or
                                    (not hasattr(check_name, 'value') and check_name is None))

                        if desc_empty and name_empty:
                            consecutive_empty += 1
                        else:
                            break

                    # 如果连续3列都为空，则停止
                    if consecutive_empty >= 3:
                        break

            descriptions.append(desc_str)
            field_names.append(name_str)

            # 如果指定了max_columns，检查是否已达到限制
            if max_columns and len(field_names) >= max_columns:
                break

        return {
            'descriptions': descriptions,
            'field_names': field_names
        }

    @classmethod
    def search(
        cls,
        file_path: str,
        pattern: str,
        sheet_name: Optional[str] = None,
        case_sensitive: bool = False,
        whole_word: bool = False,
        use_regex: bool = False,
        include_values: bool = True,
        include_formulas: bool = False,
        range: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        @intention 在Excel文件中搜索单元格内容（VSCode风格搜索选项）

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            pattern: 搜索模式（正则表达式或字面字符串）
            sheet_name: 工作表名称 (可选)
            case_sensitive: 大小写敏感
            whole_word: 全词匹配
            use_regex: 启用正则表达式
            include_values: 是否搜索单元格值
            include_formulas: 是否搜索公式内容
            range: 搜索范围表达式

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            search_type = "正则" if use_regex else ("全词" if whole_word else "字符串")
            case_info = "大小写敏感" if case_sensitive else "忽略大小写"
            logger.info(f"{cls._LOG_PREFIX} 开始{search_type}搜索({case_info}): {pattern}")

        try:
            from ..core.excel_search import ExcelSearcher
            import re

            searcher = ExcelSearcher(file_path)

            # 构建正则表达式模式
            if use_regex:
                # 直接使用用户提供的正则表达式
                regex_pattern = pattern
            else:
                # 将字面字符串转义为正则表达式
                escaped_pattern = re.escape(pattern)

                # 如果是全词匹配，添加单词边界
                if whole_word:
                    regex_pattern = r'\b' + escaped_pattern + r'\b'
                else:
                    regex_pattern = escaped_pattern

            # 构建正则表达式标志
            regex_flags = "" if case_sensitive else "i"

            result = searcher.regex_search(regex_pattern, regex_flags, include_values, include_formulas, sheet_name, range)
            return format_operation_result(result)

        except Exception as e:
            search_type = "正则" if use_regex else ("全词" if whole_word else "字符串")
            error_msg = f"{search_type}搜索失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def search_directory(
        cls,
        directory_path: str,
        pattern: str,
        case_sensitive: bool = False,
        whole_word: bool = False,
        use_regex: bool = False,
        include_values: bool = True,
        include_formulas: bool = False,
        recursive: bool = True,
        file_extensions: Optional[List[str]] = None,
        file_pattern: Optional[str] = None,
        max_files: int = 100
    ) -> Dict[str, Any]:
        """
        @intention 在目录下的所有Excel文件中搜索内容（VSCode风格搜索选项）

        Args:
            directory_path: 目录路径
            pattern: 搜索模式（正则表达式或字面字符串）
            case_sensitive: 大小写敏感
            whole_word: 全词匹配
            use_regex: 启用正则表达式
            其他参数同search方法

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            search_type = "正则" if use_regex else ("全词" if whole_word else "字符串")
            case_info = "大小写敏感" if case_sensitive else "忽略大小写"
            logger.info(f"{cls._LOG_PREFIX} 开始目录{search_type}搜索({case_info}): {directory_path}")

        try:
            from ..core.excel_search import ExcelSearcher
            import re

            # 构建正则表达式模式
            if use_regex:
                # 直接使用用户提供的正则表达式
                regex_pattern = pattern
            else:
                # 将字面字符串转义为正则表达式
                escaped_pattern = re.escape(pattern)

                # 如果是全词匹配，添加单词边界
                if whole_word:
                    regex_pattern = r'\b' + escaped_pattern + r'\b'
                else:
                    regex_pattern = escaped_pattern

            # 构建正则表达式标志
            regex_flags = "" if case_sensitive else "i"

            result = ExcelSearcher.search_directory_static(
                directory_path, regex_pattern, regex_flags, include_values, include_formulas,
                recursive, file_extensions, file_pattern, max_files
            )
            return format_operation_result(result)

        except Exception as e:
            search_type = "正则" if use_regex else ("全词" if whole_word else "字符串")
            error_msg = f"目录{search_type}搜索失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def get_sheet_headers(cls, file_path: str) -> Dict[str, Any]:
        """
        @intention 获取Excel文件中所有工作表的双行表头信息（字段描述+字段名）

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)

        Returns:
            Dict: 包含所有工作表的双行表头信息
            {
                'success': bool,
                'sheets_with_headers': [
                    {
                        'name': str,
                        'headers': List[str],       # 字段名（兼容性）
                        'descriptions': List[str],  # 字段描述（第1行）
                        'field_names': List[str],   # 字段名（第2行）
                        'header_count': int
                    }
                ],
                'file_path': str,
                'total_sheets': int
            }
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始获取所有工作表双行表头: {file_path}")

        try:
            # 步骤1: 获取所有工作表列表
            sheets_result = cls.list_sheets(file_path)
            if not sheets_result.get('success'):
                return sheets_result

            # 步骤2: 获取每个工作表的双行表头
            sheets_with_headers = []
            sheets = sheets_result.get('sheets', [])

            for sheet_name in sheets:
                try:
                    header_result = cls.get_headers(file_path, sheet_name, header_row=1)

                    if header_result.get('success'):
                        headers = header_result.get('headers', [])
                        descriptions = header_result.get('descriptions', [])
                        field_names = header_result.get('field_names', [])

                        # 如果没有获取到field_names，使用headers作为fallback
                        if not field_names and headers:
                            field_names = headers

                        sheets_with_headers.append({
                            'name': sheet_name,
                            'headers': field_names,         # 兼容性字段，使用字段名
                            'descriptions': descriptions,   # 字段描述（第1行）
                            'field_names': field_names,     # 字段名（第2行）
                            'header_count': len(field_names)
                        })
                    else:
                        sheets_with_headers.append({
                            'name': sheet_name,
                            'headers': [],
                            'descriptions': [],
                            'field_names': [],
                            'header_count': 0,
                            'error': header_result.get('error', '未知错误')
                        })

                except Exception as e:
                    sheets_with_headers.append({
                        'name': sheet_name,
                        'headers': [],
                        'descriptions': [],
                        'field_names': [],
                        'header_count': 0,
                        'error': str(e)
                    })

            return format_operation_result({
                'success': True,
                'sheets_with_headers': sheets_with_headers,
                'file_path': file_path,
                'total_sheets': len(sheets)
            })

        except Exception as e:
            error_msg = f"获取工作表表头失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def insert_rows(
        cls,
        file_path: str,
        sheet_name: str,
        row_index: int,
        count: int = 1
    ) -> Dict[str, Any]:
        """
        @intention 在指定位置插入空行

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            sheet_name: 工作表名称
            row_index: 插入位置 (1-based)
            count: 插入行数

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始插入行: {sheet_name} 第{row_index}行")

        try:
            from ..core.excel_writer import ExcelWriter
            writer = ExcelWriter(file_path)
            result = writer.insert_rows(sheet_name, row_index, count)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"插入行操作失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def insert_columns(
        cls,
        file_path: str,
        sheet_name: str,
        column_index: int,
        count: int = 1
    ) -> Dict[str, Any]:
        """
        @intention 在指定位置插入空列

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            sheet_name: 工作表名称
            column_index: 插入位置 (1-based)
            count: 插入列数

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始插入列: {sheet_name} 第{column_index}列")

        try:
            from ..core.excel_writer import ExcelWriter
            writer = ExcelWriter(file_path)
            result = writer.insert_columns(sheet_name, column_index, count)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"插入列操作失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def export_to_csv(
        cls,
        file_path: str,
        output_path: str,
        sheet_name: Optional[str] = None,
        encoding: str = "utf-8"
    ) -> Dict[str, Any]:
        """
        @intention 将Excel工作表导出为CSV文件

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            output_path: 输出CSV文件路径
            sheet_name: 工作表名称 (默认使用活动工作表)
            encoding: 文件编码 (默认: utf-8)

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始导出为CSV: {output_path}")

        try:
            from ..core.excel_converter import ExcelConverter
            converter = ExcelConverter(file_path)
            result = converter.export_to_csv(output_path, sheet_name, encoding)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"导出为CSV失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def import_from_csv(
        cls,
        csv_path: str,
        output_path: str,
        sheet_name: str = "Sheet1",
        encoding: str = "utf-8",
        has_header: bool = True
    ) -> Dict[str, Any]:
        """
        @intention 从CSV文件导入数据创建Excel文件

        Args:
            csv_path: CSV文件路径
            output_path: 输出Excel文件路径
            sheet_name: 工作表名称
            encoding: CSV文件编码
            has_header: 是否包含表头行

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始从CSV导入: {csv_path}")

        try:
            from ..core.excel_converter import ExcelConverter
            result = ExcelConverter.import_from_csv(csv_path, output_path, sheet_name, encoding, has_header)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"从CSV导入失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def convert_format(
        cls,
        input_path: str,
        output_path: str,
        target_format: str = "xlsx"
    ) -> Dict[str, Any]:
        """
        @intention 转换Excel文件格式

        Args:
            input_path: 输入文件路径
            output_path: 输出文件路径
            target_format: 目标格式

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始格式转换: {input_path} -> {output_path}")

        try:
            from ..core.excel_converter import ExcelConverter
            result = ExcelConverter.convert_format(input_path, output_path, target_format)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"文件格式转换失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def merge_files(
        cls,
        input_files: List[str],
        output_path: str,
        merge_mode: str = "sheets"
    ) -> Dict[str, Any]:
        """
        @intention 合并多个Excel文件

        Args:
            input_files: 输入文件路径列表
            output_path: 输出文件路径
            merge_mode: 合并模式

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始合并文件: {len(input_files)}个文件")

        try:
            from ..core.excel_converter import ExcelConverter
            result = ExcelConverter.merge_files(input_files, output_path, merge_mode)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"合并Excel文件失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def get_file_info(cls, file_path: str) -> Dict[str, Any]:
        """
        @intention 获取Excel文件的详细信息

        Args:
            file_path: Excel文件路径

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始获取文件信息: {file_path}")

        try:
            from ..core.excel_manager import ExcelManager
            result = ExcelManager.get_file_info(file_path)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"获取文件信息失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def create_sheet(
        cls,
        file_path: str,
        sheet_name: str,
        index: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        @intention 在文件中创建新工作表

        Args:
            file_path: Excel文件路径
            sheet_name: 新工作表名称
            index: 插入位置

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始创建工作表: {sheet_name}")

        try:
            from ..core.excel_manager import ExcelManager
            manager = ExcelManager(file_path)
            result = manager.create_sheet(sheet_name, index)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"创建工作表失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def delete_sheet(cls, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """
        @intention 删除指定工作表

        Args:
            file_path: Excel文件路径
            sheet_name: 要删除的工作表名称

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始删除工作表: {sheet_name}")

        try:
            from ..core.excel_manager import ExcelManager
            manager = ExcelManager(file_path)
            result = manager.delete_sheet(sheet_name)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"删除工作表失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def rename_sheet(
        cls,
        file_path: str,
        old_name: str,
        new_name: str
    ) -> Dict[str, Any]:
        """
        @intention 重命名工作表

        Args:
            file_path: Excel文件路径
            old_name: 当前工作表名称
            new_name: 新工作表名称

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始重命名工作表: {old_name} -> {new_name}")

        try:
            from ..core.excel_manager import ExcelManager
            manager = ExcelManager(file_path)
            result = manager.rename_sheet(old_name, new_name)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"重命名工作表失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def delete_rows(
        cls,
        file_path: str,
        sheet_name: str,
        row_index: int,
        count: int = 1
    ) -> Dict[str, Any]:
        """
        @intention 删除指定行

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            row_index: 起始行号 (1-based)
            count: 删除行数

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始删除行: {sheet_name} 第{row_index}行")

        try:
            from ..core.excel_writer import ExcelWriter
            writer = ExcelWriter(file_path)
            result = writer.delete_rows(sheet_name, row_index, count)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"删除行操作失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def delete_columns(
        cls,
        file_path: str,
        sheet_name: str,
        column_index: int,
        count: int = 1
    ) -> Dict[str, Any]:
        """
        @intention 删除指定列

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            column_index: 起始列号 (1-based)
            count: 删除列数

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始删除列: {sheet_name} 第{column_index}列")

        try:
            from ..core.excel_writer import ExcelWriter
            writer = ExcelWriter(file_path)
            result = writer.delete_columns(sheet_name, column_index, count)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"删除列操作失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def format_cells(
        cls,
        file_path: str,
        sheet_name: str,
        range: str,
        formatting: Optional[Dict[str, Any]] = None,
        preset: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        @intention 设置单元格格式

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            range: 目标范围
            formatting: 自定义格式配置
            preset: 预设样式

        Returns:
            Dict: 标准化的操作结果
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始格式化单元格: {range}")

        try:
            from ..core.excel_writer import ExcelWriter
            writer = ExcelWriter(file_path)
            # 处理预设格式
            if preset:
                preset_formats = {
                    'title': {'font': {'name': '微软雅黑', 'size': 14, 'bold': True}, 'alignment': {'horizontal': 'center'}},
                    'header': {'font': {'name': '微软雅黑', 'size': 11, 'bold': True}, 'fill': {'color': 'D9D9D9'}},
                    'data': {'font': {'name': '微软雅黑', 'size': 10}},
                    'highlight': {'fill': {'color': 'FFFF00'}},
                    'currency': {'number_format': '¥#,##0.00'}
                }
                formatting = preset_formats.get(preset, formatting or {})

            # 构建完整的range表达式
            if '!' not in range:
                range_expression = f"{sheet_name}!{range}"
            else:
                range_expression = range

            result = writer.format_cells(range_expression, formatting or {})
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"单元格格式化失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    # --- 错误处理 ---
    @classmethod
    def _format_error_result(cls, error_message: str) -> Dict[str, Any]:
        """创建标准化的错误响应"""
        return {
            'success': False,
            'error': error_message,
            'data': None
        }

    # --- 单元格操作扩展 ---
    @classmethod
    def merge_cells(cls, file_path: str, sheet_name: str, range: str) -> Dict[str, Any]:
        """
        @intention 合并指定范围的单元格
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始合并单元格: {range}")

        try:
            from ..core.excel_writer import ExcelWriter
            writer = ExcelWriter(file_path)
            result = writer.merge_cells(range, sheet_name)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"合并单元格失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def unmerge_cells(cls, file_path: str, sheet_name: str, range: str) -> Dict[str, Any]:
        """
        @intention 取消合并指定范围的单元格
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始取消合并单元格: {range}")

        try:
            from ..core.excel_writer import ExcelWriter
            writer = ExcelWriter(file_path)
            result = writer.unmerge_cells(range, sheet_name)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"取消合并单元格失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def set_borders(cls, file_path: str, sheet_name: str, range: str,
                   border_style: str = "thin") -> Dict[str, Any]:
        """
        @intention 为指定范围设置边框样式
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始设置边框: {range}, 样式: {border_style}")

        try:
            from ..core.excel_writer import ExcelWriter
            writer = ExcelWriter(file_path)
            result = writer.set_borders(range, border_style, sheet_name)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"设置边框失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def set_row_height(cls, file_path: str, sheet_name: str, row_index: int,
                      height: float, count: int = 1) -> Dict[str, Any]:
        """
        @intention 调整指定行的高度
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始调整行高: 行{row_index}, 高度{height}, 数量{count}")

        try:
            from ..core.excel_writer import ExcelWriter
            writer = ExcelWriter(file_path)

            # ExcelWriter.set_row_height(row_number, height, sheet_name)
            for i in range(count):
                row_num = row_index + i
                result = writer.set_row_height(row_num, height, sheet_name)
                if not result.success:
                    break

            return format_operation_result(result)

        except Exception as e:
            error_msg = f"调整行高失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def set_column_width(cls, file_path: str, sheet_name: str, column_index: int,
                        width: float, count: int = 1) -> Dict[str, Any]:
        """
        @intention 调整指定列的宽度
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始调整列宽: 列{column_index}, 宽度{width}, 数量{count}")

        try:
            from ..core.excel_writer import ExcelWriter
            from openpyxl.utils import get_column_letter

            writer = ExcelWriter(file_path)

            # ExcelWriter.set_column_width(column, width, sheet_name)
            for i in range(count):
                col_idx = column_index + i
                column_letter = get_column_letter(col_idx)
                result = writer.set_column_width(column_letter, width, sheet_name)
                if not result.success:
                    break

            return format_operation_result(result)

        except Exception as e:
            error_msg = f"调整列宽失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def compare_sheets(cls, file1_path: str, sheet1_name: str, file2_path: str,
                      sheet2_name: str, id_column: Union[int, str] = 1,
                      header_row: int = 1) -> Dict[str, Any]:
        """
        @intention 比较两个Excel工作表，识别ID对象的新增、删除、修改
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始比较工作表: {file1_path}:{sheet1_name} vs {file2_path}:{sheet2_name}")

        try:
            from ..core.excel_compare import ExcelComparer
            from ..models.types import ComparisonOptions

            # 创建比较选项
            options = ComparisonOptions()
            comparer = ExcelComparer(options)

            # 执行比较 - 使用正确的参数顺序
            result = comparer.compare_sheets(
                file1_path, sheet1_name, file2_path, sheet2_name, options
            )
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"比较工作表失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    # --- 公式操作扩展 ---
    @classmethod
    def set_formula(cls, file_path: str, sheet_name: str, cell_range: str,
                   formula: str) -> Dict[str, Any]:
        """
        @intention 设置指定单元格或区域的公式
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始设置公式: {cell_range} = {formula}")

        try:
            from ..core.excel_writer import ExcelWriter
            writer = ExcelWriter(file_path)
            result = writer.set_formula(sheet_name, cell_range, formula)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"设置公式失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def evaluate_formula(cls, formula: str, context_sheet: Optional[str] = None) -> Dict[str, Any]:
        """
        @intention 计算公式的值，不修改文件
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始计算公式: {formula}")

        try:
            from ..core.excel_writer import ExcelWriter
            writer = ExcelWriter("")  # 临时实例，不需要文件
            result = writer.evaluate_formula(formula, context_sheet)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"公式计算失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def compare_files(cls, file1_path: str, file2_path: str) -> Dict[str, Any]:
        """
        @intention 比较两个Excel文件的所有工作表
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始比较文件: {file1_path} vs {file2_path}")

        try:
            from ..models.types import ComparisonOptions
            from ..core.excel_compare import ExcelComparer

            # 标准文件比较配置
            options = ComparisonOptions(
                compare_values=True,
                compare_formulas=False,
                compare_formats=False,
                ignore_empty_cells=True,
                case_sensitive=True,
                structured_comparison=False
            )

            comparer = ExcelComparer(options)
            result = comparer.compare_files(file1_path, file2_path)
            return format_operation_result(result)

        except Exception as e:
            error_msg = f"比较文件失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def find_last_row(
        cls,
        file_path: str,
        sheet_name: str,
        column: Optional[Union[str, int]] = None
    ) -> Dict[str, Any]:
        """
        @intention 查找表格中最后一行有数据的位置

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            sheet_name: 工作表名称
            column: 指定列来查找最后一行（可选）
                - None: 查找整个工作表的最后一行
                - 整数: 列索引 (1-based，1=A列)
                - 字符串: 列名 (A, B, C...)

        Returns:
            Dict: 包含 success、last_row、message 等信息

        Example:
            # 查找整个工作表的最后一行
            result = ExcelOperations.find_last_row("data.xlsx", "Sheet1")
            # 查找A列的最后一行有数据的位置
            result = ExcelOperations.find_last_row("data.xlsx", "Sheet1", "A")
            # 查找第3列的最后一行有数据的位置
            result = ExcelOperations.find_last_row("data.xlsx", "Sheet1", 3)
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始查找最后一行: {sheet_name}")

        try:
            from ..core.excel_reader import ExcelReader
            reader = ExcelReader(file_path)

            # 获取工作簿和工作表
            workbook = reader._get_workbook(read_only=True, data_only=True)
            sheet = reader._get_worksheet(workbook, sheet_name)

            last_row = 0
            search_info = ""

            if column is None:
                # 查找整个工作表的最后一行
                last_row = sheet.max_row
                # 从后往前查找真正有数据的最后一行
                for row_num in range(sheet.max_row, 0, -1):
                    has_data = False
                    for col_num in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=row_num, column=col_num).value
                        if cell_value is not None and str(cell_value).strip():
                            has_data = True
                            break
                    if has_data:
                        last_row = row_num
                        break
                else:
                    last_row = 0  # 整个工作表都没有数据
                search_info = "整个工作表"
            else:
                # 查找指定列的最后一行
                from openpyxl.utils import column_index_from_string, get_column_letter

                # 转换列参数为列索引
                if isinstance(column, str):
                    try:
                        col_index = column_index_from_string(column.upper())
                    except ValueError:
                        reader.close()
                        return cls._format_error_result(f"无效的列名: {column}")
                elif isinstance(column, int):
                    if column < 1:
                        reader.close()
                        return cls._format_error_result("列索引必须大于等于1")
                    col_index = column
                else:
                    reader.close()
                    return cls._format_error_result("列参数必须是字符串或整数")

                # 查找指定列的最后一行有数据
                for row_num in range(sheet.max_row, 0, -1):
                    cell_value = sheet.cell(row=row_num, column=col_index).value
                    if cell_value is not None and str(cell_value).strip():
                        last_row = row_num
                        break

                col_letter = get_column_letter(col_index)
                search_info = f"{col_letter}列"

            reader.close()

            return {
                'success': True,
                'data': {
                    'last_row': last_row,
                    'sheet_name': sheet_name,
                    'column': column,
                    'search_scope': search_info
                },
                'last_row': last_row,  # 兼容性字段
                'message': f"成功查找{search_info}最后一行: 第{last_row}行" if last_row > 0 else f"{search_info}没有数据"
            }

        except Exception as e:
            error_msg = f"查找最后一行失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def assess_operation_impact(
        cls,
        file_path: str,
        range_expression: str,
        operation_type: str = "update",
        preview_data: Optional[List[List[Any]]] = None
    ) -> Dict[str, Any]:
        """
        @intention 评估Excel操作的数据影响范围，提供安全分析

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            range_expression: 范围表达式，必须包含工作表名
            operation_type: 操作类型 ('update', 'delete', 'insert', 'format')
            preview_data: 预览数据（对于更新操作）

        Returns:
            Dict: 包含影响分析的安全评估结果
            {
                'success': bool,
                'impact_analysis': {
                    'affected_cells': int,           # 将影响的单元格数量
                    'affected_rows': int,            # 将影响的行数
                    'affected_columns': int,         # 将影响的列数
                    'non_empty_cells': int,          # 非空单元格数量
                    'existing_data_summary': dict,   # 现有数据摘要
                    'operation_risk_level': str,     # 操作风险等级 (low/medium/high/critical)
                    'warnings': List[str],           # 警告信息
                    'recommendations': List[str]     # 安全建议
                },
                'preview_data': List[List],         # 当前数据预览
                'safe_execution_plan': dict         # 安全执行计划
            }
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始评估操作影响: {range_expression}, 类型: {operation_type}")

        try:
            # 步骤1: 验证参数格式
            validation_result = cls._validate_range_format(range_expression)
            if not validation_result['valid']:
                return cls._format_error_result(validation_result['error'])

            # 步骤2: 解析范围表达式
            range_info = cls._parse_range_expression(range_expression)
            if not range_info['success']:
                return cls._format_error_result(range_info['error'])

            # 步骤3: 获取当前数据预览
            reader = ExcelReader(file_path)
            current_data_result = reader.get_range(range_expression)
            reader.close()

            if not current_data_result.success:
                return cls._format_error_result(f"无法获取当前数据: {current_data_result.message}")

            current_data = current_data_result.data or []

            # 步骤4: 分析影响范围
            impact_analysis = cls._analyze_operation_impact(
                range_info, current_data, operation_type, preview_data
            )

            # 步骤5: 生成安全执行计划
            safe_execution_plan = cls._generate_safe_execution_plan(
                operation_type, impact_analysis, range_expression
            )

            return {
                'success': True,
                'impact_analysis': impact_analysis,
                'preview_data': current_data,
                'safe_execution_plan': safe_execution_plan,
                'message': f"操作影响评估完成，风险等级: {impact_analysis['operation_risk_level']}"
            }

        except Exception as e:
            error_msg = f"操作影响评估失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)

    @classmethod
    def _parse_range_expression(cls, range_expression: str) -> Dict[str, Any]:
        """解析范围表达式，提取工作表名和范围信息"""
        try:
            if '!' not in range_expression:
                return {'success': False, 'error': '范围表达式必须包含工作表名'}

            sheet_name, range_part = range_expression.split('!', 1)
            sheet_name = sheet_name.strip('\'"')  # 处理带引号的工作表名

            # 解析范围部分
            if ':' in range_part:
                # 范围格式如 A1:C10
                start_cell, end_cell = range_part.split(':', 1)
                start_info = cls._parse_cell_reference(start_cell.strip())
                end_info = cls._parse_cell_reference(end_cell.strip())

                if not start_info['success'] or not end_info['success']:
                    return {'success': False, 'error': '无效的范围格式'}

                return {
                    'success': True,
                    'sheet_name': sheet_name,
                    'start_row': start_info['row'],
                    'start_col': start_info['col'],
                    'end_row': end_info['row'],
                    'end_col': end_info['col'],
                    'is_range': True
                }
            else:
                # 单元格格式如 A1 或 A 或 1
                cell_info = cls._parse_cell_reference(range_part.strip())
                if cell_info['success']:
                    return {
                        'success': True,
                        'sheet_name': sheet_name,
                        'start_row': cell_info['row'],
                        'start_col': cell_info['col'],
                        'end_row': cell_info['row'],
                        'end_col': cell_info['col'],
                        'is_range': False
                    }
                else:
                    return {'success': False, 'error': '无效的单元格引用'}

        except Exception as e:
            return {'success': False, 'error': f'解析范围表达式失败: {str(e)}'}

    @classmethod
    def _parse_cell_reference(cls, cell_ref: str) -> Dict[str, Any]:
        """解析单元格引用，返回行列信息"""
        try:
            import re

            # 匹配单元格格式 (如 A1, B10, AA1)
            cell_pattern = r'^([A-Za-z]+)(\d*)$'
            match = re.match(cell_pattern, cell_ref)

            if match:
                col_str, row_str = match.groups()

                # 转换列字母为数字
                col_num = 0
                for char in col_str.upper():
                    col_num = col_num * 26 + (ord(char) - ord('A') + 1)

                # 处理行号
                if row_str:
                    row_num = int(row_str)
                else:
                    row_num = None  # 仅列引用

                return {
                    'success': True,
                    'col': col_num,
                    'row': row_num
                }

            # 匹配纯数字（仅行引用）
            elif cell_ref.isdigit():
                return {
                    'success': True,
                    'col': None,  # 整行
                    'row': int(cell_ref)
                }

            else:
                return {'success': False, 'error': '无效的单元格引用格式'}

        except Exception as e:
            return {'success': False, 'error': f'解析单元格引用失败: {str(e)}'}

    @classmethod
    def _analyze_operation_impact(
        cls,
        range_info: Dict[str, Any],
        current_data: List[List[Any]],
        operation_type: str,
        preview_data: Optional[List[List[Any]]]
    ) -> Dict[str, Any]:
        """分析操作影响"""
        try:
            # 计算影响范围
            if range_info['start_row'] and range_info['end_row']:
                affected_rows = range_info['end_row'] - range_info['start_row'] + 1
            else:
                affected_rows = 1  # 默认影响1行

            if range_info['start_col'] and range_info['end_col']:
                affected_columns = range_info['end_col'] - range_info['start_col'] + 1
            else:
                affected_columns = 1  # 默认影响1列

            affected_cells = affected_rows * affected_columns

            # 分析现有数据
            non_empty_cells = 0
            data_types = {}
            formula_count = 0

            for row_idx, row in enumerate(current_data):
                for col_idx, cell in enumerate(row):
                    if cell is not None:
                        non_empty_cells += 1

                        # 统计数据类型
                        if hasattr(cell, 'value'):
                            # CellInfo对象
                            cell_value = cell.value
                            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                                formula_count += 1
                        else:
                            # 普通值
                            cell_value = cell

                        if cell_value is not None:
                            data_type = type(cell_value).__name__
                            data_types[data_type] = data_types.get(data_type, 0) + 1

            # 现有数据摘要
            existing_data_summary = {
                'non_empty_cells': non_empty_cells,
                'total_cells': affected_cells,
                'data_completeness': f"{(non_empty_cells / affected_cells * 100):.1f}%" if affected_cells > 0 else "0%",
                'formula_count': formula_count,
                'data_types': data_types
            }

            # 评估风险等级
            risk_level, warnings, recommendations = cls._assess_risk_level(
                operation_type, affected_cells, non_empty_cells, formula_count, preview_data
            )

            return {
                'affected_cells': affected_cells,
                'affected_rows': affected_rows,
                'affected_columns': affected_columns,
                'non_empty_cells': non_empty_cells,
                'existing_data_summary': existing_data_summary,
                'operation_risk_level': risk_level,
                'warnings': warnings,
                'recommendations': recommendations
            }

        except Exception as e:
            logger.error(f"{cls._LOG_PREFIX} 分析操作影响失败: {str(e)}")
            return {
                'affected_cells': 0,
                'affected_rows': 0,
                'affected_columns': 0,
                'non_empty_cells': 0,
                'operation_risk_level': 'high',
                'warnings': [f"影响分析失败: {str(e)}"],
                'recommendations': ['建议取消操作或检查参数']
            }

    @classmethod
    def _assess_risk_level(
        cls,
        operation_type: str,
        affected_cells: int,
        non_empty_cells: int,
        formula_count: int,
        preview_data: Optional[List[List[Any]]]
    ) -> tuple[str, List[str], List[str]]:
        """评估操作风险等级和生成建议"""
        warnings = []
        recommendations = []

        # 基于操作类型的基础风险评估
        risk_factors = {
            'update': 2,      # 更新操作中等风险
            'delete': 4,      # 删除操作高风险
            'insert': 1,      # 插入操作低风险
            'format': 1       # 格式化操作低风险
        }

        base_risk = risk_factors.get(operation_type, 3)

        # 基于影响范围调整风险
        if affected_cells > 1000:
            base_risk += 2
            warnings.append("操作范围超过1000个单元格，属于大规模操作")
        elif affected_cells > 100:
            base_risk += 1
            warnings.append("操作范围超过100个单元格，请谨慎确认")

        # 基于现有数据调整风险
        if non_empty_cells > 0:
            data_ratio = non_empty_cells / affected_cells if affected_cells > 0 else 0
            if data_ratio > 0.8:
                base_risk += 2
                warnings.append(f"目标区域包含大量现有数据（{non_empty_cells}个非空单元格）")
            elif data_ratio > 0.5:
                base_risk += 1
                warnings.append(f"目标区域包含较多现有数据（{non_empty_cells}个非空单元格）")

        # 基于公式数量调整风险
        if formula_count > 0:
            base_risk += min(formula_count, 2)  # 最多增加2级风险
            warnings.append(f"目标区域包含{formula_count}个公式，操作可能影响计算结果")

        # 确定最终风险等级
        if base_risk <= 2:
            risk_level = "low"
            recommendations.append("操作风险较低，可以安全执行")
        elif base_risk <= 4:
            risk_level = "medium"
            recommendations.append("操作风险中等，建议先预览再执行")
        elif base_risk <= 6:
            risk_level = "high"
            recommendations.append("操作风险较高，强烈建议创建备份")
        else:
            risk_level = "critical"
            recommendations.append("操作风险极高，必须创建备份并获得确认")

        # 操作特定建议
        if operation_type == "update" and non_empty_cells > 0:
            recommendations.append("建议使用insert_mode=True避免覆盖现有数据")
        elif operation_type == "delete":
            recommendations.append("删除操作不可逆，请三思而后行")
        elif operation_type == "format" and formula_count > 0:
            recommendations.append("格式化可能影响公式显示，请谨慎操作")

        # 如果没有警告，添加默认信息
        if not warnings:
            warnings.append("目标区域当前无数据，操作风险较低")

        return risk_level, warnings, recommendations

    @classmethod
    def _generate_safe_execution_plan(
        cls,
        operation_type: str,
        impact_analysis: Dict[str, Any],
        range_expression: str
    ) -> Dict[str, Any]:
        """生成安全执行计划"""
        risk_level = impact_analysis['operation_risk_level']

        plan = {
            'operation_type': operation_type,
            'range_expression': range_expression,
            'risk_level': risk_level,
            'required_steps': [],
            'safety_measures': [],
            'rollback_available': True
        }

        # 基础步骤
        plan['required_steps'].append("1. 验证文件路径和权限")
        plan['required_steps'].append("2. 检查文件锁定状态")
        plan['required_steps'].append("3. 确认操作范围和数据")

        # 基于风险等级添加安全措施
        if risk_level in ['medium', 'high', 'critical']:
            plan['safety_measures'].append("创建操作前备份")
            plan['required_steps'].append("4. 创建自动备份文件")

        if risk_level in ['high', 'critical']:
            plan['safety_measures'].append("要求用户明确确认")
            plan['required_steps'].append("5. 显示操作预览和影响摘要")
            plan['required_steps'].append("6. 等待用户确认")

        if risk_level == 'critical':
            plan['safety_measures'].append("多重确认机制")
            plan['required_steps'].append("7. 二次确认操作意图")
            plan['required_steps'].append("8. 执行操作")
            plan['required_steps'].append("9. 验证操作结果")

        # 操作特定步骤
        if operation_type == "update":
            plan['required_steps'].append(f"10. 使用安全的insert_mode=True模式更新数据")
        elif operation_type == "delete":
            plan['safety_measures'].append("删除操作不可逆，将显示详细警告")

        # 回滚计划
        plan['rollback_steps'] = [
            "1. 停止当前操作",
            "2. 从备份文件恢复数据",
            "3. 验证恢复结果",
            "4. 记录操作日志"
        ]

        return plan

    @classmethod
    def _generate_safety_warnings(
        cls,
        operation_type: str,
        range_expression: str,
        impact_analysis: Dict[str, Any],
        preview_data: Optional[List[List[Any]]],
        insert_mode: bool = True
    ) -> Dict[str, Any]:
        """生成详细的安全警告信息"""
        try:
            risk_level = impact_analysis['operation_risk_level']
            affected_cells = impact_analysis['affected_cells']
            non_empty_cells = impact_analysis['non_empty_cells']
            existing_data_summary = impact_analysis['existing_data_summary']

            warnings = []
            critical_warnings = []
            preventative_measures = []
            visual_indicators = []

            # 基础警告信息
            if risk_level == 'critical':
                critical_warnings.append("🚨 极高风险操作：可能导致数据永久丢失！")
                warnings.append("此操作影响范围巨大，强烈建议取消并重新评估")
                visual_indicators.append("🔴 红色警告：高风险操作")
            elif risk_level == 'high':
                critical_warnings.append("⚠️ 高风险操作：可能影响重要数据")
                warnings.append("操作风险较高，建议创建备份后再执行")
                visual_indicators.append("🟡 黄色警告：中高风险操作")
            elif risk_level == 'medium':
                warnings.append("操作存在一定风险，建议仔细检查参数")
                visual_indicators.append("🟠 橙色提示：中等风险操作")
            else:
                visual_indicators.append("🟢 绿色标识：低风险操作")

            # 大规模操作警告
            if affected_cells > 1000:
                critical_warnings.append(f"📊 超大规模操作：将影响 {affected_cells:,} 个单元格")
                preventative_measures.append("建议分批处理，每次操作不超过100个单元格")
            elif affected_cells > 100:
                warnings.append(f"📋 大规模操作：将影响 {affected_cells:,} 个单元格")

            # 数据覆盖警告
            if non_empty_cells > 0:
                data_ratio = non_empty_cells / affected_cells if affected_cells > 0 else 0
                if data_ratio > 0.8:
                    critical_warnings.append(f"💥 大量数据覆盖：目标区域 {non_empty_cells:,} 个单元格包含数据")
                    if not insert_mode:
                        critical_warnings.append("🔄 覆盖模式将永久删除现有数据！")
                        preventative_measures.append("强烈建议使用 insert_mode=True 避免数据丢失")
                else:
                    warnings.append(f"📝 数据影响：将覆盖 {non_empty_cells:,} 个现有数据单元格")

            # 公式警告
            formula_count = existing_data_summary.get('formula_count', 0)
            if formula_count > 0:
                if formula_count > 10:
                    critical_warnings.append(f"🧮 大量公式：目标区域包含 {formula_count} 个公式")
                else:
                    warnings.append(f"🧮 公式影响：目标区域包含 {formula_count} 个公式")
                preventative_measures.append("操作可能影响公式计算结果，建议检查依赖关系")

            # 操作类型特定警告
            if operation_type == 'update':
                if not insert_mode:
                    critical_warnings.append("🔄 覆盖模式警告：现有数据将被永久替换")
                    preventative_measures.append("考虑使用插入模式 (insert_mode=True) 保护现有数据")
                else:
                    preventative_measures.append("插入模式相对安全，将向下移动现有数据")
            elif operation_type == 'delete':
                critical_warnings.append("🗑️ 删除操作警告：数据删除后无法撤销")
                preventative_measures.append("删除前务必确认已创建备份")
            elif operation_type == 'format':
                if formula_count > 0:
                    warnings.append("🎨 格式化可能影响公式显示效果")

            # 数据完整性警告
            if preview_data:
                preview_rows = len(preview_data)
                preview_cols = max(len(row) for row in preview_data) if preview_data else 0
                expected_cells = preview_rows * preview_cols

                if expected_cells != affected_cells:
                    warnings.append(f"📏 数据范围不匹配：预览数据({expected_cells})与目标范围({affected_cells})大小不一致")
                    preventative_measures.append("请检查数据范围是否正确")

            # 生成可视化表示
            visual_representation = cls._generate_operation_visualization(
                range_expression, impact_analysis
            )

            return {
                'risk_level': risk_level,
                'critical_warnings': critical_warnings,
                'general_warnings': warnings,
                'preventative_measures': preventative_measures,
                'visual_indicators': visual_indicators,
                'visual_representation': visual_representation,
                'impact_summary': {
                    'affected_cells': f"{affected_cells:,}",
                    'non_empty_cells': f"{non_empty_cells:,}",
                    'data_coverage_ratio': f"{(non_empty_cells / affected_cells * 100):.1f}%" if affected_cells > 0 else "0%",
                    'operation_mode': '安全插入模式' if insert_mode else '风险覆盖模式'
                }
            }

        except Exception as e:
            logger.error(f"{cls._LOG_PREFIX} 生成安全警告失败: {str(e)}")
            return {
                'risk_level': 'unknown',
                'critical_warnings': [f"警告生成失败: {str(e)}"],
                'general_warnings': [],
                'preventative_measures': ['建议取消操作并检查参数'],
                'visual_indicators': ['❓ 未知风险'],
                'visual_representation': '',
                'impact_summary': {}
            }

    @classmethod
    def _generate_operation_visualization(
        cls,
        range_expression: str,
        impact_analysis: Dict[str, Any]
    ) -> str:
        """生成操作范围的可视化表示"""
        try:
            affected_cells = impact_analysis['affected_cells']
            non_empty_cells = impact_analysis['non_empty_cells']
            risk_level = impact_analysis['operation_risk_level']

            # 简单的文本可视化
            risk_symbols = {
                'low': '🟢',
                'medium': '🟠',
                'high': '🟡',
                'critical': '🔴'
            }

            symbol = risk_symbols.get(risk_level, '❓')

            # 创建可视化网格
            if affected_cells <= 50:
                # 小范围操作，显示详细网格
                grid_size = min(int(affected_cells ** 0.5) + 1, 10)
                grid = []
                for i in range(grid_size):
                    row = []
                    for j in range(grid_size):
                        if i * grid_size + j < affected_cells:
                            if i * grid_size + j < non_empty_cells:
                                row.append('█')  # 有数据的单元格
                            else:
                                row.append('░')  # 空单元格
                        else:
                            row.append(' ')   # 范围外
                    grid.append(''.join(row))

                visualization = f"\n{symbol} 操作范围可视化 ({affected_cells} 个单元格):\n"
                visualization += "┌" + "─" * len(grid[0]) + "┐\n"
                for row in grid:
                    visualization += "│" + row + "│\n"
                visualization += "└" + "─" * len(grid[0]) + "┘\n"
                visualization += f"█ = 有数据 ({non_empty_cells})  ░ = 空单元格 ({affected_cells - non_empty_cells})"
            else:
                # 大范围操作，显示统计信息
                empty_cells = affected_cells - non_empty_cells
                data_ratio = (non_empty_cells / affected_cells * 100) if affected_cells > 0 else 0

                bar_length = 20
                filled_length = int(bar_length * data_ratio / 100)
                bar = '█' * filled_length + '░' * (bar_length - filled_length)

                visualization = f"\n{symbol} 大规模操作统计:\n"
                visualization += f"总单元格: {affected_cells:,}\n"
                visualization += f"有数据: {non_empty_cells:,} ({data_ratio:.1f}%)\n"
                visualization += f"空白: {empty_cells:,} ({100-data_ratio:.1f}%)\n"
                visualization += f"数据密度: [{bar}] {data_ratio:.1f}%"

            return visualization

        except Exception as e:
            logger.error(f"{cls._LOG_PREFIX} 生成操作可视化失败: {str(e)}")
            return f"❓ 无法生成可视化: {str(e)}"

    @classmethod
    def check_file_status(cls, file_path: str) -> Dict[str, Any]:
        """
        @intention 检查Excel文件的状态，验证文件是否可安全操作

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)

        Returns:
            Dict: 包含文件状态检查结果
            {
                'success': bool,
                'file_status': {
                    'exists': bool,              # 文件是否存在
                    'readable': bool,            # 文件是否可读
                    'writable': bool,           # 文件是否可写
                    'locked': bool,              # 文件是否被锁定
                    'locked_by': Optional[str],  # 锁定文件的程序
                    'file_size': int,            # 文件大小（字节）
                    'last_modified': str,        # 最后修改时间
                    'file_format': str,          # 文件格式
                    'backup_available': bool,    # 是否有可用备份
                    'corruption_risk': str,      # 损坏风险等级
                    'safety_recommendations': List[str]
                },
                'permissions': dict,            # 文件权限信息
                'security_warnings': List[str]   # 安全警告
            }
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始检查文件状态: {file_path}")

        try:
            import os
            import time
            from pathlib import Path

            file_status = {}
            security_warnings = []
            safety_recommendations = []

            # 步骤1: 检查文件是否存在
            if not os.path.exists(file_path):
                return {
                    'success': False,
                    'error': f'文件不存在: {file_path}',
                    'file_status': {
                        'exists': False,
                        'readable': False,
                        'writable': False,
                        'locked': True,
                        'corruption_risk': 'high'
                    },
                    'security_warnings': ['文件不存在，无法执行操作'],
                    'safety_recommendations': ['请检查文件路径是否正确']
                }

            file_status['exists'] = True

            # 步骤2: 检查文件基本信息
            path_obj = Path(file_path)
            file_status['file_size'] = path_obj.stat().st_size
            file_status['last_modified'] = time.strftime(
                '%Y-%m-%d %H:%M:%S',
                time.localtime(path_obj.stat().st_mtime)
            )

            # 检查文件格式
            file_extension = path_obj.suffix.lower()
            valid_formats = ['.xlsx', '.xlsm']
            if file_extension not in valid_formats:
                security_warnings.append(f"不支持的文件格式: {file_extension}")
                file_status['corruption_risk'] = 'high'
                safety_recommendations.append("请使用.xlsx或.xlsm格式的Excel文件")
            else:
                file_status['file_format'] = file_extension
                file_status['corruption_risk'] = 'low'

            # 步骤3: 检查文件权限
            file_status['readable'] = os.access(file_path, os.R_OK)
            file_status['writable'] = os.access(file_path, os.W_OK)

            if not file_status['readable']:
                security_warnings.append("文件不可读，可能权限不足")
                file_status['corruption_risk'] = 'high'
                safety_recommendations.append("检查文件读取权限")

            if not file_status['writable']:
                security_warnings.append("文件不可写，可能权限不足或被占用")
                safety_recommendations.append("检查文件写入权限或关闭占用程序")

            # 步骤4: 检查文件锁定状态
            lock_info = cls._check_file_lock_status(file_path)
            file_status.update(lock_info)

            if file_status.get('locked', False):
                locked_by = file_status.get('locked_by', '未知程序')
                security_warnings.append(f"文件已被 {locked_by} 锁定，无法安全操作")
                file_status['corruption_risk'] = 'critical'
                safety_recommendations.extend([
                    f"关闭 {locked_by} 程序后重试",
                    "或创建文件副本进行操作"
                ])

            # 步骤5: 检查文件大小和完整性
            size_mb = file_status['file_size'] / (1024 * 1024)
            if size_mb > 50:  # 大于50MB
                security_warnings.append(f"文件较大({size_mb:.1f}MB)，操作可能较慢")
                safety_recommendations.append("考虑分批处理或优化文件大小")
            elif size_mb == 0:
                security_warnings.append("文件为空，可能已损坏")
                file_status['corruption_risk'] = 'high'
                safety_recommendations.append("检查文件完整性或使用备份文件")

            # 步骤6: 尝试读取文件验证完整性
            integrity_check = cls._verify_file_integrity(file_path)
            if not integrity_check['valid']:
                security_warnings.append("文件完整性验证失败")
                file_status['corruption_risk'] = 'high'
                safety_recommendations.extend(integrity_check['recommendations'])

            # 步骤7: 检查备份可用性
            backup_info = cls._check_backup_availability(file_path)
            file_status['backup_available'] = backup_info['available']
            if not backup_info['available']:
                safety_recommendations.append("建议在操作前创建手动备份")

            # 步骤8: 生成权限信息
            permissions = {
                'can_read': file_status['readable'],
                'can_write': file_status['writable'],
                'can_execute': os.access(file_path, os.X_OK),
                'owner': path_obj.stat().st_uid if hasattr(path_obj.stat(), 'st_uid') else None,
                'group': path_obj.stat().st_gid if hasattr(path_obj.stat(), 'st_gid') else None
            }

            # 步骤9: 综合安全评估
            if not security_warnings:
                safety_recommendations.append("文件状态良好，可以安全操作")

            return {
                'success': True,
                'file_status': file_status,
                'permissions': permissions,
                'security_warnings': security_warnings,
                'safety_recommendations': safety_recommendations,
                'message': f"文件状态检查完成，风险等级: {file_status['corruption_risk']}"
            }

        except Exception as e:
            error_msg = f"文件状态检查失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return {
                'success': False,
                'error': error_msg,
                'file_status': {
                    'exists': False,
                    'locked': True,
                    'corruption_risk': 'unknown'
                },
                'security_warnings': [error_msg],
                'safety_recommendations': ['请检查文件路径和权限']
            }

    @classmethod
    def _check_file_lock_status(cls, file_path: str) -> Dict[str, Any]:
        """检查文件锁定状态"""
        try:
            import os
            import stat

            lock_info = {
                'locked': False,
                'locked_by': None,
                'lock_type': None
            }

            # 方法1: 尝试以独占模式打开文件
            try:
                # Windows系统检查
                if os.name == 'nt':
                    import msvcrt
                    try:
                        # 尝试以独占模式打开
                        file_handle = open(file_path, 'r+b')
                        file_handle.close()
                    except IOError:
                        lock_info['locked'] = True
                        lock_info['locked_by'] = 'Excel或其他程序'
                        lock_info['lock_type'] = 'exclusive'
                else:
                    # Unix/Linux系统检查
                    # 尝试重命名文件来检测锁定
                    test_path = file_path + '.test_lock'
                    try:
                        os.rename(file_path, test_path)
                        os.rename(test_path, file_path)
                    except OSError:
                        lock_info['locked'] = True
                        lock_info['locked_by'] = '系统进程'
                        lock_info['lock_type'] = 'filesystem'

            except Exception as e:
                logger.debug(f"{cls._LOG_PREFIX} 锁定检查方法1失败: {str(e)}")

            # 方法2: 检查临时锁定文件
            lock_file_patterns = [
                file_path + '.lock',
                file_path + '~',
                file_path.replace('.xlsx', '.~$xlsx'),
                file_path.replace('.xlsm', '.~$xlsm')
            ]

            for pattern in lock_file_patterns:
                if os.path.exists(pattern):
                    lock_info['locked'] = True
                    lock_info['locked_by'] = 'Excel锁定文件'
                    lock_info['lock_type'] = 'temp_file'
                    break

            # 方法3: 检查文件权限变化
            try:
                current_stat = os.stat(file_path)
                if hasattr(current_stat, 'st_mtime'):
                    # 检查文件是否最近被修改过（可能正在使用）
                    import time
                    time_diff = time.time() - current_stat.st_mtime
                    if time_diff < 60:  # 1分钟内修改过
                        if not lock_info['locked']:
                            lock_info['locked'] = True
                            lock_info['locked_by'] = '可能被程序占用'
                            lock_info['lock_type'] = 'recent_access'
            except Exception as e:
                logger.debug(f"{cls._LOG_PREFIX} 锁定检查方法3失败: {str(e)}")

            return lock_info

        except Exception as e:
            logger.error(f"{cls._LOG_PREFIX} 文件锁状态检查失败: {str(e)}")
            return {
                'locked': True,
                'locked_by': '未知',
                'lock_type': 'check_failed'
            }

    @classmethod
    def _verify_file_integrity(cls, file_path: str) -> Dict[str, Any]:
        """验证文件完整性"""
        try:
            from ..core.excel_reader import ExcelReader

            integrity_result = {
                'valid': False,
                'error': None,
                'recommendations': []
            }

            # 尝试打开并读取文件
            try:
                reader = ExcelReader(file_path)
                workbook = reader._get_workbook(read_only=True)

                # 检查工作簿是否可以正常访问
                if workbook.worksheets:
                    integrity_result['valid'] = True
                else:
                    integrity_result['error'] = '文件不包含任何工作表'
                    integrity_result['recommendations'].append('检查文件是否为有效的Excel文件')

                reader.close()

            except Exception as e:
                integrity_result['error'] = str(e)
                if "zip" in str(e).lower():
                    integrity_result['recommendations'].append('文件可能已损坏，尝试使用Excel修复功能')
                elif "permission" in str(e).lower():
                    integrity_result['recommendations'].append('检查文件权限，确保有读取权限')
                else:
                    integrity_result['recommendations'].append('尝试使用Excel打开文件验证完整性')

            return integrity_result

        except Exception as e:
            return {
                'valid': False,
                'error': f'完整性验证失败: {str(e)}',
                'recommendations': ['文件可能严重损坏，建议使用备份文件']
            }

    @classmethod
    def _check_backup_availability(cls, file_path: str) -> Dict[str, Any]:
        """检查备份文件可用性"""
        try:
            import os
            from pathlib import Path

            backup_patterns = [
                file_path.replace('.xlsx', '.backup.xlsx'),
                file_path.replace('.xlsx', '.bak.xlsx'),
                file_path.replace('.xlsm', '.backup.xlsm'),
                file_path.replace('.xlsm', '.bak.xlsm'),
                file_path + '.backup',
                file_path + '.bak'
            ]

            # 检查同目录下的备份文件
            path_obj = Path(file_path)
            parent_dir = path_obj.parent

            # 查找最近修改的备份文件
            available_backups = []
            for pattern in backup_patterns:
                if os.path.exists(pattern):
                    backup_stat = os.stat(pattern)
                    available_backups.append({
                        'path': pattern,
                        'size': backup_stat.st_size,
                        'modified': backup_stat.st_mtime
                    })

            # 检查Excel自动创建的临时文件
            temp_patterns = [
                path_obj.stem + '~$' + path_obj.suffix,
                path_obj.stem + '.~' + path_obj.suffix
            ]

            for temp_file in parent_dir.glob('*'):
                if temp_file.name.startswith(path_obj.stem) and any(x in temp_file.name for x in ['~$', '.~']):
                    if temp_file.is_file():
                        temp_stat = temp_file.stat()
                        available_backups.append({
                            'path': str(temp_file),
                            'size': temp_stat.st_size,
                            'modified': temp_stat.st_mtime,
                            'type': 'auto_recovery'
                        })

            return {
                'available': len(available_backups) > 0,
                'backup_count': len(available_backups),
                'backups': available_backups[:5]  # 最多返回5个最近的备份
            }

        except Exception as e:
            logger.error(f"{cls._LOG_PREFIX} 备份检查失败: {str(e)}")
            return {
                'available': False,
                'backup_count': 0,
                'backups': []
            }

    @classmethod
    def cancel_operation(
        cls,
        operation_id: str,
        reason: str = "用户取消操作",
        user_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        @intention 取消正在进行的Excel操作

        Args:
            operation_id: 操作ID
            reason: 取消原因
            user_id: 用户ID（用于权限验证）

        Returns:
            Dict: 取消结果
        """
        try:
            operation_manager = OperationManager()

            # 验证用户权限（如果提供了user_id）
            if user_id:
                operation_status = operation_manager.get_operation_status(operation_id)
                if operation_status.get('user_id') != user_id:
                    return {
                        'success': False,
                        'error': 'PERMISSION_DENIED',
                        'message': '您没有权限取消此操作'
                    }

            # 执行取消
            cancel_result = operation_manager.cancel_operation(operation_id, reason)

            if cancel_result['success']:
                return {
                    'success': True,
                    'operation_id': operation_id,
                    'message': cancel_result['message'],
                    'cancelled_at': cancel_result['cancelled_at'],
                    'reason': reason
                }
            else:
                return {
                    'success': False,
                    'error': cancel_result['error'],
                    'message': cancel_result['message']
                }

        except Exception as e:
            error_msg = f"取消操作失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return {
                'success': False,
                'error': error_msg,
                'message': '取消操作时发生异常'
            }

    @classmethod
    def get_operation_status(
        cls,
        operation_id: str,
        user_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        @intention 获取操作状态和进度

        Args:
            operation_id: 操作ID
            user_id: 用户ID（用于权限验证）

        Returns:
            Dict: 操作状态信息
        """
        try:
            import time

            operation_manager = OperationManager()
            operation_status = operation_manager.get_operation_status(operation_id)

            if operation_status.get('status') == 'not_found':
                return {
                    'success': False,
                    'error': 'OPERATION_NOT_FOUND',
                    'message': f'操作不存在: {operation_id}'
                }

            # 验证用户权限（如果提供了user_id）
            if user_id and operation_status.get('user_id') != user_id:
                return {
                    'success': False,
                    'error': 'PERMISSION_DENIED',
                    'message': '您没有权限查看此操作的状态'
                }

            # 格式化状态信息
            status_info = {
                'success': True,
                'operation_id': operation_id,
                'operation_type': operation_status.get('type'),
                'file_path': operation_status.get('file_path'),
                'range_expression': operation_status.get('range_expression'),
                'status': operation_status.get('status'),
                'progress': operation_status.get('progress', 0),
                'cancellable': operation_status.get('cancellable', True),
                'start_time': operation_status.get('start_time'),
                'end_time': operation_status.get('end_time'),
                'duration': None
            }

            # 计算持续时间
            if operation_status.get('start_time'):
                if operation_status.get('end_time'):
                    status_info['duration'] = operation_status['end_time'] - operation_status['start_time']
                else:
                    status_info['duration'] = time.time() - operation_status['start_time']

            # 添加状态特定的信息
            if operation_status.get('status') == 'cancelled':
                status_info['cancel_reason'] = operation_status.get('cancel_reason')
                status_info['message'] = '操作已被取消'
            elif operation_status.get('status') == 'completed':
                status_info['message'] = '操作已完成'
                status_info['result'] = operation_status.get('result')
            elif operation_status.get('status') == 'failed':
                status_info['error'] = operation_status.get('error')
                status_info['message'] = '操作失败'
            elif operation_status.get('status') == 'pending':
                status_info['message'] = '操作等待执行'
            else:
                status_info['message'] = '操作正在进行中'

            # 添加当前消息（如果有）
            if operation_status.get('current_message'):
                status_info['current_message'] = operation_status['current_message']

            return status_info

        except Exception as e:
            error_msg = f"获取操作状态失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return {
                'success': False,
                'error': error_msg,
                'message': '获取操作状态时发生异常'
            }

    @classmethod
    def list_operations(
        cls,
        user_id: Optional[str] = None,
        include_history: bool = False,
        limit: int = 50
    ) -> Dict[str, Any]:
        """
        @intention 列出操作状态和历史

        Args:
            user_id: 用户ID（过滤特定用户的操作）
            include_history: 是否包含历史操作
            limit: 返回结果数量限制

        Returns:
            Dict: 操作列表
        """
        try:
            operation_manager = OperationManager()

            # 获取活跃操作
            active_operations = operation_manager.list_active_operations(user_id)

            result = {
                'success': True,
                'active_operations': active_operations,
                'active_count': len(active_operations),
                'message': f"当前有 {len(active_operations)} 个活跃操作"
            }

            # 如果包含历史，添加历史操作
            if include_history:
                history_operations = operation_manager.list_operation_history(limit, user_id)
                result['history_operations'] = history_operations
                result['history_count'] = len(history_operations)
                result['message'] += f"，最近 {len(history_operations)} 个历史操作"

            return result

        except Exception as e:
            error_msg = f"列出操作失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return {
                'success': False,
                'error': error_msg,
                'message': '列出操作时发生异常'
            }

    @classmethod
    def _create_cancellable_operation(
        cls,
        operation_type: str,
        file_path: str,
        range_expression: str,
        user_id: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        @intention 创建可取消的操作
        """
        operation_manager = OperationManager()
        return operation_manager.start_operation(
            operation_type, file_path, range_expression, user_id, metadata
        )

    @classmethod
    def _check_operation_cancelled(cls, operation_id: str) -> bool:
        """
        @intention 检查操作是否已被取消
        """
        operation_manager = OperationManager()
        return operation_manager.check_cancelled(operation_id)

    @classmethod
    def _update_operation_progress(cls, operation_id: str, progress: int, message: Optional[str] = None):
        """
        @intention 更新操作进度
        """
        operation_manager = OperationManager()
        operation_manager.update_progress(operation_id, progress, message)

    @classmethod
    def _complete_operation(cls, operation_id: str, result: Optional[Dict[str, Any]] = None):
        """
        @intention 标记操作完成
        """
        operation_manager = OperationManager()
        operation_manager.complete_operation(operation_id, result)

    @classmethod
    def _fail_operation(cls, operation_id: str, error: str):
        """
        @intention 标记操作失败
        """
        operation_manager = OperationManager()
        operation_manager.fail_operation(operation_id, error)

    @classmethod
    def confirm_operation(
        cls,
        file_path: str,
        range_expression: str,
        operation_type: str,
        preview_data: Optional[List[List[Any]]] = None,
        operation_params: Optional[Dict[str, Any]] = None,
        confirmation_token: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        @intention 创建操作确认步骤，为危险操作提供明确的确认机制

        Args:
            file_path: Excel文件路径 (.xlsx/.xlsm)
            range_expression: 范围表达式，必须包含工作表名
            operation_type: 操作类型 ('update', 'delete', 'insert', 'format')
            preview_data: 预览数据（对于更新操作）
            operation_params: 操作参数（如insert_mode等）
            confirmation_token: 确认令牌（用于验证用户意图）

        Returns:
            Dict: 包含确认流程的结果
            {
                'success': bool,
                'confirmation_required': bool,
                'operation_summary': dict,      # 操作摘要
                'risk_assessment': dict,       # 风险评估
                'confirmation_steps': List[str], # 确认步骤
                'user_confirmation': dict,     # 用户确认信息
                'safety_guarantees': List[str] # 安全保证
            }
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始操作确认流程: {operation_type} - {range_expression}")

        try:
            # 步骤1: 生成操作摘要
            operation_summary = cls._generate_operation_summary(
                file_path, range_expression, operation_type, preview_data, operation_params
            )

            # 步骤2: 执行风险评估
            impact_assessment = cls.assess_operation_impact(
                file_path, range_expression, operation_type, preview_data
            )

            if not impact_assessment['success']:
                return {
                    'success': False,
                    'error': f"风险评估失败: {impact_assessment.get('error', '未知错误')}",
                    'confirmation_required': True
                }

            # 步骤3: 生成确认步骤
            confirmation_steps = cls._generate_confirmation_steps(
                operation_type, impact_assessment['impact_analysis']
            )

            # 步骤4: 验证确认令牌（如果提供）
            confirmation_valid = False
            if confirmation_token:
                confirmation_valid = cls._validate_confirmation_token(
                    confirmation_token, operation_summary, impact_assessment
                )

            # 步骤5: 生成安全保证
            safety_guarantees = cls._generate_safety_guarantees(
                operation_type, impact_assessment['impact_analysis']
            )

            # 步骤6: 确定是否需要确认
            risk_level = impact_assessment['impact_analysis']['operation_risk_level']
            requires_confirmation = risk_level in ['medium', 'high', 'critical']

            # 步骤7: 构建确认结果
            result = {
                'success': True,
                'confirmation_required': requires_confirmation,
                'operation_summary': operation_summary,
                'risk_assessment': impact_assessment,
                'confirmation_steps': confirmation_steps,
                'safety_guarantees': safety_guarantees,
                'risk_level': risk_level
            }

            # 如果提供了确认令牌，添加验证结果
            if confirmation_token:
                result['user_confirmation'] = {
                    'token_valid': confirmation_valid,
                    'confirmation_status': 'confirmed' if confirmation_valid else 'invalid_token',
                    'message': '用户确认已验证，操作可以执行' if confirmation_valid else '确认令牌无效，请重新确认'
                }
                result['can_proceed'] = confirmation_valid
            else:
                result['user_confirmation'] = {
                    'token_valid': False,
                    'confirmation_status': 'pending',
                    'message': '等待用户确认操作'
                }
                result['can_proceed'] = not requires_confirmation

            return result

        except Exception as e:
            error_msg = f"操作确认流程失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return {
                'success': False,
                'error': error_msg,
                'confirmation_required': True,
                'message': '确认流程异常，建议取消操作'
            }

    @classmethod
    def _generate_operation_summary(
        cls,
        file_path: str,
        range_expression: str,
        operation_type: str,
        preview_data: Optional[List[List[Any]]],
        operation_params: Optional[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """生成操作摘要"""
        try:
            import os
            from pathlib import Path

            # 文件信息
            path_obj = Path(file_path)
            file_info = {
                'name': path_obj.name,
                'directory': str(path_obj.parent),
                'size': f"{path_obj.stat().st_size / (1024*1024):.2f}MB" if path_obj.exists() else "文件不存在"
            }

            # 操作信息
            operation_info = {
                'type': operation_type,
                'type_description': cls._get_operation_description(operation_type),
                'target_range': range_expression,
                'parameters': operation_params or {}
            }

            # 数据信息
            data_info = {}
            if preview_data:
                data_info = {
                    'preview_rows': len(preview_data),
                    'preview_columns': max(len(row) for row in preview_data) if preview_data else 0,
                    'total_cells': len(preview_data) * max(len(row) for row in preview_data) if preview_data else 0,
                    'has_data': len(preview_data) > 0 and any(any(cell for cell in row) for row in preview_data)
                }

            # 安全信息
            safety_info = {
                'insert_mode': operation_params.get('insert_mode', True) if operation_params else True,
                'preserve_formulas': operation_params.get('preserve_formulas', True) if operation_params else True,
                'backup_recommended': operation_type in ['delete', 'update']
            }

            return {
                'file_info': file_info,
                'operation_info': operation_info,
                'data_info': data_info,
                'safety_info': safety_info,
                'timestamp': cls._get_current_timestamp()
            }

        except Exception as e:
            logger.error(f"{cls._LOG_PREFIX} 生成操作摘要失败: {str(e)}")
            return {
                'file_info': {'name': '未知', 'directory': '未知', 'size': '未知'},
                'operation_info': {'type': operation_type, 'target_range': range_expression},
                'error': f"摘要生成失败: {str(e)}"
            }

    @classmethod
    def _get_operation_description(cls, operation_type: str) -> str:
        """获取操作类型描述"""
        descriptions = {
            'update': '更新数据',
            'delete': '删除数据',
            'insert': '插入数据',
            'format': '格式化单元格',
            'merge': '合并单元格',
            'unmerge': '取消合并',
            'create_sheet': '创建工作表',
            'delete_sheet': '删除工作表'
        }
        return descriptions.get(operation_type, f'{operation_type}操作')

    @classmethod
    def _generate_confirmation_steps(
        cls,
        operation_type: str,
        impact_analysis: Dict[str, Any]
    ) -> List[str]:
        """生成确认步骤"""
        risk_level = impact_analysis['operation_risk_level']
        affected_cells = impact_analysis['affected_cells']
        non_empty_cells = impact_analysis['non_empty_cells']

        steps = []

        # 基础确认步骤
        steps.append("1. 仔细检查操作范围和目标区域")
        steps.append("2. 确认操作类型和参数设置正确")
        steps.append("3. 查看影响分析了解操作后果")

        # 基于风险等级的步骤
        if risk_level in ['medium', 'high', 'critical']:
            steps.append("4. 检查现有数据，确认覆盖范围可接受")
            steps.append("5. 确认重要数据已备份")

        if risk_level in ['high', 'critical']:
            steps.append("6. 二次确认操作意图和影响范围")
            steps.append("7. 确认没有其他程序正在使用该文件")

        if risk_level == 'critical':
            steps.append("8. 最终确认：理解操作不可逆的后果")
            steps.append("9. 确认具备恢复操作的能力")

        # 基于操作类型的特殊步骤
        if operation_type == 'delete':
            steps.append(f"10. 确认删除 {affected_cells} 个单元格的数据不可恢复")
        elif operation_type == 'update' and non_empty_cells > 0:
            steps.append(f"10. 确认覆盖 {non_empty_cells} 个现有数据单元格")
        elif operation_type == 'format':
            steps.append("10. 确认格式化不会影响数据计算")

        return steps

    @classmethod
    def _validate_confirmation_token(
        cls,
        token: str,
        operation_summary: Dict[str, Any],
        impact_assessment: Dict[str, Any]
    ) -> bool:
        """验证确认令牌"""
        try:
            import hashlib
            import json

            # 生成预期的令牌
            token_data = {
                'operation': operation_summary.get('operation_info', {}),
                'risk_level': impact_assessment.get('impact_analysis', {}).get('operation_risk_level'),
                'timestamp': operation_summary.get('timestamp', '')
            }

            # 创建确定性哈希
            token_string = json.dumps(token_data, sort_keys=True, separators=(',', ':'))
            expected_token = hashlib.sha256(token_string.encode()).hexdigest()[:16]

            return token == expected_token

        except Exception as e:
            logger.error(f"{cls._LOG_PREFIX} 确认令牌验证失败: {str(e)}")
            return False

    @classmethod
    def _generate_safety_guarantees(
        cls,
        operation_type: str,
        impact_analysis: Dict[str, Any]
    ) -> List[str]:
        """生成安全保证"""
        guarantees = []
        risk_level = impact_analysis['operation_risk_level']

        # 基础保证
        guarantees.append("✓ 操作前已完成文件状态检查")
        guarantees.append("✓ 已验证文件格式和完整性")
        guarantees.append("✓ 已分析操作影响范围")

        # 基于风险等级的保证
        if risk_level in ['medium', 'high', 'critical']:
            guarantees.append("✓ 大数据操作已触发安全预警")
            guarantees.append("✓ 提供了详细的操作预览")

        if risk_level in ['high', 'critical']:
            guarantees.append("✓ 建议创建操作前备份")
            guarantees.append("✓ 系统已记录操作日志")

        if risk_level == 'critical':
            guarantees.append("✓ 极高风险操作需要多重确认")
            guarantees.append("✓ 提供了完整的回滚方案")

        # 操作特定保证
        if operation_type == 'update':
            guarantees.append("✓ 默认使用安全的插入模式")
        elif operation_type == 'delete':
            guarantees.append("✓ 删除操作已明确标注不可逆")
        elif operation_type == 'format':
            guarantees.append("✓ 格式化操作不会影响数据内容")

        return guarantees

    @classmethod
    def _get_current_timestamp(cls) -> str:
        """获取当前时间戳"""
        import time
        return time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())

    @classmethod
    def get_safe_operation_guidance(
        cls,
        operation_goal: str,
        file_path: str,
        context_info: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        @intention 为LLM提供安全的工具调用序列建议

        Args:
            operation_goal: 操作目标描述（如"更新技能表数据"、"删除重复行"等）
            file_path: 目标Excel文件路径
            context_info: 上下文信息（如用户偏好、历史操作等）

        Returns:
            Dict: 安全操作指导，包含推荐的工具调用序列和安全注意事项
        """
        try:
            guidance = {
                'success': True,
                'operation_goal': operation_goal,
                'file_path': file_path,
                'safety_level': 'standard',
                'recommended_sequence': [],
                'safety_notes': [],
                'risk_assessments': [],
                'alternative_approaches': []
            }

            # 步骤1: 分析操作目标和风险
            risk_analysis = cls._analyze_operation_risk(operation_goal)
            guidance.update(risk_analysis)

            # 步骤2: 根据操作类型生成推荐序列
            if cls._is_data_modification_operation(operation_goal):
                guidance['recommended_sequence'] = cls._get_safe_modification_sequence(
                    file_path, operation_goal, context_info
                )
            elif cls._is_data_analysis_operation(operation_goal):
                guidance['recommended_sequence'] = cls._get_safe_analysis_sequence(
                    file_path, operation_goal, context_info
                )
            elif cls._is_file_management_operation(operation_goal):
                guidance['recommended_sequence'] = cls._get_safe_management_sequence(
                    file_path, operation_goal, context_info
                )
            else:
                guidance['recommended_sequence'] = cls._get_generic_safe_sequence(
                    file_path, operation_goal, context_info
                )

            # 步骤3: 添加安全注意事项
            guidance['safety_notes'] = cls._generate_safety_notes(operation_goal, guidance['safety_level'])

            # 步骤4: 提供替代方案
            guidance['alternative_approaches'] = cls._suggest_alternative_approaches(operation_goal)

            return guidance

        except Exception as e:
            error_msg = f"生成安全操作指导失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return {
                'success': False,
                'error': error_msg,
                'message': '无法生成安全操作指导'
            }

    @classmethod
    def _analyze_operation_risk(cls, operation_goal: str) -> Dict[str, Any]:
        """分析操作风险"""
        risk_keywords = {
            'high_risk': ['删除', '清空', '覆盖', '替换', '移除', 'drop', 'delete', 'remove', 'clear'],
            'medium_risk': ['更新', '修改', '更改', '编辑', '插入', '添加', 'update', 'modify', 'insert'],
            'low_risk': ['查看', '读取', '搜索', '分析', '检查', 'read', 'search', 'analyze', 'check']
        }

        operation_lower = operation_goal.lower()

        for level, keywords in risk_keywords.items():
            if any(keyword in operation_lower for keyword in keywords):
                if level == 'high_risk':
                    return {'safety_level': 'high', 'risk_category': 'data_modification'}
                elif level == 'medium_risk':
                    return {'safety_level': 'medium', 'risk_category': 'data_change'}
                else:
                    return {'safety_level': 'low', 'risk_category': 'data_access'}

        return {'safety_level': 'medium', 'risk_category': 'unknown'}

    @classmethod
    def _is_data_modification_operation(cls, operation_goal: str) -> bool:
        """判断是否为数据修改操作"""
        modification_keywords = [
            '更新', '修改', '更改', '编辑', '插入', '添加', '删除', '清空', '覆盖',
            'update', 'modify', 'change', 'edit', 'insert', 'add', 'delete', 'clear', 'overwrite'
        ]
        return any(keyword in operation_goal.lower() for keyword in modification_keywords)

    @classmethod
    def _is_data_analysis_operation(cls, operation_goal: str) -> bool:
        """判断是否为数据分析操作"""
        analysis_keywords = [
            '查看', '读取', '搜索', '分析', '检查', '比较', '统计', '汇总',
            'read', 'search', 'analyze', 'check', 'compare', 'count', 'summary'
        ]
        return any(keyword in operation_goal.lower() for keyword in analysis_keywords)

    @classmethod
    def _is_file_management_operation(cls, operation_goal: str) -> bool:
        """判断是否为文件管理操作"""
        management_keywords = [
            '创建', '删除文件', '重命名', '复制', '移动', '备份', '工作表',
            'create', 'delete file', 'rename', 'copy', 'move', 'backup', 'worksheet'
        ]
        return any(keyword in operation_goal.lower() for keyword in management_keywords)

    @classmethod
    def _get_safe_modification_sequence(
        cls, file_path: str, operation_goal: str, context_info: Optional[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """获取数据修改的安全操作序列"""
        sequence = []

        # 步骤1: 文件状态检查
        sequence.append({
            'step': 1,
            'action': '文件状态检查',
            'tool': 'excel_check_file_status',
            'purpose': '确认文件可用且未被锁定',
            'parameters': {'file_path': file_path},
            'critical': True
        })

        # 步骤2: 操作影响评估
        sequence.append({
            'step': 2,
            'action': '影响评估',
            'tool': 'excel_assess_operation_impact',
            'purpose': '分析操作影响范围和风险',
            'parameters': {
                'file_path': file_path,
                'range_expression': '需要根据具体操作确定',
                'operation_type': 'update/delete/insert'
            },
            'critical': True
        })

        # 步骤3: 操作确认（高风险操作）
        if '删除' in operation_goal or '清空' in operation_goal:
            sequence.append({
                'step': 3,
                'action': '操作确认',
                'tool': 'excel_confirm_operation',
                'purpose': '获得用户明确确认',
                'parameters': {
                    'file_path': file_path,
                    'operation_type': 'delete',
                    'require_confirmation': True
                },
                'critical': True
            })

        # 步骤4: 备份（高风险操作）
        if any(keyword in operation_goal for keyword in ['删除', '清空', '覆盖']):
            sequence.append({
                'step': 4,
                'action': '创建备份',
                'tool': 'excel_create_backup',
                'purpose': '操作前自动创建备份',
                'parameters': {
                    'file_path': file_path,
                    'backup_name': 'auto_backup_before_operation'
                },
                'critical': True
            })

        # 步骤5: 执行实际操作
        sequence.append({
            'step': 5,
            'action': '执行操作',
            'tool': '根据具体操作选择工具',
            'purpose': '执行用户请求的操作',
            'parameters': {
                'file_path': file_path,
                'insert_mode': True,  # 默认使用安全模式
                'skip_safety_checks': False
            },
            'critical': True
        })

        # 步骤6: 验证结果
        sequence.append({
            'step': 6,
            'action': '结果验证',
            'tool': 'excel_get_range',
            'purpose': '验证操作结果是否符合预期',
            'parameters': {
                'file_path': file_path,
                'range_expression': '操作范围'
            },
            'critical': False
        })

        return sequence

    @classmethod
    def _get_safe_analysis_sequence(
        cls, file_path: str, operation_goal: str, context_info: Optional[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """获取数据分析的安全操作序列"""
        sequence = []

        # 步骤1: 文件信息获取
        sequence.append({
            'step': 1,
            'action': '获取文件信息',
            'tool': 'excel_get_file_info',
            'purpose': '了解文件基本信息',
            'parameters': {'file_path': file_path},
            'critical': False
        })

        # 步骤2: 工作表列表
        sequence.append({
            'step': 2,
            'action': '列出工作表',
            'tool': 'excel_list_sheets',
            'purpose': '获取所有工作表信息',
            'parameters': {'file_path': file_path},
            'critical': False
        })

        # 步骤3: 数据读取（只读操作）
        sequence.append({
            'step': 3,
            'action': '读取数据',
            'tool': 'excel_get_range',
            'purpose': '获取需要分析的数据',
            'parameters': {
                'file_path': file_path,
                'range_expression': '目标范围',
                'include_formatting': False
            },
            'critical': False
        })

        return sequence

    @classmethod
    def _get_safe_management_sequence(
        cls, file_path: str, operation_goal: str, context_info: Optional[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """获取文件管理的安全操作序列"""
        sequence = []

        if '创建' in operation_goal:
            # 创建文件操作
            sequence.append({
                'step': 1,
                'action': '创建文件',
                'tool': 'excel_create_file',
                'purpose': '创建新的Excel文件',
                'parameters': {
                    'file_path': file_path,
                    'sheet_names': ['Sheet1']
                },
                'critical': False
            })
        elif '备份' in operation_goal:
            # 备份操作
            sequence.append({
                'step': 1,
                'action': '创建备份',
                'tool': 'excel_create_backup',
                'purpose': '创建文件备份',
                'parameters': {
                    'file_path': file_path,
                    'backup_name': 'manual_backup'
                },
                'critical': False
            })

        return sequence

    @classmethod
    def _get_generic_safe_sequence(
        cls, file_path: str, operation_goal: str, context_info: Optional[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """获取通用的安全操作序列"""
        sequence = []

        # 默认安全序列
        sequence.append({
            'step': 1,
            'action': '文件状态检查',
            'tool': 'excel_check_file_status',
            'purpose': '确认文件状态正常',
            'parameters': {'file_path': file_path},
            'critical': True
        })

        sequence.append({
            'step': 2,
            'action': '影响评估',
            'tool': 'excel_assess_operation_impact',
            'purpose': '评估操作影响',
            'parameters': {
                'file_path': file_path,
                'operation_type': 'unknown'
            },
            'critical': True
        })

        return sequence

    @classmethod
    def _generate_safety_notes(cls, operation_goal: str, safety_level: str) -> List[str]:
        """生成安全注意事项"""
        notes = []

        # 基础安全提示
        notes.append("⚠️ 始终在操作前检查文件状态")
        notes.append("📋 大范围操作建议分批进行")
        notes.append("💾 重要操作前手动创建备份")

        # 基于风险等级的提示
        if safety_level == 'high':
            notes.extend([
                "🚨 高风险操作：可能造成数据永久丢失",
                "✋ 必须获得用户明确确认才能执行",
                "🔄 强烈建议使用insert_mode=True避免覆盖",
                "📝 记录操作日志以便追踪和回滚"
            ])
        elif safety_level == 'medium':
            notes.extend([
                "⚠️ 中等风险操作：建议预览影响范围",
                "🔍 检查现有数据避免意外覆盖",
                "💡 考虑使用测试数据验证操作"
            ])
        else:
            notes.extend([
                "✅ 低风险操作：可以安全执行",
                "👀 仍建议查看操作预览"
            ])

        # 基于操作类型的提示
        if '删除' in operation_goal:
            notes.append("🗑️ 删除操作不可逆，请三思而后行")
        elif '更新' in operation_goal or '修改' in operation_goal:
            notes.append("✏️ 更新操作建议使用insert_mode保护现有数据")
        elif '格式' in operation_goal:
            notes.append("🎨 格式化可能影响公式显示，请谨慎操作")

        return notes

    @classmethod
    def _suggest_alternative_approaches(cls, operation_goal: str) -> List[Dict[str, Any]]:
        """建议替代方案"""
        alternatives = []

        # 通用替代方案
        if '大范围' in operation_goal or '批量' in operation_goal:
            alternatives.append({
                'approach': '分批操作',
                'description': '将大范围操作分解为多个小批量操作',
                'benefits': ['降低风险', '提高可控性', '便于错误定位'],
                'implementation': '每次操作不超过100个单元格'
            })

        if '删除' in operation_goal:
            alternatives.append({
                'approach': '先备份后删除',
                'description': '创建完整备份后再执行删除操作',
                'benefits': ['可恢复', '降低风险', '安全可靠'],
                'implementation': '使用excel_create_backup创建备份'
            })

        if '更新' in operation_goal or '修改' in operation_goal:
            alternatives.append({
                'approach': '使用插入模式',
                'description': '使用insert_mode=True避免覆盖现有数据',
                'benefits': ['保护现有数据', '可回滚', '更安全'],
                'implementation': '设置insert_mode=True参数'
            })

        return alternatives

    @classmethod
    def create_auto_backup(
        cls,
        file_path: str,
        backup_name: Optional[str] = None,
        backup_reason: str = "自动备份",
        user_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        @intention 创建自动备份文件

        Args:
            file_path: 原始Excel文件路径
            backup_name: 备份名称（可选）
            backup_reason: 备份原因描述
            user_id: 用户ID

        Returns:
            Dict: 备份结果，包含备份文件路径和相关信息
        """
        try:
            import os
            import shutil
            import time
            from pathlib import Path

            # 检查源文件是否存在
            if not os.path.exists(file_path):
                return {
                    'success': False,
                    'error': 'SOURCE_FILE_NOT_FOUND',
                    'message': f'源文件不存在: {file_path}'
                }

            # 检查源文件状态
            file_status = cls.check_file_status(file_path)
            if not file_status['success']:
                return {
                    'success': False,
                    'error': 'FILE_STATUS_CHECK_FAILED',
                    'message': f'无法验证文件状态: {file_status.get("error", "未知错误")}'
                }

            if file_status['file_status'].get('locked', False):
                return {
                    'success': False,
                    'error': 'FILE_LOCKED',
                    'message': f'文件被锁定，无法创建备份: {file_status["file_status"]["locked_by"]}'
                }

            # 生成备份文件路径
            path_obj = Path(file_path)
            backup_dir = path_obj.parent / "auto_backups"
            backup_dir.mkdir(exist_ok=True)

            # 生成备份文件名
            timestamp = time.strftime('%Y%m%d_%H%M%S', time.localtime())
            if backup_name:
                backup_filename = f"{backup_name}_{timestamp}{path_obj.suffix}"
            else:
                backup_filename = f"{path_obj.stem}_auto_backup_{timestamp}{path_obj.suffix}"

            backup_path = backup_dir / backup_filename

            # 创建备份
            try:
                shutil.copy2(file_path, backup_path)

                # 验证备份文件
                if not os.path.exists(backup_path):
                    return {
                        'success': False,
                        'error': 'BACKUP_CREATION_FAILED',
                        'message': '备份文件创建失败'
                    }

                backup_size = os.path.getsize(backup_path)
                original_size = os.path.getsize(file_path)

                if backup_size != original_size:
                    # 删除不完整的备份
                    os.remove(backup_path)
                    return {
                        'success': False,
                        'error': 'BACKUP_INCOMPLETE',
                        'message': f'备份文件大小不匹配: 原文件{original_size}字节, 备份{backup_size}字节'
                    }

                # 创建备份元数据
                backup_metadata = {
                    'original_file': str(path_obj),
                    'backup_file': str(backup_path),
                    'backup_name': backup_name or 'auto_backup',
                    'backup_reason': backup_reason,
                    'created_at': time.time(),
                    'created_by': user_id or 'system',
                    'file_size': backup_size,
                    'original_checksum': cls._calculate_file_checksum(file_path),
                    'backup_checksum': cls._calculate_file_checksum(str(backup_path)),
                    'file_status': file_status['file_status']
                }

                # 保存备份元数据
                metadata_path = backup_path.with_suffix('.json')
                import json
                with open(metadata_path, 'w', encoding='utf-8') as f:
                    json.dump(backup_metadata, f, ensure_ascii=False, indent=2)

                # 清理旧备份（保留最近10个）
                cls._cleanup_old_backups(backup_dir, path_obj.stem, keep_count=10)

                logger.info(f"{cls._LOG_PREFIX} 自动备份创建成功: {backup_path}")

                return {
                    'success': True,
                    'backup_file': str(backup_path),
                    'backup_name': backup_name or 'auto_backup',
                    'backup_reason': backup_reason,
                    'file_size': backup_size,
                    'created_at': time.time(),
                    'metadata_file': str(metadata_path),
                    'message': f'自动备份创建成功: {backup_filename}'
                }

            except Exception as e:
                # 清理可能创建的不完整文件
                if backup_path.exists():
                    try:
                        os.remove(backup_path)
                    except:
                        pass

                return {
                    'success': False,
                    'error': 'BACKUP_COPY_FAILED',
                    'message': f'复制文件失败: {str(e)}'
                }

        except Exception as e:
            error_msg = f"创建自动备份失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return {
                'success': False,
                'error': error_msg,
                'message': '创建自动备份时发生异常'
            }

    @classmethod
    def restore_from_backup(
        cls,
        backup_file: str,
        target_file: Optional[str] = None,
        user_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        @intention 从备份文件恢复原始文件

        Args:
            backup_file: 备份文件路径
            target_file: 目标文件路径（可选，默认从备份元数据获取）
            user_id: 用户ID

        Returns:
            Dict: 恢复结果
        """
        try:
            import os
            import shutil
            import json
            from pathlib import Path

            # 检查备份文件是否存在
            if not os.path.exists(backup_file):
                return {
                    'success': False,
                    'error': 'BACKUP_FILE_NOT_FOUND',
                    'message': f'备份文件不存在: {backup_file}'
                }

            # 读取备份元数据
            metadata_path = Path(backup_file).with_suffix('.json')
            if not metadata_path.exists():
                return {
                    'success': False,
                    'error': 'BACKUP_METADATA_NOT_FOUND',
                    'message': f'备份元数据不存在: {metadata_path}'
                }

            try:
                with open(metadata_path, 'r', encoding='utf-8') as f:
                    backup_metadata = json.load(f)
            except Exception as e:
                return {
                    'success': False,
                    'error': 'BACKUP_METADATA_READ_FAILED',
                    'message': f'读取备份元数据失败: {str(e)}'
                }

            # 确定目标文件路径
            if target_file is None:
                target_file = backup_metadata['original_file']

            # 检查目标文件状态
            target_exists = os.path.exists(target_file)
            if target_exists:
                target_status = cls.check_file_status(target_file)
                if not target_status['file_status'].get('locked', False):
                    # 如果目标文件未被锁定，创建恢复前备份
                    pre_restore_backup = cls.create_auto_backup(
                        target_file,
                        backup_name="pre_restore_backup",
                        backup_reason="恢复前自动备份",
                        user_id=user_id
                    )
                    if not pre_restore_backup['success']:
                        return {
                            'success': False,
                            'error': 'PRE_RESTORE_BACKUP_FAILED',
                            'message': '创建恢复前备份失败，为安全起见取消恢复操作'
                        }
                else:
                    return {
                        'success': False,
                        'error': 'TARGET_FILE_LOCKED',
                        'message': f'目标文件被锁定，无法恢复: {target_status["file_status"]["locked_by"]}'
                    }

            # 验证备份文件完整性
            current_backup_checksum = cls._calculate_file_checksum(backup_file)
            if current_backup_checksum != backup_metadata.get('backup_checksum'):
                return {
                    'success': False,
                    'error': 'BACKUP_CORRUPTED',
                    'message': '备份文件可能已损坏，校验和不匹配'
                }

            # 执行恢复
            try:
                shutil.copy2(backup_file, target_file)

                # 验证恢复结果
                if not os.path.exists(target_file):
                    return {
                        'success': False,
                        'error': 'RESTORE_FAILED',
                        'message': '文件恢复失败'
                    }

                restored_checksum = cls._calculate_file_checksum(target_file)
                if restored_checksum != backup_metadata.get('original_checksum'):
                    # 恢复的文件校验和不匹配
                    os.remove(target_file)
                    return {
                        'success': False,
                        'error': 'RESTORE_VERIFICATION_FAILED',
                        'message': '恢复的文件校验和不匹配，可能恢复不完整'
                    }

                logger.info(f"{cls._LOG_PREFIX} 文件恢复成功: {target_file}")

                return {
                    'success': True,
                    'restored_file': target_file,
                    'backup_file': backup_file,
                    'backup_created_at': backup_metadata.get('created_at'),
                    'backup_reason': backup_metadata.get('backup_reason'),
                    'pre_restore_backup_available': target_exists and not backup_metadata.get('file_status', {}).get('locked', False),
                    'restored_at': time.time(),
                    'message': f'文件恢复成功: {Path(target_file).name}'
                }

            except Exception as e:
                return {
                    'success': False,
                    'error': 'RESTORE_COPY_FAILED',
                    'message': f'恢复文件失败: {str(e)}'
                }

        except Exception as e:
            error_msg = f"从备份恢复失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return {
                'success': False,
                'error': error_msg,
                'message': '恢复文件时发生异常'
            }

    @classmethod
    def list_backups(
        cls,
        file_path: Optional[str] = None,
        backup_name: Optional[str] = None,
        limit: int = 50
    ) -> Dict[str, Any]:
        """
        @intention 列出可用的备份文件

        Args:
            file_path: 原始文件路径（过滤特定文件的备份）
            backup_name: 备份名称（过滤特定名称的备份）
            limit: 返回结果数量限制

        Returns:
            Dict: 备份文件列表
        """
        try:
            import os
            import json
            from pathlib import Path

            backups = []

            # 确定备份目录
            if file_path:
                path_obj = Path(file_path)
                backup_dir = path_obj.parent / "auto_backups"
            else:
                # 搜索当前目录及子目录中的所有备份
                backup_dir = Path.cwd() / "auto_backups"

            if not backup_dir.exists():
                return {
                    'success': True,
                    'backups': [],
                    'total_count': 0,
                    'message': '没有找到备份目录'
                }

            # 搜索备份文件
            for backup_file in backup_dir.glob("*.xlsx"):
                metadata_file = backup_file.with_suffix('.json')
                if metadata_file.exists():
                    try:
                        with open(metadata_file, 'r', encoding='utf-8') as f:
                            metadata = json.load(f)

                        # 应用过滤条件
                        if file_path and metadata.get('original_file') != str(Path(file_path).resolve()):
                            continue

                        if backup_name and metadata.get('backup_name') != backup_name:
                            continue

                        # 添加备份信息
                        backup_info = {
                            'backup_file': str(backup_file),
                            'original_file': metadata.get('original_file'),
                            'backup_name': metadata.get('backup_name'),
                            'backup_reason': metadata.get('backup_reason'),
                            'created_at': metadata.get('created_at'),
                            'created_by': metadata.get('created_by'),
                            'file_size': metadata.get('file_size'),
                            'checksum_valid': cls._calculate_file_checksum(str(backup_file)) == metadata.get('backup_checksum'),
                            'metadata_file': str(metadata_file)
                        }

                        backups.append(backup_info)

                    except Exception as e:
                        logger.warning(f"读取备份元数据失败 {metadata_file}: {str(e)}")
                        continue

            # 按创建时间排序（最新的在前）
            backups.sort(key=lambda x: x.get('created_at', 0), reverse=True)

            # 限制返回数量
            limited_backups = backups[:limit]

            return {
                'success': True,
                'backups': limited_backups,
                'total_count': len(backups),
                'returned_count': len(limited_backups),
                'backup_directory': str(backup_dir),
                'message': f'找到 {len(backups)} 个备份文件'
            }

        except Exception as e:
            error_msg = f"列出备份文件失败: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return {
                'success': False,
                'error': error_msg,
                'message': '列出备份文件时发生异常'
            }

    @classmethod
    def _calculate_file_checksum(cls, file_path: str) -> str:
        """计算文件校验和"""
        try:
            import hashlib

            hash_md5 = hashlib.md5()
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)

            return hash_md5.hexdigest()

        except Exception as e:
            logger.error(f"{cls._LOG_PREFIX} 计算文件校验和失败 {file_path}: {str(e)}")
            return ""

    @classmethod
    def _cleanup_old_backups(cls, backup_dir: Path, file_stem: str, keep_count: int = 10):
        """清理旧备份文件，保留最新的几个"""
        try:
            backup_files = list(backup_dir.glob(f"{file_stem}_*_auto_backup_*.xlsx"))

            # 按修改时间排序（最新的在前）
            backup_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)

            # 删除多余的备份
            for old_backup in backup_files[keep_count:]:
                try:
                    old_backup.unlink()
                    # 同时删除元数据文件
                    metadata_file = old_backup.with_suffix('.json')
                    if metadata_file.exists():
                        metadata_file.unlink()
                    logger.debug(f"{cls._LOG_PREFIX} 清理旧备份: {old_backup}")
                except Exception as e:
                    logger.warning(f"{cls._LOG_PREFIX} 清理备份失败 {old_backup}: {str(e)}")

        except Exception as e:
            logger.error(f"{cls._LOG_PREFIX} 清理旧备份失败: {str(e)}")

    @classmethod
    def check_duplicate_ids(
        cls,
        file_path: str,
        sheet_name: str,
        id_column: Union[int, str] = 1,
        header_row: int = 1
    ) -> Dict[str, Any]:
        """
        检查Excel工作表中的ID重复情况

        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            id_column: ID列位置 (1-based数字或列名)
            header_row: 表头行号 (1-based)

        Returns:
            Dict: 包含success、has_duplicates、duplicate_count、total_ids、unique_ids、duplicates、message
        """
        if cls.DEBUG_LOG_ENABLED:
            logger.info(f"{cls._LOG_PREFIX} 开始检查ID重复: {sheet_name}")

        try:
            from collections import Counter
            from openpyxl import load_workbook

            # 参数验证
            if not file_path or not sheet_name:
                return {
                    'success': False,
                    'message': '文件路径和工作表名不能为空',
                    'has_duplicates': False,
                    'duplicate_count': 0,
                    'total_ids': 0,
                    'unique_ids': 0,
                    'duplicates': []
                }

            # 加载工作簿
            try:
                wb = load_workbook(file_path, read_only=True)
            except FileNotFoundError:
                return {
                    'success': False,
                    'message': f'文件不存在: {file_path}',
                    'has_duplicates': False,
                    'duplicate_count': 0,
                    'total_ids': 0,
                    'unique_ids': 0,
                    'duplicates': []
                }
            except Exception as e:
                return {
                    'success': False,
                    'message': f'无法加载文件: {str(e)}',
                    'has_duplicates': False,
                    'duplicate_count': 0,
                    'total_ids': 0,
                    'unique_ids': 0,
                    'duplicates': []
                }

            # 检查工作表是否存在
            if sheet_name not in wb.sheetnames:
                return {
                    'success': False,
                    'message': f'工作表不存在: {sheet_name}',
                    'has_duplicates': False,
                    'duplicate_count': 0,
                    'total_ids': 0,
                    'unique_ids': 0,
                    'duplicates': []
                }

            ws = wb[sheet_name]

            # 处理列索引
            if isinstance(id_column, str):
                from openpyxl.utils import column_index_from_string
                try:
                    col_idx = column_index_from_string(id_column)
                except Exception:
                    return {
                        'success': False,
                        'message': f'无效的列名: {id_column}',
                        'has_duplicates': False,
                        'duplicate_count': 0,
                        'total_ids': 0,
                        'unique_ids': 0,
                        'duplicates': []
                    }
            else:
                col_idx = id_column

            # 检查表头行是否存在
            if header_row < 1 or header_row > ws.max_row:
                return {
                    'success': False,
                    'message': f'表头行不存在: {header_row}',
                    'has_duplicates': False,
                    'duplicate_count': 0,
                    'total_ids': 0,
                    'unique_ids': 0,
                    'duplicates': []
                }

            # 检查列是否存在
            if col_idx < 1 or col_idx > ws.max_column:
                return {
                    'success': False,
                    'message': f'列不存在或索引超出范围: {col_idx}',
                    'has_duplicates': False,
                    'duplicate_count': 0,
                    'total_ids': 0,
                    'unique_ids': 0,
                    'duplicates': []
                }

            # 收集ID数据
            ids_with_rows = []
            for row in range(header_row + 1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value is not None:  # 跳过空值
                    ids_with_rows.append((cell_value, row))

            # 统计ID出现次数
            id_counter = Counter([id_val for id_val, _ in ids_with_rows])
            total_ids = len(ids_with_rows)
            unique_ids = len(id_counter)

            # 查找重复的ID
            duplicates = []
            duplicate_count = 0

            for id_value, count in id_counter.items():
                if count > 1:
                    duplicate_count += 1
                    # 找到该ID的所有行号
                    rows = [row for id_val, row in ids_with_rows if id_val == id_value]
                    # 使用绝对行号（Excel中的实际行号）
                    absolute_rows = rows

                    duplicates.append({
                        'id_value': id_value,
                        'count': count,
                        'rows': absolute_rows
                    })

            has_duplicates = duplicate_count > 0

            # 构建返回结果
            message = f"共检查{total_ids}个ID，发现{duplicate_count}个重复ID" if has_duplicates else f"共检查{total_ids}个ID，无重复ID"

            return {
                'success': True,
                'has_duplicates': has_duplicates,
                'duplicate_count': duplicate_count,
                'total_ids': total_ids,
                'unique_ids': unique_ids,
                'duplicates': duplicates,
                'message': message
            }

        except Exception as e:
            error_msg = f"检查ID重复时发生错误: {str(e)}"
            logger.error(f"{cls._LOG_PREFIX} {error_msg}")
            return cls._format_error_result(error_msg)
