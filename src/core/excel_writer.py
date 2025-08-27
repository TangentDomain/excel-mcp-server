"""
Excel MCP Server - Excel写入模块

提供Excel文件写入和修改功能
"""

import logging
import tempfile
import os
from typing import List, Any, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.utils import range_boundaries

from ..models.types import RangeInfo, ModifiedCell, OperationResult, RangeType
from ..utils.validators import ExcelValidator
from ..utils.parsers import RangeParser
from ..utils.exceptions import SheetNotFoundError, DataValidationError

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Excel文件写入器"""

    def __init__(self, file_path: str):
        """
        初始化Excel写入器

        Args:
            file_path: Excel文件路径
        """
        self.file_path = ExcelValidator.validate_file_path(file_path)

    def update_range(
        self,
        range_expression: str,
        data: List[List[Any]],
        preserve_formulas: bool = True
    ) -> OperationResult:
        """
        修改Excel文件中指定范围的数据

        Args:
            range_expression: 范围表达式
            data: 要写入的二维数据数组
            preserve_formulas: 是否保留现有的公式

        Returns:
            OperationResult: 修改操作的结果
        """
        try:
            # 解析范围表达式
            range_info = RangeParser.parse_range_expression(range_expression)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, range_info.sheet_name)

            # 根据范围类型处理不同的范围格式
            cell_range_for_boundaries = self._convert_to_cell_range(
                range_info.cell_range, range_info.range_type, sheet, data
            )

            # 获取范围边界
            min_col, min_row, max_col, max_row = range_boundaries(cell_range_for_boundaries)

            # 获取范围维度（允许数据大小不匹配，Excel会自动处理）
            range_rows = max_row - min_row + 1
            range_cols = max_col - min_col + 1
            # 注意: 不再严格验证数据维度，允许数据超出或不足范围

            # 写入数据
            modified_cells = self._write_data(
                sheet, data, min_row, min_col, preserve_formulas
            )

            # 保存文件
            workbook.save(self.file_path)

            return OperationResult(
                success=True,
                data=modified_cells,
                metadata={
                    'file_path': self.file_path,
                    'range': range_expression,
                    'sheet_name': sheet.title,
                    'modified_cells_count': len(modified_cells)
                }
            )

        except Exception as e:
            logger.error(f"更新范围数据失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def insert_rows(
        self,
        sheet_name: Optional[str] = None,
        row_index: int = 1,
        count: int = 1
    ) -> OperationResult:
        """
        在Excel文件中插入空白行

        Args:
            sheet_name: 工作表名称
            row_index: 插入行的位置（1-based）
            count: 要插入的行数

        Returns:
            OperationResult: 插入操作的结果
        """
        try:
            # 验证参数
            ExcelValidator.validate_row_operations(row_index, count)
            ExcelValidator.validate_sheet_name(sheet_name)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, sheet_name)

            # 记录操作前的信息
            original_max_row = sheet.max_row

            # 插入行
            sheet.insert_rows(row_index, count)

            # 保存文件
            workbook.save(self.file_path)

            return OperationResult(
                success=True,
                message=f"成功在第{row_index}行前插入{count}行",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet.title,
                    'inserted_at_row': row_index,
                    'inserted_count': count,
                    'original_max_row': original_max_row,
                    'new_max_row': sheet.max_row
                }
            )

        except Exception as e:
            logger.error(f"插入行失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def insert_columns(
        self,
        sheet_name: Optional[str] = None,
        column_index: int = 1,
        count: int = 1
    ) -> OperationResult:
        """
        在Excel文件中插入空白列

        Args:
            sheet_name: 工作表名称
            column_index: 插入列的位置（1-based）
            count: 要插入的列数

        Returns:
            OperationResult: 插入操作的结果
        """
        try:
            # 验证参数
            ExcelValidator.validate_column_operations(column_index, count)
            ExcelValidator.validate_sheet_name(sheet_name)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, sheet_name)

            # 记录操作前的信息
            original_max_column = sheet.max_column

            # 插入列
            sheet.insert_cols(column_index, count)

            # 保存文件
            workbook.save(self.file_path)

            return OperationResult(
                success=True,
                message=f"成功插入{count}列",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet.title,
                    'inserted_at_column': column_index,
                    'inserted_count': count,
                    'original_max_column': original_max_column,
                    'new_max_column': sheet.max_column
                }
            )

        except Exception as e:
            logger.error(f"插入列失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def delete_rows(
        self,
        sheet_name: Optional[str] = None,
        start_row: int = 1,
        count: int = 1
    ) -> OperationResult:
        """
        删除Excel文件中的行

        Args:
            sheet_name: 工作表名称
            start_row: 开始删除的行号
            count: 要删除的行数

        Returns:
            OperationResult: 删除操作的结果
        """
        try:
            # 验证参数
            ExcelValidator.validate_row_operations(start_row, count)
            ExcelValidator.validate_sheet_name(sheet_name)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, sheet_name)

            # 记录操作前的信息
            original_max_row = sheet.max_row

            # 验证删除范围
            if start_row > original_max_row:
                raise DataValidationError(
                    f"起始行号({start_row})超过工作表最大行数({original_max_row})"
                )

            # 计算实际删除的行数
            actual_count = min(count, original_max_row - start_row + 1)

            # 删除行
            sheet.delete_rows(start_row, actual_count)

            # 保存文件
            workbook.save(self.file_path)

            return OperationResult(
                success=True,
                message=f"成功从第{start_row}行开始删除{actual_count}行",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet.title,
                    'deleted_start_row': start_row,
                    'actual_deleted_count': actual_count,
                    'original_max_row': original_max_row,
                    'new_max_row': sheet.max_row
                }
            )

        except Exception as e:
            logger.error(f"删除行失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def delete_columns(
        self,
        sheet_name: Optional[str] = None,
        start_column: int = 1,
        count: int = 1
    ) -> OperationResult:
        """
        删除Excel文件中的列

        Args:
            sheet_name: 工作表名称
            start_column: 开始删除的列号
            count: 要删除的列数

        Returns:
            OperationResult: 删除操作的结果
        """
        try:
            # 验证参数
            ExcelValidator.validate_column_operations(start_column, count)
            ExcelValidator.validate_sheet_name(sheet_name)

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 确定工作表
            sheet = self._get_worksheet(workbook, sheet_name)

            # 记录操作前的信息
            original_max_column = sheet.max_column

            # 验证删除范围
            if start_column > original_max_column:
                raise DataValidationError(
                    f"起始列号({start_column})超过工作表最大列数({original_max_column})"
                )

            # 计算实际删除的列数
            actual_count = min(count, original_max_column - start_column + 1)

            # 删除列
            sheet.delete_cols(start_column, actual_count)

            # 保存文件
            workbook.save(self.file_path)

            return OperationResult(
                success=True,
                message=f"成功删除{actual_count}列",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet.title,
                    'deleted_start_column': start_column,
                    'actual_deleted_count': actual_count,
                    'original_max_column': original_max_column,
                    'new_max_column': sheet.max_column
                }
            )

        except Exception as e:
            logger.error(f"删除列失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def _convert_to_cell_range(self, cell_range: str, range_type: RangeType, sheet, data: List[List[Any]]) -> str:
        """
        将不同类型的范围表达式转换为标准的单元格范围格式
        
        Args:
            cell_range: 原始范围表达式
            range_type: 范围类型
            sheet: 工作表对象
            data: 数据数组（用于确定需要的列数）
            
        Returns:
            标准的单元格范围表达式
        """
        from openpyxl.utils import get_column_letter
        
        if range_type == RangeType.ROW_RANGE:
            # 处理行范围，如 "1:1" 或 "1250:1250"
            if ':' in cell_range:
                start_row, end_row = cell_range.split(':')
                start_row, end_row = int(start_row), int(end_row)
            else:
                start_row = end_row = int(cell_range)
                
            # 根据数据确定需要的列数，如果没有数据则使用第一列到数据宽度
            if data and len(data) > 0:
                data_cols = len(data[0]) if data[0] else 1
                end_col = data_cols
            else:
                end_col = 1  # 默认只使用第一列
                
            start_col = 1
            start_col_letter = get_column_letter(start_col)
            end_col_letter = get_column_letter(end_col)
            
            return f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
            
        elif range_type == RangeType.COLUMN_RANGE:
            # 处理列范围，如 "A:A" 或 "A:C"
            if ':' in cell_range:
                start_col, end_col = cell_range.split(':')
            else:
                start_col = end_col = cell_range
                
            # 根据数据确定需要的行数
            if data:
                data_rows = len(data)
                end_row = data_rows
            else:
                end_row = 1  # 默认只使用第一行
                
            start_row = 1
            return f"{start_col}{start_row}:{end_col}{end_row}"
            
        elif range_type == RangeType.SINGLE_ROW:
            # 处理单行，如 "1"
            row_num = int(cell_range.split(':')[0])  # 规范化后是 "1:1" 格式
            
            # 根据数据确定列数
            if data and len(data) > 0:
                data_cols = len(data[0]) if data[0] else 1
                end_col = data_cols
            else:
                end_col = 1
                
            start_col_letter = get_column_letter(1)
            end_col_letter = get_column_letter(end_col)
            
            return f"{start_col_letter}{row_num}:{end_col_letter}{row_num}"
            
        elif range_type == RangeType.SINGLE_COLUMN:
            # 处理单列，如 "A"
            col_letter = cell_range.split(':')[0]  # 规范化后是 "A:A" 格式
            
            # 根据数据确定行数
            if data:
                data_rows = len(data)
                end_row = data_rows
            else:
                end_row = 1
                
            return f"{col_letter}1:{col_letter}{end_row}"
            
        else:
            # 其他情况（CELL_RANGE）直接返回
            return cell_range

    def _get_worksheet(self, workbook, sheet_name: Optional[str]):
        """获取工作表 - 强制要求指定工作表名称"""
        if not sheet_name or not sheet_name.strip():
            raise SheetNotFoundError(f"工作表名称不能为空，必须明确指定工作表")
        
        if not workbook.sheetnames:
            raise SheetNotFoundError(f"Excel文件中没有任何工作表")
            
        if sheet_name not in workbook.sheetnames:
            raise SheetNotFoundError(f"工作表不存在: {sheet_name}，可用工作表: {', '.join(workbook.sheetnames)}")
            
        return workbook[sheet_name]

    def _write_data(
        self,
        sheet,
        data: List[List[Any]],
        start_row: int,
        start_col: int,
        preserve_formulas: bool
    ) -> List[ModifiedCell]:
        """写入数据到工作表"""
        modified_cells = []

        for row_offset, row_data in enumerate(data):
            for col_offset, value in enumerate(row_data):
                row_idx = start_row + row_offset
                col_idx = start_col + col_offset
                cell = sheet.cell(row=row_idx, column=col_idx)

                # 保留公式检查
                if preserve_formulas and cell.data_type == 'f':
                    continue

                old_value = cell.value
                cell.value = value

                modified_cells.append(ModifiedCell(
                    coordinate=cell.coordinate,
                    old_value=old_value,
                    new_value=value
                ))

        return modified_cells

    def set_formula(
        self,
        cell_address: str,
        formula: str,
        sheet_name: Optional[str] = None
    ) -> OperationResult:
        """
        设置单元格公式

        Args:
            cell_address: 目标单元格地址（如"A1"）
            formula: Excel公式（不包含等号）
            sheet_name: 目标工作表名称

        Returns:
            OperationResult: 公式设置结果
        """
        try:
            # 验证公式格式（简单验证）
            if not formula.strip():
                return OperationResult(
                    success=False,
                    error="公式不能为空"
                )

            # 确保公式不以等号开头（openpyxl会自动添加）
            if formula.startswith('='):
                formula = formula[1:]

            # 验证单元格地址格式
            from openpyxl.utils.cell import coordinate_from_string
            try:
                coordinate_from_string(cell_address)
            except ValueError as e:
                return OperationResult(
                    success=False,
                    error=f"单元格地址格式错误: {cell_address}"
                )

            # 加载工作簿并设置公式
            workbook = load_workbook(self.file_path)
            sheet = self._get_worksheet(workbook, sheet_name)

            # 设置公式
            cell = sheet[cell_address]
            old_value = cell.value
            old_formula = cell.formula if hasattr(cell, 'formula') else None

            cell.value = f"={formula}"

            # 保存文件
            workbook.save(self.file_path)
            workbook.close()

            # 重新读取以获取计算值
            workbook_read = load_workbook(self.file_path, data_only=True)
            sheet_read = self._get_worksheet(workbook_read, sheet_name)
            calculated_value = sheet_read[cell_address].value
            workbook_read.close()

            logger.info(f"成功设置公式: {cell_address} = {formula}")

            return OperationResult(
                success=True,
                message=f"公式设置成功",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet.title,
                    'cell_address': cell_address,
                    'formula': formula,
                    'calculated_value': calculated_value,
                    'old_value': old_value,
                    'old_formula': old_formula
                }
            )

        except Exception as e:
            logger.error(f"设置公式失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def format_cells(
        self,
        range_expression: str,
        formatting: dict,
        sheet_name: Optional[str] = None
    ) -> OperationResult:
        """
        设置单元格格式

        Args:
            range_expression: 目标范围
            formatting: 格式配置字典
            sheet_name: 目标工作表名

        Returns:
            OperationResult: 格式应用结果
        """
        try:
            from openpyxl.styles import Font, PatternFill, Border, Alignment

            # 解析范围表达式
            range_info = RangeParser.parse_range_expression(range_expression)

            # 加载工作簿
            workbook = load_workbook(self.file_path)
            sheet = self._get_worksheet(workbook, sheet_name or range_info.sheet_name)

            # 获取范围边界
            if range_info.range_type in [RangeType.COLUMN_RANGE, RangeType.SINGLE_COLUMN, RangeType.ROW_RANGE, RangeType.SINGLE_ROW]:
                # 处理整行或整列
                cells_range = sheet[range_expression.replace(f"{sheet.title}!", "")]
            else:
                min_col, min_row, max_col, max_row = range_boundaries(range_info.cell_range)
                cells_range = sheet.iter_rows(
                    min_row=min_row, max_row=max_row,
                    min_col=min_col, max_col=max_col
                )

            formatted_count = 0

            # 应用格式
            for row in cells_range:
                if isinstance(row, tuple):
                    for cell in row:
                        self._apply_cell_format(cell, formatting)
                        formatted_count += 1
                else:
                    self._apply_cell_format(row, formatting)
                    formatted_count += 1

            # 保存文件
            workbook.save(self.file_path)
            workbook.close()

            logger.info(f"成功格式化{formatted_count}个单元格")

            return OperationResult(
                success=True,
                message=f"成功格式化{formatted_count}个单元格",
                metadata={
                    'file_path': self.file_path,
                    'sheet_name': sheet.title,
                    'range': range_expression,
                    'formatted_count': formatted_count,
                    'formatting_applied': formatting
                }
            )

        except Exception as e:
            logger.error(f"格式化失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def evaluate_formula(
        self,
        formula: str,
        context_sheet: Optional[str] = None
    ) -> OperationResult:
        """
        临时执行Excel公式并返回计算结果，不修改文件
        使用缓存机制提升性能

        Args:
            formula: Excel公式（不包含等号）
            context_sheet: 公式执行的上下文工作表

        Returns:
            OperationResult: 公式执行结果
        """
        try:
            import tempfile
            import os
            from openpyxl import Workbook
            import time
            from ..utils.formula_cache import get_formula_cache

            start_time = time.time()

            # 确保公式不以等号开头
            if formula.startswith('='):
                formula = formula[1:]

            # 验证公式格式
            if not formula.strip():
                return OperationResult(
                    success=False,
                    error="公式不能为空"
                )

            # 尝试从缓存获取结果
            cache = get_formula_cache()
            cached_result = cache.get(self.file_path, formula, context_sheet)

            if cached_result is not None:
                execution_time = round((time.time() - start_time) * 1000, 2)
                logger.info(f"缓存命中，公式: {formula} = {cached_result}")

                # 确定结果类型
                result_type = self._get_result_type(cached_result)

                return OperationResult(
                    success=True,
                    message="公式执行成功（缓存）",
                    data=cached_result,
                    metadata={
                        'formula': formula,
                        'result': cached_result,
                        'result_type': result_type,
                        'execution_time_ms': execution_time,
                        'context_sheet': context_sheet or "default",
                        'cached': True,
                        'cache_stats': cache.get_stats()
                    }
                )

            # 缓存未命中，尝试获取缓存的工作簿
            cached_workbook_data = cache.get_cached_workbook(self.file_path)

            if cached_workbook_data:
                temp_workbook, temp_file_path = cached_workbook_data
                logger.debug("使用缓存的工作簿进行计算")
            else:
                # 创建新的临时工作簿
                temp_workbook, temp_file_path = self._create_temp_workbook(context_sheet, cache)

            try:
                # 使用xlcalculator计算公式
                calculated_value = self._calculate_with_xlcalculator(
                    temp_file_path, formula, temp_workbook
                )

            except ImportError:
                return OperationResult(
                    success=False,
                    error="需要安装xlcalculator库来支持公式计算: pip install xlcalculator"
                )
            except Exception as calc_error:
                # 如果xlcalculator失败，尝试基础的手动解析
                logger.warning(f"xlcalculator计算失败，尝试基础解析: {calc_error}")
                calculated_value = self._fallback_calculation(temp_file_path, formula)

            # 缓存计算结果
            cache.put(self.file_path, formula, calculated_value, context_sheet)

            # 确定结果类型
            result_type = self._get_result_type(calculated_value)

            execution_time = round((time.time() - start_time) * 1000, 2)
            logger.info(f"成功计算公式: {formula} = {calculated_value}")

            return OperationResult(
                success=True,
                message="公式执行成功",
                data=calculated_value,
                metadata={
                    'formula': formula,
                    'result': calculated_value,
                    'result_type': result_type,
                    'execution_time_ms': execution_time,
                    'context_sheet': context_sheet or "default",
                    'cached': False,
                    'cache_stats': cache.get_stats()
                }
            )

        except Exception as e:
            logger.error(f"公式执行失败: {e}")
            return OperationResult(
                success=False,
                error=f"公式执行失败: {str(e)}"
            )

    def _create_temp_workbook(
        self,
        context_sheet: Optional[str],
        cache
    ) -> tuple:
        """创建临时工作簿用于计算"""
        # 加载原始工作簿（用于提供数据上下文）
        original_workbook = load_workbook(self.file_path, data_only=False)

        # 创建临时工作簿进行计算
        temp_workbook = Workbook()
        temp_sheet = temp_workbook.active
        temp_sheet.title = "Calculation"

        # 选择要复制的源工作表
        if context_sheet and context_sheet in original_workbook.sheetnames:
            source_sheet = original_workbook[context_sheet]
        else:
            # 使用活动工作表或第一个工作表
            source_sheet = original_workbook.active

        # 复制数据到临时工作表（只复制有数据的区域以提升性能）
        if source_sheet.max_row > 1 or source_sheet.max_column > 1:
            for row in source_sheet.iter_rows(
                max_row=min(source_sheet.max_row, 1000),  # 限制复制范围，提升性能
                max_col=min(source_sheet.max_column, 100)
            ):
                for cell in row:
                    if cell.value is not None:
                        target_cell = temp_sheet.cell(
                            row=cell.row,
                            column=cell.column
                        )
                        target_cell.value = cell.value

        # 保存到临时文件
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()

        # 保存工作簿
        temp_workbook.save(temp_file.name)
        original_workbook.close()

        # 缓存工作簿
        cache.cache_workbook(self.file_path, temp_workbook, temp_file.name)

        return temp_workbook, temp_file.name

    def _calculate_with_xlcalculator(
        self,
        temp_file_path: str,
        formula: str,
        temp_workbook
    ) -> any:
        """使用xlcalculator进行计算"""
        from xlcalculator import ModelCompiler, Evaluator

        # 在临时单元格中设置要计算的公式
        temp_sheet = temp_workbook.active
        calc_cell = temp_sheet['Z1']  # 使用Z1作为计算单元格
        calc_cell.value = f"={formula}"

        # 保存更新后的工作簿
        temp_workbook.save(temp_file_path)

        # 编译模型
        compiler = ModelCompiler()
        model = compiler.read_and_parse_archive(temp_file_path)
        evaluator = Evaluator(model)

        # 计算Z1位置的公式
        calculated_value = evaluator.evaluate('Calculation!Z1')
        return calculated_value

    def _fallback_calculation(self, temp_file_path: str, formula: str) -> any:
        """备用计算方法"""
        # 重新加载工作簿获取数据
        data_workbook = load_workbook(temp_file_path, data_only=True)
        data_sheet = data_workbook["Calculation"]

        # 尝试基础的公式解析
        calculated_value = self._basic_formula_parse(formula, data_sheet)
        data_workbook.close()

        return calculated_value

    def _get_result_type(self, value) -> str:
        """确定结果类型"""
        if value is None:
            return "null"
        elif isinstance(value, (int, float)):
            return "number"
        elif isinstance(value, str):
            return "text"
        elif isinstance(value, bool):
            return "boolean"
        else:
            try:
                # 检查是否是日期
                from datetime import datetime, date
                if isinstance(value, (datetime, date)):
                    return "date"
            except:
                pass
            return "unknown"

    def _basic_formula_parse(self, formula: str, sheet) -> any:
        """增强的基础公式解析器 - 支持numpy统计函数"""
        import re

        formula = formula.strip()

        # ==================== 基础统计函数 ====================

        # SUM函数
        sum_match = re.match(r'SUM\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if sum_match:
            start_cell, end_cell = sum_match.groups()
            return self._calculate_range_sum(sheet, start_cell, end_cell)

        # AVERAGE函数
        avg_match = re.match(r'AVERAGE\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if avg_match:
            start_cell, end_cell = avg_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_average(values)

        # COUNT函数
        count_match = re.match(r'COUNT\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if count_match:
            start_cell, end_cell = count_match.groups()
            return self._calculate_range_count(sheet, start_cell, end_cell)

        # MIN函数
        min_match = re.match(r'MIN\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if min_match:
            start_cell, end_cell = min_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_min(values)

        # MAX函数
        max_match = re.match(r'MAX\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if max_match:
            start_cell, end_cell = max_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_max(values)

        # ==================== 高级统计函数 (numpy支持) ====================

        # MEDIAN函数
        median_match = re.match(r'MEDIAN\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if median_match:
            start_cell, end_cell = median_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_median(values)

        # STDEV函数 (样本标准差)
        stdev_match = re.match(r'STDEV(?:\.S)?\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if stdev_match:
            start_cell, end_cell = stdev_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_stdev(values)

        # VAR函数 (样本方差)
        var_match = re.match(r'VAR(?:\.S)?\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if var_match:
            start_cell, end_cell = var_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_var(values)

        # PERCENTILE函数
        percentile_match = re.match(r'PERCENTILE\(([A-Z]+\d+):([A-Z]+\d+),\s*([0-9.]+)\)', formula, re.IGNORECASE)
        if percentile_match:
            start_cell, end_cell, percentile = percentile_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_percentile(values, float(percentile))

        # QUARTILE函数
        quartile_match = re.match(r'QUARTILE\(([A-Z]+\d+):([A-Z]+\d+),\s*([0-3])\)', formula, re.IGNORECASE)
        if quartile_match:
            start_cell, end_cell, quartile = quartile_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_quartile(values, int(quartile))

        # ==================== 条件统计函数 ====================

        # COUNTIF函数
        countif_match = re.match(r'COUNTIF\(([A-Z]+\d+):([A-Z]+\d+),\s*"?([^"]+)"?\)', formula, re.IGNORECASE)
        if countif_match:
            start_cell, end_cell, condition = countif_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_countif(values, condition)

        # SUMIF函数
        sumif_match = re.match(r'SUMIF\(([A-Z]+\d+):([A-Z]+\d+),\s*"?([^"]+)"?\)', formula, re.IGNORECASE)
        if sumif_match:
            start_cell, end_cell, condition = sumif_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_sumif(values, condition)

        # AVERAGEIF函数
        avgif_match = re.match(r'AVERAGEIF\(([A-Z]+\d+):([A-Z]+\d+),\s*"?([^"]+)"?\)', formula, re.IGNORECASE)
        if avgif_match:
            start_cell, end_cell, condition = avgif_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_averageif(values, condition)

        # ==================== 特殊统计函数 ====================

        # MODE函数 (众数)
        mode_match = re.match(r'MODE(?:\.SNGL)?\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if mode_match:
            start_cell, end_cell = mode_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_mode(values)

        # SKEW函数 (偏度)
        skew_match = re.match(r'SKEW\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if skew_match:
            start_cell, end_cell = skew_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_skew(values)

        # KURT函数 (峰度)
        kurt_match = re.match(r'KURT\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if kurt_match:
            start_cell, end_cell = kurt_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_kurtosis(values)

        # GEOMEAN函数 (几何平均数)
        geomean_match = re.match(r'GEOMEAN\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if geomean_match:
            start_cell, end_cell = geomean_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_geomean(values)

        # HARMEAN函数 (调和平均数)
        harmean_match = re.match(r'HARMEAN\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
        if harmean_match:
            start_cell, end_cell = harmean_match.groups()
            values = self._get_range_values(sheet, start_cell, end_cell)
            return self._numpy_harmean(values)

        # ==================== 其他函数 ====================

        # 简单的数学表达式
        if re.match(r'^[\d\+\-\*\/\s\(\)\.]+$', formula):
            try:
                return eval(formula)  # 注意：这在生产环境中需要更安全的实现
            except:
                pass

        # IF函数简单实现
        if_match = re.match(r'IF\((.+),\s*"?([^,"]+)"?,\s*"?([^,"]+)"?\)', formula, re.IGNORECASE)
        if if_match:
            condition, true_val, false_val = if_match.groups()
            # 简单条件判断
            if '>' in condition:
                parts = condition.split('>')
                if len(parts) == 2:
                    left = float(parts[0].strip())
                    right = float(parts[1].strip())
                    return true_val if left > right else false_val
            elif '<' in condition:
                parts = condition.split('<')
                if len(parts) == 2:
                    left = float(parts[0].strip())
                    right = float(parts[1].strip())
                    return true_val if left < right else false_val

        # CONCATENATE函数
        concat_match = re.match(r'CONCATENATE\((.+)\)', formula, re.IGNORECASE)
        if concat_match:
            args = concat_match.group(1).split(',')
            result = ""
            for arg in args:
                arg = arg.strip().strip('"')
                result += arg
            return result

        return None

    def _get_range_values(self, sheet, start_cell: str, end_cell: str) -> list:
        """获取范围内的数值列表"""
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
        values = []

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    values.append(float(cell.value))

        return values

    # ==================== Numpy统计函数实现 ====================

    def _numpy_average(self, values: list) -> float:
        """计算平均值"""
        try:
            import numpy as np
            if not values:
                return 0
            return float(np.mean(values))
        except:
            return sum(values) / len(values) if values else 0

    def _numpy_min(self, values: list) -> float:
        """计算最小值"""
        try:
            import numpy as np
            if not values:
                return 0
            return float(np.min(values))
        except:
            return min(values) if values else 0

    def _numpy_max(self, values: list) -> float:
        """计算最大值"""
        try:
            import numpy as np
            if not values:
                return 0
            return float(np.max(values))
        except:
            return max(values) if values else 0

    def _numpy_median(self, values: list) -> float:
        """计算中位数"""
        try:
            import numpy as np
            if not values:
                return 0
            return float(np.median(values))
        except:
            sorted_values = sorted(values)
            n = len(sorted_values)
            if n == 0:
                return 0
            elif n % 2 == 1:
                return sorted_values[n // 2]
            else:
                return (sorted_values[n // 2 - 1] + sorted_values[n // 2]) / 2

    def _numpy_stdev(self, values: list) -> float:
        """计算样本标准差"""
        try:
            import numpy as np
            if len(values) < 2:
                return 0
            return float(np.std(values, ddof=1))
        except:
            if len(values) < 2:
                return 0
            mean = sum(values) / len(values)
            variance = sum((x - mean) ** 2 for x in values) / (len(values) - 1)
            return variance ** 0.5

    def _numpy_var(self, values: list) -> float:
        """计算样本方差"""
        try:
            import numpy as np
            if len(values) < 2:
                return 0
            return float(np.var(values, ddof=1))
        except:
            if len(values) < 2:
                return 0
            mean = sum(values) / len(values)
            return sum((x - mean) ** 2 for x in values) / (len(values) - 1)

    def _numpy_percentile(self, values: list, percentile: float) -> float:
        """计算百分位数"""
        try:
            import numpy as np
            if not values:
                return 0
            return float(np.percentile(values, percentile * 100))
        except:
            if not values:
                return 0
            sorted_values = sorted(values)
            k = percentile * (len(sorted_values) - 1)
            f = int(k)
            c = k - f
            if f + 1 < len(sorted_values):
                return sorted_values[f] * (1 - c) + sorted_values[f + 1] * c
            else:
                return sorted_values[f]

    def _numpy_quartile(self, values: list, quartile: int) -> float:
        """计算四分位数"""
        quartile_map = {0: 0, 1: 0.25, 2: 0.5, 3: 0.75}
        return self._numpy_percentile(values, quartile_map.get(quartile, 0.5))

    def _numpy_countif(self, values: list, condition: str) -> int:
        """条件计数"""
        try:
            import numpy as np
            arr = np.array(values)

            if condition.startswith('>'):
                threshold = float(condition[1:])
                return int(np.sum(arr > threshold))
            elif condition.startswith('<'):
                threshold = float(condition[1:])
                return int(np.sum(arr < threshold))
            elif condition.startswith('>='):
                threshold = float(condition[2:])
                return int(np.sum(arr >= threshold))
            elif condition.startswith('<='):
                threshold = float(condition[2:])
                return int(np.sum(arr <= threshold))
            elif condition.startswith('='):
                threshold = float(condition[1:])
                return int(np.sum(arr == threshold))
            else:
                threshold = float(condition)
                return int(np.sum(arr == threshold))
        except:
            # 回退到基础实现
            count = 0
            for value in values:
                if condition.startswith('>'):
                    threshold = float(condition[1:])
                    if value > threshold:
                        count += 1
                elif condition.startswith('<'):
                    threshold = float(condition[1:])
                    if value < threshold:
                        count += 1
                # 添加更多条件...
            return count

    def _numpy_sumif(self, values: list, condition: str) -> float:
        """条件求和"""
        try:
            import numpy as np
            arr = np.array(values)

            if condition.startswith('>'):
                threshold = float(condition[1:])
                return float(np.sum(arr[arr > threshold]))
            elif condition.startswith('<'):
                threshold = float(condition[1:])
                return float(np.sum(arr[arr < threshold]))
            elif condition.startswith('>='):
                threshold = float(condition[2:])
                return float(np.sum(arr[arr >= threshold]))
            elif condition.startswith('<='):
                threshold = float(condition[2:])
                return float(np.sum(arr[arr <= threshold]))
            elif condition.startswith('='):
                threshold = float(condition[1:])
                return float(np.sum(arr[arr == threshold]))
            else:
                threshold = float(condition)
                return float(np.sum(arr[arr == threshold]))
        except:
            # 回退到基础实现
            total = 0
            for value in values:
                if condition.startswith('>'):
                    threshold = float(condition[1:])
                    if value > threshold:
                        total += value
                # 添加更多条件...
            return total

    def _numpy_averageif(self, values: list, condition: str) -> float:
        """条件平均值"""
        try:
            import numpy as np
            arr = np.array(values)

            if condition.startswith('>'):
                threshold = float(condition[1:])
                filtered = arr[arr > threshold]
                return float(np.mean(filtered)) if len(filtered) > 0 else 0
            elif condition.startswith('<'):
                threshold = float(condition[1:])
                filtered = arr[arr < threshold]
                return float(np.mean(filtered)) if len(filtered) > 0 else 0
            # 添加更多条件...
        except:
            # 回退到基础实现
            filtered_values = []
            for value in values:
                if condition.startswith('>'):
                    threshold = float(condition[1:])
                    if value > threshold:
                        filtered_values.append(value)
            return sum(filtered_values) / len(filtered_values) if filtered_values else 0

    def _numpy_mode(self, values: list) -> float:
        """计算众数"""
        try:
            from scipy import stats
            if not values:
                return 0
            mode_result = stats.mode(values, keepdims=True)
            return float(mode_result[0][0])
        except:
            # 简单实现：返回最频繁出现的值
            if not values:
                return 0
            from collections import Counter
            counts = Counter(values)
            return float(counts.most_common(1)[0][0])

    def _numpy_skew(self, values: list) -> float:
        """计算偏度"""
        try:
            from scipy import stats
            if len(values) < 3:
                return 0
            return float(stats.skew(values))
        except:
            return 0

    def _numpy_kurtosis(self, values: list) -> float:
        """计算峰度"""
        try:
            from scipy import stats
            if len(values) < 4:
                return 0
            return float(stats.kurtosis(values))
        except:
            return 0

    def _numpy_geomean(self, values: list) -> float:
        """计算几何平均数"""
        try:
            from scipy import stats
            if not values or any(v <= 0 for v in values):
                return 0
            return float(stats.gmean(values))
        except:
            if not values or any(v <= 0 for v in values):
                return 0
            import math
            product = 1
            for v in values:
                product *= v
            return product ** (1.0 / len(values))

    def _numpy_harmean(self, values: list) -> float:
        """计算调和平均数"""
        try:
            from scipy import stats
            if not values or any(v <= 0 for v in values):
                return 0
            return float(stats.hmean(values))
        except:
            if not values or any(v <= 0 for v in values):
                return 0
            return len(values) / sum(1.0 / v for v in values)

    def _calculate_range_sum(self, sheet, start_cell: str, end_cell: str) -> float:
        """计算范围求和"""
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
        total = 0

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    total += cell.value

        return total

    def _calculate_range_count(self, sheet, start_cell: str, end_cell: str) -> int:
        """计算范围内数值个数"""
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
        count = 0

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    count += 1

        return count

    def _apply_cell_format(self, cell, formatting: dict):
        """应用单元格格式"""
        from openpyxl.styles import Font, PatternFill, Border, Alignment

        # 字体格式
        if 'font' in formatting:
            font_config = formatting['font']
            cell.font = Font(
                name=font_config.get('name', cell.font.name),
                size=font_config.get('size', cell.font.size),
                bold=font_config.get('bold', cell.font.bold),
                italic=font_config.get('italic', cell.font.italic),
                color=font_config.get('color', cell.font.color)
            )

        # 背景颜色
        if 'fill' in formatting:
            fill_config = formatting['fill']
            cell.fill = PatternFill(
                start_color=fill_config.get('color', 'FFFFFF'),
                end_color=fill_config.get('color', 'FFFFFF'),
                fill_type='solid'
            )

        # 对齐方式
        if 'alignment' in formatting:
            align_config = formatting['alignment']
            cell.alignment = Alignment(
                horizontal=align_config.get('horizontal', cell.alignment.horizontal),
                vertical=align_config.get('vertical', cell.alignment.vertical)
            )
