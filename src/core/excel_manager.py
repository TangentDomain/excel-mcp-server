"""
Excel MCP Server - Excel管理模块

提供Excel文件和工作表管理功能
"""

import logging
from typing import List, Optional
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from ..models.types import SheetInfo, OperationResult
from ..utils.validators import ExcelValidator
from ..utils.exceptions import SheetNotFoundError, DataValidationError

# 设置日志级别为INFO以便调试
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)


class ExcelManager:
    """Excel文件和工作表管理器"""

    @classmethod
    def create_file(
        cls,
        file_path: str,
        sheet_names: Optional[List[str]] = None
    ) -> OperationResult:
        """
        创建新的Excel文件

        Args:
            file_path: 要创建的Excel文件路径
            sheet_names: 工作表名称列表

        Returns:
            OperationResult: 创建操作的结果
        """
        try:
            # 验证路径，默认允许覆盖已存在的文件
            validated_path = ExcelValidator.validate_file_for_creation(file_path, overwrite=True)

            # 创建工作簿
            workbook = Workbook()

            # 处理工作表
            if sheet_names:
                # 删除默认工作表
                default_sheet = workbook.active
                workbook.remove(default_sheet)

                # 创建指定的工作表
                created_sheets = []
                for i, sheet_name in enumerate(sheet_names):
                    if not sheet_name or not sheet_name.strip():
                        raise DataValidationError(f"工作表名称不能为空: 索引 {i}")

                    sheet = workbook.create_sheet(title=sheet_name.strip())
                    created_sheets.append(SheetInfo(
                        index=i,
                        name=sheet.title,
                        is_active=i == 0,
                        max_row=1,
                        max_column=1,
                        max_column_letter='A'
                    ))

                # 设置第一个工作表为活动工作表
                if created_sheets:
                    workbook.active = workbook[created_sheets[0].name]
            else:
                # 使用默认工作表
                created_sheets = [SheetInfo(
                    index=0,
                    name='Sheet1',
                    is_active=True,
                    max_row=1,
                    max_column=1,
                    max_column_letter='A'
                )]

            # 确保目录存在
            Path(validated_path).parent.mkdir(parents=True, exist_ok=True)

            # 保存文件
            workbook.save(validated_path)

            return OperationResult(
                success=True,
                data=created_sheets,
                message=f"成功创建Excel文件，包含{len(created_sheets)}个工作表",
                metadata={
                    'file_path': validated_path,
                    'total_sheets': len(created_sheets)
                }
            )

        except Exception as e:
            logger.error(f"创建Excel文件失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def __init__(self, file_path: str):
        """
        初始化Excel管理器

        Args:
            file_path: Excel文件路径
        """
        self.file_path = ExcelValidator.validate_file_path(file_path)

    def create_sheet(
        self,
        sheet_name: str,
        index: Optional[int] = None
    ) -> OperationResult:
        """
        在Excel文件中创建新工作表，支持中文字符

        Args:
            sheet_name: 新工作表名称
            index: 插入位置索引

        Returns:
            OperationResult: 创建操作的结果
        """
        try:
            # 验证工作表名称
            if not sheet_name or not sheet_name.strip():
                raise DataValidationError("工作表名称不能为空")

            # 规范化工作表名称，处理中文字符
            sheet_name = self._normalize_sheet_name(sheet_name.strip())

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 检查工作表名称是否已存在
            if sheet_name in workbook.sheetnames:
                raise DataValidationError(f"工作表名称已存在: {sheet_name}")

            # 验证索引
            total_sheets = len(workbook.sheetnames)
            if index is not None:
                if index < 0 or index > total_sheets:
                    raise DataValidationError(f"索引超出范围: {index}，应在 0-{total_sheets} 之间")

            # 创建新工作表，使用安全的编码处理
            try:
                new_sheet = workbook.create_sheet(title=sheet_name, index=index)
            except Exception as sheet_error:
                # 如果直接创建失败，尝试使用ASCII兼容的名称
                logger.warning(f"创建工作表失败，尝试备用方法: {sheet_error}")
                fallback_name = self._create_fallback_name(sheet_name, workbook.sheetnames)
                new_sheet = workbook.create_sheet(title=fallback_name, index=index)
                logger.info(f"使用备用名称创建工作表: {fallback_name}")
                sheet_name = fallback_name

            # 保存文件
            workbook.save(self.file_path)

            # 获取新工作表信息
            sheet_info = SheetInfo(
                index=workbook.sheetnames.index(sheet_name),
                name=new_sheet.title,
                is_active=new_sheet == workbook.active,
                max_row=1,
                max_column=1,
                max_column_letter='A'
            )

            return OperationResult(
                success=True,
                data=sheet_info,
                message=f"成功创建工作表: {sheet_name}",
                metadata={
                    'file_path': self.file_path,
                    'total_sheets': len(workbook.sheetnames),
                    'all_sheets': workbook.sheetnames
                }
            )

        except Exception as e:
            logger.error(f"创建工作表失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def _normalize_sheet_name(self, name: str) -> str:
        """
        规范化工作表名称，确保与Excel兼容

        Args:
            name: 原始工作表名称

        Returns:
            str: 规范化后的名称
        """
        import re

        # Excel工作表名称限制：
        # 1. 不能超过31个字符
        # 2. 不能包含: / \ ? * [ ] :
        # 3. 不能以单引号开头或结尾

        # 移除或替换无效字符
        invalid_chars = r'[/\\?*\[\]:]'
        name = re.sub(invalid_chars, '_', name)

        # 移除首尾的单引号和空格
        name = name.strip("' \t\n\r")

        # 限制长度（考虑中文字符占用更多字节）
        if len(name) > 31:
            # 对于中文字符，保守截取到25个字符
            name = name[:25] + "..."

        # 确保名称不为空
        if not name:
            name = "Sheet"

        return name

    def _create_fallback_name(self, original_name: str, existing_names: list) -> str:
        """
        创建备用工作表名称

        Args:
            original_name: 原始名称
            existing_names: 已存在的名称列表

        Returns:
            str: 备用名称
        """
        import re

        # 尝试创建ASCII兼容的名称
        fallback_base = "Sheet"

        # 尝试从原始名称中提取英文字符
        ascii_chars = re.findall(r'[a-zA-Z0-9]', original_name)
        if ascii_chars:
            fallback_base = ''.join(ascii_chars)[:10]  # 最多取10个字符

        # 确保名称唯一
        counter = 1
        fallback_name = fallback_base
        while fallback_name in existing_names:
            fallback_name = f"{fallback_base}_{counter}"
            counter += 1

        return fallback_name

    def delete_sheet(self, sheet_name: str) -> OperationResult:
        """
        删除Excel文件中的工作表

        Args:
            sheet_name: 要删除的工作表名称

        Returns:
            OperationResult: 删除操作的结果
        """
        try:
            # 验证工作表名称
            if not sheet_name or not sheet_name.strip():
                raise DataValidationError("工作表名称不能为空")

            sheet_name = sheet_name.strip()

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 检查工作表是否存在
            if sheet_name not in workbook.sheetnames:
                raise SheetNotFoundError(f"工作表不存在: {sheet_name}")

            # 检查是否为最后一个工作表
            if len(workbook.sheetnames) <= 1:
                raise DataValidationError("无法删除最后一个工作表，Excel文件至少需要一个工作表")

            # 记录删除前的信息
            deleted_sheet_index = workbook.sheetnames.index(sheet_name)
            was_active = workbook[sheet_name] == workbook.active
            deleted_sheet_name = sheet_name  # 保存要删除的工作表名称
            
            logger.info(f"准备删除工作表: {deleted_sheet_name}, 索引: {deleted_sheet_index}")
            logger.info(f"删除前工作表列表: {workbook.sheetnames}")

            # 删除工作表
            workbook.remove(workbook[sheet_name])
            
            logger.info(f"删除后工作表列表: {workbook.sheetnames}")

            # 如果删除的是活动工作表，设置新的活动工作表
            if was_active:
                if deleted_sheet_index < len(workbook.sheetnames):
                    workbook.active = deleted_sheet_index
                else:
                    workbook.active = deleted_sheet_index - 1

            # 保存文件
            workbook.save(self.file_path)

            # 返回更新后的工作表信息
            remaining_sheet_infos = []
            for i, sheet_name_iter in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name_iter]
                remaining_sheet_infos.append(SheetInfo(
                    index=i,
                    name=sheet_name_iter,
                    is_active=sheet == workbook.active,
                    max_row=sheet.max_row,
                    max_column=sheet.max_column,
                    max_column_letter=get_column_letter(sheet.max_column)
                ))

            return OperationResult(
                success=True,
                data=remaining_sheet_infos,  # 添加data字段
                message=f"成功删除工作表: {deleted_sheet_name}",
                metadata={
                    'file_path': self.file_path,
                    'deleted_sheet': deleted_sheet_name,
                    'deleted_index': deleted_sheet_index,
                    'was_active': was_active,
                    'new_active_sheet': workbook.active.title,
                    'remaining_sheets': workbook.sheetnames,
                    'total_sheets': len(workbook.sheetnames)
                }
            )

        except Exception as e:
            logger.error(f"删除工作表失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )

    def rename_sheet(self, old_name: str, new_name: str) -> OperationResult:
        """
        重命名Excel文件中的工作表

        Args:
            old_name: 原工作表名称
            new_name: 新工作表名称

        Returns:
            OperationResult: 重命名操作的结果
        """
        try:
            # 验证工作表名称
            if not old_name or not old_name.strip():
                raise DataValidationError("原工作表名称不能为空")
            if not new_name or not new_name.strip():
                raise DataValidationError("新工作表名称不能为空")

            old_name = old_name.strip()
            new_name = new_name.strip()

            if old_name == new_name:
                raise DataValidationError("新名称与原名称相同，无需重命名")

            # 加载Excel文件
            workbook = load_workbook(self.file_path)

            # 检查原工作表是否存在
            if old_name not in workbook.sheetnames:
                raise SheetNotFoundError(f"原工作表不存在: {old_name}")

            # 检查新名称是否已存在
            if new_name in workbook.sheetnames:
                raise DataValidationError(f"新工作表名称已存在: {new_name}")

            # 获取工作表
            sheet = workbook[old_name]
            old_index = workbook.sheetnames.index(old_name)
            was_active = sheet == workbook.active

            # 重命名工作表
            sheet.title = new_name

            # 保存文件
            workbook.save(self.file_path)

            # 构建重命名后的工作表信息
            renamed_sheet_info = SheetInfo(
                index=old_index,
                name=new_name,
                is_active=was_active,
                max_row=sheet.max_row,
                max_column=sheet.max_column,
                max_column_letter=get_column_letter(sheet.max_column) if sheet.max_column > 0 else 'A'
            )

            return OperationResult(
                success=True,
                message=f"成功将工作表 '{old_name}' 重命名为 '{new_name}'",
                data=renamed_sheet_info,
                metadata={
                    'file_path': self.file_path,
                    'old_name': old_name,
                    'new_name': new_name,
                    'sheet_index': old_index,
                    'is_active': was_active,
                    'all_sheets': workbook.sheetnames
                }
            )

        except Exception as e:
            logger.error(f"重命名工作表失败: {e}")
            return OperationResult(
                success=False,
                error=str(e)
            )
