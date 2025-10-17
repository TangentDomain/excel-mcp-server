#!/usr/bin/env python3
"""
Excel MCP Server - 基于 FastMCP 和 openpyxl 实现

重构后的服务器文件，只包含MCP接口定义，具体实现委托给核心模块

主要功能：
1. 正则搜索：在Excel文件中搜索符合正则表达式的单元格
2. 范围获取：读取指定范围的Excel数据
3. 范围修改：修改指定范围的Excel数据
4. 工作表管理：创建、删除、重命名工作表
5. 行列操作：插入、删除行列

技术栈：
- FastMCP: 用于MCP服务器框架
- openpyxl: 用于Excel文件操作
"""

import logging
import os
import shutil
from datetime import datetime
from enum import Enum
from typing import Optional, List, Dict, Any, Union

try:
    from mcp.server.fastmcp import FastMCP
except ImportError as e:
    print(f"Error: 缺少必要的依赖包: {e}")
    print("请运行: pip install fastmcp openpyxl")
    exit(1)

# 导入API模块
from .api.excel_operations import ExcelOperations

# ==================== 操作日志系统 ====================
class OperationLogger:
    """操作日志记录器，用于跟踪所有Excel操作"""

    def __init__(self):
        self.log_file = None
        self.current_session = []

    def start_session(self, file_path: str):
        """开始新的操作会话"""
        self.log_file = os.path.join(
            os.path.dirname(file_path),
            ".excel_mcp_logs",
            f"operations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        )

        os.makedirs(os.path.dirname(self.log_file), exist_ok=True)

        self.current_session = [{
            'session_id': datetime.now().isoformat(),
            'file_path': file_path,
            'operations': []
        }]

        self._save_log()

    def log_operation(self, operation: str, details: Dict[str, Any]):
        """记录操作"""
        if not self.current_session:
            return

        operation_record = {
            'timestamp': datetime.now().isoformat(),
            'operation': operation,
            'details': details
        }

        self.current_session[0]['operations'].append(operation_record)
        self._save_log()

    def _save_log(self):
        """保存日志到文件"""
        if not self.log_file:
            return

        try:
            import json
            with open(self.log_file, 'w', encoding='utf-8') as f:
                json.dump(self.current_session, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logger.error(f"保存操作日志失败: {e}")

    def get_recent_operations(self, limit: int = 10) -> List[Dict[str, Any]]:
        """获取最近的操作记录"""
        if not self.current_session:
            return []

        operations = self.current_session[0]['operations']
        return operations[-limit:] if len(operations) > limit else operations

# 全局操作日志器
operation_logger = OperationLogger()

# ==================== 配置和初始化 ====================
# 开启详细日志用于调试
logging.basicConfig(
    level=logging.DEBUG,  # 改为DEBUG级别获取更多信息
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
    ]
)
logger = logging.getLogger(__name__)

# 创建FastMCP服务器实例，开启调试模式和详细日志
mcp = FastMCP(
    name="excel-mcp",
    instructions=r"""🔒 Excel安全操作专家 - 数据安全第一的Excel配置管理

## 🛡️ 安全第一原则
• **数据保护优先**：所有操作都以保护用户数据安全为首要原则
• **预览必做**：任何修改操作前必须使用 `excel_preview_operation` 预览影响
• **自动备份**：重要操作前自动创建备份，支持 `excel_create_backup`
• **安全默认**：`excel_update_range` 默认使用 `insert_mode=True` 防止数据覆盖
• **操作确认**：高风险操作前进行风险评估和用户确认
• **完整日志**：记录所有操作历史，支持 `excel_get_operation_history` 追踪

## 🔍 安全操作工作流程

### 标准安全更新流程
1. **📋 操作预览**：`excel_preview_operation` → 分析影响范围和风险
2. **💾 创建备份**：`excel_create_backup` → 自动创建操作前备份
3. **📊 数据评估**：`excel_assess_data_impact` → 全面评估操作影响
4. **⚠️ 风险确认**：查看安全建议，确认操作风险等级
5. **✏️ 安全执行**：使用安全默认参数执行操作
6. **✅ 结果验证**：重新读取确认操作结果
7. **📝 日志记录**：操作自动记录到历史日志

### 风险评估工作流
```
🛡️ 安全操作流程:
excel_assess_data_impact() → 获取风险等级
↓ 检查安全建议
🔴 高风险: 强烈建议备份 + 用户确认
🟡 中风险: 建议备份 + 操作验证
🟢 低风险: 常规安全操作
```

### 备份与恢复工作流
```
💾 数据保护流程:
excel_create_backup() → 创建时间戳备份
↓ 执行操作
excel_restore_backup() → 如有问题可立即恢复
excel_list_backups() → 查看所有可用备份
```

## 🛡️ 安全操作指南

### 数据保护最佳实践
```
✅ 安全操作准则:
🔍 操作前预览: excel_preview_operation() → 了解影响范围
💾 重要数据备份: excel_create_backup() → 创建安全备份
📊 风险评估: excel_assess_data_impact() → 获取风险等级
⚠️ 确认后执行: 根据风险等级决定是否继续
📝 操作追踪: excel_get_operation_history() → 查看操作记录
```

### 范围格式安全规范
```
📐 安全范围表达式:
单元格: "Sheet1!A1:C10"       # 标准矩形范围
整行:   "数据表!5:10"          # 第5-10行
整列:   "统计表!B:F"           # B到F列
单行:   "配置表!1"             # 仅第1行
单列:   "记录表!D"             # 仅D列
```

### 安全工具使用指南

| 安全工具 | 用途 | 何时使用 |
|---------|------|---------|
| `excel_preview_operation` | 预览操作影响 | 任何修改前 |
| `excel_assess_data_impact` | 全面风险评估 | 重要操作前 |
| `excel_create_backup` | 创建备份 | 高风险操作前 |
| `excel_restore_backup` | 恢复数据 | 操作失误后 |
| `excel_get_operation_history` | 查看操作记录 | 问题排查时 |

## ⚠️ 风险等级说明

### 🔴 高风险操作
- 影响超过1000个单元格
- 覆盖大量现有数据
- 删除整行或整列
- **强制要求**: 备份 + 用户确认

### 🟡 中风险操作
- 影响100-1000个单元格
- 部分数据覆盖
- 格式化大范围
- **建议**: 备份 + 操作验证

### 🟢 低风险操作
- 影响少于100个单元格
- 空白区域操作
- 读取和查询操作
- **标准**: 常规安全流程

🔒 **安全第一**: 所有Excel操作都将以数据安全为最高优先级，确保用户数据万无一失""",
    debug=True,                    # 开启调试模式
    log_level="DEBUG"              # 设置日志级别为DEBUG
)


# ==================== MCP 工具定义 ====================

@mcp.tool()
def excel_list_sheets(file_path: str) -> Dict[str, Any]:
    """
    列出Excel文件中所有工作表名称

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)

    Returns:
        Dict: 包含success、sheets、total_sheets

    Example:
        # 列出工作表名称
        result = excel_list_sheets("data.xlsx")
        # 返回: {
        #   'success': True,
        #   'sheets': ['Sheet1', 'Sheet2'],
        #   'total_sheets': 2
        # }
    """
    return ExcelOperations.list_sheets(file_path)


@mcp.tool()
def excel_get_sheet_headers(file_path: str) -> Dict[str, Any]:
    """
    获取Excel文件中所有工作表的双行表头信息（游戏开发专用）

    这是 excel_get_headers 的便捷封装，用于批量获取所有工作表的双行表头。
    专为游戏配置表设计，同时获取字段描述和字段名。

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)

    Returns:
        Dict: 包含所有工作表的双行表头信息
        {
            'success': bool,
            'sheets_with_headers': [
                {
                    'name': str,
                    'headers': List[str],       # 字段名（兼容性）
                    'descriptions': List[str],  # 字段描述（第1行）
                    'field_names': List[str],   # 字段名（第2行）
                    'header_count': int
                }
            ],
            'file_path': str,
            'total_sheets': int
        }

    游戏配置表批量分析:
        一次性获取所有配置表的结构信息，包括字段描述和字段名，便于快速了解整个配置文件的结构。

    Example:
        # 获取游戏配置文件中所有表的双行表头
        result = excel_get_sheet_headers("game_config.xlsx")
        for sheet in result['sheets_with_headers']:
            print(f"表名: {sheet['name']}")
            print(f"字段描述: {sheet['descriptions']}")
            print(f"字段名: {sheet['field_names']}")
            print("---")

        # 返回示例: {
        #   'success': True,
        #   'sheets_with_headers': [
        #     {
        #       'name': '技能配置表',
        #       'headers': ['skill_id', 'skill_name', 'skill_type'],
        #       'descriptions': ['技能ID描述', '技能名称描述', '技能类型描述'],
        #       'field_names': ['skill_id', 'skill_name', 'skill_type'],
        #       'header_count': 3
        #     },
        #     {
        #       'name': '装备配置表',
        #       'headers': ['item_id', 'item_name', 'item_quality'],
        #       'descriptions': ['装备ID描述', '装备名称描述', '装备品质描述'],
        #       'field_names': ['item_id', 'item_name', 'item_quality'],
        #       'header_count': 3
        #     }
        #   ],
        #   'total_sheets': 2
        # }
    """
    return ExcelOperations.get_sheet_headers(file_path)


@mcp.tool()
def excel_search(
    file_path: str,
    pattern: str,
    sheet_name: Optional[str] = None,
    case_sensitive: bool = False,
    whole_word: bool = False,
    use_regex: bool = False,
    include_values: bool = True,
    include_formulas: bool = False,
    range: Optional[str] = None
) -> Dict[str, Any]:
    """
    在Excel文件中搜索单元格内容（VSCode风格搜索选项）

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        pattern: 搜索模式。当use_regex=True时为正则表达式，否则为字面字符串
        sheet_name: 工作表名称 (可选，不指定时搜索所有工作表)
        case_sensitive: 大小写敏感 (默认False，即忽略大小写)
        whole_word: 全词匹配 (默认False，即部分匹配)
        use_regex: 启用正则表达式 (默认False，即字面字符串搜索)
        include_values: 是否搜索单元格值
        include_formulas: 是否搜索公式内容
        range: 搜索范围表达式，支持多种格式：
            - 单元格范围: "A1:C10" 或 "Sheet1!A1:C10"
            - 行范围: "3:5" 或 "Sheet1!3:5" (第3行到第5行)
            - 列范围: "B:D" 或 "Sheet1!B:D" (B列到D列)
            - 单行: "7" 或 "Sheet1!7" (仅第7行)
            - 单列: "C" 或 "Sheet1!C" (仅C列)

    Returns:
        Dict: 包含 success、matches(List[Dict])、match_count、searched_sheets

    Example:
        # 普通字符串搜索（默认忽略大小写）
        result = excel_search("data.xlsx", "总计")
        # 大小写敏感搜索
        result = excel_search("data.xlsx", "Total", case_sensitive=True)
        # 全词匹配搜索（只匹配完整单词）
        result = excel_search("data.xlsx", "sum", whole_word=True)
        # 正则表达式搜索邮箱格式
        result = excel_search("data.xlsx", r'\\w+@\\w+\\.\\w+', use_regex=True)
        # 正则表达式搜索数字（大小写敏感）
        result = excel_search("data.xlsx", r'\\d+', use_regex=True, case_sensitive=True)
        # 搜索指定范围
        result = excel_search("data.xlsx", "金额", range="Sheet1!A1:C10", whole_word=True)
        # 搜索指定工作表
        result = excel_search("data.xlsx", "error", sheet_name="Sheet1", case_sensitive=True)
        # 搜索数字并包含公式
        result = excel_search("data.xlsx", r'\\d+', use_regex=True, include_formulas=True)
    """
    return ExcelOperations.search(file_path, pattern, sheet_name, case_sensitive, whole_word, use_regex, include_values, include_formulas, range)


@mcp.tool()
def excel_search_directory(
    directory_path: str,
    pattern: str,
    case_sensitive: bool = False,
    whole_word: bool = False,
    use_regex: bool = False,
    include_values: bool = True,
    include_formulas: bool = False,
    recursive: bool = True,
    file_extensions: Optional[List[str]] = None,
    file_pattern: Optional[str] = None,
    max_files: int = 100
) -> Dict[str, Any]:
    """
    在目录下的所有Excel文件中搜索内容（VSCode风格搜索选项）

    Args:
        directory_path: 目录路径
        pattern: 搜索模式。当use_regex=True时为正则表达式，否则为字面字符串
        case_sensitive: 大小写敏感 (默认False，即忽略大小写)
        whole_word: 全词匹配 (默认False，即部分匹配)
        use_regex: 启用正则表达式 (默认False，即字面字符串搜索)
        include_values: 是否搜索单元格值
        include_formulas: 是否搜索公式内容
        recursive: 是否递归搜索子目录
        file_extensions: 文件扩展名过滤，如[".xlsx", ".xlsm"]
        file_pattern: 文件名正则模式过滤
        max_files: 最大搜索文件数限制

    Returns:
        Dict: 包含 success、matches(List[Dict])、total_matches、searched_files

    Example:
        # 普通字符串搜索目录
        result = excel_search_directory("./data", "总计")
        # 大小写敏感搜索
        result = excel_search_directory("./data", "Error", case_sensitive=True)
        # 全词匹配搜索
        result = excel_search_directory("./data", "sum", whole_word=True)
        # 正则表达式搜索邮箱格式
        result = excel_search_directory("./data", r'\\w+@\\w+\\.\\w+', use_regex=True)
        # 搜索特定文件名模式
        result = excel_search_directory("./reports", r'\\d+', use_regex=True, file_pattern=r'.*销售.*')
    """
    return ExcelOperations.search_directory(directory_path, pattern, case_sensitive, whole_word, use_regex, include_values, include_formulas, recursive, file_extensions, file_pattern, max_files)


@mcp.tool()
def excel_get_range(
    file_path: str,
    range: str,
    include_formatting: bool = False
) -> Dict[str, Any]:
    """
    读取Excel指定范围的数据

    Args:
        file_path (str): Excel文件路径 (.xlsx/.xlsm) [必需]
        range (str): 范围表达式，必须包含工作表名 [必需]
            支持格式：
            - 标准单元格范围: "Sheet1!A1:C10"、"TrSkill!A1:Z100"
            - 行范围: "Sheet1!1:1"、"数据!5:10"
            - 列范围: "Sheet1!A:C"、"统计!B:E"
            - 单行/单列: "Sheet1!5"、"数据!C"
        include_formatting (bool, 可选): 是否包含单元格格式，默认 False

    Returns:
        Dict: 包含 success、data(List[List])、range_info

    注意:
        为保持API一致性和清晰度，range必须包含工作表名。
        这消除了参数间的条件依赖，提高了可预测性。

    Example:
        # 读取单元格范围
        result = excel_get_range("data.xlsx", "Sheet1!A1:C10")
        # 读取整行
        result = excel_get_range("data.xlsx", "Sheet1!1:1")
        # 读取列范围
        result = excel_get_range("data.xlsx", "数据!A:C")
    """
    # 增强参数验证
    from .utils.validators import ExcelValidator, DataValidationError

    try:
        # 验证范围表达式格式
        range_validation = ExcelValidator.validate_range_expression(range)

        # 验证操作规模
        scale_validation = ExcelValidator.validate_operation_scale(range_validation['range_info'])

        # 记录验证成功到调试日志
        logger.debug(f"范围验证成功: {range_validation['normalized_range']}")

    except DataValidationError as e:
        # 记录验证失败
        logger.error(f"范围验证失败: {str(e)}")

        return {
            'success': False,
            'error': 'VALIDATION_FAILED',
            'message': f"范围表达式验证失败: {str(e)}"
        }

    # 调用原始函数
    result = ExcelOperations.get_range(file_path, range, include_formatting)

    # 如果成功，添加验证信息到结果中
    if result.get('success'):
        result['validation_info'] = {
            'normalized_range': range_validation['normalized_range'],
            'sheet_name': range_validation['sheet_name'],
            'range_type': range_validation['range_info']['type'],
            'scale_assessment': scale_validation
        }

    return result


@mcp.tool()
def excel_get_headers(
    file_path: str,
    sheet_name: str,
    header_row: int = 1,
    max_columns: Optional[int] = None
) -> Dict[str, Any]:
    """
    获取Excel工作表的双行表头信息（游戏开发专用）

    专为游戏配置表设计，同时获取字段描述（第1行）和字段名（第2行）

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        header_row: 表头起始行号 (1-based，默认从第1行开始获取两行)
        max_columns: 最大读取列数限制 (可选)
            - 指定数值: 精确读取指定列数，如 max_columns=10 读取A-J列
            - None(默认): 读取前100列范围 (A-CV列)，然后截取到第一个空列

    Returns:
        Dict: 包含双行表头信息
        {
            'success': bool,
            'data': List[str],          # 字段名列表（兼容性）
            'headers': List[str],       # 字段名列表（兼容性）
            'descriptions': List[str],  # 字段描述列表（第1行）
            'field_names': List[str],   # 字段名列表（第2行）
            'header_count': int,
            'sheet_name': str,
            'header_row': int,
            'message': str
        }

    游戏配置表标准格式:
        第1行（descriptions）: ['技能ID描述', '技能名称描述', '技能类型描述', '技能等级描述']
        第2行（field_names）:   ['skill_id', 'skill_name', 'skill_type', 'skill_level']

    Example:
        # 获取技能配置表的双行表头
        result = excel_get_headers("skills.xlsx", "技能配置表")
        print(result['descriptions'])  # ['技能ID描述', '技能名称描述', ...]
        print(result['field_names'])   # ['skill_id', 'skill_name', ...]

        # 获取装备表第3-4行作为表头，精确读取8列
        result = excel_get_headers("items.xlsx", "装备配置表", header_row=3, max_columns=8)
    """
    return ExcelOperations.get_headers(file_path, sheet_name, header_row, max_columns)


@mcp.tool()
def excel_update_range(
    file_path: str,
    range: str,
    data: List[List[Any]],
    preserve_formulas: bool = True,
    insert_mode: bool = True,
    require_confirmation: bool = False,
    skip_safety_checks: bool = False
) -> Dict[str, Any]:
    """
更新Excel指定范围的数据。默认使用安全的插入模式。

Args:
    file_path: Excel文件路径 (.xlsx/.xlsm)
    range: 范围表达式，必须包含工作表名，支持格式：
        - 标准单元格范围: "Sheet1!A1:C10"、"TrSkill!A1:Z100"
        - 不支持行范围格式，必须使用明确单元格范围
    data: 二维数组数据 [[row1], [row2], ...]
    preserve_formulas: 保留已有公式 (默认值: True)
        - True: 如果目标单元格包含公式，则保留公式不覆盖
        - False: 覆盖所有内容，包括公式
    insert_mode: 数据写入模式 (默认值: True - 安全优先)
        - True: 插入模式，在指定位置插入新行然后写入数据（默认安全）
        - False: 覆盖模式，直接覆盖目标范围的现有数据（谨慎使用）
    require_confirmation: 是否需要用户确认 (默认值: False)
        - True: 高风险操作需要用户确认
        - False: 自动进行安全检查和风险评估
    skip_safety_checks: 跳过安全检查 (默认值: False)
        - True: 跳过所有安全检查（仅限系统维护使用）
        - False: 执行完整的安全检查流程

Returns:
    Dict: 包含 success、updated_cells(int)、message

⚠️ 安全提示:
    - 默认使用插入模式防止数据覆盖
    - 如需覆盖现有数据，请明确设置 insert_mode=False
    - 建议先使用 excel_get_range 预览当前数据

Example:
    data = [["姓名", "年龄"], ["张三", 25]]
    # 安全插入模式（默认）
    result = excel_update_range("test.xlsx", "Sheet1!A1:B2", data)
    # 覆盖模式（需要明确指定）
    result = excel_update_range("test.xlsx", "Sheet1!A1:B2", data, insert_mode=False)
    """
    # 增强参数验证
    from .utils.validators import ExcelValidator, DataValidationError

    try:
        # 验证范围表达式格式
        range_validation = ExcelValidator.validate_range_expression(range)

        # 验证操作规模
        scale_validation = ExcelValidator.validate_operation_scale(range_validation['range_info'])

        # 如果有警告信息，记录到操作日志
        if scale_validation.get('warning'):
            logger.warning(f"操作规模警告: {scale_validation['warning']}")

    except DataValidationError as e:
        # 记录验证失败
        operation_logger.start_session(file_path)
        operation_logger.log_operation("validation_failed", {
            "operation": "update_range",
            "range": range,
            "error": str(e)
        })

        return {
            'success': False,
            'error': 'VALIDATION_FAILED',
            'message': f"参数验证失败: {str(e)}"
        }

    # 开始操作会话
    operation_logger.start_session(file_path)

    # 记录操作日志
    operation_logger.log_operation("update_range", {
        "range": range,
        "validated_range": range_validation['normalized_range'],
        "data_rows": len(data),
        "insert_mode": insert_mode,
        "preserve_formulas": preserve_formulas,
        "scale_info": scale_validation
    })

    try:
        result = ExcelOperations.update_range(file_path, range, data, preserve_formulas, insert_mode, require_confirmation, skip_safety_checks)

        # 记录操作结果
        operation_logger.log_operation("operation_result", {
            "success": result.get('success', False),
            "updated_cells": result.get('updated_cells', 0),
            "message": result.get('message', '')
        })

        return result

    except Exception as e:
        # 记录错误
        operation_logger.log_operation("operation_error", {
            "error": str(e),
            "message": f"更新操作失败: {str(e)}"
        })

        return {
            'success': False,
            'error': 'OPERATION_FAILED',
            'message': f"更新操作失败: {str(e)}"
        }


@mcp.tool()
def excel_preview_operation(
    file_path: str,
    range: str,
    operation_type: str = "update",
    data: Optional[List[List[Any]]] = None
) -> Dict[str, Any]:
    """
    预览Excel操作的影响范围和当前数据，确保安全操作

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        range: 范围表达式，必须包含工作表名
        operation_type: 操作类型 ("update", "delete", "format")
        data: 对于更新操作，提供将要写入的数据

    Returns:
        Dict: 包含预览信息、当前数据、影响评估

    Example:
        # 预览更新操作
        result = excel_preview_operation("data.xlsx", "Sheet1!A1:C10", "update", new_data)
        # 预览删除操作
        result = excel_preview_operation("data.xlsx", "Sheet1!5:10", "delete")
    """
    # 读取当前数据
    current_data = ExcelOperations.get_range(file_path, range)

    if not current_data.get('success'):
        return {
            'success': False,
            'error': 'PREVIEW_FAILED',
            'message': f"无法预览操作: {current_data.get('message', '未知错误')}"
        }

    # 分析影响
    data_rows = len(current_data.get('data', []))
    data_cols = len(current_data.get('data', [])) if data_rows > 0 else 0
    total_cells = data_rows * data_cols

    # 检查是否包含非空数据
    has_data = any(
        any(cell is not None and str(cell).strip() for cell in row)
        for row in current_data.get('data', [])
    )

    # 安全评估
    risk_level = "LOW"
    if has_data:
        if total_cells > 100:
            risk_level = "HIGH"
        elif total_cells > 20:
            risk_level = "MEDIUM"
        else:
            risk_level = "LOW"

    return {
        'success': True,
        'operation_type': operation_type,
        'range': range,
        'current_data': current_data.get('data', []),
        'impact_assessment': {
            'rows_affected': data_rows,
            'columns_affected': data_cols,
            'total_cells': total_cells,
            'has_existing_data': has_data,
            'risk_level': risk_level
        },
        'recommendations': _get_safety_recommendations(operation_type, has_data, risk_level),
        'safety_warning': _generate_safety_warning(operation_type, has_data, risk_level)
    }


def _get_safety_recommendations(operation_type: str, has_data: bool, risk_level: str) -> List[str]:
    """获取安全操作建议"""
    recommendations = []

    if operation_type == "update":
        if has_data:
            recommendations.append("⚠️ 范围内已有数据，建议使用 insert_mode=True")
            if risk_level == "HIGH":
                recommendations.append("🔴 大范围数据操作，强烈建议先备份")
            recommendations.append("📊 建议先预览完整数据再操作")
        else:
            recommendations.append("✅ 范围为空，可以安全操作")

    elif operation_type == "delete":
        recommendations.append("🗑️ 删除操作不可逆，请确认")
        if has_data:
            recommendations.append("⚠️ 将删除现有数据，请仔细检查")

    return recommendations


def _generate_safety_warning(operation_type: str, has_data: bool, risk_level: str) -> str:
    """生成安全警告"""
    if risk_level == "HIGH":
        return f"🔴 高风险警告: {operation_type}操作将影响大量数据，请谨慎操作"
    elif risk_level == "MEDIUM":
        return f"🟡 中等风险: {operation_type}操作将影响部分数据，建议先备份"
    else:
        return f"✅ 低风险: {operation_type}操作影响较小，可以安全执行"


@mcp.tool()
def excel_assess_data_impact(
    file_path: str,
    range: str,
    operation_type: str = "update",
    data: Optional[List[List[Any]]] = None
) -> Dict[str, Any]:
    """
    全面评估Excel操作对数据的潜在影响

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        range: 范围表达式，必须包含工作表名
        operation_type: 操作类型 ("update", "delete", "format")
        data: 对于更新操作，提供将要写入的数据

    Returns:
        Dict: 包含详细的数据影响评估报告

    Example:
        # 评估更新操作的影响
        result = excel_assess_data_impact("data.xlsx", "Sheet1!A1:C10", "update", new_data)
        # 评估删除操作的影响
        result = excel_assess_data_impact("data.xlsx", "Sheet1!5:10", "delete")
    """
    from .utils.validators import ExcelValidator, DataValidationError

    try:
        # 验证范围表达式
        range_validation = ExcelValidator.validate_range_expression(range)
        range_info = range_validation['range_info']

        # 获取当前数据
        current_data_result = ExcelOperations.get_range(file_path, range)

        if not current_data_result.get('success'):
            return {
                'success': False,
                'error': 'DATA_RETRIEVAL_FAILED',
                'message': f"无法获取当前数据: {current_data_result.get('message', '未知错误')}"
            }

        current_data = current_data_result.get('data', [])

        # 分析当前数据内容
        data_analysis = _analyze_current_data(current_data)

        # 计算操作规模
        scale_info = ExcelValidator.validate_operation_scale(range_info)

        # 评估操作风险
        risk_assessment = _assess_operation_risk(
            operation_type,
            data_analysis,
            scale_info,
            data
        )

        # 生成建议
        recommendations = _generate_safety_recommendations(
            operation_type,
            data_analysis,
            risk_assessment,
            scale_info
        )

        # 预测结果
        prediction = _predict_operation_result(
            operation_type,
            current_data,
            data,
            scale_info
        )

        return {
            'success': True,
            'operation_type': operation_type,
            'range': range,
            'validation_info': range_validation,
            'current_data_analysis': data_analysis,
            'scale_assessment': scale_info,
            'risk_assessment': risk_assessment,
            'safety_recommendations': recommendations,
            'result_prediction': prediction,
            'impact_summary': {
                'total_cells': scale_info['total_cells'],
                'non_empty_cells': data_analysis['non_empty_cell_count'],
                'data_type_distribution': data_analysis['data_types'],
                'potential_data_loss': data_analysis['non_empty_cell_count'] if operation_type in ['delete', 'update'] else 0,
                'overall_risk_level': risk_assessment['overall_risk']
            }
        }

    except DataValidationError as e:
        return {
            'success': False,
            'error': 'VALIDATION_FAILED',
            'message': f"参数验证失败: {str(e)}"
        }

    except Exception as e:
        return {
            'success': False,
            'error': 'ASSESSMENT_FAILED',
            'message': f"数据影响评估失败: {str(e)}"
        }


def _analyze_current_data(data: List[List[Any]]) -> Dict[str, Any]:
    """分析当前数据内容"""
    if not data:
        return {
            'row_count': 0,
            'column_count': 0,
            'total_cells': 0,
            'non_empty_cell_count': 0,
            'empty_cell_count': 0,
            'data_types': {},
            'has_formulas': False,
            'has_numeric_data': False,
            'has_text_data': False,
            'has_dates': False,
            'completeness_rate': 0.0
        }

    total_cells = len(data) * max(len(row) for row in data) if data else 0
    non_empty_cells = 0
    data_types = {}
    has_formulas = False
    has_numeric_data = False
    has_text_data = False
    has_dates = False

    for row in data:
        for cell in row:
            if cell is not None and str(cell).strip():
                non_empty_cells += 1

                # 分析数据类型
                if isinstance(cell, str):
                    if cell.startswith('='):
                        has_formulas = True
                        data_types['formulas'] = data_types.get('formulas', 0) + 1
                    else:
                        has_text_data = True
                        data_types['text'] = data_types.get('text', 0) + 1
                elif isinstance(cell, (int, float)):
                    has_numeric_data = True
                    data_types['numeric'] = data_types.get('numeric', 0) + 1
                else:
                    data_types['other'] = data_types.get('other', 0) + 1

    return {
        'row_count': len(data),
        'column_count': max(len(row) for row in data) if data else 0,
        'total_cells': total_cells,
        'non_empty_cell_count': non_empty_cells,
        'empty_cell_count': total_cells - non_empty_cells,
        'data_types': data_types,
        'has_formulas': has_formulas,
        'has_numeric_data': has_numeric_data,
        'has_text_data': has_text_data,
        'has_dates': has_dates,
        'completeness_rate': (non_empty_cells / total_cells * 100) if total_cells > 0 else 0.0
    }


def _assess_operation_risk(
    operation_type: str,
    data_analysis: Dict[str, Any],
    scale_info: Dict[str, Any],
    new_data: Optional[List[List[Any]]] = None
) -> Dict[str, Any]:
    """评估操作风险"""
    risk_factors = []
    risk_score = 0

    # 基于操作类型的风险
    if operation_type == "delete":
        risk_factors.append("删除操作不可逆")
        risk_score += 30
    elif operation_type == "update":
        if data_analysis['non_empty_cell_count'] > 0:
            risk_factors.append("将覆盖现有数据")
            risk_score += 20
    elif operation_type == "format":
        risk_factors.append("格式化操作")
        risk_score += 10

    # 基于数据量的风险
    if scale_info['total_cells'] > 10000:
        risk_factors.append("大范围操作")
        risk_score += 25
    elif scale_info['total_cells'] > 1000:
        risk_factors.append("中等范围操作")
        risk_score += 15

    # 基于数据内容的风险
    if data_analysis['has_formulas']:
        risk_factors.append("包含公式数据")
        risk_score += 15

    if data_analysis['completeness_rate'] > 80:
        risk_factors.append("高密度数据区域")
        risk_score += 10

    # 确定整体风险等级
    if risk_score >= 60:
        overall_risk = "HIGH"
    elif risk_score >= 30:
        overall_risk = "MEDIUM"
    else:
        overall_risk = "LOW"

    return {
        'risk_score': risk_score,
        'overall_risk': overall_risk,
        'risk_factors': risk_factors,
        'requires_backup': overall_risk in ["HIGH", "MEDIUM"],
        'requires_confirmation': overall_risk == "HIGH"
    }


def _generate_safety_recommendations(
    operation_type: str,
    data_analysis: Dict[str, Any],
    risk_assessment: Dict[str, Any],
    scale_info: Dict[str, Any]
) -> List[str]:
    """生成安全建议"""
    recommendations = []

    # 基础建议
    if risk_assessment['requires_backup']:
        recommendations.append("🔴 强烈建议在操作前创建备份")

    if risk_assessment['requires_confirmation']:
        recommendations.append("⚠️ 高风险操作，请仔细确认后再执行")

    # 基于数据内容的建议
    if data_analysis['has_formulas']:
        recommendations.append("📊 检测到公式数据，建议验证公式的正确性")

    if data_analysis['completeness_rate'] > 50:
        recommendations.append("💾 数据密度较高，建议先导出重要数据")

    # 基于操作类型的建议
    if operation_type == "delete":
        recommendations.append("🗑️ 删除操作不可逆，请确认数据不再需要")
    elif operation_type == "update":
        if data_analysis['non_empty_cell_count'] > 0:
            recommendations.append("✏️ 将覆盖现有数据，建议使用insert_mode=True")

    # 性能建议
    if scale_info['total_cells'] > 5000:
        recommendations.append("⏱️ 大范围操作可能需要较长时间，请耐心等待")

    return recommendations


def _predict_operation_result(
    operation_type: str,
    current_data: List[List[Any]],
    new_data: Optional[List[List[Any]]],
    scale_info: Dict[str, Any]
) -> Dict[str, Any]:
    """预测操作结果"""
    prediction = {
        'affected_cells': scale_info['total_cells'],
        'data_overwrite_count': 0,
        'data_insert_count': 0,
        'estimated_time': "minimal"
    }

    if operation_type == "update" and new_data:
        prediction['data_overwrite_count'] = len([cell for row in current_data for cell in row if cell is not None])
        prediction['data_insert_count'] = len([cell for row in new_data for cell in row if cell is not None])
    elif operation_type == "delete":
        prediction['data_overwrite_count'] = len([cell for row in current_data for cell in row if cell is not None])

    # 估算执行时间
    if scale_info['total_cells'] > 10000:
        prediction['estimated_time'] = "long"
    elif scale_info['total_cells'] > 1000:
        prediction['estimated_time'] = "medium"

    return prediction


@mcp.tool()
def excel_check_danger_level(
    file_path: str,
    operation_type: str = "update",
    range: Optional[str] = None,
    data: Optional[List[List[Any]]] = None,
    data_count: Optional[int] = None
) -> Dict[str, Any]:
    """
    检查操作的危险等级，为大范围操作提供特别警告

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        operation_type: 操作类型 ("update", "delete", "format", "insert", "delete_rows", "delete_columns")
        range: 范围表达式 (可选，用于精确分析)
        data: 对于更新操作，提供将要写入的数据 (可选)
        data_count: 数据行数 (可选，用于快速评估)

    Returns:
        Dict: 包含危险等级评估和详细警告信息

    Example:
        # 检查操作危险等级
        result = excel_check_danger_level("data.xlsx", "update", "Sheet1!A1:Z1000")
        # 快速检查数据行数
        result = excel_check_danger_level("data.xlsx", "update", data_count=5000)
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            return {
                'success': False,
                'error': 'FILE_NOT_FOUND',
                'message': f"文件不存在: {file_path}"
            }

        # 分析操作规模
        total_cells = 0
        estimated_data_loss = 0
        risk_factors = []
        risk_score = 0

        if range:
            # 基于范围计算
            from .utils.validators import ExcelValidator, DataValidationError
            try:
                range_validation = ExcelValidator.validate_range_expression(range)
                range_info = range_validation['range_info']
                scale_info = ExcelValidator.validate_operation_scale(range_info)
                total_cells = scale_info['total_cells']
            except DataValidationError as e:
                return {
                    'success': False,
                    'error': 'INVALID_RANGE',
                    'message': f"无效的范围表达式: {str(e)}"
                }
        elif data_count:
            # 基于数据行数计算 (假设每行10个单元格)
            total_cells = data_count * 10

        # 获取文件信息用于风险评估
        file_info = os.stat(file_path)
        file_size_mb = file_info.st_size / (1024 * 1024)

        # 危险因素评估
        risk_factors = []
        risk_score = 0

        # 基于单元格数量的风险
        if total_cells > 50000:
            risk_factors.append("🔴 极大范围操作 (>50,000单元格)")
            risk_score += 40
        elif total_cells > 10000:
            risk_factors.append("🟠 大范围操作 (>10,000单元格)")
            risk_score += 30
        elif total_cells > 1000:
            risk_factors.append("🟡 中等范围操作 (>1,000单元格)")
            risk_score += 20
        elif total_cells > 100:
            risk_factors.append("🟢 小范围操作 (>100单元格)")
            risk_score += 10

        # 基于文件大小的风险
        if file_size_mb > 100:
            risk_factors.append("🔴 大文件操作 (>100MB)")
            risk_score += 15
        elif file_size_mb > 50:
            risk_factors.append("🟡 中等文件大小 (>50MB)")
            risk_score += 10
        elif file_size_mb > 10:
            risk_factors.append("🟢 较大文件 (>10MB)")
            risk_score += 5

        # 基于操作类型的风险
        if operation_type in ["delete", "delete_rows", "delete_columns"]:
            risk_factors.append("🔴 删除操作不可逆")
            risk_score += 25
        elif operation_type == "update":
            risk_factors.append("⚠️ 更新操作可能覆盖数据")
            risk_score += 15
        elif operation_type == "format":
            risk_factors.append("⚡️ 格式化操作")
            risk_score += 10

        # 特殊情况：文件锁定检测
        try:
            # 检查文件是否被其他程序锁定
            import time
            import msvcrt
            try:
                # 尝试重命名文件来检测锁定状态
                temp_path = file_path + ".lock_check_" + str(int(time.time()))
                os.rename(file_path, temp_path)
                os.rename(temp_path, file_path)
            except PermissionError:
                risk_factors.append("🔴 文件可能被其他程序锁定")
                risk_score += 20
            except OSError:
                risk_factors.append("🟡 文件访问受限")
                risk_score += 10
        except Exception:
            # 忽略文件锁定检测错误
            pass

        # 确定危险等级
        if risk_score >= 80:
            danger_level = "EXTREME"
            emoji = "🚨"
            urgency = "立即停止"
        elif risk_score >= 60:
            danger_level = "HIGH"
            emoji = "🔴"
            urgency = "强烈建议检查"
        elif risk_score >= 40:
            danger_level = "MEDIUM"
            emoji = "🟡"
            urgency = "建议谨慎操作"
        elif risk_score >= 20:
            danger_level = "LOW"
            emoji = "🟢"
            urgency = "可以安全操作"
        else:
            danger_level = "MINIMAL"
            emoji = "✅"
            urgency = "安全操作"

        # 生成警告消息
        warning_messages = []
        if risk_score >= 60:
            warning_messages.append("🔴 高风险警告：此操作可能影响大量数据")
        if risk_score >= 40:
            warning_messages.append("⚠️ 建议先创建备份再执行操作")
        if total_cells > 1000:
            warning_messages.append(f"📊 将影响约 {total_cells:,} 个单元格")

        # 生成操作建议
        recommendations = []
        if risk_score >= 60:
            recommendations.append("🛑️ 立即停止，使用 excel_create_backup 创建备份")
            recommendations.append("📋 重新评估操作范围，考虑分批处理")
            recommendations.append("🔍 使用 excel_preview_operation 预览具体影响")
        elif risk_score >= 40:
            recommendations.append("💾 建议创建备份后再继续")
            recommendations.append("📊 仔细检查操作范围和参数")
            recommendations.append("🔍 使用 excel_assess_data_impact 全面评估")
        elif risk_score >= 20:
            recommendations.append("📊 确认操作参数正确")
            recommendations.append("📋 考虑使用预览功能")

        return {
            'success': True,
            'danger_level': danger_level,
            'risk_score': risk_score,
            'risk_factors': risk_factors,
            'total_cells': total_cells,
            'file_size_mb': round(file_size_mb, 2),
            'warning_emoji': emoji,
            'urgency': urgency,
            'warning_messages': warning_messages,
            'recommendations': recommendations,
            'can_proceed': risk_score < 60,
            'requires_backup': risk_score >= 40,
            'requires_confirmation': risk_score >= 60,
            'estimated_execution_time': _estimate_execution_time(total_cells, file_size_mb),
            'operation_type': operation_type
        }

    except Exception as e:
        return {
            'success': False,
            'error': 'DANGER_CHECK_FAILED',
            'message': f"危险等级检查失败: {str(e)}"
        }


def _estimate_execution_time(total_cells: int, file_size_mb: float) -> str:
    """估算操作执行时间"""
    if total_cells > 50000:
        return "very_long"  # >30秒
    elif total_cells > 10000:
        return "long"  # 10-30秒
    elif total_cells > 1000:
        return "medium"  # 5-10秒
    elif total_cells > 100:
        return "short"  # 1-5秒
    else:
        return "minimal"  # <1秒


@mcp.tool()
def excel_check_file_status(
    file_path: str,
    check_locks: bool = True,
    check_permissions: bool = True,
    check_integrity: bool = True
) -> Dict[str, Any]:
    """
    全面检查Excel文件状态，验证文件是否被其他程序锁定

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        check_locks: 是否检查文件锁定状态
        check_permissions: 是否检查文件读写权限
        check_integrity: 是否检查文件完整性

    Returns:
        Dict: 包含文件状态详细检查结果

    Example:
        # 全面检查文件状态
        result = excel_check_file_status("data.xlsx")
        # 仅检查锁定状态
        result = excel_check_file_status("data.xlsx", check_locks=True, check_permissions=False, check_integrity=False)
    """
    import platform
    import time
    import tempfile

    try:
        # 基础检查：文件是否存在
        if not os.path.exists(file_path):
            return {
                'success': False,
                'error': 'FILE_NOT_FOUND',
                'message': f"文件不存在: {file_path}",
                'file_status': 'missing'
            }

        # 初始化检查结果
        status_info = {
            'file_path': file_path,
            'file_exists': True,
            'file_size': os.path.getsize(file_path),
            'modified_time': datetime.fromtimestamp(os.path.getmtime(file_path)),
            'checks_performed': []
        }

        # 权限检查
        if check_permissions:
            permission_status = _check_file_permissions(file_path)
            status_info['permissions'] = permission_status
            status_info['checks_performed'].append('permissions')

        # 文件锁定检查
        lock_status = {'locked': False, 'lock_type': None, 'lock_details': None}
        if check_locks:
            lock_status = _check_file_lock_status(file_path)
            status_info['lock_status'] = lock_status
            status_info['checks_performed'].append('locks')

        # 文件完整性检查
        integrity_status = {'valid': True, 'issues': []}
        if check_integrity:
            integrity_status = _check_file_integrity(file_path)
            status_info['integrity'] = integrity_status
            status_info['checks_performed'].append('integrity')

        # 系统信息
        status_info['system_info'] = {
            'platform': platform.system(),
            'python_version': platform.python_version(),
            'check_time': datetime.now().isoformat()
        }

        # 生成总体状态评估
        overall_status = _assess_overall_file_status(status_info)
        status_info['overall_status'] = overall_status

        # 生成建议
        recommendations = _generate_file_status_recommendations(status_info)
        status_info['recommendations'] = recommendations

        # 检查是否可以安全操作
        can_operate = (
            not lock_status['locked'] and
            permission_status.get('readable', False) and
            permission_status.get('writable', False) and
            integrity_status['valid']
        )

        status_info['can_safely_operate'] = can_operate

        return {
            'success': True,
            'file_status': status_info,
            'can_safely_operate': can_operate,
            'overall_assessment': overall_status['status'],
            'urgency_level': overall_status['urgency'],
            'recommendations': recommendations,
            'message': f"文件状态检查完成: {overall_status['description']}"
        }

    except Exception as e:
        return {
            'success': False,
            'error': 'FILE_STATUS_CHECK_FAILED',
            'message': f"文件状态检查失败: {str(e)}",
            'file_path': file_path
        }


def _check_file_permissions(file_path: str) -> Dict[str, Any]:
    """检查文件读写权限"""
    permissions = {
        'readable': False,
        'writable': False,
        'executable': False,
        'owner_info': None,
        'permission_bits': None
    }

    try:
        # 检查读权限
        if os.access(file_path, os.R_OK):
            permissions['readable'] = True

        # 检查写权限
        if os.access(file_path, os.W_OK):
            permissions['writable'] = True

        # 检查执行权限
        if os.access(file_path, os.X_OK):
            permissions['executable'] = True

        # 获取文件权限信息
        import stat
        file_stat = os.stat(file_path)
        permissions['permission_bits'] = oct(file_stat.st_mode)[-3:]

        # 尝试获取所有者信息
        try:
            import pwd
            permissions['owner_info'] = {
                'uid': file_stat.st_uid,
                'gid': file_stat.st_gid,
                'user': pwd.getpwuid(file_stat.st_uid).pw_name
            }
        except (ImportError, KeyError):
            permissions['owner_info'] = {
                'uid': file_stat.st_uid,
                'gid': file_stat.st_gid
            }

        # 权限问题诊断
        permission_issues = []
        if not permissions['readable']:
            permission_issues.append("无法读取文件")
        if not permissions['writable']:
            permission_issues.append("无法写入文件")

        permissions['issues'] = permission_issues
        permissions['sufficient_for_excel'] = permissions['readable'] and permissions['writable']

    except Exception as e:
        permissions['error'] = str(e)
        permissions['sufficient_for_excel'] = False

    return permissions


def _check_file_lock_status(file_path: str) -> Dict[str, Any]:
    """检查文件锁定状态"""
    lock_info = {
        'locked': False,
        'lock_type': None,
        'lock_details': None,
        'potential_lockers': []
    }

    try:
        import platform

        if platform.system() == "Windows":
            # Windows系统锁定检查
            lock_info = _check_windows_file_lock(file_path)
        elif platform.system() in ["Linux", "Darwin"]:
            # Unix系统锁定检查
            lock_info = _check_unix_file_lock(file_path)
        else:
            # 通用检查方法
            lock_info = _check_generic_file_lock(file_path)

        # 检查可能的锁定进程
        if lock_info['locked']:
            lock_info['potential_lockers'] = _find_potential_lockers(file_path)

    except Exception as e:
        lock_info['error'] = str(e)
        lock_info['locked'] = True  # 保守策略：检查失败时认为被锁定

    return lock_info


def _check_windows_file_lock(file_path: str) -> Dict[str, Any]:
    """Windows系统文件锁定检查"""
    lock_info = {'locked': False, 'lock_type': None, 'lock_details': None}

    try:
        import msvcrt

        # 方法1：尝试重命名文件
        original_name = file_path
        temp_name = file_path + f".lock_test_{int(time.time())}"

        try:
            os.rename(original_name, temp_name)
            os.rename(temp_name, original_name)
            lock_info['locked'] = False
        except PermissionError:
            lock_info['locked'] = True
            lock_info['lock_type'] = 'permission_denied'
            lock_info['lock_details'] = '文件被其他程序锁定，无法重命名'
            return lock_info
        except OSError as e:
            if e.winerror == 32:  # ERROR_SHARING_VIOLATION
                lock_info['locked'] = True
                lock_info['lock_type'] = 'sharing_violation'
                lock_info['lock_details'] = '文件共享冲突，可能被Excel打开'
            else:
                lock_info['locked'] = True
                lock_info['lock_type'] = 'os_error'
                lock_info['lock_details'] = f'系统错误: {e.winerror}'
            return lock_info

        # 方法2：尝试以独占模式打开文件
        try:
            fd = os.open(file_path, os.O_RDWR | os.O_EXCL)
            os.close(fd)
            lock_info['locked'] = False
        except OSError:
            lock_info['locked'] = True
            lock_info['lock_type'] = 'exclusive_access_denied'
            lock_info['lock_details'] = '无法获得独占访问权限'

    except ImportError:
        # msvcrt不可用，使用替代方法
        lock_info = _check_generic_file_lock(file_path)

    return lock_info


def _check_unix_file_lock(file_path: str) -> Dict[str, Any]:
    """Unix系统文件锁定检查"""
    lock_info = {'locked': False, 'lock_type': None, 'lock_details': None}

    try:
        # 方法1：检查文件描述符
        import subprocess
        result = subprocess.run(
            ['lsof', file_path],
            capture_output=True,
            text=True,
            timeout=5
        )

        if result.returncode == 0:
            lines = result.stdout.strip().split('\n')
            if len(lines) > 1:  # 有输出说明文件被打开
                lock_info['locked'] = True
                lock_info['lock_type'] = 'process_open'
                lock_info['lock_details'] = f'文件被 {len(lines)-1} 个进程打开'

                # 解析进程信息
                processes = []
                for line in lines[1:]:
                    parts = line.split()
                    if len(parts) >= 2:
                        processes.append({
                            'pid': parts[1],
                            'command': parts[0] if parts else 'unknown'
                        })
                lock_info['processes'] = processes

        # 方法2：创建临时文件测试
        temp_dir = os.path.dirname(file_path)
        temp_file = os.path.join(temp_dir, f".lock_test_{int(time.time())}")

        try:
            with open(temp_file, 'w') as f:
                f.write('test')
            os.remove(temp_file)
            # 如果成功，说明目录可写
        except PermissionError:
            lock_info['locked'] = True
            lock_info['lock_type'] = 'directory_permission'
            lock_info['lock_details'] = '目录权限不足，可能影响文件操作'

    except (subprocess.TimeoutExpired, subprocess.SubprocessError, FileNotFoundError):
        # lsof不可用，使用通用方法
        lock_info = _check_generic_file_lock(file_path)

    return lock_info


def _check_generic_file_lock(file_path: str) -> Dict[str, Any]:
    """通用文件锁定检查方法"""
    lock_info = {'locked': False, 'lock_type': None, 'lock_details': None}

    try:
        # 方法1：尝试打开文件进行写入
        test_data = b'lock_test'
        temp_path = file_path + f".test_{int(time.time())}"

        try:
            # 复制原文件
            import shutil
            shutil.copy2(file_path, temp_path)

            # 尝试写入测试数据
            with open(temp_path, 'r+b') as f:
                original_pos = f.tell()
                f.write(test_data)
                f.seek(original_pos)
                original_data = f.read(len(test_data))

            # 恢复原数据
            with open(temp_path, 'r+b') as f:
                f.seek(original_pos)
                f.write(original_data)

            os.remove(temp_path)
            lock_info['locked'] = False

        except (PermissionError, OSError) as e:
            lock_info['locked'] = True
            lock_info['lock_type'] = 'write_blocked'
            lock_info['lock_details'] = f'写入测试失败: {str(e)}'

            # 清理临时文件
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            except:
                pass

    except Exception as e:
        lock_info['locked'] = True
        lock_info['lock_type'] = 'check_failed'
        lock_info['lock_details'] = f'锁定检查失败: {str(e)}'

    return lock_info


def _find_potential_lockers(file_path: str) -> List[str]:
    """查找可能锁定文件的程序"""
    potential_lockers = []

    try:
        import platform

        if platform.system() == "Windows":
            # Windows：检查Excel进程
            try:
                import psutil
                for proc in psutil.process_iter(['pid', 'name']):
                    try:
                        if 'excel' in proc.info['name'].lower():
                            potential_lockers.append(f"Excel进程 (PID: {proc.info['pid']})")
                    except (psutil.NoSuchProcess, psutil.AccessDenied):
                        continue
            except ImportError:
                potential_lockers.append("Excel程序可能正在运行")
        else:
            # Unix系统：使用lsof查找
            try:
                import subprocess
                result = subprocess.run(
                    ['lsof', file_path],
                    capture_output=True,
                    text=True,
                    timeout=5
                )
                if result.returncode == 0:
                    lines = result.stdout.strip().split('\n')
                    for line in lines[1:]:
                        parts = line.split()
                        if len(parts) >= 2:
                            potential_lockers.append(f"进程 {parts[0]} (PID: {parts[1]})")
            except (subprocess.SubprocessError, FileNotFoundError):
                pass

        # 通用猜测
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext in ['.xlsx', '.xlsm', '.xls']:
            if not potential_lockers:
                potential_lockers.append("Microsoft Excel可能正在打开此文件")

    except Exception:
        potential_lockers.append("无法确定锁定程序")

    return potential_lockers


def _check_file_integrity(file_path: str) -> Dict[str, Any]:
    """检查Excel文件完整性"""
    integrity = {
        'valid': True,
        'issues': [],
        'file_type': None,
        'size_check': True,
        'structure_check': True
    }

    try:
        # 检查文件大小
        file_size = os.path.getsize(file_path)
        if file_size == 0:
            integrity['valid'] = False
            integrity['issues'].append("文件大小为0")
            integrity['size_check'] = False

        # 检查文件扩展名
        file_ext = os.path.splitext(file_path)[1].lower()
        integrity['file_type'] = file_ext

        if file_ext not in ['.xlsx', '.xlsm', '.xls']:
            integrity['valid'] = False
            integrity['issues'].append(f"不支持的文件格式: {file_ext}")

        # 尝试用openpyxl读取文件结构
        try:
            from openpyxl import load_workbook
            # 只加载文件结构，不加载全部数据
            wb = load_workbook(file_path, read_only=True, data_only=False)

            # 检查工作表数量
            if len(wb.worksheets) == 0:
                integrity['issues'].append("文件中没有工作表")
                integrity['structure_check'] = False

            wb.close()

        except Exception as e:
            integrity['valid'] = False
            integrity['structure_check'] = False
            integrity['issues'].append(f"文件结构损坏或不是有效的Excel文件: {str(e)}")

        # 检查文件最后修改时间
        mtime = os.path.getmtime(file_path)
        now = time.time()
        if now - mtime < 1:  # 文件刚刚被修改
            integrity['issues'].append("文件刚刚被修改，可能正在被其他程序操作")

    except Exception as e:
        integrity['valid'] = False
        integrity['issues'].append(f"完整性检查失败: {str(e)}")

    return integrity


def _assess_overall_file_status(status_info: Dict[str, Any]) -> Dict[str, Any]:
    """评估整体文件状态"""
    assessment = {
        'status': 'unknown',
        'urgency': 'low',
        'description': '文件状态未知',
        'blocking_issues': [],
        'warnings': []
    }

    # 检查锁定状态
    lock_status = status_info.get('lock_status', {})
    if lock_status.get('locked', False):
        assessment['status'] = 'blocked'
        assessment['urgency'] = 'high'
        assessment['description'] = '文件被锁定，无法操作'
        assessment['blocking_issues'].append('文件锁定')

    # 检查权限
    permissions = status_info.get('permissions', {})
    if not permissions.get('sufficient_for_excel', True):
        assessment['status'] = 'permission_denied'
        assessment['urgency'] = 'high'
        assessment['description'] = '权限不足，无法操作'
        assessment['blocking_issues'].append('权限不足')

    # 检查完整性
    integrity = status_info.get('integrity', {})
    if not integrity.get('valid', True):
        assessment['status'] = 'corrupted'
        assessment['urgency'] = 'high'
        assessment['description'] = '文件损坏或格式无效'
        assessment['blocking_issues'].append('文件损坏')

    # 检查警告
    if integrity.get('issues'):
        assessment['warnings'].extend(integrity['issues'])

    if lock_status.get('locked'):
        assessment['warnings'].append('文件可能被Excel或其他程序打开')

    # 如果没有问题，状态为良好
    if not assessment['blocking_issues']:
        assessment['status'] = 'good'
        assessment['urgency'] = 'low'
        assessment['description'] = '文件状态良好，可以安全操作'

    return assessment


def _generate_file_status_recommendations(status_info: Dict[str, Any]) -> List[str]:
    """生成文件状态建议"""
    recommendations = []

    # 基于锁定状态的建议
    lock_status = status_info.get('lock_status', {})
    if lock_status.get('locked', False):
        recommendations.append("🔒 关闭可能正在打开此文件的Excel程序")
        if lock_status.get('potential_lockers'):
            recommendations.append(f"🔍 检查以下进程: {', '.join(lock_status['potential_lockers'][:3])}")
        recommendations.append("⏳ 等待文件解锁后重试")
        recommendations.append("🔄 重启电脑可能解决文件锁定问题")

    # 基于权限的建议
    permissions = status_info.get('permissions', {})
    if not permissions.get('writable', False):
        recommendations.append("📝 检查文件写入权限")
        recommendations.append("👤 确认当前用户有修改文件的权限")
        recommendations.append("🔐 尝试以管理员身份运行程序")

    # 基于完整性的建议
    integrity = status_info.get('integrity', {})
    if not integrity.get('valid', True):
        recommendations.append("🔧 文件可能损坏，尝试使用Excel的修复功能")
        recommendations.append("💾 从备份恢复文件")
        recommendations.append("📋 重新创建或导出数据到新文件")

    # 通用建议
    if not recommendations:
        recommendations.append("✅ 文件状态良好，可以安全进行Excel操作")
        recommendations.append("💾 建议在重要操作前创建备份")
        recommendations.append("📊 使用预览功能确认操作范围")

    return recommendations


@mcp.tool()
def excel_confirm_operation(
    file_path: str,
    operation_type: str,
    range: str,
    risk_assessment: Dict[str, Any],
    confirmation_token: Optional[str] = None,
    force_proceed: bool = False
) -> Dict[str, Any]:
    """
    为危险操作创建确认步骤，确保用户明确理解风险

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        operation_type: 操作类型 ("update", "delete", "format", "insert", "delete_rows", "delete_columns")
        range: 范围表达式
        risk_assessment: 来自 excel_assess_data_impact 或 excel_check_danger_level 的风险评估
        confirmation_token: 确认令牌，用于验证用户确认
        force_proceed: 强制继续操作（仅用于特殊情况）

    Returns:
        Dict: 包含确认状态和操作许可

    Example:
        # 获取风险评估
        risk = excel_assess_data_impact("data.xlsx", "Sheet1!A1:C100", "update", new_data)
        # 请求用户确认
        result = excel_confirm_operation("data.xlsx", "update", "Sheet1!A1:C100", risk)
    """
    try:
        # 验证风险评估
        if not risk_assessment or not isinstance(risk_assessment, dict):
            return {
                'success': False,
                'error': 'INVALID_RISK_ASSESSMENT',
                'message': '风险评估无效或缺失'
            }

        # 提取风险信息
        risk_level = risk_assessment.get('overall_risk', risk_assessment.get('danger_level', 'UNKNOWN'))
        risk_score = risk_assessment.get('risk_score', 0)
        requires_backup = risk_assessment.get('requires_backup', False)
        requires_confirmation = risk_assessment.get('requires_confirmation', False)

        # 生成确认令牌
        import uuid
        session_token = str(uuid.uuid4())

        # 创建确认信息
        confirmation_info = {
            'session_token': session_token,
            'operation_details': {
                'file_path': file_path,
                'operation_type': operation_type,
                'range': range,
                'risk_level': risk_level,
                'risk_score': risk_score
            },
            'confirmation_required': True,
            'status': 'pending'
        }

        # 根据风险等级设置确认要求
        if risk_level in ['EXTREME', 'HIGH']:
            confirmation_info['confirmation_required'] = True
            confirmation_info['warning_level'] = 'critical'
            confirmation_info['user_action_required'] = 'explicit_confirmation'
        elif risk_level == 'MEDIUM':
            confirmation_info['confirmation_required'] = True
            confirmation_info['warning_level'] = 'warning'
            confirmation_info['user_action_required'] = 'acknowledgement'
        else:
            confirmation_info['confirmation_required'] = False
            confirmation_info['warning_level'] = 'info'
            confirmation_info['user_action_required'] = 'none'

        # 生成确认消息
        confirmation_messages = _generate_confirmation_messages(
            operation_type, range, risk_assessment
        )
        confirmation_info['messages'] = confirmation_messages

        # 检查是否有确认令牌
        if confirmation_token:
            token_validation = _validate_confirmation_token(
                confirmation_token, confirmation_info
            )
            if token_validation['valid']:
                confirmation_info['status'] = 'confirmed'
                confirmation_info['confirmed_at'] = datetime.now().isoformat()
            else:
                confirmation_info['status'] = 'invalid_token'
                confirmation_info['error'] = token_validation['error']

        # 检查强制继续标志
        if force_proceed:
            confirmation_info['status'] = 'forced_proceed'
            confirmation_info['force_reason'] = '用户强制继续操作'
            confirmation_info['warning'] = '⚠️ 用户选择强制继续，跳过安全确认'

        # 生成操作许可
        operation_permission = _generate_operation_permission(confirmation_info)

        # 记录确认会话
        operation_logger.start_session(file_path)
        operation_logger.log_operation("operation_confirmation", {
            "session_token": session_token,
            "operation_type": operation_type,
            "range": range,
            "risk_level": risk_level,
            "confirmation_status": confirmation_info['status'],
            "permission_granted": operation_permission['granted']
        })

        return {
            'success': True,
            'confirmation_info': confirmation_info,
            'operation_permission': operation_permission,
            'can_proceed': operation_permission['granted'],
            'session_token': session_token,
            'next_steps': _get_next_steps(confirmation_info),
            'message': _format_confirmation_message(confirmation_info)
        }

    except Exception as e:
        return {
            'success': False,
            'error': 'CONFIRMATION_FAILED',
            'message': f"操作确认失败: {str(e)}"
        }


def _generate_confirmation_messages(
    operation_type: str,
    range: str,
    risk_assessment: Dict[str, Any]
) -> Dict[str, Any]:
    """生成确认消息"""
    risk_level = risk_assessment.get('overall_risk', risk_assessment.get('danger_level', 'UNKNOWN'))
    risk_score = risk_assessment.get('risk_score', 0)
    risk_factors = risk_assessment.get('risk_factors', [])
    total_cells = risk_assessment.get('total_cells', 0)

    messages = {
        'title': '',
        'warning': '',
        'risk_summary': '',
        'consequences': [],
        'required_actions': [],
        'safety_recommendations': []
    }

    # 根据风险等级生成标题
    if risk_level == 'EXTREME':
        messages['title'] = '🚨 极高风险操作确认'
        messages['warning'] = '此操作可能导致数据永久丢失，请极度谨慎！'
    elif risk_level == 'HIGH':
        messages['title'] = '🔴 高风险操作确认'
        messages['warning'] = '此操作可能影响大量数据，强烈建议先备份！'
    elif risk_level == 'MEDIUM':
        messages['title'] = '🟡 中等风险操作确认'
        messages['warning'] = '此操作将影响部分数据，建议先备份。'
    else:
        messages['title'] = '🟢 低风险操作确认'
        messages['warning'] = '此操作风险较低，但仍需谨慎。'

    # 风险摘要
    messages['risk_summary'] = f"""
操作类型: {operation_type}
影响范围: {range}
风险等级: {risk_level} (评分: {risk_score})
影响单元格: {total_cells:,} 个
主要风险: {', '.join(risk_factors[:3]) if risk_factors else '无'}
    """.strip()

    # 后果描述
    if operation_type in ['delete', 'delete_rows', 'delete_columns']:
        messages['consequences'] = [
            '🗑️ 数据将被永久删除，无法撤销',
            '📊 相关公式和引用可能失效',
            '🔗 依赖此数据的其他工作表可能受影响'
        ]
    elif operation_type == 'update':
        messages['consequences'] = [
            '✏️ 现有数据将被新数据覆盖',
            '📊 公式可能被保留或覆盖（取决于设置）',
            '🎨 格式可能发生变化'
        ]
    elif operation_type == 'format':
        messages['consequences'] = [
            '🎨 单元格格式将被修改',
            '📊 数据内容不会改变',
            '👁️ 视觉显示将发生变化'
        ]

    # 必要行动
    if risk_level in ['EXTREME', 'HIGH']:
        messages['required_actions'] = [
            '✅ 必须创建备份 (使用 excel_create_backup)',
            '✅ 必须预览操作结果 (使用 excel_preview_operation)',
            '✅ 必须确认理解风险后果',
            '✅ 必须明确提供确认令牌'
        ]
    elif risk_level == 'MEDIUM':
        messages['required_actions'] = [
            '💾 建议创建备份',
            '📋 建议预览操作结果',
            '✅ 必须确认操作范围正确'
        ]
    else:
        messages['required_actions'] = [
            '📊 确认操作参数正确',
            '📋 建议使用预览功能'
        ]

    # 安全建议
    messages['safety_recommendations'] = [
        '💾 重要操作前总是创建备份',
        '📊 使用预览功能确认操作范围',
        '🔍 检查文件状态 (使用 excel_check_file_status)',
        '⏰ 在非高峰时间执行大型操作',
        '📝 记录操作日志以便追踪'
    ]

    return messages


def _validate_confirmation_token(
    token: str,
    confirmation_info: Dict[str, Any]
) -> Dict[str, Any]:
    """验证确认令牌"""
    # 在实际实现中，这里可以检查令牌的有效性、过期时间等
    # 为了演示，我们使用简单的验证逻辑

    if not token or not isinstance(token, str):
        return {
            'valid': False,
            'error': '确认令牌无效'
        }

    # 检查令牌长度（简单验证）
    if len(token) < 10:
        return {
            'valid': False,
            'error': '确认令牌格式错误'
        }

    # 检查会话状态
    if confirmation_info.get('status') == 'confirmed':
        return {
            'valid': False,
            'error': '此确认会话已完成'
        }

    return {
        'valid': True,
        'message': '确认令牌有效'
    }


def _generate_operation_permission(confirmation_info: Dict[str, Any]) -> Dict[str, Any]:
    """生成操作许可"""
    permission = {
        'granted': False,
        'reason': '',
        'conditions': [],
        'expires_at': None,
        'restrictions': []
    }

    status = confirmation_info.get('status', 'pending')
    risk_level = confirmation_info.get('operation_details', {}).get('risk_level', 'LOW')
    warning_level = confirmation_info.get('warning_level', 'info')

    # 根据状态决定是否授权
    if status == 'confirmed':
        permission['granted'] = True
        permission['reason'] = '用户已确认操作'

        # 设置条件
        if risk_level in ['EXTREME', 'HIGH']:
            permission['conditions'] = [
                '必须已创建备份',
                '必须已预览操作结果',
                '用户明确承担风险'
            ]
        elif risk_level == 'MEDIUM':
            permission['conditions'] = [
                '建议已创建备份',
                '建议已预览操作结果'
            ]

    elif status == 'forced_proceed':
        permission['granted'] = True
        permission['reason'] = '用户强制继续操作'
        permission['restrictions'] = [
            '用户承担全部责任',
            '跳过安全检查',
            '建议记录操作日志'
        ]

    elif status == 'pending':
        if warning_level == 'info':
            # 低风险操作可以直接进行
            permission['granted'] = True
            permission['reason'] = '低风险操作，无需确认'
        else:
            permission['granted'] = False
            permission['reason'] = f'等待用户确认 (风险等级: {risk_level})'

    elif status == 'invalid_token':
        permission['granted'] = False
        permission['reason'] = '确认令牌无效'

    # 设置过期时间（高风险操作限时较短）
    import time
    if risk_level in ['EXTREME', 'HIGH']:
        permission['expires_at'] = time.time() + 300  # 5分钟
    elif risk_level == 'MEDIUM':
        permission['expires_at'] = time.time() + 1800  # 30分钟
    else:
        permission['expires_at'] = time.time() + 3600  # 1小时

    return permission


def _get_next_steps(confirmation_info: Dict[str, Any]) -> List[str]:
    """获取下一步操作指引"""
    status = confirmation_info.get('status', 'pending')
    risk_level = confirmation_info.get('operation_details', {}).get('risk_level', 'LOW')
    warning_level = confirmation_info.get('warning_level', 'info')

    next_steps = []

    if status == 'confirmed':
        next_steps = [
            '✅ 确认完成，可以执行操作',
            '📊 执行操作后验证结果',
            '📝 检查操作日志确认执行'
        ]
    elif status == 'forced_proceed':
        next_steps = [
            '⚠️ 强制继续，用户承担风险',
            '📊 执行操作时密切监控',
            '💾 建议立即创建备份'
        ]
    elif status == 'pending':
        if warning_level in ['critical', 'warning']:
            next_steps = [
                '🔒 需要用户确认才能继续',
                '💾 创建备份文件',
                '📋 预览操作结果',
                '✅ 提供有效确认令牌'
            ]
        else:
            next_steps = [
                '✅ 可以直接执行操作',
                '📊 建议预览操作结果',
                '💾 建议创建备份'
            ]
    elif status == 'invalid_token':
        next_steps = [
            '❌ 确认令牌无效',
            '🔄 重新获取确认令牌',
            '📋 检查令牌格式和有效期'
        ]

    return next_steps


def _format_confirmation_message(confirmation_info: Dict[str, Any]) -> str:
    """格式化确认消息"""
    status = confirmation_info.get('status', 'pending')
    risk_level = confirmation_info.get('operation_details', {}).get('risk_level', 'LOW')

    if status == 'confirmed':
        return f"✅ 操作已确认，可以安全执行 ({risk_level}风险)"
    elif status == 'forced_proceed':
        return f"⚠️ 用户强制继续操作 ({risk_level}风险)"
    elif status == 'pending':
        return f"🔒 等待用户确认 ({risk_level}风险)"
    elif status == 'invalid_token':
        return "❌ 确认令牌无效，请重新确认"
    else:
        return f"📋 确认状态: {status}"


@mcp.tool()
def excel_generate_operation_summary(
    file_path: str,
    operation_type: str,
    range: str,
    current_data: Optional[List[List[Any]]] = None,
    new_data: Optional[List[List[Any]]] = None,
    include_analysis: bool = True
) -> Dict[str, Any]:
    """
    生成操作摘要，显示操作前后的数据对比和变化分析

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        operation_type: 操作类型 ("update", "delete", "format", "insert")
        range: 范围表达式
        current_data: 当前数据（可选，如果不提供将自动读取）
        new_data: 新数据（对于更新操作）
        include_analysis: 是否包含详细分析

    Returns:
        Dict: 包含操作摘要和对比分析

    Example:
        # 生成更新操作摘要
        result = excel_generate_operation_summary(
            "data.xlsx", "update", "Sheet1!A1:C10", current_data, new_data
        )
        # 生成删除操作摘要
        result = excel_generate_operation_summary("data.xlsx", "delete", "Sheet1!5:10")
    """
    try:
        # 如果没有提供当前数据，读取当前数据
        if current_data is None:
            current_result = ExcelOperations.get_range(file_path, range)
            if not current_result.get('success'):
                return {
                    'success': False,
                    'error': 'FAILED_TO_READ_CURRENT_DATA',
                    'message': f"无法读取当前数据: {current_result.get('message', '未知错误')}"
                }
            current_data = current_result.get('data', [])

        # 分析范围信息
        from .utils.validators import ExcelValidator, DataValidationError
        try:
            range_validation = ExcelValidator.validate_range_expression(range)
            range_info = range_validation['range_info']
        except DataValidationError as e:
            return {
                'success': False,
                'error': 'INVALID_RANGE',
                'message': f"无效的范围表达式: {str(e)}"
            }

        # 生成基础摘要
        summary = {
            'operation_info': {
                'file_path': file_path,
                'operation_type': operation_type,
                'range': range,
                'range_info': range_info,
                'timestamp': datetime.now().isoformat()
            },
            'data_analysis': {},
            'changes': {},
            'statistics': {},
            'recommendations': []
        }

        # 分析当前数据
        current_analysis = _analyze_data_content(current_data, "当前数据")
        summary['data_analysis']['current'] = current_analysis

        # 分析新数据（如果提供）
        if new_data is not None:
            new_analysis = _analyze_data_content(new_data, "新数据")
            summary['data_analysis']['new'] = new_analysis

            # 生成变化对比
            changes = _compare_data_changes(current_data, new_data, operation_type)
            summary['changes'] = changes

        # 生成统计信息
        statistics = _generate_operation_statistics(current_data, new_data, operation_type, range_info)
        summary['statistics'] = statistics

        # 生成操作影响分析
        if include_analysis:
            impact_analysis = _generate_impact_analysis(current_data, new_data, operation_type, range_info)
            summary['impact_analysis'] = impact_analysis

        # 生成建议
        recommendations = _generate_operation_recommendations(summary, operation_type)
        summary['recommendations'] = recommendations

        # 生成摘要描述
        summary_description = _format_summary_description(summary, operation_type)
        summary['summary_description'] = summary_description

        return {
            'success': True,
            'operation_summary': summary,
            'has_changes': bool(summary.get('changes', {}).get('data_changes')),
            'risk_level': _assess_operation_risk(summary, operation_type),
            'message': f"操作摘要生成完成: {summary_description}"
        }

    except Exception as e:
        return {
            'success': False,
            'error': 'SUMMARY_GENERATION_FAILED',
            'message': f"操作摘要生成失败: {str(e)}"
        }


def _analyze_data_content(data: List[List[Any]], data_label: str) -> Dict[str, Any]:
    """分析数据内容"""
    if not data:
        return {
            'label': data_label,
            'row_count': 0,
            'column_count': 0,
            'total_cells': 0,
            'non_empty_cells': 0,
            'data_types': {},
            'has_formulas': False,
            'numeric_summary': {},
            'text_summary': {},
            'sample_data': []
        }

    total_rows = len(data)
    total_cols = max(len(row) for row in data) if data else 0
    total_cells = total_rows * total_cols

    non_empty_cells = 0
    data_types = {'text': 0, 'numeric': 0, 'formula': 0, 'boolean': 0, 'empty': 0, 'other': 0}
    numeric_values = []
    text_values = []
    formulas = []

    for row in data:
        for cell in row:
            if cell is None or (isinstance(cell, str) and cell.strip() == ''):
                data_types['empty'] += 1
            elif isinstance(cell, str):
                if cell.startswith('='):
                    data_types['formula'] += 1
                    formulas.append(cell)
                    non_empty_cells += 1
                else:
                    data_types['text'] += 1
                    text_values.append(cell)
                    non_empty_cells += 1
            elif isinstance(cell, bool):
                data_types['boolean'] += 1
                non_empty_cells += 1
            elif isinstance(cell, (int, float)):
                data_types['numeric'] += 1
                numeric_values.append(cell)
                non_empty_cells += 1
            else:
                data_types['other'] += 1
                non_empty_cells += 1

    # 数值摘要
    numeric_summary = {}
    if numeric_values:
        numeric_summary = {
            'count': len(numeric_values),
            'min': min(numeric_values),
            'max': max(numeric_values),
            'average': sum(numeric_values) / len(numeric_values),
            'sum': sum(numeric_values)
        }

    # 文本摘要
    text_summary = {}
    if text_values:
        text_lengths = [len(str(text)) for text in text_values]
        text_summary = {
            'count': len(text_values),
            'avg_length': sum(text_lengths) / len(text_lengths),
            'max_length': max(text_lengths),
            'min_length': min(text_lengths)
        }

    # 获取样本数据（前5行）
    sample_data = []
    for i, row in enumerate(data[:5]):
        sample_row = []
        for j, cell in enumerate(row[:10]):  # 前10列
            sample_row.append(str(cell) if cell is not None else '')
        sample_data.append(sample_row)

    return {
        'label': data_label,
        'row_count': total_rows,
        'column_count': total_cols,
        'total_cells': total_cells,
        'non_empty_cells': non_empty_cells,
        'empty_cells': total_cells - non_empty_cells,
        'data_types': data_types,
        'has_formulas': len(formulas) > 0,
        'formula_count': len(formulas),
        'numeric_summary': numeric_summary,
        'text_summary': text_summary,
        'sample_data': sample_data
    }


def _compare_data_changes(
    current_data: List[List[Any]],
    new_data: List[List[Any]],
    operation_type: str
) -> Dict[str, Any]:
    """比较数据变化"""
    changes = {
        'data_changes': [],
        'structural_changes': {},
        'content_changes': {
            'added_cells': 0,
            'modified_cells': 0,
            'deleted_cells': 0,
            'added_rows': 0,
            'deleted_rows': 0
        },
        'type_changes': {},
        'value_changes': []
    }

    if operation_type == 'update':
        # 更新操作的详细比较
        max_rows = max(len(current_data), len(new_data))
        max_cols = 0
        if current_data:
            max_cols = max(max_cols, max(len(row) for row in current_data))
        if new_data:
            max_cols = max(max_cols, max(len(row) for row in new_data))

        for i in range(max_rows):
            for j in range(max_cols):
                current_val = current_data[i][j] if i < len(current_data) and j < len(current_data[i]) else None
                new_val = new_data[i][j] if i < len(new_data) and j < len(new_data[i]) else None

                cell_addr = f"R{i+1}C{j+1}"  # 1-based索引

                if current_val != new_val:
                    change_detail = {
                        'cell_address': cell_addr,
                        'row': i + 1,
                        'column': j + 1,
                        'old_value': current_val,
                        'new_value': new_val,
                        'change_type': 'modified'
                    }

                    if current_val is None and new_val is not None:
                        change_detail['change_type'] = 'added'
                        changes['content_changes']['added_cells'] += 1
                    elif new_val is None and current_val is not None:
                        change_detail['change_type'] = 'deleted'
                        changes['content_changes']['deleted_cells'] += 1
                    else:
                        change_detail['change_type'] = 'modified'
                        changes['content_changes']['modified_cells'] += 1

                    # 分析数据类型变化
                    old_type = _get_data_type(current_val)
                    new_type = _get_data_type(new_val)
                    if old_type != new_type:
                        change_detail['type_change'] = {'old': old_type, 'new': new_type}

                    changes['data_changes'].append(change_detail)

    elif operation_type == 'delete':
        # 删除操作分析
        changes['content_changes']['deleted_cells'] = sum(
            1 for row in current_data for cell in row
            if cell is not None and str(cell).strip() != ''
        )
        changes['content_changes']['deleted_rows'] = len(current_data)

    elif operation_type == 'insert':
        # 插入操作分析
        changes['content_changes']['added_cells'] = sum(
            1 for row in new_data for cell in row
            if cell is not None and str(cell).strip() != ''
        )
        changes['content_changes']['added_rows'] = len(new_data)

    # 结构变化
    current_rows = len(current_data)
    new_rows = len(new_data) if new_data else 0
    current_cols = max(len(row) for row in current_data) if current_data else 0
    new_cols = max(len(row) for row in new_data) if new_data else 0

    changes['structural_changes'] = {
        'row_change': new_rows - current_rows,
        'column_change': new_cols - current_cols,
        'old_dimensions': f"{current_rows}x{current_cols}",
        'new_dimensions': f"{new_rows}x{new_cols}"
    }

    return changes


def _get_data_type(value: Any) -> str:
    """获取数据类型"""
    if value is None or (isinstance(value, str) and value.strip() == ''):
        return 'empty'
    elif isinstance(value, str):
        if value.startswith('='):
            return 'formula'
        return 'text'
    elif isinstance(value, bool):
        return 'boolean'
    elif isinstance(value, (int, float)):
        return 'numeric'
    else:
        return 'other'


def _generate_operation_statistics(
    current_data: List[List[Any]],
    new_data: Optional[List[List[Any]]],
    operation_type: str,
    range_info: Dict[str, Any]
) -> Dict[str, Any]:
    """生成操作统计信息"""
    stats = {
        'operation_type': operation_type,
        'range_info': range_info,
        'data_volume': {},
        'impact_metrics': {},
        'performance_estimate': {}
    }

    # 数据量统计
    stats['data_volume'] = {
        'current_data_cells': len(current_data) * (max(len(row) for row in current_data) if current_data else 0),
        'new_data_cells': len(new_data) * (max(len(row) for row in new_data) if new_data else 0) if new_data else 0,
        'affected_cells': 0,
        'data_density_current': 0
    }

    # 计算数据密度
    total_current_cells = stats['data_volume']['current_data_cells']
    if total_current_cells > 0:
        non_empty_current = sum(
            1 for row in current_data for cell in row
            if cell is not None and str(cell).strip() != ''
        )
        stats['data_volume']['data_density_current'] = (non_empty_current / total_current_cells) * 100

    # 影响指标
    if operation_type == 'update' and new_data:
        stats['impact_metrics'] = {
            'cells_to_update': len(new_data) * len(new_data[0]) if new_data else 0,
            'data_overlap': 0,  # 计算重叠区域
            'new_data_percentage': 0
        }
        stats['data_volume']['affected_cells'] = stats['impact_metrics']['cells_to_update']

    elif operation_type == 'delete':
        stats['impact_metrics'] = {
            'cells_to_delete': stats['data_volume']['current_data_cells'],
            'data_loss_risk': 'high' if stats['data_volume']['data_density_current'] > 50 else 'medium'
        }
        stats['data_volume']['affected_cells'] = stats['impact_metrics']['cells_to_delete']

    elif operation_type == 'insert':
        stats['impact_metrics'] = {
            'cells_to_add': stats['data_volume']['new_data_cells'],
            'growth_percentage': 0
        }
        stats['data_volume']['affected_cells'] = stats['impact_metrics']['cells_to_add']

    # 性能估算
    total_affected = stats['data_volume']['affected_cells']
    if total_affected > 50000:
        stats['performance_estimate'] = {
            'execution_time': 'very_long',
            'memory_usage': 'high',
            'recommended_approach': 'batch_processing'
        }
    elif total_affected > 10000:
        stats['performance_estimate'] = {
            'execution_time': 'long',
            'memory_usage': 'medium',
            'recommended_approach': 'monitor_progress'
        }
    elif total_affected > 1000:
        stats['performance_estimate'] = {
            'execution_time': 'medium',
            'memory_usage': 'low',
            'recommended_approach': 'standard'
        }
    else:
        stats['performance_estimate'] = {
            'execution_time': 'fast',
            'memory_usage': 'minimal',
            'recommended_approach': 'direct'
        }

    return stats


def _generate_impact_analysis(
    current_data: List[List[Any]],
    new_data: Optional[List[List[Any]]],
    operation_type: str,
    range_info: Dict[str, Any]
) -> Dict[str, Any]:
    """生成影响分析"""
    analysis = {
        'data_integrity_risk': 'low',
        'formula_impact': 'none',
        'dependency_risk': 'low',
        'rollback_complexity': 'low',
        'business_impact': {}
    }

    # 数据完整性风险评估
    current_density = sum(
        1 for row in current_data for cell in row
        if cell is not None and str(cell).strip() != ''
    ) / (len(current_data) * max(len(row) for row in current_data) if current_data else 1)

    if operation_type == 'delete' and current_density > 0.7:
        analysis['data_integrity_risk'] = 'high'
    elif operation_type == 'update' and current_density > 0.5:
        analysis['data_integrity_risk'] = 'medium'

    # 公式影响分析
    formula_count = sum(
        1 for row in current_data for cell in row
        if isinstance(cell, str) and cell.startswith('=')
    )

    if formula_count > 0:
        if operation_type in ['delete', 'update']:
            analysis['formula_impact'] = 'high'
            analysis['formula_count'] = formula_count
        else:
            analysis['formula_impact'] = 'medium'

    # 依赖关系风险评估
    if range_info.get('type') == 'full_sheet' or range_info.get('range_size', {}).get('total_cells', 0) > 1000:
        analysis['dependency_risk'] = 'medium'
        analysis['dependency_reason'] = '大型操作可能影响其他工作表'

    # 回滚复杂性
    if operation_type == 'delete':
        analysis['rollback_complexity'] = 'high'
        analysis['rollback_requirement'] = '需要完整备份'
    elif operation_type == 'update':
        analysis['rollback_complexity'] = 'medium'
        analysis['rollback_requirement'] = '需要数据备份'

    # 业务影响
    analysis['business_impact'] = {
        'data_availability': 'temporarily_affected' if operation_type in ['update', 'delete'] else 'expanded',
        'user_experience': 'minimal' if analysis['data_integrity_risk'] == 'low' else 'significant',
        'recovery_time': 'minimal' if analysis['rollback_complexity'] == 'low' else 'extended'
    }

    return analysis


def _generate_operation_recommendations(
    summary: Dict[str, Any],
    operation_type: str
) -> List[str]:
    """生成操作建议"""
    recommendations = []

    risk_level = summary.get('risk_level', 'low')
    impact_analysis = summary.get('impact_analysis', {})
    statistics = summary.get('statistics', {})

    # 基于风险等级的建议
    if risk_level in ['high', 'critical']:
        recommendations.append("🔴 高风险操作，强烈建议创建完整备份")
        recommendations.append("⚠️ 考虑在非工作时间执行此操作")
        recommendations.append("📊 准备回滚计划")

    # 基于操作类型的建议
    if operation_type == 'delete':
        recommendations.append("🗑️ 删除操作不可逆，请仔细确认")
        if impact_analysis.get('data_integrity_risk') == 'high':
            recommendations.append("💾 数据密度高，建议先导出重要数据")
    elif operation_type == 'update':
        recommendations.append("✏️ 建议使用预览功能确认更新范围")
        if summary.get('data_analysis', {}).get('current', {}).get('has_formulas'):
            recommendations.append("📊 检测到公式，更新后请验证公式正确性")
    elif operation_type == 'insert':
        recommendations.append("➕ 确认插入位置不会破坏现有数据结构")

    # 基于性能的建议
    perf_estimate = statistics.get('performance_estimate', {})
    if perf_estimate.get('execution_time') in ['long', 'very_long']:
        recommendations.append("⏱️ 大型操作预计耗时较长，请耐心等待")
        recommendations.append("💾 确保有足够的内存和磁盘空间")

    # 通用建议
    recommendations.append("📋 执行操作前检查文件状态")
    recommendations.append("📝 记录操作以便追踪")
    recommendations.append("✅ 操作完成后验证结果")

    return recommendations


def _assess_operation_risk(summary: Dict[str, Any], operation_type: str) -> str:
    """评估操作风险等级"""
    risk_score = 0

    # 基于操作类型的基础风险
    if operation_type == 'delete':
        risk_score += 30
    elif operation_type == 'update':
        risk_score += 20
    elif operation_type == 'insert':
        risk_score += 10

    # 基于数据量的风险
    current_data = summary.get('data_analysis', {}).get('current', {})
    total_cells = current_data.get('total_cells', 0)
    non_empty_cells = current_data.get('non_empty_cells', 0)

    if total_cells > 10000:
        risk_score += 25
    elif total_cells > 1000:
        risk_score += 15
    elif total_cells > 100:
        risk_score += 5

    # 基于数据密度的风险
    if total_cells > 0:
        density = (non_empty_cells / total_cells) * 100
        if density > 80:
            risk_score += 20
        elif density > 50:
            risk_score += 10

    # 基于公式的风险
    if current_data.get('has_formulas', False):
        risk_score += 15

    # 基于影响分析的风险
    impact_analysis = summary.get('impact_analysis', {})
    if impact_analysis.get('data_integrity_risk') == 'high':
        risk_score += 20
    elif impact_analysis.get('data_integrity_risk') == 'medium':
        risk_score += 10

    # 确定风险等级
    if risk_score >= 70:
        return 'critical'
    elif risk_score >= 50:
        return 'high'
    elif risk_score >= 30:
        return 'medium'
    else:
        return 'low'


def _format_summary_description(summary: Dict[str, Any], operation_type: str) -> str:
    """格式化摘要描述"""
    current_data = summary.get('data_analysis', {}).get('current', {})
    changes = summary.get('changes', {})
    statistics = summary.get('statistics', {})

    # 基础信息
    range_info = summary.get('operation_info', {}).get('range_info', {})
    range_str = summary.get('operation_info', {}).get('range', 'Unknown')

    description_parts = [
        f"操作类型: {operation_type}",
        f"影响范围: {range_str}",
        f"当前数据: {current_data.get('row_count', 0)}行 x {current_data.get('column_count', 0)}列"
    ]

    # 添加变化信息
    if operation_type == 'update' and changes.get('content_changes'):
        content_changes = changes['content_changes']
        description_parts.extend([
            f"修改单元格: {content_changes.get('modified_cells', 0)}",
            f"新增单元格: {content_changes.get('added_cells', 0)}"
        ])
    elif operation_type == 'delete':
        description_parts.append(f"将删除 {current_data.get('non_empty_cells', 0)} 个非空单元格")

    # 添加风险等级
    risk_level = summary.get('risk_level', 'low')
    risk_emoji = {'critical': '🚨', 'high': '🔴', 'medium': '🟡', 'low': '🟢'}
    description_parts.append(f"风险等级: {risk_emoji.get(risk_level, '❓')} {risk_level}")

    # 添加执行时间预估
    perf_estimate = statistics.get('performance_estimate', {})
    exec_time = perf_estimate.get('execution_time', 'unknown')
    time_emoji = {'very_long': '⏳', 'long': '🕐', 'medium': '⏱️', 'fast': '⚡'}
    description_parts.append(f"预计执行时间: {time_emoji.get(exec_time, '❓')} {exec_time}")

    return " | ".join(description_parts)


@mcp.tool()
def excel_visualize_operation_range(
    file_path: str,
    range: str,
    operation_type: str = "update",
    current_data: Optional[List[List[Any]]] = None,
    visualization_mode: str = "text",
    include_context: bool = True
) -> Dict[str, Any]:
    """
    生成操作范围可视化，清晰标识将要影响的区域

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        range: 范围表达式
        operation_type: 操作类型 ("update", "delete", "format", "insert")
        current_data: 当前数据（可选，如果不提供将自动读取）
        visualization_mode: 可视化模式 ("text", "detailed", "summary", "matrix")
        include_context: 是否包含上下文信息

    Returns:
        Dict: 包含范围可视化信息

    Example:
        # 文本模式可视化
        result = excel_visualize_operation_range("data.xlsx", "Sheet1!A1:C10", "update")
        # 详细模式可视化
        result = excel_visualize_operation_range("data.xlsx", "Sheet1!A1:C10", "delete", visualization_mode="detailed")
    """
    try:
        # 如果没有提供当前数据，读取当前数据
        if current_data is None:
            current_result = ExcelOperations.get_range(file_path, range)
            if not current_result.get('success'):
                return {
                    'success': False,
                    'error': 'FAILED_TO_READ_CURRENT_DATA',
                    'message': f"无法读取当前数据: {current_result.get('message', '未知错误')}"
                }
            current_data = current_result.get('data', [])

        # 解析范围信息
        from .utils.validators import ExcelValidator, DataValidationError
        try:
            range_validation = ExcelValidator.validate_range_expression(range)
            range_info = range_validation['range_info']
        except DataValidationError as e:
            return {
                'success': False,
                'error': 'INVALID_RANGE',
                'message': f"无效的范围表达式: {str(e)}"
            }

        # 生成可视化
        visualization = {
            'operation_info': {
                'file_path': file_path,
                'operation_type': operation_type,
                'range': range,
                'range_info': range_info,
                'visualization_mode': visualization_mode
            },
            'range_details': {},
            'impact_map': {},
            'context_info': {}
        }

        # 生成范围详细信息
        range_details = _generate_range_details(range_info, current_data)
        visualization['range_details'] = range_details

        # 生成影响映射
        impact_map = _generate_impact_map(current_data, operation_type, range_info)
        visualization['impact_map'] = impact_map

        # 生成上下文信息
        if include_context:
            context_info = _generate_context_info(file_path, range_info, range_details)
            visualization['context_info'] = context_info

        # 根据可视化模式生成具体内容
        if visualization_mode == "text":
            viz_content = _generate_text_visualization(visualization, operation_type)
        elif visualization_mode == "detailed":
            viz_content = _generate_detailed_visualization(visualization, operation_type)
        elif visualization_mode == "summary":
            viz_content = _generate_summary_visualization(visualization, operation_type)
        elif visualization_mode == "matrix":
            viz_content = _generate_matrix_visualization(visualization, operation_type)
        else:
            viz_content = _generate_text_visualization(visualization, operation_type)

        visualization['visualization_content'] = viz_content

        # 生成可视化摘要
        viz_summary = _generate_visualization_summary(visualization, operation_type)
        visualization['summary'] = viz_summary

        return {
            'success': True,
            'visualization': visualization,
            'visualization_type': visualization_mode,
            'impact_cells': visualization['impact_map'].get('total_affected_cells', 0),
            'message': f"操作范围可视化完成: {viz_summary['description']}"
        }

    except Exception as e:
        return {
            'success': False,
            'error': 'VISUALIZATION_FAILED',
            'message': f"范围可视化失败: {str(e)}"
        }


def _generate_range_details(range_info: Dict[str, Any], current_data: List[List[Any]]) -> Dict[str, Any]:
    """生成范围详细信息"""
    range_type = range_info.get('type', 'unknown')
    range_size = range_info.get('range_size', {})

    details = {
        'range_type': range_type,
        'dimensions': {
            'rows': 0,
            'columns': 0,
            'total_cells': 0
        },
        'bounds': {},
        'data_content': {}
    }

    if range_type == 'cell_range':
        # 单元格范围
        bounds = range_info.get('bounds', {})
        start_row = bounds.get('start_row', 1)
        end_row = bounds.get('end_row', 1)
        start_col = bounds.get('start_col', 1)
        end_col = bounds.get('end_col', 1)

        details['dimensions'] = {
            'rows': end_row - start_row + 1,
            'columns': end_col - start_col + 1,
            'total_cells': (end_row - start_row + 1) * (end_col - start_col + 1)
        }

        details['bounds'] = {
            'start_row': start_row,
            'end_row': end_row,
            'start_col': start_col,
            'end_col': end_col,
            'start_cell': f"{_col_num_to_letter(start_col)}{start_row}",
            'end_cell': f"{_col_num_to_letter(end_col)}{end_row}"
        }

    elif range_type == 'row_range':
        # 行范围
        bounds = range_info.get('bounds', {})
        start_row = bounds.get('start_row', 1)
        end_row = bounds.get('end_row', 1)

        details['dimensions'] = {
            'rows': end_row - start_row + 1,
            'columns': len(current_data[0]) if current_data else 0,
            'total_cells': (end_row - start_row + 1) * len(current_data[0]) if current_data else 0
        }

        details['bounds'] = {
            'start_row': start_row,
            'end_row': end_row,
            'row_count': end_row - start_row + 1
        }

    elif range_type == 'column_range':
        # 列范围
        bounds = range_info.get('bounds', {})
        start_col = bounds.get('start_col', 1)
        end_col = bounds.get('end_col', 1)

        details['dimensions'] = {
            'rows': len(current_data) if current_data else 0,
            'columns': end_col - start_col + 1,
            'total_cells': len(current_data) * (end_col - start_col + 1) if current_data else 0
        }

        details['bounds'] = {
            'start_col': start_col,
            'end_col': end_col,
            'col_count': end_col - start_col + 1,
            'start_letter': _col_num_to_letter(start_col),
            'end_letter': _col_num_to_letter(end_col)
        }

    # 分析数据内容
    if current_data:
        non_empty_cells = sum(
            1 for row in current_data for cell in row
            if cell is not None and str(cell).strip() != ''
        )

        details['data_content'] = {
            'total_rows': len(current_data),
            'total_cols': max(len(row) for row in current_data) if current_data else 0,
            'non_empty_cells': non_empty_cells,
            'empty_cells': details['dimensions']['total_cells'] - non_empty_cells,
            'data_density': (non_empty_cells / details['dimensions']['total_cells'] * 100) if details['dimensions']['total_cells'] > 0 else 0,
            'has_formulas': any(
                isinstance(cell, str) and cell.startswith('=')
                for row in current_data for cell in row
            )
        }

    return details


def _generate_impact_map(current_data: List[List[Any]], operation_type: str, range_info: Dict[str, Any]) -> Dict[str, Any]:
    """生成影响映射"""
    impact_map = {
        'total_affected_cells': 0,
        'data_cells_affected': 0,
        'empty_cells_affected': 0,
        'formula_cells_affected': 0,
        'impact_zones': [],
        'risk_zones': []
    }

    if not current_data:
        return impact_map

    total_cells = 0
    data_cells = 0
    empty_cells = 0
    formula_cells = 0

    # 分析每个单元格的影响
    for i, row in enumerate(current_data):
        for j, cell in enumerate(row):
            total_cells += 1

            cell_info = {
                'row': i + 1,
                'column': j + 1,
                'cell_address': f"{_col_num_to_letter(j + 1)}{i + 1}",
                'value': cell,
                'impact_type': _determine_impact_type(cell, operation_type),
                'risk_level': 'low'
            }

            if cell is None or (isinstance(cell, str) and cell.strip() == ''):
                empty_cells += 1
                cell_info['data_type'] = 'empty'
            elif isinstance(cell, str) and cell.startswith('='):
                formula_cells += 1
                cell_info['data_type'] = 'formula'
                cell_info['risk_level'] = 'high' if operation_type in ['delete', 'update'] else 'medium'
            elif isinstance(cell, (int, float)):
                data_cells += 1
                cell_info['data_type'] = 'numeric'
            elif isinstance(cell, bool):
                data_cells += 1
                cell_info['data_type'] = 'boolean'
            else:
                data_cells += 1
                cell_info['data_type'] = 'text'

            impact_map['impact_zones'].append(cell_info)

    impact_map['total_affected_cells'] = total_cells
    impact_map['data_cells_affected'] = data_cells
    impact_map['empty_cells_affected'] = empty_cells
    impact_map['formula_cells_affected'] = formula_cells

    # 生成风险区域
    risk_zones = []
    for zone in impact_map['impact_zones']:
        if zone['risk_level'] in ['high', 'medium']:
            risk_zones.append(zone)

    impact_map['risk_zones'] = risk_zones

    return impact_map


def _determine_impact_type(cell_value: Any, operation_type: str) -> str:
    """确定影响类型"""
    if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ''):
        if operation_type == 'insert':
            return 'data_creation'
        else:
            return 'empty_operation'
    elif operation_type == 'delete':
        return 'data_deletion'
    elif operation_type == 'update':
        return 'data_modification'
    elif operation_type == 'format':
        return 'format_change'
    else:
        return 'general_operation'


def _generate_context_info(file_path: str, range_info: Dict[str, Any], range_details: Dict[str, Any]) -> Dict[str, Any]:
    """生成上下文信息"""
    context_info = {
        'file_info': {},
        'sheet_info': {},
        'surrounding_data': {},
        'dependency_info': {}
    }

    # 文件信息
    try:
        import os
        if os.path.exists(file_path):
            stat = os.stat(file_path)
            context_info['file_info'] = {
                'file_size': stat.st_size,
                'modified_time': datetime.fromtimestamp(stat.st_mtime),
                'file_name': os.path.basename(file_path)
            }
    except Exception:
        pass

    # 工作表信息
    try:
        sheets_result = ExcelOperations.list_sheets(file_path)
        if sheets_result.get('success'):
            context_info['sheet_info'] = {
                'total_sheets': sheets_result.get('total_sheets', 0),
                'sheet_names': sheets_result.get('sheets', [])
            }
    except Exception:
        pass

    # 周围数据信息
    bounds = range_details.get('bounds', {})
    if bounds:
        # 扩展范围以获取上下文
        context_ranges = _generate_context_ranges(bounds, range_info.get('type'))
        context_info['surrounding_data'] = {
            'above_range': context_ranges.get('above'),
            'below_range': context_ranges.get('below'),
            'left_of_range': context_ranges.get('left'),
            'right_of_range': context_ranges.get('right')
        }

    return context_info


def _generate_context_ranges(bounds: Dict[str, Any], range_type: str) -> Dict[str, Any]:
    """生成上下文范围"""
    context_ranges = {}

    if range_type == 'cell_range':
        start_row = bounds.get('start_row', 1)
        end_row = bounds.get('end_row', 1)
        start_col = bounds.get('start_col', 1)
        end_col = bounds.get('end_col', 1)

        # 上方范围（前3行）
        if start_row > 1:
            above_start = max(1, start_row - 3)
            above_end = start_row - 1
            context_ranges['above'] = f"R{above_start}:R{above_end}"

        # 下方范围（后3行）
        context_ranges['below'] = f"R{end_row + 1}:R{end_row + 3}"

        # 左侧范围（前3列）
        if start_col > 1:
            left_start = max(1, start_col - 3)
            left_end = start_col - 1
            context_ranges['left'] = f"C{left_start}:C{left_end}"

        # 右侧范围（后3列）
        context_ranges['right'] = f"C{end_col + 1}:C{end_col + 3}"

    return context_ranges


def _generate_text_visualization(visualization: Dict[str, Any], operation_type: str) -> Dict[str, Any]:
    """生成文本模式可视化"""
    range_details = visualization.get('range_details', {})
    impact_map = visualization.get('impact_map', {})

    content = {
        'header': '',
        'range_display': '',
        'impact_summary': '',
        'risk_indicators': [],
        'ascii_map': ''
    }

    # 标题
    content['header'] = f"📊 操作范围可视化 - {operation_type.upper()}操作"

    # 范围显示
    bounds = range_details.get('bounds', {})
    if bounds.get('start_cell') and bounds.get('end_cell'):
        content['range_display'] = f"📍 影响范围: {bounds['start_cell']} : {bounds['end_cell']}"
    else:
        content['range_display'] = f"📍 影响范围: {visualization['operation_info']['range']}"

    # 影响摘要
    dimensions = range_details.get('dimensions', {})
    content['impact_summary'] = (
        f"📐 尺寸: {dimensions['rows']}行 x {dimensions['columns']}列 "
        f"({dimensions['total_cells']}个单元格) | "
        f"📊 数据: {impact_map['data_cells_affected']}个 | "
        f"📝 公式: {impact_map['formula_cells_affected']}个"
    )

    # 风险指标
    if impact_map['formula_cells_affected'] > 0:
        content['risk_indicators'].append("⚠️ 包含公式")
    if impact_map['data_cells_affected'] > 100:
        content['risk_indicators'].append("📈 大量数据")
    if operation_type in ['delete', 'update']:
        content['risk_indicators'].append("🔄 修改操作")

    # ASCII地图（小范围）
    if dimensions['rows'] <= 20 and dimensions['columns'] <= 10:
        content['ascii_map'] = _generate_ascii_map(visualization, operation_type)

    return content


def _generate_detailed_visualization(visualization: Dict[str, Any], operation_type: str) -> Dict[str, Any]:
    """生成详细模式可视化"""
    text_viz = _generate_text_visualization(visualization, operation_type)

    content = {
        'basic_info': text_viz,
        'cell_details': [],
        'data_analysis': {},
        'risk_analysis': {}
    }

    # 单元格详细信息
    impact_zones = visualization.get('impact_map', {}).get('impact_zones', [])
    for zone in impact_zones[:50]:  # 限制显示数量
        cell_detail = {
            'address': zone['cell_address'],
            'position': f"行{zone['row']}, 列{zone['column']}",
            'current_value': str(zone['value']) if zone['value'] is not None else '[空]',
            'data_type': zone['data_type'],
            'impact_type': zone['impact_type'],
            'risk_level': zone['risk_level']
        }
        content['cell_details'].append(cell_detail)

    # 数据分析
    range_details = visualization.get('range_details', {})
    data_content = range_details.get('data_content', {})
    content['data_analysis'] = {
        'data_density': f"{data_content.get('data_density', 0):.1f}%",
        'has_formulas': data_content.get('has_formulas', False),
        'non_empty_ratio': f"{(data_content.get('non_empty_cells', 0) / max(1, range_details.get('dimensions', {}).get('total_cells', 1)) * 100):.1f}%"
    }

    # 风险分析
    impact_map = visualization.get('impact_map', {})
    content['risk_analysis'] = {
        'high_risk_cells': len([z for z in impact_zones if z['risk_level'] == 'high']),
        'medium_risk_cells': len([z for z in impact_zones if z['risk_level'] == 'medium']),
        'formula_risk': '高' if impact_map['formula_cells_affected'] > 0 else '低',
        'data_volume_risk': '高' if impact_map['data_cells_affected'] > 100 else '中' if impact_map['data_cells_affected'] > 10 else '低'
    }

    return content


def _generate_summary_visualization(visualization: Dict[str, Any], operation_type: str) -> Dict[str, Any]:
    """生成摘要模式可视化"""
    content = {
        'quick_overview': '',
        'key_metrics': {},
        'risk_assessment': '',
        'recommendations': []
    }

    # 快速概览
    range_details = visualization.get('range_details', {})
    dimensions = range_details.get('dimensions', {})
    content['quick_overview'] = f"{operation_type.upper()} {dimensions['rows']}x{dimensions['columns']} 区域"

    # 关键指标
    impact_map = visualization.get('impact_map', {})
    content['key_metrics'] = {
        '总单元格': dimensions['total_cells'],
        '数据单元格': impact_map['data_cells_affected'],
        '空单元格': impact_map['empty_cells_affected'],
        '公式单元格': impact_map['formula_cells_affected']
    }

    # 风险评估
    risk_score = 0
    if impact_map['formula_cells_affected'] > 0:
        risk_score += 30
    if impact_map['data_cells_affected'] > 100:
        risk_score += 20
    if operation_type in ['delete', 'update']:
        risk_score += 20

    if risk_score >= 50:
        content['risk_assessment'] = "🔴 高风险"
    elif risk_score >= 30:
        content['risk_assessment'] = "🟡 中等风险"
    else:
        content['risk_assessment'] = "🟢 低风险"

    # 建议
    if risk_score >= 50:
        content['recommendations'] = ["建议创建备份", "需要用户确认"]
    elif risk_score >= 30:
        content['recommendations'] = ["建议预览操作", "检查公式依赖"]
    else:
        content['recommendations'] = ["可以安全操作"]

    return content


def _generate_matrix_visualization(visualization: Dict[str, Any], operation_type: str) -> Dict[str, Any]:
    """生成矩阵模式可视化"""
    content = {
        'matrix_grid': [],
        'legend': {},
        'dimensions': {}
    }

    # 获取数据
    impact_zones = visualization.get('impact_map', {}).get('impact_zones', [])
    if not impact_zones:
        return content

    # 确定矩阵尺寸
    max_row = max(z['row'] for z in impact_zones)
    max_col = max(z['column'] for z in impact_zones)

    # 限制显示尺寸
    display_rows = min(max_row, 20)
    display_cols = min(max_col, 15)

    content['dimensions'] = {
        'display_rows': display_rows,
        'display_cols': display_cols,
        'actual_rows': max_row,
        'actual_cols': max_col
    }

    # 创建矩阵网格
    matrix = []
    for i in range(display_rows):
        row = []
        for j in range(display_cols):
            # 找到对应的单元格信息
            cell_info = next((z for z in impact_zones if z['row'] == i + 1 and z['column'] == j + 1), None)

            if cell_info:
                if cell_info['data_type'] == 'formula':
                    row.append('F')  # 公式
                elif cell_info['data_type'] == 'empty':
                    row.append('.')  # 空
                else:
                    row.append('D')  # 数据
            else:
                row.append(' ')  # 超出范围

        matrix.append(row)

    content['matrix_grid'] = matrix

    # 图例
    content['legend'] = {
        'D': '数据单元格',
        'F': '公式单元格',
        '.': '空单元格',
        ' ': '超出显示范围'
    }

    return content


def _generate_ascii_map(visualization: Dict[str, Any], operation_type: str) -> str:
    """生成ASCII地图"""
    impact_zones = visualization.get('impact_map', {}).get('impact_zones', [])
    if not impact_zones:
        return ""

    # 获取边界
    max_row = max(z['row'] for z in impact_zones)
    max_col = max(z['column'] for z in impact_zones)

    # 限制显示尺寸
    display_rows = min(max_row, 15)
    display_cols = min(max_col, 12)

    # 创建地图
    map_lines = []
    map_lines.append("    " + "".join(f"{c:2}" for c in range(1, display_cols + 1)))
    map_lines.append("   " + "—" * (display_cols * 2 + 1))

    for i in range(display_rows):
        row_line = f"{i+1:2} |"
        for j in range(display_cols):
            cell_info = next((z for z in impact_zones if z['row'] == i + 1 and z['column'] == j + 1), None)

            if cell_info:
                if cell_info['risk_level'] == 'high':
                    row_line += "⚠️ "
                elif cell_info['risk_level'] == 'medium':
                    row_line += "⚡ "
                elif cell_info['data_type'] == 'formula':
                    row_line += "F "
                elif cell_info['data_type'] == 'empty':
                    row_line += ". "
                else:
                    row_line += "◼ "
            else:
                row_line += "  "

        map_lines.append(row_line)

    return "\n".join(map_lines)


def _col_num_to_letter(col_num: int) -> str:
    """将列号转换为字母"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + (col_num % 26)) + result
        col_num //= 26
    return result


def _generate_visualization_summary(visualization: Dict[str, Any], operation_type: str) -> Dict[str, Any]:
    """生成可视化摘要"""
    range_details = visualization.get('range_details', {})
    impact_map = visualization.get('impact_map', {})
    dimensions = range_details.get('dimensions', {})

    summary = {
        'description': '',
        'complexity': 'simple',
        'visual_impact': 'low',
        'recommended_view': 'text'
    }

    # 描述
    summary['description'] = (
        f"{operation_type.upper()}操作将影响{dimensions['rows']}行{dimensions['columns']}列 "
        f"(共{dimensions['total_cells']}个单元格)，其中{impact_map['data_cells_affected']}个包含数据"
    )

    # 复杂度评估
    if dimensions['total_cells'] > 1000:
        summary['complexity'] = 'complex'
        summary['recommended_view'] = 'summary'
    elif dimensions['total_cells'] > 100:
        summary['complexity'] = 'medium'
        summary['recommended_view'] = 'detailed'
    else:
        summary['complexity'] = 'simple'
        summary['recommended_view'] = 'matrix'

    # 视觉影响
    if impact_map['formula_cells_affected'] > 0:
        summary['visual_impact'] = 'high'
    elif impact_map['data_cells_affected'] > 50:
        summary['visual_impact'] = 'medium'
    else:
        summary['visual_impact'] = 'low'

    return summary


@mcp.tool()
def excel_get_operation_history(
    file_path: Optional[str] = None,
    limit: int = 20
) -> Dict[str, Any]:
    """
    获取Excel操作历史记录

    Args:
        file_path: 文件路径 (可选，用于过滤特定文件的操作)
        limit: 返回的操作记录数量 (默认20)

    Returns:
        Dict: 包含操作历史和统计信息

    Example:
        # 获取所有操作历史
        result = excel_get_operation_history()
        # 获取特定文件的操作历史
        result = excel_get_operation_history("data.xlsx", 10)
    """
    try:
        recent_operations = operation_logger.get_recent_operations(limit)

        # 如果指定了文件路径，过滤操作
        if file_path:
            recent_operations = [
                op for op in recent_operations
                if op.get('details', {}).get('file_path') == file_path
            ]

        # 统计信息
        total_operations = len(recent_operations)
        operation_types = {}
        for op in recent_operations:
            op_type = op.get('operation', 'unknown')
            operation_types[op_type] = operation_types.get(op_type, 0) + 1

        # 统计成功/失败
        success_count = sum(1 for op in recent_operations
                          if op.get('operation') == 'operation_result' and
                          op.get('details', {}).get('success', False))

        error_count = sum(1 for op in recent_operations
                        if op.get('operation') == 'operation_error')

        return {
            'success': True,
            'file_path': file_path,
            'operations': recent_operations,
            'statistics': {
                'total_operations': total_operations,
                'operation_types': operation_types,
                'success_count': success_count,
                'error_count': error_count,
                'success_rate': f"{(success_count / (success_count + error_count) * 100):.1f}%" if (success_count + error_count) > 0 else "0%"
            },
            'message': f"找到 {total_operations} 条操作记录"
        }

    except Exception as e:
        return {
            'success': False,
            'error': 'HISTORY_RETRIEVAL_FAILED',
            'message': f"获取操作历史失败: {str(e)}"
        }


@mcp.tool()
def excel_create_backup(
    file_path: str,
    backup_dir: Optional[str] = None
) -> Dict[str, Any]:
    """
    为Excel文件创建自动备份

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        backup_dir: 备份目录 (可选，默认在文件同目录下创建.backup文件夹)

    Returns:
        Dict: 包含备份结果和备份文件路径

    Example:
        # 创建备份
        result = excel_create_backup("data.xlsx")
        # 指定备份目录
        result = excel_create_backup("data.xlsx", "./backups")
    """
    if not os.path.exists(file_path):
        return {
            'success': False,
            'error': 'FILE_NOT_FOUND',
            'message': f"源文件不存在: {file_path}"
        }

    try:
        # 创建备份目录
        if backup_dir is None:
            base_dir = os.path.dirname(file_path)
            backup_dir = os.path.join(base_dir, ".excel_mcp_backups")

        os.makedirs(backup_dir, exist_ok=True)

        # 生成备份文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.basename(file_path)
        name, ext = os.path.splitext(filename)
        backup_filename = f"{name}_backup_{timestamp}{ext}"
        backup_path = os.path.join(backup_dir, backup_filename)

        # 创建备份
        shutil.copy2(file_path, backup_path)

        # 检查备份大小
        original_size = os.path.getsize(file_path)
        backup_size = os.path.getsize(backup_path)

        return {
            'success': True,
            'original_file': file_path,
            'backup_file': backup_path,
            'backup_directory': backup_dir,
            'file_size': {
                'original': original_size,
                'backup': backup_size
            },
            'timestamp': timestamp,
            'message': f"备份创建成功: {backup_filename}"
        }

    except Exception as e:
        return {
            'success': False,
            'error': 'BACKUP_FAILED',
            'message': f"备份创建失败: {str(e)}"
        }


@mcp.tool()
def excel_restore_backup(
    backup_path: str,
    target_path: Optional[str] = None
) -> Dict[str, Any]:
    """
    从备份恢复Excel文件

    Args:
        backup_path: 备份文件路径
        target_path: 目标文件路径 (可选，默认恢复到原始位置)

    Returns:
        Dict: 包含恢复结果

    Example:
        # 恢复备份
        result = excel_restore_backup("./backups/data_backup_20250117_143022.xlsx")
        # 恢复到指定位置
        result = excel_restore_backup("./backups/data_backup_20250117_143022.xlsx", "restored_data.xlsx")
    """
    if not os.path.exists(backup_path):
        return {
            'success': False,
            'error': 'BACKUP_NOT_FOUND',
            'message': f"备份文件不存在: {backup_path}"
        }

    try:
        # 确定目标路径
        if target_path is None:
            # 尝试从备份文件名推断原始文件名
            filename = os.path.basename(backup_path)
            if "_backup_" in filename:
                # 移除备份时间戳
                parts = filename.split("_backup_")
                target_path = parts[0] + os.path.splitext(backup_path)[1]
            else:
                target_path = filename.replace("_backup_", ".")

        # 创建目标目录
        target_dir = os.path.dirname(target_path)
        if target_dir:
            os.makedirs(target_dir, exist_ok=True)

        # 检查目标文件是否存在
        target_exists = os.path.exists(target_path)

        # 执行恢复
        shutil.copy2(backup_path, target_path)

        return {
            'success': True,
            'backup_file': backup_path,
            'target_file': target_path,
            'target_existed': target_exists,
            'message': f"文件恢复成功: {os.path.basename(target_path)}"
        }

    except Exception as e:
        return {
            'success': False,
            'error': 'RESTORE_FAILED',
            'message': f"恢复失败: {str(e)}"
        }


@mcp.tool()
def excel_list_backups(
    file_path: str,
    backup_dir: Optional[str] = None
) -> Dict[str, Any]:
    """
    列出指定文件的所有备份

    Args:
        file_path: 原始Excel文件路径
        backup_dir: 备份目录 (可选)

    Returns:
        Dict: 包含备份文件列表

    Example:
        result = excel_list_backups("data.xlsx")
    """
    try:
        # 确定备份目录
        if backup_dir is None:
            base_dir = os.path.dirname(file_path)
            backup_dir = os.path.join(base_dir, ".excel_mcp_backups")

        if not os.path.exists(backup_dir):
            return {
                'success': True,
                'backups': [],
                'message': "备份目录不存在"
            }

        # 获取文件名
        filename = os.path.basename(file_path)
        name, ext = os.path.splitext(filename)
        backup_pattern = f"{name}_backup_*{ext}"

        # 查找备份文件
        backup_files = []
        for file in os.listdir(backup_dir):
            if file.startswith(f"{name}_backup_") and file.endswith(ext):
                full_path = os.path.join(backup_dir, file)
                stat = os.stat(full_path)
                backup_files.append({
                    'filename': file,
                    'path': full_path,
                    'size': stat.st_size,
                    'created_time': datetime.fromtimestamp(stat.st_ctime),
                    'modified_time': datetime.fromtimestamp(stat.st_mtime)
                })

        # 按时间排序
        backup_files.sort(key=lambda x: x['created_time'], reverse=True)

        return {
            'success': True,
            'original_file': file_path,
            'backup_directory': backup_dir,
            'backups': backup_files,
            'total_backups': len(backup_files)
        }

    except Exception as e:
        return {
            'success': False,
            'error': 'LIST_BACKUPS_FAILED',
            'message': f"列出备份失败: {str(e)}"
        }


@mcp.tool()
def excel_insert_rows(
    file_path: str,
    sheet_name: str,
    row_index: int,
    count: int = 1
) -> Dict[str, Any]:
    """
    在指定位置插入空行

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        row_index: 插入位置 (1-based，即第1行对应Excel中的第1行)
        count: 插入行数 (默认值: 1，即插入1行)

    Returns:
        Dict: 包含 success、inserted_rows、message

    Example:
        # 在第3行插入1行（使用默认count=1）
        result = excel_insert_rows("data.xlsx", "Sheet1", 3)
        # 在第5行插入3行（明确指定count）
        result = excel_insert_rows("data.xlsx", "Sheet1", 5, 3)
    """
    return ExcelOperations.insert_rows(file_path, sheet_name, row_index, count)


@mcp.tool()
def excel_insert_columns(
    file_path: str,
    sheet_name: str,
    column_index: int,
    count: int = 1
) -> Dict[str, Any]:
    """
    在指定位置插入空列

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        column_index: 插入位置 (1-based，即第1列对应Excel中的A列)
        count: 插入列数 (默认值: 1，即插入1列)

    Returns:
        Dict: 包含 success、inserted_columns、message

    Example:
        # 在第2列插入1列（使用默认count=1，即在B列前插入1列）
        result = excel_insert_columns("data.xlsx", "Sheet1", 2)
        # 在第1列插入2列（明确指定count，即在A列前插入2列）
        result = excel_insert_columns("data.xlsx", "Sheet1", 1, 2)
    """
    return ExcelOperations.insert_columns(file_path, sheet_name, column_index, count)


@mcp.tool()
def excel_find_last_row(
    file_path: str,
    sheet_name: str,
    column: Optional[Union[str, int]] = None
) -> Dict[str, Any]:
    """
    查找表格中最后一行有数据的位置

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        column: 指定列来查找最后一行（可选）
            - None: 查找整个工作表的最后一行
            - 整数: 列索引 (1-based，1=A列)
            - 字符串: 列名 (A, B, C...)

    Returns:
        Dict: 包含 success、last_row、message 等信息

    Example:
        # 查找整个工作表的最后一行
        result = excel_find_last_row("data.xlsx", "Sheet1")
        # 查找A列的最后一行有数据的位置
        result = excel_find_last_row("data.xlsx", "Sheet1", "A")
        # 查找第3列的最后一行有数据的位置
        result = excel_find_last_row("data.xlsx", "Sheet1", 3)
    """
    return ExcelOperations.find_last_row(file_path, sheet_name, column)


@mcp.tool()
def excel_create_file(
    file_path: str,
    sheet_names: Optional[List[str]] = None
) -> Dict[str, Any]:
    """
    创建新的Excel文件

    Args:
        file_path: 新文件路径 (必须以.xlsx或.xlsm结尾)
        sheet_names: 工作表名称列表 (默认值: None)
            - None: 创建包含一个默认工作表"Sheet1"的文件
            - []: 创建空的工作簿
            - ["名称1", "名称2"]: 创建包含指定名称工作表的文件

    Returns:
        Dict: 包含 success、file_path、sheets

    Example:
        # 创建简单文件（使用默认sheet_names=None，会有一个"Sheet1"）
        result = excel_create_file("new_file.xlsx")
        # 创建包含多个工作表的文件
        result = excel_create_file("report.xlsx", ["数据", "图表", "汇总"])
    """
    return ExcelOperations.create_file(file_path, sheet_names)


@mcp.tool()
def excel_export_to_csv(
    file_path: str,
    output_path: str,
    sheet_name: Optional[str] = None,
    encoding: str = "utf-8"
) -> Dict[str, Any]:
    """
    将Excel工作表导出为CSV文件

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        output_path: 输出CSV文件路径
        sheet_name: 工作表名称 (默认使用活动工作表)
        encoding: 文件编码 (默认: utf-8，可选: gbk)

    Returns:
        Dict: 包含 success、output_path、row_count、message

    Example:
        # 导出活动工作表为CSV
        result = excel_export_to_csv("data.xlsx", "output.csv")
        # 导出指定工作表
        result = excel_export_to_csv("report.xlsx", "summary.csv", "汇总", "gbk")
    """
    return ExcelOperations.export_to_csv(file_path, output_path, sheet_name, encoding)


@mcp.tool()
def excel_import_from_csv(
    csv_path: str,
    output_path: str,
    sheet_name: str = "Sheet1",
    encoding: str = "utf-8",
    has_header: bool = True
) -> Dict[str, Any]:
    """
    从CSV文件导入数据创建Excel文件

    Args:
        csv_path: CSV文件路径
        output_path: 输出Excel文件路径
        sheet_name: 工作表名称 (默认: Sheet1)
        encoding: CSV文件编码 (默认: utf-8，可选: gbk)
        has_header: 是否包含表头行

    Returns:
        Dict: 包含 success、output_path、row_count、sheet_name

    Example:
        # 从CSV创建Excel文件
        result = excel_import_from_csv("data.csv", "output.xlsx")
        # 指定编码和工作表名
        result = excel_import_from_csv("sales.csv", "report.xlsx", "销售数据", "gbk")
    """
    return ExcelOperations.import_from_csv(csv_path, output_path, sheet_name, encoding, has_header)


@mcp.tool()
def excel_convert_format(
    input_path: str,
    output_path: str,
    target_format: str = "xlsx"
) -> Dict[str, Any]:
    """
    转换Excel文件格式

    Args:
        input_path: 输入文件路径
        output_path: 输出文件路径
        target_format: 目标格式，可选值: "xlsx", "xlsm", "csv", "json"

    Returns:
        Dict: 包含 success、input_format、output_format、file_size

    Example:
        # 将xlsm转换为xlsx
        result = excel_convert_format("macro.xlsm", "data.xlsx", "xlsx")
        # 转换为JSON格式
        result = excel_convert_format("data.xlsx", "data.json", "json")
    """
    return ExcelOperations.convert_format(input_path, output_path, target_format)


@mcp.tool()
def excel_merge_files(
    input_files: List[str],
    output_path: str,
    merge_mode: str = "sheets"
) -> Dict[str, Any]:
    """
    合并多个Excel文件

    Args:
        input_files: 输入文件路径列表
        output_path: 输出文件路径
        merge_mode: 合并模式，可选值:
            - "sheets": 将每个文件作为独立工作表
            - "append": 将数据追加到单个工作表中
            - "horizontal": 水平合并（按列）

    Returns:
        Dict: 包含 success、merged_files、total_sheets、output_path

    Example:
        # 将多个文件合并为多个工作表
        files = ["file1.xlsx", "file2.xlsx", "file3.xlsx"]
        result = excel_merge_files(files, "merged.xlsx", "sheets")

        # 将数据追加合并
        result = excel_merge_files(files, "combined.xlsx", "append")
    """
    return ExcelOperations.merge_files(input_files, output_path, merge_mode)


@mcp.tool()
def excel_get_file_info(file_path: str) -> Dict[str, Any]:
    """
    获取Excel文件的详细信息

    Args:
        file_path: Excel文件路径

    Returns:
        Dict: 包含文件信息，如大小、创建时间、工作表数量、格式等

    Example:
        # 获取文件详细信息
        result = excel_get_file_info("data.xlsx")
        # 返回: {
        #   'success': True,
        #   'file_size': 12345,
        #   'created_time': '2025-01-01 10:00:00',
        #   'modified_time': '2025-01-02 15:30:00',
        #   'format': 'xlsx',
        #   'sheet_count': 3,
        #   'has_macros': False
        # }
    """
    return ExcelOperations.get_file_info(file_path)


@mcp.tool()
def excel_create_sheet(
    file_path: str,
    sheet_name: str,
    index: Optional[int] = None
) -> Dict[str, Any]:
    """
    在文件中创建新工作表

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 新工作表名称 (不能与现有工作表重复)
        index: 插入位置 (0-based，默认值: None)
            - None: 在所有工作表的最后位置创建
            - 0: 在第一个位置创建
            - 1: 在第二个位置创建，以此类推

    Returns:
        Dict: 包含 success、sheet_name、total_sheets

    Example:
        # 创建新工作表到末尾（使用默认index=None）
        result = excel_create_sheet("data.xlsx", "新数据")
        # 创建新工作表到第一个位置（index=0）
        result = excel_create_sheet("data.xlsx", "首页", 0)
    """
    return ExcelOperations.create_sheet(file_path, sheet_name, index)


@mcp.tool()
def excel_delete_sheet(
    file_path: str,
    sheet_name: str
) -> Dict[str, Any]:
    """
    删除指定工作表

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 要删除的工作表名称

    Returns:
        Dict: 包含 success、deleted_sheet、remaining_sheets

    Example:
        # 删除指定工作表
        result = excel_delete_sheet("data.xlsx", "临时数据")
    """
    # 开始操作会话
    operation_logger.start_session(file_path)

    # 记录删除工作表操作日志
    operation_logger.log_operation("delete_sheet", {
        "sheet_name": sheet_name
    })

    try:
        result = ExcelOperations.delete_sheet(file_path, sheet_name)

        # 记录操作结果
        operation_logger.log_operation("operation_result", {
            "success": result.get('success', False),
            "deleted_sheet": result.get('deleted_sheet', ''),
            "remaining_sheets": result.get('remaining_sheets', 0),
            "message": result.get('message', '')
        })

        return result

    except Exception as e:
        # 记录错误
        operation_logger.log_operation("operation_error", {
            "error": str(e),
            "message": f"删除工作表操作失败: {str(e)}"
        })

        return {
            'success': False,
            'error': 'DELETE_SHEET_FAILED',
            'message': f"删除工作表操作失败: {str(e)}"
        }


@mcp.tool()
def excel_rename_sheet(
    file_path: str,
    old_name: str,
    new_name: str
) -> Dict[str, Any]:
    """
    重命名工作表

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        old_name: 当前工作表名称
        new_name: 新工作表名称 (不能与现有重复)

    Returns:
        Dict: 包含 success、old_name、new_name

    Example:
        # 重命名工作表
        result = excel_rename_sheet("data.xlsx", "Sheet1", "主数据")
    """
    return ExcelOperations.rename_sheet(file_path, old_name, new_name)


@mcp.tool()
def excel_delete_rows(
    file_path: str,
    sheet_name: str,
    row_index: int,
    count: int = 1
) -> Dict[str, Any]:
    """
    删除指定行

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        row_index: 起始行号 (1-based，即第1行对应Excel中的第1行)
        count: 删除行数 (默认值: 1，即删除1行)

    Returns:
        Dict: 包含 success、deleted_rows、message

    Example:
        # 删除第5行（使用默认count=1）
        result = excel_delete_rows("data.xlsx", "Sheet1", 5)
        # 删除第3-5行（删除3行，从第3行开始）
        result = excel_delete_rows("data.xlsx", "Sheet1", 3, 3)
    """
    # 开始操作会话
    operation_logger.start_session(file_path)

    # 记录删除操作日志
    operation_logger.log_operation("delete_rows", {
        "sheet_name": sheet_name,
        "row_index": row_index,
        "count": count
    })

    try:
        result = ExcelOperations.delete_rows(file_path, sheet_name, row_index, count)

        # 记录操作结果
        operation_logger.log_operation("operation_result", {
            "success": result.get('success', False),
            "deleted_rows": result.get('deleted_rows', 0),
            "message": result.get('message', '')
        })

        return result

    except Exception as e:
        # 记录错误
        operation_logger.log_operation("operation_error", {
            "error": str(e),
            "message": f"删除行操作失败: {str(e)}"
        })

        return {
            'success': False,
            'error': 'DELETE_ROWS_FAILED',
            'message': f"删除行操作失败: {str(e)}"
        }


@mcp.tool()
def excel_delete_columns(
    file_path: str,
    sheet_name: str,
    column_index: int,
    count: int = 1
) -> Dict[str, Any]:
    """
    删除指定列

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        column_index: 起始列号 (1-based，即第1列对应Excel中的A列)
        count: 删除列数 (默认值: 1，即删除1列)

    Returns:
        Dict: 包含 success、deleted_columns、message

    Example:
        # 删除第2列（使用默认count=1，即删除B列）
        result = excel_delete_columns("data.xlsx", "Sheet1", 2)
        # 删除第1-3列（删除3列，从A列开始删除A、B、C列）
        result = excel_delete_columns("data.xlsx", "Sheet1", 1, 3)
    """
    # 开始操作会话
    operation_logger.start_session(file_path)

    # 记录删除列操作日志
    operation_logger.log_operation("delete_columns", {
        "sheet_name": sheet_name,
        "column_index": column_index,
        "count": count
    })

    try:
        result = ExcelOperations.delete_columns(file_path, sheet_name, column_index, count)

        # 记录操作结果
        operation_logger.log_operation("operation_result", {
            "success": result.get('success', False),
            "deleted_columns": result.get('deleted_columns', 0),
            "message": result.get('message', '')
        })

        return result

    except Exception as e:
        # 记录错误
        operation_logger.log_operation("operation_error", {
            "error": str(e),
            "message": f"删除列操作失败: {str(e)}"
        })

        return {
            'success': False,
            'error': 'DELETE_COLUMNS_FAILED',
            'message': f"删除列操作失败: {str(e)}"
        }

# 暂时注释掉, 以后可能会用到
# @mcp.tool()
def excel_set_formula(
    file_path: str,
    sheet_name: str,
    cell_address: str,
    formula: str
) -> Dict[str, Any]:
    """
    设置单元格公式

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        cell_address: 单元格地址 (如"A1")
        formula: Excel公式 (不包含等号)

    Returns:
        Dict: 包含 success、formula、calculated_value

    Example:
        # 设置求和公式
        result = excel_set_formula("data.xlsx", "Sheet1", "D10", "SUM(D1:D9)")
        # 设置平均值公式
        result = excel_set_formula("data.xlsx", "Sheet1", "E1", "AVERAGE(A1:A10)")
    """
    return ExcelOperations.set_formula(file_path, sheet_name, cell_address, formula)

# 暂时注释掉, 以后可能会用到
# @mcp.tool()
def excel_evaluate_formula(
    file_path: str,
    formula: str,
    context_sheet: Optional[str] = None
) -> Dict[str, Any]:
    """
    临时执行Excel公式并返回计算结果，不修改文件

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        formula: Excel公式 (不包含等号，如"SUM(A1:A10)")
        context_sheet: 公式执行的上下文工作表名称

    Returns:
        Dict: 包含 success、formula、result、result_type

    Example:
        # 计算A1:A10的和
        result = excel_evaluate_formula("data.xlsx", "SUM(A1:A10)")
        # 计算特定工作表的平均值
        result = excel_evaluate_formula("data.xlsx", "AVERAGE(B:B)", "Sheet1")
    """
    return ExcelOperations.evaluate_formula(formula, context_sheet)


@mcp.tool()
def excel_format_cells(
    file_path: str,
    sheet_name: str,
    range: str,
    formatting: Optional[Dict[str, Any]] = None,
    preset: Optional[str] = None
) -> Dict[str, Any]:
    """
    设置单元格格式（字体、颜色、对齐等）- 支持自定义和预设两种模式

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        range: 目标范围 (如"A1:C10")
        formatting: 自定义格式配置字典（可选）：
            - font: {'name': '宋体', 'size': 12, 'bold': True, 'color': 'FF0000'}
            - fill: {'color': 'FFFF00'}
            - alignment: {'horizontal': 'center', 'vertical': 'center'}
        preset: 预设样式（可选），可选值: "title", "header", "data", "highlight", "currency"

    注意: formatting 和 preset 必须指定其中一个，如果同时指定，preset 优先

    Returns:
        Dict: 包含 success、formatted_count、message

    Example:
        # 使用预设样式
        result = excel_format_cells("data.xlsx", "Sheet1", "A1:D1", preset="title")

        # 使用自定义格式
        result = excel_format_cells("data.xlsx", "Sheet1", "A1:D1",
            formatting={'font': {'bold': True, 'color': 'FF0000'}})
    """
    return ExcelOperations.format_cells(file_path, sheet_name, range, formatting, preset)


@mcp.tool()
def excel_merge_cells(
    file_path: str,
    sheet_name: str,
    range: str
) -> Dict[str, Any]:
    """
    合并指定范围的单元格

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        range: 要合并的范围 (如"A1:C3")

    Returns:
        Dict: 包含 success、message、merged_range

    Example:
        # 合并A1:C3范围的单元格
        result = excel_merge_cells("data.xlsx", "Sheet1", "A1:C3")
        # 合并标题行
        result = excel_merge_cells("report.xlsx", "Summary", "A1:E1")
    """
    return ExcelOperations.merge_cells(file_path, sheet_name, range)


@mcp.tool()
def excel_unmerge_cells(
    file_path: str,
    sheet_name: str,
    range: str
) -> Dict[str, Any]:
    """
    取消合并指定范围的单元格

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        range: 要取消合并的范围 (如"A1:C3")

    Returns:
        Dict: 包含 success、message、unmerged_range

    Example:
        # 取消合并A1:C3范围的单元格
        result = excel_unmerge_cells("data.xlsx", "Sheet1", "A1:C3")
    """
    return ExcelOperations.unmerge_cells(file_path, sheet_name, range)


@mcp.tool()
def excel_set_borders(
    file_path: str,
    sheet_name: str,
    range: str,
    border_style: str = "thin"
) -> Dict[str, Any]:
    """
    为指定范围设置边框样式

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        range: 目标范围 (如"A1:C10")
        border_style: 边框样式，可选值: "thin", "thick", "medium", "double", "dotted", "dashed"

    Returns:
        Dict: 包含 success、message、styled_range

    Example:
        # 为表格添加细边框
        result = excel_set_borders("data.xlsx", "Sheet1", "A1:E10", "thin")
        # 为标题添加粗边框
        result = excel_set_borders("data.xlsx", "Sheet1", "A1:E1", "thick")
    """
    return ExcelOperations.set_borders(file_path, sheet_name, range, border_style)


@mcp.tool()
def excel_set_row_height(
    file_path: str,
    sheet_name: str,
    row_index: int,
    height: float,
    count: int = 1
) -> Dict[str, Any]:
    """
    调整指定行的高度

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        row_index: 起始行号 (1-based)
        height: 行高 (磅值，如15.0)
        count: 调整行数 (默认值: 1)

    Returns:
        Dict: 包含 success、message、affected_rows

    Example:
        # 调整第1行高度为25磅
        result = excel_set_row_height("data.xlsx", "Sheet1", 1, 25.0)
        # 调整第2-4行高度为18磅
        result = excel_set_row_height("data.xlsx", "Sheet1", 2, 18.0, 3)
    """
    return ExcelOperations.set_row_height(file_path, sheet_name, row_index, height, count)


@mcp.tool()
def excel_set_column_width(
    file_path: str,
    sheet_name: str,
    column_index: int,
    width: float,
    count: int = 1
) -> Dict[str, Any]:
    """
    调整指定列的宽度

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        column_index: 起始列号 (1-based，1=A列)
        width: 列宽 (字符单位，如12.0)
        count: 调整列数 (默认值: 1)

    Returns:
        Dict: 包含 success、message、affected_columns

    Example:
        # 调整A列宽度为15字符
        result = excel_set_column_width("data.xlsx", "Sheet1", 1, 15.0)
        # 调整B-D列宽度为12字符
        result = excel_set_column_width("data.xlsx", "Sheet1", 2, 12.0, 3)
    """
    return ExcelOperations.set_column_width(file_path, sheet_name, column_index, width, count)


# ==================== Excel比较功能 ====================

# @mcp.tool()
def excel_compare_files(
    file1_path: str,
    file2_path: str,
    id_column: Union[int, str] = 1,
    header_row: int = 1
) -> Dict[str, Any]:
    """
    比较两个Excel文件 - 游戏开发专用版

    专注于ID对象的新增、删除、修改检测，自动识别配置表变化。

    Args:
        file1_path: 第一个Excel文件路径
        file2_path: 第二个Excel文件路径
        id_column: ID列位置（1-based数字或列名），默认第一列
        header_row: 表头行号（1-based），默认第一行

    Returns:
        Dict: 比较结果，包含新增、删除、修改的ID对象信息
        - 🆕 新增对象：ID在文件2中新出现
        - 🗑️ 删除对象：ID在文件1中存在但文件2中消失
        - 🔄 修改对象：ID存在于两文件中但属性发生变化
    """
    return ExcelOperations.compare_files(file1_path, file2_path)


@mcp.tool()
def excel_check_duplicate_ids(
    file_path: str,
    sheet_name: str,
    id_column: Union[int, str] = 1,
    header_row: int = 1
) -> Dict[str, Any]:
    """
    检查Excel工作表中ID列的重复值

    专为游戏配置表设计，快速识别ID重复问题，确保配置数据的唯一性。

    Args:
        file_path: Excel文件路径 (.xlsx/.xlsm)
        sheet_name: 工作表名称
        id_column: ID列位置（1-based数字或列名），默认第一列
        header_row: 表头行号（1-based），默认第一行

    Returns:
        Dict: 查重结果
        {
            "success": true,
            "has_duplicates": true,
            "duplicate_count": 2,
            "total_ids": 100,
            "unique_ids": 98,
            "duplicates": [
                {
                    "id_value": "100001",
                    "count": 3,
                    "rows": [5, 15, 25]
                },
                {
                    "id_value": "100002",
                    "count": 2,
                    "rows": [8, 18]
                }
            ],
            "message": "发现2个重复ID，涉及5行数据"
        }

    Example:
        # 检查技能配置表ID重复
        result = excel_check_duplicate_ids("skills.xlsx", "技能配置表")
        # 检查装备表第2列ID重复
        result = excel_check_duplicate_ids("items.xlsx", "装备配置表", id_column=2)
    """
    return ExcelOperations.check_duplicate_ids(file_path, sheet_name, id_column, header_row)


@mcp.tool()
def excel_compare_sheets(
    file1_path: str,
    sheet1_name: str,
    file2_path: str,
    sheet2_name: str,
    id_column: Union[int, str] = 1,
    header_row: int = 1
) -> Dict[str, Any]:
    """
    比较两个Excel工作表，识别ID对象的新增、删除、修改。

    专为游戏配置表设计，使用紧凑数组格式提高传输效率。

    Args:
        file1_path: 第一个Excel文件路径
        sheet1_name: 第一个工作表名称
        file2_path: 第二个Excel文件路径
        sheet2_name: 第二个工作表名称
        id_column: ID列位置（1-based数字或列名），默认第一列
        header_row: 表头行号（1-based），默认第一行

    Returns:
        Dict: 比较结果
        {
            "success": true,
            "message": "成功比较工作表，发现3处差异",
            "data": {
                "sheet_name": "TrSkill vs TrSkill",
                "total_differences": 3,
                "row_differences": [
                    // 字段定义
                    ["row_id", "difference_type", "row_index1", "row_index2", "sheet_name", "field_differences"],

                    // 新增行
                    ["100001", "row_added", 0, 5, "TrSkill", null],

                    // 删除行
                    ["100002", "row_removed", 8, 0, "TrSkill", null],

                    // 修改行 - 包含变化的字段
                    ["100003", "row_modified", 10, 10, "TrSkill",
                        // field_differences: 变化的字段数组，每个元素格式 [字段名, 旧值, 新值, 变化类型]
                        [["技能名称", "火球术", "冰球术", "text_change"]]
                    ]
                ],
                "structural_changes": {
                    "max_row": {"sheet1": 100, "sheet2": 101, "difference": 1}
                }
            }
        }

    数据解析：
        row_differences[0] = 字段定义（索引说明）
        row_differences[1+] = 实际数据行

        对于row_modified类型：
        - field_differences: 变化的字段数组
          格式：[[字段名, 旧值, 新值, 变化类型], ...]
          变化类型："text_change" | "numeric_change" | "formula_change"

        对于row_added/row_removed类型：
        - field_differences为null，因为整行都是变化

    Example:
        result = excel_compare_sheets("old.xlsx", "Sheet1", "new.xlsx", "Sheet1")
        differences = result['data']['row_differences']
        for row in differences[1:]:  # 跳过字段定义行
            row_id, diff_type = row[0], row[1]
            print(f"{diff_type}: {row_id}")
    """
    return ExcelOperations.compare_sheets(file1_path, sheet1_name, file2_path, sheet2_name, id_column, header_row)
# ==================== 主程序 ====================
if __name__ == "__main__":
    # 运行FastMCP服务器
    mcp.run()
