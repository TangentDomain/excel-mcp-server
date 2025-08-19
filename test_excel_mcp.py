#!/usr/bin/env python3
"""
Excel MCP Server 功能测试脚本
测试所有三个核心功能的正确性
"""

import os
import sys
import tempfile
from pathlib import Path

# 添加当前目录到Python路径以导入server模块
current_dir = Path(__file__).parent
sys.path.insert(0, str(current_dir))

import openpyxl
from server import excel_regex_search, excel_get_range, excel_update_range, excel_list_sheets

def create_complex_test_file():
    """创建一个复杂的测试Excel文件"""
    wb = openpyxl.Workbook()

    # 第一个工作表：基础数据测试
    ws1 = wb.active
    ws1.title = "基础数据"

    # 填充基础测试数据
    test_data = [
        ["姓名", "邮箱", "电话", "价格"],
        ["张三", "zhang.san@company.com", "138-0000-1234", 100.50],
        ["李四", "li.si@example.org", "139-1111-2345", 200.75],
        ["王五", "wang.wu@test.net", "186-2222-3456", 300.25],
        ["赵六", "zhao.liu@sample.com", "187-3333-4567", 400.00]
    ]

    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)

    # 添加一些公式
    ws1['E2'] = '=D2*1.1'  # 加10%
    ws1['E3'] = '=D3*1.1'
    ws1['E4'] = '=D4*1.1'
    ws1['E5'] = '=D5*1.1'

    # 第二个工作表：复杂数据测试
    ws2 = wb.create_sheet("复杂数据")
    ws2['A1'] = "产品代码"
    ws2['B1'] = "描述"
    ws2['A2'] = "PROD-001"
    ws2['B2'] = "这是一个产品描述，包含特殊字符：#@%&*"
    ws2['A3'] = "PROD-002"
    ws2['B3'] = "另一个产品，价格$99.99"
    ws2['A4'] = "PROD-003"
    ws2['B4'] = "第三个产品，邮箱联系：contact@product.com"

    return wb

def test_regex_search(file_path):
    """测试正则搜索功能"""
    print("\n🔍 测试正则搜索功能...")

    # 测试1: 搜索邮箱地址
    print("测试1: 搜索邮箱地址")
    result = excel_regex_search(
        file_path=file_path,
        pattern=r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b",
        flags="i"
    )

    if result['success']:
        print(f"  ✅ 找到 {result['total_matches']} 个邮箱地址")
        for match in result['matches']:
            print(f"    📍 {match['sheet']}.{match['cell']}: {match['match']}")
    else:
        print(f"  ❌ 搜索失败: {result['error']}")

    # 测试2: 搜索电话号码
    print("\n测试2: 搜索电话号码")
    result = excel_regex_search(
        file_path=file_path,
        pattern=r"\d{3}-\d{4}-\d{4}",
        flags=""
    )

    if result['success']:
        print(f"  ✅ 找到 {result['total_matches']} 个电话号码")
        for match in result['matches']:
            print(f"    📍 {match['sheet']}.{match['cell']}: {match['match']}")
    else:
        print(f"  ❌ 搜索失败: {result['error']}")

    # 测试3: 搜索产品代码
    print("\n测试3: 搜索产品代码")
    result = excel_regex_search(
        file_path=file_path,
        pattern=r"PROD-\d{3}",
        flags=""
    )

    if result['success']:
        print(f"  ✅ 找到 {result['total_matches']} 个产品代码")
        for match in result['matches']:
            print(f"    📍 {match['sheet']}.{match['cell']}: {match['match']}")
    else:
        print(f"  ❌ 搜索失败: {result['error']}")

def test_sheet_list(file_path):
    """测试工作表列表功能"""
    print("\n📋 测试工作表列表功能...")

    result = excel_list_sheets(file_path=file_path)

    if result['success']:
        print(f"  ✅ 成功获取工作表列表，共 {result['total_sheets']} 个工作表")
        print(f"    🎯 当前活动工作表: {result['active_sheet']}")
        for sheet in result['sheets']:
            active_marker = "🎯" if sheet['is_active'] else "📄"
            print(f"    {active_marker} {sheet['index']+1}. {sheet['name']} (数据范围: {sheet['max_column_letter']}{sheet['max_row']})")
    else:
        print(f"  ❌ 获取失败: {result['error']}")

def test_row_column_access(file_path):
    """测试行列访问功能"""
    print("\n🔢 测试行列访问功能...")

    # 测试1: 获取第1行数据
    print("测试1: 获取第1行数据 (1:1)")
    result = excel_get_range(
        file_path=file_path,
        range_expression="1:1",
        include_formatting=False
    )

    if result['success']:
        print(f"  ✅ 成功获取第1行，维度: {result['dimensions']['rows']}x{result['dimensions']['columns']}")
        print(f"    📊 范围类型: {result['range_type']}")
        if result['data'] and len(result['data'][0]) > 0:
            values = [cell['value'] for cell in result['data'][0][:5]]  # 只显示前5列
            print(f"    📋 前5列数据: {values}")
    else:
        print(f"  ❌ 获取失败: {result['error']}")

    # 测试2: 获取A列数据
    print("\n测试2: 获取A列数据 (A:A)")
    result = excel_get_range(
        file_path=file_path,
        range_expression="A:A",
        include_formatting=False
    )

    if result['success']:
        print(f"  ✅ 成功获取A列，维度: {result['dimensions']['rows']}x{result['dimensions']['columns']}")
        print(f"    📊 范围类型: {result['range_type']}")
        if result['data'] and len(result['data']) > 0:
            values = [row[0]['value'] for row in result['data'][:5]]  # 只显示前5行
            print(f"    📋 前5行数据: {values}")
    else:
        print(f"  ❌ 获取失败: {result['error']}")

    # 测试3: 获取第3行数据 (单行模式)
    print("\n测试3: 获取第3行数据 (3)")
    result = excel_get_range(
        file_path=file_path,
        range_expression="3",
        include_formatting=False
    )

    if result['success']:
        print(f"  ✅ 成功获取第3行，维度: {result['dimensions']['rows']}x{result['dimensions']['columns']}")
        print(f"    📊 范围类型: {result['range_type']}")
    else:
        print(f"  ❌ 获取失败: {result['error']}")

    # 测试4: 获取B列数据 (单列模式)
    print("\n测试4: 获取B列数据 (B)")
    result = excel_get_range(
        file_path=file_path,
        range_expression="B",
        include_formatting=False
    )

    if result['success']:
        print(f"  ✅ 成功获取B列，维度: {result['dimensions']['rows']}x{result['dimensions']['columns']}")
        print(f"    📊 范围类型: {result['range_type']}")
    else:
        print(f"  ❌ 获取失败: {result['error']}")

def test_range_get(file_path):
    """测试范围获取功能"""
    print("\n📊 测试范围获取功能...")

    # 测试1: 获取基础数据表头
    print("测试1: 获取基础数据表头 (A1:D1)")
    result = excel_get_range(
        file_path=file_path,
        range_expression="基础数据!A1:D1",
        include_formatting=False
    )

    if result['success']:
        print(f"  ✅ 成功获取范围，维度: {result['dimensions']['rows']}x{result['dimensions']['columns']}")
        for row in result['data']:
            values = [cell['value'] for cell in row]
            print(f"    📋 {values}")
    else:
        print(f"  ❌ 获取失败: {result['error']}")

    # 测试2: 获取完整数据区域
    print("\n测试2: 获取数据区域 (A1:E5)")
    result = excel_get_range(
        file_path=file_path,
        range_expression="A1:E5"
    )

    if result['success']:
        print(f"  ✅ 成功获取范围，维度: {result['dimensions']['rows']}x{result['dimensions']['columns']}")
        print(f"    📍 工作表: {result['sheet_name']}")
        print(f"    📋 数据行数: {len(result['data'])}")
    else:
        print(f"  ❌ 获取失败: {result['error']}")

def test_range_update(file_path):
    """测试范围修改功能"""
    print("\n✏️ 测试范围修改功能...")

    # 测试1: 修改单个单元格
    print("测试1: 修改单个单元格 (F1)")
    result = excel_update_range(
        file_path=file_path,
        range_expression="F1",
        data=[["测试修改"]],
        preserve_formulas=True
    )

    if result['success']:
        print(f"  ✅ 成功修改 {result['modified_cells_count']} 个单元格")
        for cell in result['modified_cells']:
            print(f"    📝 {cell['coordinate']}: {cell['old_value']} → {cell['new_value']}")
    else:
        print(f"  ❌ 修改失败: {result['error']}")

    # 测试2: 批量修改范围
    print("\n测试2: 批量修改范围 (F2:H3)")
    result = excel_update_range(
        file_path=file_path,
        range_expression="F2:H3",
        data=[
            ["批量1", "批量2", "批量3"],
            ["数据A", "数据B", "数据C"]
        ],
        preserve_formulas=True
    )

    if result['success']:
        print(f"  ✅ 成功修改 {result['modified_cells_count']} 个单元格")
        print(f"    📍 工作表: {result['sheet_name']}")
    else:
        print(f"  ❌ 修改失败: {result['error']}")

    # 验证修改结果
    print("\n验证修改结果...")
    verify_result = excel_get_range(
        file_path=file_path,
        range_expression="F1:H3"
    )

    if verify_result['success']:
        print("  ✅ 修改验证成功:")
        for row in verify_result['data']:
            values = [cell['value'] for cell in row]
            print(f"    📋 {values}")

def main():
    """主测试流程"""
    print("🧪 Excel MCP Server 功能测试")
    print("=" * 50)

    # 创建测试文件
    test_file = "test-data.xlsx"
    print(f"📁 创建测试文件: {test_file}")

    try:
        wb = create_complex_test_file()
        wb.save(test_file)
        print("  ✅ 测试文件创建成功")

        # 获取文件绝对路径
        file_path = str(Path(test_file).absolute())

        # 运行所有测试
        test_sheet_list(file_path)          # 新增：测试工作表列表
        test_row_column_access(file_path)   # 新增：测试行列访问
        test_regex_search(file_path)        # 原有：测试正则搜索
        test_range_get(file_path)           # 原有：测试范围获取
        test_range_update(file_path)        # 原有：测试范围修改

        print("\n" + "=" * 50)
        print("🎉 所有功能测试完成!")

        # 显示最终文件状态
        print(f"\n📄 最终测试文件位置: {file_path}")
        print("💡 您可以打开Excel文件查看修改结果")

    except Exception as e:
        print(f"\n❌ 测试过程发生错误: {e}")
        import traceback
        traceback.print_exc()

    finally:
        # 清理临时文件（可选）
        # if os.path.exists(test_file):
        #     os.remove(test_file)
        pass

if __name__ == "__main__":
    main()
