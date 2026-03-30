#!/usr/bin/env python3
"""
测试 apply_formula 的参数验证修复
"""

import json
import sys
import os
from pathlib import Path
import openpyxl

# 添加项目路径
sys.path.insert(0, str(Path(__file__).parent / "src"))

from excel_mcp_server_fastmcp.server import excel_set_formula

def create_test_excel():
    """创建测试Excel文件"""
    test_file = "/tmp/test_formula_validation.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["B1"] = 20
    wb.save(test_file)

    print(f"✅ 创建测试文件: {test_file}")
    return test_file

def main():
    """测试公式参数验证"""
    print("🔍 测试 apply_formula 参数验证修复")
    print("=" * 60)

    test_file = create_test_excel()

    try:
        # 测试 1: 空字符串
        print("\n1️⃣ 测试空字符串 formula...")
        result1 = excel_set_formula(
            file_path=test_file,
            sheet_name="Sheet1",
            cell_address="C1",
            formula=""
        )
        print(f"结果: {json.dumps(result1, indent=2, ensure_ascii=False)}")
        if not result1.get('success') and '不能为空' in result1.get('message', ''):
            print("✅ 空字符串验证通过")
        else:
            print("❌ 空字符串验证失败")

        # 测试 2: 只有空格的字符串
        print("\n2️⃣ 测试只有空格的 formula...")
        result2 = excel_set_formula(
            file_path=test_file,
            sheet_name="Sheet1",
            cell_address="C2",
            formula="   "
        )
        print(f"结果: {json.dumps(result2, indent=2, ensure_ascii=False)}")
        if not result2.get('success') and '不能为空' in result2.get('message', ''):
            print("✅ 只有空格的字符串验证通过")
        else:
            print("❌ 只有空格的字符串验证失败")

        # 测试 3: 有效公式
        print("\n3️⃣ 测试有效公式...")
        result3 = excel_set_formula(
            file_path=test_file,
            sheet_name="Sheet1",
            cell_address="C3",
            formula="=A1+B1"
        )
        print(f"结果: {json.dumps(result3, indent=2, ensure_ascii=False)}")
        if result3.get('success'):
            print("✅ 有效公式验证通过")
        else:
            print("❌ 有效公式验证失败")

        # 测试 4: 不带等号的公式（可能被视为无效）
        print("\n4️⃣ 测试不带等号的 formula...")
        result4 = excel_set_formula(
            file_path=test_file,
            sheet_name="Sheet1",
            cell_address="C4",
            formula="A1+B1"
        )
        print(f"结果: {json.dumps(result4, indent=2, ensure_ascii=False)}")
        # 这种情况可能成功（openpyxl会自动处理），也可能失败（安全验证）
        print("⚠️ 不带等号的公式结果取决于实现")

        print("\n" + "=" * 60)
        print("✅ 测试完成")

    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)
            print(f"🧹 清理测试文件: {test_file}")

if __name__ == "__main__":
    main()
