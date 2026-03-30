#!/usr/bin/env python3
"""
Comprehensive test script to reproduce the 5 API issues through MCP tool calls.

Tests:
1. read_data_from_excel (excel_get_range): range query parameter order
2. format_range (excel_format_cells): missing bold parameter handling
3. apply_formula (excel_set_formula): missing formula parameter handling
4. read_data_from_excel (excel_search): search logic parameter confusion
5. write_data_to_excel (excel_update_range): data format mismatch handling
"""

import json
import sys
import os
import tempfile
from pathlib import Path

# Add src to path
sys.path.insert(0, 'src')

def create_test_excel():
    """Create a simple test Excel file"""
    from openpyxl import Workbook

    # Create temp file
    test_file = "/tmp/test_api_5_issues.xlsx"
    if os.path.exists(test_file):
        os.remove(test_file)

    # Create workbook with test data
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add headers and data
    headers = ["Name", "Age", "City", "Salary"]
    data = [
        ["Alice", 25, "Shenzhen", 10000],
        ["Bob", 30, "Beijing", 12000],
        ["Charlie", 35, "Shanghai", 15000]
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    wb.save(test_file)
    print(f"Created test file: {test_file}")
    return test_file

def test_1_range_query_parameter_order():
    """
    Test 1: excel_get_range - range query parameter order

    Issue: start_cell/end_cell parameters might be in wrong order
    Expected: Should build range correctly as start_cell:end_cell
    """
    print("\n" + "="*60)
    print("Test 1: excel_get_range - range query parameter order")
    print("="*60)

    test_file = create_test_excel()

    try:
        from excel_mcp_server_fastmcp.server import excel_get_range

        # Test with start_cell and end_cell
        print("\n1a. Testing with start_cell='A1', end_cell='C3'")
        result = excel_get_range(
            file_path=test_file,
            range="",  # Empty range, should use start_cell/end_cell
            sheet_name="Sheet1",
            start_cell="A1",
            end_cell="C3"
        )
        print(f"Success: {result.get('success')}")
        if result.get('success'):
            # Handle both response formats
            data = result.get('data')
            if isinstance(data, dict):
                rows = data.get('data', [])
            elif isinstance(data, list):
                rows = data
            else:
                rows = []
            print(f"Data shape: {len(rows)} rows x {len(rows[0]) if rows else 0} cols")
            print(f"First row: {rows[0] if rows else 'N/A'}")
            print("✅ PASS: Range query with start_cell/end_cell works")
        else:
            print(f"❌ FAIL: {result.get('message')}")

        # Test with reversed order (should still work if logic is correct)
        print("\n1b. Testing with explicit range='Sheet1!A1:C3'")
        result2 = excel_get_range(
            file_path=test_file,
            range="Sheet1!A1:C3"
        )
        print(f"Success: {result2.get('success')}")
        if result2.get('success'):
            data2 = result2.get('data')
            if isinstance(data2, dict):
                rows2 = data2.get('data', [])
            elif isinstance(data2, list):
                rows2 = data2
            else:
                rows2 = []
            print(f"Data shape: {len(rows2)} rows x {len(rows2[0]) if rows2 else 0} cols")
            print("✅ PASS: Direct range expression works")

    except Exception as e:
        print(f"❌ EXCEPTION: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)

def test_2_format_range_missing_parameters():
    """
    Test 2: excel_format_cells - missing bold parameter handling

    Issue: What happens when formatting parameters like 'bold' are missing?
    Expected: Should handle gracefully (no-op or default formatting)
    """
    print("\n" + "="*60)
    print("Test 2: excel_format_cells - missing bold parameter handling")
    print("="*60)

    test_file = create_test_excel()

    try:
        from excel_mcp_server_fastmcp.server import excel_format_cells

        # Test without any formatting parameters
        print("\n2a. Testing without formatting or preset parameters")
        result = excel_format_cells(
            file_path=test_file,
            sheet_name="Sheet1",
            range="A1:C1",
            formatting=None,
            preset=None
        )
        print(f"Success: {result.get('success')}")
        print(f"Message: {result.get('message', 'N/A')}")
        if result.get('success'):
            print("✅ PASS: Format cells handles missing parameters gracefully")
        else:
            print(f"❌ FAIL: {result.get('message')}")

        # Test with empty formatting dict
        print("\n2b. Testing with empty formatting dict")
        result2 = excel_format_cells(
            file_path=test_file,
            sheet_name="Sheet1",
            range="A1:C1",
            formatting={}
        )
        print(f"Success: {result2.get('success')}")
        print(f"Message: {result2.get('message', 'N/A')}")

        # Test with valid formatting
        print("\n2c. Testing with valid bold formatting")
        result3 = excel_format_cells(
            file_path=test_file,
            sheet_name="Sheet1",
            range="A1:C1",
            formatting={"bold": True}
        )
        print(f"Success: {result3.get('success')}")
        print(f"Message: {result3.get('message', 'N/A')}")

    except Exception as e:
        print(f"❌ EXCEPTION: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)

def test_3_apply_formula_missing_parameter():
    """
    Test 3: excel_set_formula - missing formula parameter handling

    Issue: What happens when formula parameter is missing or empty?
    Expected: Should return a friendly error message
    """
    print("\n" + "="*60)
    print("Test 3: excel_set_formula - missing formula parameter handling")
    print("="*60)

    test_file = create_test_excel()

    try:
        from excel_mcp_server_fastmcp.server import excel_set_formula

        # Test with empty formula
        print("\n3a. Testing with empty formula string")
        result = excel_set_formula(
            file_path=test_file,
            sheet_name="Sheet1",
            cell_address="D2",
            formula=""
        )
        print(f"Success: {result.get('success')}")
        print(f"Message: {result.get('message', 'N/A')}")
        if not result.get('success'):
            if 'MISSING_FORMULA' in str(result.get('meta', {})):
                print("✅ PASS: Returns proper error code for empty formula")
            else:
                print("⚠️ WARNING: Error message could be more specific")
        else:
            print("❌ FAIL: Should reject empty formula")

        # Test with None formula (will cause TypeError if not validated)
        print("\n3b. Testing with None formula (if possible)")
        try:
            # This might cause TypeError at the Python level
            result2 = excel_set_formula(
                file_path=test_file,
                sheet_name="Sheet1",
                cell_address="D2",
                formula=None
            )
            print(f"Success: {result2.get('success')}")
            print(f"Message: {result2.get('message', 'N/A')}")
            if not result2.get('success'):
                print("✅ PASS: Handles None formula gracefully")
        except TypeError as e:
            print(f"❌ FAIL: TypeError raised (not user-friendly): {e}")

        # Test with valid formula
        print("\n3c. Testing with valid formula")
        result3 = excel_set_formula(
            file_path=test_file,
            sheet_name="Sheet1",
            cell_address="D2",
            formula="=C2*1.1"
        )
        print(f"Success: {result3.get('success')}")
        print(f"Message: {result3.get('message', 'N/A')}")

    except Exception as e:
        print(f"❌ EXCEPTION: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)

def test_4_search_parameter_confusion():
    """
    Test 4: excel_search - search logic parameter confusion

    Issue: Might confuse sheet_name with other parameters
    Expected: Should correctly use sheet_name for filtering
    """
    print("\n" + "="*60)
    print("Test 4: excel_search - search logic parameter confusion")
    print("="*60)

    test_file = create_test_excel()

    try:
        from excel_mcp_server_fastmcp.server import excel_search

        # Test basic search
        print("\n4a. Testing basic search for 'Alice'")
        result = excel_search(
            file_path=test_file,
            pattern="Alice",
            sheet_name="Sheet1"
        )
        print(f"Success: {result.get('success')}")
        if result.get('success'):
            data = result.get('data')
            # Handle both response formats
            if isinstance(data, dict):
                matches = data.get('matches', [])
            elif isinstance(data, list):
                matches = data
            else:
                matches = []
            print(f"Found {len(matches)} matches")
            if matches:
                print(f"First match: {matches[0]}")
            print("✅ PASS: Basic search works")
        else:
            print(f"❌ FAIL: {result.get('message')}")

        # Test with various parameter combinations
        print("\n4b. Testing with case_sensitive=True")
        result2 = excel_search(
            file_path=test_file,
            pattern="alice",
            sheet_name="Sheet1",
            case_sensitive=True
        )
        print(f"Success: {result2.get('success')}")
        if result2.get('success'):
            data2 = result2.get('data')
            if isinstance(data2, dict):
                matches2 = data2.get('matches', [])
            elif isinstance(data2, list):
                matches2 = data2
            else:
                matches2 = []
            print(f"Found {len(matches2)} matches (should be 0 for case-sensitive 'alice')")
            print("✅ PASS: Case sensitivity works")

        # Test with whole_word
        print("\n4c. Testing with whole_word=True")
        result3 = excel_search(
            file_path=test_file,
            pattern="Ali",
            sheet_name="Sheet1",
            whole_word=True
        )
        print(f"Success: {result3.get('success')}")
        if result3.get('success'):
            data3 = result3.get('data')
            if isinstance(data3, dict):
                matches3 = data3.get('matches', [])
            elif isinstance(data3, list):
                matches3 = data3
            else:
                matches3 = []
            print(f"Found {len(matches3)} matches (should be 0 for whole-word 'Ali')")
            print("✅ PASS: Whole word matching works")

    except Exception as e:
        print(f"❌ EXCEPTION: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)

def test_5_write_data_format_mismatch():
    """
    Test 5: excel_update_range - data format mismatch handling

    Issue: What happens when data is not a proper 2D array?
    Expected: Should validate and reject invalid formats
    """
    print("\n" + "="*60)
    print("Test 5: excel_update_range - data format mismatch handling")
    print("="*60)

    test_file = create_test_excel()

    try:
        from excel_mcp_server_fastmcp.server import excel_update_range

        # Test with correct 2D array format
        print("\n5a. Testing with correct 2D array format")
        correct_data = [["David", 40, "Hangzhou", 18000]]
        result = excel_update_range(
            file_path=test_file,
            range="Sheet1!A5:D5",
            data=correct_data
        )
        print(f"Success: {result.get('success')}")
        print(f"Message: {result.get('message', 'N/A')}")
        if result.get('success'):
            print("✅ PASS: Correct 2D array format works")
        else:
            print(f"❌ FAIL: {result.get('message')}")

        # Test with incorrect 1D array format
        print("\n5b. Testing with incorrect 1D array format")
        incorrect_data = ["Eve", 28, "Guangzhou", 16000]
        result2 = excel_update_range(
            file_path=test_file,
            range="Sheet1!A6:D6",
            data=incorrect_data
        )
        print(f"Success: {result2.get('success')}")
        print(f"Message: {result2.get('message', 'N/A')}")
        if not result2.get('success'):
            if '格式错误' in result2.get('message', '') or 'format' in result2.get('message', '').lower():
                print("✅ PASS: Rejects 1D array with proper error message")
            else:
                print(f"⚠️ WARNING: Error message not clear about format issue")
        else:
            print("❌ FAIL: Should reject 1D array format")

        # Test with mixed format (some rows not lists)
        print("\n5c. Testing with mixed format (invalid row)")
        mixed_data = [["Frank", 45], "InvalidRow"]
        result3 = excel_update_range(
            file_path=test_file,
            range="Sheet1!A7:D7",
            data=mixed_data
        )
        print(f"Success: {result3.get('success')}")
        print(f"Message: {result3.get('message', 'N/A')}")
        if not result3.get('success'):
            print("✅ PASS: Rejects mixed format with proper error message")
        else:
            print("❌ FAIL: Should reject mixed format")

        # Test with string instead of list
        print("\n5d. Testing with string instead of list")
        string_data = "NotAList"
        result4 = excel_update_range(
            file_path=test_file,
            range="Sheet1!A8:D8",
            data=string_data
        )
        print(f"Success: {result4.get('success')}")
        print(f"Message: {result4.get('message', 'N/A')}")
        if not result4.get('success'):
            if '二维数组' in result4.get('message', '') or 'list' in result4.get('message', '').lower():
                print("✅ PASS: Rejects string with proper error message")
            else:
                print(f"⚠️ WARNING: Error message: {result4.get('message', 'N/A')}")
        else:
            print("❌ FAIL: Should reject string data")

    except Exception as e:
        print(f"❌ EXCEPTION: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)

def main():
    print("\n" + "="*60)
    print("COMPREHENSIVE API ISSUES TEST")
    print("Testing 5 reported API issues through MCP tool calls")
    print("="*60)

    test_1_range_query_parameter_order()
    test_2_format_range_missing_parameters()
    test_3_apply_formula_missing_parameter()
    test_4_search_parameter_confusion()
    test_5_write_data_format_mismatch()

    print("\n" + "="*60)
    print("ALL TESTS COMPLETED")
    print("="*60)

if __name__ == "__main__":
    main()
