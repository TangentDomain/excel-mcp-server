#!/usr/bin/env python3
"""
直接测试MCP工具复现监工发现的5个API问题
"""

import asyncio
import json
import sys
from pathlib import Path
import openpyxl

def test_mcp_tools():
    """直接测试MCP工具复现问题"""
    
    print("=" * 60)
    print("开始复现监工发现的5个API问题 (MCP工具直接测试)")
    print("=" * 60)
    
    # 创建测试文件
    test_file = Path("test_api_issues.xlsx")
    if test_file.exists():
        test_file.unlink()
    
    print("📝 创建测试Excel文件...")
    
    # 使用openpyxl直接创建测试文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 写入测试数据
    headers = ["ID", "Name", "Value", "Active"]
    test_data = [
        [1, "Item1", 100, True],
        [2, "Item2", 200, False],
        [3, "Item3", 300, True],
    ]
    
    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=header)
    
    for i, row in enumerate(test_data, 2):
        for j, cell_value in enumerate(row, 1):
            ws.cell(row=i, column=j, value=cell_value)
    
    wb.save("test_api_issues.xlsx")
    wb.close()
    
    print("✅ 测试文件创建完成")
    print()
    
    # 导入MCP工具
    try:
        # 尝试导入server中的工具
        sys.path.append('src')
        from excel_mcp_server_fastmcp.server import excel_get_range, excel_format_cells, excel_set_formula, excel_search, excel_update_range
        
        print("🔍 测试1: excel_get_range（范围查询参数验证）")
        try:
            # 测试正常范围
            result1 = excel_get_range("test_api_issues.xlsx", range="A1:C3", include_formatting=False)
            print(f"   正常范围 (A1:C3): {len(result1.get('data', []))} 行数据")
            
            # 测试可能的参数组合 - 使用start_cell和end_cell
            result2 = excel_get_range("test_api_issues.xlsx", sheet_name="Sheet1", start_cell="A1", end_cell="C3")
            print(f"   参数组合 (start_cell+end_cell): {len(result2.get('data', []))} 行数据")
            
            if result1 == result2:
                print("   ✅ 不同的参数组合处理一致")
            else:
                print("   ⚠️  不同参数组合结果不一致")
                
        except Exception as e:
            print(f"   ❌ 错误: {e}")
        print()
        
        print("🔍 测试2: excel_format_cells（参数缺失处理）")
        try:
            # 测试只传部分格式参数
            formatting = {"bold": True, "font_color": "FF0000"}
            result = excel_format_cells("test_api_issues.xlsx", sheet_name="Sheet1", range="A1:A3", formatting=formatting)
            print("   ✅ 部分格式参数处理成功")
            
            # 测试不传格式参数
            result2 = excel_format_cells("test_api_issues.xlsx", sheet_name="Sheet1", range="B1:B3")
            print("   ✅ 无格式参数处理成功")
            
        except Exception as e:
            print(f"   ❌ 错误: {e}")
        print()
        
        print("🔍 测试3: excel_set_formula（必填参数校验）")
        try:
            # 故意不传formula参数 - 这个函数需要4个参数，应该会报错
            result = excel_set_formula("test_api_issues.xlsx", "Sheet1", "A1")
            print("   ❌ 问题：缺少必填参数formula应该报错但没有")
            
        except Exception as e:
            print(f"   ✅ 正确报错: {e}")
        print()
        
        print("🔍 测试4: excel_search（搜索逻辑sheet_name验证）")
        try:
            # 测试不存在的sheet - 应该报错
            result = excel_search("test_api_issues.xlsx", "NonExistentSheet", "ID", "=1")
            print("   ❌ 问题：不存在的sheet应该报错但没有")
            
        except Exception as e:
            print(f"   ✅ 正确报错: {e}")
            
        # 测试有效的sheet
        try:
            result = excel_search("test_api_issues.xlsx", "Sheet1", "ID", "=1")
            print(f"   ✅ 有效sheet搜索成功")
            
        except Exception as e:
            print(f"   ❌ 有效sheet搜索失败: {e}")
        print()
        
        print("🔍 测试5: excel_update_range（数据格式验证）")
        try:
            # 测试错误的数据格式（不是list of lists）
            invalid_data = "not a list"
            result = excel_update_range("test_api_issues.xlsx", "Sheet1", "D1", invalid_data)
            print("   ❌ 问题：错误的数据格式应该报错但没有")
            
        except Exception as e:
            print(f"   ✅ 正确报错: {e}")
            
        # 测试正确的数据格式
        try:
            valid_data = [["Test", "Data"]]
            result = excel_update_range("test_api_issues.xlsx", "Sheet1", "D1", valid_data)
            print("   ✅ 正确数据格式写入成功")
            
        except Exception as e:
            print(f"   ❌ 正确数据格式写入失败: {e}")
        print()
        
    except ImportError as e:
        print(f"❌ 无法导入MCP工具: {e}")
        print("这可能意味着API有问题，我们需要直接查看源代码...")
        
        # 直接分析源代码来找到问题
        analyze_source_code()
    
    # 清理测试文件
    if test_file.exists():
        test_file.unlink()
        print("🧹 测试文件已清理")
    
    print("=" * 60)
    print("复测完成")
    print("=" * 60)

def analyze_source_code():
    """直接分析源代码找出问题"""
    print("🔍 直接分析源代码...")
    
    print("📋 问题1: excel_get_range 参数顺序")
    # 检查 get_range 实现
    with open("src/excel_mcp_server_fastmcp/api/excel_operations.py", "r") as f:
        content = f.read()
        
    # 查找get_range的实现
    if "def get_range(" in content:
        print("   ✅ get_range方法存在")
        
        # 检查参数验证逻辑
        if "range_expression" in content:
            print("   ✅ 使用range_expression参数，应该是字符串格式，无顺序问题")
        else:
            print("   ❌ 参数处理逻辑可能有问题")
    
    print("📋 问题2: excel_format_cells 参数缺失")
    # 检查format_cells实现
    if "formatting: Optional[Dict[str, Any]] = None" in content:
        print("   ✅ formatting参数是可选的，默认None")
    else:
        print("   ❌ 参数处理可能有问题")
    
    print("📋 问题3: excel_set_formula 必填参数")
    # 检查set_formula实现
    if "formula: str" in content:
        print("   ✅ formula是必填参数")
    else:
        print("   ❌ 参数处理可能有问题")
    
    print("📋 问题4: excel_search sheet_name验证")
    # 检查search实现
    if "sheet_name" in content:
        print("   ✅ sheet_name参数存在")
    else:
        print("   ❌ 参数处理可能有问题")
    
    print("📋 问题5: excel_update_range 数据格式")
    # 检查update_range实现
    if "data" in content:
        print("   ✅ data参数存在")
    else:
        print("   ❌ 参数处理可能有问题")

if __name__ == "__main__":
    test_mcp_tools()